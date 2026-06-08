from pdb import run
from any import *
from .model import *
from sqlalchemy.sql import func,not_,or_,and_
import time
from datetime import datetime
import json,os
from .__default__ import get_user_path
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Border, Side, Alignment, Font, Alignment, NamedStyle,PatternFill,Color
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", 'openpyxl'])
    
SYS_FIELDS = ['sid','rid','uid','ctime','mtime','has_att','modi_uid','wf_status','wf_unit','pid','archived']


# 客户拜访的编辑界面加载校验
@any_route('/api/saier/bol_apply/sfcx/change', methods=['POST', 'GET'])
@require_token
async def view_saier_bol_apply_sfcx_change(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        fphm = j.get('fphm', '')
        kh_id = j.get('kh_id', '')
        msg = '操作成功'
        code = 1
        cxcs = 0
        cxts = 0
        d = run_sql(f"select cxts from fdsq1 where (sfcx='是') and (kh_id='{kh_id}') and (fphm<>'{fphm}')")
        for r in d:
            cxts += int(r.get('cxts', 0))
            cxcs += 1
           
        return json_result(code, msg, {'cxcs':cxcs,'cxts':cxts})
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 客户拜访的编辑界面加载校验
@any_route('/api/saier/bol_apply/fphm/change', methods=['POST', 'GET'])
@require_token
async def view_saier_bol_apply_fphm_change(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        fphm = j.get('fphm', '')
        xybx = j.get('xybx', '')
        khmc = j.get('khmc', '')
        ywry = j.get('ywry', '')
        kh_id = j.get('kh_id', '')
        path = j.get('path', '')
        path2 = path.replace('\\','\\\\')
        rid = j.get('rid', '')
        msg = '操作成功'
        code = 1
        flag = 0
        tshz = '否'
        d = s.query(fdsq1).filter(fdsq1.fphm==fphm, fdsq1.rid!=rid).first()
        if d:
            flag = 1
        else:
            d = s.query(fdsq1sheet).filter(fdsq1sheet.fphm==fphm, fdsq1sheet.pid!=rid).first()
            if d:
                flag = 1
        if flag==1:
            return json_result(-1, '此外销发票号已有放单')  
        
        fpsb = 0
        zg_list = []
        data = {}
        d = run_sql(f"select khmc,chyrq,sjcy1,wxje,htje hzhj,yjje,ywry,ywbm,mjzj,RMBkh khpd,ayjje,myjje,zdry,xybx,kh_id,eta,hxdh,jhfs,dfrq from cymx \
            where (fphm='{fphm}') and (ifnull(chyrq,'')<>'') limit 1")
        if len(d) > 0:
            data = d[0]
            fpsb = 1
        if len(data)==0:
            d = run_sql(f"select khmc,chyrq,ywry,khpd,kh_id,jhfs,xybx,yjje,wxje,mjzj,ayjje,myjje,hzhj from chyjh where (wxfp='{fphm}') limit 1")
            if len(d) > 0:
                tshz = '是'
                data = d[0]
                fpsb = 1
                u = run_sql(f"select bm from ywrybiao where (yhm='{str(data.get('ywry'))}') limit 1")
                if len(u) > 0:
                    data['ywbm'] = u[0].get('bm', '')
        if len(data)==0:
            d = run_sql(f"select khmc,chyrq,ywry,khpd,kh_id,jhfs,xybx,yjje,wxje,mjzj,ayjje,myjje from pgjh where (wxfp='{fphm}') limit 1")
            if len(d) > 0:
                tshz = '是'
                data = d[0]
                fpsb = 1
                u = run_sql(f"select bm from ywrybiao where (yhm='{str(data.get('ywry'))}') limit 1")
                if len(u) > 0:
                    data['ywbm'] = u[0].get('bm', '')
        tjzg = ''
        fkzg = ''
        zjl = ''
        cw_list = []
        if 'BEST PRICE' in khmc.upper():
            c = run_sql(f"select dlr from spwt where ((yhms=2) or (yhms=1))  and ('{path2}' like '%fdsq%') and (dlr<>'{user.username}') order by yhms desc limit 1")
            if len(c) > 0:
                tjzg = c[0].get('dlr', '')     
        path1 = ''
        if ywry!='' and ywry!=None:
            org = get_user_path(ywry)
            path1 = org.get('path','')
            logger.error(path1)
        # path3 = path1.replace('\\','\\\\')
        if path1!='' and path!=path1:
            if 'BEST PRICE' in khmc.upper():
                c = run_sql(f"select dlr from spwt where ((yhms=2) or (yhms=1))  and ('{path1}' like '%fdsq%') and (dlr<>'{user.username}') order by yhms desc")
                zg_list = [r.get('dlr', '') for r in c]
            else:
                c = run_sql(f"select dlr from spwt where ((yhms<11)  and ('{path1}' like '%fdsq%') and (dlr<>'{user.username}') order by yhms desc")
                zg_list = [r.get('dlr', '') for r in c]
            c = run_sql(f"select dlr from spwt where (yhms>20) and (yhms<31) and ('{path1}' like '%fdsq%') order by yhms desc limit 1")
            if len(c) > 0:
                fkzg = c[0].get('dlr', '')
            c = run_sql(f"select dlr from spwt where ((yhms=2) or (yhms=1)) and ('{path1}' like '%fdsq%') order by yhms desc limit 1")
            if len(c) > 0:
                zjl = c[0].get('dlr', '')
        
        if 'BEST PRICE' in khmc.upper():
            c = run_sql(f"select dlr from spwt where (yhms=11) and ('{path1}' like '%fdsq%') order by yhms desc")
            cw_list = [r.get('dlr', '') for r in c]
        else:
            c = run_sql(f"select dlr from spwt where (yhms>11) and (yhms<21) and ('{path1}' like '%fdsq%') order by yhms desc")
            cw_list = [r.get('dlr', '') for r in c]

        data['tshz'] = tshz
        data['fpsb'] = fpsb
        data['tjzg'] = tjzg
        data['fkzg'] = fkzg
        data['zjl'] = zjl
        data['zg_list'] = zg_list
        data['cw_list'] = cw_list
        data['path1'] = path1

        return json_result(code, msg, data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 客户拜访的编辑界面加载校验
@any_route('/api/saier/bol_apply/child/fphm/change', methods=['POST', 'GET'])
@require_token
async def view_saier_bol_apply_child_fphm_change(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        fphm = j.get('fphm', '')
        zbfp = j.get('zbfp', '')
        xybx = j.get('xybx', '')
        kh_id = j.get('kh_id', '')
        rmbkh = j.get('rmbkh', '')
        rid = j.get('rid', '')
        msg = '操作成功'
        code = 1
        flag = 0
        if zbfp!='' and zbfp!=None and fphm!=zbfp:
            d = s.query(fdsq1).filter(fdsq1.fphm==fphm, fdsq1.rid!=rid).first()
            if d:
                flag = 1
            else:
                d = s.query(fdsq1sheet).filter(fdsq1sheet.fphm==fphm, fdsq1sheet.pid!=rid).first()
                if d:
                    flag = 1
        if flag==1:
            return json_result(-1, '此外销发票号已有放单')  
        data = {}
        d = run_sql(f"select khmc,chyrq,sjcy1,wxje,htje hzhj,yjje,ywry,ywbm,mjzj,RMBkh khpd,ayjje,myjje,zdry,xybx,kh_id,eta,hxdh,jhfs,dfrq from cymx \
            where (fphm='{fphm}') and (RMBkh='{rmbkh}') and (xybx='{xybx}') and (kh_id='{kh_id}') and (ifnull(chyrq,'')<>'') limit 1")
        if len(d) > 0:
            data = d[0]
        if len(data)==0:
            d = run_sql(f"select khmc,chyrq,ywry,khpd,kh_id,jhfs,xybx,yjje,wxje,mjzj,ayjje,myjje,hzhj from chyjh where (wxfp='{fphm}') limit 1")
            if len(d) > 0:
                data = d[0]
                u = run_sql(f"select bm from ywrybiao where (yhm='{str(data.get('ywry'))}') limit 1")
                if len(u) > 0:
                    data['ywbm'] = u[0].get('bm', '')
        if len(data)==0:
            d = run_sql(f"select khmc,chyrq,ywry,khpd,kh_id,jhfs,xybx,yjje,wxje,mjzj,ayjje,myjje from pgjh where (wxfp='{fphm}') limit 1")
            if len(d) > 0:
                data = d[0]
                u = run_sql(f"select bm from ywrybiao where (yhm='{str(data.get('ywry'))}') limit 1")
                if len(u) > 0:
                    data['ywbm'] = u[0].get('bm', '')
        hjsh = 0
        c = run_sql(f"select ifnull(sum(ifnull(sydje,0)),0) as sjsh1 from khdjsheet where (wxht like '%%{fphm}%%')")
        if len(c) > 0:
            hjsh = c[0].get('sjsh1', 0)
        c = run_sql(f"select ifnull(sum(ifnull(sydje2,0)),0) as syhj from krshsheet where (fphm='{fphm}') or (fphm=concat('{fphm}','A')) \
            or (fphm=concat('{fphm}','B')) or (fphm=concat('{fphm}','C')) or (fphm=concat('{fphm}','D')) or (fphm=concat('{fphm}','E')) or (fphm=concat('{fphm}','F'))")
        if len(c) > 0:
            hjsh += c[0].get('syhj', 0)
        data['hjsh'] = hjsh

        return json_result(code, msg, data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 客户拜访的记录保存前校验
@any_route('/api/saier/bol_apply/child/delete/before', methods=['POST', 'GET'])
@require_token
async def view_saier_bol_apply_child_delete_before(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        fphm = j.get('fphm', '') 
        if fphm=='' or fphm==None:
            return json_result(-1, '发票号码不能为空')
        d = s.query(cymx).filter(cymx.fphm==fphm,func.ifnull(cymx.fphm,"")!="").all()
        for r in d:
            r.hxdate='否'
            r.modi_uid=user.rid
            r.mtime=time.strftime('%Y-%m-%d %H:%M:%S')
            s.add(r)
        s.commit()
        return json_result(1, '操作成功')
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 客户拜访的编辑界面加载校验
@any_route('/api/saier/bol_apply/load/check', methods=['POST', 'GET'])
@require_token
async def view_saier_bol_apply_load_check(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        path = j.get('path', '')
        tjzg = j.get('tjzg', '')
        tjfk = j.get('tjfk', '')
        tjcw = j.get('tjcw', '')
        zjl = j.get('zjl', '')
        zjlsp = j.get('zjlsp', '')
        cwsp = j.get('cwsp', '')
        fksp = j.get('fksp', '')
        zgsp = j.get('zgsp', '')
        shje = j.get('shje', 0)
        khmc = j.get('khmc', '')
        kh_id = j.get('kh_id', '')
        fphm = j.get('fphm', '')
        lines = j.get('lines', [])
        cw_list = []
        zg_list = []
        qg_list = []
        cxcs = 0
        syrq = ''
        ljsh = 0
        ljsh1 = 0   
        ljwsh = 0
        data = {}
        if path=='' or path==None:
            org = get_user_path(user.username)
            path = org.get('path','')
        if path!='' and path!=None:
            path1 = path.replace('\\','_')
            if tjzg =='' or tjzg==None:
                if 'BEST PRICE' in khmc.upper():
                    
                    c = run_sql(f"select dlr from spwt where ((yhms=2) or (yhms=1))  and ('{path1}' like CONCAT('%', replace(fdsq, '\\\\', '_'), '%')) and (dlr<>'{user.username}') order by yhms desc")
                    zg_list = [r.get('dlr', '') for r in c]
                    tjzg = c[0].get('dlr', '') if len(c)>0 else ''
                else:
                    c = run_sql(f"select dlr from spwt where (yhms<11)  and ('{path1}' like CONCAT('%', replace(fdsq, '\\\\', '_'), '%')) and (dlr<>'{user.username}') order by yhms desc")
                    zg_list = [r.get('dlr', '') for r in c]
                    tjzg = c[0].get('dlr', '') if len(c)>0 else ''
            if tjfk =='' or tjfk==None:    
                c = run_sql(f"select dlr from spwt where (yhms>20) and (yhms<31) and ('{path1}' like CONCAT('%', replace(fdsq, '\\\\', '_'), '%')) order by yhms desc limit 1")
                tjfk = c[0].get('dlr', '') if len(c) > 0 else ''
            if zjl =='' or zjl==None:
                c = run_sql(f"select dlr from spwt where ((yhms=2) or (yhms=1)) and ('{path1}' like CONCAT('%', replace(fdsq, '\\\\', '_'), '%')) order by yhms desc limit 1")
                zjl = c[0].get('dlr', '') if len(c) > 0 else ''
            if 'BEST PRICE' in khmc.upper():
                c = run_sql(f"select dlr from spwt where (yhms=11) and ('{path1}' like CONCAT('%', replace(fdsq, '\\\\', '_'), '%')) order by yhms desc")
                cw_list = [r.get('dlr', '') for r in c]
            else:
                c = run_sql(f"select dlr from spwt where (yhms>11) and (yhms<21) and ('{path1}' like CONCAT('%', replace(fdsq, '\\\\', '_'), '%')) order by yhms desc")
                cw_list = [r.get('dlr', '') for r in c]
        if fphm!='' and fphm!=None:
            qg_list = run_sql(f"select scrq,zlmc,scry,zlwy from zlsc where (fphm='{fphm}') and (ly='放单申请')")

        if zgsp == '待审批' and tjcw == user.username and cwsp != '通过' and shje == 0:
            c = run_sql(f"select syrq from krshsheet where (fphm='{fphm}') or (fphm=concat('{fphm}','A')) \
                or (fphm=concat('{fphm}','B')) or (fphm=concat('{fphm}','C')) or (fphm=concat('{fphm}','D')) or (fphm=concat('{fphm}','E')) \
                or (fphm=concat('{fphm}','F')) order by syrq desc limit 1")
            if len(c) > 0:
                syrq = c[0].get('syrq', None)
            c = run_sql(f"select ifnull(cswsh, 0) cswsh from kh where (kh_id='{kh_id}') limit 1")
            if len(c) > 0:
                ljwsh = c[0].get('cswsh', 0)
            c = run_sql(f"select ifnull(sum(ifnull(wshje,0)),0) as wshje1 from fdsq1 where (kh_id='{kh_id}') and (fphm<>'{fphm}') and (cpjq<>'是' or ifnull(cpjq,'')='')")
            if len(c) > 0:
                ljwsh = c[0].get('wshje1', 0) + ljwsh
            for r in lines:
                ljsh1 = 0
                fphm = r.get('fphm', '')
                c = run_sql(f"select ifnull(sum(ifnull(sydje,0)),0) as sjsh1 from khdjsheet where (wxht like '%{fphm}%')")
                if len(c) > 0:
                    ljsh += c[0].get('sjsh1', 0)
                    ljsh1 += c[0].get('sjsh1', 0)
                c = run_sql(f"select ifnull(sum(ifnull(sydje2,0)),0) as syhj from krshsheet where (fphm='{fphm}') or (fphm=concat('{fphm}','A')) \
                    or (fphm=concat('{fphm}','B')) or (fphm=concat('{fphm}','C')) or (fphm=concat('{fphm}','D')) or (fphm=concat('{fphm}','E')) or (fphm=concat('{fphm}','F'))")
                if len(c) > 0:
                    ljsh += c[0].get('syhj', 0)
                    ljsh1 += c[0].get('syhj', 0)
                data[r.get('rid','')] = ljsh1

        if zjlsp == '待审批' and tjfk == user.username and fksp != '通过' and zgsp == '通过':
            c = run_sql(f"select count(cxts) cxcs from fdsq1 where (sfcx='是') and (kh_id='{kh_id}')")
            cxcs = c[0].get('cxcs', 0) if len(c)>0 else 0

        return json_result(1, '操作成功', {'tjzg':tjzg,'tjfk':tjfk,'zjl':zjl,'cw_list':cw_list,'zg_list':zg_list,'path':path,'qg_list':qg_list,'cxcs':cxcs,'ljsh':ljsh,'ljsh1':ljsh1,'ljwsh':ljwsh,'syrq':syrq,'data':data})
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 客户拜访的记录保存后校验
@any_route('/api/saier/bol_apply/batch/confirm', methods=['POST', 'GET'])
@require_token
async def view_saier_bol_apply_batch_confirm(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        c = run_sql(f"select rid from sys_user where (username='{user.username}') and ((position like '%财务%') or (memo like '%财务%')) limit 1")
        if len(c) == 0:
            return json_result(-1, '不好意思,只有财务人员才能操作此功能')
        c = run_sql(f"select distinct fphm from fdsq1sheet where (cpjq='否') or (ifnull(cpjq,'')='')")
        for r in c:
            d = run_sql(f"select fphm from krshsheet where (sfjq='是') and (fphm like '%{r.get('fphm','')}%')")
            if len(d) == 0:
                continue
            f = s.query(fdsq1sheet).filter(fdsq1sheet.fphm==bfdh).all()
            for l in f:
                l.cpjq = '是'
                s.add(l)
                rid = str(l.pid)
                m = s.query(fdsq1sheet).filter(fdsq1sheet.pid==rid,or_(func.ifnull(fdsq1sheet.cpjq,"")=="",func.ifnull(fdsq1sheet.cpjq,"")=="否")).first()
                if not m:
                    continue
                n = s.query(fdsq1).filter(fdsq1.rid==rid).first()
                if n:
                    n.cpjq = '是'
                    s.add(n)
        s.commit()

        return json_result(1, '操作成功')
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 客户拜访的记录保存后校验
@any_route('/api/saier/bol_apply/batch/incomes', methods=['POST', 'GET'])
@require_token
async def view_saier_bol_apply_batch_incomes(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        fphm = j.get('fphm', '')
        kh_id = j.get('kh_id', '')
        lines = j.get('lines', [])
        syrq = ''
        ljwsh = 0
        c = run_sql(f"select syrq from krshsheet where (fphm='{fphm}') or (fphm=concat('{fphm}','A')) \
            or (fphm=concat('{fphm}','B')) or (fphm=concat('{fphm}','C')) or (fphm=concat('{fphm}','D')) or (fphm=concat('{fphm}','E')) \
            or (fphm=concat('{fphm}','F')) order by syrq desc limit 1")
        if len(c) > 0:
            syrq = c[0].get('syrq', None)
        c = run_sql(f"select cswsh from kh where (kh_id='{kh_id}') limit 1")
        if len(c) > 0:
            ljwsh = c[0].get('cswsh', 0)

        data = {}
        ljsh = 0
        for r in lines:
            ljsh1 = 0
            fphm = r.get('fphm','')
            c = run_sql(f"select ifnull(sum(ifnull(sydje,0)),0) as sjsh1 from khdjsheet where (wxht like '%{fphm}%')")
            if len(c) > 0:
                ljsh += c[0].get('sjsh1', 0)
                ljsh1 += c[0].get('sjsh1', 0)
            c = run_sql(f"select ifnull(sum(ifnull(sydje2,0)),0) as syhj from krshsheet where (fphm='{fphm}') or (fphm=concat('{fphm}','A')) \
                or (fphm=concat('{fphm}','B')) or (fphm=concat('{fphm}','C')) or (fphm=concat('{fphm}','D')) or (fphm=concat('{fphm}','E')) or (fphm=concat('{fphm}','F'))")
            if len(c) > 0:
                ljsh += c[0].get('syhj', 0)
                ljsh1 += c[0].get('syhj', 0)
            data[r.get('rid','')] = ljsh1

        return json_result(1, '操作成功', {'syrq':syrq,'ljwsh':ljwsh,'ljsh':ljsh,'data':data})
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 客户拜访的记录保存后校验
@any_route('/api/saier/bol_apply/batch/shipment', methods=['POST', 'GET'])
@require_token
async def view_saier_bol_apply_batch_shipment(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        code = 1
        msg = '操作成功'
        rid = j.get('rid', '')
        lines = j.get('lines', [])
        data = []
        flag = 0
        for r in lines:
            dzry = r.get('dzry', '')
            fphm = r.get('fphm','')
            if user.username!=dzry:
                continue
            c = s.query(cymx).filter(cymx.fphm==fphm).first()
            if c :
                flag = 1
                dfrq = time.strftime('%Y-%m-%d')
                if c.dfrq==None or c.dfrq=='':
                    c.hxdate1 = '是'
                    c.dfrq = dfrq
                    c.modi_uid=user.rid
                    c.mtime=time.strftime('%Y-%m-%d %H:%M:%S')
                    s.add(c)
                else:
                    dfrq = str(c.dfrq)
                d = s.query(fdsq1).filter(fdsq1.rid==rid, fdsq1.fphm==fphm).first()
                if d:
                    d.dfrq = dfrq
                    d.modi_uid=user.rid
                    d.mtime=time.strftime('%Y-%m-%d %H:%M:%S')
                    s.add(d)
                m = s.query(fdsq1sheet).filter(fdsq1sheet.pid==rid,fdsq1sheet.fphm==fphm).all()
                for d in m:
                    d.dfrq = dfrq
                    d.dzdf = '是'
                    d.modi_uid=user.rid
                    d.mtime=time.strftime('%Y-%m-%d %H:%M:%S')
                    s.add(d)
            data.append({'fphm':fphm,'cdmc':str(c.cdmc),'chyrq':str(c.chyrq)[:10]})
        if flag==1:
            d = s.query(sys_task).filter(sys_task.module=='放单申请',sys_task.pid==rid).all()
            for l in d:
                s.query(sys_task_receiver).filter(sys_task_receiver.pid==l.rid).delete()
                s.delete(l)

            s.commit()

        sbs_path = ''
        filename = ''
        if len(data) > 0:
            # 创建字体
            font_ = Font(
                name="Arial",
                size=12,
                bold=False
            )
            # 创建单元格样式
            alight_ = Alignment(
                horizontal='center',  # 水平对齐方式:center, left, right
                vertical='center',  # 垂直对齐方式: center, top, bottom
                wrap_text=True
            )
            # 创建边框样式
            border_ = Border(left=Side(style='thin',color='FF000000'),
                    right=Side(style='thin',color='FF000000'),
                    top=Side(style='thin',color='FF000000'),
                    bottom=Side(style='thin',color='FF000000'))
        
            wb = Workbook() # 新建Excel文件
            ws = wb.active # wb.create_sheet() # 默认激活第一个sheet
            ws.column_dimensions['A'].width = 14.3
            ws['A1'] = '发票号码'
            ws.column_dimensions['B'].width = 28
            ws['B1'] = '货代名称'
            ws.column_dimensions['C'].width = 14.3
            ws['C1'] = '出运日期'
            ws.column_dimensions['D'].width = 14.3
            ws['D1'] = '制单人员'
            ws.column_dimensions['E'].width = 14.3
            i = 1
            for r in data:
                i += 1
                ws[f'A{i}'] = r.get('fphm','')
                ws[f'B{i}'] = r.get('cdmc','')
                ws[f'C{i}'] = r.get('chyrq','')
                ws[f'D{i}'] = user.username
            for r in tuple(ws['A%d:S%d'%(1,i)]):
                for c in r:
                    # c.font = font_
                    c.alignment = alight_
                    # c.border=border_

            path = config.tmp_path
            if not os.path.exists(path):
                make_dirs(path)
            filename=str(get_uuid())+'.xlsx'
            fn = os.path.join(path, filename)
            sbs_path = path[-10:] + '/' + filename
            wb.save(fn)
        else:
            code = 0
            msg = '未找到符合条件的放单记录,操作取消'

        return json_result(code, msg, filename)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 放单申请的提交主管改变取数校验
@any_route('/api/saier/bol_apply/tjzg/change', methods=['POST', 'GET'])
@require_token
async def view_saier_bol_apply_tjzg_change(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        code = 1
        msg = '操作成功'
        tjzg = j.get('tjzg', '')
        bjdl = j.get('bjdl', '')
        bjdl = bjdl.replace('\\','_')
        d = run_sql(f"select rid from spwt where (dlr='{tjzg}') and ('{bjdl}' like concat('%', replace(fdsq, '\\\\', '_'), '%')) and (yhms<11) order by yhms desc limit 1")
        logger.error(d)
        if len(d) == 0:
            return json_result(-1, '此人没有审批权限,请重新选择')

        return json_result(code, msg)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 放单申请的提交财务改变取数校验
@any_route('/api/saier/bol_apply/tjcw/change', methods=['POST', 'GET'])
@require_token
async def view_saier_bol_apply_tjcw_change(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        code = 1
        msg = '操作成功'
        rid = j.get('rid', '')
        tjzg = j.get('tjzg', '')
        bjdl = j.get('bjdl', '')
        khmc = j.get('khmc', '')
        bjdl = bjdl.replace('\\','_')
        tjcw = j.get('tjcw', '')
        flag = 0
        if khmc != '' and khmc != None and 'BEST PRICE' in khmc.upper():
            flag = 1
        if (tjzg == '' or tjzg == None) and flag==1:
            d = run_sql(f"select dlr from spwt where ('{bjdl}' like concat('%', replace(fdsq, '\\\\', '_'), '%')) and (yhms=2) order by yhms desc limit 1")
            if len(d) > 0:
                return json_result(-1, '此人没有审批权限,请重新选择')
            tjzg = d[0].get('dlr', '')
        old_tjcw = ''
        d = run_sql(f"select tjcw from fdsq1 where rid = '{rid}' limit 1")
        if len(d) > 0:
            old_tjcw = d[0].get('tjcw', '')
        if old_tjcw != tjcw:
            d = run_sql(f"select dlr from spwt where (dlr='{tjcw}') and ('{bjdl}' like concat('%', replace(fdsq, '\\\\', '_'), '%')) and (yhms>10) and (yhms<21) order by yhms desc limit 1")
            if len(d) == 0:
                return json_result(-1, '此人没有审批权限,请重新选择')

        return json_result(code, msg, {'tjzg': tjzg, 'tjcw': tjcw})
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 放单申请的提交风控改变取数校验
@any_route('/api/saier/bol_apply/tjfk/change', methods=['POST', 'GET'])
@require_token
async def view_saier_bol_apply_tjfk_change(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        code = 1
        msg = '操作成功'
        tjfk = j.get('tjfk', '')
        bjdl = j.get('bjdl', '')
        bjdl = bjdl.replace('\\','_')
        d = run_sql(f"select rid from spwt where (dlr='{tjfk}') and ('{bjdl}' like concat('%', replace(fdsq, '\\\\', '_'), '%')) and (yhms<31) and (yhms>20) order by yhms desc limit 1")
        logger.error(d)
        if len(d) == 0:
            return json_result(-1, '此人没有审批权限,请重新选择')
            
        return json_result(code, msg)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 放单申请的提交总经理改变取数校验
@any_route('/api/saier/bol_apply/tjzjl/change', methods=['POST', 'GET'])
@require_token
async def view_saier_bol_apply_tjzjl_change(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        code = 1
        msg = '操作成功'
        tjzjl = j.get('tjzjl', '')
        bjdl = j.get('bjdl', '')
        bjdl = bjdl.replace('\\','_')
        # rid = j.get('rid', '')
        d = run_sql(f"select rid from spwt where (dlr='{tjzjl}') and ('{bjdl}' like concat('%', replace(fdsq, '\\\\', '_'), '%')) and (yhms=1) order by yhms desc limit 1")
        logger.error(d)
        if len(d) == 0:
            return json_result(-1, '此人没有审批权限,请重新选择')
        
        # old_tjzjl = ''
        # d = run_sql(f"select tjzjl from fdsq1 where rid = '{rid}' limit 1")
        # if len(d) > 0:
        #     old_tjzjl = d[0].get('tjzjl', '')
        # if old_tjzjl != tjzjl:
        #     d = run_sql(f"select dlr from spwt where (dlr='{tjzjl}') and ('{bjdl}' like concat('%', replace(fdsq, '\\\\', '_'), '%')) and (yhms=1) order by yhms desc limit 1")
        #     if len(d) == 0:
        #         return json_result(-1, '此人没有审批权限,请重新选择')
            
        return json_result(code, msg)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 放单申请的提交总经理改变取数校验
@any_route('/api/saier/bol_apply/cwsp/change', methods=['POST', 'GET'])
@require_token
async def view_saier_bol_apply_cwsp_change(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        code = 1
        msg = '操作成功'
        khmc = j.get('khmc', '')
        zm = ''
        ms = 0
        d = run_sql(f"select xm from cyzglsheet where (bz='{khmc}') and (zm='放单免审公司') limit 1")
        logger.error(d)
        if len(d) > 0:
            zm = d[0].get('xm', '')
            ms = 1
            
        return json_result(code, msg, {'zm': zm, 'ms': ms})
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()