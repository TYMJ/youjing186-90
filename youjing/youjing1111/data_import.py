# -*- coding:utf-8 -*-
'''
Author: Grays
Date: 2022-11-02 13:31:26
LastEditors: Grays
LastEditTime: 2022-11-02 17:56:09
Description:
'''
from dataclasses import field
from time import sleep
from any import *
# from openpyxl import *
from starlette.background import BackgroundTask
from asyncio.log import logger
import os

from .order_apply import order_apply_one_new
from .model import *
import subprocess

try:
    from PIL import Image
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", 'PIL'])
try:
    from openpyxl import load_workbook
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", 'openpyxl'])

_IMPORT_TOPIC = 'WHALE_IMPORT'

def _make_msg(msg,code=0,position=0,max=0):
    return{
        'msg':msg,
        'code':code,
        'position':position,
        'max':max
    }

async def _import_data(fn,user,form):
    check = form.get('option.check','true')
    check_child = form.get('option.check_child','true')
    image_main = form.get('option.image','true')
    image_child = form.get('option.image_child','true')
    if not os.path.exists(fn):
        messages.pub(_make_msg('无法找到文件',-1))
        return
    wb = load_workbook(fn)
    path = config.get_today_upload_path()
    if not os.path.exists(path):
        make_dirs(path)
    errors = []
    ms = {}
    data = {}
    for sheet in wb.worksheets:
        where_field = ''
        check_json = {}
        name = sheet.title
        if sheet.max_row<=1:
            errors.append(f'无法找到{name}对应的表')
            continue

        names = name.split('.')
        table = ''
        parent =''
        parent_field =''
        keys = {}
        if(len(names))==1:
            m = get_module(name)
            table = m.table_name
        elif(len(names)==2):
            m = ms.get(names[0],None)
            group = m.group_by_name(names[1])
            if not group:
                errors.append(f'无法找到{name}对应的表')
                continue
            else:
                if not group.is_table:
                    errors.append(f'{name}不是表')
                    continue
                table = group.db.table_name
                parent = names[0]
                parent_field = sheet.cell(1,1).value

        if m==None:
            errors.append(f'无法找到{name}对应的表')
            continue
        from_column =1
        if parent!='':
            from_column = 2
        fields =[]
        if parent=='':
            image_col = 'A' 
        else:
            image_col = 'B' 
        num = len(sheet._images)
        i = 0
        # 遍历sheet中的图片
        # await messages.pub(_make_msg('正在导入数据',code=2,position=1,max=sheet.max_row-1),topic=_IMPORT_TOPIC)  
        if (image_main=='true' or image_child=='true'):
            for image1 in sheet._images:
                i += 1
                image_anchor = image1.anchor
                image= image_anchor._from
                col, colOff, row, rowOff=image.col,image.colOff,image.row,image.rowOff
                # print(col,colOff,row,rowOff)
                #print(image_anchor)
                # 将图像数据保存到文件
                img = Image.open(image1.ref).convert("RGB")
                # img = np.array(img)
                image_name = sheet[image_col+str(row+1)].value
                file_name = os.path.join(path,str(image_name)+'.png')
                img.save(file_name)
                # await messages.pub(_make_msg('正在导出图片',code=2,position=i,max=num),topic=_IMPORT_TOPIC)
        
        j = 1
        _name = sheet.cell(1,j).value
        while _name !="" and _name != None:
        # for col in range(from_column,sheet.max_column+1):
            field = m.field_by_full_name(f'{name}.{_name}')
            if parent=='':
                if j == 1:
                    where_field = field.db.name
            else:
                if j == 2:
                    where_field = field.db.name
            if field!=None:
                fields.append({
                    'name':_name,
                    'field':field.db.name,
                    'col':j
                })
            j+=1
            _name = sheet.cell(1,j).value
                
        records = []
        i = 2
        row_str = sheet.cell(i,1).value
        # for row in range(2,sheet.max_row+1):
        while row_str != "" and row_str != None:
            if parent=='':
                where_value = str(sheet.cell(i,1).value)
            else:
                where_value = str(sheet.cell(i,2).value)
            rid = get_uuid()
            if not where_value in check_json.keys():
                check_json[where_value]=rid
            # else :
            #     rid = check_json[where_value]
            where_value = ''
            _data = {
                'rid':rid,
                'uid':user.rid,
            }
            if parent!='':
                key = sheet.cell(i,1).value
                __p = data.get(parent,None)
                if __p==None:
                    continue
                pid = __p['keys'].get(key,None)
                if pid==None:
                    msg_str ='工作表['+ name + ']中'  + parent_field + '为' + str(key) + '的数据，与' + names[0] + parent_field +'不匹配！请修改后再导入！'
                    await messages.pub(_make_msg(msg_str,code=-1),topic=_IMPORT_TOPIC)
                    return
                _data['pid'] = pid
            col_index = 0
            for f in fields:
                col_index += 1
                # logger.error('-------file_name-------')
                if ((image_main=='true' and parent=='') or (parent!='' and image_child=='true')) and col_index==len(fields):
                    image_name = sheet[image_col+str(i)].value
                    file_name = os.path.join(path,str(image_name)+'.png')
                    # logger.error(file_name)
                    if os.path.exists(file_name):
                        photo_list = []
                        sbs_path = path[-10:]
                        photo_list.append({"src":str(sbs_path)+"/"+str(image_name)+".png","name":str(image_name)+".png"})
                        Barcode = str(photo_list)
                        _data[f['field']]=Barcode.replace("'",'"')
                    else:
                        _data[f['field']]=None
                else:
                    v = sheet.cell(i,f['col']).value
                    _data[f['field']] = v
                if f['col']==1:
                    keys[v] = _data['rid']

            records.append(_data)
            i += 1
            row_str = sheet.cell(i,1).value
            pass
        data[name]= {
                'table':table,
                'parent':parent,
                'name':sheet.title,
                'parent_field':parent_field,
                'fields':fields,
                'keys':keys,
                'records':records,
                'where_field':where_field
                }
        ms[name] = m
        
    count = 0
    # print(data)
    for v in data.values():
        count =count + len(v['records'])
    process = 0
    s = Session()
    count = count +1 #算上commit的
    try:
        i=0
        j=0
        row_list = []
        for v in data.values():
            model = get_model_by_table_name(v['table'])
            where_field = v['where_field']
            check_list = []
            repeat_list = []
            for d in v['records']:
                if (v['parent']=='' and (check=='true' or check==True )) or (v['parent']!='' and (check_child=='true' or check_child==True )):
                    if d[where_field] in check_list:
                        j = j + 1
                        repeat_list.append(str(d[where_field]))
                        continue
                    else:
                        check_list.append(str(d[where_field]))

                    l_sql = 'select rid from '+str(v['table']) +' where '+ str(where_field) + '="'+ str(d[where_field]) +'" limit 1'
                    sql_return = run_sql(l_sql)
                    if len(sql_return)>0:
                        repeat_list.append(str(d[where_field]))
                        j = j + 1
                        continue

                o = model()
                for key  in d.keys():
                    setattr(o,key,d[key])
                s.add(o)
                i = i + 1
                process = process+1
                if process %5==0:
                    await messages.pub(_make_msg('正在导入',code=1,position=process,max=count),topic=_IMPORT_TOPIC)
            logger.error(repeat_list)
            row_list.append(v['name']+':'+','.join(repeat_list))

        # await messages.pub(_make_msg('正在导入',code=1,position=process,max=count),topic=_IMPORT_TOPIC)
        s.commit()
        sleep(0.5)
        if (j>0):
            msg_str = '成功导入'+str(i)+'条记录,重复'+str(j)+'条记录,'+';'.join(row_list)
        else:
            msg_str = '成功导入'+str(i)+'条记录'
        await messages.pub(_make_msg(msg_str,code=2,position=count,max=count),topic=_IMPORT_TOPIC)
    except Exception as e:
        logger.error(trace_error())
        await messages.pub(_make_msg(trace_error(),code=-1),topic=_IMPORT_TOPIC)
        s.rollback()
    finally:
        s.close()
    pass


@any_route('/api/whale/data/import', methods=['POST', 'GET'])
@require_token
async def api_whale_data_import(request):
    form = await request.form()
    # print(form)
    user = request.current_user
    file  = form.get('file.raw',None)
    file_name = form.get('file.name','')
    file_size = form.get('file.size',0)
    # option = form.get('option', None)

    if is_none(file):
        return json_result(ERR_PARAM_NOT_ENOUGH)
    fn = get_tmp_abs_file_name(get_suffix(file_name))
    logger.info(fn)
    data = await file.read()
    write_file(fn,data,'wb',encoding=None)
    task =BackgroundTask(_import_data,fn,user,form)
    return json_result(task = task)
    pass


async def _import_sub_child(fn,user,form):
    # check_child = form.get('option.check_child','true')
    if not os.path.exists(fn):
        messages.pub(_make_msg('无法找到文件',-1))
        return
    wb = load_workbook(fn)
    path = config.get_today_upload_path()
    if not os.path.exists(path):
        make_dirs(path)
    errors = []
    # ms = {}
    data = {}
    for sheet in wb.worksheets:
        name = sheet.title
        names = name.split('.')
        keys = {}

        if(len(names)<3):
            errors.append(f'无法找到{names}对应的子子表')
            continue
        m = get_module(names[0])

        if m==None:
            errors.append(f'无法找到{name}对应的表')
            continue
        c_group = m.group_by_name(names[1])

        if not c_group:
            errors.append(f'无法找到{c_group}对应的表')
            continue
        if not c_group.is_table:
            errors.append(f'{c_group}不是表')
            continue

        s_group = m.group_by_name(names[2])
        if not s_group:
            errors.append(f'无法找到{s_group}对应的表')
            continue
        if not s_group.is_table:
            errors.append(f'{s_group}不是表')
            continue
        
        m_table = m.table_name
        c_table = c_group.db.table_name
        s_table = s_group.db.table_name
        i = 0
        fields =[]
        j = 1
        
        _name = sheet.cell(1,j).value
        field = m.field_by_full_name(f'{names[0]}.{_name}')
        m_where_field = field.db.name
        fields.append({
            'name':_name,
            'field':m_where_field,
            'col':j
        })       
        j = 2
        _name = sheet.cell(1,j).value
        field = m.field_by_full_name(f'{names[0]}.{names[1]}.{_name}')
        c_where_field = field.db.name
        fields.append({
            'name':_name,
            'field':c_where_field,
            'col':j
        })
        j = 3
        _name = sheet.cell(1,j).value
        field = m.field_by_full_name(f'{names[0]}.{names[2]}.{_name}')
        s_where_field = field.db.name
        fields.append({
            'name':_name,
            'field':s_where_field,
            'col':j
        })
        j = 4
        _name = sheet.cell(1,j).value
        while _name !="" and _name != None:
            # logger.error(fields)
            field = m.field_by_full_name(f'{names[0]}.{names[2]}.{_name}')
            if field!=None:
                fields.append({
                    'name':_name,
                    'field':field.db.name,
                    'col':j
                })
            j+=1
            _name = sheet.cell(1,j).value
        records = []
        i = 2
        row_str = sheet.cell(i,1).value
        # for row in range(2,sheet.max_row+1):
        
        while row_str != "" and row_str != None:
            rid = get_uuid()
            _data = {
                'rid':rid,
                'uid':user.rid,
                # 'pid':'',
                # 'mid':''
            }
            # m_key = sheet.cell(i,1).value
            # p_key = sheet.cell(i,2).value
            # s_key = sheet.cell(i,3).value
            for f in fields:
                # logger.error(f)
                v = sheet.cell(i,f['col']).value
                _data[f['field']] = v
            
            # logger.error(_data)
            records.append(_data)
            i += 1
            row_str = sheet.cell(i,1).value
            pass
        data[name]= {
                's_table':s_table,
                'm_table':m_table,
                'c_table':c_table,
                'name':sheet.title,
                'fields':fields,
                'keys':keys,
                'records':records,
                's_where_field':s_where_field,
                'c_where_field':c_where_field,
                'm_where_field':m_where_field
                }
        # ms[name] = m
        
    count = 0
    # print(data)
    for v in data.values():
        count =count + len(v['records'])
    # logger.error(data)
    process = 0
    s = Session()
    count = count +1 #算上commit的
    try:
        i=0
        j=0
        row_list = []
        for v in data.values():
            model = get_model_by_table_name(v['s_table'])
            m_where_field = v['m_where_field']
            s_where_field = v['s_where_field']
            c_where_field = v['c_where_field']
            m_table = v['m_table']
            s_table = v['s_table']
            c_table = v['c_table']
            # check_list = []
            repeat_list = []
            for d in v['records']:
                logger.error(d) 
                m_sql = 'select rid from '+str(v['m_table']) +' where '+ str(m_where_field) + '="'+ str(d[m_where_field]) +'" limit 1'
                logger.error(d) 
                m_return = run_sql(m_sql)
                if len(m_return)==0:
                    errors.append(str(d[c_where_field]))
                    j = j + 1
                    continue
                mid = m_return[0]['rid']
                c_sql = 'select rid from '+str(v['c_table']) +' where pid="'+str(mid)+'" and '+ str(c_where_field) + '="'+ str(d[c_where_field]) +'" limit 1'
                c_return = run_sql(c_sql)
                if len(c_return)==0:
                    errors.append(str(d[c_where_field]))
                    j = j + 1
                    continue
                pid=c_return[0]['rid']
                # if (check_child=='true' or check_child==True ):
                #     if d[c_where_field] in check_list:
                #         j = j + 1
                #         repeat_list.append(str(d[c_where_field]))
                #         continue
                #     else:
                #         check_list.append(str(d[c_where_field]))

                #     l_sql = 'select rid from '+str(v['c_table']) +' where '+ str(c_where_field) + '="'+ str(d[c_where_field]) +'" limit 1'
                #     sql_return = run_sql(l_sql)
                #     if len(sql_return)>0:
                #         repeat_list.append(str(d[c_where_field]))
                #         j = j + 1
                #         continue

                o = model()
                for key  in d.keys():
                    setattr(o,key,d[key])
                setattr(o,'mid',mid)
                setattr(o,'pid',pid)
                s.add(o)
                i = i + 1
                process = process+1
                if process %5==0:
                    await messages.pub(_make_msg('正在导入',code=1,position=process,max=count),topic=_IMPORT_TOPIC)
            # logger.error(repeat_list)
            row_list.append(v['name']+':'+','.join(repeat_list))

        # await messages.pub(_make_msg('正在导入',code=1,position=process,max=count),topic=_IMPORT_TOPIC)
        s.commit()
        sleep(0.5)
        if (j>0):
            msg_str = '成功导入'+str(i)+'条记录,重复'+str(j)+'条记录,'+';'.join(row_list)
        else:
            msg_str = '成功导入'+str(i)+'条记录'
        await messages.pub(_make_msg(msg_str,code=2,position=count,max=count),topic=_IMPORT_TOPIC)
    except Exception as e:
        logger.error(trace_error())
        await messages.pub(_make_msg(trace_error(),code=-1),topic=_IMPORT_TOPIC)
        s.rollback()
    finally:
        s.close()
    pass


@any_route('/api/whale/data/import/sub/child', methods=['POST', 'GET'])
@require_token
async def api_whale_data_import_sub_child(request):
    form = await request.form()
    # print(form)
    user = request.current_user
    file  = form.get('file.raw',None)
    file_name = form.get('file.name','')
    file_size = form.get('file.size',0)
    # option = form.get('option', None)

    if is_none(file):
        return json_result(ERR_PARAM_NOT_ENOUGH)
    fn = get_tmp_abs_file_name(get_suffix(file_name))
    logger.info(fn)
    data = await file.read()
    write_file(fn,data,'wb',encoding=None)
    task =BackgroundTask(_import_sub_child,fn,user,form)
    return json_result(task = task)
    pass


async def _import_update_data(fn):
    logger.error('---------row_str----------')
    if not os.path.exists(fn):
        messages.pub(_make_msg('无法找到文件',-1))
        return
    wb = load_workbook(fn)
    try:
        s = Session()
        logger.error('---------row_str----------')
        for sheet in wb.worksheets:
            i = 2
            j = 0
            row_str = sheet.cell(i,1).value
            logger.error(row_str)
            while row_str != "" and row_str != None:
                logger.error(row_str)
                j += 1
                BStock = sheet.cell(i,2).value
                logger.error(BStock)
                update_json = {'BStock': BStock}
                s.query(ProductSKUAttrList.rid).filter(ProductSKUAttrList.skuId==str(row_str)).update(update_json)
                i += 1
                row_str = sheet.cell(i,1).value
        
            s.commit()
            sleep(0.5)
            msg_str = '成功更新'+str(j)+'条记录'
            await messages.pub(_make_msg(msg_str,code=2),topic=_IMPORT_TOPIC)
    except Exception as e:
        logger.error(trace_error())
        await messages.pub(_make_msg(trace_error(),code=-1),topic=_IMPORT_TOPIC)
        s.rollback()
    finally:
        s.close()
        wb.close()
    pass


@any_route('/api/whale/data/import/update/data', methods=['POST', 'GET'])
@require_token
async def api_whale_data_import_update_data(request):
    form = await request.form()
    # print(form)
    user = request.current_user
    file  = form.get('file.raw',None)
    file_name = form.get('file.name','')
    file_size = form.get('file.size',0)
    # option = form.get('option', None)

    if is_none(file):
        return json_result(ERR_PARAM_NOT_ENOUGH)
    fn = get_tmp_abs_file_name(get_suffix(file_name))
    logger.info(fn)
    data = await file.read()
    write_file(fn,data,'wb',encoding=None)
    task =BackgroundTask(_import_update_data,fn)
    logger.error('---------row_str----------')
    return json_result(task = task)
    pass


async def _import_photo_from_excel(fn,user,group,kfield,pfield,request,option='update',heard=1):
    try:
        # wb = load_workbook(fn,data_only=True)
        wb = load_workbook(fn)
        ws = wb.active
        await messages.message(user.username,{'drts':'正在准备导入Excel文件...','total':99,'progress':0,'kind':'0'},MSG_KIND_NORMAL,request) 
        path = config.get_today_upload_path()
        data = []
        col = 1
        fields = []
        d = group.split('.')
        if len(d)!=2:
            await messages.message(user.username,{'total':99,'progress':99,'kind':'3','err':'参数错误！'},MSG_KIND_NORMAL,request)
            return
        m = get_module(d[0])
        val = ws.cell(heard,col).value
        while val != None and val != '':
            val = val.strip().replace('*','')
            if m.field_by_full_name(f'{group}.{val}')!=None:
                f = m.field_by_full_name(f'{group}.{val}')
                field = {'name':f.db.name,'val':val,'kind':f.db.kind,'col':col}
                fields.append(field)
            col += 1
            val = ws.cell(heard,col).value

        photo_list=[]
        param = {}
        for photo in ws._images:
            await messages.message(user.username,{'drts':'开始处理图片信息...','total':99,'progress':0,'kind':'1'},MSG_KIND_NORMAL,request)
            await asyncio.sleep(0.002)
            image_anchor = photo.anchor
            image = image_anchor._from
            col, colOff, row, rowOff=image.col,image.colOff,image.row,image.rowOff
            # 将图像数据保存到文件
            img = Image.open(photo.ref).convert("RGB")
            image_name = str(user.sid)+str(get_uuid())
            file_name = os.path.join(path,str(image_name)+'.png')
            if os.path.exists(file_name):  # 检查文件是否存在
                os.remove(file_name)  # 删除文件
            img.save(file_name)
            # 获取文件大小
            size = os.path.getsize(file_name)
            val = ws.cell(row+1,1).value
            file_name = os.path.join(path,str(image_name)+'.png')
            if os.path.exists(file_name)==True:
                photo_list = []
                sbs_path = path[-10:]
                photo_list.append({"src":str(sbs_path)+"/"+str(image_name)+".png","name":str(image_name)+".png"})
                p = str(photo_list)
                param[val]= p.replace("'",'"')

        heard +=1
        val = ws.cell(heard,1).value
        record = {}
        logger.error(param)
        while val != None and val != '':
            record[val]= {}
            r = {}
            logger.error('-------row_str-------')
            logger.error(fields)
            logger.error(val)
            for f in fields:
                col = f.get('col')
                kind = f.get('kind')
                v = ws.cell(heard,col).value
                if v==None or v=='':
                    if kind == 1 or kind == 2:
                        v = 0
                    elif kind == 3:
                        v = None
                    else:
                        v = ''
                else:
                    if kind != 1 and kind !=2 and kind !=3:
                        v = str(v).strip()
                if f.get('name')==pfield:
                    if val in param:
                        r[f['name']] = param[val]
                    else:
                        r[f['name']] = None
                else:
                    r[f['name']] = v

            record[val] = r
            data.append(r)
            heard +=1
            val = ws.cell(heard,1).value
            await messages.message(user.username,{'drts':'正在处理第'+str(heard-1)+'行数据...','total':99,'progress': math.floor(heard / ws.max_row * 99),'kind':'1'},MSG_KIND_NORMAL,request) 
            await asyncio.sleep(0.002)

        if option=='update':
            d = record
        else:
            d = data

        logger.error(d)
        await messages.message(user.username,{'total':99,'progress':99,'kind':'2','kfield':kfield,'pfield':pfield,'data':d,'option':option},MSG_KIND_NORMAL,request) 
    except:
        await messages.message(user.username,{'total':99,'progress':99,'kind':'3','err':str(trace_error())},MSG_KIND_NORMAL,request)


@any_route('/api/saier/import/photo/from/excel', methods=['POST', 'GET'])
@require_token
async def api_import_photo_from_excel(request):
    form = await request.form()
    option = form.get('option','')
    heard = form.get('heard',1)
    group = form.get('group','')
    kfield = form.get('kfield','')
    pfield = form.get('pfield','')
    path = config.get_today_upload_path()
    if not os.path.exists(path):
        make_dirs(path)

    user = request.current_user
    file  = form.get('file.raw',None)
    if is_none(file):
        return json_result(ERR_PARAM_NOT_ENOUGH)

    filename=str(user.sid)+str(get_uuid())+'.xlsx'
    fn = os.path.join(path, filename)
    d = await file.read()
    write_file(fn,d,'wb',encoding=None)

    async_thread_run(_import_photo_from_excel,fn, user, group, kfield, pfield, request, option, heard)
    return json_result()



async def order_apply_from_excel(fn,user,wfgs_n,request):
    s = Session()
    try:
        if wfgs_n == '' or wfgs_n == None:
            wfgs_n = '宁波优景进出口有限公司'

        wb = load_workbook(fn)
        ws = wb.active
        await messages.message(user.username,{'drts':'正在准备导入Excel文件...','total':99,'progress':0,'kind':'0'},MSG_KIND_NORMAL,request) 
        heard = 2
        val = ws.cell(heard,1).value
        errs = []
        csd = ''
        d = run_sql(f"select bz from cyzglsheet where (zm='合同申请公司') and (xm='{wfgs_n}') limit 1")
        if len(d)>0:
            csd = d[0].get('bz','')
        # logger.error(f"csd 1111 s: {csd}")
        while val != None and val != '':
            r = {}
            r['wfgs'] = wfgs_n
            r['khmc'] = ws.cell(heard,1).value
            r['mdck'] = ws.cell(heard,2).value
            r['sfxkr'] = ws.cell(heard,3).value
            r['krly'] = ws.cell(heard,4).value
            r['bz'] = ws.cell(heard,5).value
            r['csd'] = csd
            if r.get('sfxkr') != '是':
                r['sfxkr'] = '否'
            logger.error(r)
            # logger.error(r.get('sfxkr'))
            # logger.error(r.get('krly'))
            if (r.get('krly','') == '' or r.get('krly')==None) and r.get('sfxkr') == '是':
                errs.append(f"请注意客人: " + r.get('khmc','') + "是新客人但无来源")  # 如果不是客户则不生成合同申请单
                heard +=1
                val = ws.cell(heard,1).value
                continue
            
            if ws.cell(heard,6).value==None or ws.cell(heard,6).value=='':
                r['ts'] = 1
            else:
                r['ts'] = int(ws.cell(heard,6).value)
            # for i in range(1,int(r['ts'])):
            res = await order_apply_one_new(r,user,s)
            if res.get('code')!=1:
                s.rollback()
                await messages.message(user.username,{'total':99,'progress':99,'kind':'3','err':str(trace_error())},MSG_KIND_NORMAL,request)
                return
                
            heard +=1
            val = ws.cell(heard,1).value
            await messages.message(user.username,{'drts':'正在处理第'+str(heard-1)+'行数据...','total':99,'progress': math.floor(heard / ws.max_row * 99),'kind':'1'},MSG_KIND_NORMAL,request) 
            await asyncio.sleep(0.002)

        s.commit()
        await messages.message(user.username,{'total':99,'progress':99,'kind':'2','err':'\n'.join(errs)},MSG_KIND_NORMAL,request) 
    except:
        s.rollback()
        logger.error(trace_error())
        await messages.message(user.username,{'total':99,'progress':99,'kind':'3','err':str(trace_error())},MSG_KIND_NORMAL,request)
    finally:
        s.close()

@any_route('/api/saier/order_apply/from/excel', methods=['POST', 'GET'])
@require_token
async def api_order_apply_from_excel(request):
    form = await request.form()
    path = config.get_today_upload_path()
    if not os.path.exists(path):
        make_dirs(path)
    wfgs_n = form.get('wfgs','')
    user = request.current_user
    file  = form.get('file.raw',None)
    if is_none(file):
        return json_result(ERR_PARAM_NOT_ENOUGH)

    filename=str(user.sid)+str(get_uuid())+'.xlsx'
    fn = os.path.join(path, filename)
    d = await file.read()
    write_file(fn,d,'wb',encoding=None)

    async_thread_run(order_apply_from_excel,fn, user,wfgs_n,request)
    return json_result()


async def order_apply_download_excel(user):
    try:
        wb = Workbook() # 新建Excel文件
        ws = wb.active # wb.create_sheet() # 默认激活第一个sheet
        heard = 1
        ws.column_dimensions['A'].width = 28
        ws.cell(heard,1).value = '客户名称'
        ws.column_dimensions['B'].width = 14.3
        ws.cell(heard,2).value = '目的仓库'
        ws.column_dimensions['C'].width = 14.3
        ws.cell(heard,3).value = '是否新客人'
        ws.column_dimensions['D'].width = 14.3
        ws.cell(heard,4).value = '客人来源'
        ws.column_dimensions['E'].width = 28
        ws.cell(heard,5).value = '备注'
        ws.column_dimensions['F'].width = 14.3
        ws.cell(heard,6).value = '生成条数'
        path = config.get_today_upload_path()
        if not os.path.exists(path):
            make_dirs(path)
        filename=str(get_uuid())+'_.xlsx'
        fn = os.path.join(path, filename)
        wb.save(fn)
        # wb.close()
        sbs_path = path[-10:]

        return {'code':1,'msd':'成功','data':str(sbs_path)+'/'+filename}
    except:
        logger.error(trace_error())
        return {'code':-1,'msd':trace_error()}


@any_route('/api/saier/order_apply/download/excel', methods=['POST', 'GET'])
@require_token
async def api_order_apply_download_excel(request):
    j = await request.json()
    user = request.current_user
    try:
        res = await order_apply_download_excel(user)

        return json_result(res.get('code'),res.get('msd'),res.get('data'))
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1,trace_error())
    
# async def _excel_image_export2(fn,user):
#     s = Session()
#     try:
#         if not os.path.exists(fn):
#             messages.pub(_make_msg('无法找到文件',-1))
#             return

#         wb = load_workbook(fn)
#         ws = wb['Sheet1']
#         num = ws.max_row  # 总行数，用来遍历读取和图片命名
#         path = config.get_today_upload_path()
#         if not os.path.exists(path):
#             make_dirs(path)
#         num = len(ws._images)
#         i = 0
#         # 遍历sheet中的图片
#         for image1 in ws._images:
#             i += 1
#             image_anchor = image1.anchor
#             image= image_anchor._from

#             col, colOff, row, rowOff=image.col,image.colOff,image.row,image.rowOff
#             # print(col,colOff,row,rowOff)
#             #print(image_anchor)
#             # 将图像数据保存到文件
#             img = Image.open(image1.ref).convert("RGB")
#             # img = np.array(img)
#             image_name = ws['A'+str(row+1)].value
#             file_name = path + '/'+ str(image_name)+'.png'
#             img.save(file_name)

#         # image_loader = SheetImageLoader(ws)
#         # for i in range(2, num + 1):  # 从第2行开始
#         #     image_name = ws['A'+str(i)].value
#         #     file_name = path + '/'+ str(image_name)+'.png'
#         #     image = image_loader.get('B' + str(i))  # B列的图片
#         #     logger.error(image)
#         #     image.save(file_name)  # 行号-2命名

#             await messages.pub(_make_msg('正在导出图片',code=2,position=i,max=num),topic=_IMPORT_TOPIC)

#         wb.close()
#     except Exception as e:
#         logger.error(trace_error())
#         await messages.pub(_make_msg(trace_error(),code=-1),topic=_IMPORT_TOPIC)
#         s.rollback()
#     finally:
#         s.close()
#     pass


# from PIL import ImageGrab

# import win32com.client as win32


# async def _excel_image_export3(fn,user):
#     s = Session()
#     try:
#         if not os.path.exists(fn):
#             messages.pub(_make_msg('无法找到文件',-1))
#             return

#         # 能把所有图片按原图象素导出来

#         excel = win32.gencache.EnsureDispatch('Excel.Application')

#         workbook = excel.Workbooks.Open(r'C:\\Users\\admin\\Desktop\\picexcel.xlsx')

#         for sheet in workbook.Worksheets:

#             for i, shape in enumerate(sheet.Shapes):

#                 if shape.Name.startswith('Picture'):

#                     shape.Copy()

#                     image = ImageGrab.grabclipboard()

#                     image.save('{}.jpg'.format(i+1), 'jpeg')

#         excel.Quit()

#         await messages.pub(_make_msg(msg_str,code=2,position=count,max=count),topic=_IMPORT_TOPIC)
        
    # except Exception as e:
    #     logger.error(trace_error())
    #     await messages.pub(_make_msg(trace_error(),code=-1),topic=_IMPORT_TOPIC)
    #     s.rollback()
    # finally:
    #     s.close()
    # pass