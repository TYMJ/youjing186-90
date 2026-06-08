
import os
import uuid
import math
import zipfile
from any import *
from .model import *
from decimal import Decimal, ROUND_HALF_UP
from sqlalchemy.sql import text
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.utils import coordinate_to_tuple
from .__default__ import get_user_path


#出运明细  优景list 分单价
@any_route('/api/saier/shipment/price_list/export', methods=['POST'])
@require_token
async def api_saier_export_uv_price_list(request):
    """
    优景list分单价 导出
    转译自 Delphi: 合并同产品多单价、处理化学品名富文本下标渲染
    """
    j = await request.json()
    s = Session()
    try:
        record_id = j.get('record_id')
        currency_type = j.get('currency_type', '1') # 1: 美金, 2: 人民币。前端传入

        user = request.current_user
        username = user.username

        # 1. 权限校验
        auth_sql = text("""
            SELECT 1 FROM sys_user
            WHERE username = :name AND (path LIKE '%外销%' OR position LIKE '%单证%' OR position LIKE '%外销%')
        """)
        if not s.execute(auth_sql, {"name": username}).fetchone():
            return json_result(-1, "权限不足")

        # 2. 获取主表信息
        master = s.execute(text("SELECT * FROM cymx WHERE rid=:id"), {"id": record_id}).fetchone()
        if not master:
            return json_result(-1, "未找到该单据数据")
            
        if master.RMBkh == '是' and currency_type == '1':
             # 这里只是为了匹配原代码的逻辑，前端如果提示了可以忽略
             logger.warning(f"单据 {record_id} 是RMB客户，但选择了美金导出")

        curr_symbol = '￥' if currency_type == '2' else '$'

        # 3. 预处理数据：合并多级价格 (在内存中完成，不再 update 数据库)
        # 获取第 4 层明细 (商品层级)
        sheet4_rows = s.execute(text("SELECT * FROM cymxsheet4 WHERE pid=:id"), {"id": record_id}).fetchall()
        
        # 获取底层明细 (包含价格的层级)
        sheet_rows = s.execute(text("SELECT * FROM cymxsheet WHERE pid=:id"), {"id": record_id}).fetchall()
        
        # 预加载 tspmb (特殊品名及下标)
        cpbh_list = list(set([r.cpbh for r in sheet4_rows if r.cpbh]))
        tspm_map = {}
        # if cpbh_list:
        #     tspm_rows = s.execute(text("SELECT cpbh, djpmy, djyxb FROM tspmb WHERE cpbh IN :ids AND djpmy<>''"), {"ids": tuple(cpbh_list)}).fetchall()
        #     tspm_map = {r.cpbh: r for r in tspm_rows}

        # 构建价格映射字典: {cpbh: "￥1.50,￥1.60"}

        # 【必须补上这句】准备 UPDATE 语句
        update_hbdj_sql = text("""
            UPDATE cymxsheet4 
            SET hbdj = :hbdj 
            WHERE rid = :rid
        """)
        price_map = {}
        for row4 in sheet4_rows:
            cpbh = row4.cpbh
            row4_id = row4.rid
            prices_set = set() # 用 set 去重
            for row in sheet_rows:
                if row.zycpbh == cpbh:
                    if currency_type == '2': # RMB
                        if row.mjdj1: prices_set.add(str(row.mjdj1))
                    else: # USD
                        if row.wxjg and row.zmyhl:
                            # 计算外销价格*汇率，保留3位小数
                            val = Decimal(str(row.wxjg)) * Decimal(str(row.zmyhl))
                            val_rounded = val.quantize(Decimal('0.000'), rounding=ROUND_HALF_UP)
                            prices_set.add(str(val_rounded))
                            
            # 拼接带符号的价格字符串
            formatted_prices = [f"{curr_symbol}{p}" for p in prices_set]
            hbdj_str = ",".join(formatted_prices) if formatted_prices else ""
            # 存入字典给生成 Excel 填单元格用
            price_map[cpbh] = hbdj_str
            
            # 执行更新，将计算好的字符串回写到 cymxsheet4 对应的行
            s.execute(update_hbdj_sql, {
                "hbdj": hbdj_str, 
                "rid": row4_id
            })
        s.commit()
        # 4. 生成 Excel (不使用本地模板)
        r_path = os.path.join(config.data_upload_path, 'template')
        fn = os.path.join(r_path, 'uvlist.xlsx') #
        if not os.path.exists(fn):
            return json_result(-1, f"未找到报表模板: uvlist.xlsx")
        wb = load_workbook(fn)
        ws = wb.active
        ws.title = "优景list分单价"

        # 写入头部信息 (原代码写入 A6 和 D6，我们向上收紧一点)
        fprqy = str(master.fprqy).upper() if master.fprqy else ""
        chrqy = str(master.chrqy).upper() if master.chrqy else ""
        htjx = master.htjx or ""
        
       # 填在第 6 行和第 7 行的空白处
        ws['A6'] = f"Date: {fprqy}"
        ws['D6'] = f"VALID TILL: {chrqy}"
        ws['B7'] = f"Contract No: {htjx}"
        
        
        # 5. 循环写入明细数据
        # start_row = 5
        # font_subscript = InlineFont(vertAlign='subscript') # 定义下标样式
        # font_normal = InlineFont() # 定义普通样式

        # for i, row4 in enumerate(sheet4_rows):
        #     r_idx = start_row + i
        #     cpbh = row4.cpbh
            
        #     ws[f'A{r_idx}'] = i + 1
        #     ws[f'B{r_idx}'] = str(cpbh).upper() if cpbh else ""
            
            # --- 核心：处理品名富文本 (富文本下标渲染) ---
            # tspm_info = tspm_map.get(cpbh)
            # djpmy = str(row4.djpmy).upper() if row4.djpmy else ""
            # djpmw = str(row4.djpmw).upper() if row4.djpmw else ""
            # full_desc = f"{djpmy}/{djpmw}"
            
            # if tspm_info and tspm_info.djyxb:
            #     # 解析下标位置字符串，例如 "2, 5, 8"
            #     idx_str = str(tspm_info.djyxb).replace('，', ',')
            #     sub_indexes = []
            #     for x in idx_str.split(','):
            #         x = x.strip()
            #         if x.isdigit():
            #             # 注意：原Delphi代码索引是1-based，且是基于第一个字符串。
            #             # 这里我们转换为 0-based 索引。
            #             sub_indexes.append(int(x) - 1)
                
            #     # 构建 RichText 对象
            #     rich_text_elements = []
            #     for char_idx, char in enumerate(full_desc):
            #         if char_idx in sub_indexes:
            #             rich_text_elements.append(TextBlock(font_subscript, char))
            #         else:
            #             rich_text_elements.append(TextBlock(font_normal, char))
                        
            #     ws[f'C{r_idx}'] = CellRichText(*rich_text_elements)
            # else:
            #     ws[f'C{r_idx}'] = full_desc
            # ----------------------------------------------
            # --- 核心：直接获取明细表品名并拼接 ---
            # djpmy = str(row4.djpmy).upper() if row4.djpmy else ""
            # djpmw = str(row4.djpmw).upper() if row4.djpmw else ""
            # full_desc = f"{djpmy}/{djpmw}"  # 为空 因为sheet4  数据为空字符
            
            # # 写入 C 列
            # ws[f'C{r_idx}'] = full_desc
            # ws[f'D{r_idx}'] = price_map.get(cpbh, "")
            # ws[f'E{r_idx}'] = str(row4.jldw).upper() if row4.jldw else ""
            
            # 样式
            # for col in range(1, 6):
            #     ws.cell(row=r_idx, column=col).alignment = Alignment(vertical='center')
            # ws[f'C{r_idx}'].alignment = Alignment(wrapText=True, vertical='center')
        start_row = 9 

        for i, row4 in enumerate(sheet4_rows):
            r_idx = start_row + i
            cpbh = row4.cpbh
            
            ws[f'A{r_idx}'] = i + 1                                     # 序号 (#)
            ws[f'B{r_idx}'] = str(cpbh).upper() if cpbh else ""         # 产品编号 (ARTICLE)
            
            djpmy = str(row4.djpmy).upper() if row4.djpmy else ""
            djpmw = str(row4.djpmw).upper() if row4.djpmw else ""
            full_desc = f"{djpmy}/{djpmw}" if djpmy or djpmw else ""    
            
            ws[f'C{r_idx}'] = full_desc                                 # 品名 (DESCRIPTION OF GOODS)
            ws[f'D{r_idx}'] = price_map.get(cpbh, "")                   # 合并单价 (UNIT PRICE)
            ws[f'E{r_idx}'] = str(row4.jldw).upper() if row4.jldw else "" # 单位 (UNIT)
            
            # 设置垂直居中，C列(品名)允许自动换行防遮挡
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{r_idx}'].alignment = Alignment(vertical='center')
            ws[f'C{r_idx}'].alignment = Alignment(wrapText=True, vertical='center')
        # 6. 保存并返回文件
        s_path = config.tmp_path
        file_name = f"{master.fphm or 'Unknown'}_Price_List.xlsx"
        wb.save(os.path.join(s_path, file_name))
        return json_result(1, '导出成功', file_name)

    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f"生成分单价报价单失败: {str(e)}")
    finally:
        s.close()



# 俄文版 报关单new
@any_route('/api/saier/shipment/russian_declaration/export', methods=['POST'])
@require_token
async def api_export_russian_customs_declaration(request):

    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        record_id = j.get('rid')
        if not record_id:
            return json_result(-1, "缺少单据ID")

        # 1. 权限校验
        auth_sql = text("""
            SELECT 1 FROM sys_user
            WHERE username = :name 
            AND (path LIKE '%外销%' OR position LIKE '%单证%')
        """)
        if not s.execute(auth_sql, {"name": user.username}).fetchone():
            return json_result(-1, "权限不足!")

        # 2. 查询出运明细主表 (cymx)
        cymx = s.execute(text("SELECT * FROM cymx WHERE rid = :id"), {"id": record_id}).fetchone()
        if not cymx:
            return json_result(-1, "未找到该出运单据")

        # 3. 获取相关的字典翻译数据 (价格条款, 运输方式, 口岸)
        jgtk_desc = s.execute(text("SELECT sm FROM jgtk WHERE jgtk = :val"), {"val": cymx.jgtk}).scalar() or ""
        ysfs_desc = s.execute(text("SELECT ww FROM ysfs WHERE ysfs = :val"), {"val": cymx.ysfs}).scalar() or ""
        cyka_desc = s.execute(text("SELECT wwmc FROM cyka WHERE ywmc = :val OR cyka = :val"), {"val": cymx.cyka}).scalar() or cymx.cyka
        mdka_desc = s.execute(text("SELECT mdkaw FROM mdka WHERE mdka = :val"), {"val": cymx.mdka}).scalar() or cymx.mdka

        # 获取模板配置
        report_config = s.execute(
            text("SELECT bz, bz1, bz2, bz3 FROM cyzglsheet WHERE xm = :xm AND zm = '俄文报关单客人区别'"),
            {"xm": cymx.khmc}
        ).fetchone()
        
        template_name = report_config.bz if report_config and report_config.bz else 'bge'
        bz1 = report_config.bz1 if report_config else "Входящий номер："
        bz2 = report_config.bz2 if report_config else "Код таможни："
        bz3 = report_config.bz3 if report_config else "Номер страницы/количество страниц:"

        # 4. 获取需要导出的明细数据 (cymxsheet4)
        sheet4_rows = s.execute(
            text("SELECT * FROM cymxsheet4 WHERE pid = :id AND (xs > 0 OR sfpx = '是')"),
            {"id": record_id}
        ).fetchall()
        
        if not sheet4_rows:
            return json_result(-1, "该单据没有有效的产品明细")

        # --- 预加载字典数据，防止 N+1 查询 ---
        cpbh_list = [r.cpbh for r in sheet4_rows]
        hgbm_list = [r.hgbm2 for r in sheet4_rows]
        hyd_list = [r.hyd for r in sheet4_rows]
        jldw_list = [r.jldw for r in sheet4_rows]
        #  原先的wypm  表  不存在了  
        # wypm_map = {}
        # if cpbh_list:
        #     wypm_rows = s.execute(text("SELECT cpbh, krhh, djpmw FROM wypm WHERE cpbh IN :ids OR krhh IN :ids"), {"ids": tuple(cpbh_list)}).fetchall()
        #     for r in wypm_rows:
        #         wypm_map[r.cpbh] = r.djpmw
        #         if r.krhh: wypm_map[r.krhh] = r.djpmw

        hyd_map = {}
        if hyd_list:
            hyd_rows = s.execute(text("SELECT hyd, hydwy, hydbm FROM hyd WHERE hyd IN :ids"), {"ids": tuple(hyd_list)}).fetchall()
            hyd_map = {r.hyd: r for r in hyd_rows}

        hgbm_map = {}
        if hgbm_list:
            hgbm_rows = s.execute(text("SELECT hgbm, hwmc, ecznr FROM hgbmbsheet WHERE hgbm IN :ids"), {"ids": tuple(hgbm_list)}).fetchall()
            hgbm_map = {f"{r.hgbm}_{r.hwmc}": r.ecznr for r in hgbm_rows}

        jldw_map = {}
        if jldw_list:
            jldw_rows = s.execute(text("SELECT dwdm, ywfs, wwdw FROM jldw WHERE dwdm IN :ids OR ywfs IN :ids"), {"ids": tuple(jldw_list)}).fetchall()
            for r in jldw_rows:
                jldw_map[r.dwdm] = r.wwdw
                if r.ywfs: jldw_map[r.ywfs] = r.wwdw

        total_mzhj = sum(round(float(r.zmz or 0), 1) for r in sheet4_rows)
        total_jzhj = sum(round(float(r.zjz or 0), 1) for r in sheet4_rows)

        # ================== [修改说明 1: 模板路径加载逻辑] ==================

        r_path = os.path.join(config.data_upload_path, 'template')
        fn = os.path.join(r_path, str(template_name) + '.xlsx')
        
        if not os.path.exists(fn):
            return json_result(-1, f"未找到报表模板")
            
        wb = load_workbook(fn)
        base_sheet = wb.active
        for sheet_name in wb.sheetnames:
            if sheet_name != base_sheet.title:
                del wb[sheet_name]
        
        # 顺手把第一页重命名一下，显得更专业
        base_sheet.title = "第1页"
        # ====================================================================
        
        # 计算需要多少个 Sheet (每页 6 个商品)
        total_items = len(sheet4_rows)
        total_sheets = math.ceil(total_items / 6)

        # 复制出所需的 Sheet 数量
        sheets = [base_sheet]
        for i in range(1, total_sheets):
            new_sheet = wb.copy_worksheet(base_sheet)
            new_sheet.title = f"第{i+1}页"
            sheets.append(new_sheet)

        # 6. 开始填充数据
        is_rmb_customer = (cymx.RMBkh == '是')
        fprq_str = str(cymx.fprq)[:10].replace("-", "") if cymx.fprq else "" 
        
        for idx, row4 in enumerate(sheet4_rows):
            sheet_idx = idx // 6
            item_in_sheet = idx % 6
            ws = sheets[sheet_idx]

            if item_in_sheet == 0:
                if cymx.bgdh:
                    ws['A2'] = f"{bz1}({cymx.bgdh})                                {bz2}{cymx.bgdh}            \n(Таможня Хайш)"
                else:
                    ws['A2'] = f"{bz1}                                            \n                            {bz2}                                                     (Таможня Хайш)"
                
                ws['O2'] = f"{bz3}{sheet_idx + 1}/{total_sheets}"
                if mdka_desc: ws['L10'] = mdka_desc
                
                ws['H4'] = cymx.chyrq or ""
                ws['L4'] = fprq_str
                ws['H6'] = f"{str(cymx.cmin or '').upper()}/{str(cymx.hangci or '').upper()}"
                ws['L6'] = cymx.tdh or ""
                ws['A10'] = cymx.htjx or ""
                ws['J12'] = jgtk_desc
                ws['E4'] = cyka_desc
                ws['P10'] = cyka_desc
                ws['E12'] = cymx.xshj or ""
                ws['F12'] = total_mzhj
                ws['H12'] = total_jzhj
                ws['A17'] = f"Номер контейнера：{cymx.xh or ''}"

            r1 = 19 + (item_in_sheet * 3)  
            r2 = r1 + 1                    
            r3 = r1 + 2                    

            # pm = wypm_map.get(row4.cpbh) or wypm_map.get(row4.krhh) or str(row4.djpmw or "").upper()
            pm = str(row4.djpmw or row4.zwbgpm or "").upper()
            hyd_info = hyd_map.get(row4.hyd)
            hyd_name = hyd_info.hydwy if hyd_info else 'Другое,Чжэцзян'
            hyd_code = f"({hyd_info.hydbm})" if (hyd_info and hyd_info.hydbm) else '(33909)'
            hg_key = f"{row4.hgbm2}_{row4.zwbgpm}"
            ecznr = hgbm_map.get(hg_key, "")
            unit = jldw_map.get(row4.jldw, "")

            # === 写入第 1 行 ===
            ws[f'A{r1}'] = idx + 1
            ws[f'B{r1}'] = str(row4.krcode or "")
            
            # ================= [修改说明 2: 修复原代码中的细节] =================
            # 还原 Delphi 的 NumberFormatlocal := '@'
            # 强制条码列为文本格式，防止 0 开头的条码被 Excel 变成数字
            ws[f'B{r1}'].number_format = '@'
            # ====================================================================

            ws[f'C{r1}'] = pm
            ws[f'G{r1}'] = f"{row4.zjz or ''}кг"
            ws[f'I{r1}'] = row4.khrmb if is_rmb_customer else row4.wxdj
            ws[f'K{r1}'] = 'Китай'
            ws[f'N{r1}'] = 'Российская федерация'
            ws[f'Q{r1}'] = hyd_name
            ws[f'S{r1}'] = 'Налог в соответствии с правилами'

            # === 写入第 2 行 ===
            if ecznr: ws[f'C{r2}'] = ecznr
            ws[f'G{r2}'] = f"{row4.chsl or ''}{unit}"
            ws[f'I{r2}'] = row4.kfrmbz if is_rmb_customer else row4.wxzj
            ws[f'K{r2}'] = '(КНР)'
            ws[f'N{r2}'] = '(РФ)'
            ws[f'Q{r2}'] = hyd_code
            ws[f'S{r2}'] = '(1)'

            # === 写入第 3 行 ===
            ws[f'I{r3}'] = 'Юань' if is_rmb_customer else 'Доллары'

        # 7. 保存文件并返回下载地址
        save_dir = config.tmp_path
        if not os.path.exists(save_dir):
            os.makedirs(save_dir, exist_ok=True)
            
        file_name = f"{cymx.fphm or 'Unknown'}_俄文报关单.xlsx"
        save_path = os.path.join(save_dir, file_name)
        
        wb.save(save_path)
        s.commit()
        wb.close()

        return json_result(1, "生成成功", file_name)

    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f"生成俄文报关单失败: {str(e)}")
    finally:
        s.close()




# 俄文版INV PL

RUSSIAN_MONTHS = {
    '01': 'Январь', '02': 'Февраль', '03': 'Март', '04': 'Апрель',
    '05': 'Май', '06': 'Июнь', '07': 'Июль', '08': 'Август',
    '09': 'Сентябрь', '10': 'Октябрь', '11': 'Ноябрь', '12': 'Декабрь'
}
@any_route('/api/saier/shipment/inv_pl/export', methods=['POST'])
@require_token
async def api_export_russian_inv_pl(request):
    """
    导出俄文版商业发票(INVOICE) 与 装箱单(PACKING LIST)
    重构自原 Delphi 的 "俄文版INV PL.txt"
    """
    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        record_id = j.get('rid')
        hbdm_input = j.get('currency_type', '1') # 1:人民币(R), 2:美金(M)
        
        if not record_id: return json_result(-1, "缺少单据ID")

        # 1. 权限校验
        auth_sql = text("""
            SELECT 1 FROM sys_user
            WHERE username = :name AND (path LIKE '%外销%' OR position LIKE '%单证%')
        """)
        if not s.execute(auth_sql, {"name": '李辉'}).fetchone():
            return json_result(-1, "权限不足")

        # 2. 查询出运明细主表 (cymx)
        cymx = s.execute(text("SELECT * FROM cymx WHERE rid = :id"), {"id": record_id}).fetchone()
        if not cymx: return json_result(-1, "未找到该出运单据")

        # 3. 客户专属排版路由判断 (BEST PRICE 逻辑)
        is_best_price = 'BEST PRICE' in str(cymx.khmc or '').upper()
        hbdm = 'R' if hbdm_input == '1' else 'M'
        
        # 动态列映射机制
        col = {
            'AA': 'B' if is_best_price else 'A',
            'B': 'C' if is_best_price else 'B',
            'C': 'D' if is_best_price else 'C',
            'DD': 'E' if is_best_price else 'D',
            'E': 'F' if is_best_price else 'E',
            'F': 'G' if is_best_price else 'F',
            'G': 'H' if is_best_price else 'G',
            'H': 'I' if is_best_price else 'H',
            'II': 'J' if is_best_price else 'I',
            'JJ': 'K' if is_best_price else 'J',
            'KK': 'L' if is_best_price else 'K'
        }

        # 4. 日期转换与港口翻译
        fprq_str = str(cymx.fprq)[:10] if cymx.fprq else ""
        qsn, qsy1, qsr = "", "", ""
        qsy_rus = ""
        if len(fprq_str) == 10:
            qsn, qsy1, qsr = fprq_str[0:4], fprq_str[5:7], fprq_str[8:10]
            qsy_rus = RUSSIAN_MONTHS.get(qsy1, '')

        # 港口翻译 (简化版兜底)
        mdka_val = cymx.mdka or ""
        mdka_rus_map = {
            'NINGBO, CHINA': 'Нинбо, Китай',
            'SHANTOU, CHINA': 'Шаньтоу, Китай',
            'XIAMEN, CHINA': 'Сямэнь, Китай'
        }
        ew_port = mdka_rus_map.get(mdka_val.upper(), mdka_val)
        
        cyka_trans = s.execute(text("SELECT wwmc FROM cyka WHERE ywmc = :val OR cyka = :val"), {"val": mdka_val}).scalar()
        if cyka_trans: ew_port = cyka_trans
        g14_port = mdka_val
        mdka_trans = s.execute(text("SELECT mdkaw FROM mdka WHERE mdka = :val"), {"val": mdka_val}).scalar()
        if mdka_trans: g14_port = mdka_trans
        # 5. 获取明细数据并计算汇总 (cymxsheet4)
        sheet4_rows = s.execute(
            text("SELECT * FROM cymxsheet4 WHERE pid = :id AND (xs > 0 OR sfpx = '是')"),
            {"id": record_id}
        ).fetchall()
        
        if not sheet4_rows: return json_result(-1, "该单据没有有效的产品明细")

        total_mzhj = sum(round(float(r.zmz or 0), 1) for r in sheet4_rows)
        total_jzhj = sum(round(float(r.zjz or 0), 1) for r in sheet4_rows)
        total_tjhj = sum(round(float(r.ztj or 0), 2) for r in sheet4_rows)

        # 预加载计量单位字典
        jldw_list = [r.jldw for r in sheet4_rows]
        jldw_map = {}
        if jldw_list:
            jldw_rows = s.execute(text("SELECT dwdm, wwdw FROM jldw WHERE dwdm IN :ids"), {"ids": tuple(jldw_list)}).fetchall()
            jldw_map = {str(r.dwdm).upper(): r.wwdw for r in jldw_rows}

        # ==================== 核心开始：生成 INVOICE ====================
        inv_tpl_prefix = "x" if is_best_price else ""
        inv_tpl_currency = "rmbuvinvd" if hbdm == 'R' else "uvinvd"
        inv_tpl_name = f"{inv_tpl_prefix}{inv_tpl_currency}.xlsx"
        
        r_path = os.path.join(config.data_upload_path, 'template')
        inv_fn = os.path.join(r_path, inv_tpl_name)
        if not os.path.exists(inv_fn): return json_result(-1, f"找不到发票模板: {inv_tpl_name}")

        wb_inv = load_workbook(inv_fn)
        base_sheet_inv = wb_inv.active
        # 清理空白的sheet2
        for sheet_name in wb_inv.sheetnames:
            if sheet_name != base_sheet_inv.title:
                del wb_inv[sheet_name]

        base_sheet_inv.sheet_view.showGridLines = True
        total_items = len(sheet4_rows)
        inv_pages = math.ceil(total_items / 8) # 发票每页 8 个
        inv_sheets = [base_sheet_inv]
        for i in range(1, inv_pages):
            inv_sheets.append(wb_inv.copy_worksheet(base_sheet_inv))

        for idx, row in enumerate(sheet4_rows):
            page_idx = idx // 8
            item_idx = idx % 8
            ws = inv_sheets[page_idx]

            # 写入单页表头
            if item_idx == 0:
                ws[f'{col["JJ"]}8'] = cymx.htjx or ""
                ws[f'{col["JJ"]}9'] = f"{qsr} {qsy_rus} {qsn}"
                ws[f'{col["JJ"]}10'] = cymx.fporder or ""
                ws[f'{col["JJ"]}14'] = f"{cymx.fporder or ''}S"
                ws[f'{col["JJ"]}15'] = f"{qsr} {qsy_rus} {qsn}"
                ws[f'{col["JJ"]}11'] = str(page_idx + 1)
                ws[f'{col["II"]}20'] = ew_port

                # 底部银行与总计信息
                # ws[f'{col["DD"]}54'] = 'РЕКВИЗИТЫ БАНКА:'
                # ws[f'{col["DD"]}55'] = 'CHINA EVERBRIGHT BANK, NINGBO BR.'
                # ws[f'{col["DD"]}56'] = 'A/C: 7680-14-88-0000104-23'
                # ws[f'{col["G"]}57'] = 'ОБЩАЯ СУММА:'
                # ws[f'{col["II"]}57'] = float(cymx.mjzj or 0) if hbdm == 'R' else float(cymx.wxje or 0)
            if idx == total_items - 1:
                ws[f'{col["DD"]}54'] = 'РЕКВИЗИТЫ БАНКА:'
                ws[f'{col["DD"]}55'] = 'CHINA EVERBRIGHT BANK, NINGBO BR.'
                ws[f'{col["DD"]}56'] = 'A/C: 7680-14-88-0000104-23'
                ws[f'{col["G"]}57'] = 'ОБЩАЯ СУММА:'
                ws[f'{col["II"]}57'] = float(cymx.mjzj or 0) if hbdm == 'R' else float(cymx.wxje or 0)
            # 写入 4 行 1 个商品的明细
            # Delphi d 从 1 开始(第一次循环d=d+4变为1), row是 21+d -> 22起
            start_row = 22 + (item_idx * 4)
            
            # 品名使用明细表的数据兜底
            pm = str(row.djpmw or row.zwbgpm or "").upper()
            unit = jldw_map.get(str(row.jldw or '').upper(), str(row.jldw or '').upper())

            if is_best_price: ws[f'{col["AA"]}{start_row}'] = str(idx + 1)
            ws[f'{col["B"]}{start_row}'] = row.cpbh
            ws[f'{col["DD"]}{start_row}'] = pm
            ws[f'{col["DD"]}{start_row + 1}'] = f"КОД ТНВЭД:{row.krcode or ''}"
            ws[f'{col["G"]}{start_row}'] = float(row.chsl or 0)
            ws[f'{col["H"]}{start_row}'] = unit
            
            if hbdm == 'R':
                ws[f'{col["II"]}{start_row}'] = float(row.khrmb or 0)
                ws[f'{col["KK"]}{start_row}'] = float(row.kfrmbz or 0)
            else:
                ws[f'{col["II"]}{start_row}'] = float(row.wxdj or 0)
                ws[f'{col["KK"]}{start_row}'] = float(row.wxzj or 0)

        inv_save_name = f"{cymx.fphm or 'Unknown'}_INVOICE_перевод_ЮС.xlsx"
        inv_save_path = os.path.join(config.tmp_path, inv_save_name)
        wb_inv.save(inv_save_path)
        wb_inv.close()

        # ==================== 核心开始：生成 PACKING LIST ====================
        pl_tpl_name = f"{'x' if is_best_price else ''}uvpld.xlsx"
        pl_fn = os.path.join(r_path, pl_tpl_name)
        if not os.path.exists(pl_fn): return json_result(-1, f"找不到装箱单模板: {pl_tpl_name}")

        wb_pl = load_workbook(pl_fn)
        base_sheet_pl = wb_pl.active
        # 清理空白的sheet2
        base_sheet_pl.sheet_view.showGridLines = True
        for sheet_name in wb_pl.sheetnames:
            if sheet_name != base_sheet_pl.title:
                del wb_pl[sheet_name]
        pl_pages = math.ceil(total_items / 14) # 装箱单每页 14 个
        pl_sheets = [base_sheet_pl]
        for i in range(1, pl_pages):
            pl_sheets.append(wb_pl.copy_worksheet(base_sheet_pl))

        red_font = Font(color="FF0000") # 异常红色字体

        for idx, row in enumerate(sheet4_rows):
            page_idx = idx // 14
            item_idx = idx % 14
            ws = pl_sheets[page_idx]

            # 写入单页表头
            if item_idx == 0:
                ws[f'H7'] = cymx.htjx or ""
                ws[f'H8'] = f"{qsr} {qsy_rus} {qsn}"
                ws[f'H9'] = cymx.fporder or ""
                ws[f'H10'] = str(page_idx + 1)
                ws[f'B14'] = ew_port
                ws[f'G14'] = g14_port
                ws[f'{col["DD"]}18'] = f"{cymx.xh or ''}/{cymx.fh or ''}"

                # 底部合计信息
                # ws[f'{col["DD"]}54'] = f"ОБЩАЯ СУММА:        {cymx.xshj2 or ''} КОР     {total_mzhj} ВЕС (КГ)  {total_jzhj} ВЕС (КГ)  {total_tjhj} КУБ М    "
            if idx == total_items - 1:
                ws[f'{col["DD"]}54'] = f"ОБЩАЯ СУММА:        {cymx.xshj2 or ''} КОР     {total_mzhj} ВЕС (КГ)  {total_jzhj} ВЕС (КГ)  {total_tjhj} КУБ М    "
            # 写入 2 行 1 个商品的明细
            # Delphi d 从 1 开始(第一次循环d=d+2变为1), row是 24+d -> 25起
            start_row = 25 + (item_idx * 2)
            
            pm = str(row.djpmw or row.zwbgpm or "").upper()
            unit = jldw_map.get(str(row.jldw or '').upper(), str(row.jldw or '').upper())

            if is_best_price: ws[f'{col["AA"]}{start_row}'] = str(idx + 1)
            ws[f'{col["B"]}{start_row}'] = row.cpbh
            ws[f'{col["DD"]}{start_row}'] = pm
            
            # 异常校验逻辑 (装箱率计算不匹配标红)
            chsl = float(row.chsl or 0)
            wxrl = float(row.wxrl or 0)
            xs = float(row.xs or 0)
            is_warning = (wxrl > 0 and abs((chsl / wxrl) - xs) > 0.01)

            cell_e = ws.cell(row=start_row, column=ws[f'{col["E"]}{start_row}'].column)
            cell_f = ws.cell(row=start_row, column=ws[f'{col["F"]}{start_row}'].column)
            cell_g = ws.cell(row=start_row, column=ws[f'{col["G"]}{start_row}'].column)
            
            cell_e.value = xs
            cell_f.value = wxrl
            cell_g.value = chsl
            
            if is_warning:
                cell_e.font = red_font
                cell_f.font = red_font
                cell_g.font = red_font

            ws[f'{col["H"]}{start_row}'] = unit
            ws[f'{col["II"]}{start_row}'] = round(float(row.zmz or 0), 1)
            ws[f'{col["JJ"]}{start_row}'] = round(float(row.zjz or 0), 1)
            ws[f'{col["KK"]}{start_row}'] = round(float(row.ztj or 0), 2)

        pl_save_name = f"{cymx.fphm or 'Unknown'}_PACKING_LIST_перевод_ЮС.xlsx"
        pl_save_path = os.path.join(config.tmp_path, pl_save_name)
        wb_pl.save(pl_save_path)
        wb_pl.close()


        zip_file_name = f"{cymx.fphm or 'Unknown'}_INV_PL.zip"
        zip_file_path = os.path.join(config.tmp_path, zip_file_name)
        # 打包zip
        with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # 参数1: 要被打包的文件绝对路径
            # 参数2: 打包到 zip 里面后的文件名称 (如果不传参数2，压缩包里会包含服务器的完整层级目录)
            zipf.write(inv_save_path, inv_save_name)
            zipf.write(pl_save_path, pl_save_name)

        # (可选) 打包完成后，清理掉单独的 Excel 临时文件，节省服务器空间
        try:
            os.remove(inv_save_path)
            os.remove(pl_save_path)
        except Exception:
            pass

        # 6. 返回唯一的 ZIP 文件名
        return json_result(1, "生成成功", zip_file_name)

    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f"生成俄文INV/PL失败: {str(e)}")
    finally:
        s.close()




# 优景财务出运清单


def apply_subscript(text_val, index_str):
    """根据下标位置字符串(如'1,3,5')，将对应位置的字符转为下标RichText"""
    if not text_val or not index_str:
        return text_val
    indices = []
    # 解析逗号分隔的数字 (Delphi中是 1-based, 转换为 Python 的 0-based)
    for x in re.split(r'[,，]', str(index_str)):
        x = x.strip()
        if x.isdigit():
            indices.append(int(x) - 1)
            
    elements = []
    font_sub = InlineFont(vertAlign='subscript')
    font_norm = InlineFont()
    
    for i, char in enumerate(str(text_val)):
        if i in indices:
            elements.append(TextBlock(font_sub, char))
        else:
            elements.append(TextBlock(font_norm, char))
    return CellRichText(*elements)

# ================= 辅助函数：绘制边框 =================
def set_border(ws, row_idx, start_col, end_col):
    """为指定行的区间设置细实线边框 (对应 Delphi 的 Borders.LineStyle := 1)"""
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for col in range(start_col, end_col + 1):
        ws.cell(row=row_idx, column=col).border = border

def safe_write(ws, coord, value, num_format=None):
    """
    智能穿透写入：如果目标单元格被合并了，自动找到该合并区域的左上角主单元格进行写入。
    完美解决 MergedCell is read-only 报错！
    """
    row, col = coordinate_to_tuple(coord)
    target_cell = ws[coord]
    
    # 检查该坐标是否在某个合并单元格范围内
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            # 如果被合并了，就把目标切换为该区域的左上角真实单元格
            target_cell = ws.cell(row=mr.min_row, column=mr.min_col)
            break
            
    target_cell.value = value
    if num_format:
        target_cell.number_format = num_format


@any_route('/api/saier/shipment/finance_list/export', methods=['POST'])
@require_token
async def api_export_uv_finance_list(request):
    """
     - 财务出运清单 导出 (支持部门拆分、双模板、高精度、电商杂费)
    """
    j = await request.json()
    s = Session()
    try:
        record_id = j.get('record_id')
        if not record_id: return json_result(-1, "缺少单据ID")
    
        user = request.current_user
        
        # 1. 调用全局函数获取动态权限与岗位
        user_info = get_user_path(user.username)
        user_path = str(user_info.get('path', ''))
        user_pos = str(user_info.get('position', ''))
        
        # 权限校验：外销、单证、财务
        if not any(role in user_path + user_pos for role in ["外销", "单证", "财务"]):
            return json_result(-1, "权限不足！")
            
        # 智能模板路由：外销用业务版，其他用财务版
        is_sales = "外销" in user_path
        tpl_name = 'uvcymxyw.xlsx' if is_sales else 'uvcymx.xlsx'

        # 2. 接收保留小数位数 (满血版特性，默认2位)
        djws = str(j.get('decimal_places', '2'))
        if djws not in ['1', '2', '3', '4']: djws = '2'
        dd_map = {'1': '0.0', '2': '0.00', '3': '0.000', '4': '0.0000'}
        dd = dd_map[djws]

        # 3. 获取主表数据
        master = s.execute(text("SELECT * FROM cymx WHERE rid=:id"), {"id": record_id}).fetchone()
        if not master: return json_result(-1, "未找到该出运单据")
        
        if not master.chyrq and "BEST PRICE" not in str(master.khmc or "").upper():
            return json_result(-1, "请设置货物状态为已出运且出运日期不能为空！")

        # 货币与景业特殊逻辑判断
        jysb = '1' if master.wfgs == '宁波景业国际贸易有限公司' else ''
        hbdm_raw = str(master.hbdm or '').upper()
        hbdm = 'USD' if hbdm_raw in ('USD', 'USD$') else hbdm_raw
        hbdmfh = s.execute(text("SELECT bjdh FROM hbdm WHERE bjdid=:bjdid"), {"bjdid": hbdm}).scalar() or ''
        rmbkh = (master.RMBkh == '是')

        # 4. 部门拆分判定逻辑 (满血版核心特性)
        cyzgl = s.execute(text("SELECT 1 FROM cyzglsheet WHERE xm=:xm AND zm='财务出运清单分部门'"), {"xm": master.khmc}).fetchone()
        fcsb = '1' if cyzgl else ''
        
        dept_rows = s.execute(text("SELECT wxdbm1 FROM cymxsheet WHERE pid=:id GROUP BY wxdbm1"), {"id": record_id}).fetchall()
        wxzbm = [r.wxdbm1 for r in dept_rows if r.wxdbm1]
        
        # 即使有多个部门，如果不符合分表条件，也只打印总表
        if not wxzbm or len(wxzbm) <= 1 or not fcsb:
            wxzbm = ['1']
        else:
            # 原逻辑：如果分表，要在列表最前面插入 '1' 作为总表
            if '1' not in wxzbm:
                wxzbm.insert(0, '1')

        # 5. 加载模板并复制 Sheet
        r_path = os.path.join(config.data_upload_path, 'template')
        fn = os.path.join(r_path, tpl_name)
        if not os.path.exists(fn): return json_result(-1, f"找不到模板: {tpl_name}")

        wb = load_workbook(fn)
        base_sheet = wb.active
        for sheet_name in wb.sheetnames:
            if sheet_name != base_sheet.title: del wb[sheet_name]
        base_sheet.sheet_view.showGridLines = True

        sheets_data = []
        for i, dept in enumerate(wxzbm):
            ws = base_sheet if i == 0 else wb.copy_worksheet(base_sheet)
            ws.sheet_view.showGridLines = True
            ws.title = '总表' if dept == '1' else str(dept)
            sheets_data.append((dept, ws))

        # ================= 6. 预加载全局数据字典 (彻底消灭 N+1) =================
        all_gcmc_rows = s.execute(
            text("SELECT DISTINCT gcmc FROM cymxsheet WHERE pid=:id UNION SELECT DISTINCT gcmc FROM bgmxdsheet WHERE pid IN (SELECT rid FROM bgmxd WHERE fphm=:fphm)"), 
            {"id": record_id, "fphm": master.fphm}
        ).fetchall()
        factory_names = [r.gcmc for r in all_gcmc_rows if r.gcmc]
        
        zycs_map = {}
        if factory_names:
            zycs_rows = s.execute(text("SELECT company_name, cymch, csjc FROM zycs WHERE company_name IN :names OR cymch IN :names"), {"names": tuple(factory_names)}).fetchall()
            for zr in zycs_rows:
                if zr.csjc and zr.csjc != '无':
                    zycs_map[zr.company_name] = zr.csjc
                    zycs_map[zr.cymch] = zr.csjc

        bg_rmb_map = {}
        bgmxd_info = s.execute(text("SELECT rid FROM bgmxd WHERE fphm=:fphm"), {"fphm": master.fphm}).fetchone()
        if bgmxd_info:
            old_orders = s.execute(text("SELECT fphm, RMBkh FROM cymx WHERE fphm IN (SELECT DISTINCT yfph FROM bgmxdsheet WHERE pid=:pid AND yfph!='')"), {"pid": bgmxd_info.rid}).fetchall()
            bg_rmb_map = {r.fphm: (r.RMBkh == '是') for r in old_orders}

        # ================= 7. 多部门 Sheet 遍历渲染 =================
        for dept, ws in sheets_data:
            
            # 【扫雷行动】：解除第 3 行以下的所有阻碍写入的合并单元格
            for mr in list(ws.merged_cells.ranges):
                if mr.min_row >= 3:
                    ws.unmerge_cells(str(mr))

            # 写入表头 (修复原版中 ysfp 与 fphm 的互相覆盖 Bug)
            if dept == '1': 
                safe_write(ws, 'G1', str(master.fphm or '').strip())
                
            safe_write(ws, 'K1', master.xh, '@')
            safe_write(ws, 'S1', master.fh, '@')
            safe_write(ws, 'V1', str(master.khmc or '').strip())
            safe_write(ws, 'Y1', master.zdry)
            safe_write(ws, 'Z1', master.shqx)
            safe_write(ws, 'AA1', master.qxnr)
            safe_write(ws, 'AK1', master.jhfs)
            
            if master.xybx == '是': 
                safe_write(ws, 'BC1', f"信保({float(master.xbfl or 0) * 100}%)")
                
            if rmbkh:
                ws.column_dimensions['J'].width = 12
                safe_write(ws, 'J2', '客户人民币单价')
                ws.column_dimensions['V'].width = 12
                safe_write(ws, 'V2', '客户人民币总价')
                
            if jysb == '1':
                safe_write(ws, 'BD2', '景业单价')
                safe_write(ws, 'BE2', '头程费用')
                
            safe_write(ws, 'BF2', '客人品牌')

            # ----------------- 数据分流 -----------------
            if dept == '1':
                all_details = s.execute(text("SELECT * FROM cymxsheet WHERE pid=:id"), {"id": record_id}).fetchall()
            else:
                all_details = s.execute(text("SELECT * FROM cymxsheet WHERE pid=:id AND wxdbm1=:dept"), {"id": record_id, "dept": dept}).fetchall()
            
            real_products = [r for r in all_details if r.cpbh or float(r.chsl or 0) > 0]
            empty_fees = [r for r in all_details if not r.cpbh and float(r.chsl or 0) == 0 and r.zwpm]
            pending_customs = [r for r in all_details if r.sfdb == '是']
            
            # 【满血版专属】：基于明细的退税精准聚加统计
            tax_totals = {} 
            xs = sl = 0
            zmz = zjz = ztj = wxzj = cgzj = 0.0
            
            # ----------------- A. 核心商品明细 -----------------
            k = 0
            for row in real_products:
                k += 1
                r_idx = 2 + k
                set_border(ws, r_idx, 4, 58)  # 原代码是 D列  BF列 对应着 4 ,58  对应画线 
                
                # 累加本行退税额 (修复原版 a17, a16 混乱累加)
                tsl = int(float(row.tsl or 0))
                tax_totals[tsl] = tax_totals.get(tsl, 0.0) + float(row.gczjrmb or 0)
                
                ws[f'A{r_idx}'] = row.bzcd
                ws[f'B{r_idx}'] = row.bzkd
                ws[f'C{r_idx}'] = row.bzgd
                ws[f'D{r_idx}'] = row.wxht
                ws[f'E{r_idx}'] = row.khht
                ws[f'F{r_idx}'].number_format = '@'
                ws[f'G{r_idx}'].number_format = '@'
                if row.czkrhh:
                    ws[f'F{r_idx}'] = row.czkrhh
                    ws[f'G{r_idx}'] = str(row.czkrhh) + '    '
                else:
                    ws[f'F{r_idx}'] = row.khhh
                    ws[f'G{r_idx}'] = str(row.cpbh or '') + '    '
                
                # 剔除 tspmb 后的纯净品名兜底
                ws[f'H{r_idx}'] = str(row.djpmy or '').upper()
                ws[f'AN{r_idx}'] = str(row.djpmy or '').upper()
                ws[f'AO{r_idx}'] = str(row.djpmw or '').upper()

                ws[f'I{r_idx}'] = row.zwpm
                ws[f'I{r_idx}'].alignment = Alignment(horizontal='left')
                
                # 动态位数的金额控制
                if rmbkh:
                    ws[f'J{r_idx}'] = float(row.mjdj1 or 0)
                    ws[f'J{r_idx}'].number_format = f'￥#,##{dd}'
                    val_zj = float(row.dsmjzj or 0) if jysb == '1' else float(row.mjzj or 0)
                    ws[f'V{r_idx}'] = val_zj
                    ws[f'V{r_idx}'].number_format = f'￥#,##{dd}'
                    wxzj += val_zj
                else:
                    val_dj = float(row.dswxjg or 0) if jysb == '1' else float(row.wxjg or 0)
                    ws[f'J{r_idx}'] = round(val_dj, 2) if jysb == '1' else val_dj
                    ws[f'J{r_idx}'].number_format = f'"{hbdmfh}"#,##{dd}'
                    val_zj = float(row.dswxzj or 0) if jysb == '1' else float(row.wxzj or 0)
                    ws[f'V{r_idx}'] = val_zj
                    ws[f'V{r_idx}'].number_format = f'"{hbdmfh}"#,##{dd}'
                    wxzj += val_zj
                    
                ws[f'K{r_idx}'] = float(row.gcjg or 0)
                ws[f'L{r_idx}'] = row.jldw
                ws[f'M{r_idx}'] = float(row.wxrl or 0)
                ws[f'N{r_idx}'] = int(row.chxs or 0)
                xs += int(row.chxs or 0)
                ws[f'O{r_idx}'] = float(row.wxmz or 0)
                ws[f'P{r_idx}'] = float(row.wxjz or 0)
                ws[f'Q{r_idx}'] = float(row.wxtj or 0)
                ws[f'R{r_idx}'] = float(row.zmz or 0)
                zmz += float(row.zmz or 0)
                ws[f'S{r_idx}'] = float(row.zjz or 0)
                zjz += float(row.zjz or 0)
                ws[f'T{r_idx}'] = float(row.chsl or 0)
                sl += int(row.chsl or 0)
                ws[f'U{r_idx}'] = float(row.ztj or 0)
                ztj += float(row.ztj or 0)
                
                ws[f'W{r_idx}'].number_format = '￥#,##0.00' if ('RMB' in str(row.cghbdm or '') or not row.cghbdm) else '$#,##0.00'
                ws[f'W{r_idx}'] = float(row.gczj or 0)
                cgzj += float(row.gczjrmb or 0)
                
                ws[f'X{r_idx}'] = zycs_map.get(row.gcmc, row.gcmc)
                
                ytsb = str(row.ytsb or '')
                if ytsb == '预填': ws[f'Y{r_idx}'] = f"{row.zhwbgpm}(预填)"
                elif ytsb == '预填KGS': ws[f'Y{r_idx}'] = f"{row.zhwbgpm}(预填KGS)"
                else: ws[f'Y{r_idx}'] = row.zhwbgpm
                
                ws[f'Z{r_idx}'] = row.zzsl
                ws[f'AB{r_idx}'] = row.gcdh
                ws[f'AC{r_idx}'] = row.jhrq
                
                ws[f'AD{r_idx}'].number_format = '@'
                div_val = float(row.mjzj or 0) if rmbkh else (float(row.dswxzj or 0) if jysb == '1' else float(row.wxzj or 0))
                if div_val > 0: ws[f'AD{r_idx}'] = f"{float(row.gczjrmb or 0) / div_val:.2f}"

                ws[f'AE{r_idx}'] = row.ywrya
                ws[f'AF{r_idx}'] = row.ywry
                ws[f'AG{r_idx}'] = row.gdry
                ws[f'AJ{r_idx}'] = row.zkfy
                ws[f'AK{r_idx}'] = row.bz1
                ws[f'AL{r_idx}'] = row.sfpx
                ws[f'AM{r_idx}'] = row.krcode
                ws[f'AM{r_idx}'].number_format = '@'
                
                ws.row_dimensions[r_idx].height = 20
                #  前面涉及到tspmb 表的 已经写过了
                ws[f'AP{r_idx}'] = row.sfsj
                ws[f'AQ{r_idx}'] = row.djje if float(row.djje or 0) > 0 else row.yfje
                ws[f'AR{r_idx}'] = row.tsl
                ws[f'AS{r_idx}'] = row.scrq
                ws[f'AT{r_idx}'] = row.kpgc
                ws[f'AU{r_idx}'] = row.zzjgdm
                ws[f'AV{r_idx}'] = row.kplxr
                ws[f'AW{r_idx}'] = row.kpdh
                ws[f'AX{r_idx}'] = row.yjcq
                ws[f'AY{r_idx}'] = row.hgbm
                ws[f'AY{r_idx}'].number_format = '@'
                ws[f'AZ{r_idx}'] = row.caiziz
                ws[f'BA{r_idx}'] = row.hgjldw
                ws[f'BB{r_idx}'] = master.chyrq
                ws[f'BC{r_idx}'] = row.hyd
                ws[f'BG{r_idx}'] = row.krddh
                if jysb == '1':
                    ws[f'BD{r_idx}'] = f"{float(row.dsdj or 0):.2f}"
                    ws[f'BE{r_idx}'] = f"{float(row.dsfy or 0):.2f}"
                ws[f'BF{r_idx}'] = row.krpp

            # ----------------- B. 核心商品小计 -----------------
            k += 1
            r_idx = 2 + k
            ws[f'N{r_idx}'] = xs
            # ws[f'T{r_idx}'] = sl    原delphi代码中没有的
            ws[f'R{r_idx}'] = f"{zmz:.1f}"
            ws[f'S{r_idx}'] = f"{zjz:.1f}"
            ws[f'U{r_idx}'] = f"{ztj:.2f}"
            
            ws[f'V{r_idx}'] = wxzj
            ws[f'V{r_idx}'].number_format = f'￥#,##{dd}' if rmbkh else f'"{hbdmfh}"#,##{dd}'
            
            ws[f'W{r_idx}'] = f"{cgzj:.2f}"
            ws[f'W{r_idx}'].number_format = '￥#,##0.00'
            
            ws[f'AD{r_idx}'].number_format = '@'
            div_total = float(master.mjzj or 0) if rmbkh else float(master.wxje or 0)
            if div_total > 0: ws[f'AD{r_idx}'] = f"{float(master.cghjzje or 0) / div_total:.2f}"

            # ----------------- C. 空行与基础杂费 -----------------
            for f_row in empty_fees:
                k += 1
                r_idx = 2 + k
                set_border(ws, r_idx, 4, 58)
                ws.merge_cells(f'I{r_idx}:U{r_idx}')
                ws[f'I{r_idx}'] = f_row.zwpm
                ws[f'I{r_idx}'].alignment = Alignment(horizontal='left')
                ws[f'V{r_idx}'] = float(f_row.mjzj or 0) if rmbkh else float(f_row.wxzj or 0)
                ws[f'V{r_idx}'].number_format = f'￥#,##{dd}' if rmbkh else f'"{hbdmfh}"#,##{dd}'
                ws[f'W{r_idx}'] = float(f_row.gczj or 0)
                ws[f'W{r_idx}'].number_format = '￥#,##0.00' if ('RMB' in str(f_row.cghbdm or '') or not f_row.cghbdm) else '$#,##0.00'
                ws[f'X{r_idx}'] = zycs_map.get(f_row.gcmc, f_row.gcmc)
            # ----------------- D. 满血版全局 加项/减项/明佣 -----------------
            if float(master.jxUSD or 0) > 0 or float(master.jxRMB or 0) > 0:
                k += 1
                r_idx = 2 + k
                ws[f'U{r_idx}'] = '合 计'
                if float(master.jxUSD or 0) > 0 and float(master.jxKHRMB or 0) > 0:
                    ws[f'V{r_idx}'] = f"￥{float(master.jxKHRMB):.2f}/{hbdmfh}{float(master.jxUSD):.2f}"
                else:
                    if float(master.jxUSD or 0) > 0:
                        ws[f'V{r_idx}'] = float(master.jxUSD)
                        ws[f'V{r_idx}'].number_format = '$#,##0.00'
                    if float(master.jxKHRMB or 0) > 0:
                        ws[f'V{r_idx}'] = float(master.jxKHRMB)
                        ws[f'V{r_idx}'].number_format = '￥#,##0.00'
                ws[f'W{r_idx}'] = float(master.jxRMB or 0)
                ws[f'AD{r_idx}'].number_format = '@'

            if float(master.myjje or 0) > 0:
                k += 1
                r_idx = 2 + k
                ws[f'U{r_idx}'] = '明佣金额'
                ws[f'V{r_idx}'] = f"{'￥' if rmbkh else hbdmfh}{float(master.myjje):.2f}"
                ws[f'V{r_idx}'].number_format = '@'
                ws[f'AD{r_idx}'].number_format = '@'

            # 满血版明细加减项 (jxje1..3, jjxje1..3)
            for i in range(1, 4):
                if float(getattr(master, f'jxje{i}', 0)) > 0:
                    k += 1; r_idx = 2 + k
                    ws.merge_cells(f'I{r_idx}:U{r_idx}')
                    ws[f'I{r_idx}'] = f"加项名称：{getattr(master, f'jxmc{i}', '')}"
                    ws[f'V{r_idx}'] = float(getattr(master, f'jxje{i}', 0))
                    ws[f'V{r_idx}'].number_format = f'￥#,##0.00' if rmbkh else f'"{hbdmfh}"#,##0.00'
                    ws[f'AD{r_idx}'].number_format = '@'
                    
                if float(getattr(master, f'jjxje{i}', 0)) > 0:
                    k += 1; r_idx = 2 + k
                    ws.merge_cells(f'I{r_idx}:U{r_idx}')
                    ws[f'I{r_idx}'] = f"减项名称：{getattr(master, f'jjxmc{i}', '')}"
                    ws[f'V{r_idx}'] = float(getattr(master, f'jjxje{i}', 0))
                    ws[f'V{r_idx}'].number_format = f'￥#,##0.00' if rmbkh else f'"{hbdmfh}"#,##0.00'
                    ws[f'AD{r_idx}'].number_format = '@'

            # ----------------- E. 说明及备注区 -----------------
            r_idx_6 = 6 + k
            ws[f'D{r_idx_6}'].font = Font(bold=True)
            ws[f'D{r_idx_6}'] = '注意事项:'
            ws.merge_cells(f'G{r_idx_6}:AM{r_idx_6}')
            ws.row_dimensions[r_idx_6].height = 30
            ws[f'BJ{r_idx_6}'] = str(master.zysx or '').strip() + '\n'
            ws[f'G{r_idx_6}'] = str(master.zysx or '').strip() + '\n'

            qtsm1 = str(master.qtshm or '')
            if '暗佣' not in qtsm1.upper() and float(master.ayjje or 0) > 0:
                qtsm1 += f";暗佣金额：{'￥' if rmbkh else hbdmfh}{float(master.ayjje)}"
                    
            r_idx_7 = 7 + k
            ws[f'D{r_idx_7}'].font = Font(bold=True)
            ws[f'D{r_idx_7}'] = '其它说明:'
            ws.merge_cells(f'G{r_idx_7}:AM{r_idx_7}')
            ws.row_dimensions[r_idx_7].height = 30
            ws[f'BJ{r_idx_7}'] = qtsm1.strip()
            ws[f'G{r_idx_7}'] = qtsm1.strip() + '\n'

            r_idx_8 = 8 + k
            ws[f'D{r_idx_8}'].font = Font(bold=True)
            ws[f'D{r_idx_8}'] = '出运说明:'
            ws.merge_cells(f'G{r_idx_8}:AM{r_idx_8}')
            ws.row_dimensions[r_idx_8].height = 30
            ws[f'BJ{r_idx_8}'] = str(master.cyshm or '').strip() + '\n'
            ws[f'G{r_idx_8}'] = str(master.cyshm or '').strip() + '\n'

            r_idx_9 = 9 + k
            ws[f'D{r_idx_9}'].font = Font(bold=True)
            ws[f'D{r_idx_9}'] = '单证备注:'
            ws.merge_cells(f'G{r_idx_9}:AM{r_idx_9}')
            ws.row_dimensions[r_idx_9].height = 30
            ws[f'BJ{r_idx_9}'] = str(master.dzbz or '').strip()
            ws[f'G{r_idx_9}'] = str(master.dzbz or '').strip() + '\n'

            # ----------------- F. 报关补报信息 (bgmxdsheet) -----------------
            ws[f'D{10 + k}'].font = Font(bold=True)
            ws[f'D{10 + k}'] = '补报信息:'
            if bgmxd_info:
                supp_items = s.execute(text("SELECT * FROM bgmxdsheet WHERE pid=:pid AND yfph!='' AND yfph IS NOT NULL ORDER BY zhwbgpm, gcmc"), {"pid": bgmxd_info.rid}).fetchall()
                if supp_items:
                    k += 1 
                    for bg in supp_items:
                        k += 1
                        br_idx = 9 + k
                        set_border(ws, br_idx, 4, 58)
                        
                        b_rmbkh = bg_rmb_map.get(bg.yfph, False)
                        
                        ws[f'A{br_idx}'] = float(bg.bzcd or 0)
                        ws[f'B{br_idx}'] = float(bg.bzkd or 0)
                        ws[f'C{br_idx}'] = float(bg.bzgd or 0)
                        ws[f'D{br_idx}'] = bg.yfph
                        ws[f'E{br_idx}'] = bg.khht
                        ws[f'F{br_idx}'].number_format = '@'
                        ws[f'G{br_idx}'].number_format = '@'
                        if bg.czkrhh:
                            ws[f'F{br_idx}'] = bg.czkrhh
                            ws[f'G{br_idx}'] = bg.czkrhh
                        else:
                            ws[f'F{br_idx}'] = bg.khhh
                            ws[f'G{br_idx}'] = bg.cpbh
                            
                        ws[f'H{br_idx}'] = str(bg.djpmy or '').upper()
                        ws[f'AN{br_idx}'] = str(bg.djpmy or '').upper()
                        ws[f'AO{br_idx}'] = str(bg.djpmw or '').upper()
                        
                        ws[f'I{br_idx}'] = bg.zwpm
                        ws[f'I{br_idx}'].alignment = Alignment(horizontal='left')
                        
                        if b_rmbkh:
                            ws[f'J{br_idx}'] = float(bg.mjdj1 or 0)
                            ws[f'V{br_idx}'] = float(bg.mjzj or 0)
                            ws[f'J{br_idx}'].number_format = ws[f'V{br_idx}'].number_format = f'￥#,##{dd}'
                        else:
                            if jysb == '1':
                                ds = s.execute(text("SELECT dswxjg, dswxzj FROM cymxsheet WHERE cywyzd=:cywyzd AND fpsb1='是'"), {"cywyzd": bg.cywyzd}).fetchone()
                                ws[f'J{br_idx}'] = float(ds.dswxjg or 0) if ds else 0
                                ws[f'V{br_idx}'] = float(ds.dswxzj or 0) if ds else 0
                            else:
                                ws[f'J{br_idx}'] = float(bg.wxjg or 0)
                                ws[f'V{br_idx}'] = float(bg.wxzj or 0)
                            ws[f'J{br_idx}'].number_format = ws[f'V{br_idx}'].number_format = f'"{hbdmfh}"#,##{dd}'
                            
                        ws[f'K{br_idx}'] = f"{float(bg.gczj or 0)/float(bg.chsl):.2f}" if float(bg.chsl or 0) > 0 else float(bg.gcjg or 0)
                        ws[f'L{br_idx}'] = bg.jldw
                        ws[f'M{br_idx}'] = float(bg.wxrl or 0)
                        ws[f'N{br_idx}'] = int(bg.chxs or 0)
                        ws[f'O{br_idx}'] = float(bg.wxmz or 0)
                        ws[f'P{br_idx}'] = float(bg.wxjz or 0)
                        ws[f'Q{br_idx}'] = float(bg.wxtj or 0)
                        ws[f'R{br_idx}'] = float(bg.zmz or 0)
                        ws[f'S{br_idx}'] = float(bg.zjz or 0)
                        ws[f'T{br_idx}'] = float(bg.chsl or 0)
                        ws[f'U{br_idx}'] = float(bg.ztj or 0)
                        ws[f'W{br_idx}'] = float(bg.gczj or 0)
                        ws[f'X{br_idx}'] = zycs_map.get(bg.gcmc, bg.gcmc)
                        # Y列：预填标识处理




                        ytsb = str(bg.ytsb or '')
                        if ytsb == '预填': 
                            ws[f'Y{br_idx}'] = f"{bg.zhwbgpm}(预填)"
                        elif ytsb == '预填KGS': 
                            ws[f'Y{br_idx}'] = f"{bg.zhwbgpm}(预填KGS)"
                        else: 
                            ws[f'Y{br_idx}'] = bg.zhwbgpm
                            
                        # Z到AC列：税率、合同等
                        ws[f'Z{br_idx}'] = bg.zzsl
                        ws[f'AA{br_idx}'] = ''
                        ws[f'AB{br_idx}'] = bg.gcdh
                        ws[f'AC{br_idx}'] = bg.jhrq
                        
                        # AD列：占比计算 (复用之前我们提炼的极简除法防御逻辑)
                        ws[f'AD{br_idx}'].number_format = '@'
                        # # 根据补报原单的货币类型选择分母
                        # div_val = float(bg.mjzj or 0) if b_rmbkh else float(bg.wxzj or 0)
                        # if div_val > 0: 
                        #     ws[f'AD{br_idx}'] = f"{float(bg.gczj or 0) / div_val:.2f}"
                        if b_rmbkh:
                            # 还原 Delphi 逻辑：如果是人民币客户，除以 mjzj
                            if float(bg.mjzj or 0) > 0:
                                ws[f'AD{br_idx}'] = f"{float(bg.gczj or 0) / float(bg.mjzj):.2f}"
                        else:
                            # 还原 Delphi 逻辑：如果是外币客户
                            if float(bg.wxzj or 0) > 0:
                                if jysb == '1':
                                    # 【完美复刻原作者的幽灵代码块】：什么都不做！AD列直接留空！   
                                    pass 
                                else:
                                    ws[f'AD{br_idx}'] = f"{float(bg.gczj or 0) / float(bg.wxzj):.2f}"
                        # AE到AL列：各种人员、费用、备注
                        ws[f'AE{br_idx}'] = bg.ywrya
                        ws[f'AF{br_idx}'] = bg.ywry
                        ws[f'AG{br_idx}'] = bg.gdry
                        ws[f'AJ{br_idx}'] = bg.zkfy
                        ws[f'AK{br_idx}'] = bg.bz1
                        ws[f'AL{br_idx}'] = bg.sfpx
                        
                        # AM列：KR Code
                        ws[f'AM{br_idx}'].number_format = '@'
                        ws[f'AM{br_idx}'] = bg.krcode
                        
                        # AN, AO列：废弃 tspmb，直接转大写兜底
                        ws[f'AN{br_idx}'] = str(bg.djpmy or '').upper()
                        ws[f'AO{br_idx}'] = str(bg.djpmw or '').upper()
                        
                        # AP到AX列：基础报关属性
                        ws[f'AP{br_idx}'] = bg.sfsj
                        ws[f'AQ{br_idx}'] = bg.djje
                        ws[f'AR{br_idx}'] = bg.tsl
                        ws[f'AS{br_idx}'] = bg.scrq
                        ws[f'AT{br_idx}'] = bg.kpgc
                        ws[f'AU{br_idx}'] = bg.zzjgdm
                        ws[f'AV{br_idx}'] = bg.kplxr
                        ws[f'AW{br_idx}'] = bg.kpdh
                        ws[f'AX{br_idx}'] = bg.yjcq
                        
                        # AY到BC列：海关硬指标
                        ws[f'AY{br_idx}'].number_format = '@'
                        ws[f'AY{br_idx}'] = bg.hgbm
                        ws[f'AZ{br_idx}'] = bg.caiziz
                        ws[f'BA{br_idx}'] = bg.hgjldw
                        ws[f'BB{br_idx}'] = bg.chyrq
                        ws[f'BC{br_idx}'] = bg.hyd
                        
                        # BD, BE列：景业公司专属价格
                        if jysb == '1':
                            ws[f'BD{br_idx}'] = f"{float(bg.dsdj or 0):.2f}"
                            ws[f'BE{br_idx}'] = f"{float(bg.dsfy or 0):.2f}"
                            
                        # BF, BG列：客人品牌与订单号
                        ws[f'BF{br_idx}'] = bg.krpp
                        ws[f'BG{br_idx}'] = bg.krddh

            # ----------------- G. 待补报信息 -----------------
            if pending_customs:
                k += 1
                ws[f'D{10 + k}'].font = Font(bold=True)
                ws[f'D{10 + k}'] = '待补报信息:'
                for pc in pending_customs:
                    k += 1
                    p_idx = 10 + k
                    ws.merge_cells(f'H{p_idx}:T{p_idx}')
                    ws[f'H{p_idx}'].alignment = Alignment(horizontal='left', wrap_text=True)
                    ws.row_dimensions[p_idx].height = 30
                    ws[f'H{p_idx}'] = f"货号：{pc.cpbh or ''} 品名:{pc.zhwbgpm or ''} 工厂:{pc.gcmc or ''} 等待补报"

            # ================== H. 满血版 平行排版区 ==================
            base_row = 11 + k
            
            # (左侧): 装柜与超长电商附加费 (D列)
            r_box = base_row
            ws[f'D{r_box}'] = str(master.zgdd or '') + '仓库装柜'; r_box+=1
            ws[f'D{r_box}'] = '监装人：' + str(master.jzry or ''); r_box+=1
            ws[f'D{r_box}'] = '日期：' + str(master.zgrq or ''); r_box+=1
            ws[f'D{r_box}'] = '空柜：' + str(master.hgzl or ''); r_box+=1
            ws[f'D{r_box}'] = '车子：' + str(master.czzl or ''); r_box+=1
            ws[f'D{r_box}'] = '车子+空柜：' + str(master.cjkg or ''); r_box+=1
            ws[f'D{r_box}'] = '车子+重柜：' + str(master.cjgz or ''); r_box+=1

            if float(master.fyzj1 or 0) > 0: 
                ws[f'D{r_box}'] = f"业务额外费用￥{master.fyzj1}"; r_box+=1
            if sum(float(getattr(master, f, 0)) for f in ['hdsyfM', 'qggsM', 'qgzs$', 'mdgfy']) > 0:
                ws[f'D{r_box}'] = f"电商海运费${master.hdsyfM}货代{master.Mhd or ''}"; r_box+=1
            if float(master.hdsyfR or 0) > 0: 
                ws[f'D{r_box}'] = f"电商海运费￥{master.hdsyfR}货代{master.Mhd or ''}"; r_box+=1
            if float(master.gnnlf or 0) > 0: 
                ws[f'D{r_box}'] = f"国内内陆费￥{master.gnnlf}货代{master.Rhd or ''}{master.fyzc or ''}"; r_box+=1
            if float(master.mdgfy or 0) > 0: 
                ws[f'D{r_box}'] = f"目的港费用${master.mdgfy}货代{master.Ehd or ''}"; r_box+=1
            if float(master.hybfM or 0) > 0: 
                ws[f'D{r_box}'] = f"海运保费${master.hybfM}"; r_box+=1
            if float(master.dshl or 0) > 0 and float(master.dshl) != 1: 
                ws[f'D{r_box}'] = f"电商汇率{master.dshl}"; r_box+=1

            # ----------------- H. 工厂付款方式汇总 (1:1 还原 Delphi 逐行打印版) -----------------
            # 1. 绘制表头 (包含居中对齐)
            r_head = 11 + k
            ws.merge_cells(f'H{r_head}:I{r_head}')
            ws[f'H{r_head}'] = '工厂'
            ws[f'H{r_head}'].font = Font(bold=True)
            ws[f'H{r_head}'].alignment = Alignment(horizontal='center')

            ws.merge_cells(f'J{r_head}:T{r_head}')
            ws[f'J{r_head}'] = '付款方式'
            ws[f'J{r_head}'].font = Font(bold=True)
            ws[f'J{r_head}'].alignment = Alignment(horizontal='center')

            ws[f'U{r_head}'] = '进仓日期'
            ws[f'U{r_head}'].font = Font(bold=True)
            ws[f'U{r_head}'].alignment = Alignment(horizontal='center')

            # 2. 逐行打印工厂 (严格还原 Delphi：有多少商品就印多少行，不再去重)
            r_fac = r_head
            for rp in real_products:
                r_fac += 1
                ws.merge_cells(f'H{r_fac}:I{r_fac}')
                ws[f'H{r_fac}'] = zycs_map.get(rp.gcmc, rp.gcmc)
                
                ws.merge_cells(f'J{r_fac}:T{r_fac}')
                ws[f'J{r_fac}'] = rp.jsfs
                ws[f'J{r_fac}'].alignment = Alignment(wrap_text=True, vertical='center')
                ws.row_dimensions[r_fac].height = 20
                
                ws[f'U{r_fac}'].number_format = '@'
                ws[f'U{r_fac}'] = rp.scrq


            # ----------------- I. 退税率统计表 (1:1 还原历史遗留 Bug 版) ----------------- 
            # 1. 模拟 Delphi 中的 11 个孤立变量
            a17 = tax_totals.get(17, 0.0)
            a16 = tax_totals.get(16, 0.0)
            a15 = tax_totals.get(15, 0.0)
            a13 = tax_totals.get(13, 0.0)
            a10 = tax_totals.get(10, 0.0)
            a9  = tax_totals.get(9, 0.0)
            a6  = tax_totals.get(6, 0.0)
            a5  = tax_totals.get(5, 0.0)
            a3  = tax_totals.get(3, 0.0)
            a1  = tax_totals.get(1, 0.0)
            a0  = tax_totals.get(0, 0.0)

            # 2. 绘制固定的表头
            ws[f'W{r_head}'] = '退税'; ws[f'W{r_head}'].alignment = Alignment(horizontal='center')
            ws[f'X{r_head}'] = '金额'; ws[f'X{r_head}'].alignment = Alignment(horizontal='center')
            set_border(ws, r_head, 23, 24)

            # 3. 开启原作者的“史诗级”判断分支
            if a17 > 0:
                ws[f'W{12+k}'] = '17%'; ws[f'X{12+k}'] = a17; set_border(ws, 12+k, 23, 24)
                ws[f'W{13+k}'] = '16%'; ws[f'X{13+k}'] = a16; set_border(ws, 13+k, 23, 24)
                ws[f'W{14+k}'] = '15%'; ws[f'X{14+k}'] = a15; set_border(ws, 14+k, 23, 24)
                ws[f'W{15+k}'] = '13%'; ws[f'X{15+k}'] = a13; set_border(ws, 15+k, 23, 24)
                
                set_border(ws, 16+k, 23, 24)
                if a9 == 0:
                    ws[f'W{16+k}'] = '10%'; ws[f'X{16+k}'] = a10
                else:
                    ws[f'W{16+k}'] = '9%';  ws[f'X{16+k}'] = a9
                    
                set_border(ws, 17+k, 23, 24)
                if a6 == 0:
                    ws[f'W{17+k}'] = '6%';  ws[f'X{17+k}'] = a6
                else:
                    ws[f'W{17+k}'] = '5%';  ws[f'X{17+k}'] = a5
                    
                ws[f'W{18+k}'] = '3%';  ws[f'X{18+k}'] = a3; set_border(ws, 18+k, 23, 24)
                ws[f'W{19+k}'] = '0';   ws[f'X{19+k}'] = a0; set_border(ws, 19+k, 23, 24)

            else:
                ws[f'W{12+k}'] = '16%'; ws[f'X{12+k}'] = a16; set_border(ws, 12+k, 23, 24)
                ws[f'W{13+k}'] = '15%'; ws[f'X{13+k}'] = a15; set_border(ws, 13+k, 23, 24)
                ws[f'W{14+k}'] = '13%'; ws[f'X{14+k}'] = a13; set_border(ws, 14+k, 23, 24)
                
                set_border(ws, 15+k, 23, 24)
                if a9 == 0:
                    ws[f'W{15+k}'] = '10%'; ws[f'X{15+k}'] = a10
                else:
                    ws[f'W{15+k}'] = '9%';  ws[f'X{15+k}'] = a9
                    
                set_border(ws, 16+k, 23, 24)
                if a5 == 0:
                    ws[f'W{16+k}'] = '6%';  ws[f'X{16+k}'] = a6
                else:
                    ws[f'W{16+k}'] = '5%';  ws[f'X{16+k}'] = a5
                    
                ws[f'W{17+k}'] = '3%';  ws[f'X{17+k}'] = a3; set_border(ws, 17+k, 23, 24)
                ws[f'W{18+k}'] = '1%';  ws[f'X{18+k}'] = a1; set_border(ws, 18+k, 23, 24)
                ws[f'W{19+k}'] = '0';   ws[f'X{19+k}'] = a0; set_border(ws, 19+k, 23, 24)

            for r in range(12+k, 20+k):
                ws[f'W{r}'].number_format = '@'

            # ================= 对齐最大行号 (准备接下来的打印) =================
            max_row = max(r_box, r_fac, 19+k) + 1
  
            # ----------------- I. sheet6 额外费用 (满血版新增) -----------------
            sheet6 = s.execute(text("SELECT * FROM cymxsheet6 WHERE pid=:id"), {"id": record_id}).fetchall()
            for fee in sheet6:
                ws[f'D{max_row}'] = f"费用名称:{fee.fymc or ''}  费用金额:{fee.hbdm or ''}{fee.fyje or ''}  厂商名称:{fee.csmc or ''}"
                max_row += 1
                
            # ----------------- J. sheet10 预填详情表 (满血版新增) -----------------
            if dept == '1':
                sheet10 = s.execute(text("SELECT * FROM cymxsheet10 WHERE pid=:id"), {"id": record_id}).fetchall()
            else:
                sheet10 = s.execute(text("SELECT * FROM cymxsheet10 WHERE pid=:id AND wxdbm1=:dept"), {"id": record_id, "dept": dept}).fetchall()
                
            if sheet10:
                ws[f'D{max_row}'] = '预填详情'
                headers = ['付款抬头', '货源地', '海关计量单位', '中文品名', '增值税率', '退税率', '数量', '金额', '', '', '', '工厂名称']
                for col_offset, h_name in enumerate(headers):
                    if h_name: ws.cell(row=max_row, column=7+col_offset, value=h_name)
                set_border(ws, max_row, 7, 18) # 画表头边框
                max_row += 1
                
                for pr in sheet10:
                    ws[f'G{max_row}'] = pr.fktt
                    ws[f'H{max_row}'] = pr.hyd
                    ws[f'I{max_row}'] = pr.hgjldw
                    ws[f'J{max_row}'] = pr.zwpm
                    ws[f'K{max_row}'].number_format = '@'; ws[f'K{max_row}'] = pr.zzsl
                    ws[f'L{max_row}'] = pr.zsl
                    ws[f'M{max_row}'] = pr.chsl
                    ws[f'N{max_row}'] = pr.ytje
                    ws[f'R{max_row}'] = pr.sccj
                    set_border(ws, max_row, 7, 18) # 画明细边框
                    max_row += 1

            # ----------------- K. 签名栏排版 -----------------
            sign_row = max_row + 2 
            ws.merge_cells(f'D{sign_row}:AM{sign_row}')
            sign_cell = ws[f'D{sign_row}']
            sign_cell.font = Font(bold=True, size=14)
            sign_cell.value = f"单证签字: {master.zdry or ''}                         经理签字:                                     总经理签字:                                                      改单签字:"

        # ================= 保存导出 =================
        excel_file_name = f"{master.fphm or 'Unknown'}_cymx.xlsx"
        save_dir = config.tmp_path
        if not os.path.exists(save_dir):
            os.makedirs(save_dir, exist_ok=True)
            
        excel_save_path = os.path.join(save_dir, excel_file_name)
        wb.save(excel_save_path)
        wb.close()

        return json_result(1, '导出成功', excel_file_name)

    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f"生成财务出运清单失败: {str(e)}")
    finally:
        s.close()







# 优景财务出运清单 总   总总总总 

@any_route('/api/saier/shipment/finance_list_total/export', methods=['POST'])
@require_token
async def api_export_uv_finance_list_total(request):
    """
    优景专属 - 财务出运清单(总) 导出
    【数据库规范版】：全面适配 主表rid 与 子表pid 的最新外键结构
    """
    j = await request.json()
    s = Session()
    try:
        record_id = j.get('record_id')
        if not record_id: return json_result(-1, "缺少单据ID")
            
        user = request.current_user
        user_info = get_user_path(user.username)
        # 1. 专属部门与权限校验
        user_path = str(user_info.get('path', ''))
        user_pos = str(user_info.get('position', ''))
        if "优景" not in user_path:
            return json_result(-1, "权限不足！")
            
        if not any(role in user_path + user_pos for role in ["外销", "单证", "财务"]):
            return json_result(-1, "权限不足：需要外销、单证或财务权限！")

        # 2. 接收保留小数位数 (默认2位)
        djws = str(j.get('decimal_places', '2'))
        if djws not in ['1', '2', '3', '4']: djws = '2'
        dd_map = {'1': '0.0', '2': '0.00', '3': '0.000', '4': '0.0000'}
        dd = dd_map[djws]

        # 3. 获取主表数据 ( cymx -> rid )
        master = s.execute(text("SELECT * FROM cymx WHERE rid=:id"), {"id": record_id}).fetchone()
        if not master: return json_result(-1, "未找到该出运单据")
        
        if not master.chyrq and "BEST PRICE" not in str(master.khmc or "").upper():
            return json_result(-1, "请设置货物状态为已出运且出运日期不能为空！")

        # 货币与景业判断
        jysb = '1' if master.wfgs == '宁波景业国际贸易有限公司' else ''
        hbdm_raw = str(master.hbdm or '').upper()
        hbdm = 'USD' if hbdm_raw in ('USD', 'USD$') else hbdm_raw
        hbdmfh = s.execute(text("SELECT bjdh FROM hbdm WHERE bjdid=:bjdid"), {"bjdid": hbdm}).scalar() or ''
        rmbkh = (master.RMBkh == '是')

        # 决定是否分部门拆分 Sheet ( cymxsheet -> pid )
        cyzgl = s.execute(text("SELECT 1 FROM cyzglsheet WHERE xm=:xm AND zm='财务出运清单分部门'"), {"xm": master.khmc}).fetchone()
        fcsb = '1' if cyzgl else ''
        
        dept_rows = s.execute(text("SELECT wxdbm1 FROM cymxsheet WHERE pid=:id GROUP BY wxdbm1"), {"id": record_id}).fetchall()
        wxzbm = [r.wxdbm1 for r in dept_rows if r.wxdbm1]
        if not wxzbm or len(wxzbm) <= 1 or not fcsb:
            wxzbm = ['1']

        # 4. 加载模板
        tpl_name = 'uvcymx.xlsx'
        r_path = os.path.join(config.data_upload_path, 'template')
        fn = os.path.join(r_path, tpl_name)
        if not os.path.exists(fn): return json_result(-1, f"找不到模板: {tpl_name}")

        wb = load_workbook(fn)
        base_sheet = wb.active
        for sheet_name in wb.sheetnames:
            if sheet_name != base_sheet.title: del wb[sheet_name]
        base_sheet.sheet_view.showGridLines = True

        sheets_data = []
        for i, dept in enumerate(wxzbm):
            ws = base_sheet if i == 0 else wb.copy_worksheet(base_sheet)
            ws.sheet_view.showGridLines = True
            ws.title = '总表' if dept == '1' else str(dept)
            sheets_data.append((dept, ws))

        # ================= 预加载全局字典映射 =================
        # 【修改点 1】：bgmxd 和 bgmxdsheet 的关联查询改为 rid 和 pid
        all_gcmc_rows = s.execute(
            text("SELECT DISTINCT gcmc FROM cymxsheet WHERE pid=:id UNION SELECT DISTINCT gcmc FROM bgmxdsheet WHERE pid IN (SELECT rid FROM bgmxd WHERE fphm=:fphm)"), 
            {"id": record_id, "fphm": master.fphm}
        ).fetchall()
        factory_names = [r.gcmc for r in all_gcmc_rows if r.gcmc]
        
        zycs_map = {}
        if factory_names:
            zycs_rows = s.execute(text("SELECT company_name, cymch, csjc FROM zycs WHERE company_name IN :names OR cymch IN :names"), {"names": tuple(factory_names)}).fetchall()
            for zr in zycs_rows:
                if zr.csjc and zr.csjc != '无':
                    zycs_map[zr.company_name] = zr.csjc
                    zycs_map[zr.cymch] = zr.csjc

        # 【修改点 2】：查询报关主表的 rid (原 number)
        bg_rmb_map = {}
        bgmxd_info = s.execute(text("SELECT rid FROM bgmxd WHERE fphm=:fphm"), {"fphm": master.fphm}).fetchone()
        if bgmxd_info:
            # 【修改点 3】：查询报关子表用 pid (原 father)
            old_orders = s.execute(text("SELECT fphm, RMBkh FROM cymx WHERE fphm IN (SELECT DISTINCT yfph FROM bgmxdsheet WHERE pid=:pid AND yfph!='')"), {"pid": bgmxd_info.rid}).fetchall()
            bg_rmb_map = {r.fphm: (r.RMBkh == '是') for r in old_orders}

        # ================= 开始多部门 Sheet 循环 =================
        for dept, ws in sheets_data:
            # 2. 写表头 (使用 safe_write 穿透所有的合并单元格，绝不报错！)
            if dept == '1': 
                safe_write(ws, 'G1', str(master.fphm or '').strip())
            safe_write(ws, 'K1', master.xh, '@')
            safe_write(ws, 'S1', master.fh, '@')
            safe_write(ws, 'V1', str(master.khmc or '').strip())
            safe_write(ws, 'Y1', master.zdry)
            safe_write(ws, 'Z1', master.shqx)
            safe_write(ws, 'AA1', master.qxnr)
            safe_write(ws, 'AK1', master.jhfs)
            
            if master.xybx == '是': 
                safe_write(ws, 'BC1', f"信保({float(master.xbfl or 0) * 100}%)")
                
            if rmbkh:
                ws.column_dimensions['J'].width = 12
                safe_write(ws, 'J2', '客户人民币单价')
                ws.column_dimensions['V'].width = 12
                safe_write(ws, 'V2', '客户人民币总价')
                
            if jysb == '1':
                safe_write(ws, 'BD2', '景业单价')
                safe_write(ws, 'BE2', '头程费用')
                
            safe_write(ws, 'BF2', '客人品牌')
            
            # 获取出运明细 (cymxsheet -> pid)
            if dept == '1':
                all_details = s.execute(text("SELECT * FROM cymxsheet WHERE pid=:id"), {"id": record_id}).fetchall()
            else:
                all_details = s.execute(text("SELECT * FROM cymxsheet WHERE pid=:id AND wxdbm1=:dept"), {"id": record_id, "dept": dept}).fetchall()
            
            real_products = [r for r in all_details if r.cpbh or float(r.chsl or 0) > 0]
            empty_fees = [r for r in all_details if not r.cpbh and float(r.chsl or 0) == 0 and r.zwpm]
            pending_customs = [r for r in all_details if r.sfdb == '是']
            
            tax_totals = { '17':0, '16':0, '15':0, '13':0, '10':0, '9':0, '6':0, '5':0, '3':0, '1':0, '0':0 }
            xs = sl = 0
            zmz = zjz = ztj = wxzj = cgzj = 0.0
            
            # ----------------- 1. 打印真实商品明细 -----------------
            k = 0
            for row in real_products:
                k += 1
                r_idx = 2 + k
                set_border(ws, r_idx, 4, 58) 
                
                tsl = str(row.tsl or '0')
                if tsl in tax_totals: tax_totals[tsl] += float(row.gczjrmb or 0)
                
                ws[f'A{r_idx}'] = row.bzcd
                ws[f'B{r_idx}'] = row.bzkd
                ws[f'C{r_idx}'] = row.bzgd
                ws[f'D{r_idx}'] = row.wxht
                ws[f'E{r_idx}'] = row.khht
                ws[f'F{r_idx}'].number_format = '@'
                ws[f'G{r_idx}'].number_format = '@'
                if row.czkrhh:
                    ws[f'F{r_idx}'] = row.czkrhh
                    ws[f'G{r_idx}'] = str(row.czkrhh) + '    '
                else:
                    ws[f'F{r_idx}'] = row.khhh
                    ws[f'G{r_idx}'] = str(row.cpbh or '') + '    '
                
                ws[f'H{r_idx}'] = str(row.djpmy or '').upper()
                ws[f'AN{r_idx}'] = str(row.djpmy or '').upper()
                ws[f'AO{r_idx}'] = str(row.djpmw or '').upper()

                ws[f'I{r_idx}'] = row.zwpm
                ws[f'I{r_idx}'].alignment = Alignment(horizontal='left')
                
                if rmbkh:
                    ws[f'J{r_idx}'] = float(row.mjdj1 or 0)
                    ws[f'J{r_idx}'].number_format = f'￥#,##{dd}'
                    val_zj = float(row.dsmjzj or 0) if jysb == '1' else float(row.mjzj or 0)
                    ws[f'V{r_idx}'] = val_zj
                    ws[f'V{r_idx}'].number_format = f'￥#,##{dd}'
                    wxzj += val_zj
                else:
                    val_dj = float(row.dswxjg or 0) if jysb == '1' else float(row.wxjg or 0)
                    ws[f'J{r_idx}'] = round(val_dj, 2) if jysb == '1' else val_dj
                    ws[f'J{r_idx}'].number_format = f'"{hbdmfh}"#,##{dd}'
                    val_zj = float(row.dswxzj or 0) if jysb == '1' else float(row.wxzj or 0)
                    ws[f'V{r_idx}'] = val_zj
                    ws[f'V{r_idx}'].number_format = f'"{hbdmfh}"#,##{dd}'
                    wxzj += val_zj
                    
                ws[f'K{r_idx}'] = float(row.gcjg or 0)
                ws[f'L{r_idx}'] = row.jldw
                ws[f'M{r_idx}'] = float(row.wxrl or 0)
                ws[f'N{r_idx}'] = int(row.chxs or 0)
                xs += int(row.chxs or 0)
                ws[f'O{r_idx}'] = float(row.wxmz or 0)
                ws[f'P{r_idx}'] = float(row.wxjz or 0)
                ws[f'Q{r_idx}'] = float(row.wxtj or 0)
                ws[f'R{r_idx}'] = float(row.zmz or 0)
                zmz += float(row.zmz or 0)
                ws[f'S{r_idx}'] = float(row.zjz or 0)
                zjz += float(row.zjz or 0)
                ws[f'T{r_idx}'] = float(row.chsl or 0)
                sl += int(row.chsl or 0)
                ws[f'U{r_idx}'] = float(row.ztj or 0)
                ztj += float(row.ztj or 0)
                
                ws[f'W{r_idx}'].number_format = '￥#,##0.00' if ('RMB' in str(row.cghbdm or '') or not row.cghbdm) else '$#,##0.00'
                ws[f'W{r_idx}'] = float(row.gczj or 0)
                cgzj += float(row.gczjrmb or 0)
                
                ws[f'X{r_idx}'] = zycs_map.get(row.gcmc, row.gcmc)
                
                ytsb = str(row.ytsb or '')
                if ytsb == '预填': ws[f'Y{r_idx}'] = f"{row.zhwbgpm}(预填)"
                elif ytsb == '预填KGS': ws[f'Y{r_idx}'] = f"{row.zhwbgpm}(预填KGS)"
                else: ws[f'Y{r_idx}'] = row.zhwbgpm
                
                ws[f'Z{r_idx}'] = row.zzsl
                ws[f'AB{r_idx}'] = row.gcdh
                ws[f'AC{r_idx}'] = row.jhrq
                
                ws[f'AD{r_idx}'].number_format = '@'
                div_val = float(row.mjzj or 0) if rmbkh else (float(row.dswxzj or 0) if jysb == '1' else float(row.wxzj or 0))
                if div_val > 0: ws[f'AD{r_idx}'] = f"{float(row.gczjrmb or 0) / div_val:.2f}"

                ws[f'AE{r_idx}'] = row.ywrya
                ws[f'AF{r_idx}'] = row.ywry
                ws[f'AG{r_idx}'] = row.gdry
                ws[f'AJ{r_idx}'] = row.zkfy
                ws[f'AK{r_idx}'] = row.bz1
                ws[f'AL{r_idx}'] = row.sfpx
                ws[f'AM{r_idx}'] = row.krcode
                ws[f'AM{r_idx}'].number_format = '@'
                
                ws.row_dimensions[r_idx].height = 20
                ws[f'AP{r_idx}'] = row.sfsj
                ws[f'AQ{r_idx}'] = row.djje if float(row.djje or 0) > 0 else row.yfje
                ws[f'AR{r_idx}'] = row.tsl
                ws[f'AS{r_idx}'] = row.scrq
                ws[f'AT{r_idx}'] = row.kpgc
                ws[f'AU{r_idx}'] = row.zzjgdm
                ws[f'AV{r_idx}'] = row.kplxr
                ws[f'AW{r_idx}'] = row.kpdh
                ws[f'AX{r_idx}'] = row.yjcq
                ws[f'AY{r_idx}'] = row.hgbm
                ws[f'AY{r_idx}'].number_format = '@'
                ws[f'AZ{r_idx}'] = row.caiziz
                ws[f'BA{r_idx}'] = row.hgjldw
                ws[f'BB{r_idx}'] = master.chyrq
                ws[f'BC{r_idx}'] = row.hyd
                ws[f'BG{r_idx}'] = row.krddh
                if jysb == '1':
                    ws[f'BD{r_idx}'] = f"{float(row.dsdj or 0):.2f}"
                    ws[f'BE{r_idx}'] = f"{float(row.dsfy or 0):.2f}"
                ws[f'BF{r_idx}'] = row.krpp

            # ----------------- 2. 小计合计行 -----------------
            k += 1
            r_idx = 2 + k
            ws[f'N{r_idx}'] = xs
            ws[f'T{r_idx}'] = sl
            ws[f'R{r_idx}'] = f"{float(master.mzhj or 0):.1f}"
            ws[f'S{r_idx}'] = f"{float(master.jzhj or 0):.1f}"
            ws[f'U{r_idx}'] = f"{float(master.tjhj or 0):.2f}"
            
            ws[f'V{r_idx}'] = float(master.mjzj or 0) if rmbkh else float(master.wxje or 0)
            ws[f'V{r_idx}'].number_format = f'￥#,##{dd}' if rmbkh else f'"{hbdmfh}"#,##{dd}'
            
            ws[f'W{r_idx}'] = f"{float(master.cghjzje or 0):.2f}"
            ws[f'W{r_idx}'].number_format = '￥#,##0.00'
            
            ws[f'AD{r_idx}'].number_format = '@'
            div_total = float(master.mjzj or 0) if rmbkh else float(master.wxje or 0)
            if div_total > 0: ws[f'AD{r_idx}'] = f"{float(master.cghjzje or 0) / div_total:.2f}"

            # ----------------- 3. 空行与附加费 -----------------
            for f_row in empty_fees:
                k += 1
                r_idx = 2 + k
                set_border(ws, r_idx, 4, 56)
                
                ws.merge_cells(f'I{r_idx}:U{r_idx}')
                ws[f'I{r_idx}'] = f_row.zwpm
                ws[f'I{r_idx}'].alignment = Alignment(horizontal='left')
                
                ws[f'V{r_idx}'] = float(f_row.mjzj or 0) if rmbkh else float(f_row.wxzj or 0)
                ws[f'V{r_idx}'].number_format = f'￥#,##{dd}' if rmbkh else f'"{hbdmfh}"#,##{dd}'
                
                ws[f'W{r_idx}'] = float(f_row.gczj or 0)
                ws[f'W{r_idx}'].number_format = '￥#,##0.00' if ('RMB' in str(f_row.cghbdm or '') or not f_row.cghbdm) else '$#,##0.00'

            # ----------------- 4. 全局加项与明佣 -----------------
            if float(master.jxUSD or 0) > 0 or float(master.jxRMB or 0) > 0:
                k += 1
                r_idx = 2 + k
                ws[f'U{r_idx}'] = '合 计'
                if float(master.jxUSD or 0) > 0 and float(master.jxKHRMB or 0) > 0:
                    ws[f'V{r_idx}'] = f"￥{float(master.jxKHRMB):.2f}/{hbdmfh}{float(master.jxUSD):.2f}"
                else:
                    if float(master.jxUSD or 0) > 0:
                        ws[f'V{r_idx}'] = float(master.jxUSD)
                        ws[f'V{r_idx}'].number_format = '$#,##0.00'
                    if float(master.jxKHRMB or 0) > 0:
                        ws[f'V{r_idx}'] = float(master.jxKHRMB)
                        ws[f'V{r_idx}'].number_format = '￥#,##0.00'
                ws[f'W{r_idx}'] = float(master.jxRMB or 0)
                ws[f'AD{r_idx}'].number_format = '@'

            if float(master.myjje or 0) > 0:
                k += 1
                r_idx = 2 + k
                ws[f'U{r_idx}'] = '佣  金'
                ws[f'V{r_idx}'] = f"{'￥' if rmbkh else hbdmfh}{float(master.myjje):.2f}"
                ws[f'V{r_idx}'].number_format = '@'
                ws[f'AD{r_idx}'].number_format = '@'

            # ----------------- 5. 备注及说明 -----------------
            r_idx_6 = 6 + k
            ws[f'D{r_idx_6}'].font = Font(bold=True)
            ws[f'D{r_idx_6}'] = '注意事项:'
            ws.merge_cells(f'G{r_idx_6}:AM{r_idx_6}')
            ws.row_dimensions[r_idx_6].height = 30
            ws[f'BJ{r_idx_6}'] = str(master.zysx or '').strip() + '\n'
            ws[f'G{r_idx_6}'] = str(master.zysx or '').strip() + '\n'

            qtsm1 = str(master.qtshm or '')
            if '暗佣' not in qtsm1.upper() and float(master.ayjje or 0) > 0:
                qtsm1 += f";暗佣金额：{'￥' if rmbkh else hbdmfh}{float(master.ayjje)}"
                    
            r_idx_7 = 7 + k
            ws[f'D{r_idx_7}'].font = Font(bold=True)
            ws[f'D{r_idx_7}'] = '其它说明:'
            ws.merge_cells(f'G{r_idx_7}:AM{r_idx_7}')
            ws[f'BJ{r_idx_7}'] = qtsm1.strip()
            ws[f'G{r_idx_7}'] = qtsm1.strip() + '\n'

            r_idx_8 = 8 + k
            ws[f'D{r_idx_8}'].font = Font(bold=True)
            ws[f'D{r_idx_8}'] = '出运说明:'
            ws.merge_cells(f'G{r_idx_8}:AM{r_idx_8}')
            ws[f'BJ{r_idx_8}'] = str(master.cyshm or '').strip() + '\n'
            ws[f'G{r_idx_8}'] = str(master.cyshm or '').strip() + '\n'

            r_idx_9 = 9 + k
            ws[f'D{r_idx_9}'].font = Font(bold=True)
            ws[f'D{r_idx_9}'] = '单证备注:'
            ws.merge_cells(f'G{r_idx_9}:AM{r_idx_9}')
            ws[f'BJ{r_idx_9}'] = str(master.dzbz or '').strip()
            ws[f'G{r_idx_9}'] = str(master.dzbz or '').strip() + '\n'

            # ----------------- 6. 报关补报信息 (bgmxdsheet -> pid) -----------------
            ws[f'D{10 + k}'].font = Font(bold=True)
            ws[f'D{10 + k}'] = '补报信息:'
            
            if bgmxd_info:
                # 【修改点 4】：使用 pid 替换 father
                supp_items = s.execute(text("SELECT * FROM bgmxdsheet WHERE pid=:pid AND yfph!='' AND yfph IS NOT NULL ORDER BY zhwbgpm, gcmc"), {"pid": bgmxd_info.rid}).fetchall()
                if supp_items:
                    k += 1 
                    for bg in supp_items:
                        k += 1
                        br_idx = 9 + k
                        set_border(ws, br_idx, 4, 58)
                        
                        b_rmbkh = bg_rmb_map.get(bg.yfph, False)
                        
                        ws[f'A{br_idx}'] = float(bg.bzcd or 0)
                        ws[f'B{br_idx}'] = float(bg.bzkd or 0)
                        ws[f'C{br_idx}'] = float(bg.bzgd or 0)
                        ws[f'D{br_idx}'] = bg.yfph
                        ws[f'E{br_idx}'] = bg.khht
                        ws[f'F{br_idx}'].number_format = '@'
                        ws[f'G{br_idx}'].number_format = '@'
                        if bg.czkrhh:
                            ws[f'F{br_idx}'] = bg.czkrhh
                            ws[f'G{br_idx}'] = bg.czkrhh
                        else:
                            ws[f'F{br_idx}'] = bg.khhh
                            ws[f'G{br_idx}'] = bg.cpbh
                            
                        ws[f'H{br_idx}'] = str(bg.djpmy or '').upper()
                        ws[f'AN{br_idx}'] = str(bg.djpmy or '').upper()
                        ws[f'AO{br_idx}'] = str(bg.djpmw or '').upper()
                        
                        ws[f'I{br_idx}'] = bg.zwpm
                        ws[f'I{br_idx}'].alignment = Alignment(horizontal='left')
                        
                        if b_rmbkh:
                            ws[f'J{br_idx}'] = float(bg.mjdj1 or 0)
                            ws[f'V{br_idx}'] = float(bg.mjzj or 0)
                            ws[f'J{br_idx}'].number_format = ws[f'V{br_idx}'].number_format = f'￥#,##{dd}'
                        else:
                            if jysb == '1':
                                ds = s.execute(text("SELECT dswxjg, dswxzj FROM cymxsheet WHERE cywyzd=:cywyzd AND fpsb1='是'"), {"cywyzd": bg.cywyzd}).fetchone()
                                ws[f'J{br_idx}'] = float(ds.dswxjg or 0) if ds else 0
                                ws[f'V{br_idx}'] = float(ds.dswxzj or 0) if ds else 0
                            else:
                                ws[f'J{br_idx}'] = float(bg.wxjg or 0)
                                ws[f'V{br_idx}'] = float(bg.wxzj or 0)
                            ws[f'J{br_idx}'].number_format = ws[f'V{br_idx}'].number_format = f'"{hbdmfh}"#,##{dd}'
                            
                        ws[f'K{br_idx}'] = f"{float(bg.gczj or 0)/float(bg.chsl):.2f}" if float(bg.chsl or 0) > 0 else float(bg.gcjg or 0)
                        
                        ws[f'L{br_idx}'] = bg.jldw
                        ws[f'M{br_idx}'] = float(bg.wxrl or 0)
                        ws[f'N{br_idx}'] = int(bg.chxs or 0)
                        ws[f'O{br_idx}'] = float(bg.wxmz or 0)
                        ws[f'P{br_idx}'] = float(bg.wxjz or 0)
                        ws[f'Q{br_idx}'] = float(bg.wxtj or 0)
                        ws[f'R{br_idx}'] = float(bg.zmz or 0)
                        ws[f'S{br_idx}'] = float(bg.zjz or 0)
                        ws[f'T{br_idx}'] = float(bg.chsl or 0)
                        ws[f'U{br_idx}'] = float(bg.ztj or 0)
                        ws[f'W{br_idx}'] = float(bg.gczj or 0)
                        ws[f'X{br_idx}'] = zycs_map.get(bg.gcmc, bg.gcmc)

            # ----------------- 7. 待补报信息 -----------------
            if pending_customs:
                k += 1
                ws[f'D{10 + k}'].font = Font(bold=True)
                ws[f'D{10 + k}'] = '待补报信息:'
                for pc in pending_customs:
                    k += 1
                    p_idx = 10 + k
                    ws.merge_cells(f'H{p_idx}:T{p_idx}')
                    ws[f'H{p_idx}'].alignment = Alignment(horizontal='left')
                    ws[f'H{p_idx}'] = f"货号：{pc.cpbh or ''} 品名:{pc.zhwbgpm or ''} 工厂:{pc.gcmc or ''} 等待补报"

            # ================== 平行排版区 (装柜 / 工厂 / 退税) ==================
            base_row = 11 + k
            
            # (A) 左侧: 装柜与额外费用 (D列)
            r_box = base_row
            ws[f'D{r_box}'] = str(master.zgdd or '') + '仓库装柜'; r_box+=1
            ws[f'D{r_box}'] = '监装人：' + str(master.jzry or ''); r_box+=1
            ws[f'D{r_box}'] = '日期：' + str(master.zgrq or ''); r_box+=1
            ws[f'D{r_box}'] = '空柜：' + str(master.hgzl or ''); r_box+=1
            ws[f'D{r_box}'] = '车子：' + str(master.czzl or ''); r_box+=1
            ws[f'D{r_box}'] = '车子+空柜：' + str(master.cjkg or ''); r_box+=1
            ws[f'D{r_box}'] = '车子+重柜：' + str(master.cjgz or ''); r_box+=1

            if float(master.fyzj1 or 0) > 0: ws[f'D{r_box}'] = f"业务额外费用￥{master.fyzj1}"; r_box+=1
            if sum(float(getattr(master, f, 0)) for f in ['hdsyfM', 'qggsM', 'qgzs$', 'mdgfy']) > 0:
                ws[f'D{r_box}'] = f"电商海运费${master.hdsyfM}货代{master.Mhd or ''}"; r_box+=1
            if float(master.hdsyfR or 0) > 0: ws[f'D{r_box}'] = f"电商海运费￥{master.hdsyfR}货代{master.Mhd or ''}"; r_box+=1
            if float(master.gnnlf or 0) > 0: ws[f'D{r_box}'] = f"国内内陆费￥{master.gnnlf}货代{master.Rhd or ''}{master.fyzc or ''}"; r_box+=1
            if float(master.mdgfy or 0) > 0: ws[f'D{r_box}'] = f"目的港费用${master.mdgfy}货代{master.Ehd or ''}"; r_box+=1
            if float(master.hybfM or 0) > 0: ws[f'D{r_box}'] = f"海运保费${master.hybfM}"; r_box+=1
            if float(master.dshl or 0) > 0 and float(master.dshl) != 1: ws[f'D{r_box}'] = f"电商汇率{master.dshl}"; r_box+=1

            # (B) 中间: 工厂汇总表 (H-U列)
            r_fac = base_row
            ws.merge_cells(f'H{r_fac}:I{r_fac}'); ws[f'H{r_fac}'] = '工厂'; ws[f'H{r_fac}'].font = Font(bold=True)
            ws.merge_cells(f'J{r_fac}:T{r_fac}'); ws[f'J{r_fac}'] = '付款方式'; ws[f'J{r_fac}'].font = Font(bold=True)
            ws[f'U{r_fac}'] = '进仓日期'; ws[f'U{r_fac}'].font = Font(bold=True)
            
            fac_info_map = {}
            for item in all_details:
                if item.gcmc: fac_info_map[zycs_map.get(item.gcmc, item.gcmc)] = (item.jsfs, item.scrq)
                
            r_fac += 1
            # for f_name, (f_jsfs, f_scrq) in fac_info_map.items():
            #     ws.merge_cells(f'H{r_fac}:I{r_fac}'); ws[f'H{r_fac}'] = f_name
            #     ws.merge_cells(f'J{r_fac}:T{r_fac}'); ws[f'J{r_fac}'] = f_jsfs
            #     ws[f'U{r_fac}'].number_format = '@'
            #     ws[f'U{r_fac}'] = f_scrq
            #     r_fac += 1
            for f_name, (f_jsfs, f_scrq) in fac_info_map.items():
                ws.merge_cells(f'H{r_fac}:I{r_fac}')
                ws[f'H{r_fac}'] = f_name
                
                ws.merge_cells(f'J{r_fac}:T{r_fac}')
                ws[f'J{r_fac}'] = f_jsfs
                # 【新增】：强制开启自动换行，并让文字垂直居中，绝不溢出到右边！
                ws[f'J{r_fac}'].alignment = Alignment(wrap_text=True, vertical='center')
                # 稍微把行高拉大一点，容纳多行文字
                ws.row_dimensions[r_fac].height = 30 
                
                ws[f'U{r_fac}'].number_format = '@'
                ws[f'U{r_fac}'] = f_scrq
                r_fac += 1
            # (C) 右侧: 税率瀑布流重构 (W-X列)
            r_tax = base_row
            ws[f'W{r_tax}'] = '退税'; ws[f'X{r_tax}'] = '金额'
            set_border(ws, r_tax, 23, 24)
            r_tax += 1
            
            for rate, amt in [
                ('17%', master.tshj17), ('16%', master.tshj16), ('15%', master.tshj15),
                ('13%', master.tshj13), ('10%', master.tshj10), ('9%', master.tshj9),
                ('6%', master.tshj6), ('5%', master.tshj5), ('3%', master.tshj3),
                ('1%', master.tshj1), ('0%', master.xjhj)
            ]:
                if amt and float(amt) > 0:
                    ws[f'W{r_tax}'] = rate
                    ws[f'X{r_tax}'] = float(amt)
                    set_border(ws, r_tax, 23, 24)
                    r_tax += 1

            # ================= 平行区结束，对齐最大行号 =================
            max_row = max(r_box, r_fac, r_tax)
            
            # ----------------- 8. cymxsheet6 额外费用 (改用 pid) -----------------
            # 【修改点 5】：将 father 改为 pid
            sheet6 = s.execute(text("SELECT * FROM cymxsheet6 WHERE pid=:id"), {"id": record_id}).fetchall()
            for fee in sheet6:
                ws[f'D{max_row}'] = f"费用名称:{fee.fymc or ''} 费用金额:{fee.hbdm or ''}{fee.fyje or ''} 厂商名称:{fee.csmc or ''}"
                max_row += 1
                
            # ----------------- 9. cymxsheet10 预填详情 (改用 pid) -----------------
            if dept == '1':
                # 【修改点 6】：将 father 改为 pid
                sheet10 = s.execute(text("SELECT * FROM cymxsheet10 WHERE pid=:id"), {"id": record_id}).fetchall()
            else:
                sheet10 = s.execute(text("SELECT * FROM cymxsheet10 WHERE pid=:id AND wxdbm1=:dept"), {"id": record_id, "dept": dept}).fetchall()
                
            if sheet10:
                ws[f'D{max_row}'] = '预填详情'
                headers = ['付款抬头', '货源地', '海关计量单位', '中文品名', '增值税率', '退税率', '数量', '金额', '', '', '', '工厂名称']
                for col_offset, h_name in enumerate(headers):
                    if h_name: ws.cell(row=max_row, column=7+col_offset, value=h_name)
                set_border(ws, max_row, 7, 18)
                max_row += 1
                
                for pr in sheet10:
                    ws[f'G{max_row}'] = pr.fktt
                    ws[f'H{max_row}'] = pr.hyd
                    ws[f'I{max_row}'] = pr.hgjldw
                    ws[f'J{max_row}'] = pr.zwpm
                    ws[f'K{max_row}'].number_format = '@'; ws[f'K{max_row}'] = pr.zzsl
                    ws[f'L{max_row}'] = pr.zsl
                    ws[f'M{max_row}'] = pr.chsl
                    ws[f'N{max_row}'] = pr.ytje
                    ws[f'R{max_row}'] = pr.sccj
                    set_border(ws, max_row, 7, 18)
                    max_row += 1

            # ----------------- 10. 签名栏 -----------------
            sign_row = max_row + 4 if len(fac_info_map) > 8 else max_row + 15
            ws.merge_cells(f'D{sign_row}:AM{sign_row}')
            sign_cell = ws[f'D{sign_row}']
            sign_cell.font = Font(bold=True, size=14)
            sign_cell.value = f"单证签字: {master.zdry or ''}                         经理签字:                                     总经理签字:                                                      改单签字:"

        # ================= 保存导出 =================
        excel_file_name = f"{master.fphm or 'Unknown'}_cymx.xlsx"
        save_dir = config.tmp_path
        if not os.path.exists(save_dir):
            os.makedirs(save_dir, exist_ok=True)
            
        excel_save_path = os.path.join(save_dir, excel_file_name)
        wb.save(excel_save_path)
        wb.close()

        return json_result(1, '导出成功', excel_file_name)

    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f"生成财务出运清单失败: {str(e)}")
    finally:
        s.close()




#  出货清单导入（优景）
# --- 自定义函数辅助方法 ---
def safe_float(val):
    try:
        return float(val) if val else 0.0
    except Exception:
        return 0.0

@any_route('/api/saier/shipment/shipping_list/import', methods=['POST'])
@require_token
async def api_import_shipping_list_import(request):
    try:
        # 接收前端参数
        j = await request.form()
        user = request.current_user
        uid = user.rid
        file = j.get('file',None)
        mode = j.get('mode','1')
        rid = j.get('rid')
        if not rid:
            return json_result(-1, "未获取到主单ID")
        if is_none(file):
            return json_result(-1, "未获取到文件")
            
        # 1. 保存临时文件
        temp_dir = config.tmp_path
        temp_file = os.path.join(temp_dir, f"{uuid.uuid4()}_{file.filename}")
        data = await file.read()
        write_file(temp_file, data, 'wb', encoding=None)
        
        s = Session()
        wb = None
        errors_list = []
        
        try:
            wb = load_workbook(temp_file, data_only=True) 
            ws = wb.worksheets[0]
            
            # ==========================================
            # Block 1: 解析表头与风控核查 (含亚马逊特权)
            # ==========================================
            khmc = str(ws.cell(row=1, column=22).value or '').strip() # V1
            fphm = str(ws.cell(row=1, column=7).value or '').strip() # G1
            
            if not khmc:
                return json_result(-1, "请在V1输入客人名称!")
                
            is_amazon = 'AMAZON' in khmc.upper()
            
            # 严格还原的“白名单”风控探针
            kh_sql = """
                SELECT company_name, fkgl 
                FROM kh 
                WHERE company_name = :cname 
                AND (
                    fkgl <> '是' 
                    OR fkgl IS NULL 
                    OR company_name LIKE :amazon_like
                )
            """
            kh_row = s.execute(text(kh_sql), {
                'cname': khmc,
                'amazon_like': '%AMAZON%'
            }).fetchone()
            
            if not kh_row:
                return json_result(-1, '此客人为风控客人或无此客人，不能导入！')

            # 组装表头返回给前端
            header_data = {
                "khmc": khmc,
                "ywry": str(ws.cell(row=1, column=25).value or ''), # Y1
                "fphm": fphm,
                "xh": str(ws.cell(row=1, column=11).value or ''), # K1
                "fh": str(ws.cell(row=1, column=19).value or ''), # S1
                "qxnr": str(ws.cell(row=1, column=26).value or '') # Z1
            }
            
            # ==========================================
            # Block 2: 逐行扫表与数据装载
            # ==========================================
            items = []
            row_idx = 3 # 从第3行开始 (对应 Delphi 的 A3)
            
            ytxxz_list = []      # 用于聚合发票备注
            missing_contracts = [] # 用于致命缺失熔断
            
            while True:
                # 探针字段：长、宽、高、体积
                bzcd = safe_float(ws.cell(row=row_idx, column=1).value)
                bzkd = safe_float(ws.cell(row=row_idx, column=2).value)
                bzgd = safe_float(ws.cell(row=row_idx, column=3).value)
                tj = safe_float(ws.cell(row=row_idx, column=17).value)
                
                # 结束条件：探针全空跳出循环
                if bzcd == 0 and bzkd == 0 and bzgd == 0 and tj == 0:
                    break
                    
                # 提取明细字段    对应到582----683
                wxht = str(ws.cell(row=row_idx, column=4).value or '').strip() # D
                khhh = str(ws.cell(row=row_idx, column=6).value or '').strip() # F
                cpbh = str(ws.cell(row=row_idx, column=7).value or '').strip() # G
                zwpm = str(ws.cell(row=row_idx, column=9).value or '').strip() # I

                wxrl = safe_float(ws.cell(row=row_idx, column=13).value) # M
                xs = safe_float(ws.cell(row=row_idx, column=14).value)   # N
                mz = str(ws.cell(row=row_idx, column=15).value or '')    # O
                jz = str(ws.cell(row=row_idx, column=16).value or '')    # P
                chsl = safe_float(ws.cell(row=row_idx, column=20).value) # T
                bgpm = str(ws.cell(row=row_idx, column=25).value or '')  # Y
                zzsl = safe_float(ws.cell(row=row_idx, column=26).value) # Z
                
                wyzd = str(ws.cell(row=row_idx, column=30).value or '').strip() # AD
                ywry = str(ws.cell(row=row_idx, column=31).value or '')  # AE
                cgry = str(ws.cell(row=row_idx, column=32).value or '')  # AF
                gdry = str(ws.cell(row=row_idx, column=33).value or '')  # AG
                dq1 = str(ws.cell(row=row_idx, column=35).value or '')   # AI
                bz1 = str(ws.cell(row=row_idx, column=37).value or '')   # AK
                sfpx = str(ws.cell(row=row_idx, column=38).value or '否').strip() # AL
                fktt = str(ws.cell(row=row_idx, column=41).value or '')  # AO
                cght = str(ws.cell(row=row_idx, column=50).value or '').strip()   # AX

                # 数据容错：税率强转 17->16，出货量计算
                if zzsl == 17: zzsl = 16
                zsl1 = chsl if (sfpx == '是' or sfpx != '否') else (wxrl * xs if wxrl > 0 and xs > 0 else chsl)
                
                cpbh1 = khhh if khhh else cpbh
                
                # --- 缺合同熔断收集 ---  
                #  计数器
                if not cght or not cpbh1:  
                    missing_contracts.append(cpbh1)
                    row_idx += 1
                    continue
                    
                # ------------------------------------------
                # Block 2.1: 三步连环查 (采购ID、外销唯一等溯源字段)
                # ------------------------------------------
                cghtwyzd, ytsb, wxwyzd, xjht = '', '', '', ''
                cgsl, cgxs = 0.0, 0
                ytxx = ''

                # [第一查]：查明细
                sql_detail = """
                    SELECT wyzd, cghtwyzd, ytsb, ytxx, wxwyzd, xjht, wxht 
                    FROM cghtsheet 
                    WHERE (bjhh=:bjhh OR khhh=:bjhh2) AND hthm=:hthm
                """
                params_detail = {'bjhh': cpbh, 'bjhh2': cpbh1, 'hthm': cght}
                
                if wyzd:
                    sql_detail += " AND WYZD=:wyzd"
                    params_detail['wyzd'] = wyzd
                else:
                    sql_detail += " AND wxrl=:wxrl AND xjht<>'作废'"
                    params_detail['wxrl'] = wxrl

                row_detail = s.execute(text(sql_detail), params_detail).fetchone()
                
                if row_detail:
                    wyzd = row_detail.wyzd or ''
                    cghtwyzd = row_detail.cghtwyzd or ''
                    ytsb = row_detail.ytsb or ''
                    xjht = row_detail.xjht or ''
                    ytxx = row_detail.ytxx or ''
                    
                    if ytxx: ytxxz_list.append(ytxx)

                    # [第二查]：外销唯一容错纠正
                    if '赠送' in wxht:
                        wxwyzd = ''
                    else:
                        db_wxht = row_detail.wxht or ''
                        if wxht != '' and wxht != db_wxht:
                            sql_wx = "SELECT wxwyzd FROM wxhtsheet WHERE order_id=:order_id AND bjhh=:bjhh AND khmc=:khmc"
                            row_wx = s.execute(text(sql_wx), {'order_id': wxht, 'bjhh': cpbh, 'khmc': khmc}).fetchone()
                            if row_wx:
                                wxwyzd = row_wx.wxwyzd or ''
                        else:
                            wxwyzd = row_detail.wxwyzd or ''
                else:
                    wyzd = ''
                    
                # [第三查]：查采购汇总
                sql_sum = """
                    SELECT SUM(cgsl) as cgsl1, SUM(cgxs) as cgxs1 
                    FROM cghtsheet 
                    WHERE (bjhh=:bjhh OR khhh=:bjhh2) AND hthm=:hthm
                """
                params_sum = {'bjhh': cpbh, 'bjhh2': cpbh1, 'hthm': cght}
                if wyzd:
                    sql_sum += " AND WYZD=:wyzd"
                    params_sum['wyzd'] = wyzd
                else:
                    sql_sum += " AND wxrl=:wxrl"
                    params_sum['wxrl'] = wxrl
                    
                row_sum = s.execute(text(sql_sum), params_sum).fetchone()
                if row_sum and row_sum.cgsl1 is not None:
                    cgsl = float(row_sum.cgsl1)
                    cgxs = int(row_sum.cgxs1)

                # ------------------------------------------
                # Block 2.2: 查出运历史防超发
                # ------------------------------------------
                cyhj1 = 0.0
                chhj2 = 0.0

                if '赠送' in cght:
                    pass # 赠送豁免查历史  chun2b
                else:
                    his_sql = """
                        SELECT SUM(chsl) as cyhj1 
                        FROM cymxsheet
                        WHERE zycpbh=:cpbh 
                          AND cght=:cght 
                          AND fpsb1<>'是' 
                          AND ywchy<>'是' 
                          AND (chyrq IS NULL )
                    """
                    his_params = {'cpbh': cpbh, 'cght': cght}
                    
                    if wyzd == '':
                        his_sql += " AND wxrl=:wxrl"
                        his_params['wxrl'] = wxrl
                    else:
                        his_sql += " AND pgwy=:pgwy"
                        his_params['pgwy'] = wyzd
                        
                    his_res = s.execute(text(his_sql), his_params).fetchone()
                    if his_res and his_res.cyhj1 is not None:
                        cyhj1 = float(his_res.cyhj1)
                        
                    chhj2 = cyhj1 + zsl1
                
                # --- 最终超发与作废熔断 (卫语句提前返回) ---
                if (cgsl < chhj2) or (xjht == '作废'):
                    if cgsl < chhj2:
                        errors_list.append(f"第{row_idx-2}条记录：产品编号：{cpbh1} 采购合同：{cght} 出货数量大于采购数量请核实!")
                    if xjht == '作废':
                        errors_list.append(f"第{row_idx-2}条记录：产品编号：{cpbh1} 采购合同：{cght} 作废!")
                    
                    row_idx += 1
                    continue

                # 以上对应407行代码  
                # ------------------------------------------
                # Block 2.3: 组装子表装载数据 (原 Pascal else 逻辑)      
                # ------------------------------------------
                calc_tj = (bzcd * bzkd * bzgd) / 1000000
                final_tj = calc_tj if is_amazon else round(calc_tj, 3)
                if tj > 0: final_tj = tj if is_amazon else round(tj, 3)

                zzsl_db = int(zzsl) if zzsl > 1 else int(zzsl * 100) if zzsl > 0 else 0
                
                zsong = '否'
                if '赠送' in wxht and '赠送' in cght: zsong = '是'
                elif '赠送' in wxht: zsong = '客人'
                elif '赠送' in cght: zsong = '工厂'

                item = {
                    "rid":get_uuid(),
                    "pid": rid,
                    "uid":uid,
                    "bzcd": f"{bzcd:.1f}", "bzkd": f"{bzkd:.1f}", "bzgd": f"{bzgd:.1f}",
                    "fktt": fktt, "wxtj": final_tj, "wxht": wxht, "khht": '',
                    "khmc": khmc, "khhh": khhh, "cpbh": cpbh1, "zycpbh": cpbh,
                    "yxcpbh": cpbh, "zwpm": zwpm, "bzdw": "CTNS",
                    "wxrl": wxrl, "chxs": int(xs) if xs else 0,
                    "wxmz1": mz, "wxjz1": jz, "wxmz": mz, "wxjz": jz,
                    "cgsl": cgsl, "zzsl": zzsl_db, "zhwbgpm": bgpm,
                    "chsl": zsl1, "ywry": cgry, "gdry": gdry, "sfpx": sfpx,
                    "sysl": cgsl, "cght": cght, "wxwyzd": wxwyzd, "chsl1": 0,
                    "zs": zsong, "rkdd": dq1, "ywrya": ywry, "bz1": bz1, "pgwy": wyzd,
                    "cgxs": cgxs, "cgID": cghtwyzd, "ytsb": ytsb
                }
                
                if is_amazon:
                    item.update({
                        "hjbh": str(ws.cell(row=row_idx, column=43).value or ''),
                        "dpmc": str(ws.cell(row=row_idx, column=44).value or ''),
                        "amsgj": str(ws.cell(row=row_idx, column=45).value or ''),
                        "SellerID": str(ws.cell(row=row_idx, column=46).value or ''),
                        "marketplace_id": str(ws.cell(row=row_idx, column=47).value or ''),
                        "fulfillment_network_sku": str(ws.cell(row=row_idx, column=52).value or ''),
                        "cjhjsj": str(ws.cell(row=row_idx, column=27).value or ''),
                        "hjckh": str(ws.cell(row=row_idx, column=28).value or '')
                    })
                    
                items.append(item)
                row_idx += 1
                
            # ==============================================================
            # Block 3: 循环后逻辑 (底栏提取、致命错误熔断、日志抛出)
            # ==============================================================
            err_filename = f"未导入详情{fphm}.txt"
            
            # --- 致命缺失强熔断 (原 cght1 > 0 逻辑) ---
            if missing_contracts:
                cght2_str = ";".join(missing_contracts)
                fatal_msg = f"产品编号:{cght2_str}没有采购合同号或产品编号请重新导入!"
                errors_list.append(fatal_msg)
                
                err_filepath = os.path.join(temp_dir, err_filename)
                write_file(err_filepath, '\n'.join(errors_list), 'w')
                
                # 拒绝返回 items，保护表单不被残缺数据污染
                return json_result(-1, fatal_msg, data={"duplicate_file_name": err_filename})
            
            # --- 提取底栏杂费 (仅覆盖模式处理) ---
            footer_data = {}
            if mode == '1':
                da1 = str(ws.cell(row=row_idx + 1, column=7).value or '') # G(4+i)
                da12 = str(ws.cell(row=row_idx + 2, column=7).value or '') # G(5+i)
                cysm = ""
                if da1: cysm = f"注意事项:{da1}\n"
                if da12: cysm += f"其它说明:{da12}"
                
                ytxxz_final = ",".join(ytxxz_list) # 发票备注拼接
                #  对应着 688-724
                footer_data = {
                    "zysx": da1, "cysm": cysm,
                    "jhfs": str(ws.cell(row=row_idx + 3, column=7).value or ''),
                    "tydh": str(ws.cell(row=row_idx + 4, column=7).value or ''),
                    "jzry": str(ws.cell(row=row_idx + 5, column=7).value or ''),
                    "czzl": str(ws.cell(row=row_idx + 6, column=7).value or ''),
                    "hgzl": str(ws.cell(row=row_idx + 7, column=7).value or ''),
                    "cjkg": str(ws.cell(row=row_idx + 8, column=7).value or ''),
                    "cjgz": str(ws.cell(row=row_idx + 9, column=7).value or ''),
                    "ytxxz": ytxxz_final 
                }

            # --- 组装并返回结果 ---
            result_data = {
                "header": header_data,
                "items": items,
                "footer": footer_data
            }
            
            if errors_list:
                err_filepath = os.path.join(temp_dir, err_filename)
                write_file(err_filepath, '\n'.join(errors_list), 'w')
                result_data['duplicate_file_name'] = err_filename
                logger.error(errors_list)    # 查看 errordata
            
            logger.error(result_data)
        

            return json_result(0, '解析成功', data=result_data)

        except Exception as e:
            logger.error(trace_error())
            return json_result(-1, f'解析Excel异常: {str(e)}')
        finally:
            if wb: wb.close()
            s.close()
            
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'导入请求异常: {str(e)}')







# 优景优胜 报关单INV PL 


# uae 的获取 

@any_route('/api/saier/shipment/declaration/check_uae', methods=['POST'])
@require_token
async def api_check_declaration_uae(request):
    """用于在前端打开报关单/发票/箱单导出弹窗前，检查是否为 UAE 客户"""
    j = await request.json()
    record_id = j.get('rid')
    if not record_id:
        return json_result(-1, "缺少单据ID")

    s = Session()
    try:
        # 1. 查找主表 cymx，获取客户名称 khmc
        row_cymx = s.execute(text("SELECT khmc FROM cymx WHERE rid=:rid"), {'rid': record_id}).fetchone()
        if not row_cymx or not row_cymx.khmc:
            return json_result(-1, "查询不到主表中的字段khmc")

        # 2. 检查 cyzglsheet 表中是否配置了 UAE
        khmc = str(row_cymx.khmc)
        row_uae = s.execute(
            text("SELECT BZ FROM cyzglsheet WHERE xm=:xm AND zm='UAE'"), 
            {'xm': khmc}
        ).fetchone()

        is_uae = False
        # 依据 Delphi 逻辑，只有 BZ='UAE' 时，UAE变量才置为 '1' 并在前端隐藏部分选项
        if row_uae and str(row_uae.BZ) == 'UAE':
            is_uae = True
            
        return json_result(1, "Success", data=is_uae)

    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f"查询UAE状态失败: {str(e)}")
    finally:
        s.close()






from datetime import datetime as dt


def safe_int(val):
    try: return int(val) if val else 0
    except: return 0

def apply_subscript(text, djyxb_str):
    """完美平替 Delphi 中的 msexcelworksheet.Range.Characters[zcx, 1].Font.Subscript"""
    if not text or not djyxb_str: return text
    indices = set()
    for part in str(djyxb_str).replace('，', ',').split(','):
        part = part.strip()
        if part.isdigit():
            indices.add(int(part) - 1)  
    if not indices: return text
    
    sub_font = InlineFont(vertAlign='subscript')
    elements, curr = [], ""
    for i, char in enumerate(str(text)):
        if i in indices:
            if curr:
                elements.append(curr)
                curr = ""
            elements.append(TextBlock(sub_font, char))
        else:
            curr += char
    if curr: elements.append(curr)
    return CellRichText(*elements)

def copy_row_style(ws, src_row, tgt_row):
    """用于动态在末尾插入行并复制模板行格式 (LV / UAE 专属)"""
    for col in range(1, ws.max_column + 1):
        src = ws.cell(row=src_row, column=col)
        tgt = ws.cell(row=tgt_row, column=col)
        if src.has_style:
            tgt.font, tgt.border, tgt.fill = pycopy(src.font), pycopy(src.border), pycopy(src.fill)
            tgt.number_format, tgt.alignment = src.number_format, pycopy(src.alignment)

@any_route('/api/saier/shipment/declaration/export', methods=['POST'])
@require_token
async def api_export_declaration(request):
    j = await request.json()
    record_id = j.get('rid')
    djws = str(j.get('djws', '2'))
    gs11 = str(j.get('gs11', '1'))
    ed = str(j.get('ed', '123'))
    hbdm = str(j.get('hbdm', '1'))
    hr1 = str(j.get('hr1', '1'))
    da1 = str(j.get('da1', '3'))

    dd1_map = {'1': '0.0', '2': '0.00', '3': '0.000', '4': '0.0000'}
    dd1 = dd1_map.get(djws, '0.00')
    user= request.current_user
    user_info=get_user_path(user.username)
    user_path=user_info.get('path', '')
    user_pos=user_info.get('position', '')
    is_valid_role = False
    if ('外销' in user_path) or ('单证' in user_position) or ('外销' in user_position):
        is_valid_role = True
    else:
        return json_result(-1, "您没有权限！")
    
    s = Session()
    try:
        row_cymx = s.execute(text("SELECT * FROM cymx WHERE rid=:rid"), {'rid': record_id}).fetchone()
        if not row_cymx: return json_result(-1, "未找到出运明细")

        hbdm1 = str(row_cymx.hbdm or '')
        if hbdm1 in ('USD', 'USD$'): hbdm1 = 'USD'
        
        row_hbdm = s.execute(text("SELECT bjdh FROM hbdm WHERE bjdid=:bjdid"), {'bjdid': hbdm1}).fetchone()
        hbdmfh = str(row_hbdm.bjdh) if row_hbdm else ''

        vcsb = '1' if '景驰' in str(row_cymx.wfgs or '') else ''
        
        ujsb = '1' if s.execute(text("SELECT TOP 1 xm FROM cyzglsheet WHERE zm='优景报关模板' AND xm=:xm"), {'xm': str(row_cymx.kh_id)}).fetchone() else ''

        lvpd, uae, bank = '', '', ''
        row_uae = s.execute(text("SELECT * FROM cyzglsheet WHERE xm=:xm AND zm='UAE'"), {'xm': str(row_cymx.khmc)}).fetchone()
        if row_uae:
            if str(row_uae.BZ) == 'LV': lvpd = '1'
            if str(row_uae.BZ) == 'UAE': uae = '1'
            bank = str(row_uae.BZ1 or '')

        if not str(row_cymx.eta) or not str(row_cymx.chyrq) or str(row_cymx.hwzt) != '已出运':
            return json_result(-1, '请注意预计到港或出运日期末填或货物状态非已出运')
        
        #  以上对应 前137 行 

        if uae == '1': da1 = '1'
        if not ed: ed = '123'
        if gs11 != '2': gs11 = '1'
        
        gs = str(row_cymx.wfgs) if gs11 == '2' else ('宁波优景进出口有限公司' if uae == '1' else '宁波优胜国际贸易有限公司')
        if ujsb == '1': gs = '宁波优景进出口有限公司'
        if vcsb == '1': gs = '宁波景驰进出口有限公司'
        # 对应前197

        row_wfgs = s.execute(text("SELECT TOP 1 ywmc FROM wfgs WHERE wfgs=:wfgs"), {'wfgs': gs}).fetchone()
        gsy = str(row_wfgs.ywmc).upper() if row_wfgs else ''
        
        row_tpzx = s.execute(text("SELECT sb,dyyy,bgh FROM tpzx WHERE cpbh=:cpbh"), {'cpbh': f"{gs}蓝章"}).fetchone()
        gs1 = f"{gsy}\n{str(row_tpzx.dyyy)}\n{str(row_tpzx.bgh)}" if row_tpzx else ''

        sbdate = str(row_cymx.sbdate or '')
        fprq = str(row_cymx.fprq or '')
        date_str = sbdate if sbdate else fprq
        qsn, qsy1, qsr = date_str[:4], date_str[5:7], date_str[8:10]
        
        bp = '1' if 'BEST PRICE' in str(row_cymx.khmc or '').upper() or vcsb == '1' else ''
        AA, B, C, DD, E, F, G, H, II, JJ, KK, L, M, N, O = ('B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P') if bp == '1' else ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O')
        
        #以上对应前277 

        month_map = {'01':'JAN','02':'FEB','03':'MAR','04':'APR','05':'MAY','06':'JUN','07':'JUL','08':'AUG','09':'SEP','10':'OCT','11':'NOV','12':'DEC'}
        qsy = month_map.get(qsy1, '')
        # 以上对应前327
        
        if hbdm in ('', '1'): hbdm = 'M'
        elif hbdm == '2': hbdm = 'R'
        if da1 not in ('1', '2'): da1 = '3'
        if not hr1: hr1 = '1'

        items = s.execute(text("SELECT * FROM cymxsheet4 WHERE pid=:rid"), {'rid': record_id}).fetchall()
        valid_items = [row for row in items if safe_int(row.xs) > 0 or str(row.sfpx) == '是']
        
        top_border = Border(top=Side(style='thin'))
        tpl_dir = os.path.join(config.data_upload_path, 'template')
        generated_files = []

        # ================= [ 第一大分支: 优景/景驰 ] =================
        if hr1 == '1':
            mzhj = sum([round(safe_float(row.zmz) * 1000) / 1000 for row in items])
            jzhj = sum([round(safe_float(row.zjz) * 1000) / 1000 for row in items])
            tjhj = sum([round(safe_float(row.ztj) * 1000) / 1000 for row in items])

            # 1. 报关单
            if '1' in ed:
                if da1 == '3':
                    tpl = 'ujsbbgdn.xlsx' if ujsb == '1' else ('vcsbbgdn.xlsx' if vcsb == '1' else 'bgdn.xlsx')
                elif da1 == '1':
                    tpl = 'uvbgdrmb.xlsx' if hbdm == 'R' else ('uaebgd.xlsx' if uae == '1' else 'uvbgd.xlsx')
                else:
                    tpl = 'hbbgd.xlsx'

                if not os.path.exists(os.path.join(tpl_dir, tpl)): return json_result(-1, f"找不到模板: {tpl}")
                wb = load_workbook(os.path.join(tpl_dir, tpl))  # 加载模板 
                
                i1 = len(valid_items)
                divider = 6 if da1 == '3' else (12 if da1 == '1' else 15)
                i = int(i1 / divider)
                if i == i1 / float(divider): i -= 1
                for a in range(i): wb.copy_worksheet(wb.worksheets[0])
                
                k, i5, number = 0, 0, 0
                for row in valid_items:
                    k += 1
                    number += 1
                    if k == 1:
                        ws = wb.worksheets[i5]

                        if da1 == '3':
                            if vcsb == '1': ws['A6'] = str(row_cymx.khmc)
                            bgdh = str(row_cymx.bgdh or '')
                            if not bgdh:
                                row_md = s.execute(text("SELECT bgdh FROM cymx WHERE fphm=:fphm"), {'fphm': str(row_cymx.mdka) + 'A'}).fetchone()
                                if row_md: bgdh = str(row_md.bgdh or '')
                            ws['A2'] = f"预入录编号：{bgdh}                     海关编号：{bgdh}            (海曙海关)                                                                   页码/页数:{i5+1}/{i+1}"
                            ws['L4'] = f"{qsn}{qsy1}{qsr}"
                            ws['H6'] = f"{str(row_cymx.cmin).upper()}/{str(row_cymx.hangci).upper()}"
                            ws['L6'], ws['A10'], ws['J12'] = str(row_cymx.tdh), str(row_cymx.htjx), str(row_cymx.jgtk)
                            
                            row_mdka = s.execute(text("SELECT gkzw,ssgjz,gkbm FROM mdka WHERE mdka=:mdka"), {'mdka': str(row_cymx.mdka)}).fetchone()
                            if row_mdka:
                                ws['L10'] = f"{row_mdka.gkzw}({row_mdka.ssgjz})"
                                if row_mdka.gkbm: ws['M9'] = f"({row_mdka.gkbm})"
                            else:
                                ws['L10'] = str(row_cymx.mdka)
                                
                            ws['E12'], ws['F12'], ws['H12'] = str(row_cymx.xshj), mzhj, jzhj
                            ws['A17'] = f"集装箱标箱数及号码：{str(row_cymx.xh)}"
                        elif da1 == '1':
                            ws['A5'] = str(row_cymx.cyka).upper()
                            if uae == '1': ws['A7'] = '宁波优景进出口有限公司'
                            if vcsb == '1': ws['A7'] = '宁波景驰进出口有限公司'
                            ws['G5'], ws['I5'], ws['E7'], ws['D13'] = str(row_cymx.chyrq), str(row_cymx.fprq), str(row_cymx.ysfs), str(row_cymx.jgtk)
                            ws['Y7'] = ws['F7'] = f"{str(row_cymx.cmin).upper()}/{str(row_cymx.hangci).upper()}"
                            ws['H7'], ws['E9'], ws['I9'] = str(row_cymx.tdh), str(row_cymx.myfds), 'T/T'
                            ws['H11'], ws['D11'], ws['F11'], ws['A15'] = str(row_cymx.cyka).upper(), str(row_cymx.mygb), str(row_cymx.mdka), str(row_cymx.htjx)
                            ws['D15'], ws['G15'], ws['I15'], ws['A17'] = str(row_cymx.xshj), mzhj, jzhj, f"{str(row_cymx.xh)}/{str(row_cymx.fh)}"
                        else:
                            ws['A4'], ws['F4'], ws['B4'] = gs1.upper(), f"{qsy}.{qsr},{qsn}".upper(), str(row_cymx.cyka).upper()
                            ws['A8'], ws['E12'], ws['A6'], ws['B6'] = gs1.upper(), str(row_cymx.jgtk), str(row_cymx.khmc).upper(), str(row_cymx.ysfs).upper()
                            ws['D6'], ws['F6'], ws['A10'] = f"{str(row_cymx.cmin).upper()}/{str(row_cymx.hangci).upper()}", str(row_cymx.tdh).upper(), str(row_cymx.htjx).upper()
                            ws['B10'] = ws['D10'] = str(row_cymx.mygb).upper()
                            ws['F10'], ws['I10'], ws['B12'] = str(row_cymx.mdka).upper(), str(row_cymx.cyka).upper(), str(row_cymx.xshj2)
                            ws['C12'], ws['D12'], ws['A12'] = f"{safe_float(row_cymx.mzhj):.2f}", f"{safe_float(row_cymx.jzhj):.2f}", str(row_cymx.bzzl).upper()
                            ws['A15'] = f"标记唛码及备注                VGM：{safe_float(row_cymx.hgzl) + safe_float(row_cymx.mzhj):.2f}KGS"
                            if str(row_cymx.wxmt):
                                ws['N16'] = ws['A16'] = f"{str(row_cymx.wxmt).upper()}   {str(row_cymx.xh)}/{str(row_cymx.fh)}"
                            else:
                                ws['A16'] = f"N/M   {str(row_cymx.xh)}/{str(row_cymx.fh)}"

                    # 根据 da1 分支写行级明细
                    if da1 == '3':
                        r_idx = 18 + k
                        ws[f'A{r_idx}'] = str(number)
                        ws[f'B{r_idx}'].number_format, ws[f'B{r_idx}'] = '@', str(row.krcode)
                        ws[f'C{r_idx}'], ws[f'G{r_idx}'] = str(row.zwbgpm), f"{row.zjz}千克"
                        ws[f'I{r_idx}'] = str(row.khrmb) if str(row_cymx.RMBkh) == '是' else str(row.wxdj)
                        ws[f'K{r_idx}'], ws[f'N{r_idx}'], ws[f'Q{r_idx}'], ws[f'Q{r_idx+1}'], ws[f'S{r_idx}'] = '中国', '俄罗斯联邦', '浙江其他', '(33909)', '照章征税'
                        
                        if str(row.hyd):
                            row_hyd = s.execute(text("SELECT * FROM hyd WHERE hyd=:hyd AND hydbm<>''"), {'hyd': row.hyd}).fetchone()
                            if row_hyd: ws[f'Q{r_idx}'], ws[f'Q{r_idx+1}'] = str(row_hyd.hyd), f"({row_hyd.hydbm})"
                        
                        k += 1 # Delphi K := K + 1
                        r_idx = 18 + k
                        row_hg = s.execute(text("SELECT cznr FROM hgbmbsheet WHERE hgbm=:hgbm AND hwmc=:hwmc AND cznr<>'无'"), {'hgbm': row.hgbm2, 'hwmc': row.zwbgpm}).fetchone()
                        if row_hg: ws[f'C{r_idx}'] = str(row_hg.cznr)
                        row_jl = s.execute(text("SELECT zwmc FROM jldw WHERE dwdm=:dw OR ywfs=:dw"), {'dw': row.jldw}).fetchone()
                        ws[f'G{r_idx}'] = f"{row.chsl}{row_jl.zwmc}" if row_jl else str(row.chsl)
                        ws[f'I{r_idx}'] = str(row.kfrmbz) if str(row_cymx.RMBkh) == '是' else str(row.wxzj)
                        ws[f'K{r_idx}'], ws[f'N{r_idx}'], ws[f'S{r_idx}'] = '(CHN)', '(RUS)', '(1)'
                        
                        k += 1
                        r_idx = 18 + k
                        ws[f'I{r_idx}'] = '人民币' if str(row_cymx.RMBkh) == '是' else '美金'

                        if k == 18:
                            k = 0
                            i5 += 1

                    elif da1 == '1':
                        r_idx = 22 + k
                        ws[f'A{r_idx}'].number_format, ws[f'A{r_idx}'] = '@', str(row.krcode)
                        row_tspmb = s.execute(text("SELECT djpmy,djyxb FROM tspmb WHERE cpbh=:cpbh AND djpmy<>''"), {'cpbh': row.cpbh}).fetchone()
                        rich_str = apply_subscript(str(row_tspmb.djpmy).upper(), row_tspmb.djyxb) if row_tspmb else str(row.djpmy).upper()
                        ws[f'Z{r_idx}'], ws[f'B{r_idx}'] = rich_str, rich_str
                        ws[f'F{r_idx}'], ws[f'G{r_idx}'], ws[f'H{r_idx}'] = str(row.zwbgpm), safe_float(row.chsl), str(row.jldw).upper() + 'S'
                        ws[f'J{r_idx}'] = safe_float(row.kfrmbz) if hbdm == 'R' else safe_float(row.wxzj)

                        if k == 12 or number == i1:
                            if number == i1:
                                ws[f'G{23+k}'].border = ws[f'H{23+k}'].border = ws[f'I{23+k}'].border = ws[f'J{23+k}'].border = top_border
                                ws[f'G{23+k}'] = 'TOTAL:'
                                ws[f'J{23+k}'] = safe_float(row_cymx.mjzj) if hbdm == 'R' else safe_float(row_cymx.wxje)
                            if k == 12: k, i5 = 0, i5 + 1
                    else:
                        ik = k # DA1=2 (横版) specific logic
                        r_idx = 18 + ik
                        ws.insert_rows(r_idx)
                        copy_row_style(ws, 18, r_idx)
                        ws[f'A{r_idx}'].number_format, ws[f'A{r_idx}'] = '@', f"{number} {row.krcode}"
                        row_tspmb = s.execute(text("SELECT djpmy,djyxb FROM tspmb WHERE cpbh=:cpbh AND djpmy<>''"), {'cpbh': row.cpbh}).fetchone()
                        rich_str = apply_subscript(str(row_tspmb.djpmy).upper(), row_tspmb.djyxb) if row_tspmb else str(row.djpmy).upper()
                        ws[f'B{r_idx}'], ws[f'M{r_idx}'] = rich_str, rich_str
                        ws[f'D{r_idx}'], ws[f'F{r_idx}'] = str(row.zwbgpm), f"{safe_float(row.chsl):.0f}{str(row.jldw).upper()}S"
                        
                        if hbdm == 'R':
                            ws[f'H{r_idx}'].number_format, ws[f'H{r_idx}'] = '￥#,##0.00', safe_float(row.kfrmbz)
                        else:
                            ws[f'H{r_idx}'].number_format, ws[f'H{r_idx}'] = f'"{hbdmfh}"#,##0.00', safe_float(row.wxzj)
                        
                        if k == 15 or number == i1:
                            if number == i1:
                                ws[f'F{19+ik}'].border = ws[f'H{19+ik}'].border = top_border
                                ws[f'F{19+ik}'] = 'TOTAL:'
                                ws[f'H{19+ik}'].number_format = '￥#,##0.00' if hbdm == 'R' else f'"{hbdmfh}"#,##0.00'
                                ws[f'H{19+ik}'] = safe_float(row_cymx.mjzj) if hbdm == 'R' else safe_float(row_cymx.wxje)
                            if k == 15: k, i5, ik = 0, i5 + 1, 0

                save_path = os.path.join(config.tmp_path, f"{row_cymx.fphm}报关单.xlsx")
                wb.save(save_path)
                generated_files.append(save_path)

            # 2. 发票
            if (not lvpd and uae != '1') and ('2' in ed):
                if bp == '1' or vcsb == '1':
                    if hbdm == 'R': tpl = 'ujsbxrmbuvinv.xlsx' if ujsb == '1' else ('VC-RINV.xlsx' if vcsb == '1' else 'xrmbuvinv.xlsx')
                    else: tpl = 'ujsbxuvinv.xlsx' if ujsb == '1' else ('VC-INV.xlsx' if vcsb == '1' else 'xuvinv.xlsx')
                else:
                    tpl = 'rmbuvinv.xlsx' if hbdm == 'R' else 'uvinv.xlsx'

                wb = load_workbook(os.path.join(tpl_dir, tpl))
                i1 = len(valid_items)
                i = int(i1 / 8)
                if i == i1 / 8.0: i -= 1
                for a in range(i): wb.copy_worksheet(wb.worksheets[0])
                
                k, i5, number, s_cnt, d = 0, 0, 0, 0, -3
                for row in valid_items:
                    k, d, number = k + 1, d + 4, number + 1
                    if k == 1:
                        ws = wb.worksheets[i5]
                        s_cnt += 1
                        ws[f'{JJ}8'] = str(row_cymx.htjx)
                        ws[f'{JJ}9'] = f"{qsy}.{qsr},{qsn}"
                        if vcsb != '1':
                            ws[f'{JJ}10'], ws[f'{JJ}14'], ws[f'{JJ}15'], ws[f'{JJ}11'] = str(row_cymx.fporder), f"{str(row_cymx.fporder)}S", f"{qsy}.{qsr},{qsn}", str(s_cnt)
                        else:
                            ws['C8'], ws[f'{JJ}10'] = str(row_cymx.shr), str(s_cnt)
                        ws[f'{II}20'] = f"{str(row_cymx.jgtk).upper()} {str(row_cymx.cyka).upper()}"
                        
                    if bp == '1': ws[f'{AA}{21+d}'] = str(number)
                    ws[f'{B}{21+d}'] = f"{row.czkrhh}     " if row.czkrhh else f"{row.cpbh}     "
                    
                    row_tspmb = s.execute(text("SELECT djpmy,djyxb FROM tspmb WHERE cpbh=:cpbh AND djpmy<>''"), {'cpbh': row.cpbh}).fetchone()
                    row_djpmw = s.execute(text("SELECT djpmw FROM wypm WHERE cpbh=:cpbh OR krhh=:cpbh"), {'cpbh': row.cpbh}).fetchone()
                    djpmw_val = str(row_djpmw.djpmw) if row_djpmw else str(row.djpmw)
                    djpmy_val = str(row_tspmb.djpmy) if row_tspmb else str(row.djpmy)
                    
                    rich_djpmy = apply_subscript(f"{djpmy_val.upper()}\nCUSTOMS CODE:{row.krcode}", row_tspmb.djyxb if row_tspmb else '')
                    ws[f'{DD}{21+d}'] = rich_djpmy
                    ws[f'{DD}{22+d}'] = djpmw_val
                    
                    ws[f'{G}{21+d}'], ws[f'{H}{21+d}'] = safe_float(row.chsl), str(row.jldw).upper() + 'S'
                    if hbdm == 'R':
                        ws[f'{II}{21+d}'], ws[f'{KK}{21+d}'] = safe_float(row.khrmb), safe_float(row.kfrmbz)
                    else:
                        ws[f'{II}{21+d}'], ws[f'{KK}{21+d}'] = safe_float(row.wxdj), safe_float(row.wxzj)
                        
                    if k == 8 or number == i1:
                        if number == i1:
                            if vcsb != '1':
                                ws[f'{DD}52'], ws[f'{DD}54'] = 'NOTE: ORIGIN OF GOODS: PRC', 'BANK DETAILS:'
                                if ujsb == '1':
                                    ws[f'{DD}55'] = 'BANK OF CHINA NINGBO BRANCH \nAddress: NO.255 DINGTAI ROAD NINGBO CHINA\nSWIFT CODE: BKCHCNBJ92A\nADDRESS:   350683107991'
                                    ws[f'{DD}56'] = 'A/C NO.  104332051437'
                                else:
                                    ws[f'{DD}55'] = 'HONGKONG AND SHANGHAI BANKING \nCORPORATION LIMITED (HSBC HONG KONG)'
                                    ws[f'{DD}56'] = 'A/C: 078-369527-838'
                                ws[f'{G}57'] = 'TOTAL:'
                            ws[f'{II}57'] = safe_float(row_cymx.mjzj) if hbdm == 'R' else safe_float(row_cymx.wxje)
                        if k == 8: k, d, i5 = 0, -3, i5 + 1

                save_path = os.path.join(config.tmp_path, f"{row_cymx.fphm}inv.xlsx")
                wb.save(save_path)
                generated_files.append(save_path)

            # 3. 装箱单
            if (not lvpd and uae != '1') and ('3' in ed):
                if bp == '1' or vcsb == '1': tpl = 'ujsbxuvpl.xlsx' if ujsb == '1' else ('VC-PL.xlsx' if vcsb == '1' else 'xuvpl.xlsx')
                else: tpl = 'ujsbuvpl.xlsx' if ujsb == '1' else 'uvpl.xlsx'

                wb = load_workbook(os.path.join(tpl_dir, tpl))
                i1 = len(valid_items)
                i = int(i1 / 14)
                if i == i1 / 14.0: i -= 1
                for a in range(i): wb.copy_worksheet(wb.worksheets[0])
                
                k, i5, number, s_cnt, d = 0, 0, 0, 0, -1
                for row in valid_items:
                    k, d, number = k + 1, d + 2, number + 1
                    if k == 1:
                        ws = wb.worksheets[i5]
                        s_cnt += 1
                        if vcsb == '1': ws['C7'] = str(row_cymx.shr)
                        ws[f'{H}7'], ws[f'{H}8'] = str(row_cymx.htjx), f"{qsy}.{qsr},{qsn}"
                        if vcsb != '1':
                            ws[f'{H}9'], ws[f'{H}10'], ws[f'{DD}18'] = str(row_cymx.fporder), str(s_cnt), f"{str(row_cymx.xh)}/{str(row_cymx.fh)}"
                        ws[f'{B}14'], ws[f'{G}14'] = str(row_cymx.cyka).upper(), str(row_cymx.mdka)
                        
                    if bp == '1': ws[f'{AA}{24+d}'] = str(number)
                    ws[f'{B}{24+d}'] = f"{row.czkrhh}    " if row.czkrhh else f"{row.cpbh}    "
                    row_tspmb = s.execute(text("SELECT djpmy,djyxb FROM tspmb WHERE cpbh=:cpbh AND djpmy<>''"), {'cpbh': row.cpbh}).fetchone()
                    rich_djpmy = apply_subscript(str(row_tspmb.djpmy).upper(), row_tspmb.djyxb) if row_tspmb else str(row.djpmy).upper()
                    ws[f'{DD}{24+d}'] = rich_djpmy
                    
                    chsl_v, wxrl_v, xs_v = safe_float(row.chsl), safe_float(row.wxrl), safe_float(row.xs)
                    if wxrl_v > 0 and (chsl_v / wxrl_v) != xs_v:
                        ws[f'{E}{24+d}'].font = ws[f'{F}{24+d}'].font = ws[f'{G}{24+d}'].font = pycopy(ws[f'{E}{24+d}'].font)
                        ws[f'{E}{24+d}'].font.color = ws[f'{F}{24+d}'].font.color = ws[f'{G}{24+d}'].font.color = 'FF0000'
                        
                    ws[f'{E}{24+d}'], ws[f'{F}{24+d}'], ws[f'{G}{24+d}'], ws[f'{H}{24+d}'] = str(row.xs), str(row.wxrl), str(row.chsl), str(row.jldw).upper() + 'S'
                    ws[f'{II}{24+d}'], ws[f'{JJ}{24+d}'], ws[f'{KK}{24+d}'] = round(safe_float(row.zmz), 1), round(safe_float(row.zjz), 1), round(safe_float(row.ztj), 2)
                    
                    if k == 14 or number == i1:
                        if number == i1:
                            ws[f'{DD}54'] = f"TOTAL:        {str(row_cymx.xshj)}        {mzhj}KGS        {jzhj}KGS        {tjhj}CBM       "
                        if k == 14: k, d, i5 = 0, -1, i5 + 1

                save_path = os.path.join(config.tmp_path, f"{row_cymx.fphm}PL.xlsx")
                wb.save(save_path)
                generated_files.append(save_path)
        

        # ================= [ 第二大分支: FF / 俄罗斯 (俄英双语模板) ] =================
        elif hr1 == '2':
            mzhj = sum([round(safe_float(row.zmz), 1) for row in items])
            jzhj = sum([round(safe_float(row.zjz), 1) for row in items])
            tjhj = sum([round(safe_float(row.ztj), 2) for row in items])

            # 1. 报关单
            if '1' in ed:
                tpl = 'uvbgdrmb.xlsx' if hbdm == 'R' else 'uvbgd.xlsx'
                wb = load_workbook(os.path.join(tpl_dir, tpl))
                i1 = len(valid_items)
                divider = 12
                i = math.ceil(i1 / divider) if i1 > 0 else 1
                for a in range(i - 1): wb.copy_worksheet(wb.worksheets[0])
                
                k, i5 = 0, 0
                for number, row in enumerate(valid_items, 1):
                    k += 1
                    ws = wb.worksheets[i5]
                    if k == 1:
                        ws['A5'] = str(row_cymx.cyka).upper()
                        ws['G5'], ws['I5'], ws['E7'], ws['D13'] = str(row_cymx.chyrq), str(row_cymx.fprq), str(row_cymx.ysfs), str(row_cymx.jgtk)
                        cmin_hc = f"{str(row_cymx.cmin).upper()}/{str(row_cymx.hangci).upper()}"
                        ws['Y7'] = ws['F7'] = cmin_hc
                        ws['H7'], ws['E9'], ws['I9'] = str(row_cymx.tdh), str(row_cymx.myfds), 'T/T'
                        ws['H11'], ws['D11'], ws['F11'], ws['A15'] = str(row_cymx.cyka).upper(), str(row_cymx.mygb), str(row_cymx.mdka), str(row_cymx.htjx)
                        ws['D15'], ws['G15'], ws['I15'], ws['A17'] = str(row_cymx.xshj), mzhj, jzhj, f"{str(row_cymx.xh)}/{str(row_cymx.fh)}"

                    r_idx = 22 + k
                    ws[f'A{r_idx}'].number_format, ws[f'A{r_idx}'] = '@', str(row.krcode)
                    
                    row_tspmb = s.execute(text("SELECT djpmy,djyxb FROM tspmb WHERE cpbh=:cpbh AND djpmy<>''"), {'cpbh': row.cpbh}).fetchone()
                    rich_str = apply_subscript(str(row_tspmb.djpmy).upper(), row_tspmb.djyxb) if row_tspmb else str(row.djpmy).upper()
                    ws[f'Z{r_idx}'], ws[f'B{r_idx}'] = rich_str, rich_str
                    
                    ws[f'F{r_idx}'], ws[f'G{r_idx}'], ws[f'H{r_idx}'] = str(row.zwbgpm), safe_float(row.chsl), str(row.jldw).upper() + 'S'
                    ws[f'J{r_idx}'] = safe_float(row.kfrmbz) if hbdm == 'R' else safe_float(row.wxzj)

                    if k == 12 or number == i1:
                        if number == i1:
                            for col in ['G', 'H', 'I', 'J']: ws[f'{col}{23+k}'].border = top_border
                            ws[f'G{23+k}'] = 'TOTAL:'
                            ws[f'J{23+k}'] = safe_float(row_cymx.mjzj) if hbdm == 'R' else safe_float(row_cymx.wxje)
                        if k == 12: k, i5 = 0, i5 + 1

                save_path = os.path.join(config.tmp_path, f"{row_cymx.fphm}报关单.xlsx")
                wb.save(save_path)
                generated_files.append(save_path)

            # 2. 商业发票 (俄英双语)
            if '2' in ed:
                tpl = 'rmbffinv.xlsx' if hbdm == 'R' else 'ffinv.xlsx'
                wb = load_workbook(os.path.join(tpl_dir, tpl))
                i1 = len(valid_items)
                i = math.ceil(i1 / 12) if i1 > 0 else 1
                for a in range(i - 1): wb.copy_worksheet(wb.worksheets[0])
                
                k, i5, d = 0, 0, -1
                for number, row in enumerate(valid_items, 1):
                    k, d = k + 1, d + 1
                    ws = wb.worksheets[i5]
                    if k == 1:
                        ws['A5'] = f"COMMERCIAL INVOICE / КОММЕРЧЕСКИЙ ИНВОЙС № {str(row_cymx.htjx)}"
                        ws['A6'] = f"Date / Дата: {qsr}.{qsy1}.{qsn}"
                        ws['J14'] = f"Terms of delivery / Условия поставки:   {str(row_cymx.cyka).upper()} / ФОБ Нингбо Китай"
                        ws['J16'] = f"The container’s and seal’s no / Номер контейнера и пломбы:{str(row_cymx.xh)}/{str(row_cymx.fh)}"

                    r_idx = 21 + d
                    ws[f'A{r_idx}'] = k
                    
                    row_tspmb = s.execute(text("SELECT djpmy,djyxb FROM tspmb WHERE cpbh=:cpbh AND djpmy<>''"), {'cpbh': row.cpbh}).fetchone()
                    row_djpmw = s.execute(text("SELECT djpmw FROM wypm WHERE cpbh=:cpbh OR krhh=:cpbh"), {'cpbh': row.cpbh}).fetchone()
                    djpmy_val = str(row_tspmb.djpmy).upper() if row_tspmb else str(row.djpmy).upper()
                    djpmw_val = str(row_djpmw.djpmw) if row_djpmw else str(row.djpmw)
                    
                    # 优化：原先用上下两行占据单元格，现直接组装换行富文本
                    ws[f'B{r_idx}'] = apply_subscript(f"{djpmy_val}\n{djpmw_val}", row_tspmb.djyxb if row_tspmb else '')
                    ws[f'C{r_idx}'] = f"{row.czkrhh}    " if row.czkrhh else f"{row.cpbh}    "
                    ws[f'D{r_idx}'], ws[f'E{r_idx}'], ws[f'F{r_idx}'] = 'STD', str(row.czwy), 'CHINA / КИТАЙ'
                    ws[f'G{r_idx}'], ws[f'H{r_idx}'] = 'Hello beautiful', 'Ningbo Union Vision Imp&Exp Co., Ltd'
                    ws[f'I{r_idx}'].number_format, ws[f'I{r_idx}'] = '@', str(row.krcode)
                    ws[f'J{r_idx}'], ws[f'L{r_idx}'] = safe_float(row.chsl), safe_float(row.xs)
                    ws[f'M{r_idx}'], ws[f'N{r_idx}'] = round(safe_float(row.zjz), 1), round(safe_float(row.zmz), 1)
                    
                    if hbdm == 'R': ws[f'K{r_idx}'], ws[f'O{r_idx}'] = safe_float(row.khrmb), safe_float(row.kfrmbz)
                    else: ws[f'K{r_idx}'], ws[f'O{r_idx}'] = safe_float(row.wxdj), safe_float(row.wxzj)

                    if k == 12 or number == i1:
                        if number == i1:
                            ws['J33'], ws['L33'] = safe_float(row_cymx.htzsl1), safe_float(row_cymx.xshj2)
                            ws['M33'], ws['N33'] = safe_float(row_cymx.jzhj), safe_float(row_cymx.mzhj)
                            ws['O33'] = safe_float(row_cymx.mjzj) if hbdm == 'R' else safe_float(row_cymx.wxje)
                        if k == 12: k, d, i5 = 0, -1, i5 + 1

                save_path = os.path.join(config.tmp_path, f"{row_cymx.fphm}inv.xlsx")
                wb.save(save_path)
                generated_files.append(save_path)

            # 3. 装箱单 (FF 包含两个箱单导出: PL.xlsx 和 packing.xlsx)
            if '3' in ed:
                # ------ 生成第一个 FF 专用 PL ------
                tpl_pl = 'ffpl.xlsx'
                wb_pl = load_workbook(os.path.join(tpl_dir, tpl_pl))
                i1 = len(valid_items)
                i = math.ceil(i1 / 14) if i1 > 0 else 1
                for a in range(i - 1): wb_pl.copy_worksheet(wb_pl.worksheets[0])
                
                k, i5, d = 0, 0, 0
                for number, row in enumerate(valid_items, 1):
                    k, d = k + 1, d + 1
                    ws = wb_pl.worksheets[i5]
                    if k == 1:
                        ws['A5'] = f"PACKING LIST / УПАКОВОЧНЫЙ ЛИСТ № {str(row_cymx.htjx)}"
                        ws['A6'] = f"Date / Дата: {qsr}.{qsy1}.{qsn}"
                        ws['K15'] = f"Terms of delivery / Условия поставки:   {str(row_cymx.cyka).upper()} / ФОБ Нингбо Китай"
                        ws['K17'] = f"The container’s and seal’s no / Номер контейнера и пломбы:{str(row_cymx.xh)}/{str(row_cymx.fh)}"

                    r_idx = 21 + d
                    ws[f'A{r_idx}'] = k
                    
                    row_tspmb = s.execute(text("SELECT djpmy,djyxb FROM tspmb WHERE cpbh=:cpbh AND djpmy<>''"), {'cpbh': row.cpbh}).fetchone()
                    row_djpmw = s.execute(text("SELECT djpmw FROM wypm WHERE cpbh=:cpbh OR krhh=:cpbh"), {'cpbh': row.cpbh}).fetchone()
                    ws[f'B{r_idx}'] = apply_subscript(f"{str(row_tspmb.djpmy).upper() if row_tspmb else str(row.djpmy).upper()}\n{str(row_djpmw.djpmw) if row_djpmw else str(row.djpmw)}", row_tspmb.djyxb if row_tspmb else '')
                    ws[f'C{r_idx}'] = f"{row.czkrhh}    " if row.czkrhh else f"{row.cpbh}    "
                    ws[f'D{r_idx}'], ws[f'E{r_idx}'], ws[f'F{r_idx}'] = 'STD', str(row.czwy), 'CHINA / КИТАЙ'
                    ws[f'G{r_idx}'], ws[f'H{r_idx}'] = 'Hello beautiful', 'Ningbo Union Vision Imp&Exp Co., Ltd'
                    ws[f'L{r_idx}'], ws[f'M{r_idx}'] = safe_float(row.chsl), safe_float(row.xs)
                    ws[f'N{r_idx}'], ws[f'O{r_idx}'] = round(safe_float(row.zjz), 1), round(safe_float(row.zmz), 1)

                    if k == 14 or number == i1:
                        if number == i1:
                            ws['L36'], ws['M36'] = safe_float(row_cymx.htzsl1), safe_float(row_cymx.xshj2)
                            ws['N36'], ws['O36'] = safe_float(row_cymx.jzhj), safe_float(row_cymx.mzhj)
                        if k == 14: k, d, i5 = 0, 0, i5 + 1
                
                save_path_pl = os.path.join(config.tmp_path, f"{row_cymx.fphm}PL.xlsx")
                wb_pl.save(save_path_pl)
                generated_files.append(save_path_pl)

                # ------ 生成第二个 FF 专用 packing ------
                tpl_packing = 'rmbffpacking.xlsx' if hbdm == 'R' else 'ffpacking.xlsx'
                wb_packing = load_workbook(os.path.join(tpl_dir, tpl_packing))
                i = math.ceil(i1 / 12) if i1 > 0 else 1
                for a in range(i - 1): wb_packing.copy_worksheet(wb_packing.worksheets[0])
                
                k, i5, d = 0, 0, -1
                for number, row in enumerate(valid_items, 1):
                    k, d = k + 1, d + 1
                    ws = wb_packing.worksheets[i5]
                    if k == 1:
                        ws['H13'] = f"Terms of delivery / Условия поставки:   {str(row_cymx.cyka).upper()} / ФОБ Нингбо Китай"
                        ws['H15'] = f"The container’s and seal’s no / Номер контейнера и пломбы:{str(row_cymx.xh)}/{str(row_cymx.fh)}"

                    r_idx = 21 + d
                    ws[f'A{r_idx}'] = k
                    row_tspmb = s.execute(text("SELECT djpmy,djyxb FROM tspmb WHERE cpbh=:cpbh AND djpmy<>''"), {'cpbh': row.cpbh}).fetchone()
                    row_djpmw = s.execute(text("SELECT djpmw FROM wypm WHERE cpbh=:cpbh OR krhh=:cpbh"), {'cpbh': row.cpbh}).fetchone()
                    ws[f'B{r_idx}'] = apply_subscript(f"{str(row_tspmb.djpmy).upper() if row_tspmb else str(row.djpmy).upper()}\n{str(row_djpmw.djpmw) if row_djpmw else str(row.djpmw)}", row_tspmb.djyxb if row_tspmb else '')
                    ws[f'C{r_idx}'] = f"{row.czkrhh}    " if row.czkrhh else f"{row.cpbh}    "
                    ws[f'D{r_idx}'].number_format, ws[f'D{r_idx}'] = '@', str(row.hgbm)
                    ws[f'E{r_idx}'] = str(row.ysez)
                    ws[f'F{r_idx}'].number_format, ws[f'F{r_idx}'] = '@', str(row.krcode)
                    
                    ws[f'G{r_idx}'], ws[f'H{r_idx}'] = safe_float(row.chsl), safe_float(row.xs)
                    ws[f'I{r_idx}'] = safe_float(row.chsl) / safe_float(row.xs) if safe_float(row.xs) > 0 else 0
                    ws[f'J{r_idx}'], ws[f'K{r_idx}'] = round(safe_float(row.zjz), 1), round(safe_float(row.zmz), 1)

                    if hbdm == 'R': ws[f'L{r_idx}'], ws[f'M{r_idx}'] = safe_float(row.khrmb), safe_float(row.kfrmbz)
                    else: ws[f'L{r_idx}'], ws[f'M{r_idx}'] = safe_float(row.wxdj), safe_float(row.wxzj)

                    if k == 12 or number == i1:
                        if number == i1:
                            ws['G33'], ws['H33'] = safe_float(row_cymx.htzsl1), safe_float(row_cymx.xshj2)
                            ws['J33'], ws['K33'] = safe_float(row_cymx.jzhj), safe_float(row_cymx.mzhj)
                            ws['M33'] = safe_float(row_cymx.mjzj) if hbdm == 'R' else safe_float(row_cymx.wxje)
                        if k == 12: k, d, i5 = 0, -1, i5 + 1
                
                save_path_packing = os.path.join(config.tmp_path, f"{row_cymx.fphm}packing.xlsx")
                wb_packing.save(save_path_packing)
                generated_files.append(save_path_packing)


        # ================= [ 第三大分支: 通用模板 ] =================
        elif hr1 == '3':
            mzhj = sum([round(safe_float(row.zmz), 1) for row in items])
            jzhj = sum([round(safe_float(row.zjz), 1) for row in items])
            tjhj = sum([round(safe_float(row.ztj), 2) for row in items])

            # 1. 报关单
            if '1' in ed:
                tpl = 'uvbgdrmb.xlsx' if hbdm == 'R' else 'uvbgd.xlsx'
                wb = load_workbook(os.path.join(tpl_dir, tpl))
                i1 = len(valid_items)
                i = math.ceil(i1 / 12) if i1 > 0 else 1
                for a in range(i - 1): wb.copy_worksheet(wb.worksheets[0])
                
                k, i5 = 0, 0
                for number, row in enumerate(valid_items, 1):
                    k += 1
                    ws = wb.worksheets[i5]
                    if k == 1:
                        ws['A5'] = str(row_cymx.cyka).upper()
                        ws['G5'], ws['I5'], ws['E7'], ws['D13'] = str(row_cymx.chyrq), str(row_cymx.fprq), str(row_cymx.ysfs), str(row_cymx.jgtk)
                        cmin_hc = f"{str(row_cymx.cmin).upper()}/{str(row_cymx.hangci).upper()}"
                        ws['Y7'] = ws['F7'] = cmin_hc
                        ws['H7'], ws['E9'], ws['I9'] = str(row_cymx.tdh), str(row_cymx.myfds), 'T/T'
                        ws['H11'], ws['D11'], ws['F11'], ws['A15'] = str(row_cymx.cyka).upper(), str(row_cymx.mygb), str(row_cymx.mdka), str(row_cymx.htjx)
                        ws['D15'], ws['G15'], ws['I15'], ws['A17'] = str(row_cymx.xshj), mzhj, jzhj, f"{str(row_cymx.xh)}/{str(row_cymx.fh)}"

                    r_idx = 22 + k
                    ws[f'A{r_idx}'].number_format, ws[f'A{r_idx}'] = '@', str(row.krcode)
                    
                    row_tspmb = s.execute(text("SELECT djpmy,djyxb FROM tspmb WHERE cpbh=:cpbh AND djpmy<>''"), {'cpbh': row.cpbh}).fetchone()
                    rich_str = apply_subscript(str(row_tspmb.djpmy).upper(), row_tspmb.djyxb) if row_tspmb else str(row.djpmy).upper()
                    ws[f'Z{r_idx}'], ws[f'B{r_idx}'] = rich_str, rich_str
                    
                    ws[f'F{r_idx}'], ws[f'G{r_idx}'], ws[f'H{r_idx}'] = str(row.zwbgpm), safe_float(row.chsl), str(row.jldw).upper() + 'S'
                    ws[f'J{r_idx}'] = safe_float(row.kfrmbz) if hbdm == 'R' else safe_float(row.wxzj)

                    if k == 12 or number == i1:
                        if number == i1:
                            for col in ['G', 'H', 'I', 'J']: ws[f'{col}{23+k}'].border = top_border
                            ws[f'G{23+k}'] = 'TOTAL:'
                            ws[f'J{23+k}'] = safe_float(row_cymx.mjzj) if hbdm == 'R' else safe_float(row_cymx.wxje)
                        if k == 12: k, i5 = 0, i5 + 1

                save_path = os.path.join(config.tmp_path, f"{row_cymx.fphm}报关单.xlsx")
                wb.save(save_path)
                generated_files.append(save_path)

            # 2. 发票 (通用型)
            if '2' in ed:
                tpl = 'rmbuvtyinv.xlsx' if hbdm == 'R' else 'uvtyinv.xlsx'
                wb = load_workbook(os.path.join(tpl_dir, tpl))
                i1 = len(valid_items)
                i = math.ceil(i1 / 8) if i1 > 0 else 1
                for a in range(i - 1): wb.copy_worksheet(wb.worksheets[0])
                
                k, i5, d = 0, 0, -3
                for number, row in enumerate(valid_items, 1):
                    k, d = k + 1, d + 4
                    ws = wb.worksheets[i5]
                    if k == 1:
                        ws['B8'], ws['B9'], ws['B11'] = str(row_cymx.khmc), str(row_cymx.address), str(row_cymx.phone)
                        ws['J8'], ws['J9'], ws['J10'] = str(row_cymx.htjx), f"{qsy}.{qsr},{qsn}", i5+1
                        ws['D13'] = str(row_cymx.jhfs) if str(row_cymx.jhfs) else 'T/T '
                        ws['I20'] = f"{str(row_cymx.jgtk).upper()} {str(row_cymx.cyka).upper()}"

                    r_idx = 21 + d
                    ws[f'B{r_idx}'] = f"{row.czkrhh}    " if row.czkrhh else f"{row.cpbh}    "
                    
                    row_tspmb = s.execute(text("SELECT djpmy,djyxb FROM tspmb WHERE cpbh=:cpbh AND djpmy<>''"), {'cpbh': row.cpbh}).fetchone()
                    row_djpmw = s.execute(text("SELECT djpmw FROM wypm WHERE cpbh=:cpbh OR krhh=:cpbh"), {'cpbh': row.cpbh}).fetchone()
                    djpmy_val = str(row_tspmb.djpmy).upper() if row_tspmb else str(row.djpmy).upper()
                    djpmw_val = str(row_djpmw.djpmw) if row_djpmw else str(row.djpmw)
                    
                    ws[f'D{r_idx}'] = apply_subscript(f"{djpmy_val}\nCUSTOMS CODE:{row.krcode}\n{djpmw_val}", row_tspmb.djyxb if row_tspmb else '')
                    ws[f'G{r_idx}'], ws[f'H{r_idx}'] = safe_float(row.chsl), str(row.jldw).upper() + 'S'
                    
                    if hbdm == 'R': ws[f'I{r_idx}'], ws[f'K{r_idx}'] = safe_float(row.khrmb), safe_float(row.kfrmbz)
                    else: ws[f'I{r_idx}'], ws[f'K{r_idx}'] = safe_float(row.wxdj), safe_float(row.wxzj)

                    if k == 8 or number == i1:
                        if number == i1:
                            ws['D54'], ws['D55'], ws['D56'] = 'BANK DETAILS:', 'CHINA EVERBRIGHT BANK, NINGBO BR.', 'A/C: 7680-14-88-0000104-23'
                            ws['G57'] = 'TOTAL:'
                            ws['I57'] = safe_float(row_cymx.mjzj) if hbdm == 'R' else safe_float(row_cymx.wxje)
                        if k == 8: k, d, i5 = 0, -3, i5 + 1

                save_path = os.path.join(config.tmp_path, f"{row_cymx.fphm}inv.xlsx")
                wb.save(save_path)
                generated_files.append(save_path)

            # 3. 装箱单 (通用型)
            if '3' in ed:
                tpl = 'uvtyPACKING.xlsx'
                wb = load_workbook(os.path.join(tpl_dir, tpl))
                i1 = len(valid_items)
                i = math.ceil(i1 / 14) if i1 > 0 else 1
                for a in range(i - 1): wb.copy_worksheet(wb.worksheets[0])
                
                k, i5, d = 0, 0, -1
                for number, row in enumerate(valid_items, 1):
                    k, d = k + 1, d + 2
                    ws = wb.worksheets[i5]
                    if k == 1:
                        ws['B7'], ws['B8'], ws['B10'] = str(row_cymx.khmc), str(row_cymx.address), str(row_cymx.phone)
                        ws['H7'], ws['H8'], ws['H9'] = str(row_cymx.htjx), f"{qsy}.{qsr},{qsn}", i5+1
                        ws['B13'], ws['G13'] = str(row_cymx.cyka), str(row_cymx.mdka)
                        ws['D15'] = str(row_cymx.jhfs) if str(row_cymx.jhfs) else 'T/T '
                        ws['D17'] = f"{str(row_cymx.xh)}/{str(row_cymx.fh)}"

                    r_idx = 24 + d
                    ws[f'B{r_idx}'] = f"{row.czkrhh}    " if row.czkrhh else f"{row.cpbh}    "
                    
                    row_tspmb = s.execute(text("SELECT djpmy,djyxb FROM tspmb WHERE cpbh=:cpbh AND djpmy<>''"), {'cpbh': row.cpbh}).fetchone()
                    ws[f'D{r_idx}'] = apply_subscript(str(row_tspmb.djpmy).upper() if row_tspmb else str(row.djpmy).upper(), row_tspmb.djyxb if row_tspmb else '')

                    chsl_v, wxrl_v, xs_v = safe_float(row.chsl), safe_float(row.wxrl), safe_float(row.xs)
                    if wxrl_v > 0 and (chsl_v / wxrl_v) != xs_v:
                        from openpyxl.styles import Font
                        ws[f'E{r_idx}'].font = ws[f'F{r_idx}'].font = ws[f'G{r_idx}'].font = Font(color='FF0000')

                    ws[f'E{r_idx}'], ws[f'F{r_idx}'], ws[f'G{r_idx}'], ws[f'H{r_idx}'] = str(row.xs), str(row.wxrl), str(row.chsl), str(row.jldw).upper() + 'S'
                    ws[f'I{r_idx}'], ws[f'J{r_idx}'], ws[f'K{r_idx}'] = round(safe_float(row.zmz),1), round(safe_float(row.zjz),1), round(safe_float(row.ztj),2)

                    if k == 14 or number == i1:
                        if number == i1:
                            ws['D53'] = f"TOTAL:        {str(row_cymx.xshj)}        {mzhj}KGS        {jzhj}KGS        {tjhj}CBM       "
                        if k == 14: k, d, i5 = 0, -1, i5 + 1
                
                save_path = os.path.join(config.tmp_path, f"{row_cymx.fphm}PL.xlsx")
                wb.save(save_path)
                generated_files.append(save_path)


        # ================= [ UAE / LV 特殊客户逻辑 (动态行插入) ] =================
        if lvpd == '1' or uae == '1':
            if '2' in ed:
                tpl = 'lvinv.xlsx' if lvpd == '1' else 'uaeinv.xlsx'
                wb = load_workbook(os.path.join(tpl_dir, tpl))
                ws = wb.worksheets[0]
                k, number = 0, 19
                if bank: ws['A13'] = bank
                if lvpd == '1':
                    ws['E8'], ws['E9'], ws['G17'] = f"INVOICE NO. :{str(row_cymx.htjx)}", f"DATE              :{str(row_cymx.fprq)}", f"{str(row_cymx.jgtk).upper()} {str(row_cymx.cyka).upper()}"
                else:
                    ws['H8'], ws['H9'] = f"INVOICE NO. :{str(row_cymx.htjx)}", f"DATE              :{str(row_cymx.fprq)}"
                
                for row in valid_items:
                    k += 1
                    ws.insert_rows(19 + k)
                    copy_row_style(ws, number, 19 + k)
                    ws.row_dimensions[19+k].height = 15
                    
                    row_zscp = s.execute(text("SELECT krtm,lvkrtm FROM zscp WHERE cpbh=:cpbh OR krhh=:cpbh"), {'cpbh': row.cpbh}).fetchone()
                    if row_zscp:
                        ws[f'A{19+k}'].number_format, ws[f'A{19+k}'] = '@', str(row_zscp.krtm)
                        ws[f'B{19+k}'].number_format, ws[f'B{19+k}'] = '@', str(row_zscp.lvkrtm)
                        
                    c_val = f"{row.czkrhh}    " if row.czkrhh else f"{row.cpbh}    "
                    ws[f'C{19+k}'] = c_val
                    
                    if lvpd == '1':
                        ws[f'D{19+k}'], ws[f'E{19+k}'], ws[f'F{19+k}'] = str(row.djpmy).upper(), safe_float(row.chsl), str(row.jldw).upper() + 'S'
                        if hbdm == 'R':
                            ws[f'G{19+k}'].number_format, ws[f'H{19+k}'].number_format = f'￥#,##{dd1}', '￥#,##0.00'
                            ws[f'G{19+k}'], ws[f'H{19+k}'] = safe_float(row.khrmb), safe_float(row.kfrmbz)
                        else:
                            ws[f'G{19+k}'].number_format, ws[f'H{19+k}'].number_format = f'"{hbdmfh}"#,##{dd1}', f'"{hbdmfh}"#,##0.00'
                            ws[f'G{19+k}'], ws[f'H{19+k}'] = safe_float(row.wxdj), safe_float(row.wxzj)
                    else:
                        d_val = str(row.djpmy).upper() if str(row.krpp) in ('', '无') else f"{str(row.djpmy).upper()},{str(row.krpp).upper()}"
                        ws[f'D{19+k}'], ws[f'H{19+k}'], ws[f'G{19+k}'] = d_val, safe_float(row.chsl), str(row.jldw).upper() + 'S'
                        ws[f'I{19+k}'], ws[f'J{19+k}'] = safe_float(row.wxdj), safe_float(row.wxzj)

                ws.delete_rows(number)
                save_path = os.path.join(config.tmp_path, f"{row_cymx.fphm}inv.xlsx")
                wb.save(save_path)
                generated_files.append(save_path)

            if '3' in ed:
                tpl = 'lvpl.xlsx' if lvpd == '1' else 'uaepl.xlsx'
                wb = load_workbook(os.path.join(tpl_dir, tpl))
                ws = wb.worksheets[0]
                k, number = 0, 24
                if lvpd == '1':
                    ws['H10'], ws['H11'] = f"INVOICE NO. :{str(row_cymx.htjx)}", f"DATE              :{str(row_cymx.fprq)}"
                    ws['A17'], ws['C18'] = f"FROM: {str(row_cymx.cyka).upper()}", f" {str(row_cymx.xh).upper()}/ {str(row_cymx.fh).upper()}"
                else:
                    ws['k9'], ws['k10'] = str(row_cymx.htjx), str(row_cymx.fprq)
                    ws['C17'], ws['C18'] = f"FROM: {str(row_cymx.cyka).upper()}", f"THE CONTAINER’S AND SEAL’S NO.: {str(row_cymx.xh).upper()}/ {str(row_cymx.fh).upper()}"
                
                for row in valid_items:
                    k += 1
                    ws.insert_rows(24 + k)
                    copy_row_style(ws, number, 24 + k)
                    ws.row_dimensions[24+k].height = 15
                    
                    row_zscp = s.execute(text("SELECT krtm,lvkrtm,bzcd,bzkd,bzgd,nhzd,nhkd,nhgd FROM zscp WHERE cpbh=:cpbh OR krhh=:cpbh"), {'cpbh': row.cpbh}).fetchone()
                    if row_zscp:
                        ws[f'A{24+k}'].number_format, ws[f'A{24+k}'] = '@', str(row_zscp.krtm)
                        ws[f'B{24+k}'].number_format, ws[f'B{24+k}'] = '@', str(row_zscp.lvkrtm)
                        ws[f'J{24+k}'], ws[f'K{24+k}'], ws[f'L{24+k}'] = round(safe_float(row_zscp.bzcd),1), round(safe_float(row_zscp.bzkd),1), round(safe_float(row_zscp.bzgd),1)
                        ws[f'M{24+k}'], ws[f'N{24+k}'], ws[f'O{24+k}'] = round(safe_float(row_zscp.nhzd),1), round(safe_float(row_zscp.nhkd),1), round(safe_float(row_zscp.nhgd),1)
                        
                    ws[f'C{24+k}'] = f"{row.czkrhh}    " if row.czkrhh else f"{row.cpbh}    "
                    ws[f'D{24+k}'] = str(row.djpmy).upper() if (lvpd == '1' or str(row.krpp) in ('', '无')) else f"{str(row.djpmy).upper()},{str(row.krpp).upper()}"
                    ws[f'E{24+k}'], ws[f'F{24+k}'], ws[f'G{24+k}'] = str(row.wxrl), str(row.xs), str(row.chsl)
                    ws[f'H{24+k}'], ws[f'I{24+k}'], ws[f'P{24+k}'] = round(safe_float(row.zmz),1), round(safe_float(row.zjz),1), round(safe_float(row.ztj),2)
                    
                ws.delete_rows(number)
                save_path = os.path.join(config.tmp_path, f"{row_cymx.fphm}PL.xlsx")
                wb.save(save_path)
                generated_files.append(save_path)
        



        # ================= [ 打包与返回 ] =================
        if len(generated_files) > 1:
            zip_name = f"{row_cymx.fphm}_单证集.zip"
            zip_path = os.path.join(config.tmp_path, zip_name)
            with zipfile.ZipFile(zip_path, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
                for f in generated_files:
                    zf.write(f, os.path.basename(f))
                    os.remove(f)
            return json_result(1, '单据生成成功', data=zip_name)
        elif len(generated_files) == 1:
            return json_result(1, '单据生成成功', data=os.path.basename(generated_files[0]))
        else:
            return json_result(-1, '未生成任何单据，请检查选项或数据')

    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'生成单据报错: {str(e)}')
    finally:
        s.close()
