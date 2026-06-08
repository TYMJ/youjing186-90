'''
Author: gavin
Date: 2022-08-12 16:08:53
LastEditors: gavin
LastEditTime: 2022-08-15 09:28:33
Description: 装箱单_主文件
'''
from any import *
from .model import *
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, Alignment, NamedStyle,PatternFill,Color
from openpyxl.drawing.image import Image as Image_Get
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import  XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.worksheet.protection import SheetProtection

import math
from sqlalchemy.sql import func
import os,time
import json
from datetime import datetime
# from num2words import num2words
from openpyxl.utils import get_column_letter
 
# def number_to_english(num):
#     return num2words(num, lang='en')


def extract_digits(string):
    digits = ""
    for char in string:
        if char.isdigit():
            digits += str(char)
    return digits

def number_to_english(num):
    digit_dict = {
        0: "zero", 1: "one", 2: "two", 3: "three", 4: "four",
        5: "five", 6: "six", 7: "seven", 8: "eight", 9: "nine",
        10: "ten", 11: "eleven", 12: "twelve", 13: "thirteen", 14: "fourteen",
        15: "fifteen", 16: "sixteen", 17: "seventeen", 18: "eighteen", 19: "nineteen",
        20: "twenty", 30: "thirty", 40: "forty", 50: "fifty", 60: "sixty",
        70: "seventy", 80: "eighty", 90: "ninety"
    }
 
    if num in digit_dict:
        return digit_dict[num]
 
    if num < 100:
        return digit_dict[num // 10 * 10] + " " + digit_dict[num % 10]
    
    if num < 1000:
        return digit_dict[num // 100] + " hundred " + number_to_english(num % 100)
 
    if num < 1000000:
        return number_to_english(num // 1000) + " thousand " + number_to_english(num % 1000)
    
    if num < 1000000000:
        return number_to_english(num // 1000000) + " million " + number_to_english(num % 1000000)
 
    return "Out of range"

def changeNumToChar(ch):
    if ch <= 26:
        result_char=chr(ch + 64)
    else:
        result_char=chr(64 + (ch // 26)) + chr(64 + (ch % 26))
    return result_char

# 根据单元格内容的长度自动调整行高
def auto_adjust_row_height(ws, row, col):
    # for col in range(start_col, end_col + 1):
    cell = ws.cell(row, col)
    logger.error(cell.value)
    if cell.value !=None and cell.value!="":
        # 获取单元格内容的长度，并加上单元格内上下的边距（例如：2pt）
        logger.error(cell.value)
        # logger.error(cell.char_width)
        logger.error(cell.font.underline)
        width =  cell.char_width * len(cell.value)
        # 设置行高
        ws.row_dimensions[row].height = max(ws.row_dimensions[row].height, width)

def pixels_to_points(value, dpi=96):
    """96 dpi, 72i"""
    return value * 72 / dpi


def points_to_pixels(value, dpi=96):
    return int(math.ceil(value * dpi / 72))

def offset_img(img, col, row, x_pad=4, y_pad=25):
    """精确设置图片位置，偏移量以万为单位进行微调吧，具体计算公式太麻烦了
    row column 的索引都是从0开始的，我这里要把图片插入到单元格A17
    """
    p2e = pixels_to_EMU
    h, w = img.height, img.width
    # 图像等比例缩放因子
    resize_factor = 0.8
    w_h_ratio = w/h
    resize_H = int(resize_factor * h)
    resize_W = int(resize_factor * resize_H * w_h_ratio)
    #
    # x_pad = 4
    # y_pad = 25
    # 注意这里的行、列索引从0开始, 所以需要减1
    size = XDRPositiveSize2D(p2e(w), p2e(h))
    marker = AnchorMarker(col=col-1, colOff=pixels_to_EMU(x_pad), row=row-1, rowOff=pixels_to_EMU(y_pad))
    img.anchor = OneCellAnchor(_from=marker, ext=size)

    # 图像在Excel里面的大小
    # image_size_excel = XDRPositiveSize2D(pixels_to_EMU(resize_W), pixels_to_EMU(resize_H))
    ##############################
    # 设置单元格大小，单元格默认宽度单位：字符；高度单位：point(磅)
    # cell_height = int(pixels_to_points(resize_H+10, dpi=96)) #高度上增加10个像素放大单元格
    # cell_width = int(resize_W/8) + 2 # 宽度上增加16（2*8）个像素放大单元格

    # marker = AnchorMarker(col=col, colOff=pixels_to_EMU(x_pad), row=row, rowOff=pixels_to_EMU(y_pad))
    # img.anchor = OneCellAnchor(_from=marker, ext=image_size_excel)

# 导出数据到服务器上Excel文件中
def customer_excel_futong(rids):
    s = Session()
    try:
        # wb = Workbook() # 新建Excel文件
        # ws = wb.active # wb.create_sheet() # 默认激活第一个sheet
        r_path = os.path.join(config.data_upload_path,'template')
        fn = os.path.join(r_path, '优景富通客户导入模板.xlsx')
        
        if not os.path.exists(fn):
            return {"code":-1,"msg":"未找到报表模板"}
        wb = load_workbook(fn) 
        ws = wb.get_sheet_by_name('客户')

        i = 1
        path = config.tmp_path
        for rid in rids:
            d = s.query(kh).filter(kh.rid==rid).first()
            if not d:
                return {"code":-1,"msg":"未找到客户资料记录"}
            c = alchemy_object_to_dict(d)
            i += 1
            ws['A' + str(i)] = str(c.get('kh_id')).upper().strip()
            ws['B' + str(i)] = str(c.get('company_name')).upper().strip()
            ws['C' + str(i)] = str(c.get('khjc')).upper().strip()
            ws['D' + str(i)] = str(c.get('khlx')).strip()
            ws['E' + str(i)] = str(c.get('khly')).strip()
            ws['F' + str(i)] = str(c.get('hzdj')).strip()
            ws['G' + str(i)] = str(c.get('country')).upper().strip()
            ws['H' + str(i)] = ''
            ws['I' + str(i)] = str(c.get('ywfw')).strip()
            ws['J' + str(i)] = str(c.get('syms')).strip()
            ws['K' + str(i)] = str(c.get('web','')).upper().strip()
            ws['L' + str(i)] = str(c.get('address')).upper().strip()
            ws['M' + str(i)] = str(c.get('phone')).upper().strip()
            ws['N' + str(i)] = str(c.get('mdka')).strip()
            ws['O' + str(i)] = ''
            ws['P' + str(i)] = ''
            ws['Q' + str(i)] = ''
            ws['R' + str(i)] = ''
            ws['S' + str(i)] = str(c.get('cslxr')).strip()
            ws['T' + str(i)] = str(c.get('sjhm')).strip()
            ws['U' + str(i)] = str(c.get('fax')).upper().strip()
            ws['V' + str(i)] = str(c.get('account')).upper().strip()
            ws['W' + str(i)] = str(c.get('bank')).upper().strip()
            ws['X' + str(i)] = str(c.get('ywry')).strip()
            ws['Y' + str(i)] = ''
            ws['Z' + str(i)] = ''
            ws['AA' + str(i)] = time.strftime('%Y-%m-%d')
            ws['AB' + str(i)] = str(c.get('memo')).strip()
            ws['AC' + str(i)] = ''
            ws['AD' + str(i)] = ''
            ws['AE' + str(i)] = c.get('xbed')
            ws['AF' + str(i)] = c.get('yyed')
            ws['AG' + str(i)] = c.get('kfrq')
            ws['AH' + str(i)] = str(c.get('kdzh')).strip()
            ws['AI' + str(i)] = c.get('kfrq')
            ws['AJ' + str(i)] = str(c.get('dlmc')).strip()
            ws['AK' + str(i)] = str(c.get('jgtk')).upper().strip()
            ws['AL' + str(i)] = str(c.get('zdhd')).upper().strip()
            ws['AM' + str(i)] = str(c.get('cslxr')).upper().strip()
            ws['AN' + str(i)] = str(c.get('sjhm')).strip()
            ws['AO' + str(i)] = str(c.get('email')).upper().strip()

        ws = wb.get_sheet_by_name('联系人')
        i = 1
        for rid in rids:
            d = s.query(khlxr,kh.kh_id,kh.company_name,kh.ywry).filter(kh.rid==rid).outerjoin(kh,kh.rid==khlxr.pid).first()
            if not d:
                continue
            c = alchemy_object_to_dict(d)
            i += 1
            ws['A' + str(i)] = str(c.get('kh_id','')).upper().strip()
            ws['B' + str(i)] = str(c.get('company_name')).upper().strip()
            ws['C' + str(i)] = '' 
            ws['D' + str(i)] = '' 
            ws['E' + str(i)] = str(c.get('xm')).upper().strip()
            ws['F' + str(i)] = str(c.get('xm')).upper().strip()
            ws['G' + str(i)] = str(c.get('email')).upper().strip()
            ws['H' + str(i)] = str(c.get('mphone')).strip()
            ws['I' + str(i)] = str(c.get('phone')).strip() 
            ws['J' + str(i)] = str(c.get('sex')).strip() 
            ws['K' + str(i)] =  c.get('csrq')
            ws['L' + str(i)] =  c.get('fax')
            ws['M' + str(i)] =  c.get('zw')
            ws['N' + str(i)] = '' 
            ws['O' + str(i)] = '' 
            ws['P' + str(i)] = '' 
            ws['Q' + str(i)] = '' 
            ws['R' + str(i)] = '' 
            ws['S' + str(i)] = '' 
            ws['T' + str(i)] = '' 
            ws['U' + str(i)] = '' 
            ws['V' + str(i)] = '' 
            ws['W' + str(i)] =  c.get('ywry')

        report_rid = get_uuid()
        wb.save(path + '/'+ str(report_rid)+'.xlsx')
        return {'code':1,'msg':'生成报表成功','data': str(report_rid)+'.xlsx'}
    except:
        logger.error(trace_error())
        return {'code':-1,'msg':trace_error()}
    finally:
        s.close()


'''
Description: 导出数据到 excel
return {*}
'''
@any_route('/api/saier/customer/excel/futong',methods=['POST','GET'])
@require_token
async def api_saier_customer_excel_futong(request):
    j = await request.json()
    try:
        rids = j.get('rids','')
        res = customer_excel_futong(rids)

        return json_result(res.get('code'),res.get('msd'),res.get('data'))
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1,trace_error())
    

#  导出数据到新建 excel 文件中
def customer_excel_data(rids):
    s = Session()
    try:
        wb = Workbook() # 新建Excel文件
        ws = wb.active # wb.create_sheet() # 默认激活第一个sheet
        i = 1
        ws.column_dimensions['A'].width = 14.3
        ws['A1'] = '序列号'
        ws.column_dimensions['B'].width = 14.3
        ws['B1'] = '客户编号'
        ws.column_dimensions['C'].width = 25
        ws['C1'] = '上属公司'
        ws.column_dimensions['D'].width = 17
        ws['D1'] = '联系人'
        ws.column_dimensions['E'].width = 14.3
        ws['E1'] = '职务'
        ws.column_dimensions['F'].width = 17
        ws['F1'] = '联系方式'
        ws.column_dimensions['G'].width = 14.3
        ws['G1'] = '性质'
        ws.column_dimensions['H'].width = 20
        ws['H1'] = '电子邮箱'
        ws.column_dimensions['I'].width = 14.3
        ws['I1'] = '贸易国别'
        ws.column_dimensions['J'].width = 14.3
        ws['J1'] = '所属地区'
        ws.column_dimensions['K'].width = 18
        ws['K1'] = '公司网站'
        ws.column_dimensions['L'].width = 17
        ws['L1'] = '客户来源'
        ws.column_dimensions['M'].width = 23
        ws['M1'] = '详细地址'
        ws.column_dimensions['N'].width = 8
        ws['N1'] = '客户类型'
        ws.column_dimensions['O'].width = 14.3
        ws['O1'] = '客户备注'
        ws.column_dimensions['P'].width = 14.3
        ws['P1'] = '业务人员'
        ws.column_dimensions['Q'].width = 14.3
        ws['Q1'] = '首次联系时间'
        ws.column_dimensions['R'].width = 14.3
        ws['R1'] = '支付方式'
        ws.column_dimensions['S'].width = 14.3
        ws['S1'] = '跟踪反馈'
        path = config.tmp_path

        # 创建字体
        font_ = Font(
            name="Arial",
            size=12,
            bold=False
        )
        # 创建单元格样式
        alight_ = Alignment(
            horizontal='left',  # 水平对齐方式:center, left, right
            vertical='center',  # 垂直对齐方式: center, top, bottom
            wrap_text=True
        )
        # 创建边框样式
        border_ = Border(left=Side(style='thin',color='FF000000'),
                right=Side(style='thin',color='FF000000'),
                top=Side(style='thin',color='FF000000'),
                bottom=Side(style='thin',color='FF000000'))
        
        for rid in rids:
            d = s.query(kh).filter(kh.rid==rid).first()
            if not d:
                return {"code":-1,"msg":"未找到客户资料记录"}
            c = alchemy_object_to_dict(d)
            i += 1
            ws['A' + str(i)] = i-1
            ws['B' + str(i)] = str(c.get('kh_id')).upper()
            # ws.merge_cells('A%d:B%d' % (i,i+1)) # 合并单元格 合并单元格后，后续代码不能再次赋值否则会报错
            ws['C' + str(i)] = str(c.get('company_name')).upper()
            ws['D' + str(i)] = str(c.get('cslxr')).upper()
            l = s.query(khlxr.zw).filter(khlxr.pid==rid,khlxr.xm==c.get('cslxr')).first()
            if l:
                ws['E' + str(i)] = str(l.zw).upper()
            ws['F' + str(i)] = str(c.get('phone')).upper()
            ws['G' + str(i)] = str(c.get('syms')).upper()
            ws['H' + str(i)] = str(c.get('email')).upper()
            ws['I' + str(i)] = str(c.get('country')).upper()
            ws['J' + str(i)] = str(c.get('ssdq')).upper()
            ws['K' + str(i)] = str(c.get('web')).upper()
            ws['L' + str(i)] = str(c.get('khly')).upper()
            ws['M' + str(i)] = str(c.get('address')).upper()
            ws['N' + str(i)] = str(c.get('khlx')).upper()
            ws['O' + str(i)] = str(c.get('memo')).upper()
            ws['P' + str(i)] = str(c.get('ywry')).upper()
            ws['Q' + str(i)] = str(c.get('kfrq')).upper()
            ws['R' + str(i)] = str(c.get('zffs')).upper()
            ws['S' + str(i)] = ''

        for line in tuple(ws['A%d:S%d'%(1,i)]):
                for cell in line:
                    cell.font = font_
                    cell.alignment = alight_
                    cell.border=border_

        report_rid = get_uuid()
        wb.save(path + '/'+ str(report_rid)+'.xlsx')
        return {'code':1,'msg':'生成报表成功','data': str(report_rid)+'.xlsx'}
    except:
        logger.error(trace_error())
        return {'code':-1,'msg':trace_error()}
    finally:
        s.close()


'''
Description: 导出数据到 excel
return {*}
'''
@any_route('/api/saier/customer/excel/data',methods=['POST','GET'])
@require_token
async def api_saier_customer_excel_data(request):
    j = await request.json()
    try:
        rids = j.get('rids','')
        res = customer_excel_data(rids)

        return json_result(res.get('code'),res.get('msd'),res.get('data'))
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1,trace_error())