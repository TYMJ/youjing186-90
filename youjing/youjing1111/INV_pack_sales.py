import asyncio
from concurrent.futures import ThreadPoolExecutor
from any import *  # 假设此模块包含了 Session, config, logger, get_uuid, require_token, any_route, json_result
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import os
import traceback
import io # 需要导入 io 模块来处理内存流

# --- 核心逻辑函数 ---
def _generate_excel_logic(selected_contract_rids, user_name):
    """
    生成 Excel 并返回字节流 (Bytes)，而不是保存在服务器上。
    """
    
    s = Session()
    try:
        # --- 1. 权限检查 ---
        # perm_sql = "SELECT 1 FROM cyzglsheet WHERE xm = :uname AND zm = '合同明细导出' LIMIT 1"
        # if not s.execute(perm_sql, {'uname': user_name}).fetchone():
        #     return {'code': -1, 'msg': f'权限不足：用户 [{user_name}] 没有【合同明细导出】权限', 'data': ''}

        if not selected_contract_rids:
            return {'code': -1, 'msg': '未选择任何记录', 'data': ''}

        # --- 2. 生成 Excel (在内存中) ---
        wb = Workbook()
        ws = wb.active
        ws.title = "优景 INV pack sales"

        # 设置标题行
        headers = [
            'Order NO.', 'Costomer NO.', 'Item NO.', 'item Name', '款式', '颜色', 'Pcs/Ctn', 'UNIT', 
            'Inner CTNS in 1 CTNt', 'кол-во кор (CTNS)', 'TOTAL QTY', 'USD', 'AMOUNT', 'ETD', 'G.W.', 
            'N,W.', 'Volume', '工厂', 'RMB', 'Amount(RMB)', '下单地点', '跟单人员', '工厂回签日期', 
            '工厂交期', '产前确认', 'AW确认', 'AW', 'SM确认', 'SM', 'PR-M SAMPLE', 'INSPECTION PHOTS', 
            'LOADING TIME', 'IN ORDER', 'QTY', '备注', '采购业务员', '业务员', '船期'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(col_num)].width = 14.3

        current_row = 2

        # --- 3. 数据库查询与填充 (保留你原有的复杂逻辑) ---
        # 将字符串列表转换为适合SQL IN子句的格式
        placeholders = ','.join([f':num_{i}' for i in range(len(selected_contract_rids))])
        params = {f'num_{i}': num for i, num in enumerate(selected_contract_rids)}

        # 查询主表
        query_main_contracts = f""" SELECT wxht, hthm, szdq, gdry, ywry, rid FROM cght WHERE rid IN ({placeholders}) """
        main_contracts = s.execute(query_main_contracts, params).fetchall()

        for main_contract in main_contracts:
            # 获取业务员信息
            salesperson = ''
            if main_contract.wxht:
                salesperson_query = "SELECT ywry FROM wxht WHERE order_id = :order_id"
                salesperson_result = s.execute(salesperson_query, {'order_id': main_contract.wxht}).fetchone()
                if salesperson_result:
                    salesperson = salesperson_result.ywry or ''

            # 获取明细
            details_query = """ SELECT mjdj1, khhh, cpbh, bjhh, ywpm, ks, yse, wxrl, jldw, nhwx, cgxs, cgsl, Twxdj, zje, yjcq, mz, jz, wxwyzd, bz3, tj, csmc, sccj1, cgjg, jhrq, cqqr, AWqr, SMqr, rid, pid FROM cghtsheet WHERE pid = :contract_rid """
            details = s.execute(details_query, {'contract_rid': main_contract.rid}).fetchall()

            for detail in details:
                # 获取采购跟单 (工厂回签日期)
                factory_sign_date = ''
                if main_contract.hthm:
                    factory_sign_query = "SELECT gdrq FROM cggd WHERE hthm = :hthm"
                    factory_sign_result = s.execute(factory_sign_query, {'hthm': main_contract.hthm}).fetchone()
                    if factory_sign_result:
                        factory_sign_date = factory_sign_result.gdrq or ''

                # 获取采购明细 (QTY, Volume)
                chxs_combined = ''
                tj_combined = ''
                if detail.wxwyzd or detail.wxwyzd != '':
                    cymx_query = """ SELECT chxs, wxtj FROM cymxsheet WHERE wxwyzd = :wxwyzd AND chsl1 > 0 AND cght = :cght """
                    cymx_results = s.execute(cymx_query, { 'wxwyzd': detail.wxwyzd or '', 'cght': main_contract.hthm or '' }).fetchall()
                    for idx, cymx_item in enumerate(cymx_results):
                        if idx == 0:
                            chxs_combined = str(cymx_item.chxs)
                            if cymx_item.wxtj > 0: tj_combined = str(cymx_item.wxtj)
                        else:
                            chxs_combined += '/' + str(cymx_item.chxs)
                            if cymx_item.wxtj > 0: tj_combined += '/' + str(cymx_item.wxtj)

                # 获取发票信息 (LOADING TIME, IN ORDER)
                loading_time = ''
                in_order = ''
                if detail.rid:
                    invoice_query = "SELECT zgrq, fphm FROM cymx WHERE rid = :rid"
                    invoice_result = s.execute(invoice_query, {'rid': detail.rid}).fetchone()
                    if invoice_result:
                        loading_time = invoice_result.zgrq or ''
                        in_order = invoice_result.fphm or ''


                # 7. 填充 Excel 行 (A to AL)
                # A: Order NO.
                ws.cell(row=current_row, column=1, value=main_contract.wxht or '').alignment = Alignment(horizontal='center', vertical='center')
                # B: Costomer NO.
                ws.cell(row=current_row, column=2, value=detail.khhh or '').alignment = Alignment(horizontal='center', vertical='center')
                # C: Item NO.
                item_no = detail.bjhh or detail.cpbh
                ws.cell(row=current_row, column=3, value=item_no or '').alignment = Alignment(horizontal='center', vertical='center')
                # D: item Name
                ws.cell(row=current_row, column=4, value=detail.ywpm or '').alignment = Alignment(horizontal='center', vertical='center')
                # E: 款式
                ws.cell(row=current_row, column=5, value=detail.ks or '').alignment = Alignment(horizontal='center', vertical='center')
                # F: 颜色
                ws.cell(row=current_row, column=6, value=detail.yse or '').alignment = Alignment(horizontal='center', vertical='center')
                # G: Pcs/Ctn
                ws.cell(row=current_row, column=7, value=float(detail.wxrl or 0)).alignment = Alignment(horizontal='center', vertical='center')
                # H: UNIT
                ws.cell(row=current_row, column=8, value=detail.jldw or '').alignment = Alignment(horizontal='center', vertical='center')
                # I: Inner CTNS in 1 CTNt
                ws.cell(row=current_row, column=9, value=int(detail.nhwx or 0)).alignment = Alignment(horizontal='center', vertical='center')
                # J: кол-во кор (CTNS)
                ws.cell(row=current_row, column=10, value=float(detail.cgxs or 0)).alignment = Alignment(horizontal='center', vertical='center')
                # K: TOTAL QTY
                ws.cell(row=current_row, column=11, value=float(detail.cgsl or 0)).alignment = Alignment(horizontal='center', vertical='center')
                # L: USD (价格)
                unit_price = detail.mjdj1 if detail.mjdj1 > 0 else detail.Twxdj
                ws.cell(row=current_row, column=12, value=float(unit_price or 0)).alignment = Alignment(horizontal='center', vertical='center')
                # M: AMOUNT (金额)
                amount = (unit_price or 0) * (detail.cgsl or 0)
                ws.cell(row=current_row, column=13, value=float(amount)).alignment = Alignment(horizontal='center', vertical='center')
                # N: ETD
                ws.cell(row=current_row, column=14, value=detail.yjcq or '').alignment = Alignment(horizontal='center', vertical='center')
                # O: G.W.
                ws.cell(row=current_row, column=15, value=float(detail.mz or 0)).alignment = Alignment(horizontal='center', vertical='center')
                # P: N,W.
                ws.cell(row=current_row, column=16, value=float(detail.jz or 0)).alignment = Alignment(horizontal='center', vertical='center')
                # Q: Volume
                ws.cell(row=current_row, column=17, value=tj_combined).alignment = Alignment(horizontal='center', vertical='center')
                # R: 工厂
                factory_name = detail.sccj1 or detail.csmc
                ws.cell(row=current_row, column=18, value=factory_name or '').alignment = Alignment(horizontal='center', vertical='center')
                # S: RMB
                ws.cell(row=current_row, column=19, value=float(detail.cgjg or 0)).alignment = Alignment(horizontal='center', vertical='center')
                # T: Amount(RMB)
                ws.cell(row=current_row, column=20, value=float(detail.zje or 0)).alignment = Alignment(horizontal='center', vertical='center')
                # U: 下单地点
                ws.cell(row=current_row, column=21, value=main_contract.szdq or '').alignment = Alignment(horizontal='center', vertical='center')
                # V: 跟单人员
                ws.cell(row=current_row, column=22, value=main_contract.gdry or '').alignment = Alignment(horizontal='center', vertical='center')
                # W: 工厂回签日期
                ws.cell(row=current_row, column=23, value=factory_sign_date).alignment = Alignment(horizontal='center', vertical='center')
                # X: 工厂交期
                ws.cell(row=current_row, column=24, value=detail.jhrq or '').alignment = Alignment(horizontal='center', vertical='center')
                # Y: 产前确认
                ws.cell(row=current_row, column=25, value=detail.cqqr or '').alignment = Alignment(horizontal='center', vertical='center')
                # Z: AW确认
                ws.cell(row=current_row, column=26, value=detail.AWqr or '').alignment = Alignment(horizontal='center', vertical='center')
                # AA: AW (图片占位)
                ws.cell(row=current_row, column=27, value='').alignment = Alignment(horizontal='center', vertical='center')
                # AB: SM确认
                ws.cell(row=current_row, column=28, value=detail.SMqr or '').alignment = Alignment(horizontal='center', vertical='center')
                # AC: SM (图片占位)
                ws.cell(row=current_row, column=29, value='').alignment = Alignment(horizontal='center', vertical='center')
                # AD: PR-M SAMPLE
                ws.cell(row=current_row, column=30, value='').alignment = Alignment(horizontal='center', vertical='center')
                # AE: INSPECTION PHOTS
                ws.cell(row=current_row, column=31, value='').alignment = Alignment(horizontal='center', vertical='center')
                # AF: LOADING TIME
                ws.cell(row=current_row, column=32, value=loading_time).alignment = Alignment(horizontal='center', vertical='center')
                # AG: IN ORDER
                ws.cell(row=current_row, column=33, value=in_order).alignment = Alignment(horizontal='center', vertical='center')
                # AH: QTY
                ws.cell(row=current_row, column=34, value=chxs_combined).alignment = Alignment(horizontal='center', vertical='center')
                # AI: 备注
                ws.cell(row=current_row, column=35, value=detail.bz3 or '').alignment = Alignment(horizontal='center', vertical='center')
                # AJ: 采购业务员
                ws.cell(row=current_row, column=36, value=user_name).alignment = Alignment(horizontal='center', vertical='center')
                # AK: 业务员
                ws.cell(row=current_row, column=37, value=salesperson).alignment = Alignment(horizontal='center', vertical='center')
                # AL: 船期
                ws.cell(row=current_row, column=38, value=detail.yjcq or '').alignment = Alignment(horizontal='center', vertical='center')

                current_row += 1

         # --- 5. 关键修改：不保存到磁盘，而是保存到内存 ---
        # 创建一个 BytesIO 对象
        output = io.BytesIO()
        wb.save(output)
        # 获取二进制数据
        excel_data = output.getvalue()
        output.close()
        
        return {'code': 1, 'msg': '生成成功', 'data': excel_data} # 返回二进制数据

    except Exception as e:
        logger.error(f"Excel生成错误: {traceback.format_exc()}")
        return {'code': -1, 'msg': f'生成报表出错: {str(e)}', 'data': None}
    finally:
        s.close()

# --- 接口路由 ---
@any_route('/api/saier/declaration/button/drxjsc', methods=['POST'])
@require_token
async def api_declaration_drxjsc(request):
    """
    修改后的接口：直接返回文件流，触发浏览器下载。
    """
    user = request.current_user
    username = user.username
try:
        print("Request State Attributes:", dir(request.state))
        print("Request State Content:", request.state.__dict__)
        # 1. 获取请求体
        body = await request.json()
        rids = body.get('rids', [])
        
       # --- 在你的接口函数内部 ---
  
        j = await request.json()
         # ... 其他参数提取 ...

         # --- 关键修改：获取用户名 (兼容现有 backend.txt 的风格) ---
         # 方案1: 优先尝试从 request.current_user 获取 (这是 backend.txt 中大多数接口的做法)
    if hasattr(request, 'current_user'):
        current_user_obj = request.current_user
        # 尝试获取 username 属性，如果不存在则尝试字典形式取值
        if hasattr(current_user_obj, 'username'):
            user_name = current_user_obj.username
        elif isinstance(current_user_obj, dict) and 'username' in current_user_obj:
            user_name = current_user_obj['username']
        else:
            # 如果对象里没有 username，尝试取 user_id 或 name 作为兜底
            user_name = str(getattr(current_user_obj, 'user_id', '') or getattr(current_user_obj, 'name', ''))
    else:
          # 方案2: 如果 request 没有 current_user 属性，尝试 backend.txt 中提到的全局 user 对象
        # 注意：这通常依赖于 @require_token 装饰器将用户信息挂载到了其他地方
        global_user = request.state.get('user', None) if hasattr(request.state, 'get') else None
        if global_user and hasattr(global_user, 'username'):
            user_name = global_user.username
        else:
            user_name = 'admin' # 最后的兜底
        
        # --- 后续业务逻辑 ---
        # 将 user_name 传入你的核心函数
     result = await loop.run_in_executor(executor, _your_core_function, param1, param2, user_name)

        
     # 错误处理
        
        # 3. 执行核心逻辑
        result = _generate_excel_logic(rids, current_user)
        
        if result['code'] == 1:
            # --- 关键修改：构建文件下载响应 ---
            # 创建响应对象
            response = web.Response(
                body=result['data'], # 这里是文件的二进制数据
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            # 设置下载头
            response.headers['Content-Disposition'] = 'attachment; filename="优景 INV pack sales.xlsx"'
            return response
        else:
            # 如果生成失败，返回错误信息
            return json_result(result['code'], result['msg'])
except Exception as e:
        error_msg = traceback.format_exc()
        logger.error(error_msg)
        return json_result(-1, f"服务器内部错误: {str(e)}")