# backend: purchase_contract_export.py
# 依赖：SQLAlchemy Session、openpyxl、Pillow、uuid、subprocess（可选：libreoffice/soffice）
# 参考：quotation.py 的路由风格与 Excel 生成方式。:contentReference[oaicite:5]{index=5}
import zipfile
import os,io,re,json
import time
import uuid
from sqlalchemy.sql import text as sql_text
import subprocess,httpx
from venv import logger
from any import *
from .model import *   # 你的 ORM 映射表：cght, cghtsheet, cxgc, sys_user ...
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, pixels_to_points
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.utils import get_column_letter, coordinate_to_tuple 
from openpyxl.cell.text import InlineFont
from PIL import Image
from .__default__ import get_user_path, config 
from datetime import datetime as dt

try:
    import cn2an 
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", 'cn2an'])

# --- 辅助工具函数 ---
def number_to_chinese_currency(num):
    try:        
        # 转换为中文金额
        return cn2an.an2cn(float(num), "rmb")
    except Exception as e:
        logger.error(trace_error())
        return '金额转换错误!!!'
    


def safe_write(ws, coord, value, num_format=None):
    """
    智能穿透写入：如果目标单元格被合并了，自动找到该合并区域的左上角主单元格进行写入。
    完美解决 MergedCell is read-only 报错！
    """
    row, col = coordinate_to_tuple(coord)
    target_cell = ws[coord]
    
    # 检查该坐标是否在某个合并单元格范围内
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            # 如果被合并了，就把目标切换为该区域的左上角真实单元格
            target_cell = ws.cell(row=mr.min_row, column=mr.min_col)
            break
            
    target_cell.value = value
    if num_format:
        target_cell.number_format = num_format



def calc_visual_length(text):
    """转译 Delphi 那几百行的 zsw='A' 遍历：计算文本视觉长度用于行高"""
    if not text: return 0
    en_chars = len(re.findall(r'[a-zA-Z0-9\x00-\xff]', text)) # 半角字符
    cn_chars = len(text) - en_chars # 全角字符
    return cn_chars * 2 + en_chars # 中文算2个单位，英文算1个

def offset_img(img, col, row, x_pad=4, y_pad=25):
    """精确设置图片位置，偏移量以万为单位进行微调吧，具体计算公式太麻烦了
    row column 的索引都是从0开始的，我这里要把图片插入到单元格A17
    """
    p2e = pixels_to_EMU
    h, w = img.height, img.width
    # 图像等比例缩放因子
    resize_factor = 0.8
    w_h_ratio = w/h
    resize_H = int(resize_factor * h)
    resize_W = int(resize_factor * resize_H * w_h_ratio)
    #
    # x_pad = 4
    # y_pad = 25
    # 注意这里的行、列索引从0开始, 所以需要减1
    size = XDRPositiveSize2D(p2e(w), p2e(h))
    marker = AnchorMarker(col=col-1, colOff=pixels_to_EMU(x_pad), row=row-1, rowOff=pixels_to_EMU(y_pad))
    img.anchor = OneCellAnchor(_from=marker, ext=size)

    # 图像在Excel里面的大小
    # image_size_excel = XDRPositiveSize2D(pixels_to_EMU(resize_W), pixels_to_EMU(resize_H))
    ##############################
    # 设置单元格大小，单元格默认宽度单位：字符；高度单位：point(磅)
    # cell_height = int(pixels_to_points(resize_H+10, dpi=96)) #高度上增加10个像素放大单元格
    # cell_width = int(resize_W/8) + 2 # 宽度上增加16（2*8）个像素放大单元格

    # marker = AnchorMarker(col=col, colOff=pixels_to_EMU(x_pad), row=row, rowOff=pixels_to_EMU(y_pad))
    # img.anchor = OneCellAnchor(_from=marker, ext=image_size_excel)

def convert_excel_to_pdf(excel_path, output_dir=None):
    """
    将Excel文件转换为PDF（使用 whale_report.jar）
    
    Args:
        excel_path: Excel文件的完整路径
        output_dir: PDF输出目录，默认与Excel同目录
        
    Returns:
        dict: {'success': bool, 'pdf_path': str, 'error': str}
    """
    try:
        # 确定输出目录
        if output_dir is None:
            output_dir = os.path.dirname(excel_path)
        
        # 生成PDF文件路径
        base_name = os.path.splitext(excel_path)[0]
        pdf_path = base_name + '.pdf'
        
        print(f'========== 开始PDF转换 ==========')
        print(f'源文件: {excel_path}')
        print(f'目标文件: {pdf_path}')
        
        # 检查 Java 和 JAR 文件是否存在
        if not hasattr(config, 'java_path') or not config.java_path:
            return {
                'success': False,
                'error': 'Java路径未配置（config.java_path）'
            }
        
        if not hasattr(config, 'report_jar') or not config.report_jar:
            return {
                'success': False,
                'error': 'whale_report.jar路径未配置（config.report_jar）'
            }
        
        if not os.path.exists(config.report_jar):
            return {
                'success': False,
                'error': f'whale_report.jar文件不存在: {config.report_jar}'
            }
        
        print(f'Java路径: {config.java_path}')
        print(f'JAR路径: {config.report_jar}')
        
        # 构建命令
        # console_run(config.java_path, ['-jar', config.report_jar, 'a', 'b', template, output_file, '2'])
        cmd = [
            config.java_path,
            '-jar',
            config.report_jar,
            'a',  # 占位参数
            'b',  # 占位参数
            excel_path,      # 源Excel文件
            pdf_path,        # 目标PDF文件
            '2'              # 转换类型参数
        ]
        
        print(f'执行命令: {" ".join(cmd)}')
        
        # 执行转换
        result = subprocess.run(
            cmd,
            capture_output=True,
            timeout=120,  # 2分钟超时
            text=True
        )
        
        if result.returncode != 0:
            print(f'转换失败，返回码: {result.returncode}')
            print(f'错误输出: {result.stderr}')
            return {
                'success': False,
                'error': f'PDF转换失败（返回码 {result.returncode}）：{result.stderr}'
            }
        
        # 检查PDF文件是否生成
        if not os.path.exists(pdf_path):
            return {
                'success': False,
                'error': f'PDF文件未生成（期望路径：{pdf_path}）'
            }
        
        file_size = os.path.getsize(pdf_path)
        print(f'✓ PDF转换成功')
        print(f'✓ 文件大小: {file_size:,} 字节')
        print(f'========== PDF转换完成 ==========')
        
        return {
            'success': True,
            'pdf_path': pdf_path,
            'error': None
        }
        
    except subprocess.TimeoutExpired:
        return {
            'success': False,
            'error': 'PDF转换超时（超过2分钟）'
        }
    except Exception as e:
        logger.error(trace_error())
        return {
            'success': False,
            'error': f'PDF转换异常：{str(e)}'
        }


# --- 主路由与业务逻辑   优景采购合同（图）---
@any_route('/api/saier/purchase_contract/export', methods=['POST'])
@require_token
async def api_export_purchase_contract(request):
    
    j = await request.json()
    
    # 2. 提取业务参数 (主表 ID 等，请根据前端实际传的 key 修改，这里兼容了 record_id 和 rid)
    record_id = j.get('rid')
    if not record_id:
        return json_result(-1, "参数缺失：无法获取当前合同的主键 rid")
   
    gs = str(j.get('gs')).strip()
    if not gs:
        return json_result(-1, "参数缺失：公司")
    
    if gs in ('', '1'):
        gs, gs1 = '1', '优景进出口'
    elif gs == '2':
        gs, gs1 = '2', '锐亿进出口'
    else:
        gs1 = gs   # 用户自定义抬头
        gs = '2'   # 强制复用2号模板
    
    pdf = str(j.get('pdf')).strip()
    if not pdf:
        return json_result(-1, "参数缺失：是否生成PDF")


    # 3. 提取当前登录用户名 (WhaleCloud 标准)
    user = request.current_user
    username = user.username
    # 4. 获取数据库 Session (WhaleCloud 标准)
    
    s = Session()
    
    # ================= [ Part 1: 初始化与权限预检 ] =================
    dhsb, dhsb3 = '', ''
    FF, htsm = '', ''
    tmpstr2 = []
    
    if s.execute(sql_text("SELECT 1 FROM cyzglsheet WHERE xm=:xm AND zm='采购合同图PDF'"), {'xm': username}).fetchone(): dhsb = '1'
    if s.execute(sql_text("SELECT 1 FROM cyzglsheet WHERE xm=:xm AND zm='采购合同图PDF签名'"), {'xm': username}).fetchone(): dhsb3 = '1'

    # 对 pdf 参数进行处理      原先shit code 的解读
    if dhsb == '1':
        pdf = '3'  # 有权限直接定死为 3
    else:
        if pdf not in ('2', '3'):
            pdf = '1'
        else:
            if dhsb3 == '1':
                dhsb = '1'  # 权限继承
            pdf = '3'       # 只要选了2或3，一律按3处理

    
    row_ff = s.execute(sql_text("SELECT WB1 FROM zx WHERE ly='反腐专线'")).fetchone()
    if row_ff: FF = str(row_ff.WB1 or '')
    row_htsm = s.execute(sql_text("SELECT nr FROM zx WHERE ly='采购合同签订注意要点'")).fetchone()
    if row_htsm:
        htsm = str(row_htsm.nr or '')
        if htsm != '':
            # 既然不能 showmessage，我们把它追加到提示信息的数组里
            tmpstr2.append(f"【签订注意要点】: {htsm}")      # 关于  showmessage    htsm 只是用了一次
                           
    cxje1, cxjez1, cxjezz = 0.0, 1000000.0, 0.0
    cxsb, sfsh = '', '不需要'
    row_cx = s.execute(sql_text("SELECT cs, sz1 FROM zx WHERE mc='诚信金额'")).fetchone()        
    if row_cx:
        cxje1 = float(row_cx.cs or 0)
        cxjez1 = float(row_cx.sz1 or 0)
        if cxjez1 == 0: cxjez1 = 1000000.0

    sfhs1 = '是'
    if s.execute(sql_text("SELECT * FROM cghtsheet WHERE pid=:pid AND sfhs<>'是'"), {'pid': record_id}).fetchone(): sfhs1 = '否'   

    # ================= [ Part 2: 获取主表记录 (Outermost 1) ] =================
    row_cght = s.execute(sql_text("SELECT * FROM cght WHERE rid=:rid"), {'rid': record_id}).fetchone()
    # 最外层if 语句  Outermost1
    if row_cght:   
        tedi = '是' if str(row_cght.khmc or '') == 'TEDi GmbH & Co. KG' else ''
        cps = 0
        gqsb = 0
        
        current_now = dt.now()
        rq = current_now.strftime('%Y-%m-%d')
        qsrq = f"{current_now.year}-01-01"
        d1 = current_now.replace(hour=0, minute=0, second=0, microsecond=0)
        
        htrq_str = str(row_cght.htrq or '').strip()
        D3 = dt.strptime(htrq_str[:10], '%Y-%m-%d') if htrq_str else None
            
        yjcq_str = str(row_cght.yjcq or '').strip()
        if yjcq_str != '':
            d2 = dt.strptime(yjcq_str[:10], '%Y-%m-%d')
            d = float((d2 - d1).days - 7)
        else:
            gqsb = 1

        # ================= [ Part 3: 是否通过审核 (Outermost 2) ] =================   Outermost2
        if str(row_cght.sfhz) == '通过':
            
            # 提取前端请求参数替代弹窗 (如果没有传，给默认值)
           
            # filename1 = j.get('filename1', os.path.join(TEMP_DIR, 'contract_'))   这是保存地址  不需要取


            # 签名图片 获取   对应  224-  245 code   应该使用的时候调用放后面？？？？  todo 
            img_png_path = ''
            if dhsb == '1':
                row_tpzx = s.execute(sql_text("SELECT tpmc FROM tpzx WHERE cpbh='陈妍科小'")).fetchone()
                if row_tpzx and row_tpzx.tpmc:
                    img_png_path = os.path.join(config.data_upload_path,json.loads(row_tpzx.tpmc)[0].get('src', ''))   # tpmc 中包含 transfer  形如  
              
                else:
                    return json_result(-1, "缺失：签名图片")

            # ================= [ Part 4: 基础字段提取 (Outermost 3) ] =================
        
            gqsb = 1   #  shit param   后面没再使用 
            #  tmpcomcg.SQL.Text := 'select * from cght where number=:number';
            gdry_str = str(row_cght.gdry)
            gdry3 = gdry_str if gdry_str else str(row_cght.cgry )
            gdbm3 = str(row_cght.gdbm) if gdry_str else str(row_cght.cgbm)

            bjdh = str(row_cght.hthm or '')
            ywry = str(row_cght.ywry or '')
            cgry2 = str(row_cght.cgry or '')
            gdry2 = gdry_str
            # hthm2 = str(row_cght.hthm or '')
            jhrq2 = str(row_cght.jhrq or '')
            sccj2 = str(row_cght.sccj or '')
            sccj_id2 = str(row_cght.cs_id or '')
            zygcid2 = str(row_cght.sccj1id or '')
            zygcmc2 = str(row_cght.sccj1 or '')

            cslxr, phone, fax, sjhm, twhm, fkhm, bank1, zh1 = '', '', '', '', '', '', '', ''
            if zygcid2 != '':
                row_zycs = s.execute(sql_text("SELECT * FROM zycs WHERE cs_id=:cs_id"), {'cs_id': zygcid2}).fetchone()
                if row_zycs:
                    cslxr = str(row_zycs.cslxr or '')
                    phone = str(row_zycs.phone or '')
                    fax = str(row_zycs.fax or '')
                    sjhm = str(row_zycs.sjhm or '')
                    # address = str(row_zycs.address or '')
                    twhm = str(row_zycs.twhm or '')
                    fkhm = str(row_zycs.fkhm or '')
                    bank1 = str(row_zycs.bank1 or '')   # 后面用到了 但是目前代码遗漏
                    zh1 = str(row_zycs.zh1 or '')

            htrq2 = str(row_cght.htrq or '')
            jsfs = str(row_cght.jsfs or '')
            
            #  对应  330 - 340 赋值语句   

            szdh = ''
            row_ywry = s.execute(sql_text("SELECT ssdq FROM ywrybiao WHERE yhm=:yhm"), {'yhm': cgry2}).fetchone()
            if row_ywry: szdh = str(row_ywry.ssdq)

            bzyq = '1.出口五层牛皮纸双瓦楞标准纸箱包装。2.要求无钉箱，胶带工字封箱...' if tedi == '是' else str(row_cght.bzyq)
            csmc2, csbh2 = (zygcmc2, zygcid2) if zygcmc2 else (sccj2, sccj_id2)

            row_count = s.execute(sql_text("SELECT COUNT(*) as sl1s FROM cghtsheet WHERE pid=:pid"), {'pid': record_id}).fetchone()
            if row_count: cps = int(row_count.sl1s or 0)

            htjez, htzslz, htzxsz, htzmzz, htzjzz, htztjz = 0.0, 0.0, 0.0, 0.0, 0.0, 0.0

            # ================= [ Part 5: 第一次明细遍历 (金额汇总) ] =================
            items_cght = s.execute(sql_text("SELECT * FROM cghtsheet WHERE pid=:pid"), {'pid': record_id}).fetchall()
            zje1, zxs1 = 0.0, 0.0
            i3 = 0
            
            if items_cght:
                for row_item in items_cght:
                    htjez += float(row_item.zje or 0)
                    htzslz += float(row_item.cgsl or 0)
                    htzxsz += float(row_item.cgxs or 0)
                    htzmzz += float(row_item.zmz or 0)
                    htzjzz += float(row_item.zjz or 0)
                    htztjz += float(row_item.ztj or 0)
                    
                    zje1 += float(row_item.zje or 0)
                    zxs1 += float(row_item.cgxs or 0)
                    i3 += 1
                    # 原代码中很多赋值语句  但是是在循环内，取到数据之后  不用 而是一味的覆盖  一堆shit  所以循环这里删掉


            # ================= [ Part 6: 诚信工厂风控预警 ] =================
            row_htje_sum = s.execute(sql_text("SELECT SUM(htje) as htje1 FROM cght WHERE (htrq>=:htrq) AND (sccj1id=:sccj1id) AND (sfhz='通过')"), {'htrq': qsrq, 'sccj1id': zygcid2}).fetchone()
            if row_htje_sum: cxjezz = float(row_htje_sum.htje1 or 0)

            sfsh = '不需要'
            if (htjez > cxje1) or (cxjezz > cxjez1):    # 对应源代码 大约 554   只要这张合同的金额，或者今年累计跟这家工厂合作的金额，超过了设定的警戒线
                row_cxgc = s.execute(sql_text("SELECT * FROM cxgc WHERE (gcmc=:gcmc) OR (chgc=:chgc)"), {'gcmc': zygcmc2, 'chgc': zygcmc2}).fetchone()
                # 供应商诚信报告（反腐）自动建档与过期拦截系统
                if not row_cxgc:  #系统里从来没有这家工厂的诚信档案
                    now_str, date_str = dt.now().strftime('%Y-%m-%d %H:%M:%S'), dt.now().strftime('%Y-%m-%d')
                    cxgc_rid = get_uuid() # 如果您有封装好的 get_uuid()，可以直接改为 jc_rid = get_uuid()
                    user_rid = user.rid # 获取当前用户的 rid

                    s.execute(sql_text("INSERT INTO cxgc (rid, uid, ctime, gcmc, ywry, ywz, yrrq, sfsh, yrgs, qrrq, chgc, fkry, xgry, cs_id) VALUES (:rid, :uid, :ctime, :gcmc, :ywry, :ywz, :yrrq, '否', '宁波优景进出口有限公司', '', :chgc, '', '', :cs_id)"), {'rid': cxgc_rid, 'uid': user_rid, 'ctime': now_str, 'gcmc': zygcmc2, 'ywry': gdry3, 'ywz': gdbm3, 'yrrq': date_str, 'chgc': zygcmc2, 'cs_id': zygcid2})
                    s.execute(sql_text("UPDATE zycs SET cxhc='否' WHERE cs_id=:cs_id"), {'cs_id': zygcid2})
                    tmpstr2.append(f"工厂名称:{zygcmc2}需提交诚信报告")
                    sfsh, cxsb = '待提供', '1'
                else:  #系统里有这家工厂的诚信档案
                    qrrq_str = str(row_cxgc.qrrq or '').strip()
                    qrrq_date = dt.strptime(qrrq_str[:10], '%Y-%m-%d') if qrrq_str else dt.strptime('1999-01-01', '%Y-%m-%d')
                    if D3 is not None:
                        days_diff = (D3 - qrrq_date).days
                        if days_diff > 365:
                            s.execute(sql_text("UPDATE cxgc SET yrrq=:yrrq, ywry=:ywry, ywz=:ywz, sfsh='否', yrgs='宁波优景进出口有限公司', qrrq='', chgc=:chgc, cs_id=:cs_id WHERE gcmc=:gcmc"), {'gcmc': zygcmc2, 'ywry': gdry3, 'ywz': gdbm3, 'yrrq': dt.now().strftime('%Y-%m-%d'), 'chgc': zygcmc2, 'cs_id': zygcid2})
                            s.execute(sql_text("UPDATE zycs SET cxhc='否' WHERE cs_id=:cs_id"), {'cs_id': zygcid2})
                            tmpstr2.append(f"工厂名称:{zygcmc2}需提交诚信报告")
                            sfsh, cxsb = '待提供', '1'
                        elif days_diff > 330:
                            tmpstr2.append(f"工厂名称:{zygcmc2}还有一个月需提交诚信报告")
                            cxsb = '1'
                if sfsh != '待提供': sfsh = '已提供'

            # ================= [ Part 7: 回写金额与页数计算 ] =================   对应原来的shit code  里面 大约  638 
            try:
                s.execute(sql_text("UPDATE cght SET webpd1='是', htje=:htje, htzsl=:htzsl, htzxs=:htzxs, htzmz=:htzmz, htzjz=:htzjz, htztj=:htztj, sfsh=:sfsh WHERE rid=:rid"), {'rid': record_id, 'htje': htjez, 'htzsl': htzslz, 'htzxs': htzxsz, 'htzmz': htzmzz, 'htzjz': htzjzz, 'htztj': htztjz, 'sfsh': sfsh})
                s.commit()
            except Exception as e:
                s.rolllback()
                logger.error(trace_error())
                return json_result(-1, f"生成失败: {str(e)}")
            finally:
                s.close()
        
            htje, htzxs = str(htjez), str(htzxsz)
            # 采购合同汇总页（模板 2）是每页固定排版 4 条明细。
            ys1 = int(cps / 4)
            iz = int(cps / 4)
            ys3 = ys1 if ys1 == (cps / 4.0) else ys1 + 1
            if ys1 == (cps / 4.0): iz -= 1

            ndz, dz, lxfs, cgzj, zjl, gstt = '', '', '', '', '', ''
            row_htjj_nb = s.execute(sql_text("SELECT * FROM htjj WHERE dq='宁波' AND gstt LIKE :gstt"), {'gstt': f'%{gs1}%'}).fetchone()
            if row_htjj_nb: ndz = str(row_htjj_nb.dz or '')
            # 查当前所在地区的公司高管和抬头配置
            row_htjj_sz = s.execute(sql_text("SELECT * FROM htjj WHERE dq=:dq AND gstt LIKE :gstt"), {'dq': szdh, 'gstt': f'%{gs1}%'}).fetchone()
            if row_htjj_sz:
                dz, lxfs, cgzj, zjl, gstt = str(row_htjj_sz.dz or ''), str(row_htjj_sz.lxfs or ''), str(row_htjj_sz.cgzj or ''), str(row_htjj_sz.zjl or ''), str(row_htjj_sz.gstt or '')



            # 判定 Excel 模板路径   原先 700-821  100+ 代码  如下 不超过10条......    原本 16个模板  根据之前的 变量  tedi  gs    szdh   sffl  
            sffl_str = str(row_cght.sffl or '')
            filename, filename2 = '', ''
            prefix = 'uv' if gs == '1' else 'ry'
            region = 'yw' if szdh == '义乌' else 'nb'
            fl = '-FL' if sffl_str == '是' else ''
            td = 'tedi' if tedi == '是' else ''
            xlsx_path = os.path.join(config.data_upload_path, 'template')
            filename = os.path.join(xlsx_path, f"{prefix}{region}{fl}{td}.xlsx")
            filename2 = os.path.join(xlsx_path, f"{prefix}{region}{fl}1{td}.xlsx")

            # ================= [ Part 8: 覆盖联系信息 ] =================     828 开始 

            sccj, hthm, sccj1 = sccj2, bjdh, zygcmc2
            # 提前初始化所有工厂联系变量，防止后续 Excel 填表时报“变量未定义”错误
            cslxr, phone, fax, sjhm, address = '', '', '', '', ''
            twhm, fkhm, bank1, zh1 = '', '', '', ''

            # 2. 查询主营厂商档案 (对应原代码的 tmpcom.SQL.Text := 'select * from zycs...')
            # 使用 Pythonic 的隐式布尔判断 (if sccj1 相当于 if sccj1 != '')
            if sccj1:
                sccj1id = zygcid2
                
                # 建立查询并安全提取第一条记录
                row_zycs = s.execute(
                    text("SELECT * FROM zycs WHERE cs_id=:cs_id"), 
                    {'cs_id': sccj1id}
                ).fetchone()
                
                # 如果查到了这家工厂的档案，批量赋值     拿来打底  后面若有数据 则覆盖  66666
                if row_zycs:
                    cslxr = str(row_zycs.cslxr or '')
                    phone = str(row_zycs.phone or '')
                    fax = str(row_zycs.fax or '')
                    sjhm = str(row_zycs.sjhm or '')
                    address = str(row_zycs.address or '')
                    twhm = str(row_zycs.twhm or '')
                    fkhm = str(row_zycs.fkhm or '')
                    bank1 = str(row_zycs.bank1 or '')
                    zh1 = str(row_zycs.zh1 or '')
            

            if str(row_cght.lxry or '') != '': cslxr = str(row_cght.lxry or '')
            if str(row_cght.gcdh or '') != '': phone = str(row_cght.gcdh or '')
            if str(row_cght.sjhm or '') != '': sjhm = str(row_cght.sjhm or '')
            if str(row_cght.gccz or '') != '': fax = str(row_cght.gccz or '')

            htrq, jhrq, cgry, gdry = htrq2, jhrq2, cgry2, gdry2

            lxdh, ydhm, cgjl, lxdh1, ydhm1, lxdh2, gdjl, lxdh3, ydhm2 = '', '', '', '', '', '', '', '', ''

            if cgry != '':
                row_cg = s.execute(sql_text("SELECT * FROM ywrybiao WHERE yhm=:yhm"), {'yhm': cgry}).fetchone()
                if row_cg: cgry, lxdh, ydhm, cgjl = str(row_cg.ryxm or ''), str(row_cg.lxdh or ''), str(row_cg.ydhm or ''), str(row_cg.bmjl or '')
            if cgjl != '':
                row_jl = s.execute(sql_text("SELECT * FROM ywrybiao WHERE yhm=:yhm"), {'yhm': cgjl}).fetchone()
                if row_jl: cgjl, lxdh1, ydhm1 = str(row_jl.ryxm or ''), str(row_jl.lxdh or ''), str(row_jl.ydhm or '')
            if gdry != '':
                row_gd = s.execute(sql_text("SELECT * FROM ywrybiao WHERE yhm=:yhm"), {'yhm': gdry}).fetchone()
                if row_gd: 
                    gdry, lxdh2, ydhm2, gdjl, lxdh3 = str(row_gd.ryxm or ''),str(row_gd.lxdh or ''),str(row_gd.ydhm or ''),str(row_gd.bmjl or ''),str(row_gd.jldh or '')
                        

            # ================= [ Part 9~14: 只有一条明细时 (Outermost 4 if) ] =================   对应 927     
            out_name_base = config.tmp_path
            out_xls = out_name_base + '.xlsx'
            #  如果明细里面有一条 执行下面操作 
            if cps == 1:
                if items_cght:
                    row_item = items_cght[0]
                    cgsl, cgjg, zje = float(row_item.cgsl or 0), float(row_item.cgjg or 0), float(row_item.zje or 0)
                    bjhh, zhwbzh, zwdw, cgxs = str(row_item.bjhh or ''), str(row_item.zhwbzh or ''), str(row_item.zwdw or ''), str(row_item.cgxs or '')
                    wxrl = float(row_item.wxrl or 0)
                    zzsl = float(row_item.zzsl or 0)
                    khhh, cpbh, zycpbh = str(row_item.khhh or ''), str(row_item.ksdhh or ''), str(row_item.zycpbh or '')
                    hhbz, wyzd = str(row_item.hhbz or ''), str(row_item.wyzd or '')
                    
                    cpsm = str(row_item.cpsm or '')
                    ywpm = str(row_item.zwpm or '') + '\r\n' + str(row_item.cpgg or '')
                    nbpm = ywpm
                    zhwbgpm = str(row_item.zhwbgpm or '')
                    qtsm1_val = str(row_item.qtsm1 or '')
                    zwz1, dwz, dwz1 = '', 0, 0
                    cpsm_val = str(row_item.cpsm or '')
                    # 这里的if 判断 目的是 后面计算dwz1 
                    if zzsl > 0:
                        zwz1 = '1'
                        
                        # 1. 拼接最终的说明文字
                        qtsm1 = f"{qtsm1_val}                               \r\n请注意开票货源地为：{row_item.hyd or ''}"
                        
                        # =========================================================================
                        # 2. 算半角字符，计算 Excel 标红的偏移量 (几十行 OR 语句)
                        # =========================================================================
                        
                        # 包含所有需要判断的半角字符集 (大写)
                        ascii_chars = "1 234567890ABCDEFGHIJKLMNOPQRSTUVWXYZ,.;:!@#$%*()-_=`?<>/\\\"{}[ ]&"
                        
                        # 计算 dwz
                        zwz_len = len(qtsm1_val)
                        zs3_count = sum(1 for char in qtsm1_val.upper() if char in ascii_chars)
                        dwz = int((zwz_len - zs3_count) / 2) + zs3_count + 30
                        
                        # 计算 dwz1 (用于合并后的文本)
                        text_to_check = f"{cpsm_val}\r\n{qtsm1_val}"
                        zwz1_len = len(text_to_check)
                        zs3_count1 = sum(1 for char in text_to_check.upper() if char in ascii_chars)
                        dwz1 = int((zwz1_len - zs3_count1) / 2) + zs3_count1 + 30

                    else:
                        qtsm1 = qtsm1_val + '                               '
   
                    if str(row_item.krddh ) != '': qtsm1 += '               \r\n客人订单号为:' + str(row_item.krddh or '')

                    # -----------------   后面开始第九部分   打开指定的excel 然后开始写 -----------------
                    try:
                        logger.error(f"打开Excel模板: {filename}")
                        wb = load_workbook(filename)
                        ws = wb.worksheets[0]
                        
                        result_cn = number_to_chinese_currency(htje) if htje else ''
                        if sccj == '': sccj = sccj1 if sccj1 else twhm
                        if phone != '': phone = phone + '\\' + sjhm if sjhm else phone
                        else: phone = sjhm

                        
                        is_yiwu = (szdh == '义乌')
                        is_fl = (sffl_str == '是')
                        # 因为 cps == 1，所以必定是最后一页 (模拟 Delphi 中的 i4 == i3)   i3 总页数  i4 当前页
                        is_last_page = True 
                        
                        cells = {} # 用于统一收集需赋值的坐标

                        # --- 1. 顶部表头、合同基础信息、金额 ---
                        if is_yiwu:
                            # 义乌通用头
                            cells.update({'A6': gstt, 'A7': ndz, 'A8': lxfs, 'A9': dz, 'B12': sccj, 'C11': htrq, 'M11': hthm, 'H11': szdh})
                            if not is_fl:
                                cells.update({'M12': cslxr, 'B13': phone, 'M13': fax})
                                if is_last_page: cells.update({'E18': result_cn, 'O18': htje})
                            else:
                                cells.update({'K12': cslxr, 'B13': phone, 'K13': fax})
                                cells.update({'E18': result_cn, 'O18': htje})
                        else:
                            # 宁波通用头
                            cells.update({'B6': gstt, 'B7': dz, 'B8': lxfs, 'D10': htrq})
                            if not is_fl:
                                cells.update({'C11': sccj, 'N10': hthm, 'I10': szdh, 'L11': cslxr, 'C12': phone, 'L12': fax})
                                if is_last_page: cells.update({'F17': result_cn, 'P17': htje})
                                cells['B24'] = f"四、外箱包装要求：{bzyq}"
                            else:
                                cells.update({'I10': szdh, 'N10': hthm, 'C11': sccj, 'L11': cslxr, 'C12': phone, 'L12': fax})
                                cells.update({'F17': result_cn, 'P17': htje})
                                if (zhwbgpm and zhwbgpm != '无') or (zzsl > 0):
                                    cells['O18'] = f"开票品名: {zhwbgpm}{zzsl}%"
                                else:
                                    cells['O18'] = ''
                                cells['B24'] = f"四、结算方式: {jsfs}"

                        # --- 2. 底部落款、箱数、交货期、收款账号 ---
                        if is_yiwu:
                            cells['C26'] = htzxs if '箱' in htzxs else f"{htzxs}箱"
                            cells['D27'] = jhrq
                            if not is_fl:
                                cells.update({
                                    'A31': f"七、结算方式：{jsfs}",
                                    'A48': f"跟单人员: {gdry}({lxdh2}/{ydhm2})",
                                    'I48': f"采购员: {cgry}({lxdh}/{ydhm})",
                                    'A49': f"采购经理: {cgjl}({lxdh1}/{ydhm1})",
                                    'I49': cgzj, 'A50': zjl, 'I50': FF
                                })
                            else:
                                cells.update({
                                    'A25': f"四、结算方式: {jsfs}",
                                    'A45': f"跟单人员: {gdry}({lxdh2}/{ydhm2})",
                                    'A46': cgzj, 'A47': zjl, 'I47': FF,
                                    'A49': "** 互利共赢  携手并进** 做大靠客人，做强靠厂商 客户和厂商是我们最珍贵的最根本的核心资源 感谢支持！"
                                })
                        else:
                            cells['D25'] = htzxs if '箱' in htzxs else f"{htzxs}箱"
                            cells['E26'] = jhrq
                            if not is_fl:
                                cells.update({
                                    'B30': f"七、结算方式：{jsfs}",
                                    'B47': f"跟单人员: {gdry}({lxdh2}/{ydhm2})",
                                    'J47': f"采购员: {cgry}({lxdh}/{ydhm})",
                                    'B48': f"采购经理: {cgjl}({lxdh1}/{ydhm1})",
                                    'J48': cgzj, 'B49': zjl, 'J49': FF
                                })
                            else:
                                cells.update({
                                    'B43': f"跟单人员: {gdry}({lxdh2}/{ydhm2})",
                                    'B44': cgzj, 'B45': zjl, 'J45': FF,
                                    'B47': "** 互利共赢  携手并进** 做大靠客人，做强靠厂商 客户和厂商是我们最珍贵的最根本的核心资源 感谢支持！"
                                })

                        # --- 3. 解析产品货号 (cphh2) 并赋值坐标 ---
                        if khhh: cphh2 = khhh
                        else: cphh2 = zycpbh if zycpbh else (cpbh if cpbh else bjhh)
                        
                        if is_yiwu: cells['A16'] = cphh2
                        else: cells['B15'] = cphh2

                        # --- 4. 执行所有单元格文本写入 ---
                        for cell_key, cell_val in cells.items():
                            safe_write(ws, cell_key, cell_val)

                    # ⭐⭐⭐ [义乌+宁波双轨制 明细赋值、局部红字、图片处理] ⭐⭐⭐
                        
                        # --- 5. 补充产品明细行赋值与局部红字 ---
                        # 定义红字加粗样式，用于替换原代码的 Characters.Font.ColorIndex := 3
                        red_font = InlineFont(b=True, color='FF0000') 
                        
                        if is_yiwu:
                            # ================= 【义乌版明细】 =================
                            if not is_fl:
                                ws['I16'], ws['J16'], ws['K16'], ws['L16'], ws['N16'], ws['O16'], ws['C16'], ws['E16'] = str(wxrl), zwdw, cgxs, str(cgsl), str(cgjg), str(zje), ywpm, zhwbzh
                                # ws['T17']=cpsm
                                ws['E17'] =cpsm
                                # 局部红字逻辑 (利用 dwz)
                                if zwz1 == '1' and dwz > 0:
                                    split_idx = dwz - 1 # Delphi是1-based，Python是0-based
                                    part1, part2 = qtsm1[:split_idx], qtsm1[split_idx:]
                                    rich_text = CellRichText(part1, TextBlock(red_font, part2))
                                    # ws['U17']= rich_text
                                    ws['N17'] = rich_text
                                else:

                                    # ws['U17']= qtsm1
                                    ws['N17'] = qtsm1
                                    
                                if (zhwbgpm and zhwbgpm != '无') or zzsl > 0:
                                    # ws['U19'] = 
                                    ws['N19'] = f"开票品名: {zhwbgpm}{zzsl}%"
                            else:
                                ws['C16'], ws['E16'], ws['I16'], ws['K16'], ws['M16'], ws['O16'] = ywpm, zhwbzh, str(cgsl), zwdw, str(cgjg), str(zje)
                                
                                # 局部红字逻辑 (利用 dwz1)
                                combined_text = f"{cpsm}\r\n{qtsm1}"
                                if zwz1 == '1' and dwz1 > 0:
                                    split_idx = dwz1 - 1
                                    part1, part2 = combined_text[:split_idx], combined_text[split_idx:]
                                    rich_text = CellRichText(part1, TextBlock(red_font, part2))
                                    ws['E17'], ws['R17'] = rich_text, rich_text
                                else:
                                    ws['E17'], ws['R17'] = combined_text, combined_text
                                    
                                if (zhwbgpm and zhwbgpm != '无') or zzsl > 0:
                                    ws['S19'] = ws['N19'] = f"开票品名: {zhwbgpm}{zzsl}%"
                            
                            # 义乌版行高设定
                            ws.row_dimensions[16].height = max(95, min(400, 95 + len(ywpm) * 2))
                            ws.row_dimensions[17].height = max(95, min(400, 95 + len(qtsm1)))  # 对应pascal 中的   if msexcel.Selection.height < 95 30余行代码
                            ws.row_dimensions[19].height = 16

                        else:
                            # ================= 【宁波版明细】 =================
                            # (摒弃了原 Delphi 中把内容塞进临时格子撑行高又删掉的奇葩操作，直接落子目标格)
                            if not is_fl:
                                ws['D15'], ws['F15'] = nbpm, zhwbzh
                                ws['J15'], ws['K15'], ws['L15'], ws['M15'], ws['O15'], ws['P15'] = str(wxrl), zwdw, cgxs, str(cgsl), str(cgjg), str(zje)
                                ws['F16'] = cpsm
                                
                                # 局部红字逻辑 (利用 dwz) -> 落在 N16
                                if zwz1 == '1' and dwz > 0:
                                    split_idx = dwz - 1
                                    part1, part2 = qtsm1[:split_idx], qtsm1[split_idx:]
                                    ws['N16'] = CellRichText(part1, TextBlock(red_font, part2))
                                else:
                                    ws['N16'] = qtsm1
                                    
                                if zhwbgpm and zhwbgpm != '无':
                                    ws['O18'] = f"开票品名: {zhwbgpm}{zzsl}%"
                            else:
                                ws['D15'], ws['F15'] = nbpm, zhwbzh
                                ws['J15'], ws['L15'], ws['N15'], ws['P15'] = str(cgsl), zwdw, str(cgjg), str(zje)
                                
                                # 局部红字逻辑 (利用 dwz1) -> 落在 F16
                                combined_text = f"{cpsm}\r\n{qtsm1}"
                                if zwz1 == '1' and dwz1 > 0:
                                    split_idx = dwz1 - 1
                                    part1, part2 = combined_text[:split_idx], combined_text[split_idx:]
                                    ws['F16'] = CellRichText(part1, TextBlock(red_font, part2))
                                else:
                                    ws['F16'] = combined_text

                            # 宁波版行高设定
                            ws.row_dimensions[15].height = max(95, min(400, 95 + len(nbpm) * 2))   # 这样替换  是否影响格式   flag
                            ws.row_dimensions[16].height = max(95, min(400, 95 + len(qtsm1)))
                            ws.row_dimensions[18].height = 16

                        # --- 6. 解析 JSON 路径并插入图片 (适配现有项目规范) --- (兼容义乌 A17 与 宁波 B16) ---  对应的是  代码 1567--1708
                        yytp_json_str = ''
                        
                        # 查询获取 JSON 字符串 (已使用 zscp 表和 cpbh 条件)
                        if not is_fl and cpbh != '':
                            row_img = s.execute(sql_text("SELECT yytp FROM zscp WHERE cpbh=:cpbh"), {'cpbh': cpbh}).fetchone()
                            if row_img and row_img.yytp: 
                                yytp_json_str = str(row_img.yytp)
                        elif is_fl and wyzd != '' and hhbz != '':
                            row_img = s.execute(sql_text("SELECT yytp FROM cghtsheet WHERE wyzd=:wyzd"), {'wyzd': wyzd}).fetchone()
                            if row_img and row_img.yytp: 
                                yytp_json_str = str(row_img.yytp)
                                
                        # 解析 JSON 并插入图片    // 代复原？ 原先的delphi代码？ 
                        # --- 6. 还原 Delphi 动态居中算法并插入图片 ---                  
                        '''
                        if yytp_json_str and yytp_json_str not in ('', '[]', 'None'):
                            try:
                                photo_data = json.loads(yytp_json_str)
                                if photo_data and len(photo_data) > 0:
                                    file_path = photo_data[0].get('src', '')
                                    if file_path:
                                        fn = os.path.join(config.data_upload_path, str(file_path))
                                        if os.path.exists(fn):
                                            img = XLImage(fn)
                                            
                                            # 获取图片原始尺寸
                                            orig_w = img.width
                                            orig_h = img.height
                                            
                                            # 1. 定义目标单元格的物理可用尺寸 (对应 Delphi 的 msexcel.Selection.width/height)
                                            # 假设您的模板里，图片存放格子的标准宽为120，高为100
                                            cell_w, cell_h = 120.0, 100.0 
                                            
                                            # 2. 还原 Delphi 里的减 3 魔法 (留出安全边距)
                                            fwidth = cell_w - 3
                                            fheight = cell_h - 3
                                            
                                            # 3. 还原 Delphi 核心缩放比较公式：
                                            # if MsExcel.Selection.ShapeRange.height >= MsExcel.Selection.ShapeRange.width / (fwidth / fheight)
                                            # 交叉相乘避免除数为0： height * fwidth >= width * fheight
                                            if orig_h * fwidth >= orig_w * fheight:
                                                # 图片偏高瘦：以高度为基准缩放
                                                bz2 = fheight / orig_h
                                                img.height = int(fheight)
                                                img.width = int(orig_w * bz2)
                                                
                                                # 还原 bz1 计算水平偏移
                                                bz1 = int((fwidth - img.width) / 2)
                                                
                                                # 还原初始的 +2 魔法：左移 = 2 + 居中偏移，下移 = 2
                                                x_offset = 2 + bz1
                                                y_offset = 2
                                            else:
                                                # 图片偏矮胖：以宽度为基准缩放
                                                bz2 = fwidth / orig_w
                                                img.width = int(fwidth)
                                                img.height = int(orig_h * bz2)
                                                
                                                # 还原 bz1 计算垂直偏移
                                                bz1 = int((fheight - img.height) / 2)
                                                
                                                # 还原初始的 +2 魔法：左移 = 2，下移 = 2 + 居中偏移
                                                x_offset = 2
                                                y_offset = 2 + bz1
                                                
                                            # 4. 召唤您的黑科技定位函数！
                                            col_idx = 1 if is_yiwu else 2  
                                            row_idx = 17 if is_yiwu else 16
                                            
                                            # 完美注入老代码算出来的 x_offset 和 y_offset
                                            offset_img(img, col_idx, row_idx, x_pad=x_offset, y_pad=y_offset)
                                            
                                            # 绝对不可带单元格坐标
                                            ws.add_image(img)
                                            
                            except Exception as img_err:
                                logger.error(f"单页明细图片插入失败: {trace_error()}")
                        '''
       
                        if yytp_json_str and yytp_json_str not in ('', '[]', 'None'):
                            try:
                                photo_data = json.loads(yytp_json_str)
                                if photo_data and len(photo_data) > 0:
                                    file_path = photo_data[0].get('src', '')
                                    if file_path:
                                        fn = os.path.join(config.data_upload_path, str(file_path))
                                        if os.path.exists(fn):
                                            img = XLImage(fn) # 选用 openpyxl 的 Image (兼容您给的 Image_Get)
                                            
                                            # 仿照您的标准：留白 4 像素
                                            img.width = 116  # 预估列宽转换像素后减4 (原代码大概120)
                                            img.height = 96  # 预估行高(100)减4
                                            
                                            # target_cell = 'A17' if is_yiwu else 'B16'
    
                                            col_idx = 1 if is_yiwu else 2
                                            row_idx = 17 if is_yiwu else 16
                                            offset_img(img, col_idx, row_idx, 15, 10)
                                            ws.add_image(img)
                                            
                                            # 否则直接按单元格锚点插入
                                            # ws.add_image(img, target_cell)
                            except Exception as img_err:
                                logger.error(f"单页明细图片插入失败: {trace_error()}")

                        # --- 7. 执行排版格式化 (控制行高、隐藏行、对其方式) ---
                        # 控制收款账户显示与隐藏
                        if sfhs1 != '是':
                            bank_info = f"收款户名:{fkhm}      开户银行:{bank1}      银行账号:{zh1}"
                            if is_yiwu: ws['A20'] = bank_info
                            else: ws['B19'] = bank_info
                        else:
                            row_to_hide = 20 if is_yiwu else 19
                            ws.row_dimensions[row_to_hide].height = 0 
                        
                        # 结算方式内容过长时的行高自适应
                        if len(str(jsfs)) > 94:
                            if is_yiwu: autofit_row = 31 if not is_fl else 25
                            else: autofit_row = 30 if not is_fl else 24
                            ws.row_dimensions[autofit_row].height = 31.5 
                            
                        # 宁波版的强制居中与换行
                        if not is_yiwu:
                            for row_cells in ws['B16:AE16']:
                                for c in row_cells:
                                    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                        # ⭐⭐⭐ [最新高级重构结束] ⭐⭐⭐

                        # --- 8. 插入电子签名 (原 Delphi dhsb = '1' 逻辑) ---
                        if dhsb == '1' and img_png_path and os.path.exists(img_png_path):
                            try:
                                sig_img = XLImage(img_png_path)
                                # 适当缩放签名图片大小 (可以根据实际效果微调)
                                sig_img.width, sig_img.height = 120, 50 
                                
                                if is_yiwu:
                                    if not is_fl:
                                        col_idx, row_idx = 2, 46 # B46
                                    else:
                                        col_idx, row_idx = 3, 43 # C43
                                else:
                                    if not is_fl:
                                        col_idx, row_idx = 4, 45 # D45
                                    else:
                                        col_idx, row_idx = 4, 41 # D41
                                        
                                # 补漏 1：还原 Delphi 里的 left+2, top+13 精确偏移！
                                # (注意：您若没有引入 offset_img，可继续用 ws.add_image(sig_img, f"{get_column_letter(col_idx)}{row_idx}"))
                                offset_img(sig_img, col_idx, row_idx, x_pad=2, y_pad=13)
                                ws.add_image(sig_img)
                            except Exception as e:
                                logger.error(f"插入签名图片失败: {trace_error()}")  

                        # 保存 Excel
                        excel_org_path = config.tmp_path
                        # 提取基础文件名，方便后面复用
                        file_base_name = f"{ywry}-{sccj}{hthm}" 
                        # 拼装 Excel 的绝对物理路径
                        excel_save_path = os.path.join(excel_org_path, f"{file_base_name}.xlsx")
                        wb.save(excel_save_path)
                        
                        if pdf == '1':
                            return json_result(1, '生成采购合同成功', data={
                                'path': f"{file_base_name}.xlsx", # 直接返回文件名即可！     
                                'name': file_base_name
                            })
                        # 现代化的 PDF 转换方案 (无缝平替原系统的 doPDF 打印机)  
                        if pdf in ['2', '3']:
                            # 第一个参数是 Excel文件绝对路径，第二个参数是 输出目录
                            pdf_result = convert_excel_to_pdf(excel_save_path, excel_org_path)
                            
                            if pdf_result.get('success'):
                                pdf_name = f"{file_base_name}.pdf"
                                
                                # 💡 优良传统：过河拆桥！转换成功后删掉 Excel 底稿
                                if os.path.exists(excel_save_path):
                                    os.remove(excel_save_path)
                                    
                                logger.info(f'========== 合同(PDF)导出完成 ==========\n')
                                
                                return json_result(1, '生成采购合同PDF成功', data={
                                    'path': f"{file_base_name}.pdf",
                                    'name': file_base_name
                                })
                            else:
                                logger.error(f'✗ PDF转换失败: {pdf_result.get("error")}')
                                return json_result(-1, f"PDF转换失败: {pdf_result.get('error')}")
                    except Exception as e:
                        logger.error(trace_error())
                        return  json_result(-1, "生成单页明细失败")

            # ================= [ Part 15~16: 多条明细时 (Outermost 4 else) ] =================
            else:
                try:
                    logger.error(f"打开Excel模板: {filename2}")
                    wb = load_workbook(filename2)
                    ws_cover = wb.worksheets[0] # 多页模板的第一页是【封面】
                    
                    if htje != '':
                        result_cn = number_to_chinese_currency(zje1) if zje1 else ''
                    else:
                        result_cn = ''

                    if sccj == '': sccj = sccj1 if sccj1 else twhm
                    if phone != '': phone = phone + '\\' + sjhm if sjhm else phone
                    else: phone = sjhm

                    # ⭐⭐⭐ [多页模板封面 (Cover Page) 赋值映射] ⭐⭐⭐
                    is_yiwu = (szdh == '义乌')
                    is_fl = (sffl_str == '是')
                    cover_cells = {}
                    
                    if is_yiwu:
                        cover_cells.update({'A6': gstt, 'A7': ndz, 'A8': lxfs, 'A9': dz, 'B12': sccj, 'M11': hthm, 'H11': szdh, 'C26': f"{zxs1}箱", 'D27': jhrq, 'A17': f"具体见附页共{ys3}页"})
                        if not is_fl:
                            cover_cells.update({'M12': cslxr, 'B13': phone, 'M13': fax, 'C11': htrq, 'K16': 0, 'L16': 0, 'O16': 0, 'E18': result_cn, 'O18': zje1})
                            cover_cells.update({'A31': f"七、结算方式：{jsfs}", 'A48': f"跟单人员: {gdry}({lxdh2}/{ydhm2})", 'I48': f"采购员: {cgry}({lxdh}/{ydhm})", 'A49': f"采购经理: {cgjl}({lxdh1}/{ydhm1})", 'I49': cgzj, 'A50': zjl, 'I50': FF, 'A25': f"四、外箱包装要求{bzyq}"})
                            if len(str(jsfs)) > 94: ws_cover.row_dimensions[31].height = 31.5
                        else:
                            cover_cells.update({'C11': htrq, 'K12': cslxr, 'B13': phone, 'K13': fax, 'K16': 0, 'L16': 0, 'O16': 0, 'E18': result_cn, 'O18': zje1})
                            cover_cells.update({'A25': f"四、结算方式: {jsfs}", 'A45': f"跟单人员: {gdry}({lxdh2}/{ydhm2})", 'A46': cgzj, 'A47': zjl, 'I47': FF, 'A49': "** 互利共赢  携手并进** 做大靠客人，做强靠厂商 客户和厂商是我们最珍贵的最根本的核心资源 感谢支持！"})
                            if len(str(jsfs)) > 94: ws_cover.row_dimensions[25].height = 31.5
                    else:
                        # 宁波封面
                        if not is_fl:
                            cover_cells.update({'B6': gstt, 'B7': dz, 'B8': lxfs, 'C11': sccj, 'N10': hthm, 'I10': szdh, 'L11': cslxr, 'C12': phone, 'L12': fax, 'D10': htrq, 'K16': 0, 'L16': 0, 'O16': 0, 'F17': result_cn, 'P17': zje1})
                            cover_cells.update({'B24': f"四、外箱包装要求：{bzyq}", 'D25': f"{zxs1}箱", 'E26': jhrq, 'B30': f"七、结算方式：{jsfs}", 'B47': f"跟单人员: {gdry}({lxdh2}/{ydhm2})", 'J47': f"采购员: {cgry}({lxdh}/{ydhm})", 'B48': f"采购经理: {cgjl}({lxdh1}/{ydhm1})", 'J48': cgzj, 'B49': zjl, 'J49': FF, 'B16': f"具体见附页共{ys3}页"})
                            if len(str(jsfs)) > 94: ws_cover.row_dimensions[30].height = 31.5
                        else:
                            cover_cells.update({'D10': htrq, 'I10': szdh, 'N10': hthm, 'C11': sccj, 'L11': cslxr, 'C12': phone, 'L12': fax, 'K16': 0, 'L16': 0, 'O16': 0, 'F17': result_cn, 'P17': zje1})
                            cover_cells.update({'B24': f"四、结算方式: {jsfs}", 'D25': f"{zxs1}箱", 'E26': jhrq, 'B16': f"具体见附页共{ys3}页", 'B43': f"跟单人员: {gdry}({lxdh2}/{ydhm2})", 'B44': cgzj, 'B45': zjl, 'J45': FF, 'B47': "** 互利共赢  携手并进** 做大靠客人，做强靠厂商 客户和厂商是我们最珍贵的最根本的核心资源 感谢支持！"})
                            if (zhwbgpm and zhwbgpm != '无') or zzsl > 0: 
                                cover_cells['O18'] = f"开票品名: {zhwbgpm}{zzsl}%"
                            if len(str(jsfs)) > 94: ws_cover.row_dimensions[24].height = 31.5

                    # 统一写入封面
                    for k_cell, v_cell in cover_cells.items():
                        safe_write(ws_cover, k_cell, v_cell)
                      
                        
                    # 封面收款账户
                    if sfhs1 != '是':
                        bank_info = f"收款户名:{fkhm}      开户银行:{bank1}      银行账号:{zh1}"
                        if is_yiwu: ws_cover['A20'] = bank_info
                        else: ws_cover['B19'] = bank_info
                    else:
                        row_to_hide = 20 if is_yiwu else 19
                        ws_cover.row_dimensions[row_to_hide].height = 0
                        
                    # 封面签名
                    if dhsb == '1' and img_png_path and os.path.exists(img_png_path):
                        try:
                            sig_img = XLImage(img_png_path)
                            # 保留您优秀的尺寸防爆机制
                            sig_img.width, sig_img.height = 120, 50 
                            # 确定坐标并转化为 offset_img 需要的数字索引 (列, 行)
                            if is_yiwu:
                                if not is_fl:
                                    col_idx, row_idx = 2, 46  # B46
                                else:
                                    col_idx, row_idx = 3, 43  # C43
                            else:
                                if not is_fl:
                                    col_idx, row_idx = 4, 45  # D45
                                else:
                                    col_idx, row_idx = 4, 41  # D41
                            
                            offset_img(sig_img, col_idx, row_idx, x_pad=2, y_pad=10)
                            # 绝对不能传 sig_cell 给 add_image，否则会覆盖 offset 偏移量！
                            ws_cover.add_image(sig_img)
                            
                        except Exception as e:
                            logger.error(f"插入封面签名图片失败: {trace_error()}")
                    # ⭐⭐⭐ [封面逻辑结束] ⭐⭐⭐

                    # ================= [ 动态复制明细分页 ] =================
                    # 原系统 filename2 中，sheet[0]是封面，sheet[1]是明细模板
                    detail_template = wb.worksheets[1]
                    for az in range(iz): wb.copy_worksheet(detail_template)
                    
                    k, i4, i5 = 0, 0, 0
                    page_kppm_list = [] # 收集当前页的开票品名
                    red_font = InlineFont(b=True, color='FF0000') 
                    
                    for row_item in items_cght:
                        k += 1
                        i5 += 1
                        if k == 1:
                            i4 += 1
                            ws_detail = wb.worksheets[i4] # 取出新一页的明细 sheet
                            ws_detail['A5'], ws_detail['K5'] = f"我司合同号:{bjdh}", f"★共{ys3}页,第{i4}页"
                            page_kppm_list.clear()
                        
                        # 提取变量
                        khhh = str(row_item.khhh or '')
                        cphh2 = str(row_item.khhh or '') or str(row_item.ksdhh or '') or str(row_item.zycpbh or '') or str(row_item.bjhh or '')
                        ywpm = str(row_item.zwpm or '') + '\r\n' + str(row_item.cpgg or '')
                        zhwbzh = str(row_item.zhwbzh or '')
                        cpsm = str(row_item.cpsm or '')
                        qtsm1_val = str(row_item.qtsm1 or '')
                        zzsl = float(row_item.zzsl or 0)
                        cpbh = str(row_item.ksdhh or '')
                        
                        # ⭐⭐⭐ [附页明细填入] ⭐⭐⭐
                        row_idx = 8 + k
                        ws_detail[f'A{row_idx}'] = cphh2
                        ws_detail[f'A{row_idx}'].number_format = '@'
                        ws_detail[f'C{row_idx}'] = ywpm
                        ws_detail[f'D{row_idx}'] = zhwbzh
                        ws_detail[f'E{row_idx}'] = str(row_item.wxrl or 0)
                        ws_detail[f'F{row_idx}'] = str(row_item.zwdw or '')
                        ws_detail[f'G{row_idx}'] = str(row_item.cgxs or '')
                        ws_detail[f'H{row_idx}'] = str(row_item.cgsl or 0)
                        ws_detail[f'I{row_idx}'] = str(row_item.cgjg or 0)
                        ws_detail[f'J{row_idx}'] = float(row_item.zje or 0)
                        


                            # --- 3. 解析产品货号 (cphh2) 并赋值坐标 ---
                     
                        # 收集开票品名
                        zhwbgpm = str(row_item.zhwbgpm or '')    
                        if zhwbgpm and zhwbgpm not in page_kppm_list:
                            page_kppm_list.append(zhwbgpm)
                        
                        if khhh: cphh2 = khhh
                        else: cphh2 = zycpbh if zycpbh else (cpbh if cpbh else bjhh)

                        # ⭐⭐⭐ [附页局部红字逻辑] ⭐⭐⭐
                       # 1. 提前准备好客人订单号的后缀 (全局共用)
                        krddh_val = str(row_item.krddh or '')
                        krddh_suffix = f"                              \r\n客人订单号为:{krddh_val}" if krddh_val else ""

                        if zzsl > 0:
                            # 严格还原老代码：富文本截断点 dwz1 只计算最原始的 cpsm 和 qtsm1
                            ascii_chars = "1 234567890ABCDEFGHIJKLMNOPQRSTUVWXYZ,.;:!@#$%*()-_=`?<>/\\{}[+]&"
                            calc_text = f"{cpsm}\r\n{qtsm1_val}"
                            zs3_count1 = sum(1 for char in calc_text.upper() if char in ascii_chars)
                            dwz1 = int((len(calc_text) - zs3_count1) / 2) + zs3_count1 + 30
                            
                            # 2. 组装红字后缀：货源地 + 客人订单号 (完美契合老代码的追加顺序)
                            hyd_suffix = f"                              \r\n请注意开票货源地为：{row_item.hyd or ''}"
                            total_suffix = hyd_suffix + krddh_suffix
                            
                            # 3. 富文本切割与拼装
                            if dwz1 >= len(calc_text):
                                part1 = calc_text
                                part2 = total_suffix
                            else:
                                part1 = calc_text[:dwz1]
                                part2 = calc_text[dwz1:] + total_suffix
                            
                            # 写入富文本 (前面黑字，后面红字)
                            ws_detail[f'K{row_idx}'] = CellRichText(part1, TextBlock(red_font, part2))
                            
                        else:
                            # zzsl <= 0 的情况：没有货源地，只有原始说明 + 客人订单号
                            base_text = f"{cpsm}\r\n{qtsm1_val}                              "
                            
                            # 直接拼接全局的订单号后缀写入
                            ws_detail[f'K{row_idx}'] = base_text + krddh_suffix
                        
                        # 明细行格式与行高保护
                        ws_detail.row_dimensions[row_idx].height = max(95, min(400, 95 + len(ywpm)*2))  # 对应2609--2632

                        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        for row_cells in ws_detail[f'A{row_idx}:AE{row_idx}']:
                            for cell in row_cells:
                                cell.alignment = center_alignment

                        # ⭐⭐⭐ [附页零 IO 图片插入] ⭐⭐⭐     待修改   3-23 - 15：06 
                        hhbz, wyzd = str(row_item.hhbz or ''), str(row_item.wyzd or '')
                        # ⭐⭐⭐ [附页图片插入 (基于 JSON 路径)] ⭐⭐⭐
                        yytp_json_str = ''
                        if not is_fl and cpbh != '':
                            row_img = s.execute(sql_text("SELECT yytp FROM zscp WHERE cpbh=:cpbh"), {'cpbh': cpbh}).fetchone()
                            if row_img and row_img.yytp: 
                                yytp_json_str = str(row_img.yytp)
                        elif is_fl and wyzd != '' and hhbz != '':
                            row_img = s.execute(sql_text("SELECT yytp FROM cghtsheet WHERE wyzd=:wyzd"), {'wyzd': wyzd}).fetchone()
                            if row_img and row_img.yytp: 
                                yytp_json_str = str(row_img.yytp)
                                
                        if yytp_json_str and yytp_json_str not in ('', '[]', 'None'):
                            try:
                                photo_data = json.loads(yytp_json_str)
                                if photo_data and len(photo_data) > 0:
                                    file_path = photo_data[0].get('src', '')
                                    if file_path:
                                        fn = os.path.join(config.data_upload_path, str(file_path))
                                        if os.path.exists(fn):
                                            img = XLImage(fn)
                                            
                                            # 控制尺寸
                                            img.width = 116 
                                            img.height = 96
                                            
                                            # 附页的图片全部插在 B 列对应的行
                                            target_cell = f'B{row_idx}'
                                            col_idx = 2
                                            offset_img(img, col_idx, row_idx, x_pad=15, y_pad=10)
                                            
                                            # ⭐️ 3. 直接插入
                                            ws_detail.add_image(img)
                            except Exception as img_err:
                                logger.error(f"附页明细图片插入失败: {trace_error()}")
                        # ⭐⭐⭐ [附页表尾处理 (排满 4 条或最后一条)] ⭐⭐⭐
                        if k == 4 or i5 == cps:
                            
                            # 1. 一行代码降维合并开票品名
                            kppm_str = ";".join([f"开票品名: {pm}" for pm in page_kppm_list])
                            if kppm_str: 
                                ws_detail['I14'] = kppm_str
                                # 对应原代码的 AutoFit，利用 openpyxl 的换行属性让其自适应
                                ws_detail['I14'].alignment = Alignment(wrap_text=True)
                            
                            # 2. 还原“奇葩覆盖”的人员排版
                            if not is_fl:
                                ws_detail['A17'] = f"采购员: {cgry}({lxdh}/{ydhm})"
                                ws_detail['D17'] = f"采购经理: {cgjl}({lxdh1}/{ydhm1})"
                            else:
                                ws_detail['A17'] = f"跟单人员: {gdry}({lxdh2}/{ydhm2})"
                                
                            ws_detail['A18'], ws_detail['A19'], ws_detail['D19'] = cgzj, zjl, FF
                            
                            # 3. 附页电子签名 (找回丢失的 +2 偏移)
                            if dhsb == '1' and img_png_path and os.path.exists(img_png_path):
                                try:
                                    sig_img = XLImage(img_png_path)
                                    sig_img.width, sig_img.height = 120, 50 
                                    
                                    # 完美还原 Delphi: Range['b16'].left + 2, top + 0
                                    offset_img(sig_img, col_idx=2, row_idx=16, x_pad=2, y_pad=0) 
                                    ws_detail.add_image(sig_img)
                                except Exception as e:
                                    logger.error(f"附页签名插入失败: {trace_error()}")

                            # 4. 彻底还原：清空品名列表，重置当前页商品计数器，准备排下一页
                            page_kppm_list.clear() 
                            k = 0
                    
                    # 导出与拆分逻辑
                    excel_org_path= config.tmp_path
                    file_base_name= f"{ywry}-{sccj}{hthm}"
                    excel_save_path= os.path.join(excel_org_path,f"{file_base_name}.xlsx")
               

                    wb.save(excel_save_path) 
                    if pdf == '1':
                        return  json_result(1, '生成采购合同成功', data= {
                            'path': f"{file_base_name}.xlsx", 
                            'name': file_base_name}
                        )
                        
                    if pdf in ['2', '3']:    
                        generated_pdfs = []  
                        for ts in range(1, int(ys3) + 2):
                            # 1. 确定单页的临时 XLSX 绝对路径
                            if ts == 1:
                                out_temp = os.path.join(excel_org_path, f"{file_base_name}_cover_temp.xlsx")
                            else:
                                out_temp = os.path.join(excel_org_path, f"{file_base_name}副本{ts - 1}.xlsx")
                            
                            # 2. 内存级安全拆分
                            temp_wb = load_workbook(excel_save_path)
                            target_sheet = temp_wb.worksheets[ts - 1]
                            
                            # 建议采用这种最稳妥的写法
                            for sht in list(temp_wb.worksheets):
                                if sht is not target_sheet:
                                    temp_wb.remove(sht)
                            
                            temp_wb.save(out_temp)
                            temp_wb.close()
                            
                           # 3. 统一调用项目自带的 Java/JAR 转换引擎！
                            pdf_result = convert_excel_to_pdf(out_temp, excel_org_path)
                            
                            # 如果转换失败，记录日志并跳过/退出
                            if not pdf_result.get('success'):
                                logger.error(f"⚠️ 多页PDF转换失败(第{ts}页): {pdf_result.get('error')}")
                                # 可以根据业务需求选择 continue 跳过，或者直接返回报错
                                continue 
                                
                            generated_pdfs.append(pdf_result.get('pdf_path'))

                            
                            try:
                                if os.path.exists(out_temp):
                                    os.remove(out_temp)
                            except Exception:
                                logger.warning(f"删除临时 xlsx 失败: {out_temp}")
                            # 4. 封面重命名还原
                            # if ts == 1 and pdf_generated and os.path.exists(pdf_generated):
                            #     pdf_target = os.path.join(excel_org_path, f"{file_base_name}.pdf")
                            #     if os.path.exists(pdf_target):
                            #         os.remove(pdf_target) 
                            #     os.rename(pdf_generated, pdf_target)

                        if generated_pdfs:
                            zip_path = os.path.join(excel_org_path, f"{file_base_name}.zip")
                            try:
                                with zipfile.ZipFile(zip_path, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
                                    for p in generated_pdfs:
                                        if p and os.path.exists(p):
                                            zf.write(p, os.path.basename(p))
                                            try:
                                                os.remove(p)
                                            except Exception:
                                                logger.warning(f"删除临时 pdf 失败: {p}")
                            except Exception as e:
                                logger.error(f"打包 ZIP 失败: {trace_error()}")
                                # 若打包失败，退回到原来的行为：将第一份 PDF 作为 target_pdf（保持兼容）
                                try:
                                    cover_pdf = generated_pdfs[0]
                                    target_pdf = os.path.join(excel_org_path, f"{file_base_name}.pdf")
                                    if os.path.exists(target_pdf):
                                        os.remove(target_pdf)
                                    os.rename(cover_pdf, target_pdf)
                                    return json_result(1, "生成主采购合同PDF成功", data={'path': f"{file_base_name}.pdf", 'name': file_base_name})
                                except Exception:
                                    return json_result(-1, "生成PDF/ZIP失败")
        

                        if os.path.exists(excel_save_path): 
                            os.remove(excel_save_path)
                        
                            
                        # cover_temp_xls = os.path.join(excel_org_path, f"{file_base_name}_cover_temp.xlsx")
                        # if os.path.exists(cover_temp_xls): 
                        #     os.remove(cover_temp_xls)
                            
                        # for ts in range(2, int(ys3) + 2):
                        #     temp_fuben_xls = os.path.join(excel_org_path, f"{file_base_name}副本{ts - 1}.xlsx")
                        #     if os.path.exists(temp_fuben_xls): 
                        #         os.remove(temp_fuben_xls)

                        logger.info(f'========== 多页合同(PDF)拆分导出完成 ==========\n')

                        # return json_result(1, "生成采购合同PDF成功",
                        #     data ={
                        #         'path': f"{file_base_name}.xlsx", # 直接返回文件名即可！
                        #         'name': file_base_name
                        #     }
                        # )
                        return json_result(1, "生成采购合同ZIP成功", data={'path': f"{file_base_name}.zip", 'name': file_base_name})

                 
                except Exception as e:
                    logger.error(f"多页模板生成失败: {trace_error()}")
                    return json_result(-1, "多页模板生成失败")

            
    # 诚信报告预警提示落地
    warning_msg = ""
    warning_path=""
    if cxsb == '1':
        warning_path = os.path.join(config.data_upload_path, f"{dt.now().strftime('%Y-%m-%d')}_诚信报告.txt") 
        with open(warning_path, 'w', encoding='utf-8') as f: 
            f.write("\n".join(tmpstr2))
        warning_msg = f"有需提交诚信报告的工厂，已生成日志: {warning_path}"

    return json_result(1, "采购合同生成完毕！",  {"warning": warning_msg, "warning_path": warning_path})



# 上传到领星采购单

@any_route('/api/saier/purchase_contract/upload_to_lingxing', methods=['POST'])
@require_token
async def purchase_contract_upload_to_lingxing(request):
    s = Session()
    user_name = request.current_user.username
    j = await request.json()
    record_id = j.get('rid') # 获取当前单据的 id
    if not record_id:
        return json_result(-1, "缺少rid编号")

    try:
        # 1. 数据库查询逻辑 (对应 Delphi 的 tmpcom SQL)
        # 条件：业务员/采购员/归档员为当前用户，且合同号含 AMZ，且审核通过
        sql = sql_text("""
            SELECT hthm FROM cght 
            WHERE ((ywry = :u) OR (cgry = :u) OR (gdry = :u)) 
              AND (rid = :record_id) 
              AND (hthm LIKE '%AMZ%') 
              AND (sfhz = '通过')
            LIMIT 1
        """)
        
        result = s.execute(sql, {'u': user_name, 'record_id': record_id}).fetchone()

        if not result:
            return json_result(-1, "未找到符合条件的采购合同")

        hthm = result.hthm # 获取合同名

        # 2. 发送 POST 请求到领星 (对应 Delphi 的 SendPost)
        lx_url = "http://221.12.55.50:8413/lx/cght/"
        try:
            # 根据 Delphi 的 O.Send(params)，这里直接发送字符串内容
            body = {'hthm': hthm}
            lx_response = httpx.post(lx_url, json=body, timeout=200)
            logger.error(f"领星响应: {lx_response}")
            if lx_response.status_code == 200:
                return json_result(1,lx_response.text)
            else:
                return json_result(-1, f"上传失败，状态码：{lx_response.status_code}，响应内容：{lx_response.text}")
                        # 3. 结果判断 (对应 Delphi 的 POS('success', returndto) > 0)
            # if 'success' in return_dto.lower():
            #     return json_result(1, "完成")
            # else:
            #     return json_result(0, f"失败一条, 采购单号: {hthm}, 报错内容为: {return_dto}")

        except httpx.HTTPStatusError as e:
            return json_result(-1, f"领星响应错误：{e}")

    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f"系统错误: {str(e)}")
    finally:
        s.close()





# 爱马仕采购出货统计

@any_route('/api/saier/hermes/stats/export', methods=['POST'])
@require_token
async def api_hermes_purchase_export(request):

    j = await request.json()
    # 对应 Delphi 中的 da1, da2
    start_date = j.get('start_date')
    end_date = j.get('end_date')
    
    if not start_date or not end_date:
        return json_result(-1, "请选择统计的起始日期和结束日期")

    user_name = request.current_user.username
    s = Session()
    try:
        # 1. 权限校验 (对照 Delphi: select * from cyzglsheet where xm=:mc and zm="爱马士采购出货统计产品")
        # 直接使用字符串，不用 text()

        #侯柳红
        perm_sql = """
            SELECT 1 FROM cyzglsheet 
            WHERE xm = :name AND zm = '爱马士采购出货统计产品'
            LIMIT 1
        """
        perm_check = s.execute(perm_sql, {'name': user_name}).fetchone()
        # perm_check = s.execute(perm_sql, {'u': user_name}).fetchone()
        if not perm_check:
            return json_result(-1, "对不起，您没有‘爱马士采购出货统计’的导出权限")

        # 2. 核心业务逻辑 SQL (严格还原 Delphi 的聚合逻辑)
        # 采用 Left Join 关联出货明细，计算下单数、出货数、剩余数
        query_sql = """
        SELECT 
            cghtsheet.bjhh, 
            cghtsheet.khhh, 
            SUM(cghtsheet.cgsl) AS sl1, 
            COALESCE(cy.chslz, 0) AS chslz1, 
            SUM(cghtsheet.cgsl) - COALESCE(cy.chslz, 0) AS sysl 
        FROM cghtsheet 
        LEFT OUTER JOIN (
            SELECT SUM(chsl) AS chslz, zycpbh, khhh 
            FROM cymxsheet 
            WHERE (fksb = '是') AND (cght LIKE :amz) 
            AND (htrq >= :da1) AND (htrq <= :da2) 
            AND (htrq IS NOT NULL) 
            GROUP BY zycpbh, khhh
        ) AS cy ON cy.zycpbh = cghtsheet.bjhh AND cy.khhh = cghtsheet.khhh 
        WHERE (cghtsheet.hthm LIKE :amz) 
        AND (cghtsheet.htrq >= :da1) AND (cghtsheet.htrq <= :da2) 
        AND (cghtsheet.htrq IS NOT NULL) 
        AND (cghtsheet.sfhz = '通过') 
        GROUP BY cghtsheet.bjhh, cghtsheet.khhh, cy.chslz 
        ORDER BY sysl, cghtsheet.bjhh, cghtsheet.khhh
        """
        
        # 注意：在原生 SQL 中使用 LIKE 时，有些环境需要 %% 来转义 %
        records = s.execute(query_sql, {'da1': start_date, 'da2': end_date, 'amz': '%AMZ%'}).fetchall()

        if not records:
            return json_result(-1, f"在 {start_date} 至 {end_date} 期间未找到符合条件(AMZ)的数据")

        # 3. 生成 Excel 文件 (替代 Delphi 的 msexcel variant 操作)
        wb = Workbook()
        ws = wb.active
        ws.title = "爱马仕采购出货统计"
        
        # 写入表头 (对应 Delphi 逻辑)
        headers = ['产品编号', '客人货号', '下单数量', '出货数量', '剩余数量']
        ws.append(headers)
        
        # 设置基础列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

        # 循环写入数据行
        for r in records:
            ws.append([
                r.bjhh, 
                r.khhh, 
                float(r.sl1 or 0), 
                float(r.chslz1 or 0), 
                float(r.sysl or 0)
            ])


        # 4. 保存并返回
        s_path = config.tmp_path
        report_rid = get_uuid()
        filename=str(report_rid)+'.xlsx'
        wb.save(s_path + '/'+ str(report_rid)+'.xlsx')
        return json_result(1, '导出成功', filename)
    

    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f"执行统计失败: {str(e)}")
    finally:
        s.close()




# 出运计划  --  电商询价表导出
@any_route('/api/saier/shipping/e_commerce_inquiry/export', methods=['POST'])
@require_token
async def api_saier_ecommerce_inquiry_export(request):
    j = await request.json()
    s = Session()
    try:
        record_id = j.get('rid') # 对应 Delphi 的 self.getnumber
        if not record_id:
            return json_result(-1, "缺少出运计划记录 ID")

        # 1. 查询主表 (出运计划物流信息)
        sql_master = sql_text("""
            SELECT fphm, hhrq, cyka, hglx, mdck, xshj2, mzhj, tjhj 
            FROM chyjh 
            WHERE rid=:id
        """)
        master = s.execute(sql_master, {"id": record_id}).fetchone()
        
        if not master:
            return json_result(-1, "未找到该出运计划数据")

        # 2. 查询明细表 (用于去重聚合品名)
        sql_detail = sql_text("""
            SELECT kpgc, zhwbgpm, djpmy 
            FROM chyjhsheet 
            WHERE pid=:id
        """)
        details = s.execute(sql_detail, {"id": record_id}).fetchall()

        # 3. 核心逻辑：利用 Python Set 进行品名去重聚合 (替代 Delphi TStringList)
        seen_items = set()
        kpgc_list, zh_list, en_list = [], [], []

        for row in details:
            # 处理 None 值
            kpgc = row.kpgc or ''
            zh_name = row.zhwbgpm or ''
            en_name = row.djpmy or ''
            
            # 如果三者都为空，则跳过
            if not kpgc and not zh_name and not en_name:
                continue
                
            # 使用 Tuple 作为去重唯一键
            unique_key = (kpgc, zh_name, en_name)
            
            if unique_key not in seen_items:
                seen_items.add(unique_key)
                kpgc_list.append(kpgc)
                zh_list.append(zh_name)
                en_list.append(en_name)

        wb = Workbook()
        ws = wb.active
        ws.title = "电商询价表"
        headers = [
            '发票号码', '货好日期', '开票工厂', '出运口岸', '柜型', 
            '目的仓库', '中文报关品名', '英文品名', '总箱数', '总毛重', 
            '备注(预留)', '总体积'
        ]
        ws.append(headers)
        # 表头样式美化 (加粗并居中)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        # 5. 填充 Excel 数据 (Delphi 逻辑中 i=0, i+1 即第 2 行)
        row_num = 2 
        
        ws[f'A{row_num}'] = str(master.fphm).upper() if master.fphm else ''
        ws[f'B{row_num}'] = master.hhrq or ''
        ws[f'D{row_num}'] = master.cyka or ''
        # LCL 转拼箱逻辑
        ws[f'E{row_num}'] = '拼箱' if str(master.hglx).strip().upper() == 'LCL' else '整柜'
        ws[f'F{row_num}'] = master.mdck or ''
        ws[f'I{row_num}'] = master.xshj2 or ''
        ws[f'J{row_num}'] = master.mzhj or ''
        ws[f'L{row_num}'] = master.tjhj or ''

        # 填充去重聚合后的品名数据 (使用 \n 拼接)
        ws[f'C{row_num}'] = '\n'.join(kpgc_list)
        ws[f'G{row_num}'] = '\n'.join(zh_list)
        ws[f'H{row_num}'] = '\n'.join(en_list)

        # 针对包含多行文本的单元格，开启自动换行
        for col in ['C', 'G', 'H']:
            ws[f'{col}{row_num}'].alignment = Alignment(wrapText=True, vertical='center')
        # 顺手设置一下体验更好的基础列宽
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 18
        ws.column_dimensions['L'].width = 12
        ws.column_dimensions['C'].width = 30 # 工厂名通常较长
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 30 # 英文名通常较长
        # 6. 生成并保存临时文件
    
        s_path = config.tmp_path
        report_rid = get_uuid()
        filename=str(report_rid)+'.xlsx'
        wb.save(s_path + '/'+ str(report_rid)+'.xlsx')

        return json_result(1, "电商询价表生成成功", filename)

    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f"生成电商询价表失败: {str(e)}")
    finally:
        s.close()



# 出运计划 -- lv出运清单 


# --- 辅助函数：纯内存等比缩放插入图片 ---
def process_and_insert_image_from_file(ws, cell_coordinate, file_path, max_width=100, max_height=70):
    if not os.path.exists(file_path):
        return
    try:
        # 直接通过硬盘路径打开图片
        img = Image.open(file_path)
        orig_w, orig_h = img.size
        
        # 等比缩放计算
        ratio = min(max_width / orig_w, max_height / orig_h)
        new_w, new_h = int(orig_w * ratio), int(orig_h * ratio)
        
        # 转换为 Openpyxl 图像对象
        xl_img = XLImage(file_path)
        xl_img.width, xl_img.height = new_w, new_h
        ws.add_image(xl_img, cell_coordinate)
    except Exception as e:
        logger.warning(f"图片处理失败 {file_path}: {e}")


@any_route('/api/saier/shipping/lv/export', methods=['POST'])
@require_token
async def api_saier_shipping_lv_export(request):
    """
    LV 出运清单导出 - 重构版
    """
    j = await request.json()
    s = Session()
    try:
        record_id = j.get('rid')
        if not record_id:
            return json_result(-1, "缺少出运计划记录 ID")

        # 1. 查询主表外销发票
        master = s.execute(sql_text("SELECT wxfp FROM chyjh  WHERE rid=:id"), {"id": record_id}).fetchone()
        if not master:
            return json_result(-1, "未找到该出运数据")
        wxfp_val = str(master.wxfp).upper() if master.wxfp else ""

        # 2. 查询出运明细
        details = s.execute(sql_text("SELECT * FROM chyjhsheet  WHERE pid=:id"), {"id": record_id}).fetchall()
        
        # 3. 批量预加载关联数据 (干掉 Delphi 的 N+1 循环查询)
        bjhh_list = list(set([d.bjhh for d in details if d.bjhh]))
        
        # 将关联数据查出并存入字典，实现 O(1) 匹配
        zscp_map, wypm_map, pic_map = {}, {}, {}
        if bjhh_list:
            # 产品资料
            zscp_rows = s.execute(sql_text(
                "SELECT cpbh, cpfl, krtm, lvkrtm, ggwy, cpgg, czwy, chpkzh FROM zscp  WHERE cpbh IN :ids"
            ), {"ids": tuple(bjhh_list)}).fetchall()
            zscp_map = {r.cpbh: r for r in zscp_rows}

            # 外语品名   该 表 不在数据库里面了 
            # wypm_rows = s.execute(sql_text(
            #     "SELECT cpbh, ggly, czly FROM wypm  WHERE cpbh IN :ids"
            # ), {"ids": tuple(bjhh_list)}).fetchall()
            # wypm_map = {r.cpbh: r for r in wypm_rows}
            wypm_map = {}
            # 图片
            pic_rows = s.execute(sql_text(
                "SELECT cpbh, yytp FROM cjcp  WHERE cpbh IN :ids"
            ), {"ids": tuple(bjhh_list)}).fetchall()
            # 将产品编号(cpbh)作为 key，JSON字符串(yytp)作为 value
            pic_map = {r.cpbh: r.yytp for r in pic_rows}

        # 4. 从零构建 Excel (不再依赖需要手动维护的 LV出运清单.xls 模板)
        wb = Workbook()
        ws = wb.active
        ws.title = "LV出运清单"

        # 设置表头
        ws['C1'] = wxfp_val
        ws['C1'].font = Font(bold=True, size=14)
        
        # 精准还原 Delphi 输出的列位 (注意 Delphi 中跳过了 F 列)
        headers = [
            '产品分类', '外销合同', '图片', '客人条码', 'LV条码', '', '产品编号', '规格外语', 
            '产品规格', '规格来源', '材质外语', '材质来源', '外箱容量', '单位', '内盒外箱', 
            '内盒容量', '出货箱数', '外箱毛重', '外箱净重', '空重', '外箱体积', '总毛重', 
            '总净重', '总数量', '总体积', '海关编码', '预计船期'
        ]
        ws.append([]) # 空出第2行
        ws.append(headers) # 第3行写表头

        # 设置表头样式及列宽
        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            ws[f"{col_letter}3"].font = Font(bold=True)
            ws[f"{col_letter}3"].alignment = Alignment(horizontal='center')
            ws.column_dimensions[col_letter].width = 20
        ws.column_dimensions['I'].width = 60 # 图片列宽一点
        ws.column_dimensions['K'].width = 80 # 图片列宽一点
        # 5. 循环填充数据
        start_row = 4
        for i, d in enumerate(details):
            row_num = start_row + i
            bjhh = d.bjhh
            z_info = zscp_map.get(bjhh)
            # w_info = wypm_map.get(bjhh)
            
            # 设置行高 (Delphi 中设置的 RowHeight := 24, 这里为了图片显示更好可适当加大)
            ws.row_dimensions[row_num].height = 60

            # 填充数据 (精准对应原系统字母列)
            ws[f'A{row_num}'] = z_info.cpfl if z_info else ''
            ws[f'B{row_num}'] = d.wxht or ''
            # C列为图片，最后处理
            ws[f'D{row_num}'] = z_info.krtm if z_info else ''
            ws[f'E{row_num}'] = z_info.lvkrtm if z_info else ''
            # F列跳过
            ws[f'G{row_num}'] = z_info.cpbh if z_info else ''
            ws[f'H{row_num}'] = z_info.ggwy if z_info else ''
            ws[f'I{row_num}'] = z_info.cpgg if z_info else ''
            # ws[f'J{row_num}'] = str(w_info.ggly) if w_info and w_info.ggly else ''
            ws[f'J{row_num}'] = ''  # 原规格来源 (ggly)，现已废弃留空
            ws[f'K{row_num}'] = z_info.czwy if z_info else ''
            # ws[f'L{row_num}'] = str(w_info.czly) if w_info and w_info.czly else ''
            ws[f'L{row_num}'] = ''  # 原材质来源 (czly)，现已废弃留空
            
            ws[f'M{row_num}'] = float(d.wxrl or 0)
            ws[f'N{row_num}'] = d.jldw or ''
            ws[f'O{row_num}'] = int(d.nhwx or 0)
            ws[f'P{row_num}'] = float(d.nhrl or 0)
            ws[f'Q{row_num}'] = int(d.chxs2 or 0) # 箱数
            ws[f'R{row_num}'] = float(d.wxmz or 0) # 单毛重
            ws[f'S{row_num}'] = float(d.wxjz or 0) # 单净重
            ws[f'T{row_num}'] = float(z_info.chpkzh or 0) if z_info else 0
            ws[f'U{row_num}'] = float(d.wxtj or 0) # 单体积
            
            # 写入原生 Excel 计算公式 (完全复刻 Delphi 的 Formula 特性)
            ws[f'V{row_num}'] = f"=Q{row_num}*R{row_num}" # 总毛重
            ws[f'W{row_num}'] = f"=Q{row_num}*S{row_num}" # 总净重
            ws[f'X{row_num}'] = f"=Q{row_num}*M{row_num}" # 总数量
            ws[f'Y{row_num}'] = f"=Q{row_num}*U{row_num}" # 总体积
            
            ws[f'Z{row_num}'] = d.hgbm or ''
            ws[f'AA{row_num}'] = d.yjcq or ''

            # 处理图片
            yytp_json_str = pic_map.get(bjhh)
            # 判空及防止空 JSON 数组
            if yytp_json_str and yytp_json_str.strip() not in ['', '[]']:
                try:
                    photo_list = json.loads(yytp_json_str)
                    if isinstance(photo_list, list) and len(photo_list) > 0:
                        file_path = photo_list[0].get('src')
                        if file_path:
                            # 核心路由：判断 src 中是否包含 'upload'
                            if 'upload' in file_path:
                                fn = file_path # 视为绝对路径直接使用
                            else:
                                # 相对路径，进行前缀拼接
                                fn = os.path.join(config.data_upload_path, str(file_path))
                            
                            # 校验并插入
                            if os.path.exists(fn):
                                process_and_insert_image_from_file(ws, f'C{row_num}', fn, max_width=100, max_height=70)
                except Exception as e:
                    # 容错处理，防止单条数据 JSON 异常导致整个报表崩溃
                    logger.warning(f"解析产品 {bjhh} 图片JSON失败: {e}")


            # 居中对齐所有单元格
            for col in range(1, 28):
                ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='center', vertical='center')




        # 6. 生成并返回文件
        s_path = config.tmp_path
        report_rid = get_uuid()
        filename=str(report_rid)+'.xlsx'
        wb.save(s_path + '/'+ str(report_rid)+'.xlsx')
        return json_result(1, "LV出运清单生成成功",  filename)

    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f"生成清单失败: {str(e)}")
    finally:
        s.close()



