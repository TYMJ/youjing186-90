from math import e
from any import *
from .model import *
from sqlalchemy.sql import func, not_, or_, and_
import time
import json
import os
from .__default__ import get_user_path, module_xxck_new, module_share_new, user_task_new,user_task_delete
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, Alignment, NamedStyle,PatternFill,Color
from openpyxl.drawing.image import Image as Image_Get
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import  XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.xml.constants import SHEET_MAIN_NS
import zipfile
import re
from xml.etree import ElementTree as ET


SYS_FIELDS = ['sid', 'rid', 'uid', 'ctime', 'mtime', 'has_att','modi_uid', 'wf_status', 'wf_unit', 'pid', 'archived']


# 出运计划的产品资料采购人员、业务人员变更时，自动带出对应的部门编号和路径
@any_route('/api/saier/shipping/user/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_user_change(request):
    s = Session()
    j = await request.json()
    try:
        bm = ''
        username = j.get('username', '')
        org = get_user_path(username)
        path = org.get('path', '')
        d = run_sql(f"select bm from ywrybiao where yhm='{username}' limit 1")
        if len(d) > 0:
            bm = d[0].get('bm', '')

        return json_result(1, '操作成功', {'bm': bm, 'path': path})
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运计划的产品资料专业货号变更时，自动带出对应的单据品名


@any_route('/api/saier/shipping/zyhh/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_zyhh_change(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        cpbh = j.get('cpbh', '')
        data = {}
        d = run_sql(
            f"select djpmy,djpmw,djpm,wypp from zscp where (cpbh='{cpbh}') limit 1")
        if len(d) > 0:
            data = d[0]

        return json_result(1, '操作成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 出运计划的产品资料采购合同变更时，自动带出对应的付款抬头
@any_route('/api/saier/shipping/cght/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_cght_change(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        hthm = j.get('hthm', '')
        data = {}
        d = run_sql(
            f"select fktt from fkspsheet3 where (hthm='{hthm}') limit 1")
        if len(d) > 0:
            data = d[0]

        return json_result(1, '操作成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运计划的产品资料海关编码变更时


@any_route('/api/saier/shipping/hgbm/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_hgbm_change(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        hscode = j.get('hscode', '')
        jzsb = '否'
        d = run_sql(
            f"select rid from cyzglsheet where (xm='{hscode}' and zm='禁止出运海关编码') limit 1")
        if len(d) > 0:
            jzsb = '禁止出运'
        if jzsb == '否':
            d = run_sql(
                f"select rid from cyzglsheet where ('{hscode}' like concat(xm,'%') and zm='禁止出运海关编码') limit 1")
            if len(d) > 0:
                jzsb = '禁止出运'

        return json_result(1, '操作成功', jzsb)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运计划的产品资料出货数量变更时
@any_route('/api/saier/shipping/chsl/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_chsl_change(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        hthm = j.get('hthm', '')
        cpbh = j.get('cpbh', '')
        wyzd = j.get('wyzd', '')
        wxrl = j.get('wxrl', '')
        cywyzd = j.get('cywyzd', '')
        rid = j.get('rid', '')
        cysl = 0
        cgsl = 0
        yjhs = 0
        if wyzd !='' and wyzd != None and hthm != '' and hthm != None:
            if wyzd == '' and wyzd == None:
                d = run_sql(f"select ifnull(sum(ifnull(chsl,0)),0) as chsl from cymxsheet where (zycpbh='{cpbh}') and (wxrl='{wxrl}') \
                    and (cght='{hthm}') and (fpsb1<>'是') and (ifnull(ywchy,'')<>'是')  and (ifnull(chyrq,'')='') limit 1")
            else:
                d = run_sql(f"select ifnull(sum(ifnull(chsl,0)),0) as chsl from cymxsheet where (zycpbh='{cpbh}') and (pgwy='{wyzd}') \
                    and (cght='{hthm}') and (fpsb1<>'是') and (ifnull(ywchy,'')<>'是')  and (ifnull(chyrq,'')='') limit 1")
            if len(d) > 0:
                cysl = d[0].get('chsl', 0)

            d = run_sql(f"select ifnull(cgsl,0) as cgsl from cghtsheet where (wyzd='{wyzd}') and (hthm='{hthm}') and (xjht<>'作废') limit 1")
            if len(d) > 0:
                cgsl = d[0].get('cgsl', 0)
            if cywyzd != '' and cywyzd != None:
                d = run_sql(f"select ifnull(sum(ifnull(chsl,0)),0) as chsl from chyjhsheet where wyzd='{wyzd}' and ifnull(rid,'')<>'{rid}' limit 1")
            else:
                d = run_sql(f"select ifnull(sum(ifnull(chsl,0)),0) as chsl from chyjhsheet where wyzd='{wyzd}' limit 1")
            if len(d) > 0:
                yjhs = d[0].get('chsl', 0)

        return json_result(1, '操作成功', {'cysl': cysl, 'cgsl': cgsl, 'yjhs': yjhs})
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


@any_route('/api/saier/shipping/wyzd/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_wyzd_change(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        hthm = j.get('hthm', '')
        # cpbh = j.get('cpbh', '')
        wyzd = j.get('wyzd', '')
        # wxrl = j.get('wxrl', '')
        cywyzd = j.get('cywyzd', '')
        cysl = 0
        cgsl = 0
        yjhs = 0
        # if wyzd !='':
        #     d = run_sql(f"select ifnull(sum(ifnull(chsl,0)),0) as chsl from cymxsheet where (zycpbh='{cpbh}') and (wxrl='{wxrl}') and (cght='{hthm}') and (fpsb1<>'是') and (ifnull(ywchy,'')<>'')  and (ifnull(chyrq,'')='') limit 1")
        # else:
        # (zycpbh='{cpbh}') and
        d = run_sql(
            f"select ifnull(sum(ifnull(chsl,0)),0) as chsl from cymxsheet where (pgwy='{wyzd}') and (cght='{hthm}') and (fpsb1<>'是') and (ifnull(ywchy,'')<>'')  and (ifnull(chyrq,'')='') limit 1")
        if len(d) > 0:
            cysl = d[0].get('chsl', 0)
        d = run_sql(
            f"select ifnull(cgsl,0) as cgsl from cghtsheet where (wyzd='{wyzd}') and (hthm='{hthm}') and (xjht<>'作废') limit 1")
        if len(d) > 0:
            cgsl = d[0].get('cgsl', 0)

        d = run_sql(
            f"select ifnull(sum(ifnull(chsl,0)),0) as chsl from chyjhsheet where (wyzd='{wyzd}') and (ifnull(cywyzd,'')<>'{cywyzd}') limit 1")
        if len(d) > 0:
            yjhs = d[0].get('chsl', 0)

        return json_result(1, '操作成功', {'cysl': cysl, 'cgsl': cgsl, 'yjhs': yjhs})
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运计划的产品资料替换合同变更时
@any_route('/api/saier/shipping/thht/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_thht_change(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        thht = j.get('thht', '')
        wyzd = j.get('wyzd', '')
        rid = j.get('rid', '')
        fphm = j.get('fphm', '')
        qtchht = ''
        ht_list = []
        if thht != '' and thht != None:
            ht_list.append(f"合同明细：")
            ht_list.append(f"外销发票:{fphm}/外销合同:{thht}")
        d = run_sql(
            f"select thht from chyjhsheet where (wyzd='{wyzd}') and ifnull(thht,'')<>'' and rid<>'{rid}' group by thht")
        for r in d:
            if len(ht_list) > 0:
                ht_list.append(f"外销发票:{fphm}/外销合同:{r.get('thht','')}")
            else:
                ht_list.append(f"合同明细：")
                ht_list.append(f"外销发票:{fphm}/外销合同:{r.get('thht','')}")
        if len(ht_list) > 0:
            qtchht = '\n'.join(ht_list)
        s.query(cghtsheet).filter(cghtsheet.wyzd == wyzd,).update(
            {'qtchht': qtchht}, synchronize_session=False)
        s.commit()
        return json_result(1, '操作成功')
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 出运计划的补报信息字段变更时
@any_route('/api/saier/shipping/bbxx/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_bbxx_change(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        cghth = j.get('cghth', '')
        fphm = j.get('fphm', '')
        cpbh = j.get('cpbh', '')
        wxwyzd = j.get('wxwyzd', '')
        cywyzd = j.get('cywyzd', '')
        wxfp = j.get('wxfp', '')

        if cghth != '' and cghth != None and fphm != '' and fphm != None and cpbh != '' and cpbh != None and wxwyzd != '' and wxwyzd != None and cywyzd != '' and cywyzd != None:
            s.query(dbbqd).filter(dbbqd.cghth == cghth, dbbqd.cywyzd == cywyzd, dbbqd.wxwyzd == wxwyzd,
                dbbqd.cpbh == cpbh, dbbqd.fphm == fphm).update({'ck': '否', 'bbfp': wxfp}, synchronize_session=False)
            s.commit()

        return json_result(1, '操作成功')
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 出运计划的产品资料佣金比例变更
@any_route('/api/saier/shipping/yjbl/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_yjbl_change(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        flag = 0
        khmc = j.get('khmc', '')
        d = run_sql(
            f"select bz from cyzglsheet where xm='{str(khmc)}' and zm='按佣金单价计算佣金' limit 1")
        if len(d) > 0:
            flag = 1

        return json_result(1, '操作成功', flag)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运计划的产品资料佣金比例变更


@any_route('/api/saier/shipping/zs/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_zs_change(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        wyzd = j.get('wyzd', '')
        wxwyzd = j.get('wxwyzd', '')
        cg_data = {}
        wx_data = {}
        d = run_sql(
            f"select cgjg,hthm,sl,ljrk,zzsl,zhwbgpm,kpgc,zzjgdm,kplxr,kpdh,hyd from cghtsheet where wyzd='{wyzd}' limit 1")
        if len(d) > 0:
            cg_data = d[0]
        d = run_sql(
            f"select wxdj,Twxdj,mjdj1,pkRMB,order_id from wxhtsheet where wxwyzd='{wxwyzd}' limit 1")
        if len(d) > 0:
            wx_data = d[0]

        return json_result(1, '操作成功', {'cg_data': cg_data, 'wx_data': wx_data})
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运计划的目的口岸变更


@any_route('/api/saier/shipping/mdka/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_mdka_change(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        kamc = j.get('kamc', '')
        ssgjz = ''
        d = run_sql(f"select ssgjz from mdka where mdka='{kamc}' limit 1")
        if len(d) == 0:
            return json_result(-1, '目的口岸错请,重选')
        ssgjz = d[0].get('ssgjz', '')

        d = run_sql(f"select rid from mygb where zwmc='{ssgjz}' limit 1")
        if len(d) == 0:
            return json_result(-1, '此港口中文国家有误,请更改')

        return json_result(1, '操作成功')
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运计划的已收定金变更
@any_route('/api/saier/shipping/ysdj/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_ysdj_change(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        fphm = j.get('fphm', '')
        cydh = j.get('cydh', '')
        lines = j.get('lines', [])
        sydje = 0
        fp_list = []
        fp_list.append(fphm)
        if fphm != '' and fphm != None and fphm == cydh:
            for r in lines:
                xshth = r.get('wxht', '')
                if xshth != '' and xshth != None:
                    fp_list.append(xshth)

        for fp in fp_list:
            d = run_sql(
                f"select ifnull(sum(ifnull(sydje2,0)),0) as sydje from krshsheet where fphm='{fp}' limit 1")
            if len(d) > 0:
                sydje = sydje + float(d[0].get('sydje', 0))

        return json_result(1, '操作成功', sydje)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 出运计划单证人员变更按钮校验
@any_route('/api/saier/shipping/dzry/check', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_dzry_check(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        rid = j.get('rid', '')
        d = s.query(chyjh.wf_status,chyjh.dzzt).filter(chyjh.rid == rid).first()
        if d and d.wf_status == 2 and d.dzzt == 2:
            a = 1
            # 该功能未完成


        return json_result(1, '操作成功')
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运计划的LC信保变更
@any_route('/api/saier/shipping/lcxb/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_lcxb_change(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        khmc = j.get('khmc', '')
        lcxb = j.get('lcxb', '')
        xbfl = 0
        if lcxb == '是':
            d = run_sql(
                f"select ifnull(LCfl,0) xbfl from fkkh where khmc='{khmc}' limit 1")
            if len(d) > 0:
                xbfl = float(d[0].get('xbfl', 0))
        else:
            d = run_sql(
                f"select ifnull(xbfl,0) xbfl from kh where company_name='{khmc}' limit 1")
            if len(d) > 0:
                xbfl = float(d[0].get('xbfl', 0))

        return json_result(1, '操作成功', xbfl)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 出运计划的LC信保变更
@any_route('/api/saier/shipping/sfxb/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_lcxb_change(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        khmc = j.get('khmc', '')
        sfxb = j.get('sfxb', '')
        xbfl = 0
        flag = 0
        if sfxb == '是' and khmc != '' and khmc != None:
            d = run_sql(
                f"select ifnull(xbfl,0) xbfl from kh where company_name='{khmc}' limit 1")
            if len(d) > 0:
                xbfl = float(d[0].get('xbfl', 0))
            d = run_sql(
                f"select ifnull(LCfl,0) xbfl from fkkh where khmc='{khmc}' limit 1")
            if len(d) > 0 and float(d[0].get('xbfl', 0)) > 0:
                flag = 1

        return json_result(1, '操作成功', {'xbfl': xbfl, 'flag': flag})
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运计划的出运单号变更


@any_route('/api/saier/shipping/cydh/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_cydh_change(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        zdry = j.get('zdry', '')
        org = get_user_path(zdry)
        path = org.get('path', '')
        logger.info(f'path: {path}')
        cysh = ''
        if path != '' and path != None:
            path1 = path.replace('\\', '_')
            d = run_sql(
                f"select rid from spwt where '{path1}' like concat('%',replace(cydl,'\\\\','_'),'%') limit 1")
            if len(d) > 0:
                cysh = '123'

        return json_result(1, '操作成功', {'path': path, 'cysh': cysh})
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运计划的合同号码变更


@any_route('/api/saier/shipping/wxht/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_wxht_change(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        order_id = j.get('order_id', '')
        data = {}
        if order_id != '' and order_id != None:
            d = run_sql(
                f"select ssdq,hglx,hgcc,kfzy,kffp,jgtk,jhfs,dforder_id,hbdm,customer,ysfs,cyka,mdka,ywry,mt,kh_id,htbz,zgbz,country,yjcq,khpd from wxht where order_id='{order_id}' limit 1")
            if len(d) > 0:
                data = d[0]

        return json_result(1, '操作成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运计划的外销发票码变更


@any_route('/api/saier/shipping/wxfp/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_wxfp_change(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        fphm = j.get('fphm', '')
        data = {}
        if fphm != '' and fphm != None:
            if not 'CSD-' in fphm:
                d = run_sql(
                    f"select bz from cyzglsheet where ('{fphm}' like concat('%',xm,'%')) and (zm='外销合同识别公司') limit 1")
                if len(d) > 0:
                    data['gsmc'] = d[0].get('bz', '')

            d = run_sql(
                f"select SNID,StuffingDate,SealNo,LoadingSupervi,zggr,ContainerNo from delivery where fphm='{fphm}' limit 1")
            if len(d) > 0:
                data.update(d[0])
            d = run_sql(
                f"select ifnull(sum(ifnull(sydje2,0)),0) as sydje from krshsheet where fphm='{fphm}' limit 1")
            if len(d) > 0:
                data['ysdj'] = float(d[0].get('sydje', 0))

        return json_result(1, '操作成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运计划的品名数变更
@any_route('/api/saier/shipping/pms/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_pms_change(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        pms = j.get('pms', 0)
        rid = j.get('rid', '')
        data = {}
        d = s.query(chyjh.khpd).filter(chyjh.rid == rid).first()
        if not d:
            return json_result(-1, '请先保存记录')
        if d.khpd == '是':
            field = 'mjzj'
        else:
            field = 'wxzj'
        pm_list = []
        d = run_sql(
            f"select ywbgpm,sum({field}) as je from chyjhsheet where pid='{rid}' group by ywbgpm order by sum({field}) desc limit {pms}")
        for r in d:
            pm_list.append(r.get('ywbgpm', ''))

        return json_result(1, '操作成功', pm_list)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运计划的定金查看变更
@any_route('/api/saier/shipping/djck/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_djck_change(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        ht_list = j.get('ht_list', [])
        rid = j.get('rid', '')
        data = []
        for ht in ht_list:
            d = run_sql(
                f"select pid father,fphm Field1,wxht Field2,sydje Field3,syrq Field4,hbdm Field5,yhhl Field6,ywry Field7,ywbm Field8,sybz Field9,sfdj Field10 from krshsheet where (wxht='{ht}') and (sfdj='是')")
            for r in d:
                c = s.query(krsh.djdh).filter(
                    krsh.rid == r.get('father', '')).first()
                if c:
                    r['Field'] = c.djdh
                    r['rid'] = get_uuid()
                    r['pid'] = rid
                    r['ctime'] = time.strftime('%Y-%m-%d %H:%M:%S')
                    r['uid'] = user.rid
                data.append(r)

        return json_result(1, '操作成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运计划的客户名称变更
@any_route('/api/saier/shipping/khmc/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_khmc_change(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        khmc = j.get('khmc', '')
        khbh = j.get('khbh', '')
        dl_list = []
        pid = ''
        fkksj = ''
        data = {}
        data['sfxk'] = ''
        data['xksj'] = ''
        data['ttr'] = {}
        data['cdgs'] = {}
        data['dl_list'] = []
        data['kh_data'] = {'xksj': '', 'xbdm': '', 'bglyq': '', 'zfqx': 0,'xbed': 0, 'yyed': 0, 'syed': 0, 'xbzffs': '', 'xbqx': '', 'ydqx': ''}
        d = run_sql(
            f"select yjds,fkgl,rid,kh_id,khjc,xybx,zffs,ssdq,country from kh where company_name='{khmc}' limit 1")
        if len(d) > 0:
            pid = d[0].get('rid', '')
            data['kh_data'].update(d[0])
        d = run_sql(
            f"select xksj,xbdm,bglyq,zfqx,xbed,yyed,syed,xbzffs,xbqx,ydqx,fxksj from fkkh where (khbh='{khbh}') limit 1")
        if len(d) > 0:
            data['kh_data'].update(d[0])
            fkksj = d[0].get('fxksj', '')

        username = user.username
        username = '张露洁'
        org = get_user_path(username)
        path = org.get('path', '')
        field = 'cydl'
        if 'AMAZON' in khmc.upper():
            field = 'dxdl'
        path = path.replace('\\', '_')
        c = run_sql(f"select dlr,yhms from spwt where '{path}' like concat('%',replace({field},'\\\\','_'),'%') and dlr<>'{username}' order by yhms desc")
        for r in c:
            if not r.get('dlr', '') in dl_list:
                dl_list.append(r.get('dlr', ''))

        data['dl_list'] = dl_list
        if pid != '' and pid != None:
            c = run_sql(
                f"select ttr,tzr,shr from khttxx where pid='{pid}' order by sid limit 1")
            if len(c) > 0:
                data['ttr'] = c[0]

            c = run_sql(
                f"select hdxx,cygs,Forwarder from customerforwarder where pid='{pid}' order by sid limit 1")
            if len(c) > 0:
                data['cdgs'] = c[0]

        sn = 0
        c = run_sql(f"select mc from zx where ly='老客新出' order by sid")
        if len(c) > 0:
            sn = c[0].get('mc', '')
        xksb = 1
        c = run_sql(
            f"select sjcy1 from cymx where (cysb='是') and (ifnull(sjcy1,'')<>'') and (kh_id='{pid}') order by sjcy1 limit 1")
        if len(c) > 0:
            if c[0].get('sjcy1', '') != '' and c[0].get('sjcy1', '') != None and c[0].get('sjcy1', '')[:4] != time.strftime('%Y'):
                xksb = 1
            else:
                xksb = 2
        sjcy = str(int(time.strftime('%Y')) - int(sn)) + str('-01-01')
        c = run_sql(f"select sjcy1,wxjez from cymx where (cysb='是') and (ifnull(sjcy1,'')<>'') and (sjcy1>'{sjcy}') \
            and (sjcy1<'{time.strftime('%Y-01-01')}') and (kh_id='{khbh}') order by sjcy1 limit 1")
        if len(c) == 0:
            if xksb == 2:
                data['sfxk'] = '老客新出'

        if xksb == 1:
            data['sfxk'] = '新客新出'
        if fkksj != '' and fkksj != None and fkksj[:4] == time.strftime('%Y'):
            data['sfxk'] = '否'
        else:
            c = run_sql(
                f"select xksj from fkkh where (khbh='{khbh}') and (sfxk='是') limit 1")
            if len(c) > 0:
                xksj = time.strftime('%Y')
                if c[0].get('xksj', '') != '' and c[0].get('xksj', '') != None:
                    xksj = c[0].get('xksj', '')[:4]
                if xksj == time.strftime('%Y'):
                    data['sfxk'] = '是'
                    data['xksj'] = c[0].get('xksj')

        return json_result(1, '操作成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 出运计划的风控人员变更
@any_route('/api/saier/shipping/fkry/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_fkry_change(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        fkry = j.get('fkry')
        szgs = j.get('szgs', '')
        d = run_sql(f"select fkdz,fksd from spwt where (dlr='{fkry}') and (szgs='{szgs}') and (fkyw='有')")
        if len(d) == 0:
            return json_result(-1, '此人非风控人员,请重新选择')

        return json_result(1, '操作成功')
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运计划的单证人员变更
@any_route('/api/saier/shipping/dzry/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_dzry_change(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        dzry = j.get('dzry')
        d = run_sql(f"select rid from ywrybiao where (yhm='{dzry}') and (zw like '%单证%') limit 1")
        if len(d) == 0:
            return json_result(-1, '此人非单证人员,请重新选择')

        return json_result(1, '操作成功')
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运计划的数值变更
@any_route('/api/saier/shipping/sz/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_sz_change(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        data = {}
        d = run_sql(f"select sz1,sz2 from zx where (mc='换汇成本上下限') limit 1")
        if len(d) > 0:
            data = d[0]

        return json_result(1, '操作成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 出运计划的出运申请变更
@any_route('/api/saier/shipping/cysq/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_cysq_change(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        rid = j.get('rid')
        spsq = j.get('spsq','')
        dzry = j.get('dzry')
        lines = j.get('lines',[])
        data = {}
        jzz = []
        cywyzdsl = 0
        i = 0
        bgdsc = ''
        error1 = []
        error2 = []
        for r in lines:
            i = i + 1
            a = r.get('bjhh') +  r.get('krddh') +  r.get('zwpm') +  r.get('cght') + str(r.get('chsl')) + str(r.get('gczj')) + str(r.get('wxzj')) + str(r.get('mjzj')) + str(r.get('tsl')) +  r.get('wyzd') +  r.get('dpmc')
            if a not in jzz:
                jzz.append(a)
            else:
                cywyzdsl = cywyzdsl + 1
            tssb = 0
            i3 = 0
            i4 = 0
            zzsl = r.get('zzsl')
            tsl = r.get('tsl')
            hwmc = r.get('zhwbgpm')
            zs = r.get('zs')
            sccj1id = r.get('sccj1id')
            skfm = r.get('skfm')
            bank = r.get('bank')
            zh = r.get('zh')
            kpgc = r.get('kpgc')
            hyd = r.get('hyd')
            sfdk = r.get('ewchy')
            dkds = r.get('dkds')
            hbdm = r.get('hbdm')
            jsfs = r.get('jsfs')
            jcsj = r.get('jcsj')
            if zzsl>0 and tsl>0 and hwmc!='' and hwmc!='无' and hwmc!=None:
                d = run_sql(f"select rid from hgbmbsheet where (hwmc='{hwmc}') and (zzsl='{zzsl}') and (tsl='{tsl}') limit 1")
                if len(d) > 0:
                    tssb = 1
            if (zzsl == 0 or zzsl==None) and zs!='是' and zs!='工厂' and sccj1id!='' and sccj1id!=None and (hwmc=='' or hwmc=='无' or hwmc==None) and (skfm=='' or bank=='' or zh=='' or skfm==None or bank==None or zh==None) and i4==0:
                error1.append(i)
                i4 = i4 + 1
            if zzsl != 0 and zzsl!=None and (hwmc=='' or hwmc=='无' or hwmc==None or kpgc=='' or hyd=='' or kpgc==None or hyd==None or ((tsl == 0 or tsl==None) and tssb==0)) and i3==0:
                error2.append(i)
                i3 = i3 + 1
            if  hwmc!='' and hwmc!='无' and hwmc!=None and (kpgc=='' or hyd=='' or kpgc==None or hyd==None or zzsl == 0 or zzsl==None or ((tsl == 0 or tsl==None) and tssb==0)) and i3==0:
                error2.append(i)
                i3 = i3 + 1
            if  sfdk=='是' and hwmc!='无' and (dkds == 0 or dkds==None) and i3==0:
                error2.append(i)
                i3 = i3 + 1
            if  jsfs!='' and jsfs!=None and '于进仓' in jsfs and (jcsj == '' or jcsj==None) and i3==0:
                error2.append(i)
                i3 = i3 + 1
            if  hbdm!='' and hbdm!='RMB' and i3==0:
                bgdsc = '待上传'

        data = {'bgdsc':bgdsc,'error1':error1,'error2':error2, 'cywyzdsl': cywyzdsl}
        if len(error1) or len(error2) or cywyzdsl>0:
            return json_result(-1,'校验失败', data)
        
        org = get_user_path(spsq)
        path = org.get('path')
        if path == '' or path == None:
            return json_result(-1,'无此人员,请重新选择')
        d = s.query(chyjh.rid).filter(chyjh.rid==rid).first()
        if not d:
            return json_result(-1,'请先保存记录再重新选择')
        org = get_user_path(user.username)
        path2 = org.get('path')
        path2 = path2.replace('\\','_')

        d = run_sql(f"select rid from spwt where dlr='{spsq}' and ('{path2}' like concat('%',replace(cydl,'\\\\','_'),'%') or '{path2}' like concat('%',replace(dxdl,'\\\\','_'),'%') or \
            cydl like concat('%,','{user.username}',',%') or dxdl like concat('%,','{user.username}',',%')) limit 1")
        if len(d) == 0:
            return json_result(-1,'此人没有审批权限,请重新选择')
        cyhz = ''
        fkqr = ''
        zjhz = ''
        if dzry!='' and dzry!=None:
            cyhz = '通过'
            fkqr = '待审批'
        else:
            d = run_sql(f"select tsdl  from spwt where (dlr='{spsq}') and (tsdl<>'') limit 1")
            if len(d) > 0:
                zjhz = d[0].get('tsdl', '')

        data['cyhz'] = cyhz
        data['fkqr'] = fkqr
        data['zjhz'] = zjhz
        return json_result(1, '操作成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 出运计划的单证确认状态变更按钮
@any_route('/api/saier/shipping/update/dzzt', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_update_dzzt(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        wtms = ''
        d = run_sql(f"select wtms from faq where (sqbh='24-01-00002') limit 1")
        if len(d) > 0:
            wtms = d[0].get('wtms', '')

        return json_result(1, '操作成功', wtms)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 出运计划的外销发票码变更
@any_route('/api/saier/shipping/open/fkkh', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_open_fkkh(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        khbh = j.get('khbh', '')
        rid = ''
        d = run_sql(
            f"select rid from fkkh where khbh='{khbh}' and ckqx='有' limit 1")
        if len(d) > 0:
            rid = d[0].get('rid', '')

        return json_result(1, '操作成功', rid)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运计划的发票号码变更


@any_route('/api/saier/shipping/update/fphm', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_update_fphm(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        lines = j.get('lines', [])
        data = {}
        for r in lines:
            wyzd = r.get('wyzd', '')
            rid = r.get('rid', '')
            data[rid] = {'wxht': '', 'hthm': ''}
            d = run_sql(
                f"select wxht,hthm from cghtsheet where wyzd='{wyzd}' limit 1")
            if len(d) > 0:
                data[rid] = d[0]
                s.query(chyjhsheet).filter(chyjhsheet.wyzd == wyzd).update({'wxht': d[0].get(
                    'wxht', ''), 'cght': d[0].get('hthm', '')}, synchronize_session=False)

        s.commit()
        return json_result(1, '操作成功', data)
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运计划的产品资料料.更新现金工厂开票按钮
@any_route('/api/saier/shipping/items/update/supplier', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_items_update_supplier(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        lines = j.get('lines', [])
        cs_data = {}
        for r in lines:
            cs_id = r.get('sccj1id', '')
            zzsl = r.get('zzsl', 0)
            bgpm = r.get('zhwbgpm', '')
            skfm = r.get('skfm', '')
            bank = r.get('bank', '')
            zh = r.get('zh', '')
            if cs_id == '' or cs_id == None:
                continue
            if zzsl == 0 or bgpm == '' or bgpm == '无' or skfm == '' or bank == '' or zh == '':
                if cs_id in cs_data:
                    continue
                d = run_sql(
                    f"select fkhm,bank1,zh1 from zycs where cs_id='{cs_id}' limit 1")
                if len(d) > 0:
                    cs_data[cs_id] = d[0]

        return json_result(1, '操作成功', cs_data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 出运计划的改单申请按钮
@any_route('/api/saier/shipping/update/apply', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_update_apply(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        rid = j.get('rid', '')
        wxfp = j.get('wxfp', '')
        spsq = j.get('spsq', '')
        org = get_user_path(spsq)
        path = org.get('path', '')
        date = str(time.strftime("%Y-%m-%d"))
        xxnr = f"出运计划发票为:{wxfp}申请改单,日期:{date}"
        row = {
            "xxly": '出运计划',
            "bjdh": '',
            "wxht": '',
            "cght": '',
            "yhdh": wxfp,
            "xxnr": xxnr,
            "jsr": str(spsq),
            "sys_path": path,
            "spsq": spsq
        }
        res = module_xxck_new([row], user, s)
        if res.get('code',1) != 1:
            return json_result(-1, res.get('msg'))
        
        res = user_task_new('出运计划', rid, '外销发票', '出运计划[外销发票]','出运计划发票为:' + str(wxfp) + '申请改单,日期:' + str(time.strftime("%Y-%m-%d")), user, s, [spsq], True)
        logger.error(res)
        if res.get('code') != 1:
            s.rollback()
            return json_result(res.get('code'), res.get('msg'))

        s.commit()
        return json_result(1, '操作成功')
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运计划的产品资料料.更新货源地按钮
@any_route('/api/saier/shipping/items/update/hyd', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_items_update_hyd(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        lines = j.get('lines', [])
        cs_data = {}
        for r in lines:
            wyzd = r.get('wyzd', '')
            kpgc = r.get('kpgc', '')
            hyd = r.get('hyd', '')
            rid = r.get('rid', '')
            cs_data[rid] = {}
            if kpgc != '' and kpgc != None and (hyd == '' or hyd == None):
                d = run_sql(
                    f"select hyd from zycs where (company_name=='{kpgc}' limit 1")
                if len(d) > 0:
                    cs_data[rid]['hyd'] = d[0].get('hyd', '')
            elif (kpgc == '' or kpgc == None) and wyzd != '' and kpgc != None:
                d = run_sql(
                    f"select hyd,kpgc from cggdsheet where (wyzd=='{wyzd}' limit 1")
                if len(d) > 0:
                    cs_data[rid]['hyd'] = d[0].get('hyd', '')
                    cs_data[rid]['kpgc'] = d[0].get('kpgc', '')

        return json_result(1, '操作成功', cs_data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 出运计划的产品资料料.更新货源地按钮
@any_route('/api/saier/shipping/items/update/data', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_items_update_data(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        lines = j.get('lines', [])
        kh_id = j.get('kh_id', '')
        item_data = {}
        for r in lines:
            wyzd = r.get('wyzd', '')
            wxwyzd = r.get('wxwyzd', '')
            order_id = r.get('order_id', '')
            bjhh = r.get('bjhh', '')
            zyhh = r.get('zyhh', '')
            number1 = r.get('number1', '')

            rid = r.get('rid', '')
            item_data[rid] = {}
            d = run_sql(f"select hbdm,bgjldw,zhwbgpm,zzsl,sl,bz3,jhrq,khhh,hgbm,wcsl,mz,fktt,jz,bzcd,bzkd,bzgd,hyd,tj,zkfy,sjqk,kpgc,sccj1,csmc,zzjgdm,kplxr,\
                kpdh,jcsj,qtsm1,yjcq,zycpbh,cgry,gdry,ljrk,cgjg,cgsl,cgxs,ksdhh from cghtsheet where (wyzd='{wyzd}' or rid='{number1}') limit 1")
            if len(d) > 0:
                item_data[rid] = d[0]
                pid = d[0].get('rid', '')
                c = run_sql(
                    f"select rkdd,cgry,gdry,jsfs from cght where rid='{pid}' limit 1")
                if len(c) > 0:
                    item_data[rid].update(c[0])

            if wxwyzd != '' and wxwyzd != None:
                c = run_sql(
                    f"select wxdj,Twxdj,mjdj1,wxwyzd,aybl,asl6,yjbl,sl6,khpd,bjyw,krddh,zmyhl,khhh,pkRMB from wxhtsheet where (wxwyzd='{wxwyzd}') limit 1")
            else:
                c = run_sql(
                    f"select wxdj,Twxdj,mjdj1,wxwyzd,aybl,asl6,yjbl,sl6,khpd,bjyw,krddh,zmyhl,khhh,pkRMB from wxhtsheet where (order_id='{order_id}' and bjhh='{bjhh}') limit 1")
            if len(c) > 0:
                item_data[rid].update(c[0])
            if zyhh != '' and zyhh != None:
                c = run_sql(
                    f"select djpm,caizi,ywpm,sfsq,krhh,krtm,wypp,krCODE,djpmy,djpmw from zscp where (cpbh='{zyhh}' or krhh='{zyhh}') limit 1")
                if len(c) > 0:
                    item_data[rid].update(c[0])
                c = run_sql(
                    f"select krtm from zscpsheet7 where (cpbh='{zyhh}' and krID='{kh_id}') limit 1")
                if len(c) > 0:
                    if c[0].get('krtm', '') != '' and c[0].get('krtm', '') != None:
                        item_data[rid]['krtm'] = c[0].get('krtm', '')

        return json_result(1, '操作成功', item_data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运计划的开票工厂变更


@any_route('/api/saier/shipping/items/delete', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_items_delete(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        group = j.get('group', '')
        wyzd = j.get('wyzd', '')
        hthm = j.get('hthm', '')
        wxwyzd = j.get('wxwyzd', '')
        jcrq = j.get('jcrq', '')
        ckmc = j.get('ckmc', '')
        cpbh = j.get('cpbh', '')
        cgwyzd = j.get('cgwyzd', '')
        wyzd = j.get('wyzd', '')
        code = 1
        msg = '操作成功'
        if group == '进仓资料':
            if jcrq != '' and jcrq != None:
                jcrq = jcrq[:10]
            d = run_sql(
                f"select rid,sysl from cghtsheet where (wyzd='{wyzd}') limit 1")
            for r in d:
                c = run_sql(
                    f"select rid from storageline where (pid='{r.get('rid')}') and (ItemNo='{cpbh}') and (wxwyzd='{wxwyzd}') and (cgwyzd='{cgwyzd}')  and (wyzd='{wyzd}') and (ifnull(ReturnCartonQty,'')=0)")
                if len(c) > 0:
                    code = -1
                    msg = f"不好意思请先通知仓库删除后在删除此记录!"
                    break

        return json_result(code, msg)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 采购跟单的保存之前操作
@any_route('/api/saier/shipping/save/before', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_save_before(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        order_id = j.get('order_id', '')
        wxfp = j.get('wxfp', '')
        cydh = j.get('cydh', '')
        fkry = j.get('fkry', '')
        dzry = j.get('dzry', '')
        rid = j.get('rid', '')
        sdsm = j.get('sdsm', '')
        fksm = j.get('fksm', '')
        dzsm = j.get('dzsm', '')
        fkzk = j.get('fkzk', '')
        sfxk = j.get('sfxk', '')
        khbh = j.get('khbh', '')
        xksj = j.get('xksj', '')
        sdqr = j.get('sdqr', '')
        cysq = j.get('sqsq', '')
        ysdz = ''
        yfkz = ''
        ydzz = ''
        data = {}
        fksb = 0
        dzsb = 0
        xx = ''
        read = 1
        if fkzk == '是':
            read = 0
            xx = '风控暂扣'
        if wxfp != '' and wxfp != None and wxfp == cydh:
            s.query(wxht).filter(wxht.order_id == order_id).update(
                {'pgxq': '是'}, synchronize_session=False)
        d = s.query(chyjh.fkry,chyjh.dzry,chyjh.dzsm,chyjh.sdsm,chyjh.fksm,chyjh.cysq).filter(chyjh.rid==rid).first()
        if not d:
                fksb = 1
                dzsb = 1
        else:
            if d.fkry != fkry:
                fksb = 1
            if d.dzry != dzry:
                dzsb = 1
            ydzr = d.dzry
            yfkr = d.fkry
            ydzz = d.dzsm
            ysdz = d.sdsm
            ysqr = d.cysq

        if (fkry != '' and fkry != None) or (dzry != '' and dzry != None):
            if fksb == 1 :
                date = str(time.strftime("%Y-%m-%d"))
                xxnr = f"出运计划发票为:{wxfp}请审单,日期:{date}"
                row = {
                    "xxly": '出运计划',
                    "bjdh": '',
                    "wxht": '',
                    "cght": '',
                    "yhdh": wxfp,
                    "xxnr": xxnr,
                    'zt': '待审',
                    "jsr": str(fkry),
                    "sys_path": "",
                    "spsq": fkry
                }
                res = module_xxck_new([row], user, s)
                if res.get('code',1) != 1:
                    return json_result(-1, res.get('msg'))
                
                res = user_task_new('出运计划', rid, '外销发票', '出运计划[外销发票]申请审单','【审批】' + '出运计划发票为:' + str(wxfp) + xx + '请审单,日期:' + str(time.strftime("%Y-%m-%d")), user, s, [fkry], True)
                logger.error(res)
                if res.get('code') != 1:
                    s.rollback()
                    return json_result(res.get('code'), res.get('msg'))
                
            if dzsb == 1 and dzry != user.username:
                date = str(time.strftime("%Y-%m-%d"))
                xxnr = f"出运计划发票为:{wxfp}请查看,日期:{date}"
                row = {
                    "xxly": '出运计划',
                    "bjdh": '',
                    "wxht": '',
                    "cght": '',
                    "yhdh": wxfp,
                    "xxnr": xxnr,
                    "jsr": str(dzry),
                    'zt': '',
                    "sys_path": "",
                    "spsq": dzry
                }
                res = module_xxck_new([row], user, s)
                if res.get('code',1) != 1:
                    return json_result(-1, res.get('msg'))
                
                res = user_task_new('出运计划', rid, '外销发票', '出运计划[外销发票]','出运计划发票为:' + str(wxfp) + xx + '请查看,日期:' + str(time.strftime("%Y-%m-%d")), user, s, [dzry], True)
                logger.error(res)
                if res.get('code') != 1:
                    s.rollback()
                    return json_result(res.get('code'), res.get('msg'))

            if dzry != '' and dzry != None and ydzr != dzry:
                res = user_task_delete('出运计划', rid, s, [ydzr])
                if res.get('code') != 1:
                    s.rollback()
                    return json_result(res.get('code'), res.get('msg'))
                
            if cysq != '' and cysq != None and cysq != ysqr and (dzry == '' or dzry == None):
                date = str(time.strftime("%Y-%m-%d"))
                xxnr = f"出运计划发票为:{wxfp}请审单,日期:{date}"
                row = {
                    "xxly": '出运计划',
                    "bjdh": '',
                    "wxht": '',
                    "cght": '',
                    "yhdh": wxfp,
                    "xxnr": xxnr,
                    "jsr": str(cysq),
                    'zt': '',
                    "sys_path": "",
                    "spsq": cysq
                }
                res = module_xxck_new([row], user, s)
                if res.get('code',1) != 1:
                    return json_result(-1, res.get('msg'))
                
                res = user_task_new('出运计划', rid, '外销发票', '出运计划[外销发票]','出运计划发票为:' + str(wxfp) + xx + '请审单,日期:' + str(time.strftime("%Y-%m-%d")), user, s, [cysq], True)
                logger.error(res)
                if res.get('code') != 1:
                    s.rollback()
                    return json_result(res.get('code'), res.get('msg'))

            if dzry !='' and dzry != None and cysq != '' and cysq != None:
                res = user_task_delete('出运计划', rid, s, [cysq])
                if res.get('code') != 1:
                    s.rollback()
                    return json_result(res.get('code'), res.get('msg'))

            if cysq != '' and cysq != None and ydzr != cysq and (dzry == '' or dzry == None):
                res = user_task_delete('出运计划', rid, s, [cysq])
                if res.get('code') != 1:
                    s.rollback()
                    return json_result(res.get('code'), res.get('msg'))
                
            u = {}
            if (sdsm != '' and sdsm != None and sdsm != ysdz) or ((sdsm=='' or sdsm==None) and ysdz!='' and ysdz!=None):
                u['sdsm'] = sdsm
            if (fksm != '' and fksm != None and fksm != yfkz) or ((fksm == '' or fksm == None) and yfkz != '' and yfkz!=None):
                u['fksm'] = fksm
            if (dzsm != '' and dzsm != None and dzsm != ydzz) or ((dzsm == '' or dzsm == None) and ydzz != '' and ydzz != None):
                u['dzsm'] = dzsm
            if len(u) > 0:
                s.query(cymx).filter(cymx.fphm == wxfp).update(u, synchronize_session=False)
            if dzry != '' and dzry != None:
                org = get_user_path(dzry)
                uid = org.get('rid', '')
                c = s.query(sys_chyjh_tag).filter(sys_chyjh_tag.record_id == rid, sys_chyjh_tag.uid == uid).first()
                if c:
                    c.read = read
                else:
                    c = sys_chyjh_tag()
                    c.record_id = rid
                    c.uid = uid
                    c.read = read
                s.add(c)
        
        if fkry != '' and fkry != None and fkry==user.username:
            if xksj == '' or xksj == None:
                xksj = time.strftime('%Y-%m-%d')
            if sfxk == '否':
                s.query(cyzglsheet).filter(cyzglsheet.xm == khbh, cyzglsheet.zm == '新客时间', cyzglsheet.bz.like(f"%%{time.strftime('%Y')}%%")).delete()
                s.query(fkkh).filter(fkkh.khbh == khbh).update({'sfxk': '否', 'fxksj': time.strftime('%Y-%m-%d')}, synchronize_session=False)
                s.query(kh).filter(kh.khbh == khbh).update({'sfxk': '否', 'fkxk': '否','fkxksj':None}, synchronize_session=False)
            elif sfxk == '是':
                c = s.query(fkkh).filter(fkkh.khbh == khbh, fkkh.sfxk == '是', fkkh.xksj.like(f'%%{str(int(xksj[:4]) - 1)}%%')).first()
                if c:
                    return json_result(0, f"请注意此客人为去年新客，请去风控客人中更新状态")
                else:
                    s.query(fkkh).filter(fkkh.khbh == khbh).update({'sfxk': '是', 'xksj': xksj, 'xkfp':wxfp,'fxksj':xksj}, synchronize_session=False)
                    c = s.query(cyzglsheet).filter(
                        cyzglsheet.xm == khbh, 
                        cyzglsheet.zm == '新客时间', 
                        # cyzglsheet.bz.like(f"%%{str(int(xksj[:4]))}%%")
                    ).first()
                    if not c:
                        c = cyzglsheet()
                        c.rid = get_uuid()
                        c.uid = user.rid
                        c.ctime = time.strftime('%Y-%m-%d %H:%M:%S')
                        c.pid = '94'
                        c.xm = khbh
                        c.zm = '新客时间'
                        c.qxzl = '新客时间'
                        c.bz = xksj
                        c.bz1 = ''
                        c.sz = 0
                        s.add(c)

        res = await shipping_sdqr_check(sdqr, wxfp, rid, user, s)
        if res.get('code') == -1:
            s.rollback()
            return json_result(res.get('code'), res.get('msg'))
        
        s.commit()
        return json_result(1, '操作成功', data)
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 出运计划的保存之前操作
@any_route('/api/saier/shipping/load/check', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_load_check(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        fphm = j.get('fphm', '')
        khmc = j.get('khmc', '')
        data = {'ysdj': 0, 'cybz': 0, 'yw':'', 'dl_list': []}
        if fphm != '' and fphm != None:
            d = run_sql(f"select ifnull(sum(ifnull(sydje2,0)),0) as sydje1 from krshsheet where (fphm='{fphm}') and (ifnull(sfdj,'')='是') limit 1")
            if len(d) > 0:
                data['ysdj'] = float(d[0].get('sydje1', 0))
            d = run_sql(f"select rid from cymx where (fphm='{fphm}') limit 1")
            if len(d) > 0:
                data['cybz'] = 1
        org = get_user_path(user.username)
        data['yw'] = org.get('path', '')
        field = 'cydl'
        if 'AMAZON' in khmc.upper():
            field = 'dxdl'
        path = org.get('path', '').replace('\\', '_')
        dl_list = []
        c = run_sql(f"select dlr,yhms from spwt where '{path}' like concat('%',replace({field},'\\\\','_'),'%') and dlr<>'{user.username}' order by yhms desc")
        for r in c:
            if not r.get('dlr', '') in dl_list:
                dl_list.append(r.get('dlr', ''))
        data['dl_list'] = dl_list

        return json_result(1, '操作成功', data)
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 出运计划的订柜模板下载
@any_route('/api/saier/shipping/download/dgmb', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_download_dgmb(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        wb = Workbook() # 新建Excel文件
        ws = wb.active # wb.create_sheet() # 默认激活第一个sheet
        i = 1
        ws.column_dimensions['A'].width = 14.3
        ws['A1'] = '外销发票'
        ws.column_dimensions['B'].width = 14.3
        ws['B1'] = '对应客人'
        ws.column_dimensions['C'].width = 25
        ws['C1'] = '柜型'
        ws.column_dimensions['D'].width = 17
        ws['D1'] = '装柜地点'
        ws.column_dimensions['E'].width = 14.3
        ws['E1'] = '货代'
        ws.column_dimensions['F'].width = 17
        ws['F1'] = '预计船期'
        ws.column_dimensions['G'].width = 14.3
        ws['G1'] = '出运口岸'
        ws.column_dimensions['H'].width = 20
        ws['H1'] = '目的口岸'
        ws.column_dimensions['I'].width = 14.3
        ws['I1'] = '装柜日期'
        ws.column_dimensions['J'].width = 14.3
        ws['J1'] = '对应单证'
        ws.column_dimensions['K'].width = 18
        ws['K1'] = '备注'
        rids = j.get('rids', [])
        for rid in rids :
            d = run_sql(f"select wxfp,khmc,hglx,zgdd,cdmc,etd,cyka,mdka,zgrq,dzry,zysx from chyjh where (rid='{rid}') limit 1")
            if len(d) == 0:
                continue
            c = d[0]
            i += 1
            ws['A' + str(i)] = c.get('wxfp')
            ws['B' + str(i)] = c.get('khmc')
            ws['C' + str(i)] = c.get('hglx')
            ws['D' + str(i)] = c.get('zgdd')
            ws['E' + str(i)] = c.get('cdmc')
            ws['F' + str(i)] = c.get('etd')
            ws['G' + str(i)] = c.get('cyka')
            ws['H' + str(i)] = c.get('mdka')
            ws['I' + str(i)] = c.get('zgrq')
            ws['J' + str(i)] = c.get('dzry')
            ws['K' + str(i)] = c.get('zysx')

        path = config.tmp_path
        report_rid = get_uuid()
        wb.save(path + '/'+ str(report_rid)+'.xlsx')

        return json_result(1,'生成报表成功',{'path': str(report_rid)+'.xlsx', 'name':'出运计划'+str(time.strftime("%Y%m%d%H%M%S"))+'.xlsx'})
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# sudo yum install -y libreoffice
# sudo yum install -y libreoffice-pyuno
# 出运计划的订柜模板下载
@any_route('/api/saier/shipping/download/test', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_download_test(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        combox = j.get('combox', '复选框 245')
        # res = set_checkbox_value(combox, True)
        # 使用示例
        r_path = os.path.join(config.data_upload_path,'template')
        fn = os.path.join(r_path, '优景最新版BOOKING格式.xlsx')
            
        if not os.path.exists(fn):
            return {"code":-1,"msg":"未找到报表模板"}
    
        # 使用示例

        # 先检测
        # control_type = detect_control_type(fn)
        # print(f"\n控件类型: {control_type}")
        # 使用
        # 使用
        set_excel_checkboxes_uno(fn)
        return json_result(1,'生成报表成功')
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
 
# import uno
# from com.sun.star.beans import PropertyValue

def set_excel_checkboxes_uno(excel_path):
    """
    使用LibreOffice/OpenOffice的UNO API
    """
    # 连接到LibreOffice
    local_context = uno.getComponentContext()
    resolver = local_context.ServiceManager.createInstanceWithContext(
        "com.sun.star.bridge.UnoUrlResolver", local_context)
    
    try:
        # 连接到运行中的LibreOffice实例
        context = resolver.resolve(
            "uno:socket,host=localhost,port=8100;urp;StarOffice.ComponentContext")
        
        desktop = context.ServiceManager.createInstanceWithContext(
            "com.sun.star.frame.Desktop", context)
        
        # 打开Excel文件
        properties = []
        prop = PropertyValue()
        prop.Name = "Hidden"
        prop.Value = True
        properties.append(prop)
        
        doc_url = uno.systemPathToFileUrl(os.path.abspath(excel_path))
        doc = desktop.loadComponentFromURL(doc_url, "_blank", 0, tuple(properties))
        
        # 访问表单控件
        draw_page = doc.getDrawPage()
        forms = draw_page.getForms()
        
        checkbox_settings = {
            '复选框 245': True,  # FCL柜货
            '复选框 246': True,  # LCL散货
            '复选框 247': True,  # Collect到付
            '复选框 248': True,  # Prepaid预付
            '复选框 249': True,  # FOB
            '复选框 250': True,  # CNF
            '复选框 251': True,  # CIF
            '复选框 252': True,  # Yes有
            '复选框 253': True,  # No没有
            '复选框 254': True,  # 空运
            '复选框 255': True,  # 海运
            '复选框 256': True,  # 铁路运输
            '复选框 257': True,  # Yes有
            '复选框 258': True,  # No没有
            '复选框 259': True,  # Yes有
            '复选框 260': True,  # No没有
        }
        
        # 遍历所有形状查找控件
        for i in range(draw_page.getCount()):
            shape = draw_page.getByIndex(i)
            shape_name = shape.getName()
            
            if shape_name in checkbox_settings:
                try:
                    # 尝试获取控件
                    control = shape.getControl()
                    if hasattr(control, 'State'):
                        control.State = 1 if checkbox_settings[shape_name] else 0
                        print(f"设置 {shape_name} = {checkbox_settings[shape_name]}")
                except:
                    print(f"无法设置 {shape_name}")
        
        # 保存
        output_url = uno.systemPathToFileUrl(
            os.path.abspath(excel_path.replace('.xls', '_uno.xls')))
        doc.storeAsURL(output_url, ())
        doc.close(True)
        
    except Exception as e:
        print(f"错误: {e}")

def set_checkboxes_with_xlwings(input_file, output_file=None):
    """
    使用xlwings处理Excel表单控件（最可靠的方法）
    xlwings可以正确保留表单控件
    """
    if output_file is None:
        output_file = input_file.replace('.xlsx', '_xlwings.xlsx')
    
    # 启动Excel应用
    app = xw.App(visible=False)  # 设置为True可以看到操作过程
    
    try:
        # 打开工作簿
        wb = app.books.open(os.path.abspath(input_file))
        ws = wb.sheets['booking form']
        
        print("正在查找并设置复选框...")
        
        # 方法1：通过名称查找形状
        # 首先，我们需要知道复选框的名称
        print("工作表中的形状:")
        for shape in ws.shapes:
            print(f"  形状: {shape.name}, 类型: {shape.type}")
            
            # 尝试获取控件值
            try:
                if hasattr(shape, 'control_format'):
                    cf = shape.control_format
                    print(f"    控件格式: {cf}")
                    print(f"    值: {cf.value}")
                    print(f"    链接单元格: {cf.linked_cell}")
            except:
                pass
        
        # 方法2：根据位置设置（更可靠）
        # 根据你的表格结构，这些位置可能有复选框
        checkbox_positions = [
            # (行, 列) - 根据你的表格调整
            (21, 5),   # E21 - 提单是否电放
            (21, 12),  # L21 - Insurance是否做保险
            (44, 3),   # C44 - ★是否信保
            (46, 3),   # C46 - ★财务审批
            (47, 3),   # C47 - ★风控审批
            (44, 13),  # M44 - 拼柜
            (45, 13),  # M45 - 小柜
            (46, 13),  # M46 - 平柜
            (47, 13),  # M47 - 高柜
        ]
        
        print("\n通过位置设置:")
        for row, col in checkbox_positions:
            cell = ws.range((row, col))
            print(f"  单元格 ({row}, {col}): {cell.address} = {cell.value}")
            
            # 设置值为TRUE（选中）
            cell.value = True
        
        # 保存文件
        wb.save(os.path.abspath(output_file))
        wb.close()
        
        print(f"\n✓ 处理完成!")
        print(f"✓ 输出文件: {output_file}")
        
        return output_file
        
    except Exception as e:
        print(f"处理失败: {e}")
        import traceback
        traceback.print_exc()
        return None
        
    finally:
        # 关闭Excel应用
        app.quit()



def detect_control_type(file_path):
    """检测Excel中的控件类型"""
    print(f"检测文件: {file_path}")
    
    # 1. 检查文件是否包含控件
    with zipfile.ZipFile(file_path, 'r') as z:
        file_list = z.namelist()
        
        print("文件包含的内容:")
        for f in sorted(file_list):
            if 'drawing' in f or 'vml' in f or 'control' in f:
                print(f"  ✓ {f}")
        
        # 检查是否有VML绘图（表单控件）
        vml_files = [f for f in file_list if 'vmlDrawing' in f]
        if vml_files:
            print(f"\n发现表单控件（VML格式）: {len(vml_files)} 个文件")
            return 'form_control'
        
        # 检查是否有ActiveX控件
        control_files = [f for f in file_list if 'ctrl' in f or 'activeX' in f]
        if control_files:
            print(f"\n发现ActiveX控件: {len(control_files)} 个文件")
            return 'activex'
        
        print("\n未发现控件文件，可能是符号或数据验证")
        return 'symbol'

def find_linked_cells_from_vml(file_path):
    """从VML文件中提取链接单元格"""
    linked_cells = {}
    
    # 直接读取zip中的VML文件
    with zipfile.ZipFile(file_path, 'r') as z:
        # 查找VML文件
        vml_files = [f for f in z.namelist() if f.endswith('.vml') and 'vmlDrawing' in f]
        
        for vml_file in vml_files:
            try:
                content = z.read(vml_file).decode('utf-8', errors='ignore')
                
                # 查找所有表单控件及其链接
                # 模式: <x:FmlaLink>Sheet1!$E$21</x:FmlaLink>
                pattern = r'<x:FmlaLink>([^<]+)</x:FmlaLink>'
                links = re.findall(pattern, content)
                
                for link in links:
                    # 链接格式可能是: Sheet1!$E$21 或 $E$21
                    if '!' in link:
                        sheet_cell = link.split('!')
                        sheet = sheet_cell[0]
                        cell = sheet_cell[1].replace('$', '')
                    else:
                        sheet = 'booking form'  # 默认工作表
                        cell = link.replace('$', '')
                    
                    if sheet not in linked_cells:
                        linked_cells[sheet] = []
                    
                    if cell not in linked_cells[sheet]:
                        linked_cells[sheet].append(cell)
                        
                        # 同时查找控件的当前值
                        # 查找这个链接附近的Value
                        val_pattern = rf'{re.escape(link)}.*?<x:Val>([01])</x:Val>'
                        val_match = re.search(val_pattern, content, re.DOTALL)
                        current_value = int(val_match.group(1)) if val_match else 0
                        
                        print(f"找到链接: {sheet}!{cell} = {current_value}")
                        
            except Exception as e:
                print(f"读取VML文件出错: {e}")
    
    return linked_cells

def set_checkbox_via_linked_cells(file_path):
    """通过链接单元格设置复选框"""
    # 1. 查找链接单元格
    print("查找链接单元格...")
    linked_cells = find_linked_cells_from_vml(file_path)
    
    if not linked_cells:
        print("未找到链接单元格，尝试手动指定...")
        # 根据你的表格结构手动指定
        linked_cells = {
            'booking form': [
                'E21',   # 提单是否电放
                'L21',   # Insurance 是否做保险
                'C44',   # ★是否信保
                'C46',   # ★财务审批
                'C47',   # ★风控审批
                'M44',   # 拼柜
                'M45',   # 小柜
                'M46',   # 平柜
                'M47',   # 高柜
            ]
        }
    
    # 2. 使用openpyxl设置单元格值
    wb = load_workbook(file_path)
    
    for sheet_name, cells in linked_cells.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n设置工作表: {sheet_name}")
            
            for cell_addr in cells:
                if cell_addr in ws:
                    # 设置值为TRUE表示选中
                    ws[cell_addr].value = True
                    print(f"  ✓ {cell_addr} = TRUE")
                else:
                    print(f"  ✗ {cell_addr} 不存在")
    
    # 3. 保存
    output_path = file_path.replace('.xlsx', '_linked.xlsx')
    wb.save(output_path)
    
    print(f"\n✓ 处理完成!")
    print(f"✓ 新文件: {output_path}")
    
    return output_path



def set_booking_checkboxes_simple():
    """最简单的设置方法 - 直接指定单元格位置"""
    # file_path = 'BOOKING格式.xlsx'
    r_path = os.path.join(config.data_upload_path,'template')
    file_path = os.path.join(r_path, '优景最新版BOOKING格式.xlsx')
    # 加载工作簿
    wb = load_workbook(file_path)
    ws = wb['booking form']
    
    # 根据表格内容，手动确定复选框位置
    # 这需要你打开Excel文件查看实际的复选框位置
    
    # 示例：假设以下单元格是复选框
    checkbox_cells = {
        '提单是否电放': 'D21',   # 需要根据实际情况调整
        '是否做保险': 'L21',     # 需要根据实际情况调整
        '是否信保': 'B45',       # 需要根据实际情况调整
        '财务审批': 'B47',       # 需要根据实际情况调整
        '风控审批': 'B48',       # 需要根据实际情况调整
        '拼柜': 'M44',          # 需要根据实际情况调整
        '小柜': 'M45',          # 需要根据实际情况调整
        '平柜': 'M46',          # 需要根据实际情况调整
        '高柜': 'M47',          # 需要根据实际情况调整
    }
    
    print("设置复选框:")
    for label, cell_addr in checkbox_cells.items():
        if cell_addr in ws:
            current = ws[cell_addr].value
            # 设置新值
            ws[cell_addr].value = '是'  # 或者 True，或者 '✓'
            print(f"  {label}: {cell_addr} = {current} -> 是")
        else:
            print(f"  {label}: {cell_addr} 不存在")
    
    # 保存
    output_path = file_path.replace('.xlsx', '_checked.xlsx')
    wb.save(output_path)
    print(f"\n文件已保存: {output_path}")
    
    return checkbox_cells



class BookingFormProcessor:
    """订舱单表格处理器"""
    
    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = load_workbook(file_path)
        self.ws = self.wb['booking form']  # 根据你提供的工作表名称
        
    def find_checkbox_symbols(self):
        """查找复选框符号（□ ☐ ☑ ✓ 等）"""
        checkboxes = []
        
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell_text = str(cell.value)
                    
                    # 查找复选框符号
                    symbols = ['□', '☐', '☑', '✓', '√', '[ ]', '[x]', '[X]', '[√]']
                    
                    for symbol in symbols:
                        if symbol in cell_text:
                            # 查找这个符号附近的布尔值单元格
                            bool_cell = self._find_adjacent_bool_cell(cell)
                            
                            checkboxes.append({
                                'cell': cell.coordinate,
                                'text': cell_text,
                                'symbol': symbol,
                                'bool_cell': bool_cell,
                                'type': 'symbol'
                            })
                            break
        
        return checkboxes
    
    def find_star_checkboxes(self):
        """查找带★的复选框（如★是否信保、★财务审批、★风控审批）"""
        star_checkboxes = []
        
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell_text = str(cell.value)
                    
                    if '★' in cell_text:
                        # 获取列字母和行号
                        col = cell.column_letter
                        row_num = cell.row
                        
                        # 在右侧查找可能的复选框（通常是同一行的下一列或下两列）
                        for offset in [1, 2, 3]:
                            try:
                                check_cell = self.ws[f"{chr(ord(col) + offset)}{row_num}"]
                                if check_cell.value in [True, False, '是', '否', 'Y', 'N', 'y', 'n']:
                                    star_checkboxes.append({
                                        'label_cell': cell.coordinate,
                                        'label_text': cell_text,
                                        'checkbox_cell': check_cell.coordinate,
                                        'value': check_cell.value,
                                        'offset': offset
                                    })
                                    break
                            except:
                                continue
        
        return star_checkboxes
    
    def find_yes_no_checkboxes(self):
        """查找是否类复选框"""
        yes_no_checkboxes = []
        
        # 根据你的表格，需要查找的标签
        target_labels = [
            '提单是否电放',
            'Insurance 是否做保险', 
            '★是否信保',
            '拼柜',
            '小柜',
            '平柜',
            '高柜'
        ]
        
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell_text = str(cell.value).strip()
                    
                    for label in target_labels:
                        if label in cell_text:
                            # 在右侧查找对应的复选框
                            col = cell.column_letter
                            row_num = cell.row
                            
                            found = False
                            for offset in range(1, 5):  # 向右搜索最多4列
                                try:
                                    check_col = chr(ord(col) + offset)
                                    check_cell = self.ws[f"{check_col}{row_num}"]
                                    
                                    # 检查是否包含复选框符号或布尔值
                                    if check_cell.value in [True, False, '是', '否', 'Y', 'N', 'y', 'n', '□', '☐', '☑', '✓']:
                                        yes_no_checkboxes.append({
                                            'label_cell': cell.coordinate,
                                            'label_text': cell_text,
                                            'checkbox_cell': check_cell.coordinate,
                                            'value': check_cell.value,
                                            'label_type': label
                                        })
                                        found = True
                                        break
                                    
                                    # 或者检查单元格文本
                                    if check_cell.value and isinstance(check_cell.value, str):
                                        check_text = str(check_cell.value)
                                        if any(sym in check_text for sym in ['□', '☐', '☑', '✓']):
                                            yes_no_checkboxes.append({
                                                'label_cell': cell.coordinate,
                                                'label_text': cell_text,
                                                'checkbox_cell': check_cell.coordinate,
                                                'value': check_text,
                                                'label_type': label
                                            })
                                            found = True
                                            break
                                        
                                except:
                                    continue
                            
                            if found:
                                break
        
        return yes_no_checkboxes
    
    def _find_adjacent_bool_cell(self, text_cell, search_radius=3):
        """查找相邻的布尔值单元格"""
        row = text_cell.row
        col = text_cell.column
        
        for dr in range(-search_radius, search_radius + 1):
            for dc in range(-search_radius, search_radius + 1):
                if dr == 0 and dc == 0:
                    continue
                
                try:
                    check_row = row + dr
                    check_col = col + dc
                    
                    if 1 <= check_row <= self.ws.max_row and 1 <= check_col <= self.ws.max_column:
                        check_cell = self.ws.cell(row=check_row, column=check_col)
                        
                        # 检查布尔值或复选框符号
                        if check_cell.value in [True, False, '是', '否', 'Y', 'N', 'y', 'n']:
                            return {
                                'cell': check_cell.coordinate,
                                'value': check_cell.value
                            }
                        
                        # 检查是否包含复选框符号
                        if check_cell.value and isinstance(check_cell.value, str):
                            check_text = str(check_cell.value)
                            if any(sym in check_text for sym in ['□', '☐', '☑', '✓']):
                                return {
                                    'cell': check_cell.coordinate,
                                    'value': check_text
                                }
                except:
                    continue
        
        return None
    
    def analyze_form_structure(self):
        """分析表格结构"""
        print(f"\n{'='*60}")
        print(f"订舱单表格分析 - {self.file_path}")
        print(f"{'='*60}")
        
        # 1. 查找所有可能包含复选框的行
        checkbox_rows = []
        for row in self.ws.iter_rows(min_row=1, max_row=50, values_only=False):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell_text = str(cell.value)
                    
                    # 检查是否包含复选框相关的关键词
                    checkbox_keywords = ['是否', '审批', '信保', '电放', '保险', '拼柜', '小柜', '平柜', '高柜']
                    
                    if any(keyword in cell_text for keyword in checkbox_keywords):
                        checkbox_rows.append({
                            'row': cell.row,
                            'cell': cell.coordinate,
                            'text': cell_text
                        })
        
        print(f"\n找到 {len(checkbox_rows)} 个可能包含复选框的行:")
        for item in checkbox_rows:
            print(f"  行 {item['row']}: {item['cell']} - {item['text']}")
            
            # 显示这一行的内容
            row_cells = []
            for col in range(1, 15):  # 查看前14列
                cell = self.ws.cell(row=item['row'], column=col)
                if cell.value:
                    row_cells.append(f"{cell.coordinate}: {cell.value}")
            
            if row_cells:
                print(f"      该行内容: {', '.join(row_cells)}")
        
        # 2. 查找符号复选框
        symbol_checkboxes = self.find_checkbox_symbols()
        print(f"\n找到 {len(symbol_checkboxes)} 个符号复选框:")
        for cb in symbol_checkboxes:
            bool_info = f" -> {cb['bool_cell']['cell']}={cb['bool_cell']['value']}" if cb['bool_cell'] else ""
            print(f"  {cb['cell']}: {cb['text']}{bool_info}")
        
        # 3. 查找带★的复选框
        star_checkboxes = self.find_star_checkboxes()
        print(f"\n找到 {len(star_checkboxes)} 个带★的复选框:")
        for cb in star_checkboxes:
            print(f"  {cb['label_cell']} ({cb['label_text']}) -> {cb['checkbox_cell']}={cb['value']}")
        
        # 4. 查找是否类复选框
        yes_no_checkboxes = self.find_yes_no_checkboxes()
        print(f"\n找到 {len(yes_no_checkboxes)} 个是否类复选框:")
        for cb in yes_no_checkboxes:
            print(f"  {cb['label_cell']} ({cb['label_text']}) -> {cb['checkbox_cell']}={cb['value']}")
        
        return {
            'symbol_checkboxes': symbol_checkboxes,
            'star_checkboxes': star_checkboxes,
            'yes_no_checkboxes': yes_no_checkboxes,
            'checkbox_rows': checkbox_rows
        }
    
    def set_checkbox_value(self, cell_address, value, symbol_mode=False):
        """
        设置复选框值
        
        cell_address: 单元格地址
        value: True/False 或 '是'/'否'
        symbol_mode: 是否使用符号模式（□/☑）
        """
        if cell_address not in self.ws:
            print(f"单元格 {cell_address} 不存在")
            return False
        
        cell = self.ws[cell_address]
        current_value = cell.value
        
        if symbol_mode:
            # 符号模式
            if value:
                new_value = cell.value.replace('□', '☑').replace('☐', '☑') if cell.value else '☑'
            else:
                new_value = cell.value.replace('☑', '□').replace('✓', '□') if cell.value else '□'
        else:
            # 布尔/文本模式
            if isinstance(value, bool):
                new_value = value
            elif value in ['是', 'Y', 'y']:
                new_value = '是'
            elif value in ['否', 'N', 'n']:
                new_value = '否'
            else:
                new_value = value
        
        cell.value = new_value
        
        # 添加视觉反馈
        fill_color = "C6EFCE" if value else "FFC7CE"  # 绿色/红色
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        print(f"已设置 {cell_address}: {current_value} -> {new_value}")
        return True
    
    def set_checkbox_by_label(self, label_text, value=True):
        """通过标签文本设置复选框"""
        label_text = label_text.strip()
        
        # 查找标签
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if label_text in str(cell.value):
                        print(f"找到标签: {cell.coordinate} = {cell.value}")
                        
                        # 在右侧查找复选框
                        col = cell.column_letter
                        row_num = cell.row
                        
                        for offset in range(1, 5):
                            try:
                                check_col = chr(ord(col) + offset)
                                check_cell = self.ws[f"{check_col}{row_num}"]
                                
                                if check_cell.value is not None:
                                    print(f"  候选复选框: {check_cell.coordinate} = {check_cell.value}")
                                    return self.set_checkbox_value(check_cell.coordinate, value)
                                    
                            except:
                                continue
        
        print(f"未找到标签: {label_text}")
        return False
    
    def save(self, output_path=None):
        """保存文件"""
        if output_path is None:
            import os
            base_name, ext = os.path.splitext(self.file_path)
            output_path = f"{base_name}_processed{ext}"
        
        self.wb.save(output_path)
        print(f"文件已保存: {output_path}")
        return output_path
    
    def print_form_preview(self, max_rows=40):
        """打印表格预览"""
        print(f"\n{'='*80}")
        print(f"订舱单表格预览 (前{max_rows}行)")
        print(f"{'='*80}")
        
        for row_idx in range(1, min(max_rows, self.ws.max_row) + 1):
            row_data = []
            for col_idx in range(1, 16):  # 前15列
                cell = self.ws.cell(row=row_idx, column=col_idx)
                
                if cell.value is None:
                    display = ""
                elif isinstance(cell.value, bool):
                    display = "✓" if cell.value else "□"
                elif isinstance(cell.value, str):
                    # 简化显示
                    text = str(cell.value)
                    if len(text) > 10:
                        display = text[:8] + "..."
                    else:
                        display = text
                else:
                    display = str(cell.value)
                
                row_data.append(f"{display:<12}")
            
            print(f"行{row_idx:3}: {' | '.join(row_data)}")


    


def set_activex_checkbox(file_path, sheet_name, checkbox_name, value=True):
    """
    设置 ActiveX 复选框的值
    """

    r_path = os.path.join(config.data_upload_path,'template')
    fn = os.path.join(r_path, '优景最新版BOOKING格式.xlsx')
        
    if not os.path.exists(fn):
        return {"code":-1,"msg":"未找到报表模板"}
    wb = load_workbook(fn) 
    ws = wb['booking form']
    # 查找控件
    for drawing in ws._drawing_part.drawing.charts or []:
        for shape in drawing.sp or []:
            if shape.nvSpPr.cNvPr.name == checkbox_name:
                # 修改控件的 XML 属性
                if hasattr(shape, 'spPr'):
                    # 这里是简化处理，实际需要修改更复杂的 XML 结构
                    pass

    path = config.tmp_path
    report_rid = get_uuid()
    wb.save(path + '/'+ str(report_rid)+'.xlsx')

def set_checkbox_value(checkbox_name='复选框 245', value=True):
    """
    设置复选框的值
    :param file_path: Excel文件路径
    :param sheet_name: 工作表名称
    :param checkbox_name: 复选框名称（如 '复选框 255'）
    :param value: True/False
    """
    excel = win32.Dispatch('Excel.Application')
    excel.Visible = False  # 设置为 True 可以看到操作过程
    excel.DisplayAlerts = False

    r_path = os.path.join(config.data_upload_path,'template')
    fn = os.path.join(r_path, '优景最新版BOOKING格式.xlsx')
        
    if not os.path.exists(fn):
        return {"code":-1,"msg":"未找到报表模板"}
    wb = excel.Workbooks.Open(fn)
    ws = wb.Sheets('booking form')
    # wb = load_workbook(fn) 
    # ws = wb['booking form']
    # 查找并设置复选框
    # for shape in ws._shapes:
    #     if shape.name == checkbox_name:
    #         # 对于表单控件，可能需要修改链接的单元格值
    #         if hasattr(shape, 'control') and hasattr(shape.control, 'linked_cell'):
    #             linked_cell = shape.control.linked_cell
    #             if linked_cell:
    #                 ws[linked_cell].value = value
    #         break
    ws.Shapes(checkbox_name).OLEFormat.Object.Value = value
    path = config.tmp_path
    report_rid = get_uuid()
    wb.Save(path + '/'+ str(report_rid)+'.xlsx')
    wb.Close()

# 出运计划的审单状态操作
async def shipping_sdqr_check(sdqr, wxfp, rid, user, s):
    try:
        if sdqr != '结束':
            return {'code': 0, 'msg': '无需更新'}
        
        s.query(xxck).filter(xxck.xxly=='出运计划',xxck.yhdh==wxfp,or_(xxck.zt=='待审批',func.ifnull(xxck.zt,'')==''),xxck.jsr==user.username).update({'zt':'结束'}, synchronize_session=False)
        res = user_task_delete('出运计划', rid, s, [user.username])
        if res.get('code') != 1:
            s.rollback()
            return {'code': res.get('code'), 'msg': res.get('msg')}
            
        return {'code':1, 'msg': '操作成功'}
    except:
        s.rollback()
        logger.error(trace_error())
        return {'code':-1, 'msg': trace_error()}
    

def shipping_insert_shipment(rid, s, user):
    try:
        data = s.query(chyjh).filter(chyjh.rid==rid).first()
        if not data:
            return {'code':-1, 'msg':'未找到记录信息，不能进行此操作'}
        if data.dzqr1 != '结束' and data.wf_status != 2:
            return {'code':-1, 'msg':'出运计划未审批通过，不能进行此操作'}
        if data.dzry != user.username and data.ywry != user.username and data.zdry != user.username:
            return {'code':-1, 'msg':' 只有出运计划的业务人员、制单人员和单证人员才能进行此操作'}
        m = alchemy_object_to_dict(data)
        d = s.query(cymx).filter(cymx.fphm==m.get('wxfp'), cymx.chydh==m.get('fphm')).first()
        if d:
            return {'code':-1, 'msg':'已存在对应的出运明细记录，不能进行此操作'}
        flag = False
        new_rid = get_uuid()
        c = s.query(chyjhsheet).filter(chyjhsheet.pid==data.rid).all()
        for r in c:
            if r.sdjy != '不通过' and r.cywyzd != '' and (float(r.chsl) > 0 or float(r.chxs) > 0 or float(r.mjzj) != 0 or float(r.gczj) != 0 or float(r.wxzj) != 0):
                # l = s.query(cymxsheet).filter(cymxsheet.pid==d.rid, cymxsheet.cywyzd12==r.cywyzd).first()
                # if not l:
                l = cymxsheet()
                for k,v in alchemy_object_to_dict(r).items():
                    if k in SYS_FIELDS or k in ['bz','jhrq1']:
                        continue
                    if hasattr(l,k):
                        setattr(l,k,v)
                l.rid = get_uuid()
                l.pid = new_rid
                l.uid = user.rid
                l.ctime = time.strftime("%Y-%m-%d %H:%M:%S")
                l.cywyzd12 = r.cywyzd
                l.cywyzd123 = '新增资料'
                l.cght = r.cght
                l.cpbh = r.bjhh
                s.add(l)
                flag = True
        # d.qzgx20 = '出运更改识别'
        # d.zdzt = '否'
        # d.LCxb = m.get('LCxb')
        # d.xybx = m.get('xybx')
        # d.ysdj = m.get('ysdj', 0)
        # d.djyc = m.get('djyc', 0)
        # s.add(d)
        # else:
        #     dzry = m.get('dzry','')
        #     d = s.query(cymx).filter(cymx.fphm==m.get('wxfp'), or_(cymx.chydh==m.get('fphm'), func.ifnull(cymx.chydh,'')=='')).first()
        #     if not d:
        if flag==False:
            return {'code':-1, 'msg':'未找到符合生成出运明细的产品资料记录，请确认出货数量、出货箱数、采购总价、外销总价、客户RMB总价是否为0以及审单建议是否为不通过'}
        d = cymx()
        for k,v in m.items():
            if k in SYS_FIELDS or k in ['sjtx']:
                continue
            if hasattr(d,k):
                setattr(d,k,v)
        d.rid = new_rid
        d.uid = user.rid
        d.ctime = time.strftime("%Y-%m-%d %H:%M:%S")
        d.fphm = m.get('wxfp')
        d.chydh = m.get('fphm')
        d.zdry = m.get('dzry')
        d.zdzt = '否'
        d.xxyr = '引'
        d.tjry = m.get('zdry','')
        d.cpyr = '否1'
        d.LCxb = m.get('LCxb')
        d.xybx = m.get('xybx')
        d.ysdj = m.get('ysdj', 0)
        d.djyc = m.get('djyc', 0)
        d.order_id = m.get('order_id','')
        d.fyzj1 = m.get('ywewfy', 0)
        # d.chydh = m.get('fphm','')

        d.htzsl1 = m.get('htzsl', 0)
        # d.items2 = 0
        s.add(d)

        
        return {'code':1, 'msg':'操作成功'}
    except Exception as e:
        logger.error(trace_error())
        s.rollback()
        return {'code':-1, 'msg':trace_error()}


# 出运计划的订柜模板下载
@any_route('/api/saier/shipping/new/shipment', methods=['POST', 'GET'])
@require_token
async def view_saier_shipping_new_shipment(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        rid = j.get('rid', '')
        res = shipping_insert_shipment(rid, s, user)
        if res.get('code') != 1:
            s.rollback()
            return json_result(res.get('code'), res.get('msg'))
        
        s.commit()
        return json_result(1,'生成出运明细成功')
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()