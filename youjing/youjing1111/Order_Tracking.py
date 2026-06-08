import asyncio
from concurrent.futures import ThreadPoolExecutor
from any import * # 假设此模块包含了 Session, config, logger, get_uuid, require_token, any_route, json_result
from any import *
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import os
import traceback

# 用于异步执行耗时任务的线程池
executor = ThreadPoolExecutor(max_workers=1)


def _generate_contract_report_on_thread(cgry_filter, gdry_filter, user_name):
    """
    在后台线程中执行的核心逻辑，包括数据库查询和 Excel 生成。
    这个函数是同步的，因为它将在一个单独的线程中运行。
    """
    s = Session()
    try:
        # --- 1. 权限检查 (模拟 Delphi 代码) ---
        # 检查用户是否在 cyzglsheet 表中有记录，或者拥有特定职位
              # --- 1. 权限检查 (仅检查 cyzglsheet 表) ---
        # 检查用户是否在 cyzglsheet 表中有记录，该表使用 'xm' 字段存储姓名
        permission_check_query = """
            SELECT 1 FROM cyzglsheet WHERE xm = :username AND zm = '跟单统计表' LIMIT 1;
        """
        permission_result = s.execute(permission_check_query, {'username': user_name}).fetchone()
        
        if not permission_result:
            return {'code': -1, 'msg': '权限不足，无法生成报表。', 'data': None}
        # permission_result = s.execute(permission_check_query, {'username': user_name}).fetchone()
        
        # # if not permission_result:
        #     return {'code': -1, 'msg': '权限不足，无法生成报表。', 'data': None}

        # --- 2. 构建动态SQL查询条件 ---
        conditions = ["1=1"] # 一个始终为真的条件，方便后续拼接
        params = {'cgry_filter': cgry_filter, 'gdry_filter': gdry_filter}

        # 如果指定了采购人员，则添加过滤条件
        if cgry_filter:
            conditions.append("cgry = :cgry_filter")
        # 如果指定了跟单人员，则添加过滤条件
        if gdry_filter:
            conditions.append("gdry = :gdry_filter")

        where_clause = "WHERE " + " AND ".join(conditions)

        # --- 3. 查询符合条件的合同数据 ---
        # 查询所有满足条件的合同，用于后续按采购员和跟单员分组
        query = f"""
        SELECT hthm, htrq, htje, cgry, gdry
        FROM cght
        {where_clause}
        ORDER BY htrq
        """
        all_filtered_contracts = s.execute(query, params).fetchall()

        if not all_filtered_contracts:
            return {'code': -1, 'msg': '没有找到符合筛选条件的合同数据', 'data': None}

        # --- 4. 按采购员和跟单员分组数据 ---
        # 使用字典来存储每个采购员/跟单员的数据
        data_by_purchaser = {}
        data_by_handler = {}

        for row in all_filtered_contracts:
            # 按采购员分组
            if row.cgry:
                if row.cgry not in data_by_purchaser:
                    data_by_purchaser[row.cgry] = []
                data_by_purchaser[row.cgry].append(row)

            # 按跟单员分组
            if row.gdry:
                if row.gdry not in data_by_handler:
                    data_by_handler[row.gdry] = []
                data_by_handler[row.gdry].append(row)

        # --- 5. 生成 Excel 报告 ---
        from openpyxl import Workbook
        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        
        # --- 第一个工作表：采购人员统计 ---
        ws_purchase = wb.active
        ws_purchase.title = "采购人员统计"
        
        # 设置标题行和列宽
        headers = ['合同号码', '合同日期', '采购人员', '票数', '合同金额￥']
        for col_num, header in enumerate(headers, 1):
            cell = ws_purchase.cell(row=1, column=col_num, value=header)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for col_idx in range(1, len(headers) + 1):
            ws_purchase.column_dimensions[get_column_letter(col_idx)].width = 14.3

        current_row = 2
        # 遍历按采购员分组的数据
        for purchaser, contracts in sorted(data_by_purchaser.items()): # 排序以保证输出顺序一致
            total_amount_for_purchaser = 0.0
            count_for_summary = 0
            
            for contract in contracts:
                ws_purchase.cell(row=current_row, column=1, value=contract.hthm).alignment = Alignment(horizontal='center', vertical='center')
                ws_purchase.cell(row=current_row, column=2, value=str(contract.htrq)).alignment = Alignment(horizontal='center', vertical='center')
                ws_purchase.cell(row=current_row, column=3, value=purchaser).alignment = Alignment(horizontal='center', vertical='center')
                ws_purchase.cell(row=current_row, column=4, value=1).alignment = Alignment(horizontal='center', vertical='center')
                
                amount = float(contract.htje) if contract.htje else 0.0
                ws_purchase.cell(row=current_row, column=5, value=amount).alignment = Alignment(horizontal='center', vertical='center')
                
                total_amount_for_purchaser += amount
                count_for_summary += 1
                current_row += 1
            
            # 写入小计行
            if count_for_summary > 0:
                ws_purchase.cell(row=current_row, column=3, value="小计").alignment = Alignment(horizontal='center', vertical='center')
                ws_purchase.cell(row=current_row, column=4, value=count_for_summary).alignment = Alignment(horizontal='center', vertical='center')
                ws_purchase.cell(row=current_row, column=5, value=total_amount_for_purchaser).alignment = Alignment(horizontal='center', vertical='center')
                current_row += 2 # 留空一行，准备下一个采购员的数据


        # --- 第二个工作表：跟单人员统计 ---
        ws_handle = wb.create_sheet(title="跟单人员统计")
        
        # 设置标题行和列宽
        for col_num, header in enumerate(headers, 1): # headers 与采购表相同
            cell = ws_handle.cell(row=1, column=col_num, value=header.replace('采购人员', '跟单人员')) # 替换标题
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for col_idx in range(1, len(headers) + 1):
            ws_handle.column_dimensions[get_column_letter(col_idx)].width = 14.3

        current_row = 2
        # 遍历按跟单员分组的数据
        for handler, contracts in sorted(data_by_handler.items()): # 排序以保证输出顺序一致
            total_amount_for_handler = 0.0
            count_for_summary = 0
            
            for contract in contracts:
                ws_handle.cell(row=current_row, column=1, value=contract.hthm).alignment = Alignment(horizontal='center', vertical='center')
                ws_handle.cell(row=current_row, column=2, value=str(contract.htrq)).alignment = Alignment(horizontal='center', vertical='center')
                ws_handle.cell(row=current_row, column=3, value=handler).alignment = Alignment(horizontal='center', vertical='center')
                ws_handle.cell(row=current_row, column=4, value=1).alignment = Alignment(horizontal='center', vertical='center')
                
                amount = float(contract.htje) if contract.htje else 0.0
                ws_handle.cell(row=current_row, column=5, value=amount).alignment = Alignment(horizontal='center', vertical='center')
                
                total_amount_for_handler += amount
                count_for_summary += 1
                current_row += 1
            
            # 写入小计行
            if count_for_summary > 0:
                ws_handle.cell(row=current_row, column=3, value="小计").alignment = Alignment(horizontal='center', vertical='center')
                ws_handle.cell(row=current_row, column=4, value=count_for_summary).alignment = Alignment(horizontal='center', vertical='center')
                ws_handle.cell(row=current_row, column=5, value=total_amount_for_handler).alignment = Alignment(horizontal='center', vertical='center')
                current_row += 2 # 留空一行，准备下一个跟单员的数据


        # --- 6. 保存文件并返回路径 ---
        path = config.tmp_path
        os.makedirs(path, exist_ok=True) # 确保目录存在
        report_rid = get_uuid() # 假设这是一个生成唯一ID的函数
        output_filename = f'{report_rid}_跟单统计报表.xlsx'
        full_output_path = os.path.join(path, output_filename)
        
        wb.save(full_output_path)

        return {'code': 1, 'msg': '报表生成成功', 'data': output_filename}

    except Exception as e:
        logger.error(traceback.format_exc())
        return {'code': -1, 'msg': f'生成报表时发生错误: {str(e)}', 'data': None}
    finally:
        s.close()


# --- FastAPI 路由处理器 ---
@any_route('/api/saier/contract/report/excel', methods=['POST'])
@require_token
async def api_saier_contract_report_excel(request):
    """
    API 接口，接收请求并触发后台的 Excel 生成任务。
    """
    user = request.current_user
    try:
        # 1. 从 POST 请求体中获取参数
        data = await request.json()
        cgry_filter = data.get('cgry_filter', '').strip() or None # 如果为空字符串则设为None
        gdry_filter = data.get('gdry_filter', '').strip() or None # 如果为空字符串则设为None
        
        # 获取当前用户
        
        user_name = request.state.user.get('user.username') if hasattr(request.state, 'user') else None
        # if not user_name:
        #     raise AttributeError("无法从请求上下文获取当前用户名。")

        # 2. 将所有参数传递给后台线程执行
        loop = asyncio.get_event_loop()
        res = await loop.run_in_executor(
            executor,
            _generate_contract_report_on_thread,
            cgry_filter, gdry_filter, user_name
        )
        return json_result(res.get('code'), res.get('msg'), res.get('data'))
    
    except AttributeError as ae:
        error_msg = f"获取用户信息失败: {str(ae)}"
        logger.error(error_msg)
        return json_result(-1, error_msg)
    except Exception as e:
        logger.error(traceback.format_exc())
        return json_result(-1, traceback.format_exc())