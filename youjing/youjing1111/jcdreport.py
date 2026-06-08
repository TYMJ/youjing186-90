from any import *
from .model import *
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors, Alignment
from openpyxl.drawing.image import Image as Image_Get
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import  XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
import math
from sqlalchemy.sql import func
import os
        
        # ✅ 改进的模板保存逻辑
import base64
import shutil
import datetime,json,time
try:
    from PIL import Image
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", 'PIL'])
try:
    from openpyxl import load_workbook
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", 'openpyxl'])

_IMPORT_TOPIC = 'WHALE_IMPORT'


# ============ PDF转换功能 ============
def convert_excel_to_pdf(excel_path, output_dir=None):
    """
    将Excel文件转换为PDF（使用 whale_report.jar）
    
    Args:
        excel_path: Excel文件的完整路径
        output_dir: PDF输出目录，默认与Excel同目录
        
    Returns:
        dict: {'success': bool, 'pdf_path': str, 'error': str}
    """
    try:
        # 确定输出目录
        if output_dir is None:
            output_dir = os.path.dirname(excel_path)
        
        # 生成PDF文件路径
        base_name = os.path.splitext(excel_path)[0]
        pdf_path = base_name + '.pdf'
        
        print(f'========== 开始PDF转换 ==========')
        print(f'源文件: {excel_path}')
        print(f'目标文件: {pdf_path}')
        
        # 检查 Java 和 JAR 文件是否存在
        if not hasattr(config, 'java_path') or not config.java_path:
            return {
                'success': False,
                'error': 'Java路径未配置（config.java_path）'
            }
        
        if not hasattr(config, 'report_jar') or not config.report_jar:
            return {
                'success': False,
                'error': 'whale_report.jar路径未配置（config.report_jar）'
            }
        
        if not os.path.exists(config.report_jar):
            return {
                'success': False,
                'error': f'whale_report.jar文件不存在: {config.report_jar}'
            }
        
        print(f'Java路径: {config.java_path}')
        print(f'JAR路径: {config.report_jar}')
        
        # 构建命令
        # console_run(config.java_path, ['-jar', config.report_jar, 'a', 'b', template, output_file, '2'])
        cmd = [
            config.java_path,
            '-jar',
            config.report_jar,
            'a',  # 占位参数
            'b',  # 占位参数
            excel_path,      # 源Excel文件
            pdf_path,        # 目标PDF文件
            '2'              # 转换类型参数
        ]
        
        print(f'执行命令: {" ".join(cmd)}')
        
        # 执行转换
        result = subprocess.run(
            cmd,
            capture_output=True,
            timeout=120,  # 2分钟超时
            text=True
        )
        
        if result.returncode != 0:
            print(f'转换失败，返回码: {result.returncode}')
            print(f'错误输出: {result.stderr}')
            return {
                'success': False,
                'error': f'PDF转换失败（返回码 {result.returncode}）：{result.stderr}'
            }
        
        # 检查PDF文件是否生成
        if not os.path.exists(pdf_path):
            return {
                'success': False,
                'error': f'PDF文件未生成（期望路径：{pdf_path}）'
            }
        
        file_size = os.path.getsize(pdf_path)
        print(f'✓ PDF转换成功')
        print(f'✓ 文件大小: {file_size:,} 字节')
        print(f'========== PDF转换完成 ==========')
        
        return {
            'success': True,
            'pdf_path': pdf_path,
            'error': None
        }
        
    except subprocess.TimeoutExpired:
        return {
            'success': False,
            'error': 'PDF转换超时（超过2分钟）'
        }
    except Exception as e:
        logger.error(trace_error())
        return {
            'success': False,
            'error': f'PDF转换异常：{str(e)}'
        }

# ============ 图片处理功能 ============
def offset_img(img, col, row, x_pad=4, y_pad=25):
    """
    精确设置图片在Excel单元格中的位置
    
    Args:
        img: openpyxl的Image对象
        col: 列位置（从1开始）
        row: 行位置（从1开始）
        x_pad: 左边距（像素）
        y_pad: 上边距（像素）
    """
    p2e = pixels_to_EMU
    h, w = img.height, img.width
    
    size = XDRPositiveSize2D(p2e(w), p2e(h))
    marker = AnchorMarker(
        col=col-1, 
        colOff=p2e(x_pad), 
        row=row-1, 
        rowOff=p2e(y_pad)
    )
    img.anchor = OneCellAnchor(_from=marker, ext=size)


def get_product_image(cpbh, temp_dir='/tmp'):
    """
    获取产品图片并保存为临时文件
    
    Args:
        cpbh: 产品编号
        temp_dir: 临时目录
        
    Returns:
        str: 图片文件路径，如果没有图片返回None
    """
    try:
        # 查询图片数据（tpmc字段可能是base64或JSON路径）
        # ✅ 修复：使用MySQL的LENGTH函数替代SQL Server的DATALENGTH
        sql = """
            SELECT tpmc FROM tpzx 
            WHERE cpbh=:cpbh 
            AND LENGTH(tpmc) > 5
            LIMIT 1
        """
        result = run_sql(sql, {"cpbh": cpbh})
        
        if len(result) == 0:
            return None
        
        tpmc = result[0].get('tpmc', '')
        if not tpmc:
            return None
        
        # 尝试解析JSON格式（如果是文件路径）
        try:
            photo_data = json.loads(tpmc)
            if isinstance(photo_data, list) and len(photo_data) > 0:
                # 格式：[{"src": "2024/01/01/xxx.jpg"}]
                file_path = photo_data[0].get('src', '')
                full_path = os.path.join(config.data_upload_path, file_path)
                
                if os.path.exists(full_path):
                    print(f'找到图片文件: {full_path}')
                    return full_path
        except:
            pass
        
        # 如果是Base64编码
        if tpmc.startswith('data:image') or len(tpmc) > 100:
            import base64
            import re
            
            # 提取base64数据
            base64_match = re.search(r'base64,(.+)', tpmc)
            if base64_match:
                base64_data = base64_match.group(1)
            else:
                base64_data = tpmc
            
            # 解码并保存
            try:
                image_data = base64.b64decode(base64_data)
                temp_file = os.path.join(temp_dir, f'{cpbh}.png')
                
                with open(temp_file, 'wb') as f:
                    f.write(image_data)
                
                print(f'Base64图片已保存: {temp_file}')
                return temp_file
            except Exception as e:
                print(f'Base64解码失败: {e}')
        
        return None
        
    except Exception as e:
        print(f'获取产品图片失败: {e}')
        return None
    
@any_route('/api/ravencloud/export_jcexcel', methods=['POST','GET']) 
@require_token
async def fnview_export_jcexcel(request):
    try:
        JSONRes = await request.json()
        s = Session()
        rid = JSONRes.get("rid", "")
      
        
        lsql='''select * from delivery where rid=:rid'''
        msql = run_sql(lsql, {"rid": rid})
        if len(msql) > 0:
            sfqy = msql[0]["sfqy"]
            zyck = msql[0]["zyck"]
            jcrq = msql[0]["jcrq"]
            asql = 'select * from deliverysheet where pid =:pid'
            asqlm = run_sql(asql, {"pid": rid})
            if len(asqlm) == 0:
                return json_result(code=-1, msg='没有进仓编号数据！')
            print('asqlm------------------------',asqlm)
            print('rid------------------------',rid)
                
            jcbh_list= []
            for item in asqlm:
               jcbh_list.append({
                'jcbh': item.get('SNID', ''),
                'cght': item.get('cght', ''),
                'csmc': item.get('SupplierShortName', ''),
                'gdry': item.get('gdry', '')
              })
               
             # 查询产品资料子表 deliveryline
            bsql = 'SELECT SNID, ItemNo, OutCartonQty1, ckmz, cktj FROM deliveryline WHERE pid=:pid'
            bsqlm = run_sql(bsql, {"pid": rid})
        
            # 组装 cpzl_list（和前端传来的格式一样）
            cpzl_list = []
            for item in bsqlm:
               cpzl_list.append({
                'jcdh': item.get('SNID', ''),
                'cpbh': item.get('ItemNo', ''),
                'ckxs': item.get('OutCartonQty1', 0),
                'ckmz': item.get('ckmz', 0),
                'cktj': item.get('cktj', 0)
            })
        
   
        # 确定模板配置
        template_config = {
            "龙和": {"file": "longhe.xlsx", "name": "宁波龙和", "start_row": 17},
            "志恒": {"file": "zhiheng.xlsx", "name": "宁波志恒", "start_row": 18},
            "汕头": {"file": "shangtouban.xlsx", "name": "汕头办", "start_row": 18},
            "义乌": {"file": "dafu.xlsx", "name": "义乌大富", "start_row": 19},
            "柳青": {"file": "dafu.xlsx", "name": "义乌大富", "start_row": 19}
        }
        
        mb = None
        for key, val in template_config.items():
            if key in zyck:
                mb = val
                break
        
        if not mb:
            return json_result(code=-1, msg='未找到对应仓库模板')
        
        template_path = os.path.join(os.path.dirname(__file__), 'templates', mb['file'])
        
        if not os.path.exists(template_path):
            return json_result(code=-1, msg=f'模板文件不存在: {mb["file"]}')
        
        path = config.get_today_upload_path()
        if not os.path.exists(path):
            make_dirs(path)
        
        # 定义样式
        border_ = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        alight_ = Alignment(horizontal='center', vertical='center')
        
        files = []
        
        for idx, jcbh in enumerate(jcbh_list):
            sql1 = "SELECT lxry, lxdh, sjhm FROM cggd WHERE hthm=:hthm AND sccj=:sccj"
            contact = run_sql(sql1, {"hthm": jcbh['cght'], "sccj": jcbh['csmc']})
            
            lxry = contact[0]['lxry'] if len(contact) > 0 else ''
            lxdh = contact[0]['lxdh'] if len(contact) > 0 else ''
            sjhm = contact[0]['sjhm'] if len(contact) > 0 else ''
            
            sql2 = "SELECT lxdh, ydhm FROM ywrybiao WHERE yhm=:yhm"
            gd_info = run_sql(sql2, {"yhm": jcbh['gdry']})
            
            gddh = gd_info[0]['lxdh'] if len(gd_info) > 0 else ''
            gdsj = gd_info[0]['ydhm'] if len(gd_info) > 0 else ''
            
            wb = load_workbook(template_path)
            ws = wb.active
            
            # 填充表头
            if '龙和' in zyck:
                ws['C4'] = jcbh['cght']
                ws['I4'] = jcbh['jcbh']
                ws['C8'] = jcbh['gdry']
                ws['E8'] = f"{gddh}/{gdsj}"
                ws['F13'] = jcrq
                ws['C14'] = jcbh['csmc']
                ws['I14'] = f"{lxry}/{lxdh}/{sjhm}"
            elif '志恒' in zyck or '汕头' in zyck:
                ws['C4'] = jcbh['cght']
                ws['I4'] = jcbh['jcbh']
                ws['C9'] = jcbh['gdry']
                ws['G9'] = f"{gddh}/{gdsj}"
                ws['C15'] = jcbh['csmc']
                ws['G14'] = jcrq
                ws['I15'] = f"{lxry}/{lxdh}/{sjhm}"
            elif '义乌' in zyck or '柳青' in zyck:
                ws['C4'] = jcbh['cght']
                ws['I4'] = jcbh['jcbh']
                ws['C10'] = jcbh['gdry']
                ws['F10'] = f"{gddh}/{gdsj}"
                ws['C16'] = jcbh['csmc']
                ws['G15'] = jcrq
                ws['H16'] = f"{lxry}/{lxdh}/{sjhm}"
            
            # 产品明细
            start_row = mb['start_row']
            items = [cpzl for cpzl in cpzl_list if cpzl['jcdh'] == jcbh['jcbh']]
            
            print(f'===== 开始处理进仓单号: {jcbh["jcbh"]} =====')
            print(f'总共找到 {len(items)} 条产品数据')
            
            # 第一步：保存模板行的样式
            from copy import copy
            template_row_height = ws.row_dimensions[start_row].height
            template_row_styles = []
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(start_row, col)
                template_row_styles.append({
                    'font': copy(cell.font) if cell.font else None,
                    'border': copy(cell.border) if cell.border else None,
                    'fill': copy(cell.fill) if cell.fill else None,
                    'alignment': copy(cell.alignment) if cell.alignment else None,
                    'number_format': cell.number_format,
                    'protection': copy(cell.protection) if cell.protection else None
                })
            
            # 第二步：如果有多条数据，插入额外的行
            print(f'模板起始行: {start_row}')
            if len(items) > 1:
                insert_count = len(items) - 1
                print(f'需要插入 {insert_count} 行')
                ws.insert_rows(start_row + 1, insert_count)
            
            # 第三步：取消所有数据行范围内的合并单元格
            end_row = start_row + len(items) - 1
            print(f'数据行范围: {start_row} 到 {end_row}')
            
            # 收集需要取消的合并单元格
            merged_to_remove = []
            for merged_range in ws.merged_cells.ranges:
                # 检查合并单元格是否与数据行有交集
                if (merged_range.min_row <= end_row and merged_range.max_row >= start_row):
                    merged_to_remove.append(str(merged_range))
            
            print(f'需要取消的合并单元格: {merged_to_remove}')
            for mr in merged_to_remove:
                try:
                    ws.unmerge_cells(mr)
                    print(f'已取消合并: {mr}')
                except Exception as e:
                    print(f'取消合并失败 {mr}: {e}')
            
            # 第四步：为所有行复制样式
            print(f'为 {len(items)} 行复制样式')
            for i in range(len(items)):
                row = start_row + i
                # 设置行高
                if template_row_height:
                    ws.row_dimensions[row].height = template_row_height
                # 复制每个单元格的样式
                for col_idx, style_data in enumerate(template_row_styles, start=1):
                    cell = ws.cell(row, col_idx)
                    if style_data['font']:
                        cell.font = copy(style_data['font'])
                    if style_data['border']:
                        cell.border = copy(style_data['border'])
                    if style_data['fill']:
                        cell.fill = copy(style_data['fill'])
                    if style_data['alignment']:
                        cell.alignment = copy(style_data['alignment'])
                    if style_data['number_format']:
                        cell.number_format = style_data['number_format']
                    if style_data['protection']:
                        cell.protection = copy(style_data['protection'])
            
            # 第五步：填充数据
            print(f'开始填充数据')
            for i, cpzl in enumerate(items):
                row = start_row + i
                print(f'正在处理第{i+1}行，row={row}，产品编号={cpzl["cpbh"]}')
                
                ckxs = float(cpzl.get('ckxs', 0) or 0)
                ckmz = float(cpzl.get('ckmz', 0) or 0)
                cktj = float(cpzl.get('cktj', 0) or 0)
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                # C列 - 产品编号 (C-D合并，写入C列)
                ws.cell(row, 3).value = cpzl['cpbh']
                ws.cell(row, 3).alignment = alight_
                ws.cell(row, 3).border = border_
                # 合并C-D列
                ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
                
                # E列 - 箱数
                ws.cell(row, 5).value = f"{int(ckxs)}CTNS"
                ws.cell(row, 5).alignment = alight_
                ws.cell(row, 5).border = border_
                ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
                # G列 - 毛重 (G-H合并，写入G列)
                ws.cell(row, 7).value = f"{ckxs * ckmz:.2f}KGS"
                ws.cell(row, 7).alignment = alight_
                ws.cell(row, 7).border = border_
                # 合并G-H列
                ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
                
                # I列 - 体积
                ws.cell(row, 9).value = f"{cktj * ckxs:.2f}CBM"
                ws.cell(row, 9).alignment = alight_
                ws.cell(row, 9).border = border_
                
                # J列 - 进仓单号 (J-K合并，写入J列)
                ws.cell(row, 10).value = cpzl['jcdh']
                ws.cell(row, 10).alignment = alight_
                ws.cell(row, 10).border = border_
                # 合并J-K列
                ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=11)
                
                print(f'===== 填充完成第{i+1}条数据到第{row}行 =====')
            
            file_name = f"{idx + 1}{mb['name']}{jcbh['jcbh']}.xlsx"
            wb.save(path + '/' + file_name)
            
            sbs_path = path[-10:]
            files.append(sbs_path + '/' + file_name)
            print('jcbh_list------------------------',jcbh_list)
        
        if len(files) == 1:
            return json_result(code=1, msg='生成报表成功', data={
                'files': files[0],
                'sfqy': sfqy,
                'zyck': zyck,'jcbh_list':jcbh_list
            })
        else:
            return json_result(code=1, msg=f'生成{len(files)}个报表成功', data={
                'files': files,
                'sfqy': sfqy,
                'zyck': zyck,'jcbh_list':jcbh_list
            })
        
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(code=-1, msg=trace_error())
    finally:
        s.close()
  
  
@any_route('/api/ravencloud/import_ckd_excel', methods=['POST'])
@require_token
async def fnview_import_ckd_excel(request):
    try:
        # 获取表单数据
        form = await request.form()
        
        # 尝试多种方式获取文件
        file = form.get('file', None)
        rid = form.get('rid', None)
        if not file:
            file = form.get('file.raw', None)
        
        if not file:
            return json_result(code=-1, msg='请选择文件')
        
        # 读取文件内容
        file_content = await file.read()
        print(f'文件大小: {len(file_content)} 字节')
        
        # 直接从内存读取Excel，不需要保存到磁盘
        from io import BytesIO
        wb = load_workbook(BytesIO(file_content))
        ws = wb.active
        
        print(f'Excel已打开，工作表: {ws.title}')
        
        result = []
        ckmc = ''  # 仓库名称
        i = 2  # 从第2行开始
        
        while True:
            cght = ws[f'D{i}'].value  # 采购合同
            cpbh = ws[f'E{i}'].value  # 产品编号
            rkrq = ws[f'T{i}'].value  # 入库日期
            jcdh = ws[f'W{i}'].value  # 进仓单号
            rksj = ws[f'V{i}'].value  # 入库时间
            ckrq = ws[f'U{i}'].value  # 出库日期
            
            # 没有数据就结束
            if not cght or not cpbh or not rkrq or not jcdh or not rksj:
                break
            
            # 其他字段
            ckxs = ws[f'K{i}'].value or 0  # 出库箱数
            csts = ws[f'I{i}'].value or ''  # 托数
            syts = ws[f'J{i}'].value or ''  # 剩余托数
            xcw = ws[f'S{i}'].value or ''  # 转移仓位
            Memo = ws[f'L{i}'].value or ''  # 备注
            wyzd2 = ws[f'AE{i}'].value or ''  # 唯一字段
            
            # 格式化入库时间
            if rksj:
                try:
                    if hasattr(rksj, 'strftime'):
                        rksj = rksj.strftime('%Y-%m-%d')
                    else:
                        rksj = str(rksj)[:10]
                except:
                    rksj = str(rksj)
            
            # 格式化日期
            if rkrq:
                try:
                    if hasattr(rkrq, 'strftime'):
                        rkrq = rkrq.strftime('%Y-%m-%d')
                    else:
                        rkrq = str(rkrq)[:10]
                except:
                    rkrq = str(rkrq)
            
            if ckrq:
                try:
                    if hasattr(ckrq, 'strftime'):
                        ckrq = ckrq.strftime('%Y-%m-%d')
                    else:
                        ckrq = str(ckrq)[:10]
                except:
                    ckrq = str(ckrq)
            else:
                from datetime import datetime
                ckrq = datetime.now().strftime('%Y-%m-%d')
            
            # 查询库存明细
            if wyzd2:
                sql = '''SELECT * FROM inventorydetail 
                         WHERE StorageDate=:StorageDate AND StorageTime=:StorageTime 
                         AND wyzd=:wyzd AND PurchaseOrderNo=:PurchaseOrderNo 
                         AND ItemNo=:ItemNo AND SNID=:SNID'''
                data = run_sql(sql, {
                    "StorageDate": rkrq,
                    "StorageTime": rksj,
                    "wyzd": wyzd2,
                    "PurchaseOrderNo": cght,
                    "ItemNo": cpbh,
                    "SNID": jcdh
                })
            else:
                sql = '''SELECT * FROM inventorydetail 
                         WHERE StorageDate=:StorageDate AND StorageTime=:StorageTime 
                         AND PurchaseOrderNo=:PurchaseOrderNo 
                         AND ItemNo=:ItemNo AND SNID=:SNID'''
                data = run_sql(sql, {
                    "StorageDate": rkrq,
                    "StorageTime": rksj,
                    "PurchaseOrderNo": cght,
                    "ItemNo": cpbh,
                    "SNID": jcdh
                })
            
            if len(data) > 0:
                row = data[0]
                ckmc = row.get('WarehouseName', '')
                
                # 计算出库总体积和总毛重
                ckxs_int = int(float(ckxs or 0))
                ck_volume = row.get('OuterVolume', 0) or 0
                ck_weight = row.get('OuterGrossWeight', 0) or 0
                
                result.append({
                    'rid': get_uuid(),
                    'pid': rid,
                    'StorageDate': row.get('StorageDate', ''),
                    'StorageTime': row.get('StorageTime', ''),
                    'InCartonQty': row.get('InCartonQty', 0),
                    'CartonQty': row.get('InCartonQty', 0),
                    'ReturnCartonQty2': row.get('ReturnCartonQty2', 0),
                    'OutCartonQty': row.get('OutCartonQty', 0),
                    'PurchaseOrderNo': row.get('PurchaseOrderNo', ''),
                    'ItemNo': row.get('ItemNo', ''),
                    'zwmc': row.get('zwmc', ''),
                    'SupplierShortName': row.get('SupplierShortName', ''),
                    'CartonQty': row.get('CartonQty', 0),
                    'Collator': row.get('Collator', ''),
                    'ckrq': ckrq,
                    'OuterLengt': row.get('OuterLength1', 0),
                    'OuterWidth': row.get('OuterWidth1', 0),
                    'OuterHeight': row.get('OuterHeight1', 0),
                    'ckzd': row.get('OuterLength', 0),
                    'ckkd': row.get('OuterWidth', 0),
                    'ckgd': row.get('OuterHeight', 0),
                    'OuterGrossWeight': row.get('OuterGrossWeight1', 0),
                    'OuterNetWeight': row.get('OuterNetWeight1', 0),
                    'ckmz': row.get('OuterGrossWeight', 0),
                    'ckjz': row.get('OuterNetWeight', 0),
                    'OuterVolume': row.get('OuterVolume1', 0),
                    'Volume': row.get('Volume1', 0),
                    'GrossWeight': row.get('GrossWeight1', 0),
                    'cktj': row.get('OuterVolume', 0),
                    'ckztj1': row.get('Volume', 0),
                    'ckzmz1': row.get('GrossWeight', 0),
                    'ckztj': ckxs_int * float(ck_volume),
                    'ckzmz': ckxs_int * float(ck_weight),
                    'PalletQty': csts,
                    'syts': syts,
                    'WarehousePosition': row.get('WarehousePosition', ''),
                    'WarehousePosition1': xcw if xcw else row.get('WarehousePosition', ''),
                    'LotNumber': row.get('LotNumber', ''),
                    'gdry': row.get('gdry', ''),
                    'Salesman': row.get('Salesman', ''),
                    'SNID': row.get('SNID', ''),
                    'cgwyzd': row.get('cgwyzd', ''),
                    'wxwyzd': row.get('wxwyzd', ''),
                    'Exporter': row.get('Exporter', ''),
                    'PurchasingAgent': row.get('PurchasingAgent', ''),
                    'SalesOrderNo': row.get('SalesOrderNo', ''),
                    'wyzd': row.get('wyzd', ''),
                    'Operator': row.get('Operator', ''),
                    'rkdd': row.get('rkdd', ''),
                    'Memo': Memo,
                    'jcf': row.get('jcf', 0),
                    'ywpath': row.get('ywpath', ''),
                    'cgpath': row.get('cgpath', ''),
                    'khmc': row.get('khmc', ''),
                    'OutCartonQty': ckxs_int
                })
            
            i += 1
        
        print(f'读取完成，共 {len(result)} 条数据')
        
        return json_result(code=1, data={'ckmc': ckmc, 'items': result}, msg='success')
        
    except Exception as e:
        
        logger.error(trace_error())
        return json_result(code=-1, msg=trace_error())
    
    
    
    


@any_route('/api/Ravencloud/export_inventorydetail_excel', methods=['POST', 'GET'])
@require_token
async def fnview_export_inventorydetail_excel(request):
    """导出库存明细到Excel"""
    try:
        JSONRes = await request.json()
        rid_list = JSONRes.get('rid_list', [])  # 前端传入的rid列表
        username = JSONRes.get('username', '')
        
        # 获取用户路径
        path = ''
        org = get_user_path(name)
        path = org.get('path','')
        path = path[:100] 
       
        # 查询数据
        data_list = []
        for rid in rid_list:
            sSql = """
                SELECT * FROM inventorydetail 
                WHERE sys_path = :sys_path AND rid = :rid
            """
            result = run_sql(sSql, {'sys_path': path, 'rid': rid})
            
            if result:
                for row in result:
                    # 查询业务部门
                    ywbm = ''
                    ywbm_sql = "SELECT bm FROM ywrybiao WHERE yhm = :yhm"
                    ywbm_result = run_sql(ywbm_sql, {'yhm': row.get('Salesman', '')})
                    if ywbm_result:
                        ywbm = ywbm_result[0]['bm']
                    
                    data_list.append({
                        'OuterLength': row.get('OuterLength', ''),
                        'OuterWidth': row.get('OuterWidth', ''),
                        'OuterHeight': row.get('OuterHeight', ''),
                        'PurchaseOrderNo': row.get('PurchaseOrderNo', ''),
                        'ItemNo': row.get('ItemNo', ''),
                        'OuterGrossWeight': row.get('OuterGrossWeight', ''),
                        'OuterNetWeight': row.get('OuterNetWeight', ''),
                        'CartonQty': row.get('CartonQty', ''),
                        'PalletQty': row.get('PalletQty', ''),
                        'Memo': row.get('Memo', ''),
                        'OuterVolume': row.get('OuterVolume', ''),
                        'Volume': row.get('Volume', ''),
                        'GrossWeight': row.get('GrossWeight', ''),
                        'Collator': row.get('Collator', ''),
                        'LotNumber': row.get('LotNumber', ''),
                        'WarehousePosition': row.get('WarehousePosition', ''),
                        'StorageDate': row.get('StorageDate', ''),
                        'StorageTime': row.get('StorageTime', ''),
                        'SNID': row.get('SNID', ''),
                        'SupplierShortName': row.get('SupplierShortName', ''),
                        'PurchasingAgent': row.get('PurchasingAgent', ''),
                        'gdry': row.get('gdry', ''),
                        'Salesman': row.get('Salesman', ''),
                        'ywbm': ywbm,
                        'wyzd': row.get('wyzd', '')
                    })
        
        # 创建Excel文件
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '出库单'
        
        # 写入表头
        headers = [
            '仓库长度', '仓库宽度', '仓库高度', '采购合同', '产品编号', '仓库毛重', '仓库净重',
            '在仓箱数1', '托数', '剩余托数', '出库箱数', '备注信息', '仓库体积', '在仓总体积',
            '在仓总毛重', '理货员', '批号', '仓位', '转移仓位', '入库日期', '出库日期',
            '入库时间', '进仓单号', '厂商名称', '采购员', '跟单人员', '业务员', '业务部门',
            '1', '2', '唯一字段'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # 写入数据
        for row_idx, data in enumerate(data_list, start=2):
            ws.cell(row=row_idx, column=1, value=data['OuterLength'])
            ws.cell(row=row_idx, column=2, value=data['OuterWidth'])
            ws.cell(row=row_idx, column=3, value=data['OuterHeight'])
            ws.cell(row=row_idx, column=4, value=data['PurchaseOrderNo'])
            ws.cell(row=row_idx, column=5, value=data['ItemNo'])
            ws.cell(row=row_idx, column=6, value=data['OuterGrossWeight'])
            ws.cell(row=row_idx, column=7, value=data['OuterNetWeight'])
            ws.cell(row=row_idx, column=8, value=data['CartonQty'])
            ws.cell(row=row_idx, column=9, value=data['PalletQty'])
            ws.cell(row=row_idx, column=10, value='')  # 剩余托数
            ws.cell(row=row_idx, column=11, value='')  # 出库箱数
            ws.cell(row=row_idx, column=12, value=data['Memo'])
            ws.cell(row=row_idx, column=13, value=data['OuterVolume'])
            ws.cell(row=row_idx, column=14, value=data['Volume'])
            ws.cell(row=row_idx, column=15, value=data['GrossWeight'])
            ws.cell(row=row_idx, column=16, value=data['Collator'])
            ws.cell(row=row_idx, column=17, value=data['LotNumber'])
            ws.cell(row=row_idx, column=18, value=data['WarehousePosition'])
            ws.cell(row=row_idx, column=19, value='')  # 转移仓位
            ws.cell(row=row_idx, column=20, value=data['StorageDate'])
            ws.cell(row=row_idx, column=21, value='')  # 出库日期
            ws.cell(row=row_idx, column=22, value=data['StorageTime'])
            ws.cell(row=row_idx, column=23, value=data['SNID'])
            ws.cell(row=row_idx, column=24, value=data['SupplierShortName'])
            ws.cell(row=row_idx, column=25, value=data['PurchasingAgent'])
            ws.cell(row=row_idx, column=26, value=data['gdry'])
            ws.cell(row=row_idx, column=27, value=data['Salesman'])
            ws.cell(row=row_idx, column=28, value=data['ywbm'])
            ws.cell(row=row_idx, column=29, value='')
            ws.cell(row=row_idx, column=30, value='')
            ws.cell(row=row_idx, column=31, value=data['wyzd'])
        
        # 添加合计行
        total_row = len(data_list) + 2
        ws.cell(row=total_row, column=1, value='合计')
        ws.cell(row=total_row, column=8, value=f'=SUM(H2:H{total_row-1})')
        ws.cell(row=total_row, column=14, value=f'=SUM(N2:N{total_row-1})')
        ws.cell(row=total_row, column=15, value=f'=SUM(O2:O{total_row-1})')
        
        # 设置列宽
        column_widths = {
            'A': 10, 'B': 10, 'C': 10, 'D': 25, 'E': 18, 'F': 10, 'G': 10,
            'H': 10, 'I': 10, 'J': 10, 'K': 10, 'L': 50, 'M': 10, 'N': 10,
            'O': 10, 'P': 10, 'Q': 10, 'R': 10, 'S': 10, 'T': 10, 'U': 10,
            'V': 10, 'W': 16, 'X': 30, 'Y': 10, 'Z': 10, 'AA': 10, 'AB': 23
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 设置边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws[f'A2:AB{total_row}']:
            for cell in row:
                cell.border = thin_border
        
        # 保存文件
        today = datetime.now().strftime('%Y-%m-%d')
        filename = f'/tmp/{today}出库单.xlsx'
        wb.save(filename)
        
        return json_result(code=1, data={'filename': filename}, msg='导出成功')
        
    except Exception as e:
        logger.error(trace_error())
        return json_result(code=-1, msg=trace_error())
    
    
@any_route('/api/Ravencloud/export_cggd_report', methods=['POST','GET']) 
@require_token
async def fnview_export_cggd_report(request):
    """导出采购跟单进仓单报表 - 完全对应Delphi逻辑"""
    try:
        JSONRes = await request.json()
        s = Session()
        rid = JSONRes.get("rid", "")
        tp = JSONRes.get("tp", "1")
        print('========== 开始导出采购跟单进仓单报表 ==========')
        print(f'导出类型: {tp}')
        print(f'RID: {rid}')
 
        
        if not rid:
            return json_result(code=-1, msg='单号不能为空')
        
        print(f'========== 开始导出报表 ==========')
        print(f'导出类型: {tp}')
        print(f'RID: {rid}')
        
        # 查询主表数据
        lsql = 'SELECT * FROM cggd WHERE rid=:rid'
        msql = run_sql(lsql, {"rid": rid})
        
        if len(msql) == 0:
            return json_result(code=-1, msg='未找到相关数据')
        print(f'msql {msql}')
        main_row = msql[0]
        ckmc = main_row.get('ckmc', '')
        jcbh = main_row.get('jcbh', '')
        wfgs = main_row.get('wfgs', '')
        hthm = main_row.get('hthm', '')
        gdry = main_row.get('gdry', '')
        jcsj = main_row.get('jcsj', '')
        sccj = main_row.get('sccj', '')
        cs_id = main_row.get('cs_id', '')
        lxry = main_row.get('lxry', '')
        lxdh = main_row.get('lxdh', '')
        sjhm = main_row.get('sjhm', '')
        shc = main_row.get('shc', '')
        
        print(f'仓库名称: {ckmc}')
        print(f'进仓编号: {jcbh}')
        
        if not ckmc:
            return json_result(code=-1, msg='仓库名称为空')
        
        # 确定模板配置
        template_config = {
            "宁波龙和": {"file": "ningbolh.xlsx", "start_row": 17},
            "宁波万纬": {"file": "ningboww.xlsx", "start_row": 17},
            "宁波志恒": {"file": "nbzh.xlsx", "start_row": 18},
            "汕头仓库": {"file": "shantou.xlsx", "start_row": 18},
            "义乌仓库": {"file": "yiwu.xlsx", "start_row": 19}
        }
        
        mb = template_config.get(ckmc)
        if not mb:
            return json_result(code=-1, msg=f'未找到仓库 {ckmc} 对应的模板')
        
        template_path = os.path.join(os.path.dirname(__file__), 'templates', mb['file'])
        
        if not os.path.exists(template_path):
            return json_result(code=-1, msg=f'模板文件不存在: {mb["file"]}')
        
        path = config.get_today_upload_path()
        if not os.path.exists(path):
            make_dirs(path)
        
        # 查询厂商简称
        csjc_sql = 'SELECT csjc FROM zycs WHERE company_name=:company_name OR cs_id=:cs_id'
        csjc_data = run_sql(csjc_sql, {"company_name": sccj, "cs_id": cs_id})
        csjc = csjc_data[0]['csjc'] if len(csjc_data) > 0 else ''
        
        # 查询跟单人员信息
        sql_gd = "SELECT lxdh, ydhm FROM ywrybiao WHERE yhm=:yhm"
        gd_info = run_sql(sql_gd, {"yhm": gdry})
        gd_ydhm = gd_info[0]['ydhm'] if len(gd_info) > 0 else ''
        
        # 查询明细数据
        print(f'========== 开始查询明细数据 ==========')
        asql = 'SELECT cpbh, jcxs, htzmz, htztj FROM cggdsheet1 WHERE pid=:father'
        asqlm = run_sql(asql, {"father": rid})
        
        if len(asqlm) == 0:
            asql = 'SELECT cpbh, jcxs, htzmz, htztj FROM cggdsheet1 WHERE pid=:pid'
            asqlm = run_sql(asql, {"pid": rid})
        
        if len(asqlm) == 0:
            return json_result(code=-1, msg='未找到明细数据')
        
        print(f'查询到 {len(asqlm)} 条明细数据')
        
        # 加载模板
        wb = load_workbook(template_path)
        ws = wb.active
        
        print(f'模板加载完成: {ws.title}')
        
        # 填充主表数据
        ws['A1'] = wfgs
        ws['C4'] = hthm
        ws['I4'] = jcbh
        
        contact_info = f"{lxry}/{lxdh}/{sjhm}"
        start_row = mb['start_row']
        number = start_row
        
        print(f'起始行号: {number}')
        
        # 根据不同仓库填充表头
        if ckmc == '义乌仓库':
            ws['C10'] = gdry
            ws['E10'] = gd_ydhm
            ws['G15'] = jcsj
            ws['C16'] = sccj
            ws['N16'] = sccj
            ws['O16'] = contact_info
            ws['H16'] = contact_info
            if shc:
                ws['C17'] = shc
        elif ckmc in ['宁波志恒', '汕头仓库']:
            ws['C9'] = gdry
            ws['E9'] = gd_ydhm
            ws['G14'] = jcsj
            ws['C15'] = sccj
            ws['N15'] = sccj
            ws['O15'] = contact_info
            ws['I15'] = contact_info
            if shc:
                ws['C16'] = shc
        elif ckmc in ['宁波龙和', '宁波万纬']:
            ws['C8'] = gdry
            ws['E8'] = gd_ydhm
            ws['F13'] = jcsj
            ws['C14'] = sccj
            ws['N14'] = sccj
            ws['O14'] = contact_info
            ws['I14'] = contact_info
            if shc:
                ws['C15'] = shc
        
        print(f'========== 表头填充完成，开始处理明细 ==========')
        
        # 处理明细数据
        from copy import copy
        
        k = 0
        
        for i, item in enumerate(asqlm):
            k = k + 1
            
            print(f'\n========== 处理第 {k} 条数据 ==========')
            
            # 插入新行
            insert_pos = number + k
            print(f'在位置 {insert_pos} 插入新行')
            ws.insert_rows(insert_pos, 1)
            
            # 复制模板行样式
            for col in range(1, 15):
                source_cell = ws.cell(number, col)
                target_cell = ws.cell(insert_pos, col)
                
                if source_cell.font:
                    target_cell.font = copy(source_cell.font)
                if source_cell.border:
                    target_cell.border = copy(source_cell.border)
                if source_cell.fill:
                    target_cell.fill = copy(source_cell.fill)
                if source_cell.alignment:
                    target_cell.alignment = copy(source_cell.alignment)
                if source_cell.number_format:
                    target_cell.number_format = source_cell.number_format
            
            # 设置行高
            ws.row_dimensions[insert_pos].height = 46
            
            # 获取数据
            cpbh = item.get('cpbh', '')
            jcxs = float(item.get('jcxs', 0) or 0)
            htzmz = float(item.get('htzmz', 0) or 0)
            htztj = float(item.get('htztj', 0) or 0)
            
            # 查询产品品名
            jcpm_sql = 'SELECT jcpm FROM zscp WHERE cpbh=:cpbh OR krhh=:krhh'
            jcpm_data = run_sql(jcpm_sql, {"cpbh": cpbh, "krhh": cpbh})
            jcpm = jcpm_data[0]['jcpm'] if len(jcpm_data) > 0 else ''
            
            # 填充数据
            ws.cell(insert_pos, 1).value = jcpm
            ws.cell(insert_pos, 3).value = cpbh
            ws.cell(insert_pos, 5).value = f"{int(jcxs)}CTNS"
            ws.cell(insert_pos, 7).value = f"{htzmz:.2f}KGS"
            ws.cell(insert_pos, 9).value = f"{htztj:.2f}CBM"
            ws.cell(insert_pos, 10).value = jcbh
            ws.cell(insert_pos, 13).value = cpbh
            
            print(f'第 {k} 条数据填充完成')
            
            # ========== 图片插入逻辑（仅普通类型） ==========
            if tp not in ['2', '3', '4']:
                print(f'开始处理产品图片: {cpbh}')
                
                image_path = get_product_image(cpbh, temp_dir=path)
                
                if image_path and os.path.exists(image_path):
                    try:
                        # 加载图片
                        img = Image_Get(image_path)
                        
                        # 调整图片大小
                        img.width = 120
                        img.height = 80
                        
                        # 插入到A列（产品名称旁边）
                        offset_img(img, 2, insert_pos, 5, 5)
                        ws.add_image(img)
                        
                        print(f'  ✓ 图片已插入')
                        
                        # 如果是临时文件，使用后删除
                        if image_path.startswith('/tmp') or image_path.startswith(path):
                            try:
                                os.remove(image_path)
                            except:
                                pass
                                
                    except Exception as e:
                        print(f'  ✗ 图片插入失败: {e}')
                else:
                    print(f'  - 未找到产品图片')
        
        print(f'\n========== 所有数据填充完成 ==========')
        
        # 删除模板行
        ws.delete_rows(number, 1)
        print(f'模板行已删除')
        
        # 生成文件
        from datetime import datetime
        sj = datetime.now().strftime('%Y%m%d%H%M%S')
        
        file_name = f"{csjc}{jcbh}{ckmc}{sj}.xlsx"
        full_path = os.path.join(path, file_name)
        
        print(f'\n========== 保存文件 ==========')
        print(f'文件名: {file_name}')
        
        wb.save(full_path)
        print(f'✓ Excel文件已保存: {full_path}')
        
        # ========== PDF转换逻辑 ==========
        final_file_path = full_path
        final_file_name = file_name
        
        if tp in ['2', '4']:
            print(f'\n========== 开始PDF转换（类型: {tp}） ==========')
            
            pdf_result = convert_excel_to_pdf(full_path, path)
            
            if pdf_result['success']:
                pdf_full_path = pdf_result['pdf_path']
                pdf_file_name = file_name.replace('.xlsx', '.pdf')
                
                final_file_path = pdf_full_path
                final_file_name = pdf_file_name
                
                print(f'✓ PDF生成成功')
            else:
                print(f'✗ PDF转换失败: {pdf_result["error"]}')
                return json_result(code=-1, msg=pdf_result['error'])
        
        # 生成返回路径
        sbs_path = path[-10:]
        file_path = sbs_path + '/' + final_file_name
        
        print(f'最终返回路径: {file_path}')
        print(f'========== 导出完成 ==========\n')
        
        return json_result(code=1, msg='生成报表成功', data={
            'files': file_path,
            'ckmc': ckmc,
            'jcbh': jcbh,
            'tp': tp
        })
        
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(code=-1, msg=trace_error())
    finally:
        s.close()
        

    """将阿拉伯数字金额转换为中文大写"""
    try:
        amount = float(amount)
        
        # 中文数字
        chn_num = ['零', '壹', '贰', '叁', '肆', '伍', '陆', '柒', '捌', '玖']
        chn_unit = ['', '拾', '佰', '仟', '万', '拾万', '佰万', '仟万', '亿']
        
        # 分离整数和小数部分
        int_part = int(amount)
        dec_part = int(round((amount - int_part) * 100))
        
        # 转换整数部分
        result = ''
        if int_part == 0:
            result = '零'
        else:
            str_int = str(int_part)
            for i, digit in enumerate(str_int):
                unit_idx = len(str_int) - i - 1
                if digit != '0':
                    result += chn_num[int(digit)] + (chn_unit[unit_idx] if unit_idx < len(chn_unit) else '')
                elif i > 0 and str_int[i-1] != '0':
                    result += '零'
        
        result += '元'
        
        # 转换小数部分
        if dec_part == 0:
            result += '整'
        else:
            jiao = dec_part // 10
            fen = dec_part % 10
            
            if jiao > 0:
                result += chn_num[jiao] + '角'
            if fen > 0:
                result += chn_num[fen] + '分'
        
        return result
    except:
        return f'{amount}元'