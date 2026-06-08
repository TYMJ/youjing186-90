import asyncio
from concurrent.futures import ThreadPoolExecutor
from any import *  
from .model import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors, Alignment
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
import traceback
import io # 需要导入 io 模块来处理内存流
from openpyxl.utils import coordinate_to_tuple
from email import errors
from math import e
from pdb import run
from webbrowser import get
from any import *
from .model import *
from sqlalchemy.sql import func,not_,or_,and_
import time
import json,os
from .__default__ import get_user_path,module_xxck_new,module_share_new
from openpyxl.drawing.image import Image as Image_Get
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, Alignment
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from PIL import Image as PILImage  # 重命名避免冲突

from PIL import JpegImagePlugin
import logging



executor = ThreadPoolExecutor(max_workers=4)


def crate_special_report_logic(recordset,zm):
    """
    同步的报表生成逻辑
    """
   
    s = Session()
    
    tlist1=[]
    temp_files = []
    main_info={}
    d=s.query(bgmxd).filter(bgmxd.rid==recordset.get('rids')[0]).first()
    if d:
        main_info = alchemy_object_to_dict(d)

    logger.error(f"报表生成开始，参数 zm: {zm}, recordset:zm {recordset.get('zm', '')}, rids: {recordset.get('rids', [])}")
    try:
        zm=recordset.get('zm', '')
            
        if zm =='':
            zm = 'DOLLARAMA 南美-INV&PL'
        a = run_sql("select zm from cyzgl where qxzl='特殊报表打印出运明细'")
        if a :
            for i in a:
                tlist1.append(i['zm'])
        b = run_sql("select tpl,ywxh,qsh,sfdhy,rid,invsz,plsz,wb from cyzgl where (qxzl='特殊报表打印出运明细') and (zm=:zm)", {'zm': zm})
        if b:
            
            qnd=b[0]
            
            
            qsh = int(qnd.get('qsh', 0))
            tpl = qnd.get('tpl', '')
            ywxh = qnd.get('ywxh', '')
            sfdhy = qnd.get('sfdhy', '')
            number = qnd.get('rid', '')
            invsz = int(qnd.get('invsz', 0))
            plsz = int(qnd.get('plsz', 0))
            wa = qnd.get('wb', '')
        
        template_tempfile = None
       
        rid_val = []
        logger.error(f"参数 zm: {zm}")
        f = run_sql("SELECT rid FROM cyzgl WHERE qxzl like '%特殊报表打印%'AND zm=:zm limit 1" , { "zm": zm})
        logger.error(f"查询到的记录: {f}")
        rid_val=(f[0].get('rid', ''))
        if rid_val:
            e = run_sql("SELECT path FROM sys_attachment WHERE pid=:rid AND path IS NOT NULL LIMIT 1", {"rid":rid_val})
            report_path = e[0].get('path', '')
        logger.error(f"查询到的报表路径: {report_path}")
        if not report_path:
            return json_result(-1, f"未找到报表模板: {zm}")
       
        
        fn = os.path.join(config.data_upload_path,report_path)
        logger.error(f"构建的文件路径: {fn}")
        if not os.path.exists(fn):
            return json_result(-1,"未找到报表模板")

        wb = load_workbook(fn)
        # c=run_sql("select seq from  :wb  where pid=:rids ",{"rids": recordset.get('rids'),"wb":wa})
      
        c = run_sql(f"select seq from {wa} where pid =:rids", {"rids": recordset.get('rids')})

        i1z = 0  # 初始化行数计数器
        if c:
   
            i1z = len(c)
        ws = wb.worksheets[0] 
        for si in range(1, 2):
            i4 = 0  # 累计偏移量计数器
            
          
            if si == 1:
                i = qsh - 1
                i2 = invsz  # INV 表的列数/行数配置
            else:
                i = qsh - 1
                i2 = plsz   # PL 表的列数/行数配置
            
         
            if sfdhy != '是':
                i2 += 1
            logger.error(f"当前工作表索引: {si}, 初始行号: {i}, 需要插入的行数: {i1z - 3}, 每行数据占用的行数: {i2}")
           
            ws = wb.worksheets[si - 1]
            
            if i1z > 3:
              
                for i1 in range(1, i1z - 2):  
                  
                    for i3 in range(1, i2 + 1):
                        i4 += 1
                       
                        insert_row_index = qsh + 1 
                        ws.insert_rows(insert_row_index)
                        
                        source_row = ws[qsh]
                        target_row = ws[qsh + 1]
                        
                       
                        for cell in source_row:
                            target_cell = ws.cell(row=qsh + 1, column=cell.column)
                            if cell.has_style:
                                target_cell.font = copy(cell.font)
                                target_cell.border = copy(cell.border)
                                target_cell.fill = copy(cell.fill)
                                target_cell.number_format = cell.number_format
                                target_cell.alignment = copy(cell.alignment)
                        ws[f'A{qsh + 1}'] = ''
        d =run_sql("select * from cymx where rid=:rid",{"rid": recordset.get('rids')})
        if d:
            main_record = d[0]
           
            e = """
                select xm, bz, bz1, bz3, bz2, sz, sz1, bz12 
                from cyzglsheet 
                where (zm=:zm) and (sz=:sz) and (pid=:rid) and (bz6='主')
            """
            mapping_data = run_sql(e, {
                "zm": zm, 
                "rid": number, 
                "sz": si
            })
            logger.error(f"查询到的映射数据: {mapping_data}")
            if mapping_data:  # 对应 Delphi 的 if tmpcom1.isempty=false
                for idx in mapping_data:  # 对应 while not tmpcom1.eof 循环
                    # 提取配置字段
                    xm = idx.get('xm', '')      # 单元格坐标 (如 "C5")
                    bz = idx.get('bz', '')      # 数据库字段名
                    bz1 = idx.get('bz1', '')    # 前缀字符串
                                    
                    ws[xm].value = bz1 + str(idx.get(bz, ''))
                logger.error(f"xm: {xm}")
            if wa == 'cymxsheet4':
                detail_sql = f"select * from {wa} where (pid=:rid) AND ((xs>0) or (sfpx='是'))"
            else:
                detail_sql = f"select * from {wa} where (pid=:rid) AND ((chxs>0) or (sfpx='是'))"

            detail_data = run_sql(detail_sql, {"rid": recordset.get('rids')})

            if detail_data:
                # i 代表当前 Excel 的实际行号，初始值由之前的逻辑决定（通常是 qsh）
                for row_data in detail_data:  # 对应 while not tmpcom.eof
                    
                    # 查询当前明细行在 Excel 中的字段映射规则 (bz6="子" 代表明细行配置)
                    detail_mapping_sql = """
                        select xm, bz, bz1, bz3, bz2, sz, sz1, bz12, bz5, bz4 
                        from cyzglsheet 
                        where (zm=:zm) and (sz=:sz) and (pid=:rid) and (bz6='子')
                    """
                    detail_configs = run_sql(detail_mapping_sql, {
                        "zm": zm, "rid": number, "sz": si
                    })
                    logger.error(f"查询到的明细映射配置: {detail_configs}")
                    if detail_configs:
                        for item in detail_configs:
                            bz5 = item.get('bz5', '')
                            bz3 = item.get('bz3', '') 
                            bz = item.get('bz', '')   
                            xm = item.get('xm', '')    
                            bz1 = item.get('bz1', '')  
                            bz4 = item.get('bz4', '')  
                            if bz5 == '加':
                                i += 1   
                                logger.error(f"插入行后当前行号: {i}")                   
                            ws.row_dimensions[i].auto_size = True
                            if bz: 
                                if bz3 == 'float':  
                                    target_field = 'wxdj' if (bz == 'wxjg' and wa == 'cymxsheet4') else bz3
                                    ws[f"{xm}{i}"].value = float(row_data.get(target_field, 0))
                                elif bz3 == 'djpmw':  
                                    pm_sql = """
                                        select djpmw, djpmw1, cpbh from wypm 
                                        where ((cpbh=:cpbh) or ((krhh=:cpbh1) and (krhh<>'') and (krhh IS NOT NULL)))
                                    """
                                    pm_data = run_sql(pm_sql, {
                                        "cpbh": row_data.get('zycpbh', ''),
                                        "cpbh1": row_data.get('cpbh', '')
                                    })
                                    if pm_data:
                                        ws[f"{xm}{i}"].value = pm_data[0].get('djpmw')
                                    else:
                                        ws[f"{xm}{i}"].value = bz1 + str(row_data.get(bz3, '')) + bz4
                                else:
                                    ws[f"{xm}{i}"].value = bz1 + str(row_data.get(bz3, '')) + bz4
                            else:
                                ws[f"{xm}{i}"].value = bz1
                    if sfdhy != '是':
                        i += 1
        output_filename = os.path.join( f"{zm}_output.xlsx")
        wb.save(output_filename)
        wb.close()
        path = config.tmp_path
        os.makedirs(path, exist_ok=True) # 确保目录存在
        full_output_path = os.path.join(path, output_filename)
        s_path = config.tmp_path
        wb.save(full_output_path)
        return {'code': 1, 'msg': '报表生成成功', 'data': output_filename}
    except Exception as e:
        logger.error(f"生成Excel失败: {str(e)}")
        logger.error(traceback.format_exc())
        return {'code': -1, 'msg': f'生成Excel失败: {str(e)}', 'data': ''}
    finally:
        s.close()
       

# API 路由 (修正版)
@any_route('/api/shipping/export/detailed/special/list', methods=['POST'])
@require_token
async def special_report_list(request):
    """
    特殊报表打印出运明细 (修正版)
        """
    try:
        j = await request.json()
        zm=j.get('zm', '')
        result = crate_special_report_logic(j, zm)
        
        if result['code'] == 1:
            # 这里假设返回的是文件路径，实际可能需要读取文件流返回
            file_path = result['data']
            if os.path.exists(file_path):
               
                return json_result(1, result['msg'],  file_path)
            else:
                return json_result(-1, "文件生成失败或未找到", {})
        else:
            return json_result(result['code'], result['msg'], {})
            
    except Exception as e:
        logger.error(f"API调用失败: {str(e)}")
        logger.error(traceback.format_exc())
        return json_result(-1, f'API调用失败: {str(e)}', {})