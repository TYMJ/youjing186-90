from ast import mod
from webbrowser import get

from any import *
from .model import *
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Optional, Tuple
import os
from .__default__ import module_share_new, module_xxck_new, get_user_path, user_task_new
from sqlalchemy.sql import func, not_, or_, and_
from collections import OrderedDict,defaultdict
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from datetime import datetime
import time
import re
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill


# 安装中文数字转换库

try:
    import cn2an
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "cn2an"])
    import cn2an
    

def split_payment_date(fkrq):
    """把付款日期拆分为年、月、日字符串。"""
    text = str(fkrq or "").strip()
    if text == "":
        return "", "", ""

    digits = re.sub(r"[^0-9]", "", text)
    if len(digits) >= 8:
        return digits[0:4], digits[4:6], digits[6:8]
    return "", "", ""

# 日期批量小写
@any_route("/api/bill/advance_payment/sync_payment_date_parts", methods=["POST", "GET"])
@require_token
async def advance_payment_sync_payment_date_parts(request):
    """
    预付货款-同步付款日期拆分字段（nsz/ysz/rsz）。
    """
    try:
        j = await request.json()
        user = request.current_user
        user_name = user.username

        rids = j.get("rids", [])
        if not isinstance(rids, list):
            rids = []
        if len(rids) == 0:
            rid = j.get("rid", "")
            if rid:
                rids = [rid]
        if len(rids) == 0:
            return json_result(-1, "请选择需要处理的数据")

        user_name_safe = esc_sql(user_name)
        sql_user = f"""select * from sys_user where username='{user_name_safe}'"""
        rows_user = run_sql(sql_user) or []
        if len(rows_user) == 0:
            return json_result(-1, "无操作权限")

        user_rid = str(rows_user[0].get("rid", "") or "")
        if user_rid == "":
            return json_result(-1, "无操作权限")

        sql_user_role = f"""select * from sys_user_role where user_id='{esc_sql(user_rid)}'"""
        rows_user_role = run_sql(sql_user_role) or []
        if len(rows_user_role) == 0:
            return json_result(-1, "无操作权限")

        role_rid = str(rows_user_role[0].get("role_id", "") or "")
        sql_role = f"""select * from sys_role where rid='{esc_sql(role_rid)}' and ((name like '%财务%'))"""
        rows_role = run_sql(sql_role) or []
        if len(rows_role) == 0:
            return json_result(-1, "无操作权限")

        form_fields = j.get("form_fields", {})
        if not isinstance(form_fields, dict):
            form_fields = {}

        updated_count = 0
        skipped_count = 0

        for rid in rids:
            rid_safe = esc_sql(rid)
            if rid_safe == "":
                skipped_count += 1
                continue

            sql_bill = f"""select * from yfhk where rid='{rid_safe}'"""
            rows_bill = run_sql(sql_bill) or []
            if len(rows_bill) == 0:
                skipped_count += 1
                continue

            row = rows_bill[0]
            fkrq = str(row.get("fkrq", "") or "")

            # 单条处理时优先使用前端实时值，避免保存前界面值丢失。
            if len(rids) == 1:
                fkrq_form = str(form_fields.get("fkrq", "") or "")
                if fkrq_form != "":
                    fkrq = fkrq_form

            nsz, ysz, rsz = split_payment_date(fkrq)
            if nsz == "" or ysz == "" or rsz == "":
                skipped_count += 1
                continue

            sql_update = f"""
                update yfhk
                set nsz='{esc_sql(nsz)}', ysz='{esc_sql(ysz)}', rsz='{esc_sql(rsz)}'
                where rid='{rid_safe}'
            """
            run_sql(sql_update)
            updated_count += 1

        data = {
            "updated_count": updated_count,
            "skipped_count": skipped_count
        }
        return json_result(1, "操作成功", data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        pass



# 预付货款导出
@any_route("/api/bill/advance_payment/yfhk/export_excel", methods=["POST"])
@require_token
async def advance_payment_export_excel(request):
    """按 oldCode 逻辑导出预付货款数据到 Excel。"""
    try:
        j = await request.json()
        rid_list = j.get("rids") or []

        if not rid_list:
            return json_result(-1, "未选择导出记录", None)
        
        template_path = os.path.join(config.data_upload_path, "template/预付货款导出.xlsx")
        if not os.path.exists(template_path):
            return json_result(-1, "预付货款导出模板文件不存在", None)

        wb = load_workbook(template_path)
        ws = wb.worksheets[0]
        ws.title = "预付货款导出"

        # Pascal 从第 3 行开始写数据
        start_row = 3
        field_order = [
            "sb", "cght", "fkje", "htje", "sqje", "fkbz", "fkrq",
            "jsrm", "sccj", "gcmc1", "yhzh", "khh", "sfjq", "bzsm"
        ]

        for idx, rid_value in enumerate(rid_list):
            row_no = start_row + idx
            row_data = query_yfhk_by_rid(rid_value)
            if not row_data:
                continue

            for col_idx, field_name in enumerate(field_order, start=1):
                logger.error(field_name)
                ws.cell(row=row_no, column=col_idx, value=row_data.get(field_name, ""))

        file_name = f"预付货款导出_{get_uuid()}.xlsx"
        file_path = os.path.join(config.tmp_path, file_name)
        wb.save(file_path)

        return json_result(1, "success", {
            "file_name": "预付货款导出.xlsx",
            "file_path": file_name
        })

    except Exception:
        logger.error(trace_error())
        return json_result(-1, trace_error(), None)


def query_yfhk_by_rid(rid_value):
    """按迁移规则将 number 语义映射为 rid 查询单条记录。"""
    rid_safe = esc_sql(rid_value)
    sql = f"""
        select rid,sb,cght,fkje,htje,sqje,fkbz,fkrq,jsrm,sccj,gcmc1,yhzh,khh,sfjq,bzsm,chrq,wxfp,fkdq
        from yfhk
        where rid='{rid_safe}'
        limit 1
    """
    rows = run_sql(sql)
    if not rows:
        return None
    return rows[0]


def esc_sql(v):
    # 第0步：对动态字符串做单引号转义，避免 SQL 拼接时语法异常
    return str(v or "").replace("'", "''")

def to_str(v, default=""):
    """安全字符串转换。"""
    if v is None:
        return str(default)
    try:
        return str(v)
    except Exception:
        return str(default)


#经手人名1
@any_route("/api/bill/advance_payment/operator_name_one", methods=["POST", "GET"])
@require_token
async def advance_payment_operator_name_one(request):
    """
    经手人名1
    """
    try:
        j = await request.json()
        user = request.current_user
        user_name = user.username

        form_fields = j.get("form_fields", {})
        if not isinstance(form_fields, dict):
            form_fields = {}

        operator_name = str(form_fields.get("operator_name", "") or "").strip()
        if operator_name == "":
            return json_result(1, "success", {"clear_operator": 0})

        is_yj_user = 0
        # sql_user = f"""select * from sys_user where username='{esc_sql(user_name)}' and (position like '%优景%')"""
        # rows_user = run_sql(sql_user) or []
        org = get_user_path(user.username)
        path = org.get("path", "")
        logger.error('-----position-------')
        logger.error(path)
        if '优景' in path:
        # if len(rows_user) > 0:
            is_yj_user = 1

        business_department = ""
        business_area = ""
        clear_operator = 0
        error_msg = ""

        if is_yj_user == 1:
            sql_bm = f"""select bm from ywrybiao where yhm='{esc_sql(operator_name)}'"""
            rows_bm = run_sql(sql_bm) or []
            if len(rows_bm) > 0:
                business_department = str(rows_bm[0].get("bm", "") or "")
            else:
                clear_operator = 1
                error_msg = "不好意思,无此人员,请重新输入,谢谢!"

        sql_ssdq = f"""select ssdq from ywrybiao where yhm='{esc_sql(operator_name)}'"""
        rows_ssdq = run_sql(sql_ssdq) or []
        if len(rows_ssdq) > 0:
            business_area = str(rows_ssdq[0].get("ssdq", "") or "")

        if business_area == "":
            sql_dq = f"""select dq from ywrylx where ryxm='{esc_sql(operator_name)}'"""
            rows_dq = run_sql(sql_dq) or []
            if len(rows_dq) > 0:
                business_area = str(rows_dq[0].get("dq", "") or "")

        data = {
            "business_department": business_department,
            "business_area": business_area,
            "clear_operator": clear_operator,
            "error_msg": error_msg
        }
        return json_result(1, "success", data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        pass


#付款编号
@any_route("/api/bill/advance_payment/payment_no", methods=["POST", "GET"])
@require_token
async def advance_payment_payment_no(request):
    """
    付款编号
    """
    try:
        j = await request.json()
        form_fields = j.get("form_fields", {})
        if not isinstance(form_fields, dict):
            form_fields = {}

        payment_no = str(form_fields.get("payment_no", "") or "").strip()
        if payment_no == "":
            return json_result(1, "success", {})

        sql_fksp = f"""select * from fksp where fkbh='{esc_sql(payment_no)}'"""
        rows_fksp = run_sql(sql_fksp) or []
        if len(rows_fksp) == 0:
            return json_result(1, "success", {})

        row = rows_fksp[0]
        data = {
            "outer_sales_contract": str(row.get("wxfp", "") or ""),
            "outer_invoice_no": str(row.get("fphm", "") or ""),
            "purchase_contract": str(row.get("hthm", "") or ""),
            "contract_amount": row.get("jehj", 0) or 0,
            "request_payment_date": str(row.get("sqrq1", "") or ""),
            "request_amount": row.get("seje", 0) or 0,
            "payment_currency": str(row.get("fkbz", "") or ""),
            "estimated_shipping_date": str(row.get("chrq", "") or ""),
            "operator_name": str(row.get("jbry", "") or ""),
            "manufacturer": str(row.get("csmc", "") or ""),
            "shipping_factory": str(row.get("csmc", "") or ""),
            "bank_account": str(row.get("yhzh", "") or ""),
            "bank_name": str(row.get("khh", "") or ""),
            "factory_code": str(row.get("gcbh", "") or ""),
            "remark": str(row.get("bz1", "") or ""),
            "payment_area": str(row.get("fkdq", "") or ""),
            "chinese_product_name": str(row.get("zwpm", "") or ""),
            "business_department": str(row.get("ywbm", "") or ""),
            "contact": str(row.get("rxfs", "") or ""),
            "payment_type": str(row.get("hklx", "") or ""),
            "payment_form": str(row.get("fkxs", "") or ""),
            "fksp_rid": str(row.get("rid", "") or "")
        }
        return json_result(1, "success", data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        pass


# 付款日期
@any_route("/api/bill/advance_payment/payment_date", methods=["POST"])
@require_token
async def advance_payment_payment_date(request):
    """同步付款编号对应的付款日期到关联表。"""
    try:
        j = await request.json()
        form_fields = j.get("form_fields", {}) if isinstance(j, dict) else {}
        payment_no = esc_sql(to_str(form_fields.get("付款编号", "")).strip())
        payment_date = esc_sql(to_str(form_fields.get("付款日期", "")).strip())

        if not payment_no or not payment_date:
            return json_result(1, "success", {"updated": 0})

        sql_fksp = f"""update fksp set fkrq='{payment_date}' where fkbh='{payment_no}'"""
        sql_fkspsheet3 = f"""update fkspsheet3 set fkrq='{payment_date}' where fkbh='{payment_no}'"""
        run_sql(sql_fksp)
        run_sql(sql_fkspsheet3)

        return json_result(1, "success", {"updated": 1})
    except Exception:
        logger.error(trace_error())
        return json_result(-1, trace_error(), None)
    finally:
        pass


    #经手人名
@any_route("/api/bill/advance_payment/operator_name", methods=["POST", "GET"])
@require_token
async def advance_payment_operator_name(request):
    """
    经手人名
    """
    try:
        j = await request.json()
        user = request.current_user
        user_name = user.username

        form_fields = j.get("form_fields", {})
        if not isinstance(form_fields, dict):
            form_fields = {}

        purchase_contract = str(form_fields.get("purchase_contract", "") or "").strip()
        operator_name = str(form_fields.get("operator_name", "") or "").strip()
        if purchase_contract == "":
            return json_result(1, "success", {"operator_name": operator_name})
        if operator_name != "":
            return json_result(1, "success", {"operator_name": operator_name})

        yjgs = ""
        # sql_user = f"""select * from sys_user where username='{esc_sql(user_name)}' and (position like '%优胜%')"""
        # rows_user = run_sql(sql_user) or []
        # if len(rows_user) > 0:
        org = get_user_path(user.username)
        path = org.get("path", "")
        if '优胜' in path:
            yjgs = "1"

        sql_cght = f"""select top 1 cgry,cgbm,gdry from cght where hthm='{esc_sql(purchase_contract)}'"""
        rows_cght = run_sql(sql_cght) or []
        if len(rows_cght) > 0:
            row = rows_cght[0]
            if yjgs == "1" and str(row.get("gdry", "") or "") != "":
                operator_name = str(row.get("gdry", "") or "")
            else:
                operator_name = str(row.get("cgry", "") or "")

        return json_result(1, "success", {"operator_name": operator_name})
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        pass


def esc_sql(v):
    """SQL 安全转义。"""
    return str(v or "").replace("'", "''")


def to_str(v, default=""):
    """安全字符串转换。"""
    if v is None:
        return str(default)
    try:
        return str(v)
    except Exception:
        return str(default)


def to_num(v, default=0):
    """统一数值转换。"""
    try:
        if v is None or v == "":
            return int(default)
        return int(float(v))
    except Exception:
        return int(default)


def add_days(date_text, days):
    """yyyy-mm-dd 字符串加天数并返回 yyyy-mm-dd。"""
    try:
        dt = datetime.strptime(str(date_text), "%Y-%m-%d")
        return (dt + timedelta(days=to_num(days, 0))).strftime("%Y-%m-%d")
    except Exception:
        return ""
    

#预付款通知
@any_route("/api/bill/advance_payment/prepayment_notice", methods=["POST", "GET"])
@require_token
async def advance_payment_prepayment_notice(request):
    """
    预付款通知
    """
    s = Session()
    try:
        j = await request.json()
        user = request.current_user
        user_name = user.username

        form_fields = j.get("form_fields", {})
        if not isinstance(form_fields, dict):
            form_fields = {}

        payment_amount_num = 0.0
        try:
            payment_amount_num = float(form_fields.get("payment_amount", 0) or 0)
        except:
            payment_amount_num = 0.0
        if payment_amount_num <= 0:
            return json_result(1, "success", {"inserted": 0})

        hthm = str(form_fields.get("purchase_contract", "") or "")
        gdry = str(form_fields.get("operator_name", "") or "")
        gcmc = str(form_fields.get("manufacturer", "") or "")
        chgc = str(form_fields.get("shipping_factory", "") or "")
        fkje = str(form_fields.get("payment_amount", "") or "")

        now_dt = datetime.now()
        sys_created = now_dt.strftime("%Y-%m-%d %H:%M:%S")
        sys_lastmodified = sys_created
        fsrq = now_dt.strftime("%Y-%m-%d")
        xxnr = f"合同号码:{hthm}工厂:{gcmc}/{chgc}的预付款共计{fkje}已安排支付,日期:{fsrq}"

        row = {
            "xxly": '预付款',
            "bjdh": '',
            "wxht": '',
            "cght": '',
            "yhdh": '',
            "xxnr": xxnr,
            "jsr": gdry,
            "sys_path": "",
            "spsq": gdry
        }
        res = module_xxck_new([row],user,s)
        if res.get('code')!=1:
            s.rollback()
            return json_result(res.get('code'), res.get('code'))
        s.commit()
        return json_result(1, "success", {"inserted": 1})
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        pass


#工厂定金管理
@any_route("/api/bill/advance_payment/factory_deposit_management", methods=["POST", "GET"])
@require_token
async def advance_payment_factory_deposit_management(request):
    """
    工厂定金管理：保存前同步采购合同与付款审批数据。
    """
    try:
        j = await request.json()
        form_fields = j.get("form_fields", {})
        if not isinstance(form_fields, dict):
            form_fields = {}

        purchase_contract_id = str(form_fields.get("purchase_contract_id", "") or "").strip()
        purchase_contract = str(form_fields.get("purchase_contract", "") or "").strip()
        payment_amount = str(form_fields.get("payment_amount", "") or "").strip()
        payment_date = str(form_fields.get("payment_date", "") or "").strip()
        payment_approval_id = str(form_fields.get("payment_approval_id", "") or "").strip()

        updated_cght = 0
        updated_fksp = 0

        if purchase_contract_id != "":
            sql_update_cght = f"""
                update cght
                set djzje='{esc_sql(payment_amount)}', djrq='{esc_sql(payment_date)}'
                where (rid='{esc_sql(purchase_contract_id)}') and (hthm='{esc_sql(purchase_contract)}')
            """
            run_sql(sql_update_cght)
            updated_cght = 1

        if payment_approval_id != "":
            sql_update_fksp = f"""
                update fksp
                set yfje='{esc_sql(payment_amount)}', djrq='{esc_sql(payment_date)}'
                where (rid='{esc_sql(payment_approval_id)}') and (hthm='{esc_sql(purchase_contract)}')
            """
            run_sql(sql_update_fksp)
            updated_fksp = 1

        data = {
            "updated_cght": updated_cght,
            "updated_fksp": updated_fksp
        }
        return json_result(1, "success", data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        pass


# 正式付款日期
@any_route("/api/bill/advance_payment/formal_payment_date", methods=["POST"])
@require_token
async def advance_payment_formal_payment_date(request):
    """根据提前付款和预计出货日期计算正式付款日期。"""
    try:
        j = await request.json()
        form_fields = j.get("form_fields", {}) if isinstance(j, dict) else {}
        early_payment = to_str(form_fields.get("提前付款", "")).strip()
        expected_ship_date = to_str(form_fields.get("预计出货日期", "")).strip()
        factory_name = to_str(form_fields.get("生产厂家", "")).strip()

        user = request.current_user
        user_name = user.username
        if not user_name:
            return json_result(-1, "无法获取当前用户", None)

        if early_payment != "是":
            return json_result(1, "success", {"formal_payment_date": ""})

        if expected_ship_date == "":
            return json_result(1, "success", {"formal_payment_date": "2100-12-31"})

        a = 0
        b = 0

        # sql_user = f"""select * from sys_user where username='{esc_sql(user_name)}' and (position like '%优胜%')"""
        # rows_user = run_sql(sql_user)
        org = get_user_path(user.username)
        path = org.get("path", "")
        if '优胜' in path:
        # if rows_user and len(rows_user) > 0:
            sql_rule = f"""select * from cwjxrq where (szgs like '%优胜%')"""
            rows_rule = run_sql(sql_rule)
            if rows_rule and len(rows_rule) > 0:
                row = rows_rule[0]
                if isinstance(row, dict):
                    a = to_num(row.get("kpjxts", 0), 0)
                    b = to_num(row.get("xjjxts", 0), 0)
                else:
                    a = to_num(getattr(row, "kpjxts", 0), 0)
                    b = to_num(getattr(row, "xjjxts", 0), 0)

        d = a if factory_name != "" else b
        formal_payment_date = add_days(expected_ship_date, d)
        if formal_payment_date == "":
            formal_payment_date = "2100-12-31"

        return json_result(1, "success", {"formal_payment_date": formal_payment_date})

    except Exception:
        logger.error(trace_error())
        return json_result(-1, trace_error(), None)
    finally:
        pass


# 预付单号
@any_route("/api/bill/advance_payment/prepayment_no", methods=["POST"])
@require_token
async def advance_payment_prepayment_no(request):
    """按预付货款规则生成预付单号。"""
    try:
        j = await request.json()
        form_fields = j.get("form_fields", {}) if isinstance(j, dict) else {}
        current_no = str(form_fields.get("预付单号", "") or "").strip()
        if current_no:
            return json_result(1, "success", {"prepayment_no": current_no})

        user = request.current_user
        user_name = ""
        if user:
            user_name = str(getattr(user, "name", "") or getattr(user, "username", "") or "").strip()
        if not user_name:
            return json_result(-1, "无法获取当前用户", None)

        t1 = datetime.now().strftime("%y%m%d")
        sql = f"""select rid from yfhk order by rid desc"""
        rows = run_sql(sql)

        if rows and len(rows) > 0:
            row = rows[0]
            if isinstance(row, dict):
                last_number = str(row.get("rid", "") or "")
            else:
                last_number = str(getattr(row, "rid", "") or "")
            prepayment_no = f"{user_name}{t1}{last_number}"
        else:
            prepayment_no = f"{user_name}{t1}00001"

        return json_result(1, "success", {"prepayment_no": prepayment_no})
    except Exception:
        logger.error(trace_error())
        return json_result(-1, trace_error(), None)
    finally:
        pass



# 付款金额
@any_route("/api/advance_payment/process_before_save_fkje", methods=["POST"])
@require_token
async def advance_payment_process_before_save(request):
    """
    预付货款保存前统一处理接口
    包含：
    1. 金额转中文大写
    2. 根据经手人查询部门
    """
    try:
        j = await request.json()
        s = Session()
        
        # 准备返回数据结构
        data = {
            "chinese_amount": "",  # 转换后的大写金额
            "department": "",      # 查询到的部门
            "warnings": []         # 非阻断性警告信息
        }

        # --- 逻辑1：金额转中文大写 ---
        payment_amount = j.get("payment_amount", "")
        if payment_amount:
            try:
                # 使用cn2an库进行金额转换
                data["chinese_amount"] = cn2an.an2cn(float(payment_amount), "rmb")
            except Exception as e:
                logger.error("金额转换失败：%s", str(e))
                data["warnings"].append(f"金额转换失败: {str(e)}")
                # 降级：返回原值加"元"
                data["chinese_amount"] = f"{payment_amount}元"

        # --- 逻辑2：根据经手人查询部门 ---
        # 仅当前端明确标记需要查询时才执行
        need_dept_query = j.get("need_dept_query", False)
        handler_name = j.get("handler_name", "")
        
        if need_dept_query and handler_name:
            try:
                # SQL: select bm from ywrybiao where yhm=:yhm
                row = s.query(ywrybiao.bm).filter(
                    ywrybiao.yhm == handler_name
                ).first()

                if row and row.bm:
                    data["department"] = row.bm
                else:
                    data["warnings"].append(f"未找到经手人[{handler_name}]的部门信息")
            except Exception as e:
                logger.error("部门查询失败：%s", trace_error())
                data["warnings"].append("部门查询发生系统错误")

        return json_result(1, "success", data)
            
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()



# 所属部门
@any_route("/api/bill/advance_payment/owner_department", methods=["POST"])
@require_token
async def advance_payment_owner_department(request):
    """新建预付货款时，读取当前用户所属部门并返回。"""
    try:
        user = request.current_user
        user_name = ""
        if user:
            user_name = str(getattr(user, "name", "") or getattr(user, "username", "") or "").strip()

        if not user_name:
            return json_result(-1, "无法获取当前用户", None)

        user_name_safe = esc_sql(user_name)
        org = get_user_path(user_name_safe)
        path = org.get("path", "") if org else ""
        owner_department = str(path or "")[:100]

        return json_result(1, "success", {"owner_department": owner_department})
    except Exception:
        logger.error(trace_error())
        return json_result(-1, trace_error(), None)


# 生成编码
@any_route("/api/bill/advance_payment/generate_code", methods=["POST"])
@require_token
async def advance_payment_generate_code(request):
    """新建预付货款时按规则生成识别编码。"""
    try:
        user = request.current_user
        user_name = ""
        if user:
            user_name = str(getattr(user, "name", "") or getattr(user, "username", "") or "").strip()
        if not user_name:
            return json_result(-1, "无法获取当前用户", None)

        t1 = datetime.now().strftime("%y%m%d")
        sql = f"""select rid from yfhk order by rid desc"""
        rows = run_sql(sql)

        if rows and len(rows) > 0:
            last_number = ""
            row = rows[0]
            if isinstance(row, dict):
                last_number = str(row.get("rid", "") or "")
            else:
                last_number = str(getattr(row, "rid", "") or "")
            identify_code = f"{user_name}{t1}{last_number}"
        else:
            identify_code = f"{user_name}{t1}00001"
        
        return json_result(1, "success", {"identify_code": identify_code})
    except Exception:
        logger.error(trace_error())
        return json_result(-1, trace_error(), None)
    finally:
        pass

@any_route('/api/saier/advance_payment/bank_apply_status', methods=['POST', 'GET'])
@require_token
async def view_saier_advance_payment_bank_apply_status(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        khyh = j.get('khyh','')
        wfyh = j.get('wfyh','')
        yhdm = ''
        tc1 = ''
        th1 = ''
        tc2 = ''
        th2 = ''
        c = s.query(newcwcs.yhdm).filter(newcwcs.bank == str(khyh)).first()
        if c is not None:
            yhdm = c.yhdm
        c = s.query(yhbm.province, yhbm.city, yhbm.ssyh).filter((yhbm.yhmc == str(khyh)) | ((yhbm.yhdm == str(yhdm)) & (yhbm.yhdm != '') & (yhbm.yhdm.isnot(None)))).first()
        if c:
            tc1 = c.province + '' + c.city
            th1 = c.ssyh
        c = s.query(yhbm.province, yhbm.city, yhbm.ssyh).filter(yhbm.yhmc == str(wfyh)).first()
        if c:
            tc2 = c.province + '' + c.city
            th2 = c.ssyh
        yhdm1 = ''
        c = s.query(yhbm.yhdm).filter(yhbm.yhmc == str(wfyh)).first()
        if c:
            yhdm1 = c.yhdm
        data = {"tc1":tc1,"th1":th1,"tc2":tc2,"th2":th2,"yhdm1":yhdm1}
        
        return json_result(1, '操作成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

@any_route('/api/saier/payment_approval/user_check', methods=['POST', 'GET'])
@require_token
async def view_saier_payment_approval_user_check(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        org = get_user_path(user.username)
        postion = org.get('position','')
        if not '财务' in postion:
            return json_result(-1, '不好意思，您没有权限操作！')
        
        return json_result(1, '校验成功')
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

def new_payment_merge_data(user, data, module, s):
    try:
        flag = 0
        lyrid = data.get('rid','')
        # d = s.query(fkhc).filter(fkhc.rid == lyrid).first()
        # if d:
        #     return {'code':0, 'msg':'数据已存在！'}
        d = fkhc()
        d.rid = get_uuid()
        d.uid = user.rid
        d.ctime = time.strftime("%Y-%m-%d %H:%M:%S")
        d.custId = data.get('custId','')
        d.serialNo = ''
        d.corpCode = data.get('corpCode','')
        d.zh = data.get('yhzh','')
        d.wfyhzh = data.get('wfyhzh','')
        d.bank = data.get('khh','')
        # d.yhdm = data.get('yhdm','') # 这个银行代码字段系统找不到对应数据，暂时不填
        d.gcmc = data.get('sccj','')
        d.yt = data.get('yt','货款')
        d.hkje = data.get('fkje',0)
        d.yyzf = data.get('yyzf','')
        d.thzz = data.get('thzz','')
        d.tcyd = data.get('tcyd','')
        d.yhstate = '无'
        d.wfyhmc = data.get('wfyhmc','')
        d.tjrq = time.strftime("%Y-%m-%d")
        d.numberhz = data.get('rid')
        d.jsrm = data.get('jsrm','')
        if data.get('fpwk'):
            d.fpwk = data.get('fpwk')
        d.fkhz = data.get('fkhz','采购合同:' + data.get('cght','') + '付款金额' + str(data.get('hkje',0)))
        d.ly = module
        d.wstt = data.get('wstt','')
        s.add(d)
        return {'code':1, 'msg':'操作成功'}
    except:
        logger.error(trace_error())
        return {'code':-1, 'msg':trace_error()}

@any_route('/api/bill/advance_payment/new_payment_merge', methods=['POST', 'GET'])
@require_token
async def view_saier_payment_approval_new_payment_merge(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        flag = 0
        m = j.get('main',{})
        fkdq = m.get('fkdq','')
        module = j.get('module','')
        u = s.query(sys_user_role).filter(sys_user_role.user_id == user.rid).all()
        role_ids = [r.role_id for r in u]
        for r in role_ids:
            c = s.query(sys_role).filter(sys_role.rid == r).first()
            if c and '义乌财务' in c.name:
                flag = 1
                break
        if flag == 1 and fkdq == '宁波':
            return json_result(-1, '义乌财务没有权限操作宁波付款单！')
        yhzt = m.get('Field','')
        if yhzt != '待提交':
            return json_result(0, '无需处理')
        m['fkhz'] = '采购合同:' + m.get('cght','') + '付款金额' + str(m.get('hkje',0))
        m['ly'] = module
        m['yt'] = '货款'
        res = new_payment_merge_data(user, m, module, s)
        if res.get('code') != 1:
            return json_result(-1, res.get('msg',''))
        
        s.commit()
        return json_result(1, '校验成功')
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

@any_route('/api/saier/advance_payment/batch_bank_apply_check', methods=['POST', 'GET'])
@require_token
async def view_saier_advance_payment_batch_bank_apply_check(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        rids = j.get('rids',[])
        module = j.get('module','')
        org = get_user_path(user.username)
        postion = org.get('position','')
        if not '财务' in postion:
            return json_result(-1, '不好意思，您没有权限操作！')
        fkdq = ''
        wstt1 = ''
        new_rids = []
        data = {}
        d = s.query(cyzglsheet).filter(cyzglsheet.zm=='特殊银行付款').all()
        banks = list(OrderedDict.fromkeys([r.xm for r in d]))
        t = get_module(module)
        o = get_model_by_table_name(t.table_name)
        for rid in rids:
            d = s.query(o).filter(o.rid==rid, or_(o.yhstate=='无', func.ifnull(o.yhstate,'')=='')).first()
            if module == '预付货款':
                if d and d.khh != '' and d.khh != None and d.yhzh != '' and d.yhzh != None and d.fkje > 1:
                    if wstt1 == '' and d.wstt != '' and d.wstt != None:
                        wstt1 = d.wstt if d.wstt else ''
                        fkdq = d.fkdq if d.fkdq else ''
                    new_rids.append(rid)
            else:
                if d and d.bank != '' and d.zh != None and d.yt != '' and d.bank != None and d.zh != None and d.yt != None and (d.hkje > 1):
                    if wstt1 == '' and d.wstt != '' and d.wstt != None:
                        wstt1 = d.wstt if d.wstt else ''
                        fkdq = d.fkdq if d.fkdq else ''
                    new_rids.append(rid)

        if len(new_rids)==0:
            return json_result(-1, '没有符合条件的预付款记录')
        d = s.query(fkyh.payAcc).filter(fkyh.wstt == wstt1, fkyh.fkdq == fkdq).all()
        yhzh = [r.payAcc for r in d]
        data = {'fkdq':fkdq,'wstt':wstt1,'banks':banks,'new_rids':new_rids, 'yhzh':yhzh}

        return json_result(1, '校验成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


@any_route('/api/saier/advance_payment/batch_bank_apply_update', methods=['POST', 'GET'])
@require_token
async def view_saier_advance_payment_batch_bank_apply_update(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        rids = j.get('new_rids',[])
        yhzh = j.get('yhzh', '')
        banks = j.get('banks', [])
        module = j.get('module','')
        qb1 = []
        new_data = {}
        xx = ''
        t = get_module(module)
        o = get_model_by_table_name(t.table_name)
        for rid in rids:
            zhsb = ''
            xgqd = ''
            tc1 = ''
            tc2 = ''
            th1 = ''
            th2 = ''
            thzz = ''
            tcyd = ''
            d = s.query(o).filter(o.rid == rid).first()
            if d:
                zhsb = ''
                if d.bzsm in banks:
                    zhsb = '1'
                if zhsb == '':
                    xgqd = xgqd
                c = s.query(fkyh).filter(fkyh.wstt == d.wstt, fkyh.fkdq == d.fkdq, fkyh.payAcc == yhzh).first()
                if c:
                    val = getattr(d, 'bank', '') if hasattr(d, 'bank') else getattr(d, 'khh', '')
                    yhdm = ''
                    a = s.query(newcwcs.yhdm).filter(newcwcs.bank == val).first()
                    if a:
                        yhdm = a.yhdm
                    if yhdm == '':
                        a = s.query(yhbm.yhdm).filter(yhbm.yhmc == val).first()
                        if a:
                            yhdm = a.yhdm
                    a = s.query(yhbm.province, yhbm.city, yhbm.ssyh).filter((yhbm.yhmc == val) | ((yhbm.yhdm == yhdm) & (yhbm.yhdm != '') & (yhbm.yhdm.isnot(None)))).first()
                    if a:
                        tc1 = a.province + '' + a.city
                        th1 = a.ssyh
                    a = s.query(yhbm.province, yhbm.city, yhbm.ssyh).filter(yhbm.yhmc == d.wfyhmc).first()
                    if a:
                        tc2 = a.province + '' + a.city
                        th2 = a.ssyh
                    if (tc1 != '') and (tc2 != ''):
                        if (tc1 == tc2):
                            tcyd = '同城'
                        else:
                            tcyd = '异地'
                    if (th1 != '') and (th2 != ''):
                        if (th1 == th2):
                            thzz = '同行'
                        else:
                            thzz = '他行'
                    else:
                        if (val == d.wfyhmc):
                            thzz = '同行'
                        else:
                            thzz = '他行'

                    logger.error(f"银行申请批次更新校验结果：rid={rid}, tc1={tc1}, th1={th1}, tc2={tc2}, th2={th2}, tcyd={tcyd}, thzz={thzz}")
                    if thzz != '' and tcyd != '':
                        if module == '预付货款':
                            hb_str = d.yhzh +'_' + d.sccj + '_' + d.jsrm + '_' + d.wstt + '_' + d.wstt
                        else:
                            hb_str = d.zh +'_' + d.gcmc + '_' + d.jsrm + '_' + d.wstt + '_' + d.wstt
                        if not hb_str in new_data:
                            new_data[hb_str] = alchemy_object_to_dict(d)
                            if d.yyzf == '' or d.yyzf == None:
                                new_data[hb_str]['yyzf'] = '非预约'
                            if module == '预付货款':
                                new_data[hb_str]['fkhz1'] = '采购合同:' + d.cght + '付款金额' + str(d.fkje)
                            else:
                                new_data[hb_str]['fkhz1'] = '外销发票:' + d.wxfp + '付款金额' + str(d.hkje)
                            new_data[hb_str]['numberhz'] = [d.rid]
                        else:
                            new_data[hb_str]['fkje'] += d.fkje
                            new_data[hb_str]['numberhz'].append(d.rid)
                            # new_data[hb_str]['fkhz1'].append('采购合同:' + d.cght + '付款金额' + str(d.fkje))

                        d.thzz = thzz
                        d.tcyd = tcyd
                        d.xgqd = xgqd
                        d.modi_uid = user.rid
                        d.modi_time = time.strftime("%Y-%m-%d %H:%M:%S")
                        d.subaAccountSerial = c.subaAccountSerial
                        d.corpCode = c.corpCode
                        d.custId = c.custId
                        d.wfyhzh = c.payAcc
                        d.wfyhmc = c.wfyhmc
                        d.thzz = thzz
                        d.tcyd = tcyd
                        s.add(d)
                    else:
                        if module == '预付货款':
                            xx = '采购合同:' + str(d.cght) + '生产厂家' + str(d.sccj) + '开户银行' + str(val) + '我司抬头' + str(d.wstt)
                        else:
                            xx = '发票号:' + str(d.wxfp) + '生产厂家' + str(d.gcmc) + '开户银行' + str(val) + '我司抬头' + str(d.wstt) + '发票为空:' + str(d.fpwk)
        
        for k,v in new_data.items():
            v['fkhz'] = '采购合同:' + v.get('cght','') + '付款金额' + str(v.get('hkje',0))
            v['ly'] = module
            v['yt'] = '货款'
            res = new_payment_merge_data(user, v, module, s)
            if res.get('code') != 1:
                s.rollback()
                return json_result(-1, res.get('msg',''))
            
        s.commit()
        return json_result(1, '校验成功', xx)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 批量付款批次信息导出
@any_route("/api/advance_payment/batch_export_excel", methods=["POST"])
@require_token
async def view_advance_payment_batch_export_excel(request):
    """按 oldCode 逻辑批量付款批次信息导出到 Excel。"""
    try:
        j = await request.json()
        rid_list = j.get("rids") or []
        if not rid_list:
            return json_result(-1, "未选择导出记录", None)
        wb = Workbook()
        ws = wb.active
        ws.title = "付款批次"
        field_tilte = ["rid", "识别码", "采购合同", "付款金额", "是否结清", "付款日期", "出运日期", "生产厂家", "出货工厂", "合同金额", "经手人名", "备注说明", "外销发票", "付款地区"]
        for col_idx, title in enumerate(field_tilte, start=1):
            ws.cell(row=1, column=col_idx, value=title)
        # Pascal 从第 3 行开始写数据
        start_row = 2
        field_order = [
            "rid", "sb", "cght", "fkje", "sfjq", "fkrq", "chrq", "sccj", "gcmc1", "htje", "jsrm", "bzsm", "wxfp", "fkdq"
        ]
        for idx, rid_value in enumerate(rid_list):
            row_no = start_row + idx
            row_data = query_yfhk_by_rid(rid_value)
            if not row_data:
                continue
            for col_idx, field_name in enumerate(field_order, start=1):
                ws.cell(row=row_no, column=col_idx, value=row_data.get(field_name, ""))
        
        file_name = f"付款批次导出_{get_uuid()}.xlsx"
        file_path = os.path.join(config.tmp_path, file_name)
        wb.save(file_path)

        return json_result(1, "success", {
            "file_name": f"付款批次导出{time.strftime('%Y-%m-%d_%H-%M-%S')}.xlsx",
            "file_path": file_name
        })
    except Exception:
        logger.error(trace_error())
        return json_result(-1, trace_error(), None)
    

"""
导入付款批次资料Excel
对应原Pascal: 付款批次回导功能
"""
@any_route('/api/advance_payment/batch_update_payment', methods=['POST'])
@require_token
async def view_advance_payment_batch_update_payment(request):
    # 1. 检查用户是否是财务岗位 (对应 tmpcom2 查询)
    s = Session()
    try:
        j = await request.form()
        module = form_value(j, 'module', '预付货款')
        logger.error(f"Batch update payment - module: {module}")
        logger.error(f"Form data keys: {list(j.keys())}")
        user = request.current_user
        org = get_user_path(user.username)
        position = org.get('position')
        if '财务' not in position:
            return json_result(-1, '只有财务岗位用户才能执行此操作')
        # 2. 获取上传的Excel文件
        file = form_value(j,'file',None)
        if is_none(file):
            return json_result(ERR_PARAM_NOT_ENOUGH)
        # 保存临时文件
        temp_dir = config.tmp_path
        temp_file = os.path.join(temp_dir, f"{uuid.uuid4()}_{file.filename}")
        # logger.error(f"Saving uploaded file to: {temp_file}")
        data = await file.read()
        write_file(temp_file, data, 'wb', encoding=None)
        
        # 3. 读取Excel文件
        wb = load_workbook(temp_file)
        ws = wb.worksheets[0]  # 读取第一个工作表
        # 4. 处理每一行数据
        # duplicates = []  # 存储重复记录信息
        row_idx = 2  # 从第2行开始（第1行是表头）
        
        if module == '预付货款':
            o = yfhk
        else:
            o = gchk
        while True:
            # 读取Excel单元格数据
            rid = str(ws.cell(row=row_idx, column=1).value or '').strip()  # A列 - rid
            sb = str(ws.cell(row=row_idx, column=2).value or '')  # B列 - 识别码
            fkpc1 = str(ws.cell(row=row_idx, column=17).value or '')  # Q列 -  付款批次
            # 如果rid为空，结束循环
            if not rid:
                break
            d = s.query(o).filter(o.rid == rid, o.sb == sb).first()
            if d:
                d.fkpc1 = fkpc1
                d.modi_uid = user.rid
                d.modi_time = time.strftime("%Y-%m-%d %H:%M:%S")
                s.add(d)
            row_idx += 1
            
        s.commit()
        return json_result(0, "批量更新成功")
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()
        wb.close()


# ==================== 辅助函数 ====================
def getnum(value):
    """提取数字"""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r'[\d.]+', str(value))
    return float(match.group()) if match else 0.0

def get_column_letter(col_num):
    """将列号转换为 Excel 列字母（1-based）"""
    result = ""
    while col_num > 0:
        col_num -= 1
        result = chr(col_num % 26 + 65) + result
        col_num //= 26
    return result
    
# ==================== 第二步：读取并更新数据 ====================
def process_excel_data(ws, mapping, user, s):
    """处理 Excel 数据并更新数据库"""
    # 获取各字段的列字母
    A = mapping.get('识别码')
    B = mapping.get('采购合同')
    C = mapping.get('付款金额')
    D = mapping.get('合同金额')
    E = mapping.get('申请金额')
    F = mapping.get('付款币种')
    G = mapping.get('付款日期')
    H = mapping.get('经手人名')
    I3 = mapping.get('生产厂家')
    J3 = mapping.get('出货工厂')
    K = mapping.get('银行帐号')
    L = mapping.get('开户银行')
    M = mapping.get('是否结清')
    N = mapping.get('备注说明')
    
    # 检查必需的列是否存在
    if not A:
        logger.error("错误：未找到'识别码'列")
        return {'code': -1, 'msg': "未找到'识别码'列"}
    
    def get_cell_value(col_letter, row_num):
        if col_letter:
            return ws[col_letter + str(row_num)].value
        return None
    
    i = 3
    # 读取第3行数据
    sb = get_cell_value(A, i)
    hthm = get_cell_value(B, i)
    fkje = get_cell_value(C, i)
    htje = get_cell_value(D, i)
    sqje = get_cell_value(E, i)
    fkbz = get_cell_value(F, i)
    fkrq = get_cell_value(G, i)
    jsrm = get_cell_value(H, i)
    sccj = get_cell_value(I3, i)
    gcmc1 = get_cell_value(J3, i)
    yhzh = get_cell_value(K, i)
    khh = get_cell_value(L, i)
    sfjq = get_cell_value(M, i)
    bzsm = get_cell_value(N, i)
    # 循环处理
    while sb:
        # 更新数据库
        d = s.query(yfhk).filter(yfhk.sb == str(sb)).first()
        if d:
            xgqd = ''
            d.cght = hthm
            d.fkje = getnum(fkje)
            d.htje = getnum(htje)
            d.sqje = getnum(sqje)
            d.fkbz = fkbz
            d.fkrq = fkrq
            d.jsrm = jsrm
            d.sccj = sccj
            d.gcmc1 = gcmc1
            d.yhzh = yhzh
            d.khh = khh
            d.sfjq = sfjq
            d.bzsm = bzsm
            if d.xgqd != '' and d.xgqd is not None:
                xgqd = f"{d.xgqd};{user.username}回导日期{datetime.now().strftime('%Y-%m-%d')}"
            else:
                xgqd = f"{user.username}回导日期{datetime.now().strftime('%Y-%m-%d')}"
            d.xgqd = xgqd
            s.add(d)

        i += 1
        sb = get_cell_value(A, i)
        hthm = get_cell_value(B, i)
        fkje = get_cell_value(C, i)
        htje = get_cell_value(D, i)
        sqje = get_cell_value(E, i)
        fkbz = get_cell_value(F, i)
        fkrq = get_cell_value(G, i)
        jsrm = get_cell_value(H, i)
        sccj = get_cell_value(I3, i)
        gcmc1 = get_cell_value(J3, i)
        yhzh = get_cell_value(K, i)
        khh = get_cell_value(L, i)
        sfjq = get_cell_value(M, i)
        bzsm = get_cell_value(N, i)
    
    return {'code': 1, 'msg': '批量更新成功'}

"""
对应原Pascal: 批量预付款回导
"""
@any_route('/api/advance_payment/batch_update_excel', methods=['POST'])
@require_token
async def view_advance_payment_batch_update_excel(request):
    # 1. 检查用户是否是财务岗位 (对应 tmpcom2 查询)
    s = Session()
    try:
        j = await request.form()
        user = request.current_user
        org = get_user_path(user.username)
        position = org.get('position')
        if '财务' not in position:
            return json_result(-1, '只有财务岗位用户才能执行此操作')
        # 2. 获取上传的Excel文件
        file = form_value(j,'file',None)
        if is_none(file):
            return json_result(ERR_PARAM_NOT_ENOUGH)
        # 保存临时文件
        temp_dir = config.tmp_path
        temp_file = os.path.join(temp_dir, f"{uuid.uuid4()}_{file.filename}")
        # logger.error(f"Saving uploaded file to: {temp_file}")
        data = await file.read()
        write_file(temp_file, data, 'wb', encoding=None)
        
        # 3. 读取Excel文件
        wb = load_workbook(temp_file)
        ws = wb.worksheets[0]  # 读取第一个工作表
        mapping = {
            '识别码': None,
            '采购合同': None,
            '付款金额': None,
            '合同金额': None,
            '申请金额': None,
            '付款币种': None,
            '付款日期': None,
            '经手人名': None,
            '生产厂家': None,
            '出货工厂': None,
            '银行帐号': None,
            '开户银行': None,
            '是否结清': None,
            '备注说明': None
        }
        # ==================== 第一步：获取列字母映射 ====================
        def get_column_mapping(ws, mapping):
            """扫描第2行，获取各字段对应的列字母"""
            for col in range(1, 48):
                cell_value = ws.cell(row=2, column=col).value
                column_letter = get_column_letter(col)
                
                if cell_value in mapping:
                    mapping[cell_value] = column_letter
            
            return mapping
        mapping = get_column_mapping(ws, mapping)
        res = process_excel_data(ws, mapping, user, s)
        if res.get('code') != 1:
            s.rollback()
            return json_result(-1, res.get('msg',''))
        
        s.commit()
        return json_result(0, "批量更新成功")
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()
        wb.close()


# ==================== 辅助函数 ====================
def set_cell_style(ws, cell_address: str, value: Any):
    """
    设置单元格的值和样式（居中、自动换行）
    """
    cell = ws[cell_address]
    cell.value = value
    cell.alignment = ALIGN_CENTER


def write_payment_row(ws, row_num: int, data: Dict, config: Dict):
    """
    写入一行付款数据到 Excel
    
    Args:
        worksheet: Excel 工作表对象
        row_num: 行号
        data: 付款数据 (wxfp, fkje, sccj)
        config: 配置参数 (rq, rqn, rqy, pzh, kmdm, bank, bankdm, b, gcsb)
    """
    # 基础数据
    fkje = data['fkje']
    
    # 第一行数据
    _write_row_data(ws, row_num, config, fkje, is_first_row=True)
    
    # 第二行数据（借方行）
    _write_row_data(ws, row_num + 1, config, fkje, is_first_row=False)


def _write_row_data(ws, row_num: int, config: Dict, fkje: float, is_first_row: bool):
    """
    写入单行数据
    
    Args:
        is_first_row: 是否为第一行（贷方行），否则为借方行
    """
    row = str(row_num)
    
    # 定义每列的数据映射
    if is_first_row:
        # 贷方行
        column_data = {
            'A': config['rq'], 'B': config['nf'], 'C': config['yf'], 'D': '转', 'E': config['pzh'],
            'F': config['kmdm'], 'G': '应付帐款', 'H': 'RMB', 'I': '人民币',
            'J': fkje, 'K': fkje, 'L': 0, 'M': config['username'],
            'N': 'NONE', 'O': 'NONE', 'P': 'NONE', 'R': '*',
            'T': f"{config['bank']}支付货款{config.get('wxfp', '')} {config.get('sccj', '')}",
            'U': 0, 'V': '*', 'W': 0, 'Y': config['rq'],
            'AA': 0, 'AB': 208330, 'AE': '公司汇率', 'AF': 1,
            'AG': config['i'] - 1,
            'AH': f"供应商---{config['kmdm']}---{config['sccj']}" if config['kmdm'] else '',
            'AI': 0, 'AL': config['gcsb']
        }
    else:
        # 借方行
        column_data = {
            'A': config['rq'], 'B': config['nf'], 'C': config['yf'], 'D': '转', 'E': config['pzh'],
            'F': config['bankdm'], 'G': config['BANK'], 'H': 'RMB', 'I': '人民币',
            'J': fkje, 'K': 0, 'L': fkje, 'M': config['username'],
            'N': 'NONE', 'O': 'NONE', 'P': 'NONE', 'R': '*',
            'T': f"{config['bank']}支付货款{config.get('wxfp', '')} {config.get('sccj', '')}",
            'U': 0, 'V': '*', 'W': 0, 'Y': config['rq'],
            'AA': 0, 'AB': 208330, 'AE': '公司汇率', 'AF': 1,
            'AG': config['i'] - 1, 'AH': '', 'AI': 0, 'AL': config['gcsb']
        }
    
    # 应用样式和值
    for col, value in column_data.items():
        set_cell_style(ws, f"{col}{row}", value)

def query_supplier_code(s, supplier_name: str) -> tuple:
    """
    查询供应商代码
    
    Returns:
        (gcsb, kmdm) - gcsb 默认为 '1'，kmdm 为供应商代码
    """
    if not supplier_name:
        return '', ''
    
    sql = text("SELECT dm FROM jdgysdm WHERE mc = :mc")
    result = s.execute(sql, {"mc": supplier_name})
    row = result.fetchone()
    
    if row:
        return '1', row[0]
    return '', ''


def query_payment_data(s, rid, module: str) -> Optional[Dict]:
    """
    查询付款数据
    
    Returns:
        包含 wxfp, fkje, sccj 的字典，如果没有数据返回 None
    """
    if module == '预付货款':
        sql = text("SELECT wxfp, fkje, sccj FROM yfhk WHERE rid = :rid")
    else:
        sql = text("SELECT wxfp, hkje fkje, gcmc sccj FROM gchk WHERE rid = :rid")
    result = s.execute(sql, {"rid": rid})
    row = result.fetchone()
    
    if row:
        return {
            'wxfp': row[0],
            'fkje': row[1],
            'sccj': row[2]
        }
    return None

# ==================== 主处理函数 ====================
def kindee_excel_export(
    s,
    ws,
    rids: list,
    rq: str,
    nf: str,
    yf: str,
    pzh: str,
    bank: str,
    bankdm: str,
    username: str,
    kmdm: str = '',
    module: str = ''
):
    """
    主处理函数：根据编号列表查询数据并写入 Excel
    
    Args:
        worksheet: Excel 工作表对象
        numbers: 编号列表（对应 Delphi 的 bgpmstr1.Strings）
        rq, nf, yf, pzh, bank, bankdm, username, kmdm: 配置参数
    """
    i = 1
    
    for rid in rids:
        if not rid:
            continue
            
        # 查询付款数据
        payment_data = query_payment_data(s, rid, module)
        if not payment_data:
            continue
        
        # 查询供应商代码
        gcsb, kmdm = query_supplier_code(s, payment_data['sccj'])
        
        # 准备配置参数
        config = {
            'rq': rq, 'nf': nf, 'yf': yf, 'pzh': pzh,
            'bank': bank, 'bankdm': bankdm, 'username': username, 'kmdm': kmdm,
            'kmdm': kmdm, 'gcsb': gcsb, 'i': i + 1,
            'wxfp': payment_data['wxfp'], 'sccj': payment_data['sccj'],
            'BANK': bank,  # 需要根据实际情况定义
            'module': module
        }
        
        # 写入数据行（每次写入两行）
        write_payment_row(ws, i + 1, payment_data, config)
        
        # 更新行索引
        i += 2
    
    # 自动调整所有写入的行
    for row_num in range(1, i + 2):
        ws.row_dimensions[row_num].auto_size = True


# 金蝶付款（单笔）
@any_route("/api/advance_payment/kingdee_export_excel", methods=["POST"])
@require_token
async def view_advance_payment_kingdee_export_excel(request):
    s = Session()
    try:
        j = await request.json()
        rids = j.get("rids") or []
        if not rids:
            return json_result(-1, "未选择导出记录", None)
        module = j.get('module','')
        user = request.current_user
        org = get_user_path(user.username)
        position = org.get('position')
        bank = j.get("bank", "")
        pzh = j.get("pzh", "")
        kmdm = j.get("kmdm", "")
        if '财务' not in position:
            return json_result(-1, '只有财务岗位用户才能执行此操作')
        if pzh == '' or pzh is None:
            return json_result(-1, "请输入凭证号", None)
        if bank == '' or bank is None:
            return json_result(-1, "请输入银行", None)
        if kmdm == '' or kmdm is None:
            return json_result(-1, "请输入科目代码", None)
        
        rq = time.strftime("%Y-%m-%d")
        nf = time.strftime("%Y")
        yf = int(time.strftime("%m"))

        template_path = os.path.join(config.data_upload_path, "template/金蝶付款.xlsx")
        if not os.path.exists(template_path):
            return json_result(-1, "金蝶付款导出模板文件不存在", None)
        wb = load_workbook(template_path)
        ws = wb.active
        d = s.query(jdyhdm.yhdm).filter(jdyhdm.yhmc == bank).first()
        bankdm = d.yhdm if d else ''

        res = kindee_excel_export(
            s,
            ws,
            rids,
            rq,
            nf,
            yf,
            pzh,
            bank,
            bankdm,
            user.username,
            kmdm,
            module=module
        )

        file_name = f"金蝶付款导出_{get_uuid()}.xlsx"
        file_path = os.path.join(config.tmp_path, file_name)
        wb.save(file_path)

        return json_result(1, "success", {
            "file_name": f"金蝶付款导出{time.strftime('%Y-%m-%d_%H-%M-%S')}.xlsx",
            "file_path": file_name
        })
    except Exception:
        logger.error(trace_error())
        return json_result(-1, trace_error(), None)
    finally:
        s.close()


# ==================== 辅助函数 ====================
def calc_difference(sqje: float, scje: float) -> float:
    """计算差额（申请金额 - 实付金额），负数返回0"""
    diff = sqje - scje
    return diff if diff > 0 else 0


# ==================== 样式常量 ====================
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center', wrap_text=True)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
MONTH_FILL = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
MONTH_FONT = Font(bold=True, size=12, color="0066CC")
SUMMARY_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
SUMMARY_FONT = Font(bold=True)

# 列映射（A-S列，注意跳过T列）
COLUMN_HEADERS = {
    'A': '采购合同', 'B': '合同金额', 'C': '生产厂家', 'D': '工厂名称',
    'E': '申请金额', 'F': '产品名称', 'G': '预计出库', 'I': '出货日期',
    'J': '实付金额', 'K': '经手人', 'L': '差额', 'M': '合同日期',
    'N': '付款金额', 'O': '付款日期', 'P': '结算方式', 'Q': '付款地区',
    'R': '扣点%', 'S': '扣点金额'
}

# 需要应用边框的列（排除 T 列）
BORDER_COLUMNS = list('ABCDEFGHIJKLMNOPQRS')


def record_to_row(record: dict) -> dict:
    """将记录字典转换为行数据字典"""
    return {
        'A': record.get('cght', ''),
        'B': record.get('htje', 0),
        'C': record.get('sccj', ''),
        'D': record.get('gcmc1', ''),
        'E': record.get('sqje', 0),
        'F': record.get('cpmc', ''),
        'G': record.get('yjch', ''),
        'I': record.get('chrq', ''),
        'J': record.get('scje', 0),
        'K': record.get('jsrm', ''),
        'L': calc_difference(record.get('sqje', 0), record.get('scje', 0)),
        'M': record.get('htrq', ''),
        'N': record.get('fkje', 0),
        'O': record.get('fkrq', ''),
        'P': record.get('jsfs', ''),
        'Q': record.get('fkdq', ''),
        'R': record.get('kd', ''),
        'S': record.get('kdje', '')
    }


def auto_fit_columns(ws, max_col: int = 19):
    """自动调整列宽（只调整A-S列）"""
    columns_to_fit = list('ABCDEFGHIJKLMNOPQRS')
    for col_letter in columns_to_fit:
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell_value = ws[f"{col_letter}{row}"].value
            if cell_value:
                length = len(str(cell_value))
                max_length = max(max_length, length)
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width


def apply_borders(ws, start_row: int, end_row: int):
    """应用边框样式（排除 T 列）"""
    for row in range(start_row, end_row + 1):
        for col_letter in BORDER_COLUMNS:
            cell = ws[f"{col_letter}{row}"]
            cell.border = THIN_BORDER
            if isinstance(cell.value, (int, float)):
                cell.alignment = ALIGN_RIGHT
            else:
                cell.alignment = ALIGN_LEFT


def extract_month_from_date(date_str: str) -> str:
    """
    从日期字符串中提取月份（YYYY-MM格式）
    """
    if not date_str:
        return ''
    
    date_str = str(date_str).strip()
    
    # 尝试不同的日期格式
    for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y%m%d', '%Y-%m', '%Y/%m']:
        try:
            if len(date_str) <= 7:
                dt = datetime.strptime(date_str, fmt)
            else:
                dt = datetime.strptime(date_str[:10], fmt)
            return dt.strftime('%Y-%m')
        except ValueError:
            continue
    
    return ''


def group_records_by_month(records: List[dict]) -> dict:
    """
    按月度分组记录
    返回格式: {'2023-01': [records], '2023-02': [records]}
    """
    groups = defaultdict(list)
    
    for record in records:
        # 优先使用付款日期，如果没有则使用合同日期
        date_str = record.get('fkrq') or record.get('htrq') or ''
        month_key = extract_month_from_date(date_str)
        
        if month_key:
            groups[month_key].append(record)
        else:
            groups['未分类'].append(record)
    
    # 按月份排序
    sorted_groups = {}
    for key in sorted(groups.keys()):
        if key != '未分类':
            sorted_groups[key] = groups[key]
    if '未分类' in groups:
        sorted_groups['未分类'] = groups['未分类']
    
    return sorted_groups


def calculate_month_summary(records: List[dict]) -> dict:
    """
    计算月度汇总数据
    """
    summary = {
        'total_sqje': 0,      # 申请金额合计
        'total_fkje': 0,      # 付款金额合计
        'total_kdje': 0,      # 扣点金额合计
        'record_count': 0     # 记录数量
    }
    
    for record in records:
        sqje = record.get('sqje', 0) or 0
        fkje = record.get('fkje', 0) or 0
        kdje = float(record.get('kdje', 0)) if record.get('kdje') else 0
        
        summary['total_sqje'] += sqje
        summary['total_fkje'] += fkje
        summary['total_kdje'] += kdje
        summary['record_count'] += 1
    
    return summary


def write_month_separator(ws, current_row: int, month_key: str) -> int:
    """
    写入月份分隔行
    """
    # 合并A到S列
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=19)
    cell = ws.cell(row=current_row, column=1, value=f"【{month_key}】")
    cell.font = MONTH_FONT
    cell.fill = MONTH_FILL
    cell.alignment = ALIGN_LEFT
    cell.border = THIN_BORDER
    
    return current_row + 1


def write_month_summary(ws, current_row: int, month_key: str, summary: dict) -> int:
    """
    写入月度汇总行（只有合计，没有付款方式明细）
    """
    # 合计行背景色
    for col_letter in BORDER_COLUMNS:
        cell = ws[f"{col_letter}{current_row}"]
        cell.fill = SUMMARY_FILL
        cell.font = SUMMARY_FONT
        cell.border = THIN_BORDER
    
    # 合计信息（只显示合计，不显示付款方式明细）
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    ws[f"A{current_row}"] = f"{month_key}合计"
    ws[f"C{current_row}"] = f"记录数: {summary['record_count']}"
    ws[f"E{current_row}"] = f"申请金额合计: {summary['total_sqje']:.2f}"
    ws[f"N{current_row}"] = f"付款金额合计: {summary['total_fkje']:.2f}"
    ws[f"S{current_row}"] = f"扣点金额合计: {summary['total_kdje']:.2f}"
    
    return current_row + 1


# ==================== 核心生成函数 ====================
def create_report_workbook(
    records: List[dict],
    template_path: str = None,
    title: str = "预付货款及带款提货"
) -> openpyxl.Workbook:
    """
    创建报表工作簿（按月度汇总）
    """
    # 创建工作簿
    if template_path and os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
    else:
        wb = openpyxl.Workbook()
    
    # 处理第一个工作表
    ws1 = wb.worksheets[0]
    _write_sheet_data(ws1, records, title)
    
    # 处理第二个工作表
    if len(wb.worksheets) > 1:
        ws2 = wb.worksheets[1]
    else:
        ws2 = wb.create_sheet("Sheet2")
    
    _write_sheet_data(ws2, records, f"{title} - 副本")
    
    return wb


def _write_sheet_data(ws, records: List[dict], title: str):
    """写入工作表数据（按月度分组）"""
    current_row = 1
    
    # 写入标题
    # ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=19)
    # cell = ws.cell(row=current_row, column=1, value=title)
    # cell.font = Font(bold=True, size=14)
    # cell.alignment = ALIGN_CENTER
    current_row += 1
    
    # 写入表头
    # for col_letter, header in COLUMN_HEADERS.items():
    #     cell = ws[f"{col_letter}{current_row}"]
    #     cell.value = header
    #     cell.font = HEADER_FONT
    #     cell.fill = HEADER_FILL
    #     cell.alignment = ALIGN_CENTER
    #     cell.border = THIN_BORDER
    current_row += 1
    
    # 按月度分组
    grouped_records = group_records_by_month(records)
    
    # 遍历每个月度
    for month_key, month_records in grouped_records.items():
        # 写入月份分隔行
        # current_row = write_month_separator(ws, current_row, month_key)
        
        # 写入该月的数据行
        start_data_row = current_row
        for record in month_records:
            row_data = record_to_row(record)
            for col_letter, value in row_data.items():
                cell = ws[f"{col_letter}{current_row}"]
                if col_letter in ['I','O']:
                    cell.value = str(value)[:10]
                else:
                    cell.value = value
                cell.border = THIN_BORDER
                if isinstance(value, (int, float)):
                    cell.alignment = ALIGN_RIGHT
                else:
                    cell.alignment = ALIGN_LEFT
            current_row += 1
        
        # 应用边框（排除 T 列）
        if month_records:
            apply_borders(ws, start_data_row, current_row - 1)
        
        # 计算并写入月度汇总
        month_summary = calculate_month_summary(month_records)
        current_row = write_month_summary(ws, current_row, month_key, month_summary)
        
        # 月度之间空一行
        # current_row += 1
    
    # 自动调整列宽
    auto_fit_columns(ws)


async def prepay_generate_report(
    records: List[dict],
    output_path: str = None,
    title: str = "预付货款及带款提货",
    template_path: str = None
) -> Tuple[bool, str, Optional[openpyxl.Workbook]]:
    """
    生成报表并保存（按月度汇总）
    """
    try:
        # 创建工作簿
        wb = create_report_workbook(records, template_path, title)
        
        # 保存文件
        if not output_path:
            output_path = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb.save(output_path)
        
        return True, f"报表生成成功: {output_path}", wb
        
    except Exception as e:
        return False, f"生成失败: {str(e)}", None


# ==================== 路由接口 ====================
@any_route("/api/advance_payment/prepay_export_excel", methods=["POST"])
@require_token
async def view_advance_payment_prepay_export_excel(request):
    s = Session()
    try:
        j = await request.json()
        user = request.current_user
        org = get_user_path(user.username)
        position = org.get('position')
        ksrq = j.get("ksrq", "")
        jsrq = j.get("jsrq", "")

        if ksrq == '' or ksrq is None:
            return json_result(-1, "请输入开始日期", None)
        if jsrq == '' or jsrq is None:
            return json_result(-1, "请输入结束日期", None)
        if ksrq > jsrq:
            return json_result(-1, "开始日期不能大于结束日期", None)
        
        template_path = os.path.join(config.data_upload_path, "template/cwb1.xlsx")
        if not os.path.exists(template_path):
            return json_result(-1, "预付款导出模板文件不存在", None)
        
        # 查询数据
        d = s.query(yfhk).filter(yfhk.fkrq >= ksrq).filter(yfhk.fkrq <= jsrq).all()
        records = alchemy_object_list_to_dict(d)
        
        # 生成文件名和路径
        file_name = f"预付款导出_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = os.path.join(config.tmp_path, file_name)
        
        # 生成报表（按月度汇总）
        success, msg, wb = await prepay_generate_report(
            records=records,
            output_path=file_path,
            title=f"预付货款及带款提货 ({ksrq} 至 {jsrq})",
            template_path=template_path
        )
        
        if not success:
            return json_result(-1, msg, None)
        
        return json_result(1, "success", {
            "file_name": file_name,
            "file_path": file_name
        })
        
    except Exception:
        logger.error(trace_error())
        return json_result(-1, trace_error(), None)
    finally:
        s.close()


# 批量结清
@any_route("/api/bill/advance_payment/batch_settlement", methods=["POST"])
@require_token
async def view_bill_advance_payment_batch_settlement(request):
    s = Session()
    try:
        j = await request.json()
        rids = j.get("rids") or []
        sccj = j.get("sccj", "")
        if not rids:
            return json_result(-1, "未选择导出记录", None)
        module = j.get('module','')
        user = request.current_user
        org = get_user_path(user.username)
        position = org.get('position')
        if '财务' not in position:
            return json_result(-1, '只有财务岗位用户才能执行此操作')
        if sccj == '' or sccj is None:
            return json_result(-1, "请输入生产厂家", None)
        
        for rid in rids:
            d = s.query(yfhk).filter_by(rid=rid).first()
            if not d:
                continue
            if sccj not in d.sccj:
                continue
            d.sfjq = '是'
            d.modi_uid = user.rid
            d.modi_time = time.strftime("%Y-%m-%d %H:%M:%S")
            s.add(d)

        s.commit()
        return json_result(1, "批量结清成功")
    except Exception:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error(), None)
    finally:
        s.close()
        