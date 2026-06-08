# youjing/warehouse_report.py
from any import *
from .model import *
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import time


def get_ckfypath(name):
    if not name:
            return ''
    path = ''
    org = get_user_path(name)
    if org is None or not isinstance(org, dict):
        return '没有'
    path = org.get('path', '')
    return path[:100] if path else '没有'




@any_route('/api/RavenCloud/kcdc/report/export', methods=['POST'])
@require_token
async def export_warehouse_report(request):
    """导出出库单报表"""
    try:
        request_data = await request.json()
        rids = request_data.get("rids", [])
        
        if not rids:
            return json_result(code=-1, msg="请选择要导出的记录")
        
        # # 获取当前用户path
        # user = request.current_user
        # if (user.username):
       
        #     sys_path = get_ckfypath(user.username)
       
        
        # 按rids查询inventorydetail数据
        rid_str = ','.join([f"'{r}'" for r in rids])
        detail_sql = f"""
            SELECT * FROM inventorydetail 
            WHERE  rid IN ({rid_str})
        """
  
        details = run_sql(detail_sql)
        
        if not details:
            return json_result(code=-1, msg='未找到数据')
        
        # 创建Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "出库单"
        
        # 定义样式
        header_font = Font(bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 表头
        headers = [
            '仓库长度', '仓库宽度', '仓库高度', '采购合同', '产品编号',
            '仓库毛重', '仓库净重', '在仓箱数1', '托数', '剩余托数',
            '出库箱数', '备注信息', '仓库体积', '在仓总体积', '在仓总毛重',
            '理货员', '批号', '仓位', '转移仓位', '入库日期', '出库日期',
            '入库时间', '进仓单号', '厂商名称', '采购员', '跟单人员',
            '业务员', '业务部门', '1', '2', '唯一字段'
        ]
        
        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(1, col_idx, header)
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # 写入数据
        row_idx = 2
        for row_data in details:
            # 查询业务部门
            print('dd------')
            print(row_data.get('Salesman', ''))
           
            bm_sql = "SELECT bm FROM ywrybiao WHERE yhm =:yhm"
            bm_result = run_sql(bm_sql, {'yhm':row_data.get('Salesman', '')})
            bm = bm_result[0]['bm'] if bm_result else ''
            
            data_row = [
                row_data.get('OuterLength', '0'),
                row_data.get('OuterWidth', '0'),
                row_data.get('OuterHeight', '0'),
                row_data.get('PurchaseOrderNo', ''),
                str(row_data.get('ItemNo', '')).strip(),
                row_data.get('OuterGrossWeight', '0'),
                row_data.get('OuterNetWeight', '0'),
                row_data.get('CartonQty', '0'),
                row_data.get('PalletQty', '0'),
                '0',  # 剩余托数
                '0',  # 出库箱数
                str(row_data.get('Memo', '')).strip(),
                row_data.get('OuterVolume', '0'),
                row_data.get('Volume', '0'),
                row_data.get('GrossWeight', '0'),
                row_data.get('Collator', '0'),
                row_data.get('LotNumber', '0'),
                row_data.get('WarehousePosition', ''),
                '',  # 转移仓位
                row_data.get('StorageDate', ''),
                '',  # 出库日期
                row_data.get('StorageTime', ''),
                row_data.get('SNID', ''),
                row_data.get('SupplierShortName', ''),
                row_data.get('PurchasingAgent', ''),
                row_data.get('gdry', ''),
                row_data.get('Salesman', ''),
                bm,
                '',
                '',
                row_data.get('wyzd', '')
            ]
            
            for col_idx, value in enumerate(data_row, 1):
                cell = ws.cell(row_idx, col_idx, value)
                cell.border = border
            
            row_idx += 1
        
        # 添加合计行
        total_row = row_idx
        ws.cell(total_row, 1, '合计')
        ws.cell(total_row, 1).font = header_font
        
        # 在仓箱数合计 (H列)
        ws.cell(total_row, 8).value = f"=SUM(H2:H{row_idx-1})"
        # 在仓总体积合计 (N列)
        ws.cell(total_row, 14).value = f"=SUM(N2:N{row_idx-1})"
        # 在仓总毛重合计 (O列)
        ws.cell(total_row, 15).value = f"=SUM(O2:O{row_idx-1})"
        
        # 合计行加边框
        for col_idx in range(1, len(headers) + 1):
            ws.cell(total_row, col_idx).border = border
        
        # 设置列宽
        col_widths = {
            'A': 10, 'B': 10, 'C': 10, 'D': 25, 'E': 18,
            'F': 10, 'G': 10, 'H': 10, 'I': 10, 'J': 10,
            'K': 10, 'L': 50, 'M': 10, 'N': 10, 'O': 10,
            'P': 10, 'Q': 10, 'R': 10, 'S': 10, 'T': 10,
            'U': 10, 'V': 10, 'W': 16, 'X': 30, 'Y': 10,
            'Z': 10, 'AA': 10, 'AB': 23
        }
        
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width
       
        # 保存文件
        today = time.strftime('%Y-%m-%d')
        filename = f"{today}出库单.xlsx"
        filepath = config.get_today_upload_path()
        make_dirs(filepath)
        
        
        path = config.tmp_path
        print('执行到1')
        print('path',path)
        report_rid = get_uuid()
        wb.save(path + '/'+ str(report_rid)+'.xlsx')
        print('执行到2')
        print('filename',filename)
        # wb.save(path + '/'+ str(report_rid)+'.xlsx')
        
        # full_path = os.path.join(filepath, filename)
        # wb.save(full_path)
        
        # # 返回下载信息（和你的dgmb接口格式一致）
        # return json_result(code=1, msg="导出成功", data={
        #     'path': path,
        #     'name': filename
        # })
        return json_result(1,'生成报表成功',{'path': str(report_rid)+'.xlsx', 'name':'库存明细'+str(time.strftime("%Y%m%d%H%M%S"))+'.xlsx'})
        
    except Exception as e:
        logger.error(trace_error())
        return json_result(code=-1, msg=f"导出失败: {trace_error()}")