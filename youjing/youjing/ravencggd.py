# 只保留一个API：/api/Ravencloud/special_mark_generate
# 逻辑对应 Pascal：读取 sys_attachment 模板 + cyzglsheet 映射 + excel/database 双模式 + 输出PDF

from any import *
import os
import re
import json
import subprocess
from io import BytesIO
from openpyxl.reader.excel import load_workbook

ILLEGAL_CHARS = set('/\\:*?"<>|\r\n')


def _safe_filename(text):
    s = str(text or '').strip()
    if not s:
        return 'EMPTY'
    s = ''.join('_' if ch in ILLEGAL_CHARS else ch for ch in s)
    return s[:120]


def _cpbh_head_pascal(cpbh):
    """
    Pascal里文件名逻辑：遇到非法字符取前半段
    """
    s = str(cpbh or '')
    if not s:
        return ''
    for i, ch in enumerate(s):
        if ch in ILLEGAL_CHARS:
            return s[:i]
    return s


def _to_upload_rel_path(abs_path):
    """
    转为 /api/file/get 可下载的相对路径
    """
    abs_path = os.path.abspath(abs_path)
    base_upload = getattr(config, 'data_upload_path', '') or ''
    if base_upload:
        base_abs = os.path.abspath(base_upload)
        if abs_path.startswith(base_abs):
            rel = abs_path[len(base_abs):].lstrip('/\\')
            return rel.replace('\\', '/')
    return abs_path.replace('\\', '/')


def _convert_excel_to_pdf(excel_path):
    """
    使用 whale_report.jar 转 PDF
    """
    pdf_path = os.path.splitext(excel_path)[0] + '.pdf'

    if not hasattr(config, 'java_path') or not config.java_path:
        return {'success': False, 'error': '未配置 config.java_path'}
    if not hasattr(config, 'report_jar') or not config.report_jar:
        return {'success': False, 'error': '未配置 config.report_jar'}
    if not os.path.exists(config.report_jar):
        return {'success': False, 'error': f'JAR文件不存在: {config.report_jar}'}

    cmd = [
        config.java_path,
        '-jar',
        config.report_jar,
        'a',
        'b',
        excel_path,
        pdf_path,
        '2'
    ]

    try:
        rs = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
        if rs.returncode != 0:
            return {'success': False, 'error': rs.stderr or f'returncode={rs.returncode}'}
        if not os.path.exists(pdf_path):
            return {'success': False, 'error': 'PDF未生成'}
        return {'success': True, 'pdf_path': pdf_path}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def _get_mapping_rows(template_number):
    """
    兼容 pid/father 两种字段
    """
    rows = run_sql("""
        SELECT bz, bz1 FROM cyzglsheet
        WHERE pid=:father
    """, {'father': template_number})

    if len(rows) == 0:
        rows = run_sql("""
            SELECT bz, bz1 FROM cyzglsheet
            WHERE father=:father
        """, {'father': template_number})

    return rows


def _collect_order_numbers(selected_numbers):
    """
    对应 Pascal MultiReport：从 cggd 收集唯一 number
    """
    nums = []
    for x in selected_numbers:
        rows = run_sql("""
            SELECT rid FROM cggd
            WHERE rid=:v 
        """, {'v': str(x)})
        for r in rows:
            n = str(r.get('number') or r.get('rid') or '')
            if n and n not in nums:
                nums.append(n)
    return nums


def _new_template_wb_ws(template_bytes):
    """
    从模板字节创建 Workbook
    注意：openpyxl 只支持 xlsx；若是旧版 xls 会报错
    """
    try:
        wb = load_workbook(BytesIO(template_bytes))
        ws = wb.active
        return wb, ws
    except Exception as e:
        raise Exception('模板不是可读的xlsx格式，请将特殊唛头模板另存为xlsx后上传') from e


@any_route("/api/Ravencloud/special_mark_generate", methods=['POST'])
async def api_special_mark_generate(request):
    s = Session()
    try:
        form = await request.form()

        khmc = (form.get('khmc') or '').strip()
        req_mode = (form.get('mode') or '').strip().lower()  # excel/database/''
        filename_prefix = _safe_filename(form.get('filename_prefix') or '特殊唛头_')
        number_list_str = form.get('number_list', '[]')

        if not khmc:
            return json_result(code=-1, msg='客人全称不能为空')

        # 1) 查配置（按 qxzl+zm 定位）
        zm = f'{khmc}特殊唛头'
        cfg = run_sql("""
            SELECT rid number, zm, qxzl
            FROM cyzgl
            WHERE qxzl=:qxzl AND zm=:zm
            LIMIT 1
        """, {'qxzl': '特殊报表打印', 'zm': zm})

        if len(cfg) == 0:
            return json_result(code=-1, msg=f'未找到报表模板: {zm}')

        cfg_row = cfg[0]
        template_number = cfg_row.get('number', '')
        qxzl = (cfg_row.get('qxzl') or '').strip().lower()

        report_path = ''
        if template_number:
            att_rows = run_sql("""
                SELECT path
                FROM sys_attachment
                WHERE pid=:rid AND path IS NOT NULL
                LIMIT 1
            """, {'rid': template_number})
            if len(att_rows) > 0:
                report_path = str(att_rows[0].get('path') or '').strip()

        if not report_path:
            return json_result(code=-1, msg=f'未找到报表模板: {zm}')

        base_upload = getattr(config, 'data_upload_path', '') or ''
        r_path = os.path.join(base_upload, report_path.lstrip('/\\'))
        if not os.path.exists(r_path):
            return json_result(code=-1, msg='未找到报表模板')

        ext = os.path.splitext(r_path)[1].lower()
        if ext in ('.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp'):
            return json_result(code=-1, msg='模板附件是图片格式，请上传xlsx模板文件')

        with open(r_path, 'rb') as f:
            template_bytes = f.read()
        if not template_bytes:
            return json_result(code=-1, msg='报表模板为空')

        # 2) 模式判定：前端传了就用前端，没传就按配置 qxzl
        if req_mode in ('excel', 'database'):
            mode = req_mode
        else:
            mode = 'excel' if qxzl == 'excel' else 'database'

        # 3) 映射规则（对应 Pascal: cyzglsheet bz/bz1）
        mappings = _get_mapping_rows(template_number)
        if len(mappings) == 0:
            return json_result(code=-1, msg='未找到字段映射规则(cyzglsheet)')

        # 4) 输出目录
        output_dir = config.get_today_upload_path()
        if not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)

        pdf_files = []
        index = 0

        # =========================
        # Excel 模式（对应 Pascal qxzl='excel'）
        # =========================
        if mode == 'excel':
            excel_file = form.get('excel_file', None) or form.get('excel_file.raw', None)
            if not excel_file:
                return json_result(code=-1, msg='excel模式请上传excel_file')

            excel_bytes = await excel_file.read()
            src_wb = load_workbook(BytesIO(excel_bytes), data_only=True)
            src_ws = src_wb.active

            row_idx = 2  # Pascal 从 A2 开始
            while True:
                sb = src_ws[f'A{row_idx}'].value
                if sb is None or str(sb).strip() == '':
                    break

                index += 1
                wb, ws = _new_template_wb_ws(template_bytes)

                for m in mappings:
                    src_rule = str(m.get('bz') or '').strip()    # 如 A
                    dst_cell = str(m.get('bz1') or '').strip()   # 如 C4
                    if not src_rule or not dst_cell:
                        continue

                    try:
                        # Pascal: source = bz + row
                        if re.match(r'^[A-Za-z]+$', src_rule):
                            v = src_ws[f'{src_rule}{row_idx}'].value
                        elif re.match(r'^[A-Za-z]+\\d+$', src_rule):
                            v = src_ws[src_rule].value
                        else:
                            v = ''
                        ws[dst_cell] = v
                    except Exception:
                        pass

                out_name = f'{filename_prefix}{index}_{_safe_filename(sb)}'
                temp_xlsx = os.path.join(output_dir, out_name + '.xlsx')
                wb.save(temp_xlsx)
                wb.close()

                pdf_res = _convert_excel_to_pdf(temp_xlsx)
                try:
                    os.remove(temp_xlsx)
                except Exception:
                    pass

                if pdf_res.get('success'):
                    pdf_files.append(_to_upload_rel_path(pdf_res['pdf_path']))

                row_idx += 1

            src_wb.close()

        # =========================
        # 数据库模式（对应 Pascal else 分支）
        # =========================
        else:
            try:
                selected_numbers = json.loads(number_list_str) if isinstance(number_list_str, str) else number_list_str
            except Exception:
                selected_numbers = []

            if not selected_numbers:
                return json_result(code=-1, msg='database模式请传number_list')

            order_numbers = _collect_order_numbers(selected_numbers)

            for order_no in order_numbers:
                # 兼容 pid/father 两种字段
                details = run_sql("""
                    SELECT * FROM cggdsheet WHERE pid=:father
                """, {'father': str(order_no)})

         

                for detail in details:
                    index += 1
                    wb, ws = _new_template_wb_ws(template_bytes)

                    for m in mappings:
                        src_field = str(m.get('bz') or '').strip()   # Pascal: 字段名
                        dst_cell = str(m.get('bz1') or '').strip()
                        if not src_field or not dst_cell:
                            continue
                        ws[dst_cell] = detail.get(src_field, '')

                    cpbh_raw = detail.get('cpbh', '')
                    cpbh_head = _cpbh_head_pascal(cpbh_raw) or str(cpbh_raw or 'NOCPBH')
                    out_name = f'{filename_prefix}{index}_{_safe_filename(cpbh_head)}'

                    temp_xlsx = os.path.join(output_dir, out_name + '.xlsx')
                    wb.save(temp_xlsx)
                    wb.close()

                    pdf_res = _convert_excel_to_pdf(temp_xlsx)
                    try:
                        os.remove(temp_xlsx)
                    except Exception:
                        pass

                    if pdf_res.get('success'):
                        pdf_files.append(_to_upload_rel_path(pdf_res['pdf_path']))

        return json_result(code=1, msg=f'成功生成{len(pdf_files)}个唛头标签', data={
            'mode': mode,
            'files': pdf_files,
            'count': len(pdf_files)
        })

    except Exception:
        logger.error(trace_error())
        return json_result(code=-1, msg=trace_error())
    finally:
        s.close()
        
        
      