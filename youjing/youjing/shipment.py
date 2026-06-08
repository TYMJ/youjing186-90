from webbrowser import get

from any import *
from .model import *
from sqlalchemy.sql import func, not_, or_, and_
import time
from datetime import datetime
import json
import os
import string
from .__default__ import get_user_path, module_xxck_new, module_share_new, user_task_new, user_task_delete
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, Alignment, NamedStyle, PatternFill, Color
from openpyxl.drawing.image import Image as Image_Get
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.xml.constants import SHEET_MAIN_NS
import zipfile
import re


SYS_FIELDS = ['sid', 'rid', 'uid', 'ctime', 'mtime', 'has_att',
    'modi_uid', 'wf_status', 'wf_unit', 'pid', 'archived']
suffixes = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']

# 出运明细的客户名称联动
@any_route('/api/saier/shipment/khmc/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_khmc_change(request):
    s = Session()
    j = await request.json()
    try:
        khmc = j.get('khmc', '')
        d = run_sql(
            f"select ydqx,RMBkh,zfqx,fkys,ydqx,LCydqx from kh where company_name='{khmc}' limit 1")

        return json_result(1, '操作成功', d)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运明细的目的口岸联动


@any_route('/api/saier/shipment/mdka/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_mdka_change(request):
    s = Session()
    j = await request.json()
    try:
        kamc = j.get('kamc', '')
        d = run_sql(
            f"select ssgjz,ssgj from mdka where (mdka='{kamc}') limit 1")
        if len(d) == 0:
            return json_result(-1, '不好意思,口岸错误请重选')
        d = run_sql(
            f"select zwmc from mygb where (zwmc='{d[0].get('ssgjz')}') limit 1")
        if len(d) == 0:
            return json_result(-1, '不好意思,此港口中文国家有误,请更改')

        return json_result(1, '操作成功')
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运明细的汇率联动


@any_route('/api/saier/shipment/hl/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_hl_change(request):
    s = Session()
    j = await request.json()
    try:
        hl = 0
        d = run_sql(
            f"select hhl from hbdm where (hbdm='USD$' or hbdm='USD') limit 1")
        if len(d) > 0:
            hl = d[0]['hhl']

        return json_result(1, '操作成功', hl)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 出运明细的业务人员联动
@any_route('/api/saier/shipment/ywry/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_ywry_change(request):
    s = Session()
    j = await request.json()
    try:
        ywry = j.get('ywry', '')
        hxfp = j.get('hxfp', '')
        fphm = j.get('fphm', '')
        ywbm = ''
        if ywry != '' and ywry != '':
            d = run_sql(
                f"select bm from ywrybiao where (yhm='{ywry}') limit 1")
            if len(d) > 0:
                ywbm = d[0]['bm']
        if hxfp != '' and fphm != '':
            s.query(fysqsheet).filter(fysqsheet.wxfp == fphm).update(
                {fysqsheet.ysfp: hxfp}, synchronize_session=False)
            s.query(hdfy).filter(hdfy.fphm == fphm).update(
                {hdfy.ysfp: hxfp}, synchronize_session=False)
            d = run_sql(
                f"select bz,xm from cyzglsheet where  (zm='核算发票业务组对照表')")
            for r in d:
                if r.get('xm') in hxfp.upper():
                    ywbm = r.get('bz')
                    break
            s.commit()
        return json_result(1, '操作成功', ywbm)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 出运明细的装柜日期联动
@any_route('/api/saier/shipment/zgrq/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_zgrq_change(request):
    s = Session()
    j = await request.json()
    try:
        fphm = j.get('fphm', '')
        zgrq = j.get('zgrq', '')
        zgrqj = j.get('zgrqj', '')
        s.query(bgmxd).filter(bgmxd.fphm == fphm).update(
            {bgmxd.zgrq: zgrq, bgmxd.zgrqj: zgrqj})
        s.commit()
        return json_result(1, '操作成功')
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运明细的装柜日期联动


@any_route('/api/saier/shipment/get/dshl', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_get_dshl(request):
    s = Session()
    j = await request.json()
    try:
        hl = 0
        xszb = j.get('xszb', '')
        fphm = j.get('fphm', '')
        if fphm == '' or xszb == '':
            return json_result(0, '参数为空')
        d = s.query(dshlb.hl).filter(dshlb.hbdm == xszb).first()
        if d:
            hl = float(d.hl)

        return json_result(1, '操作成功', hl)
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 出运明细的装柜日期联动
@any_route('/api/saier/shipment/jgrq/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_jgrq_change(request):
    s = Session()
    j = await request.json()
    try:
        cydh = j.get('cydh', '')
        jgrq = j.get('jgrq', None)
        if cydh == '' or jgrq is None:
            return json_result(0, '参数为空')

        params = [
            str(cydh) + letter
            for letter in string.ascii_uppercase[:19]
        ]
        params.append(cydh)
        s.query(cymx).filter(cymx.fphm.in_(params)).update(
            {cymx.jgrq: jgrq}, synchronize_session=False)
        s.query(bgmxd).filter(bgmxd.fphm.in_(params)).update(
            {bgmxd.jgrq: jgrq}, synchronize_session=False)

        s.commit()
        return json_result(1, '操作成功')
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运明细的是否信保联动


@any_route('/api/saier/shipment/sfxb/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_sfxb_change(request):
    s = Session()
    j = await request.json()
    try:
        khbh = j.get('khbh', '')
        sfxb = j.get('sfxb', '')
        data = {}
        if sfxb == '是':
            d = run_sql(
                f"select LCfl,LCydqx,ydqx from fkkh where (khbh='{khbh}') limit 1")
        else:
            d = run_sql(
                f"select zfqx,fkys from kh where (kh_id='{khbh}') limit 1")
        if len(d) > 0:
            data.update(d[0])

        return json_result(1, '操作成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运明细的产品资料出运唯一字段联动


@any_route('/api/saier/shipment/cywyzd/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_cywyzd_change(request):
    s = Session()
    j = await request.json()
    try:
        cywyzd = j.get('cywyzd', '')
        data = {}
        d = run_sql(
            f"select bjhh,khhh from chyjhsheet where (cywyzd='{cywyzd}') limit 1")
        if len(d) > 0:
            data = d[0]

        return json_result(1, '操作成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运明细的产品资料退税率联动


@any_route('/api/saier/shipment/tsl/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_tsl_change(request):
    s = Session()
    j = await request.json()
    try:
        cywyzd = j.get('cywyzd', '')

        sb = 0
        d = run_sql(
            f"select rid from kaiptzsheet where (cywyzd='{cywyzd}') and ((thdz='否') or (ifnull(thdz,'')='')) limit 1")
        if len(d) > 0:
            sb = 1
        d = run_sql(
            f"select rid from xjhcsheet where (cywyzd='{cywyzd}') and ((thdz='否') or (ifnull(thdz,'')='')) limit 1")
        if len(d) > 0:
            sb = 1

        return json_result(1, '操作成功', sb)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运明细的产品资料海关编码联动


@any_route('/api/saier/shipment/hgbm/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_hgbm_change(request):
    s = Session()
    j = await request.json()
    try:
        hgbm = j.get('hgbm', '')
        data = {}
        d = run_sql(
            f"select tsl,zzsl,cznr from hgbmbsheet where (hgbm='{hgbm}') and (sfpz='是') limit 1")
        if len(d) > 0:
            data = d[0]

        return json_result(1, '操作成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运明细的产品资料采购单价联动


@any_route('/api/saier/shipment/cgdj/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_cgdj_change(request):
    s = Session()
    j = await request.json()
    try:
        data = {}
        d = run_sql(
            f"select sz,sz1 from cyzglsheet where (xm='出运明细清关单价') and (zm='公式计算') limit 1")
        if len(d) > 0:
            data = d[0]

        return json_result(1, '操作成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运明细的产品资料更新总体按钮


@any_route('/api/saier/shipment/item/update/volume', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_item_update_volume(request):
    s = Session()
    j = await request.json()
    try:
        data = []
        bm_list = []
        lines = j.get('lines', [])
        tjhj = 0
        rid = ''
        for line in lines:
            wxbm = line.get('wxbm1', '')
            if not wxbm in bm_list:
                bm_list.append(wxbm)
                d = run_sql(
                    f"select sz,sz1 from cyzglsheet where (xm='{wxbm}') and (zm='单证体积判断') limit 1")
                if len(d) > 0:
                    data.append(wxbm)
                else:
                    tjhj += float(line.get('tjhj', 0))
                    if line.get('ztj', 0) > 0:
                        rid = line.get('rid', '')

        return json_result(1, '操作成功', {'data': data, 'tjhj': tjhj, 'rid': rid})
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运明细的产品资料待开点数联动


@any_route('/api/saier/shipment/item/dkds/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_item_dkds_change(request):
    s = Session()
    j = await request.json()
    try:
        dkds = 3.24
        d = run_sql(f"select cs from zx where ly='开票点数3' limit 1")
        if len(d) > 0:
            dkds = float(d[0]['cs'])

        return json_result(1, '操作成功', dkds)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运明细的货物状态改变


@any_route('/api/saier/shipment/hhzt/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_hhzt_change(request):
    s = Session()
    j = await request.json()
    try:
        hglx = j.get('hglx', '')
        ysfs = j.get('ysfs', '')
        fpsb = j.get('fpsb', '')
        wxfp = j.get('wxfp', '')
        khmc = j.get('khmc', '')
        hwzt = j.get('hwzt', '')
        gdsb = 0
        bptjx = 0
        bptjd = 0
        if hwzt == '已出运' and 'BEST PRICE' in khmc.upper():
            d = s.query(cyzglsheet.sz, cyzglsheet.sz1, cyzglsheet.bz1).filter(
                cyzglsheet.xm == hglx,
                cyzglsheet.zm == 'BP装柜体积范围',
                cyzglsheet.bz.like(f'%{ysfs}%'),
                ~cyzglsheet.bz1.like(f'%{fpsb}%')
            ).first()
            if d:
                bptjx = float(d.sz)
                bptjd = float(d.sz1)

        d = run_sql(
            f"select rid from chyjh where (wxfp='{wxfp}') and (fkzk='是') limit 1")
        if len(d) > 0:
            gdsb = 3

        return json_result(1, '操作成功', {'gdsb': gdsb, 'bptjx': bptjx, 'bptjd': bptjd})
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 出运明细的产品资料出货数量联动


@any_route('/api/saier/shipment/item/chsl/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_item_chsl_change(request):
    s = Session()
    j = await request.json()
    try:
        khmc = j.get('khmc', '')
        flag = 0
        d = run_sql(
            f"select bz from cyzglsheet where (xm='{khmc}') and (zm='按佣金单价计算佣金') limit 1")
        if len(d) > 0:
            flag = 1

        return json_result(1, '操作成功', flag)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 出运明细的产品资料毛重按钮
@any_route('/api/saier/shipment/item/update/weight', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_item_update_weight(request):
    s = Session()
    j = await request.json()
    try:
        check_list = []
        rids_json = {}
        cjzg = float(j.get('cjzg', 0))
        cjkg = float(j.get('cjkg', 0))
        mzhj = float(j.get('mzhj', 0))
        xshj1 = float(j.get('xshj1', 0))
        lines = j.get('lines', [])
        i1 = 0
        i2 = 0
        mz2 = 0
        zxs = 0
        mz_js = []
        for line in lines:
            wxbm = line.get('wxbm1', '')
            zycpbh = line.get('zycpbh', '')
            rid = line.get('rid', '')
            if not str(wxbm) + '_' + str(zycpbh) in check_list:
                check_list.append(str(wxbm) + '_' + str(zycpbh))
                d = run_sql(
                    f"select rid from cyzglsheet where ((xm='{wxbm}') or (xm='{zycpbh}')) and (zm='单证毛重判断') limit 1")
                if len(d) > 0:
                    mz_js.append(str(wxbm) + '_' + str(zycpbh))

            if str(wxbm) + '_' + str(zycpbh) in mz_js:
                rids_json[rid] = '否'
                zxs = line.get('chxs', 0) + zxs
            else:
                rids_json[rid] = '是'

        if (cjzg - cjkg > mzhj):
            mz2 = (cjzg - cjkg - mzhj) / (xshj1 - zxs)
            i1 = i1 + 1
        else:
            mz2 = (cjkg + mzhj - cjzg) / (xshj1 - zxs)
            i2 = i2 + 1
        if mz2 > 0.1:
            return json_result(-1, f"毛重差距等于{round(mz2,2)}大于0.1KG,请先向业务员核实后,手动更改")

        return json_result(1, '操作成功', {'rids_json': rids_json, 'zxs': zxs, 'mz2': round(mz2, 2)})
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 出运明细的产品资料更新出运信息按钮
@any_route('/api/saier/shipment/item/update/check', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_item_update_check(request):
    s = Session()
    j = await request.json()
    user = request.current_user
    try:
        kind = j.get('kind', 0)
        job = j.get('job', 0)
        rid = j.get('rid', '')
        dz1 = 0
        dz = 0
        sb = 0
        khmc = j.get('khmc', '')
        # if kind == 0:
        #     d = run_sql(f"select rid from sys_user where (username='{user.username}') and (position like '%单证%') limit 1")
        #     if len(d) == 0:
        #         return json_result(-1, '当前用户不是单证人员，不能执行此操作')
        # else:
        #     if job != 1:
        #         dz1 = 1
        #     else:
        #         d = s.query(cymx.uid).filter(cymx.rid==rid).first()
        #         if d:
        #             d = run_sql(f"select rid from sys_user where (rid='{d.uid}') and (position like '%单证%') limit 1")
        #             if len(d) > 0:
        #                 dz1 = 1
        #     if dz1 == 0 and dz == 0:
        #         return json_result(-1, '当前记录的用户不是单证人员，不能执行此操作')

        #     d = run_sql(f"select bz from cyzglsheet where (xm='{khmc}') and (zm='按佣金单价计算佣金') limit 1")
        #     if len(d) > 0:
        #         sb = 1

        return json_result(1, '操作成功', {'dz1': dz1, 'dz': dz, 'sb': sb})
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 出运明细的产品资料更新信息按钮
@any_route('/api/saier/shipment/item/update/info', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_item_update_info(request):
    s = Session()
    j = await request.json()
    user = request.current_user
    try:
        pid = j.get('rid', '')
        khmc = j.get('khmc', '')
        lines = j.get('lines', [])
        krID = j.get('krID', '')
        selected = j.get('selected', '所有记录')
        data = {}
        data['nxkh'] = ''
        data['sz'] = 0
        if khmc != '' and khmc != None:
            d = run_sql(
                f"select nxkh from kh where company_name='{khmc}' limit 1")
            if len(d) > 0:
                data['nxkh'] = d[0].get('nxkh', '')
        d = run_sql(
            f"select sz from cyzglsheet where (xm='公司') and (zm='抹零金额') limit 1")
        if len(d) > 0:
            data['sz'] = d[0].get('sz', 0)
        amazon = 1 if 'AMAZON' in khmc.upper() else 0
        data['mlgc'] = {}
        data['hgbm'] = {}
        data['ywry'] = {}
        data['skxx'] = {}
        data['kpgc'] = {}
        for r in lines:
            rid = r.get('rid', '')
            sfgd = r.get('sfgd', '')
            hthm = r.get('cght', '')
            ywpath = r.get('ywpath', '')
            ywry = r.get('ywry', '')
            wxwyzd = r.get('wxwyzd', '')
            order_id = r.get('wxht', '')
            wxrl = r.get('wxrl', 0)
            zzsl = r.get('zzsl', 0)
            tsl = r.get('tsl', 0)
            hgdm = r.get('hgbm', '')
            pgwy = r.get('pgwy', '')
            if selected == '所有解锁数据' and sfgd != '解锁':
                continue
            bjhh = r.get('bjhh', '')
            cpbh = r.get('cpbh', '')
            logger.error('zycpbh aaa: %s', r.get('zycpbh', ''))
            logger.error('bjhh aaa: %s', bjhh)
            zyhh = r.get('zycpbh', '')
            zycpbh = r.get('zycpbh', '')
            gcmc = r.get('gcmc', '')
            hyd = r.get('hyd', '')
            if zyhh == '' or zyhh == None:
                if bjhh != '' and bjhh != None:
                    zyhh = bjhh
                else:
                    zyhh = '无货号'
            hthm = r.get('cght', '')
            data[rid] = {}
            data[rid]['wxpd'] = 0
            data[rid]['ttly'] = ''
            data[rid]['ytfy'] = []
            if hthm != '' and hthm != None:
                if hthm not in data:
                    data[hthm] = {}
                    data[hthm]['cgsb'] = 0
                    data[hthm]['ttly'] = ''
                    data[hthm]['items'] = []
                    data[hthm]['main'] = {}
                logger.error('aaa----aaaaa')
                logger.error(data[hthm]['main'])
                if data[hthm]['main'] == None or data[hthm]['main'] == {}:
                    logger.error('bbb----bbb')
                    c = run_sql(
                        f"select cs_id,sccj,sccj1id,sccj1,lxry,gcdh,sjhm,gccz,jsfs,cgry,gdry,rkdd,sccj1id,szxq,zh,bank,skfm from cght where (hthm='{hthm}') and (sfhz='通过') limit 1")
                    if len(c) > 0:
                        data[hthm]['main'] = c[0]
                        data[hthm]['cgsb'] = 1
                    c = run_sql(
                        f"select fktt from fkspsheet3 where (hthm='{hthm}') limit 1")
                    if len(c) > 0:
                        data[hthm]['fktt'] = c[0]['fktt']
                        data[hthm]['ttly'] = '预付申请'
                
                if data[hthm].get('cgsb', 0) == 1:
                    if pgwy != '' and pgwy != None:
                        m = s.query(cghtsheet).filter(cghtsheet.hthm == hthm, or_(
                            cghtsheet.bjhh == zyhh, cgbjsheet.khhh == bjhh), cghtsheet.wyzd == pgwy).first()
                    else:
                        m = s.query(cghtsheet).filter(cghtsheet.hthm == hthm, or_(
                            cghtsheet.bjhh == zyhh, cgbjsheet.khhh == bjhh), cghtsheet.wxrl == wxrl).order_by(cghtsheet.sid.desc()).first()
                    if m:
                        m.dzsd = user.username
                        m.sfhz = '通过'
                        s.add(m)
                        c = alchemy_object_to_dict(m)
                        data[rid]['cgcp'] = c
                        data[rid]['cgsb'] = data[hthm].get('cgsb', 0)
                        data[rid]['cght'] = hthm
                        data[rid]['ttly'] = data[rid]['ttly'] + '采购合同'
                else:
                    data[hthm]['sfgd'] = '解锁1'
                    data[hthm]['wgxyy'] = '采购合同末审批'
                if wxwyzd != '' and wxwyzd != None and order_id != '' and order_id != None and zyhh != '' and zyhh != None:
                    if wxwyzd != '' and wxwyzd != None:
                        c = run_sql(f"select khmc,bjyw,Twxdj,hhbz,wxdj,jldw,yjcq,mjdj1,zycpbh,wxbm1,khhh,cply1,pkrmb,ywpm,dkje,dkds,yjbl,yj,sj,khpd,htsl,aybl,asl6,ayj,sl6,krddh,sfsq,mdck,htbj,hydj,hyRMBdj,\
                            wxwyzd,jdrq from wxhtsheet where (wxwyzd='{wxwyzd}') limit 1")
                    else:
                        c = run_sql(f"select khmc,bjyw,Twxdj,hhbz,wxdj,jldw,yjcq,mjdj1,zycpbh,wxbm1,khhh,cply1,pkrmb,ywpm,dkje,dkds,yjbl,yj,sj,khpd,htsl,aybl,asl6,ayj,sl6,krddh,sfsq,mdck,htbj,hydj,hyRMBdj,\
                            wxwyzd,jdrq from wxhtsheet where (order_id='{order_id}') and (bjhh='{zycpbh}') limit 1")
                    if len(c) > 0:
                        data[rid]['wxcp'] = c[0]
                        if c[0].get('jdrq', '') != '' and c[0].get('jdrq', '') != None and c[0].get('jdrq') != '1900-01-01':
                            data[rid]['tsl'] = 1
                        s.query(wxhtsheet).filter(wxhtsheet.wxwyzd == wxwyzd).update(
                            {wxhtsheet.dzsd: user.username}, synchronize_session=False)
                    else:
                        data[rid]['wxpd'] = 1
                if cpbh != '' and cpbh != None:
                    c = run_sql(
                        f"select * from zscp where (cpbh='{zyhh}') or (krhh='{cpbh}') limit 1")
                    logger.error('ccc----ccc')
                    logger.error(c)
                    if len(c) > 0:
                        data[rid]['zscp'] = c[0]
                    c = run_sql(
                        f"select krhh,krtm,mjfob,Twxdj,rmbdj,pkRMB,djpm,djpmy,djpmw,krcode,krsl,wypp,ggwy from zscpsheet7 where ((cpbh='{zyhh}') or (krhh='{cpbh}')) and (krID='{krID}') limit 1")
                    if len(c) > 0:
                        data[rid]['zscpsheet'] = c[0]

                if gcmc != '' and gcmc != None and gcmc not in data['mlgc']:
                    c = run_sql(
                        f"select rid from bmlgc where (gcmc='{gcmc}') limit 1")
                    if len(c) > 0:
                        data['mlgc'][gcmc] = 0
                    else:
                        data['mlgc'][gcmc] = 1

                if gcmc != '' and gcmc != None and gcmc not in data['kpgc']:
                    c = run_sql(
                        f"select fkhm,bank1,zh1 from zycs where (company_name='{gcmc}') limit 1")
                    if len(c) > 0:
                        data['kpgc'][gcmc] = c[0]

                if hgdm != '' and hgdm != None and zzsl > 4 and tsl != 16 and hgdm not in data['hgbm']:
                    c = run_sql(
                        f"select tsl,zzsl,sfsj from hgbmb where hgbmb.hgbm='{r.get('hgbm', '')}' limit 1")
                    if len(c) > 0:
                        data['hgbm'][hgdm] = c[0]

                if amazon == 1:
                    c = run_sql(
                        f"select *,1 amazon from dsytfysheet where (cght='{hthm}') and (ifnull(cght,'')<>'') and (sccj='{gcmc}') and (hyd='{hyd}') and (zzsl={zzsl})")
                else:
                    c = run_sql(
                        f"select *,0 amazon from ytfysheet where (cght='{hthm}') and (ifnull(cght,'')<>'') and (sccj='{gcmc}') and (hyd='{hyd}') and (zzsl={zzsl})")
                if len(c) > 0:
                    data[rid]['ytfy'] = c
                    if amazon == 1:
                        data[rid]['ttly'] = data[rid]['ttly'] + '电商预填'
                    else:
                        data[rid]['ttly'] = data[rid]['ttly'] + '预填'

                if (ywpath == '' or ywpath == None) and ywry != '' and ywry != None and ywry not in data['ywry']:
                    org = get_user_path(ywry)
                    data[rid]['ywry'] = org.get('path', '')

        s.commit()
        return json_result(1, '操作成功', data)
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 出运明细的产品资料更新出运信息按钮
@any_route('/api/saier/shipment/item/update/items', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_item_update_items(request):
    s = Session()
    j = await request.json()
    try:
        rids_json = {}
        lines = j.get('lines', [])
        fphm = j.get('fphm')
        d = s.query(chyjh.rid).filter(chyjh.wxfp == fphm).first()
        if not d:
            return json_result(-1, '未找到出运计划记录')
        pid = d.rid
        for line in lines:
            cywyzd = line.get('cywyzd')
            cywyzd12 = line.get('cywyzd12')
            cywyzd123 = line.get('cywyzd123')
            sfgd = line.get('sfgd')
            rid = line.get('rid')
            if cywyzd12 == '' or cywyzd12 == None or (sfgd != '是' and cywyzd123 != '新增资料'):
                continue
            rids_json[rid] = {'selected': '是'}
            d = run_sql(
                f"select * from chyjhsheet where (cywyzd='{cywyzd}') and (ifnull(sdjy,'')='' or sdjy<>'不通过') and pid='{pid}' limit 1")
            if len(d) > 0:
                rids_json[rid]['cyjh'] = d[0]

        return json_result(1, '操作成功', rids_json)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 出运明细的产品资料批量解锁按钮
@any_route('/api/saier/shipment/item/update/unlock', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_item_update_unlock(request):
    s = Session()
    j = await request.json()
    user = request.current_user
    try:
        kind = j.get('kind', '')
        lines = j.get('lines', [])
        flag = 0
        rids = []
        await messages.message(user.username, {'msg': '正在处理中。。。', 'kind': '0'}, MSG_KIND_NORMAL)
        index = 0
        for line in lines:
            cywyzd = line.get('cywyzd', '')
            bjhh = line.get('bjhh', '')
            zyhh = line.get('zycpbh', '')
            wxwyzd = line.get('wxywzd', '')
            pgwy = line.get('pgwy', '')
            wxrl = line.get('wxrl', 0)
            hthm = line.get('cght', '')
            rid = line.get('rid', '')
            khhh = line.get('khhh', '')
            index = index + 1
            await messages.message(user.username, {'msg': f"正在处理采购合同:{str(hthm)}", 'total': 99, 'progress': math.floor(index / len(lines) * 99), "kind": "0"}, MSG_KIND_NORMAL)
            if zyhh == '':
                if bjhh != '':
                    zyhh = bjhh
                else:
                    zyhh = '无货号'
            if kind == '全部' or kind == '采购':
                d = run_sql(
                    f"select rid from kaiptzsheet where (cywyzd='{cywyzd}') and ((ifnull(thdz,'')='否') or (ifnull(thdz,'')='')) limit 1")
                if len(d) > 0:
                    await messages.message(user.username, {'msg': f"此票已生成开票通知，更改请通知财务", 'total': 99, 'progress': math.floor(index / len(lines) * 99), "kind": "2"}, MSG_KIND_NORMAL)
                    return json_result(-1, '此票已生成开票通知，更改请通知财务')
                d = run_sql(
                    f"select rid from xjhcsheet where (cywyzd='{cywyzd}') and ((ifnull(thdz,'')='否') or (ifnull(thdz,'')='')) limit 1")
                if len(d) > 0:
                    await messages.message(user.username, {'msg': f"此票已生成信息合成，更改请通知财务", 'total': 99, 'progress': math.floor(index / len(lines) * 99), "kind": "2"}, MSG_KIND_NORMAL)
                    return json_result(-1, '此票已生成信息合成，更改请通知财务')
            title = '货号:' + str(zyhh[:45]) + '/客户货号:' + str(khhh[:20]) + '申请解锁'
            res = user_task_delete('出运明细', rid, s, [user.username], title)
            if res.get('code') != 1:
                return json_result(-1, res.get('msg'))
            if (hthm == '' or hthm == None):
                continue
            s.query(cymxsheet).filter(cymxsheet.wxwyzd == wxwyzd, cymxsheet.cght == hthm, func.ifnull(cymxsheet.sjcy, '') == '', or_(
                cymxsheet.fpsb1 == '是', cymxsheet.fksb == '是')).update({cymxsheet.sfgd: '解锁'}, synchronize_session=False)
            if kind == '外销':
                s.query(wxhtsheet).filter(wxhtsheet.wxywzd == wxwyzd).update(
                    {wxhtsheet.hthm: ''}, synchronize_session=False)
            if kind == '采购':
                s.query(cghtsheet).filter(cghtsheet.hthm == hthm, or_(cghtsheet.wyzd == pgwy, cghtsheet.wxrl == wxrl), func.ifnull(
                    cghtsheet.wyzd, '') != '').update({cghtsheet.dzsd: ''}, synchronize_session=False)
            flag = 1
            rids.append(rid)

        if flag == 1:
            s.commit()
        await messages.message(user.username, {'msg': f"处理完成", 'total': 100, 'progress': 100, "kind": "1"}, MSG_KIND_NORMAL)
        return json_result(1, '操作成功', rids)
    except:
        s.rollback()
        logger.error(trace_error())
        await messages.message(user.username, {'msg': f"trace_error()", 'total': 99, 'progress': 99, "kind": "2"}, MSG_KIND_NORMAL)
        return json_result(-1, trace_error())
    finally:
        s.close()


def shipment_lines_update_invoice(lines, fphm, cydh, s):
    s = Session()
    try:
        for line in lines:
            s.query(cymxsheet).filter(cymxsheet.chydh == fphm, cymxsheet.rid == line.get(
                'rid')).update({cymxsheet.fpsfytq: '是'}, synchronize_session=False)

        return {'code': 1, 'msg': '操作成功'}
    except:
        logger.error(trace_error())
        return {'code': -1, 'msg': trace_error()}


# 出运明细的产品资料记录删除校验
@any_route('/api/saier/shipment/item/delete/check', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_delete_check(request):
    s = Session()
    j = await request.json()
    try:
        cywyzd = j.get('cywyzd', '')
        hthm = j.get('hthm', '')
        wxwyzd = j.get('wxwyzd', '')
        cpbh = j.get('cpbh', '')
        rid = j.get('rid', '')
        fphm = j.get('fphm', '')
        sfdb = j.get('sfdb', '否')
        fksb = j.get('fksb', '否')
        cydh = j.get('cydh', '')
        cydh2 = j.get('cydh2', '')
        fphm2 = j.get('fphm2', '')
        if cywyzd == '' or cywyzd == None or sfdb == '是':
            return json_result(0, '新纪录校验跳过')
        m = s.query(cymxsheet).filter(cymxsheet.rid == rid).first()
        if not m:
            return json_result(0, '新纪录校验跳过')
        d = run_sql(
            f"select cywyzd from kaiptzsheet where (cywyzd='{cywyzd}') limit 1")
        if len(d) > 0:
            return json_result(-1, '请注意此票已生成开票通知，不能删除')
        d = run_sql(
            f"select rid from bgmxdsheet where  (cght='{hthm}') and (cpbh='{cpbh}') and (cywyzd='{cywyzd}') and (wxwyzd='{wxwyzd}') and (ifnull(yfph,'')<>'') limit 1")
        if len(d) > 0:
            return json_result(-1, '请注意此票在报关明细已补报，不能删除')
        if fphm != '' and cydh != '' and cydh2 != fphm2:
            m.fpsfytq = '否'
            s.add(m)

        s.query(dbbqd).filter(dbbqd.cght == hthm, dbbqd.cpbh == cpbh, dbbqd.cywyzd == cywyzd,
                dbbqd.wxwyzd == wxwyzd, dbbqd.fphm == fphm).delete(synchronize_session=False)
        if fksb == '是':
            s.query(gcdjsheet2).filter(gcdjsheet2.cght == hthm,
                    gcdjsheet2.cywyzd == cywyzd).delete(synchronize_session=False)
        cyxq = []
        d = run_sql(
            f"select chxs,fphm,sjcy from cymxsheet where (wxwyzd='{wxwyzd}') and (cght='{hthm}') and (ifnull(sjcy,'')<>'') and (cywyzd<>'{cywyzd}') AND (fksb='是')")
        for r in d:
            cyxq.append(
                f"{r.get('chxs',0)}\\{r.get('sjcy','')}({r.get('fphm','')})")
        s.query(wxhtsheet).filter(wxhtsheet.wxwyzd == wxwyzd).update(
            {wxhtsheet.dzsd: ''}, synchronize_session=False)
        s.query(cghtsheet).filter(cghtsheet.wxwyzd == wxwyzd, cghtsheet.hthm == hthm).update(
            {cghtsheet.cyxq: ''.join(cyxq), cghtsheet.dzsd: ''}, synchronize_session=False)

        s.commit()
        return json_result(1, '操作成功')
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


def shipment_items_sfdb_update(mains, lines, bgzt, user, s, batch=0):
    try:
        new_data = {}
        for r in lines:
            rid = r.get('rid')
            if batch == 1:
                if r.get('sfdb') == bgzt:
                    continue
                sfdb = bgzt
            else:
                sfdb = r.get('sfdb', '否')
            logger.error(f"sfdb:{sfdb}")
            cywyzd = r.get('cywyzd')
            zzsl = r.get('zzsl', 0)
            cpbh = r.get('cpbh')
            wxwyzd = r.get('wxwyzd')
            hthm = r.get('cght')
            cgdj = r.get('cgdj', 0)
            sfpx = r.get('sfpx', '否')
            wxrl = r.get('wxrl', 0)
            wxmz = r.get('wxmz', 0)
            wxjz = r.get('wxjz', 0)
            wxtj = r.get('wxtj', 0)
            mjzj = r.get('mjzj', 0)
            wxdj = r.get('wxdj', 0)
            cgdj = r.get('cgdj', 0)
            if sfdb == '' or sfdb == None:
                continue
            sb1 = 0
            if cywyzd == '' or cywyzd == None:
                cywyzd = rid
            c = run_sql(
                f"select rid from kaiptzsheet where (cywyzd='{cywyzd}') and (thdz<>'是') limit 1")
            if len(c) > 0:
                sb1 = 1
            if sb1 != 1:
                c = run_sql(
                    f"select rid from xjhcsheet where (cywyzd='{cywyzd}') and (thdz<>'是') limit 1")
                if len(c) > 0:
                    sb1 = 1
            logger.error(f"sb1:{sb1}")
            if sb1 == 1:
                c = run_sql(
                    f"select sfdb from cymxsheet where (rid='{rid}') limit 1")
                if len(c) > 0:
                    if c[0].get('sfdb') != sfdb:
                        new_data[rid] = {'sfdb': c[0].get(
                            'sfdb'), 'msg': '请注意此票已有开票或现金合成不能再选'}
                        sfdb = ''
                continue
            if zzsl <= 0:
                sfdb = ''
                new_data[rid] = {'sfdb': '否',
                    'bbxq': '', 'msg': '增值税率必须大于0才能选择待报'}

            if sfdb == '是':
                chsl = 0
                bbwc = 0
                chslz = r.get('chsl', 0)
                d = s.query(cymxsheet.rid).filter(cymxsheet.cpbh == cpbh, cymxsheet.cywyzd ==cywyzd, cymxsheet.wxwyzd == wxwyzd, cymxsheet.sfdb == '是').first()
                if not d:
                    c = s.query(bgmxdsheet.rid).filter(bgmxdsheet.cpbh == cpbh, bgmxdsheet.cywyzd == cywyzd, bgmxdsheet.wxwyzd == wxwyzd, func.ifnull(bgmxdsheet.yfph, '') == '').first()
                    if c:
                        new_data[rid] = {'msg': f"请注意此合同{hthm}已在报关明细"}
                    c = s.query(bgmxdsheet.rid).filter(bgmxdsheet.cpbh == cpbh, bgmxdsheet.cywyzd ==cywyzd, bgmxdsheet.wxwyzd == wxwyzd, func.ifnull(bgmxdsheet.yfph, '') != '').all()
                    for l in c:
                        if l.bbwc == '是':
                            bbwc = 1
                        if cgdj > 0:
                            chsl = chsl + round(l.gczj / cgdj)
                        else:
                            if float(l.gcjg) > 0:
                                chsl = chsl + \
                                    math.trunc(float(l.gczj) / float(l.gcjg))
                            else:
                                chsl = chsl + l.chsl
                    logger.error(f"chsl:{chsl}, chslz:{chslz}, bbwc:{bbwc}")
                    logger.error(f"chslz - chsl: {chslz - chsl}")
                    if bbwc != 1 and (chslz - chsl) > 0:
                        s.query(dbbqd).filter(dbbqd.cpbh == cpbh, dbbqd.cywyzd == cywyzd,
                            dbbqd.wxwyzd == wxwyzd).delete(synchronize_session=False)
                        m = dbbqd()
                        m.rid = get_uuid()
                        for k, v in mains.items():
                            if k in SYS_FIELDS or k in ['Field']:
                                continue
                            if hasattr(m, k):
                                setattr(m, k, v)
                        for k, v in r.items():
                            if k in SYS_FIELDS or k in ['Field']:
                                continue
                            if hasattr(m, k):
                                setattr(m, k, v)
                        m.uid = user.rid
                        m.ctime = time.strftime('%Y-%m-%d %H:%M:%S')
                        m.chsl = r.get('chsl', 0) - chsl
                        m.name = 'name'
                        m.ck = '是'
                        m.wxfp = ''
                        m.Field1 = r.get('bfmy', 0)
                        if sfpx == '否' or sfpx == '' or sfpx == None:
                            if wxrl != 0:
                                m.chxs = (chslz - chsl) / wxrl
                                m.zmz = round(
                                    ((chslz - chsl) / wxrl) * wxmz, 2)
                                m.zjz = round(
                                    ((chslz - chsl) / wxrl) * wxjz, 2)
                                m.ztj = round(
                                    ((chslz - chsl) / wxrl) * wxtj, 3)
                            else:
                                m.chxs = 0
                                m.zmz = 0
                                m.zjz = 0
                                m.ztj = 0
                        else:
                            m.chxs = 0
                            m.zmz = 0
                            m.zjz = 0
                            m.ztj = 0
                        m.mjzj = (chslz - chsl) * mjzj
                        m.wxzj = (chslz - chsl) * wxdj
                        m.gczj = (chslz - chsl) * cgdj
                        m.kh_id = mains.get('kh_id', '')
                        m.chyrq = mains.get('chyrq')
                        m.cghbdm = r.get('cghbdm', '')
                        s.add(m)
                        new_data[rid] = {'sfdb': '是', 'msg': ""}
                        sfdb = ''
                else:
                    new_data[rid] = {'sfdb': '否',
                        'msg': f"请注意此票在出运明细{mains.get('fphm', '')}已待报不能再选!!"
                    }
                    sfdb = ''

            elif sfdb == '否':
                db = 0
                c = run_sql(
                    f"select fphm from dbbqd where (cpbh='{cpbh}') and (cywyzd='{cywyzd}') and (wxwyzd='{wxwyzd}') and (ck='否') limit 1")
                if len(c) > 0:
                    db = 1
                    new_data[rid] = {'sfdb': '是',
                        'msg': f"请注意此票在报关明细已报或已选不能再取消!!"}
                    sfdb = ''
                if db != 1:
                    c = run_sql(
                        f"select chsl,gczj,gcjg,fphm from bgmxdsheet where  (cpbh='{cpbh}') and (cywyzd='{cywyzd}') and (wxwyzd='{wxwyzd}')  and (ifnull(yfph,'')<>'') limit 1")
                    if len(c) > 0:
                        db = 1
                        new_data[rid] = {
                            'sfdb': '是', 'msg': f"请注意此票在报关明细{c[0]['fphm']}已报不能再取消!!"}
                        sfdb = ''
                if db != 1:
                    s.query(dbbqd).filter(dbbqd.cywyzd == cywyzd, dbbqd.wxwyzd == wxwyzd).delete(
                        synchronize_session=False)
                    if batch == 1:
                        new_data[rid] = {'sfdb': '否', 'bbxq': '', 'msg': ""}
                    else:
                        new_data[rid] = {'sfdb': '', 'bbxq': '', 'msg': ""}

        return {'code': 1, 'msg': '操作成功', 'data': new_data}
    except:
        logger.error(trace_error())
        return {'code': -1, 'msg': trace_error()}


# 出运明细的产品资料是否待报更新
@any_route('/api/saier/shipment/item/update/sfdb', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_item_update_sfdb(request):
    s = Session()
    j = await request.json()
    user = request.current_user
    try:
        lines = j.get('lines', [])
        mains = j.get('mains', {})
        bgzt = j.get('bgzt', '')
        batch = j.get('batch', 0)
        res = shipment_items_sfdb_update(mains, lines, bgzt, user, s, batch)
        if res.get('code') != 1:
            s.rollback()
            return json_result(-1, res.get('msg'))

        s.commit()
        return json_result(1, '操作成功', res.get('data'))
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


def shipment_update_by_sea_data(main, user, s):
    try:
        u = run_sql(
            f"select rid from sys_user where (username='{user.username}') and (position like '%单证%' or memo like '%单证%') limit 1")
        if len(u) == 0:
            return {'code': 0, 'msg': '只有单证用户才能执行该功能'}
        cydh = main.get('cydh', '')
        fphm = main.get('fphm', '')
        tdh = main.get('fttdh', '')
        cdgs = main.get('FTShipCompany', '')
        gqdm = main.get('FTPortAreaCode', '')
        shr = main.get('shr', '')
        kh_id = main.get('kh_id', '')
        sjcy = main.get('sjcy1', '')
        webpd = main.get('webpd', '')
        rmbkh = main.get('rmbkh', '')
        wyzd = main.get('wyzd', '')
        if tdh != '' and tdh != None and fphm != '' and fphm != None and cdgs != '' and cdgs != None and gqdm != '' and gqdm != None:
            d = s.query(hyxx).filter(hyxx.tdh == tdh).first()
            if not d:
                d = hyxx()
                d.rid = get_uuid()
                d.ctime = time.strftime('%Y-%m-%d %H:%M:%S')
                org = get_user_path('谢晓霞')
                d.uid = org.get('uid', '') if org.get(
                    'uid', '') != '' else user.rid
                d.fphm = fphm
                d.tdh = tdh
                d.zdry = main.get('zdry', '')
                d.ywry = main.get('ywry', '')
                d.hyxh = main.get('hyxh', '')
                d.hyfh = main.get('hyfh', '')
                d.FTShipCompany = cdgs
                d.FTPortAreaCode = gqdm
                d.cymxfather = main.get('rid', '')
                d.hyfh = main.get('hyfh', '')
            else:
                m = s.query(hyxx).filter(
                    or_(hyxx.tdh == tdh, hyxx.fphm == fphm)).all()
                for d in m:
                    if d.hyxh == '' or d.hyxh == None or d.hyfh == '' or d.hyfh == None:
                        d.hyxh = main.get('hyxh', '')
                        d.hyfh = main.get('hyfh', '')
                    d.zdry = main.get('zdry', '')
                    d.ywry = main.get('ywry', '')
                    d.sjcy1 = main.get('sjcy1', '')
                    s.add(d)

        if fphm != cydh and fphm != None:
            if rmbkh == '是':
                rmbkh = 'RMB'
            else:
                rmbkh = 'USD'
            d = s.query(gdqkb).filter(gdqkb.yfphm == wyzd).first()
            if not d:
                d = gdqkb()
                d.rid = get_uuid()
                d.ctime = time.strftime('%Y-%m-%d %H:%M:%S')
                org = get_user_path('赵波')
                d.uid = org.get('uid', '') if org.get(
                    'uid', '') != '' else user.rid
                d.fphm = main.get('fphm', '')
                d.khmc = main.get('khmc', '')
                d.chyrq = main.get('chyrq')
                d.sjcy1 = main.get('sjcy1')
                d.mygb = main.get('mygb', '')
                d.ywry = main.get('ywry', '')
                d.ywbm = main.get('ywbm', '')
                d.xRMBkh = rmbkh
                d.RMBkh = rmbkh
                d.yfphm = wyzd
            else:
                if (d.xhtje != main.get('htje', 0)) or (d.xRMBkh != rmbkh) or (d.fphm != main.get('fphm', '')):
                    bgjl = '原发票号' + main.get('ychydh', '') + '在' + time.strftime('%Y-%m-%d') + '变更发票号为' + main.get(
                        'fphm', '') + '货币代码' + main.get('rmbkh', '') + '货值合计' + str(main.get('htje', 0))
                    d.modi_uid = user.rid
                    d.mtime = time.strftime('%Y-%m-%d %H:%M:%S')
                    d.xRMBkh = rmbkh
                    d.bgrq = time.strftime('%Y-%m-%d')
                    d.xhtje = main.get('htje', 0)
                    d.fphm = main.get('fphm', '')
                    d.bgjl = bgjl
                    user_list = []
                    c = run_sql(
                        f"select xm,bz from cyzglsheet with (nolock) where (zm='风控改单通知')")
                    for r in c:
                        if r.get('xm') not in user_list:
                            user_list.append(r.get('xm'))
                        row = {
                            "xxly": '改单情况表',
                            "bjdh": '',
                            "wxht": main.get('fphm', ''),
                            "cght": '',
                            "yhdh": '',
                            "xxnr": bgjl,
                            "jsr": str(r),
                            "sys_path": "",
                            "spsq": user.username
                        }
                        res = module_xxck_new([row], user, s)
                        if res.get('code', 1) != 1:
                            return {'code': -1, 'msg': res.get('msg')}

                    res = user_task_new(
                        '改单情况表', d.rid, '发票号码', f"发票号{main.get('fphm', '')}在{time.strftime('%Y-%m-%d')}有变更", bgjl, user, s, user_list)
                    if res.get('code') != 1:
                        s.rollback()
                        return {'code': -1, 'msg': res.get('msg')}
                    s.add(d)
            if webpd == '是' and sjcy != '' and sjcy != None:
                flag = 0
                d = s.query(kh).filter(kh.kh_id == kh_id).first()
                if d:
                    if shr != '' and shr != None:
                        d.shr = shr[:250]
                        flag = 1
                    if d.hzdj == "一级" or d.hzdj == "潜在客人":
                        d.hzdj = '三级'
                        flag = 1
                    if flag == 1:
                        s.add(d)

            c = s.query(cymxsheet.zsbh).filter(func.ifnull(cymxsheet.zsbh, '') != '',
                        cymxsheet.fksb == '是', cymxsheet.fphm == fphm).group_by(cymxsheet.zsbh).all()
            l = [r.zsbh for r in c]
            for r in l:
                d = s.query(cymxsheet).filter(cymxsheet.sjcy >= '2023-12-01', cymxsheet.fphm == fphm, cymxsheet.fphm == cymxsheet.chydh,
                    cymxsheet.zsbh == r, func.ifnull(cymxsheet.sjcy, '') != '', or_(
                        cymxsheet.xpsb == '是', and_(cymxsheet.rid < 527724, func.length(cymxsheet.rid) < 30))
                ).order_by(cymxsheet.sjcy.asc()).first()
                if d:
                    r.xpsb = '是'
                    r.je1 = r.gczj
                    s.add(d)
                    s.query(zscp).filter(zscp.zycphhp == r.zsbh).update(
                        {zscp.chrq: sjcy}, synchronize_session=False)

        return {'code': 1, 'msg': '操作成功'}
    except:
        logger.error(trace_error())
        return {'code': -1, 'msg': trace_error()}


# 出运明细的记录保存后执行
@any_route('/api/saier/shipment/after/save', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_after_save(request):
    s = Session()
    j = await request.json()
    user = request.current_user
    try:
        # lines = j.get('lines', [])
        # fphm = j.get('fphm', '')
        # cydh = j.get('cydh', '')
        main = j.get('main', {})
        res = shipment_update_by_sea_data(main, user, s)
        if res.get('code') == -1:
            s.rollback()
            return json_result(-1, res.get('msg'))

        s.commit()
        return json_result(1, '操作成功')
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 出运明细的单据确认按钮
@any_route('/api/saier/shipment/document/confirm', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_document_confirm(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        rid = j.get('rid', '')
        hxdh = str(j.get('hxdh', '')).strip()
        if hxdh == '':
            return json_result(-1, '请输入单据确认日期,日期格式2010-01-18')

        try:
            hxdh_date = datetime.strptime(hxdh, '%Y-%m-%d')
        except:
            return json_result(-1, '日期格式错误,请按2010-01-18格式输入')

        # u = s.query(sys_user).filter(sys_user.username == user.username).first()
        org = get_user_path(user.username)
        path = org.get('path', '')
        if '外销' not in str(path or ''):
            return json_result(-1, '只有外销岗位才能执行此操作')

        d = None
        if rid != '':
            d = s.query(cymx).filter(cymx.rid == rid).first()
        if not d:
            return json_result(-1, '没有找到对应的出运明细记录')

        xgqd = str(d.xgqd or '')
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        xg_item = f"{user.username}单据确认({hxdh}){now_str}"
        d.hxdh = hxdh_date
        d.xgqd = f"{xgqd}{xg_item}" if xgqd != '' else xg_item
        d.modi_uid = user.rid
        d.mtime = datetime.now()
        fphm = str(d.fphm or '')
        if fphm != '':
            s.query(fdsq1).filter(fdsq1.fphm == fphm).update(
                {fdsq1.hxdh: hxdh_date}, synchronize_session=False)
            s.query(fdsq1sheet).filter(fdsq1sheet.fphm == fphm).update(
                {fdsq1sheet.hxdh: hxdh_date}, synchronize_session=False)

        s.add(d)
        s.commit()

        return json_result(1, '操作成功', {'hxdh': hxdh, 'fphm': fphm, 'xgqd': d.xgqd})
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 出运明细 - 出运单号.change
# 说明：
# 1. 本接口仅负责“联动计算/建议值返回”，不直接修改主表记录。
# 2. 前端根据返回 data 决定是否回填字段。
# 3. 查询风格保持与项目既有 run_sql 用法一致，避免影响历史行为。
@any_route('/api/saier/shipment/cydh/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_cydh_change(request):
	"""
	用途：承接前端“出运明细.出运单号.change”联动请求。
	入参：
	- cydh: 出运单号
	- fphm: 发票号码
	- wyzd: 唯一字段（用于判断是否需要重算）

	返回 data 主要字段：
	- has_error / error_msg: 基础校验结果
	- hxfp: 核算发票建议值
	- duplicate_ysfp: 发票号是否被其它记录占用为核算发票
	- zgfy: 装柜费用
	- cpd: 出票判断
	- fporder: 发票order
	- new_wyzd: 新唯一字段
	- chyjh: 联动主数据（用于前端批量回填）
	"""
	s = Session()
	user = request.current_user
	j = await request.json()
	try:
		# 前端主输入：发票号、出运单号、唯一字段
		# 统一做 strip，兼容用户输入前后空格
		fphm = str(j.get('fphm', '') or '').strip()
		cydh = str(j.get('cydh', '') or '').strip()
		wyzd = str(j.get('wyzd', '') or '').strip()

		# 统一返回结构：前端可按字段存在与否决定是否回填
		data = {
			'has_error': False,
			'error_msg': '',
			'duplicate_ysfp': False,
			'hxfp': '',
			'zgfy': 0,
			'cpd': '否',
			'fporder': '',
			'new_wyzd': '',
			'chyjh': {}
		}

		if cydh != '':
			# 1) 出运单号末位字符校验（A-J 禁止）
			# 原 Pascal 规则：末位是 A~J 时直接提示并中断后续联动
			last_ch = cydh[-1:].upper()
			if last_ch in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
				data['has_error'] = True
				data['error_msg'] = '出运单号最后位不能为字符请检查后输入！'
				return json_result(1, '操作成功', data)

		if fphm != '':
			# 2) 核算发票号校对表取值 + ysfp 冲突检查
			# hxfp 来源：cyzglsheet(zm='核算发票号校对', xm=发票号)
			d = run_sql(
			    f"select bz from cyzglsheet where (xm='{fphm}') and (zm='核算发票号校对') limit 1")
			if len(d) > 0:
				data['hxfp'] = str(d[0].get('bz', '') or '')

			# duplicate_ysfp：当前发票号已被其它出运明细作为“核算发票(ysfp)”占用
			ys = s.query(cymx.rid).filter(cymx.ysfp == fphm, cymx.fphm != fphm).first()
			if ys:
				data['duplicate_ysfp'] = True

		if cydh != '':
			# 3) 读取装柜费用
			# zgfy 来源：delivery.fphm=出运单号
			d = run_sql(f"select zgfy from delivery where (fphm='{cydh}') limit 1")
			if len(d) > 0:
				try:
					data['zgfy'] = float(d[0].get('zgfy', 0) or 0)
				except:
					data['zgfy'] = 0

		# 4) 出票判断：发票号 == 出运单号 -> 是，否则 否
		data['cpd'] = '是' if (fphm != '' and fphm == cydh) else '否'

		lookup_no = fphm if fphm != '' else cydh
		if lookup_no != '':
			# 4) 取 chyjh 主记录，前端据此回填大量字段
			# 规则：优先用发票号查询；若发票号为空，回退用出运单号
			d = run_sql(f"select * from chyjh where (wxfp='{lookup_no}') limit 1")
			if len(d) > 0:
				data['chyjh'] = d[0]

		if fphm != '':
			# 5) 发票order：取发票号中第一个 '-' 之后的字符串
			idx = fphm.find('-')
			if idx >= 0 and idx < len(fphm) - 1:
				data['fporder'] = fphm[idx + 1:]

		if wyzd != '':
			# 5) 生成新的唯一字段（保留旧系统规则）
			# 规则：基准单号(发票号优先) + 用户名 + 当前时间
			base_no = fphm if fphm != '' else cydh
			if base_no != '':
				data['new_wyzd'] = f"{base_no}{user.username}{time.strftime('%Y-%m-%d %H:%M:%S')}"

		return json_result(1, '操作成功', data)
	except:
		# 保持与项目统一的异常日志方式
		logger.error(trace_error())
		return json_result(-1, trace_error())
	finally:
		# 显式关闭 Session
		s.close()


def _parse_date(v):
	"""将字符串/日期对象转换为 datetime，转换失败返回 None。"""
	if not v:
		return None
	if isinstance(v, datetime):
		return v
	s = str(v).strip()
	if s == '':
		return None
	for fmt in ('%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d', '%Y/%m/%d %H:%M:%S'):
		try:
			return datetime.strptime(s[:19], fmt)
		except Exception:
			continue
	return None


def _date_to_str(v):
	"""把 datetime/date 格式化为 yyyy-mm-dd。"""
	dt = _parse_date(v)
	if not dt:
		return ''
	return dt.strftime('%Y-%m-%d')


def _safe_float(v, default=0.0):
	"""安全转 float，避免 None/空值导致异常。"""
	try:
		if v is None or v == '':
			return float(default)
		return float(v)
	except Exception:
		return float(default)


def _safe_int(v, default=0):
	"""安全转 int。"""
	try:
		if v is None or v == '':
			return int(default)
		return int(v)
	except Exception:
		return int(default)


def _set_attrs(obj, data):
	"""仅在字段存在时赋值，避免模型字段不一致导致报错。"""
	for k, v in data.items():
		if hasattr(obj, k):
			setattr(obj, k, v)


def _invoice_candidates(chydh, fphm):
	"""生成出运单号对应的发票候选列表：主号 + A~S 后缀。"""
	chydh = (chydh or '').strip()
	fphm = (fphm or '').strip()
	if chydh == '':
		return [fphm] if fphm else []
	suffix = [f'{chydh}{letter}' for letter in string.ascii_uppercase[:19]]
	res = [chydh] + suffix
	if fphm and fphm not in res:
		res.append(fphm)
	return res


@any_route('/api/saier/shipment/sjcy/change', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_sjcy_change(request):
	"""
	出运明细“实际出运.change”迁移接口（Pascal -> WhaleCloud）。
	核心要求：insert/update 全部使用 ORM 实现，不直接执行 update/insert SQL。
	"""
	s = Session()
	try:
		j = await request.json()

		# ------------------------
		# 1. 读取并校验入参
		# ------------------------
		chydh = (j.get('chydh', '') or '').strip()
		ychydh = (j.get('ychydh', '') or '').strip()
		fphm = (j.get('fphm', '') or '').strip()
		wxfp = (j.get('wxfp', '') or '').strip()
		kh_id = (j.get('kh_id', '') or '').strip()
		khmc = (j.get('khmc', '') or '').strip()
		ywry = (j.get('ywry', '') or '').strip()
		sjcy1 = (j.get('sjcy1', '') or '').strip()
		yjdg = (j.get('yjdg', '') or '').strip()
		ydqx = _safe_int(j.get('ydqx', 0), 0)

		if sjcy1 == '' or fphm == '' or sjcy1 == None:
			return json_result(0, '实际出运或发票号码为空，跳过处理')

		# 核算发票锁定校验：锁定则禁止执行后续更新
		locked = s.query(fpgl.sid).filter(
			fpgl.wxfp == wxfp,
			(fpgl.sfjd == '是') | (fpgl.cysd == '锁定')
		).first()
		if locked:
			return json_result(0, '核算发票已结单或锁定，禁止更新')

		if ychydh == '':
			ychydh = chydh

		actual_dt = _parse_date(sjcy1)
		actual_year = sjcy1[:4] if len(sjcy1) >= 4 else ''
		actual_year_month = sjcy1[:7] if len(sjcy1) >= 7 else ''

		# ------------------------
		# 2. 同步 kaiptz（ORM update）
		# ------------------------
		if fphm:
			s.query(kaiptz).filter(kaiptz.fphm == fphm).update(
				{kaiptz.sjcy1: sjcy1}, synchronize_session=False
			)

		# ------------------------
		# 3. 老客新出年度逻辑（kh.lkxc）
		# ------------------------
		sn = 0
		zx_row = s.query(zx).filter(zx.ly == '老客新出').first()
		if zx_row and getattr(zx_row, 'mc', None):
			sn = _safe_int(zx_row.mc, 0)

		xksb = '1'
		first_ship_row = s.query(cymx).filter(
			cymx.cysb == '是',
			func.ifnull(cymx.sjcy1, '') != '',
			cymx.kh_id == kh_id
		).order_by(cymx.sjcy1).first()

		if first_ship_row and getattr(first_ship_row, 'sjcy1', ''):
			first_year = str(first_ship_row.sjcy1)[:4]
			xksb = '1' if first_year == actual_year else '2'

		if actual_year and actual_dt:
			sn1 = f'{_safe_int(actual_year, 0) - sn}-01-01'
			sn2 = f'{actual_year}-01-01'
			year_ship = s.query(cymx).filter(
				cymx.cysb == '是',
				func.ifnull(cymx.sjcy1, '') != '',
				cymx.sjcy1 > sn1,
				cymx.sjcy1 < sn2,
				cymx.kh_id == kh_id
			).first()
			if (not year_ship) and xksb == '2':
				s.query(kh).filter(kh.kh_id == kh_id).update(
					{kh.lkxc: actual_year}, synchronize_session=False
				)

		# ------------------------
		# 4. 信保费率计算
		# ------------------------
		xbdm1 = ''
		if 'BEST PRICE LLC' in khmc.upper():
			xbdm1 = 'RUS/190983'

		fkkh_row = s.query(fkkh).filter(fkkh.khbh == kh_id).first()
		if fkkh_row and getattr(fkkh_row, 'xbdm', ''):
			xbdm1 = fkkh_row.xbdm

		insurance_rate = 0.0
		if xbdm1 and actual_dt:
			fkxbfl_row = s.query(fkxbfl).filter(
				fkxbfl.qsrq <= sjcy1,
				fkxbfl.jsrq >= sjcy1,
				fkxbfl.xbdm == xbdm1
			).first()
			if fkxbfl_row and getattr(fkxbfl_row, 'xbfl', None) is not None:
				insurance_rate = _safe_float(fkxbfl_row.xbfl, 0.0)

		# ------------------------
		# 5. 计算目标发票集合
		# ------------------------
		invoice_scope = _invoice_candidates(chydh, fphm)
		if not invoice_scope:
			return json_result(0, '出运单号和发票号码都为空，无法处理')

		# 如果存在 A~S 拆分发票，则仅处理拆分发票；否则处理当前发票
		split_invoices = []
		if chydh:
			split_candidates = [f'{chydh}{letter}' for letter in string.ascii_uppercase[:19]]
			split_rows = s.query(cymx.fphm).filter(
				cymx.chydh == chydh,
				cymx.fphm.in_(split_candidates)
			).all()
			split_invoices = [r[0] for r in split_rows if r and r[0]]

		process_invoices = sorted(set(split_invoices)) if split_invoices else [fphm]

		# 业务人员组织路径
		user_path = ''
		user_rid = ''
		if ywry:
			org = get_user_path(ywry)
			user_path = (org or {}).get('path', '')[:100]
			user_rid = (org or {}).get('rid', '')

		# ------------------------
		# 6. shqk upsert（ORM insert/update）
		# ------------------------
		sample_cymx = None
		for invoice in process_invoices:
			if not invoice:
				continue

			mx = s.query(cymx).filter(cymx.chydh == chydh, cymx.fphm == invoice).first()
			if not mx:
				continue
			if sample_cymx is None:
				sample_cymx = mx

			djjez = _safe_float(
				sum([_safe_float(x[0], 0.0) for x in s.query(krshsheet.sydje2).filter(krshsheet.fphm == invoice).all()]),
				0.0
			)
			ysr = _safe_float(
				sum([_safe_float(x[0], 0.0) for x in s.query(krshsheet.sydje).filter(
					krshsheet.fphm == invoice,
					(krshsheet.hbdm == 'RMB') | (krshsheet.hbdm == 'RMB￥')
				).all()]),
				0.0
			)
			ysm = _safe_float(
				sum([_safe_float(x[0], 0.0) for x in s.query(krshsheet.sydje).filter(
					krshsheet.fphm == invoice,
					krshsheet.hbdm != 'RMB',
					krshsheet.hbdm != 'RMB￥'
				).all()]),
				0.0
			)

			# Pascal 逻辑：CSD- 发票不处理 shqk
			if 'CSD-' in fphm.upper():
				continue

			record = s.query(shqk).filter(shqk.fphm == invoice).first()
			is_insert = False
			if not record:
				record = shqk()
				record.rid = get_uuid()
				is_insert = True

			myjje = _safe_float(getattr(mx, 'myjje', 0), 0.0)
			htje = _safe_float(getattr(mx, 'htje', 0), 0.0)
			htjer = _safe_float(getattr(mx, 'htjer', 0), 0.0)
			htjem = _safe_float(getattr(mx, 'htjem', 0), 0.0)
			ysdj = _safe_float(getattr(mx, 'ysdj', 0), 0.0)

			htjez = round(htje - myjje, 3)
			wsjez = round(htje - myjje - djjez, 3)
			wsje_rmb = round(htjer - myjje - ysr, 3)
			wsje_foreign = round(htjem - myjje - ysm, 3)

			rmbkh = getattr(mx, 'RMBkh', '') or '否'
			if rmbkh == '是':
				calc_htje = round(htjer - myjje, 3)
				calc_htjem = htjem
				sfjq = '是' if (-1 < wsje_rmb < 200) else '否'
			else:
				calc_htje = htjer
				calc_htjem = round(htjem - myjje, 3)
				sfjq_num = round((htjem - myjje - ysm) * 1000)
				sfjq = '是' if (-1 < sfjq_num < 30) else '否'

			yssj = ''
			if actual_dt:
				yssj = _date_to_str(actual_dt + timedelta(days=ydqx))

			data_map = {
				'kh_id': getattr(mx, 'kh_id', ''),
				'khmc': getattr(mx, 'khmc', ''),
				'xybx': getattr(mx, 'xybx', ''),
				'webpd': '是',
				'htjez': htjez,
				'djjez': djjez,
				'wsjez': wsjez,
				'RMBkh': rmbkh,
				'fphm': invoice,
				'ysfp': getattr(mx, 'ysfp', ''),
				'htje': calc_htje,
				'djje': ysr,
				'wsje': wsje_rmb,
				'htjem': calc_htjem,
				'djjem': ysm,
				'wsjem': wsje_foreign,
				'jgtk': getattr(mx, 'jgtk', ''),
				'jhfs': getattr(mx, 'jhfs', ''),
				'ydqx': ydqx,
				'sjcy1': sjcy1,
				'yssj': yssj,
				'ywry': getattr(mx, 'ywry', ''),
				'ywpath': user_path,
				'sfjq': sfjq,
				'sfdj': '是' if ysdj > 0 else '否',
				'djm': 0,
				'djr': 0,
				'ysdj': ysdj,
			}

			# 仅新增时初始化的字段
			if is_insert:
				data_map.update({
					'shrq': None,
					'yqts': 0,
					'rid': get_uuid(),
					'uid': user_rid,
					'ctime':time.strftime('%Y-%m-%d %H:%M:%S'),
				})

			_set_attrs(record, data_map)
			if is_insert:
				s.add(record)

		# ------------------------
		# 7. 计算并回写收汇期限
		# ------------------------
		receivable_deadline = ''
		fkts = ydqx
		if fkts == 0 and sample_cymx is not None:
			fkts = _safe_int(getattr(sample_cymx, 'fkys', 0), 0)

		if actual_dt and fkts > 0:
			receivable_deadline = _date_to_str(actual_dt + timedelta(days=fkts))
		elif yjdg:
			receivable_deadline = _date_to_str(yjdg)

		if receivable_deadline == '' and actual_dt:
			receivable_deadline = _date_to_str(actual_dt + timedelta(days=30))

		khmc_upper = khmc.upper()
		special_customers = ['BEST PRICE LLC', 'FIX PRICE GENERAL TRADING LLC', 'SIA FP LV']
		if actual_dt and any(name in khmc_upper for name in special_customers):
			receivable_deadline = _date_to_str(actual_dt + timedelta(days=120))

		# ------------------------
		# 8. 批量回写 cymx / cymxsheet（ORM update）
		# ------------------------
		cymx_rows = s.query(cymx).filter(cymx.fphm.in_(invoice_scope)).all()
		for row in cymx_rows:
			_set_attrs(row, {
				'webpd': '是',
				'sjny': actual_year_month,
				'shqx': receivable_deadline,
				'xbfl': insurance_rate,
				'sjcy1': sjcy1,
				'ychydh': ychydh,
			})

		cymxsheet_rows = s.query(cymxsheet).filter(cymxsheet.fphm.in_(invoice_scope)).all()
		for row in cymxsheet_rows:
			_set_attrs(row, {'sjcy': sjcy1})

		# ------------------------
		# 9. 超期提醒（与 Pascal 一致：超过30天提醒）
		# ------------------------
		warning_msg = ''
		now_dt = datetime.now()
		if actual_dt and (now_dt - actual_dt).days > 30:
			warning_msg = '请注意时间超30天!'

		s.commit()
		return json_result(1, '操作成功', {
			'original_cydh': ychydh,
			'actual_year': actual_year,
			'actual_year_month': actual_year_month,
			'insurance_rate': insurance_rate,
			'receivable_deadline': receivable_deadline,
			'warning_msg': warning_msg,
			'processed_invoices': process_invoices,
		})
	except Exception:
		s.rollback()
		logger.error(trace_error())
		return json_result(-1, trace_error())
	finally:
		s.close()



@any_route('/api/saier/shipment/date/check', methods=['POST'])
@require_token
async def api_sairer_shpipment_date_check(request):
    """
    实际出运对照按钮
    对应原Pascal: 实际出运对照
    """
    user = request.current_user
    # 1. 检查用户是否是财务岗位 (对应 tmpcom2 查询)
    s = Session()
    try:
        # 2. 获取上传的Excel文件
        j = await request.form()
        file = form_value(j,'file',None)
        if is_none(file):
            return json_result(ERR_PARAM_NOT_ENOUGH)
        # 保存临时文件
        temp_dir = config.tmp_path
        temp_file = os.path.join(temp_dir, f"{uuid.uuid4()}_{file.filename}")
        logger.error(f"Saving uploaded file to: {temp_file}")
        data = await file.read()
        write_file(temp_file, data, 'wb', encoding=None)
        
        # 3. 读取Excel文件
        wb = load_workbook(temp_file)
        ws = wb.worksheets[0]  # 读取第一个工作表
        
        # 4. 处理每一行数据
        duplicates = []  # 存储重复记录信息
        row_idx = 2  # 从第2行开始（第1行是表头）
        while True:
            # 读取Excel单元格数据
            wxfp = str(ws.cell(row=row_idx, column=3).value or '').strip()  # C列 - 发票号码
            # bank = str(ws.cell(row=row_idx, column=2).value or '')  # B列 - 银行
            # zh = str(ws.cell(row=row_idx, column=3).value or '')  # C列 - 账号
            # shui = str(ws.cell(row=row_idx, column=4).value or '')  # D列 - 税号
            
            # 如果公司名称为空，结束循环
            if not wxfp:
                break
            
            d = s.query(cymx.rid).filter(cymx.fphm==wxfp, func.ifnull(cymx.sjcy1,'') != '').first()
            if not d:                
                # 记录重复信息
                duplicates.append(f'请注意发票号码: {wxfp} 没填实际出运日期')
            
            row_idx += 1
        
        # 如果有重复记录，生成文本文件
        result_data = {
            'has_duplicates': len(duplicates) > 0,
        }
        
        if duplicates:
            # 生成重复记录文件
            dup_filename = f"无实际出运日期_{datetime.now().strftime('%Y-%m-%d')}.txt"
            dup_filepath = os.path.join(temp_dir, dup_filename)
            val ='\n'.join(duplicates)
            logger.error(f"Writing duplicates to: {dup_filepath}")
            write_file(dup_filepath, val, 'w')

            # 这里应该上传文件到可下载的位置，返回URL
            # 简化处理：返回文件内容或保存到特定目录
            result_data['duplicate_file_name'] = dup_filename
        
        # 清理临时文件
        # os.remove(temp_file)
        
        return json_result(0, data=result_data)
        
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()
        wb.close()

@any_route('/api/saier/shipment/zdry/change', methods=['POST'])
@require_token
async def api_sairer_shpipment_zdry_change(request):
    s = Session()
    j = await request.json()
    user = request.current_user
    s = Session()
    try:
        # 获取请求参数
        zdry = j.get('zdry')
        fphm = j.get('fphm')
        # 3. 检查制单人员是否为当前用户，如果是则跳过
        if zdry == user.username:
            s.close()
            return json_result(0, '制单人员未变更')

        # 4. 查询制单人员是否为单证人员 (sys_user position like "%单证%")
        sb  = 0
        sb1 = 0
        org = get_user_path(zdry)
        position = (org or {}).get('position', '') if org else ''
        if '单证' in position:
            sb = 1
            if '经理' in position:
                sb1 = 1

        if sb == 0:
            s.close()
            return json_result(-1, '此人非单证人员,请重新选择')
        org = get_user_path(user.username)
        position = (org or {}).get('position', '') if org else ''
        if '单证' not in position :
            if sb1 != 1:
                s.close()
                return json_result(-1, '此人无权分配单证人员，请重新选择')
        else:
            d = s.query(cymx.rid).filter(cymx.fphm == fphm).first()
            if  d:                    
                zdry1 = d.uid
                c = s.query(ywrybiao.yhm).filter(ywrybiao.yhm == user.username, ywrybiao.zw.like('%单证%'),
                        sys_user.rid == zdry1
                    ).outerjoin(sys_user,sys_user.yhm==ywrybiao.yhm).first()
                if not c:
                    return json_result(-1, '请先更改发票号,保存后在继续')
            
            d = s.query(ywrybiao.yhm).filter(ywrybiao.yhm == user.username, ywrybiao.zw.like('%单证%')).first()
            if not d:
                return json_result(-1, '请先更改发票号,保存后在继续')
                
        return json_result(1, '操作成功', {'sb1': sb1})

    except Exception as e:
        s.rollback()
        s.close()
        logger.error(f"制单人员变更失败: {trace_error()}")
        return json_result(-1, f'操作失败: {str(e)}')
    finally:
        s.close()



@any_route('/api/saier/shipment/fphm/change', methods=['POST'])
@require_token
async def view_saier_shipment_fphm_change(request):
    """
    发票号码 change 联动接口

    Request JSON:
        fphm   (str)  发票号码（前端已trim）
        chydh  (str)  出运单号
        hzf    (str)  末位字母（前端计算，非数字时非空）
        khmc   (str)  客户名称
        sid    (int)  记录主键 ID，对应 Pascal self.getnumber
        zdry   (str)  当前制单人员字段值
        hwzt   (str)  货物状态
        qzgx   (str)  强制更新标志
        wyzd1  (str)  唯一字段当前值
    Response data（前端按key按需应用）:
        zjz_no      bool   True时前端将"自动净重"设为"否"
        htjc        str    合同简写
        hxfp        str    核算发票
        wfgs        str    我方公司
        chydh_new   str    出运单号新值（原为空时）
        cpjd        str    出票判断 "是"/"否"
        djyc        float  定金预测
        enable_ysj  bool   True时解锁"已收定金"
        ywry        str    业务人员
        khht        str    客户合同
        jdrq        str    接单日期
        jgtk        str    价格条款
        mdck        str    目的仓库
        fporder     str    发票order
        new_wyzd    str    唯一字段新值
        new_zdry    str    制单人员新值（需要更新时才有）
        should_save bool   True时前端触发 checksave(false, true)
    """
    s = Session()
    try:
        j = await request.json()
        user = request.current_user
        current_user_name = user.username   # 登录用户名
        # ── 接收并清洗前端参数 ──────────────────────────────────────────────
        fphm  = (j.get('fphm')  or '').strip()
        chydh = (j.get('chydh') or '').strip()
        hzf   = (j.get('hzf')   or '').strip()   # 末位字母（前端已计算）
        khmc  = (j.get('khmc')  or '').strip()
        # job   = int(j.get('job') or 0)
        zdry  = (j.get('zdry')  or '').strip()
        hwzt  = (j.get('hwzt')  or '').strip()
        qzgx  = (j.get('qzgx')  or '').strip()
        wyzd1 = (j.get('wyzd1') or '').strip()
        data = {}   # 最终返回给前端的字段更新字典
        # ── A. 自动净重规则 ─────────────────────────────────────────────────
        # Pascal: select ... from cyzglsheet
        #         where (fphm like xm+'%') and (zm='出运明细非自动净重')
        zjz_rows = run_sql(
            f"select xm from cyzglsheet "
            f"where ('{fphm}' like concat(xm, '%')) "
            f"and (zm='出运明细非自动净重') limit 1"
        )
        if zjz_rows:
            data['zjz_no'] = True   # 前端将"自动净重"置"否"
        # ── B. 合同简写 ──────────────────────────────────────────────────────
        # Pascal 逻辑：
        #   1) 查 cyzglsheet 特殊合同简写规则 → jxsb
        #   2) fphm 含 -N2- 或 -DMD- → jxsb1
        #   3) 有规则时：找第一个"UV"位置截取
        #      - jxsb=1 且 jxsb1≠1 → 从 UV 前两位截取
        #      - 否则 → 从 UV 当前位置截取
        #   4) 无规则 → htjc = fphm
        jxsb_rows = run_sql(
            f"select sz from cyzglsheet "
            f"where ('{fphm}' like concat(xm, '%')) "
            f"and (zm='特殊合同简写') limit 1"
        )
        jxsb = '1' if jxsb_rows else ''
        fphm_upper = fphm.upper()
        jxsb1 = '1' if ('-N2-' in fphm_upper or '-DMD-' in fphm_upper) else ''
        htjc = ''
        if jxsb == '1' or jxsb1 == '1':
            uv_pos = fphm_upper.find('UV')
            if uv_pos >= 0:
                if jxsb == '1' and jxsb1 != '1':
                    # Pascal: copy(fphm, zs2-2, zs)  从 UV 前两位开始
                    htjc = fphm[max(0, uv_pos - 2):]
                else:
                    # Pascal: copy(fphm, zs2, zs)  从 UV 当前位置开始
                    htjc = fphm[uv_pos:]
            data['htjc'] = htjc
        else:
            data['htjc'] = fphm   # 无规则时合同简写等于发票号码
        # ── C. 核算发票 ──────────────────────────────────────────────────────
        # Pascal 逻辑：
        #   1) 默认 hxfp = fphm
        #   2) 查 cyzglsheet 核算发票号校对（等值匹配）→ bz 非空则覆盖
        #   3) 若 hzf 非空：查 cymx ysfp，拼接 ysfp+hzf
        hxfp = fphm
        hxfp_rows = run_sql(
            f"select bz from cyzglsheet "
            f"where (xm='{fphm}') and (zm='核算发票号校对') limit 1"
        )
        if hxfp_rows and hxfp_rows[0].get('bz'):
            hxfp = hxfp_rows[0]['bz']
        # Pascal SQL: select ysfp,fphm,chydh from cymx
        #             where fphm=:chydh and chydh=:chydh and fphm<>:ysfp
        # 注：Pascal 中 fphm 参数和 ysfp 参数均被赋值为 chydh，
        #     且条件 fphm<>ysfp 实际等价于 fphm<>'' 此处保持等价语义
        if hzf and chydh:
            cymx_hzf = run_sql(
                f"select ysfp from cymx "
                f"where fphm='{chydh}' and chydh='{chydh}' and fphm<>'' limit 1"
            )
            if cymx_hzf and cymx_hzf[0].get('ysfp'):
                hxfp = cymx_hzf[0]['ysfp'] + hzf
        data['hxfp'] = hxfp
        # ── D. 我方公司 ──────────────────────────────────────────────────────
        # Pascal 优先级：CSD- 前缀 > cyzglsheet 识别 > kh.Vendorgs（后者覆盖前者）
        gsmc = None
        if 'CSD-' in fphm:
            gsmc = '宁波可思达进出口有限公司'
        else:
            # D1. cyzglsheet 外销合同识别公司（反向 LIKE：fphm LIKE '%xm%'）
            gsmc_rows = run_sql(
                f"select bz from cyzglsheet "
                f"where ('{fphm}' like concat('%', xm, '%')) "
                f"and (zm='外销合同识别公司') limit 1"
            )
            if gsmc_rows and gsmc_rows[0].get('bz'):
                gsmc = gsmc_rows[0]['bz']
            # D2. kh.Vendorgs（按客户名称匹配，会覆盖 D1 结果，与 Pascal 一致）
            if khmc:
                kh_rows = run_sql(
                    f"select Vendorgs from kh "
                    f"where (company_name='{khmc}') "
                    f"and (Vendorgs<>'') and (Vendorgs is not null) limit 1"
                )
                if kh_rows and kh_rows[0].get('Vendorgs'):
                    gsmc = kh_rows[0]['Vendorgs']

        if gsmc is not None:
            data['wfgs'] = gsmc
        # ── E. 强制更新≠'是' 时的额外处理 ──────────────────────────────────
        # 对应 Pascal: if 强制更新<>'是' then begin ... end
        if qzgx != '是':
            # E1. 出运单号：为空时填入发票号码（并解锁字段）
            # Pascal: if 出运单号='' then enabled=true setfielddataasstring(fphm)
            if not chydh:
                data['chydh_new'] = fphm
            effective_chydh = chydh if chydh else fphm   # 用于后续计算
            # E2. 出票判断（是/否）
            # Pascal: 若 fphm == chydh 或 fphm == chydh+[A-S] → '是'
            sb1 = ''
            for suffix in 'ABCDEFGHIJKLMNOPQRS':
                if fphm == effective_chydh + suffix:
                    sb1 = '1'
                    break
            data['cpjd'] = '是' if (fphm == effective_chydh or sb1 == '1') else '否'
            # E3. 定金预测 —— ORM SUM
            # Pascal: select sum(sydje2) as sydje1 from krshsheet where fphm=:fphm
            djyc_val = s.query(func.sum(krshsheet.sydje2)).filter(
                krshsheet.fphm == fphm
            ).scalar()
            data['djyc'] = float(djyc_val or 0)
            data['enable_ysj'] = True   # 解锁"已收定金"字段
            # E4. wxht 联动
            # Pascal: select mdka,dforder_id,jgtk,ywry,jdrq from wxht
            #         where order_id=:chydh or order_id=:fphm
            wxht_rows = run_sql(
                f"select mdka,dforder_id,jgtk,ywry,jdrq from wxht "
                f"where (order_id='{effective_chydh}') or (order_id='{fphm}') limit 1"
            )
            if wxht_rows:
                w = wxht_rows[0]
                data['ywry'] = w.get('ywry') or ''
                data['khht'] = w.get('dforder_id') or ''
                data['jdrq'] = str(w.get('jdrq') or '')
                data['jgtk'] = w.get('jgtk') or ''
                data['mdck'] = w.get('mdka') or ''
            # E5. 发票order：取第一个"-"之后的部分
            # Pascal: 遍历字符找第一个"-"，然后截取
            fporder = ''
            dash_idx = fphm.find('-')
            if dash_idx >= 0:
                fporder = fphm[dash_idx + 1:]
            data['fporder'] = fporder
            # E8. 唯一字段 / 制单人员 / checksave
            # Pascal: if wyzd1<>'' then begin
            #           唯一字段 = fphm + user + datetime
            #           if 制单人员<>user and 货物状态<>'已出运' then 制单人员=user
            #           if 强制更新<>'是1' then checksave(false, true)
            #         end
            if zdry != current_user_name and hwzt != '已出运':
                data['new_zdry'] = current_user_name

        return json_result(1, data=data)

    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'操作失败: {str(e)}')
    finally:
        s.close()

@any_route('/api/saier/shipment/load/check', methods=['POST'])
@require_token
async def view_saier_shipment_load_check(request):
    """
    出运明细记录加载接口
    Request JSON:
        fphm             (str)  发票号码
        sid              (int)  记录主键 ID（对应 Pascal self.getnumber）
        actual_shipment  (str)  实际出运字段当前值
        item_count       (int)  前端 ITEM数量 字段值

    Response data:
        is_zdz              bool  登录用户是否为单证人员
        is_zdz_mgr          bool  登录用户是否为单证经理
        is_dz_check_user    bool  是否"单证查验"用户（需隐藏大量字段）
        multi_guihao        bool  cymxsheet 中 fphm 数>1（两票不同柜号警告）
        item_count_mismatch bool  cymxsheet 行数与 ITEM数量 字段不一致
        shipment_apply_locked bool  实际出运有"申请中/已批复"记录
        is_readonly         bool  只读状态为'是'
        has_tongji_date     bool  统计日期字段非空（需禁用出运日期）
        need_save_warning   bool  产品引入刚被重置（需提示保存）
    """
    try:
        j = await request.json()
        user = request.current_user
        current_user_name = user.username   # 登录用户名
        job            = int(j.get('job') or 0)  
        fphm            = (j.get('fphm') or '').strip()
        rid             = (j.get('rid') or '')
        actual_shipment = (j.get('actual_shipment') or '').strip()
        item_count_fe   = int(j.get('item_count') or 0)   # 前端 ITEM数量 字段值

        data = {}

        # ── 1. 查询当前用户 position ────────────────────────────────────────
        # 对应 Pascal:
        #   select position from sys_users where name=:name
        #   select * from sys_users where name=:name and position like '%单证经理%'
        #   select * from sys_users where name=:name and position like '%单证%'
        org = get_user_path(current_user_name)
        # user_rows = run_sql(
        #     f"select position from sys_user "
        #     f"where (username='{current_user_name}') limit 1"
        # )
        position = org.get('position')
        # if user_rows:
        #     position = (user_rows[0].get('position') or '').strip()

        is_zdz     = '单证' in position                         # 是否单证人员
        is_zdz_mgr = '单证经理' in position                     # 是否单证经理（含单证经理时两者均为 True）

        data['is_zdz']     = is_zdz
        data['is_zdz_mgr'] = is_zdz_mgr

        # ── 2. 单证查验用户判断 ─────────────────────────────────────────────
        # 对应 Pascal:
        #   select * from cyzglsheet where xm=:name and zm='单证查验'
        dz_check_rows = run_sql(
            f"select xm from cyzglsheet "
            f"where (xm='{current_user_name}') and (zm='单证查验') limit 1"
        )
        data['is_dz_check_user'] = bool(dz_check_rows)

        # ── 3. cymxsheet 柜号数量校验（仅发票号码非空且 sid>0 时）───────────
        # 对应 Pascal:
        #   select fphm from cymxsheet where father=:father group by fphm
        multi_guihao        = False
        item_count_mismatch = False
        if fphm and job == 1:
            # 3a. 统计不同 fphm 数
            guihao_rows = run_sql(
                f"select count(distinct fphm) as cnt from cymxsheet "
                f"where (pid='{rid}')"
            )
            guihao_cnt = int((guihao_rows[0].get('cnt') or 0)) if guihao_rows else 0
            if guihao_cnt > 1:
                multi_guihao = True

            # 3b. ITEM数量与实际 cymxsheet 行数对比
            # 对应 Pascal: i1 = recordset.Table_Count('产品资料')
            item_real_rows = run_sql(
                f"select count(*) as cnt from cymxsheet "
                f"where (pid='{rid}')"
            )
            item_real = int((item_real_rows[0].get('cnt') or 0)) if item_real_rows else 0
            if item_real != item_count_fe and item_real > 0 and item_count_fe > 0:
                item_count_mismatch = True

        data['multi_guihao']        = multi_guihao
        data['item_count_mismatch'] = item_count_mismatch

        # ── 4. 实际出运申请状态查询 ─────────────────────────────────────────
        # 对应 Pascal:
        #   select state from shipmentapply
        #   where invoiceNo=:invoiceNo and (state='申请中' or state='已批复')
        shipment_apply_locked = False
        if actual_shipment and fphm:
            apply_rows = run_sql(
                f"select state from shipmentapply "
                f"where (invoiceNo='{fphm}') "
                f"and ((state='申请中') or (state='已批复')) limit 1"
            )
            if apply_rows:
                shipment_apply_locked = True
        data['shipment_apply_locked'] = shipment_apply_locked

        # ── 5. 只读状态与统计日期
        # 这两个字段值存在记录本身，前端已可读取；
        # 此处返回 False 作为默认值，实际由前端直接判断记录字段。
        # 若需要后端校验可在此扩展。
        data['is_readonly']      = False  # 由前端判断 '只读状态' 字段
        data['has_tongji_date']  = False  # 由前端判断 '统计日期' 字段
        data['need_save_warning'] = False  # 由前端 sb12 变量判断

        return json_result(0, data=data)

    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'操作失败: {str(e)}')


@any_route('/api/saier/shipment/before/save', methods=['POST'])
@require_token
async def api_sairer_shpipment_before_save(request):
    s = Session()
    j = await request.json()
    user = request.current_user
    s = Session()
    try:
        # 获取请求参数
        # zdry = j.get('zdry')
        module = j.get('module', '出运明细')
        rid = j.get('rid')
        fphm = j.get('fphm')
        cydh = j.get('cydh')
        tjry = j.get('tjry')
        dywry = j.get('dywry')
        khmc12 = j.get('khmc12')
        csz = 0
        btsb = 0
        dd = 0
        sb = 0
        sdsb = 0
        fksb = '否'
        cysb = '否' 
        sfdz = 0
        dzjy = 0 
        jh_data = {}
        t = get_module(module)
        o = get_model_by_table_name(t.table_name)
        d = s.query(chyjh).filter(chyjh.wxfp == fphm, chyjh.dzqr1 == '再审').first()
        if d:
            return json_result(-1, '此票在出运计划为再审状态,请先处理出运计划')
        
        c = run_sql(f"select bz from cyzglsheet where ('{khmc12}' like concat('%', xm, '%')) and (zm='出运组出运客人') limit 1")
        if len(c) > 0:
            csz = 0
            btsb = 1
        c = run_sql(f"select xybx,sz,zysx,qtshm,hjbh,fksm,sdsm,dzsm,jxmc3,jxje3,jjxmc3,jjxje3,myjje,zpyj,ayjje from chyjh where (wxfp='{fphm}') and (fphm='{cydh}') limit 1")
        if len(c) > 0:
            if csz == 0:
                csz = c[0].get('sz', 0) or 1
            jh_data = c[0]
        # if csz == 0:
        #     csz = 0
        c = s.query(o.rid).filter(o.fphm == fphm, o.rid!=rid).first()
        if c:
            return json_result(-1, '发票号码已存在,请重新输入')
        
        m = s.query(o.zdry,o.fphm).filter(o.rid == rid).first()
        if m :
            zdry = m.zdry
        else:
            zdry = user.username

        # yw_org = get_user_path(dywry)
        # ywpath = yw_org.get('path', '') if yw_org else ''
        zd_org = get_user_path(zdry)
        zdpath = zd_org.get('path', '') if zd_org else ''
        if '单证' in zdpath:
            dd = 1
            sb = 1
            sfdz = 1
        else:
            sfdz = 0
            u = run_sql(f"select yhm from ywrybiao where yhm='{zdry}' and zw like '%单证%' limit 1")
            if len(u) == 0:
                dzjy = 0
            else:
                dzjy = 1
        fphm_list = [str(cydh) + suffix for suffix in suffixes]
        if cydh == fphm:
            d = s.query(o).filter(o.fphm.in_(fphm_list)).first()
            if not d:
                fksb = '是'
                cysb = '是'
        else:
            for r in suffixes:
                if fphm == str(cydh) + str(r):
                    fksb = '是'
                    cysb = '是'
                    break
        if module == '出运明细': 
            s.query(o).filter(o.fphm.in_(fphm_list)).update({o.fksb: '是', o.cysb: '是'}, synchronize_session=False)
            s.query(o).filter(o.fphm.notin_(fphm_list),o.chydh==cydh).update({o.fksb: '否', o.cysb: '否'}, synchronize_session=False) 
            s.query(cymxsheet).filter(cymxsheet.fphm.in_(fphm_list)).update({cymxsheet.fksb: '是', cymxsheet.fpsb1: '是'}, synchronize_session=False)
            s.query(cymxsheet).filter(cymxsheet.fphm.notin_(fphm_list),cymxsheet.chydh==cydh).update({cymxsheet.fksb: '否', cymxsheet.fpsb1: '否'}, synchronize_session=False)
            if m: 
                old_fphm = m.fphm
                old_zdry = m.zdry
                # E7. shqk 同步发票号 —— ORM UPDATE
                # 条件：旧发票号≠当前发票号 且 当前制单人员与记录制单人员一致
                # Pascal: update shqk set fphm=:fphm where fphm=:fphm1
                if old_fphm and old_fphm != fphm and zdry == old_zdry:
                    s.query(shqk).filter(shqk.fphm == old_fphm).update(
                        {shqk.fphm: fphm}, synchronize_session=False
                    )

            if m and m.zdry != zdry:
                if zdry != user.username and zdry != '' and zdry != None:
                    xxnr = f"{module}发票号为:{fphm},请查看,日期:{time.strftime('%Y-%m-%d')}"
                    row = {
                        "xxly": module,
                        "bjdh": '',
                        "wxht": fphm,
                        "cght": '',
                        "yhdh": '',
                        'cymx': fphm,''
                        "xxnr": xxnr,
                        "jsr": str(zdry),
                        "sys_path": "",
                        "spsq": user.username
                    }
                    res = module_xxck_new([row], user, s)
                    if res.get('code', 1) != 1:
                        return {'code': -1, 'msg': res.get('msg')}
                if tjry != user.username and tjry != '' and tjry != None:
                    xxnr = f"{module}发票号为:{fphm},已分配给:{zdry},请查看,日期:{time.strftime('%Y-%m-%d')}"
                    row = {
                        "xxly": module,
                        "bjdh": '',
                        "wxht": fphm,
                        "cght": '',
                        "yhdh": '',
                        'cymx': fphm,''
                        "xxnr": xxnr,
                        "jsr": str(tjry),
                        "sys_path": "",
                        "spsq": user.username
                    }
                    res = module_xxck_new([row], user, s)
                    if res.get('code', 1) != 1:
                        return {'code': -1, 'msg': res.get('msg')}
        data = {
            'csz': csz,
            'btsb': btsb,
            'dd': dd,
            'sb': sb,
            'sfdz': sfdz,
            'dzjy': dzjy,
            'jh_data': jh_data,
            'fksb': fksb,
            'cysb': cysb,
            'sdsb': sdsb
        }
        s.commit()  
        return json_result(1, '操作成功', data)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'操作失败: {str(e)}')
    finally:
        s.close()

def shipment_dfrq_batch_update(rids, date, user):
    s = Session()
    try:
        flag = 0
        for rid in rids:
            d = s.query(cymx).filter(cymx.rid == rid).first()
            if not d:
                continue  # 记录不存在时跳过
            if d.zdry != user.username:
                continue  # 非制单人员时跳过
            d.dfrq = date
            d.hxdate1 = '是'
            d.modi_uid = user.id  # 更新制单人员为当前用户
            d.mtime = time.strftime('%Y-%m-%d %H:%M:%S')  # 更新时间
            s.add(d)
            if d.fphm != '' and d.fphm != None:
                s.query(fdsq1).filter(fdsq1.fphm == d.fphm).update(
                    {fdsq1.dfrq: date, fdsq1.mtime: time.strftime('%Y-%m-%d %H:%M:%S'), fdsq1.modi_uid: user.id}, synchronize_session=False
                )
                s.query(fdsq1sheet).filter(fdsq1sheet.fphm == d.fphm).update(
                    {fdsq1sheet.dfrq: date, fdsq1sheet.dzdf: '是', fdsq1sheet.mtime: time.strftime('%Y-%m-%d %H:%M:%S'), fdsq1sheet.modi_uid: user.id}, synchronize_session=False
                )
            flag = 1

        if flag == 0:
            return json_result(-1, '没有找到需要修改的记录')
        s.commit()  
        return json_result(1, '操作成功')
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'操作失败: {str(e)}')
    finally:
        s.close()

@any_route('/api/saier/shipment/dfrq/update', methods=['POST'])
@require_token
async def api_sairer_shpipment_dfrq_update(request):
    s = Session()
    j = await request.json()
    user = request.current_user
    s = Session()
    try:
        rid = j.get('rid')
        date = j.get('date')
        d = s.query(cymx).filter(cymx.rid == rid).first()
        if not d:
            return json_result(-1, '记录不存在')
        if d.zdry != user.username:
            return json_result(-1, '只有制单人员可以修改电放日期')
        d.dfrq = date
        d.hxdate1 = '是'
        s.add(d)
        if d.fphm != '' and d.fphm != None:
            s.query(fdsq1).filter(fdsq1.fphm == d.fphm).update(
                {fdsq1.dfrq: date}, synchronize_session=False
            )
            s.query(fdsq1sheet).filter(fdsq1sheet.fphm == d.fphm).update(
                {fdsq1sheet.dfrq: date, fdsq1sheet.dzdf: '是'}, synchronize_session=False
            )

        s.commit()  
        return json_result(1, '操作成功')
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'操作失败: {str(e)}')
    finally:
        s.close()


@any_route('/api/saier/shipment/sfdf/update', methods=['POST'])
@require_token
async def api_sairer_shpipment_sfdf_update(request):
    user = request.current_user
    # 1. 检查用户是否是财务岗位 (对应 tmpcom2 查询)
    s = Session()
    j = await request.form()
    try:
        # 2. 获取上传的Excel文件
        file = form_value(j,'file',None)
        if is_none(file):
            return json_result(ERR_PARAM_NOT_ENOUGH)
        # 保存临时文件
        temp_dir = config.tmp_path
        temp_file = os.path.join(temp_dir, f"{uuid.uuid4()}_{file.filename}")
        logger.error(f"Saving uploaded file to: {temp_file}")
        data = await file.read()
        write_file(temp_file, data, 'wb', encoding=None)
        
        # 3. 读取Excel文件
        wb = load_workbook(temp_file)
        ws = wb.worksheets[0]  # 读取第一个工作表
        # 4. 处理每一行数据
        duplicates = []  # 存储重复记录信息
        row_idx = 2  # 从第2行开始（第1行是表头）
        while True:
            # 读取Excel单元格数据
            fphm = str(ws.cell(row=row_idx, column=1).value or '').strip()  # C列 - 发票号码
            sb = str(ws.cell(row=row_idx, column=2).value or '').strip()  # A列 - 标识（任意非空值表示该行有效）
            # bank = str(ws.cell(row=row_idx, column=2).value or '')  # B列 - 银行
            # zh = str(ws.cell(row=row_idx, column=3).value or '')  # C列 - 账号
            # shui = str(ws.cell(row=row_idx, column=4).value or '')  # D列 - 税号
            
            # 如果公司名称为空，结束循环
            if fphm == '' or fphm == None:
                break
            if sb == '' or sb == None:
                sb = '是'
            d = s.query(cymx.rid).filter(cymx.fphm==fphm, func.ifnull(cymx.sjcy1,'') != '').first()
            if not d:                
                # 记录重复信息
                duplicates.append(f'请注意发票号码: {fphm} 没填实际出运日期')
            
            row_idx += 1
        
        # 如果有重复记录，生成文本文件
        result_data = {
            'has_duplicates': len(duplicates) > 0,
        }
        
        if duplicates:
            # 生成重复记录文件
            dup_filename = f"无实际出运日期_{datetime.now().strftime('%Y-%m-%d')}.txt"
            dup_filepath = os.path.join(temp_dir, dup_filename)
            val ='\n'.join(duplicates)
            logger.error(f"Writing duplicates to: {dup_filepath}")
            write_file(dup_filepath, val, 'w')

            # 这里应该上传文件到可下载的位置，返回URL
            # 简化处理：返回文件内容或保存到特定目录
            result_data['duplicate_file_name'] = dup_filename
        
        # 清理临时文件
        # os.remove(temp_file)
        
        return json_result(0, data=result_data)
        
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()
        wb.close()

# 出运明细的批量电放按钮
@any_route('/api/shipment/import_sfdf_excel', methods=['POST'])
@require_token
async def view_shpipment_import_sfdf_excel(request):
    user = request.current_user
    s = Session()
    try:
        org = get_user_path(user.username)
        position = org.get('position')
        if '电放' not in position:
            return json_result(-1, '只有电放岗位用户才能执行此操作')
        # 2. 获取上传的Excel文件
        j = await request.form()
        file = form_value(j,'file',None)
        if is_none(file):
            return json_result(ERR_PARAM_NOT_ENOUGH)
        # 保存临时文件
        temp_dir = config.tmp_path
        temp_file = os.path.join(temp_dir, f"{uuid.uuid4()}_{file.filename}")
        data = await file.read()
        write_file(temp_file, data, 'wb', encoding=None)
        # 3. 读取Excel文件
        wb = load_workbook(temp_file)
        ws = wb.worksheets[0]  # 读取第一个工作表
        # 4. 处理每一行数据
        row_idx = 2  # 从第2行开始（第1行是表头）
        while True:
            # 读取Excel单元格数据
            fphm = str(ws.cell(row=row_idx, column=1).value or '').strip()  # A列 - 发票号码
            hxdate = str(ws.cell(row=row_idx, column=2).value or '')  # B列 - 银行            
            # 如果发票号码为空，结束循环
            if not fphm:
                break
            if hxdate == '' or hxdate == None:
                hxdate = '是'
            s.query(cymx).filter(cymx.fphm==fphm).update({cymx.hxdate: str(hxdate)}, synchronize_session=False)            
            row_idx += 1
        
        # 提交事务
        s.commit()
        return json_result(0, '操作成功')
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()

# 出运明细的批量电放按钮
@any_route('/api/saier/shipment/return/cyz', methods=['POST'])
@require_token
async def view_shpipment_return_cyz(request):
    user = request.current_user
    s = Session()
    j = await request.json()
    try:
        org = get_user_path(user.username)
        position = org.get('position')
        # if '单证' not in position:
        #     return json_result(-1, '只有单证岗位用户才能执行此操作')
        s.query(cymx).filter(cymx.rid==j.get('rid')).update({cymx.tjpd: '有'}, synchronize_session=False)
        
        # 提交事务
        s.commit()
        return json_result(1, '操作成功')
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()

def shipment_update_bgmx(d, o, c, user, flag=0):
    d.mtime = time.strftime('%Y-%m-%d %H:%M:%S')
    d.modi_uid = user.rid
    d.wxwyzd = getattr(c, o.field_by_full_name('产品资料.外销唯一字段').db.name)
    d.cpbh = getattr(c, o.field_by_full_name('产品资料.产品编号').db.name)
    d.zwpm = getattr(c, o.field_by_full_name('产品资料.中文品名').db.name)
    d.hgbm = getattr(c, o.field_by_full_name('产品资料.海关编码').db.name)
    d.gcmc = getattr(c, o.field_by_full_name('产品资料.工厂名称').db.name)
    d.jsfs = getattr(c, o.field_by_full_name('产品资料.结算方式').db.name)
    d.scrq = getattr(c, o.field_by_full_name('产品资料.进仓日期').db.name) if getattr(c, o.field_by_full_name('产品资料.进仓日期').db.name) != '' else None
    d.gcjg = getattr(c, o.field_by_full_name('产品资料.采购单价').db.name)
    d.gczj = getattr(c, o.field_by_full_name('产品资料.采购总价').db.name)
    d.wxrl = getattr(c, o.field_by_full_name('产品资料.外箱容量').db.name)
    d.chxs = getattr(c, o.field_by_full_name('产品资料.箱    数').db.name)
    d.chsl = getattr(c, o.field_by_full_name('产品资料.出货数量').db.name)
    if flag == 0:
        d.kpgc = getattr(c, o.field_by_full_name('产品资料.开票工厂').db.name)
        d.sbys = getattr(c, o.field_by_full_name('产品资料.申报要素').db.name)
        d.zzsl = getattr(c, o.field_by_full_name('产品资料.增值税率').db.name)
        d.tsl = getattr(c, o.field_by_full_name('产品资料.退 税 率').db.name)
        d.tse = getattr(c, o.field_by_full_name('产品资料.退 税 额').db.name)
        d.xddd = getattr(c, o.field_by_full_name('产品资料.下单地点').db.name)
        d.hyd = getattr(c, o.field_by_full_name('产品资料.货 源 地').db.name)
        d.zhwbgpm = getattr(c, o.field_by_full_name('产品资料.中文报关品名').db.name)
        d.wxjg = getattr(c, o.field_by_full_name('产品资料.外销单价').db.name)
        d.wxzj = getattr(c, o.field_by_full_name('产品资料.外销总价').db.name)
        d.mjdj1 = getattr(c, o.field_by_full_name('产品资料.客户RMB单价').db.name)
        d.mjzj = getattr(c, o.field_by_full_name('产品资料.客户RMB总价').db.name)
        d.djpmw = getattr(c, o.field_by_full_name('产品资料.单据品名外').db.name)
        d.djpmy = getattr(c, o.field_by_full_name('产品资料.单据品名英').db.name)
        d.krcode = getattr(c, o.field_by_full_name('产品资料.客人CODE').db.name)
        d.sfsj = getattr(c, o.field_by_full_name('产品资料.是否商检').db.name)
        d.ewchy = getattr(c, o.field_by_full_name('产品资料.是否代开').db.name)

    return d

# 出运明细的同步到报关明细按钮
@any_route('/api/saier/shipment/update/bgmx', methods=['POST'])
@require_token
async def view_shpipment_update_bgmx(request):
    user = request.current_user
    s = Session()
    j = await request.json()
    try:
        rid = j.get('rid')
        lines = j.get('lines', [])
        m = s.query(cymx).filter(cymx.rid==rid).first()
        if not m:
            return json_result(-1, '出运明细记录不存在')
        # if user.username != m.zdry:
        #     return json_result(-1, '只有制单人员可以执行此操作')
        fkhl = 0
        # 获取付款汇率
        h = s.query(fkhlb.hl).filter(fkhlb.djny == str(m.chyrq)[:7]).first()
        if h:
            fkhl = h.hl or 0
        ywywhlv = m.huilv or 0
        fphm = str(m.fphm)
        RMBkh = str(m.RMBkh)

        fphm_list = [fphm + suffix for suffix in suffixes]
        o = get_module('出运明细')
        # g = o.group_by_name('产品资料')
        for rid in lines:
            c = s.query(cymxsheet).filter(cymxsheet.rid==rid).first()
            if not c:
                 return json_result(-1, f'产品资料记录不存在')
            cywyzd = c.cywyzd
            # 同步到现金合成
            d = s.query(xjhcsheet).filter(xjhcsheet.cywyzd==cywyzd).first()
            if d:
                d = shipment_update_bgmx(d, o, c, user, 1)
                s.add(d)
            # 同步到出运明细的产品资料（可能有多行，需循环）
            a = s.query(cymxsheet).filter(cymxsheet.cywyzd==cywyzd).filter(cymxsheet.fphm.in_(fphm_list)).all()
            for d in a:
                d = shipment_update_bgmx(d, o, c, user)
                s.add(d)
            # 同步到报关明细的产品资料
            d = s.query(bgmxdsheet).filter(bgmxdsheet.cywyzd==cywyzd).first()
            if d:
                d = shipment_update_bgmx(d, o, c, user)
                s.add(d)
            # 同步到未处理清单
            d = s.query(dbbqd).filter(dbbqd.cywyzd==cywyzd).first()
            if d:
                d = shipment_update_bgmx(d, o, c, user)
                s.add(d)
        # 更新出运明细的汇总金额等字段
        c = s.query(cymx).filter(cymx.fphm.in_(fphm_list)).all()
        for d in c:
            rmbche = s.query(func.sum(func.ifnull(cymxsheet.gczj, 0))).filter(cymxsheet.pid==d.rid, cymxsheet.ewchy!='是').scalar() or 0
            d.hjje = rmbche
            mjzj1 = 0
            wxje1 = 0
            cbzje = 0
            a = s.query(cymxsheet.mjzj,cymxsheet.wxzj,cymxsheet.RMBkh).filter(cymxsheet.pid==d.rid).all()
            for r in a:
                if r.RMBkh == '是':
                    mjzj1 = mjzj1 + (r.mjzj or 0)
                else:
                    wxje1 = wxje1 + (r.wxzj or 0)
            mjzj1 = round(mjzj1*100)/100
            wxje1 = round(wxje1*100)/100
            if d.yzaf == None or d.yzaf == 0:
                cbzje = rmbche + d.jxRMB + d.zgfy + d.ckfy + d.ygnlf + d.qtrmb + d.kpfyz
            else:
                cbzje = rmbche + d.jxRMB + d.zgfy + d.ckfy + d.yzaf + d.qtrmb + d.kpfyz
            if RMBkh != '是':
                htje = round((wxje1 + d.jxje1 + d.jxje2 + d.jxje3 - d.jjxje1 - d.jjxje2 - d.jjxje3 + mjzj1 / ywywhlv + d.jxKHRMB / ywywhlv + d.ewhj + d.jxUSD) * 100) / 100
                bf1 = round(((wxje1 - d.myjje - d.ysdj) * d.xbfl) * 100) / 100
                wxjez = htje
                hzhjz = htje
                fkje = htje
            else:
                htje = round((mjzj1 + d.jxje1 + d.jxje2 + d.jxje3 - d.jjxje1 - d.jjxje2 - d.jjxje3 + wxje1 * ywywhlv + d.jxKHRMB + d.jxUSD * ywywhlv + d.ewhj) * 100) / 100
                bf = round(((htje - d.myjje - d.ysdj) * d.xbfl) * 100) / 100
                if ywywhlv > 0:
                    wxjez = round((htje / ywywhlv) * 100) / 100
                    if fkhl!= 0:
                        fkje = round(((round(((mjzj1 + d.jxje1 + d.jxje2 + d.jxje3 - d.jjxje1 - d.jjxje2 - d.jjxje3 + wxje1 * d.zmyhl * fkhl + d.jxKHRMB + d.jxUSD * d.zmyhl * fkhl + d.ewhj)) * 100) / 100) / fkhl) * 100) / 100
                    else:
                        fkje = wxjez
                    hzhjz = wxjez
            if RMBkh == '是':
                htjer = round(((mjzj1 + d.jxje1 + d.jxje2 + d.jxje3 - d.jjxje1 - d.jjxje2 - d.jjxje3 + d.jxKHRMB + d.ewhj)) * 100) / 100
                htjem = round(((wxje1 + d.jxUSD)) * 100) / 100
            else:
                htjem = round(((wxje1 + d.jxje1 + d.jxje2 + d.jxje3 - d.jjxje1 - d.jjxje2 - d.jjxje3 + d.jxUSD + d.ewhj)) * 100) / 100
                htjer = round((mjzj1 + d.jxKHRMB) * 100) / 100
            d.htjem = htjem
            d.htjer = htjer
            d.fkje = fkje
            d.hzhjz = hzhjz
            d.wxjez = wxjez
            d.cbzje = cbzje
            d.cghjzje = rmbche
            d.htje = htje
            d.bfmy = bf1
            d.bf = bf
            d.htje = htje
            d.wxje = wxje1
            d.mjzj = mjzj1
            if RMBkh == '是':
                d.hzhjDR = htje - d.myjje
                d.hzhjDM = 0
                d.mjzjs = round(((htje - d.myjje) / d.huilv) * 100) / 100
            else:
                d.hzhjDR = 0
                d.hzhjDM = (htje - d.myjje) * d.zmyhl
                d.mjzjs = (htje - d.myjje) * d.zmyhl
            s.add(d)

        # 提交事务
        s.commit()
        return json_result(1, '操作成功')
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()


# 出运明细的业务停止收汇提醒按钮
@any_route('/api/saier/shipment/stop_incomes_notice', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_stop_incomes_notice(request):
    s = Session()
    j = await request.json()
    user = request.current_user
    try:
        rid = j.get('rid', '')
        d = s.query(cymx).filter(cymx.rid == rid).first()
        if not d:
            return json_result(-1, '记录不存在，请选中需要操作的记录')
        # if d.ywry != user.username:
        #     return json_result(-1, '业务人员不是当前用户，无法执行此操作')
        fttdh = d.fttdh or ''
        tztxnr = d.tztxnr or ''
        if not '收汇提醒' in d.fttdh:
            if tztxnr == '' or tztxnr == None:
                tztxnr = '收汇提醒'
            else:
                tztxnr = tztxnr + ',收汇提醒'
        if fttdh != '' and fttdh != None and tztxnr != '' and tztxnr != None:
            s.query(cymx).filter(cymx.fttdh == fttdh, func.ifnull(cymx.fttdh,"") != "").update({cymx.tztxnr: tztxnr}, synchronize_session=False)
            s.commit()

        return json_result(1, '操作成功')
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

def api_fdsq1sheet_new(pid, d, user, j, s):
    for c in d:
        m = fdsq1sheet()
        for r,v in c.items():
            if r in SYS_FIELDS:
                continue
            setattr(m, r, v)
        m.rid = get_uuid()
        m.pid = pid
        m.ctime = time.strftime('%Y-%m-%d %H:%M:%S')
        m.sqry = user.username
        m.sqrq = time.strftime("%Y-%m-%d")
        m.sqbz = ''
        m.ywqr = ''
        m.dzdf = ''
        m.qrry = ''
        m.cpjq = '否'
        for c,v in j.items():
            if c in SYS_FIELDS:
                continue
            setattr(m, c, v)
        s.add(m)


def api_fdsq1_new(rid, c, user, j, s):
    m = fdsq1()
    for r,v in c.items():
        if r in SYS_FIELDS:
            continue
        setattr(m, r, v)
    m.rid = rid
    m.uid = user.rid
    m.ctime = time.strftime('%Y-%m-%d %H:%M:%S')
    m.sqrq = time.strftime("%Y-%m-%d")
    m.sqbz = ''
    m.htjez = c.get('htje', 0)
    m.tjcw = ''
    m.cwsp = '待审批'
    m.jlsp = '待审批'
    m.fksp = '待审批'
    m.zjlsp = '待审批'
    m.sftj = '否'
    m.cpjq = '否'
    m.sftssp = '否'
    for c,v in j.items():
        if c in SYS_FIELDS:
            continue
        setattr(m, c, v)
    s.add(m)


# 出运明细的批量放单申请按钮
@any_route('/api/saier/shipment/batch/fdsq', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_batch_fdsq(request):
    s = Session()
    j = await request.json()
    user = request.current_user
    try:
        # d =  run_sql(f"select mc from zx where ly='批量放单申请' and mc='{user.username}' limit 1")
        # if len(d) == 0:
        #     return json_result(-1, '权限校验失败')
        f = ['rid', 'khmc', 'chyrq', 'sjcy1', 'wxje', 'htje', 'yjje', 'ywry', 'ywbm', 'mjzj', 'RMBkh', 'ayjje', 'myjje', 'zdry', 'xybx', 'kh_id', 'eta', 'fphm', 'hxdh', 'jhfs', 'dfrq', 'kh_idf', 'hxdate']
        fields = [getattr(cymx, k) for k in f]
        rids = j.get('rids', [])
        key_data = {}
        for rid in rids:
            d = s.query(*fields).filter(cymx.rid == rid).first()
            if not d:
                continue
            if d.chyrq== None or d.chyrq == '' or d.hxdate == '是':
                continue
            if d.khmc == None or (not ('BEST PRICE LLC' in d.khmc or d.khmc == 'SIA FP LV' or d.khmc == 'FIX PRICE GENERAL TRADING LLC')):
                continue
            c = s.query(fdsq1sheet.rid).filter(fdsq1sheet.fphm == d.fphm).first()
            if c:
                continue
            c = alchemy_object_to_dict(d)
            key_str = f"{d.kh_idf}_{str(d.RMBkh)}_{d.xybx}"
            if key_str not in key_data:
                c = alchemy_object_to_dict(d)
                key_data[key_str]['main'] = alchemy_object_to_dict(c)
                key_data[key_str]['lines'] = [alchemy_object_to_dict(c)]
            else:
                key_data[key_str]['lines'].append(alchemy_object_to_dict(c))
                l = key_data[key_str].get('main', {})
                l['wxje'] = round(l.get('wxje', 0) or 0 + c.get('wxje', 0) or 0, 3)
                l['htje'] = round(l.get('htje', 0) or 0 + c.get('htje', 0) or 0, 3)
                l['yjje'] = round(l.get('yjje', 0) or 0 + c.get('yjje', 0) or 0, 3)
                l['mjzj'] = round(l.get('mjzj', 0) or 0 + c.get('mjzj', 0) or 0, 3)
                l['ayjje'] = round(l.get('ayjje', 0) or 0 + c.get('ayjje', 0) or 0, 3)
                l['myjje'] = round(l.get('myjje', 0) or 0 + c.get('myjje', 0) or 0, 3)
        user_list = []
        for k,v in key_data.items():
            n = v.get('main')
            l = v.get('lines', [])
            pid = get_uuid()
            ywry = n.get('ywry', '')
            tjjl = ''
            tjfk = ''
            tjzjl = ''
            u =s.query(ywrybiao.yhm).filter(ywrybiao.yhm != ywry, ywrybiao.bmjl == ywry, func.ifnull(ywrybiao.zt,"") != "离职", func.ifnull(ywrybiao.zt,"")!="").first()
            if u:
                ywry = u.yhm
            org = get_user_path(ywry)
            yw = org.get('path', '')
            yw1 = yw.replace('\\', '_')
            u = run_sql(f"select dlr from spwt where (yhms=2) and ('{yw1} like concat('%', replace(fdsq,'\\\\','_'), '%') and (dlr<>'{user.username}') order by yhms desc")
            if len(u)>0:
              tjjl = u[0].get('dlr')
              if tjjl!=None and tjjl!= '' and tjjl not in user_list:
                user_list.append(tjjl)
            u = run_sql(f"select dlr from spwt where (yhms>20) and (yhms<31) and ('{yw1} like concat('%', replace(fdsq,'\\\\','_'), '%') and (dlr<>'{user.username}') order by yhms desc")
            if len(u)>0:
                tjfk = u[0].get('dlr') 
                if tjfk!=None and tjfk!= '' and tjfk not in user_list:
                    user_list.append(tjfk)
            u = run_sql(f"select dlr from spwt where (yhms=1) and ('{yw1} like concat('%', replace(fdsq,'\\\\','_'), '%') and (dlr<>'{user.username}') order by yhms desc")
            if len(u)>0:
                tjzjl = u[0].get('dlr')
                if tjzjl!=None and tjzjl!= '' and tjzjl not in user_list:
                    user_list.append(tjzjl)
            fdbh = auto_number.generate(s,'放单申请.放单编号',{'rid':pid})
            api_fdsq1_new(pid, n, user, {'fdbh': fdbh, 'tjjl': tjjl, 'tjfk': tjfk, 'tjzjl': tjzjl, 'yw': yw}, s)
            api_fdsq1sheet_new(pid, l, user, {'yw': yw}, s)

            for u in user_list:
                row = {
                    "xxly": '放单申请',
                    "bjdh": '',
                    "wxht": main.get('fphm', ''),
                    "cght": '',
                    "yhdh": '',
                    "xxnr": f"{user.username}的放单申请{fdbh}需校对,日期:{time.strftime('%Y-%m-%d')}",
                    "jsr": str(u),
                    "sys_path": "",
                    "spsq": user.username
                }
                res = module_xxck_new([row], user, s)
                if res.get('code', 1) != 1:
                    return {'code': -1, 'msg': res.get('msg')}

            res = user_task_new(
                '放单申请', d.rid, '放单编号', f"放单编号{fdbh}需校对", f"{user.username}的放单申请{fdbh}需校对,日期:{time.strftime('%Y-%m-%d')}", user, s, user_list)
            if res.get('code') != 1:
                s.rollback()
                return {'code': -1, 'msg': res.get('msg')}
            
        s.commit()
        return json_result(1, '操作成功')
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()



# 出运明细的批量放单申请按钮
@any_route('/api/saier/shipment/cyjh_return', methods=['POST', 'GET'])
@require_token
async def view_saier_shipment_cyjh_return(request):
    s = Session()
    j = await request.json()
    user = request.current_user
    try:
        rid = j.get('rid', '')
        bgyy = j.get('bgyy', '')
        flag = j.get('flag', 0)
        cydh = ''
        filters = []
        if flag == 0:
            d = s.query(cymx).filter(cymx.rid==rid).first()
            if not d:
                return json_result(-1, '记录不存在')
            # if d.zdry != user.username:
            #     return json_result(-1, '制单人员不是当前用户，无法执行此操作')
            cydh = d.chydh
            d.zdzt = '是'
            s.add(d)
            filters.append(chyjh.wxfp==cydh)
            if d.zdry != user.username:
                return json_result(-1, '只有制单人员可以执行此操作')
        else:
            filters.append(chyjh.rid==rid)

        c = s.query(chyjh).filter(*filters).first()
        if c:
            if c.dzry != user.username and c.fkdz != user.username:
                return json_result(-1, '只有单证人员或单证主管可以执行此操作')
            # if c.wf_status !=2:
            #     return json_result(-1, '审批未通过时不能执行此操作')
            cydh = c.wxfp
            c.jlhz = '待审批'
            c.cysq = ''
            c.zjhz = ''
            c.dzqr1 = '再审'
            c.wf_status = 3
            s.add(c)
            # 如果当天没有记录，则新增一条记录
            if bgyy != '' and bgyy != None:
                b = s.query(chyjhsheet10).filter(chyjhsheet10.pid==c.rid,func.ifnull(chyjhsheet10.sj, '')==time.strftime('%Y-%m-%d')).first()
                if not b:
                    a = chyjhsheet10()
                    a.sj = time.strftime('%Y-%m-%d')
                    a.rid = get_uuid()
                    a.pid = c.rid
                    a.uid = user.rid
                    a.ctime = time.strftime('%Y-%m-%d %H:%M:%S')
                    a.bgyy = bgyy
                    a.bgtj = user.username
                    s.add(a)

            res = user_task_new(
                '出运计划', c.rid, '出运单号', f"出运计划发票号码{c.wxfp}需要更改", f"请注意发票号码为{cydh}需要更改,日期:{time.strftime('%Y-%m-%d')}", user, s, [c.zdry])
            if res.get('code') != 1:
                s.rollback()
                return {'code': -1, 'msg': res.get('msg')}
              
        s.commit()
        return json_result(1, '操作成功')
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()