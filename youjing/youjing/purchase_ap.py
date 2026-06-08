import gc
from math import trunc
import os
from any import *
from .model import *
from sqlalchemy.sql import func, not_, or_, and_
from .__default__ import user_task_delete, user_task_new, module_xxck_new, get_user_path
import time
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, Alignment, NamedStyle,PatternFill,Color
from openpyxl.drawing.image import Image as Image_Get
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import  XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.worksheet.protection import SheetProtection

# 预填费用的预填金额改变时
@any_route('/api/saier/purchase_ap/ytje/change', methods=['POST', 'GET'])
@require_token
async def view_saier_purchase_ap_ytje_change(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        gcmc = j.get('gcmc')
        ytje = j.get('ytje', 0)
        mlsb = '是'
        je = 0
        d = s.query(bmlgc.rid).filter(bmlgc.gcmc == gcmc).first()
        if d:
            mlsb = '否'

        d = s.query(cyzglsheet.sz).filter(cyzglsheet.xm == '公司', cyzglsheet.zm == '抹零金额').first()
        if d:
            if ytje != None and ytje != 0 and float(ytje) <= float(d.sz):
                mlsb = '否'
        d = s.query(cyzglsheet.sz).filter(cyzglsheet.xm == user.username, cyzglsheet.zm == '财务预填费用').first()
        if not d:
            if ytje != None and ytje > 10 and mlsb == '是':
                je = trunc(float(ytje) / 10) * 10
        
        return json_result(1, '操作成功', je)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

@any_route('/api/saier/purchase_ap/load/check', methods=['POST', 'GET'])
@require_token
async def view_saier_purchase_ap_load_check(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        sqry = j.get('sqry', '')
        sqbm = j.get('sqbm', '')
        cwsb = 0
        flag = j.get('flag', 0)
        if sqry == '' or sqry == None:
            sqry = user.username
        bm = ''
        bmdm = ''
        if sqbm == '' or sqbm == None:
            d = s.query(ywrybiao.bm).filter(ywrybiao.yhm == sqry).first()
            if d:
                bm = d.bm
                c = s.query(ywbdzb.dyywzm).filter(ywbdzb.ywb == bm).first()
                if c:
                    bmdm = c.dyywzm
        if flag == 0:
            d = s.query(cyzglsheet.sz).filter(cyzglsheet.xm == user.username, cyzglsheet.zm == '财务预填费用').first()
            if d:
                cwsb = 1

        return json_result(1, '操作成功', {'cwsb': cwsb, 'bm': bm, 'bmdm': bmdm})
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


@any_route('/api/saier/purchase_ap/return/back', methods=['POST', 'GET'])
@require_token
async def view_saier_purchase_ap_return_back(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        rid = j.get('rid')
        cwsb = 0
        sb = 0
        d = s.query(ytfy).filter(ytfy.rid == rid, ytfy.tjjl == user.username, or_(ytfy.zjlhz == '待审批', ytfy.zjlhz == '不通过', func.ifnull(ytfy.zjlhz,"") == '')).first()
        if d:
            d.tjjl=""
            d.jlhz="待审批"
            d.zjlhz="待审批"
            d.tjzjl=""
            sb = 1
        c = s.query(cyzglsheet.sz).filter(cyzglsheet.xm == user.username, cyzglsheet.zm == '财务预填费用').first()
        if c:
            cwsb = 1
        if cwsb==1 and d.zjlhz == '通过':
            d.tjjl=""
            d.jlhz="待审批"
            d.zjlhz="待审批"
            d.tjzjl=""
            d.cwsd=""
            d.kfdy="不可"
            sb = 1
        if sb == 1:
            s.add(d)
            s.commit()
        
        return json_result(1, '操作成功', sb)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


@any_route('/api/saier/purchase_ap/save/before', methods=['POST', 'GET'])
@require_token
async def view_saier_purchase_ap_save_before(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        data = j.get('data')
        lines = j.get('lines')
        rid = data.get('rid')
        fktt = data.get('fktt')
        sqbm = data.get('sqbm')
        ytdh = data.get('ytdh', '')
        zjlhz = data.get('zjlhz', '')
        module = j.get('module', '预填费用')
        m = get_module(module)
        modi_rids = []
        spsq = ''
        tjzjl = data.get('tjzjl', '')
        tjjl = data.get('tjjl', '')
        Field1 = data.get('field1', '')
        Field2 = data.get('field2', '')
        Field3 = data.get('field3', '')
        ttsb = 0
        i2 = 0
        for r in lines:
            ttsb1 = 0
            i2 = i2 + 1
            c = s.query(fkspsheet3.fktt).filter(fkspsheet3.hthm == r.get('cght')).first()
            if c:
                r['fktt'] = c.fktt
                ttsb1 = 1
            c = s.query(bgmxdsheet.fktt).filter(bgmxdsheet.cght == r.get('cght')).first()
            if c:
                r['fkfs'] = c.fktt
                ttsb1 = 1
            if ttsb1== 1 and r.get('fktt') != fktt:
                ttsb = 1
            if ttsb1 != 1:
                r['fktt'] = fktt
            r['ytbh'] = data.get(m.field_by_full_name(module+'.预填单号').db.name, '')
            r['sqrq'] = data.get(m.field_by_full_name(module+'.申请日期').db.name, '')
            r['sccj'] = data.get(m.field_by_full_name(module+'.生产厂家').db.name, '')
            r['zzsl'] = data.get(m.field_by_full_name(module+'.增值税率').db.name, 0)
            r['tsl'] = data.get(m.field_by_full_name(module+'.退 税 率').db.name, 0)
            r['hyd'] = data.get(m.field_by_full_name(module+'.货 源 地').db.name, '')

            modi_rids.append(r.get('rid'))
        if ttsb == 1:
            return json_result(-1, '详情信息中存在付款抬头和主要信息不一致的情况请核实后再保存')
        if i2 == 0:
            return json_result(-1, '无详情信息不能保存')
        # if i2 > 0:
        if data.get('Field') != '' and data.get('Field') != None:
            spsq = data.get('sqry', '')
        elif Field1 != '' and Field1 != None:
            c = s.query(cyzglsheet.xm).filter(cyzglsheet.bz.like('%' + sqbm + '%'), cyzglsheet.zm == '财务预填费用').first()
            if c:
                spsq = c.xm
        elif Field3 != '' and Field3 != None:
            spsq = tjzjl
        elif Field2 != '' and Field2 != None:
            spsq = tjjl

        if spsq != '' and spsq != None:
            res = user_task_delete(module, rid, s, [spsq])
            if res.get('code', 1) != 1:
                s.rollback()
                return json_result(-1, res.get('msg'))
            
            content = user.username + '预填费用:' + str(ytdh) + '审批申请,请查看:' + time.strftime('%Y-%m-%d')
            title = str(user.username) + '预填费用' + str(ytdh) + '审批通知'
            res = user_task_new(module, rid, '预填单号', title, content, user, s, [spsq], True)
            if res.get('code') != 1:
                logger.error(res.get('msg','创建任务失败'))
                return json_result(-1, res.get('msg'))
            row = {
                "xxly": module,
                "bjdh": '',
                "wxht": '',
                "cght": ytdh,
                "yhdh": '',
                "xxnr": content,
                "jsr": str(spsq),
                "sys_path": '',
                "spsq": user.username,
            }
            res = module_xxck_new([row], user, s)
            if res.get('code',1) != 1:
                return json_result(-1, res.get('msg'))
        if zjlhz == '不通过':
            spsq1 = data.get('sqry')
            res = user_task_delete(module, rid, s, [spsq1])
            if res.get('code', 1) != 1:
                s.rollback()
                return json_result(-1, res.get('msg'))
            
            content = user.username + '预填费用:' + str(ytdh) + '审批不通过,原因:' + data.get('zjlyj') + ',请查看:' + time.strftime('%Y-%m-%d')
            title = str(user.username) + '预填费用' + str(ytdh) + '审批不通过通知'
            res = user_task_new(module, rid, '预填单号', title, content, user, s, [spsq1], True)
            if res.get('code') != 1:
                logger.error(res.get('msg','创建任务失败'))
                return json_result(-1, res.get('msg'))
            row = {
                "xxly": module,
                "bjdh": '',
                "wxht": '',
                "cght": ytdh,
                "yhdh": '',
                "xxnr": content,
                "jsr": str(spsq1),
                "sys_path": '',
                "spsq": user.username,
            }
            res = module_xxck_new([row], user, s)
            if res.get('code',1) != 1:
                return json_result(-1, res.get('msg'))
        
        s.commit()
        return json_result(1, '操作成功', {'modi_rids': modi_rids, 'lines': lines})
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


@any_route('/api/saier/purchase_ap/billed/download', methods=['POST', 'GET'])
@require_token
async def view_saier_purchase_ap_billed_download(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        rids = j.get('rids', [])
        kind = j.get('kind', '')
        filename17l = '锐亿17j.xlsx'
        filename31l = '锐亿3j.xlsx'
        tp = 1
        if kind == 'PDF' or kind == '批量PDF':
            filename17l = '锐亿17.xlsx'
            filename31l = '锐亿3.xlsx'
            tp = 2
        
        f17l_path = os.path.join(config.data_upload_path, 'template/' + filename17l)
        f31l_path = os.path.join(config.data_upload_path, 'template/' + filename31l)
        if not os.path.exists(f17l_path) or not os.path.exists(f31l_path):
            return json_result(-1, '开票信息导出模板文件不存在，请联系管理员')
        
        ywcs = 0
        nbcs = 0
        t = run_sql("select cs from zx where (ly='合同签订日期')")
        if len(t) > 0:
            ywcs = t[0].get('cs')
            nbcs = t[0].get('cs')

        wb = load_workbook(f17l_path)
        ws = wb.active
        for rid in rids:
            yssb = 0
            ywsb = 0
            i1 = 0
            ps = 0
            c = 0
            d = s.query(ytfy).filter(ytfy.rid == rid, ytfy.kfdy=='可以').first()
            if not d:
                return json_result(-1, '未找到对应的预填费用记录')
            if '优景' in d.fktt:
                yssb = 1
            d = alchemy_object_to_dict(d)
            chyrq = d.get('chrq') if d.get('chrq') else ''
            hyd = d.get('hyd') if d.get('hyd') else ''
            fktt = d.get('fktt') if d.get('fktt') else ''
            wstt = d.get('wstt') if d.get('wstt') else ''
            sfdl = d.get('fphm', '')[:2] if d.get('fphm') else ''
            jjdz = ''
            wstt1 = ''
            gs = ''
            dz = ''
            sw = ''
            bank = ''
            wstt2 = ''
            kpcxdh = ''
            p = run_sql(f"select * from kpnr")
            for r in p:
                wstt2 =r.get('gsjc')
                if wstt2 in  fktt:
                    gs =r.get('wfgs')
                    dz =r.get('kpdz')
                    sw =r.get('kpsh')
                    bank =r.get('kpyh')
                    jjdz =r.get('kpjjdz')
                    kpcxdh =r.get('kpcxdh')
                    wstt1 =r.get('gsjc')
                    break

            if '义乌' in fktt:
                ywsb = 1

            if jjdz == '' or jjdz == None:
                if '义乌' in wstt:
                    jjdz = '寄件地址：义乌市宗泽北路531号赛尔集团二楼   /邮编：322000/ 吴春燕 收/0579-85096055'
                else:
                    jjdz = '寄件地址：宁波高新区光华路288号赛尔大厦25层  邮编315103/财务收/0574-27833931'

            cx1 = d.get('fphm', '').upper() if d.get('fphm') else ''
            cx1z = d.get('fphm', '').upper() if d.get('fphm') else ''
            zs = len(d.get('fphm', '')) if d.get('fphm') else 0
            zs2 = 0
            zs3 = 0
            zsw = ''
            sbcx = ''
            for zs1 in range(1, zs + 1):
                zs2 += 1
                zsw = cx1[zs2 - 1]  # 字符串索引从0开始，所以要减1
                if (zsw == '/' or zsw == '\n' or zsw == '\r' or zsw == '\\' or 
                    zsw == ':' or zsw == '*' or zsw == '?' or zsw == '"' or 
                    zsw == '<' or zsw == '>' or zsw == '|' or zsw == ''):
                        sbcx = '1'
                        if zs3 == 0:
                            zs3 = zs2
            if zs3 >= 1:
                if sbcx == '1':
                    cx1z = d.get('fphm', '')[:zs3 - 1]
                else:
                    cx1z = d.get('fphm', '')
            lxfs = d.get('lxfs') if d.get('lxfs') else ''
            if d.get('yjch') != '' and d.get('yjch') != None:
                cx2 = d.get('yjch')
            else:
                cx2 = d.get('chrq')
            cx3 = d.get('ywry')
            cx4 = d.get('lxdh')
            cx = d.get('sccj')
            zs = len(d.get('sccj'))
            zs2 = 0
            zs3 = 0
            zsw = ''
            sbcx = ''
            for zs1 in range(1, zs + 1):
                zs2 += 1
                zsw = cx[zs2 - 1]
                if (zsw == '/' or zsw == '\n' or zsw == '\r' or zsw == '\\' or 
                    zsw == ':' or zsw == '*' or zsw == '?' or zsw == '"' or 
                    zsw == '<' or zsw == '>' or zsw == '|' or zsw == ''):
                        sbcx = '1'
                        if zs3 == 0:
                            zs3 = zs2
            if zs3 >= 1:
                if sbcx == '1':
                    cx = cx[:zs3 - 1]
                else:
                    cx = d.get('sccj')
            number = d.get('rid')
            i1 = i1 + 1
            zb1 = ''

            if d.get('zsl') >= 0:
                ps = ps + 1
                i = 1
                c = c + 1
                tppath = ''
                u = run_sql(f"select * from ywrylx where (ryxm='{d.get('gdry')}')")
                if len(u) > 0 and u[0].get('Msn') != '' and u[0].get('Msn') != None:
                    jjdz = u[0].get('Msn')
                cpbh = gs + '合同章'
                t = run_sql(f"select tpmc,tpmc1 from tpzx where (cpbh='{cpbh}') and (length(tpmc) > 5)")
                if len(t) == 0:
                    tppath = ''
                else:
                    if tp==2 :
                        tppath = t[0].get('tpmc')
                    else:
                        tppath = t[0].get('tpmc1')

                ws['A8'] = gs
                ws['C14'] = gs
                ws['C15'] = dz
                ws['C16'] = sw
                ws['C17'] = bank
                # msexcelworksheet.Rows[inttostr(11)].AutoFit
                ws['G11'] = cx1
                ws['G12'] = ''  # cx7
                ws['G44'] = cx1
                # msexcelworksheet.Rows[inttostr(21)].AutoFit
                ws['C21'] = d.get('zwpm')
                ws['G24'] = cx2
                ws['B22'] = d.get('chsl')
                ws['E22'] = d.get('hgjldw')
                ws['E24'] = '0'
                ws['G23'] = d.get('ytje')
                if d.get('zsl') > 3:
                    ws['B23'] = round((d.get('ytje')) / 1.13 * 100) / 100
                    ws['G22'] = '13%'
                    ws['E23'] = d.get('ytje') - round((d.get('ytje')) / 1.13 * 100) / 100
                if d.get('zsl') == 3:
                    ws['B23'] = round((d.get('ytje')) / 1.03 * 100) / 100
                    ws['G22'] = '3%'
                    ws['E23'] = d.get('ytje') - round((d.get('ytje')) / 1.03 * 100) / 100
                    
                if d.get('zsl') == 1:
                    ws['B23'] = round((d.get('ytje')) / 1.01 * 100) / 100
                    ws['G22'] = '1%'
                    ws['E23'] = d.get('ytje') - round((d.get('ytje')) / 1.01 * 100) / 100

                # if tp == 2:
                #     msexcel.Run('Sheet1.A')

                ws['G45'] = '(预填单号:' + d.get('ytbh') + ')'

                ws['A20'] = '发票快件查询 资料不清晰请拨打电话:' + str(kpcxdh)
                # msexcelworksheet.Rows[inttostr(3)].AutoFit

                # ws['A9'].font.color = clred
                ws['A9'] = '★' + jjdz
                ws['C18'] = d.get('gdry')

                if d.get('lxdh') != '' and d.get('lxdh') != None:
                    ws['G18'] = d.get('lxdh')
                else:
                    t = run_sql(f"select zjdh from ywrylx where (ryxm='{d.get('gdry')}')")
                    if len(t) > 0:
                        ws['G18'] = t[0].get('zjdh')

                ws['B24'] = d.get('hyd')
                if ywsb == 1:
                    ws['c25'] = d.get('cz')
                    ws['A16'] = '统一社会信用代码'
                    ws['A19'] = str(d.get('zsl')) + '%'
                    ws['A11'] = '合同签订日期:'
                    ws['B11'] = (chyrq - ywcs).strftime('%Y-%m-%d')
                else:
                    ws['G21'] = str(d.get('zsl')) + '%'

                # c1 = c
                # if uv == '1':
                #     msexcelworksheet = null
                #     msexcelworkbook = null
                #     msexcelworksheet = msexcel.workbooks[1].worksheets[2]
                #     msexcelworksheet.Activate
                #     msexcel.visible = false
                #     ws['G4'] = cx1
                #     msexcelworksheet.Rows[inttostr(7)].AutoFit
                #     ws['J7'] = d.get('htxs')
                #     ws['c7'] = d.get('htxs')
                #     ws['c8'] = d.get('sccj')
                #     ws['c9'] = floattostr(d.get('zsl').Asfloat) + '%'
                #     ws['c10'] = chyrq
                #     msexcelworksheet.Rows[inttostr(11)].AutoFit
                #     if yssb == '1':
                #         ws['C11'].font.color = clred
                #         ws['J11'] = jjdz
                #         ws['c11'] = jjdz
                #     else:
                #         ws['J11'] = jjdz
                #         ws['c11'] = jjdz
                save_dir = config.tmp_path
                if not os.path.exists(save_dir):
                    os.makedirs(save_dir, exist_ok=True)
                
                file_path = get_unique(16) + '.xlsx'
                file_name = f"{zb1}{cx}{cx1z}{i1}.xlsx"
                save_path = os.path.join(save_dir, file_path)
                wb.save(save_path)

        return json_result(1, '操作成功', {'path': file_path, 'name': file_name})
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()