import asyncio
from concurrent.futures import ThreadPoolExecutor
from any import *  
from .model import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors, Alignment
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
import traceback
import io # 需要导入 io 模块来处理内存流
from openpyxl.utils import coordinate_to_tuple
from email import errors
from math import e
from pdb import run
from webbrowser import get
from any import *
from .model import *
from sqlalchemy.sql import func,not_,or_,and_
import time
import json,os
from .__default__ import get_user_path,module_xxck_new,module_share_new
from openpyxl.drawing.image import Image as Image_Get
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, Alignment
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from PIL import Image as PILImage  # 重命名避免冲突
# jpeg图片出错的解决方法
from PIL import JpegImagePlugin
# 第2，3，4工作表英文品名部分添加未完成（表tspmb没有）
# 创建线程池执行器
# executor = ThreadPoolExecutor(max_workers=4)
# 获取事件循环
# loop = asyncio.get_event_loop()
import copy
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries, coordinate_to_tuple


def _fmt_ref(min_c, min_r, max_c, max_r):
    a = f"{get_column_letter(min_c)}{min_r}"
    b = f"{get_column_letter(max_c)}{max_r}"
    return a if a == b else f"{a}:{b}"


def _shift_ref(ref, idx, n):
    min_c, min_r, max_c, max_r = range_boundaries(ref)
    if max_r < idx:
        return _fmt_ref(min_c, min_r, max_c, max_r)
    if min_r >= idx:
        return _fmt_ref(min_c, min_r + n, max_c, max_r + n)
    # 插入点在区域内部：扩展底边
    return _fmt_ref(min_c, min_r, max_c, max_r + n)


def _shift_refs(refs, idx, n):
    return " ".join(_shift_ref(r, idx, n) for r in str(refs).split() if r)


def insert_row_keep_layout(ws, idx, n=1):
    # 1) 先记录需要手动修复的结构
    old_merges = [str(r) for r in ws.merged_cells.ranges]
    old_tables = {k: t.ref for k, t in ws.tables.items()}
    old_filter = ws.auto_filter.ref if ws.auto_filter and ws.auto_filter.ref else None
    old_dv = [(dv, str(dv.sqref)) for dv in getattr(ws.data_validations, "dataValidation", [])]

    cf = ws.conditional_formatting
    old_cf = list(getattr(cf, "_cf_rules", {}).items()) if hasattr(cf, "_cf_rules") else []

    # 2) 记录模板行样式（默认复制插入点上一行）
    src_row = max(1, idx - 1)
    style_map = {c: copy(ws.cell(src_row, c)._style) for c in range(1, ws.max_column + 1)}
    src_dim = ws.row_dimensions[src_row]

    # 3) 插入
    ws.insert_rows(idx, n)

    # 4) 给新插入行补样式，避免“格式断层”
    for r in range(idx, idx + n):
        rd = ws.row_dimensions[r]
        rd.height = src_dim.height
        rd.hidden = src_dim.hidden
        rd.outlineLevel = src_dim.outlineLevel
        rd.collapsed = src_dim.collapsed
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c)._style = copy(style_map[c])

    # 5) 合并单元格重建
    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))
    for mr in old_merges:
        ws.merge_cells(_shift_ref(mr, idx, n))

    # 6) 图片/图表锚点下移
    def _shift_obj(obj):
        anc = obj.anchor
        if isinstance(anc, str):
            row, col = coordinate_to_tuple(anc)
            if row >= idx:
                obj.anchor = f"{get_column_letter(col)}{row + n}"
            return
        if hasattr(anc, "_from"):
            if anc._from.row + 1 >= idx:
                anc._from.row += n
            if hasattr(anc, "_to") and anc._to and anc._to.row + 1 >= idx:
                anc._to.row += n

    for img in getattr(ws, "_images", []):
        _shift_obj(img)
    for ch in getattr(ws, "_charts", []):
        _shift_obj(ch)

    # 7) Table / 自动筛选 / 数据验证 / 条件格式
    for name, ref in old_tables.items():
        ws.tables[name].ref = _shift_ref(ref, idx, n)

    if old_filter:
        ws.auto_filter.ref = _shift_ref(old_filter, idx, n)

    for dv, ref in old_dv:
        dv.sqref = _shift_refs(ref, idx, n)

    if old_cf:
        cf._cf_rules.clear()
        for ref_obj, rules in old_cf:
            new_ref = _shift_refs(str(ref_obj), idx, n)
            for rule in rules:
                cf.add(new_ref, rule)



def safe_write(ws, coord, value, num_format=None):
    """
    智能穿透写入：如果目标单元格被合并了，自动找到该合并区域的左上角主单元格进行写入。
    完美解决 MergedCell is read-only 报错！
    """
    row, col = coordinate_to_tuple(coord)
    target_cell = ws[coord]
    
    # 检查该坐标是否在某个合并单元格范围内
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            # 如果被合并了，就把目标切换为该区域的左上角真实单元格
            target_cell = ws.cell(row=mr.min_row, column=mr.min_col)
            break
            
    target_cell.value = value
    if num_format:
        target_cell.number_format = num_format
def offset_img(img, col, row, x_pad=4, y_pad=25):
    """精确设置图片位置，偏移量以万为单位进行微调吧，具体计算公式太麻烦了
    row column 的索引都是从0开始的，我这里要把图片插入到单元格A17
    """
    p2e = pixels_to_EMU
    h, w = img.height, img.width
    # 图像等比例缩放因子
    resize_factor = 0.8
    w_h_ratio = w/h
    resize_H = int(resize_factor * h)
    resize_W = int(resize_factor * resize_H * w_h_ratio)
    #
    # x_pad = 4
    # y_pad = 25
    # 注意这里的行、列索引从0开始, 所以需要减1
    size = XDRPositiveSize2D(p2e(w), p2e(h))
    marker = AnchorMarker(col=col-1, colOff=pixels_to_EMU(x_pad), row=row-1, rowOff=pixels_to_EMU(y_pad))
    img.anchor = OneCellAnchor(_from=marker, ext=size)


def copy_row_style(source_row_cells, target_row_idx):
    """
    将 source_row_cells (一行单元格对象) 的格式复制到 target_row_idx (目标行号)
    """
    ws: Workbook = source_row_cells[0].parent  # 获取工作表对象
    # 修正点：enumerate 默认从 0 开始，但 Excel 列是从 1 开始的 (A=1)
    # 所以我们需要 col_idx + 1
    for col_idx, source_cell in enumerate(source_row_cells):
        
        # 获取目标单元格
        # 注意：openpyxl 的 cell 方法 column 参数接受整数 (1-based)
        target_cell = ws.cell(row=target_row_idx, column=col_idx + 1)
        
        # 1. 复制值 (如果需要)
        # target_cell.value = source_cell.value 
        
        # 2. 复制样式
        if source_cell.has_style:
            # 关键点：必须使用 copy.copy() 将 StyleProxy 转换为真实对象
            # 否则直接赋值会报 'unhashable type: StyleProxy' 错误
            target_cell.font = copy.copy(source_cell.font)
            target_cell.border = copy.copy(source_cell.border)
            target_cell.fill = copy.copy(source_cell.fill)
            target_cell.number_format = copy.copy(source_cell.number_format)
            target_cell.alignment = copy.copy(source_cell.alignment)

# --- 核心逻辑函数 ---
def _generate_excel_logic(recordset):
#     """
#     生成 Excel 并返回字节流 (Bytes)，而不是保存在服务器上。
#     """

    hr = '2'
    r_path = os.path.join(config.data_upload_path, 'template')
    if hr == '1':
        template_file = 'bestorebips.xlsx'
    else:
        template_file = 'bestorebips1.xlsx'

    fn = os.path.join(r_path, template_file)
    if not os.path.exists(fn):
        return json_result(-1, f"未找到报表模板: {template_file}")
    wb = load_workbook(fn)
    ws = wb.active
  

    
    s = Session()
    try:
        
        # --- 1. 数据处理 ---
        main_info={}
        merged_products = []
        customs_forms = []
        d=s.query(bgmxd).filter(bgmxd.rid==recordset.get('rids')[0]).first()
        if d:
            main_info = alchemy_object_to_dict(d)
        d=s.query(bgmxdhbcp).filter(bgmxdhbcp.pid==recordset.get('rids')[0]).all()
        if d:
            merged_products = alchemy_object_list_to_dict(d)
        
        d=s.query(bgmxdsheet2).filter(bgmxdsheet2.pid==recordset.get('rids')[0]).all()
        if d:
            customs_forms = alchemy_object_list_to_dict(d)
    
        # 检查是否为RMB客户
        if main_info.get('RMBkh', '是') == '否':
            return {'code': -1, 'msg': '此客人为USD客人，请选USD格式', 'data': ''}
        
        # 初始化变量
        ls = ''
        hr = '2'
        hbdm = main_info.get('hbdm', '')
        if hbdm == 'RMB￥':
            hbdm = 'RMB'
        
        # 获取报关公司
        gs=''
        gs1 = main_info.get('bggs', '')
        gsyw = ''
        # 数据库查询
        if gs1:
            j = run_sql(f"SELECT  ywmc FROM wfgs WHERE (wfgs ='{main_info.get('bggs', '')}') limit 1")
           
        
        # 处理日期
        invoice_date = main_info.get('fprq', '')
        qsn = invoice_date[:4] if len(invoice_date) >= 4 else ''
        qsy1 = invoice_date[5:7] if len(invoice_date) >= 7 else ''
        qsr = invoice_date[8:10] if len(invoice_date) >= 10 else ''
        
        # 月份转换
        month_map = {
            '01': 'JAN', '02': 'FEB', '03': 'MAR', '04': 'APR',
            '05': 'MAY', '06': 'JUN', '07': 'JUL', '08': 'AUG',
            '09': 'SEP', '10': 'OCT', '11': 'NOV', '12': 'DEC'
        }
        qsy = month_map.get(qsy1, '')
        
        # 初始化统计变量
        a1 = 0
        bgxshb = 0
        bgbgzsl = 0
        bgbgzmz = 0
        bgbgzjz = 0
        bgbgztj = 0
        bgbgzje = 0
        i6 = 0
        bgdmzhj = 0
        bgdjzhj = 0
        
        # 处理合并产品数据
        for product in merged_products:
            customer_name = main_info.get('khmc', '')
            if 'BEST PRICE LLC' in customer_name:
                bgdmzhj += round(product.get('zmz', '0') * 10) / 10
                bgdjzhj += round(product.get('zjz', '0') * 10) / 10
            else:
                bgdmzhj += round(product.get('zmz', '0') * 100) / 100
                bgdjzhj += round(product.get('zjz', '0') * 100) / 100
            
            bgxshb += float(product.get('chxs', '0'))
            bgbgzsl += product.get('chsl', '0')
            bgbgzmz += product.get('zmz', '0')
            bgbgzjz += product.get('zjz', '0')
            bgbgztj += product.get('ztj', '0')
            bgbgzje += product.get('wxzj', '0')
        
#         # --- 2. 生成 Excel (在内存中) ---
        output = io.BytesIO()

        ws = wb.worksheets[0]    
           
        # ws.title = "报关单"
        i=+1
        cpbh = main_info.get('wfgs', '') + '报关专用章'
        g = run_sql(f"select tpmc from tpzx where (cpbh='{cpbh}') and (LENGTH(tpmc) > 5)")
        if len(g)>0 and g[0].get('tpmc','') != '' and g[0].get('tpmc','') != None and g[0].get('tpmc','') != '[]':
            cpbh = g[0].get('tpmc','')
            Photo = json.loads(str(cpbh))
            if Photo != None:
                file_path = Photo[0]['src']
                fn = os.path.join(config.data_upload_path, str(file_path))
                if (os.path.exists(fn)):
                    img = Image_Get(fn) #选择图片
                    # img.width = 150  # 设置图像宽度
                    # img.height = 68  # 设置图像高度
                    col_width = (ws.column_dimensions['D'].width + ws.column_dimensions['E'].width)*7
                    row_height = 100 # (ws.row_dimensions[i].height)*1.3333
                    img.width=col_width-4 # 转换为像素
                    img.height=row_height-4 # 转换为像素
                    x_offset = 15 # (col_width-img.width)/2
                    y_offset = 10 # (row_height-img.height)/2
                    col = 5 # E列
                    row = 25
                    offset_img(img, col, row, x_offset, y_offset) #col为列位置，row为行位置,x_offset为左边边距,y_offset为上边距
                    ws.add_image(img)  #添加图片
        cpbh = main_info.get('wfgs', '') + '蓝章'
        j = run_sql(f"select sb, dyyy, bgh from tpzx where (cpbh='{cpbh}')")
        # 3. 初始化变量

        gs1 = ""
        sb_val = ""
        dyyy_val = ""
        bgh_val = ""

        # 4. 判断是否有结果
        if j and len(j) > 0:
            # 获取第一行数据
            row = j[0]
            
            # 提取字段值，防止为 None
            sb_val = row.get('sb') or ""
            dyyy_val = row.get('dyyy') or ""
            bgh_val = row.get('bgh') or ""

            # 5. 拼接字符串
            # 注意：这行代码必须缩进在 if 内部，或者确保变量已定义
            gs1 = f"{sb_val}\n{dyyy_val}\n{bgh_val}"

        ws['A4'] = gs1
        ws.row_dimensions[4].height = 39
        ws['A6'] = main_info.get('khmc', '')
        ws['A8'] = gs1
        ws.row_dimensions[8].height = 39
        ws['B4'] = main_info.get('cyka', '')
        ws['B6'] = main_info.get('ysfs', '')
        ws['D6'] = main_info.get('ysgj', '')
        ws['F6'] = main_info.get('tdh', '')
        ws['F8'] = main_info.get('xkzh', '')
        ws['A10'] = main_info.get('ysfp', '')
        ws['B10'] = main_info.get('kagjy', '')
        ws['D10'] = main_info.get('kagjy', '')
        ws['F10'] = main_info.get('mdka', '')
        ws['I10'] = main_info.get('cyka', '')
        ws['E12'] = main_info.get('jgtk', '')
        
        # CNF运费处理
        if main_info.get('jgtk', '') == 'CNF':
            ws['F12'] = merged_products.get('CNFyf$', '')
        
        ws['B12'] = main_info.get('bgxshb', '')
        ws['C12'] = str(bgdmzhj)
        ws['D12'] = str(bgdjzhj)
        ws['A12'] = main_info.get('bzzl', '')
        
        # VGM标记唛码
        vgm_weight = main_info.get('hgzl', 0) + bgdmzhj
        ws['A15'] = f'标记唛码及备注                VGM：{vgm_weight:.2f}KGS'
        
        # 唛头信息
        mark_head = main_info.get('wxmt', '')
        if mark_head:
            ws['N16'] = mark_head
            ws['A16'] = mark_head
        else:
            ws['A16'] = 'N/M'
        
        # 填充报关单项信息
        start_row = 18
        source_row = ws[start_row]
        i=1
        num = len(customs_forms)
        ws.insert_rows(start_row, num)
        while i <= num:
        
            target_idx = 18 + num
            copy_row_style( source_row,target_idx)
            i=i+1
        
        # insert_row_keep_layout(ws, 19, 1)
        logger.info(f"从哪行插入: {start_row+1}")
        logger.error(f"报关单项数量: {num}")
        for idx, form in enumerate(customs_forms):
            row_num = start_row + idx + 1   # 在当前行插入新行，原有内容下移
            safe_write(ws, f'A{row_num}', f"{idx + 1} {form.get('hgbm', '')}")
            safe_write(ws, f'B{row_num}', form.get('zwpm', ''))

            # 数量信息
            unit = form.get('hgjldw', '')
            if unit != 'KGS' and unit != '千克':
                quantity_str = f"{form.get('chsl', 0):.0f}{unit}/{form.get('zjz', 0)}千克"
            else:
                quantity_str = f"{form.get('chsl', 0):.0f}/{form.get('zjz', 0)}千克"
            safe_write(ws, f'D{row_num}', quantity_str)
            
            safe_write(ws, f'H{row_num}', '中国')
            
            price = form.get('mjzj', 0)
            if price > 0:
                safe_write(ws, f'F{row_num}', price, '$#,##0.000')
            else:
                safe_write(ws, f'F{row_num}', '')
            safe_write(ws, f'J{row_num}', form.get('hyd', ''))
            safe_write(ws, f'K{row_num}', main_info.get('zmxz', ''))
            # ws.insert_rows(row_num+1)
        
        # 底部信息
        final_row = start_row + len(customs_forms) + 3
        container_info = f"箱/封号：{main_info.get('jzxh', '')}/{main_info.get('fh', '')}"
        safe_write(ws, f'B{final_row}', container_info)
        
        bl_no = f"提单号：{main_info.get('tdh', '')}"
        # safe_write(ws, f'B{final_row + 1 }', bl_no)

        thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        ws = wb.worksheets[1]       
        # ws.title = "发票"
        cpbh = main_info.get('wfgs', '') + '蓝章名'
        logger.error("123shjakdh")
        g = run_sql(f"select tpmc from tpzx where (cpbh='{cpbh}') and (LENGTH(tpmc) > 5)")
        if len(g)>0 and g[0].get('tpmc','') != '' and g[0].get('tpmc','') != None and g[0].get('tpmc','') != '[]':
            cpbh = g[0].get('tpmc','')
            Photo = json.loads(str(cpbh))
            if Photo != None:
                file_path = Photo[0]['src']
                fn = os.path.join(config.data_upload_path, str(file_path))
                if (os.path.exists(fn)):
                    img = Image_Get(fn) #选择图片
                    # img.width = 150  # 设置图像宽度
                    # img.height = 68  # 设置图像高度
                    col_width = (ws.column_dimensions['D'].width + ws.column_dimensions['E'].width)*7
                    row_height = 100 # (ws.row_dimensions[i].height)*1.3333
                    img.width=col_width-4 # 转换为像素
                    img.height=row_height-4 # 转换为像素
                    x_offset = 15 # (col_width-img.width)/2
                    y_offset = 10 # (row_height-img.height)/2
                    col = 6 # F列
                    row = 48
                    logger.error("123456789djajkh")
                    offset_img(img, col, row, x_offset, y_offset) #col为列位置，row为行位置,x_offset为左边边距,y_offset为上边距
                    ws.add_image(img)  #添加图片
                  
         # 基本信息
        safe_write(ws, f'A6',main_info.get('gsyw', ''))
        safe_write(ws, f'C15', main_info.get('ysfp', ''))
        date_str = f"{qsy}.{qsr},{qsn}"
        safe_write(ws, f'C17', date_str)
        safe_write(ws, f'B18', main_info.get('cyka', ''))
        safe_write(ws, f'B21', main_info.get('mdka', ''))
        safe_write(ws, f'F22', main_info.get('jhfs', ''))
        safe_write(ws, f'E25', f"{main_info.get('jgtk', '')} {main_info.get('cyka', '')}")
        safe_write(ws, f'G25', hbdm)
        
        # 收货人处理
        consignee = main_info.get('shr', '')
        if not consignee:
            safe_write(ws, f'A10', 'BEST PRICE')
        elif 'BEST PRICE' in main_info.get('wfgs', ''):
            safe_write(ws, f'A10', main_info.get('khmc', ''))
        else:
            safe_write(ws, f'A10', main_info.get('shry', ''))
        
        # 填充合并产品信息
        start_row = 26
        for idx, product in enumerate(merged_products):
            row_num = start_row + idx + 1
            
            # 英文品名
            english_desc = product.get('ywhj', '') or product.get('ywpm', '')
            if english_desc:
                safe_write(ws, f'B{row_num}', english_desc)
            
            # 其他信息
            safe_write(ws, f'C{row_num}', product.get('chsl', 0))
            safe_write(ws, f'D{row_num}', f"{product.get('jldw', '')}S")
            safe_write(ws, f'E{row_num}', hbdm)
            safe_write(ws, f'F{row_num}', f"{product.get('price', 0):.3f}")
            safe_write(ws, f'G{row_num}', hbdm)
            safe_write(ws, f'H{row_num}', product.get('wxzj', 0))
        
        # 唛头信息合并
        mark_head = main_info.get('wxmt', '')
        if mark_head:
            ws.merge_cells(f'A27:A{start_row + len(merged_products) + 1}')
            ws['A27'] = mark_head
        else:
            ws.merge_cells(f'A27:A{start_row + len(merged_products) + 1}')
            ws['A27'] = 'N/M'

        ws = wb.worksheets[2]       
        # ws.title = "装箱单"
        cpbh = main_info.get('wfgs', '') + '蓝章名'
        g = run_sql(f"select tpmc from tpzx where (cpbh='{cpbh}') and (LENGTH(tpmc) > 5)")
        if len(g)>0 and g[0].get('tpmc','') != '' and g[0].get('tpmc','') != None and g[0].get('tpmc','') != '[]':
                cpbh = g[0].get('tpmc','')
                Photo = json.loads(str(cpbh))
                if Photo != None:
                    file_path = Photo[0]['src']
                    fn = os.path.join(config.data_upload_path, str(file_path))
                    if (os.path.exists(fn)):
                        img = Image_Get(fn) #选择图片
                        # img.width = 150  # 设置图像宽度
                        # img.height = 68  # 设置图像高度
                        col_width = (ws.column_dimensions['D'].width + ws.column_dimensions['E'].width)*7
                        row_height = 100 # (ws.row_dimensions[i].height)*1.3333
                        img.width=col_width-4 # 转换为像素
                        img.height=row_height-4 # 转换为像素
                        x_offset = 15 # (col_width-img.width)/2
                        y_offset = 10 # (row_height-img.height)/2
                        col = 4 # F列
                        row = 52
                        offset_img(img, col, row, x_offset, y_offset) #col为列位置，row为行位置,x_offset为左边边距,y_offset为上边距
                        ws.add_image(img)  #添加图片
               
     # 基本信息
        ws['A6'] = gsyw
        ws['C15'] = main_info.get('ysfp', '')
        date_str = f"{qsy}.{qsr},{qsn}"
        ws['C17'] = date_str
        ws['B18'] = main_info.get('cyka', '')
        ws['B21'] = main_info.get('mdka', '')
        ws['D22'] = main_info.get('jhfs', '')
        
        # 收货人处理
        consignee = main_info.get('shr', '')
        if not consignee:
            safe_write(ws, f'A10', 'BEST PRICE')
        elif 'BEST PRICE' in main_info.get('wfgs', ''):
            safe_write(ws, f'A10', main_info.get('khmc', ''))
        else:
            safe_write(ws, f'A10', main_info.get('shry', ''))
        
        # 填充合并产品信息
        start_row = 26
        for idx, product in enumerate(merged_products):
            row_num = start_row + idx + 1
            
            # 英文品名
            english_desc = product.get('ywhj', '') or product.get('ywpm', '')
            if english_desc:
                safe_write(ws, f'B{row_num}', english_desc)
            
            # 数量和重量信息
            safe_write(ws, f'C{row_num}', product.get('chxs', 0))
            
            # 按客户类型处理重量精度
            customer_name = main_info.get('khmc', '')
            if 'BEST PRICE LLC' in customer_name:
                gross_weight = round(product.get('zmz', 0) * 10) / 10
                net_weight = round(product.get('zjz', 0) * 10) / 10
            else:
                gross_weight = round(product.get('zmz', 0) * 100) / 100
                net_weight = round(product.get('zjz', 0) * 100) / 100
            
            safe_write(ws, f'D{row_num}', gross_weight)
            safe_write(ws, f'E{row_num}', net_weight)
            safe_write(ws, f'F{row_num}', product.get('ztj', 0))
        
        # 唛头信息合并
        mark_head = main_info.get('wxmt', '')
        if mark_head:
            ws.merge_cells(f'A27:A{start_row + len(merged_products) + 1}')
            ws['A27'] = mark_head
        else:
            ws.merge_cells(f'A27:A{start_row + len(merged_products) + 1}')
            ws['A27'] = 'N/M'

        ws = wb.worksheets[3]
        # ws.title = "商业发票"

        cpbh =main_info.get('khmc', '') 
        g = run_sql(f"select tpmc from kh where (company_name='{main_info.get('company_name', '')}') and (LENGTH(tpmc) > 5)")
        if len(g)>0 and g[0].get('tpmc','') != '' and g[0].get('tpmc','') != None and g[0].get('tpmc','') != '[]':
                cpbh = g[0].get('tpmc','')
                Photo = json.loads(str(cpbh))
                if Photo != None:
                    file_path = Photo[0]['src']
                    fn = os.path.join(config.data_upload_path, str(file_path))
                    if (os.path.exists(fn)):
                        img = Image_Get(fn) #选择图片
                        # img.width = 150  # 设置图像宽度
                        # img.height = 68  # 设置图像高度
                        col_width = (ws.column_dimensions['D'].width + ws.column_dimensions['E'].width)*7
                        row_height = 100 # (ws.row_dimensions[i].height)*1.3333
                        img.width=col_width-4 # 转换为像素
                        img.height=row_height-4 # 转换为像素
                        x_offset = 15 # (col_width-img.width)/2
                        y_offset = 10 # (row_height-img.height)/2
                        col = 2# F列
                        row = 31
                        offset_img(img, col, row, x_offset, y_offset) #col为列位置，row为行位置,x_offset为左边边距,y_offset为上边距
                        ws.add_image(img)  #添加图片
        cpbh = main_info.get('wfgs', '') + '蓝章名'
        g = run_sql(f"select tpmc from tpzx where (cpbh='{cpbh}') and (LENGTH(tpmc) > 5)")
        if len(g)>0 and g[0].get('tpmc','') != '' and g[0].get('tpmc','') != None and g[0].get('tpmc','') != '[]':
                cpbh = g[0].get('tpmc','')
                Photo = json.loads(str(cpbh))
                if Photo != None:
                    file_path = Photo[0]['src']
                    fn = os.path.join(config.data_upload_path, str(file_path))
                    if (os.path.exists(fn)):
                        img = Image_Get(fn) #选择图片
                        # img.width = 150  # 设置图像宽度
                        # img.height = 68  # 设置图像高度
                        col_width = (ws.column_dimensions['D'].width + ws.column_dimensions['E'].width)*7
                        row_height = 100 # (ws.row_dimensions[i].height)*1.3333
                        img.width=col_width-4 # 转换为像素
                        img.height=row_height-4 # 转换为像素
                        x_offset = 15 # (col_width-img.width)/2
                        y_offset = 10 # (row_height-img.height)/2
                        col = 9 # F列
                        row = 51
                        offset_img(img, col, row, x_offset, y_offset) #col为列位置，row为行位置,x_offset为左边边距,y_offset为上边距
                        ws.add_image(img)  #添加图片



        ws['I2'] = main_info.get('ysfp', '')
        date_str = f"{qsy}.{qsr},{qsn}"
        ws['I3'] = date_str
        ws['G9'] = f"{main_info.get('jgtk', '')} {main_info.get('cyka', '')}"
        
        # 收货人处理
        consignee = main_info.get('shr', '')
        if not consignee:
            ws['C6'] = 'BEST PRICE'
        elif 'BEST PRICE' in main_info.get('wfgs', ''):
            ws['C6'] = main_info.get('khmc', '')
        else:
            ws['C6'] = main_info.get('shry', '')
        
        # 填充合并产品信息
        start_row = 9
        for idx, product in enumerate(merged_products):
            row_num = start_row + idx + 1
            
            # 英文品名
            english_desc = product.get('ywhj', '') or product.get('ywpm', '')
            if english_desc:
                safe_write(ws, f'A{row_num}', english_desc)
                safe_write(ws, f'M{row_num}', english_desc)
            else:
                safe_write(ws, f'A{row_num}', product.get('ywpm', ''))
                safe_write(ws, f'M{row_num}', product.get('ywpm', ''))
            
            # 其他信息
            safe_write(ws, f'E{row_num}', product.get('chsl', 0))
            safe_write(ws, f'F{row_num}', f"{product.get('jldw', '')}S")
            safe_write(ws, f'G{row_num}', hbdm)
            safe_write(ws, f'H{row_num}', f"{product.get('price', 0):.3f}")
            safe_write(ws, f'I{row_num}', hbdm)
            safe_write(ws, f'J{row_num}', product.get('wxzj', 0))
        
        if row_num<= 29:
        
            start_row = 29
        elif row_num > 29: 
            start_row =  start_row + idx + 1
        final_row = start_row 
        logger.error("Final row for bottom info: " + str( len(merged_products))) 
        logger.error( str(len(merged_products)))
        safe_write(ws, f'I{final_row + 1}', hbdm)
        safe_write(ws, f'D{final_row + 3}', main_info.get('zyqx', ''))
        safe_write(ws, f'E{final_row + 4}', main_info.get('cyka', ''))
        safe_write(ws, f'I{final_row + 4}', main_info.get('mdka', ''))
        safe_write(ws, f'A{final_row + 8 }', main_info.get('jhfs', ''))
        
        # 唛头和备注信息
        mark_head = main_info.get('wxmt', '')
        if mark_head:
            safe_write(ws, f'D{final_row + 9}', mark_head)
        else:
            safe_write(ws, f'D{final_row + 9}', 'N/M')
        
        safe_write(ws, f'C{final_row + 11}', main_info.get('bza', ''))



        output.close()
        path = config.tmp_path
        os.makedirs(path, exist_ok=True) # 确保目录存在
        report_rid = get_uuid() # 假设这是一个生成唯一ID的函数
        output_filename = f'{main_info.get("ysfp")}BIPS.xlsx'
        full_output_path = os.path.join(path, output_filename)
        s_path = config.tmp_path
        # file_name = f"{master.fphm or 'Unknown'}优景_INV_pack_sales.xlsx"
        # wb.save(os.path.join(s_path, file_name))
        # return json_result(1, '导出成功', file_name)

        
        wb.save(full_output_path)

        return {'code': 1, 'msg': '报表生成成功', 'data': output_filename}

    except Exception as e:
        logger.error(f"生成Excel失败: {str(e)}")
        logger.error(traceback.format_exc())
        return {'code': -1, 'msg': f'生成Excel失败: {str(e)}', 'data': ''}
    finally:
        s.close()

# --- API路由 ---
@any_route('/api/saier/export/detailed/INVpacksales/list2', methods=['POST'])
@require_token
async def inv_pack_sales_export(request):
    """
    导出INV pack sales详细数据
    """
    j = await request.json()
    try:
        result = _generate_excel_logic(j)
        logger.error(result)
        if result['code'] == 1:
            return json_result(1, result['msg'], result['data'])
        else:
            return json_result(result['code'], result['msg'], {})
    except Exception as e:
        logger.error(f"API调用失败: {str(e)}")
        logger.error(traceback.format_exc())
        return json_result(-1, f'API调用失败: {str(e)}', {})
    