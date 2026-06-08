import asyncio
from concurrent.futures import ThreadPoolExecutor
from any import *
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import os
import traceback
from datetime import datetime

# 用于异步执行耗时任务的线程池
executor = ThreadPoolExecutor(max_workers=1)


def _generate_excel_on_thread(qsrq, jsrq, user_name):
   
    """
    在后台线程中执行的核心逻辑，包括数据库查询和 Excel 生成。
    这个函数是同步的，因为它将在一个单独的线程中运行。
    """
    # --- 1. 参数校验 ---
    # 校验日期格式是否有效，避免将空字符串等无效值传递给数据库
    if not qsrq or not jsrq:
        return {'code': -1, 'msg': '缺少必要参数：开始日期或结束日期不能为空。', 'data': None}
    
    # 尝试将字符串解析为日期对象，以验证其格式
    try:
        datetime.strptime(qsrq, '%Y-%m-%d')  # 假设日期格式为 YYYY-MM-DD
        datetime.strptime(jsrq, '%Y-%m-%d')
        if qsrq > jsrq:
             return {'code': -1, 'msg': '开始日期不能晚于结束日期。', 'data': None}
    except ValueError:
        return {'code': -1, 'msg': '日期格式无效，请使用 YYYY-MM-DD 格式。', 'data': None}


    # --- 2. 权限检查 (模拟 Delphi 代码) ---
    # 检查用户是否为经理、总监或采购
    s = Session()
    try:
        is_manager_query = """
            SELECT 1 FROM sys_user
            WHERE username = :username AND (
                position LIKE '%经理%' OR position LIKE '%总%' OR position LIKE '%采购%' OR
                memo LIKE '%经理%' OR memo LIKE '%总%' OR memo LIKE '%采购%'
            )
        """
        is_manager_result = s.execute(is_manager_query, {'username': user_name}).fetchone()
        
        if not is_manager_result:
            return {'code': -1, 'msg': '权限不足，无法生成报表。', 'data': None}

        # --- 3. 数据查询与处理 (核心逻辑) ---
        
        # a. 查询符合条件的主表数据 (bj)
        # 与 Delphi 代码逻辑一致：查询日期范围内、已通过的报价单
        main_query = """
            SELECT ywry, rid
            FROM bj
            WHERE dateis >= :qsrq AND dateis <= :jsrq AND bjjg = '通过'
        """
        main_results = s.execute(main_query, {'qsrq': qsrq, 'jsrq': jsrq}).fetchall()

        if not main_results:
            return {'code': -1, 'msg': '该日期范围内没有符合条件的报价数据', 'data': None}

        # b. 提取所有主表ID，用于查询子表
        main_rids = [row.rid for row in main_results]

        # c. 批量查询子表数据 (bjsheet)
        # 与 Delphi 代码逻辑一致：获取所有相关报价单的明细
        detail_query = """
            SELECT pid, bjry, wxxz, bjbjwyzd, cply1, wxbm
            FROM bjsheet
            WHERE pid IN :main_rids
        """
        detail_results = s.execute(detail_query, {'main_rids': tuple(main_rids)}).fetchall()

        # d. 内存聚合数据 (替代 Delphi 的 TStringList.IndexOf)
        # 使用字典来高效地按业务员聚合数据
        # 结构: stats = {ywry: {stat_metrics}}
        stats = {}

        for detail_row in detail_results:
            pid = detail_row.pid
            bjry = detail_row.bjry
            wxxz = detail_row.wxxz
            bjbjwyzd = detail_row.bjbjwyzd
            cply1 = detail_row.cply1
            wxbm = detail_row.wxbm

            # 寻找对应的业务员 (ywry) -> (pid, bjry) -> ywry
            corresponding_main = next((mr for mr in main_results if mr.rid == pid), None)
            if not corresponding_main:
                continue # 数据不一致，跳过
            ywry = corresponding_main.ywry

            # 如果业务员不在统计字典中，初始化
            if ywry not in stats:
                stats[ywry] = {
                    'cgbj': 0, 'wxbj': 0, 'wxbmbj': '', 'wxbm_dict': {},
                    'FD': 0, 'FD1': 0, 'XJ': 0, 'XJ1': 0,
                    'CGTJ': 0, 'CGTJ1': 0, 'WXTJ': 0, 'WXTJ1': 0,
                    'KHZX': 0, 'KHZX1': 0, 'YPLL': 0, 'YPLL1': 0,
                    'wxbm_set': set() # 用于去重 wxbm
                }

            node = stats[ywry]
            node['cgbj'] += 1 # 报价产品数

            is_selected = (wxxz == '是')
            if is_selected:
                node['wxbj'] += 1 # 外销选中数

            # 处理外销部门统计 (wxbm, bjbjwyzd)
            if is_selected and bjbjwyzd and wxbm:
                wxbm_str = str(wxbm).strip()
                if wxbm_str not in node['wxbm_set']:
                    node['wxbm_set'].add(wxbm_str)
                    node['wxbm_dict'][wxbm_str] = node['wxbm_dict'].get(wxbm_str, 0) + 1

            # 来源统计 (cply1)
            cply_lower = cply1.lower() if cply1 else ''
            if '返单' in cply_lower:
                node['FD'] += 1
                if is_selected: node['FD1'] += 1
            if '询价' in cply_lower:
                node['XJ'] += 1
                if is_selected: node['XJ1'] += 1
            if '采购推荐' in cply_lower:
                node['CGTJ'] += 1
                if is_selected: node['CGTJ1'] += 1
            if '外销推荐' in cply_lower:
                node['WXTJ'] += 1
                if is_selected: node['WXTJ1'] += 1
            if '客户自选' in cply_lower:
                node['KHZX'] += 1
                if is_selected: node['KHZX1'] += 1
            if '样品录入' in cply_lower:
                node['YPLL'] += 1
                if is_selected: node['YPLL1'] += 1

        # --- 4. 生成 Excel 文件 ---
        wb = Workbook()
        ws = wb.active
        ws.title = "报价统计报表"

        # 设置标题行
        headers = [
            '姓名', '报价产品数', '外销选中数', '外销报价命中率', '返单', '询价',
            '采购推荐', '外销推荐', '客户自选', '样品录入', '外销合同选中数', 
            '外销合同命中率', '采购合同选中数', '采购合同命中率'
        ]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 设置列宽
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14.3

        # 从第二行开始写入数据
        current_row = 2
        for ywry, data in stats.items():
            # 计算命中率
            hit_rate = round(data['wxbj'] / data['cgbj'], 3) if data['cgbj'] > 0 else 0.000
            
            # 构建外销选中数的详细信息 (C列)
            wxbmbj_parts = [str(data['wxbj'])] # 先添加总数
            # 添加 wxbm 和数量的列表
            for wxbm_key, count in data['wxbm_dict'].items():
                 # 限制显示数量，避免内容过长
                if len(wxbmbj_parts) > 5: 
                    wxbmbj_parts.append("...")
                    break
                wxbmbj_parts.append(f"{wxbm_key} {count}")
            wxbmbj_str = "\n".join(wxbmbj_parts)

            # 写入当前业务员的数据到一行
            row_data = [
                ywry,                                    # A
                data['cgbj'],                           # B
                wxbmbj_str,                             # C
                f"{hit_rate:.3f}",                      # D
                f"{data['FD']}({data['FD1']})",         # E
                f"{data['XJ']}({data['XJ1']})",         # F
                f"{data['CGTJ']}({data['CGTJ1']})",     # G
                f"{data['WXTJ']}({data['WXTJ1']})",     # H
                f"{data['KHZX']}({data['KHZX1']})",     # I
                f"{data['YPLL']}({data['YPLL1']})",     # J
                len(data['wxbm_set']),                  # K
                "0.000",                                # L (硬编码)
                0,                                      # M (硬编码)
                "0.000",                                # N (硬编码)
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            current_row += 1

        # --- 5. 保存文件并返回路径 ---
        path = config.tmp_path
        os.makedirs(path, exist_ok=True) # 确保目录存在
        report_rid = get_uuid() # 假设这是一个生成唯一ID的函数
        output_filename = f'{report_rid}.xlsx'
        full_output_path = os.path.join(path, output_filename)
        
        wb.save(full_output_path)

        # **关键修改: 将成功码改为 1，并确保 data 是文件名字符串**
        return {'code': 1, 'msg': '报表生成成功', 'data': output_filename}

    except Exception as e:
        logger.error(traceback.format_exc())
        return {'code': -1, 'msg': f'生成报表时发生错误: {str(e)}', 'data': None}
    finally:
        s.close()


# --- FastAPI 路由处理器 ---
@any_route('/api/saier/quotation/statistics/excel', methods=['POST', 'GET'])
@require_token
async def api_saier_quotation_statistics_excel(request):
    """
    API 接口，接收请求并触发后台的 Excel 生成任务。
    """
    user = request.current_user
    try:
        # 1. 在主线程中获取请求参数
        j = await request.json()
        # **关键修改: 使用前端实际发送的参数名**
        qsrq = j.get('qsrq', '')  # 开始日期
        jsrq = j.get('jsrq', '')    # 结束日期
        
        # **关键修改: 获取当前用户，因为前端没有发送用户名**
        # 请根据您的 @require_token 装饰器实际将用户信息放置的位置来修改。
        # 常见情况如 request.state.current_user 或 request.token_user
        # 这里假设它被放在了 request.state.user
        # *** 请根据您的框架实际情况修改 ***
        user_name = user.username # request.state.user.get('user.username') if hasattr(request.state, 'user') else None
        # if not user_name:
        # #     # 或者，如果框架将用户信息直接放在 request 对象上，例如 request.current_user
        # #     # user_name = getattr(request, 'current_user', {}).get('username')
        # #     # 如果以上都不对，请提供您框架获取用户名的正确方法。
        #     raise AttributeError("当前用户名不存在。")

        # 2. 所有参数均已获得，现在将耗时的数据库和文件操作放到后台线程执行
        loop = asyncio.get_event_loop()
        res = await loop.run_in_executor(
            executor,
            _generate_excel_on_thread, # 调用新的后台函数
            qsrq, jsrq, user_name
        )
        return json_result(res.get('code'), res.get('msg'), res.get('data'))
    
    except AttributeError as ae:
        # 捕获因无法获取用户名而导致的错误
        error_msg = f"获取用户信息失败: {str(ae)}"
        logger.error(error_msg)
        return json_result(-1, error_msg)
    except Exception as e:
        logger.error(traceback.format_exc())
        return json_result(-1, traceback.format_exc())