
from .__default__ import get_user_path
from datetime import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from .model import *
from sqlalchemy.sql import text
from openpyxl.utils import get_column_letter

# 普通用户登入的话 会因为用户名而找不出数据
@any_route('/api/saier/salesorder/order_analysis/export', methods=['POST'])
@require_token
async def api_saier_product_retention_export(request):
    """
    产品接单年度对比分析 - 业务逻辑转译
    参考 Delphi 脚本中的 1/2/3 年留存分析逻辑
    """
    j = await request.json()
    s = Session()
    try:
        # 1. 获取并处理前端参数
        khmc = j.get('khmc', '')     # 客人名称
        country = j.get('country', '') # 国家
        ssdq = j.get('ssdq', '')       # 州别/地区
        qsrq = str(j.get('qsrq', '3')) # 查看年数: '1','2','3'
        yfeng = j.get('yfeng', '')     # 具体月份
        user = request.current_user     # 当前用户对象 (包含 username, path)
        org_path=get_user_path(user.username)
        # org_path = get_user_path('丁瑞莹')
        # 2. 计算日期维度 (对应 Delphi 的 dn, qn, qn1, qn2)
        curr_year = dt.now().year
        periods = []
        num_years = int(qsrq) if qsrq in ['1', '2', '3'] else 3
        
        # 格式化月份 (Delphi 逻辑: yfeng := '0' + yfeng)
        suffix = f"-{yfeng.zfill(2)}" if yfeng else ""
        
        for i in range(num_years):
            periods.append(f"{curr_year - i}{suffix}")
        
        # 核心筛选条件 (对应 Delphi 的 path 权限过滤)
        # periods 示例: ['2024-05', '2023-05', '2022-05'] 或 ['2024', '2023', '2022']
        
        # 3. 核心 SQL 逻辑：一次性找出符合“连续接单”条件的 bjhh
        # 替代 Delphi 的三重嵌套循环，使用 HAVING COUNT 提升性能
        retention_sql = """
            SELECT bjhh 
            FROM wxhtsheet
            WHERE ((bjpath LIKE :path) OR (bjyw = :username))
              AND (jdrq <> '1999-01-01' AND jdrq IS NOT NULL)
              AND (khmc LIKE :khmc)
              AND (country LIKE :country)
              AND (ssdq LIKE :ssdq)
              AND (
                """ + " OR ".join([f"jdrq LIKE :p{i}" for i in range(len(periods))]) + """
              )
            GROUP BY bjhh
            HAVING COUNT(DISTINCT LEFT(jdrq, :len)) = :count
        """
        
        sql_params = {
            "path": f"%{org_path}%",
            "username": user.username,
            "khmc": f"%{khmc}%",
            "country": f"%{country}%",
            "ssdq": f"%{ssdq}%",
            "len": 7 if yfeng else 4,
            "count": len(periods)
        }
        for i, p in enumerate(periods):
            sql_params[f"p{i}"] = f"%{p}%"

        # 执行筛选
        valid_records = s.execute(retention_sql, sql_params).fetchall()
        valid_bjhhs = [r.bjhh for r in valid_records]

        if not valid_bjhhs:
            return json_result(-1, "未找到在指定期间内连续下单的产品")

        # 4. 获取汇总数据及中文品名 (对应 Delphi 的最后填充阶段)
        # 获取这些 bjhh 在各年份的 sum(htsl)
        summary_sql = """
            SELECT w.bjhh, z.zwpm, 
                   LEFT(w.jdrq, :len) as period, 
                   SUM(w.htsl) as total_qty
            FROM wxhtsheet w 
            LEFT JOIN zscp z ON w.bjhh = z.cpbh
            WHERE w.bjhh IN :bjhhs
              AND (""" + " OR ".join([f"w.jdrq LIKE :p{i}" for i in range(len(periods))]) + """)
            GROUP BY w.bjhh, z.zwpm, LEFT(w.jdrq, :len)
        """
        sql_params["bjhhs"] = tuple(valid_bjhhs)
        final_data = s.execute(summary_sql, sql_params).fetchall()

        # 5. 生成 Excel (参考 purchase_contract_report.py 风格)
        wb = Workbook()
        ws = wb.active
        ws.title = "产品接单分析"

        # 写入动态表头 (对应 Delphi 的 C1, D1, E1, F1 逻辑)
        headers = ["产品编号", "中文品名"]
        for p in periods:
            headers.append(f"{p} 接单数")
        ws.append(headers)

        # 数据重组：将行数据转为透视格式
        # 结构: { bjhh: { 'zwpm': 'xx', '2024': 100, '2023': 50 ... } }
        data_map = {}
        for r in final_data:
            if r.bjhh not in data_map:
                data_map[r.bjhh] = {"zwpm": r.zwpm or ""}
            data_map[r.bjhh][r.period] = r.total_qty

        # 写入行
        for bjhh, info in data_map.items():
            row = [bjhh, info.get("zwpm")]
            for p in periods:
                row.append(info.get(p, 0)) # 若某期无数据补0
            ws.append(row)

        # 样式美化
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        ws.column_dimensions['A'].width = 20  # 产品编号通常是英文字母和数字，20的宽度比较合适
        ws.column_dimensions['B'].width = 20  
        for i in range(len(periods)):
            col_letter = get_column_letter(i + 3) 
            ws.column_dimensions[col_letter].width = 20 # 给动态年份列设置统一的宽度


        # 保存文件
        s_path = config.tmp_path
        report_rid = get_uuid()
        filename=str(report_rid)+'.xlsx'
        wb.save(s_path + '/'+ str(report_rid)+'.xlsx')
        return json_result(1, '导出成功', filename)

    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f"导出接单分析失败: {str(e)}")
    finally:
        s.close()






# ======================= ======================== 业务询价中的   采购报价统计 =================================

@any_route('/api/saier/quotation/generate', methods=['POST'])
@require_token
async def api_saier_quotation_generate(request):
    """
    询价单转报价单 (转译自 Delphi 采购报价统计脚本)
    """
    j = await request.json()
    s = Session()
    try:
        # 1. 解析前端传来的当前单据参数 (替代 Delphi 的 recordset)
        jzrq = j.get('jzrq', '')          # 截止日期
        xpdh = j.get('xpdh', '')          # 询盘单号
        ywry = j.get('ywry', '')          # 业务人员 (客户经理)
        khbh = j.get('khbh', '')          # 客户编号
        bjbz = j.get('bjbz', '')          # 报价备注
        shsc = j.get('shsc', '')          # 适合市场
        selected_items = j.get('items', []) # 选中的产品明细列表
        
        user = request.current_user
        username = user.username
        user_path = getattr(user, 'path', '')

        # 2. 时效校验 (对比服务器当前时间与截止日期)
        # 对应 Delphi: if (截止日期 + ' 23:59:59') > rq1
        now_str = dt.now().strftime('%Y-%m-%d %H:%M:%S')
        deadline_str = f"{jzrq} 23:59:59" if jzrq else ""
        
        if deadline_str and deadline_str <= now_str:
            return json_result(-1, "不好意思,此报价已过截止日期!")

        if not selected_items:
            return json_result(-1, "没有选中任何需要报价的产品明细!")

        # 3. 构造新的报价单号
        # 对应 Delphi: bjdh := 询盘单号 + username + 年月日时分秒
        time_suffix = dt.now().strftime('%Y%m%d%H%M%S')
        bjdh = f"{xpdh}-{username}{time_suffix}"
        
        # 4. 插入消息系统 (对应 Delphi 的 xxck 表记录)
        # 通知业务员某人已开始准备报价
        sql_xxck = text("""
            INSERT INTO xxck (sys_owner, sys_path, sys_created, sys_lastmodified, fsrq, fsr, xxly, bjdh, xxnr, jsr)
            VALUES (:sys_owner, :sys_path, :sys_created, :sys_lastmodified, :fsrq, :fsr, :xxly, :bjdh, :xxnr, :jsr)
        """)
        s.execute(sql_xxck, {
            "sys_owner": ywry,
            "sys_path": "我的公司\\",
            "sys_created": now_str,
            "sys_lastmodified": now_str,
            "fsrq": dt.now().strftime('%Y-%m-%d'),
            "fsr": username,
            "xxly": "业务询价",
            "bjdh": xpdh,
            "xxnr": f"业务询价:{xpdh}已由{username}准备报价,时间{now_str}",
            "jsr": ywry
        })

        # 5. 获取业务员部门、公司、基础配置字典
        wfgs, bm = '', ''
        ywry_info = s.execute(text("SELECT wfgs, bm FROM ywrybiao WHERE yhm=:yhm"), {"yhm": username}).fetchone()
        if ywry_info:
            wfgs, bm = ywry_info.wfgs or '', ywry_info.bm or ''

        # 获取结算方式
        jsfs_query = text("SELECT top 1 jsfs FROM jsfs WHERE szgs=:szgs") if wfgs else text("SELECT top 1 jsfs FROM jsfs")
        jsfs_row = s.execute(jsfs_query, {"szgs": wfgs}).fetchone()
        jsfs = jsfs_row.jsfs if jsfs_row else ''

        # 获取常用变量 (出运卡、运输方式、汇率)
        cyka = s.execute(text("SELECT top 1 ywmc FROM cyka")).scalar() or ''
        ysfs = s.execute(text("SELECT top 1 ysfs FROM ysfs")).scalar() or ''
        huilv = s.execute(text("SELECT top 1 hhl FROM hbdm WHERE hbdm='USD$'")).scalar() or 0.0

        # 6. 生成报价单主表 (bj) 
        # 检查是否已存在 (防重复生成)
        existing_bj = s.execute(text("SELECT number FROM bj WHERE bj_id=:bj_id"), {"bj_id": bjdh}).fetchone()
        if existing_bj:
            return json_result(-1, "该报价单已存在，请勿重复操作!")

        sql_bj = text("""
            INSERT INTO bj(sys_owner, sys_path, yw, sys_created, sys_lastmodified, bj_id, dateis, 
                           ywz, bjsh, wyzd, qsn, qsy, bjzt, hbdm, bjjg, bjql, ykbz, yjbl, bjql1, 
                           tgsb, fxsb, sfms, xgck, cgqx, cyka, ysfs, huilv, bxjc, bxbl, xj_dh, kh_id, customer, bjbz, xjht, shsc)
            VALUES(:sys_owner, :sys_path, :yw, :sys_created, :sys_lastmodified, :bj_id, :dateis, 
                   :ywz, :bjsh, :wyzd, :qsn, :qsy, :bjzt, :hbdm, :bjjg, :bjql, :ykbz, :yjbl, :bjql1, 
                   :tgsb, :fxsb, :sfms, :xgck, :cgqx, :cyka, :ysfs, :huilv, :bxjc, :bxbl, :xj_dh, :kh_id, :customer, :bjbz, :xjht, :shsc)
        """)
        
        # 判断公司审核流级别
        bjsh = '123' if wfgs == '宁波优景进出口有限公司' else '1'
        wyzd = f"bj{bjdh}{username}{time_suffix}"
        
        s.execute(sql_bj, {
            "sys_owner": username, "sys_path": user_path, "yw": user_path,
            "sys_created": now_str, "sys_lastmodified": now_str,
            "bj_id": bjdh, "dateis": dt.now().strftime('%Y-%m-%d'),
            "ywz": bm, "bjsh": bjsh, "wyzd": wyzd,
            "qsn": str(dt.now().year), "qsy": str(dt.now().month).zfill(2),
            "bjzt": "使用中", "hbdm": "USD$", "bjjg": "待审批", "bjql": "待审批",
            "ykbz": "否", "yjbl": 1.0, "bjql1": "待审批", "tgsb": "否", "fxsb": "否", 
            "sfms": "否", "xgck": "否", "cgqx": "有",
            "cyka": cyka, "ysfs": ysfs, "huilv": huilv, "bxjc": 110.0, "bxbl": 0.78,
            "xj_dh": xpdh, "kh_id": khbh, "customer": ywry, "bjbz": bjbz, "xjht": "旧", "shsc": shsc
        })

        # 获取刚插入主表的系统流水号 (number / father)
        father_row = s.execute(text("SELECT number FROM bj WHERE bj_id=:bj_id"), {"bj_id": bjdh}).fetchone()
        father = father_row.number if father_row else None

        if not father:
            s.rollback()
            return json_result(-1, "主表生成失败")

        # 7. 循环插入报价单明细表 (bjsheet)
        # Delphi 的 while 循环，在这里转变为 Python 的 for 循环
        sql_sheet = text("""
            INSERT INTO bjsheet(father, zwpm, ywpm, cpgg, nhrl, nhwx, wxrl, zhwbzh, cszl, cz, 
                                cpfl, yjfl, ejfl, flmc, cply1, mbjg, yjsl, ndzs, bz, hhbz, xjhh, xjsx, bjpath)
            VALUES(:father, :zwpm, :ywpm, :cpgg, :nhrl, :nhwx, :wxrl, :zhwbzh, :cszl, :cz, 
                   :cpfl, :yjfl, :ejfl, :flmc, :cply1, :mbjg, :yjsl, :ndzs, :bz, :hhbz, :xjhh, :xjsx, :bjpath)
        """)

        # 为了极致性能，收集参数后批量执行 (Bulk Insert)
        sheet_params = []
        for item in selected_items:
            sheet_params.append({
                "father": father,
                "bjpath": user_path,
                "zwpm": item.get('zwpm', ''),        # 中文品名
                "ywpm": item.get('ywpm', ''),        # 英文品名
                "cszl": item.get('cszl', ''),        # 测试种类
                "cpfl": item.get('cpfl', ''),        # 产品大类
                "yjfl": item.get('yjfl', ''),        # 一级分类
                "ejfl": item.get('ejfl', ''),        # 二级分类
                "flmc": item.get('flmc', ''),        # 分类名称
                "zhwbzh": item.get('zhwbzh', ''),    # 中文包装
                "cpgg": item.get('cpgg', ''),        # 产品规格
                "nhrl": float(item.get('nhrl', 0)),  # 内盒装箱量
                "wxrl": float(item.get('wxrl', 0)),  # 外箱装箱量
                "nhwx": int(item.get('nhwx', 0) or 0), # 内盒/外箱
                "cz": item.get('cz', ''),            # 材质
                "cply1": item.get('cply1', ''),      # 产品来源
                "mbjg": float(item.get('mbjg', 0)),  # 目标价格
                "yjsl": int(item.get('yjsl', 0) or 0), # 预计数量
                "ndzs": int(item.get('ndzs', 0) or 0), # 难度指数
                "hhbz": item.get('xjhh', ''),        # 询价货号
                "bz": item.get('bz', ''),            # 备注
                "xjhh": item.get('xjhh', ''),
                "xjsx": item.get('xjsx', '')         # 询价属性
            })
        
        # 批量执行明细插入
        if sheet_params:
            s.execute(sql_sheet, sheet_params)

        # 8. 提交事务
        s.commit()

        # 9. 返回成功及新单号给前端，不再使用 Delphi 的 xia.txt 文本文件跳转机制
        return json_result(1, "生成报价单成功", {"new_quote_id": bjdh, "number": father})

    except Exception as e:
        s.rollback() # 发生异常必须回滚
        logger.error(trace_error())
        return json_result(-1, f"生成报价单失败: {str(e)}")
    finally:
        s.close() # 严格遵循你代码中的 finally 关闭 session



# 业务询价中的    业务报价统计




#  业务询价中的     采购询价完成统计