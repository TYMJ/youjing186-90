# backend: purchase_followup_export.py
# 采购跟单 - 优景采购合同批量（图）
# 源码：source.pas；单份合同 Excel/PDF 逻辑在本文件完整实现，不跨文件 import。
# 跟单特有：cggd 收集 hthm、da2 当前/批量、多合同 ZIP。
import io
import zipfile
import os
import shutil
import json
from sqlalchemy.sql import text as sql_text
import subprocess
from any import *
from .model import *

try:
    import xlwings as xw
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "xlwings"])
    import xlwings as xw

from .__default__ import config
from datetime import datetime as dt

try:
    import cn2an
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "cn2an"])

try:
    from openpyxl import load_workbook as oxl_load_workbook
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import load_workbook as oxl_load_workbook
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont


# --- 辅助工具函数 ---
def number_to_chinese_currency(num):
    try:
        # 转换为中文金额
        return cn2an.an2cn(float(num), "rmb")
    except Exception as e:
        logger.error(trace_error())
        return "金额转换错误!!!"


def safe_write(ws, coord, value, num_format=None):
    """合并单元格安全写入：coord 若在合并区内则写入该区域左上角主单元格。"""
    rng = ws.range(coord)
    area = rng.merge_area
    rows, cols = area.shape
    text_val = "" if value is None else str(value)
    if rows * cols > 1:
        cell = area(1, 1)
        if num_format:
            cell.number_format = num_format
        cell.value = text_val
    else:
        if num_format:
            ws[coord].number_format = num_format
        ws[coord].value = text_val


def copy_row_style(ws, source_row, target_row):
    ws.range(source_row, source_row).copy()
    ws.range(target_row, target_row).paste(paste="formats")


def insert_row_at(ws, at_row):
    ws.range(at_row, at_row).api.EntireRow.Insert()


def _add_picture_at(ws, coord, path, x_pad, y_pad, width, height):
    anchor = ws.range(coord)
    pic = ws.pictures.add(path, left=anchor.left + x_pad, top=anchor.top + y_pad)
    if width is not None:
        pic.width = width
    if height is not None:
        pic.height = height
    return pic


def _set_row_height(ws, row, height):
    ws.range(f"{row}:{row}").row_height = height


def _set_range_center_wrap(ws, coord_range):
    rng = ws.range(coord_range)
    api = rng.api
    api.WrapText = True
    api.HorizontalAlignment = -4108
    api.VerticalAlignment = -4108


def _set_cell_left_wrap(ws, coord):
    """单元格左对齐 + 自动换行。"""
    rng = ws.range(coord).merge_area
    api = rng(1, 1).api
    api.WrapText = True
    api.HorizontalAlignment = -4131


def _set_wrap_text(ws, coord):
    ws.range(coord).api.WrapText = True


def _autofit_row(ws, row):
    """对照 Pascal Rows[n].AutoFit。"""
    try:
        ws.range(row, row).api.EntireRow.AutoFit()
    except Exception:
        pass


def _autofit_row_clamped(ws, row, min_h=95, max_h=400):
    """AutoFit 后限制行高（对应 Pascal Selection.height 95~400 逻辑）。"""
    _autofit_row(ws, row)
    try:
        h = ws.range(row, row).row_height or min_h
    except Exception:
        h = min_h
    if h < min_h:
        h = min_h
    elif h > max_h:
        h = max_h
    _set_row_height(ws, row, h)


def _get_row_height(ws, row):
    try:
        return float(ws.range(row, row).row_height or 0)
    except Exception:
        return 0.0


def _measure_row_selection_height(ws, row):
    """读取整行选中后的实际行高（source.pas Selection.height）。"""
    try:
        ws.range(f"{row}:{row}").api.Select()
        return float(ws.api.Application.Selection.RowHeight)
    except Exception:
        return 0.0


def _estimate_text_row_height(text, col_units=24, pt_per_line=15, min_h=95, max_h=400):
    """合并单元格 AutoFit 不准时，按文本折行估算行高。"""
    norm = _normalize_cell_text(text)
    if not norm:
        return min_h
    total_lines = 0
    for line in norm.split("\n"):
        if not line:
            total_lines += 1
            continue
        en = sum(1 for ch in line if ord(ch) < 128)
        cn = len(line) - en
        width_units = cn * 2 + en
        total_lines += max(1, (width_units + col_units - 1) // col_units)
    return max(min_h, min(max_h, total_lines * pt_per_line + 8))


def _autofit_nb_row15(ws, nbpm="", zhwbzh=""):
    """宁波第15行：Q/R 临时格 + 整行换行 + Selection 行高（source.pas 1628-1664）。"""
    ws["Q15"].value = nbpm
    ws["R15"].value = zhwbzh
    _set_range_center_wrap(ws, "B15:AE15")

    for _ in range(2):
        _autofit_row(ws, 15)

    h = _measure_row_selection_height(ws, 15)
    if h <= 0:
        h = _get_row_height(ws, 15)

    est = max(_estimate_text_row_height(nbpm, col_units=20), _estimate_text_row_height(zhwbzh, col_units=18))
    h = max(h, est, 95)
    if h > 400:
        h = 400

    _set_row_height(ws, 15, h)
    ws["Q15"].value = ""
    ws["R15"].value = ""


def _autofit_nb_row16(ws, is_fl, cpsm="", qtsm1=""):
    """宁波第16行：S/T 临时格 + 换行 + AutoFit（source.pas 1667-1742）。"""
    _set_range_center_wrap(ws, "B16:AE16")
    if not is_fl:
        ws["S16"].value = cpsm
        ws["T16"].value = qtsm1
    else:
        ws["S16"].value = f"{cpsm}\r\n{qtsm1}"

    _autofit_row(ws, 16)
    h = _get_row_height(ws, 16)
    if h < 95:
        h = 95
    elif h > 400:
        for ts1 in range(1, 6):
            try:
                ws.api.Columns("M").ColumnWidth = 25 + 5 * ts1
                s_w = (34 + 25 + 5 * ts1) if not is_fl else (54.25 + 25 + 5 * ts1)
                ws.api.Columns("S").ColumnWidth = s_w
            except Exception:
                pass
            _autofit_row(ws, 16)
            h = _get_row_height(ws, 16)
            if h <= 400:
                break
        if h > 400:
            for font_size in (11, 10):
                try:
                    ws.range("F16").api.Font.Size = font_size
                    if not is_fl:
                        ws.range("S16").api.Font.Size = font_size
                    _autofit_row(ws, 16)
                    h = _get_row_height(ws, 16)
                except Exception:
                    pass
                if h <= 400:
                    break
        h = min(h, 400)

    _set_row_height(ws, 16, h)
    if not is_fl:
        ws["S16"].value = ""
        ws["T16"].value = ""
    else:
        ws["S16"].value = ""


def _autofit_nb_row18(ws, kppm_line, is_fl=False):
    """宁波第18行：开票品名写入 + AutoFit，行高 16~400（source.pas 1747-1778）。"""
    if not kppm_line:
        return
    kppm_coords = ["T18"] if is_fl else ["O18", "U18"]
    for coord in kppm_coords:
        _set_cell_left_wrap(ws, coord)
    _autofit_row(ws, 18)
    if is_fl:
        safe_write(ws, "T18", kppm_line)
    else:
        safe_write(ws, "O18", kppm_line)
        safe_write(ws, "U18", kppm_line)
    for coord in kppm_coords:
        _set_cell_left_wrap(ws, coord)
    _autofit_row(ws, 18)
    h = _measure_row_selection_height(ws, 18)
    if h <= 0:
        h = _get_row_height(ws, 18)
    est = _estimate_text_row_height(kppm_line, col_units=20, pt_per_line=15, min_h=16, max_h=400)
    h = max(h, est, 16)
    if h > 400:
        h = 400
    _set_row_height(ws, 18, h)


def _load_detail_yytp_json(s, is_fl, hhbz, wyzd):
    """按 source.pas 查产品图 yytp：普通→lscp.lshh=hhbz，辅料→cghtsheet.wyzd。"""
    hhbz = str(hhbz or "").strip()
    wyzd = str(wyzd or "").strip()
    if not is_fl and hhbz:
        row_img = s.execute(
            sql_text("SELECT yytp FROM lscp WHERE lshh=:lshh AND LENGTH(yytp) > 5"), {"lshh": hhbz}
        ).fetchone()
    elif is_fl and wyzd and hhbz:
        row_img = s.execute(
            sql_text("SELECT yytp FROM cghtsheet WHERE wyzd=:wyzd AND LENGTH(yytp) > 5"), {"wyzd": wyzd}
        ).fetchone()
    else:
        return ""
    if row_img and row_img.yytp:
        return str(row_img.yytp)
    return ""


def _format_zzsl_pct(zzsl):
    """对应 Pascal floattostr(zzsl)，整数税率不带小数点。"""
    try:
        v = float(zzsl or 0)
    except (TypeError, ValueError):
        return "0"
    if v == int(v):
        return str(int(v))
    return str(v).rstrip("0").rstrip(".")


def _should_write_kppm(zhwbgpm, zzsl):
    name = str(zhwbgpm or "").strip()
    return (name and name != "无") or float(zzsl or 0) > 0


def _kppm_invoice_line(zhwbgpm, zzsl):
    """开票品名行：source.pas 『开票品名: 』+ zhwbgpm + floattostr(zzsl) + 『%』。"""
    if not _should_write_kppm(zhwbgpm, zzsl):
        return ""
    name = str(zhwbgpm or "").strip()
    if name == "无":
        name = ""
    return f"开票品名: {name}{_format_zzsl_pct(zzsl)}%"


def _pick_nonempty(new_val, old_val):
    """zycs 字段为空时不覆盖 cs 已有值（避免 phone=「无」被空串盖掉）。"""
    new_val = str(new_val or "")
    return new_val if new_val != "" else str(old_val or "")


def _load_factory_contact(s, row_cght, sccj2, sccj_id2, zygcmc2, zygcid2):
    """加载工厂联系信息（source.pas 821-874）。"""
    cslxr, phone, fax, sjhm, address = "", "", "", "", ""
    twhm, fkhm, bank1, zh1 = "", "", "", ""

    if sccj2 != "":
        row_cs = s.execute(sql_text("SELECT * FROM cs WHERE cs_id=:cs_id"), {"cs_id": sccj_id2}).fetchone()
        if row_cs:
            cslxr = str(row_cs.cslxr or "")
            phone = str(row_cs.phone or "")
            fax = str(row_cs.fax or "")
            sjhm = str(row_cs.sjhm or "")
            address = str(row_cs.address or "")
            twhm = str(row_cs.twhm or "")

    if zygcmc2 != "":
        row_zycs = s.execute(sql_text("SELECT * FROM zycs WHERE cs_id=:cs_id"), {"cs_id": zygcid2}).fetchone()
        if row_zycs:
            cslxr = _pick_nonempty(row_zycs.cslxr, cslxr)
            phone = _pick_nonempty(row_zycs.phone, phone)
            fax = _pick_nonempty(row_zycs.fax, fax)
            sjhm = _pick_nonempty(row_zycs.sjhm, sjhm)
            address = _pick_nonempty(row_zycs.address, address)
            twhm = _pick_nonempty(row_zycs.twhm, twhm)
            fkhm = str(row_zycs.fkhm or "")
            bank1 = str(row_zycs.bank1 or "")
            zh1 = str(row_zycs.zh1 or "")

    if str(row_cght.lxry or "") != "":
        cslxr = str(row_cght.lxry or "")
    if str(row_cght.gcdh or "") != "":
        phone = str(row_cght.gcdh or "")
    if str(row_cght.sjhm or "") != "":
        sjhm = str(row_cght.sjhm or "")
    if str(row_cght.gccz or "") != "":
        fax = str(row_cght.gccz or "")

    return cslxr, phone, fax, sjhm, address, twhm, fkhm, bank1, zh1


PHONE_CELL_FORMAT = "@"


def _parse_jhrq(jhrq):
    """解析交货期字符串为 datetime。"""
    s = str(jhrq or "").strip()
    if not s:
        return None
    for fmt, ln in (
        ("%Y-%m-%d %H:%M:%S", 19),
        ("%Y-%m-%d %H:%M", 16),
        ("%Y/%m/%d %H:%M:%S", 19),
        ("%Y/%m/%d %H:%M", 16),
        ("%Y-%m-%d", 10),
        ("%Y/%m/%d", 10),
    ):
        try:
            return dt.strptime(s[:ln], fmt)
        except ValueError:
            continue
    return None


def _format_jhrq_for_cell(jhrq):
    """交货期写入 Excel：显示 yyyy/m/d，如 2026/5/16（对应模板 E26/D27）。"""
    parsed = _parse_jhrq(jhrq)
    if parsed:
        return parsed
    return str(jhrq or "")


JHRQ_NUM_FORMAT = "yyyy/m/d"


def _normalize_cell_text(text):
    return str(text or "").replace("_x000D_\n", "\n").replace("\r\n", "\n")


def _calc_red_start_1based(text, dwz_fallback):
    """标红起始位置：优先按关键字，回退 dwz（Pascal Characters 起点）。"""
    norm = _normalize_cell_text(text)
    marker = "请注意开票货源地为："
    idx = norm.find(marker)
    if idx >= 0:
        return idx + 1
    return dwz_fallback


class _RedTextPatches:
    """合并单元格上 xlwings Characters 无效，保存后用 openpyxl 写富文本红字。"""

    def __init__(self):
        self._items = []

    def add(self, coord, text, start_1based, sheet_index=0):
        if text and start_1based > 0:
            self._items.append((sheet_index, coord, text, start_1based))

    def apply(self, xlsx_path):
        if not self._items or not xlsx_path or not os.path.exists(xlsx_path):
            return
        try:
            red_font = InlineFont(b=True, color="00FF0000")
            wb = oxl_load_workbook(xlsx_path)
            for sheet_index, coord, text, start_1based in self._items:
                ws = wb.worksheets[sheet_index] if sheet_index < len(wb.worksheets) else wb.active
                norm = _normalize_cell_text(text)
                split_idx = _calc_red_start_1based(norm, start_1based) - 1
                if split_idx <= 0:
                    ws[coord] = norm
                else:
                    ws[coord] = CellRichText(norm[:split_idx], TextBlock(red_font, norm[split_idx:]))
            wb.save(xlsx_path)
            wb.close()
        except Exception:
            logger.error(f"openpyxl 局部红字写入失败: {trace_error()}")


def _load_htjj_info(session, gs1, szdh):
    """查询合同抬头/地址/联系方式（source.pas htjj 两次查询）。"""
    ndz, dz, lxfs, cgzj, zjl, gstt = "", "", "", "", "", ""
    row_nb = session.execute(
        sql_text("SELECT * FROM htjj WHERE dq='宁波' AND gstt LIKE :gstt"), {"gstt": f"%{gs1}%"}
    ).fetchone()
    if row_nb:
        ndz = str(row_nb.dz or "")
    dq = str(szdh or "").strip() or "宁波"
    row = session.execute(
        sql_text("SELECT * FROM htjj WHERE dq=:dq AND gstt LIKE :gstt"), {"dq": dq, "gstt": f"%{gs1}%"}
    ).fetchone()
    if not row and dq != "宁波":
        row = session.execute(
            sql_text("SELECT * FROM htjj WHERE dq='宁波' AND gstt LIKE :gstt"), {"gstt": f"%{gs1}%"}
        ).fetchone()
    if row:
        dz = str(row.dz or "")
        lxfs = str(row.lxfs or "")
        cgzj = str(row.cgzj or "")
        zjl = str(row.zjl or "")
        gstt = str(row.gstt or "")
    return ndz, dz, lxfs, cgzj, zjl, gstt


def _start_excel_app():
    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    app.screen_updating = False
    return app


def _open_workbook(app, path):
    return app.books.open(path)


def _save_and_close(wb, path):
    wb.save(path)
    wb.close()


def convert_excel_to_pdf(excel_path, output_dir=None):
    """
    将Excel文件转换为PDF（使用 whale_report.jar）

    Args:
        excel_path: Excel文件的完整路径
        output_dir: PDF输出目录，默认与Excel同目录

    Returns:
        dict: {'success': bool, 'pdf_path': str, 'error': str}
    """
    try:
        # 确定输出目录
        if output_dir is None:
            output_dir = os.path.dirname(excel_path)

        # 生成PDF文件路径
        base_name = os.path.splitext(excel_path)[0]
        pdf_path = base_name + ".pdf"

        print("========== 开始PDF转换 ==========")
        print(f"源文件: {excel_path}")
        print(f"目标文件: {pdf_path}")

        # 检查 Java 和 JAR 文件是否存在
        if not hasattr(config, "java_path") or not config.java_path:
            return {"success": False, "error": "Java路径未配置（config.java_path）"}

        if not hasattr(config, "report_jar") or not config.report_jar:
            return {"success": False, "error": "whale_report.jar路径未配置（config.report_jar）"}

        if not os.path.exists(config.report_jar):
            return {"success": False, "error": f"whale_report.jar文件不存在: {config.report_jar}"}

        print(f"Java路径: {config.java_path}")
        print(f"JAR路径: {config.report_jar}")

        # 构建命令
        # console_run(config.java_path, ['-jar', config.report_jar, 'a', 'b', template, output_file, '2'])
        cmd = [
            config.java_path,
            "-jar",
            config.report_jar,
            "a",  # 占位参数
            "b",  # 占位参数
            excel_path,  # 源Excel文件
            pdf_path,  # 目标PDF文件
            "2",  # 转换类型参数
        ]
        # cmd = console_run(config.java_path, ["-jar", config.report_jar, "a", "b", template, output_file, "2"])

        print(f"执行命令: {' '.join(cmd)}")

        # 执行转换
        result = subprocess.run(
            cmd,
            capture_output=True,
            timeout=120,  # 2分钟超时
            text=True,
        )

        print(f"返回码: {result.returncode}")
        if result.stdout:
            print(f"标准输出: {result.stdout}")
        if result.stderr:
            print(f"错误输出: {result.stderr}")

        if result.returncode != 0:
            return {
                "success": False,
                "error": f"PDF转换失败（返回码 {result.returncode}）：stdout={result.stdout} stderr={result.stderr}",
            }

        # 检查PDF文件是否生成
        if not os.path.exists(pdf_path):
            # 有时jar会把PDF生成到同目录但文件名不同，尝试查找
            alt_candidates = [
                os.path.join(output_dir, os.path.basename(pdf_path)),
                pdf_path + ".pdf",
                os.path.splitext(excel_path)[0] + ".pdf",
            ]
            found_alt = None
            for alt in alt_candidates:
                if os.path.exists(alt) and alt != pdf_path:
                    found_alt = alt
                    break
            if found_alt:
                shutil.move(found_alt, pdf_path)
                print(f"[INFO] 找到并移动PDF: {found_alt} -> {pdf_path}")
            else:
                return {
                    "success": False,
                    "error": f"PDF文件未生成（期望路径：{pdf_path}），stdout={result.stdout} stderr={result.stderr}",
                }

        file_size = os.path.getsize(pdf_path)
        print("[OK] PDF转换成功")
        print(f"[OK] 文件大小: {file_size:,} 字节")
        print("========== PDF转换完成 ==========")

        return {"success": True, "pdf_path": pdf_path, "error": None}

    except subprocess.TimeoutExpired:
        return {"success": False, "error": "PDF转换超时（超过2分钟）"}
    except Exception as e:
        logger.error(trace_error())
        return {"success": False, "error": f"PDF转换异常：{str(e)}"}


# --- 采购跟单：合同号收集（source.pas MultiReport / da2）---


async def _parse_followup_request_body(request):
    """
    解析请求体：支持 application/json 与 multipart/form-data。
    RavenJLL-cggdreport.js 使用 FormData；purchase_followup_export.js 使用 JSON。
    """
    j = {}
    try:
        body = await request.json()
        if isinstance(body, dict):
            j = body
    except Exception:
        pass

    if not j:
        try:
            form = await request.form()
            if form:
                j = {
                    "da2": form_value(form, "da2", ""),
                    "rid": form_value(form, "rid", "") or form_value(form, "record_id", ""),
                    "rid_list": form_value(form, "rid_list", "") or form_value(form, "rids", ""),
                    "gs": form_value(form, "gs", ""),
                    "company": form_value(form, "company", ""),
                    "pdf": form_value(form, "pdf", ""),
                }
        except Exception:
            pass

    if not j:
        return {}

    # 旧前端 FormData：自定义简称时曾传 gs=2 + company=景驰
    gs_raw = str(j.get("gs") or "").strip()
    company = str(j.get("company") or "").strip()
    if gs_raw == "2" and company and company not in ("锐亿进出口", "优景进出口"):
        j["gs"] = company

    if not str(j.get("da2") or "").strip():
        if j.get("rid_list") or j.get("rids"):
            j["da2"] = "2"
        elif j.get("rid") or j.get("record_id"):
            j["da2"] = "1"

    return j


def _parse_rid_list(raw):
    """解析前端传入的 rid 列表（JSON 数组或逗号分隔）。"""
    if raw is None:
        return []
    if isinstance(raw, list):
        return [str(x).strip() for x in raw if str(x).strip()]
    s = str(raw).strip()
    if not s:
        return []
    if s.startswith("["):
        try:
            arr = json.loads(s)
            if isinstance(arr, list):
                return [str(x).strip() for x in arr if str(x).strip()]
        except Exception:
            pass
    return [p.strip() for p in s.split(",") if p.strip()]


def _multi_report_collect_hthm(session, cggd_rid, username):
    """
    对应 Pascal MultiReport(ANumber, a)：
    从 cggd 按 rid 查出记录，仅保留 gdry=当前用户的 hthm（去重）。
    """
    hthms = []
    rows = session.execute(
        sql_text("SELECT hthm, gdry FROM cggd WHERE rid=:rid"), {"rid": str(cggd_rid).strip()}
    ).fetchall()
    for row in rows:
        if str(row.gdry or "").strip() != username:
            continue
        h = str(row.hthm or "").strip()
        if h and h not in hthms:
            hthms.append(h)
    return hthms


def _collect_hthm_list(session, da2, rid, rid_list, username):
    """
    步骤 A（source.pas 约 61-176 行）：
    da2='1' 当前跟单 rid；da2='2' 批量 rid_list + MultiReport。
    """
    bgpmstr = []
    da2 = str(da2 or "1").strip()
    if da2 != "2":
        da2 = "1"
    if da2 == "1":
        if not rid:
            return [], "参数缺失：当前模式需要 rid（采购跟单主键）"
        bgpmstr = _multi_report_collect_hthm(session, rid, username)
        if not bgpmstr:
            row = session.execute(
                sql_text("SELECT hthm FROM cggd WHERE rid=:rid LIMIT 1"), {"rid": str(rid).strip()}
            ).fetchone()
            if row and str(row.hthm or "").strip():
                return [], "当前记录跟单人员与登录用户不一致，无法导出"
            return [], "未找到可导出的采购跟单记录"
    else:
        if not rid_list:
            return [], "批量模式需要 rid_list"
        for cggd_rid in rid_list:
            for h in _multi_report_collect_hthm(session, cggd_rid, username):
                if h not in bgpmstr:
                    bgpmstr.append(h)
        if not bgpmstr:
            return [], "批量导出未匹配到合同号，请确认跟单人员为当前用户"
    return bgpmstr, ""


def _resolve_cght_rid(session, hthm):
    row = session.execute(
        sql_text("SELECT rid FROM cght WHERE hthm=:hthm AND sfhz='通过' LIMIT 1"), {"hthm": hthm}
    ).fetchone()
    return str(row.rid).strip() if row and row.rid else ""


def _abs_output_path(rel_path):
    if not rel_path:
        return ""
    if os.path.isabs(rel_path):
        return rel_path
    base = getattr(config, "tmp_path", "") or ""
    return os.path.join(base, os.path.basename(rel_path))


def _init_pdf_permissions(s, username, pdf):
    """Part 1 权限与 pdf 参数（source.pas 约 85-126 行）。"""
    dhsb, dhsb3 = "", ""
    if s.execute(sql_text("SELECT 1 FROM cyzglsheet WHERE xm=:xm AND zm='采购合同图PDF'"), {"xm": username}).fetchone():
        dhsb = "1"
    if s.execute(
        sql_text("SELECT 1 FROM cyzglsheet WHERE xm=:xm AND zm='采购合同图PDF签名'"), {"xm": username}
    ).fetchone():
        dhsb3 = "1"
    if dhsb == "1":
        pdf = "3"
    else:
        if pdf not in ("2", "3"):
            pdf = "1"
        else:
            if dhsb3 == "1":
                dhsb = "1"
            pdf = "3"
    return pdf, dhsb, dhsb3


def _export_single_contract(s, user, record_id, gs, gs1, pdf, username, tmpstr2, dhsb, dhsb3):
    """
    单份采购合同（图）导出核心逻辑（源码 source.pas 内层循环，与优景采购合同图导出一致）。
    返回 dict: code, msg, data；不关闭 session，便于跟单批量循环。
    """
    if not record_id:
        return {"code": -1, "msg": "参数缺失：采购合同 rid", "data": {}}

    FF, htsm = "", ""
    cxsb = ""

    row_ff = s.execute(sql_text("SELECT WB1 FROM zx WHERE ly='反腐专线'")).fetchone()
    if row_ff:
        FF = str(row_ff.WB1 or "")
    row_htsm = s.execute(sql_text("SELECT nr FROM zx WHERE ly='采购合同签订注意要点'")).fetchone()
    if row_htsm:
        htsm = str(row_htsm.nr or "")
        if htsm != "":
            # 既然不能 showmessage，我们把它追加到提示信息的数组里
            tmpstr2.append(f"【签订注意要点】: {htsm}")  # 关于  showmessage    htsm 只是用了一次

    cxje1, cxjez1, cxjezz = 0.0, 1000000.0, 0.0
    cxsb, sfsh = "", "不需要"
    row_cx = s.execute(sql_text("SELECT cs, sz1 FROM zx WHERE mc='诚信金额'")).fetchone()
    if row_cx:
        cxje1 = float(row_cx.cs or 0)
        cxjez1 = float(row_cx.sz1 or 0)
        if cxjez1 == 0:
            cxjez1 = 1000000.0

    sfhs1 = "是"
    if s.execute(sql_text("SELECT * FROM cghtsheet WHERE pid=:pid AND sfhs<>'是'"), {"pid": record_id}).fetchone():
        sfhs1 = "否"

    # ================= [ Part 2: 获取主表记录 (Outermost 1) ] =================
    row_cght = s.execute(sql_text("SELECT * FROM cght WHERE rid=:rid"), {"rid": record_id}).fetchone()
    # 最外层if 语句  Outermost1
    if row_cght:
        tedi = "是" if str(row_cght.khmc or "") == "TEDi GmbH & Co. KG" else ""
        cps = 0
        gqsb = 0

        current_now = dt.now()
        rq = current_now.strftime("%Y-%m-%d")
        qsrq = f"{current_now.year}-01-01"
        d1 = current_now.replace(hour=0, minute=0, second=0, microsecond=0)

        htrq_str = str(row_cght.htrq or "").strip()
        D3 = dt.strptime(htrq_str[:10], "%Y-%m-%d") if htrq_str else None

        yjcq_str = str(row_cght.yjcq or "").strip()
        if yjcq_str != "":
            d2 = dt.strptime(yjcq_str[:10], "%Y-%m-%d")
            d = float((d2 - d1).days - 7)
        else:
            gqsb = 1

        # ================= [ Part 3: 是否通过审核 (Outermost 2) ] =================   Outermost2
        if str(row_cght.sfhz) == "通过":
            # 提取前端请求参数替代弹窗 (如果没有传，给默认值)

            # filename1 = j.get('filename1', os.path.join(TEMP_DIR, 'contract_'))   这是保存地址  不需要取

            # 签名图片 获取   对应  224-  245 code   应该使用的时候调用放后面？？？？  todo
            img_png_path = ""
            if dhsb == "1":
                row_tpzx = s.execute(sql_text("SELECT tpmc FROM tpzx WHERE cpbh='陈妍科小'")).fetchone()
                if row_tpzx and row_tpzx.tpmc:
                    img_png_path = os.path.join(
                        config.data_upxpath, json.loads(row_tpzx.tpmc)[0].get("src", "")
                    )  # tpmc 中包含 transfer  形如

                else:
                    return {"code": -1, "msg": "缺失：签名图片", "data": {}}

            # ================= [ Part 4: 基础字段提取 (Outermost 3) ] =================

            gqsb = 1  #  shit param   后面没再使用
            #  tmpcomcg.SQL.Text := 'select * from cght where number=:number';
            gdry_str = str(row_cght.gdry)
            gdry3 = gdry_str if gdry_str else str(row_cght.cgry)
            gdbm3 = str(row_cght.gdbm) if gdry_str else str(row_cght.cgbm)

            bjdh = str(row_cght.hthm or "")
            ywry = str(row_cght.ywry or "")
            cgry2 = str(row_cght.cgry or "")
            gdry2 = gdry_str
            # hthm2 = str(row_cght.hthm or '')
            jhrq2 = str(row_cght.jhrq or "")
            sccj2 = str(row_cght.sccj or "")
            sccj_id2 = str(row_cght.cs_id or "")
            zygcid2 = str(row_cght.sccj1id or "")
            zygcmc2 = str(row_cght.sccj1 or "")

            cslxr, phone, fax, sjhm, twhm, fkhm, bank1, zh1 = "", "", "", "", "", "", "", ""
            if zygcid2 != "":
                row_zycs = s.execute(sql_text("SELECT * FROM zycs WHERE cs_id=:cs_id"), {"cs_id": zygcid2}).fetchone()
                if row_zycs:
                    cslxr = str(row_zycs.cslxr or "")
                    phone = str(row_zycs.phone or "")
                    fax = str(row_zycs.fax or "")
                    sjhm = str(row_zycs.sjhm or "")
                    # address = str(row_zycs.address or '')
                    twhm = str(row_zycs.twhm or "")
                    fkhm = str(row_zycs.fkhm or "")
                    bank1 = str(row_zycs.bank1 or "")  # 后面用到了 但是目前代码遗漏
                    zh1 = str(row_zycs.zh1 or "")

            htrq2 = str(row_cght.htrq or "")
            jsfs = str(row_cght.jsfs or "")

            #  对应  330 - 340 赋值语句

            szdh = ""
            row_ywry = s.execute(sql_text("SELECT ssdq FROM ywrybiao WHERE yhm=:yhm"), {"yhm": cgry2}).fetchone()
            if row_ywry:
                szdh = str(row_ywry.ssdq or "").strip()
            if not szdh:
                szdh = "宁波"

            bzyq = (
                "1.出口五层牛皮纸双瓦楞标准纸箱包装。2.要求无钉箱，胶带工字封箱..."
                if tedi == "是"
                else str(row_cght.bzyq)
            )
            csmc2, csbh2 = (zygcmc2, zygcid2) if zygcmc2 else (sccj2, sccj_id2)

            row_count = s.execute(
                sql_text("SELECT COUNT(*) as sl1s FROM cghtsheet WHERE pid=:pid"), {"pid": record_id}
            ).fetchone()
            if row_count:
                cps = int(row_count.sl1s or 0)

            htjez, htzslz, htzxsz, htzmzz, htzjzz, htztjz = 0.0, 0.0, 0.0, 0.0, 0.0, 0.0

            # ================= [ Part 5: 第一次明细遍历 (金额汇总) ] =================
            items_cght = s.execute(sql_text("SELECT * FROM cghtsheet WHERE pid=:pid"), {"pid": record_id}).fetchall()
            zje1, zxs1 = 0.0, 0.0
            i3 = 0

            if items_cght:
                for row_item in items_cght:
                    htjez += float(row_item.zje or 0)
                    htzslz += float(row_item.cgsl or 0)
                    htzxsz += float(row_item.cgxs or 0)
                    htzmzz += float(row_item.zmz or 0)
                    htzjzz += float(row_item.zjz or 0)
                    htztjz += float(row_item.ztj or 0)

                    zje1 += float(row_item.zje or 0)
                    zxs1 += float(row_item.cgxs or 0)
                    i3 += 1
                    # 原代码中很多赋值语句  但是是在循环内，取到数据之后  不用 而是一味的覆盖  一堆shit  所以循环这里删掉

            # ================= [ Part 6: 诚信工厂风控预警 ] =================
            row_htje_sum = s.execute(
                sql_text(
                    "SELECT SUM(htje) as htje1 FROM cght WHERE (htrq>=:htrq) AND (sccj1id=:sccj1id) AND (sfhz='通过')"
                ),
                {"htrq": qsrq, "sccj1id": zygcid2},
            ).fetchone()
            if row_htje_sum:
                cxjezz = float(row_htje_sum.htje1 or 0)

            sfsh = "不需要"
            if (htjez > cxje1) or (
                cxjezz > cxjez1
            ):  # 对应源代码 大约 554   只要这张合同的金额，或者今年累计跟这家工厂合作的金额，超过了设定的警戒线
                row_cxgc = s.execute(
                    sql_text("SELECT * FROM cxgc WHERE (gcmc=:gcmc) OR (chgc=:chgc)"),
                    {"gcmc": zygcmc2, "chgc": zygcmc2},
                ).fetchone()
                # 供应商诚信报告（反腐）自动建档与过期拦截系统
                if not row_cxgc:  # 系统里从来没有这家工厂的诚信档案
                    now_str, date_str = (
                        dt.now().strftime("%Y-%m-%d %H:%M:%S") or None,
                        dt.now().strftime("%Y-%m-%d") or None,
                    )
                    cxgc_rid = get_uuid()  # 如果您有封装好的 get_uuid()，可以直接改为 jc_rid = get_uuid()
                    user_rid = user.rid  # 获取当前用户的 rid

                    s.execute(
                        sql_text(
                            "INSERT INTO cxgc (rid, uid, ctime, gcmc, ywry, ywz, yrrq, sfsh, yrgs, qrrq, chgc, fkry, xgry, cs_id) \
                        VALUES (:rid, :uid, :ctime, :gcmc, :ywry, :ywz, :yrrq, '否', '宁波优景进出口有限公司', null, :chgc, '', '', :cs_id)"
                        ),
                        {
                            "rid": cxgc_rid,
                            "uid": user_rid,
                            "ctime": now_str,
                            "gcmc": zygcmc2,
                            "ywry": gdry3,
                            "ywz": gdbm3,
                            "yrrq": date_str,
                            "chgc": zygcmc2,
                            "cs_id": zygcid2,
                        },
                    )
                    s.execute(sql_text("UPDATE zycs SET cxhc='否' WHERE cs_id=:cs_id"), {"cs_id": zygcid2})
                    tmpstr2.append(f"工厂名称:{zygcmc2}需提交诚信报告")
                    sfsh, cxsb = "待提供", "1"
                else:  # 系统里有这家工厂的诚信档案
                    qrrq_str = str(row_cxgc.qrrq or "").strip()
                    qrrq_date = (
                        dt.strptime(qrrq_str[:10], "%Y-%m-%d") if qrrq_str else dt.strptime("1999-01-01", "%Y-%m-%d")
                    )
                    if D3 is not None:
                        days_diff = (D3 - qrrq_date).days
                        if days_diff > 365:
                            s.execute(
                                sql_text(
                                    "UPDATE cxgc SET yrrq=:yrrq, ywry=:ywry, ywz=:ywz, sfsh='否', yrgs='宁波优景进出口有限公司', qrrq=null, chgc=:chgc, cs_id=:cs_id WHERE gcmc=:gcmc"
                                ),
                                {
                                    "gcmc": zygcmc2,
                                    "ywry": gdry3,
                                    "ywz": gdbm3,
                                    "yrrq": dt.now().strftime("%Y-%m-%d"),
                                    "chgc": zygcmc2,
                                    "cs_id": zygcid2,
                                },
                            )
                            s.execute(sql_text("UPDATE zycs SET cxhc='否' WHERE cs_id=:cs_id"), {"cs_id": zygcid2})
                            tmpstr2.append(f"工厂名称:{zygcmc2}需提交诚信报告")
                            sfsh, cxsb = "待提供", "1"
                        elif days_diff > 330:
                            tmpstr2.append(f"工厂名称:{zygcmc2}还有一个月需提交诚信报告")
                            cxsb = "1"
                if sfsh != "待提供":
                    sfsh = "已提供"

            # ================= [ Part 7: 回写金额与页数计算 ] =================   对应原来的shit code  里面 大约  638
            try:
                s.execute(
                    sql_text(
                        "UPDATE cght SET webpd1='是', htje=:htje, htzsl=:htzsl, htzxs=:htzxs, htzmz=:htzmz, htzjz=:htzjz, htztj=:htztj, sfsh=:sfsh WHERE rid=:rid"
                    ),
                    {
                        "rid": record_id,
                        "htje": htjez,
                        "htzsl": htzslz,
                        "htzxs": htzxsz,
                        "htzmz": htzmzz,
                        "htzjz": htzjzz,
                        "htztj": htztjz,
                        "sfsh": sfsh,
                    },
                )
                s.commit()
            except Exception as e:
                s.rolllback()
                logger.error(trace_error())
                return {"code": -1, "msg": f"生成失败: {str(e)}", "data": {}}
            htje, htzxs = str(htjez), str(htzxsz)
            # 采购合同汇总页（模板 2）是每页固定排版 4 条明细。
            ys1 = int(cps / 4)
            iz = int(cps / 4)
            ys3 = ys1 if ys1 == (cps / 4.0) else ys1 + 1
            if ys1 == (cps / 4.0):
                iz -= 1

            ndz, dz, lxfs, cgzj, zjl, gstt = _load_htjj_info(s, gs1, szdh)

            # 判定 Excel 模板路径   原先 700-821  100+ 代码  如下 不超过10条......    原本 16个模板  根据之前的 变量  tedi  gs    szdh   sffl
            sffl_str = str(row_cght.sffl or "")
            filename, filename2 = "", ""
            prefix = "uv" if gs == "1" else "ry"
            region = "yw" if szdh == "义乌" else "nb"
            fl = "-FL" if sffl_str == "是" else ""
            td = "tedi" if tedi == "是" else ""
            xlsx_path = os.path.join(config.data_upload_path, "template")
            filename = os.path.join(xlsx_path, f"{prefix}{region}{fl}{td}.xlsx")
            filename2 = os.path.join(xlsx_path, f"{prefix}{region}{fl}1{td}.xlsx")

            # ================= [ Part 8: 覆盖联系信息 ] =================     828 开始

            sccj, hthm, sccj1 = sccj2, bjdh, zygcmc2
            cslxr, phone, fax, sjhm, address, twhm, fkhm, bank1, zh1 = _load_factory_contact(
                s, row_cght, sccj2, sccj_id2, zygcmc2, zygcid2
            )

            htrq, jhrq, cgry, gdry = htrq2, jhrq2, cgry2, gdry2
            jhrq_display = _format_jhrq_for_cell(jhrq)

            lxdh, ydhm, cgjl, lxdh1, ydhm1, lxdh2, gdjl, lxdh3, ydhm2 = "", "", "", "", "", "", "", "", ""

            if cgry != "":
                row_cg = s.execute(sql_text("SELECT * FROM ywrybiao WHERE yhm=:yhm"), {"yhm": cgry}).fetchone()
                if row_cg:
                    cgry, lxdh, ydhm, cgjl = (
                        str(row_cg.ryxm or ""),
                        str(row_cg.lxdh or ""),
                        str(row_cg.ydhm or ""),
                        str(row_cg.bmjl or ""),
                    )
            if cgjl != "":
                row_jl = s.execute(sql_text("SELECT * FROM ywrybiao WHERE yhm=:yhm"), {"yhm": cgjl}).fetchone()
                if row_jl:
                    cgjl, lxdh1, ydhm1 = str(row_jl.ryxm or ""), str(row_jl.lxdh or ""), str(row_jl.ydhm or "")
            if gdry != "":
                row_gd = s.execute(sql_text("SELECT * FROM ywrybiao WHERE yhm=:yhm"), {"yhm": gdry}).fetchone()
                if row_gd:
                    gdry, lxdh2, ydhm2, gdjl, lxdh3 = (
                        str(row_gd.ryxm or ""),
                        str(row_gd.lxdh or ""),
                        str(row_gd.ydhm or ""),
                        str(row_gd.bmjl or ""),
                        str(row_gd.jldh or ""),
                    )

            # ================= [ Part 9~14: 只有一条明细时 (Outermost 4 if) ] =================   对应 927
            out_name_base = config.tmp_path
            out_xls = out_name_base + ".xlsx"
            #  如果明细里面有一条 执行下面操作
            if cps == 1:
                if items_cght:
                    row_item = items_cght[0]
                    cgsl, cgjg, zje = float(row_item.cgsl or 0), float(row_item.cgjg or 0), float(row_item.zje or 0)
                    bjhh, zhwbzh, zwdw, cgxs = (
                        str(row_item.bjhh or ""),
                        str(row_item.zhwbzh or ""),
                        str(row_item.zwdw or ""),
                        str(row_item.cgxs or ""),
                    )
                    wxrl = float(row_item.wxrl or 0)
                    zzsl = float(row_item.zzsl or 0)
                    khhh, cpbh, zycpbh = str(row_item.khhh or ""), str(row_item.ksdhh or ""), str(row_item.zycpbh or "")
                    hhbz, wyzd = str(row_item.hhbz or ""), str(row_item.wyzd or "")

                    cpsm = str(row_item.cpsm or "")
                    ywpm = str(row_item.zwpm or "") + "\r\n" + str(row_item.cpgg or "")
                    nbpm = ywpm
                    zhwbgpm = str(row_item.zhwbgpm or "")
                    qtsm1_val = str(row_item.qtsm1 or "")
                    zwz1, dwz, dwz1 = "", 0, 0
                    cpsm_val = str(row_item.cpsm or "")
                    # 这里的if 判断 目的是 后面计算dwz1
                    if zzsl > 0:
                        zwz1 = "1"

                        # 1. 拼接最终的说明文字
                        qtsm1 = (
                            f"{qtsm1_val}                               \r\n请注意开票货源地为：{row_item.hyd or ''}"
                        )

                        # =========================================================================
                        # 2. 算半角字符，计算 Excel 标红的偏移量 (几十行 OR 语句)
                        # =========================================================================

                        # 包含所有需要判断的半角字符集 (大写)
                        ascii_chars = '1 234567890ABCDEFGHIJKLMNOPQRSTUVWXYZ,.;:!@#$%*()-_=`?<>/\\"{}[ ]&'

                        # 计算 dwz
                        zwz_len = len(qtsm1_val)
                        zs3_count = sum(1 for char in qtsm1_val.upper() if char in ascii_chars)
                        dwz = int((zwz_len - zs3_count) / 2) + zs3_count + 30

                        # 计算 dwz1 (用于合并后的文本)
                        text_to_check = f"{cpsm_val}\r\n{qtsm1_val}"
                        zwz1_len = len(text_to_check)
                        zs3_count1 = sum(1 for char in text_to_check.upper() if char in ascii_chars)
                        dwz1 = int((zwz1_len - zs3_count1) / 2) + zs3_count1 + 30

                    else:
                        qtsm1 = qtsm1_val + "                               "

                    if str(row_item.krddh) != "":
                        qtsm1 += "               \r\n客人订单号为:" + str(row_item.krddh or "")

                    # -----------------   后面开始第九部分   打开指定的excel 然后开始写 -----------------
                    app = _start_excel_app()
                    wb = None
                    red_patches = _RedTextPatches()
                    try:
                        logger.error(f"打开Excel模板: {filename}")
                        wb = _open_workbook(app, filename)
                        ws = wb.sheets[0]

                        result_cn = number_to_chinese_currency(htje) if htje else ""
                        if sccj == "":
                            sccj = sccj1 if sccj1 else twhm
                        if phone != "":
                            if sjhm != "":
                                phone = phone + "\\" + sjhm
                        elif sjhm != "":
                            phone = sjhm

                        is_yiwu = szdh == "义乌"
                        is_fl = sffl_str == "是"
                        # 因为 cps == 1，所以必定是最后一页 (模拟 Delphi 中的 i4 == i3)   i3 总页数  i4 当前页
                        is_last_page = True

                        cells = {}  # 用于统一收集需赋值的坐标
                        logger.error(f"开始填充Excel内容，合同号: {hthm}, 客户: {sccj}, 金额: {htje}")
                        # --- 1. 顶部表头、合同基础信息、金额 ---
                        if is_yiwu:
                            logger.error("使用义乌模板")
                            # 义乌通用头
                            cells.update(
                                {
                                    "A6": gstt,
                                    "A7": ndz,
                                    "A8": lxfs,
                                    "A9": dz,
                                    "B12": sccj,
                                    "C11": htrq,
                                    "M11": hthm,
                                    "H11": szdh,
                                }
                            )
                            if not is_fl:
                                cells.update({"M12": cslxr, "B13": phone, "M13": fax})
                                if is_last_page:
                                    cells.update({"E18": result_cn, "O18": htje})
                            else:
                                cells.update({"K12": cslxr, "B13": phone, "K13": fax})
                                cells.update({"E18": result_cn, "O18": htje})
                        else:
                            logger.error("使用宁波模板")
                            # 宁波通用头
                            cells.update({"B6": gstt, "B7": dz, "B8": lxfs, "D10": htrq})
                            if not is_fl:
                                cells.update(
                                    {"C11": sccj, "N10": hthm, "I10": szdh, "L11": cslxr, "C12": phone, "L12": fax}
                                )
                                if is_last_page:
                                    cells.update({"F17": result_cn, "P17": htje})
                                cells["B24"] = f"四、外箱包装要求：{bzyq}"
                            else:
                                cells.update(
                                    {"I10": szdh, "N10": hthm, "C11": sccj, "L11": cslxr, "C12": phone, "L12": fax}
                                )
                                cells.update({"F17": result_cn, "P17": htje})
                                cells["B24"] = f"四、结算方式: {jsfs}"

                        # --- 2. 底部落款、箱数、交货期、收款账号 ---
                        logger.error(f"填充底部信息，结算方式: {jsfs}, 箱数: {htzxs}, 交货期: {jhrq}")
                        if is_yiwu:
                            logger.error("义乌模板底部赋值")
                            cells["C26"] = htzxs if "箱" in htzxs else f"{htzxs}箱"
                            cells["D27"] = jhrq_display
                            if not is_fl:
                                cells.update(
                                    {
                                        "A31": f"七、结算方式：{jsfs}",
                                        "A48": f"跟单人员: {gdry}({lxdh2}/{ydhm2})",
                                        "I48": f"采购员: {cgry}({lxdh}/{ydhm})",
                                        "A49": f"采购经理: {cgjl}({lxdh1}/{ydhm1})",
                                        "I49": cgzj,
                                        "A50": zjl,
                                        "I50": FF,
                                    }
                                )
                            else:
                                logger.error("义乌模板底部赋值（福利）")
                                cells.update(
                                    {
                                        "A25": f"四、结算方式: {jsfs}",
                                        "A45": f"跟单人员: {gdry}({lxdh2}/{ydhm2})",
                                        "A46": cgzj,
                                        "A47": zjl,
                                        "I47": FF,
                                        "A49": "** 互利共赢  携手并进** 做大靠客人，做强靠厂商 客户和厂商是我们最珍贵的最根本的核心资源 感谢支持！",
                                    }
                                )
                        else:
                            logger.error("宁波模板底部赋值")
                            cells["D25"] = htzxs if "箱" in htzxs else f"{htzxs}箱"
                            cells["E26"] = jhrq_display
                            if not is_fl:
                                cells.update(
                                    {
                                        "B30": f"七、结算方式：{jsfs}",
                                        "B47": f"跟单人员: {gdry}({lxdh2}/{ydhm2})",
                                        "J47": f"采购员: {cgry}({lxdh}/{ydhm})",
                                        "B48": f"采购经理: {cgjl}({lxdh1}/{ydhm1})",
                                        "J48": cgzj,
                                        "B49": zjl,
                                        "J49": FF,
                                    }
                                )
                            else:
                                logger.error("宁波模板底部赋值（福利）")
                                cells.update(
                                    {
                                        "B43": f"跟单人员: {gdry}({lxdh2}/{ydhm2})",
                                        "B44": cgzj,
                                        "B45": zjl,
                                        "J45": FF,
                                        "B47": "** 互利共赢  携手并进** 做大靠客人，做强靠厂商 客户和厂商是我们最珍贵的最根本的核心资源 感谢支持！",
                                    }
                                )

                        # --- 3. 解析产品货号 (cphh2) 并赋值坐标 ---
                        logger.error(f"解析产品货号，客户货号: {khhh}, 采购订单货号: {cpbh}, 主要产品货号: {zycpbh}")
                        if khhh:
                            cphh2 = khhh
                        else:
                            cphh2 = zycpbh if zycpbh else (cpbh if cpbh else bjhh)

                        if is_yiwu:
                            cells["A16"] = cphh2
                        else:
                            cells["B15"] = cphh2

                        # --- 4. 执行所有单元格文本写入 ---
                        logger.error(f"开始写入Excel单元格，待写入内容: {cells}")
                        jhrq_cells = {"E26", "D27"}
                        phone_cells = {"C12", "B13"}
                        for cell_key, cell_val in cells.items():
                            if cell_key in jhrq_cells:
                                safe_write(ws, cell_key, cell_val, num_format=JHRQ_NUM_FORMAT)
                            elif cell_key in phone_cells:
                                safe_write(ws, cell_key, cell_val, num_format=PHONE_CELL_FORMAT)
                            else:
                                safe_write(ws, cell_key, cell_val)

                        # ⭐⭐⭐ [义乌+宁波双轨制 明细赋值、局部红字、图片处理] ⭐⭐⭐

                        # --- 5. 补充产品明细行赋值与局部红字 ---
                        logger.error(
                            f"开始填充明细信息，cgsl: {cgsl}, cgjg: {cgjg}, zje: {zje}, cpsm: {cpsm}, qtsm1: {qtsm1}"
                        )
                        if is_yiwu:
                            # ================= 【义乌版明细】 =================
                            if not is_fl:
                                ws["I16"].value = str(wxrl)
                                ws["J16"].value = zwdw
                                ws["K16"].value = cgxs
                                ws["L16"].value = str(cgsl)
                                ws["N16"].value = str(cgjg)
                                ws["O16"].value = str(zje)
                                ws["C16"].value = ywpm
                                ws["E16"].value = zhwbzh
                                # ws['T17']=cpsm
                                ws["E17"].value = cpsm
                                safe_write(ws, "N17", qtsm1)
                                if zwz1 == "1" and dwz > 0:
                                    red_patches.add("N17", qtsm1, dwz)

                                if _should_write_kppm(zhwbgpm, zzsl):
                                    kppm_line = _kppm_invoice_line(zhwbgpm, zzsl)
                                    ws["N19"].value = kppm_line
                            else:
                                ws["C16"].value = ywpm
                                ws["E16"].value = zhwbzh
                                ws["I16"].value = str(cgsl)
                                ws["K16"].value = zwdw
                                ws["M16"].value = str(cgjg)
                                ws["O16"].value = str(zje)

                                combined_text = f"{cpsm}\r\n{qtsm1}"
                                safe_write(ws, "E17", combined_text)
                                safe_write(ws, "R17", combined_text)
                                if zwz1 == "1" and dwz1 > 0:
                                    red_patches.add("E17", combined_text, dwz1)
                                    red_patches.add("R17", combined_text, dwz1)

                                if _should_write_kppm(zhwbgpm, zzsl):
                                    kppm_line = _kppm_invoice_line(zhwbgpm, zzsl)
                                    ws["S19"].value = kppm_line
                                    ws["N19"].value = kppm_line

                            # 义乌版行高：先开换行再 AutoFit
                            _set_range_center_wrap(ws, "A16:AE16")
                            _set_range_center_wrap(ws, "A17:AE17")
                            _autofit_row_clamped(ws, 16)
                            _autofit_row_clamped(ws, 17)
                            _set_row_height(ws, 19, 16)

                        else:
                            logger.error("使用宁波模板填充明细")
                            # ================= 【宁波版明细】 =================
                            # (摒弃了原 Delphi 中把内容塞进临时格子撑行高又删掉的奇葩操作，直接落子目标格)
                            if not is_fl:
                                ws["D15"].value = nbpm
                                ws["F15"].value = zhwbzh
                                ws["J15"].value = str(wxrl)
                                ws["K15"].value = zwdw
                                ws["L15"].value = cgxs
                                ws["M15"].value = str(cgsl)
                                ws["O15"].value = str(cgjg)
                                ws["P15"].value = str(zje)
                                ws["F16"].value = cpsm

                                safe_write(ws, "N16", qtsm1)
                                if zwz1 == "1" and dwz > 0:
                                    red_patches.add("N16", qtsm1, dwz)
                            else:
                                ws["D15"].value = nbpm
                                ws["F15"].value = zhwbzh
                                ws["J15"].value = str(cgsl)
                                ws["L15"].value = zwdw
                                ws["N15"].value = str(cgjg)
                                ws["P15"].value = str(zje)

                                combined_text = f"{cpsm}\r\n{qtsm1}"
                                safe_write(ws, "F16", combined_text)
                                if zwz1 == "1" and dwz1 > 0:
                                    red_patches.add("F16", combined_text, dwz1)

                        # --- 6. 解析 JSON 路径并插入图片 (适配现有项目规范) --- (兼容义乌 A17 与 宁波 B16) ---  对应的是  代码 1567--1708
                        yytp_json_str = _load_detail_yytp_json(s, is_fl, hhbz, wyzd)
                        logger.error(
                            f"查询图片数据，hhbz: {hhbz}, wyzd: {wyzd}, is_fl: {is_fl}, yytp_len: {len(yytp_json_str)}"
                        )

                        # 解析 JSON 并插入图片    // 代复原？ 原先的delphi代码？
                        # --- 6. 还原 Delphi 动态居中算法并插入图片 ---
                        """
                        if yytp_json_str and yytp_json_str not in ('', '[]', 'None'):
                            try:
                                photo_data = json.loads(yytp_json_str)
                                if photo_data and len(photo_data) > 0:
                                    file_path = photo_data[0].get('src', '')
                                    if file_path:
                                        fn = os.path.join(config.data_upload_path, str(file_path))
                                        if os.path.exists(fn):
                                            img = XLImage(fn)
                                            
                                            # 获取图片原始尺寸
                                            orig_w = img.width
                                            orig_h = img.height
                                            
                                            # 1. 定义目标单元格的物理可用尺寸 (对应 Delphi 的 msexcel.Selection.width/height)
                                            # 假设您的模板里，图片存放格子的标准宽为120，高为100
                                            cell_w, cell_h = 120.0, 100.0 
                                            
                                            # 2. 还原 Delphi 里的减 3 魔法 (留出安全边距)
                                            fwidth = cell_w - 3
                                            fheight = cell_h - 3
                                            
                                            # 3. 还原 Delphi 核心缩放比较公式：
                                            # if MsExcel.Selection.ShapeRange.height >= MsExcel.Selection.ShapeRange.width / (fwidth / fheight)
                                            # 交叉相乘避免除数为0： height * fwidth >= width * fheight
                                            if orig_h * fwidth >= orig_w * fheight:
                                                # 图片偏高瘦：以高度为基准缩放
                                                bz2 = fheight / orig_h
                                                img.height = int(fheight)
                                                img.width = int(orig_w * bz2)
                                                
                                                # 还原 bz1 计算水平偏移
                                                bz1 = int((fwidth - img.width) / 2)
                                                
                                                # 还原初始的 +2 魔法：左移 = 2 + 居中偏移，下移 = 2
                                                x_offset = 2 + bz1
                                                y_offset = 2
                                            else:
                                                # 图片偏矮胖：以宽度为基准缩放
                                                bz2 = fwidth / orig_w
                                                img.width = int(fwidth)
                                                img.height = int(orig_h * bz2)
                                                
                                                # 还原 bz1 计算垂直偏移
                                                bz1 = int((fheight - img.height) / 2)
                                                
                                                # 还原初始的 +2 魔法：左移 = 2，下移 = 2 + 居中偏移
                                                x_offset = 2
                                                y_offset = 2 + bz1
                                                
                                            # 4. 召唤您的黑科技定位函数！
                                            col_idx = 1 if is_yiwu else 2  
                                            row_idx = 17 if is_yiwu else 16
                                            
                                            # 完美注入老代码算出来的 x_offset 和 y_offset
                                            offset_img(img, col_idx, row_idx, x_pad=x_offset, y_pad=y_offset)
                                            
                                            # 绝对不可带单元格坐标
                                            ws.add_image(img)
                                            
                            except Exception as img_err:
                                logger.error(f"单页明细图片插入失败: {trace_error()}")
                        """
                        logger.error(f"准备插入图片，yytp_json_str: {yytp_json_str}")

                        # --- 7. 执行排版格式化 (控制行高、隐藏行、对其方式) ---
                        # 控制收款账户显示与隐藏
                        logger.error(f"控制收款账户显示，sfhs1: {sfhs1}, fkhm: {fkhm}, bank1: {bank1}, zh1: {zh1}")
                        if sfhs1 != "是":
                            bank_info = f"收款户名:{fkhm}      开户银行:{bank1}      银行账号:{zh1}"
                            if is_yiwu:
                                ws["A20"].value = bank_info
                            else:
                                ws["B19"].value = bank_info
                        else:
                            row_to_hide = 20 if is_yiwu else 19
                            _set_row_height(ws, row_to_hide, 0)

                        # 结算方式内容过长时的行高自适应
                        logger.error(f"控制结算方式行高，jsfs: {jsfs}")
                        if len(str(jsfs)) > 94:
                            if is_yiwu:
                                autofit_row = 31 if not is_fl else 25
                            else:
                                autofit_row = 30 if not is_fl else 24
                            _set_row_height(ws, autofit_row, 31.5)

                        # 宁波版的强制居中与换行 + 行高（须在换行开启后再 AutoFit）
                        logger.error("执行宁波版局部居中与换行设置")
                        if not is_yiwu:
                            _autofit_nb_row15(ws, nbpm, zhwbzh)
                            _autofit_nb_row16(ws, is_fl, cpsm, qtsm1)
                            kppm_line = _kppm_invoice_line(zhwbgpm, zzsl)
                            _autofit_nb_row18(ws, kppm_line, is_fl)

                        # 行高确定后再插图（source.pas 先算行高再 AddPicture）
                        if yytp_json_str and yytp_json_str not in ("", "[]", "None"):
                            try:
                                photo_data = json.loads(yytp_json_str)
                                if photo_data and len(photo_data) > 0:
                                    file_path = photo_data[0].get("src", "")
                                    if file_path:
                                        fn = os.path.join(config.data_upload_path, str(file_path))
                                        if os.path.exists(fn):
                                            pic_coord = "A17" if is_yiwu else "B16"
                                            _add_picture_at(ws, pic_coord, fn, x_pad=15, y_pad=10, width=116, height=96)
                                            if not is_yiwu:
                                                img_min_h = max(_get_row_height(ws, 16), 100)
                                                _set_row_height(ws, 16, img_min_h)
                            except Exception as img_err:
                                logger.error(f"单页明细图片插入失败: {trace_error()}")

                        # ⭐⭐⭐ [最新高级重构结束] ⭐⭐⭐

                        # --- 8. 插入电子签名 (原 Delphi dhsb = '1' 逻辑) ---
                        logger.error(f"准备插入签名图片，dhsb: {dhsb}, is_fl: {is_fl}, szdh: {szdh}")
                        if dhsb == "1" and img_png_path and os.path.exists(img_png_path):
                            try:
                                if is_yiwu:
                                    if not is_fl:
                                        sig_coord = "B46"
                                    else:
                                        sig_coord = "C43"
                                else:
                                    if not is_fl:
                                        sig_coord = "D45"
                                    else:
                                        sig_coord = "D41"
                                _add_picture_at(ws, sig_coord, img_png_path, x_pad=2, y_pad=13, width=120, height=50)
                            except Exception as e:
                                logger.error(f"插入签名图片失败: {trace_error()}")
                        logger.error("单页明细内容填充完成，准备保存文件")
                        # 保存 Excel
                        excel_org_path = config.tmp_path
                        # 提取基础文件名，方便后面复用
                        file_base_name = f"{ywry}-{sccj}{hthm}"
                        # 拼装 Excel 的绝对物理路径
                        excel_save_path = os.path.join(excel_org_path, f"{file_base_name}.xlsx")
                        logger.error(f"保存Excel文件，路径: {excel_save_path}")
                        _save_and_close(wb, excel_save_path)
                        red_patches.apply(excel_save_path)
                        logger.error(f"========== 合同(Excel)导出完成 ==========\n")
                        wb = None

                        if pdf == "1":
                            logger.error(
                                f"========== 合同(Excel)导出完成，客户要求Excel格式，返回文件路径 ==========\n"
                            )
                            return {
                                "code": 1,
                                "msg": "生成采购合同成功",
                                "data": {
                                    "path": f"{file_base_name}.xlsx",  # 直接返回文件名即可！
                                    "name": file_base_name,
                                },
                            }
                        # 现代化的 PDF 转换方案 (无缝平替原系统的 doPDF 打印机)
                        if pdf in ["2", "3"]:
                            logger.error(f"客户要求生成PDF，开始转换Excel到PDF，路径: {excel_save_path}")
                            # 第一个参数是 Excel文件绝对路径，第二个参数是 输出目录
                            pdf_result = convert_excel_to_pdf(excel_save_path, excel_org_path)
                            logger.error(f"PDF转换结果: {pdf_result}")
                            if pdf_result.get("success"):
                                pdf_name = f"{file_base_name}.pdf"

                                # 💡 优良传统：过河拆桥！转换成功后删掉 Excel 底稿
                                if os.path.exists(excel_save_path):
                                    os.remove(excel_save_path)

                                logger.info(f"========== 合同(PDF)导出完成 ==========\n")

                                return {
                                    "code": 1,
                                    "msg": "生成采购合同PDF成功",
                                    "data": {"path": f"{file_base_name}.pdf", "name": file_base_name},
                                }
                            else:
                                logger.error(f"[X] PDF转换失败: {pdf_result.get('error')}")
                                return {"code": -1, "msg": f"PDF转换失败: {pdf_result.get('error')}", "data": {}}
                    except Exception as e:
                        logger.error(trace_error())
                        return {"code": -1, "msg": "生成单页明细失败", "data": {trace_error()}}
                    finally:
                        if wb is not None:
                            wb.close()
                        app.quit()

            # ================= [ Part 15~16: 多条明细时 (Outermost 4 else) ] =================
            else:
                app = _start_excel_app()
                wb = None
                # 封面 O18 开票品名：取首条明细（与 Pascal 循环末值一致，单明细时即该条）
                zhwbgpm, zzsl = "", 0.0
                if items_cght:
                    _first_item = items_cght[0]
                    zhwbgpm = str(_first_item.zhwbgpm or "")
                    zzsl = float(_first_item.zzsl or 0)
                try:
                    logger.error(f"打开Excel模板: {filename2}")
                    wb = _open_workbook(app, filename2)
                    ws_cover = wb.sheets[0]  # 多页模板的第一页是【封面】

                    if htje != "":
                        result_cn = number_to_chinese_currency(zje1) if zje1 else ""
                    else:
                        result_cn = ""

                    if sccj == "":
                        sccj = sccj1 if sccj1 else twhm
                    if phone != "":
                        if sjhm != "":
                            phone = phone + "\\" + sjhm
                    elif sjhm != "":
                        phone = sjhm

                    # ⭐⭐⭐ [多页模板封面 (Cover Page) 赋值映射] ⭐⭐⭐
                    is_yiwu = szdh == "义乌"
                    is_fl = sffl_str == "是"
                    cover_cells = {}

                    if is_yiwu:
                        cover_cells.update(
                            {
                                "A6": gstt,
                                "A7": ndz,
                                "A8": lxfs,
                                "A9": dz,
                                "B12": sccj,
                                "M11": hthm,
                                "H11": szdh,
                                "C26": f"{zxs1}箱",
                                "D27": jhrq_display,
                                "A17": f"具体见附页共{ys3}页",
                            }
                        )
                        if not is_fl:
                            cover_cells.update(
                                {
                                    "M12": cslxr,
                                    "B13": phone,
                                    "M13": fax,
                                    "C11": htrq,
                                    "K16": 0,
                                    "L16": 0,
                                    "O16": 0,
                                    "E18": result_cn,
                                    "O18": zje1,
                                }
                            )
                            cover_cells.update(
                                {
                                    "A31": f"七、结算方式：{jsfs}",
                                    "A48": f"跟单人员: {gdry}({lxdh2}/{ydhm2})",
                                    "I48": f"采购员: {cgry}({lxdh}/{ydhm})",
                                    "A49": f"采购经理: {cgjl}({lxdh1}/{ydhm1})",
                                    "I49": cgzj,
                                    "A50": zjl,
                                    "I50": FF,
                                    "A25": f"四、外箱包装要求{bzyq}",
                                }
                            )
                            if len(str(jsfs)) > 94:
                                _set_row_height(ws_cover, 31, 31.5)
                        else:
                            cover_cells.update(
                                {
                                    "C11": htrq,
                                    "K12": cslxr,
                                    "B13": phone,
                                    "K13": fax,
                                    "K16": 0,
                                    "L16": 0,
                                    "O16": 0,
                                    "E18": result_cn,
                                    "O18": zje1,
                                }
                            )
                            cover_cells.update(
                                {
                                    "A25": f"四、结算方式: {jsfs}",
                                    "A45": f"跟单人员: {gdry}({lxdh2}/{ydhm2})",
                                    "A46": cgzj,
                                    "A47": zjl,
                                    "I47": FF,
                                    "A49": "** 互利共赢  携手并进** 做大靠客人，做强靠厂商 客户和厂商是我们最珍贵的最根本的核心资源 感谢支持！",
                                }
                            )
                            if len(str(jsfs)) > 94:
                                _set_row_height(ws_cover, 25, 31.5)
                    else:
                        # 宁波封面
                        if not is_fl:
                            cover_cells.update(
                                {
                                    "B6": gstt,
                                    "B7": dz,
                                    "B8": lxfs,
                                    "C11": sccj,
                                    "N10": hthm,
                                    "I10": szdh,
                                    "L11": cslxr,
                                    "C12": phone,
                                    "L12": fax,
                                    "D10": htrq[:10],
                                    "K16": 0,
                                    "L16": 0,
                                    "O16": 0,
                                    "F17": result_cn,
                                    "P17": zje1,
                                }
                            )
                            cover_cells.update(
                                {
                                    "B24": f"四、外箱包装要求：{bzyq}",
                                    "D25": f"{zxs1}箱",
                                    "E26": jhrq_display,
                                    "B30": f"七、结算方式：{jsfs}",
                                    "B47": f"跟单人员: {gdry}({lxdh2}/{ydhm2})",
                                    "J47": f"采购员: {cgry}({lxdh}/{ydhm})",
                                    "B48": f"采购经理: {cgjl}({lxdh1}/{ydhm1})",
                                    "J48": cgzj,
                                    "B49": zjl,
                                    "J49": FF,
                                    "B16": f"具体见附页共{ys3}页",
                                }
                            )
                            if len(str(jsfs)) > 94:
                                _set_row_height(ws_cover, 30, 31.5)
                            kppm_line = _kppm_invoice_line(zhwbgpm, zzsl)
                            if kppm_line:
                                cover_cells["O18"] = kppm_line
                        else:
                            cover_cells.update(
                                {
                                    "D10": htrq,
                                    "I10": szdh,
                                    "N10": hthm,
                                    "C11": sccj,
                                    "L11": cslxr,
                                    "C12": phone,
                                    "L12": fax,
                                    "K16": 0,
                                    "L16": 0,
                                    "O16": 0,
                                    "F17": result_cn,
                                    "P17": zje1,
                                }
                            )
                            cover_cells.update(
                                {
                                    "B24": f"四、结算方式: {jsfs}",
                                    "D25": f"{zxs1}箱",
                                    "E26": jhrq_display,
                                    "B16": f"具体见附页共{ys3}页",
                                    "B43": f"跟单人员: {gdry}({lxdh2}/{ydhm2})",
                                    "B44": cgzj,
                                    "B45": zjl,
                                    "J45": FF,
                                    "B47": "** 互利共赢  携手并进** 做大靠客人，做强靠厂商 客户和厂商是我们最珍贵的最根本的核心资源 感谢支持！",
                                }
                            )
                            if _should_write_kppm(zhwbgpm, zzsl):
                                cover_cells["O18"] = _kppm_invoice_line(zhwbgpm, zzsl)
                            if len(str(jsfs)) > 94:
                                _set_row_height(ws_cover, 24, 31.5)

                    # 统一写入封面
                    jhrq_cells = {"E26", "D27"}
                    phone_cells = {"C12", "B13"}
                    for k_cell, v_cell in cover_cells.items():
                        if k_cell in jhrq_cells:
                            safe_write(ws_cover, k_cell, v_cell, num_format=JHRQ_NUM_FORMAT)
                        elif k_cell in phone_cells:
                            safe_write(ws_cover, k_cell, v_cell, num_format=PHONE_CELL_FORMAT)
                        else:
                            safe_write(ws_cover, k_cell, v_cell)

                    # 封面收款账户
                    if sfhs1 != "是":
                        bank_info = f"收款户名:{fkhm}      开户银行:{bank1}      银行账号:{zh1}"
                        if is_yiwu:
                            ws_cover["A20"] = bank_info
                        else:
                            ws_cover["B19"] = bank_info
                    else:
                        row_to_hide = 20 if is_yiwu else 19
                        _set_row_height(ws_cover, row_to_hide, 0)

                    # 封面签名
                    if dhsb == "1" and img_png_path and os.path.exists(img_png_path):
                        try:
                            if is_yiwu:
                                if not is_fl:
                                    sig_coord = "B46"
                                else:
                                    sig_coord = "C43"
                            else:
                                if not is_fl:
                                    sig_coord = "D45"
                                else:
                                    sig_coord = "D41"
                            _add_picture_at(ws_cover, sig_coord, img_png_path, x_pad=2, y_pad=10, width=120, height=50)

                        except Exception as e:
                            logger.error(f"插入封面签名图片失败: {trace_error()}")
                    # ⭐⭐⭐ [封面逻辑结束] ⭐⭐⭐

                    # ================= [ 动态复制明细分页 ] =================
                    # 原系统 filename2 中，sheet[0]是封面，sheet[1]是明细模板
                    detail_template = wb.sheets[1]
                    for az in range(iz):
                        detail_template.copy(after=wb.sheets[-1])

                    k, i4, i5 = 0, 0, 0
                    page_kppm_list = []  # 收集当前页开票品名 (zhwbgpm, zzsl)

                    for row_item in items_cght:
                        k += 1
                        i5 += 1
                        if k == 1:
                            i4 += 1
                            ws_detail = wb.sheets[i4]  # 取出新一页的明细 sheet
                            ws_detail["A5"], ws_detail["K5"] = f"我司合同号:{bjdh}", f"★共{ys3}页,第{i4}页"
                            page_kppm_list.clear()

                        # 提取变量
                        khhh = str(row_item.khhh or "")
                        cphh2 = (
                            str(row_item.khhh or "")
                            or str(row_item.ksdhh or "")
                            or str(row_item.zycpbh or "")
                            or str(row_item.bjhh or "")
                        )
                        ywpm = str(row_item.zwpm or "") + "\r\n" + str(row_item.cpgg or "")
                        zhwbzh = str(row_item.zhwbzh or "")
                        cpsm = str(row_item.cpsm or "")
                        qtsm1_val = str(row_item.qtsm1 or "")
                        zzsl = float(row_item.zzsl or 0)
                        cpbh = str(row_item.ksdhh or "")

                        # ⭐⭐⭐ [附页明细填入] ⭐⭐⭐
                        row_idx = 8 + k
                        ws_detail[f"A{row_idx}"] = cphh2
                        ws_detail.range(f"A{row_idx}").number_format = "@"
                        ws_detail[f"C{row_idx}"] = ywpm
                        ws_detail[f"D{row_idx}"] = zhwbzh
                        ws_detail[f"E{row_idx}"] = str(row_item.wxrl or 0)
                        ws_detail[f"F{row_idx}"] = str(row_item.zwdw or "")
                        ws_detail[f"G{row_idx}"] = str(row_item.cgxs or "")
                        ws_detail[f"H{row_idx}"] = str(row_item.cgsl or 0)
                        ws_detail[f"I{row_idx}"] = str(row_item.cgjg or 0)
                        ws_detail[f"J{row_idx}"] = float(row_item.zje or 0)

                        # --- 3. 解析产品货号 (cphh2) 并赋值坐标 ---

                        # 收集开票品名
                        zhwbgpm = str(row_item.zhwbgpm or "")
                        if _should_write_kppm(zhwbgpm, zzsl):
                            entry = (zhwbgpm, zzsl)
                            if entry not in page_kppm_list:
                                page_kppm_list.append(entry)

                        if khhh:
                            cphh2 = khhh
                        else:
                            cphh2 = zycpbh if zycpbh else (cpbh if cpbh else bjhh)
                            # 1. 提前准备好客人订单号的后缀 (全局共用)

                        # ⭐⭐⭐ [附页局部红字逻辑] ⭐⭐⭐
                        krddh_val = str(row_item.krddh or "")
                        krddh_suffix = (
                            f"                              \r\n客人订单号为:{krddh_val}" if krddh_val else ""
                        )

                        if zzsl > 0:
                            # 严格还原老代码：富文本截断点 dwz1 只计算最原始的 cpsm 和 qtsm1
                            ascii_chars = "1 234567890ABCDEFGHIJKLMNOPQRSTUVWXYZ,.;:!@#$%*()-_=`?<>/\\{}[+]&"
                            calc_text = f"{cpsm}\r\n{qtsm1_val}"
                            zs3_count1 = sum(1 for char in calc_text.upper() if char in ascii_chars)
                            dwz1 = int((len(calc_text) - zs3_count1) / 2) + zs3_count1 + 30

                            # 2. 组装红字后缀：货源地 + 客人订单号 (完美契合老代码的追加顺序)
                            hyd_suffix = f"                              \r\n请注意开票货源地为：{row_item.hyd or ''}"
                            total_suffix = hyd_suffix + krddh_suffix
                            safe_write(ws_detail, f"K{row_idx}", calc_text + total_suffix)

                        else:
                            # zzsl <= 0 的情况：没有货源地，只有原始说明 + 客人订单号
                            base_text = f"{cpsm}\r\n{qtsm1_val}                              "

                            # 直接拼接全局的订单号后缀写入
                            ws_detail[f"K{row_idx}"] = base_text + krddh_suffix

                        # 明细行格式与行高保护
                        _set_row_height(ws_detail, row_idx, max(95, min(400, 95 + len(ywpm) * 2)))

                        _set_range_center_wrap(ws_detail, f"A{row_idx}:AE{row_idx}")

                        # ⭐⭐⭐ [附页零 IO 图片插入] ⭐⭐⭐     待修改   3-23 - 15：06
                        hhbz, wyzd = str(row_item.hhbz or ""), str(row_item.wyzd or "")
                        yytp_json_str = _load_detail_yytp_json(s, is_fl, hhbz, wyzd)

                        if yytp_json_str and yytp_json_str not in ("", "[]", "None"):
                            try:
                                photo_data = json.loads(yytp_json_str)
                                if photo_data and len(photo_data) > 0:
                                    file_path = photo_data[0].get("src", "")
                                    if file_path:
                                        fn = os.path.join(config.data_upload_path, str(file_path))
                                        if os.path.exists(fn):
                                            _add_picture_at(
                                                ws_detail, f"B{row_idx}", fn, x_pad=15, y_pad=10, width=116, height=96
                                            )
                            except Exception as img_err:
                                logger.error(f"附页明细图片插入失败: {trace_error()}")
                        # ⭐⭐⭐ [附页表尾处理 (排满 4 条或最后一条)] ⭐⭐⭐
                        if k == 4 or i5 == cps:
                            # 1. 一行代码降维合并开票品名
                            kppm_str = ";".join(
                                [_kppm_invoice_line(pm, sl) for pm, sl in page_kppm_list if _kppm_invoice_line(pm, sl)]
                            )
                            if kppm_str:
                                ws_detail["I14"] = kppm_str
                                _set_wrap_text(ws_detail, "I14")

                            # 2. 还原“奇葩覆盖”的人员排版
                            if not is_fl:
                                ws_detail["A17"] = f"采购员: {cgry}({lxdh}/{ydhm})"
                                ws_detail["D17"] = f"采购经理: {cgjl}({lxdh1}/{ydhm1})"
                            else:
                                ws_detail["A17"] = f"跟单人员: {gdry}({lxdh2}/{ydhm2})"

                            ws_detail["A18"], ws_detail["A19"], ws_detail["D19"] = cgzj, zjl, FF

                            # 3. 附页电子签名 (找回丢失的 +2 偏移)
                            if dhsb == "1" and img_png_path and os.path.exists(img_png_path):
                                try:
                                    _add_picture_at(
                                        ws_detail, "B16", img_png_path, x_pad=2, y_pad=0, width=120, height=50
                                    )
                                except Exception as e:
                                    logger.error(f"附页签名插入失败: {trace_error()}")

                            # 4. 彻底还原：清空品名列表，重置当前页商品计数器，准备排下一页
                            page_kppm_list.clear()
                            k = 0

                    # 导出与拆分逻辑
                    excel_org_path = config.tmp_path
                    file_base_name = f"{ywry}-{sccj}{hthm}"
                    excel_save_path = os.path.join(excel_org_path, f"{file_base_name}.xlsx")

                    _save_and_close(wb, excel_save_path)
                    wb = None
                    if pdf == "1":
                        return {
                            "code": 1,
                            "msg": "生成采购合同成功",
                            "data": {"path": f"{file_base_name}.xlsx", "name": file_base_name},
                        }

                    if pdf in ["2", "3"]:
                        generated_pdfs = []
                        for ts in range(1, int(ys3) + 2):
                            # 1. 确定单页的临时 XLSX 绝对路径
                            if ts == 1:
                                out_temp = os.path.join(excel_org_path, f"{file_base_name}_cover_temp.xlsx")
                            else:
                                out_temp = os.path.join(excel_org_path, f"{file_base_name}副本{ts - 1}.xlsx")

                            # 2. 内存级安全拆分
                            temp_wb = _open_workbook(app, excel_save_path)
                            try:
                                target_idx = ts - 1
                                for i in range(len(temp_wb.sheets) - 1, -1, -1):
                                    if i != target_idx:
                                        temp_wb.sheets[i].delete()
                                _save_and_close(temp_wb, out_temp)
                                temp_wb = None
                            except Exception:
                                if temp_wb is not None:
                                    temp_wb.close()
                                raise

                            # 3. 统一调用项目自带的 Java/JAR 转换引擎！
                            pdf_result = convert_excel_to_pdf(out_temp, excel_org_path)

                            # 如果转换失败，记录日志并跳过/退出
                            if not pdf_result.get("success"):
                                logger.error(f"⚠️ 多页PDF转换失败(第{ts}页): {pdf_result.get('error')}")
                                # 可以根据业务需求选择 continue 跳过，或者直接返回报错
                                continue

                            generated_pdfs.append(pdf_result.get("pdf_path"))

                            try:
                                if os.path.exists(out_temp):
                                    os.remove(out_temp)
                            except Exception:
                                logger.warning(f"删除临时 xlsx 失败: {out_temp}")
                            # 4. 封面重命名还原
                            # if ts == 1 and pdf_generated and os.path.exists(pdf_generated):
                            #     pdf_target = os.path.join(excel_org_path, f"{file_base_name}.pdf")
                            #     if os.path.exists(pdf_target):
                            #         os.remove(pdf_target)
                            #     os.rename(pdf_generated, pdf_target)

                        if generated_pdfs:
                            zip_path = os.path.join(excel_org_path, f"{file_base_name}.zip")
                            try:
                                with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                                    for p in generated_pdfs:
                                        if p and os.path.exists(p):
                                            zf.write(p, os.path.basename(p))
                                            try:
                                                os.remove(p)
                                            except Exception:
                                                logger.warning(f"删除临时 pdf 失败: {p}")
                            except Exception as e:
                                logger.error(f"打包 ZIP 失败: {trace_error()}")
                                # 若打包失败，退回到原来的行为：将第一份 PDF 作为 target_pdf（保持兼容）
                                try:
                                    cover_pdf = generated_pdfs[0]
                                    target_pdf = os.path.join(excel_org_path, f"{file_base_name}.pdf")
                                    if os.path.exists(target_pdf):
                                        os.remove(target_pdf)
                                    os.rename(cover_pdf, target_pdf)
                                    return {
                                        "code": 1,
                                        "msg": "生成主采购合同PDF成功",
                                        "data": {"path": f"{file_base_name}.pdf", "name": file_base_name},
                                    }
                                except Exception:
                                    return {"code": -1, "msg": "生成PDF/ZIP失败", "data": {}}

                        if os.path.exists(excel_save_path):
                            os.remove(excel_save_path)

                        # cover_temp_xls = os.path.join(excel_org_path, f"{file_base_name}_cover_temp.xlsx")
                        # if os.path.exists(cover_temp_xls):
                        #     os.remove(cover_temp_xls)

                        # for ts in range(2, int(ys3) + 2):
                        #     temp_fuben_xls = os.path.join(excel_org_path, f"{file_base_name}副本{ts - 1}.xlsx")
                        #     if os.path.exists(temp_fuben_xls):
                        #         os.remove(temp_fuben_xls)

                        logger.info(f"========== 多页合同(PDF)拆分导出完成 ==========\n")

                        # return json_result(1, "生成采购合同PDF成功",
                        #     data ={
                        #         'path': f"{file_base_name}.xlsx", # 直接返回文件名即可！
                        #         'name': file_base_name
                        #     }
                        # )
                        return {
                            "code": 1,
                            "msg": "生成采购合同ZIP成功",
                            "data": {"path": f"{file_base_name}.zip", "name": file_base_name},
                        }

                except Exception as e:
                    logger.error(f"多页模板生成失败: {trace_error()}")
                    return {"code": -1, "msg": "多页模板生成失败", "data": {}}
                finally:
                    if wb is not None:
                        wb.close()
                    app.quit()
    return {"code": -1, "msg": "合同未通过审核或不存在", "data": {}}


@any_route("/api/saier/purchase_followup/export", methods=["POST"])
@require_token
async def api_export_purchase_followup(request):
    """
    采购跟单 - 优景采购合同批量（图）导出。

    参数：da2, rid, rid_list, gs, pdf（含义同 Delphi 弹窗与采购合同导出）
    """
    logger.error("========== 收到采购跟单合同(图)导出请求 ==========")
    j = await _parse_followup_request_body(request)
    if not j:
        return json_result(-1, "请求参数为空，请使用 JSON 或表单提交")

    da2 = str(j.get("da2") or "1").strip()
    rid = j.get("rid") or j.get("record_id")
    rid_list = _parse_rid_list(j.get("rid_list") or j.get("rids"))

    gs = str(j.get("gs") or "").strip()
    if not gs:
        return json_result(-1, "参数缺失：公司 gs")
    if gs in ("", "1"):
        gs, gs1 = "1", "优景进出口"
    elif gs == "2":
        gs, gs1 = "2", "锐亿进出口"
    else:
        gs1 = gs
        gs = "2"

    pdf = str(j.get("pdf") or "").strip()
    if not pdf:
        return json_result(-1, "参数缺失：导出格式 pdf")

    user = request.current_user
    username = user.username
    tmpstr2 = []

    s = Session()
    try:
        pdf, dhsb, dhsb3 = _init_pdf_permissions(s, username, pdf)

        hthm_list, err = _collect_hthm_list(s, da2, rid, rid_list, username)
        if err:
            return json_result(-1, err)

        generated_files = []
        errors = []

        for hthm in hthm_list:
            cght_rid = _resolve_cght_rid(s, hthm)
            if not cght_rid:
                errors.append(f"{hthm}: 未找到已审核通过的采购合同")
                continue
            try:
                result = _export_single_contract(s, user, cght_rid, gs, gs1, pdf, username, tmpstr2, dhsb, dhsb3)
            except Exception as e:
                logger.error(trace_error())
                errors.append(f"{hthm}: {str(e)}")
                continue
            if result.get("code") != 1:
                errors.append(f"{hthm}: {result.get('msg', '导出失败')}")
                continue
            data = result.get("data") or {}
            rel = str(data.get("path") or "").strip()
            abs_path = _abs_output_path(rel)
            if abs_path and os.path.exists(abs_path):
                generated_files.append(
                    {
                        "hthm": hthm,
                        "path": rel,
                        "abs_path": abs_path,
                        "name": data.get("name") or os.path.splitext(os.path.basename(rel))[0],
                    }
                )
            else:
                errors.append(f"{hthm}: 文件未生成")

        if not generated_files:
            return json_result(-1, f"导出失败: {'; '.join(errors) if errors else '无成功文件'}")

        warning_msg = ""
        warning_path = ""
        if any("诚信" in x for x in tmpstr2):
            warning_path = os.path.join(config.data_upload_path, f"{dt.now().strftime('%Y-%m-%d')}_诚信报告.txt")
            with io.open(warning_path, "w", encoding="utf-8") as f:
                f.write("\n".join(tmpstr2))
            warning_msg = f"有需提交诚信报告的工厂，已生成日志: {warning_path}"

        if len(generated_files) == 1:
            one = generated_files[0]
            data_out = {"path": one["path"], "name": one["name"], "hthm": one["hthm"]}
            if warning_msg:
                data_out["warning"] = warning_msg
                data_out["warning_path"] = warning_path
            msg = "采购跟单合同导出成功"
            if errors:
                msg += f"（跳过: {'; '.join(errors)}）"
            return json_result(1, msg, data_out)

        zip_base = f"采购跟单合同_{username}_{dt.now().strftime('%Y%m%d%H%M%S')}"
        zip_rel = f"{zip_base}.zip"
        zip_abs = os.path.join(config.tmp_path, zip_rel)
        with zipfile.ZipFile(zip_abs, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for item in generated_files:
                zf.write(item["abs_path"], os.path.basename(item["abs_path"]))
                try:
                    os.remove(item["abs_path"])
                except Exception:
                    pass
        data_out = {
            "path": zip_rel,
            "name": zip_base,
            "count": len(generated_files),
            "hthm_list": [x["hthm"] for x in generated_files],
        }
        if warning_msg:
            data_out["warning"] = warning_msg
            data_out["warning_path"] = warning_path
        msg = f"批量导出成功，共 {len(generated_files)} 份"
        if errors:
            msg += f"；跳过: {'; '.join(errors[:5])}"
        return json_result(1, msg, data_out)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f"采购跟单导出异常: {str(e)}")
    finally:
        try:
            s.close()
        except Exception:
            pass
