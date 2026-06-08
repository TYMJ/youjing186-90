
from any import *
from .model import *
from .__default__ import get_user_path, module_share_new, module_xxck_new, user_task_delete, user_task_new
import os,math,re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU, pixels_to_points
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from sqlalchemy import text, func
from decimal import Decimal,ROUND_HALF_UP
from typing import List, Dict, Any
from datetime import datetime, timedelta
import calendar



"""
开票通知.唯一字段
对应原Pascal: 唯一字段
"""
@any_route('/api/saier/billed_notice/save/before', methods=['POST'])
@require_token
async def view_billed_notice_save_before(request):
    s = Session()
    try:
        j = await request.json()
        user = request.current_user
        x = j.get('main', {})
        ysfp = x.get('ysfp', '')
        sccj = x.get('sccj', '')
        rid = x.get('rid', '')
        kfdy = x.get('kfdy', '')
        lines = j.get('lines', []) # 对应取合并资料列表
        items = j.get('items', []) # 对应取开票详情列表
        code = 1
        msg = '操作成功'
        wstt = x.get('wstt', '')
        org = get_user_path('zjnblh')
        position = org.get('position', '')
        bm1 = position[:1]
        kpxhn = time.strftime('%Y')[2:4]
        kpxhy = time.strftime('%m')
        kpxh1 = str(kpxhn) + str(kpxhy)
        kpxh2 = str(bm1) + str(kpxh1)
        index = 0
        kpxh_json = {}
        kdy = ''
        c = s.query(kaiptz).filter(kaiptz.ysfp == ysfp, kaiptz.sccj == sccj, kaiptz.rid != rid).first()
        if c:
            code = -1
            msg = '此生产厂家在这个外销发票号下已有开票通知请检查'
            kdy = c.kfdy
        for l in lines:
            index = index + 1
            kpxh = l.get('kpxh', '') if l.get('kpxh', '') else ''
            if kpxh != '' and kpxh != None:
                c = s.query(kaiptzsheet1).filter(kaiptzsheet1.kpxh == kpxh, kaiptzsheet1.pid != rid).first()
                if c:
                    l['kpxh'] = ''
                    kpxh = ''
            if kpxh == '' or kpxh == None:
                c = s.query(kaiptzsheet1.kpxh).filter(kaiptzsheet1.kpxh.like(f'{kpxh2}%')).order_by(kaiptzsheet1.kpxh.desc()).first()
                if c:
                    last_kpxh = c.kpxh if c.kpxh else '00000'
                    last_index_str = int(last_kpxh[5:10]) + 1
                    kpxh = kpxh2 + str(last_index_str).zfill(5) + str(index).zfill(3)
                    l['kpxh'] = kpxh
                    kpxh_json[l.get('rid')] = kpxh

            if kdy != kfdy and kfdy == '可以':
                ywry = l.get('ywrya', '') if l.get('ywrya', '') else ''
                if ywry == '' or ywry == None or ywry == user.username:
                    continue
                
                xxnr = '外销发票号码为:' + str(l.get('ysfp', '')) + '中文品名:' + str(l.get('zwpm', '')) + '工厂:' + str(l.get('sccj', '')) + '开票资料可打印' + time.strftime('%Y-%m-%d %H:%M:%S')
                row = {
                    "xxly": '开票通知',
                    "gdht": '',
                    "wxht": '',
                    "cght": '',
                    "yhdh": '',
                    "xxnr": xxnr,
                    "jsr": str(ywry),
                    "sys_path": "",
                    "spsq": user.username
                }
                res = module_xxck_new([row],user,s)
                if res.get('code')!=1:
                    s.rollback()
                    return json_result(res.get('code'), res.get('code'))
                
        for l in items:
            gcfp = l.get('gcfp', '') if l.get('gcfp', '') else ''
            kpgc = l.get('kpgc', '') if l.get('kpgc', '') else ''
            c = s.query(zlbinvoice).filter(zlbinvoice.NumberOfInvoice == gcfp, zlbinvoice.kpgcsh == kpgc, zlbinvoice.NameOfPurchaser == wstt).first()
            if c:
                l['fpdm'] = c.NumberOfInvoice
                l['fpje'] = float(c.TaxIncludedAmountInFigures) if c.TaxIncludedAmountInFigures else 0.0
                l['bhsj'] = float(c.TotalAmountExcludingTax) if c.TotalAmountExcludingTax else 0.0
                l['se'] = float(c.TotalTaxAmount) if c.TotalTaxAmount else 0.0
                l['sh'] = c.kpgcsh
                l['UniqueCodeOfInvoice'] = c.UniqueCodeOfInvoice

                if l.get('UniqueCodeOfInvoice') != '' and l.get('UniqueCodeOfInvoice') != None:
                    c.fis_use = '是'
                    s.add(c)

        s.commit()

        return json_result(code, msg, lines)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'保存校验失败: {str(e)}')
    finally:
        s.close()

"""
开票通知.开票详情.delete
对应原Pascal: 开票详情.delete
"""
@any_route('/api/saier/billed_notice/child/before/delete', methods=['POST'])
@require_token
async def view_billed_notice_child_before_delete(request):
    s = Session()
    try:
        j = await request.json()
        user = request.current_user
        x = j.get('main', {})
        wstt = x.get('wstt', '')
        items = j.get('items', []) # 对应取开票详情列表
        code = 1
        msg = '操作成功'

        s.query(zlbinvoice).filter(zlbinvoice.UniqueCodeOfInvoice==j.get('fpyz')).update({zlbinvoice.fis_use: '否'}, synchronize_session=False)

        s.commit()

        return json_result(code, msg)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'保存校验失败: {str(e)}')
    finally:
        s.close()



"""
开票通知.唯一字段
对应原Pascal: 唯一字段
"""
@any_route('/api/saier/billed_notice/save/after', methods=['POST'])
@require_token
async def view_billed_notice_save_after(request):
    s = Session()
    try:
        j = await request.json()
        user = request.current_user
        x = j.get('main', {})
        wstt = x.get('wstt', '')
        items = j.get('items', []) # 对应取开票详情列表
        code = 1
        msg = '操作成功'

        for l in items:
            gcfp = l.get('gcfp', '') if l.get('gcfp', '') else ''
            kpgc = l.get('kpgc', '') if l.get('kpgc', '') else ''
            c = s.query(zlbinvoice).filter(zlbinvoice.NumberOfInvoice == gcfp, zlbinvoice.kpgcsh == kpgc, zlbinvoice.NameOfPurchaser == wstt).first()
            if c:
                l['fpdm'] = c.NumberOfInvoice
                l['fpje'] = float(c.TaxIncludedAmountInFigures) if c.TaxIncludedAmountInFigures else 0.0
                l['bhsj'] = float(c.TotalAmountExcludingTax) if c.TotalAmountExcludingTax else 0.0
                l['se'] = float(c.TotalTaxAmount) if c.TotalTaxAmount else 0.0
                l['sh'] = c.kpgcsh
                l['UniqueCodeOfInvoice'] = c.UniqueCodeOfInvoice

                if l.get('UniqueCodeOfInvoice') != '' and l.get('UniqueCodeOfInvoice') != None:
                    c.fis_use = '是'
                    s.add(c)

        s.commit()

        return json_result(code, msg)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'保存校验失败: {str(e)}')
    finally:
        s.close()

"""
开票通知.义乌确认
对应原Pascal: 义乌确认
"""
@any_route('/api/saier/billed_notice/load/check', methods=['POST'])
@require_token
async def view_billed_notice_load_check(request):
    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        code = 1
        msg = '操作成功'
        data = {}
        d = s.query(zx.cs).filter(zx.ly == '合同收回金额').first()
        if d:
            data['htsh'] = d.cs

        org = get_user_path(user.username)
        data['position'] = org.get('position', '')
        
        return json_result(code, msg, data)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'加载检查失败: {str(e)}')
    finally:
        s.close()


"""
开票通知.合并资料 新建、复制后
对应原Pascal: 合并资料 新建、复制后
"""
@any_route('/api/saier/billed_notice/merge/new/copy/after', methods=['POST'])
@require_token
async def view_billed_notice_merge_new_copy_after(request):
    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        code = 1
        msg = '操作成功'
        data = {'cs': 30}
        d = s.query(zx.cs).filter(zx.ly == '开票逾期').first()
        if d:
            data['cs'] = d.cs

        return json_result(code, msg, data)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'出货日期修改检查失败: {str(e)}')
    finally:        
        s.close()


"""
开票通知.出货日期
对应原Pascal: 出货日期
"""
@any_route('/api/saier/billed_notice/shipping_date/change', methods=['POST'])
@require_token
async def view_billed_notice_shipping_date_change(request):
    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        code = 1
        msg = '操作成功'
        data = 30
        d = s.query(zx.cs).filter(zx.ly == '开票逾期').first()
        if d:
            data = d.cs

        return json_result(code, msg, data)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'出货日期修改检查失败: {str(e)}')
    finally:
        s.close()



"""
开票通知.跟单人员
对应原Pascal: 跟单人员
"""
@any_route('/api/saier/billed_notice/gdry/change', methods=['POST'])
@require_token
async def view_billed_notice_gdry_change(request):
    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        x = j.get('main', {})
        rid = x.get('rid', '')
        gdry = x.get('gdry', '')
        code = 1
        msg = '操作成功'

        d = s.query(kaiptz).filter(kaiptz.rid == rid).first()
        if d and d.uid != gdry and gdry != None and gdry != '':
            uid = ''
            u = s.query(sys_user.rid).filter(sys_user.username == gdry).first()
            if not u:
                code = -1
                msg = '跟单人员对应的用户名不存在，请检查'
                return json_result(code, msg)
            uid = u.rid
            d.uid = uid
            s.add(d)
            s.query(kaiptzsheet).filter(kaiptzsheet.pid == rid, kaiptzsheet.ywrya != gdry).update({kaiptzsheet.ywrya: gdry}, synchronize_session=False)
            s.query(kaiptzsheet1).filter(kaiptzsheet1.pid == rid, kaiptzsheet1.ywrya != gdry).update({kaiptzsheet1.ywrya: gdry}, synchronize_session=False)
            s.query(kaiptzxq).filter(kaiptzxq.pid == rid, kaiptzxq.jsrm != gdry).update({kaiptzxq.jsrm: gdry}, synchronize_session=False)

        s.commit()
        return json_result(code, msg)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'出货日期修改检查失败: {str(e)}')
    finally:
        s.close()


"""
开票通知.资料合并.完成开票
对应原Pascal: 资料合并.完成开票
"""
@any_route('/api/saier/billed_notice/wckp/change', methods=['POST'])
@require_token
async def view_billed_notice_wckp_change(request):
    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        wxfp = j.get('wxfp', '')
        ywfp = j.get('ywfp', '')
        wckp = j.get('wckp', '')
        kpxh = j.get('kpxh', '')
        rid = j.get('rid', '')
        code = 1
        msg = '操作成功'
        data = {'position':'宁波'}
        if wckp == '是':
            org = get_user_path(user.username)
            data['position'] = org.get('position', '')

            if wxfp == '' or wxfp is None:
                return json_result(0, ' 跳过更新')
            d = s.query(kaiptzsheet1.rid).filter(kaiptzsheet1.ysfp == wxfp, or_(func.ifnull(kaiptzsheet1.wckp, '') == '', func.ifnull(kaiptzsheet1.wckp, '') != '是')).first()
            if not d:
                s.query(gchk).filter(gchk.ysfp == wxfp, func.ifnull(gchk.fpwk, '') == '否').update({gchk.jdsj: time.strftime("%Y-%m-%d")}, synchronize_session=False)

            res = user_task_delete('开票通知', rid, s, [], '没到票')
            if res.get('code') != 1:
                s.rollback()
        else:
            s.query(gchk).filter(gchk.ysfp == wxfp, func.ifnull(gchk.fpwk, '') == '否').update({gchk.jdsj: None}, synchronize_session=False)

        s.query(gchk).filter(gchk.kpxh == kpxh).update({gchk.wckp: wckp}, synchronize_session=False)

        s.commit()
        return json_result(code, msg, data)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'出货日期修改检查失败: {str(e)}')
    finally:
        s.close()


"""
开票通知.资料合并.抹零识别
对应原Pascal: .资料合并.抹零识别
"""
@any_route('/api/saier/billed_notice/mlsb/change', methods=['POST'])
@require_token
async def view_billed_notice_mlsb_change(request):
    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        zwpm = j.get('zwpm', '')
        mlsb = j.get('mlsb', '')
        lines = j.get('lines', [])
        code = 1
        msg = '操作成功'
        if zwpm == '' or zwpm is None:
            return json_result(0, ' 跳过更新')
        
        for r in lines:
            if r.get('zwpm', '') != zwpm:
                continue
            d = s.query(cymxsheet).filter(cymxsheet.cght == r.get('cght', ''), cymxsheet.cywyzd == r.get('cywyzd', '')).all()
            for l in d:
                l.mlsb = mlsb
                s.add(l)

        s.commit()
        return json_result(code, msg)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'出货日期修改检查失败: {str(e)}')
    finally:
        s.close()


"""
开票通知.资料合并.商品分类
对应原Pascal: .资料合并.商品分类
"""
@any_route('/api/saier/billed_notice/spfl/change', methods=['POST'])
@require_token
async def view_billed_notice_spfl_change(request):
    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        zwpm = j.get('zwpm', '')
        spflmc = j.get('spflmc', '')
        code = 1
        msg = '操作成功'
        if zwpm == '' or zwpm is None or spflmc == '' or spflmc is None:
            return json_result(0, ' 跳过更新')
        
        d = run_sql(f"select spflbm,spflmc,rid from spbmflb where spflmc='{spflmc}' limit 1")
        if len(d) == 0:
            return json_result(0, '跳过更新')
        
        spflbm = d[0].get('spflbm', '')
        pid = d[0].get('rid', '')
        c = s.query(spbmflbsheet).filter(spbmflbsheet.zwpm == zwpm).first()
        if c:
            c.spflmc = spflmc
            c.pid = pid
            c.spflbm = spflbm
            c.modi_uid = user.rid
            c.mtime = time.strftime("%Y-%m-%d %H:%M:%S")
            s.add(c)
        else:
            c = spbmflbsheet()
            c.rid = get_uuid()
            c.uid = user.rid
            c.ctime = time.strftime("%Y-%m-%d %H:%M:%S")
            c.zwpm = zwpm
            c.spflmc = spflmc
            c.pid = pid
            c.spflbm = spflbm
            s.add(c)

        s.commit()
        return json_result(code, msg)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'出货日期修改检查失败: {str(e)}')
    finally:
        s.close()


"""
开票通知.开票详情.工厂发票
对应原Pascal: 开票详情.工厂发票
"""
@any_route('/api/saier/billed_notice/gcfp/change', methods=['POST'])
@require_token
async def view_billed_notice_gcfp_change(request):
    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        gcfp = j.get('gcfp', '')
        wstt = j.get('wstt', '')
        data = {}
        if gcfp == '' or gcfp is None or wstt == '' or wstt is None:
            return json_result(0, ' 跳过更新', data)

        code = 1
        msg = '操作成功'
        d = s.query(zlbinvoice).filter(zlbinvoice.NumberOfInvoice == gcfp, zlbinvoice.NameOfPurchaser == wstt).first()
        if d:
            data = {
                'NumberOfInvoice': d.NumberOfInvoice,
                'CodeOfInvoice': d.CodeOfInvoice,
                'TaxIncludedAmountInFigures': d.TaxIncludedAmountInFigures,
                'DateOfIssue': d.DateOfIssue,
                'TotalAmountExcludingTax': d.TotalAmountExcludingTax,
                'TotalTaxAmount': d.TotalTaxAmount,
                'kpgcsh': d.kpgcsh,
                'UniqueCodeOfInvoice': d.UniqueCodeOfInvoice
            }

        return json_result(code, msg, data)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'出货日期修改检查失败: {str(e)}')
    finally:
        s.close()


class InvoiceExporter:
    def __init__(self, session):
        # self.session = session
        self.rids = []  # 存储 number 列表
        self.z2 = []        # 存储发票号码列表
        self.lstallnumbers = []  # 存储所有单据号码

    def get_number_list(self) -> List[str]:
        """从 datacenter 获取号码列表"""
        # 根据您的实际业务逻辑实现
        return []

    def export_to_excel(self, rids: List[str], tempfile:str, savefile: str, s) -> dict:
        """主导出函数"""
        if not tempfile or not savefile:
            return {'code': -1, 'msg': '模板文件路径和保存文件路径不能为空'}
        self.rids = rids

        logger.error(f"导出函数被调用，rids: {rids}, tempfile: {tempfile}, savefile: {savefile}")
        # 加载或创建 Excel 工作簿
        if not os.path.exists(tempfile):
            logger.error(f"模板文件不存在: {tempfile}")
            return {'code': -1, 'msg': f'模板文件{tempfile}不存在'}
        
        wb = load_workbook(filename=tempfile)
        ws = wb.active
        

        # 收集所有 fphm 到 z2
        for rid in self.rids:
            result = s.execute(
                text("SELECT * FROM kaiptz WHERE rid = :rid"),
                {"rid": rid}
            ).fetchall()

            for row in result:
                fphm = row.fphm if hasattr(row, 'fphm') else row[1]  # 根据实际列索引调整
                if fphm not in self.z2:
                    self.z2.append(fphm)
        
        e = 1  # 起始行
        for l1, fphm in enumerate(self.z2):
            z1 = []          # 商品名称列表
            z1sl = []        # 商品名称+税率+hyd 组合
            zsl = []         # 税率列表
            zzsl = []        # 增值税率列表
            hyd = []         # hyd 列表
            zwpmbb = []      # 补报商品列表

            # 查询发票信息
            for rid in self.rids:
                result = s.execute(
                    text("SELECT * FROM kaiptz WHERE rid = :rid AND fphm = :fphm"),
                    {"rid": rid, "fphm": fphm}
                ).fetchall()
                if result:
                    row = result[0]
                    ws.row_dimensions[e].height = 40
                    ws.row_dimensions[e + 1].height = 22
                    ws.row_dimensions[e + 2].height = 22

                    kprq = row.kprq if hasattr(row, 'kprq') else row[0]
                    ws[f'I{e + 1}'].value = str(kprq)[:10] if kprq else ''
                    ws.merge_cells(f'I{e + 1}:K{e + 1}')
                    wstt = row.wstt if hasattr(row, 'wstt') else row[0]
                    if wstt and str(wstt).strip():
                        ws[f'C{e + 1}'].value = wstt
                        ws.merge_cells(f'C{e + 1}:F{e + 1}')

                    yfph = row.yfph if hasattr(row, 'yfph') else row[0]
                    if not yfph or str(yfph).strip() == '':
                        chrq = row.chrq if hasattr(row, 'chrq') else row[0]
                        ws[f'C{e + 2}'].value = str(chrq)[:10] if chrq else ''
                        ws.merge_cells(f'C{e + 2}:E{e + 2}')

                    ws[f'F{e + 2}'].value = fphm
                    ws.merge_cells(f'F{e + 2}:K{e + 2}')

            # 查询 bgmxd 表
            bgmxd_result = s.execute(
                text("SELECT * FROM bgmxd WHERE fphm = :fphm"),
                {"fphm": fphm}
            ).fetchall()

            pid = ""
            rmb = ""
            hl = 1.0
            sl = 0.0
            sl1 = 0.0

            if bgmxd_result:
                row = bgmxd_result[0]
                pid = row.rid if hasattr(row, 'rid') else row[0]
                rmb_flag = row.RMBkh if hasattr(row, 'RMBkh') else row[0]
                if rmb_flag == '是':
                    rmb = '是'
                else:
                    # 查询汇率
                    hl_result = s.execute(
                        text("SELECT tjhl FROM kpnr WHERE tjhl > 0 ORDER BY ctime DESC LIMIT 1")
                    ).fetchone()
                    if hl_result:
                        hl = float(hl_result.tjhl if hasattr(hl_result, 'tjhl') else hl_result[0])

                # 查询 bgmxdhbcp 表计算金额
                dhbcp_result = s.execute(
                    text("SELECT * FROM bgmxdhbcp WHERE pid = :pid AND zwpm <> '' ORDER BY sid ASC"),
                    {"pid": pid}
                ).fetchall()

                for item in dhbcp_result:
                    wxzj = float(item.wxzj if hasattr(item, 'wxzj') else item[0])
                    sl1 += wxzj
                    if rmb == '是':
                        sl += wxzj
                    else:
                        sl += wxzj * hl

            # 查询商品信息
            product_result = s.execute(
                text("SELECT * FROM bgmxdhbcp WHERE pid = :pid AND zwpm1 <> ''"),
                {"pid": pid}
            ).fetchall()

            for item in product_result:
                zwpm1 = item.zwpm1 if hasattr(item, 'zwpm1') else item[0]
                tsl = str(item.tsl if hasattr(item, 'tsl') else item[0])
                hyd_val = item.hyd if hasattr(item, 'hyd') else item[0]
                zzsl_val = str(item.zzsl if hasattr(item, 'zzsl') else item[0])

                key = zwpm1 + tsl + hyd_val
                if key not in z1sl:
                    z1.append(zwpm1)
                    z1sl.append(key)
                    zsl.append(tsl)
                    zzsl.append(zzsl_val)
                    hyd.append(hyd_val)

            # 查询 kaiptzsheet1 表处理补报商品
            sheet1_result = s.execute(
                text("SELECT * FROM kaiptzsheet1 WHERE fphm = :fphm"),
                {"fphm": fphm}
            ).fetchall()

            for item in sheet1_result:
                zwpm = item.zwpm if hasattr(item, 'zwpm') else item[0]
                zsl_val = str(item.zsl if hasattr(item, 'zsl') else item[0])
                hyd_val = item.hyd if hasattr(item, 'hyd') else item[0]
                key = zwpm + zsl_val + hyd_val

                if key not in z1sl:
                    if key not in zwpmbb:
                        zwpmbb.append(key)

            # 处理明细行
            i = e - 1
            hj = 0.0
            cb = 0.0
            tse = 0.0

            for l_idx, (key) in enumerate(z1sl):
                tsl_val = zsl[l_idx]
                if int(float(tsl_val)) > 4:
                    detail_result = s.execute(
                        text("""SELECT * FROM kaiptzsheet1 
                               WHERE fphm = :fphm AND zwpm1 = :zwpm1 
                               AND zsl = :zsl AND hyd = :hyd"""),
                        {"fphm": fphm, "zwpm1": z1[l_idx], "zsl": int(tsl_val), "hyd": hyd[l_idx]}
                    ).fetchall()
                else:
                    detail_result = s.execute(
                        text("""SELECT * FROM kaiptzsheet1 
                               WHERE fphm = :fphm AND zwpm1 = :zwpm1 
                               AND (zsl = 3 OR zsl = 0) AND hyd = :hyd"""),
                        {"fphm": fphm, "zwpm1": z1[l_idx], "hyd": hyd[l_idx]}
                    ).fetchall()

                if detail_result:
                    for item in detail_result:
                        gcfp = ""
                        zzslz = 0
                        kpxh = item.kpxh if hasattr(item, 'kpxh') else item[0]

                        if kpxh and str(kpxh).strip():
                            xq_result = s.execute(
                                text("SELECT gcfp, zzsl FROM kaiptzxq WHERE kpxh = :kpxh"),
                                {"kpxh": kpxh}
                            ).fetchall()

                            for xq in xq_result:
                                zzslz = 1 + float(xq.zzsl if hasattr(xq, 'zzsl') else xq[0]) / 100
                                gcfp_val = xq.gcfp if hasattr(xq, 'gcfp') else xq[0]
                                if not gcfp:
                                    gcfp = gcfp_val
                                else:
                                    gcfp += f';{gcfp_val}'

                        i += 1

                        if zzslz == 0:
                            zzslz = 1 + float(item.zzsl if hasattr(item, 'zzsl') else item[0]) / 100

                        sjkpze = float(item.sjkpze if hasattr(item, 'sjkpze') else item[0])
                        hj += sjkpze
                        zsl_val = float(item.zsl if hasattr(item, 'zsl') else item[0])

                        # 成本计算逻辑
                        if zsl_val <= 4:
                            if zsl_val == 0:
                                cb += sjkpze
                            elif zsl_val == 3:
                                cb += sjkpze - (sjkpze / 1.03 * 0.03)
                                tse += (sjkpze / 1.03 * 0.03)
                            elif zsl_val == 1:
                                cb += sjkpze - (sjkpze / 1.01 * 0.01)
                                tse += (sjkpze / 1.01 * 0.01)
                        else:
                            if zsl_val == 17:
                                cb += sjkpze - (sjkpze / 1.17 * (zsl_val / 100))
                                tse += (sjkpze / 1.17 * (zsl_val / 100))
                            else:
                                cb += sjkpze - (sjkpze / zzslz * (zsl_val / 100))
                                tse += (sjkpze / zzslz * (zsl_val / 100))

                        # 写入 Excel
                        row_num = 4 + i
                        ws.row_dimensions[row_num].height = 22

                        sccj = item.sccj if hasattr(item, 'sccj') else item[0]
                        zwpm_val = item.zwpm if hasattr(item, 'zwpm') else item[0]
                        bgsl = float(item.bgsl if hasattr(item, 'bgsl') else item[0])
                        hgjldw = item.hgjldw if hasattr(item, 'hgjldw') else item[0]
                        chs = float(item.chs if hasattr(item, 'chs') else item[0])
                        ywrya = item.ywrya if hasattr(item, 'ywrya') else item[0]
                        lxdh = item.lxdh if hasattr(item, 'lxdh') else item[0]
                        ykfp = item.ykfp if hasattr(item, 'ykfp') else item[0]

                        # 设置字体颜色（红色用于 yfph 不为空的情况）
                        yfph_val = item.yfph if hasattr(item, 'yfph') else item[0]
                        if yfph_val and str(yfph_val).strip() != '':
                            red_font = Font(color="FF0000")
                            ws[f'B{row_num}'].font = red_font
                            ws[f'C{row_num}'].font = red_font
                            ws[f'D{row_num}'].font = red_font
                            ws[f'E{row_num}'].font = red_font
                            ws[f'F{row_num}'].font = red_font
                            ws[f'G{row_num}'].font = red_font
                            ws[f'H{row_num}'].font = red_font
                            ws[f'I{row_num}'].font = red_font
                            ws[f'J{row_num}'].font = red_font
                            ws[f'K{row_num}'].font = red_font
                            ws[f'L{row_num}'].font = red_font
                            ws[f'M{row_num}'].font = red_font
                            ws[f'N{row_num}'].font = red_font

                        ws[f'B{row_num}'].value = sccj
                        ws[f'C{row_num}'].value = f"{zsl_val}%"
                        ws[f'D{row_num}'].value = zwpm_val
                        ws[f'E{row_num}'].value = bgsl
                        ws[f'F{row_num}'].value = hgjldw
                        ws[f'G{row_num}'].value = sjkpze
                        ws[f'H{row_num}'].value = chs
                        ws[f'I{row_num}'].value = gcfp
                        ws[f'J{row_num}'].value = hyd[l_idx]

                        # uv 变量需要从外部传入
                        uv = getattr(self, 'uv', '0')
                        if uv == '1':
                            cz = item.cz if hasattr(item, 'cz') else ''
                            ws[f'K{row_num}'].value = f"{ywrya}/{cz}"
                        else:
                            ws[f'K{row_num}'].value = ywrya

                        ws[f'L{row_num}'].value = lxdh
                        if yfph_val and str(yfph_val).strip() != '':
                            ws[f'M{row_num}'].value = ykfp
                        ws[f'N{row_num}'].value = kpxh

                        ws.row_dimensions[row_num].auto_size = True

                else:
                    # 补报行
                    i += 1
                    row_num = 4 + i
                    ws.row_dimensions[row_num].height = 22
                    red_font = Font(color="FF0000")
                    ws[f'B{row_num}'].font = red_font
                    ws[f'B{row_num}'].value = '补报'
                    ws[f'C{row_num}'].font = red_font
                    ws[f'C{row_num}'].value = z1[l_idx]

            # 添加合计行
            ws.row_dimensions[5 + i].height = 22
            ws[f'F{5 + i}'].value = '合计'
            ws[f'G{5 + i}'].value = hj

            # 合并单元格
            ws.merge_cells(f'A{6 + i}:L{6 + i}')

            # 计算毛利率
            ml = 0.0
            if sl > 0:
                ml = (sl - cb) / sl * 100

            if sl1 == 0:
                sl1 = 0.0000001

            ws.row_dimensions[6 + i].height = 30

            if rmb == '是':
                ws[f'A{6 + i}'].value = f"收入:{sl:.2f}成本:{cb:.2f}利润:{sl - cb:.2f}毛利率:{ml:.1f}%"
            else:
                ws[f'A{6 + i}'].value = f"报关金额:{sl1:.2f}换汇成本:{cb / sl1:.2f}收入:{sl:.2f}成本:{cb:.2f}利润:{sl - cb:.2f}毛利率:{ml:.1f}%"

            # 重置变量
            sl = 0
            cb = 0
            ml = 0
            sl1 = 0
            e = 7 + i

        # 删除模板行
        # ws.delete_rows(1, 29)

        # 添加制单人员
        # user_name = IApplication.LoginInfo.userinfo.name
        # ws['B1'].value = f"制单人员:{user_name}"
        ws['B1'].value = "制单人员:系统用户"

        # 保存文件
        savename = f"开票汇总_{int(time.time())}.xlsx"
        wb.save(os.path.join(config.tmp_path, savename))
        return {'code': 1, 'msg': '操作成功', 'data': savename}


"""
开票通知.批量开票汇总
对应原Pascal: 批量开票汇总
"""
@any_route('/api/saier/billed_notice/batch/invoice', methods=['POST'])
@require_token
async def view_billed_notice_batch_invoice(request):
    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        rids = j.get('rids', '')
        data = {}
        org = get_user_path('zjnblh')
        position = org.get('position', '')
        uv = '0'  # 设置 uv 标志
        if position != '' and position != None and position[:1] == 'E':
            uv = '1'
        data['position'] = org.get('position', '')
        code = 1
        msg = '操作成功'
        exporter = InvoiceExporter(s)
        exporter.uv = uv  # 设置 uv 标志
        tempfile = os.path.join(config.data_upload_path, 'template/开票汇总.xlsx')  # 模板文件路径
        savefile = os.path.join(config.tmp_path, f'开票汇总_{int(time.time())}.xlsx')  # 保存文件路径
        res = exporter.export_to_excel(tempfile=tempfile, savefile=savefile, rids=rids, s=s)
        if res.get('code') != 1:
            return json_result(res.get('code', -1), res.get('msg', '导出失败'))
        savename = res.get('data', '')

        return json_result(code, msg, savename)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'批量开票汇总失败: {str(e)}')
    finally:
        s.close()


def export_wxfp_summary_orm(wxfp: str, s):
    """使用 ORM 方式导出"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "外销发票汇总"
    
    # 设置表头
    headers = ['商品名称', '数量', '单位', '总价']
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header).font = Font(bold=True)
    
    # 查询 bgmxd
    c = s.query(bgmxd).filter(bgmxd.ysfp == wxfp).first()
    if not c:
        logger.error(f"未找到 ysfp='{wxfp}' 的记录")
        # wb.save(os.path.join(output_dir, f'{wxfp}汇总表{int(time.time())}.xlsx'))
        return {'code': -1, 'msg': f"未找到 ysfp='{wxfp}' 的记录", 'data': None}
    
    father4 = c.rid  # 订单号
    wxfp1 = c.fphm      # 发票号
    
    # 查询商品列表（去重）
    products = s.query(
        bgmxdhbcp.zwpm, bgmxdhbcp.hgjldw
    ).filter(
        bgmxdhbcp.pid == father4,
        bgmxdhbcp.zwpm != ''
    ).distinct().all()
    
    current_row = 2
    red_font = Font(color="FF0000")
    
    for zwpm, hgjldw in products:
        # 汇总查询
        summary = s.query(
            func.sum(kaiptzsheet1.bgsl).label('bgsl'),
            func.sum(kaiptzsheet1.gczj).label('gczj')
        ).filter(
            kaiptzsheet1.fphm == wxfp1,
            kaiptzsheet1.zwpm == zwpm
        ).first()
        
        bgsl = summary.bgsl or 0
        gczj = summary.gczj or 0
        
        if bgsl > 0 or gczj > 0:
            ws.cell(row=current_row, column=1, value=zwpm)
            ws.cell(row=current_row, column=2, value=float(bgsl))
            ws.cell(row=current_row, column=3, value=hgjldw or '')
            ws.cell(row=current_row, column=4, value=round(float(gczj or 0), 2))
        else:
            cell = ws.cell(row=current_row, column=1, value=f"{zwpm}(待补报)")
            cell.font = red_font
            ws.cell(row=current_row, column=3, value=hgjldw or '')
        
        current_row += 1
    
    # 添加合计行
    if current_row > 2:
        ws.cell(row=current_row, column=1, value="合计").font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=f"=SUM(B2:B{current_row-1})")
        ws.cell(row=current_row, column=4, value=f"=SUM(D2:D{current_row-1})")
    
    # 保存
    output_dir = config.tmp_path
    # os.makedirs(output_dir, exist_ok=True)
    output_name = f'{wxfp}汇总表{int(time.time())}.xlsx'
    output_file = os.path.join(output_dir, output_name)
    wb.save(output_file)
    logger.error(f"已保存: {output_file}")
    
    return {'code': 1, 'msg': '操作成功', 'data': output_name}


"""
开票通知.预录单对照表
对应原Pascal: 预录单对照表
"""
@any_route('/api/saier/billed_notice/compare_list', methods=['POST'])
@require_token
async def view_billed_notice_compare_list(request):
    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        wxfp = j.get('wxfp', '')
        code = 1
        msg = '操作成功'

        res = export_wxfp_summary_orm(wxfp, s=s)
        if res.get('code') != 1:
            return json_result(res.get('code', -1), res.get('msg', '导出失败'))
        savename = res.get('data', '')

        return json_result(code, msg, savename)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'预录单对照表失败: {str(e)}')
    finally:
        s.close()


"""
批量货源地
对应原Pascal: 批量货源地
"""
@any_route('/api/saier/billed_notice/source_place', methods=['POST'])
@require_token
async def view_billed_notice_source_place(request):
    # 1. 检查用户是否是财务岗位 (对应 tmpcom2 查询)
    s = Session()
    try:
        j = await request.form()
        module = form_value(j, 'module', '开票通知')
        user = request.current_user
        org = get_user_path(user.username)
        position = org.get('position')
        if '财务' not in position:
            return json_result(-1, '只有财务岗位用户才能执行此操作')
        # 2. 获取上传的Excel文件
        file = form_value(j,'file',None)
        if is_none(file):
            return json_result(ERR_PARAM_NOT_ENOUGH)
        # 保存临时文件
        temp_dir = config.tmp_path
        temp_file = os.path.join(temp_dir, f"{uuid.uuid4()}_{file.filename}")
        # logger.error(f"Saving uploaded file to: {temp_file}")
        data = await file.read()
        write_file(temp_file, data, 'wb', encoding=None)
        
        # 3. 读取Excel文件
        wb = load_workbook(temp_file)
        ws = wb.worksheets[0]  # 读取第一个工作表
        # 4. 处理每一行数据
        row_idx = 1  # 从第1行开始（第1行是表头）
        while True:
            # 读取Excel单元格数据
            kpxh = str(ws.cell(row=row_idx, column=1).value or '').strip()  # A列 - 开票序号
            hyd = str(ws.cell(row=row_idx, column=2).value or '')  # B列 - 货源地
            # 如果rid为空，结束循环
            if not fkxh:
                break
            d = s.query(kaiptzsheet1).filter(kaiptzsheet1.fkxh == fkxh, kaiptzsheet1.kpxh == kpxh).update({kaiptzsheet1.hyd: hyd}, synchronize_session=False)
            row_idx += 1
            
        s.commit()
        return json_result(0, "批量更新成功")
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()
        wb.close()


"""
开票通知.开票统计表
对应原Pascal: 开票统计表
"""
@any_route('/api/saier/billed_notice/invoice_summary', methods=['POST'])
@require_token
async def view_billed_notice_invoice_summary(request):
    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        wxfp = j.get('wxfp', '')
        hl = j.get('hl', 0)
        code = 1
        msg = '操作成功'
        if hl is None or hl == 0:
            d = s.query(hbdm.hhl).filter(hbdm.hbdm == 'USD$').first()
            if d and d.hhl :
                hl = float(d.hhl)

        res = export_invoice_summary(wxfp, hl, s=s)
        if res.get('code') != 1:
            return json_result(res.get('code', -1), res.get('msg', '导出失败'))
        savename = res.get('data', '')

        return json_result(code, msg, savename)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'预录单对照表失败: {str(e)}')
    finally:
        s.close()



def export_invoice_summary(wxfp: str, hl: float, s):
    """
    导出开票统计表
    
    Args:
        wxfp: 外销发票号/运单号 (如 'aa')
        hl: 汇率
        filename3: 输出目录
        session: SQLAlchemy Session
    """
    
    # 1. 清理文件名中的非法字符
    wxfp1 = clean_filename(wxfp)
    logger.error(f"原始 wxfp: {wxfp}, 清理后: {wxfp1}")
    
    # 2. 初始化变量
    je = 0.0
    zje = 0.0
    se = 0.0
    zse = 0.0
    tse = 0.0
    ztse = 0.0
    fpze = 0.0
    bg = 0.0
    
    # 3. 准备模板文件
    template_dir = config.data_upload_path
    template_file = os.path.join(template_dir, 'template/优胜开票统计.xlsx')
    
    # 确保目录存在
    if not os.path.exists(template_dir):
        os.makedirs(template_dir)
        # 这里需要复制模板文件，根据实际情况调整源路径
        # source_template = r'源模板路径\优胜开票统计.xls'
        # shutil.copy2(source_template, template_file)
    
    # 4. 打开 Excel 模板
    if not os.path.exists(template_file):
        logger.error(f"模板文件不存在: {template_file}")
        return {'code': -1, 'msg': f'模板文件{template_file}不存在', 'data': None}
    
    wb = load_workbook(template_file)
    ws = wb.active
    ws.title = "开票统计"
    
    # 5. 获取汇率（如果未提供）
    if hl is None or hl == 0:
        result = s.execute(
            text("SELECT hhl FROM hbdm WHERE hbdm = 'USD$'")
        ).fetchone()
        if result:
            hl = float(result[0]) if result[0] else 1.0
    
    # 6. 查询 bgmxd 表
    bgmxd_result = s.execute(
        text("SELECT bgbgzje, sbdate, fphm FROM bgmxd WHERE ysfp = :ysfp"),
        {"ysfp": wxfp}
    ).fetchall()
    
    wxfp2 = ""
    sbdate = ""
    
    if len(bgmxd_result) == 0:
        logger.error(f"未找到 ysfp='{wxfp}' 的记录")
        return {'code': -1, 'msg': f"未找到 ysfp='{wxfp}' 的报关明细记录", 'data': None}
    
    row = bgmxd_result[0]
    wxfp2 = row.fphm
    sbdate = row.sbdate
    bg = float(row.bgbgzje or 0)
        
    # 写入 A1、C1、E1
    ws['A1'] = bg
    ws['C1'] = hl
    ws['E1'] = bg * hl
    
    # 7. 查询 kaiptzsheet1 获取数据行数
    count_result = s.execute(
        text("SELECT COUNT(*) FROM kaiptzsheet1 WHERE fphm = :fphm AND bgsl > 0"),
        {"fphm": wxfp2}
    ).fetchone()
    
    i10 = count_result[0] if count_result else 0
    
    # 8. 如果有多行数据，插入复制的行
    if i10 > 1:
        # 获取模板第3行作为复制源
        template_row = 3
        for k in range(1, i10):
            # 在第 template_row + k 位置插入新行
            ws.insert_rows(template_row + k)
            # 复制第3行的样式和值到新行
            copy_row(ws, template_row, template_row + k)
    
    # 9. 查询并填充数据
    detail_result = s.execute(
        text("SELECT * FROM kaiptzsheet1 WHERE fphm = :fphm AND bgsl > 0"),
        {"fphm": wxfp2}
    ).fetchall()
    logger.error(f"查询到 {len(detail_result)} 条明细数据")
    i = 0
    if detail_result:
        for idx, row in enumerate(detail_result):
            i = idx + 1
            current_row = i + 2  # 从第3行开始（第1行是标题，第2行可能是表头）
            # 获取字段值
            yfph = row.yfph
            ysfp = row.ysfp
            ysfp1 = row.ysfp1
            ywrya = row.ywrya
            zsl = float(row.zsl or 0)
            sccj = row.sccj
            zwpm = row.zwpm
            bgsl = float(row.bgsl or 0)
            gczj = float(row.gczj or 0)
            xgzf = float(row.xgzf or 0)
            zzsl = float(row.zzsl or 0)
            # 自动调整行高
            ws.row_dimensions[current_row].height = 15
            
            # A列: ysfp
            if not yfph or str(yfph).strip() == '':
                ws[f'A{current_row}'] = ysfp
            else:
                ws[f'A{current_row}'] = f"{ysfp}({ysfp1})"
            
            # B列: 业务人员
            ws[f'B{current_row}'] = ywrya
            
            # C列: 申报日期减3天
            if sbdate:
                try:
                    if isinstance(sbdate, str):
                        sb_date_obj = datetime.strptime(sbdate, '%Y-%m-%d')
                    else:
                        sb_date_obj = sbdate
                    c_date = sb_date_obj - timedelta(days=3)
                    ws[f'C{current_row}'] = c_date.strftime('%Y-%m-%d')
                except:
                    ws[f'C{current_row}'] = sbdate
            
            # D列: 生产厂家 + 税率（小规模纳税人和普票）
            if zsl <= 4:
                ws[f'D{current_row}'] = f"{sccj}{int(zsl)}%"
            else:
                ws[f'D{current_row}'] = sccj
            
            # E列: 商品名称
            ws[f'E{current_row}'] = zwpm
            
            # F列: 数量
            ws[f'F{current_row}'] = bgsl
            
            # G列: 价税合计 - 相关费用
            gjze = gczj - xgzf
            ws[f'G{current_row}'] = round(gjze, 2)
            fpze += gjze
            
            # H列: 不含税金额
            if zsl <= 4:
                if zsl == 3:
                    je_value = gjze / 1.03
                    ws[f'H{current_row}'] = round(je_value, 2)
                    zje += je_value
                elif zsl == 1:
                    je_value = gjze / 1.01
                    ws[f'H{current_row}'] = round(je_value, 2)
                    zje += je_value
            else:
                if zsl == 17:
                    je_value = gjze / 1.17
                    ws[f'H{current_row}'] = round(je_value, 2)
                    zje += je_value
                else:
                    je_value = gjze / (1 + zzsl / 100)
                    ws[f'H{current_row}'] = round(je_value, 2)
                    zje += je_value
            
            # I列: 税额
            if zsl <= 4:
                if zsl == 3:
                    se_value = gjze - (gjze / 1.03)
                    ws[f'I{current_row}'] = round(se_value, 2)
                    zse += se_value
                elif zsl == 1:
                    se_value = gjze - (gjze / 1.01)
                    ws[f'I{current_row}'] = round(se_value, 2)
                    zse += se_value
            else:
                if zsl == 17:
                    se_value = gjze - (gjze / 1.17)
                    ws[f'I{current_row}'] = round(se_value, 2)
                    zse += se_value
                else:
                    se_value = gjze - (gjze / (1 + zzsl / 100))
                    ws[f'I{current_row}'] = round(se_value, 2)
                    zse += se_value
            
            # J列: 税率/100
            ws[f'J{current_row}'] = zsl / 100
            
            # K列: 退税额
            if zsl <= 4:
                if zsl == 3:
                    tse_value = (gjze / 1.03) * (zsl / 100)
                    ws[f'K{current_row}'] = round(tse_value, 2)
                    ztse += tse_value
                elif zsl == 1:
                    tse_value = (gjze / 1.01) * (zsl / 100)
                    ws[f'K{current_row}'] = round(tse_value, 2)
                    ztse += tse_value
            else:
                if zsl == 17:
                    tse_value = (gjze / 1.17) * (zsl / 100)
                    ws[f'K{current_row}'] = round(tse_value, 2)
                    ztse += tse_value
                else:
                    tse_value = (gjze / (1 + zzsl / 100)) * (zsl / 100)
                    ws[f'K{current_row}'] = round(tse_value, 2)
                    ztse += tse_value
    
    # 10. 添加合计行
    total_row = i + 3  # 因为 i 从0开始，加上偏移
    
    ws[f'G{total_row}'] = round(fpze, 2)
    ws[f'H{total_row}'] = round(zje, 2)
    ws[f'I{total_row}'] = round(zse, 2)
    ws[f'K{total_row}'] = round(ztse, 2)
    
    if bg > 0:
        ws[f'E{total_row}'] = round(fpze / bg, 2)
    
    # 11. 设置底部汇总行
    ws['G1'] = round(zje + zse - ztse, 2)
    ws['J1'] = round((bg * hl) - (zje + zse - ztse), 2)
    
    if (bg * hl) != 0:
        ws['H1'] = round(((bg * hl) - (zje + zse - ztse)) / (bg * hl), 4)
    
    # 12. 保存文件
    filename3 = config.tmp_path
    output_name = f'{wxfp1}开票统计{time.strftime("%Y-%m-%d_%H-%M-%S")}.xlsx'
    output_file = os.path.join(filename3, output_name)
    
    # 确保输出目录存在
    output_dir = os.path.dirname(output_file)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    wb.save(output_file)
    print(f"文件已保存: {output_file}")
    
    return {'code': 1, 'msg': '导出成功', 'data': output_name}


def clean_filename(filename: str) -> str:
    """清理文件名中的非法字符"""
    # Windows 文件名非法字符
    illegal_chars = r'[\/:*?"<>|\r\n]'
    
    # 替换非法字符为空
    cleaned = re.sub(illegal_chars, '', filename)
    
    # 如果有非法字符被移除，返回移除后的名称
    if cleaned != filename:
        return cleaned
    
    return filename


def copy_row(ws, source_row: int, target_row: int):
    """复制行的样式和值"""
    max_col = ws.max_column
    
    for col in range(1, max_col + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)
        
        # 复制值
        target_cell.value = source_cell.value
        
        # 复制样式
        if source_cell.has_style:
            target_cell.font = source_cell.font.copy() if source_cell.font else None
            target_cell.border = source_cell.border.copy() if source_cell.border else None
            target_cell.fill = source_cell.fill.copy() if source_cell.fill else None
            target_cell.number_format = source_cell.number_format
            target_cell.alignment = source_cell.alignment.copy() if source_cell.alignment else None



"""
开票通知.(优胜/优景)财务开票信息导出
对应原Pascal: (优胜/优景)财务开票信息导出
"""
@any_route('/api/saier/billed_notice/invoice_export', methods=['POST'])
@require_token
async def view_billed_notice_invoice_export(request):
    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        rids = j.get('rids', [])
        code = 1
        msg = '操作成功'
        org = get_user_path('zjnblh')
        position = org.get('position', '')
        uv = '0'  # 设置 uv 标志
        if position != '' and position != None and position[:1] == 'D':
            uv = '1'
        ywcs = 0
        nbcs = 0
        d = s.query(zx).filter(zx.ly == '合同签订日期').first()
        if d:
            ywcs = d.cs if d.cs else 0
            nbcs = d.cs if d.cs else 0

        res = export_ruiyi_invoices(rids, uv, ywcs, nbcs, s=s)
        if res.get('code') != 1:
            return json_result(res.get('code', -1), res.get('msg', '导出失败'))
        savename = res.get('data', '')

        return json_result(code, msg, savename)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'预录单对照表失败: {str(e)}')
    finally:
        s.close()


def round_to_two(value: float) -> float:
    """四舍五入保留两位小数"""
    return round(value * 100) / 100


def offset_img(img, col, row, x_pad=4, y_pad=25):
    """精确设置图片位置，偏移量以万为单位进行微调吧，具体计算公式太麻烦了
    row column 的索引都是从0开始的，我这里要把图片插入到单元格A17
    """
    p2e = pixels_to_EMU
    h, w = img.height, img.width
    # 图像等比例缩放因子
    resize_factor = 0.8
    w_h_ratio = w/h
    resize_H = int(resize_factor * h)
    resize_W = int(resize_factor * resize_H * w_h_ratio)
    #
    # x_pad = 4
    # y_pad = 25
    # 注意这里的行、列索引从0开始, 所以需要减1
    size = XDRPositiveSize2D(p2e(w), p2e(h))
    marker = AnchorMarker(col=col-1, colOff=pixels_to_EMU(x_pad), row=row-1, rowOff=pixels_to_EMU(y_pad))
    img.anchor = OneCellAnchor(_from=marker, ext=size)

    # 图像在Excel里面的大小
    # image_size_excel = XDRPositiveSize2D(pixels_to_EMU(resize_W), pixels_to_EMU(resize_H))
    ##############################
    # 设置单元格大小，单元格默认宽度单位：字符；高度单位：point(磅)
    # cell_height = int(pixels_to_points(resize_H+10, dpi=96)) #高度上增加10个像素放大单元格
    # cell_width = int(resize_W/8) + 2 # 宽度上增加16（2*8）个像素放大单元格

    # marker = AnchorMarker(col=col, colOff=pixels_to_EMU(x_pad), row=row, rowOff=pixels_to_EMU(y_pad))
    # img.anchor = OneCellAnchor(_from=marker, ext=image_size_excel)


def export_ruiyi_invoices(rids, uv, ywcs, nbcs, s):
    """
    导出锐亿171格式的发票
    
    Args:
        session: SQLAlchemy Session
        filename_template: 模板文件路径
        output_dir: 输出目录
    """
    

    filename_template = os.path.join(config.data_upload_path, 'template/锐亿171.xlsx')
    if not os.path.exists(filename_template):
        logger.error(f"模板文件不存在: {filename_template}")
        return {'code': -1, 'msg': f'模板文件{filename_template}不存在', 'data': None}
    
    files = []
    c = 0  # 计数器
    for rid in rids:
        
        # 查询 kaiptz 表
        d = s.query(kaiptz).filter(kaiptz.rid == rid).first()
        if not d:
            continue
        
        # 获取基本信息
        hyd = d.hyd or ''
        dkpd = d.cpmc or ''
        ywsb = ''
        jjdz = ''
        wstt1 = ''
        gs = ''
        dz = ''
        sw = ''
        bank = ''
        wstt2 = ''
        chyrq = ''
        zb = ''
        ywcs = 0  # 需要从外部获取
        
        # 判断是否为补报
        if d.cpmc == '补报':
            zb = '补'
            if d.chrq:
                try:
                    d1 = d.chrq
                    chyrq = d1.strftime('%Y-%m-%d')
                except:
                    chyrq = str(d.chrq) if d.chrq else ''
        else:
            chyrq = str(d.chrq) if d.chrq else ''
        
        # 查询 kpnr 表获取开票内容
        kpcxdh = ''
        kpnr_list = s.query(kpnr).all()
        for r in kpnr_list:
            gsjc = r.gsjc or ''
            if gsjc and gsjc in (d.wstt or ''):
                gs = r.wfgs or ''
                dz = r.kpdz or ''
                sw = r.kpsh or ''
                bank = r.kpyh or ''
                wstt1 = gsjc
                jjdz = r.kpjjdz or ''
                kpcxdh = r.kpcxdh or ''
        
        # 判断是否为义乌
        if d.wstt and '义乌' in d.wstt:
            ywsb = '1'
        
        # 设置寄件地址
        if not jjdz:
            if d.wstt and '义乌' in d.wstt:
                jjdz = '寄件地址：义乌市宗泽北路531号赛尔集团二楼   /邮编：322000/ 吴春燕 收/0579-85096055'
            else:
                jjdz = '寄件地址：宁波高新区光华路288号赛尔大厦25层  邮编315103/财务收/0574-27833931'
        
        # 处理各种字段
        cx1 = (d.ysfp or '').upper()
        cx7 = (d.yysfp or '').upper()
        lxfs = d.lxfs or ''
        
        if d.htrq:
            cx2 = d.htrq.strftime('%Y-%m-%d') if hasattr(d.htrq, 'strftime') else str(d.htrq)
        else:
            cx2 = chyrq
        
        cx3 = d.ywry or ''
        cx4 = d.lxdh or ''
        
        # 清理生产厂家名称中的非法字符
        cx = d.sccj or ''
        cx = clean_filename(cx)
        
        number = d.rid
        
        # 查询 kaiptzsheet1 表
        sheet1_list = s.query(kaiptzsheet1).filter(
            kaiptzsheet1.pid == number,
            kaiptzsheet1.bgsl > 0
        ).all()
        if sheet1_list:
            for sheet1 in sheet1_list:
                # 处理 yfph 字段生成 zb1
                zb1 = ''
                if sheet1.yfph:
                    if len(sheet1.yfph) > 1:
                        zb1_char = sheet1.yfph[1] if len(sheet1.yfph) > 1 else ''
                        if zb1_char not in '0123456789':
                            zb1 = zb + sheet1.yfph[:2] + '组'
                        else:
                            zb1 = zb + sheet1.yfph[:1] + '组'
                
                if sheet1.zsl >= 0:
                    c += 1
                    
                    # 加载 Excel 模板
                    wb = load_workbook(filename_template)
                    ws = wb.active
                    
                    # 查询业务员联系方式
                    y = s.query(ywrylx).filter(ywrylx.ryxm == sheet1.ywrya).first()
                    if y and y.Msn:
                        jjdz = y.Msn
                    
                    # 处理合同章图片
                    cpbh = f"{gs}合同章"
                    t = s.query(tpzx).filter(
                        tpzx.cpbh == cpbh,
                    ).first()
                    if t and t.tpmc:
                        if t.tpmc != None and t.tpmc.strip() != '' and t.tpmc != '[]':
                            photos = json.loads(t.tpmc)
                            src = photos[0].get('src') if len(photos) > 0 else None
                            img_path = os.path.join(config.data_upload_path, src)
                            if os.path.exists(img_path):
                                img = Image(img_path)
                                img = XLImage(img_path) # 选用 openpyxl 的 Image (兼容您给的 Image_Get)
                                            
                                # 仿照您的标准：留白 4 像素
                                img.width = 116  # 预估列宽转换像素后减4 (原代码大概120)
                                img.height = 96  # 预估行高(100)减4
                                col_idx = 1
                                row_idx = 17
                                offset_img(img, col_idx, row_idx, x_pad=2, y_pad=10)
                                ws.add_image(img)  # 位置可能需要调整
                    
                    # 写入数据到 Excel
                    ws['G45'] = f'(开票序号:{sheet1.kpxh})'
                    ws['A8'] = gs
                    ws['C14'] = gs
                    ws['C15'] = dz
                    ws['C16'] = sw
                    ws['C17'] = bank
                    ws.row_dimensions[11].height = 15
                    ws['G11'] = cx1
                    ws['G12'] = cx7
                    ws['G44'] = cx1
                    ws.row_dimensions[21].height = 15
                    ws['C21'] = sheet1.zwpm or ''
                    ws['G24'] = cx2
                    ws['B22'] = float(sheet1.bgsl or 0)
                    ws['E22'] = sheet1.hgjldw or ''
                    ws['E24'] = '0'
                    ws['G23'] = float(sheet1.sjkpze or 0)
                    
                    # 根据税率计算
                    zsl = float(sheet1.zsl or 0)
                    sjkpze = float(sheet1.sjkpze or 0)
                    
                    if zsl > 3:
                        je = round_to_two(sjkpze / 1.13)
                        ws['B23'] = je
                        ws['G22'] = '13%'
                        ws['E23'] = round_to_two(sjkpze - je)
                    elif zsl == 3:
                        je = round_to_two(sjkpze / 1.03)
                        ws['B23'] = je
                        ws['G22'] = '3%'
                        ws['E23'] = round_to_two(sjkpze - je)
                    elif zsl == 1:
                        je = round_to_two(sjkpze / 1.01)
                        ws['B23'] = je
                        ws['G22'] = '1%'
                        ws['E23'] = round_to_two(sjkpze - je)
                    
                    # 义乌特殊处理
                    if ywsb == '1':
                        ws['c25'] = sheet1.cz or ''
                        ws['A16'] = '统一社会信用代码'
                        ws['A19'] = f'{int(zsl)}%'
                        ws['A11'] = '合同签订日期:'
                        
                        if chyrq:
                            try:
                                chyrq_date = datetime.strptime(chyrq, '%Y-%m-%d') if isinstance(chyrq, str) else chyrq
                                b11_date = chyrq_date - timedelta(days=ywcs)
                                ws['B11'] = b11_date.strftime('%Y-%m-%d')
                            except:
                                ws['B11'] = chyrq
                    else:
                        ws['G21'] = f'{int(zsl)}%'
                    
                    # 其他设置
                    ws.row_dimensions[3].height = 15
                    ws['A20'] = f'发票快件查询 资料不清晰请拨打电话:{kpcxdh}'
                    
                    # 寄件地址（红色字体）
                    cell_a9 = ws['A9']
                    cell_a9.font = Font(color='FF0000')
                    cell_a9.value = f'★{jjdz}'
                    
                    ws['C18'] = sheet1.ywrya or ''
                    
                    # 联系电话
                    if sheet1.lxdh:
                        ws['G18'] = sheet1.lxdh
                    else:
                        y = s.query(ywrylx).filter(ywrylx.ryxm == sheet1.ywrya).first()
                        if y and y.zjdh:
                            ws['G18'] = y.zjdh
                    
                    ws['B24'] = sheet1.hyd or ''
                    
                    # 如果 uv == '1'，处理第二个工作表
                    if uv == '1':
                        if len(wb.worksheets) > 1:
                            ws2 = wb.worksheets[1]  # 第二个工作表
                            # ws2.activate()
                            ws2['G4'] = cx1
                            ws2.row_dimensions[7].height = 15
                            ws2['J7'] = sheet1.htxs or ''
                            ws2['c7'] = sheet1.htxs or ''
                            ws2['c8'] = sheet1.sccj or ''
                            ws2['c9'] = f'{int(zsl)}%'
                            ws2['c10'] = chyrq
                            ws2.row_dimensions[11].height = 15
                            ws2['J11'] = jjdz
                            ws2['c11'] = jjdz
                    
                    # 构建文件名并保存
                    filename_base = f"{zb1}{d.cpmc}{c}{wstt1}{sheet1.dq}{cx}{int(zsl)}"
                    filename_clean = clean_filename(filename_base) + f"_{time.strftime('%Y-%m-%d_%H-%M-%S')}_{uuid.uuid4()}.xlsx"
                    output_dir = config.tmp_path
                    output_file = os.path.join(output_dir, filename_clean)
                    
                    wb.save(output_file)
                    logger.error(f"已保存: {output_file}")
                    files.append(filename_clean)
                    # 更新 kaiptzsheet1 的 sfyk 字段
                    s.query(kaiptzsheet1).filter(
                        kaiptzsheet1.pid == number
                    ).update({"sfyk": "是"}, synchronize_session=False)
        
    s.commit()
    zip_filename = ''
    logger.error(f"总共生成 {len(files)} 个文件")
    logger.error(f"文件列表: {files}")
    if len(files) > 0:
        zip_filename = f"锐亿发票_{time.strftime('%Y-%m-%d_%H-%M-%S')}.zip"
        zip_filepath = os.path.join(config.tmp_path, zip_filename)
        zipFile = zipfile.ZipFile(zip_filepath, 'w')
        for f in files:
            fn = os.path.join(config.tmp_path, f)
            if os.path.exists(fn):
                zipFile.write(fn, f, zipfile.ZIP_DEFLATED)
        zipFile.close()

    return {'code': 1, 'msg': '导出成功', 'data': zip_filename}


"""
开票通知.开票信息导出
对应原Pascal: 开票信息导出
"""
@any_route('/api/saier/billed_notice/record_export', methods=['POST'])
@require_token
async def view_billed_notice_invoice_export(request):
    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        rids = j.get('rids', [])
        code = 1
        msg = '操作成功'
        org = get_user_path('zjnblh')
        position = org.get('position', '')
        uv = '0'  # 设置 uv 标志
        cwsb = '0'  # 设置 cwsb 标志
        if position != '' and position != None and position[:1] == 'D':
            uv = '1'
        if ('U' in user.username or 'JY' in user.username):
            if ('US' in user.username or 'UY' in user.username):
                cwsb = '1'
        else:
            cwsb = '1'

        ywcs = 0
        nbcs = 0
        d = s.query(zx).filter(zx.ly == '合同签订日期').first()
        if d:
            ywcs = d.cs if d.cs else 0
            nbcs = d.cs if d.cs else 0

        res = export_ruiyi17_invoices(rids, uv, ywcs, nbcs, cwsb, user, s=s)
        if res.get('code') != 1:
            return json_result(res.get('code', -1), res.get('msg', '导出失败'))
        savename = res.get('data', '')

        return json_result(code, msg, savename)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'预录单对照表失败: {str(e)}')
    finally:
        s.close()



# ==================== 主函数 ====================
def export_ruiyi17_invoices(
    rids: List[str],uv: str,ywcs: int,nbcs: int,cwsb: str,user,s
):
    """
    导出锐亿17格式的发票
    
    Args:
        session: SQLAlchemy Session
        number_list: 单据号码列表，默认 ['a', 'b']
        filename_template: 模板文件路径
        output_dir: Excel输出目录
        pdf_dir: PDF输出目录
        uv: UV标志
        ywcs: 业务参数
        cwsb: CWSB标志
    """
    
    app_path = config.data_upload_path
    filename_template = os.path.join(app_path, 'template', '锐亿17.xlsx')

    if not os.path.exists(filename_template):
        logger.error(f"模板文件不存在: {filename_template}")
        return {'code': -1, 'msg': f'模板文件{filename_template}不存在', 'data': None}
    
    # 计数器
    c = 0
    ps = 0
    i1 = 0
    
    for l,  rid in enumerate(rids):
        # 查询 kaiptz 表
        d = session.query(kaiptz).filter(
            kaiptz.rid == rid,
            kaiptz.kfdy == '可以'
        ).first()
        
        if not d:
            continue
        
        # 初始化变量
        yssb = ''
        ywsb = ''
        chyrq = ''
        yw = 0
        dkpd = ''
        zb = ''
        fssb = ''
        
        hyd = d.hyd or ''
        dkpd = d.cpmc or ''
        
        # 处理补报
        if d.cpmc == '补报':
            zb = '补'
            if d.chrq:
                chyrq = d.chrq.strftime('%Y-%m-%d') if hasattr(d.chrq, 'strftime') else str(d.chrq)
        else:
            chyrq = str(d.chrq) if d.chrq else ''
        
        # 获取 ysfp 前缀
        sfdl = (d.ysfp or '')[:2]
        sfdl1 = (d.yysfp or '')[:2]
        ywdyjl = d.ywdyjl or ''
        
        # 判断是否为义乌
        yw = 1 if d.kpdq == '义乌' else 0
        
        # 检查预填
        if d.hyd == '有预填':
            print("请注意此票有预填")
        
        # 处理 wstt 相关
        yssb = '1' if d.wstt and '优景' in d.wstt else ''
        
        # 获取开票内容
        jjdz = ''
        wstt1 = ''
        gs = ''
        dz = ''
        sw = ''
        bank = ''
        kpcxdh = ''
        
        kpnr_list = s.query(Kpnr).all()
        for kpnr in kpnr_list:
            gsjc = kpnr.gsjc or ''
            if gsjc and d.wstt and gsjc in d.wstt:
                gs = kpnr.wfgs or ''
                dz = kpnr.kpdz or ''
                sw = kpnr.kpsh or ''
                bank = kpnr.kpyh or ''
                wstt1 = gsjc
                jjdz = kpnr.kpjjdz or ''
                kpcxdh = kpnr.kpcxdh or ''
        
        # 义乌判断
        if d.wstt and '义乌' in d.wstt:
            ywsb = '1'
        
        # 设置寄件地址
        if not jjdz:
            if d.wstt and '义乌' in d.wstt:
                jjdz = '寄件地址：义乌市宗泽北路531号赛尔集团二楼   /邮编：322000/ 吴春燕 收/0579-85096055'
            else:
                jjdz = '寄件地址：宁波高新区光华路288号赛尔大厦25层  邮编315103/财务收/0574-27833931'
        
        # 处理 ysfp
        cx1 = (d.ysfp or '').upper()
        cx1z = clean_filename(cx1)
        
        # 处理 yfph 或 yysfp
        if yw == 1:
            cx7 = (d.yfph or '').upper()
        else:
            cx7 = (d.yysfp or '').upper()
        
        lxfs = d.lxfs or ''
        
        # 处理日期
        if d.htrq:
            cx2 = d.htrq.strftime('%Y-%m-%d') if hasattr(d.htrq, 'strftime') else str(d.htrq)
        else:
            cx2 = chyrq
        
        cx3 = d.ywry or ''
        cx4 = d.lxdh or ''
        
        # 处理生产厂家
        cx = clean_filename(d.sccj or '')
        
        wxfp = d.ysfp or ''
        number = d.rid
        
        # 查询 kaiptzsheet1 表
        sheet1_list = s.query(kaiptzsheet1).filter(
            kaiptzsheet1.pid == number,
            kaiptzsheet1.bgsl > 0
        ).all()
        
        if not sheet1_list:
            continue
        
        files = []
        for sheet1 in sheet1_list:
            i1 += 1
            
            # 处理 zb1
            zb1 = ''
            if sheet1.yfph:
                if len(sheet1.yfph) > 1:
                    zb1_char = sheet1.yfph[1] if len(sheet1.yfph) > 1 else ''
                    if zb1_char not in '0123456789':
                        zb1 = zb + sheet1.yfph[:2] + '组'
                    else:
                        zb1 = zb + sheet1.yfph[:1] + '组'
            
            if sheet1.zsl >= 0:
                ps += 1
                c += 1
                
                # 加载模板
                wb = load_workbook(filename_template)
                ws = wb.active
                
                # 确定合同章类型
                cpbh = f"{gs}合同章"
                if 'CSD-' in cx1 and 'VC0' in cx1.upper() and 'VC1' in cx1.upper():
                    cpbh = f"{gs}合同章景驰"
                elif 'CSD-' in cx1 and 'UV0' in cx1.upper() and 'UV1' in cx1.upper():
                    cpbh = f"{gs}合同章优景"
                
                # 检查是否禁止开票
                cyzgl = s.query(cyzglsheet).filter(
                    cyzglsheet.xm == cpbh,
                    cyzglsheet.zm == "禁止开票合同章"
                ).first()
                
                if not cyzgl:
                    # 添加合同章图片
                    t = s.query(tpzx).filter(
                        tpzx.cpbh == cpbh
                    ).first()
                    if t and t.tpmc:
                        if t.tpmc != None and t.tpmc.strip() != '' and t.tpmc != '[]':
                            photos = json.loads(t.tpmc)
                            src = photos[0].get('src') if len(photos) > 0 else None
                            img_path = os.path.join(config.data_upload_path, src)
                            if os.path.exists(img_path):
                                img = Image(img_path)
                                img = XLImage(img_path) # 选用 openpyxl 的 Image (兼容您给的 Image_Get)
                                            
                                # 仿照您的标准：留白 4 像素
                                img.width = 116  # 预估列宽转换像素后减4 (原代码大概120)
                                img.height = 96  # 预估行高(100)减4
                                col_idx = 1
                                row_idx = 17
                                offset_img(img, col_idx, row_idx, x_pad=2, y_pad=10)
                                ws.add_image(img)  # 位置可能需要调整
                
                # 写入 Excel 数据
                ws['A8'] = gs
                ws['C14'] = gs
                ws['C15'] = dz
                ws['C16'] = sw
                ws['C17'] = bank
                ws.row_dimensions[11].height = 15
                ws['G11'] = cx1
                ws['G12'] = cx7
                
                if yw == 0:
                    ws['C22'] = cx2
                
                ws['G44'] = cx1
                ws.row_dimensions[21].height = 15
                ws['C21'] = sheet1.zwpm or ''
                ws['G24'] = cx2
                ws['B22'] = float(sheet1.bgsl or 0)
                ws['E22'] = sheet1.hgjldw or ''
                ws['E24'] = '0'
                ws['G23'] = float(sheet1.sjkpze or 0)
                
                # 税率计算
                zsl = float(sheet1.zsl or 0)
                sjkpze = float(sheet1.sjkpze or 0)
                
                if zsl > 3 or zsl == 0:
                    je = round_to_two(sjkpze / 1.13)
                    ws['B23'] = je
                    ws['G22'] = '13%'
                    ws['E23'] = round_to_two(sjkpze - je)
                elif zsl == 3:
                    je = round_to_two(sjkpze / 1.03)
                    ws['B23'] = je
                    ws['G22'] = '3%'
                    ws['E23'] = round_to_two(sjkpze - je)
                elif zsl == 1:
                    je = round_to_two(sjkpze / 1.01)
                    ws['B23'] = je
                    ws['G22'] = '1%'
                    ws['E23'] = round_to_two(sjkpze - je)
                
                # 义乌特殊处理
                if ywsb == '1':
                    ws['c25'] = sheet1.cz or ''
                    ws['A16'] = '统一社会信用代码'
                    ws['A19'] = f"{int(zsl)}%"
                    ws['A11'] = '合同签订日期:'
                    if chyrq:
                        try:
                            chyrq_date = datetime.strptime(chyrq, '%Y-%m-%d') if isinstance(chyrq, str) else chyrq
                            ws['B11'] = (chyrq_date - timedelta(days=ywcs)).strftime('%Y-%m-%d')
                        except:
                            ws['B11'] = chyrq
                else:
                    ws['G21'] = f"{int(zsl)}%"
                
                ws['G45'] = f'(开票序号:{sheet1.kpxh})'
                ws['A20'] = f'发票快件查询 资料不清晰请拨打电话:{kpcxdh}'
                ws.row_dimensions[3].height = 15
                
                # 寄件地址
                if uv != '1':
                    cell_a9 = ws['A9']
                    cell_a9.font = Font(color='FF0000')
                    cell_a9.value = f'★{jjdz}'
                
                # 可思达特殊处理
                if '可思达' not in gs:
                    ws['C18'] = sheet1.ywrya or ''
                    
                    if sheet1.lxdh:
                        ws['G18'] = sheet1.lxdh
                    else:
                        y = s.query(ywrylx).filter(ywrylx.ryxm == sheet1.ywrya).first()
                        if y and y.zjdh:
                            ws['G18'] = y.zjdh
                
                ws['B24'] = sheet1.hyd or ''
                ws['G50'] = cx1
                ws.row_dimensions[53].height = 15
                ws['k53'] = sheet1.htxs or ''
                ws['c53'] = sheet1.htxs or ''
                ws['c54'] = sheet1.sccj or ''
                ws['c55'] = f"{int(zsl)}%"
                
                if zsl < 4:
                    ws['A71'] = '重要提醒：请贵司提供准确对公账户收款信息，以便我司及时支付款项，若未开立对公账户，请尽快开立。收款账户信息请与发票一并寄回我司．'
                
                ws['c56'] = chyrq
                ws.row_dimensions[57].height = 15
                
                # 寄件地址颜色
                cell_c57 = ws['c57']
                if yssb == '1':
                    cell_c57.font = Font(color='FF0000')
                ws['k57'] = jjdz
                ws['c57'] = jjdz
                
                # AMZ 特殊处理 - 第二个工作表
                if 'AMZ' in wxfp:
                    if len(wb.worksheets) > 1:
                        ws2 = wb.worksheets[1]
                        ws2.activate()
                        
                        # 设置列宽
                        ws2.column_dimensions['A'].width = 20
                        ws2.column_dimensions['B'].width = 20
                        ws2.column_dimensions['C'].width = 10
                        ws2.column_dimensions['D'].width = 15
                        
                        # 设置表头
                        for col, header in [('A', '中文品名'), ('B', '货号'), ('C', '出货数量'), ('D', '采购合同')]:
                            cell = ws2[f'{col}1']
                            cell.value = header
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # 查询 kaiptzsheet 表
                        sheet_list = s.query(kaiptzsheet).filter(
                            kaiptzsheet.pid == number
                        ).all()
                        
                        iz = 1
                        for item in sheet_list:
                            iz += 1
                            ws2.row_dimensions[iz].height = 15
                            
                            # 查询 cymxsheet
                            cymx = s.query(cymxsheet).filter(
                                cymxsheet.cywyzd == item.cywyzd
                            ).first()
                            
                            if cymx:
                                ws2[f'A{iz}'] = cymx.zwpm or ''
                                ws2[f'A{iz}'].alignment = Alignment(horizontal='center', vertical='center')
                            
                            ws2[f'B{iz}'] = item.cpbh or ''
                            ws2[f'B{iz}'].alignment = Alignment(horizontal='center', vertical='center')
                            ws2[f'C{iz}'] = float(item.chsl or 0)
                            ws2[f'C{iz}'].alignment = Alignment(horizontal='center', vertical='center')
                            ws2[f'D{iz}'] = item.cght or ''
                            ws2[f'D{iz}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # 保存 Excel 文件
                if cwsb == '1':
                    filename_base = f"{zb1}{cx}{cx1z}{sheet1.zwpm}({i1})"
                else:
                    filename_base = f"{zb1}{cx}{cx1z}({i1})"
                
                filename_clean = clean_filename(filename_base) + f"_{time.strftime('%Y-%m-%d_%H-%M-%S')}_{uuid.uuid4()}.xlsx"
                excel_path = os.path.join(config.tpm_path, filename_clean)
                files.append(filename_clean)
                wb.save(excel_path)
                print(f"已保存 Excel: {excel_path}")
                
                # 导出 PDF
                if cwsb == '1':
                    pdf_filename = f"{cx}{cx1}{sheet1.zwpm}{ps}{wstt1}.pdf"
                else:
                    pdf_filename = f"{cx}{cx1}{ps}{wstt1}.pdf"
                
                # pdf_path = os.path.join(pdf_dir, clean_filename(pdf_filename))
                # 注意：openpyxl 不支持直接导出 PDF，需要使用其他库如 win32com 或 pdfkit
                # 这里只是记录路径，实际转换需要额外处理
                # print(f"PDF 导出路径: {pdf_path}")
                
        
        # 更新 kaiptz 表的打印状态
        if ywdyjl == '是':
            ywdyjl = f"{user.username}:{datetime.now().strftime('%Y-%m-%d')}"
        else:
            ywdyjl = f"{ywdyjl}/ {user.username}:{datetime.now().strftime('%Y-%m-%d')}"
        
        d.sfdy = '可以'
        d.ywdyjl = ywdyjl
        s.add(d)

    s.commit()
    zip_filename = ''
    if len(files) > 0:
        zip_filename = f"锐亿发票_{time.strftime('%Y-%m-%d_%H-%M-%S')}.zip"
        zip_filepath = os.path.join(config.tmp_path, zip_filename)
        zipFile = zipfile.ZipFile(zip_filepath, 'w')
        for f in files:
            fn = os.path.join(config.tmp_path, f)
            if os.path.exists(fn):
                zipFile.write(fn, f, zipfile.ZIP_DEFLATED)
        zipFile.close()

    return {'code': 1, 'msg': '导出成功', 'data': zip_filename}
    

    """
开票通知.开票打印批量许可
对应原Pascal: 开票打印批量许可
"""

@any_route('/api/saier/billed_notice/print_permit', methods=['POST'])
@require_token
async def view_billed_notice_print_permit(request):
    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        rids = j.get('rids', [])
        code = 1
        msg = '操作成功'
        position = j.get('position', '')
        kind = 0
        if '义乌' in position:
            kind = 1

        for rid in rids:
            d = s.query(kaiptz).filter(kaiptz.rid == rid).first()
            if not d:
                continue
            flag = 0
            wstt = d.wstt or ''
            if kind == 1 and d.sfdy == '可以' and '义乌' in wstt:
                d.kfdy = '可以'
                d.ywqr = '可以'
                flag = 1
            if kind == 0 and d.sfdy != '可以':
                d.kfdy = '可以'
                d.kprq = time.strftime('%Y-%m-%d')
                flag = 1
            if flag == 1:
                d.xgry = user.username
                d.mtime = time.strftime('%Y-%m-%d %H:%M:%S')
                d.modi_uid = user.rid
                s.add(d)
                c = s.query(kaiptzsheet1).filter(kaiptzsheet1.pid == d.rid).all()
                for r in c:
                    username = r.ywrya or ''
                    if user.username != username and username != '' and username != None:
                        xxnr = '外销发票号码为:' + d.fphm + '中文品名:' + d.zwpm + '工厂:' + d.sccj + '开票资料可打印' + time.strftime('%Y-%m-%d %H:%M:%S')
                        data = {"type":"success","title":"采购付款核对通知",
                            "msg": xxnr,
                            "module": "开票通知",
                            "rid": d.rid, "rids":[],
                            "btns":[{"title":"打开","icon":"any-rmb-full","name":"open","color":"red"}]}
                        
                        row = {
                            "xxly": '开票通知',
                            "gdht": '',
                            "wxht": '',
                            "cght": '',
                            "yhdh": '',
                            "xxnr": xxnr,
                            "jsr": str(username),
                            "sys_path": "",
                            "spsq": user.username
                        }
                        res = module_xxck_new([row],user,s)
                        if res.get('code')!=1:
                            s.rollback()
                            return json_result(res.get('code'), res.get('code'))
                        await messages.message_to_user(username,data,MSG_KIND_NOTICE_RECORD,request)
            
        s.commit()
        return json_result(code, msg)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'预录单对照表失败: {str(e)}')
    finally:
        s.close()


@any_route('/api/saier/billed_notice/kfdy/change', methods=['POST'])
@require_token
async def view_billed_notice_kfdy_change(request):
    s = Session()
    try:
        # user = request.current_user
        # j = await request.json()
        code = 1
        msg = '操作成功'
        data = 30
        d = s.query(zx.cs).filter(zx.ly=='开票逾期').first()
        if d and d.cs:
            data = int(d.cs)

        return json_result(code, msg, data)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'预录单对照表失败: {str(e)}')
    finally:
        s.close()


def getnum(value):
    """将值转换为浮点数"""
    if value is None or value == '':
        return 0.0
    try:
        return float(value)
    except:
        return 0.0

def getint(value):
    """将值转换为整数"""
    if value is None or value == '':
        return 0
    try:
        return int(float(value))
    except:
        return 0

def round_decimal(value, decimal_places=2):
    """四舍五入"""
    return Decimal(str(value)).quantize(Decimal('0.' + '0' * decimal_places), rounding=ROUND_HALF_UP)

def process_invoice_data(da1, da2, filename, ly, user, s):
    """
    处理发票数据的主函数
    da1: '2026-05-01'
    da2: '2026-05-31'
    filename: 源Excel文件路径
    filename14: 目标模板Excel文件路径
    filename2: 输出文件路径前缀
    ly: 业务地区
    """
    
    # 打开源Excel文件
    wb_source = load_workbook(filename)
    ws_source = wb_source.active
    
    # 初始化列表
    cght1 = []  # 购货方列表
    fpdm1 = []  # 发票代码列表
    zzsl1 = []  # 税率列表
    
    # 读取第一行数据
    i = 0
    gcfp = ws_source['A2'].value
    fpje = ws_source['H2'].value
    bhsj = ws_source['I2'].value
    zzsl = ws_source['J2'].value
    se = ws_source['K2'].value
    sh = ws_source['O2'].value
    fpdm = ws_source['Q2'].value
    kprq = str(ws_source['S2'].value)[:10] if ws_source['S2'].value else ''
    yt = ws_source['P2'].value or '货款'
    
    while gcfp != None and gcfp != '':
        kpgc = ''
        # 获取开票公司名称
        if sh:
            company = s.query(newcwcs).filter(newcwcs.shui == sh).first()
            if company:
                kpgc = company.company_name
        i += 1
        # 添加到列表
        if gcfp not in cght1:
            cght1.append(gcfp)
            fpdm1.append(fpdm)
            zzsl1.append(zzsl)
        
        # 计算有效期
        if kprq:
            kprq_date = datetime.strptime(kprq, '%Y-%m-%d')
            yx = (kprq_date + timedelta(days=180)).strftime('%Y-%m-%d')
        else:
            yx = None
        
        # 查询kaiptzxq表中是否存在记录
        existing_record = s.query(kaiptzxq).filter(
            func.date_format(kaiptzxq.dprq, '%Y-%m-%d') >= da1,
            func.date_format(kaiptzxq.dprq, '%Y-%m-%d') <= da2,
            kaiptzxq.gcfp == gcfp,
            (kaiptzxq.fpdm == fpdm) | (kaiptzxq.fpdm == ',') | (kaiptzxq.fpdm == '0')
        ).first()
        logger.error(f"查询kaiptzxq表: da1={da1}, da2={da2}")    
        logger.error(f"查询kaiptzxq表: gcfp={gcfp}, fpdm={fpdm}, existing_record={existing_record}")
        if existing_record:
            # 更新记录
            update_data = {
                'gcfp': gcfp,
                'fpje': getnum(fpje),
                'bhsj': getnum(bhsj),
                'zzsl': getint(zzsl),
                'se': getnum(se),
                'sh': sh,
                'fpdm': fpdm,
                'kprq': kprq,
                'yt': yt,
                'rzyxq': yx,
                'sfrz': '是',
                'mtime': time.strftime('%Y-%m-%d %H:%M:%S'),
                'modi_uid': user.rid
            }
            
            if kpgc:
                update_data['kpgc'] = kpgc
            
            # 处理税率转换
            if getint(zzsl) == 16 and existing_record.tsl == 17:
                update_data['tsl'] = 16
                update_data['tse'] = getnum(se)
            else:
                if getint(zzsl) == existing_record.tsl:
                    update_data['tse'] = getnum(se)
                    update_data['tsl'] = existing_record.tsl
                else:
                    update_data['tsl'] = existing_record.tsl
                    # 计算调整后的税额
                    tse_value = getnum(fpje) / (1 + getint(zzsl) / 100) * existing_record.tsl / 100
                    update_data['tse'] = float(round_decimal(tse_value))
            
            # 执行更新
            s.query(kaiptzxq).filter(
                func.date_format(kaiptzxq.dprq, '%Y-%m-%d') >= da1,
                func.date_format(kaiptzxq.dprq, '%Y-%m-%d') <= da2,
                kaiptzxq.gcfp == gcfp,
                kaiptzxq.rid == existing_record.rid
            ).update(update_data, synchronize_session=False)
            
            # 更新zlbinvoice表
            s.query(zlbinvoice).filter(
                zlbinvoice.CodeOfInvoice == fpdm,
                zlbinvoice.NumberOfInvoice == gcfp
            ).update({'fis_use': '是'})
        
        # 查询kpwy和kpxh
        kpwy_record = s.query(kaiptzxq).filter(
            func.date_format(kaiptzxq.dprq, '%Y-%m-%d') >= da1,
            func.date_format(kaiptzxq.dprq, '%Y-%m-%d') <= da2,
            kaiptzxq.gcfp == gcfp,
            (kaiptzxq.fpdm == fpdm) | (kaiptzxq.fpdm == ',')
        ).first()
        
        if kpwy_record and kpwy_record.kpxh:
            fpwy = kpwy_record.kpxh
            
            # 查询金额合计
            sum_fpje = s.query(func.sum(kaiptzxq.fpje)).filter(
                kaiptzxq.kpxh == fpwy
            ).scalar() or 0
            
            # 查询kaiptzsheet1
            sheet_record = s.query(kaiptzsheet1).filter(
                kaiptzsheet1.kpxh == fpwy
            ).first()
            
            if sheet_record:
                update_sheet_data = {
                    'rzje': float(sum_fpje),
                    'zzsl': getint(zzsl)
                }
                
                if kpgc:
                    update_sheet_data['kpgc'] = kpgc
                
                if sheet_record.dpjez > 0:
                    update_sheet_data['frzje'] = sheet_record.dpjez - float(sum_fpje)
                else:
                    update_sheet_data['frzje'] = 0
                
                s.query(kaiptzsheet1).filter(
                    kaiptzsheet1.kpxh == fpwy
                ).update(update_sheet_data, synchronize_session=False)
        
        # 读取下一行数据
        gcfp = ws_source['A' + str(2 + i)].value
        fpje = ws_source['H' + str(2 + i)].value
        bhsj = ws_source['I' + str(2 + i)].value
        zzsl = ws_source['J' + str(2 + i)].value
        se = ws_source['K' + str(2 + i)].value
        sh = ws_source['O' + str(2 + i)].value
        fpdm = ws_source['Q' + str(2 + i)].value
        kprq_val = ws_source['S' + str(2 + i)].value
        kprq = str(kprq_val)[:10] if kprq_val else ''
        yt = ws_source['P' + str(2 + i)].value or '货款'
    
    
    # 打开目标Excel模板
    tmp_path = os.path.join(config.data_upload_path, 'template', '发票导入.xlsx')
    wb_target = load_workbook(tmp_path)
    ws_target = wb_target.active
    # 设置表头
    headers = {
        'O1': '税   号(销方名称)',
        'AG1': '我司抬头',
        'AI1': '开票序号',
        'AJ1': '是否电汇',
        'AK1': '盖章收回',
        'AL1': '合同详情',
        'AM1': '是否电汇',
        'AN1': '盖章收回',
        'AO1': '出货工厂',
        'AP1': '微 信 号',
        'AQ1': '盖章日期',
        'AR1': '电汇日期',
        'AS1': '合同收回',
        'AT1': '付款详情',
        'AU1': '收支详情',
        'AV1': '折扣费用',
        'AW1': '业务地区'
    }
    
    for cell, value in headers.items():
        ws_target[cell] = value
    
    # 处理数据
    e = 0
    for l in range(len(cght1)):
        records = s.query(kaiptzxq).filter(
            kaiptzxq.dprq >= da1,
            kaiptzxq.dprq <= da2,
            kaiptzxq.gcfp == cght1[l],
            kaiptzxq.fpdm == fpdm1[l]
        ).all()
        
        for record in records:
            e += 1
            
            # 查询kaiptz
            r = s.query(kaiptz).filter(
                kaiptz.rid == record.pid
            ).first()
            wstt = r.wstt if r else ''
            
            # 查询kaiptzsheet1
            sheet_record = s.query(kaiptzsheet1).filter(
                kaiptzsheet1.kpxh == record.kpxh
            ).first()
            
            if sheet_record:
                yfph = sheet_record.yfph[:99] if sheet_record.yfph else ''
                if not yfph:
                    yfph = record.ysfp or ''
                
                row_num = e + 1
                
                # 填充数据
                ws_target['A' + str(row_num)] = record.gcfp or ''
                ws_target['B' + str(row_num)] = yfph
                ws_target['C' + str(row_num)] = record.chrq or ''
                ws_target['D' + str(row_num)] = record.cghbdm or 'RMB'
                ws_target['E' + str(row_num)] = float(record.fkje) if record.fkje else 0
                ws_target['F' + str(row_num)] = float(record.htje) if record.htje else 0
                ws_target['G' + str(row_num)] = record.fkrq or ''
                ws_target['H' + str(row_num)] = float(record.fpje) if record.fpje else 0
                ws_target['I' + str(row_num)] = float(record.bhsj) if record.bhsj else 0
                
                # 税率处理
                if record.tsl == 3:
                    ws_target['J' + str(row_num)] = '3'
                else:
                    ws_target['J' + str(row_num)] = str(record.zzsl) if record.zzsl else ''
                
                ws_target['K' + str(row_num)] = float(record.se) if record.se else 0
                ws_target['L' + str(row_num)] = str(record.tsl) if record.tsl else ''
                ws_target['M' + str(row_num)] = float(record.tse) if record.tse else 0
                ws_target['N' + str(row_num)] = record.jsrm or ''
                ws_target['O' + str(row_num)] = record.sh or ''
                ws_target['P' + str(row_num)] = record.yt or ''
                ws_target['Q' + str(row_num)] = record.fpdm or ''
                ws_target['R' + str(row_num)] = (record.bzsm or '')[:240]
                ws_target['S' + str(row_num)] = record.dprq or ''
                ws_target['T' + str(row_num)] = record.ysfp or ''
                ws_target['U' + str(row_num)] = record.fpwk or ''
                ws_target['V' + str(row_num)] = ly
                ws_target['W' + str(row_num)] = record.chgc or ''
                ws_target['X' + str(row_num)] = ''
                ws_target['Z' + str(row_num)] = record.tqfk or ''
                ws_target['AA' + str(row_num)] = record.ywbm or ''
                ws_target['AB' + str(row_num)] = record.sqrq or ''
                ws_target['AC' + str(row_num)] = float(record.sqje) if record.sqje else 0
                ws_target['AD' + str(row_num)] = record.sfjq or ''
                ws_target['AE' + str(row_num)] = record.zwpm or ''
                ws_target['AF' + str(row_num)] = record.kpgc or ''
                ws_target['AG' + str(row_num)] = wstt
                ws_target['AI' + str(row_num)] = record.kpxh or ''
                ws_target['AJ' + str(row_num)] = record.fkbz or ''
                ws_target['AK' + str(row_num)] = record.drrq or ''
                ws_target['AL' + str(row_num)] = (sheet_record.htxs or '')[:240]
                ws_target['AM' + str(row_num)] = sheet_record.sfdh or ''
                ws_target['AN' + str(row_num)] = sheet_record.gzsh or ''
                ws_target['AO' + str(row_num)] = sheet_record.sccj2 or ''
                ws_target['AP' + str(row_num)] = sheet_record.wxh or ''
                ws_target['AQ' + str(row_num)] = sheet_record.gzrq1 or ''
                ws_target['AR' + str(row_num)] = sheet_record.dhrq1 or ''
                ws_target['AS' + str(row_num)] = sheet_record.htsh or ''
                ws_target['AT' + str(row_num)] = record.fkxq or ''
                ws_target['AU' + str(row_num)] = record.szxq or ''
                ws_target['AV' + str(row_num)] = float(record.zkfy1) if record.zkfy1 else 0
                ws_target['AW' + str(row_num)] = record.dq or ''
    
    # 保存文件
    output_filename = f"发票导入{user.username}{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    output_path = os.path.join(config.tmp_path, output_filename)
    wb_target.save(output_path)
    
    # 提交所有更改
    s.commit()
    # 关闭工作簿
    wb_source.close()
    wb_target.close()
    return output_filename


"""
导入发票生成
对应原Pascal: 导入发票生成
"""
@any_route('/api/saier/billed_notice/import_invoice', methods=['POST'])
@require_token
async def view_billed_notice_import_invoice(request):
    # 1. 检查用户是否是财务岗位 (对应 tmpcom2 查询)
    s = Session()
    try:
        j = await request.form()
        module = form_value(j, 'module', '开票通知')
        ksrq = form_value(j, 'ksrq', '')
        jsrq = form_value(j, 'jsrq', '')
        logger.error(f"Received request for import_invoice with j: {j}")
        logger.error(f"Received ksrq: {ksrq}, jsrq: {jsrq}")
        user = request.current_user
        org = get_user_path(user.username)
        ly = '宁波'
        position = org.get('position')
        if '义乌' in position:
            ly = '义乌'

        if ksrq == '' or jsrq == '' or ksrq == None or jsrq == None:
            return json_result('到票开始日和到票结束日不能为空')
        if ksrq > jsrq:
            return json_result('到票开始日不能大于到票结束日')
        
        # 2. 获取上传的Excel文件
        file = form_value(j,'file',None)
        if is_none(file):
            return json_result(ERR_PARAM_NOT_ENOUGH)
        # 保存临时文件
        temp_dir = config.tmp_path
        temp_file = os.path.join(temp_dir, f"{uuid.uuid4()}_{file.filename}")
        # logger.error(f"Saving uploaded file to: {temp_file}")
        data = await file.read()
        write_file(temp_file, data, 'wb', encoding=None)
        
        filename = process_invoice_data(ksrq, jsrq, temp_file, ly, user, s)

        return json_result(0, "操作成功", filename)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()

"""
批量到票信息导入
对应原Pascal: 批量到票信息导入
"""
@any_route('/api/saier/billed_notice/invoice_batch_import', methods=['POST'])
@require_token
async def view_billed_notice_invoice_batch_import(request):
    # 1. 检查用户是否是财务岗位 (对应 tmpcom2 查询)
    s = Session()
    try:
        j = await request.form()
        module = form_value(j, 'module', '开票通知')
        ly = '宁波'
        position = form_value(j, 'position', '宁波')
        if '义乌' in position:
            ly = '义乌'

        # 2. 获取上传的Excel文件
        file = form_value(j,'file',None)
        if is_none(file):
            return json_result(ERR_PARAM_NOT_ENOUGH)
        # 保存临时文件
        temp_dir = config.tmp_path
        temp_file = os.path.join(temp_dir, f"{uuid.uuid4()}_{file.filename}")
        # logger.error(f"Saving uploaded file to: {temp_file}")
        data = await file.read()
        write_file(temp_file, data, 'wb', encoding=None)
        
        filename = batch_excel_invoice_import(s, user, temp_file, ly)

        return json_result(0, "操作成功", filename)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()

# 获取列映射
def get_column_letter_for_index(idx: int) -> str:
    """根据索引获取Excel列字母 (1->A, 27->AA)"""
    if idx <= 26:
        return chr(64 + idx)
    else:
        first = idx // 26
        second = idx % 26
        if second == 0:
            first -= 1
            second = 26
        return chr(64 + first) + chr(64 + second)

def batch_excel_invoice_import(s, user, filename: str, ly: str):
    """
    将Delphi代码转换为Python
    
    参数:
        filename: 第一个Excel文件路径（发票导入）
        filename14: 第二个Excel文件路径（导入到系统）
        filename2: 输出文件路径前缀
        ly: 来源标识
        IApplication: 应用对象（用于获取用户信息）
        cght1, fpdm1: 传入的列表（引用传递）
        tmpstr1: 临时字符串列表
    """
    
    # ==================== 第一部分：读取发票数据并处理 ====================
    wb = load_workbook(filename)
    ws = wb.worksheets[0]  # worksheets[1] 在openpyxl中是0-based
    tmpstr1 = []
    fpdm1 = []
    cght1 = []
    # 激活第一个工作表（openpyxl不需要activate，直接操作即可）
    # 在openpyxl中不需要activate，直接使用ws
    
    # 获取第一行的所有列（A到AT，共46列）
    A, B, C, D, E1, F, G, H, I1, J, K, L, M, N, O = ['CC'] * 15
    P, Q, R, S, T, U, V, W, X, Y, Z, AA, AB, AC, AD = ['CC'] * 15
    AE, AF, AG, AH, AI, AJ, AK, AL, AM, AN, AO, AP, AQ, AR, AS1, AT = ['CC'] * 16
    try:
        # 读取第一行，找到对应字段的列位置
        zs1 = []
        for col_idx in range(1, 47):  # 46列
            col_letter = get_column_letter_for_index(col_idx)
            cell_value = ws[f"{col_letter}1"].value
            zs1.append(cell_value if cell_value else '')
        
        # 建立字段到列的映射
        for i2, header in enumerate(zs1):
            col_letter = get_column_letter_for_index(i2 + 1)
            if header == '工厂发票':
                A = col_letter
            elif header == '开票序号':
                B = col_letter
            elif header == '发票金额':
                C = col_letter
            elif header == '完成开票':
                D = col_letter
            elif header == '开票日期':
                E1 = col_letter
            elif header == '不含税价':
                F = col_letter
            elif header == '增值税率':
                G = col_letter
            elif header == '税    额':
                H = col_letter
            elif header in ('税   号', '销方名称', '税   号(销方名称)'):
                I1 = col_letter
            elif header == '用途':
                J = col_letter
            elif header == '发票代码':
                K = col_letter
            elif header == '是否电汇':
                L = col_letter
            elif header == '盖章收回':
                M = col_letter
            elif header == '到票日期':
                N = col_letter
            elif header == '认证有效期':
                O = col_letter
        
        # 处理数据行
        i = 0
        sb = ''
        
        while True:
            row_num = 2 + i
            gcfp = ws[f"{A}{row_num}"].value if ws[f"{A}{row_num}"].value else ''
            kpxh = ws[f"{B}{row_num}"].value if ws[f"{B}{row_num}"].value else ''
            
            if not gcfp and not kpxh:
                break
            
            # 读取各字段值
            dpje = ws[f"{C}{row_num}"].value if ws[f"{C}{row_num}"].value else ''
            wckp = ws[f"{D}{row_num}"].value if ws[f"{D}{row_num}"].value else ''
            if not wckp:
                wckp = '否'
            
            kprq = ws[f"{E1}{row_num}"].value if ws[f"{E1}{row_num}"].value else ''
            bhsj = ws[f"{F}{row_num}"].value if ws[f"{F}{row_num}"].value else ''
            zzsl = ws[f"{G}{row_num}"].value if ws[f"{G}{row_num}"].value else ''
            se = ws[f"{H}{row_num}"].value if ws[f"{H}{row_num}"].value else ''
            sh = ws[f"{I1}{row_num}"].value if ws[f"{I1}{row_num}"].value else ''
            yt = ws[f"{J}{row_num}"].value if ws[f"{J}{row_num}"].value else ''
            fpdm = ws[f"{K}{row_num}"].value if ws[f"{K}{row_num}"].value else ''
            fkbz = ws[f"{L}{row_num}"].value if ws[f"{L}{row_num}"].value else ''
            drrq = ws[f"{M}{row_num}"].value if ws[f"{M}{row_num}"].value else ''
            dprq = ws[f"{N}{row_num}"].value if ws[f"{N}{row_num}"].value else ''
            rzyxq = ws[f"{O}{row_num}"].value if ws[f"{O}{row_num}"].value else ''
            
            # 处理空值
            if not kprq:
                kprq = datetime.now().strftime('%Y-%m-%d')
            if not dprq:
                dprq = datetime.now().strftime('%Y-%m-%d')
            
            if not rzyxq and kprq:
                # 计算认证有效期 = 开票日期 + 180天
                kprq_date = datetime.strptime(str(kprq), '%Y-%m-%d')
                rzyxq = (kprq_date + timedelta(days=180)).strftime('%Y-%m-%d')
            
            if not yt:
                yt = '货款'
            if not fpdm:
                fpdm = ','
            
            # 验证开票序号格式
            xhsw = str(kpxh)[0] if kpxh else ''
            if xhsw and not xhsw.isalpha():
                tmpstr1.append(f'开票序号: {kpxh}非标准格式，发票号：{gcfp}导入无效')
                sb = '1'
            else:
                # 查询 kaiptzsheet1
                d = s.query(kaiptzsheet1).filter_by(kpxh=kpxh).first()
                
                if not d:
                    tmpstr1.append(f'开票序号: {kpxh}不存在，发票号：{gcfp}导入无效')
                    sb = '1'
                else:
                    number1 = d.rid
                    dpxq = d.dpxq or ''
                    kpgc = d.kpgc or ''
                    sjkpze = float(d.sjkpze or 0)
                    dpjez = float(d.dpjez or 0)
                    bgsl = float(d.bgsl or 0)
                    bgslz = float(d.dpslz or 0)
                    
                    # 查询 kaiptz
                    c = s.query(kaiptz).filter_by(rid=d.rid).first()
                    chrq = kaiptz.chrq if kaiptz else ''
                    
                    # 检查是否已存在
                    existing = s.query(kaiptzxq).filter_by(
                        pid=d.rid,
                        gcfp=gcfp
                    ).first()
                    
                    if existing:
                        tmpstr1.append(f'开票序号: {kpxh}发票号：{gcfp}已存在，不能导入')
                        sb = '1'
                    else:
                        # 插入 kaiptzxq 记录
                        new_kaiptzxq = kaiptzxq(
                            pid=d.pid,
                            gcfp=gcfp,
                            chrq=chrq,
                            ckhwbgdh=d.ckhwbgdh,
                            cghbdm=d.cghbdm,
                            jsrm=d.ywrya,
                            kprq=kprq,
                            wxfp=d.fphm,
                            ysfp=d.ysfp,
                            fpwk='否',
                            dpdd=ly,
                            chgc=d.sccj,
                            zwpm=d.zwpm,
                            dprq=dprq,
                            rzyxq=rzyxq,
                            kpwy=d.kpwy,
                            htje=float(d.gczj or 0),
                            ywbm=d.ywbm,
                            kpgc=d.kpgc,
                            cght=d.cght,
                            wstt=d.wstt,
                            dlfp=gcfp,
                            dlje=float(dpje) if dpje else 0,
                            bzsm=d.bzsm,
                            sfrz='是',
                            djsj=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                            kpxh=kpxh,
                            fpdm=fpdm,
                            bhsj=float(bhsj) if bhsj else 0,
                            zzsl=int(zzsl) if zzsl else 0,
                            sh=sh,
                            se=float(se) if se else 0,
                            yt=yt,
                            fkbz=d.sfdh,
                            drrq=d.gzsh,
                            cz=d.cz,
                            fpje=float(dpje) if dpje else 0,
                            fkxq=d.fkxq,
                            szxq=d.szxq,
                            dq=d.dq,
                            zkfy1=float(d.zkfy1 or 0)
                        )
                        
                        # 处理税率逻辑
                        zsl = int(kaiptzsheet1.zsl or 0)
                        zzsl_int = int(zzsl) if zzsl else 0
                        
                        if zsl == 17 and zzsl_int == 16:
                            new_kaiptzxq.tsl = 16
                            new_kaiptzxq.tse = float(se) if se else 0
                        elif zzsl_int == zsl:
                            new_kaiptzxq.tse = float(se) if se else 0
                            new_kaiptzxq.tsl = zsl
                        else:
                            new_kaiptzxq.tsl = zsl
                            if dpje and zzsl_int:
                                dpje_float = float(dpje)
                                new_kaiptzxq.tse = round(dpje_float / (1 + zzsl_int / 100) * zsl / 100 * 100) / 100
                        
                        s.add(new_kaiptzxq)
                        
                        # 更新 zlbinvoice
                        s.query(zlbinvoice).filter_by(
                            CodeOfInvoice=fpdm,
                            NumberOfInvoice=gcfp
                        ).update({'fis_use': '是'}, synchronize_session=False)
                        
                        # 更新 kaiptzsheet1
                        yqrq = ''
                        if d.yqrq and dprq > d.yqrq:
                            yqrq = '是'
                        
                        s.query(kaiptzsheet1).filter_by(rid=number1).update({
                            'dlfp': gcfp,
                            'dlje': float(dpje) if dpje else 0,
                            'gcfp': '',
                            'fpyq': yqrq,
                            'dpje': 0,
                            'dpsl': 0,
                            'sfdp': '是',
                            'zzsl': int(zzsl) if zzsl else 0,
                            'wckp': '是' if (sjkpze - dpjez - (float(dpje) if dpje else 0)) < 10 else wckp,
                            'dpjez': dpjez + (float(dpje) if dpje else 0),
                            'dpdd': ly,
                            'dprq': dprq
                        }, synchronize_session=False)
                        
                        # 更新 dpxq
                        new_dpxq = f"工厂发票:{gcfp}到票日期:{dprq}开票工厂:{kpgc}到票金额:{dpje}"
                        if dpxq:
                            new_dpxq = f"{dpxq};\n{new_dpxq}"
                        s.query(kaiptzsheet1).filter_by(rid=number1).update({'dpxq': new_dpxq}, synchronize_session=False)
            
            i += 1
        
        # ==================== 第二部分：生成导出Excel ====================
        file_path = os.path.join(config.data_upload_path, "template", "发票导入.xlsx")
        wb_out = load_workbook(file_path)
        ws_out = wb_out.worksheets[0]
        
        # 添加表头
        headers = {
            'O1': '税   号(销方名称)',
            'AG1': '我司抬头',
            'AI1': '开票序号',
            'AJ1': '是否电汇',
            'AK1': '盖章收回',
            'AL1': '合同详情',
            'AM1': '是否电汇',
            'AN1': '盖章收回',
            'AO1': '出货工厂',
            'AP1': '微 信 号',
            'AQ1': '盖章日期',
            'AR1': '电汇日期',
            'AS1': '合同收回',
            'AT1': '付款详情',
            'AU1': '收支详情',
            'AV1': '折扣费用',
            'AW1': '业务地区'
        }
        
        for cell, value in headers.items():
            ws_out[cell].value = value
        
        # 查询并填充数据
        e = 0
        for l1 in range(len(cght1)):
            gcfp_val = cght1[l1]
            kpxh_val = fpdm1[l1]
            
            records = s.query(kaiptzxq).filter_by(gcfp=gcfp_val, kpxh=kpxh_val).all()
            
            for record in records:
                e += 1
                row = e + 1
                
                # 查询 kaiptz
                kaiptz = s.query(kaiptz).filter_by(rid=record.pid).first()
                wstt = kaiptz.wstt if kaiptz else ''
                
                # 查询 kaiptzsheet1
                c = s.query(kaiptzsheet1).filter_by(kpxh=record.kpxh).first()
                
                yfph = ''
                if c:
                    yfph = (c.ysfp1 or '')[:99]
                    if not yfph:
                        yfph = record.ysfp or ''
                
                # 填充数据
                data_mapping = {
                    f'A{row}': record.gcfp,
                    f'B{row}': yfph,
                    f'C{row}': record.chrq,
                    f'D{row}': record.cghbdm or 'RMB',
                    f'E{row}': float(record.fkje) if hasattr(record, 'fkje') and record.fkje else '',
                    f'F{row}': float(record.htje) if record.htje else '',
                    f'G{row}': record.fkrq if hasattr(record, 'fkrq') else '',
                    f'H{row}': float(record.fpje) if record.fpje else '',
                    f'I{row}': float(record.bhsj) if record.bhsj else '',
                    f'J{row}': '3' if record.tsl == 3 else str(record.zzsl) if record.zzsl else '',
                    f'K{row}': float(record.se) if record.se else '',
                    f'L{row}': str(record.tsl) if record.tsl else '',
                    f'M{row}': float(record.tse) if record.tse else '',
                    f'N{row}': record.jsrm,
                    f'O{row}': record.sh,
                    f'P{row}': record.yt,
                    f'Q{row}': record.fpdm,
                    f'R{row}': (record.bzsm or '')[:240],
                    f'S{row}': record.dprq,
                    f'T{row}': record.ysfp,
                    f'U{row}': record.fpwk,
                    f'V{row}': ly,
                    f'W{row}': record.chgc,
                    f'X{row}': '',
                    f'Z{row}': record.tqfk if hasattr(record, 'tqfk') else '',
                    f'AA{row}': record.ywbm,
                    f'AB{row}': record.sqrq if hasattr(record, 'sqrq') else '',
                    f'AC{row}': float(record.sqje) if hasattr(record, 'sqje') and record.sqje else '',
                    f'AD{row}': record.sfjq if hasattr(record, 'sfjq') else '',
                    f'AE{row}': record.zwpm,
                    f'AG{row}': wstt,
                    f'AI{row}': record.kpxh,
                    f'AJ{row}': record.fkbz,
                    f'AK{row}': record.drrq,
                    f'AL{row}': (c.htxs or '')[:240] if c else '',
                    f'AM{row}': c.sfdh if c else '',
                    f'AN{row}': c.gzsh if c else '',
                    f'AO{row}': c.sccj2 if c else '',
                    f'AP{row}': c.wxh if c else '',
                    f'AQ{row}': c.gzrq1 if c else '',
                    f'AR{row}': c.dhrq1 if c else '',
                    f'AS{row}': c.htsh if c else '',
                    f'AT{row}': record.fkxq.strip() if record.fkxq else '',
                    f'AU{row}': record.szxq.strip() if record.szxq else '',
                    f'AV{row}': float(record.zkfy1) if record.zkfy1 else '',
                    f'AW{row}': record.dq.strip() if record.dq else ''
                }
                
                for cell, value in data_mapping.items():
                    ws_out[cell].value = value
        
        # 保存文件
        output_filename = f"批量到票信息导入{user.username}{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        output_path = os.path.join(config.tmp_path, output_filename)
        wb_out.save(output_path)
    
        msg = '发票回导成功'
        if len(tmpstr1) > 0:
            msg = '存在未导入发票清单,正在生成txt文件...'
            out_name = f"发票回导失败记录_{time.strftime('%Y-%m-%d_%H-%M-%S')}.txt"
            out_path = os.path.join(config.tmp_path, out_name)
            val = '\n'.join(tmpstr1)
            write_file(out_path, val, 'w')
            output_filename = f"发票回导结果_{user.username}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.zip"
            zipfile = zipfile.ZipFile(os.path.join(config.tmp_path, output_filename), 'w')
            zipfile.write(output_path, os.path.basename(output_path))
            zipfile.write(out_path, os.path.basename(out_path))
            zipfile.close()

        s.commit()
        return {'code': 0, 'msg': msg, 'data': output_filename}
    except Exception as e:
        logger.error(trace_error())
        return {'code': -1, 'msg': f'处理失败: {str(e)}', 'data': None}


"""
表头生成
对应原Pascal: 表头生成
"""
@any_route('/api/saier/billed_notice/excel_title_export', methods=['POST'])
@require_token
async def view_billed_notice_excel_title_export(request):
    # 1. 检查用户是否是财务岗位 (对应 tmpcom2 查询)
    s = Session()
    try:
        j = await request.json()
        module = j.get('module', '开票通知')
        rids = j.get('rids', [])
        # postion = j.get('position', '')
        # user = request.current_user
        # org = get_user_path('zjnblh')
        # bm1 = org.get('position')[:1]
        fp_list = []
        for rid in rids:
            d = s.query(kaiptz.fphm).filter(kaiptz.rid == rid).first()
            if not d:
                continue
            if d.fphm != None and d.fphm.strip() != '' and not d.fphm in fp_list:
                fp_list.append(d.fphm)

        res = excel_title_report_files(fp_list, user, s)

        return json_result(res.get('code', -1), res.get('msg', ''), res.get('data', ''))
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()



def sanitize_filename(filename):
    """清理文件名中的非法字符"""
    # Windows文件名非法字符
    illegal_chars = r'[\\/:*?"<>|\r\n]'
    sanitized = re.sub(illegal_chars, '', filename)
    return sanitized

def excel_title_report_files(bgpmstr1, user, s):
    """
    处理报表文件生成
    
    Args:
        bgpmstr1: 发票号列表
        filename3l: Excel模板文件路径
        filename2: 输出文件路径前缀
        s: SQLAlchemy Session对象
        download: 是否返回ZIP文件供下载
    
    Returns:
        如果download=True，返回ZIP文件的字节数据；否则返回生成的文件列表
    """
    generated_files = []
    try:
        for lz, fphm_value in enumerate(bgpmstr1):
            k = 0
            fpje = 0.0
            bhsje = 0.0
            se = 0.0
            tse = 0.0
            BGJEH = 0.0
            CNF1 = 0.0
            CNF = 0.0
            hl = 0.0
            zmyhl = 1.0
            hl1 = 0.0
            father = ''
            bgje = 0.0
            bggs = ''
            BGRQ = ''
            dkfp = 0.0
            
            # 查询bgmxd表
            bgmxd_record = s.query(bgmxd).filter(
                (bgmxd.ysfp == fphm_value) | (bgmxd.fphm == fphm_value)
            ).first()
            
            if bgmxd_record:
                father = bgmxd_record.rid
                bgje = getnum(bgmxd_record.bgbgzje)
                bggs = bgmxd_record.bggs or ''
                CNF = getnum(bgmxd_record.CNFyf)
                zmyhl = getnum(bgmxd_record.zmyhl)
                BGRQ = bgmxd_record.zgrqj or ''
                
                # 确定汇率
                if (bgmxd_record.RMBkh == '是') or (bgmxd_record.hbdm == 'RMB'):
                    hl = 1.0
                else:
                    if bgmxd_record.hbdm not in ['USD$', 'USD']:
                        kpnrsheet_record = s.query(kpnrsheet).filter(
                            kpnrsheet.wfgs == bggs,
                            kpnrsheet.hbdm == bgmxd_record.hbdm
                        ).first()
                        if kpnrsheet_record:
                            hl = getnum(kpnrsheet_record.hl)
                            hl1 = getnum(kpnrsheet_record.hl)
            

            tmp_path = os.path.join(config.data_upload_path, 'template', '开票表头.xlsx')
            # 加载Excel模板
            wb = load_workbook(tmp_path)
            ws = wb.active
            
            # 设置基础信息
            ws['B3'] = bggs
            ws['M2'] = CNF
            
            # 查询汇率
            kpnr_record = s.query(kpnr).filter(kpnr.wfgs == bggs).first()
            if kpnr_record:
                if hl1 == 0:
                    hl1 = getnum(kpnr_record.tjhl)
                if hl != 1:
                    hl = hl1
                    ws['A2'] = '$'
                    ws['B2'].number_format = '0.00'
                    ws['B2'] = bgje - CNF
                    ws['D2'] = hl
                    ws['F2'].number_format = '0.00'
                    ws['F2'] = (bgje - CNF) * hl
                else:
                    ws['B1'] = '报人民币'
                    ws['A2'] = '￥'
                    ws['B2'].number_format = '0.00'
                    ws['B2'] = bgje
                    ws['D2'] = hl
                    ws['F2'].number_format = '0.00'
                    ws['F2'] = (bgje - CNF) * hl
            
            # 查询kaiptzsheet1表
            sheet_records = s.query(kaiptzsheet1).filter(
                kaiptzsheet1.fphm == fphm_value
            ).all()
            
            if sheet_records:
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for record in sheet_records:
                    k += 1
                    row_num = 3 + k
                    
                    # 设置边框
                    for col in range(1, 14):  # A-M列
                        cell = ws.cell(row=row_num, column=col)
                        cell.border = thin_border
                    
                    # 设置行高自动调整
                    ws.row_dimensions[row_num].height = None
                    
                    # 填充数据
                    ws.cell(row=row_num, column=1, value=k)  # A列
                    
                    # B列：发票号
                    if record.yfph:
                        ws.cell(row=row_num, column=2, value=f"{record.ysfp}\\{record.ysfp1}" if record.ysfp1 else record.ysfp)
                    else:
                        ws.cell(row=row_num, column=2, value=record.ysfp or '')
                    
                    ws.cell(row=row_num, column=3, value=record.ywrya or '')  # C列
                    ws.cell(row=row_num, column=4, value=record.sccj or '')  # D列
                    ws.cell(row=row_num, column=5, value=record.zwpm or '')  # E列
                    
                    # F列：数量+单位
                    sl_value = f"{getnum(record.bgsl)}{record.hgjldw or ''}" if record.bgsl else ''
                    ws.cell(row=row_num, column=6, value=sl_value)
                    
                    # G列：金额
                    sjkpze = getnum(record.sjkpze)
                    ws.cell(row=row_num, column=7, value=sjkpze)
                    ws.cell(row=row_num, column=7).number_format = '0.00'
                    
                    # 检查是否非正
                    kaiptz_record = s.query(kaiptz).filter(
                        kaiptz.rid == record.pid,
                        kaiptz.cpmc == "非正"
                    ).first()
                    if kaiptz_record:
                        dkfp += sjkpze
                    
                    # 计算比例
                    bl = 0.0
                    sum_record = s.query(func.sum(kaiptzsheet1.sjkpze)).filter(
                        kaiptzsheet1.fphm == fphm_value,
                        kaiptzsheet1.zwpm1 == record.zwpm1,
                        kaiptzsheet1.zsl == record.zsl,
                        kaiptzsheet1.hyd == record.hyd,
                        kaiptzsheet1.bgbh == record.bgbh
                    ).scalar()
                    
                    if sum_record and sum_record > 0:
                        bl = float(sjkpze) / float(sum_record)
                    
                    fpje += sjkpze
                    zsl_value = getnum(record.zsl)
                    
                    # 根据税率计算
                    if zsl_value <= 4:
                        if zsl_value == 3:
                            bhsj_value = float(sjkpze) / 1.03
                            bhsje += bhsj_value
                            se_value = float(sjkpze) - bhsj_value
                            se += se_value
                            tse_value = bhsj_value * 0.03
                            tse += tse_value
                            
                            ws.cell(row=row_num, column=8, value=bhsj_value)  # H列
                            ws.cell(row=row_num, column=9, value=se_value)    # I列
                            ws.cell(row=row_num, column=11, value=tse_value)  # K列
                            
                        elif zsl_value == 1:
                            bhsj_value = float(sjkpze) / 1.01
                            bhsje += bhsj_value
                            se_value = float(sjkpze) - bhsj_value
                            se += se_value
                            tse_value = bhsj_value * 0.01
                            tse += tse_value
                            
                            ws.cell(row=row_num, column=8, value=bhsj_value)
                            ws.cell(row=row_num, column=9, value=se_value)
                            ws.cell(row=row_num, column=11, value=tse_value)
                            
                    else:
                        if zsl_value == 17:
                            bhsj_value = float(sjkpze) / 1.17
                            bhsje += bhsj_value
                            se_value = float(sjkpze) - bhsj_value
                            se += se_value
                            tse_value = float(bhsj_value) * 0.17
                            tse += tse_value
                            
                            ws.cell(row=row_num, column=8, value=bhsj_value)
                            ws.cell(row=row_num, column=9, value=se_value)
                            ws.cell(row=row_num, column=11, value=tse_value)
                        else:
                            rate = zsl_value / 100
                            bhsj_value = float(sjkpze) / (1 + rate)
                            bhsje += bhsj_value
                            se_value = float(sjkpze) - bhsj_value
                            se += se_value
                            tse_value = float(bhsj_value) * rate
                            tse += tse_value
                            
                            ws.cell(row=row_num, column=8, value=bhsj_value)
                            ws.cell(row=row_num, column=9, value=se_value)
                            ws.cell(row=row_num, column=11, value=tse_value)
                    
                    # J列：税率
                    ws.cell(row=row_num, column=10, value=zsl_value / 100)
                    ws.cell(row=row_num, column=10).number_format = '0.00'
                    
                    # L列：利润
                    if father:
                        # 尝试带条件的查询
                        bgmxdhbcp_record = s.query(bgmxdhbcp).filter(
                            bgmxdhbcp.pid == father,
                            bgmxdhbcp.zwpm == record.zwpm1,
                            bgmxdhbcp.tsl == record.zsl,
                            bgmxdhbcp.bgbh == record.bgbh,
                            bgmxdhbcp.hyd.like(f'%{record.hyd}%')
                        ).first()
                        
                        if bgmxdhbcp_record and record.hyd:
                            profit = (getnum(bgmxdhbcp_record.wxzj) - getnum(bgmxdhbcp_record.CNFyf)) * bl
                            ws.cell(row=row_num, column=12, value=profit)
                            ws.cell(row=row_num, column=12).number_format = '0.00'
                        else:
                            # 无条件查询
                            bgmxdhbcp_record2 = s.query(bgmxdhbcp).filter(
                                bgmxdhbcp.pid == father,
                                bgmxdhbcp.zwpm == record.zwpm1
                            ).first()
                            
                            if bgmxdhbcp_record2:
                                profit = (getnum(bgmxdhbcp_record2.wxzj) - getnum(bgmxdhbcp_record2.CNFyf)) * bl
                                ws.cell(row=row_num, column=12, value=profit)
                                ws.cell(row=row_num, column=12).number_format = '0.00'
                            else:
                                ws.cell(row=row_num, column=12, value=0)
                                ws.cell(row=row_num, column=12).number_format = '0.00'
                    
                    # M列：公式
                    if hl != 1:
                        ws.cell(row=row_num, column=13, value=f'=(G{row_num}-K{row_num})/L{row_num}')
                    else:
                        ws.cell(row=row_num, column=13, value=f'=(G{row_num}-K{row_num})/(L{row_num}/{hl1})')
                    ws.cell(row=row_num, column=13).number_format = '0.00'
                
                # 添加合计行
                k += 1
                total_row = 3 + k
                
                ws.cell(row=total_row, column=7, value=fpje)
                ws.cell(row=total_row, column=7).number_format = '0.00'
                
                ws.cell(row=total_row, column=8, value=bhsje)
                ws.cell(row=total_row, column=8).number_format = '0.00'
                
                ws.cell(row=total_row, column=9, value=se)
                ws.cell(row=total_row, column=9).number_format = '0.00'
                
                ws.cell(row=total_row, column=11, value=tse)
                ws.cell(row=total_row, column=11).number_format = '0.00'
                
                ws.cell(row=total_row, column=5, value=fpje / bgje if bgje != 0 else 0)
                ws.cell(row=total_row, column=5).number_format = '0.00'
                
                # 设置底部汇总
                ws['H2'].number_format = '0.00'
                ws['H2'] = bhsje + se - tse
                
                ws['K2'].number_format = '0.00'
                ws['K2'] = f'=(F2-H2)'
                
                ws['I2'].number_format = '0.00'
                ws['I2'] = f'=(K2/F2)'
                
                if hl != 1:
                    ws['I1'].number_format = '0.00'
                    ws['I1'] = (fpje - tse) / (bgje - CNF) if (bgje - CNF) != 0 else 0
                else:
                    ws['I1'].number_format = '0.00'
                    ws['I1'] = (fpje - tse) / ((bgje - CNF) / hl1) if (bgje - CNF) != 0 and hl1 != 0 else 0
            
            # 设置其他信息
            ws['L1'] = BGRQ
            ws['C1'] = f'D总:{dkfp:.2f}'
            
            # 清理文件名
            wxfp1z = sanitize_filename(fphm_value)
            
            # 保存文件
            output_filename = f"开票表头{wxfp1z}{time.strftime('%Y-%m-%d_%H-%M-%S')}.xls"
            output_path = os.path.join(config.tmp_path, output_filename)
            wb.save(output_path)
            generated_files.append(output_filename)
            wb.close()

        zip_filename = None
        if len(generated_files) > 0:
            zip_filename = f"表头生成_{time.strftime('%Y-%m-%d_%H-%M-%S')}.zip"
            zip_filepath = os.path.join(config.tmp_path, zip_filename)
            zipFile = zipfile.ZipFile(zip_filepath, 'w')
            for f in generated_files:
                fn = os.path.join(config.tmp_path, f)
                if os.path.exists(fn):
                    zipFile.write(fn, f, zipfile.ZIP_DEFLATED)
            zipFile.close()

        return {'code': 0, 'msg': '生成成功', 'data': zip_filename}
    except Exception as e:
        logger.error(trace_error())
        return {'code': -1, 'msg': f'生成失败: {str(e)}', 'data': None}


"""
未齐发票
对应原Pascal: 未齐发票
"""
@any_route('/api/saier/billed_notice/incomplete', methods=['POST'])
@require_token
async def view_billed_notice_incomplete(request):
    # 1. 检查用户是否是财务岗位 (对应 tmpcom2 查询)
    s = Session()
    try:
        j = await request.json()
        user = request.current_user
        module = j.get('module', '开票通知')
        rids = j.get('rids', [])
        # org = get_user_path(user.username)
        # postion = j.get('position', '')
        # if not '财务' in postion:
        #     return json_result(-1, '只有财务岗位可以查看未齐发票')

        res = generate_incomplete_invoice_report(rids, user, s)

        return json_result(res.get('code', -1), res.get('msg', ''), res.get('data', ''))
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()

def generate_incomplete_invoice_report(rids, user, s):
    """
    生成未齐发票报告
    
    Args:
        rids: 记录ID列表
        s: SQLAlchemy Session对象
    
    Returns:
        生成的Excel文件路径
    """
    
    # 创建新的Excel工作簿
    wb = Workbook()
    ws = wb.active
    try:
        # 设置列宽
        column_widths = {
            'A': 14.3,  # 业务部门
            'B': 20.3,  # 外销发票号
            'C': 20.3,  # 原发票号
            'D': 14.3,  # 开票资料日期
            'E': 14.3,  # 业务人员
            'F': 20.3,  # 生产厂家
            'G': 14.3,  # 退税率
            'H': 18.3,  # 中文品名
            'I': 14.3,  # 报关数量
            'J': 14.3,  # 海关计量单位
            'K': 14.3,  # 实际开票总额
            'L': 14.3,  # 到票金额总
            'M': 20.3   # 我司抬头
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 设置表头样式
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_font = Font(bold=True)
        
        # 设置表头
        headers = {
            'A1': '业务部门',
            'B1': '外销发票号',
            'C1': '原发票号',
            'D1': '开票资料日期',
            'E1': '业务人员',
            'F1': '生产厂家',
            'G1': '退税率',
            'H1': '中文品名',
            'I1': '报关数量',
            'J1': '海关计量单位',
            'K1': '实际开票总额',
            'L1': '到票金额总',
            'M1': '我司抬头'
        }
        
        for cell, value in headers.items():
            ws[cell] = value
            ws[cell].alignment = header_alignment
            ws[cell].font = header_font
        
        # 记录当前行号（从第2行开始，因为第1行是表头）
        current_row = 1  # 将在循环中递增
        # 处理 lstallnumbers（调用 MultiReport 的逻辑）
        # 注意：原Delphi代码中的 MultiReport 函数未提供实现
        # 这里假设 MultiReport 会向 rids 添加数据或做其他处理
        # 根据原代码逻辑，需要实现 MultiReport 的功能
        # 由于原代码中没有 MultiReport 的具体实现，这里提供一个占位
        # 实际使用时需要根据业务逻辑实现

        uid_json = {}
        org = get_user_path(user.username)
        path = org.get('path', '')
        # 处理 rids 列表
        for a1, father_value in enumerate(rids):
            # 查询 kaiptzsheet1 表中未齐发票的记录 (wckp="否")
            sheet_records = s.query(kaiptzsheet1).filter(
                kaiptzsheet1.pid == father_value,
                kaiptzsheet1.wckp == '否'
            ).all()
            
            for record in sheet_records:
                current_row += 1
                
                # 查询业务部门信息
                kaiptz_record = s.query(kaiptz).filter(
                    kaiptz.rid == father_value
                ).first()
                if not kaiptz_record:
                    continue
                if kaiptz_record.uid not in uid_json:
                    uid_json[kaiptz_record.uid] = 0
                    o = get_user_path(kaiptz_record.uid)
                    p = o.get('path', '')
                    if p.startswith(path):
                        uid_json[kaiptz_record.uid] = 1
                if uid_json.get(kaiptz_record.uid, 0) == 0:
                    continue
                # 设置单元格样式
                cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # A列：业务部门
                cell_a = ws[f'A{current_row}']
                if kaiptz_record:
                    cell_a.value = kaiptz_record.ywbm or ''
                cell_a.alignment = cell_alignment
                
                # B列：外销发票号
                cell_b = ws[f'B{current_row}']
                cell_b.value = record.ysfp or ''
                cell_b.alignment = cell_alignment
                
                # C列：原发票号
                cell_c = ws[f'C{current_row}']
                cell_c.value = record.ysfp1 or ''
                cell_c.alignment = cell_alignment
                
                # D列：开票资料日期
                cell_d = ws[f'D{current_row}']
                # 处理日期格式
                if hasattr(record, 'kprq') and record.kprq:
                    if isinstance(record.kprq, datetime):
                        cell_d.value = record.kprq.strftime('%Y-%m-%d')
                    else:
                        cell_d.value = str(record.kprq)
                else:
                    cell_d.value = ''
                cell_d.alignment = cell_alignment
                
                # E列：业务人员
                cell_e = ws[f'E{current_row}']
                cell_e.value = record.ywrya or ''
                cell_e.alignment = cell_alignment
                
                # F列：生产厂家
                cell_f = ws[f'F{current_row}']
                cell_f.value = record.sccj or ''
                cell_f.alignment = cell_alignment
                
                # G列：退税率
                cell_g = ws[f'G{current_row}']
                cell_g.value = record.zsl if hasattr(record, 'zsl') else ''
                cell_g.alignment = cell_alignment
                
                # H列：中文品名
                cell_h = ws[f'H{current_row}']
                cell_h.value = record.zwpm or ''
                cell_h.alignment = cell_alignment
                
                # I列：报关数量
                cell_i = ws[f'I{current_row}']
                cell_i.value = record.bgsl if hasattr(record, 'bgsl') else ''
                cell_i.alignment = cell_alignment
                
                # J列：海关计量单位
                cell_j = ws[f'J{current_row}']
                cell_j.value = record.hgjldw or ''
                cell_j.alignment = cell_alignment
                
                # K列：实际开票总额
                cell_k = ws[f'K{current_row}']
                cell_k.value = record.sjkpze if hasattr(record, 'sjkpze') else 0.0
                cell_k.alignment = cell_alignment
                # 设置数字格式
                cell_k.number_format = '0.00'
                
                # L列：到票金额总
                cell_l = ws[f'L{current_row}']
                cell_l.value = record.dpjez if hasattr(record, 'dpjez') else 0.0
                cell_l.alignment = cell_alignment
                cell_l.number_format = '0.00'
                
                # M列：我司抬头
                cell_m = ws[f'M{current_row}']
                cell_m.value = record.wstt if hasattr(record, 'wstt') else ''
                cell_m.alignment = cell_alignment
        
        # 保存文件
        current_date = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        output_filename = f"未齐发票{current_date}.xlsx"  # 使用.xlsx扩展名（openpyxl默认格式）
        output_path = os.path.join(config.tmp_path, output_filename)
        
        wb.save(output_path)
        wb.close()
        
        return {'code': 0, 'msg': '生成成功', 'data': output_filename}
    except Exception as e:
        logger.error(trace_error())
        return {'code': -1, 'msg': f'生成失败: {str(e)}', 'data': None}



"""
出库单批量
对应原Pascal: 出库单批量
"""
@any_route('/api/saier/billed_notice/out_stock_export', methods=['POST'])
@require_token
async def view_billed_notice_out_stock_export(request):
    # 1. 检查用户是否是财务岗位 (对应 tmpcom2 查询)
    s = Session()
    try:
        j = await request.json()
        user = request.current_user
        module = j.get('module', '开票通知')
        gd = j.get('gd', 12.75)
        rids = j.get('rids', [])
        org = get_user_path('zjnblh')
        postion = org.get('position', '')
        uv = '0'
        if postion[:1] == 'C' or postion[:1] == 'D':
            uv = '1'
        fp_list = []
        fphm = []
        for rid in rids:
            d = s.query(kaiptz.fphm, kaiptz.khmc, kaiptz.cxnf).filter(kaiptz.rid == rid).first()
            if not d:
                continue
            if d.fphm != None and d.fphm.strip() != '' and not d.fphm in fphm:
                fphm.append(d.fphm)
                fp_list.append({'fphm': d.fphm, 'khmc': d.khmc, 'cxnf': d.cxnf})
        res = process_finance_delivery_notes_orm(fp_list, gd, uv,user, s)

        return json_result(res.get('code', -1), res.get('msg', ''), res.get('data', ''))
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()


def process_finance_delivery_notes_orm(
    fp_list: List[dict],
    uv: str,
    gd: float,
    user,
    s  # SQLAlchemy Session
):
    """使用ORM方式处理"""
    ts1 = 0
    ts = 0
    try:
        for r in fp_list:
            # 处理发票号码，过滤非法字符
            cx1z = re.sub(r'[\\/*?:"<>|\n\r]', '', r['fphm'])
            khmc1 = r['khmc'] if r['khmc'] else ''
            cxnf1 = r['cxnf'] if r['cxnf'] else ''
            
            # ORM查询获取客户名称和出库年份
            if not khmc1 or not cxnf1:
                kaiptz = s.query(kaiptz).filter(
                    kaiptz.fphm == r['fphm'],
                    kaiptz.khmc != '',
                    kaiptz.cxnf > '2011-01-01'
                ).first()
                
                if kaiptz:
                    if not khmc1:
                        khmc1 = kaiptz.khmc
                    if not cxnf1:
                        cxnf1 = kaiptz.cxnf
            
            if not khmc1:
                b = s.query(bgmxd).filter(bgmxd.fphm == r['fphm']).first()
                if b:
                    khmc1 = b.khmc
            
            if not cxnf1 or cxnf1 == '' or cxnf1 == None:
                cxnf1 = datetime.now().strftime('%Y-%m-%d')
            
            # ORM查询物料信息
            records = s.query(kaiptzsheet1).filter(
                kaiptzsheet1.fphm == r['fphm']
            ).all()
            files = []
            if records:
                filename3 = os.path.join(config.data_upload_path, 'template', '财务出入库.xlsx')
                wb = load_workbook(filename3)
                ws = wb.active
                ws.sheet_view.showGridLines = False
                
                bold_font = Font(bold=True)
                center_alignment = Alignment(horizontal='center', vertical='center')
                
                for record in records:
                    # 查询海关记录单位
                    hgjldw_record = s.query(kaiptzsheet1).filter(
                        kaiptzsheet1.fphm == r['fphm'],
                        kaiptzsheet1.zwpm == record.zwpm,
                        kaiptzsheet1.hgjldw != ''
                    ).first()
                    hgjldw = hgjldw_record.hgjldw if hgjldw_record else ''
                    
                    ts += 1
                    if ts == 8:
                        ts = 1
                        ts1 += 1
                    
                    if ts == 1:
                        start_row = 17 * ts1 + 1
                        
                        # 设置表头行
                        ws.merge_cells(f'A{start_row}:C{start_row}')
                        ws.row_dimensions[start_row].height = 26
                        cxnf_date = datetime.strptime(str(cxnf1)[:10], '%Y-%m-%d')
                        ws[f'A{start_row}'].value = (cxnf_date - timedelta(days=3)).strftime('%Y-%m-%d')
                        
                        ws.merge_cells(f'G{start_row}:H{start_row}')
                        ws[f'G{start_row}'].value = khmc1
                        
                        # 第2行
                        ws.row_dimensions[start_row + 1].height = 34
                        
                        # 第3行（内容行）
                        ws.merge_cells(f'A{start_row + 2}:C{start_row + 2}')
                        ws.row_dimensions[start_row + 2].height = 26
                        ws[f'A{start_row + 2}'].value = record.zwpm
                        ws[f'E{start_row + 2}'].value = hgjldw
                        ws[f'F{start_row + 2}'].value = float(record.bgsl) if record.bgsl else 0
                        
                        h3_cell = ws[f'H{start_row + 2}']
                        h3_cell.alignment = center_alignment
                        h3_cell.value = r['fphm']
                        
                        # 设置空行的行高
                        for offset in range(3, 12):
                            row_num = start_row + offset
                            ws.merge_cells(f'A{row_num}:C{row_num}')
                            if offset == 9:
                                ws.row_dimensions[row_num].height = 21.75
                            elif offset == 10:
                                ws.row_dimensions[row_num].height = 16.5
                            elif offset == 11:
                                ws.row_dimensions[row_num].height = 24.75
                            else:
                                ws.row_dimensions[row_num].height = 26
                        
                        # 签名行
                        row_12 = start_row + 11
                        sign_name = '闫绍贵' if uv == '1' else '唐红岗'
                        ws[f'F{row_12}'].font = bold_font
                        ws[f'F{row_12}'].value = sign_name
                        ws[f'H{row_12}'].font = bold_font
                        ws[f'H{row_12}'].value = sign_name
                        
                        # 底部空行
                        for offset in range(12, 17):
                            row_num = start_row + offset
                            ws.merge_cells(f'A{row_num}:C{row_num}')
                            if offset == 15:
                                ws.row_dimensions[row_num].height = 12.75
                            elif offset == 16:
                                ws.row_dimensions[row_num].height = float(gd) if gd else 15
                            else:
                                ws.row_dimensions[row_num].height = 15
                    
                    else:
                        row_num = 17 * ts1 + ts + 2
                        ws[f'A{row_num}'].font = bold_font
                        ws[f'A{row_num}'].value = record.zwpm
                        ws[f'E{row_num}'].font = bold_font
                        ws[f'E{row_num}'].value = hgjldw
                        ws[f'F{row_num}'].font = bold_font
                        ws[f'F{row_num}'].value = float(record.bgsl) if record.bgsl else 0
                
                # 保存文件
                safe_filename = re.sub(r'[\\/*?:"<>|]', '', cx1z)
                output_filename = f"财务出库单{safe_filename}{user.username}{datetime.now().strftime('%Y-%m-%d %H-%M-%S')}.xlsx"
                wb.save(os.path.join(config.tmp_path, output_filename))
                files.append(output_filename)
                wb.close()

        output_name = ''
        if len(files) > 0:
            output_name = f"财务出库单_{user.username}{datetime.now().strftime('%Y-%m-%d %H-%M-%S')}.zip"
            zip_filepath = os.path.join(config.tmp_path, output_name)
            with zipfile.ZipFile(zip_filepath, 'w') as zip_file:
                for f in files:
                    file_path = os.path.join(config.tmp_path, f)
                    if os.path.exists(file_path):
                        zip_file.write(file_path, f, zipfile.ZIP_DEFLATED)
                zip_file.close()
        
        return {'code': 0, 'msg': '生成成功', 'data': output_name}
    except Exception as e:
        logger.error(trace_error())
        return {'code': -1, 'msg': f'处理失败: {str(e)}', 'data': None}
    

"""
入库单批量
对应原Pascal: 入库单批量
"""
@any_route('/api/saier/billed_notice/in_stock_export', methods=['POST'])
@require_token
async def view_billed_notice_in_stock_export(request):
    # 1. 检查用户是否是财务岗位 (对应 tmpcom2 查询)
    s = Session()
    try:
        j = await request.json()
        user = request.current_user
        module = j.get('module', '开票通知')
        gd = j.get('gd', 12.75)
        rids = j.get('rids', [])
        org = get_user_path('zjnblh')
        postion = org.get('position', '')
        uv = '0'
        if postion[:1] == 'C' or postion[:1] == 'D':
            uv = '1'
        fp_list = []
        fphm = []
        for rid in rids:
            d = s.query(kaiptz.fphm, kaiptz.khmc, kaiptz.cxnf).filter(kaiptz.rid == rid).first()
            if not d:
                continue
            if d.fphm != None and d.fphm.strip() != '' and not d.fphm in fphm:
                fphm.append(d.fphm)
                fp_list.append({'fphm': d.fphm, 'khmc': d.khmc, 'cxnf': d.cxnf})

        res = process_finance_receipt_notes_orm(fp_list, gd, uv, user, s)

        return json_result(res.get('code', -1), res.get('msg', ''), res.get('data', ''))
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()


def process_finance_receipt_notes_orm(
    fp_list: List[dict],
    uv: str,
    gd: float,
    user,
    s
):
    """使用ORM方式处理财务入库单"""
    try:
        files = []
        for r in fp_list:
            # 处理发票号码，过滤非法字符
            cx1z = re.sub(r'[\\/*?:"<>|\n\r]', '', r['fphm'])
            
            # 查询出库年份
            c = s.query(kaiptz).filter(
                kaiptz.fphm == r['fphm'],
                kaiptz.cxnf > '2011-01-01'
            ).first()
            
            cxnf1 = c.cxnf if c else datetime.now().strftime('%Y-%m-%d')
            
            # 查询所有开票序号
            kpxh_list = s.query(distinct(kaiptzsheet1.kpxh)).filter(
                kaiptzsheet1.fphm == r['fphm'],
                kaiptzsheet1.kpxh != ''
            ).all()
            kpxh_list = [k[0] for k in kpxh_list if k[0]]
            
            # 查询所有工厂名称
            gcmc_list = []
            for kpxh in kpxh_list:
                gcmc_records = s.query(distinct(gchk.gcmc)).filter(
                    gchk.kpxh == kpxh,
                    gchk.gcmc != ''
                ).all()
                for record in gcmc_records:
                    if record[0] and record[0] not in gcmc_list:
                        gcmc_list.append(record[0])
            
            # 创建Excel工作簿
            filename3 = os.path.join(config.data_upload_path, 'template', '财务出入库.xlsx')
            wb = load_workbook(filename3)
            ws = wb.active
            ws.sheet_view.showGridLines = False
            
            bold_font = Font(bold=True)
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            ts1 = 0
            row_data = []
            
            # 遍历工厂名称和开票序号收集数据
            for gcmc in gcmc_list:
                ts = 0
                for kpxh in kpxh_list:
                    # 验证 gchk 表中是否存在该组合
                    g = s.query(gchk).filter(
                        gchk.kpxh == kpxh,
                        gchk.gcmc == gcmc
                    ).first()
                    
                    if g:
                        # 查询物料信息
                        records = s.query(kaiptzsheet1).filter(
                            kaiptzsheet1.kpxh == kpxh
                        ).all()
                        
                        for record in records:
                            ts += 1
                            if ts == 8:
                                ts = 1
                                ts1 += 1
                            
                            row_data.append({
                                'ts': ts,
                                'ts1': ts1,
                                'gcmc': gcmc,
                                'zwpm': record.zwpm,
                                'hgjldw': record.hgjldw,
                                'chsl': record.chsl,
                                'is_first_in_batch': ts == 1
                            })
                ts1 += 1
            
            # 按批次和行顺序处理数据
            row_data.sort(key=lambda x: (x['ts1'], x['ts']))
            
            for data in row_data:
                ts = data['ts']
                ts1 = data['ts1']
                
                if data['is_first_in_batch']:
                    start_row = 17 * ts1 + 1
                    
                    # 设置表头行
                    ws.merge_cells(f'A{start_row}:C{start_row}')
                    ws.row_dimensions[start_row].height = 26
                    cxnf_date = datetime.strptime(cxnf1, '%Y-%m-%d')
                    ws[f'A{start_row}'].value = (cxnf_date - timedelta(days=5)).strftime('%Y-%m-%d')
                    
                    ws.merge_cells(f'G{start_row}:H{start_row}')
                    ws[f'G{start_row}'].value = data['gcmc']
                    
                    # 第2行
                    ws.row_dimensions[start_row + 1].height = 34
                    
                    # 第3行（内容行）
                    ws.merge_cells(f'A{start_row + 2}:C{start_row + 2}')
                    ws.row_dimensions[start_row + 2].height = 26
                    ws[f'A{start_row + 2}'].font = bold_font
                    ws[f'A{start_row + 2}'].value = data['zwpm']
                    ws[f'E{start_row + 2}'].font = bold_font
                    ws[f'E{start_row + 2}'].value = data['hgjldw']
                    ws[f'F{start_row + 2}'].font = bold_font
                    ws[f'F{start_row + 2}'].value = float(data['chsl']) if data['chsl'] else 0
                    
                    ws[f'H{start_row + 2}'].alignment = center_alignment
                    ws[f'H{start_row + 2}'].value = r['fphm']
                    
                    # 设置空行
                    for offset in range(3, 12):
                        row_num = start_row + offset
                        ws.merge_cells(f'A{row_num}:C{row_num}')
                        if offset == 9:
                            ws.row_dimensions[row_num].height = 21.75
                        elif offset == 10:
                            ws.row_dimensions[row_num].height = 16.5
                        elif offset == 11:
                            ws.row_dimensions[row_num].height = 24.75
                        else:
                            ws.row_dimensions[row_num].height = 26
                    
                    # 签名行
                    row_12 = start_row + 11
                    sign_name = '闫绍贵' if uv == '1' else '唐红岗'
                    ws[f'F{row_12}'].font = bold_font
                    ws[f'F{row_12}'].value = sign_name
                    ws[f'H{row_12}'].font = bold_font
                    ws[f'H{row_12}'].value = sign_name
                    
                    # 底部空行
                    for offset in range(12, 17):
                        row_num = start_row + offset
                        ws.merge_cells(f'A{row_num}:C{row_num}')
                        if offset == 15:
                            ws.row_dimensions[row_num].height = 12.75
                        elif offset == 16:
                            ws.row_dimensions[row_num].height = float(gd) if gd else 15
                        else:
                            ws.row_dimensions[row_num].height = 15
                
                else:
                    start_row = 17 * ts1 + 1
                    row_num = start_row + ts + 1
                    
                    ws[f'A{row_num}'].font = bold_font
                    ws[f'A{row_num}'].value = data['zwpm']
                    ws[f'E{row_num}'].font = bold_font
                    ws[f'E{row_num}'].value = data['hgjldw']
                    ws[f'F{row_num}'].font = bold_font
                    ws[f'F{row_num}'].value = float(data['chsl']) if data['chsl'] else 0
            
            # 保存文件
            safe_filename = re.sub(r'[\\/*?:"<>|]', '', cx1z)
            output_filename = f"财务入库单{safe_filename}{user.username}{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.xlsx"
            files.append(output_filename)
            wb.save(os.path.join(config.tmp_path, output_filename))
            wb.close()
        
        output_name = ''
        if len(files) > 0:
            output_name = f"财务出库单_{user.username}{datetime.now().strftime('%Y-%m-%d %H-%M-%S')}.zip"
            zip_filepath = os.path.join(config.tmp_path, output_name)
            with zipfile.ZipFile(zip_filepath, 'w') as zip_file:
                for f in files:
                    file_path = os.path.join(config.tmp_path, f)
                    if os.path.exists(file_path):
                        zip_file.write(file_path, f, zipfile.ZIP_DEFLATED)
                zip_file.close()

        return {'code': 0, 'msg': '生成成功', 'data': output_name}
    except Exception as e:
        logger.error(trace_error())
        return {'code': -1, 'msg': f'处理失败: {str(e)}', 'data': None}



"""
金蝶预估入库
对应原Pascal: 金蝶预估入库
"""
@any_route('/api/saier/billed_notice/kingdee_in_stock', methods=['POST'])
@require_token
async def view_billed_notice_kingdee_in_stock(request):
    # 1. 检查用户是否是财务岗位 (对应 tmpcom2 查询)
    s = Session()
    try:
        j = await request.json()
        user = request.current_user
        rids = j.get('rids', [])
        tjrq = j.get('tjrq', '')
        fp_list = []
        j2 = 0
        j3 = 0
        for rid in rids:
            j3 = j3 + 1
            d = s.query(kaiptz.ysfp).filter(kaiptz.rid == rid).first()
            if not d:
                continue
            if d.ysfp != None and d.ysfp.strip() != '' and not d.ysfp in fp_list:
                fp_list.append(d.ysfp)
                j2 = j2 + 1
        res = process_accounting_voucher(tjrq, fp_list, j2, j3, user, s)

        return json_result(res.get('code', -1), res.get('msg', ''), res.get('data', ''))
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()


def is_leap_year(year: int) -> bool:
    """判断是否为闰年"""
    return (year % 4 == 0 and year % 100 != 0) or (year % 400 == 0)

def get_month_end_date(year: int, month: int) -> str:
    """获取指定年月的最后一天"""
    last_day = calendar.monthrange(year, month)[1]
    return f"{last_day:02d}"

def format_sequence_number(seq: int, digits: int) -> str:
    """
    格式化序列号，根据位数补零
    digits: 2,3,4,5 表示总位数
    """
    if digits == 2:
        return f"{seq:02d}"
    elif digits == 3:
        return f"{seq:03d}"
    elif digits == 4:
        return f"{seq:04d}"
    elif digits == 5:
        return f"{seq:05d}"
    else:
        return str(seq)

def process_accounting_voucher(
    tjrq: str,             # 日期/期间字符串 (格式: YYYY-MM-DD 或 其他)
    fphm: List[str],     # 发票号码/原始凭证号列表
    j2: int,             # 用于计算 j4 的数值
    j3: int,             # 用于计算 j5 的数值
    user,
    s                    # SQLAlchemy Session
):
    """
    处理会计凭证生成
    对应原Delphi代码的凭证生成逻辑
    """
    
    if not tjrq:
        return
    
    # 解析 tjrq 字符串
    # 假设 tjrq 格式类似 "2024-01-15" 或其他格式，这里根据原代码逻辑处理
    n = tjrq[2:4]      # 年份后两位
    n1 = tjrq[:4]      # 年份前四位
    y = tjrq[5:7]      # 月份
    
    year_int = int(n1)
    month_int = int(y)
    
    # 确定月份和月底日期
    month_map = {
        '01': ('1', '01-31'),
        '02': ('2', f'02-{get_month_end_date(year_int, 2)}'),
        '03': ('3', '03-31'),
        '04': ('4', '04-30'),
        '05': ('5', '05-31'),
        '06': ('6', '06-30'),
        '07': ('7', '07-31'),
        '08': ('8', '08-31'),
        '09': ('9', '09-30'),
        '10': ('10', '10-31'),
        '11': ('11', '11-30'),
        '12': ('12', '12-31'),
    }
    
    y1, r1 = month_map.get(y, ('1', '01-31'))
    
    # 特殊处理闰年2月
    if y == '02' and is_leap_year(year_int):
        r1 = '02-29'
    
    # 计算 j4 和 j5
    if j2 < 99:
        j4 = 2
    elif j2 < 999:
        j4 = 3
    else:
        j4 = 4
    
    if j3 < 99:
        j5 = 3
    elif j3 < 999:
        j5 = 4
    else:
        j5 = 5
    
    # 创建Excel工作簿
    filename = os.path.join(config.data_upload_path, 'template', '预估入库.xlsx')
    wb = load_workbook(filename)
    ws = wb.active
    try:
        i4 = 0
        i5 = 0
        i = 0  # 全局行计数器
        i1 = 0  # 发票计数器
        
        for fphm_value in fphm:
            i1 += 1
            # 生成凭证编号 zb
            zb = format_sequence_number(i1, j4)
            zb = n + y + zb
            
            # 查询 wstt 和账簿信息
            wstt = ''
            zpbm = ''
            zpbm1 = ''
            
            z = s.query(kaiptz).filter(
                kaiptz.ysfp == fphm_value,
                kaiptz.wstt != '',
                kaiptz.wstt.isnot(None)
            ).first()
            
            if z:
                wstt = z.wstt
                # 查询账簿编码
                x = s.query(zx).filter(
                    zx.ly == '金蝶账簿编码',
                    zx.mc == wstt
                ).first()
                if x:
                    zpbm = x.wb1 or ''
                    zpbm1 = x.wb2 or ''
            
            # 查询 kaiptzsheet1 明细
            details = s.query(kaiptzsheet1).filter(
                kaiptzsheet1.ysfp == fphm_value
            ).all()
            
            bgzj = 0.0  # 暂估应付款累计
            
            for idx, detail in enumerate(details):
                i2 = idx + 1
                i += 1
                # 生成分录编号 fb
                fb = format_sequence_number(i, j5)
                fb = n + y + fb
                
                # 计算不含税金额
                sjkpze = float(detail.sjkpze or 0)
                zzsl = float(detail.zzsl or 0)
                bgsl = float(detail.bgsl or 0)
                
                tax_rate = 1 + zzsl / 100
                amount_ex_tax = round(sjkpze / tax_rate * 100) / 100
                unit_price = round(amount_ex_tax / bgsl * 100) / 100 if bgsl > 0 else 0
                
                bgzj += amount_ex_tax
                
                row_num = i + 2
                
                # 查询物料分类
                spfl = s.query(spbmflbsheet).filter(
                    spbmflbsheet.zwpm == detail.zwpm
                ).first()
                
                # 查询计量单位编码
                w = s.query(jldw).filter(
                    jldw.zwmc == detail.hgjldw
                ).first()
                
                if i2 == 1:
                    # 第一行：资产分录
                    ws[f'A{row_num}'] = '1' + zb
                    ws[f'B{row_num}'] = zpbm
                    ws[f'C{row_num}'] = f"{wstt}账簿"
                    ws[f'D{row_num}'] = f"{n1}-{r1}"
                    ws[f'E{row_num}'] = 'PRE001'
                    ws[f'F{row_num}'] = '记'
                    ws[f'H{row_num}'] = zpbm1
                    ws[f'I{row_num}'] = wstt
                    ws[f'J{row_num}'] = 'False'
                    ws[f'K{row_num}'] = n1
                    ws[f'L{row_num}'] = y1
                    ws[f'N{row_num}'] = '1' + fb
                    ws[f'O{row_num}'] = f"{fphm_value}预估入库"
                    ws[f'P{row_num}'] = '1405.03'
                    ws[f'Q{row_num}'] = '外销商品'
                    
                    if spfl:
                        ws[f'AD{row_num}'] = spfl.spflbm
                        ws[f'AE{row_num}'] = spfl.spflmc
                    
                    ws[f'BD{row_num}'] = 'PRE001'
                    ws[f'BE{row_num}'] = '人民币'
                    ws[f'BF{row_num}'] = 'HLTX01_SYS'
                    ws[f'BG{row_num}'] = '固定汇率'
                    ws[f'BH{row_num}'] = '1'
                    
                    if w:
                        ws[f'BI{row_num}'] = w.dwbm
                    
                    ws[f'BJ{row_num}'] = detail.hgjldw
                    ws[f'BK{row_num}'] = unit_price
                    ws[f'BL{row_num}'] = bgsl if bgsl > 0 else 0
                    ws[f'BM{row_num}'] = amount_ex_tax
                    ws[f'BN{row_num}'] = amount_ex_tax
                
                else:
                    # 后续行：资产分录（不含A/B/C等字段）
                    ws[f'N{row_num}'] = '1' + fb
                    ws[f'O{row_num}'] = f"{fphm_value}预估入库"
                    ws[f'P{row_num}'] = '1405.03'
                    ws[f'Q{row_num}'] = '外销商品'
                    
                    if spfl:
                        ws[f'AD{row_num}'] = spfl.spflbm
                        ws[f'AE{row_num}'] = spfl.spflmc
                    
                    ws[f'BD{row_num}'] = 'PRE001'
                    ws[f'BE{row_num}'] = '人民币'
                    ws[f'BF{row_num}'] = 'HLTX01_SYS'
                    ws[f'BG{row_num}'] = '固定汇率'
                    ws[f'BH{row_num}'] = '1'
                    
                    if w:
                        ws[f'BI{row_num}'] = w.dwbm
                    
                    ws[f'BJ{row_num}'] = detail.hgjldw
                    ws[f'BK{row_num}'] = unit_price
                    ws[f'BL{row_num}'] = bgsl if bgsl > 0 else 0
                    ws[f'BM{row_num}'] = amount_ex_tax
                    ws[f'BN{row_num}'] = amount_ex_tax
            
            # 添加负债分录（暂估应付款）
            i += 1
            fb = format_sequence_number(i, j5)
            fb = n + y + fb
            row_num = i + 2
            
            ws[f'N{row_num}'] = '1' + fb
            ws[f'O{row_num}'] = f"{fphm_value}预估入库"
            ws[f'P{row_num}'] = '2202.02'
            ws[f'Q{row_num}'] = '暂估应付款'
            ws[f'BD{row_num}'] = 'PRE001'
            ws[f'BE{row_num}'] = '人民币'
            ws[f'BF{row_num}'] = 'HLTX01_SYS'
            ws[f'BG{row_num}'] = '固定汇率'
            ws[f'BH{row_num}'] = '1'
            ws[f'BM{row_num}'] = bgzj
            ws[f'BO{row_num}'] = bgzj
        
        # 保存文件
        output_filename = f"会计凭证_{user.username}{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(os.path.join(config.tmp_path, output_filename))
        wb.close()
        
        return {'code': 0, 'msg': '成功', 'data': output_filename}
    except Exception as e:
        logger.error(trace_error())
        return {'code': -1, 'msg': f'处理失败: {str(e)}', 'data': None}

"""
金蝶合同销售
对应原Pascal: 金蝶合同销售
"""
@any_route('/api/saier/billed_notice/kingdee_contract', methods=['POST'])
@require_token
async def view_billed_notice_kingdee_contract(request):
    # 1. 检查用户是否是财务岗位 (对应 tmpcom2 查询)
    s = Session()
    try:
        j = await request.json()
        user = request.current_user
        rids = j.get('rids', [])
        position = j.get('position', '')
        nbyw = '宁波'
        if '义乌' in position:
            nbyw = '义乌'
        org = get_user_path('zjnblh')
        bm1 = org.get('position', '')[:1]
        tjrq = j.get('tjrq', '')
        fp_list = []
        j2 = 0
        j3 = 0
        for rid in rids:
            j3 = j3 + 1
            d = s.query(kaiptz.ysfp, kaiptz.khmc, kaiptz.cxnf).filter(kaiptz.rid == rid).first()
            if not d:
                continue
            if d.ysfp != None and d.ysfp.strip() != '' and not d.ysfp in fp_list:
                fp_list.append(d.ysfp)
                j2 = j2 + 1

        res = process_sales_voucher(tjrq, fp_list, j2, j3, nbyw, bm1, user, s)

        return json_result(res.get('code', -1), res.get('msg', ''), res.get('data', ''))
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()


def process_sales_voucher(
    tjrq: str,             # 日期/期间字符串 (格式: YYYY-MM-DD)
    fphm: List[str],     # 发票号码/原始凭证号列表
    j2: int,             # 用于计算 j4 的数值
    j3: int,             # 用于计算 j5 的数值
    nbyw: str,           # 业务标识
    bm1: str,            # 部门标识
    user,
    s                    # SQLAlchemy Session
):
    """
    处理销售凭证生成（合同销售）
    对应原Delphi代码的凭证生成逻辑
    """
    
    if not tjrq:
        return
    
    # 解析 tjrq 字符串，格式如 "2024-03-15"
    n = tjrq[2:4]      # 年份后两位
    n1 = tjrq[:4]      # 年份前四位
    y = tjrq[5:7]      # 月份
    
    year_int = int(n1)
    month_int = int(y)
    
    # 确定月份和月底日期
    month_map = {
        '01': ('1', '01-31'),
        '02': ('2', f'02-{get_month_end_date(year_int, 2)}'),
        '03': ('3', '03-31'),
        '04': ('4', '04-30'),
        '05': ('5', '05-31'),
        '06': ('6', '06-30'),
        '07': ('7', '07-31'),
        '08': ('8', '08-31'),
        '09': ('9', '09-30'),
        '10': ('10', '10-31'),
        '11': ('11', '11-30'),
        '12': ('12', '12-31'),
    }
    
    y1, r1 = month_map.get(y, ('1', '01-31'))
    
    # 特殊处理闰年2月
    if y == '02' and is_leap_year(year_int):
        r1 = '02-29'
    
    # 计算 j4 和 j5
    if j2 < 99:
        j4 = 2
    elif j2 < 999:
        j4 = 3
    else:
        j4 = 4
    
    if j3 < 99:
        j5 = 3
    elif j3 < 999:
        j5 = 4
    else:
        j5 = 5
    
    # 查询汇率
    hl = 0.0
    start_date = f"{n1}-{y}-01"
    end_date = f"{n1}-{r1}"
    
    # 从 xshl 表查询汇率
    xshl_record = s.execute(
        "SELECT hl FROM xshl WHERE hsrq >= :start_date AND hsrq <= :end_date AND ly = :ly",
        {'start_date': start_date, 'end_date': end_date, 'ly': f"{nbyw}金蝶"}
    ).first()
    
    if xshl_record:
        hl = float(xshl_record.hl)
    
    # 从 cwhshl 表查询汇率
    if hl == 0:
        cwhshl_record = s.execute(
            "SELECT hl FROM cwhshl WHERE hsny = :hsny",
            {'hsny': f"{n1}-{y}"}
        ).first()
        if cwhshl_record:
            hl = float(cwhshl_record.hl)
    
    if hl == 0:
        hl = 1.0
    
    # 创建Excel工作簿
    filename = os.path.join(config.data_upload_path, 'template', '合同销售.xlsx')
    wb = load_workbook(filename)
    ws = wb.active
    try:
        i = 0  # 全局行计数器
        i1 = 0  # 发票计数器
        
        for fphm_value in fphm:
            i1 += 1
            i += 1
            
            # 生成凭证编号 zb
            zb = format_sequence_number(i1, j4)
            zb = n + y + zb
            
            # 生成分录编号 fb
            fb = format_sequence_number(i, j5)
            fb = n + y + fb
            
            # 查询 wstt 和账簿信息
            wstt = ''
            zpbm = ''
            zpbm1 = ''
            
            d = s.query(kaiptz).filter(
                kaiptz.ysfp == fphm_value,
                kaiptz.wstt != '',
                kaiptz.wstt.isnot(None)
            ).first()
            
            if d:
                wstt = d.wstt
                # 查询账簿编码
                z = s.query(zx).filter(
                    zx.ly == '金蝶账簿编码',
                    zx.mc == wstt
                ).first()
                if z:
                    zpbm = z.wb1 or ''
                    zpbm1 = z.wb2 or ''
            
            row_num = i + 2
            
            # 写入基础分录
            ws[f'A{row_num}'] = '2' + zb
            ws[f'B{row_num}'] = zpbm
            ws[f'C{row_num}'] = f"{wstt}账簿"
            ws[f'D{row_num}'] = f"{n1}-{r1}"
            ws[f'E{row_num}'] = 'PRE001'
            ws[f'F{row_num}'] = '记'
            ws[f'H{row_num}'] = zpbm1
            ws[f'I{row_num}'] = wstt
            ws[f'J{row_num}'] = 'False'
            ws[f'K{row_num}'] = n1
            ws[f'L{row_num}'] = y1
            ws[f'N{row_num}'] = '2' + fb
            ws[f'O{row_num}'] = f"{fphm_value}合同销售"
            ws[f'P{row_num}'] = '1122.01'
            ws[f'Q{row_num}'] = '外销'
            
            # 查询 bgmxd 表获取金额和币种信息
            b = s.query(bgmxd).filter(bgmxd.ysfp == fphm_value).first()
            
            if b:
                bgbgzje = float(b.bgbgzje or 0)
                hbdm = b.hbdm or ''
                rmbkh = b.RMBkh or ''
                
                # 判断是否需要处理汇率
                if rmbkh != '是' or bm1 != 'D':
                    if hbdm in ('USD$', 'USD'):
                        # 美元处理
                        ws[f'BB{row_num}'] = 'PRE007'
                        ws[f'BC{row_num}'] = '美元'
                        ws[f'BD{row_num}'] = 'HLTX02_SYS'
                        ws[f'BF{row_num}'] = hl
                        ws[f'BL{row_num}'] = round(hl * bgbgzje * 100) / 100
                    else:
                        # 其他币种，从 cwhshlsheet 和 hbdm 查询
                        cwhshl = s.execute(
                            """SELECT cwhshlsheet.hl, cwhshlsheet.hbdm, 
                                    hb.jdPRE, hb.jdHLTX, hb.zwmc 
                            FROM cwhshlsheet 
                            LEFT JOIN hbdm AS hb ON hb.hbdm = cwhshlsheet.hbdm 
                            WHERE cwhshlsheet.hsny = :hsny AND cwhshlsheet.hbdm = :hbdm""",
                            {'hsny': f"{n1}-{y}", 'hbdm': hbdm}
                        ).first()
                        
                        if cwhshl:
                            hl1 = float(cwhshl.hl or 0)
                            ws[f'BB{row_num}'] = cwhshl.jdPRE or ''
                            ws[f'BC{row_num}'] = cwhshl.zwmc or ''
                            ws[f'BD{row_num}'] = cwhshl.jdHLTX or ''
                            ws[f'BF{row_num}'] = hl1
                            ws[f'BL{row_num}'] = round(hl1 * bgbgzje * 100) / 100
                    
                    ws[f'BE{row_num}'] = '即期汇率'
                    ws[f'BK{row_num}'] = bgbgzje
                    
                    # 添加收入分录
                    i += 1
                    fb = format_sequence_number(i, j5)
                    fb = n + y + fb
                    row_num2 = i + 2
                    
                    ws[f'N{row_num2}'] = '2' + fb
                    ws[f'O{row_num2}'] = f"{fphm_value}合同销售"
                    ws[f'P{row_num2}'] = '6001.01.01'
                    ws[f'Q{row_num2}'] = '外销'
                    ws[f'BB{row_num2}'] = 'PRE001'
                    ws[f'BC{row_num2}'] = '人民币'
                    ws[f'BD{row_num2}'] = 'HLTX01_SYS'
                    ws[f'BE{row_num2}'] = '固定汇率'
                    ws[f'BF{row_num2}'] = 1
                    
                    if hbdm in ('USD$', 'USD'):
                        ws[f'BK{row_num2}'] = round(hl * bgbgzje * 100) / 100
                        ws[f'BM{row_num2}'] = round(hl * bgbgzje * 100) / 100
                    else:
                        hl1 = hl1 if 'hl1' in dir() else 0
                        ws[f'BK{row_num2}'] = round(hl1 * bgbgzje * 100) / 100
                        ws[f'BM{row_num2}'] = round(hl1 * bgbgzje * 100) / 100
                
                else:
                    # 人民币直接处理
                    ws[f'BB{row_num}'] = 'PRE001'
                    ws[f'BC{row_num}'] = '人民币'
                    ws[f'BD{row_num}'] = 'HLTX01_SYS'
                    ws[f'BE{row_num}'] = '固定汇率'
                    ws[f'BF{row_num}'] = 1
                    ws[f'BK{row_num}'] = bgbgzje
                    ws[f'BL{row_num}'] = bgbgzje
                    
                    # 添加收入分录
                    i += 1
                    fb = format_sequence_number(i, j5)
                    fb = n + y + fb
                    row_num2 = i + 2
                    
                    ws[f'N{row_num2}'] = '2' + fb
                    ws[f'O{row_num2}'] = f"{fphm_value}合同销售"
                    ws[f'P{row_num2}'] = '6001.01.01'
                    ws[f'Q{row_num2}'] = '外销'
                    ws[f'BB{row_num2}'] = 'PRE001'
                    ws[f'BC{row_num2}'] = '人民币'
                    ws[f'BD{row_num2}'] = 'HLTX01_SYS'
                    ws[f'BE{row_num2}'] = '固定汇率'
                    ws[f'BF{row_num2}'] = 1
                    ws[f'BK{row_num2}'] = bgbgzje
                    ws[f'BM{row_num2}'] = bgbgzje
        
        # 保存文件
        output_filename = f"销售凭证_{user.username}{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(os.path.join(config.tmp_path, output_filename))
        wb.close()
        
        return {'code': 0, 'msg': '操作成功', 'data': output_filename}
    except Exception as e:
        logger.error(trace_error())
        return {'code': -1, 'msg': '操作失败', 'data': None}


"""
财务报关表头: 校验用户权限 
对应原Pascal: 财务报关表头
"""
@any_route('/api/saier/billed_notice/user_check', methods=['POST'])
@require_token
async def view_billed_notice_user_check(request):
    s = Session()
    try:
        j = await request.json()
        zm = j.get('zm', '财务报关表头')
        user = request.current_user
        d = s.query(cyzglsheet).filter(cyzglsheet.xm == user.username, cyzglsheet.zm == zm).first()
        # if not d:
        #     return json_result(-1, '用户权限校验失败')
        
        return json_result(0, '权限校验通过')
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'权限校验失败: {str(e)}')
    finally:        
        s.close()


"""
财务报关表头: 选中记录方式 
对应原Pascal: 财务报关表头
"""
@any_route('/api/saier/billed_notice/finace_title_record', methods=['POST'])
@require_token
async def view_billed_notice_finace_title_record(request):
    # 1. 检查用户是否是财务岗位 (对应 tmpcom2 查询)
    s = Session()
    try:
        j = await request.json()
        user = request.current_user
        rids = j.get('rids', [])
        fp_list = []
        for rid in rids:
            d = s.query(kaiptz.ysfp).filter(kaiptz.rid == rid).first()
            if not d:
                continue
            if d.ysfp != None and d.ysfp.strip() != '' and not d.ysfp in fp_list:
                fp_list.append(str(d.ysfp))
        da2 = '1'
        res = process_financial_customs_report(fp_list, user, da2, s)

        return json_result(res.get('code', -1), res.get('msg', ''), res.get('data', ''))
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()


"""
财务报关表头: Excel导入方式 
对应原Pascal: 财务报关表头
"""
@any_route('/api/saier/billed_notice/finace_title_excel', methods=['POST'])
@require_token
async def view_billed_notice_finace_title_excel(request):
    # 1. 检查用户是否是财务岗位 (对应 tmpcom2 查询)
    s = Session()
    try:
        j = await request.form()
        user = request.current_user
        # 2. 获取上传的Excel文件
        file = form_value(j,'file',None)
        if is_none(file):
            return json_result(ERR_PARAM_NOT_ENOUGH)
        # 保存临时文件
        temp_dir = config.tmp_path
        temp_file = os.path.join(temp_dir, f"{uuid.uuid4()}_{file.filename}")
        # logger.error(f"Saving uploaded file to: {temp_file}")
        data = await file.read()
        write_file(temp_file, data, 'wb', encoding=None)
        fp_list = []
        i = 2
        wb = load_workbook(temp_file)
        ws = wb.active
        while True:
            fphm = ws.cell(row=i, column=1).value
            if fphm is None or str(fphm).strip() == '':
                break
            if not fphm in fp_list:
                fp_list.append(fphm)
            i = i + 1

        da2 = '3'
        res = process_financial_customs_report(fp_list, user, da2, s)

        return json_result(res.get('code', -1), res.get('msg', ''), res.get('data', ''))
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()


def process_financial_customs_report(
    fp_list: List[str], # 原始发票号列表
    user,           # 当前用户对象
    da2: str,            # 类型标识 ('3' 表示使用 cymx 表，否则使用 bgmxd 表)
    s                    # SQLAlchemy Session
):
    """
    处理财务报关表头生成
    对应原Delphi代码的循环生成每个发票对应的报关表Excel文件
    """
    # 创建Excel工作簿 (假设模板已存在，这里创建新工作簿)
    template_filename = os.path.join(config.data_upload_path, 'template',f"财务报关表头.xlsx")
    wb = load_workbook(template_filename)
    ws = wb.active
    try:
        # 定义边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        i = 3  # 起始行号（原Delphi中从第3行开始）
        
        for l, ysfp in enumerate(fp_list):
            # 根据 da2 选择不同的表查询
            if da2 != '3':
                records = s.query(bgmxd).filter(bgmxd.ysfp == ysfp).all()
            else:
                records = s.query(cymx).filter(cymx.ysfp == ysfp).all()
            
            for r in records:
                i += 1
                # 插入新行（openpyxl 中需要手动插入行）
                # ws.insert_rows(i)
                # 复制第18行的格式（原Delphi中复制模板行）
                # 这里假设第18行是模板行，需要预先设置好格式
                if ws.max_row >= 18:
                    for col in range(1, 47):  # A到AT共46列
                        source_cell = ws.cell(row=18, column=col)
                        target_cell = ws.cell(row=i, column=col)
                        if source_cell.has_style:
                            target_cell.font = copy_font(source_cell.font)
                            target_cell.border = source_cell.border
                            target_cell.alignment = source_cell.alignment
                            target_cell.fill = source_cell.fill
                            target_cell.number_format = source_cell.number_format
                
                # 自动调整行高
                ws.row_dimensions[i].height = ws.row_dimensions[18].height if ws.row_dimensions[18].height else 15
                if da2 != '3':
                    # 处理 bgmxd 表数据
                    ws[f'A{i}'].value = r.ywxz
                    
                    # 查询公司简称
                    g = s.query(wfgs).filter(wfgs.wfgs == r.bggs).first()
                    if g:
                        ws[f'B{i}'].value = g.gsjc
                    else:
                        ws[f'B{i}'].value = r.bggs
                    
                    ws[f'I{i}'].value = r.zwpmz
                    
                    # 设置金额格式
                    if r.RMBkh == '是':
                        ws[f'Q{i}'].number_format = '¥#,##0.00'
                    else:
                        ws[f'Q{i}'].number_format = '$#,##0.00'
                    ws[f'Q{i}'].value = float(r.bgbgzje or 0)
                    
                    ws[f'AQ{i}'].value = r.bgdh
                
                else:
                    # 处理 cymx 表数据
                    ywbm = ''
                    ywdz = ''
                    ywxz = ''
                    
                    # 查询业务员信息
                    y = s.query(ywrybiao).filter(ywrybiao.yhm == r.ywry).first()
                    if y:
                        ywbm = y.bm or ''
                    
                    # 解析业务编码，提取字母部分
                    if ywbm:
                        # 提取连续字母作为 ywxz
                        match = re.search(r'[A-Z]+', ywbm)
                        if match:
                            ywxz = match.group()
                            ywdz = ywbm.replace(ywxz, '')
                        else:
                            ywdz = ywbm
                            ywxz = ''
                    
                    ws[f'A{i}'].value = ywxz
                
                # 公共字段
                ws[f'E{i}'].value = r.ywry
                ws[f'F{i}'].value = r.zdry
                ws[f'G{i}'].value = r.ysfp
                
                # 客户简称处理
                if r.khmc and 'BEST PRICE LLC' in r.khmc:
                    ws[f'H{i}'].value = 'BP'
                elif r.khjc:
                    ws[f'H{i}'].value = r.khjc
                else:
                    ws[f'H{i}'].value = r.khmc
                
                # 处理 cymx 附加信息
                if da2 == '3':
                    c = s.query(cymx).filter(cymx.ysfp == ysfp).first()
                    if c:
                        # 设置金额格式
                        if c.RMBkh == '是':
                            ws[f'R{i}'].number_format = '¥#,##0.00'
                            ws[f'S{i}'].number_format = '¥#,##0.00'
                        else:
                            ws[f'R{i}'].number_format = '$#,##0.00'
                            ws[f'S{i}'].number_format = '$#,##0.00'
                        
                        ws[f'R{i}'].value = float(c.mjzj or 0)
                        ws[f'S{i}'].value = float(c.wxje or 0)
                        ws[f'X{i}'].value = c.jhfs
                        
                        # 设置人民币格式的列
                        for col in ['Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF']:
                            ws[f'{col}{i}'].number_format = '¥#,##0.00'
                        
                        ws[f'Y{i}'].value = float(c.yjje or 0)
                        
                        # 查询 cymxsheet 明细
                        cghjzje = 0.0
                        tshj13 = 0.0
                        tshj10 = 0.0
                        tshj6 = 0.0
                        tshj3 = 0.0
                        tshj0 = 0.0
                        tshj1 = 0.0
                        ywxj = 0.0
                        nbxj = 0.0
                        
                        # 查询主要商品名称
                        z = s.query(cymxsheet.zhwbgpm).filter(
                            cymxsheet.pid == cymx.rid,
                            cymxsheet.zzsl > 1,
                            cymxsheet.zhwbgpm != '',
                            cymxsheet.zhwbgpm.isnot(None)
                        ).order_by(cymxsheet.gczjrmb1.desc()).first()
                        
                        if z:
                            ws[f'I{i}'].value = f"{z[0]}等"
                        
                        # 汇总查询（按条件分组）
                        # 零税率汇总
                        zero_tax = s.query(
                            func.sum(cymxsheet.gczjrmb).label('total')
                        ).filter(
                            cymxsheet.pid == c.rid,
                            cymxsheet.zzsl > 1,
                            cymxsheet.tsl == 0
                        ).first()
                        
                        if zero_tax and zero_tax.total:
                            ws[f'AD{i}'].value = float(zero_tax.total)
                            tshj0 = float(zero_tax.total)
                        
                        # 分组查询（mlsb="是"）
                        mlsb_yes = s.query(
                            cymxsheet.gcmc,
                            cymxsheet.tsl,
                            cymxsheet.zzsl,
                            cymxsheet.cgdq,
                            func.sum(cymxsheet.gczjrmb).label('total')
                        ).filter(
                            cymxsheet.pid == c.rid,
                            cymxsheet.mlsb == '是'
                        ).group_by(
                            cymxsheet.gcmc,
                            cymxsheet.tsl,
                            cymxsheet.zzsl,
                            cymxsheet.cgdq
                        ).all()
                        
                        for record in mlsb_yes:
                            total_val = float(record.total) if record.total else 0
                            truncated_val = (total_val // 10) * 10
                            
                            if record.zzsl == 0:
                                if record.cgdq == '义乌':
                                    ywxj += truncated_val
                                else:
                                    nbxj += truncated_val
                            
                            tsl_val = float(record.tsl) if record.tsl else 0
                            if tsl_val == 13:
                                tshj13 += truncated_val
                            elif tsl_val == 10:
                                tshj10 += truncated_val
                            elif tsl_val == 6:
                                tshj6 += truncated_val
                            elif tsl_val == 3:
                                tshj3 += truncated_val
                            elif tsl_val == 1:
                                tshj1 += truncated_val
                            
                            cghjzje += truncated_val
                        
                        # 分组查询（mlsb="否"）
                        mlsb_no = s.query(
                            cymxsheet.gcmc,
                            cymxsheet.tsl,
                            cymxsheet.zzsl,
                            cymxsheet.cgdq,
                            func.sum(cymxsheet.gczjrmb).label('total')
                        ).filter(
                            cymxsheet.pid == c.rid,
                            cymxsheet.mlsb == '否'
                        ).group_by(
                            cymxsheet.gcmc,
                            cymxsheet.tsl,
                            cymxsheet.zzsl,
                            cymxsheet.cgdq
                        ).all()
                        
                        for record in mlsb_no:
                            total_val = float(record.total) if record.total else 0
                            
                            if record.zzsl == 0:
                                if record.cgdq == '义乌':
                                    ywxj += total_val
                                else:
                                    nbxj += total_val
                            
                            tsl_val = float(record.tsl) if record.tsl else 0
                            if tsl_val == 13:
                                tshj13 += total_val
                            elif tsl_val == 10:
                                tshj10 += total_val
                            elif tsl_val == 6:
                                tshj6 += total_val
                            elif tsl_val == 3:
                                tshj3 += total_val
                            elif tsl_val == 1:
                                tshj1 += total_val
                            
                            cghjzje += total_val
                        
                        # 写入汇总数据
                        ws[f'Z{i}'].value = tshj13
                        ws[f'AA{i}'].value = tshj10
                        ws[f'AB{i}'].value = tshj3
                        ws[f'AC{i}'].value = tshj1
                        ws[f'AD{i}'].value = tshj0
                        ws[f'AE{i}'].value = ywxj
                        ws[f'AF{i}'].value = nbxj
                        
                        # 设置 SUM 公式
                        ws[f'Y{i}'].value = f'=SUM(Z{i}:AF{i})'
                
                # 到货日期
                ws[f'AP{i}'].value = r.chyrq
                
                # 计算天数公式
                ws[f'V{i}'].value = f'=IF(U{i}=0,TODAY()-AP{i},U{i}-AP{i})'
                
                # 税额计算公式
                ws[f'AH{i}'].value = f'=Z{i}/1.13*0.13+AA{i}/1.13*0.1+AB{i}/10.3*0.03+AC{i}/1.01*0.01-AI{i}'
                
                # 合计公式
                ws[f'AJ{i}'].value = f'=Z{i}+AA{i}+AB{i}+AC{i}+AD{i}-AG{i}'
                
                # 设置边框
                for col in range(1, 47):  # A 到 AT
                    cell = ws.cell(row=i, column=col)
                    cell.border = thin_border
        
        # 删除第3行（模板行）
        # ws.delete_rows(3)
        
        # 保存文件
        output_filename = f"财务报关表头{user.username}{datetime.now().strftime('%Y-%m-%d_%H:%M:%S')}.xlsx"

        wb.save(os.path.join(config.tmp_path, output_filename))
        wb.close()
        
        return {'code': 0, 'msg': '生成成功', 'data': output_filename}
    except Exception as e:
        logger.error(trace_error())
        return {'code': -1, 'msg': f'生成失败: {str(e)}', 'data': None}

def copy_font(source_font):
    """复制字体对象"""
    if source_font:
        return Font(
            name=source_font.name,
            size=source_font.size,
            bold=source_font.bold,
            italic=source_font.italic,
            color=source_font.color
        )
    return Font()


"""
备注说明批量导入导出: 导出功能 
对应原Pascal: 备注说明批量导入导出
"""
@any_route('/api/saier/billed_notice/memo_batch_export', methods=['POST'])
@require_token
async def view_billed_notice_memo_batch_export(request):
    # 1. 检查用户是否是财务岗位 (对应 tmpcom2 查询)
    s = Session()
    try:
        j = await request.json()
        user = request.current_user
        rids = j.get('rids', [])

        res = process_remark_description_report(rids, user, s)

        return json_result(res.get('code', -1), res.get('msg', ''), res.get('data', ''))
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()


def process_remark_description_report(
    rids: List[str],     # 选中记录的ID列表
    user,
    s                    # SQLAlchemy Session
):
    """
    处理备注说明报表生成
    对应原Delphi代码的报表生成逻辑
    """
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    try:
        # 定义居中对齐样式
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # 定义边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 设置列宽
        column_widths = {
            'A': 14.3,   # 开票序号
            'B': 14.3,   # 开票类型*
            'C': 14.3,   # 我司抬头*
            'D': 14.3,   # 外销发票号
            'E': 14.3,   # 业务人员
            'F': 14.3,   # 出货日期*
            'G': 24.3,   # 备注说明
            'H': 14.3,   # 合并厂家
            'I': 14.3,   # 实际开票总额
            'J': 14.3,   # 退 税 率
            'K': 14.3,   # 中文品名
            'L': 14.3,   # 报关数量
            'M': 14.3,   # 海关计量单位
            'N': 14.3,   # 货 源 地
            'O': 14.3,   # 合同日期*
            'P': 14.3,   # 出口货物报关单号
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 设置表头
        headers = {
            'A1': '开票序号',
            'B1': '开票类型*',
            'C1': '我司抬头*',
            'D1': '外销发票号',
            'E1': '业务人员',
            'F1': '出货日期*',
            'G1': '备注说明',
            'H1': '合并厂家',
            'I1': '实际开票总额',
            'J1': '退 税 率',
            'K1': '中文品名',
            'L1': '报关数量',
            'M1': '海关计量单位',
            'N1': '货 源 地',
            'O1': '合同日期*',
            'P1': '出口货物报关单号',
        }
        
        for cell, value in headers.items():
            ws[cell].value = value
            ws[cell].alignment = center_alignment
        
        
        i = 0  # 行计数器
        for rid in rids:
            i += 1
            row_num = i + 1
            
            # 设置自动换行和自动行高
            ws.row_dimensions[row_num].height = None  # 自动行高
            
            # 查询 kaiptz 主表
            d = s.query(kaiptz).filter(kaiptz.rid == rid).first()
            
            if d:
                # 查询 kaiptzsheet1 子表
                details = s.query(kaiptzsheet1).filter(
                    kaiptzsheet1.pid == d.rid,
                    kaiptzsheet1.kpxh != '',
                    kaiptzsheet1.kpxh.isnot(None)
                ).all()
                
                for detail in details:
                    # 设置单元格值
                    ws[f'B{row_num}'] = d.cpmc or ''
                    ws[f'C{row_num}'] = d.wstt or ''
                    ws[f'F{row_num}'] = d.chrq or ''
                    ws[f'O{row_num}'] = d.htrq or ''
                    
                    ws[f'A{row_num}'] = detail.kpxh or ''
                    ws[f'D{row_num}'] = detail.ysfp or ''
                    ws[f'E{row_num}'] = detail.ywrya or ''
                    ws[f'G{row_num}'] = detail.bzsm or ''
                    ws[f'H{row_num}'] = detail.sccj or ''
                    ws[f'I{row_num}'] = float(detail.sjkpze) if detail.sjkpze else 0
                    ws[f'J{row_num}'] = float(detail.zsl) if detail.zsl else 0
                    ws[f'K{row_num}'] = detail.zwpm or ''
                    ws[f'L{row_num}'] = float(detail.bgsl) if detail.bgsl else 0
                    ws[f'M{row_num}'] = detail.hgjldw or ''
                    ws[f'N{row_num}'] = detail.hyd or ''
                    
                    # 设置文本格式（防止长数字被科学计数）
                    ws[f'P{row_num}'].number_format = '@'
                    ws[f'P{row_num}'] = detail.ckhwbgdh or ''
        
        # 保存文件
        output_filename = f"备注说明_{user.username}{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        wb.save(os.path.join(config.tmp_path, output_filename))
        wb.close()
        
        return {'code':1, 'msg': '文件保存成功', 'data': output_filename}
    except Exception as e:
        logger.error(trace_error())
        return {'code': -1, 'msg': f'文件保存失败: {str(e)}', 'data': None}


"""
备注说明批量导入导出: 导入功能 
对应原Pascal: 备注说明批量导入导出
"""
@any_route('/api/saier/billed_notice/memo_batch_import', methods=['POST'])
@require_token
async def view_billed_notice_memo_batch_import(request):
    # 1. 检查用户是否是财务岗位 (对应 tmpcom2 查询)
    s = Session()
    try:
        j = await request.form()
        user = request.current_user
        file = form_value(j,'file',None)
        if is_none(file):
            return json_result(ERR_PARAM_NOT_ENOUGH)
        # 保存临时文件
        temp_dir = config.tmp_path
        temp_file = os.path.join(temp_dir, f"{uuid.uuid4()}_{file.filename}")
        # logger.error(f"Saving uploaded file to: {temp_file}")
        data = await file.read()
        write_file(temp_file, data, 'wb', encoding=None)

        res = update_remark_from_excel(temp_file, user, s)

        return json_result(res.get('code', -1), res.get('msg', ''), res.get('data', ''))
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()


# 使用原生SQL的版本（性能更好）
def update_remark_from_excel(
    filename: str,  # Excel文件路径
    user,
    s              # SQLAlchemy Session
):
    """
    从Excel文件读取备注说明和报关单号，更新到数据库（使用原生SQL）
    """
    
    # 打开Excel工作簿
    wb = load_workbook(filename)
    ws = wb.active
    try:
        # 收集所有需要更新的数据
        updates = []
        row_num = 2
        
        while True:
            kpxh = ws[f'A{row_num}'].value
            if not kpxh:
                break
            
            bzsm = ws[f'G{row_num}'].value or ''
            ckhwbgdh = ws[f'P{row_num}'].value or ''
            
            if bzsm:
                updates.append((str(kpxh), str(bzsm), str(ckhwbgdh)))
            
            row_num += 1
        
        # 使用原生SQL批量更新
        if updates:
            # 方法1：逐条更新
            for kpxh, bzsm, ckhwbgdh in updates:
                if kpxh==None or str(kpxh).strip() == '':
                    continue
                s.query(kaiptzsheet1).filter(kaiptzsheet1.kpxh == kpxh).update({
                    kaiptzsheet1.bzsm: bzsm,
                    kaiptzsheet1.ckhwbgdh: ckhwbgdh,
                    kaiptzsheet1.mtime: time.strftime('%Y-%m-%d %H:%M:%S'),
                    kaiptzsheet1.modi_uid: user.rid
                }, synchronize_session=False)
            
            # 方法2：使用 CASE WHEN 批量更新（适合大量数据）
            # 详见下方注释
        
        s.commit()
        wb.close()

        return {'code': 0, 'msg': '更新成功', 'data': None}
    except Exception as e:
        logger.error(trace_error())
        s.rollback()
        wb.close()
        return {'code': -1, 'msg': f'更新失败: {str(e)}', 'data': None}



"""
刷新逾期: 刷新逾期 
对应原Pascal: 刷新逾期
"""
@any_route('/api/saier/billed_notice/update_overdue', methods=['POST'])
@require_token
async def view_billed_notice_update_overdue(request):
    s = Session()
    try:
        j = await request.json()
        user = request.current_user
        s.query(kaiptzsheet1).filter(kaiptzsheet1.yqrq <datetime.now().strftime('%Y-%m-%d'),
            func.ifnull(kaiptzsheet1.dprq,'')=='').update({
            kaiptzsheet1.fpyq: '是',
            kaiptzsheet1.mtime: time.strftime('%Y-%m-%d %H:%M:%S'),
            kaiptzsheet1.modi_uid: user.rid
        }, synchronize_session=False)

        s.commit()  # 提交事务，确保数据一致性
        return json_result(1, '逾期刷新成功')
    except Exception as e:
        logger.error(trace_error())
        s.rollback()
        return json_result(-1, f'刷新逾期失败: {str(e)}')
    finally:
        s.close()



"""
批量更新业务地区电话: 批量更新业务地区电话 
对应原Pascal: 批量更新业务地区电话
"""
@any_route('/api/saier/billed_notice/update_source_phone', methods=['POST'])
@require_token
async def view_billed_notice_update_source_phone(request):
    s = Session()
    try:
        j = await request.json()
        user = request.current_user
        dq = j.get('dq', '')
        ywrya = j.get('ywrya', '')
        lxdh = j.get('lxdh', '')
        if ywrya == None or ywrya.strip() == '':
            return json_result(-1, '业务人员不能为空')
        
        s.execute("UPDATE kaiptzsheet1 SET dq = :dq, mtime = :mtime, modi_uid = :modi_uid \
            WHERE ywrya = :ywrya", {'dq': dq, 'mtime': datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'modi_uid': user.rid, 'ywrya': ywrya})
        
        s.execute("UPDATE kaiptzsheet1 SET lxdh = :lxdh, mtime = :mtime, modi_uid = :modi_uid \
            WHERE ywrya = :ywrya", {'lxdh': lxdh, 'mtime': datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'modi_uid': user.rid, 'ywrya': ywrya})
        
        s.commit()  # 提交事务，确保数据一致性
        return json_result(1, '业务地区电话更新成功')
    except Exception as e:
        logger.error(trace_error())
        s.rollback()
        return json_result(-1, f'业务地区电话更新失败: {str(e)}')
    finally:
        s.close()


"""
批量更新业务地区电话: 批量更新业务地区电话 
对应原Pascal: 批量更新业务地区电话
"""
@any_route('/api/saier/billed_notice/update_source_phone', methods=['POST'])
@require_token
async def view_billed_notice_update_source_phone(request):
    s = Session()
    try:
        j = await request.json()
        user = request.current_user
        dq = j.get('dq', '')
        ywrya = j.get('ywrya', '')
        lxdh = j.get('lxdh', '')
        if ywrya == None or ywrya.strip() == '':
            return json_result(-1, '业务人员不能为空')
        
        if dq != None and dq.strip() != '':
            s.execute("UPDATE kaiptzsheet1 SET dq = :dq, mtime = :mtime, modi_uid = :modi_uid \
                WHERE ywrya = :ywrya", {'dq': dq, 'mtime': datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'modi_uid': user.rid, 'ywrya': ywrya})

        if lxdh != None and lxdh.strip() != '':   
            s.execute("UPDATE kaiptzsheet1 SET lxdh = :lxdh, mtime = :mtime, modi_uid = :modi_uid \
                WHERE ywrya = :ywrya", {'lxdh': lxdh, 'mtime': datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'modi_uid': user.rid, 'ywrya': ywrya})
            
        s.commit()  # 提交事务，确保数据一致性
        return json_result(1, '业务地区电话更新成功')
    except Exception as e:
        logger.error(trace_error())
        s.rollback()
        return json_result(-1, f'业务地区电话更新失败: {str(e)}')
    finally:
        s.close()


"""
批量国税发票日期: 批量国税发票日期 
对应原Pascal: 批量国税发票日期
"""
@any_route('/api/saier/billed_notice/update_invoice_date', methods=['POST'])
@require_token
async def view_billed_notice_update_invoice_date(request):
    s = Session()
    try:
        j = await request.json()
        user = request.current_user
        rids = j.get('rids', [])
        date = j.get('date', '')
        if date == None or date == '':
            return json_result(-1, '日期不能为空')
        
        for rid in rids:
            s.query(kaiptz).filter(kaiptz.rid == rid).update({
                kaiptz.gsfprq: date,
                kaiptz.mtime: time.strftime('%Y-%m-%d %H:%M:%S'),
                kaiptz.modi_uid: user.rid
            }, synchronize_session=False)
            
        s.commit()  # 提交事务，确保数据一致性
        return json_result(1, '国税发票日期更新成功')
    except Exception as e:
        logger.error(trace_error())
        s.rollback()
        return json_result(-1, f'国税发票日期更新失败: {str(e)}')
    finally:
        s.close()


"""
刷新出货日期: 刷新出货日期 
对应原Pascal: 刷新出货日期
"""
@any_route('/api/saier/billed_notice/update_shipment_date', methods=['POST'])
@require_token
async def view_billed_notice_update_shipment_date(request):
    s = Session()
    try:
        j = await request.json()
        user = request.current_user
        # res = await update_shipment_date(s, user)

        async_thread_run(update_shipment_date, s, user)

        return json_result(1, '出货日期更新任务已启动，稍后请刷新页面查看结果')
    except Exception as e:
        logger.error(trace_error())
        s.rollback()
        return json_result(-1, f'出货日期更新失败: {str(e)}')
    finally:
        s.close()



async def update_shipment_date(s, user):
    """
    更新发票逾期相关日期（使用正则表达式版本）
    """
    
    today = datetime.now().strftime('%Y-%m-%d')
    
    # 获取逾期天数配置
    zx_record = s.query(zx).filter(zx.ly == '开票逾期').first()
    yq3 = int(zx_record.cs) if zx_record else 30
    
    # 日期格式转换的正则表达式
    def normalize_date(date_str):
        """标准化日期格式"""
        if not date_str:
            return None
        
        # 匹配各种日期格式
        # YYYY-MM-DD, YYYY/MM/DD, YYYY-M-D, YYYY/M/D
        match = re.search(r'(\d{4})[-/](\d{1,2})[-/](\d{1,2})', date_str)
        if match:
            year = match.group(1)
            month = match.group(2).zfill(2)
            day = match.group(3).zfill(2)
            return f"{year}-{month}-{day}"
        
        # YYYY-MM, YYYY/MM
        match = re.search(r'(\d{4})[-/](\d{1,2})$', date_str)
        if match:
            year = match.group(1)
            month = match.group(2).zfill(2)
            return f"{year}-{month}-01"
        
        # YYYYMMDD
        match = re.search(r'(\d{4})(\d{2})(\d{2})', date_str)
        if match:
            return f"{match.group(1)}-{match.group(2)}-{match.group(3)}"
        
        return None
    try:
        # 批量更新 chrq
        records = s.query(kaiptz).filter(
            func.ifnull(kaiptz.chrq, '') != '',
            or_(
                func.length(kaiptz.chrq) != 10,
                func.substr(kaiptz.chrq, 5, 1).notin_(['-', '/']),
                func.substr(kaiptz.chrq, 8, 1).notin_(['-', '/'])
            )
        ).all()
        
        index = 0
        for record in records:
            index += 1
            if record.chrq != None and record.chrq != '':
                new_chrq = normalize_date(str(record.chrq)[:10])
                if new_chrq and new_chrq != record.chrq:
                    d = s.query(kaiptz).filter(
                        kaiptz.rid == record.rid 
                    ).all()
                    for r in d:
                        r.chrq = new_chrq[:10]
                        r.modi_uid = user.rid
                        r.mtime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        s.add(r)
            if index % 10 == 0:
                await asyncio.sleep(0.01)  # 让出控制权，避免长时间占用事件循环

        # 更新 kaiptzsheet1 的 yqrq
        records = s.query(kaiptz).filter(
            and_(
                func.ifnull(kaiptz.chrq, '') != '',
                kaiptz.chrq.isnot(None)
            )
        ).all()
        
        index = 0
        for record in records:
            index += 1
            if record.chrq != None and record.chrq != '':
                chrq_date = datetime.strptime(str(record.chrq)[:10], '%Y-%m-%d')
                yqrq = (chrq_date + timedelta(days=yq3)).strftime('%Y-%m-%d')
                d = s.query(kaiptzsheet1).filter(
                    kaiptzsheet1.pid == record.rid
                ).all()
                for r in d:
                    r.yqrq = yqrq
                    r.modi_uid = user.rid
                    r.mtime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    s.add(r)
            if index % 10 == 0:
                await asyncio.sleep(0.01)  # 让出控制权，避免长时间占用事件循环

        index = 0
        # 更新 fpyq
        d = s.query(kaiptzsheet1).filter(
            or_(
                func.ifnull(kaiptzsheet1.dprq, '') == '',
                kaiptzsheet1.dprq.is_(None)
            ),
            kaiptzsheet1.yqrq > today
        ).all()
        for r in d:
            index += 1
            r.fpyq = '否'
            r.modi_uid = user.rid
            r.mtime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            s.add(r)
            if index % 10 == 0:
                await asyncio.sleep(0.01)  # 让出控制权，避免长时间占用事件循环

        s.commit()
        return json_result(1, '发票逾期日期更新成功')
    except Exception as e:
        logger.error(trace_error())
        s.rollback()
        return json_result(-1, f'发票逾期日期更新失败: {str(e)}')

"""
批量到票: 批量到票 
对应原Pascal: 批量到票
"""
@any_route('/api/saier/billed_notice/batch_permit', methods=['POST'])
@require_token
async def view_billed_notice_batch_permit(request):
    s = Session()
    try:
        j = await request.json()
        rids = j.get('rids', [])
        dprq = j.get('dprq', '')
        user = request.current_user
        ywbm = ''
        org = get_user_path(user.username)
        path = org.get('path', '')
        u = s.query(ywrybiao).filter(ywrybiao.yhm == user.username).first()
        if u:
            ywbm = u.bm
        for rid in rids:
            d = s.query(kaiptz).filter(kaiptz.rid == rid, kaiptz.kfdy == '可以', 
                kaiptz.uid == user.rid, 
                or_(func.ifnull(kaiptz.yjdp,"") == "", kaiptz.yjdp.is_(None))).first()
            if not d:
                continue  # 如果没有找到符合条件的记录，跳过更新
            d.yjdp = dprq
            d.mtime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            d.modi_uid = user.rid
            s.add(d)
            c = s.query(kaiptzsheet1).filter(kaiptzsheet1.pid == d.rid).all()
            for r in c:
                r.ywbm = ywbm
                r.ywpath = path
                r.yjdp = dprq
                r.mtime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                r.modi_uid = user.rid
                s.add(r)

        s.commit()  # 提交事务，确保数据一致性
        return json_result(1, '批量到票更新成功')
    except Exception as e:
        logger.error(trace_error())
        s.rollback()
        return json_result(-1, f'批量到票更新失败: {str(e)}')
    finally:
        s.close()


"""
更改批量到票: 更改批量到票 
对应原Pascal: 更改批量到票
"""
@any_route('/api/saier/billed_notice/update_permit', methods=['POST'])
@require_token
async def view_billed_notice_update_permit(request):
    s = Session()
    try:
        j = await request.json()
        rids = j.get('rids', [])
        dprq = j.get('dprq', '')
        sccj = j.get('sccj', '')
        if dprq == None or dprq.strip() == '':
            return json_result(-1, '到票日期不能为空')
        if sccj == None or sccj.strip() == '':
            return json_result(-1, '工厂名称不能为空')
        user = request.current_user
        ywbm = ''
        org = get_user_path(user.username)
        path = org.get('path', '')
        u = s.query(ywrybiao).filter(ywrybiao.yhm == user.username).first()
        if u:
            ywbm = u.bm
        for rid in rids:
            d = s.query(kaiptz).filter(kaiptz.rid == rid, kaiptz.kfdy == '可以', 
                kaiptz.uid == user.rid, 
                or_(func.ifnull(kaiptz.yjdp,"") == "", kaiptz.yjdp.is_(None))).first()
            if not d:
                continue  # 如果没有找到符合条件的记录，跳过更新
            if not sccj in (d.sccj or ''):
                continue  # 如果工厂名称已包含指定值，跳过更新

            d.yjdp = dprq
            d.mtime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            d.modi_uid = user.rid
            s.add(d)
            c = s.query(kaiptzsheet1).filter(kaiptzsheet1.pid == d.rid).all()
            for r in c:
                r.ywbm = ywbm
                r.ywpath = path
                r.yjdp = dprq
                r.mtime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                r.modi_uid = user.rid
                s.add(r)

        s.commit()  # 提交事务，确保数据一致性
        return json_result(1, '批量到票更新成功')
    except Exception as e:
        logger.error(trace_error())
        s.rollback()
        return json_result(-1, f'批量到票更新失败: {str(e)}')
    finally:
        s.close()


"""
按合同批量解除锁定
对应原Pascal: 按合同批量解除锁定
"""
@any_route('/api/saier/billed_notice/batch_unlock', methods=['POST'])
@require_token
async def view_billed_notice_batch_unlock(request):
    s = Session()
    try:
        j = await request.form()
        module = form_value(j, 'module', '开票通知')
        lock = form_value(j, 'lock', '解锁')
        kind = form_value(j, 'kind', '采购合同')
        user = request.current_user
        org = get_user_path(user.username)
        position = org.get('position')
        if '财务' not in position:
            return json_result(-1, '只有财务岗位用户才能执行此操作')
        # 2. 获取上传的Excel文件
        file = form_value(j,'file',None)
        if is_none(file):
            return json_result(ERR_PARAM_NOT_ENOUGH)
        # 保存临时文件
        temp_dir = config.tmp_path
        temp_file = os.path.join(temp_dir, f"{uuid.uuid4()}_{file.filename}")
        # logger.error(f"Saving uploaded file to: {temp_file}")
        data = await file.read()
        write_file(temp_file, data, 'wb', encoding=None)
        thdz = '否'
        if lock == '解锁':
            thdz = '是' 

        # 3. 读取Excel文件
        wb = load_workbook(temp_file)
        ws = wb.worksheets[0]  # 读取第一个工作表
        # 4. 处理每一行数据
        row_idx = 1  # 从第1行开始（第1行是表头）
        while True:
            # 读取Excel单元格数据
            hthm = str(ws.cell(row=row_idx, column=1).value or '').strip()  # A列 - 采购合同号
            cpbh = str(ws.cell(row=row_idx, column=2).value or '')  # B列 - 货号
            # 如果rid为空，结束循环
            if not hthm:
                break
            filters1 = [kaiptzsheet.cght == hthm, kaiptzsheet.cpbh == cpbh]
            filters2 = [xjhcsheet.cght == hthm, xjhcsheet.cpbh == cpbh]
            if kind == '外销合同':
                filters1 = [kaiptzsheet.wxht == hthm, kaiptzsheet.cpbh == cpbh]
                filters2 = [xjhcsheet.wxht == hthm, xjhcsheet.cpbh == cpbh]

            s.query(kaiptzsheet).filter(*filters1).update({kaiptzsheet.thdz: thdz}, synchronize_session=False)
            s.query(xjhcsheet).filter(*filters2).update({xjhcsheet.thdz: thdz}, synchronize_session=False)
            row_idx += 1
        
        s.commit()
        return json_result(0, f"批量{lock}成功")
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()
        wb.close()


"""
批量更新出货日期: 批量更新出货日期 
对应原Pascal: 批量更新出货日期
"""
@any_route('/api/saier/billed_notice/batch_shipment_date', methods=['POST'])
@require_token
async def view_billed_notice_batch_shipment_date(request):
    s = Session()
    try:
        j = await request.json()
        user = request.current_user
        rids = j.get('rids', [])
        date = j.get('date', '')
        if date == None or date.strip() == '':
            return json_result(-1, '日期不能为空')
        
        for rid in rids:
            d = s.query(kaiptz).filter(kaiptz.rid == rid).first()
            if d:
                if d.zdry != user.username:
                    continue  # 只有责任人是当前用户的记录才允许更新

                d.chrq = date
                d.mtime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                d.modi_uid = user.rid
                s.add(d)

        s.commit()  # 提交事务，确保数据一致性

        return json_result(1, '出货日期更新成功')
    except Exception as e:
        logger.error(trace_error())
        s.rollback()
        return json_result(-1, f'出货日期更新失败: {str(e)}')
    finally:
        s.close()


"""
批量核算发票
对应原Pascal: 批量核算发票
"""
@any_route('/api/saier/billed_notice/invoice_adjust', methods=['POST'])
@require_token
async def view_billed_notice_invoice_adjust(request):
    s = Session()
    try:
        j = await request.form()
        module = form_value(j, 'module', '开票通知')
        user = request.current_user
        org = get_user_path(user.username)
        position = org.get('position')
        if '财务' not in position:
            return json_result(-1, '只有财务岗位用户才能执行此操作')
        # 2. 获取上传的Excel文件
        file = form_value(j,'file',None)
        if is_none(file):
            return json_result(ERR_PARAM_NOT_ENOUGH)
        # 保存临时文件
        temp_dir = config.tmp_path
        temp_file = os.path.join(temp_dir, f"{uuid.uuid4()}_{file.filename}")
        # logger.error(f"Saving uploaded file to: {temp_file}")
        data = await file.read()
        write_file(temp_file, data, 'wb', encoding=None)
        # 3. 读取Excel文件
        wb = load_workbook(temp_file)
        ws = wb.worksheets[0]  # 读取第一个工作表
        # 4. 处理每一行数据
        row_idx = 1  # 从第1行开始（第1行是表头）
        m = [cymx, fysqsheet, cymxsheet, dbbqd, bgmxd, bgmxdsheet, ydchtjsheet, hdfy, kaiptz, kaiptzsheet, kaiptzxq, kaiptzsheet1]
        while True:
            # 读取Excel单元格数据
            fphm = str(ws.cell(row=row_idx, column=1).value or '').strip()  # A列 - 发票号码
            ysfp = str(ws.cell(row=row_idx, column=2).value or '')  # B列 - 核算发票
            # 如果rid为空，结束循环
            if not fphm or not ysfp:
                break
        
            for o in m:
                if o == kaiptzxq:
                    s.query(o).filter(o.ysfp == fphm).update({o.ysfp: ysfp}, synchronize_session=False)
                    continue
                if hasattr(o, 'wxfp') and hasattr(o, 'ysfp'):
                    s.query(o).filter(o.wxfp == fphm).update({o.ysfp: ysfp}, synchronize_session=False)
                elif hasattr(o, 'fphm') and hasattr(o, 'ysfp'):
                    s.query(o).filter(o.fphm == fphm).update({o.ysfp: ysfp}, synchronize_session=False)
                if o == kaiptzsheet or o == bgmxdsheet or o == kaiptz:
                    s.query(o).filter(o.ysfp == fphm).update({o.yysfp: ysfp}, synchronize_session=False)

            row_idx += 1
        
        s.commit()  
        return json_result(0, f"批量核算发票操作成功")
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()
        wb.close()


"""
批量离职交接人员: 批量离职交接人员 
对应原Pascal: 批量离职交接人员
"""
@any_route('/api/saier/billed_notice/batch_responsible', methods=['POST'])
@require_token
async def view_billed_notice_batch_responsible(request):
    s = Session()
    try:
        j = await request.json()
        user = request.current_user
        rids = j.get('rids', [])
        username = j.get('username', '')
        if username == None or username.strip() == '':
            return json_result(-1, '交接人员名字不能为空')
        u = s.query(sys_user.username).filter(sys_user.username == username).first()
        if not u:
            return json_result(-1, '交接人员不存在')
        for rid in rids:
            d = s.query(kaiptz).filter(kaiptz.rid == rid).first()
            if d:
                d.lzjjry = username
                d.xgry = user.username
                d.mtime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                d.modi_uid = user.rid
                s.add(d)

        s.commit()  # 提交事务，确保数据一致性

        return json_result(1, '交接人员更新成功')
    except Exception as e:
        logger.error(trace_error())
        s.rollback()
        return json_result(-1, f'交接人员更新失败: {str(e)}')
    finally:
        s.close()


"""
账面利润表: 账面利润表 
对应原Pascal: 账面利润表
"""
@any_route('/api/saier/billed_notice/batch_profit_list', methods=['POST'])
@require_token
async def view_billed_notice_batch_profit_list(request):
    s = Session()
    try:
        j = await request.json()
        user = request.current_user
        rids = j.get('rids', [])

        fp_list = []
        for rid in rids:
            d = s.query(kaiptz.fphm).filter(kaiptz.rid == rid).first()
            if not d:
                continue  # 如果没有找到符合条件的记录，跳过更新
            if d.fphm == None or d.fphm.strip() == '':
                continue  # 如果发票号码为空，跳过更新
            if not d.fphm in fp_list:
                fp_list.append(d.fphm)

        res = process_book_profit_report(fp_list, user, s)
        
        return json_result(res.get('code', -1), res.get('msg', '账面利润表生成失败'), res.get('data', None))
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'交接人员更新失败: {str(e)}')
    finally:
        s.close()



def process_book_profit_report(
    fp_list: List[str], # 原始发票号列表
    user,
    s                    # SQLAlchemy Session
):
    """
    处理账面利润表生成
    对应原Delphi代码的报表生成逻辑
    """
    
    # 创建Excel工作簿
    filename = os.path.join(config.data_upload_path, 'template', '账面利润表.xlsx')
    wb = load_workbook(filename)
    ws = wb.active
    try:
        k = 1  # 行计数器，从第1行开始
        hl = 0.0  # 汇率
        
        # 第1行设置汇率
        ws[f'D{k}'].value = hl
        for lz, ysfp in enumerate(fp_list):
            k += 1
            hl = 0.0
            hl1 = 0.0
            chyrq = ''
            
            # 查询 bgmxd 表
            d = s.query(bgmxd).filter(
                or_(
                    bgmxd.ysfp == ysfp,
                    bgmxd.fphm == ysfp
                )
            ).first()
            if d:
                # 设置基础信息
                ws[f'A{k}'].value = ysfp
                # 计算预计日期（zgrq + 10天）
                if d.zgrq and len(str(d.zgrq)) >= 10:
                    zgrq_date = datetime.strptime(str(d.zgrq)[:10], '%Y-%m-%d')
                    chyrq = (zgrq_date + timedelta(days=10)).strftime('%Y-%m-%d')
                    ws[f'J{k}'].value = str(chyrq)[:10]
                
                # 查询 kaiptz 表获取出货日期
                z = s.query(kaiptz).filter(
                    kaiptz.fphm == ysfp,
                    kaiptz.cpmc != '补报'
                ).first()
                
                if z and z.chrq:
                    ws[f'J{k}'].value = str(z.chrq)[:10]
                    chyrq = str(z.chrq)[:10]
                
                # 计算汇率
                if d.RMBkh == '是' or d.hbdm == 'RMB':
                    hl = 1.0
                else:
                    if d.hbdm not in ('USD$', 'USD'):
                        # 从 kpnrsheet 查询汇率
                        sheet = s.query(kpnrsheet).filter(
                            kpnrsheet.wfgs == d.bggs,
                            kpnrsheet.hbdm == d.hbdm
                        ).first()
                        
                        if sheet and sheet.hl:
                            hl = float(sheet.hl)
                            hl1 = float(sheet.hl)
                
                # 如果汇率为0，从 kpnr 查询
                if hl1 == 0:
                    n = s.query(kpnr).filter(kpnr.wfgs == d.bggs).first()
                    if n and n.tjhl:
                        hl1 = float(n.tjhl)
                
                if hl != 1.0:
                    hl = hl1
                
                ws[f'D{k}'].value = hl
                
                # 查询 bgmxdhbcp 表（有退税率的数据）
                bgmxdhbcp_list = s.query(
                    func.sum(bgmxdhbcp.wxzj).label('bgbgzje12'),
                    func.sum(bgmxdhbcp.CNFyf).label('CNFyf12')
                ).filter(
                    bgmxdhbcp.pid == d.rid,
                    bgmxdhbcp.tsl > 0
                ).first()
                
                if bgmxdhbcp_list:
                    bgbgzje12 = float(bgmxdhbcp_list.bgbgzje12 or 0)
                    cnfyf12 = float(bgmxdhbcp_list.CNFyf12 or 0)
                    
                    ws[f'C{k}'].value = round(bgbgzje12 * 100) / 100
                    ws[f'L{k}'].value = round(cnfyf12 * 100) / 100
                    ws[f'M{k}'].value = d.bggs
                    
                    # 查询 kaiptzsheet1 表
                    sheet2 = s.query(
                        func.sum(kaiptzsheet1.sjkpze).label('sjkpze1'),
                        func.sum(kaiptzsheet1.sjkpze / (1 + (kaiptzsheet1.zzsl / 100.0))).label('sjkpze2'),
                        func.sum((kaiptzsheet1.sjkpze / (1 + (kaiptzsheet1.zzsl / 100.0))) * (kaiptzsheet1.zzsl / 100.0)).label('sjkpze3'),
                        func.sum((kaiptzsheet1.sjkpze / (1 + (kaiptzsheet1.zzsl / 100.0))) * (kaiptzsheet1.zsl / 100.0)).label('sjkpze4')
                    ).filter(
                        kaiptzsheet1.fphm == ysfp,
                        kaiptzsheet1.zsl > 0
                    ).first()
                    
                    if sheet2:
                        sjkpze1 = float(sheet2.sjkpze1 or 0)
                        sjkpze2 = float(sheet2.sjkpze2 or 0)
                        sjkpze4 = float(sheet2.sjkpze4 or 0)
                        
                        ws[f'F{k}'].value = round(sjkpze1 * 100) / 100
                        ws[f'G{k}'].value = round(sjkpze2 * 100) / 100
                        ws[f'H{k}'].value = round(sjkpze1 * 100) / 100 - round(sjkpze2 * 100) / 100
                        ws[f'I{k}'].value = round(sjkpze4 * 100) / 100
                    
                    # 设置公式
                    ws[f'B{k}'].value = f'=(C{k}-L{k})*D{k}-E{k}'
                    ws[f'E{k}'].value = f'=F{k}-I{k}'
                    ws[f'K{k}'].value = f'=((C{k}-L{k})*D{k}-E{k})/((C{k}-L{k})*D{k})'
                
                # 查询 bgmxdhbcp 表（退税率为0的数据）
                bgmxdhbcp_zero = s.query(
                    func.sum(bgmxdhbcp.wxzj).label('bgbgzje12'),
                    func.sum(bgmxdhbcp.CNFyf).label('CNFyf12')
                ).filter(
                    bgmxdhbcp.pid == d.rid,
                    bgmxdhbcp.tsl == 0
                ).first()
                
                if bgmxdhbcp_zero and (bgmxdhbcp_zero.bgbgzje12 or 0) > 0:
                    k += 1
                    bgbgzje12_zero = float(bgmxdhbcp_zero.bgbgzje12 or 0)
                    cnfyf12_zero = float(bgmxdhbcp_zero.CNFyf12 or 0)
                    
                    ws[f'D{k}'].value = hl
                    ws[f'A{k}'].value = ysfp
                    ws[f'J{k}'].value = chyrq
                    ws[f'C{k}'].value = round(bgbgzje12_zero * 1000) / 1000
                    ws[f'L{k}'].value = round(cnfyf12_zero * 1000) / 1000
                    ws[f'M{k}'].value = d.bggs
                    
                    # 查询 kaiptzsheet1 表（退税率为0）
                    kaiptzsheet_zero = s.query(
                        func.sum(kaiptzsheet1.sjkpze).label('sjkpze1'),
                        func.sum(kaiptzsheet1.sjkpze / (1 + (kaiptzsheet1.zzsl / 100.0))).label('sjkpze2'),
                        func.sum((kaiptzsheet1.sjkpze / (1 + (kaiptzsheet1.zzsl / 100.0))) * (kaiptzsheet1.zzsl / 100.0)).label('sjkpze3')
                    ).filter(
                        kaiptzsheet1.fphm == ysfp,
                        kaiptzsheet1.zsl == 0
                    ).first()
                    
                    if kaiptzsheet_zero:
                        sjkpze1 = float(kaiptzsheet_zero.sjkpze1 or 0)
                        sjkpze2 = float(kaiptzsheet_zero.sjkpze2 or 0)
                        
                        ws[f'F{k}'].value = round(sjkpze1 * 100) / 100
                        ws[f'G{k}'].value = round(sjkpze2 * 100) / 100
                        ws[f'H{k}'].value = round(sjkpze1 * 100) / 100 - round(sjkpze2 * 100) / 100
                        ws[f'I{k}'].value = 0
                    
                    # 设置公式
                    ws[f'B{k}'].value = f'=(C{k}-L{k})*D{k}-E{k}'
                    ws[f'E{k}'].value = f'=F{k}-I{k}'
                    ws[f'K{k}'].value = f'=((C{k}-L{k})*D{k}-E{k})/((C{k}-L{k})*D{k})'
                
                # 更新 kaiptzsheet1 表的 dzsb 字段
                s.query(kaiptzsheet1).filter(
                    kaiptzsheet1.fphm == ysfp
                ).update({'dzsb': '是'}, synchronize_session=False)
        
        s.commit()
        
        # 保存文件
        logger.error(f"Saving generated report for user {user.username}")
        output_filename = f"账面利润表{user.username}{datetime.now().strftime('%Y-%m-%d%H%M%S')}.xlsx"
        logger.error(f"Saving generated report for user {output_filename}")
        wb.save(os.path.join(config.tmp_path, output_filename))
        wb.close()
    
        return {'code': 1, 'msg': '文件保存成功', 'data': output_filename}
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return {'code': -1, 'msg': f'处理失败: {str(e)}'}


"""
批量换汇成本 : 批量换汇成本  
对应原Pascal: 批量换汇成本 
"""
@any_route('/api/saier/billed_notice/batch_exchange_rate', methods=['POST'])
@require_token
async def view_billed_notice_batch_exchange_rate(request):
    s = Session()
    try:
        j = await request.json()
        user = request.current_user
        rids = j.get('rids', [])

        fp_list = []
        for rid in rids:
            d = s.query(kaiptz.fphm).filter(kaiptz.rid == rid).first()
            if not d:
                continue  # 如果没有找到符合条件的记录，跳过更新
            if d.fphm == None or d.fphm.strip() == '':
                continue  # 如果发票号码为空，跳过更新
            if not d.fphm in fp_list:
                fp_list.append(d.fphm)

        res = update_exchange_rate(fp_list, user, s)
        
        return json_result(1, '批量换汇成本更新成功', res.get('data', None))
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'批量换汇成本更新失败: {str(e)}')
    finally:
        s.close()


def update_exchange_rate(
    fp_list: list,  # 发票号码列表
    user,
    s               # SQLAlchemy Session
):
    try:
        for fphm in fp_list:
            # 查询 bgmxd 信息
            d = s.query(bgmxd).filter(bgmxd.fphm == fphm).first()
            
            if not d:
                continue
            
            hl = 1 if d.RMBkh == '是' else 0
            father = str(d.rid)
            bggs = d.bggs or ''
            
            # 查询推荐汇率
            n = s.query(kpnr).filter(kpnr.wfgs == bggs).first()
            hl1 = float(n.tjhl) if n and n.tjhl else 0
            
            # 查询所有明细
            details = s.query(kaiptzsheet1).filter(
                kaiptzsheet1.fphm == fphm
            ).all()
            
            # 预先计算每组的总金额
            group_totals = {}
            for detail in details:
                key = (detail.zwpm1, detail.zsl, detail.hyd, detail.bgbh)
                if key not in group_totals:
                    total = s.query(func.sum(kaiptzsheet1.sjkpze)).filter(
                        and_(
                            kaiptzsheet1.fphm == fphm,
                            kaiptzsheet1.zwpm1 == detail.zwpm1,
                            kaiptzsheet1.zsl == detail.zsl,
                            kaiptzsheet1.hyd == detail.hyd,
                            kaiptzsheet1.bgbh == detail.bgbh
                        )
                    ).scalar() or 0
                    group_totals[key] = float(total)
            
            # 预先查询报关金额
            bgje_cache = {}
            
            for detail in details:
                number1 = str(detail.rid)
                sjkpze = float(detail.sjkpze or 0)
                
                # 计算占比
                key = (detail.zwpm1, detail.zsl, detail.hyd, detail.bgbh)
                total = group_totals.get(key, 0)
                bl = sjkpze / total if total > 0 else 0
                
                # 计算税额
                zsl = float(detail.zsl or 0)
                zzsl = float(detail.zzsl or 0)
                
                if zsl <= 4:
                    if zsl == 3:
                        tse = (sjkpze / 1.03) * zsl / 100
                    elif zsl == 1:
                        tse = (sjkpze / 1.01) * zsl / 100
                    else:
                        tse = 0
                else:
                    tse = (sjkpze / (1 + zzsl / 100)) * zsl / 100
                
                # 获取报关金额
                cache_key = (father, detail.zwpm1, detail.zsl, detail.bgbh, detail.hyd)
                
                if cache_key not in bgje_cache:
                    # 按货源地匹配查询
                    b = s.query(bgmxdhbcp).filter(
                        and_(
                            bgmxdhbcp.father == father,
                            bgmxdhbcp.zwpm == detail.zwpm1,
                            bgmxdhbcp.tsl == detail.zsl,
                            bgmxdhbcp.bgbh == detail.bgbh,
                            bgmxdhbcp.hyd.like(f'%{detail.hyd}%')
                        )
                    ).first()
                    
                    if not b and detail.hyd:
                        b = s.query(bgmxdhbcp).filter(
                            and_(
                                bgmxdhbcp.father == father,
                                bgmxdhbcp.zwpm == detail.zwpm1
                            )
                        ).first()
                    
                    if b:
                        wxzj = float(bgmxdhbcp.wxzj or 0)
                        cnfyf = float(bgmxdhbcp.CNFyf or 0)
                        bgje_cache[cache_key] = (wxzj - cnfyf)
                    else:
                        bgje_cache[cache_key] = 0
                
                bgje = bgje_cache[cache_key] * bl
                
                # 精度处理
                sjkpze_trunc = round(sjkpze * 100) / 100
                tse_trunc = round(tse * 100) / 100
                bgje_trunc = round(bgje * 100) / 100
                
                # 计算成本率
                hhcb = 0
                if hl != 1:
                    if bgje_trunc > 0:
                        hhcb = round(((sjkpze_trunc - tse_trunc) / bgje_trunc) * 100) / 100
                else:
                    if bgje_trunc > 0 and hl1 > 0:
                        bgje_convert = round((bgje_trunc / hl1) * 100) / 100
                        if bgje_convert > 0:
                            hhcb = round(((sjkpze_trunc - tse_trunc) / bgje_convert) * 100) / 100
                
                # 更新
                s.query(kaiptzsheet1).filter(
                    and_(
                        kaiptzsheet1.fphm == fphm,
                        kaiptzsheet1.rid == number1
                    )
                ).update({'hhcb': hhcb}, synchronize_session=False)
            
        s.commit()
        return {'code': 1, 'msg': '换汇成本更新成功'}
    except Exception as e:
        logger.error(trace_error())
        s.rollback()
        return {'code': -1, 'msg': f'换汇成本更新失败: {str(e)}'}




