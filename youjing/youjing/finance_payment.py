from any import *
from .model import *
from .__default__ import get_user_path, module_share_new, module_xxck_new, user_task_delete, user_task_new
import openpyxl
from openpyxl.styles import numbers
from sqlalchemy import create_engine, Column, String, Integer, Float, Date, ForeignKey
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, relationship
from datetime import datetime


"""
财务付款核对.批量付款日期.change
对应原Pascal: 批量付款日期.change
"""
@any_route('/api/saier/finance_payment/plfkrq/change', methods=['POST'])
@require_token
async def view_finance_payment_plfkrq_change(request):
    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        plfkrq = j.get('plfkrq', '')
        lines = j.get('lines', [])
        code = 1
        msg = '操作成功'
        data = {'items':{}}
        for line in lines:
            sb = line.get('sb', '')
            ly = line.get('ly', '')
            rid = line.get('rid', '')
            # if ly == '异地':
            c = s.query(gchk.fkhj,gchk.bzsm).filter(gchk.sb==sb).first()
            if c:
                data['items'][rid] = {
                    'fkhj': float(c.fkhj) if c.fkhj else 0.0,
                    'bzsm': c.bzsm
                }
            
        return json_result(code, msg, data)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'货款类型校验失败: {str(e)}')
    finally:
        s.close()


"""
财务付款核对.截止日期.change
对应原Pascal: 截止日期.change
"""
@any_route('/api/saier/finance_payment/jzrq/change', methods=['POST'])
@require_token
async def view_finance_payment_jzrq_change(request):
    s = Session()
    try:
        code = 1
        msg = '操作成功'
        j = await request.json()
        cwry = j.get('cwry', '')
        position = ''
        if cwry != '' and cwry is not None:
            org = get_user_path(cwry)
            position = org.get('position', '')

        return json_result(code, msg, position)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'货款类型校验失败: {str(e)}')
    finally:
        s.close()


"""
财务付款核对.付款资料.delete
对应原Pascal: 付款资料.delete
"""
@any_route('/api/saier/finance_payment/child/delete', methods=['POST'])
@require_token
async def view_finance_payment_child_delete(request):
    s = Session()
    try:
        code = 1
        msg = '操作成功'
        j = await request.json()
        sb = j.get('sb', '')
        jzrq = j.get('jzrq', '')
        s.query(cgfkhdsheet).filter(cgfkhdsheet.sb == sb, cgfkhdsheet.jzrq == jzrq).delete()
        s.commit()

        return json_result(code, msg)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'货款类型校验失败: {str(e)}')
    finally:
        s.close()


"""
财务付款核对.批量付款资料导出
对应原Pascal: 批量付款资料导出
"""
@any_route('/api/saier/finance_payment/payment/export', methods=['POST'])
@require_token
async def view_finance_payment_payment_export(request):
    s = Session()
    try:
        code = 1
        msg = '操作成功'
        j = await request.json()
        rid = j.get('rid', '')
        user = request.current_user
        org = get_user_path(user.username)
        path = org.get('path', '')
        position = org.get('position', '')
        if '财务' not in position:
            return json_result(-1, '当前用户非财务岗位，无权限导出', data={})
        output_template = config.data_upload_path

        filename = os.path.join(output_template, 'template', 'yw采购付款.xlsx')
        res = export_to_excel(s, user, rid, path, filename)

        return res
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'货款类型校验失败: {str(e)}')
    finally:
        s.close()


# 主函数
def export_to_excel(s, user, rid, path, filename):
    # 创建 Excel 应用对象（openpyxl 方式）
    wb = load_workbook(filename)
    ws = wb.active
    ws.title = 'Sheet1'
    try:
        # 设置表头
        headers = {
            'AV2': '我司抬头', 'AM2': '银行编码', 'AW2': '是否电汇',
            'AX2': '盖章收回', 'AY2': '诚信报告', 'AZ2': '到票日期',
            'BA2': '业务核对', 'BB2': '核对日期', 'BC2': '业务不通过原因',
            'BD2': '是否暂扣', 'BE2': '暂扣金额', 'BF2': '我方银行帐号',
            'BG2': '我方银行名称', 'BH2': '银行代码', 'BI2': '预约支付',
            'BJ2': '同行转账', 'BK2': '同城异地', 'BL2': '公私标识',
            'BM2': '客户号', 'BN2': '单位code', 'BO2': '子户序号'
        }
        for cell, value in headers.items():
            ws[cell] = value
        
        i = 2
        records = s.query(cwfkhdsheet).filter(cwfkhdsheet.pid == rid).all()
        
        if records:
            for record in records:
                i += 1
                ws[f'A{i}'] = record.rid
                ws[f'B{i}'] = record.sb
                ws[f'C{i}'] = record.fpbh
                ws[f'D{i}'] = record.gcfp
                
                # 发票金额处理
                fpje = record.fpje or ''
                if fpje == '0' or fpje == '':
                    ws[f'F{i}'] = float(record.yfhj) if record.yfhj else 0
                else:
                    ws[f'F{i}'] = float(record.fpje) if record.fpje else 0
                
                ws[f'G{i}'] = '0'
                ws[f'H{i}'] = float(record.fkhj) if record.fkhj else 0
                ws[f'J{i}'] = record.sfjq
                ws[f'K{i}'] = str(record.hkrq)[:10] if record.hkrq else ''
                
                # 查询 gchk 或 ydhk
                if record.sb:
                    g = s.query(gchk).filter(gchk.sb == record.sb).first()
                    if g:
                        ws[f'L{i}'] = g.chrq
                        ws[f'AM{i}'] = g.yhdm
                        ws[f'AM{i}'].number_format = '@'  # 文本格式
                    # else:
                    #     y = s.query(ydhk).filter(ydhk.sb == record.sb).first()
                    #     if y:
                    #         ws[f'L{i}'] = y.chrq
                    #         ws[f'AM{i}'] = y.yhdm
                    #         ws[f'AM{i}'].number_format = '@'
                
                ws[f'M{i}'] = record.zzsl
                ws[f'N{i}'] = record.tsl
                ws[f'O{i}'] = record.gcmc
                yfgc = record.gcmc or ''
                ws[f'P{i}'] = record.gcmc1
                yfgc1 = record.gcmc1 or ''
                ws[f'Q{i}'] = record.fkbz
                ws[f'R{i}'] = record.htje
                ws[f'S{i}'] = record.bhsj
                ws[f'T{i}'] = record.se
                ws[f'U{i}'] = record.tse
                ws[f'V{i}'] = record.jsrm
                ws[f'W{i}'] = record.sh
                ws[f'X{i}'] = record.bank
                ws[f'Y{i}'] = record.zh
                ws[f'Z{i}'] = record.province
                ws[f'AA{i}'] = record.city
                ws[f'AB{i}'] = record.yt
                ws[f'AC{i}'] = record.bzsm
                ws[f'AD{i}'] = record.wxfp
                ws[f'AE{i}'] = record.fkdq
                ws[f'AJ{i}'] = record.sb
                ws[f'AK{i}'] = record.fkxh
                ws[f'AL{i}'] = record.dlrq
                ws[f'AV{i}'] = record.wstt
                ws[f'AY{i}'] = record.cxbg
                ws[f'BA{i}'] = record.ywhd
                ws[f'BB{i}'] = record.jzrq
                ws[f'BC{i}'] = record.ywbtgyy
                ws[f'BD{i}'] = record.sfzk
                ws[f'BE{i}'] = float(record.zkje) if record.zkje else 0
                
                # 查询预付合计
                if not yfgc1:
                    yfgc1 = '没有123'
                
                fkje = 0.0
                yfhk_records = s.query(yfhk).filter(
                    (yfhk.sccj.like(f'%{yfgc}%') | yfhk.gcmc1.like(f'%{yfgc1}%')),
                    yfhk.sfjq != '是',
                ).all()
                if yfhk_records:
                    for y in yfhk_records:
                        org = get_user_path(y.uid)
                        p = org.get('path', '')
                        if path != p:
                            continue
                        yfhk_detail = s.query(yfhk).filter(yfhk.rid == y.rid).first()
                        if yfhk_detail and yfhk_detail.fkje:
                            fkje += float(yfhk_detail.fkje)
                    
                    ws[f'AR{i}'] = f'有预付合计:{fkje}'
        
        # 保存 Excel 文件
        output_filename = f"采购付款资料_{user.username}_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        wb.save(os.path.join(config.tmp_path, output_filename))
        
        return json_result(0, '操作成功', output_filename)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'导出 Excel 失败: {str(e)}', '')
