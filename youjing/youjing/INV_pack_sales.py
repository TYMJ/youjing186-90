import asyncio
from concurrent.futures import ThreadPoolExecutor
from any import *  
from .model import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors, Alignment
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
import traceback
import io # 需要导入 io 模块来处理内存流
from openpyxl.utils import coordinate_to_tuple
from email import errors
from math import e
from pdb import run
from webbrowser import get
from any import *
from .model import *
from sqlalchemy.sql import func,not_,or_,and_
import time
import json,os
from .__default__ import get_user_path,module_xxck_new,module_share_new
from openpyxl.drawing.image import Image as Image_Get
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, Alignment
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from PIL import Image as PILImage  # 重命名避免冲突
# jpeg图片出错的解决方法
from PIL import JpegImagePlugin
# 第2，3，4工作表英文品名部分添加未完成（表tspmb没有）

# 创建线程池执行器
# executor = ThreadPoolExecutor(max_workers=4)
# 获取事件循环
# loop = asyncio.get_event_loop()
def safe_write(ws, coord, value, num_format=None):
    """
    智能穿透写入：如果目标单元格被合并了，自动找到该合并区域的左上角主单元格进行写入。
    完美解决 MergedCell is read-only 报错！
    """
    row, col = coordinate_to_tuple(coord)
    target_cell = ws[coord]
    
    # 检查该坐标是否在某个合并单元格范围内
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            # 如果被合并了，就把目标切换为该区域的左上角真实单元格
            target_cell = ws.cell(row=mr.min_row, column=mr.min_col)
            break
            
    target_cell.value = value
    if num_format:
        target_cell.number_format = num_format
def offset_img(img, col, row, x_pad=4, y_pad=25):
    """精确设置图片位置，偏移量以万为单位进行微调吧，具体计算公式太麻烦了
    row column 的索引都是从0开始的，我这里要把图片插入到单元格A17
    """
    p2e = pixels_to_EMU
    h, w = img.height, img.width
    # 图像等比例缩放因子
    resize_factor = 0.8
    w_h_ratio = w/h
    resize_H = int(resize_factor * h)
    resize_W = int(resize_factor * resize_H * w_h_ratio)
    #
    # x_pad = 4
    # y_pad = 25
    # 注意这里的行、列索引从0开始, 所以需要减1
    size = XDRPositiveSize2D(p2e(w), p2e(h))
    marker = AnchorMarker(col=col-1, colOff=pixels_to_EMU(x_pad), row=row-1, rowOff=pixels_to_EMU(y_pad))
    img.anchor = OneCellAnchor(_from=marker, ext=size)


# --- 核心逻辑函数 ---
def _generate_excel_logic(recordset):
#     """
#     生成 Excel 并返回字节流 (Bytes)，而不是保存在服务器上。
#     """

    hr = '2'
    r_path = os.path.join(config.data_upload_path, 'template')
    if hr == '1':
        template_file = 'bestorebips.xlsx'
    else:
        template_file = 'bestorebips1.xlsx'

    fn = os.path.join(r_path, template_file)
    if not os.path.exists(fn):
        return json_result(-1, f"未找到报表模板: {template_file}")
    wb = load_workbook(fn)
    ws = wb.active
   

    
    s = Session()
    try:
        
        # --- 1. 数据处理 ---
        main_info={}
        merged_products = []
        customs_forms = []
        d=s.query(bgmxd).filter(bgmxd.rid==recordset.get('rids')[0]).first()
        if d:
            main_info = alchemy_object_to_dict(d)
        d=s.query(bgmxdhbcp).filter(bgmxdhbcp.pid==recordset.get('rids')[0]).all()
        if d:
            merged_products = alchemy_object_list_to_dict(d)
        
        d=s.query(bgmxdsheet2).filter(bgmxdsheet2.pid==recordset.get('rids')[0]).all()
        if d:
            customs_forms = alchemy_object_list_to_dict(d)
    
        # 检查是否为RMB客户
        if main_info.get('RMBkh', '否') == '是':
            return {'code': -1, 'msg': '此客人为RMB客人，请选RMB格式', 'data': ''}
        
        # 初始化变量
        ls = ''
        hr = '2'
        hbdm = main_info.get('hbdm', '')
        if hbdm == 'USD$':
            hbdm = 'USD'
        
        # 获取报关公司
        gs=''
        gs1 = main_info.get('bggs', '')
        gsyw = ''
        # 数据库查询
        if gs1:
            j = run_sql(f"SELECT  ywmc FROM wfgs WHERE (wfgs ='{main_info.get('bggs', '')}') limit 1")
           
        
        # 处理日期
        invoice_date = main_info.get('fprq', '')
        qsn = invoice_date[:4] if len(invoice_date) >= 4 else ''
        qsy1 = invoice_date[5:7] if len(invoice_date) >= 7 else ''
        qsr = invoice_date[8:10] if len(invoice_date) >= 10 else ''
        
        # 月份转换
        month_map = {
            '01': 'JAN', '02': 'FEB', '03': 'MAR', '04': 'APR',
            '05': 'MAY', '06': 'JUN', '07': 'JUL', '08': 'AUG',
            '09': 'SEP', '10': 'OCT', '11': 'NOV', '12': 'DEC'
        }
        qsy = month_map.get(qsy1, '')
        
        # 初始化统计变量
        a1 = 0
        bgxshb = 0
        bgbgzsl = 0
        bgbgzmz = 0
        bgbgzjz = 0
        bgbgztj = 0
        bgbgzje = 0
        i6 = 0
        bgdmzhj = 0
        bgdjzhj = 0
        
        # 处理合并产品数据
        for product in merged_products:
            customer_name = main_info.get('khmc', '')
            if 'BEST PRICE LLC' in customer_name:
                bgdmzhj += round(product.get('zmz', '0') * 10) / 10
                bgdjzhj += round(product.get('zjz', '0') * 10) / 10
            else:
                bgdmzhj += round(product.get('zmz', '0') * 100) / 100
                bgdjzhj += round(product.get('zjz', '0') * 100) / 100
            
            bgxshb += float(product.get('chxs', '0'))
            bgbgzsl += product.get('chsl', '0')
            bgbgzmz += product.get('zmz', '0')
            bgbgzjz += product.get('zjz', '0')
            bgbgztj += product.get('ztj', '0')
            bgbgzje += product.get('wxzj', '0')
        
#         # --- 2. 生成 Excel (在内存中) ---
        output = io.BytesIO()

        ws = wb.worksheets[0]    
           
        # ws.title = "报关单"
        i=+1
        cpbh = main_info.get('wfgs', '') + '报关专用章'
        g = run_sql(f"select tpmc from tpzx where (cpbh='{cpbh}') and (LENGTH(tpmc) > 5)")
        if len(g)>0 and g[0].get('tpmc','') != '' and g[0].get('tpmc','') != None and g[0].get('tpmc','') != '[]':
            cpbh = g[0].get('tpmc','')
            Photo = json.loads(str(cpbh))
            if Photo != None:
                file_path = Photo[0]['src']
                fn = os.path.join(config.data_upload_path, str(file_path))
                if (os.path.exists(fn)):
                    img = Image_Get(fn) #选择图片
                    # img.width = 150  # 设置图像宽度
                    # img.height = 68  # 设置图像高度
                    col_width = (ws.column_dimensions['D'].width + ws.column_dimensions['E'].width)*7
                    row_height = 100 # (ws.row_dimensions[i].height)*1.3333
                    img.width=col_width-4 # 转换为像素
                    img.height=row_height-4 # 转换为像素
                    x_offset = 15 # (col_width-img.width)/2
                    y_offset = 10 # (row_height-img.height)/2
                    col = 5 # E列
                    row = 25
                    offset_img(img, col, row, x_offset, y_offset) #col为列位置，row为行位置,x_offset为左边边距,y_offset为上边距
                    ws.add_image(img)  #添加图片
        cpbh = main_info.get('wfgs', '') + '蓝章'
        j = run_sql(f"select sb, dyyy, bgh from tpzx where (cpbh='{cpbh}')")
        # 3. 初始化变量

        gs1 = ""
        sb_val = ""
        dyyy_val = ""
        bgh_val = ""

        # 4. 判断是否有结果
        if j and len(j) > 0:
            # 获取第一行数据
            row = j[0]
            
            # 提取字段值，防止为 None
            sb_val = row.get('sb') or ""
            dyyy_val = row.get('dyyy') or ""
            bgh_val = row.get('bgh') or ""

            # 5. 拼接字符串
            # 注意：这行代码必须缩进在 if 内部，或者确保变量已定义
            gs1 = f"{sb_val}\n{dyyy_val}\n{bgh_val}"

        ws['A4'] = gs1
        ws.row_dimensions[4].height = 39
        ws['A6'] = main_info.get('khmc', '')
        ws['A8'] = gs1
        ws.row_dimensions[8].height = 39
        ws['B4'] = main_info.get('cyka', '')
        ws['B6'] = main_info.get('ysfs', '')
        ws['D6'] = main_info.get('ysgj', '')
        ws['F6'] = main_info.get('tdh', '')
        ws['F8'] = main_info.get('xkzh', '')
        ws['A10'] = main_info.get('ysfp', '')
        ws['B10'] = main_info.get('kagjy', '')
        ws['D10'] = main_info.get('kagjy', '')
        ws['F10'] = main_info.get('mdka', '')
        ws['I10'] = main_info.get('cyka', '')
        ws['E12'] = main_info.get('jgtk', '')
        
        # CNF运费处理
        if main_info.get('jgtk', '') == 'CNF':
            ws['F12'] = merged_products.get('CNFyf$', '')
        
        ws['B12'] = main_info.get('bgxshb', '')
        ws['C12'] = str(bgdmzhj)
        ws['D12'] = str(bgdjzhj)
        ws['A12'] = main_info.get('bzzl', '')
        
        # VGM标记唛码
        vgm_weight = main_info.get('hgzl', 0) + bgdmzhj
        ws['A15'] = f'标记唛码及备注                VGM：{vgm_weight:.2f}KGS'
        
        # 唛头信息
        mark_head = main_info.get('wxmt', '')
        if mark_head:
            ws['N16'] = mark_head
            ws['A16'] = mark_head
        else:
            ws['A16'] = 'N/M'
        
        logger.error('12222222')
        # 填充报关单项信息2
        start_row = 18
        for idx, form in enumerate(customs_forms):
            row_num = start_row + idx + 1
            logger.error('111111111111111111111')
            logger.error(row_num)
            logger.error(form.get('hgbm', ''))
            safe_write(ws, f'A{row_num}', f"{idx + 1} {form.get('hgbm', '')}")
            safe_write(ws, f'B{row_num}', form.get('zwpm', ''))
            
            # 数量信息
            unit = form.get('hgjldw', '')
            if unit != 'KGS' and unit != '千克':
                quantity_str = f"{form.get('chsl', 0):.0f}{unit}/{form.get('zjz', 0)}千克"
            else:
                quantity_str = f"{form.get('chsl', 0):.0f}/{form.get('zjz', 0)}千克"
            safe_write(ws, f'D{row_num}', quantity_str)
            
            safe_write(ws, f'H{row_num}', '中国')
            
            price = form.get('wxzj', 0)
            if price > 0:
                safe_write(ws, f'F{row_num}', price, '$#,##0.000')
            else:
                safe_write(ws, f'F{row_num}', '')
            safe_write(ws, f'J{row_num}', form.get('hyd', ''))
            safe_write(ws, f'K{row_num}', main_info.get('zmxz', ''))
        
        # 底部信息
        final_row = start_row + len(customs_forms) + 3
        container_info = f"箱/封号：{main_info.get('jzxh', '')}/{main_info.get('fh', '')}"
        safe_write(ws, f'B{final_row}', container_info)
        
        bl_no = f"提单号：{main_info.get('tdh', '')}"
        safe_write(ws, f'B{final_row + 1}', bl_no)

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        ws = wb.worksheets[1]       
        # ws.title = "发票"
        cpbh = main_info.get('wfgs', '') + '蓝章'
        logger.error("123shjakdh")
        g = run_sql(f"select tpmc from tpzx where (cpbh='{cpbh}') and (LENGTH(tpmc) > 5)")
        if len(g)>0 and g[0].get('tpmc','') != '' and g[0].get('tpmc','') != None and g[0].get('tpmc','') != '[]':
            cpbh = g[0].get('tpmc','')
            Photo = json.loads(str(cpbh))
            if Photo != None:
                file_path = Photo[0]['src']
                fn = os.path.join(config.data_upload_path, str(file_path))
                if (os.path.exists(fn)):
                    img = Image_Get(fn) #选择图片
                    # img.width = 150  # 设置图像宽度
                    # img.height = 68  # 设置图像高度
                    col_width = (ws.column_dimensions['D'].width + ws.column_dimensions['E'].width)*7
                    row_height = 100 # (ws.row_dimensions[i].height)*1.3333
                    img.width=col_width-4 # 转换为像素
                    img.height=row_height-4 # 转换为像素
                    x_offset = 15 # (col_width-img.width)/2
                    y_offset = 10 # (row_height-img.height)/2
                    col = 6 # F列
                    row = 48
                    logger.error("123456789djajkh")
                    offset_img(img, col, row, x_offset, y_offset) #col为列位置，row为行位置,x_offset为左边边距,y_offset为上边距
                    ws.add_image(img)  #添加图片
                  



         # 基本信息
        safe_write(ws, f'A6',main_info.get('gsyw', ''))
        safe_write(ws, f'C15', main_info.get('ysfp', ''))
        date_str = f"{qsy}.{qsr},{qsn}"
        safe_write(ws, f'C17', date_str)
        safe_write(ws, f'B18', main_info.get('cyka', ''))
        safe_write(ws, f'B21', main_info.get('mdka', ''))
        safe_write(ws, f'F22', main_info.get('jhfs', ''))
        safe_write(ws, f'E25', f"{main_info.get('jgtk', '')} {main_info.get('cyka', '')}")
        safe_write(ws, f'G25', hbdm)
#           msexcelworksheet.Rows[inttostr(25+k+1)].AutoFit;

#   if recordset.getfielddataasstring('合并产品','英文合计')<>'' then
#   begin
#   msexcelworksheet.Range['B'+inttostr(25+k+1)].Value:=UpperCase(recordset.getfielddataasstring('合并产品','英文合计'));
#   end
#   else
#   begin
#   msexcelworksheet.Range['B'+inttostr(25+k+1)].Value:=UpperCase(recordset.getfielddataasstring('合并产品','英文品名')); 
#   tmpcomt.close;
#  tmpcomt.SQL.Text:='select djpmy,djyxb from tspmb where (djpmy=:djpmy) and (djpmy<>"")';
#  tmpcomt.ParamByName('djpmy').AsString:=recordset.getfielddataasstring('合并产品','英文品名');
#  tmpcomt.open;
#  if tmpcomt.isempty=false then                           
#      begin  
     
  
#      zs2:=0;
#  zs3:=0;
#  zsw:='';
#  zs4:=1;
#  zs6:=0;
#  zs7:=0;
#  zs:=0;
#  zs:=length(tmpcomt.FieldByName('djyxb').asstring);
#      for zs1:=1 to zs do
#      begin 
#     zs2:=zs2+1;
#     zs6:=zs6+1;
#      zsw:=copy(tmpcomt.FieldByName('djyxb').asstring,zs2,1); 
#       if (zsw=',') or (zsw='，') or (zsw='，') or (zsw='，') or (zsw=',') or (zs1=zs) then
#  begin 
#  zcx:=getint(trim(copy(tmpcomt.FieldByName('djyxb').asstring,zs4,zs6-1)));
#   if  (zs1=zs) then
#  begin
#  zcx:=getint(trim(copy(tmpcomt.FieldByName('djyxb').asstring,zs4,zs6)));
#  end;
#   msexcelworksheet.Range['B'+inttostr(25+k+1)].Characters[zcx, 1].Font.Subscript := True;
#  zs4:=zs2+1;
#  zs6:=0;
     
#    end; 
#    end;(英文品名插入b25行，英文合计插入b26行，自动识别下标并设置为字体的下标格式)
  
        # 收货人处理
        consignee = main_info.get('shr', '')
        if not consignee:
            safe_write(ws, f'A10', 'BEST PRICE')
        elif '可思达' in main_info.get('wfgs', ''):
            safe_write(ws, f'A10', main_info.get('khmc', ''))
        else:
            safe_write(ws, f'A10', main_info.get('shry', ''))
        
        # 填充合并产品信息
        start_row = 26
        for idx, product in enumerate(merged_products):
            row_num = start_row + idx + 1
            
            # 英文品名
            english_desc = product.get('ywhj', '') or product.get('ywpm', '')
            if english_desc:
                safe_write(ws, f'B{row_num}', english_desc)
            
            # 其他信息
            safe_write(ws, f'C{row_num}', product.get('chsl', 0))
            safe_write(ws, f'D{row_num}', f"{product.get('jldw', '')}S")
            safe_write(ws, f'E{row_num}', hbdm)
            safe_write(ws, f'F{row_num}', f"{product.get('price', 0):.3f}")
            safe_write(ws, f'G{row_num}', hbdm)
            safe_write(ws, f'H{row_num}', product.get('wxzj', 0))
        
        # 唛头信息合并
        mark_head = main_info.get('wxmt', '')
        if mark_head:
            ws.merge_cells(f'A27:A{start_row + len(merged_products) + 1}')
            ws['A27'] = mark_head
        else:
            ws.merge_cells(f'A27:A{start_row + len(merged_products) + 1}')
            ws['A27'] = 'N/M'

        ws = wb.worksheets[2]       
        # ws.title = "装箱单"
        cpbh = main_info.get('wfgs', '') + '蓝章'
        g = run_sql(f"select tpmc from tpzx where (cpbh='{cpbh}') and (LENGTH(tpmc) > 5)")
        if len(g)>0 and g[0].get('tpmc','') != '' and g[0].get('tpmc','') != None and g[0].get('tpmc','') != '[]':
                cpbh = g[0].get('tpmc','')
                Photo = json.loads(str(cpbh))
                if Photo != None:
                    file_path = Photo[0]['src']
                    fn = os.path.join(config.data_upload_path, str(file_path))
                    if (os.path.exists(fn)):
                        img = Image_Get(fn) #选择图片
                        # img.width = 150  # 设置图像宽度
                        # img.height = 68  # 设置图像高度
                        col_width = (ws.column_dimensions['D'].width + ws.column_dimensions['E'].width)*7
                        row_height = 100 # (ws.row_dimensions[i].height)*1.3333
                        img.width=col_width-4 # 转换为像素
                        img.height=row_height-4 # 转换为像素
                        x_offset = 15 # (col_width-img.width)/2
                        y_offset = 10 # (row_height-img.height)/2
                        col = 4 # F列
                        row = 52
                        offset_img(img, col, row, x_offset, y_offset) #col为列位置，row为行位置,x_offset为左边边距,y_offset为上边距
                        ws.add_image(img)  #添加图片
               
     # 基本信息
        ws['A6'] = gsyw
        ws['C15'] = main_info.get('ysfp', '')
        date_str = f"{qsy}.{qsr},{qsn}"
        ws['C17'] = date_str
        ws['B18'] = main_info.get('cyka', '')
        ws['B21'] = main_info.get('mdka', '')
        ws['D22'] = main_info.get('jhfs', '')
#           msexcelworksheet.Rows[inttostr(25+k+1)].AutoFit;

#   IF recordset.getfielddataasstring('合并产品','英文合计')<>'' THEN
#   BEGIN
#   msexcelworksheet.Range['B'+inttostr(25+k+1)].Value:=UpperCase(recordset.getfielddataasstring('合并产品','英文合计'));
#   END
#   ELSE
#   BEGIN
#    msexcelworksheet.Range['B'+inttostr(25+k+1)].Value:=UpperCase(recordset.getfielddataasstring('合并产品','英文品名'));
#      msexcelworksheet.Range['B'+inttostr(25+k+1)].Value:=UpperCase(recordset.getfielddataasstring('合并产品','英文品名')); 
#   tmpcomt.close;
#  tmpcomt.SQL.Text:='select djpmy,djyxb from tspmb where (djpmy=:djpmy) and (djpmy<>"")';
#  tmpcomt.ParamByName('djpmy').AsString:=recordset.getfielddataasstring('合并产品','英文品名');
#  tmpcomt.open;（英文品名插入b25行，英文合计插入b26行，自动识别下标并设置为字体的下标格式）
        # 收货人处理
        consignee = main_info.get('shr', '')
        if not consignee:
            safe_write(ws, f'A10', 'BEST PRICE')
        elif '可思达' in main_info.get('wfgs', ''):
            safe_write(ws, f'A10', main_info.get('khmc', ''))
        else:
            safe_write(ws, f'A10', main_info.get('shry', ''))
        
        # 填充合并产品信息
        start_row = 26
        for idx, product in enumerate(merged_products):
            row_num = start_row + idx + 1
            
            # 英文品名
            english_desc = product.get('ywhj', '') or product.get('ywpm', '')
            if english_desc:
                safe_write(ws, f'B{row_num}', english_desc)
            
            # 数量和重量信息
            safe_write(ws, f'C{row_num}', product.get('chxs', 0))
            
            # 按客户类型处理重量精度
            customer_name = main_info.get('khmc', '')
            if 'BEST PRICE LLC' in customer_name:
                gross_weight = round(product.get('zmz', 0) * 10) / 10
                net_weight = round(product.get('zjz', 0) * 10) / 10
            else:
                gross_weight = round(product.get('zmz', 0) * 100) / 100
                net_weight = round(product.get('zjz', 0) * 100) / 100
            
            safe_write(ws, f'D{row_num}', gross_weight)
            safe_write(ws, f'E{row_num}', net_weight)
            safe_write(ws, f'F{row_num}', product.get('ztj', 0))
        
        # 唛头信息合并
        mark_head = main_info.get('wxmt', '')
        if mark_head:
            ws.merge_cells(f'A27:A{start_row + len(merged_products) + 1}')
            ws['A27'] = mark_head
        else:
            ws.merge_cells(f'A27:A{start_row + len(merged_products) + 1}')
            ws['A27'] = 'N/M'

        ws = wb.worksheets[3]
#           msexcelworksheet.Rows[inttostr(9+k)].AutoFit;
#   IF recordset.getfielddataasstring('合并产品','英文合计')<>'' THEN
#   BEGIN 
#   msexcelworksheet.Range['A'+inttostr(9+k)].Value:=UpperCase(recordset.getfielddataasstring('合并产品','英文合计'));
#   msexcelworksheet.Range['M'+inttostr(9+k)].Value:=UpperCase(recordset.getfielddataasstring('合并产品','英文合计'));
#   END
#   ELSE
#   BEGIN 
#    msexcelworksheet.Range['A'+inttostr(9+k)].Value:=UpperCase(recordset.getfielddataasstring('合并产品','英文品名'));
#    msexcelworksheet.Range['M'+inttostr(9+k)].Value:=UpperCase(recordset.getfielddataasstring('合并产品','英文品名')); 
#      msexcelworksheet.Range['B'+inttostr(25+k+1)].Value:=UpperCase(recordset.getfielddataasstring('合并产品','英文品名')); 
#   tmpcomt.close;
#  tmpcomt.SQL.Text:='select djpmy,djyxb from tspmb where (djpmy=:djpmy) and (djpmy<>"")';
#  tmpcomt.ParamByName('djpmy').AsString:=recordset.getfielddataasstring('合并产品','英文品名');
#  tmpcomt.open;(英文品名插入a9行，英文合计插入a10行，自动识别下标并设置为字体的下标格式)

        cpbh =main_info.get('khmc', '') 
        g = run_sql(f"select tpmc from kh where (company_name='{main_info.get('company_name', '')}') and (LENGTH(tpmc) > 5)")
        if len(g)>0 and g[0].get('tpmc','') != '' and g[0].get('tpmc','') != None and g[0].get('tpmc','') != '[]':
                cpbh = g[0].get('tpmc','')
                Photo = json.loads(str(cpbh))
                if Photo != None:
                    file_path = Photo[0]['src']
                    fn = os.path.join(config.data_upload_path, str(file_path))
                    if (os.path.exists(fn)):
                        img = Image_Get(fn) #选择图片
                        # img.width = 150  # 设置图像宽度
                        # img.height = 68  # 设置图像高度
                        col_width = (ws.column_dimensions['D'].width + ws.column_dimensions['E'].width)*7
                        row_height = 100 # (ws.row_dimensions[i].height)*1.3333
                        img.width=col_width-4 # 转换为像素
                        img.height=row_height-4 # 转换为像素
                        x_offset = 15 # (col_width-img.width)/2
                        y_offset = 10 # (row_height-img.height)/2
                        col = 2# F列
                        row = 31
                        offset_img(img, col, row, x_offset, y_offset) #col为列位置，row为行位置,x_offset为左边边距,y_offset为上边距
                        ws.add_image(img)  #添加图片
        cpbh = main_info.get('wfgs', '') + '蓝章'
        g = run_sql(f"select tpmc from tpzx where (cpbh='{cpbh}') and (LENGTH(tpmc) > 5)")
        if len(g)>0 and g[0].get('tpmc','') != '' and g[0].get('tpmc','') != None and g[0].get('tpmc','') != '[]':
                cpbh = g[0].get('tpmc','')
                Photo = json.loads(str(cpbh))
                if Photo != None:
                    file_path = Photo[0]['src']
                    fn = os.path.join(config.data_upload_path, str(file_path))
                    if (os.path.exists(fn)):
                        img = Image_Get(fn) #选择图片
                        # img.width = 150  # 设置图像宽度
                        # img.height = 68  # 设置图像高度
                        col_width = (ws.column_dimensions['D'].width + ws.column_dimensions['E'].width)*7
                        row_height = 100 # (ws.row_dimensions[i].height)*1.3333
                        img.width=col_width-4 # 转换为像素
                        img.height=row_height-4 # 转换为像素
                        x_offset = 15 # (col_width-img.width)/2
                        y_offset = 10 # (row_height-img.height)/2
                        col = 9 # F列
                        row = 31
                        offset_img(img, col, row, x_offset, y_offset) #col为列位置，row为行位置,x_offset为左边边距,y_offset为上边距
                        ws.add_image(img)  #添加图片



        ws['I2'] = main_info.get('ysfp', '')
        date_str = f"{qsy}.{qsr},{qsn}"
        ws['I3'] = date_str
        ws['G9'] = f"{main_info.get('jgtk', '')} {main_info.get('cyka', '')}"
        
        # 收货人处理
        consignee = main_info.get('shr', '')
        if not consignee:
            ws['C6'] = 'BEST PRICE'
        elif '可思达' in main_info.get('wfgs', ''):
            ws['C6'] = main_info.get('khmc', '')
        else:
            ws['C6'] = main_info.get('shry', '')
        
        # 填充合并产品信息
        start_row = 9
        for idx, product in enumerate(merged_products):
            row_num = start_row + idx + 1
            
            # 英文品名
            english_desc = product.get('ywhj', '') or product.get('ywpm', '')
            if english_desc:
                safe_write(ws, f'A{row_num}', english_desc)
                safe_write(ws, f'M{row_num}', english_desc)
            else:
                safe_write(ws, f'A{row_num}', product.get('ywpm', ''))
                safe_write(ws, f'M{row_num}', product.get('ywpm', ''))
            
            # 其他信息
            safe_write(ws, f'E{row_num}', product.get('chsl', 0))
            safe_write(ws, f'F{row_num}', f"{product.get('jldw', '')}S")
            safe_write(ws, f'G{row_num}', hbdm)
            safe_write(ws, f'H{row_num}', f"{product.get('price', 0):.3f}")
            safe_write(ws, f'I{row_num}', hbdm)
            safe_write(ws, f'J{row_num}', product.get('wxzj', 0))
        
        if row_num<= 29:
        
            start_row = 29
        elif row_num > 29: 
            start_row =  start_row + idx + 1
        final_row = start_row 
        logger.error("Final row for bottom info: " + str( len(merged_products))) 
        logger.error( str(len(merged_products)))
        safe_write(ws, f'I{final_row + 1}', hbdm)
        safe_write(ws, f'D{final_row + 3}', main_info.get('zyqx', ''))
        safe_write(ws, f'E{final_row + 4}', main_info.get('cyka', ''))
        safe_write(ws, f'I{final_row + 4}', main_info.get('mdka', ''))
        safe_write(ws, f'A{final_row + 8 }', main_info.get('jhfs', ''))
        
        # 唛头和备注信息
        mark_head = main_info.get('wxmt', '')
        if mark_head:
            safe_write(ws, f'D{final_row + 9}', mark_head)
        else:
            safe_write(ws, f'D{final_row + 9}', 'N/M')
        
        safe_write(ws, f'C{final_row + 11}', main_info.get('bza', ''))


        output.close()
        path = config.tmp_path
        os.makedirs(path, exist_ok=True) # 确保目录存在
        report_rid = get_uuid() # 假设这是一个生成唯一ID的函数
        output_filename = f'优景_INV_pack_sales.xlsx'
        full_output_path = os.path.join(path, output_filename)
        s_path = config.tmp_path
        # file_name = f"{master.fphm or 'Unknown'}优景_INV_pack_sales.xlsx"
        # wb.save(os.path.join(s_path, file_name))
        # return json_result(1, '导出成功', file_name)

        
        wb.save(full_output_path)
        

        return {'code': 1, 'msg': '报表生成成功', 'data': output_filename}

    except Exception as e:
        logger.error(f"生成Excel失败: {str(e)}")
        logger.error(traceback.format_exc())
        return {'code': -1, 'msg': f'生成Excel失败: {str(e)}', 'data': ''}
    finally:
        s.close()

# --- API路由 ---
@any_route('/api/saier/export/detailed/INVpacksales/list1', methods=['POST'])
@require_token
async def inv_pack_sales_export(request):
    """
    导出INV pack sales详细数据
    """
    j = await request.json()
    # any_thead_run(inv_pack_sales_export, j, request)
    try:
        result = _generate_excel_logic(j)
        logger.error(result)
        if result['code'] == 1:
            return json_result(1, result['msg'], result['data'])
        else:
            return json_result(result['code'], result['msg'], {})
    
    except Exception as e:
        logger.error(f"API调用失败: {str(e)}")
        logger.error(traceback.format_exc())
        return json_result(-1, f'API调用失败: {str(e)}', {})
    
    