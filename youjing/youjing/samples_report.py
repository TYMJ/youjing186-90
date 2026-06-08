import asyncio
from concurrent.futures import ThreadPoolExecutor
from any import *  
from .model import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors, Alignment
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
import traceback
import io # 需要导入 io 模块来处理内存流
from openpyxl.utils import coordinate_to_tuple
from email import errors
from math import e
from pdb import run
from webbrowser import get
from any import *
from .model import *
from sqlalchemy.sql import func,not_,or_,and_
import time
import json,os
from .__default__ import get_user_path,module_xxck_new,module_share_new
from openpyxl.drawing.image import Image as Image_Get
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, Alignment
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from PIL import Image as PILImage  # 重命名避免冲突
# jpeg图片出错的解决方法
from PIL import JpegImagePlugin
# 第2，3，4工作表英文品名部分添加未完成（表tspmb没有）
# 创建线程池执行器
# executor = ThreadPoolExecutor(max_workers=4)
# 获取事件循环
# loop = asyncio.get_event_loop()
import copy
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries, coordinate_to_tuple
import base64
import tempfile
import uuid
import re
from concurrent.futures import ThreadPoolExecutor
import threading


def clean_filename(filename: str) -> str:
	"""清理文件名中的非法字符（Windows）。"""
	if not filename:
		return ''
	return re.sub(r'[<>:"/\\|?*\r\n]', '', str(filename))


# --- 辅助函数（与 ST_INV_pack_sales.py 风格一致） ---
def safe_write(ws, coord, value, num_format=None):
	row, col = coordinate_to_tuple(coord)
	target_cell = ws[coord]
	for mr in ws.merged_cells.ranges:
		if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
			target_cell = ws.cell(row=mr.min_row, column=mr.min_col)
			break
	target_cell.value = value
	if num_format and hasattr(target_cell, 'number_format'):
		target_cell.number_format = num_format


def offset_img(img, col, row, x_pad=4, y_pad=25):
	p2e = pixels_to_EMU
	h, w = img.height, img.width
	size = XDRPositiveSize2D(p2e(w), p2e(h))
	marker = AnchorMarker(col=col-1, colOff=pixels_to_EMU(x_pad), row=row-1, rowOff=pixels_to_EMU(y_pad))
	img.anchor = OneCellAnchor(_from=marker, ext=size)


def makerealfile_from_base64(b64text, dest_path):
	"""把可能的 base64 文本写成二进制文件并返回是否成功。"""
	try:
		# 若是 bytes，直接写入
		if isinstance(b64text, (bytes, bytearray)):
			data = bytes(b64text)
		else:
			# 尝试解码 base64；若失败，作为原始文本写入
			try:
				data = base64.b64decode(b64text)
			except Exception:
				data = str(b64text).encode('utf-8')
		with open(dest_path, 'wb') as f:
			f.write(data)
		return True
	except Exception as e:
		logger.error(f"写临时图片文件失败: {e}")
		return False


# --- 小范围转换样例：生成样品导出（仅写表头；尝试插入第一条记录图片） ---
def generate_samples_report(u_path, record_numbers: list, region: str = None, progress_hook=None):
	"""示例：按 Delphi 脚本风格生成 Excel，写入表头并尝试插入记录并插图。
	参数:
		- record_numbers: 要导出的记录编号列表
		- region: 可选地区过滤，支持 '宁波' 或 '义乌'，否则为全部/我的公司
		- progress_hook: 可选回调函数 progress_hook(percent:int)
	"""
	# logger.error(f"开始生成样品报表，记录编号: {record_numbers}, 地区: {region}, 用户路径: {u_path}")
	try:
		
		# 尝试从模板加载空白.xls（与 Delphi 路径相近），否则回退到新建 Workbook
		template_candidates = []
		try:
			template_candidates.append(os.path.join(config.data_upload_path, 'data', 'saier20050822', 'addonfiles', '空白.xlsx'))
		except Exception:
			pass
		try:
			template_candidates.append(os.path.join(config.data_upload_path, 'template', '空白.xlsx'))
		except Exception:
			pass
		template_candidates.append(os.path.join(os.getcwd(), '空白.xlsx'))

		wb = None
		found_template = None
		for fn in template_candidates:
			try:
				if fn and os.path.exists(fn):
					wb = load_workbook(fn)
					found_template = fn
					break
			except Exception:
				continue

		# 如果没有找到模板，按原 Delphi 行为直接返回错误（不要回退到新 workbook）
		if wb is None:
			# logger.error(f"未找到报表模板，候选路径: {template_candidates}")
			return {'code': -1, 'msg': '未找到报表模板，无法生成报表', 'data': ''}

		ws = wb.active

		# 记住是否使用了模板文件（如果使用模板，通常已有样式）
		template_used = True if wb and any(os.path.exists(p) for p in template_candidates) and wb is not None else False

		# 按 Delphi 脚本精确设置 A..Q 列的标题与列宽、居中
		cols = list('ABCDEFGHIJKLMNOPQ')
		headers = [
			'产品编号', '产品图片', '产品分类', '产品规格', '中文品名', '英文品名', '计量单位',
			'报关品名', '外箱容量', '退 税 率', '起 订 量', '外箱体积', '中文包装', '英文包装',
			'材    质', '颜    色', '所有者'
		]
		for c, h in zip(cols, headers):
			try:
				ws.column_dimensions[c].width = 14.3
			except Exception:
				pass
			# 使用 safe_write 以保持合并单元格写入安全
			safe_write(ws, f'{c}1', h)
			# 居中对齐
			cell = ws[f'{c}1']
			try:
				cell.alignment = Alignment(vertical='center', horizontal='center')
			except Exception:
				pass

		# 如果没有模板，应用统一的表头样式（粗体、填充、边框）
		if not template_used:
			from openpyxl.styles import PatternFill
			header_fill = PatternFill(fill_type='solid', start_color='FFF2F2F2', end_color='FFF2F2F2')
			header_font = Font(bold=True, size=11)
			thin = Side(style='thin')
			for c in cols:
				cell = ws[f'{c}1']
				try:
					cell.font = header_font
					cell.fill = header_fill
					cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
				except Exception:
					pass

		# 处理地区过滤（da3/da4）
		if region == '宁波':
			da3 = '优景宁波'
			da4 = '优景外销'
		elif region == '义乌':
			da3 = '优景义乌'
			da4 = '优景义乌'
		else:
			da3 = '我的公司'
			da4 = '我的公司'

		# 权限检查：尽量模拟 Delphi 中的 sys_users(position) 包含 '样品间' 的校验
		try:
			current_name = None
			try:
				current_name = IApplication.LoginInfo.userinfo.name
			except Exception:
				# 退回到可能的环境变量或 config
				current_name = getattr(config, 'current_user', None) or os.environ.get('USERNAME') or os.environ.get('USER')
			if current_name:
				j = run_sql(f"select position from sys_users where name='{current_name}'")
				if not j or len(j) == 0 or '样品间' not in (j[0].get('position') or ''):
					return {'code': -1, 'msg': '当前用户无样品间导出权限', 'data': ''}
		except Exception:
			# 权限检查失败时不阻塞导出，但记录日志
			logger.warning('无法确定当前用户或校验权限时出错，跳过权限校验')

		# 主循环：遍历 record_numbers，按原 Delphi 脚本逻辑填充行并插入图片
		total = len(record_numbers) if record_numbers else 0
		i1 = (100.0 / total) if total > 0 else 100.0
		if progress_hook:
			try:
				progress_hook(0)
			except Exception:
				pass
		s = Session()
		exported_cpbh_list = []
		try:
			k = 1
			tmpdir = getattr(config, 'tmp_path', tempfile.gettempdir())
			os.makedirs(tmpdir, exist_ok=True)

			# 支持批量记录（等同于原 tmpstr 列表）
			for idx, num in enumerate(record_numbers):
				k += 1
				# 带地区过滤的查询
				logger.error(f"正在查询记录 {idx+1}/{total}，编号: {da3}，用户路径: {u_path}")
				if da3 in u_path or da4 in u_path:
					q = f"""
						select
							cpbh, cpfl, cpgg, zwpm, ywpm, jldw, bgpm,
							bzrl, tsl, zxcgl, bztj, zhwbzh, bzhfsh,
							caizi, yse, ccwz, cdmc, uid, rid
						from ypgl
						where rid = '{num}'
					"""
					rows = run_sql(q)
				# logger.error(f"查询记录:{rows}")
				if not rows or len(rows) == 0:
					continue
				row = rows[0]
				# 列赋值（简化版，保持字段对应）
				cpbh = row['cpbh']
				if cpbh:
					exported_cpbh_list.append(str(cpbh))
				ws[f'A{k}'] = cpbh
				ws[f'C{k}'] = row.get('cpfl', '')
				ws[f'D{k}'] = row.get('cpgg', '')
				ws[f'E{k}'] = row.get('zwpm', '')
				ws[f'F{k}'] = row.get('ywpm', '')
				ws[f'G{k}'] = row.get('jldw', '')
				ws[f'H{k}'] = row.get('bgpm', '')
				# 数值列与格式化（使用 safe_write 以兼容合并单元格）
				try:
					bzrl_val = int(row.get('bzrl') or 0)
				except Exception:
					bzrl_val = 0
				safe_write(ws, f'I{k}', bzrl_val, num_format='0')

				try:
					tsl_val = float(row.get('tsl') or 0)
				except Exception:
					tsl_val = 0

				safe_write(ws, f'J{k}', tsl_val, num_format='0.000')

				safe_write(ws, f'K{k}', row.get('zxcgl') or 0, num_format='0.00')
				safe_write(ws, f'L{k}', row.get('bztj') or 0, num_format='0.00')

				safe_write(ws, f'M{k}', row.get('zhwbzh', ''))
				safe_write(ws, f'N{k}', row.get('bzhfsh', ''))
				safe_write(ws, f'O{k}', row.get('caizi', ''))
				safe_write(ws, f'P{k}', row.get('yse', ''))
				safe_write(ws, f'Q{k}', row.get('uid', ''))

				# 设置对齐与换行，Delphi 原脚本对 A..Y 范围做了设置
				for col_letter in list('ABCDEFGHIJKLMNOPQRSTUVWXY'):
					try:
						cell = ws[f'{col_letter}{k}']
						cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
						# 应用细线边框以贴近 Delphi 风格
						thin = Side(style='thin')
						cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
					except Exception:
						# 如果单元格不存在则忽略
						pass
				ws.row_dimensions[k].height = 24

				# 查询图片并插入（接近 Delphi：偶数/奇数行略有不同的定位）
				try:
					logger.error(f"正在处理记录 {idx+1}/{total}，rid: {row['rid']}, cpbh: {cpbh}")
					# img_rows = run_sql(f"select yytp from ypglsheet1 where pid='{row['rid']}' order by rid desc limit 1")
					# if img_rows and len(img_rows) > 0:
					# 	img_row = img_rows[0]
					# 	if img_row.get('yytp'):
					# 		tmp_name = os.path.join(tmpdir, f"{cpbh}_{uuid.uuid4().hex}.jpg")
					# 		ok = makerealfile_from_base64(img_row.get('yytp'), tmp_name)
					# 	if ok and os.path.exists(tmp_name):
					# 		try:
					# 			ws.row_dimensions[k].height = 80
					# 			img = Image_Get(tmp_name)

					# 			# 计算近似像素宽高（openpyxl 列宽单位近似乘7）
					# 			try:
					# 				fwidth = (ws.column_dimensions['B'].width or 14.3) * 7 - 5
					# 			except Exception:
					# 				fwidth = 100
					# 			try:
					# 				fheight = ws.row_dimensions[k].height - 5
					# 			except Exception:
					# 				fheight = 80

					# 			# 区分偶数/奇数（Delphi 使用 i mod 2）
					# 			if (idx % 2) == 0:
					# 				# 偶数行：简单插入并等比缩放到单元格大小
					# 				img.width = int(fwidth)
					# 				img.height = int(fheight)
					# 				offset_img(img, col=2, row=k, x_pad=2, y_pad=2)
					# 				ws.add_image(img)
					# 			else:
					# 				# 奇数行：插入并稍微偏移，模仿 Delphi 左/上偏移
					# 				img.width = int(fwidth)
					# 				img.height = int(fheight)
					# 				offset_img(img, col=2, row=k, x_pad=4, y_pad=4)
					# 				ws.add_image(img)
					# 		finally:
					# 			try:
					# 				os.remove(tmp_name)
					# 			except Exception:
					# 				pass
					g = run_sql(f"select yytp from ypglsheet1 where pid='{row['rid']}' order by rid desc limit 1")
					if len(g)>0 and g[0].get('yytp','') != '' and g[0].get('yytp','') != None and g[0].get('yytp','') != '[]':
						cpbh = g[0].get('yytp','')
						Photo = json.loads(str(cpbh))
						if Photo != None:
							file_path = Photo[0]['src']
							fn = os.path.join(config.data_upload_path, str(file_path))
							if (os.path.exists(fn)):
								img = Image_Get(fn) #选择图片
								# img.width = 150  # 设置图像宽度
								# img.height = 68  # 设置图像高度
								col_width = 30
								row_height = 30  # (msexcelworksheet.row_dimensions[cell].height)*1.5
								img.width=col_width-4 # 转换为像素
								img.height=row_height-4 # 转换为像素
								x_offset = 15 # (col_width-img.width)/2
								y_offset = 5 # (row_height-img.height)/2
								col = 2  
								row = k
								offset_img(img, col, row, x_offset, y_offset) #col为列位置，row为行位置,x_offset为左边边距,y_offset为上边距
								ws.add_image(img)  #添加图片
				except Exception:
					logger.error('图片插入失败：' + traceback.format_exc())

				# 更新进度
				try:
					if progress_hook:
						progress_hook(min(100, int(round((idx + 1) * i1))))
					else:
						logger.info(f"samples_report progress: {min(100, int(round((idx + 1) * i1)))}%")
				except Exception:
					pass

			# 结束遍历
			# 最终进度
			try:
				if progress_hook:
					progress_hook(100)
			except Exception:
				pass
		finally:
			s.close()


		# 将工作簿保存到临时目录（与 jygl_report 一致的 output_filename 方式）
		path = getattr(config, 'tmp_path', tempfile.gettempdir())
		os.makedirs(path, exist_ok=True)
		if exported_cpbh_list:
			cpbh_part = clean_filename(exported_cpbh_list[0])
		else:
			cpbh_part = uuid.uuid4().hex
		output_filename = f'{cpbh_part}样品管理(图).xlsx'
		full_output_path = os.path.join(path, output_filename)
		wb.save(full_output_path)
		return {
			'code': 1,
			'msg': '示例报表生成成功',
			'data': {
				'filename': output_filename,
				'full_path': full_output_path,
			}
		}

	except Exception as e:
		logger.error(traceback.format_exc())
		return {'code': -1, 'msg': str(e), 'data': ''}


def _make_progress_hook(rid: str):
	# logger.error(f"创建进度回调函数，rid: {rid}")
	def hook(percent: int):
		# logger.error(f"进度回调被调用，rid: {rid}, percent: {percent}")
		try:
			with export_progress_lock:
				st = export_progress.get(rid, {})
				st['percent'] = int(percent)
				st['status'] = 'running' if percent < 100 else 'finalizing'
				export_progress[rid] = st
		except Exception:
			logger.error('更新进度失败')
	return hook


def _run_export_job(u_path, rid: str, record_numbers, region: str):
	# logger.error(f"开始执行导出任务，1")
	try:
		hook = _make_progress_hook(rid)
		res = generate_samples_report(u_path, record_numbers, region=region, progress_hook=hook)
		with export_progress_lock:
			st = export_progress.get(rid, {})
			if res and res.get('code') == 1:
				st['percent'] = 100
				st['status'] = 'done'
				data = res.get('data', '')
				st['path'] = data.get('filename', '') if isinstance(data, dict) else data
				st['msg'] = res.get('msg', '')
			else:
				st['percent'] = st.get('percent', 0)
				st['status'] = 'error'
				st['msg'] = res.get('msg', '') if isinstance(res, dict) else str(res)
			export_progress[rid] = st
			logger.error(f'更新导出进度: {export_progress[rid]}')
		# 返回生成的文件名（如果成功）
		try:
			return st.get('path', '')
		except Exception:
			return ''
	except Exception as e:
		logger.error(traceback.format_exc())
		with export_progress_lock:
			export_progress[rid] = {'percent': 0, 'status': 'error', 'path': '', 'msg': str(e)}
		return ''

# ---------- 前端进度 API 集成 ----------
# 全局进度存储: rid -> {percent, status, path, msg}
export_progress = {}
export_progress_lock = threading.Lock()

# 线程池用于后台运行导出任务
_export_executor = ThreadPoolExecutor(max_workers=2)

@any_route('/api/saier/samples/export_current_product', methods=['POST'])
@require_token
async def samples_export(request):
	"""
	前端传参示例:
	{rids: "6629eedb-ecf4-0987-36a0-b6fcec0ee30d", region: "宁波"}
	"""
	j = await request.json()
	# 获取当前登录用户的path
	user = request.current_user
	org = get_user_path(user.username)
	u_path = org.get('path', '')
	# logger.error(f"收到导出请求，用户路径: {u_path}")
	try:
		rids = j.get('rids', '')
		region = j.get('region', '') or None  # 空字符串视为全部
		if not rids:
			return json_result(-1, '缺少记录编号参数 rids', '')
		# 前端可能传单个 rid 字符串，也可能传数组
		if isinstance(rids, str):
			record_numbers = [rids.strip()] if rids.strip() else []
		elif isinstance(rids, list):
			record_numbers = [str(x).strip() for x in rids if x]
		else:
			record_numbers = [str(rids).strip()] if rids else []
		if not record_numbers:
			return json_result(-1, '缺少记录编号参数 rids', '')
		# 导出任务进度用的 job id（与业务记录 rid 区分）
		job_rid = uuid.uuid4().hex
		# 初始化进度状态
		with export_progress_lock:
			export_progress[job_rid] = {'percent': 0, 'status': 'queued', 'path': '', 'msg': ''}
		# 提交后台任务并等待完成（阻塞，按方案 A）
		future = _export_executor.submit(_run_export_job, u_path, job_rid, record_numbers, region)
		try:
			outname = future.result()
		except Exception:
			logger.error(traceback.format_exc())
			outname = ''
		if not outname:
			return json_result(-1, '导出失败', '')
		return json_result(1, '导出完成', {'filename': outname})
	except Exception as e:
		logger.error(traceback.format_exc())
		return json_result(-1, f'API调用失败: {str(e)}', {})