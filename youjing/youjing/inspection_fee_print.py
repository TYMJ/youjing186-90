import time, json, os, math, csv, tempfile, shutil, zipfile, re, subprocess, sys
from any import *
from .model import *
from .__default__ import get_uuid
from sqlalchemy.sql import func, not_, or_, and_
try:
    import cn2an
    from openpyxl import load_workbook
    from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle, PatternFill, Color
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "cn2an"])
    import cn2an
    from openpyxl import load_workbook
    from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle, PatternFill, Color

SYS_FIELDS = [
    "sid",
    "rid",
    "uid",
    "ctime",
    "mtime",
    "has_att",
    "modi_uid",
    "wf_status",
    "wf_unit",
    "pid",
    "archived",
]


def process_hdfy_data(rid, i4, bgpm, gcmc, rmbsb, mjsb):
    # 货代信息
    try:
        d = run_sql(
            f"select cdmc,rid,zdhd,zrhd,bghd,bank1,zh1,bank2,zh2,bank3,zh3,sbje,sbjem,sjhd \
                from hdfy \
                where (rid='{rid}') and ((sbdy='') or (sbdy is null))"
        )
        if len(d) > 0:
            for row in d:
                i4 += 1
                bgpm.append(row.get("rid"))
                gcmc = str(row.get("sjhd", gcmc)).strip()
                rmbsb = "1" if row.get("sbje", 0) > 0 else rmbsb
                mjsb = "1" if row.get("sbjem", 0) > 0 else mjsb
        return i4, bgpm, gcmc, rmbsb, mjsb
    except:
        logger.error(trace_error())
        return {"code": -1, "msg": trace_error()}


async def view_saier_dzfy_generate(request):  # 单证费用
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        bgpm = []
        sss = []
        sss1 = []
        gcmc = ""
        mj = 0.0
        rmb = 0.0
        mjsb = ""
        rmbsb = ""
        i4 = 0
        sjhjgh = ""
        kpxhz = 0

        # 检查当前用户 单证
        d = run_sql(
            f"select * from sys_users where name='{user.username}' and (position like '%单证%')"
        )
        if len(d) == 0:
            return json_result(-1, "没有单证")

        # 选择模版文件
        r_path = os.path.join(config.data_upload_path, "template")
        file_name = os.path.join(r_path, "dzfybx.xlsx")
        if not os.path.exists(file_name):
            return {"code": -1, "msg": "未找到报表模板"}

        # 用户选择导出范围(当前/全部), 保存路径
        save_path = j.get("ieksavefile", "")
        da2 = j.get("iekedit", "1")
        rid = j.get("rid", "")
        rids = j.get("rids", [])

        if not save_path:
            return {"code": -1, "msg": "请提供文件保存路径"}

        if da2 == "2":
            for rid in rids:
                i4, bgpm, gcmc, rmbsb, mjsb = process_hdfy_data(
                    rid, i4, bgpm, gcmc, rmbsb, mjsb
                )
        else:
            i4, bgpm, gcmc, rmbsb, mjsb = process_hdfy_data(
                rid, i4, bgpm, gcmc, rmbsb, mjsb
            )

        mj = 0.0
        kpxhn = time.strftime("%y")
        kpxhy = time.strftime("%m")
        kpxhr = time.strftime("%d")
        kpxh1 = f"HDFY{kpxhn}{kpxhy}{kpxhr}"
        sh = ""
        bank = ""
        zh2 = ""
        yhdm = ""
        lhh = ""
        jgh = ""
        wyjgh = ""
        sjyh = ""
        sjlhh = ""
        yhdz = ""
        sjhjgh = ""
        jzpz = ""
        father = ""
        sl = 0
        what = ""
        mw = ""
        result = ""
        jl = ""
        d = run_sql(
            f"select cdzl.bank,cdzl.zh,cdzl.shui,yh.yhdm,yh.yhdz,yh.wyjgh,yh.lhh,yh.jgh,yh.sjyh,yh.sjlhh \
                from cdzl left join(select yhbm.yhdm,yhbm.yhdz,yhbm.wyjgh,yhbm.lhh,yhbm.jgh,yhbm.sjyh,yhbm.sjlhh,yhbm.yhmc from yhbm) as yh on yh.yhmc=cdzl.bank \
                where (cdzl.cdmc={gcmc})"
        )
        if len(d) > 0:
            sh = str(d[0].get("shui", "")).strip()
            bank = str(d[0].get("bank", "")).strip()
            zh2 = str(d[0].get("zh", "")).strip()
            yhdm = str(d[0].get("yhdm", "")).strip()
            yhdz = str(d[0].get("yhdz", "")).strip()
            wyjgh = str(d[0].get("wyjgh", "")).strip()
            lhh = str(d[0].get("lhh", "")).strip()
            jgh = str(d[0].get("jgh", "")).strip()
            sjyh = str(d[0].get("sjyh", "")).strip()
            sjlhh = str(d[0].get("sjlhh", "")).strip()

        dzfysb = ""
        d = run_sql(
            f"select * from cyzglsheet with (nolock) where (xm='{user.username}') and (zm='单证费用生成')"
        )
        dzfysb = "1" if len(d) > 0 else dzfysb
        if not dzfysb:
            d = run_sql(
                f"select * from cyzglsheet with (nolock) where (zm='单证费用生成')"
            )
            dzfysb = "1" if len(d) == 0 else dzfysb

        if dzfysb == "1":
            d = run_sql(
                f"select top 1 jzpz from dzfy where jzpz like '%{kpxh1}%' order by jzpz desc"
            )
            kpxhz = int(d[0].get("jzpz")[10:14]) + 1 if len(d) > 0 else 1
            # 生成单证费用
            jzpz = f"{kpxh1}{kpxhz:04d}"
            if jzpz:
                row = {
                    "jzpz": jzpz,
                    "jbr": user.username,
                    "gcmc": gcmc,
                    "sh": sh,
                    "bank": bank,
                    "zh": zh2,
                    "yt": "贷款",
                    "fpwk": "",
                    "sfjq": "否",
                    "sqrq": time.strftime("%Y-%m-%d"),
                    "yhdm": yhdm,
                    "lhh": lhh,
                    "jgh": jgh,
                    "wyjgh": wyjgh,
                    "sjyh": sjyh,
                    "sjlhh": sjlhh,
                    "yhdz": yhdz,
                    "sjhjgh": sjhjgh,
                    "wstt": "",
                    "fkdq": "宁波",
                    "hkje": 0,
                    "htje": 0,
                    "sqje": 0,
                    "zt": "待付款",
                    "hkrq": "",
                    "cwry": "",
                    "fpje": 0,
                }
                m = dzfy()
                for k, v in row.items():
                    if k in SYS_FIELDS:
                        continue
                    if hasattr(m, k):
                        setattr(m, k, v)
                m.rid = get_uuid()
                m.ctime = time.strftime("%Y-%m-%d %H:%M:%S")
                s.add(m)

                father = m.rid

                # sys_created = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                # sys_lastmodified = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                # sqrq = datetime.datetime.now().strftime("%Y-%m-%d")
                # run_sql(
                #     f"insert into dzfy(sys_owner,sys_path,sys_created,sys_lastmodified,jzpz,jbr,gcmc,sh,bank,zh,yt,fpwk,sfjq,sqrq,yhdm,lhh,jgh,wyjgh,sjyh,sjlhh,yhdz,sjhjgh,wstt,fkdq,hkje,htje,sqje,zt,hkrq,cwry,fpje)\
                #         values('{user.username}','我的公司\\宁波优景进出口有限公司\\优景财务\\单证部\\','{sys_created}','{sys_lastmodified}','{jzpz}','{user.username}','{gcmc}','{sh}','{bank}','{zh2}','贷款','','否',{sqrq},'{yhdm}','{lhh}','{jgh}','{wyjgh}','{sjyh}','{sjlhh}','{yhdz}','','','宁波',0,0,0,'待付款','','',0)"
                # )

                # d = run_sql(f"select rid from dzfy where jzpz='{jzpz}'")
                # if len(d) > 0:
                #     father = str(d[0].get("rid", "")).strip()

        for rid in bgpm:
            d = run_sql(
                f"select * from hdfy \
                    where (rid={rid}) and ((dzqr4='通过') and (ywqr4='通过')) and ((sbje>0) or (sbjem>0))"
            )
            fphm = str(d[0].get("fphm", "")).strip()
            sjfy = float(d[0].get("sjhd", 0.0))
            sjfym = float(d[0].get("sbjem", 0.0))
            tdh = str(d[0].get("tdh", "")).strip()
            kkje = float(d[0].get("sbje", 0.0))
            yfje1 = float(d[0].get("sbjem", 0.0))
            if len(d) > 0 and father:
                d1 = run_sql(
                    f"select rid from dzfysheet \
                        where (pid='{father}') and (fphm='{fphm}') and (sjfy='{sjfy}') and (sjfym='{sjfym}') and (fylx='商检费用')"
                )
                if len(d1) == 0:
                    row = {
                        "pid": father,
                        "fphm": fphm,
                        "tdh": tdh,
                        "zdfy": 0,
                        "zlzbfy": 0,
                        "dbgfy": 0,
                        "tbfy": 0,
                        "sjfy": sjfy,
                        "xgzf": 0,
                        "zdfym": 0,
                        "zlzbfym": 0,
                        "dbgfym": 0,
                        "tbfym": 0,
                        "sjfym": sjfym,
                        "xgzfm": 0,
                        "kkje": kkje,
                        "yfje1": yfje1,
                        "fylx": "商检费用",
                    }
                    m1 = dzfysheet()
                    for k, v in row.items():
                        if k in SYS_FIELDS:
                            continue
                        if hasattr(m1, k):
                            setattr(m1, k, v)
                    m1.rid = get_uuid()
                    m1.ctime = time.strftime("%Y-%m-%d %H:%M:%S")
                    s.add(m1)
                    # run_sql(
                    #     f"insert into dzfysheet(pid,fphm,tdh,zdfy,zlzbfy,dbgfy,tbfy,sjfy,xgzf,zdfym,zlzbfym,dbgfym,tbfym,sjfym,xgzfm,kkje,yfje1,fylx) \
                    #         values('{father}','{fphm}','{tdh}',0,0,0,0,'{sjfy}',0,0,0,0,0,'{sjfym}',0,'{kkje}','{yfje1}',0,'商检费用')"
                    # )
                    sl += 1

        if rmbsb == "1" and i4 > 0:
            wb = load_workbook(file_name)
            ws = wb.get_sheet_by_name("Sheet1")
            number = 5
            e = 0
            for rid in bgpm:
                d = run_sql(
                    f"select * from hdfy where (rid='{rid}') and ((dzqr4='通过') and (ywqr4='通过')) and (sbje>0)"
                )
                if len(d) > 0:
                    mj = mj + float(d[0].get("sbje", 0.0))
                    jl = str(d[0].get("dzsp4", "")).strip()
                    e += 1
                    # 插入新数据行
                    ws.insert_rows(5 + e)
                    # 复制样式
                    for col in range(1, ws.max_column + 1):
                        temp_cell = ws.cell(row=number, column=col)
                        new_cell = ws.cell(row=5 + e, column=col)
                        if temp_cell.has_style:
                            new_cell.font = temp_cell.font.copy()
                            new_cell.border = temp_cell.border.copy()
                            new_cell.fill = temp_cell.fill.copy()
                            new_cell.number_format = temp_cell.number_format
                            new_cell.alignment = temp_cell.alignment.copy()
                    ws.row_dimensions[5 + e].height = None

                    if str(d[0].get("ysfp", "")).strip():
                        ws["A" + str(5 + e)] = str(d[0].get("ysfp", "")).strip()
                    else:
                        ws["A" + str(5 + e)] = str(d[0].get("fphm", "")).strip()
                    ws["B" + str(5 + e)] = float(d[0].get("sbnr", "").strip())
                    ws["C" + str(5 + e)].format_number = "￥#,##0.00"
                    ws["C" + str(5 + e)] = float(d[0].get("sbje", 0.0))

                    sbdy = time.strftime("%Y-%m-%d")
                    run_sql(
                        f"Update hdfy set sbdy='{sbdy}' where (rid='{rid}') and ((sbdy='')) or (sbdy is null))"
                    )

            if mj > 0:
                what = f"{mj:.2f}"
                # mw = str(int(round(mj * 100)))[-1]
                result = cn2an.an2cn(float(what), "rmb")

            ws["B" + str(7 + e)] = result
            ws["C" + str(7 + e)] = mj
            ws["A" + str(8 + e)] = (
                f"部门经理:{jl}   财务审核 :                 报销人:{user.username}"
            )
            ws.delete_rows(number)
            wb.save(save_path)

        if sl == 0 and father:
            run_sql(f"delete from dzjf where (rid='{father}')")
        else:
            if father:
                d = run_sql(
                    f"select sum(kkje) as kkje,sum(yfje1) as yfje1 from dzfysheet where (rid='{father}')"
                )
                if len(d) > 0:
                    kkjez1 = float(d[0].get("kkjez1", 0.0))
                    yfje1 = float(d[0].get("yfje1", 0.0))
                    run_aql(
                        f"update dzfy set kkje={kkjez1}, yfje1={yfje1} where (rid='{father}')"
                    )
        s.commit()
        return json_result(1, "生成成功")
    except Exception:
        logger.error(trace_error())
        s.rollback()
        return json_result(-1, trace_error())
    finally:
        s.close()
