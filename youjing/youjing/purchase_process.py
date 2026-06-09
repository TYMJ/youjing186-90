from email import errors
from math import e
from pdb import run
from webbrowser import get
from any import *
from .model import *
from sqlalchemy.sql import func,not_,or_,and_
import time
import json,os
from .__default__ import get_user_path,module_xxck_new,module_share_new, user_task_new
from openpyxl.drawing.image import Image as Image_Get
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, Alignment
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from PIL import Image as PILImage  # 重命名避免冲突
# jpeg图片出错的解决方法
from PIL import JpegImagePlugin
JpegImagePlugin._getmp = lambda x:None

SYS_FIELDS = ['sid','rid','uid','ctime','mtime','has_att','modi_uid','wf_status','wf_unit','pid','archived']


# 采购跟单的编辑界面加载
@any_route('/api/saier/purchase_process/load/check', methods=['POST', 'GET'])
@require_token
async def view_saier_purchase_process_load_check(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        # cgry = j.get('cgry', '')
        org = get_user_path(user.username)
        path = org.get('path','')
        qc = 0
        if '验货' in path or '行政' in path:
            qc = 1
        htnr = ''
        d = run_sql(f"select nr from zx where (ly='采购合同签订注意要点') limit 1")
        if len(d)>0:
            htnr = d[0].get('nr','')

        return json_result(1, '操作成功', {'htnr':htnr, 'qc':qc})
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 采购跟单的开票工厂变更
@any_route('/api/saier/purchase_process/kpgc/change', methods=['POST', 'GET'])
@require_token
async def view_saier_purchase_process_kpgc_change(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        data = {}
        kpgc = j.get('kpgc', '')
        d = run_sql(f"select hyd from ozycs where (company_name='{kpgc}') limit 1")
        if len(d)==0:
            return json_result(-1, '此开票工厂不存在，请检查后输入!')
        
        hyd = d[0].get('hyd','')
        d = run_sql(f"select kpgc,zzjgdm,kplxr,kpdh from zycs where (kpgc='{kpgc}') limit 1")
        if len(d)>0:
            data = d[0]
        data['hyd'] = hyd

        return json_result(1, '操作成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 采购跟单的开票工厂变更
@any_route('/api/saier/purchase_process/get/company', methods=['POST', 'GET'])
@require_token
async def view_saier_purchase_process_get_company(request):
    try:
        d = run_sql(f"select wfgs from wfgs")
        data = [r.get('wfgs','') for r in d]

        return json_result(1, '操作成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())


# 采购跟单的产品图片变更
@any_route('/api/saier/purchase_process/fktt/btn', methods=['POST', 'GET'])
@require_token
async def view_saier_purchase_process_fktt_btn(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        data = {}
        rids = j.get('rids', [])
        fktt = j.get('fktt', '')
        for rid in rids:
            d = s.query(cggd).filter(cggd.rid == rid, cggd.gdry == user.username).all()
            for r in d:
                c = s.query(cggdsheet).filter(cggdsheet.pid == r.rid).all()
                for l in c:
                    l.fktt = fktt
                    s.add(l)
            for r in d:
                c = s.query(cghtsheet).filter(func.ifnull(cghtsheet.dzsd,'')=='',cght.hthm == r.hthm).outerjoin(cght,cght.rid==cghtsheet.pid).all()
                for l in c:
                    l.fktt = fktt
                    s.add(l)
        s.commit()

        return json_result(1, '操作成功', data)
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 采购跟单的产品图片变更
@any_route('/api/saier/purchase_process/image/change', methods=['POST', 'GET'])
@require_token
async def view_saier_purchase_process_image_change(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        data = {}
        wyzd = j.get('wyzd', '')
        hthm = j.get('hthm', '')
        tpscsj = j.get('tpscsj', '')
        wxwyzd = j.get('wxwyzd', '')
        zycpbh = j.get('zycpbh', '')
        d = s.query(cghtsheet).filter(cghtsheet.wyzd == wyzd, cghtsheet.hthm == hthm, func.ifnull(cghtsheet.tpscsj, '') == '').all()
        for r in d:
            r.tpscsj = tpscsj
            s.add(r)

        d = s.query(cymxsheet).filter(or_(cymxsheet.wxwyzd == wxwyzd, cymxsheet.zycpbh == zycpbh), cymxsheet.cght == hthm, func.ifnull(cymxsheet.tpscsj, '') == '').all()
        for r in d:
            r.tpscsj = tpscsj
            s.add(r)

        s.commit()
        return json_result(1, '操作成功', data)
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 采购跟单的开票工厂变更
@any_route('/api/saier/purchase_process/gdry/change', methods=['POST', 'GET'])
@require_token
async def view_saier_purchase_process_gdry_change(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        data = {}
        rid = j.get('rid', '')
        gdry = j.get('gdry', '')
        org = get_user_path(gdry)
        path = org.get('path','')
        uid = org.get('rid','')

        return json_result(1, '操作成功', {'path': path, 'uid': uid})
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


def process_hthm_str(hthm):
    """处理hthm字符串的函数"""
    if not hthm:
        return '', ''
    
    sbcx = ''
    szm = ''
    # 查找 '-N2-'
    if '-N2-' in hthm:
        szm = '-N2-'
        pos = hthm.find('-N2-')
        if pos != -1:
            sbcx = hthm[pos+4:]
    # 查找 '-DMD-'
    elif '-DMD-' in hthm:
        szm = '-DMD-'
        pos = hthm.find('-DMD-')
        if pos != -1:
            sbcx = hthm[pos+5:]
    # 默认情况
    else:
        szm = hthm[0]
        dash_pos = hthm.find('-')
        if dash_pos != -1:
            sbcx = hthm[0] + hthm[dash_pos+1:]
        else:
            sbcx = hthm[0]
    
    return szm, sbcx

# 采购跟单的仓库名称变更
@any_route('/api/saier/purchase_process/ckmc/change', methods=['POST', 'GET'])
@require_token
async def view_saier_purchase_process_ckmc_change(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        hthm = j.get('hthm', '')
        khmc = j.get('khmc', '')
        jcbh = ''
        ckmc = j.get('ckmc', '')
        data = {'kamc':'', 'jcbh':''}
        szm, sbcx = process_hthm_str(hthm)
        d = run_sql(f"select ckmc from cangkuzl where (ckmc='{ckmc}') and (zt='否') limit 1")
        if len(d)==0:
            return json_result(-1, '不好意思,无此仓库名称,请重新选择')
        
        d = run_sql(f"select bz,bz3 from cyzglsheet where (xm='{szm}') and (zm='进仓编号代码') limit 1")
        if len(d)==0:
            jcbh = hthm
        else:
            if sbcx!='' and sbcx!=None:
                jcbh = d[0].get('bz','') + str(sbcx)
            data['kamc'] = d[0].get('bz3','')

        d = run_sql(f"select bz from cyzglsheet where (zm='特殊进仓编号代码') and (xm='{khmc}')  and (bz1='无') limit 1")
        if len(d)>0:
            jcbh = hthm

        if jcbh != '' and jcbh != None:
            d = run_sql(f"select bz from cyzglsheet where (zm='特殊进仓编号代码') and (xm='{khmc}') limit 1")
            if len(d)>0:
                jcbh = d[0].get('bz','') + str(jcbh)

        data['jcbh'] = jcbh

        return json_result(1, '操作成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 采购跟单的跟单查看变更
@any_route('/api/saier/purchase_process/gdck/change', methods=['POST', 'GET'])
@require_token
async def view_saier_purchase_process_gdck_change(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        sccj = j.get('sccj', '')
        cs_id = ''
        kpgc = j.get('kpgc', '')
        data = {'cs_id':'', 'hyd':'','fktt':''}
        d = run_sql(f"select cs_id from zycs where (company_name='{sccj}') limit 1")
        if len(d)>0:
            cs_id = d[0].get('cs_id','')
        data['cs_id'] = cs_id
        
        if cs_id != '' and cs_id != None:
            d = run_sql(f"select cslxr,phone,sjhm,fax,province1,city1,kpgc,zzjgdm,kplxr,kpdh,address,fktt from zycs where (cs_id='{cs_id}') limit 1")
            if len(d)>0:
                data.update(d[0])

        if kpgc != '' and kpgc != None:
            d = run_sql(f"select hyd from ozycs where (company_name='{kpgc}') limit 1")
            if len(d)>0:
                data['hyd'] = d[0].get('bz','')

        return json_result(1, '操作成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 采购跟单的跟单查看变更
@any_route('/api/saier/purchase_process/qxjc/btn', methods=['POST', 'GET'])
@require_token
async def view_saier_purchase_process_qxjc_btn(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        hthm = j.get('hthm', '')
        rid = j.get('rid', '')
        jcrq = j.get('jcrq', '')
        data = {'cs_id':'', 'hyd':'','fktt':''}
        if jcrq != '' and jcrq != None:
            jcrq = jcrq[:10]
        rk_rid = ''
        c = s.query(storage.rid).filter(storage.PurchaseOrderNo == hthm, storage.jcrq == jcrq).first()
        if c:
            rk_rid = c.rid
        if rk_rid != '' and rk_rid != None:
            d = s.query(sys_record_lock.username).filter(sys_record_lock.rid == rk_rid, sys_record_lock.module == '入库单'). first()
            if d:
                return json_result(-1, f"此入库单已由仓库人员查看或流转到第三方仓库，不能取消，请联系仓库,详情可去仓库管理-入库单 按采购合同查看")
        
        d = s.query(storage).filter(storage.PurchaseOrderNo == hthm, storage.jcrq == jcrq).limit(1).all()
        if not d:
            return json_result(-1, f"此入库单不存在，请检查采购合同号:{hthm}及进仓日期:{jcrq}是否正确!")
        for r in d:
            c = s.query(storageline).filter(storageline.pid == r.rid).all()
            for l in c:
                s.delete(l)
            s.delete(r)

        s.commit()
        return json_result(1, '操作成功', data)
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 采购跟单的开票工厂变更
@any_route('/api/saier/purchase_process/child/delete', methods=['POST', 'GET'])
@require_token
async def view_saier_purchase_process_child_delete(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        group = j.get('group', '')
        wyzd = j.get('wyzd', '')
        hthm = j.get('hthm', '')
        wxwyzd = j.get('wxwyzd', '')
        jcrq = j.get('jcrq', '')
        ckmc = j.get('ckmc', '')
        cpbh = j.get('cpbh', '')
        cgwyzd = j.get('cgwyzd', '')
        wyzd = j.get('wyzd', '')
        code = 1
        msg = '操作成功'
        if group == '进仓资料':
            if jcrq != '' and jcrq != None:
                jcrq = jcrq[:10]
            d = run_sql(f"select rid from storage where (PurchaseOrderNo='{hthm}') and (date_format(jcrq, '%Y-%m-%d')='{jcrq}') and (WarehouseName='{ckmc}') ")
            for r in d:
                c = run_sql(f"select rid from storageline where (pid='{r.get('rid')}') and (ItemNo='{cpbh}') and (wxwyzd='{wxwyzd}') and (cgwyzd='{cgwyzd}')  and (wyzd='{wyzd}') and (ifnull(ReturnCartonQty,'')=0)")
                if len(c)>0:
                    code = -1
                    msg = f"不好意思请先通知仓库删除后在删除此记录!"
                    break

        return json_result(code, msg)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 采购跟单的批量更新工厂信息操作
@any_route('/api/saier/purchase_process/batch/update/gcxx', methods=['POST', 'GET'])
@require_token
async def view_saier_purchase_process_batch_update_gcxx(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        rids = j.get('rids', [])
        cs_json = {}
        for rid in rids:
            d = s.query(cggd).filter(cggd.rid == rid, cggd.gdry == user.username).first()
            if not d:
                continue
            cs_id = d.cs_id
            if cs_id == '' or cs_id == None:
                continue
            m = run_sql(f"select cs_id from zycs where (company_name='{d.sccj}') or (cymch='{d.sccj}') limit 1")
            if len(m)>0:
                if cs_id != m[0].get('cs_id','') and m[0].get('cs_id','') != '' and m[0].get('cs_id','') != None:
                    cs_id = m[0].get('cs_id','')
            
            if not cs_id in cs_json:
                m = run_sql(f"select cslxr lxry,phone lxdh,sjhm,fax gccz,province1,city1,kpgc,zzjgdm,kplxr,kpdh from zycs where (cs_id='{cs_id}') limit 1")
                if len(m)==0:
                    continue
                m = m[0]
                kpgc = m.get('kpgc','')
                if kpgc == '' or kpgc == None:
                    x = run_sql(f"select hyd from ozycs where (company_name='{kpgc}') limit 1")
                    if len(x)>0:
                        m['hyd'] = x[0].get('hyd','')
                cs_json[cs_id] = m

            for k, v in cs_json[cs_id].items():
                if hasattr(d, k):
                    setattr(d, k, v)
            s.add(d)

            c = s.query(cggdsheet).filter(cggdsheet.pid == d.rid).all()
            for r in c:
                r.zzjgdm = cs_json[cs_id].get('zzjgdm', '')
                r.kplxr = cs_json[cs_id].get('kplxr', '')
                r.kpdh = cs_json[cs_id].get('kpdh', '')
                r.kpgc = cs_json[cs_id].get('kpgc', '')
                s.add(r)

            c = s.query(cght).filter(cght.hthm == d.hthm).first()
            if c :
                for k, v in cs_json[cs_id].items():
                    if hasattr(c, k):
                        setattr(c, k, v)
                s.add(c)
                x = s.query(cghtsheet).filter(cghtsheet.pid == c.rid).all()
                for l in x:
                    l.zzjgdm = cs_json[cs_id].get('zzjgdm', '')
                    l.kplxr = cs_json[cs_id].get('kplxr', '')
                    l.kpdh = cs_json[cs_id].get('kpdh', '')
                    l.kpgc = cs_json[cs_id].get('kpgc', '')
                    s.add(l)
                
        s.commit()
        return json_result(1, '操作成功')
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 采购跟单的批量更新跟单查看操作
@any_route('/api/saier/purchase_process/batch/update/gdck', methods=['POST', 'GET'])
@require_token
async def view_saier_purchase_process_batch_update_gdck(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        rids = j.get('rids', [])
        date = j.get('date', '')
        if date == '' or date == None:
            date = time.strftime('%Y-%m-%d')
        cs_json = {}
        for rid in rids:
            d = s.query(cggd).filter(cggd.rid == rid, 
                    # cggd.gdry == user.username
                ).first()
            if not d:
                continue
            cs_id = d.cs_id
            if cs_id == '' or cs_id == None:
                continue
            kamc = ''
            szm, sbcx = process_hthm_str(d.hthm)
            jcbh = d.hthm
            c = run_sql(f"select bz,bz3 from cyzglsheet where (xm='{szm}') and (zm='进仓编号代码') limit 1")
            if len(c)>0 and sbcx!='' and sbcx!=None:
                jcbh = c[0].get('bz','') + str(sbcx)
                kamc = c[0].get('bz3','')

            if not cs_id in cs_json:
                m = run_sql(f"select cslxr lxry,phone lxdh,sjhm,fax gccz,province1,city1,kpgc,zzjgdm,kplxr,kpdh,fktt from zycs where (cs_id='{cs_id}') limit 1")
                if len(m)==0:
                    continue
                m = m[0]
                kpgc = m.get('kpgc','')
                if kpgc == '' or kpgc == None:
                    x = run_sql(f"select hyd from ozycs where (company_name='{kpgc}') limit 1")
                    if len(x)>0:
                        m['hyd'] = x[0].get('hyd','')
                cs_json[cs_id] = m

            for k, v in cs_json[cs_id].items():
                if hasattr(d, k):
                    setattr(d, k, v)
            d.jcbh = jcbh
            d.mdka = kamc
            d.gdck = '是'
            d.gdrq = date
            s.add(d)

            c = s.query(cggdsheet).filter(cggdsheet.pid == d.rid).all()
            for r in c:
                r.zzjgdm = cs_json[cs_id].get('zzjgdm', '')
                r.kplxr = cs_json[cs_id].get('kplxr', '')
                r.kpdh = cs_json[cs_id].get('kpdh', '')
                r.kpgc = cs_json[cs_id].get('kpgc', '')
                r.fktt = cs_json[cs_id].get('fktt', '')
                # if cs_json[cs_id].get('hyd', '') != '' and cs_json[cs_id].get('hyd', '') != None:
                r.hyd = cs_json[cs_id].get('hyd', '')
                r.mdka = kamc
                s.add(r)

            c = s.query(cght).filter(cght.hthm == d.hthm).first()
            if c :
                for k, v in cs_json[cs_id].items():
                    if hasattr(c, k):
                        setattr(c, k, v)
                c.sccj1id = cs_id
                s.add(c)
                x = s.query(cghtsheet).filter(cghtsheet.pid == c.rid).all()
                for l in x:
                    l.zzjgdm = cs_json[cs_id].get('zzjgdm', '')
                    l.kplxr = cs_json[cs_id].get('kplxr', '')
                    l.kpdh = cs_json[cs_id].get('kpdh', '')
                    l.kpgc = cs_json[cs_id].get('kpgc', '')
                    l.fktt = cs_json[cs_id].get('fktt', '')
                    # if cs_json[cs_id].get('hyd', '') != '' and cs_json[cs_id].get('hyd', '') != None:
                    l.hyd = cs_json[cs_id].get('hyd', '')
                    s.add(l)
                
        s.commit()
        return json_result(1, '操作成功')
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 采购跟单的保存之前操作
@any_route('/api/saier/purchase_process/save/before', methods=['POST', 'GET'])
@require_token
async def view_saier_purchase_process_save_before(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        ewsr = 0
        ewzc = 0
        uid = ''

        m = j.get('main')
        rid = m.get('rid', '')
        wyzd = m.get('wyzd', '')
        sfwc = m.get('sfwc', '')
        cgry = m.get('cgry', '')
        gdry = m.get('gdry', '')
        gdbm = m.get('gdbm', '')
        ywry = m.get('ywry', '')
        hyd = m.get('hyd', '')
        kpgc = m.get('kpgc', '')
        hthm = m.get('hthm', '')
        ckmc = m.get('ckmc')
        rkdd = m.get('rkdd')
        jcrq = m.get('jcrq')
        jcry = m.get('Operator')
        jcbh = m.get('jcbh')
        items = {}
        lines = j.get('lines', [])
        fees = j.get('fees', [])
        org = get_user_path(user.username)
        path = org.get('path','')
        pos = org.get('position')
        if not '跟单' in path and not '跟单' in pos:
            return json_result(-1, '不好意思,您没有权限更改此资料,请与跟单人员联系,谢谢!')
        
        org = get_user_path(gdry)
        uid = org.get('rid','')
        szxq_list = []

        d = s.query(cght).filter(cght.wyzd == wyzd).first()
        if d :
            if d.gdwc != sfwc:
                d.gdwc = sfwc
                d.mtime = time.strftime('%Y-%m-%d %H:%M:%S')
                d.modi_uid = user.rid
                d.xddd = m.get('xddd', '')
                d.gdry = gdry
                d.gdbm = m.get('gdbm', '')
                d.wfgs = m.get('wfgs', '')
                d.yhdd = m.get('yhdd', '')
                d.lxry = m.get('lxry', '')
                d.gcdh = m.get('gcdh', '')
                d.sjhm = m.get('sjhm', '')
                d.jhrq = m.get('jhrq', '')
                d.gccz = m.get('gccz', '')
                d.province1 = m.get('province1', '')
                d.city1 = m.get('city1', '')
                d.kpgc = m.get('kpgc', '')
                d.zzjgdm = m.get('zzjgdm', '')
                d.kplxr = m.get('kplxr', '')
                d.kpdh = m.get('kpdh', '')
                d.rkdd = m.get('rkdd', '')
                d.szxq = '\n'.join(szxq_list)
                d.hyd = m.get('hyd', '')
                d.skfm = m.get('skfm', '')
                d.bank = m.get('bank', '')
                d.zh = m.get('zh', '')
                logger.error(f"主要信息:aa")
            if d.kpgc != kpgc or d.hyd != hyd:
                c = cghtsheet1()
                c.rid = get_uuid()
                c.pid = d.rid
                c.xgrq = time.strftime('%Y-%m-%d')
                c.ggry = user.username
                c.ykpgc = d.kpgc
                c.kpgc = kpgc
                c.zzjgdm = d.zzjgdm
                c.kplxr = d.kplxr
                c.kpdh = d.kpdh
                c.hyd = hyd
                c.uid = user.rid
                c.ctime = time.strftime('%Y-%m-%d %H:%M:%S')
                s.add(c)
                logger.error(f"修改信息:bb")

            s.query(cghtsheet4).filter(cghtsheet4.pid == d.rid).delete()
            for f in fees:
                logger.error(f"费用信息:{f}")
                c = cghtsheet4()
                c.rid = get_uuid()
                c.pid = d.rid
                c.sfqk = f.get('sfqk', '')
                c.fymc = f.get('fymc', '')
                c.hbdm = f.get('hbdm', '')
                c.fyje = f.get('fyje', 0)
                c.uid = user.rid
                c.ctime = time.strftime('%Y-%m-%d %H:%M:%S')
                szxq_list.append('合同号码:' + str(hthm) + ' 费用名称:' + str(c.fymc) + '收付情况:' + str(c.sfqk) + '货币代码:' + str(c.hbdm) + ' 金额:' + str(c.fyje))
                if c.sfqk == '收入':
                    ewsr += float(c.fyje)
                if c.sfqk == '支出':
                    ewzc += float(c.fyje)
                s.add(c)

            if ewsr != d.ewsr or ewzc != d.ewzc:
                d.ewsr = ewsr
                d.ewzc = ewzc
            s.add(d)

        errors = []
        index = 0
        for r in lines:
            yyjq = ''
            index += 1
            qrjq = r.get('qrjq')
            if r.get('jhrq', '') != None and r.get('jhrq', '') != '':
                jhrq = r.get('jhrq', '')
            if qrjq != '' and qrjq != None:
                jhrq = qrjq

            c = s.query(cggdsheet).filter(cggdsheet.rid == r.get('rid')).first()
            if c and c.qrjq!=None and c.qrjq!='':
                yyjq = str(c.qrjq)[:10]

            c = s.query(cghtsheet).filter(cghtsheet.wyzd == r.get('wyzd')).first()
            if c and r.get('wyzd', '') != None and r.get('wyzd', '') != '' :
                if r.get('dyg1', '') != None and r.get('dyg1', '') != '' and r.get('dyg1', '') != '[]' and r.get('dyg1', '') != c.dyg1:
                    c.dyg1 = r.get('dyg1', '')
                if r.get('sjg2', '') != None and r.get('sjg2', '') != '' and r.get('sjg2', '') != '[]' and r.get('sjg2', '') != c.dyg2:
                    c.dyg2 = r.get('sjg2', '')
                if c.dzsd != None and c.dzsd != '':
                    items[r.get('rid')] = c.dzsd
                    errors.append(f"'第{str(index)}条由{str(c.dzsd)}锁定'")
                else:
                    c.hgbm = r.get('hgbm', '')
                    c.fktt = r.get('fktt', '')
                    c.xddd = m.get('xddd', '')
                    c.gcpp = r.get('gcpp', '')
                    c.gdry = gdry
                    c.gdbm = gdbm
                    c.cpgg = r.get('hwms', '')
                    c.yse = r.get('ysks', '')
                    c.nhrl = r.get('nhrl', 0)
                    c.zhwbzh = r.get('bzyq', '')
                    c.wxrl = r.get('zxl', 0)
                    c.jldw = r.get('dw', '')
                    c.cgxs = r.get('htzxs', 0)
                    c.cgsl = r.get('htzsl', 0)
                    c.nhwx = r.get('nhwx', 0)
                    c.jhrq = r.get('jhrq')
                    c.tj = r.get('wxcc', 0)
                    c.jz = r.get('jz', 0) 
                    c.mz = r.get('mz', 0)
                    c.zjz = r.get('htzjz', 0)
                    c.zmz = r.get('htzmz', 0)
                    c.ztj = r.get('htztj', 0)
                    # c.yhsl = r.get('yhsl', 0)
                    # c.blbl = r.get('blbl', 0)
                    # c.jiel = r.get('jiel','')
                    c.qrjq = qrjq
                    c.bzcd = r.get('bzcd', 0)
                    c.bzkd = r.get('bzkd', 0)
                    c.bzgd = r.get('bzgd', 0)
                    c.bzgd = r.get('bzgd', 0)
                    c.ks = r.get('ks', 0)
                    c.wcsl = r.get('wcsl', 0)
                    c.zhwbgpm = r.get('zhwbgpm', '')
                    c.zwpm = r.get('zwmc', '')
                    c.bgjldw = r.get('bgjldw', '')
                    c.sfsj = r.get('sfsj', '')
                    c.zkfy = r.get('zkfy', 0)
                    c.qtsm1 = r.get('qtsm1', '')
                    c.drbz = '否'
                    c.ljrk = r.get('ljrk', 0)
                    c.sfhs = r.get('sfhs', '')
                    c.tpscsj = r.get('tpscsj')
    
                c.kpgc = kpgc
                c.zzjgdm = m.get('zzjgdm', '')
                c.kplxr = m.get('kplxr', '')
                c.kpdh = m.get('kpdh', '')
                c.hyd = m.get('hyd', '')
                c.jcsj = jcrq
                c.AWqr = r.get('AWqr', '')
                c.SMqr = r.get('SMqr', '')
                c.sxbqr = r.get('sxbqr', '')
                c.cqqr = r.get('cqqr', '')
                c.bz3 = r.get('bz', '')
                c.tpscsj = r.get('tpscsj', '')
                s.add(c)
                y = s.query(cymxsheet).filter(cymxsheet.wxwyzd==r.get('wxwyzd'),cymxsheet.cght==hthm,func.ifnull(cymxsheet.wxrl,0)==r.get('wxrl',0)).all()
                for x in y:
                    if x.tpscsj == None or x.tpscsj == '':
                        x.tpscsj = r.get('tpscsj')
                    if x.scrq == None or x.scrq == '':
                        x.scrq = jcrq
                    x.wxrl = r.get('wxrl', 0)
                    s.add(x)
            if jhrq!='' and jhrq!=None and qrjq!='' and qrjq!=None and qrjq[:10]>jhrq[0:10] and qrjq[:10]!=yyjq:
                for u in [ywry, cgry]:
                    if u == '' or u == None or u == user.username:
                        continue
                    qrjq = qrjq[:10]
                    jhrq = jhrq[:10]
                    xxnr = f"{user.username}交货延后通知:合同号:{hthm},货号:{r.get('cpbh')}更改交货日期:{jhrq}为:{qrjq}"
                    row = {
                        "xxly": '采购合同',
                        "bjdh": '',
                        "wxht": '',
                        "cght": hthm,
                        "yhdh": '',
                        "xxnr": xxnr,
                        "jsr": str(u),
                        "sys_path": "",
                        "spsq": u
                    }
                    res = module_xxck_new([row], user, s)
                    if res.get('code',1) != 1:
                        return json_result(-1, res.get('msg'))
            if jcrq != '' and jcrq !=None:
                jcrq = jcrq[:10]
            if ckmc != '' and ckmc != None and rkdd != '' and rkdd !=None:
                d = s.query(storage).filter(storage.SNID==jcbh,storage.jcrq==jcrq,func.ifnull(storage.StorageDate,'')=='').first()
                if d :
                    d.WarehouseName = ckmc
                    d.rkdd = rkdd
                    d.mtime = time.strftime('%Y-%m-%d %H:%M:%S')
                    if jcry != '' and jcry != None:
                        org = get_user_path(jcry)
                        d.uid = org.get('rid')
                        d.modi_uid = org.get('rid')
                    s.add(d)

        data = {'uid': uid, 'path': path, 'ewsr': ewsr, 'ewzc': ewzc, 'items': items, 'errors': errors}
        s.commit()
        return json_result(1, '操作成功', data)
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 采购跟单的跟单查看变更
@any_route('/api/saier/purchase_process/batch/cgrk/check', methods=['POST', 'GET'])
@require_token
async def view_saier_purchase_process_batch_cgrk_check(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        d = run_sql(f"select rid from cyzglsheet where (xm='{user.username}') and (zm='批量入库单跟单') limit 1")
        if len(d)==0 :
            return json_result(-1, f"权限校验失败")
        
        return json_result(1, '操作成功')
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

def pixels_to_points(value, dpi=96):
    """96 dpi, 72i"""
    return value * 72 / dpi

def points_to_pixels(value, dpi=96):
    return int(math.ceil(value * dpi / 72))

def offset_img(img, col, row, x_pad=4, y_pad=25):
    """精确设置图片位置，偏移量以万为单位进行微调吧，具体计算公式太麻烦了
    row column 的索引都是从0开始的，我这里要把图片插入到单元格A17
    """
    p2e = pixels_to_EMU
    h, w = img.height, img.width
    # 图像等比例缩放因子
    resize_factor = 0.8
    w_h_ratio = w/h
    resize_H = int(resize_factor * h)
    resize_W = int(resize_factor * resize_H * w_h_ratio)
    #
    # x_pad = 4
    # y_pad = 25
    # 注意这里的行、列索引从0开始, 所以需要减1
    size = XDRPositiveSize2D(p2e(w), p2e(h))
    marker = AnchorMarker(col=col-1, colOff=pixels_to_EMU(x_pad), row=row-1, rowOff=pixels_to_EMU(y_pad))
    img.anchor = OneCellAnchor(_from=marker, ext=size)

    # 图像在Excel里面的大小
    # image_size_excel = XDRPositiveSize2D(pixels_to_EMU(resize_W), pixels_to_EMU(resize_H))
    ##############################
    # 设置单元格大小，单元格默认宽度单位：字符；高度单位：point(磅)
    # cell_height = int(pixels_to_points(resize_H+10, dpi=96)) #高度上增加10个像素放大单元格
    # cell_width = int(resize_W/8) + 2 # 宽度上增加16（2*8）个像素放大单元格

    # marker = AnchorMarker(col=col, colOff=pixels_to_EMU(x_pad), row=row, rowOff=pixels_to_EMU(y_pad))
    # img.anchor = OneCellAnchor(_from=marker, ext=image_size_excel)


def storage_main_new_data(rid, rk_pid, jcrq, SNID, user, s):
    try:
        d = s.query(cggd).filter(cggd.rid == rid).first()
        if not d:
            return {'code':-1, 'msg':'操作失败,未找到对应的采购跟单记录', 'data': None}
        if SNID == '' or SNID == None:
            SNID = d.jcbh
        if jcrq == '' or jcrq == None:
            jcrq = d.yjjc
        ckmc = d.ckmc
        rkdd = d.rkdd
        filters = [storage.PurchaseOrderNo == d.hthm, storage.jcrq == jcrq, storage.WarehouseName == ckmc]
        if ckmc == '宁波龙和':
            filters.append(storage.SNID == SNID)
        t = s.query(storage).filter(*filters).first()
        if not t:
            t = storage()
            t.rid = rk_pid
            t.uid = user.rid
            t.ctime = time.strftime('%Y-%m-%d %H:%M:%S')
            t.zyh = auto_number.generate(s,'入库单.作 业 号',{'rid':rk_pid})  
            t.jcrq = jcrq
            t.Operator = d.Operator
            t.SalesOrderNo = d.wxht
            t.Exporter = d.wfgs
            t.Salesman = d.ywry
            t.PurchaseOrderNo = d.hthm
            t.SupplierShortName = d.sccj
            t.PurchasingAgent = d.cgry
            t.WarehouseName = ckmc
            t.SNID = SNID
            t.sfjc = '否'
            t.gdry = d.gdry
            t.yzrq = d.yzrq
            t.rkdd = rkdd
            t.wyzd = rk_pid
            t.StorageTime = None
            t.StorageDate = None
            t.khmc = d.khmc
        else:
            t.jcrq = jcrq
            t.Operator = d.Operator
            t.WarehouseName = ckmc
            t.rkdd = rkdd
            rk_pid = t.rid
        
        return {'code':1, 'msg':'操作成功', 'data': t, 'rid': rk_pid}
    except:
        logger.error(trace_error())
        return {'code':-1, 'msg':'操作失败', 'data': None}
    
def storage_child_new_data(rid, jcbh, rk_pid, user, s, jc_rid=None):
    try:
        ywpath = ''
        cgpath = ''
        khmc = ''
        Volume = 0
        GrossWeight = 0
        CartonQty = 0
        d = s.query(cggd.khmc,cggd.ywry,cggd.cgry).filter(cggd.rid == rid).first()
        if d:
            khmc = d.khmc
            org = get_user_path(d.ywry)
            ywpath = org.get('path','')
            org = get_user_path(d.cgry)
            cgpath = org.get('path','')
        if jc_rid == None or jc_rid == '':
            filters = [cggdsheet1.pid == rid]
        else:
            filters = [cggdsheet1.rid == jc_rid]
        flag = 0
        c = s.query(cggdsheet1).filter(*filters).all()
        for l in c:
            if l.jcsj == None or l.jcsj == '':
                continue
            x = s.query(storageline).filter(storageline.wyzd == l.wyzd).first()
            if x:
                continue
            y = storageline()
            y.rid = get_uuid()
            y.pid = rk_pid
            y.uid = user.rid
            y.ctime = time.strftime('%Y-%m-%d %H:%M:%S')
            y.wyzd = l.rid
            y.OuterLength = l.bzcd or 0
            y.OuterWidth = l.bzkd or 0
            y.OuterHeight = l.bzgd or 0
            y.ItemNo = l.cpbh
            y.OuterGrossWeight = l.mz or 0
            y.OuterNetWeight = l.jz or 0
            y.ExpectedCartonQty = l.jcxs or 0
            y.OuterVolume = l.wxcc or 0
            y.Volume = l.htztj or 0
            y.GrossWeight = l.htzmz or 0
            y.ckztj = 0 # m.htztj or 0
            y.ckzmz = 0 # m.htzmz or 0
            y.cktj = 0 # m.htztj or 0
            y.ckmz = 0 # m.htzmz or 0
            y.ckjz = 0 # m.htzmz or 0
            y.ckzd = 0
            y.ckgd = 0
            y.ckkd = 0
            y.zwmc = l.zwmc
            y.wxwyzd = l.wxwyzd
            y.cgwyzd = l.cgwyzd
            y.ywpath = ywpath
            y.cgpath = cgpath
            y.khmc = khmc
            y.SNIDc = jcbh
            y.zxl = l.zxl or 0
            y.zpxs = 0
            y.CartonQty = 0
            s.add(y)
            Volume += y.Volume or 0
            GrossWeight += y.GrossWeight or 0
            CartonQty += y.ExpectedCartonQty or 0
            flag += 1
        if flag == 0:
            return {'code':-1, 'msg':'请注意已有进仓，详情可去仓库管理-入库单 按采购合同查看', 'data': None}

        return {'code':1, 'msg':'操作成功', 'Volume': Volume, 'GrossWeight': GrossWeight, 'CartonQty': CartonQty}
    except:
        logger.error(trace_error())
        return {'code':-1, 'msg':'操作失败', 'data': None}

def storage_new_data(rid, ckmc, user, s):
    try:
        c = s.query(cangkuzl).filter(cangkuzl.ckmc==ckmc, cangkuzl.zdrkd=='是').first()
        if not c:
            return {'code':-1, 'msg':'操作失败,该仓库未设置自动入库'}
        rk_pid = get_uuid()
        res = storage_main_new_data(rid, rk_pid, '', '', user, s)
        if res.get('code', 1) != 1:
            return res
        t = res.get('data')
        rk_pid = t.rid
        res = storage_child_new_data(rid, '', rk_pid, user, s)
        if res.get('code', 1) != 1:
            return res

        Volume = res.get('Volume') or 0
        GrossWeight = res.get('GrossWeight') or 0
        CartonQty = res.get('CartonQty') or 0
        if t:
            t.TotalGrossWeight = 0
            t.TotalVolumn = 0
            t.TotalCartons = 0
            t.yjtjhj = Volume
            t.yjmzhj = GrossWeight
            t.ExpectedCartonQty = CartonQty
            s.add(t)
        res = user_task_new('入库单', rid, '进仓单号', '进仓通知[进仓单号]','采购合同:' + str(t.PurchaseOrderNo) + '已进仓,进仓日期:' + str(t.jcrq)[:10], user, s, [t.Operator], True)
        if res.get('code') != 1:
            return res
        xxnr = f"{user.username}的进仓单号:{t.SNID},进仓日期:{str(t.jcrq)[:10]},请查看,日期:{time.strftime('%Y-%m-%d')}"
        row = {
            "xxly": '进仓通知',
            "bjdh": '',
            "wxht": '',
            "cght": '',
            "yhdh": '',
            "xxnr": xxnr,
            "jsr": str(t.Operator),
            "sys_path": "",
            "spsq": user.username
        }
        res = module_xxck_new([row], user, s)
        if res.get('code',1) != 1:
            return res

        return {'code':1, 'msg':'操作成功'}
    except:
        logger.error(trace_error())
        return {'code':-1, 'msg':'操作失败'}

@any_route('/api/saier/purchase_process/report/print', methods=['POST', 'GET'])
@require_token
async def view_saier_purchase_process_report_print(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        rid = j.get('rid')
        ckmc = j.get('ckmc')
        res = storage_new_data(rid, ckmc, user, s)
        if res.get('code', 1) != 1:
            s.rollback()
            return json_result(-1, res.get('msg'))
        
        s.commit()
        return json_result(1, '操作成功')
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 采购跟单的跟单查看变更
@any_route('/api/saier/purchase_process/batch/cgrk/btn', methods=['POST', 'GET'])
@require_token
async def view_saier_purchase_process_batch_cgrk_btn(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        path = config.tmp_path
        ckmc = j.get('ckmc', '')
        rids = j.get('rids', [])
        jcrq = j.get('date')
        if jcrq == '' or jcrq == None:
            jcrq = time.strftime('%Y-%m-%d')
        files = []
        # 是否合适的采购跟单记录校验
        flag = 0
        # 客户名称是否有批量入库单的权限校验
        check = 0
        # 仓库是否正确的校验
        ck_flag = 0
        # 数量是否正确的校验
        sl_flag = 0
        ry_flag = 0
        jczsl = 0
        htzsl = 0
        for rid in rids:
            # 需要当前用户是当前记录的跟单人员一致
            d = s.query(cggd).filter(cggd.rid == rid, 
                # cggd.gdry == user.username
            ).first()
            if d:
                ry_flag = 1
                logger.error(f"rid:{rid} 跟单人员:{d.gdry} 当前用户:{user.username}")
                kamc = ''
                khmc = d.khmc
                Operator1 = ''
                rkdd1 = ''
                c = s.query(wxht.mdka).filter(wxht.order_id == d.wxht).first()
                if c:
                    kamc = c.mdka
                c = run_sql(f"select xm from cyzglsheet where (bz='{khmc}') AND ((bz1='{kamc}') or (ifnull(bz1,'')='')) AND \
                    (bz2>='{time.strftime('%Y-%m-%d')}') AND  (bz3<='{time.strftime('%Y-%m-%d')}') and (zm='客人港口仓库对照表') limit 1")
                if len(c)>0:
                    ckmc = c[0].get('xm','')
                if ckmc != '' and ckmc != None:
                    c = run_sql(f"select ckmc,lxr,dq from cangkuzl where (ckmc='{ckmc}') limit 1")
                    if len(c)>0:
                        Operator1 = c[0].get('lxr','')
                        rkdd1 = c[0].get('dq','')
                # 校验当前记录的客户名称在系统参数表中是否有对应的记录，如果没有则无法进行批量入库单的操作
                c = run_sql(f"select rid from cyzglsheet where (xm='{khmc}') and (zm='批量入库单客人') limit 1")
                if not c:
                    continue
                check = 1
                if d.jhrq != '' and d.jhrq != None:
                    jcrq = str(d.jhrq)[:10]
                
                # 获取进仓编号
                jcbh = d.jcbh
                if jcbh == '' or jcbh == None:
                    hthm = d.hthm
                    szm, sbcx = process_hthm_str(hthm)
                    b = run_sql(f"select bz,bz3 from cyzglsheet where (xm='{szm}') and (zm='进仓编号代码') limit 1")
                    if len(b)==0:
                        jcbh = hthm
                    else:
                        if sbcx!='' and sbcx!=None:
                            jcbh = b[0].get('bz','') + str(sbcx)
                    
                    b = run_sql(f"select bz from cyzglsheet where (zm='特殊进仓编号代码') and (xm='{khmc}')  and (bz1='无') limit 1")
                    if len(b)>0:
                        jcbh = hthm

                    if jcbh != '' and jcbh != None:
                        b = run_sql(f"select bz from cyzglsheet where (zm='特殊进仓编号代码') and (xm='{khmc}') limit 1")
                        if len(b)>0:
                            jcbh = b[0].get('bz','') + str(jcbh)
                logger.error(f"进仓编号:{jcbh}")
                org = get_user_path(d.ywry)
                ywpath = org.get('path','')
                org = get_user_path(d.cgry)
                cgpath = org.get('path','')
                x = s.query(cggdsheet).filter(cggdsheet.pid == d.rid).all()
                jczsl = 0
                flag = 0
                index = 0
                t = None
                Volume = 0
                GrossWeight = 0
                CartonQty = 0
                rk_pid = get_uuid()
                for l in x:
                    l.jcsj = jcrq
                    l.wcsl = l.htzsl if l.htzsl != None else 0
                    s.add(l)
                    htzsl = float(l.htzsl) if l.htzsl != None else 0
                    jcsl = float(l.htzsl) if l.htzsl != None else 0
                    # 更新采购合同产品资料的进仓日期
                    s.query(cghtsheet).filter(cghtsheet.wyzd == l.wyzd).update({cghtsheet.jcsj: jcrq}, synchronize_session=False)
                    # 获取已进仓数量
                    jcsl1 = s.query(func.ifnull(func.sum(func.ifnull(cggdsheet1.jcsl, 0)), 0).label('jcsl1')).filter(cggdsheet1.pid == l.pid, cggdsheet1.cpbh == l.cpbh).scalar()
                    jcsl2 = s.query(func.ifnull(func.sum(func.ifnull(cggdsheet2.jcsl, 0)), 0).label('jcsl2')).filter(cggdsheet2.pid == l.pid, cggdsheet2.cpbh == l.cpbh).scalar()
                    jczsl = float(jcsl1) + float(jcsl2) + jcsl
                    logger.error(f"进仓数量:{jczsl} 订单数量:{l.htzsl}")
                    if jczsl <= htzsl:
                        # errors.append(f"合同号:{d.hthm} 货号:{l.cpbh} 进仓数量:{jczsl} 超过合同数量:{htzsl}")
                        # continue
                        y = s.query(cggdsheet1).filter(cggdsheet1.pid == l.pid, cggdsheet1.cpbh == l.cpbh).first()
                        if y:
                            # 生成采购跟单单的进仓历史
                            m = cggdsheet2()
                            z = alchemy_object_to_dict(y)
                            for k,v in z.items():
                                if k in SYS_FIELDS:
                                    continue
                                setattr(m, k, v)
                            m.rid = get_uuid()
                            m.pid = l.pid
                            m.uid = user.rid
                            m.ctime = time.strftime('%Y-%m-%d %H:%M:%S')
                            s.add(m)
                            s.delete(y)
                        # 生成采购跟单单的进仓资料
                        m = cggdsheet1()
                        z = alchemy_object_to_dict(l)
                        for k,v in z.items():
                            if k in SYS_FIELDS:
                                continue
                            setattr(m, k, v)
                        
                        jc_rid = get_uuid()
                        m.rid = jc_rid
                        m.pid = l.pid
                        m.uid = user.rid
                        m.ctime = time.strftime('%Y-%m-%d %H:%M:%S')
                        m.jcsl = jcsl
                        m.jcsj = jcrq
                        m.jcbh = jcbh
                        m.jcxs = round(jcsl / float(l.zxl) if l.zxl and float(l.zxl) != 0 else 0, 2)
                        m.htzjz = round(float(l.jz) * round(jcsl/float(l.zxl)) * 100) / 100
                        m.htzmz = round((float(l.mz) * round(jcsl / float(l.zxl))) * 100) / 100
                        m.sjzl = l.sjzl
                        m.htztj = round((float(l.wxcc) * round(jcsl / float(l.zxl))) * 1000) / 1000
                        m.Operator = Operator1
                        m.WarehouseName = ckmc
                        # m.yzrq = None
                        m.rkdd = rkdd1
                        m.wyzd = m.rid
                        m.sd = '否'
                        m.yjjc = jcrq
                        m.wxwyzd = l.wxwyzd
                        m.cgwyzd = l.wyzd
                        s.add(m)

                        index += 1
                        if (ckmc == '义乌仓库') or (ckmc != '宁波志恒') or (ckmc != '汕头仓库') or (ckmc != '宁波龙和') or (ckmc != '宁波万纬'):
                            logger.error("仓库名称校验通过 AAA")
                            if index == 1:
                                # 生成入库单
                                # t = storage()
                                # t.rid = rk_pid
                                # t.uid = user.rid
                                # t.ctime = time.strftime('%Y-%m-%d %H:%M:%S')
                                # t.jcrq = jcrq
                                # t.Operator = Operator1
                                # t.SalesOrderNo = d.wxht
                                # t.Exporter = d.wfgs
                                # t.Salesman = d.ywry
                                # t.PurchaseOrderNo = d.hthm
                                # t.SupplierShortName = d.sccj
                                # t.PurchasingAgent = d.cgry
                                # t.WarehouseName = ckmc
                                # t.zyh = auto_number.generate(s,'入库单.作 业 号',{'rid':rk_pid})
                                # t.SNID = jcbh
                                # t.sfjc = '否'
                                # t.gdry = d.gdry
                                # t.yzrq = d.yzrq
                                # t.rkdd = rkdd1
                                # t.wyzd = rk_pid
                                # t.StorageTime = None
                                # t.StorageDate = None
                                # t.khmc = d.khmc
                                res = storage_main_new_data(rid, rk_pid, jcrq, jcbh, user, s)
                                if res.get('code', 1) != 1:
                                    return json_result(-1, res.get('msg'))
                                t = res.get('data')
                                
                            # y = storageline()
                            # y.rid = get_uuid()
                            # y.pid = rk_pid
                            # y.uid = user.rid
                            # y.ctime = time.strftime('%Y-%m-%d %H:%M:%S')
                            # y.wyzd = jc_rid
                            # y.OuterLength = l.bzcd or 0
                            # y.OuterWidth = l.bzkd or 0
                            # y.OuterHeight = l.bzgd or 0
                            # y.ItemNo = l.cpbh
                            # y.OuterGrossWeight = l.mz or 0
                            # y.OuterNetWeight = l.jz or 0
                            # y.ExpectedCartonQty = m.jcxs or 0
                            # y.OuterVolume = l.wxcc or 0
                            # y.Volume = m.htztj or 0
                            # y.GrossWeight = m.htzmz or 0
                            # y.ckztj = 0 # m.htztj or 0
                            # y.ckzmz = 0 # m.htzmz or 0
                            # y.cktj = 0 # m.htztj or 0
                            # y.ckmz = 0 # m.htzmz or 0
                            # y.ckjz = 0 # m.htzmz or 0
                            # y.ckzd = 0
                            # y.ckgd = 0
                            # y.ckkd = 0
                            # y.zwmc = l.zwmc
                            # y.wxwyzd = l.wxwyzd
                            # y.cgwyzd = l.wyzd
                            # y.ywpath = ywpath
                            # y.cgpath = cgpath
                            # y.khmc = d.khmc
                            # y.SNIDc = jcbh
                            # y.zxl = l.zxl or 0
                            # y.zpxs = 0
                            # y.CartonQty = 0
                            res = storage_child_new_data(rid, jcbh, rk_pid, user, s, jc_rid)
                            if res.get('code', 1) != 1:
                                return json_result(-1, res.get('msg'))
                            y = res.get('data')
                            Volume += res.get('Volume', 0)
                            GrossWeight += res.get('GrossWeight', 0)
                            CartonQty += res.get('ExpectedCartonQty', 0)
                            # s.add(y)
                        else:
                            ck_flag = 1
                            logger.error("仓库名称校验失败 BBB")
                        # 更新采购合同标志
                        flag = 1
                    else:
                        sl_flag = 1
                if t:
                    t.TotalGrossWeight = 0
                    t.TotalVolumn = 0
                    t.TotalCartons = 0
                    t.yjtjhj = Volume
                    t.yjmzhj = GrossWeight
                    t.ExpectedCartonQty = CartonQty
                    s.add(t)
                if flag == 1:
                    mc = ckmc
                    if (ckmc != '义乌仓库') and (ckmc != '宁波志恒') and (ckmc != '汕头仓库') and (ckmc != '宁波龙和') and (ckmc != '宁波万纬'):
                        mc = '其他'
                    r_path = os.path.join(config.data_upload_path,'template')
                    fn = os.path.join(r_path, str(mc) + '.xlsx')
                    if not os.path.exists(fn):
                        return {"code":-1,"msg":"未找到报表模板"}
                    wb = load_workbook(fn) 
                    ws = wb.worksheets[0]
                    i = 1
                    number = 1
                    if (mc == '宁波龙和') or (mc == '宁波万纬') or (mc == '其他') :
                        number = 17
                    else:
                        if (mc == '宁波志恒') or (mc == '汕头仓库') :
                            number = 18
                        else :
                            if (mc == '义乌仓库') :
                                number = 19
                    csjc = ''
                    # c = run_sql(f"select csjc from zycs where (company_name='{d.sccj}') or (cs_id='{d.cs_id}') limit 1")
                    # if len(c)>0:
                    #     csjc = c[0].get('csjc','')
                    c = alchemy_object_to_dict(d)
                    ws['A' + str(i)] = str(c.get('wfgs')).upper().strip()
                    i = 4
                    ws['C' + str(i)] = str(c.get('hthm')).upper().strip()
                    ws['I' + str(i)] = jcbh
                    ydhm = ''
                    u = run_sql(f"select bm,ssdq,ydhm from ywrybiao where yhm='{c.get('gdry')}' limit 1")
                    if len(u)>0:
                        ydhm = u[0].get('ydhm', '')
                    if (mc == '义乌仓库'):
                        i = 10
                    elif (mc == '宁波志恒') or (mc == '汕头仓库'):
                        i = 9
                    elif (mc == '宁波龙和') or (mc == '宁波万纬') or (mc == '其他'):
                        i = 8

                    ws['C' + str(i)] = c.get('gdry')
                    ws['E' + str(i)] = ydhm
                    i = i + 5
                    if (mc == '宁波龙和') or (mc == '宁波万纬') or (mc == '其他'):
                        ws['F' + str(i)] = jcrq
                    else:
                        ws['G' + str(i)] = jcrq
                    i = i +1
                    ws['C' + str(i)] = c.get('sccj')
                    # ws['N' + str(i)] = c.get('sccj')
                    # ws['O' + str(i)] = c.get('lxry') + '\\' + c.get('lxdh') + '\\' + c.get('sjhm')
                    ws['H' + str(i)] = c.get('lxry') + '\\' + c.get('lxdh') + '\\' + c.get('sjhm')
                    i = i + 1
                    ws['C' + str(i)] = c.get('shc')
                    gzmc = c.get('wfgs') + '仓库章'
                    g = run_sql(f"select tpmc from tpzx where (cpbh='{gzmc}') limit 1")
                    if len(g)>0 and g[0].get('tpmc','') != '' and g[0].get('tpmc','') != None and g[0].get('tpmc','') != '[]':
                        gzmc = g[0].get('tpmc','')
                        Photo = json.loads(str(gzmc))
                        if Photo != None:
                            file_path = Photo[0]['src']
                            fn = os.path.join(config.data_upload_path, str(file_path))
                            if (os.path.exists(fn)):
                                img = Image_Get(fn) #选择图片
                                # img.width = 150  # 设置图像宽度
                                # img.height = 68  # 设置图像高度
                                col_width = (ws.column_dimensions['B'].width + ws.column_dimensions['C'].width)*7
                                row_height = 100 # (ws.row_dimensions[i].height)*1.3333
                                img.width=col_width-4 # 转换为像素
                                img.height=row_height-4 # 转换为像素
                                x_offset = 15 # (col_width-img.width)/2
                                y_offset = 10 # (row_height-img.height)/2
                                col = 2
                                row = i
                                offset_img(img, col, row, x_offset, y_offset) #col为列位置，row为行位置,x_offset为左边边距,y_offset为上边距
                                ws.add_image(img)  #添加图片
                    k = 0
                    t = s.query(cggdsheet1.cpbh,cggdsheet1.jcxs,cggdsheet1.htzmz,cggdsheet1.htztj).filter(cggdsheet1.pid == d.rid).all()
                    for g in t:
                        # msexcelworksheet.Rows[inttostr(number + k)].Insert
                        if len(t) > 2:
                            ws.insert_rows(number + k + 1)
                        i = number + k
                        # 如果有上方行，复制其格式到新行
                        # if i > 1:
                        #     source_row = i - 1
                        #     for row in range(i, i + 1):
                        #         for col in range(1, ws.max_column + 1):
                        #             source_cell = ws.cell(row=source_row, column=col)
                        #             target_cell = ws.cell(row=row, column=col)
                                    # 复制样式
                                    # if source_cell.has_style:
                                    #     target_cell.font = copy(source_cell.font)
                                    #     target_cell.border = copy(source_cell.border)
                                    #     target_cell.fill = copy(source_cell.fill)
                                    #     target_cell.number_format = copy(source_cell.number_format)
                                    #     target_cell.protection = copy(source_cell.protection)
                                    #     target_cell.alignment = copy(source_cell.alignment)

                        
                        p = run_sql(f"select jcpm from zscp where (cpbh='{g.cpbh}') or (krhh='{g.cpbh}') limit 1")
                        if len(p)>0:
                            ws['A' + str(i)] = p[0].get('jcpm')
                        ws['A' + str(i)] = g.cpbh
                        ws['C' + str(i)] = g.cpbh
                        ws['E' + str(i)] = str(g.jcxs) + 'CTNS'
                        ws['G' + str(i)] = str(g.htzmz) + 'KGS'
                        ws['I' + str(i)] = str(g.htztj) + 'CBM'
                        ws['J' + str(i)] = jcbh
                        ws['M' + str(i)] = g.cpbh
                        k = k + 1
                    
                    report_rid = get_uuid()
                    fn = str(jcbh) + str(ckmc) + str(d.sccj) + '.xlsx'
                    files.append({'name': fn, 'path': path + '/'+ str(report_rid)+'.xlsx'})
                    
                    wb.save(path + '/'+ str(report_rid)+'.xlsx')
                    d.jcsj = jcrq
                    d.yjjc = jcrq
                    if d.jcbh != jcbh:
                        d.jcbh = jcbh
                d.ckmc = ckmc
                if Operator1 != '' and Operator1 != None:
                    d.Operator = Operator1
                    d.rkdd = rkdd1
                s.add(d)
                flag = 1
        if ry_flag == 0:
            return json_result(-1, '没有需要生成入库单的合同，请确认当前用户与采购跟单记录中的跟单人员是否一致!')
        if ck_flag == 1:
            return json_result(-1, '所选记录的仓库不正确，请核对后再进行批量入库单的操作!')
        if sl_flag == 1:
            return json_result(-1, f"合同的进仓数量{jczsl}大于合同数量{htzsl}，请核对后再进行批量入库单的操作!")
        if check == 0:
            return json_result(-1, '没有需要生成入库单的合同，请确认当前客户在成员管理组是否有"批量入库单客人"权限!')
        if flag == 0:
            return json_result(-1, '没有需要生成入库单的合同，请确认所选记录的跟单人员和当前用户是否一致!')
        filename = ''
        if len(files) > 0:
            filename = str(jcbh) + time.strftime("%Y%m%d%H%M%S") + '.zip'
            sZipPath = os.path.join(path, filename) # 压缩包路径
            zipFile = zipfile.ZipFile(sZipPath, 'w') #生成一个压缩包，第二个参数默认值为'r'，表示读已经存在的zip文件，'w'表示新建一个zip文档或覆盖一个已经存在的zip文档
            for f in files:
                file_path = os.path.join(path,str(f.get("path")))
                if os.path.exists(file_path):
                    zipFile.write(file_path, f.get('name'), zipfile.ZIP_DEFLATED) #将file_path的文件重命名为sfilename

            zipFile.close()

        s.commit()
        return json_result(1, '操作成功', filename)
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 采购跟单的跟单查看变更
@any_route('/api/saier/purchase_process/save/after', methods=['POST', 'GET'])
@require_token
async def view_saier_purchase_process_batch_cgrk_btn(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        d = s.query(cggd).filter(cggd.rid==j.get('rid')).first()
        if d :
            gdry = d.gdry
            org = get_user_path(gdry)
            uid = org.get('rid','')
            if uid != d.uid:
                d.uid = uid
                s.add(d)
                s.commit()
        return json_result(1, '操作成功')
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()