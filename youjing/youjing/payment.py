from math import *
from any import *
from .__default__ import get_user_path, module_share_new, module_xxck_new, user_task_new
from .model import *
import time, calendar, re
from .advance_payment import new_payment_merge_data,to_num
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, numbers
from copy import deepcopy
from datetime import datetime, date, timedelta
from sqlalchemy import update, and_, or_, func
import zipfile
from typing import Tuple
from dateutil import parser
import json
import requests


# 采购付款.出货日期.change
@any_route("/api/saier/payment/chrq/change", methods=["POST", "GET"])
@require_token
async def view_payment_chrq_change(request):
    try:
        data = {'cs': 0, 'sz1': 0}
        sql_str = f"select cs,sz1 from zx where ly='财务付款结算期' limit 1"
        d = run_sql(sql_str) or []
        if len(d) > 0:
            data = d[0]

        return json_result(1, "success", data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        pass


# 采购付款.采购合同.change
@any_route("/api/saier/payment/cght/change", methods=["POST", "GET"])
@require_token
async def view_payment_cght_change(request):
    j = await request.json()
    username = request.current_user.username
    s = Session()
    try:
        jsrm = j.get("jsrm", "")
        hthm = j.get("hthm", "")
        data = ''
        if jsrm != "" and jsrm != None:
            return json_result(0, "success", '')
        if hthm != "" and hthm != None:
            return json_result(0, "success", '')
        d = s.query(cght.cgry, cght.gdry).filter(cght.hthm == hthm).first()
        if d:
            org = get_user_path(username)
            if '优景' in org.get('path') and d.gdry != None and d.gdry != "":
                data = d.gdry
            else:
                data = d.cgry

        return json_result(1, "success", data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        pass

# 采购付款.开户银行.change


@any_route("/api/saier/payment/user/check", methods=["POST", "GET"])
@require_token
async def view_payment_user_check(request):
    j = await request.json()
    user = request.current_user
    s = Session()
    try:
        data = 0
        field = j.get("field", "position")
        position = j.get("position", "财务修改")
        logger.error('field: %s', field)
        logger.error('position: %s', position)
        if field == 'position':
            org = get_user_path(user.username)
            if position in org.get(field):
                data = 1
        else:
            roles = []
            u = s.query(sys_user_role).filter(sys_user_role.user_id == user.rid).all()
            role_ids = [r.role_id for r in u]
            flag = 0
            for r in role_ids:
                c = s.query(sys_role).filter(sys_role.rid == r).first()
                if c and c.name and '财务' in c.name:
                    flag = 1
                    roles.append(c.name)
            # if flag == 0:
            #     return json_result(-1, "无操作权限")
            
            return json_result(1, "success", roles)
        
        return json_result(1, "success", data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        pass


# 采购付款.工厂发票.change
@any_route("/api/saier/payment/gcfp/change", methods=["POST", "GET"])
@require_token
async def view_payment_gcfp_change(request):
    s = Session()
    j = await request.json()
    username = request.current_user.username
    try:
        gcfp = j.get("gcfp", "")
        if gcfp == "" or gcfp == None:
            return json_result(0, "success")

        d = s.query(gchk.rid).filter(gchk.gcfp == gcfp).first()
        if d:
            return json_result(-1, "发票号码重复,请核对或在发票前后加符号")

        return json_result(1, "success")
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        pass


def new_cxgc_data(d, user, s):
    try:
        m = cxgc()
        m.rid = get_uuid()
        m.ctime = time.strftime('%Y-%m-%d %H:%M:%S')
        m.uid = user.rid
        m.gcmc = d.get('gcmc1')
        m.wxfp = d.get('wxfp')
        m.ywry = d.get('ywry')
        m.yrrq = time.strftime('%Y-%m-%d')
        m.sfsh = '否'
        m.yrgs = d.get('wfgs')
        m.chgc = d.get('gcmc1')
        m.fkry = ''
        m.xgry = ''
        s.add(m)
        return {'code': 1, 'msg': 'success'}
    except:
        logger.error(trace_error())
        return {'code': -1, 'msg': trace_error()}


# 采购付款.save.before
@any_route("/api/saier/payment/save/before", methods=["POST", "GET"])
@require_token
async def view_payment_save_before(request):
    j = await request.json()
    user = request.current_user
    s = Session()
    try:
        d = j.get('main', {})
        fkrq = d.get("fkrq")
        chgc = d.get("chgc", "")
        sb = d.get("sb", "")
        fpje = d.get("fpje", 0)
        htje = d.get("htje", 0)
        jsrm = d.get("jsrm", "")
        jsrbm = d.get("jsrbm", "")
        lines = j.get('lines', [])
        msg = ''
        errors = []
        data = {
            'cxbg': '不需要',
            'msg': '',
            'position': '',
            'path': '',
            'jsrbm': '',
            'ywdq': '',
            'errors': [],
            'days': 0,
            'fkxh': ''
        }
        if jsrm != "" and jsrm != None:
            c = s.query(ywrybiao.bm, ywrybiao.ssdq).filter(
                ywrybiao.yhm == jsrm).first()
            if jsrbm == "" or jsrbm == None:
                if c and c.bm != None and c.bm != "":
                    data['jsrbm'] = c.bm
            data['ywdq'] = c.ssdq if c and c.ssdq != None else ''
            if data['ywdq'] == '' or data['ywdq'] == None:
                c = s.query(ywrylx.dq).filter(ywrylx.ryxm == jsrm).first()
                data['ywdq'] = c.dq if c and c.dq != None else ''

        org = get_user_path(user.username)
        path = org.get('path', '')
        position = org.get('position', '')
        data['position'] = position
        data['path'] = path
        if fkrq == "" or fkrq == None:
            return json_result(0, "success", data)
        c = s.query(cgfkhdsheet).filter(cgfkhdsheet.sb == sb).first()
        if c and c.hkrq != fkrq:
            s.hkrq = fkrq
            s.add(c)

        if chgc == "" and chgc == None:
            s.commit()
            return json_result(0, "success", data)

        if 'YWB' in chgc or '义乌办' in chgc:
            s.commit()
            return json_result(0, "success", data)

        fkhj = d.get('fkhj', 0)
        yfhj = d.get('yfhj', 0)
        fkce = yfhj - fkhj
        fkce1 = yfhj - (int(yfhj / 10) * 10)
        gcfp = d.get('gcfp', '')
        mlsb = ''
        if (fkhj != yfhj) and (fkce < 10) and (gcfp == '' or gcfp == None):
            mlsb = '是'
        else:
            if (fkhj == yfhj) and (fkce1 > 0) and (gcfp == '' or gcfp == None):
                mlsb = '否'

        for r in lines:
            hthm = r.get('cght', '')
            if hthm != '' and hthm != None:
                c = s.query(gcdjsheet).filter(gcdjsheet.cght == hthm, or_(
                    gcdjsheet.sfjq == "", gcdjsheet.sfjq.is_(None), gcdjsheet.sfjq == "否")).first()
                if c:
                    if len(errors) == 0:
                        errors.append("请注意：以下合同号有预付没结清")
                    if not f"生产厂家: {d.get('gcmc', '')}, 工厂发票: {d.get('gcfp', '')}, 采购合同: {r.get('cght', '')}" in errors:
                        errors.append(
                            f"生产厂家: {d.get('gcmc', '')}, 工厂发票: {d.get('gcfp', '')}, 采购合同: {r.get('cght', '')}")
            cywyzd = r.get('cywyzd', '')
            if cywyzd != '' and cywyzd != None:
                flag = 0
                c = s.query(fkspsheet).filter(fkspsheet.cght ==
                                              hthm, gcdjsheet.sfjq == "是").first()
                if c:
                    if c.fkrq == None or c.fkrq == "":
                        c.fkrq = fkrq
                        flag = 1
                    if c.fkrq1 == None or c.fkrq1 == "":
                        c.fkrq1 = fkrq
                        flag = 1
                if flag == 1:
                    s.add(c)
            if mlsb != '' and mlsb != None:
                s.query(cymxsheet).filter(cymxsheet.cght == hthm, cymxsheet.cywyzd == cywyzd).update(
                    {"mlsb": mlsb}, synchronize_session=False)

        if len(errors) > 0:
            user_list = []
            c = s.query(cyzglsheet.xm).filter(or_(cyzglsheet.bz == "全部", cyzglsheet.bz == d.get(
                'wstt')), cyzglsheet.zm == "采购付款预付末结提醒").all()
            for l in c:
                row = {
                    "xxly": '采购付款',
                    "bjdh": '',
                    "wxht": '',
                    "cght": d.get('gcfp', ''),
                    "yhdh": '',
                    "xxnr": '\n'.join(errors),
                    "jsr": l.xm,
                    "sys_path": "",
                    "spsq": l.xm
                }
                res = module_xxck_new([row], user, s)
                if res.get('code') != 1:
                    s.rollback()
                    return json_result(res.get('code'), res.get('code'))

        cxje = 0
        gsmc = ''
        c = s.query(zx).filter(zx.mc == '诚信金额').first()
        if c:
            cxje = float(c.cs)
        c = s.query(ywrybiao).filter(ywrybiao.yhm == 'zjnblh').first()
        if c:
            gsmc = c.wfgs
        if cxje > 0 and (fpje > cxje or htje > cxje):
            c = s.query(cxgc).filter(cxgc.gcmc == chgc).first()
            if not c:
                cxbg = '待提供'
                d['wfgs'] = gsmc
                res = new_cxgc_data(d, user, s)
                if res.get('code') != 1:
                    s.rollback()
                    return json_result(-1, res.get('msg'))
            else:
                if c.sfsh == '是':
                    cxbg = '已提供'
                else:
                    cxbg = '待提供'
            if cxbg == '待提供':
                data['msg'] = f"工厂名称: {chgc} 需提交诚信报告"

        s.query(fpgl).filter((fpgl.wxfp == d.get('wxfp')) | (fpgl.hsfp == d.get('hsfp')),
                             fpgl.tssd == "否", fpgl.sfjd == "否").update({"webpdfk": "是"}, synchronize_session=False)
        if d.get('Field') == '待提交':
            d['fkhz'] = '外销发票:' + d.get('wxfp', '') + '付款金额' + d.get('hkje', 0)
            d['ly'] = module
            res = new_payment_merge_data(user, d, module, s)
            if res.get('code') != 1:
                s.rollback()
                return json_result(-1, res.get('msg'))
        data['cxbg'] = cxbg

        sccj = d.get('gcmc', '')
        a1 = d.get('gcmc1', '')
        hthm = d.get('cght', '')
        if a1 == '':
            a1 = 'zjnblh123456789'
        if hthm == '':
            hthm = 'zjnblhcght123456789'
        c = s.query(yfhk.rid).filter(or_(yfhk.sccj.like(f"%{sccj}%"), yfhk.gcmc1.like(
            f"%{a1}%"), yfhk.cght == hthm), yfhk.sfjq != "是").first()
        if c:
            org = get_user_path(c.uid)
            if org.get('path', '') == path:
                msg = f"该生产或出货厂家或采购合同有预付货款，您可以通过数据链接方式查看详细信息！"
        data['msg'] = msg
        if d.get('gcfp', '') != '' and d.get('gcfp', '') != None:
            c = s.query(gchk).filter(gchk.gcfp == d.get(
                'gcfp', ''), gchk.rid != d.get('rid', '')).first()
            if c:
                data['msg'] = f"请注意系统里有相同工厂发票"

        a = 0
        b = 0
        days = 0
        if '优胜' in path:
            sql_rule = f"""select * from cwjxrq where (szgs like '%优胜%')"""
            rows_rule = run_sql(sql_rule)
            if len(rows_rule) > 0:
                row = rows_rule[0]
                a = to_num(row.get("kpjxts", 0), 0)
                b = to_num(row.get("xjjxts", 0), 0)
                if d.get('gcfp') != '' and d.get('gcfp') != None:
                    days = a
                else:
                    days = b
        data['days'] = days

        new_fkxh = ''
        if (gcfp == '' or gcfp == None) and (fkrq != '' and fkrq != None) and fkje > 0:
            nf = time.strftime('%Y')[2:4]
            yf = time.strftime('%m')
            rq = time.strftime('%d')
            c = s.query(gchk).filter(gchk.fkxh.like(
                f"%{nf}{yf}{rq}%")).order_by(gchk.fkxh.desc()).first()
            if c:
                fkxh = str(c.fkxh)[6:10]
            else:
                fkxh = 0
            fkxh = format((int(fkxh) + 1), '04d')
            new_fkxh = str(nf) + str(yf) + str(rq) + str(fkxh)
            data['fkxh'] = new_fkxh

        s.commit()
        return json_result(1, "success", data)
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        pass

# 采购付款.load.check


@any_route("/api/saier/payment/load/check", methods=["POST", "GET"])
@require_token
async def view_payment_load_check(request):
    j = await request.json()
    user = request.current_user
    s = Session()
    try:
        data = {'htsh': '', 'cwsb': 0, 'position': ''}
        rid = j.get('rid', '')
        d = s.query(zx.cs).filter(zx.ly == '合同收回金额').first()
        if d:
            data['htsh'] = d.cs

        org = get_user_path(user.username)
        position = org.get('position', '')
        data['position'] = position
        if '财务修改' in position:
            data['cwsb'] = 1

        return json_result(1, "success", data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        pass


# 采购付款.确认按钮
@any_route("/api/saier/payment/confirm", methods=["POST", "GET"])
@require_token
async def view_payment_confirm(request):
    j = await request.json()
    user = request.current_user
    s = Session()
    try:
        rids = j.get('rids', [])
        for rid in rids:
            flag = 0
            d = s.query(gchk).filter(gchk.rid == rid).first()
            if not d:
                continue
            if d.wstt == None or d.wstt == "":
                if d.fpwk == "否":
                    d.fkdq = '宁波'
                    flag = 1
            else:
                wstt = d.wstt
                c = s.query(kpnr).filter(kpnr.wfgs == wstt).first()
                if c and "." and c.hgbh != None and c.hgbh != "" and c.hgbh != "":
                    d.fkdq = '义乌'
                    flag = 1
                else:
                    if d.fpwk == "否":
                        d.fkdq = '宁波'
                        flag = 1
            if flag == 1:
                d.mtime = time.strftime('%Y-%m-%d %H:%M:%S')
                d.modi_uid = user.rid
                s.add(d)

        s.commit()
        return json_result(1, "操作成功")
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        pass


# 采购付款.签收单按钮
@any_route("/api/saier/payment/receipt/update", methods=["POST", "GET"])
@require_token
async def view_payment_receipt_update(request):
    j = await request.json()
    user = request.current_user
    s = Session()
    try:
        rids = j.get('rids', [])
        gsmc = j.get('gsmc', '')
        for rid in rids:
            d = s.query(gchk).filter(gchk.rid == rid).first()
            if not d:
                continue
            Yfje = 0
            c = s.query(func.sum(func.ifnull(gchksheet.Yfje, 0)).label(
                'Yfje')).filter(gchksheet.pid == rid).all()
            for r in c:
                Yfje = Yfje + (float(r.Yfje) if r.Yfje != None else 0)
            d.modi_uid = user.rid
            d.mtime = time.strftime('%Y-%m-%d %H:%M:%S')
            d.qsgs = gsmc
            c = s.query(bmlgc.rid).filter(bmlgc.gcmc == d.gcmc).first()
            if c:
                yfhj = Yfje
            else:
                yfhj = Yfje + math.trunc(Yfje / 10) * 10
            c = s.query(cyzglsheet.sz).filter(cyzglsheet.xm == "公司", cyzglsheet.zm == "抹零金额").first()
            if c and c.sz != None and yfhj <= float(c.sz):
                yfhj = Yfje
            d.yfhj = yfhj

            s.add(d)

        s.commit()
        return json_result(1, "success")
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        pass


def copy_range(source_cells, target_cell, sheet, copy_style=False):
    """
    将指定区域(source_cells)的值复制到目标单元格(target_cell)起始的区域
    """
    # 将目标单元格转换为可操作的对象
    start_row, start_col = target_cell.row, target_cell.column

    for i, row in enumerate(source_cells):
        for j, source_cell in enumerate(row):
            # 计算目标单元格的位置
            target_row = start_row + i
            target_col = start_col + j
            target = sheet.cell(row=target_row, column=target_col)

            # 复制值
            target = source_cell

            # 如果需要且源单元格有样式，则复制样式
            if copy_style and source_cell.has_style:
                target._style = deepcopy(source_cell._style)


# 采购付款.签收单按钮
@any_route("/api/saier/payment/export/txt", methods=["POST", "GET"])
@require_token
async def view_payment_export_txt(request):
    j = await request.json()
    user = request.current_user
    s = Session()
    try:
        # gsmc = j.get('gsmc', '')
        rids = j.get('rids', [])
        hh = j.get('hh', '')
        jg = j.get('jg', '')
        zh = j.get('zh', '')
        # 模板
        # template_path = os.path.join(config.data_upload_path, 'template', '签收单1.xlsx')
        # if not os.path.exists(template_path):
        #     return json_result(-1, '模板不存在: 签收单1.xlsx')

        # i = 0
        # k = 0
        # index = 0
        # wb = load_workbook(template_path)
        # ws = wb.worksheets[0]
        # source_area = ws['A1':'J29']
        # for rid in rids:
        #     index = index + 1
        #     start_row = index * 29 + 1
        #     copy_range(source_area, ws.cell(row=start_row, column=1), ws, copy_style=False)

        # i = 0
        # k = 0
        index = 0
        txt_list = []
        date = time.strftime('%Y-%m-%d')
        s1 = date[0:1]
        s2 = date[1:2]
        s3 = date[2:3]
        s4 = date[3:4]
        s5 = date[5:6]
        s6 = date[6:7]
        s7 = date[8:9]
        s8 = date[9:10]
        yfje = 0
        for rid in rids:
            index = index + 1
            d = s.query(gchk).filter(gchk.rid == rid).first()
            if not d:
                continue
            yfje = yfje + (float(d.yfhj) if d.yfhj != None else 0)
        txt_list.append('OBSS|2|' + str(index) + '|' + str(yfje) + '|')
        index = 0
        for rid in rids:
            index = index + 1
            d = s.query(gchk).filter(gchk.rid == rid).first()
            if not d:
                continue
            hkrq = str(d.hkrq) if d.hkrq != None else ''
            d1 = hkrq[0:1]
            d2 = hkrq[1:2]
            d3 = hkrq[2:3]
            d4 = hkrq[3:4]
            d5 = hkrq[5:6]
            d6 = hkrq[6:7]
            d7 = hkrq[8:9]
            d8 = hkrq[9:10]

            fkbz = d.fkbz
            city = d.city
            yt = d.yt
            pd = '1'
            if fkbz == None or fkbz == '':
                fkbz = '0'
            if city == None or city == '':
                city = '0'
            if yt == None or yt == '':
                yt = '货款'
            pd = '1'
            if d.yhdm and d.yhdm != '':
                pd = '0'
            s9 = str(index).zfill(4)
            txt_list.append(str(s1) + str(s2) + str(s3) + str(s4) + str(s5) + str(s6) + str(s7) + str(s8) + str(s9) + '|' + (hh if hh else '') + '|' + (jg if jg else '') + '|' + (zh if zh else '')
                            + '|' + (d.zh if d.zh else '') + '|' + (d.lhh if d.lhh else '') + '|' + (d.jgh if d.jgh else '') + '|' + (d.bank if d.bank else '') + '|' + (d.gcmc if d.gcmc else '')
                            + '|' + (city if city else '0') + '|' + (fkbz) + '|' + str(float(d.hkje) if d.hkje != None else 0) + '||' 
                            + str(d1) + str(d2) + str(d3) + str(d4) + str(d5) + str(d6) + str(d7) + str(d8)
                            + '|' + str(yt) + '0' + '|' + '||||' + pd  + '|' + (str(d.yhdm).strip() if d.yhdm else '') + '|' + '0')
            # k = k + 1
            # yfhj = 0
            # d = s.query(gchk).filter(gchk.rid == rid).first()
            # if not d:
            #     continue
            # a1 = i + 1
            # ws['J' + str(a1 + 1 + 29)] = d.xjxh
            # ws['A' + str(a1 + 2 + 29)] = gsmc
            # ws['C' + str(a1 + 5 + 29)] = d.jsrm
            # ws['G' + str(a1 + 5 + 29)] = d.wxfp
            # ws['C' + str(a1 + 6 + 29)] = d.rxfs
            # ws['I' + str(a1 + 6 + 29)] = d.kh
            # ws['G' + str(a1 + 19 + 29)] = d.xshj
            # ws['H' + str(a1 + 19 + 29)] = d.zsl
            # ws['I' + str(a1 + 19 + 29)] = d.yfhj
            # ws['D' + str(a1 + 20 + 29)] = str(d.bzsm) + str(d.xjbz)
            # ws['J' + str(a1 + 21 + 29)] = d.fkhj
            # ws['C' + str(a1 + 23 + 29)] = d.chrq
            # ws['B' + str(a1 + 27 + 29)] = d.gcmc
            # ws['F' + str(a1 + 27 + 29)] = d.zh
            # pid = d.rid
            # i = i + 1
            # index = index + 1
            # f = 0
            # c = s.query(gchksheet).filter(gchksheet.pid == pid).all()
            # for r in c:
            #     f = f + 1
            #     if f <= 10:
            #         i = i + 1
            #         ws['A' + str(8 + i9)] = c.cght
            #         ws['B' + str(8 + i9)] = c.cphh
            #         ws['C' + str(8 + i9)] = c.zwpm
            #         ws['D' + str(8 + i9)] = c.dj
            #         ws['E' + str(8 + i9)] = c.dw
            #         ws['F' + str(8 + i9)] = c.bz1
            #         ws['G' + str(8 + i9)] = c.xs
            #         ws['H' + str(8 + i9)] = c.zsl
            #         ws['I' + str(8 + i9)] = c.Yfje
            #         ws['J' + str(8 + i9)] = c.gcmc
            #         yfhj = yfhj + float(c.Yfje) if c.Yfje != None else 0
            #     if f > 10:
            #         ws['I' + str(a1 + 1 + 29)] = '品名超过10个,请个别打印'

            # ws['I' + str(a1 + 19 + 29)] = yfhj
            # c = s.query(bmlgc.rid).filter(bmlgc.gcmc == d.gcmc).first()
            # if c:
            #     ws['J' + str(a1 + 21 + 29)] = yfhj
            # else:
            #     ws['J' + str(a1 + 21 + 29)] = math.trunc(yfhj / 10) * 10
            # c = s.query(cyzglsheet.sz).filter(cyzglsheet.xm == "公司", cyzglsheet.zm == "抹零金额").first()
            # if c and c.sz != None and yfhj <= float(c.sz):
            #     ws['J' + str(a1 + 21 + 29)] = yfhj

            # i = index * 29
        txt_list.append('OBSSEND')
        out_name = f"{user.username}_{time.strftime('%Y-%m-%d')}_付款明细.txt"
        out_path = os.path.join(config.tmp_path, out_name)
        val = '\n'.join(txt_list)
        # logger.error(f"Writing duplicates to: {out_path}")
        write_file(out_path, val, 'w')
        # os.makedirs(config.tmp_path, exist_ok=True)
        # wb.save(out_path)

        # s.commit()
        return json_result(1, '导出成功', out_name)
    except:
        # s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        pass

# 采购付款.标记导出日期
@any_route("/api/saier/payment/update/date_flag", methods=["POST", "GET"])
@require_token
async def view_payment_update_date_flag(request):
    j = await request.json()
    user = request.current_user
    s = Session()
    try:
        rids = j.get('rids', [])
        org = get_user_path(user.username)
        position = org.get('position', '')
        if '财务日期' not in position:
            return json_result(-1, "您没有权限操作")
        for rid in rids:
            d = s.query(gchk).filter(gchk.rid == rid).first()
            if not d:
                continue
            d.modi_uid = user.rid
            d.mtime = time.strftime('%Y-%m-%d %H:%M:%S')
            d.zzrq = time.strftime('%Y-%m-%d')
            d.zzrqj = time.strftime('%Y-%m-%d %H:%M:%S')
            s.add(d)

        s.commit()
        return json_result(1, "success")
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        pass

def ArabiaToChinese(num):
    map = {
        '0': '零',
        '1': '壹',
        '2': '贰',
        '3': '叁',
        '4': '肆',
        '5': '伍',
        '6': '陆',
        '7': '柒',
        '8': '捌',
        '9': '玖'
    }
    return map.get(num, "") or num

def format_month_to_chinese(month: str) -> str:
    """
    将月份转换为中文格式
    01 -> 零壹
    10 -> 壹拾零
    12 -> 壹拾贰
    """
    if len(month) == 1:
        month = '0' + month
    
    m1, m2 = month[0], month[1]
    
    if m1 == '0':
        return '零' + ArabiaToChinese(m2)
    elif m2 == '0':
        return ArabiaToChinese(m1) + '拾零'
    else:
        return ArabiaToChinese(m1) + '拾' + ArabiaToChinese(m2)


def format_day_to_chinese(day: str) -> str:
    """
    将日期转换为中文格式
    01 -> 零壹
    10 -> 壹拾零
    23 -> 贰拾叁
    30 -> 叁拾零
    """
    if len(day) == 1:
        day = '0' + day
    
    d1, d2 = day[0], day[1]
    
    if d1 == '0':
        return '零' + ArabiaToChinese(d2)
    elif d2 == '0':
        return ArabiaToChinese(d1) + '拾零'
    else:
        return ArabiaToChinese(d1) + '拾' + ArabiaToChinese(d2)


def parse_date_optimized(date_str: str) -> Tuple[str, str, str, str, str, str]:
    """
    将日期字符串转换为中文大写格式
    参数:
        date_str: 日期字符串，格式 'YYYY-MM-DD'
    返回:
        (年份中文, 月份中文, 日期中文, 年份数字, 月份数字, 日期数字)
    """
    if not date_str:
        return ('', '', '', '', '', '')
    # 按标准格式解析
    if isinstance(date_str, datetime):
        date_str = date_str.strftime('%Y-%m-%d')

    parts = date_str.split('-')
    if len(parts) != 3:
        return ('', '', '', '', '', '')
    
    year_str = parts[0].zfill(4)
    month_str = parts[1].zfill(2)
    day_str = parts[2].zfill(2)
    
    # 转换年份（逐位转换，不省略任何零）
    year_chinese = ''.join(ArabiaToChinese(ch) for ch in year_str)
    
    # 转换月份
    month_chinese = format_month_to_chinese(month_str)
    
    # 转换日期
    day_chinese = format_day_to_chinese(day_str)
    
    return (year_chinese, month_chinese, day_chinese, 
            year_str, month_str, day_str)


def chinese_amount(amount):
    """
    将数字金额转换为中文大写金额（财务标准格式）
    例如：120.23 -> 壹佰贰拾元贰角叁分
          100.50 -> 壹佰元伍角
          0.01   -> 零元壹分
    """
    if amount == 0:
        return "零元整"
    
    # 中文大写数字
    chinese_num = {
        0: '零', 1: '壹', 2: '贰', 3: '叁', 4: '肆',
        5: '伍', 6: '陆', 7: '柒', 8: '捌', 9: '玖'
    }
    
    # 整数部分单位
    unit_int = ['', '拾', '佰', '仟', '万', '拾', '佰', '仟', '亿']
    
    # 分离整数和小数部分
    integer_part = int(amount)
    decimal_part = int(round((amount - integer_part) * 100))
    
    result = []
    
    # 处理整数部分
    if integer_part == 0:
        result.append('零')
    else:
        int_str = str(integer_part)
        int_len = len(int_str)
        zero_flag = False
        
        for i, ch in enumerate(int_str):
            digit = int(ch)
            pos = int_len - i - 1
            
            if digit == 0:
                if not zero_flag and pos != 0:
                    result.append('零')
                    zero_flag = True
            else:
                if zero_flag:
                    zero_flag = False
                result.append(chinese_num[digit])
                if pos > 0:
                    result.append(unit_int[pos])
        
        result.append('元')
    
    # 处理小数部分
    if decimal_part == 0:
        result.append('整')
    else:
        jiao = decimal_part // 10
        fen = decimal_part % 10
        
        if jiao > 0:
            result.append(chinese_num[jiao] + '角')
        if fen > 0:
            result.append(chinese_num[fen] + '分')
    
    return ''.join(result)


# 采购付款.付款日期批量更改
@any_route("/api/saier/payment/update/payment_date", methods=["POST", "GET"])
@require_token
async def view_payment_update_payment_date(request):
    j = await request.json()
    user = request.current_user
    s = Session()
    try:
        rids = j.get('rids', [])
        date = j.get('date', '')
        org = get_user_path(user.username)
        position = org.get('position', '')
        if '财务' not in position:
            return json_result(-1, "无操作权限")
        
        d1 = ArabiaToChinese(date[0:1])
        d2 = ArabiaToChinese(date[1:2])
        d3 = ArabiaToChinese(date[2:3])
        d4 = ArabiaToChinese(date[3:4])
        fkn = d1 + d2 + d3 + d4
        d5 = ArabiaToChinese(date[5:6])
        d6 = ArabiaToChinese(date[6:7])
        dd6 = date[6:7]
        dd5 = date[5:6]
        if dd6 == '0':
            d9 = '零' + d5 + '拾'
            fky = d9
        else:
            fky = d5 + '拾' + d6
        if dd5 == '0':
            fky = d5 + d6
        d7 = ArabiaToChinese(date[8:9])
        d8 = ArabiaToChinese(date[9:10])
        dd8 = date[9:10]
        dd7 = date[8:9]
        if dd8 == '0':
            t = '零' + d7 + '拾'
            fkr = t
        else:
            fkr = d7 + '拾' + d8
        if dd7 == '0':
            fkr = d7 + d8
        tmpstr = []
        for rid in rids:
            d = s.query(gchk).filter(gchk.rid == rid).first()
            if not d:
                continue
            d.hkrq = date
            d.fksb = '是'
            d.nsz = date[:4]
            d.ysz = date[5:7]
            d.rsz = date[8:10]
            d.fkn = fkn
            d.fky = fky
            d.fkr = fkr
            d.modi_uid = user.rid
            d.mtime = time.strftime('%Y-%m-%d %H:%M:%S')
            s.add(d)
            s.query(cgfkhdsheet).filter(cgfkhdsheet.sb == d.sb).update({"hkrq": date}, synchronize_session=False)
            c = s.query(gchksheet).filter(gchksheet.pid == d.rid).all()
            for r in c:
                s.query(fkspsheet).filter(fkspsheet.cywyzd == r.cywyzd).update({"fkrq": date}, synchronize_session=False)
                r.hkrq = date
                s.add(r)

            s.query(gchksheet3).filter(gchksheet3.pid == d.rid).update({"hkrq": date}, synchronize_session=False)
            tmpstr.append('付款日期: ' + str(date) + '；付款金额: ' + str(d.hkje) + '；合同金额: ' + str(d.htje) + '；发票金额: ' + str(d.fpje) + '；付款合计: ' 
                + str(d.fkhj) + '；应付合计: ' + str(d.yfhj) + '；发票或名称: ' + str(d.gcmc) + '；外销发票: ' + str(d.wxfp) + '；采购合同： ' + str(d.cght))
        
        out_name = ''
        if len(tmpstr) > 0:
            content = '\n'.join(tmpstr)
            out_name = f"{user.username}_{time.strftime('%Y-%m-%d')}_付款明细.txt"
            out_path = os.path.join(config.tmp_path, out_name)
            write_file(out_path, content, 'w')

        s.commit()
        return json_result(1, '操作成功', out_name)
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        pass

# 采购付款.日期批量小写
@any_route("/api/saier/payment/update/date_lower", methods=["POST", "GET"])
@require_token
async def view_payment_update_date_lower(request):
    j = await request.json()
    user = request.current_user
    s = Session()
    try:
        rids = j.get('rids', [])

        u = s.query(sys_user_role).filter(sys_user_role.user_id == user.rid).all()
        role_ids = [r.role_id for r in u]
        flag = 0
        for r in role_ids:
            c = s.query(sys_role).filter(sys_role.rid == r).first()
            if c and '财务' in c.name:
                flag = 1
                break
        if flag == 0:
            return json_result(-1, "无操作权限")
        
        for rid in rids:
            d = s.query(gchk).filter(gchk.rid == rid).first()
            if not d:
                continue
            if d.hkrq == None or d.hkrq == "":
                continue
            hkrq = str(d.hkrq)
            n = hkrq[0:4]
            y = hkrq[5:7]
            r = hkrq[8:10]

            d.modi_uid = user.rid
            d.mtime = time.strftime('%Y-%m-%d %H:%M:%S')
            d.nsz = n
            d.ysz = y
            d.rsz = r
            s.add(d)

        s.commit()
        return json_result(1, "success")
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        pass


# 采购付款.当前现金审核
@any_route("/api/saier/payment/amount_audit", methods=["POST", "GET"])
@require_token
async def view_payment_amount_audit(request):
    j = await request.json()
    user = request.current_user
    s = Session()
    try:
        rids = j.get('rids', [])
        org = get_user_path(user.username)
        roles = []
        path = org.get('path', '')
        dh = 0
        if '优讯' in path or '优鼎' in path:
            dh = 1
        u = s.query(sys_user_role).filter(sys_user_role.user_id == user.rid).all()
        role_ids = [r.role_id for r in u]
        flag = 0

        for r in role_ids:
            c = s.query(sys_role).filter(sys_role.rid == r).first()
            if c and c.name and '财务' in c.name:
                flag = 1
                roles.append(c.name)

        if flag == 0:
            return json_result(-1, "无操作权限")

        
        for rid in rids:
            d = s.query(gchk).filter(gchk.rid == rid).first()
            if not d:
                continue
            if dh == 1 or '义乌财务' not in roles:
                d.fksh = '是'
                d.shrq = time.strftime('%Y-%m-%d')
                d.modi_uid = user.rid
                d.mtime = time.strftime('%Y-%m-%d %H:%M:%S')
                s.add(d)

        s.commit()
        return json_result(1, "success")
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        pass


# 采购付款.刷印费合计
@any_route("/api/saier/payment/printing_cost", methods=["POST", "GET"])
@require_token
async def view_payment_printing_cost(request):
    j = await request.json()
    user = request.current_user
    s = Session()
    try:
        rids = j.get('rids', [])
        u = s.query(sys_user_role).filter(sys_user_role.user_id == user.rid).all()
        role_ids = [r.role_id for r in u]
        flag = 0

        for r in role_ids:
            c = s.query(sys_role).filter(sys_role.rid == r).first()
            if c and c.name and '财务' in c.name:
                flag = 1
                break

        if flag == 0:
            return json_result(-1, "无操作权限")
        
        cjs1 = 0.07
        yhs1 = 0.0003
        for rid in rids:
            d = s.query(gchk).filter(gchk.rid == rid).first()
            if not d:
                continue
            flag = 0
            se = float(d.se) if d.se != None else 0
            fpje = float(d.fpje) if d.fpje != None else 0
            zzsl = float(d.zzsl) if d.zzsl != None else 0
            if se > 0 and fpje > 0 and d.fpwk == '否':
                if zzsl == 3:
                    d.cjs = round((se * cjs1) / 2 * 100) / 100
                    d.yhs = round((fpje / 1.03 * yhs1) / 2 * 100) / 100
                    d.sjhj = round(((fpje / 1.03 * yhs1) / 2 + (se * cjs1) / 2 + se) * 100) / 100
                    flag = 1
                if zzsl > 3:
                    d.sjhj = round(se * 100) / 100
                    flag = 1
            if flag == 1:
                d.modi_uid = user.rid
                d.mtime = time.strftime('%Y-%m-%d %H:%M:%S')
                s.add(d)

        s.commit()
        return json_result(1, "success")
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        pass



# 采购付款.刷合同详情
@any_route("/api/saier/payment/contract_details", methods=["POST", "GET"])
@require_token
async def view_payment_contract_details(request):
    j = await request.json()
    user = request.current_user
    s = Session()
    try:        
        d = s.query(gchk).filter(func.ifnull(gchk.kpxh,'') != '').order_by(gchk.sid.desc()).limit(100).all()
        for r in d:
            kpxh = r.kpxh
            htxq1 = None
            c = s.query(kaiptzsheet1).filter(kaiptzsheet1.kpxh == kpxh).all()
            for l in c:
                l.sfkp = '是'
                htxq1 = str(l.htxs)[:240] if l.htxs != None else ''
                s.add(l)
            r.htxq1 = htxq1
            r.modi_uid = user.rid
            r.mtime = time.strftime('%Y-%m-%d %H:%M:%S')
            s.add(r)

        s.commit()
        return json_result(1, "success")
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        pass

# 采购付款.刷新业务地区
@any_route("/api/saier/payment/business_area", methods=["POST", "GET"])
@require_token
async def view_payment_business_area(request):
    j = await request.json()
    user = request.current_user
    s = Session()
    try:   
        rids = j.get('rids', [])
        for rid in rids:
            r = s.query(gchk).filter(gchk.rid == rid).first()
            if not r:
                continue
            jsrm = r.jsrm
            if jsrm == None or jsrm == '':
                continue
            ywdq = ''
            c = s.query(ywrybiao.ssdq).filter(ywrybiao.yhm == jsrm).first()
            if c:
                ywdq = c.ssdq
            if ywdq == None or ywdq == '':
                c = s.query(ywrylx.dq).filter(ywrylx.ryxm == jsrm).first()
                if c:
                    ywdq = c.ywdq
            if ywdq == None or ywdq == '':
                continue
            r.ywdq = ywdq
            r.modi_uid = user.rid
            r.mtime = time.strftime('%Y-%m-%d %H:%M:%S')
            s.add(r)

        s.commit()
        return json_result(1, "success")
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        pass

    

# 采购付款.完成开票
@any_route("/api/saier/payment/complete_invoice", methods=["POST", "GET"])
@require_token
async def view_payment_complete_invoice(request):
    j = await request.json()
    user = request.current_user
    s = Session()
    try:   
        fphm = j.get('fphm', '')
        org = get_user_path(user.username)
        position = org.get('position', '')
        if '财务' not in position:
            return json_result(-1, "无操作权限")

        s.query(gchk).filter(gchk.wxfp == fphm, gchk.fpwk == '否').update({"jdsj": time.strftime('%Y-%m-%d'), 'mtime': time.strftime('%Y-%m-%d %H:%M:%S'), 'modi_uid': user.rid}, synchronize_session=False)

        s.commit()
        return json_result(1, "success")
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        pass


"""
批量实际出货年月Excel
对应原Pascal: 批量实际出货年月
"""
@any_route('/api/advance_payment/batch_update_year_month', methods=['POST'])
@require_token
async def view_advance_payment_batch_update_year_month(request):
    # 1. 检查用户权限
    s = Session()
    try:
        user = request.current_user
        j = await request.form()
        module = form_value(j, 'module', '采购货款')
        c = s.query(cyzglsheet.rid).filter(cyzglsheet.zm == '采购付款批量实际年月', cyzglsheet.xm == user.username).first()
        if not c:
            return json_result(-1, '您没有权限执行此操作')

        # 2. 获取上传的Excel文件
        file = form_value(j,'file',None)
        if is_none(file):
            return json_result(ERR_PARAM_NOT_ENOUGH)
        # 保存临时文件
        temp_dir = config.tmp_path
        temp_file = os.path.join(temp_dir, f"{uuid.uuid4()}_{file.filename}")
        # logger.error(f"Saving uploaded file to: {temp_file}")
        data = await file.read()
        write_file(temp_file, data, 'wb', encoding=None)
        
        # 3. 读取Excel文件
        wb = load_workbook(temp_file)
        ws = wb.worksheets[0]  # 读取第一个工作表
        # 4. 处理每一行数据
        # duplicates = []  # 存储重复记录信息
        row_idx = 2  # 从第2行开始（第1行是表头）
        
        if module == '预付货款':
            o = yfhk
        else:
            o = gchk
        while True:
            # 读取Excel单元格数据
            wxfp = str(ws.cell(row=row_idx, column=1) or '').strip()  # A列 - 外销发票
            date = str(ws.cell(row=row_idx, column=2) or '')  # B列 - 实际年月
            # 如果wxfp为空，结束循环
            if not wxfp:
                break
            d = s.query(o).filter(o.rid == wxfp).first()
            if d:
                d.sjchny = date
                d.modi_uid = user.rid
                d.modi_time = time.strftime("%Y-%m-%d %H:%M:%S")
                s.add(d)
            row_idx += 1
            
        s.commit()
        return json_result(0, "批量更新成功")
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()
        wb.close()


# 采购付款.批量更新出货日期--权限校验
@any_route("/api/saier/payment/shipment_date_check", methods=["POST", "GET"])
@require_token
async def view_payment_shipment_date_check(request):
    j = await request.json()
    user = request.current_user
    s = Session()
    try:   
        c = s.query(cyzglsheet).filter(cyzglsheet.xm == user.username, cyzglsheet.zm == "采购付款批量出货日期").first()
        if not c:
            return json_result(-1, "您没有权限操作")
        return json_result(1, "校验成功")
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        pass

def get_date_weekday(date):
    """获取日期对应的 日期 星期几(1-7)"""
    # python: 0=周一, 6=周日
    date = date if isinstance(date, datetime) else datetime.strptime(date, "%Y-%m-%d")
    python_weekday = date.weekday()  # 0-6, 周一=0
    # 转换: 周一(0)->1, 周二(1)->2, ..., 周六(5)->6, 周日(6)->7
    return python_weekday + 1

# 采购付款.批量更新出货日期-更新数据
@any_route("/api/saier/payment/shipment_date_update", methods=["POST", "GET"])
@require_token
async def view_payment_shipment_date_update(request):
    j = await request.json()
    user = request.current_user
    s = Session()
    try:   
        rids = j.get('rids', [])
        ksrq = j.get('date', '')
        # kind = j.get('kind', '当前记录') # 当前记录、全部记录
        jsts = 0
        sz1 = 0
        # if kind != '全部记录':
        #     kind = '当前记录'
        if ksrq == None or ksrq == '':
            return json_result(-1, '请输入查询起始日期,格式"2010-01-18"')
        
        c = s.query(zx).filter(zx.ly == '财务付款结算期').first()
        if c:
            jsts = float(c.cs) if c.cs != None else 0
            sz1 = float(c.sz1) if c.sz1 != None else 0
        c = s.query(cyzglsheet).filter(cyzglsheet.xm == user.username, cyzglsheet.zm == "采购付款批量出货日期").first()
        if not c:
            return json_result(-1, "您没有权限操作")
        rq_date = datetime.strptime(str(ksrq), '%Y-%m-%d')
        # 计算起始日期
        current_date = rq_date + timedelta(days=jsts)
        rq = get_date_weekday(current_date)
        i1 = 0
        # 向后查找直到找到符合条件的日期
        for i2 in range(1, 8):
            if rq != sz1:
                current_date = rq_date + timedelta(days=jsts + i2)
                rq = get_date_weekday(current_date)
                i1 += 1
        
        zsrq = (rq_date + timedelta(days=jsts + i1)).strftime('%Y-%m-%d')
        for rid in rids:
            r = s.query(gchk).filter(gchk.rid == rid).first()
            if not r:
                continue
            r.chrq = ksrq
            r.yjfk = zsrq
            r.modi_uid = user.rid
            r.mtime = time.strftime('%Y-%m-%d %H:%M:%S')
            s.add(r)

        s.commit()
        return json_result(1, "操作成功", zsrq)
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        pass


class InvoiceTaxExcelGenerator:
    """发票税金Excel生成器"""
    
    def __init__(self):
        """
        初始化
        :param db_url: 数据库连接URL，如不提供则使用默认配置
        """
        # 定义样式
        self._init_styles()
    
    def _init_styles(self):
        """初始化样式"""
        # 边框样式
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 标题行样式
        self.header_font = Font(bold=True, size=11)
        self.header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center')
        
        # 普通单元格样式
        self.cell_alignment = Alignment(horizontal='right', vertical='center')
        self.cell_alignment_center = Alignment(horizontal='center', vertical='center')
        self.cell_alignment_left = Alignment(horizontal='left', vertical='center')
        
        # 合计行样式
        self.total_font = Font(bold=True)
        self.total_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    
    def _set_currency_format(self, cell, currency_type):
        """设置货币格式"""
        if currency_type in ['USD$', 'USD', '$']:
            cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE  # '$#,##0.00'
        else:
            cell.number_format = '¥#,##0.00'  # 人民币格式
    
    def _apply_cell_style(self, cell, border=True, alignment='right', font_bold=False, fill=None):
        """应用单元格样式"""
        if border:
            cell.border = self.thin_border
        if alignment == 'right':
            cell.alignment = self.cell_alignment
        elif alignment == 'left':
            cell.alignment = self.cell_alignment_left
        elif alignment == 'center':
            cell.alignment = Alignment(horizontal='center', vertical='center')
        if font_bold:
            cell.font = Font(bold=True)
        if fill:
            cell.fill = fill
    
    def generate_from_data(self, invoice_data_list, output_path, company_info=None):
        """
        根据传入的数据数组生成税金报告
        
        :param invoice_data_list: 发票数据列表，每个元素为字典，包含以下字段：
            - wxfp: 外销发票号
            - gcfp: 发票号码
            - fpje: 总金额
            - tsl: 税率（如3表示3%）
            - tse: 税额
            - fkbz: 币种（USD$、USD、$、CNY等）
            - wstt: 外销发票号（抬头，仅第一行使用）
            - gcmc: 公司名称（抬头，仅第一行使用）
        :param output_path: 输出文件路径
        :param company_info: 公司信息字典，包含 wstt 和 gcmc
        :return: 是否成功
        """
        try:
            if not invoice_data_list:
                logger.error("没有数据")
                return {'code': -1, 'message': '没有数据'}
            
            # 创建新工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "发票税金计算"
            
            # 设置列宽
            column_widths = {
                'A': 15, 'B': 15, 'C': 14, 'D': 8, 'E': 14,
                'F': 12, 'G': 18, 'H': 12, 'I': 14,
                'N': 12, 'O': 12  # 额外公式列
            }
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # 设置行高
            ws.row_dimensions[1].height = 20
            ws.row_dimensions[2].height = 20
            ws.row_dimensions[3].height = 20
            ws.row_dimensions[4].height = 25
            
            # ========== 写入抬头信息 ==========
            if company_info:
                ws['A1'] = company_info.get('wstt', '')
                ws['A3'] = company_info.get('gcmc', '')
            else:
                # 从第一条数据中获取抬头信息
                ws['A1'] = invoice_data_list[0].get('wstt', '')
                ws['A3'] = invoice_data_list[0].get('gcmc', '')
            
            ws.merge_cells(f'A1:I1')
            ws[f'A1'].alignment = self.cell_alignment_center
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells(f'A3:C3')
            ws[f'A3'].alignment = self.cell_alignment_left
            ws['A3'].font = Font(bold=True, size=11)
            
            # ========== 创建表头（第4行）==========
            headers = ['外销发票号', '发票号码', '总金额', '税率', '税额', 
                    '城建税', '教育费附加及地方教育费附加', '印花税', '税金合计']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=4, column=col_idx, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
                cell.border = self.thin_border
            
            # ========== 写入数据行 ==========
            data_start_row = 5  # 数据从第5行开始
            first_currency = None
            
            for idx, record in enumerate(invoice_data_list):
                row_num = data_start_row + idx
                
                currency = record.get('fkbz', 'CNY')
                if first_currency is None:
                    first_currency = currency
                
                # 写入各列数据
                ws[f'A{row_num}'] = record.get('wxfp', '')
                ws[f'B{row_num}'] = record.get('gcfp', '')
                
                # C列：总金额（带货币格式）
                c_cell = ws[f'C{row_num}']
                c_cell = record.get('fpje', 0)
                self._set_currency_format(c_cell, currency)
                
                # D列：税率（百分比显示）
                d_cell = ws[f'D{row_num}']
                d_cell = (record.get('tsl', 0)) / 100
                d_cell.number_format = '0.00%'
                
                # E列：税额
                e_cell = ws[f'E{row_num}']
                e_cell = record.get('tse', 0)
                self._set_currency_format(e_cell, currency)
                
                # 如果税额未提供，自动计算
                if record.get('tse', 0) == 0 and record.get('fpje', 0) > 0:
                    tax_rate = record.get('tsl', 0) / 100
                    e_cell = record.get('fpje', 0) * tax_rate
                    self._set_currency_format(e_cell, currency)
                
                # F列：城建税 = 税额 * 0.07 / 2
                tse_value = record.get('tse', 0)
                if tse_value == 0 and record.get('fpje', 0) > 0:
                    tax_rate = record.get('tsl', 0) / 100
                    tse_value = record.get('fpje', 0) * tax_rate
                
                f_cell = ws[f'F{row_num}']
                f_cell = tse_value * 0.07 / 2
                self._set_currency_format(f_cell, currency)
                
                # G列：教育费附加 = 税额 * 0.05 / 2
                g_cell = ws[f'G{row_num}']
                g_cell = tse_value * 0.05 / 2
                self._set_currency_format(g_cell, currency)
                
                # H列：印花税 = (总金额 / 1.03 * 0.0003) / 2
                h_cell = ws[f'H{row_num}']
                h_cell = (record.get('fpje', 0) / 1.03 * 0.0003) / 2
                self._set_currency_format(h_cell, currency)
                
                # I列：税金合计公式 = E+F+G+H
                ws[f'I{row_num}'] = f'=E{row_num}+F{row_num}+G{row_num}+H{row_num}'
                
                # N列公式：I列 / C列
                ws[f'N{row_num}'] = f'=I{row_num}/C{row_num}'
                
                # O列公式：(I列 - G列) / C列
                ws[f'O{row_num}'] = f'=(I{row_num}-G{row_num})/C{row_num}'
                
                # 应用边框和对齐
                for col in range(1, 10):  # A-I列
                    self._apply_cell_style(ws.cell(row=row_num, column=col), border=True, alignment='right')
                
                # A、B列左对齐
                ws[f'A{row_num}'].alignment = self.cell_alignment_left
                ws[f'B{row_num}'].alignment = self.cell_alignment_left
            
            data_count = len(invoice_data_list)
            last_data_row = data_start_row + data_count - 1
            
            # ========== 合计行（数据行下方一行）==========
            total_row = last_data_row + 1
            
            # 合计行标签
            ws[f'A{total_row}'] = '合计'
            ws.merge_cells(f'A{total_row}:B{total_row}')
            self._apply_cell_style(ws[f'A{total_row}'], border=True, alignment='center', font_bold=True, fill=self.total_fill)
            ws[f'C{total_row}'] = '-----'
            
            # C列合计
            c_cell = ws[f'C{total_row}']
            c_cell = f'=SUM(C{data_start_row}:C{last_data_row})'
            self._set_currency_format(c_cell, first_currency)
            self._apply_cell_style(c_cell, border=True, alignment='right', font_bold=True, fill=self.total_fill)
            
            # E-I列合计
            for col_name, col_idx in [('E', 5), ('F', 6), ('G', 7), ('H', 8), ('I', 9)]:
                cell = ws.cell(row=total_row, column=col_idx)
                cell = f'=SUM({col_name}{data_start_row}:{col_name}{last_data_row})'
                self._set_currency_format(cell, first_currency)
                self._apply_cell_style(cell, border=True, alignment='right', font_bold=True, fill=self.total_fill)
            
            # A列和D列合计行样式
            for col in [1, 4]:  # A列和D列
                cell = ws.cell(row=total_row, column=col)
                self._apply_cell_style(cell, border=True, alignment='right', font_bold=True, fill=self.total_fill)
            
            # ========== 支付税金行（合计行下方一行）==========
            payment_row = total_row + 1
            
            ws[f'A{payment_row}'] = '支付税金'
            ws.merge_cells(f'A{payment_row}:D{payment_row}')
            # ws[f'A{payment_row}'].alignment = self.cell_alignment_center
            self._apply_cell_style(ws[f'A{payment_row}'], border=True, alignment='center', font_bold=True, fill=self.total_fill)
            
            # I列支付税金合计
            i_cell = ws[f'I{payment_row}']
            i_cell = f'=SUM(I{data_start_row}:I{last_data_row})'
            self._set_currency_format(i_cell, first_currency)
            self._apply_cell_style(i_cell, border=True, alignment='right', font_bold=True, fill=self.total_fill)
            
            # 其他列支付税金行样式
            for col in [1, 3, 4, 5, 6, 7, 8]:  # A, C, D, E, F, G, H列
                cell = ws.cell(row=payment_row, column=col)
                self._apply_cell_style(cell, border=True, alignment='right', font_bold=True, fill=self.total_fill)
            
            # 为签名行设置边框
            for row in range(5, payment_row + 1):
                for col in range(1, 10):
                    border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    ws.cell(row=row, column=col).border = border
            ws[f'A{payment_row - 1}'].alignment = self.cell_alignment_center
            ws[f'A{payment_row}'].alignment = self.cell_alignment_center
            # ========== 底部签名行 ==========
            sign_row = payment_row + 2
            ws[f'A{sign_row}'] = f'打印日期：{datetime.now().strftime("%Y-%m-%d")}'
            ws[f'D{sign_row}'] = '财务审核'
            ws[f'G{sign_row}'] = '业务签字'

            
            # 保存文件
            fp = config.tmp_path
            wb.save(os.path.join(fp, output_path))
            logger.error(f"报告已生成: {output_path}, 共处理 {data_count} 条数据")
            
            return {'code': 1, 'message': '报告生成成功', 'data': output_path}
        except Exception as e:    
            logger.error(f"生成报告时发生错误: {str(e)}")
            return {'code': -1, 'message': '报告生成失败', 'data': None}

    def generate_from_db(self, rids, output_path):
        """
        从数据库查询数据生成报告
        
        :param rid: gchk表中的rid值
        :param output_path: 输出文件路径
        :return: 是否成功
        """

        s = Session()
        try:
            invoice_data_list = []
            for rid in rids:
                d = s.query(gchk).filter(gchk.rid==rid).first()
                if not d:
                    continue
                invoice_data_list.append(alchemy_object_to_dict(d))
                d.fpsj = '是'  # 更新发票数据标志，表示已生成报告
                s.add(d)

            # 生成报告
            res = self.generate_from_data(invoice_data_list, output_path)
            
            # 更新数据库中的 fpsj 标志
            if res['code'] == 1:
                s.commit()
            
            return res
        except Exception as e:
            s.rollback()
            logger.error(f"处理失败: {str(e)}")
            return {'code': -1, 'message': '处理失败', 'data': None}
        finally:
            s.close()

# 采购付款.3%当前发票税金格式，包括选中记录
@any_route("/api/saier/payment/invoice_report", methods=["POST", "GET"])
@require_token
async def view_payment_invoice_report(request):
    j = await request.json()
    user = request.current_user
    s = Session()
    try:   
        rids = j.get('rids', [])
        generator = InvoiceTaxExcelGenerator()
        res = generator.generate_from_db(rids, f"3%发票税金格式{time.strftime('%Y%m%d%H%M%S')}.xlsx")

        return json_result(1, "操作成功", res.get('data', ''))
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        pass


# 采购付款.批量诚信导出
@any_route("/api/saier/payment/integrity_export", methods=["POST", "GET"])
@require_token
async def view_payment_integrity_export(request):
    j = await request.json()
    user = request.current_user
    s = Session()
    try:   
        ksrq = j.get('ksrq', '')
        jsrq = j.get('jsrq', '')
        fpje = j.get('fpje', 0)
        if ksrq == None or ksrq == '':
            ksrq = time.strftime('%Y-%m-%d')
        if jsrq == None or jsrq == '':
            jsrq = time.strftime('%Y-%m-%d')
        if fpje == '' or fpje == None:
            fpje = 0
        if ksrq > jsrq:
            return json_result(-1, '查询起始日期不能大于结束日期')
        ksrq = ksrq[:10]
        jsrq = jsrq[:10]
        index = 1
        d = run_sql(f"select sum(fpje) as fpje1,sum(fkhj) as fkhj1,gcmc1 FROM gchk WHERE  (hkrq >='{ksrq}') AND (hkrq <= '{jsrq}')  group BY gcmc1")
        if len(d) == 0:
            return json_result(-1, '没有查询到数据')
        wb = Workbook()
        ws = wb.active
        ws.title = "诚信导出"
        ws['A1'] = '生产厂家'
        ws['B1'] = '诚信识别'
        ws['C1'] = '确认日期'
        ws['D1'] = '引入日期'
        ws['E1'] = '出货工厂'
        ws['F1'] = '引入公司'
        for r in d:
            if r.get('fpje1', 0) < fpje and r.get('fkhj1', 0) < fpje:
                continue
            d = s.query(cxgc).filter(cxgc.gcmc == r.get('gcmc1')).first()
            if d:
                continue
            index += 1
            ws['A' + str(index)] = r.get('gcmc1', '')
            ws['B' + str(index)] = ''
            ws['C' + str(index)] = ''
            ws['D' + str(index)] = time.strftime('%Y-%m-%d')
            ws['E' + str(index)] = r.get('gcmc1', '')
            ws['F' + str(index)] = r.get('yrgs', '')
        if index == 1:
            return json_result(-1, '没有查询到数据')
        
        fp = config.tmp_path
        output_path = f"诚信导出{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        wb.save(os.path.join(fp, output_path))

        return json_result(1, "操作成功", output_path)
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        pass



"""
电子口岸的数据引入Excel
对应原Pascal: 电子口岸的数据引入
"""
@any_route('/api/saier/payment/electronic_export', methods=['POST'])
@require_token
async def view_payment_electronic_export(request):
    # 1. 检查用户权限
    s = Session()
    try:
        user = request.current_user
        j = await request.form()
        # 2. 获取上传的Excel文件
        file = form_value(j,'file',None)
        if is_none(file):
            return json_result(ERR_PARAM_NOT_ENOUGH)
        ksrq = form_value(j, 'ksrq', '')
        bcbh = form_value(j, 'bcbh', '')
        # 保存临时文件
        temp_dir = config.tmp_path
        temp_file = os.path.join(temp_dir, f"{uuid.uuid4()}_{file.filename}")
        # logger.error(f"Saving uploaded file to: {temp_file}")
        data = await file.read()
        write_file(temp_file, data, 'wb', encoding=None)
        # 3. 读取Excel文件
        wb = load_workbook(temp_file)
        ws = wb.worksheets[0]  # 读取第一个工作表
        # 4. 处理每一行数据
        # duplicates = []  # 存储重复记录信息
        row_idx = 2  # 从第2行开始（第1行是表头）
        i = 0
        flag = 0
        new_data = []
        while True:
            # 读取Excel单元格数据
            ZJZ = '9999999'
            CHSL2 = '999999999'
            CHSL1 = '999999999'
            ckbgdh = ws[f"A{row_idx}"] 
            ckhth = ws[f"B{row_idx}"]
            ckrq1 = ws[f"C{row_idx}"]
            hscode = ws[f"G{row_idx}"]
            zwpm = ws[f"H{row_idx}"]
            JLDW3 = ws[f"K{row_idx}"]
            CHSL = ws[f"L{row_idx}"]
            wxzj = ws[f"N{row_idx}"]
            FJLDW = ws[f"Q{row_idx}"]
            JLDW2 = ws[f"S{row_idx}"]
            zjz1 = ws[f"R{row_idx}"]
            myla = ws[f"O{row_idx}"]
            # 如果ckbgdh为空，结束循环
            if not ckbgdh:
                break
            if FJLDW == '千克':
                ZJZ = ws[f"R{row_idx}"]
                CHSL1 = '999999999'
            else:
                ZJZ = '9999999'
                CHSL1 = ws[f"R{row_idx}"]
            # 如果JLDW2是'千克'且ZJZ是'9999999'
            if JLDW2 == '千克' and ZJZ == '9999999':
                ZJZ = ws[f"T{row_idx}"]
                CHSL2 = '999999999'
            else:
                if JLDW2 !="":
                    CHSL2 = ws[f"T{row_idx}"] if ws[f"T{row_idx}"] else 0
                    if float(CHSL2) == 0:
                        CHSL2 = '999999999'
            i = i + 1
            bgdh = ''
            pid = ''
            zs = len(ckbgdh)
            bgdh = ckbgdh[1:zs-3]
            cksbdm = hscode
            c = s.query(hgbmb.ckspdm).filter(hgbmb.hgbm == hscode).first()
            if c:
                cksbdm = c.ckspdm
            c = s.query(bgmxd.rid).filter(bgmxd.bgdh == bgdh, func.ifnull(bgmxd.bgdh, '') != '').first()
            if c:
                pid = c.rid
                a = run_sql(f"select hgbm,zwpm,jldw,chsl,wxzj,zjz,ckhwbgdh FROM bgmxdhbcp where (pid='{pid}') and (hgbm='{hscode}') and (zwpm='{zwpm}') \
                    and (wxzj='{wxzj}') and (((jldw='{JLDW3}') and (chsl='{CHSL}')) or (zjz='{ZJZ}') or ((jldw='{FJLDW}') and (chsl='{CHSL1}')) or ((jldw='{JLDW2}') and (chsl='{CHSL2}')))")
                if len(a) > 0 and  ckbgdh != a[0].get('ckhwbgdh', ''):
                    s.query(kaiptzsheet1).filter(kaiptzsheet1.ckhwbgdh == a[0].get('ckhwbgdh', ''), func.ifnull(kaiptzsheet1.ckhwbgdh, '') != '').update({"ckhwbgdh": ckbgdh}, synchronize_session=False)
                    s.query(gchk).filter(gchk.ckhwbgdh == a[0].get('ckhwbgdh', ''), func.ifnull(gchk.ckhwbgdh, '') != '').update({"ckhwbgdh": ckbgdh}, synchronize_session=False)
            else:
                flag = 1
                ws[f"A{row_idx}"].font = Font(color='FF0000')
            b = {'ckhwbgdh': ckbgdh, 'sbny': ksrq, 'sbpc': bcbh,'glh': str(ksrq) + str(bcbh) + str(i).zfill(8), 'ckfphm': ckhth, 'ckrq': ckrq1, 'ckspdm': cksbdm, 'ckspmc': zwpm, 'jldw': FJLDW, 'cksl': zjz1, 'mylaj': myla}
            new_data.append(b)
            row_idx += 1

        logger.error(f"Processed {len(new_data)} rows from uploaded Excel file. flag = {flag}")
        error_file = ''
        error_path = ''
        if flag == 1:
            error_file = f"电子口岸数据导入错误{time.strftime('%Y%m%d%H%M%S')}.xlsx"
            error_path = os.path.join(config.tmp_path, error_file)
            wb.save(error_path)

        temp_path = os.path.join(config.data_upload_path,'template')
        fn = os.path.join(temp_path, '外贸企业出口退税进货明细申报表.xlsx')
        wb = load_workbook(fn)
        ws = wb.worksheets[0]  # 读取第一个工作表
        row_idx = 8
        for r in new_data:
            bhsjZ = 0
            sysl1 = 0
            sysl2 = 0
            d = run_sql(f"select SUM(bhsj) AS bhsjZ FROM gchk where (ckhwbgdh='{r.get('ckhwbgdh', '')}')")
            if len(d) > 0:
                bhsjZ = d[0].get('bhsjZ', 0)
            d = run_sql(f"select gcfp,sh,bhsj,zzsl,tsl,tse,kpxh FROM gchk where (ckhwbgdh='{r.get('ckhwbgdh', '')}')")
            for c in d:
                row_idx = row_idx + 1
                sysl = c.get('cksl', 0)
                if bhsjZ > 0 and sysl > 0:
                    sl = math.trunc((c.get('bhsj') / bhsjZ) * float(sysl))
                    sysl2 = sysl2 + (c.get('bhsj') / bhsjZ) * float(sysl) - sl
                    if sysl2 >= 1:
                        sl = sl + 1
                        sysl2 = sysl2 - 1

                    sysl1 = sysl - sl
                    if sysl1 <= 1:
                        sl = sl + sysl1
                else:
                    sl = 0
                sysl = sysl - sl
                ws[f"A{row_idx}"] = r.get('sbny', '')
                ws[f"B{row_idx}"] = r.get('sbpc', '')
                ws[f"C{row_idx}"] = r.get('glh', '')
                ws[f"D{row_idx}"] = 'V|增值税'
                ws[f"E{row_idx}"] = '02|增值税专用发票'
                ws[f"F{row_idx}"] = c.get('gcfp', '')
                a = run_sql(f"select sccght FROM zlbinvoice where (NumberOfInvoice=:NumberOfInvoice) and (kpxh=:kpxh)", NumberOfInvoice=c.get('gcfp', ''), kpxh=c.get('kpxh', ''))
                if len(a) > 0:
                    ws[f"G{row_idx}"] = a[0].get('sccght', '')

                ws[f"H{row_idx}"] = c.get('sh', '')
                ws[f"I{row_idx}"] = r.get('ckspdm', '')
                ws[f"J{row_idx}"] = r.get('ckspmc', '')
                ws[f"K{row_idx}"] = r.get('jldw', '')
                ws[f"L{row_idx}"] = float(sl)
                ws[f"M{row_idx}"] = float(c.get('bhsj', 0))
                ws[f"N{row_idx}"] = float(c.get('zzsl', 0))
                ws[f"O{row_idx}"] = float(c.get('tsl', 0))
                ws[f"P{row_idx}"] = float(round((c.get('bhsj', 0) * (c.get('tsl', 0) / 100)) * 100) / 100)
                row_idx += 1

        temp_file = f"外贸企业出口退税进货明细申报表{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        temp_path = os.path.join(config.tmp_path, temp_file)
        wb.save(temp_path)
        if flag == 1:
            out_path = f"电子口岸数据导入{time.strftime('%Y%m%d%H%M%S')}.zip"
            zip_path = os.path.join(config.tmp_path, out_path) # 压缩包路径
            zipFile = zipfile.ZipFile(zip_path, 'w') #生成一个压缩包，第二个参数默认值为'r'，表示读已经存在的zip文件，'w'表示新建一个zip文档或覆盖一个已经存在的zip文档
            if os.path.exists(temp_path):
                zipFile.write(temp_path, temp_file, zipfile.ZIP_DEFLATED) #将sPath的文件重命名为sfilename
            if os.path.exists(error_path):
                zipFile.write(error_path, error_file, zipfile.ZIP_DEFLATED) #将sPath的文件重命名为sfilename
            zipFile.close()
        else:
            out_path = temp_file
 
        return json_result(0, "批量更新成功", out_path)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()

"""
外贸企业出口退税进货明细申报表
对应原Pascal: 外贸企业出口退税进货明细
"""
@any_route('/api/saier/payment/refund_tax_export', methods=['POST'])
@require_token
async def view_payment_refund_tax_export(request):
    # 1. 检查用户权限
    s = Session()
    try:
        user = request.current_user
        j = await request.form()
        # 2. 获取上传的Excel文件
        file = form_value(j,'file',None)
        if is_none(file):
            return json_result(ERR_PARAM_NOT_ENOUGH)
        # 保存临时文件
        temp_dir = config.tmp_path
        temp_file = os.path.join(temp_dir, f"{uuid.uuid4()}_{file.filename}")
        # logger.error(f"Saving uploaded file to: {temp_file}")
        data = await file.read()
        write_file(temp_file, data, 'wb', encoding=None)
        # 3. 读取Excel文件
        wb = load_workbook(temp_file)
        ws = wb.worksheets[0]  # 读取第一个工作表
        # 4. 处理每一行数据
        # duplicates = []  # 存储重复记录信息
        row_idx = 9  # 从第2行开始（第1行是表头）
        flag = 0
        new_data = []
        while True:
            # 读取Excel单元格数据
            sbny = ws[f"A{row_idx}"] 
            sbpc = ws[f"B{row_idx}"]
            glh = ws[f"C{row_idx}"]
            ckhwbgdh = ws[f"D{row_idx}"]
            ckfphm = ws[f"F{row_idx}"]
            ckrq = ws[f"G{row_idx}"]
            ckspdm = ws[f"H{row_idx}"]
            ckspmc = ws[f"I{row_idx}"]
            jldw = ws[f"J{row_idx}"]
            cksl = ws[f"K{row_idx}"]
            mylaj = ws[f"L{row_idx}"]
            # 如果sbny为空，结束循环
            if not sbny:
                break

            b = {'sbny': sbny, 'sbpc': sbpc, 'glh': glh, 'ckhwbgdh': ckhwbgdh, 'ckfphm': ckfphm, 'ckrq': ckrq, 'ckspdm': ckspdm, 'ckspmc': ckspmc, 'jldw': jldw, 'cksl': cksl, 'mylaj': mylaj}
            new_data.append(b)
            row_idx += 1

        temp_path = os.path.join(config.data_upload_path,'template')
        fn = os.path.join(temp_path, '外贸企业出口退税进货明细申报表.xlsx')
        wb = load_workbook(fn)
        ws = wb.worksheets[0]  # 读取第一个工作表
        row_idx = 8
        for r in new_data:
            bhsjZ = 0
            sysl1 = 0
            sysl2 = 0
            d = run_sql(f"select SUM(bhsj) AS bhsjZ FROM gchk where (ckhwbgdh='{r.get('ckhwbgdh', '')}')")
            if len(d) > 0:
                bhsjZ = d[0].get('bhsjZ', 0)
            d = run_sql(f"select gcfp,sh,bhsj,zzsl,tsl,tse,kpxh FROM gchk where (ckhwbgdh='{r.get('ckhwbgdh', '')}')")
            for c in d:
                row_idx = row_idx + 1
                sysl = c.get('cksl', 0)
                if bhsjZ > 0 and sysl > 0:
                    sl = math.trunc((c.get('bhsj') / bhsjZ) * float(sysl))
                    sysl2 = sysl2 + (c.get('bhsj') / bhsjZ) * float(sysl) - sl
                    if sysl2 >= 1:
                        sl = sl + 1
                        sysl2 = sysl2 - 1

                    sysl1 = sysl - sl
                    if sysl1 <= 1:
                        sl = sl + sysl1
                else:
                    sl = 0
                sysl = sysl - sl
                ws[f"A{row_idx}"] = r.get('sbny', '')
                ws[f"B{row_idx}"] = r.get('sbpc', '')
                ws[f"C{row_idx}"] = r.get('glh', '')
                ws[f"D{row_idx}"] = 'V|增值税'
                ws[f"E{row_idx}"] = '02|增值税专用发票'
                ws[f"F{row_idx}"] = c.get('gcfp', '')
                a = run_sql(f"select sccght FROM zlbinvoice where (NumberOfInvoice=:NumberOfInvoice) and (kpxh=:kpxh)", NumberOfInvoice=c.get('gcfp', ''), kpxh=c.get('kpxh', ''))
                if len(a) > 0:
                    ws[f"G{row_idx}"] = a[0].get('sccght', '')

                ws[f"H{row_idx}"] = c.get('sh', '')
                ws[f"I{row_idx}"] = r.get('ckspdm', '')
                ws[f"J{row_idx}"] = r.get('ckspmc', '')
                ws[f"K{row_idx}"] = r.get('jldw', '')
                ws[f"L{row_idx}"] = float(sl)
                ws[f"M{row_idx}"] = float(c.get('bhsj', 0))
                ws[f"N{row_idx}"] = float(c.get('zzsl', 0))
                ws[f"O{row_idx}"] = float(c.get('tsl', 0))
                ws[f"P{row_idx}"] = float(round((c.get('bhsj', 0) * (c.get('tsl', 0) / 100)) * 100) / 100)
                row_idx += 1

        temp_file = f"外贸企业出口退税进货明细申报表{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        temp_path = os.path.join(config.tmp_path, temp_file)
        wb.save(temp_path)
 
        return json_result(0, "批量更新成功", temp_file)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()


"""
风控供应商
对应原Pascal: 风控供应商
"""
@any_route('/api/saier/payment/risk_supplier_export', methods=['POST'])
@require_token
async def view_payment_risk_supplier_export(request):
    # 1. 检查用户权限
    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        ksrq = j.get('ksrq', '')
        jsrq = j.get('jsrq', '')
        jlsl = j.get('jlsl', 0)
        if ksrq == None or ksrq == '':
            ksrq = '2025-01-01'
        if jsrq == None or jsrq == '':
            jsrq = time.strftime('%Y-%m-%d')
        if ksrq > jsrq:
            return json_result(-1, '查询起始日期不能大于结束日期')

        d = s.query(cyzglsheet).filter(cyzglsheet.xm == user.username, cyzglsheet.zm == "风控供应商统计表").first()
        if not d:
            return json_result(-1, '没有风控供应商统计表的权限')
        
        wb = Workbook()
        ws = wb.active
        ws.title = "风控供应商"
        ws['A1'] = '部门'
        ws['B1'] = '业务员'
        ws['C1'] = '出货工厂'
        ws['D1'] = '合同金额'
        ws['E1'] = '发票金额'
        ws['F1'] = '付款金额'
        ws['G1'] = '其中WM金额'
        ws['H1'] = '地址'
        ws['I1'] = '联系人'
        ws['J1'] = '联系电话'
        ws['K1'] = '生产工家'
        gc_list = []
        m = s.query(gchk.gcmc, gchk.gcmc1).filter(gchk.fkdc >= ksrq, gchk.fkdc <= jsrq, func.ifnull(gchk.fkdc, '') != '').group_by(gchk.gcmc, gchk.gcmc1).all()
        for r in m:
            if r.gcmc + '_' + r.gcmc1 in gc_list:
                continue
            gc_list.append(r.gcmc + '_' + r.gcmc1)

        d = s.query(gchk).filter(gchk.chrq >= ksrq, gchk.chrq <= jsrq, gchk.fpwk == "否").all()
        for r in d:
            r.tjje = r.fpje
            s.add(r)
        d = s.query(gchk).filter(gchk.chrq >= ksrq, gchk.chrq <= jsrq, gchk.fpwk == "是").all()
        for r in d:
            r.tjje = r.yfhj
            s.add(r)
        i = 0
        row_idx = 2
        bm_json = {}
        d = s.query(gchk.gcmc, gchk.gcmc1, func.sum(func.ifnull(gchk.fkhj, 0)).label('fkhj1'), func.sum(func.ifnull(gchk.yfhj, 0)).label('yfhj1'), 
            func.sum(func.ifnull(gchk.fpje, 0)).label('fpje1'), func.sum(func.ifnull(gchk.tjje, 0)).label('htje1')
        ).filter(gchk.chrq >= ksrq, gchk.chrq <= jsrq).group_by(gchk.gcmc, gchk.gcmc1).order_by(func.sum(gchk.tjje).desc()).limit(400+jlsl).all()
        for r in d:
            if r.gcmc + '_' + r.gcmc1 in gc_list:
                continue
            i = i + 1
            if i > jlsl:
                break

            bm = ''
            c = s.query(gchk.jsrm).filter(gchk.chrq >= ksrq, gchk.chrq <= jsrq, gchk.gcmc1 == r.gcmc1, gchk.gcmc == r.gcmc).order_by(gchk.tjje.desc()).first()
            if c:
                jsrm = c.jsrm
                if jsrm in bm_json:
                    bm = bm_json[jsrm]
                else:
                    a = s.query(ywrybiao.bm).filter(ywrybiao.yhm == jsrm).first()
                    if a:
                        bm_json[jsrm] = a.bm
                        bm = a.bm

            ws[f"A{row_idx}"] = bm
            ws[f"B{row_idx}"] = jsrm
            ws[f"C{row_idx}"] = r.gcmc1
            ws[f"D{row_idx}"] = float(r.htje1) if r.htje1 else 0
            ws[f"E{row_idx}"] = float(r.fpje1) if r.fpje1 else 0
            ws[f"F{row_idx}"] = float(r.yfhj1) if r.yfhj1 else 0
            if 'WM' in r.gcmc1:
                ws[f"G{row_idx}"] = float(r.fkhj1) if r.fkhj1 else 0
            else:
                ws[f"G{row_idx}"] = 0
            ws[f"K{row_idx}"] = r.gcmc
            if (r.gcmc1 != '' and r.gcmc1 != None) and (r.gcmc != '' and r.gcmc != None):
                a = s.query(ozycs.address,ozycs.cslxr,ozycs.phone,ozycs.sjhm).filter(or_(ozycs.company_name == r.gcmc1, ozycs.cymch == r.gcmc1)).first()
                if a:
                    ws[f"H{row_idx}"] = a.address
                    ws[f"I{row_idx}"] = a.cslxr
                    ws[f"J{row_idx}"] = str(a.phone) + '/' + str(a.sjhm)
                else:
                    a = s.query(ozycs.address,ozycs.cslxr,ozycs.phone,ozycs.sjhm).filter(or_(ozycs.company_name == r.gcmc, ozycs.cymch == r.gcmc)).first()
                    if a:
                        ws[f"H{row_idx}"] = a.address
                        ws[f"I{row_idx}"] = a.cslxr
                        ws[f"J{row_idx}"] = str(a.phone) + '/' + str(a.sjhm)

                s.query(gchk).filter(gchk.gcmc1 == r.gcmc1, gchk.gcmc == r.gcmc, gchk.chrq >= ksrq, gchk.chrq <= jsrq, func.ifnull(gchk.fkdc, '') == '').update({"fkdc": time.strftime("%Y-%m-%d")}, synchronize_session=False)
            row_idx += 1

        temp_file = f"风控供应商统计表{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        temp_path = os.path.join(config.tmp_path, temp_file)
        wb.save(temp_path)
        s.commit()
        return json_result(0, "批量更新成功", temp_file)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()


@any_route("/api/saier/payment/amount_query/check", methods=["POST", "GET"])
@require_token
async def view_payment_amount_query_check(request):
    j = await request.json()
    username = request.current_user.username
    s = Session()
    try:
        data = []
        ksrq = j.get('ksrq', '')
        jsrq = j.get('jsrq', '')
        cxrq = j.get('cxrq', '')
        if ksrq == None or ksrq == '' or jsrq == None or jsrq == '' or cxrq == None or cxrq == '':
            return json_result(-1, '查询日期不能为空')
        if ksrq > jsrq:
            return json_result(-1, '查询起始日期不能大于结束日期')
        
        field = j.get("field", "position")
        position = j.get("position", "财务")
        org = get_user_path(username)
        if not position in org.get(field):
            return json_result(-1, "没有权限查询")
        m = s.query(gchk.wstt).filter(gchk.jdsj >= ksrq, gchk.jdsj <= jsrq, 
            func.ifnull(gchk.fpwk, '') == '否', func.ifnull(gchk.wckp, '') == '是',
            or_(func.ifnull(gchk.jdstate, '') != '已上传', func.ifnull(gchk.jdstate, '') == '',func.ifnull(gchk.jdstate, '') == '无')
        ).group_by(gchk.wstt).all()
        for r in m:
            data.append(r.wstt)
        return json_result(1, "success", {'company': data, 'position': position})
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        pass


"""
报关金额金蝶编号查看
对应原Pascal: 报关金额金蝶编号查看
"""
@any_route('/api/saier/payment/amount_query', methods=['POST'])
@require_token
async def view_payment_amount_query(request):
    # 1. 检查用户权限
    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        ksrq = j.get('ksrq', '')
        jsrq = j.get('jsrq', '')
        cxrq = j.get('cxrq', '')
        wstt = j.get('wstt', '')
        if ksrq == None or ksrq == '' or jsrq == None or jsrq == '' or cxrq == None or cxrq == '':
            return json_result(-1, '查询日期不能为空')
        if ksrq > jsrq:
            return json_result(-1, '查询起始日期不能大于结束日期')
        org = get_user_path(user.username)
        position = org.get('position', '')
        # if '财务' not in position :
        #     return json_result(-1, '没有权限查询')
        
        nbyw = '宁波'
        if '义乌' in position:
            nbyw = '义乌'

        hl = 1
        c = run_sql(f"select hl from xshl where (substring(hsrq,1,7)='{cxrq[:7]}') and ly='{nbyw}金蝶' limit 1")
        hl = c[0].get('hl') if c else 1    
        xx_list = []
        gc_list = []
        m = s.query(gchk.gcmc).filter(gchk.jdsj >= ksrq, gchk.fkdc <= jsrq, gchk.wstt == wstt,
            func.ifnull(gchk.fpwk, '') == '否', func.ifnull(gchk.fpwk, '') == '是',
            or_(func.ifnull(gchk.jdstate, '') != '已上传', func.ifnull(gchk.jdstate, '') == '',func.ifnull(gchk.jdstate, '') == '无')
        ).group_by(gchk.gcmc).all()
        for r in m:
            c = s.query(newcwcs.jdbm).filter(newcwcs.company_name == r.gcmc, gchk.company_name == '个体户' + str(r.gcmc)).first()
            if not c:
                gc_list.append(r.gcmc)
        
        d = s.query(gchk.wxfp).filter(gchk.jdsj >= ksrq, gchk.fkdc <= jsrq, gchk.wstt == wstt,
            func.ifnull(gchk.fpwk, '') == '否', func.ifnull(gchk.fpwk, '') == '是',
            or_(func.ifnull(gchk.jdstate, '') != '已上传', func.ifnull(gchk.jdstate, '') == '',func.ifnull(gchk.jdstate, '') == '无')
        ).group_by(gchk.wxfp).all()
        for r in d:
            bgje = 0
            c = s.query(bgmxd.wstt).filter(bgmxd.ysfp == r.wxfp).first()
            if c:
                if c.RMBkh != '是' and c.hbdm != 'RMB': # or (bm1<>'D') then
                    if c.hbdm != 'USD$' and c.hbdm != 'USD':
                        c2 = s.query(cwhshlsheet.hl).filter(cwhshlsheet.hsny == cxrq[:7], cwhshlsheet.hbdm == c.hbdm).first()
                        if c2:
                            bgje = round(c2.hl * c.bgbgzje * 100) / 100
                    else:
                        bgje = round(hl * c.bgbgzje * 100) / 100
                else:
                    bgje = c.bgbgzje
                xx_list.append('发票号:' + r.wxfp + '报关金额￥：' + str(bgje))

        if len(xx_list) == 0 and len(gc_list) == 0:
            return json_result(-1, '查询成功,无异常数据')
        if len(xx_list) > 0:
            xx_list.insert(0, '无金蝶编码工厂: ')
        if len(gc_list) > 0:
            gc_list.insert(0, '无金蝶报关金额: ')

        txt_list = []
        out_name = f"无金蝶编码工厂或无金蝶报关金额_{time.strftime('%Y-%m-%d')}.txt"
        out_path = os.path.join(config.tmp_path, out_name)
        if len(xx_list) > 0 and len(gc_list) > 0:
            txt_list = xx_list + ['\n\n'] + gc_list
        elif len(xx_list) > 0:
            txt_list = xx_list
        else:
            txt_list = gc_list
        
        val = '\n'.join(txt_list)
        write_file(out_path, val, 'w')

        return json_result(1, "批量更新成功", out_name)
    except Exception as e:

        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()


"""
换单号导出
对应原Pascal: 换单号导出
"""
@any_route('/api/saier/payment/change_no', methods=['POST'])
@require_token
async def view_payment_change_no_export(request):
    # 1. 检查用户权限
    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        rids = j.get('rids', [])

        
        wb = Workbook()
        ws = wb.active
        # ws.title = "换单号导出"
        ws['A1'] = '换 单 号'
        ws['B1'] = '付款日期'
        ws['C1'] = '付款金额'
        index = 1
        for rid in rids:
            d = s.query(gchk).filter(gchk.rid == rid, func.ifnull(gchk.hdh, '') != '').first()
            if not d:
                continue
            index = index + 1
            ws[f"A{index}"] = d.hdh
            ws[f"B{index}"] = d.fkdc
            ws[f"C{index}"] = d.fkhj

        temp_file = f"换单号{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        temp_path = os.path.join(config.tmp_path, temp_file)
        wb.save(temp_path)

        return json_result(0, "换单号导出成功", temp_file)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()

def is_leap_year(year):
    """判断是否为闰年"""
    return (year % 4 == 0 and year % 100 != 0) or (year % 400 == 0)

def get_weekday_of_date(dt):
    """
    获取日期是星期几
    返回: 1=星期一, 2=星期二, ..., 7=星期日
    """
    # Python中星期一=0, 星期日=6，转换为1-7
    weekday = dt.weekday()  # 0=周一, 6=周日
    return weekday + 1  # 转换为1=周一, 7=周日

def calculate_day_of_year(year, month, day):
    """
    计算某天是当年的第几天（考虑闰年）
    返回: 第几天
    """
    month_days = [31, 29 if is_leap_year(year) else 28, 31, 30, 31, 30,
                  31, 31, 30, 31, 30, 31]
    return sum(month_days[:month-1]) + day

def get_week_number_and_code(current_date=None):
    """
    根据当前日期计算周次编号（fkpc）
    
    参数:
        current_date: datetime.date 对象，如果为None则使用当前日期
    
    返回:
        fkpc: 格式为 "年份后两位 + 周次编号" 的字符串
        week_number: 当年的第几周（按自定义规则）
    """
    if current_date is None:
        current_date = date.today()
    # 如果传入的是datetime对象或字符串，转换为date对象
    if isinstance(current_date, datetime):
        current_date = current_date.date()
    elif isinstance(current_date, str):
        current_date = datetime.strptime(current_date, "%Y-%m-%d").date()

    # 1. 格式化日期
    year = current_date.year
    month = current_date.month
    day = current_date.day
    
    # 2. 计算当前日期是当年的第几天
    day_of_year = calculate_day_of_year(year, month, day)
    
    # 3. 获取当年1月1日是星期几
    jan_1 = date(year, 1, 1)
    ts1 = get_weekday_of_date(jan_1)
    
    # 4. 转换星期几（使星期一=1，星期日=7）
    if ts1 == 1:  # 1月1日是星期日
        ts = 7
    else:  # 1月1日是星期一至星期六
        ts = ts1 - 1
    
    # 5. 计算调整后的天数
    if day_of_year > (8 - ts):
        number = day_of_year - 8 + ts
        z = None  # 将在后面计算
    else:
        number = day_of_year
        z = 1
    
    # 6. 判断闰年并调整周次
    leap = is_leap_year(year)
    
    if z != 1:  # 说明 day_of_year > (8 - ts)，即日期在第一个完整周之后
        # 闰年且月份>=3时，需要+1（因为前面按28天计算了2月）
        if leap and month >= 3:
            number += 1
        
        # 计算周次 (number 是从第一周起始日开始计数的天数)
        # 原Delphi逻辑：如果整除7则 z = number/7 + 1，否则 z = number/7 + 2
        if number % 7 == 0:
            z = number // 7 + 1
        else:
            z = number // 7 + 2
    # 如果 z == 1，则保持 z = 1
    
    # 7. 生成 fkpc：年份后两位 + 周次编号
    fkpc = f"{year % 100:02d}{z}"
    
    return fkpc, z

# def get_week_number_simplified(current_date=None):
#     """
#     简化版本：使用ISO周编号标准
#     ISO周编号：周一为一周开始，第一周包含当年的第一个星期四
    
#     返回: (ISO年份, ISO周次)
#     """
#     if current_date is None:
#         current_date = date.today()
    
#     # Python内置的isocalendar()返回 (ISO年份, ISO周次, ISO星期几)
#     iso_year, iso_week, iso_weekday = current_date.isocalendar()
#     return f"{iso_year % 100:02d}{iso_week}", iso_week

# # 示例使用
# if __name__ == "__main__":
#     # 原始逻辑
#     print("=== 原始逻辑 ===")
#     fkpc, week_num = get_week_number_and_code()
#     print(f"当前日期: {date.today()}")
#     print(f"fkpc (原始逻辑): {fkpc}")
#     print(f"周次编号: {week_num}")
    
#     # 测试特定日期
#     test_dates = [
#         date(2026, 1, 1),   # 1月1日
#         date(2026, 2, 1),   # 2月1日
#         date(2026, 5, 5),   # 5月5日
#         date(2024, 2, 29),  # 闰年2月29日
#         date(2024, 3, 1),   # 闰年3月1日
#     ]
    
#     print("\n=== 测试特定日期 ===")
#     for dt in test_dates:
#         fkpc, week = get_week_number_and_code(dt)
#         print(f"{dt}: fkpc={fkpc}, week={week}")
    
#     # 简化版本（ISO标准）
#     print("\n=== ISO标准周次（简化版）===")
#     for dt in test_dates:
#         iso_code, iso_week = get_week_number_simplified(dt)
#         print(f"{dt}: ISO周次={iso_code}")


"""
换单收条打印
对应原Pascal: 换单收条打印
"""
@any_route('/api/saier/payment/change_receipt', methods=['POST'])
@require_token
async def view_payment_change_receipt(request):
    # 1. 检查用户权限
    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        fkzb = j.get('fkzb', 'A')
        rid = j.get('rid')
        org = get_user_path(user.username)
        position = org.get('position', '')
        if '财务' not in position and '总经理' not in position:
            return json_result(-1, '没有权限导出')
        wb = Workbook()
        ws = wb.active
        # ws.title = "换单收条打印"
        for row in range(1, 6):
            ws.row_dimensions[row].width = 24.8
        
        ws.column_dimensions['A'].width = 12.81
        ws.column_dimensions['B'].width = 24.94
        ws.column_dimensions['C'].width = 0.08
        ws.column_dimensions['D'].width = 13.69
        ws.column_dimensions['E'].width = 26.06
        ws.column_dimensions['F'].width = 0.08

        date_str = datetime.now().strftime("%Y-%m-%d")
        fkpc, week = get_week_number_and_code(date_str)
        fkxh = 1
        d = s.query(gchk.fkxh1).filter(gchk.fkpc == fkpc, gchk.fkzb == fkzb).order_by(gchk.fkxh1.desc()).first()
        if d:
            fkxh = d.fkxh1 + 1
        hthm = ''
        cgry = ''
        lxdh = ''
        gc = '0'
        d = s.query(gchk).filter(gchk.rid == rid).first()
        if d:
            hthm = d.cght
            if d.fkpc != None and d.fkpc != "":
                if d.fkpc != fkpc:
                    fkpc = d.fkpc
                if d.fkzb != fkzb:
                    d.fkzb = fkzb
                    d.fkxh1 = fkxh
            else:
                d.fkpc = fkpc
                d.fkzb = fkzb
                d.fkxh1 = fkxh

            if hthm != '' and hthm != None:
                org = get_user_path(user.username)
                if '优景' in org.get('path', ''):
                    gc = '1'
                cgry = ''
                c = s.query(cght).filter(cght.hthm == hthm).first()
                if c:
                    if gc == '1':
                        cgry = c.gdry
                    else:
                        cgry = c.cgry
                if cgry != '' and cgry != None:
                    c = s.query(ywrybiao.ryxm, ywrybiao.lxdh).filter(ywrybiao.yhm == cgry).first()
                    if c:
                        cgry = c.ryxm
                        lxdh = c.lxdh

                ws['B2'] = d.fkpc
                ws['E2'] = d.fkzb + str(d.fkxh1)
                ws['B3'] = d.wxfp
                ws['E3'] = d.cght
                ws['E3'].alignment = Alignment(wrap_text=True)
                ws['B4'] = str(d.chrq)[:10] if d.chrq else ''
                ws['E4'] = str(d.chrq)[:10] if d.chrq else ''
                ws['E4'].alignment = Alignment(wrap_text=True)
                ws['B5'] = cgry
                ws['E5'] = lxdh
                ws['E5'].alignment = Alignment(wrap_text=True)
                ws['B7'] = d.gcmc
                ws['E7'] = d.bank
                ws['B8'] = d.gcmc1
                ws['E8'] = d.zh
                ws['E8'].alignment = Alignment(wrap_text=True)
                ws['B9'] = d.hkje
                ws['E9'] = str(d.hkrq)[:10] if d.hkrq else ''
                ws['E9'].alignment = Alignment(wrap_text=True)
                ws['B10'] = d.fkdx
                ws['E12'] = date_str
                s.add(d)

        temp_file = f"换单收条{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        temp_path = os.path.join(config.tmp_path, temp_file)
        wb.save(temp_path)

        s.commit()
        return json_result(0, "换单收条导出成功", temp_file)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'导出失败: {str(e)}')
    finally:
        s.close()


def generate_zb(n, y, j4, i1):
    """
    生成带前导零的编号
    参数:
        n: 前缀1（如年份）
        y: 前缀2（如月份）
        j4: 总位数（2、3 或 4）
        i1: 序号
    
    返回:
        zb: 格式化的编号字符串
    """
    # 使用 Python 的字符串格式化功能，一行代码搞定
    # :0{j4}d 表示数字补零到 j4 位
    zb = f"{n}{y}{i1:0{j4}d}"
    return zb

def generate_fb(n, y, j5, i, is_zb = True):
    """
    生成带前导零的编号 fb
    参数:
        n: 前缀1（如年份）
        y: 前缀2（如月份）
        j5: 总位数（3、4 或 5）
        i: 当前计数（函数内部会使用 i-1 作为序号）
        is_zb: 是否为 zb 编号
    返回:
        fb: 格式化的编号字符串
    """
    # 序号是 i-1
    # serial_num = i - 1
    if is_zb:
        serial_num = i - 1
    else:
        serial_num = i
    # 使用 Python 的 f-string 格式化，一行搞定
    # :0{j5}d 表示数字补零到 j5 位
    fb = f"{n}{y}{serial_num:0{j5}d}"
    
    return fb

"""
金蝶付货款
对应原Pascal: 金蝶付货款
"""
@any_route('/api/saier/payment/kingdee_payment', methods=['POST'])
@require_token
async def view_payment_kingdee_payment(request):
    # 1. 检查用户权限
    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        cxrq = j.get('cxrq', '')
        rids = j.get('rids', [])
        org = get_user_path(user.username)
        position = org.get('position', '')
        if '财务' not in position and '总经理' not in position:
            return json_result(-1, '没有权限导出')
        template_path = os.path.join(config.data_upload_path, 'template', '付货款.xlsx')
        if not os.path.exists(template_path):
            return json_result(-1, '付货款模板文件不存在')
        wb = load_workbook(template_path)
        ws = wb.active
        ws.title = "金蝶付货款"
        j2 = 0
        j3 = 0
        j4 = 0
        j5 = 0
        zb = ''
        fb = ''
        wstt = ''
        hkrq_list = []
        for rid in rids:
            d = s.query(gchk.wxfp,gchk.kpxh,gchk.gcmc,gchk.hkrq,gchk.wstt).filter(gchk.rid == rid).first()
            if not d:
                continue
            j3 = j3 + 1
            hkrq = str(d.hkrq)[:10] if d.hkrq else ''
            if d.wstt != '' and d.wstt != None:
                wstt = d.wstt
            if hkrq != '' and hkrq != None and hkrq not in hkrq_list:
                hkrq_list.append(hkrq)
                j2 += 1

        year = cxrq[:4]
        month = cxrq[5:7]
        # day = cxrq[-2:]
        weekday, days = calendar.monthrange(int(year), int(month))
        if j2 < 99:
            j4 = 2
        else:
            if j2 < 999:
                j4 = 3
            else:
                j4 = 4
        if j3 < 99:
            j5 = 3
        else:
            if j2 < 999:
                j5 = 4
            else:
                j5 = 5
        
        i = 0
        i1 = 0
        iz = 0
        for r in hkrq_list:
            zpbm = ''
            zpbm1 = ''
            i1 = i1 + 1
            i = i + 1
            iz = i
            i2 = 0
            fpje = 0
            bgpmstr1 = []
            fphm = []
            wfyhzh = []
            fphm1 = ''
            wfyhzh1 = ''
            zb = generate_zb(year[2:], month, j4, j2)
            c = s.query(gchk).filter(gchk.hkrq == r, gchk.wstt == wstt).first()
            if c:
                # fpje = float(c.fpje) if c.fpje else 0
                a = s.query(zx.wb1, zx.wb2).filter(zx.ly == "金蝶账簿编码", zx.mc == c.wstt).first()
                if a:
                    zpbm = a.wb1
                    zpbm1 = a.wb2
            c = s.query(gchk.kpxh).filter(gchk.hkrq == r, gchk.wstt == wstt, func.ifnull(gchk.fpwk, '') == '否').first()
            if c:
                if c.kpxh not in bgpmstr1:
                    bgpmstr1.append(c.kpxh)

            c = s.query(zx.wb1, zx.wb2).filter(zx.ly == "金蝶付货款", zx.mc == wstt).first()
            if c:
                fphm1 = c.wb1
                wfyhzh1 = c.wb2
            if fphm1 not in fphm:
                fphm.append(fphm1)
                wfyhzh.append(wfyhzh1)

            c = s.query(gchk.wfyhmc, gchk.wfyhzh).filter(gchk.hkrq == r, gchk.wstt == wstt).first()
            if c:
                if c.wfyhmc != '' and c.wfyhmc != None and c.wfyhmc not in fphm:
                    fphm.append(c.wfyhmc)
                    wfyhzh.append(c.wfyhzh)
            
            ws['A' + str(iz + 2)] = '1' + str(zb)
            ws['B' + str(iz + 2)] = zpbm
            ws['C' + str(iz + 2)] = str(wstt) + '账簿'
            ws['D' + str(iz + 2)] = str(year) + '-' + str(month) + '-' + str(days)
            ws['E' + str(iz + 2)] = 'PRE001'
            ws['F' + str(iz + 2)] = '记'
            ws['G' + str(iz + 2)] = ''
            ws['H' + str(iz + 2)] = '0'
            ws['I' + str(iz + 2)] = 'False'
            ws['J' + str(iz + 2)] = zpbm1
            ws['K' + str(iz + 2)] = wstt
            ws['L' + str(iz + 2)] = year
            ws['M' + str(iz + 2)] = int(month)

            for l in bgpmstr1:
                c = s.query(func.sum(func.ifnull(gchk.hkje, 0)).label('fpje1'),
                    gchk.gcmc, gchk.kpxh, gchk.wxfp
                ).filter(gchk.hkrq == r, gchk.wstt == wstt, gchk.kpxh == l, func.ifnull(gchk.fpwk, '') == '否'
                ).group_by(gchk.kpxh, gchk.gcmc, gchk.wxfp).first()
                if c:
                    i2 = i2 + 1
                    i = i + 1
                    fpje = float(c.fpje1) if c.fpje1 else 0
                    fb = generate_fb(year, month, j5, i)
                    ws['O' + str(i + 1)] = '1' + fb
                    ws['P' + str(i + 1)] = c.wxfp + '付货款'
                    ws['Q' + str(i + 1)] = '2202.01'
                    ws['R' + str(i + 1)] = '明细应付款'
                    if c.gcmc != '' and c.gcmc != None:
                        c1 = s.query(newcwcs.jdbm).filter(or_(newcwcs.company_name == c.gcmc, newcwcs.company_name == '个体户' + c.gcmc)).first()
                        if c1:
                            ws['AS' + str(i + 1)] = c1.jdbm
                        if len(c.gcmc.strip()) > 8:
                            ws['AT' + str(i + 1)] = c.gcmc
                        else:
                            ws['AT' + str(i + 1)] = '个体户' + c.gcmc
                    ws['BE' + str(i + 1)] = 'PRE001'
                    ws['BF' + str(i + 1)] = '人民币'
                    ws['BG' + str(i + 1)] = 'HLTX01_SYS'
                    ws['BH' + str(i + 1)] = '固定汇率'
                    ws['BI' + str(i + 1)] = '1'
                    ws['BN' + str(i + 1)] = round(fpje * 100) / 100
                    ws['BO' + str(i + 1)] = round(fpje * 100) / 100

            index = 0
            for l in fphm:
                c = s.query(func.sum(func.ifnull(gchk.hkje, 0)).label('fpje2')
                    ).filter(gchk.hkrq == r, gchk.wstt == wstt, gchk.wfyhmc == l, func.ifnull(gchk.fpwk, '') == '否').first()
                if c and c.fpje2 and c.fpje2 > 0:
                    fpje = float(c.fpje2) if c.fpje2 else 0
                    i2 = i2 + 1
                    i = i + 1
                    fb = generate_fb(year, month, j5, i)
                    ws['O' + str(i + 1)] = '1' + fb
                    ws['P' + str(i + 1)] = '付货款'
                    ws['Q' + str(i + 1)] = '2202.01'
                    ws['R' + str(i + 1)] = '明细应付款'
                    ws['W' + str(i + 1)] = 'YHXZ01'
                    ws['X' + str(i + 1)] = '活期'
                    if l != '' and l != None:
                        ws['AB' + str(i + 1)] = l
                    else:
                        ws['AB' + str(i + 1)] = fphm1
                    if wfyhzh[index] != '' and wfyhzh[index] != None:
                        ws['AA' + str(i + 1)] = ''
                    else:
                        ws['AA' + str(i + 1)] = wfyhzh1
                    ws['BE' + str(i + 1)] = 'PRE001'
                    ws['BF' + str(i + 1)] = '人民币'
                    ws['BG' + str(i + 1)] = 'HLTX01_SYS'
                    ws['BH' + str(i + 1)] = '固定汇率'
                    ws['BI' + str(i + 1)] = '1'
                    ws['BN' + str(i + 1)] = round(fpje * 100) / 100
                    ws['BP' + str(i + 1)] = round(fpje * 100) / 100
                index = index + 1

            i = i - 1
        temp_file = f"金蝶付款{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        temp_path = os.path.join(config.tmp_path, temp_file)
        wb.save(temp_path)

        # s.commit()
        return json_result(0, "金蝶付款导出成功", temp_file)
    except Exception as e:
        # s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'导出失败: {str(e)}')
    finally:
        s.close()



"""
金蝶合同入账
对应原Pascal: 金蝶合同入账
"""
@any_route('/api/saier/payment/kingdee_contract', methods=['POST'])
@require_token
async def view_payment_kingdee_contract(request):
    # 1. 检查用户权限
    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        cxrq = j.get('cxrq', '')
        rids = j.get('rids', [])
        org = get_user_path(user.username)
        position = org.get('position', '')
        if '财务' not in position and '总经理' not in position:
            return json_result(-1, '没有权限导出')
        template_path = os.path.join(config.data_upload_path, 'template', '合同入账.xlsx')
        if not os.path.exists(template_path):
            return json_result(-1, '合同入账模板文件不存在')
        wb = load_workbook(template_path)
        ws = wb.active
        ws.title = "金蝶合同入账"
        j2 = 0
        j3 = 0
        j4 = 0
        j5 = 0
        zb = ''
        fb = ''
        wstt = ''
        bgpmstr1 = []
        fphm = []
        for rid in rids:
            d = s.query(gchk.wxfp,gchk.kpxh,gchk.wstt).filter(gchk.rid == rid).first()
            if not d:
                continue
            wxfp = d.wxfp if d.wxfp else ''
            kpxh = d.kpxh if d.kpxh else ''
            if kpxh != '' and kpxh != None and kpxh not in bgpmstr1:
                bgpmstr1.append(kpxh)
                j3 = j3 + 1
            if wxfp != '' and wxfp != None and wxfp not in fphm:
                fphm.append(wxfp)
                j2 += 1

        year = cxrq[:4]
        month = cxrq[5:7]
        # day = cxrq[-2:]
        weekday, days = calendar.monthrange(int(year), int(month))
        if j2 < 99:
            j4 = 2
        else:
            if j2 < 999:
                j4 = 3
            else:
                j4 = 4

        if j3 < 99:
            j5 = 3
        else:
            if j2 < 999:
                j5 = 4
            else:
                j5 = 5
        
        i = 0
        i1 = 0
        iz = 0
        for r in fphm:
            zpbm = ''
            zpbm1 = ''
            i1 = i1 + 1
            i = i + 1
            iz = i
            i2 = 0
            fpje = 0
            se = 0
            c = s.query(gchk).filter(gchk.wxfp == r, func.ifnull(gchk.wstt, '') != '').first()
            if c:
                # fpje = float(c.fpje) if c.fpje else 0
                a = s.query(zx.wb1, zx.wb2).filter(zx.ly == "金蝶账簿编码", zx.mc == c.wstt).first()
                if a:
                    zpbm = a.wb1
                    zpbm1 = a.wb2
            zb = generate_zb(year[2:], month, j4, j2)
            ws['A' + str(iz + 2)] = '3' + str(zb)
            ws['B' + str(iz + 2)] = zpbm
            ws['C' + str(iz + 2)] = str(wstt) + '账簿'
            ws['D' + str(iz + 2)] = str(year) + '-' + str(month) + '-' + str(days)
            ws['E' + str(iz + 2)] = 'PRE001'
            ws['F' + str(iz + 2)] = '记'
            # ws['G' + str(iz + 2)] = ''
            ws['H' + str(iz + 2)] = zpbm1
            ws['I' + str(iz + 2)] = wstt
            ws['J' + str(iz + 2)] = 'False'
            ws['K' + str(iz + 2)] = str(year)
            ws['L' + str(iz + 2)] = str(month)
            fb = generate_fb(year, month, j5, i, False)
            ws['N' + str(iz + 2)] = '3' + fb
            ws['O' + str(iz + 2)] = str(r) + '合同入账'
            ws['P' + str(iz + 2)] = '2221.01.01.01'
            ws['Q' + str(iz + 2)] = '进项税额'
            ws['BB' + str(iz + 2)] = 'PRE001'
            ws['BC' + str(iz + 2)] = '人民币'
            ws['BD' + str(iz + 2)] = 'HLTX01_SYS'
            ws['BE' + str(iz + 2)] = '固定汇率'
            ws['BF' + str(iz + 2)] = '1'

            for l in bgpmstr1:
                c = s.query(func.sum(func.ifnull(gchk.hkje, 0)).label('fpje1'),
                    func.sum(func.ifnull(gchk.se, 0)).label('se1'),gchk.gcmc, gchk.kpxh
                ).filter(gchk.wxfp == r, func.ifnull(gchk.wstt, '') != '', gchk.kpxh == l
                ).group_by(gchk.kpxh, gchk.gcmc).first()
                if c:
                    i2 = i2 + 1
                    i = i + 1
                    fpje = float(c.fpje1) if c.fpje1 else 0
                    se = float(c.se1) if c.se1 else 0
                    fb = generate_zb(year, month, j5, i)

                    ws['N' + str(i + 1)] = '3' + fb
                    ws['O' + str(i + 1)] = str(r) + '合同入账'
                    # ws['P' + str(i + 1)] = c.wxfp + '付货款'
                    ws['P' + str(i + 1)] = '2202.01'
                    ws['Q' + str(i + 1)] = '明细应付款'
                    if c.gcmc != '' and c.gcmc != None:
                        c1 = s.query(newcwcs.jdbm).filter(or_(newcwcs.company_name == c.gcmc, newcwcs.company_name == '个体户' + c.gcmc)).first()
                        if c1:
                            ws['AP' + str(i + 1)] = c1.jdbm
                        if len(c.gcmc.strip()) > 8:
                            ws['AQ' + str(i + 1)] = c.gcmc
                        else:
                            ws['AQ' + str(i + 1)] = '个体户' + c.gcmc
                    
                    ws['BB' + str(i)] = 'PRE001'
                    ws['BC' + str(i)] = '人民币'
                    ws['BD' + str(i)] = 'HLTX01_SYS'
                    ws['BE' + str(i)] = '固定汇率'
                    ws['BF' + str(i)] = '1'
                    ws['BK' + str(i)] = round(fpje * 100) / 100
                    ws['BM' + str(i)] = round(fpje * 100) / 100
                    fpje = fpje + round(fpje * 100) / 100
                    se = se + round(se * 100) / 100

            i = i + 1     
            fb = generate_zb(year, month, j5, i)
            ws['N' + str(i)] = '3' + fb
            ws['O' + str(i)] = str(r) + '合同入账'
            ws['P' + str(i)] = '2202.02'
            ws['Q' + str(i)] = '暂估应付款'

            ws['BB' + str(i)] = 'PRE001'
            ws['BC' + str(i)] = '人民币'
            ws['BD' + str(i)] = 'HLTX01_SYS'
            ws['BE' + str(i)] = '固定汇率'
            ws['BF' + str(i)] = '1'
            ws['BK' + str(iz + 2)] = se
            ws['BL' + str(iz + 2)] = se
            ws['BK' + str(i)].font = Font(color='FF0000')
            ws['BK' + str(i)] = -(fpje - se)
            ws['BM' + str(i)].font = Font(color='FF0000')
            ws['BM' + str(i)] = -(fpje - se)

            
        temp_file = f"金蝶合同入账{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        temp_path = os.path.join(config.tmp_path, temp_file)
        wb.save(temp_path)

        # s.commit()
        return json_result(0, "金蝶合同入账导出成功", temp_file)
    except Exception as e:
        # s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'导出失败: {str(e)}')
    finally:
        s.close()



"""
金蝶合同结转
对应原Pascal: 金蝶合同结转
"""
@any_route('/api/saier/payment/kingdee_close', methods=['POST'])
@require_token
async def view_payment_kingdee_close(request):
    # 1. 检查用户权限
    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        cxrq = j.get('cxrq', '')
        rids = j.get('rids', [])
        org = get_user_path(user.username)
        position = org.get('position', '')
        if '财务' not in position and '总经理' not in position:
            return json_result(-1, '没有权限导出')
        template_path = os.path.join(config.data_upload_path, 'template', '合同结转.xlsx')
        if not os.path.exists(template_path):
            return json_result(-1, '合同结转模板文件不存在')
        wb = load_workbook(template_path)
        ws = wb.active
        ws.title = "金蝶合同结转"
        j2 = 0
        j3 = 0
        j4 = 0
        j5 = 0
        zb = ''
        fb = ''
        wstt = ''
        bgpmstr1 = []
        fphm = []
        for rid in rids:
            d = s.query(gchk.wxfp,gchk.kpxh,gchk.wstt).filter(gchk.rid == rid).first()
            if not d:
                continue
            bgpmstr1.append(rid)
            wxfp = d.wxfp if d.wxfp else ''
            j3 = j3 + 1                
            if wxfp != '' and wxfp != None and wxfp not in fphm:
                fphm.append(wxfp)
                j2 += 1

        year = cxrq[:4]
        month = cxrq[5:7]
        # day = cxrq[-2:]
        weekday, days = calendar.monthrange(int(year), int(month))
        if j2 < 99:
            j4 = 2
        else:
            if j2 < 999:
                j4 = 3
            else:
                j4 = 4

        if j3 < 99:
            j5 = 3
        else:
            if j2 < 999:
                j5 = 4
            else:
                j5 = 5
        
        i = 0
        i1 = 0
        iz = 0
        for r in fphm:
            zpbm = ''
            zpbm1 = ''
            i1 = i1 + 1
            i = i + 1
            iz = i
            i2 = 0
            tse = 0
            se = 0
            c = s.query(gchk).filter(gchk.wxfp == r, func.ifnull(gchk.wstt, '') != '').first()
            if c:
                # fpje = float(c.fpje) if c.fpje else 0
                a = s.query(zx.wb1, zx.wb2).filter(zx.ly == "金蝶账簿编码", zx.mc == c.wstt).first()
                if a:
                    zpbm = a.wb1
                    zpbm1 = a.wb2
            zb = generate_zb(year[2:], month, j4, j2)
            ws['A' + str(iz + 2)] = '4' + str(zb)
            ws['B' + str(iz + 2)] = zpbm
            ws['C' + str(iz + 2)] = str(wstt) + '账簿'
            ws['D' + str(iz + 2)] = str(year) + '-' + str(month) + '-' + str(days)
            ws['E' + str(iz + 2)] = 'PRE001'
            ws['F' + str(iz + 2)] = '记'
            ws['H' + str(iz + 2)] = zpbm1
            ws['I' + str(iz + 2)] = wstt
            ws['J' + str(iz + 2)] = 'False'
            ws['K' + str(iz + 2)] = str(year)
            ws['L' + str(iz + 2)] = str(month)
            fb = generate_fb(year, month, j5, i, False)
            ws['N' + str(iz + 2)] = '4' + fb
            ws['O' + str(iz + 2)] = str(r) + '合同结转'
            ws['P' + str(iz + 2)] = '1812.01'
            ws['Q' + str(iz + 2)] = '应收出口退税'
            ws['BB' + str(iz + 2)] = 'PRE001'
            ws['BC' + str(iz + 2)] = '人民币'
            ws['BD' + str(iz + 2)] = 'HLTX01_SYS'
            ws['BE' + str(iz + 2)] = '固定汇率'
            ws['BF' + str(iz + 2)] = '1'

            for l in bgpmstr1:
                c = s.query(func.sum(func.ifnull(gchk.tse, 0)).label('tse1'),
                    func.sum(func.ifnull(gchk.se, 0)).label('se1'),gchk.gcmc, gchk.kpxh
                ).filter(gchk.wxfp == r, func.ifnull(gchk.wstt, '') != '', gchk.rid == l
                ).group_by(gchk.kpxh, gchk.gcmc).first()
                if c:
                    i2 = i2 + 1
                    i = i + 1
                    tse1 = float(c.tse1) if c.tse1 else 0
                    se1 = float(c.se1) if c.se1 else 0
                    tse = tse + round(tse1*100) / 100
                    se = se + round(se1*100) / 100
                    
            ws['BK' + str(iz + 2)] = tse
            ws['BL' + str(iz + 2)] = se
            if se - tse > 0:
                i = i + 1
                iz = i
                fb = generate_fb(year, month, j5, i, False)
                ws['N' + str(iz + 2)] = '4' + fb
                ws['O' + str(iz + 2)] = str(r) + '合同结转'
                ws['P' + str(iz + 2)] = '6401.01.01'
                ws['Q' + str(iz + 2)] = '外销'
                ws['BB' + str(iz + 2)] = 'PRE001'
                ws['BC' + str(iz + 2)] = '人民币'
                ws['BD' + str(iz + 2)] = 'HLTX01_SYS'
                ws['BE' + str(iz + 2)] = '固定汇率'
                ws['BF' + str(iz + 2)] = '1'
                ws['BK' + str(iz + 2)] = se - tse
                ws['BL' + str(iz + 2)] = se - tse
            
            i = i + 1
            iz = i
            fb = generate_fb(year, month, j5, i, False)
            ws['N' + str(iz + 2)] = '4' + fb
            ws['O' + str(iz + 2)] = str(r) + '合同结转'
            ws['P' + str(iz + 2)] = '2221.01.01.03'
            ws['Q' + str(iz + 2)] = '出口退税'
            ws['BB' + str(iz + 2)] = 'PRE001'
            ws['BC' + str(iz + 2)] = '人民币'
            ws['BD' + str(iz + 2)] = 'HLTX01_SYS'
            ws['BE' + str(iz + 2)] = '固定汇率'
            ws['BF' + str(iz + 2)] = '1'
            ws['BK' + str(iz + 2)] = tse
            ws['BM' + str(iz + 2)] = tse

            if se - tse > 0:
                i = i + 1
                iz = i
                fb = generate_fb(year, month, j5, i, False)
                ws['N' + str(iz + 2)] = '4' + fb
                ws['O' + str(iz + 2)] = str(r) + '合同结转'
                ws['P' + str(iz + 2)] = '2221.01.01.02'
                ws['Q' + str(iz + 2)] = '进项税额转出'
                ws['BB' + str(iz + 2)] = 'PRE001'
                ws['BC' + str(iz + 2)] = '人民币'
                ws['BD' + str(iz + 2)] = 'HLTX01_SYS'
                ws['BE' + str(iz + 2)] = '固定汇率'
                ws['BF' + str(iz + 2)] = '1'
                ws['BK' + str(iz + 2)] = se - tse
                ws['BM' + str(iz + 2)] = se - tse
            
        temp_file = f"金蝶合同结转{time.strftime('%Y%m%d%H%M%S')}.xlsx"
        temp_path = os.path.join(config.tmp_path, temp_file)
        wb.save(temp_path)

        # s.commit()
        return json_result(0, "金蝶合同结转导出成功", temp_file)
    except Exception as e:
        # s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'导出失败: {str(e)}')
    finally:
        s.close()



"""
宁波批量付款回导
对应原Pascal: 宁波批量付款回导
"""
@any_route('/api/saier/payment/update_paid', methods=['POST'])
@require_token
async def view_payment_update_paid(request):
    # 1. 检查用户是否是财务岗位 (对应 tmpcom2 查询)
    s = Session()
    try:
        j = await request.form()
        module = form_value(j, 'module', '采购付款')
        roles = form_value(j, 'roles', '')
        role_list = roles.split(',') if roles else []
        fkdq = '宁波'
        for r in role_list:
            if '义乌' in r:
                fkdq = '义乌'
                break
        user = request.current_user
        org = get_user_path(user.username)
        path = org.get('path')

        # 2. 获取上传的Excel文件
        file = form_value(j,'file',None)
        if is_none(file):
            return json_result(ERR_PARAM_NOT_ENOUGH)
        # 保存临时文件
        temp_dir = config.tmp_path
        temp_file = os.path.join(temp_dir, f"{uuid.uuid4()}_{file.filename}")
        # logger.error(f"Saving uploaded file to: {temp_file}")
        data = await file.read()
        write_file(temp_file, data, 'wb', encoding=None)
        
        # 3. 读取Excel文件
        wb = load_workbook(temp_file)
        ws = wb.worksheets[0]  # 读取第一个工作表
        # 4. 处理每一行数据
        # duplicates = []  # 存储重复记录信息
        i = 2  # 从第2行开始（第1行是表头）
        sb = ws.cell(row=i, column=2)  # B列 - 识别码
        if sb != '识别码':
            return json_result(-1, '格式错误,请用宁波批量付款回导')
        xgqd = []
        i = 3
        tmpstr = []
        while True:
            # 读取Excel单元格数据
            gcfp = ws['A' + str(i)]
            sb = ws['B' + str(i)]
            gcfp1 = ws['C' + str(i)]
            bank = ws['D' + str(i)]
            yfhj = ws['F' + str(i)]
            hkje = ws['G' + str(i)]
            fkhb = ws['I' + str(i)]
            sfjq = ws['J' + str(i)]
            hkrq = ws['K' + str(i)]
            dkje1 = ws['AN' + str(i)]
            dkje = ws['AO' + str(i)]
            frq = ws['AP' + str(i)]
            qrbz1 = ws['AQ' + str(i)]
            ywqr = ws['BA' + str(i)]
            ywqrrq = ws['BB' + str(i)]
            ywbqryy = ws['BC' + str(i)]
            sfzk = ws['BD' + str(i)]
            zkje = ws['BE' + str(i)]
            wfyhzh = ws['BF' + str(i)]
            wfyhmc = ws['BG' + str(i)]
            yyzf = ws['BI' + str(i)]
            thzz = ws['BJ' + str(i)]
            tcyd = ws['BK' + str(i)]
            gsbs = ws['BL' + str(i)]
            custId = ws['BM' + str(i)]
            corpCode = ws['BN' + str(i)]
            subaAccountSerial = ws['BO' + str(i)]
            if fkhb == '' or fkhb == None:
                fkhb = '否'
            if sfjq == '' or sfjq == None:
                sfjq = '否'
            # 如果rid为空，结束循环
            if not sb:
                break

            sb1 = ''
            bzsmy = ''
            fpxh = ''
            fkhj = hkje if hkje!="" and hkje!=None else 0
            d = s.query(gchk).filter(gchk.sb == sb).first()
            if not d:
                tmpstr.Add('识别码: ' + sb + ',该记录不存在请核对!')
                tzsb = '1'
                i = i + 1
                continue
            bzsmy = d.bzsm
            c = s.query(func.sum(func.ifnull(gchksheet1.hkje,0)).label('hkjez')).filter(gchksheet1.pid == d.rid).first()
            if c and c.hkjez:
                fkhj = float(fkhj) + float(c.hkjez)
            else:
                fkhj = float(fkhj)
            if sfzk == '是':
                d.bzsm = bzsmy + '/暂扣'
                d.sfzk = '是'
                d.zkje = float(zkje) if zkje!="" and zkje!=None else 0
                d.mtime = time.strftime('%Y-%m-%d %H:%M:%S')
                d.modi_uid = user.rid

            if (hkje == '') or hkje == None:
                tmpstr.append('发票: ' + str(bank) + '付款金额未填!')
                tzsb = '1'
                i = i + 1
                continue
            if bank == '' or bank == None:
                c = s.query(gchk).filter(gchk.sb == sb, gchk.sfjq != '是', gchk.fkdq == fkdq).first()
            else:
                c = s.query(gchk).filter(gchk.sb == sb, gchk.sfjq != '是', gchk.fkdq == fkdq, gchk.fpbh == gcfp1, gchk.gcfp == bank).first()
            if not c:
                tmpstr.append('发票: ' + str(bank) + '末修改,请检查是否有权限!')
                tzsb = '1'
                i = i + 1
                continue
            if c.fpwk == '否':
                yfhj = float(c.fpje) if c.fpje else 0
            fpxh = c.kpxh if c.kpxh else ''
            sb1 = c.rid if c.rid else ''
            dlrq = c.hkrq if c.hkrq else None
            yfgc1 = c.gcmc1 if c.gcmc1 else ''
            yfgc = c.gcmc if c.gcmc else ''
            cgry = c.jsrm if c.jsrm else ''
            cght1 = c.cght if c.cght else ''
            pk = c.pk if c.pk else 0
            if pk != 0:
                tmpstr.append('发票: ' + str(bank) + ',该生产或出货厂家' + yfgc + '/' + yfgc1 + '有工厂赔款请核对,付款信息已更改!')
                tzsb = '1'
            xgqd.append(c.xgqd if c.xgqd else '')
            if dlrq == hkrq :
                tmpstr.append('发票: ' + str(bank) + '末修改,请检查是这张发票同一天是否已有付款')
                tzsb = '1'
                i = i + 1
                continue
            if yfgc1 == '':
                yfgc1 = '没有123'
            a = s.query(func.sum(func.ifnull(yfhk.fkje,0)).label('hkjez')).filter(or_(yfhk.sccj.like('%' + yfgc + '%'), yfhk.gcmc1.like('%' + yfgc1 + '%'))
                ).filter(sfjq != '是').first()
            if  a and a.hkjez and float(a.hkjez) > 0:
                tmpstr.append('发票: ' + str(bank) + '末修改,该生产或出货厂家' + yfgc + '/' + yfgc1 + '有预付货款' + str(a.hkjez))
                tzsb = '1'
                i = i + 1
                continue

            if hkje != '0' and hkje != '' and hkje != None:
                result = chinese_amount(float(hkje))

            fkn = ''
            fky = ''
            fkr = ''
            year = ''
            month = ''
            day = ''
            if hkrq != '' and hkrq != None:
                if isinstance(hkrq, datetime):
                    hkrq = hkrq.strftime('%Y-%m-%d')
                year = hkrq[:4]
                month = hkrq[5:7]
                day = hkrq[8:10]
                fky = ''.join(ArabiaToChinese(ch) for ch in year)
                fkn = format_month_to_chinese(month)
                fkr = format_month_to_chinese(day)
            yhje = ''
            yhje1 = ''
            if hkje == '' or hkje == None:
                hkje = 0
            yhje = str(float(hkje) * 100)
            i1 = len(yhje)
            for i1 in range(1, i1 + 1):
                yhje1 = yhje1 + yhje[i1-1] + ' '
            xgqd.append(user.username + time.strftime('%Y-%m-%d %H:%M:%S') + str(hkje) + str(hkrq) + '是否结清' + str(sfjq))

            if (float(yfhj) - float(fkhj) - float(pk)) < 0 :
                tmpstr.append('发票: ' + str(bank) + ',该生产或出货厂家已有付款' + str(fkhj) + '大于应付' + str(yfhj) + '付款信息末更改!')
                tzsb = '1'
            else:
                d.hkje = float(hkje)
                logger.error('-----hkrq aa---')
                logger.error(hkrq)
                d.hkrq = hkrq
                d.fkhj = fkhj
                d.sfjq = sfjq
                d.fkdx = result
                d.fkn = fkn
                d.fky = fky
                d.fkr = fkr
                d.fkhb = fkhb
                if year != '' and year != None:
                    d.nsz = year
                if month != '' and month != None:
                    d.ysz = month
                if day != '' and day != None:
                    d.rsz = day
                d.wfje = float(yfhj) - float(fkhj) - float(pk)
                d.yhje = yhje
                d.yhje1 = yhje1
                d.xgqd = "\n".join(xgqd)
                d.qrbz1 = qrbz1
                d.ywqr = ywqr
                d.ywqrrq = ywqrrq
                d.ywbqryy = ywbqryy
                d.sfzk = sfzk
                d.zkje = float(zkje) if zkje!="" and zkje!=None else 0
                d.wfyhzh = wfyhzh
                d.wfyhmc = wfyhmc
                d.yyzf = yyzf
                d.thzz = thzz
                d.tcyd = tcyd
                d.gsbs = gsbs
                d.custId = custId
                d.corpCode = corpCode
                d.subaAccountSerial = subaAccountSerial
                d.mtime = time.strftime('%Y-%m-%d %H:%M:%S')
                d.modi_uid = user.rid   


                if sb1 != '' and sb1 != None and sfjq == '是':
                    s.query(gchksheet).filter(gchksheet.pid == sb1).update({gchksheet.hkrq: hkrq}, synchronize_session=False)
                s.query(cgfkhdsheet).filter(cgfkhdsheet.sb == sb).update({cgfkhdsheet.hkrq: hkrq}, synchronize_session=False)

            if float(hkje) > 0:
                a = gchksheet1()
                a.rid = get_uuid()
                a.uid = user.rid
                a.ctime = time.strftime('%Y-%m-%d %H:%M:%S')
                a.pid = d.rid
                a.fkrq = hkrq
                a.hkje = float(hkje)
                a.czry = user.username
                a.fkxh = ''
                a.hdh = c.hdh if c.hdh else ''
                a.qsny = ''
                a.jsrybm = c.jsrybm if c.jsrybm else ''
                a.jsrm = c.jsrm if c.jsrm else ''
                s.add(a)
            b = s.query(gchksheet).filter(gchksheet.pid == d.rid).first()
            if b:
                cywyzd = b.cywyzd if b.cywyzd else ''
                if cywyzd != '':
                    s.query(fkspsheet).filter(fkspsheet.cywyzd == cywyzd, func.ifnull(fkspsheet.fkrq,"") == "").update({fkspsheet.fkrq: hkrq}, synchronize_session=False)
                    s.query(fkspsheet).filter(fkspsheet.cywyzd == cywyzd, func.ifnull(fkspsheet.fkrq1,"") == "").update({fkspsheet.fkrq1: hkrq}, synchronize_session=False)

            if cgry != '' and cgry != None and d.wxfp != '' and d.wxfp != None:
                s.query(fpgl).filter(fpgl.wxfp == d.wxfp, fpgl.tssd=='否', fpgl.sfjd=='否').update({fpgl.webpdfk: '是'}, synchronize_session=False)

            s.add(d)
            i = i + 1

        out_name = ''
        msg = '发票回导成功'
        if len(tmpstr) > 0:
            msg = '发票回导失败记录,正在生成txt文件...'
            out_name = f"发票回导失败记录_{time.strftime('%Y-%m-%d_%H-%M-%S')}.txt"
            out_path = os.path.join(config.tmp_path, out_name)
            val = '\n'.join(tmpstr)
            write_file(out_path, val, 'w')

        s.commit()
        return json_result(0, msg, out_name)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()


"""
宁波批量付款回导
对应原Pascal: 宁波批量付款回导
"""
@any_route('/api/saier/payment/batch_audit', methods=['POST'])
@require_token
async def view_payment_batch_audit(request):
    # 1. 检查用户是否是财务岗位 (对应 tmpcom2 查询)
    s = Session()
    user = request.current_user
    try:
        j = await request.json()
        module = j.get('module', '采购付款')
        rids = j.get('rids', []) 
        kind = j.get('kind', '开票')
        jzrq = j.get('jzrq', '')
        if jzrq == '' or jzrq == None:
            jzrq = time.strftime('%Y-%m-%d')
        fpwk = '是' if kind == '现金' else '否'

        org = get_user_path(user.username)
        position = org.get('position', '')
        cwpath = org.get('path', '')
        uid = user.rid

        fkdq = '宁波'
        fkdh = '义乌'
        jzrq1 = None
        if '义乌' in position:
            fkdq = '义乌'
            fkdh = '宁波'
        # 获取当前日期（去掉时间部分）
        today = datetime.now().date()
        # 查找未来7天内的第一个周三
        for i in range(1, 8):  # 对应 Delphi 的 1 to 7
            d1 = today + timedelta(days=i)
            i3 = d1.weekday()  # Python: 周一=0, 周二=1, 周三=2
            if i3 == 3:  # 周四对应的数字是 3
                jzrq1 = d1.strftime('%Y-%m-%d %H:%M:%S')  # 转为字符串，对应 Delphi 的 DateTimeToStr
                break  # 找到第一个周三后可以退出循环
        
        if jzrq == '' or jzrq == None:
            jzrq = jzrq1
        jzsj = ' 13:00:00'
        c = s.query(zx.mc).filter(zx.ly == '采购付款截止日期时间').first()
        if c and c.mc:
            jzsj = str(c.mc)
        date = time.strftime('%Y-%m-%d')
        year = date[:4]
        month = date[5:7]
        day = date[8:10]
        kpxh1 = str(year) + '-' + str(month) + '-'
        bgpmstr2 = []
        bgpmstr3 = []
        chcpzl = []
        for rid in rids:
            d = s.query(gchk).filter(gchk.rid == rid, or_(gchk.sfzk == '是',func.ifnull(gchk.sfzk,'') == ''), gchk.sfjq == '否', gchk.fkdq == fkdq).first()
            if not d:
                continue
            if d.fpwk == '否' and d.fkhj >= d.fpje:
                continue
            bzsm = d.bzsm if d.bzsm else ''
            if ('不付' in bzsm) or ('通知' in bzsm) or ('承兑' in bzsm) or ('暂扣' in bzsm) or ('合同未到' in bzsm) or ('合同模糊' in bzsm) or ('合同未盖章' in bzsm) or ('合同背面不干净' in bzsm) or ('预填' in bzsm) or ('预报' in bzsm):
                bgpmstr2.append(d.rid)
                if '合同未到' in bzsm or '合同模糊' in bzsm or '合同未盖章' in bzsm or '合同背面不干净' in bzsm:
                    bgpmstr3.append(d.rid)
                continue

            kpxh = ''
            path = ''
            pid = ''
            fkbh = ''
            kpxhz = 1
            flag = 0
            jsrm = d.jsrm if d.jsrm else ''
            if d.fpwk == fpwk:
                c = s.query(cgfkhd).filter(cgfkhd.jsrm == jsrm, cgfkhd.jzrq == jzrq, cgfkhd.gcmc == d.gcmc, cgfkhd.fpwk == fpwk).first()
                if c:
                    flag = 1
                    pid = c.rid
                    if c.ywhdzt=='完成':
                        c.ywhdzt = '未完成'
                        c.cwhdzt = '未完成'
                        c.mtime = time.strftime('%Y-%m-%d %H:%M:%S')
                        c.modi_uid = uid
                        s.add(c)

            if flag == 0:
                c = s.query(cgfkhd.fkdh).filter(cgfkhd.fkdh.like('%' + str(kpxh) + '%')).order_by(cgfkhd.fkdh.desc()).first()
                if c:
                    last_fkdh = str(c.fkdh)
                    kpxhz = int(last_fkdh[7:12]) + 1
                fkbh = kpxh1 + str(kpxhz).zfill(5)  # 将数字部分填充为5位，不足的部分用0补齐
                org = get_user_path(jsrm)
                uid = org.get('rid', '')
                path = org.get('path', '')
                m = cgfkhd()
                pid = get_uuid()
                m.rid = pid
                m.uid = uid
                m.ctime = time.strftime('%Y-%m-%d %H:%M:%S')
                m.fkdh = fkbh
                m.cwry = user.username
                m.jsrm = jsrm
                m.plhd = ''
                m.hdrq = None
                m.cwhd = ''
                m.cwhdrq = None
                m.bz = ''
                m.jzrq = jzrq
                m.jzsj = jzsj
                m.ywhdzt = '未完成'
                m.cwhdzt = '未完成'
                m.cwdq = fkdq
                m.fksb = ''
                m.plfkrq = None
                m.gcmc = d.gcmc if d.gcmc else ''
                m.plzk = ''
                m.zkyy = ''
                m.gcmyd = '一般'
                m.bmyyy = ''            
                m.fpwk = fpwk
                s.add(m)
                user_list = [jsrm]
                if user.username not in user_list:
                    user_list.append(user.username)
                
                res = module_share_new('采购付款核对',[jsrm],pid,user,s)
                if res['code'] != 1:
                    s.rollback()
                    return json_result(-1, res['msg'])
                
                for u in user_list:
                    c = s.query(khsheet3).filter(khsheet3.father1 == pid, khsheet3.module1 == '采购付款核对', khsheet3.objectnumber1 == uid, khsheet3.ywry == u).first()
                    if not c:
                        k = khsheet3()
                        k.rid = get_uuid()
                        k.father1 = pid
                        k.module1 = '采购付款核对'
                        k.objectnumber1 = uid
                        k.ywry = u
                        k.objectkind1 = 1
                        k.permission1 = '读取,更改,删除'
                        k.sys_path = path
                        k.company_name = fkbh
                        s.add(k)

                if user.username != jsrm and jsrm != '' and jsrm != None:
                    data = {"type":"success","title":"采购付款核对通知",
                        "msg":f"{d.gcmc}现金:{fpwk}的采购付款核对编号:{fkbh}需核对,日期:{time.strftime('%Y-%m-%d')}",
                        "module":"采购付款核对","rid":pid,"rids":[],
                        "btns":[{"title":"打开","icon":"any-rmb-full","name":"open","color":"red"}]}
                    await messages.message_to_user(jsrm,data,MSG_KIND_NOTICE_RECORD,request)
                    xx = f"{user.username} {d.gcmc}现金:{fpwk}的采购付款核对编号:{fkbh}需核对,日期:{time.strftime('%Y-%m-%d')}"
                    row = {
                        "xxly": '采购付款核对',
                        "bjdh": '',
                        "wxht": '',
                        "cght": '',
                        "gdht": fkbh,
                        "xxnr": xx,
                        "jsr": jsrm,
                        "sys_path": "",
                        "spsq": user.username
                    }
                    res = module_xxck_new([row],user,s)
                    if res.get('code')!=1:
                        return json_result(res.get('code'), res.get('code'))

                    res = user_task_new('采购付款核对', pid, '付款单号', '采购付款核对通知', xx, user, s, [jsrm])
                    if res.get('code') != 1:
                        s.rollback()
                        return json_result(res.get('code'), res.get('msg'))

            if pid != '' and pid != None:
                customer = ''
                wstt1 = d.wstt
                a = s.query(cgfkhdsheet).filter(cgfkhdsheet.sb == d.sb, cgfkhdsheet.jzrq == jzrq).first()
                if not a:
                    c = s.query(gchksheet).filter(gchksheet.pid == rid).all()
                    for r in c:
                        if r.cght != '' and r.cght is not None:
                            b = s.query(cght.jsfs,cght.sccj1,cght.khmc).filter(cght.hthm == r.cght).first()
                            if b and b.sccj1:
                                wstt1 = b.sccj1
                                customer = b.khmc if b.khmc else ''
                                chcpzl.append('发票号码/原发票号:' + str(r.fphm) + '/' + str(r.yfph) + '采购合同:' + str(r.cght) + '中文品名:' + str(r.zwpm) + '箱数/数量:' + str(r.xs) + '/' + str(r.zsl) + '生产厂家:' + str(b.sccj1) + '结算方式:' + str(b.jsfs) + '提前审请:' + str(r.ysqje) + '预付金额:' + str(r.yfje1))
                
                p = cgfkhdsheet()
                p.rid = get_uuid()
                p.uid = uid
                p.ctime = time.strftime('%Y-%m-%d %H:%M:%S')
                p.pid = pid
                p.fpyz = d.fpyz if d.fpyz else ''
                p.ysfp = d.ysfp if d.ysfp else ''
                p.fpbh = d.fpbh if d.fpbh else ''
                p.gcfp = d.gcfp if d.gcfp else ''
                p.wxfp = d.wxfp if d.wxfp else ''
                p.zzsl = d.zzsl if d.zzsl else 0
                p.tsl = d.tsl if d.tsl else 0
                p.hkrq = None
                p.gcmc = d.gcmc if d.gcmc else ''
                p.fkbz = d.fkbz if d.fkbz else ''
                p.hkje = 0
                p.htje = d.htje if d.htje else 0
                p.bhsj = d.bhsj if d.bhsj else 0
                p.fpje = d.fpje if d.fpje else 0
                p.bkpje = d.bkpje if d.bkpje else 0
                p.se = d.se if d.se else 0
                p.tse = d.tse if d.tse else 0
                p.jsrm = d.jsrm if d.jsrm else ''
                p.sh = d.sh if d.sh else ''
                p.bank = d.bank if d.bank else ''
                p.zh = d.zh if d.zh else ''
                p.province = d.province if d.province else ''
                p.city = d.city if d.city else ''
                p.yt = d.yt if d.yt else ''
                p.yfhj = d.yfhj if d.yfhj else 0
                p.gcmc1 = d.gcmc1 if d.gcmc1 else ''
                p.gcdh = d.gcdh if d.gcdh else ''
                p.fkdq = d.fkdq if d.fkdq else ''
                p.fkxh = d.fkxh if d.fkxh else ''
                p.sb = d.sb if d.sb else ''
                p.dlrq = d.dlrq if d.dlrq else None
                if customer == '' or customer == None:
                    p.customer = d.customer
                else:
                    p.customer = customer 
                p.wstt = wstt1
                p.cxbg = d.cxbg if d.cxbg else ''
                p.htsh = d.htsh if d.htsh else ''
                p.zkfy1 = d.zkfy1 if d.zkfy1 else 0
                p.htrq = d.htrq if d.htrq else None
                p.ywhd = ''                
                p.ywbtgyy = ''
                p.bzsm = d.bzsm if d.bzsm else ''
                p.chcpzl = '\n'.join(chcpzl)
                p.sfzk = '否'
                p.yzh = d.zh if d.zh else ''
                p.xshj = d.xshj if d.xshj else 0
                p.kkje = d.kkje if d.kkje else 0
                p.yfje1 = d.yfje1 if d.yfje1 else 0
                p.zsl = d.zsl if d.zsl else 0
                p.zkje = 0
                p.fkn = ''
                p.fky = ''
                p.fkr = ''
                p.fkdx = ''
                p.nsz = ''
                p.ysz = ''
                p.rsz = ''
                p.yhje = 0
                p.yhje1 = 0
                p.fksb = '否'
                p.fkhj = d.fkhj if d.fkhj else 0
                p.sfjq = d.sfjq if d.sfjq else ''
                p.pk = d.pk if d.pk else 0
                p.xgqd = d.xgqd if d.xgqd else ''
                p.fkdh = fkdh
                p.number1 = rid
                p.jzrq = jzrq
                p.cwpath = cwpath
                p.ly = '本地'
                s.add(p)
      
        for rid in bgpmstr3:
            jsrmz = ''
            gcmcz = ''
            c = s.query(gchk).filter(gchk.rid == rid, gchk.fpwk == fpwk).first()
            if c:
                jsrmz = c.jsrm if c.jsrm else ''
                gcmcz = c.gcmc if c.gcmc else ''
                res = user_task_new('采购付款', pid, '付款单号', f"{gcmcz}采购付款异常", f"有合同未到 合同未盖章 合同背面不干净 合同模糊问题请点击查看,日期:{time.strftime('%Y-%m-%d')}", user, s, [jsrmz])
                if res.get('code') != 1:
                    s.rollback()
                    return json_result(res.get('code'), res.get('msg'))

        for rid in bgpmstr2:
            s.query(gchk).filter(gchk.rid == rid).update({gchk.sfzk: '是'}, synchronize_session=False)


        s.commit()
        return json_result(1, '批量审核成功')
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'批量审核失败: {str(e)}')
    finally:
        s.close()



def extract_code_optimized(hthm, an2_value=''):
    """使用正则表达式的优化版本"""
    
    # 优先使用AN2的值
    if an2_value:
        return an2_value, '否'
    
    hthm_lower = hthm.lower()
    
    # 检查是否为代理业务
    if hthm_lower[:2] in ['dl', 'fw']:
        sfdl = '是'
        # 从第4位开始提取连续数字（最多3位）
        match = re.search(r'^\w{3}(\d{1,3})', hthm)
        ywbm = match.group(1) if match else ''
    else:
        sfdl = '否'
        # 从第2位开始提取连续数字（最多3位）
        match = re.search(r'^.\d+', hthm)
        if match:
            ywbm = match.group(0)[1:4]  # 去掉首位，最多取3个数字
        else:
            ywbm = ''
    
    return f"{ywbm}组" if ywbm else '', sfdl

def format_date_string(date_str: str) -> str:
    """智能解析各种日期格式"""
    if not date_str:
        return None
    
    date_str = str(date_str).strip()
    
    # 已经是标准格式且长度足够，直接返回
    if len(date_str) >= 10 and date_str[4] == '-' and date_str[7] == '-':
        return date_str
    
    try:
        # 自动解析多种格式
        dt = parser.parse(date_str, fuzzy=True)
        return dt.strftime('%Y-%m-%d')
    except:
        return date_str
    
"""
现金导入(详细)
对应原Pascal: 现金导入(详细)
"""
@any_route('/api/saier/payment/update_cash', methods=['POST'])
@require_token
async def view_payment_update_cash(request):
    # 1. 检查用户是否是财务岗位 (对应 tmpcom2 查询)
    s = Session()
    try:
        j = await request.form()
        module = form_value(j, 'module', '采购付款')
        roles = form_value(j, 'roles', '')
        role_list = roles.split(',') if roles else []
        fkdq = '宁波'
        for r in role_list:
            if '义乌' in r:
                fkdq = '义乌'
                break
        user = request.current_user
        org = get_user_path(user.username)
        path = org.get('path')
        t = time.strftime('%Y-%m-%d %H:%M:%S')
        t1 = t[2:4] + t[5:7] + t[8:10] + t[11:13] + t[14:16] + t[17:19]
        jsts = 0
        sz1 = 0
        cxje = 0
        cxje1 = 0
        gsmc = ''
        c = s.query(zx).filter(zx.ly == '财务付款结算期').first()
        if c and c.mc:
            jzsj = c.cs if c.cs else 0
            sz1 = c.sz1 if c.sz1 else 0
        org = get_user_path('zjnblh')
        position = org.get('position', '')
        bm1 = position[:1] if position else ''
        c = s.query(zx).filter(zx.ly == '诚信金额').first()
        if c and c.mc:
            cxje = c.cs if c.cs else 0
        c = s.query(wfgs).filter(ywrybiao.yhm == 'zjnblh').first()
        if c and c.wfgs:
            gsmc = c.wfgs
        i = 0
        kpjxts = 0
        xjjxts = 0
        kpxjts = 0
        xjxh = 0
        scqy = ''
        if '优胜' in path:
            c = s.query(cwjxrq).filter(cwjxrq.szgs.like('%优胜%')).first()
            if c:
                kpjxts = c.kpjxts if c.kpjxts else 0
                xjjxts = c.xjjxts if c.xjjxts else 0
        # 2. 获取上传的Excel文件
        file = form_value(j,'file',None)
        if is_none(file):
            return json_result(ERR_PARAM_NOT_ENOUGH)
        # 保存临时文件
        temp_dir = config.tmp_path
        temp_file = os.path.join(temp_dir, f"{uuid.uuid4()}_{file.filename}")
        # logger.error(f"Saving uploaded file to: {temp_file}")
        data = await file.read()
        write_file(temp_file, data, 'wb', encoding=None)
        # 3. 读取Excel文件
        wb = load_workbook(temp_file)
        ws = wb.worksheets[0]  # 读取第一个工作表
        # 4. 处理每一行数据
        # duplicates = []  # 存储重复记录信息
        i = 2  # 从第2行开始（第1行是表头）
        xgqd = []
        tmpstr = []
        while True:
            # 读取Excel单元格数据
            gcfp = ws['A' + str(i)]
            fkbz = ws['B' + str(i)]
            if fkbz == '' or fkbz == None:
                fkbz = 'RMB'
            customer = ''
            ywry = ''
            fpje = ws['C' + str(i)]
            bhsj = ws['D' + str(i)]
            zzsl = ws['E' + str(i)]
            se = ws['F' + str(i)]
            tsl = ws['G' + str(i)]
            tse = ws['H' + str(i)]
            sh = ws['I' + str(i)]
            if sh == '' or sh == None:
                sh = ws['AE' + str(i)]
            yt = ws['J' + str(i)]
            if yt == '' or yt == None:
                yt = '货款'
            pk = 0
            fpbh = ws['K' + str(i)]
            bzsm = ws['L' + str(i)]
            if bzsm == '' or bzsm == None:
                bzsm = ws['AH' + str(i)]
            kprq = ws['M' + str(i)]
            fpwk = ws['N' + str(i)]
            if fpwk == '' or fpwk == None:
                if gcfp == '' or gcfp == None:
                    fpwk = '是'
                else:
                    fpwk = '否'
            if fpwk == '否':
                kpxjts = kpjxts
            if fpwk == '是':
                kpxjts = xjjxts
            fkdq = ws['O' + str(i)]
            if fkdq == '' or fkdq == None:
                i = i + 1
                continue
            hkje = ws['P' + str(i)]
            hkrq = ws['Q' + str(i)]
            if hkrq == '' or hkrq == None:
                fksb = '否'
            if hkrq != '' and hkrq != None:
                fksb = '是'
            wxfp = ws['R' + str(i)]
            hthm = ws['S' + str(i)]
            htje = ws['T' + str(i)]
            chrq = ws['U' + str(i)]
            chrq1 = ws['U' + str(i)]
            jsrm = ws['V' + str(i)]
            xx = ws['W' + str(i)]
            cphh = ws['X' + str(i)]
            bz = ws['AH' + str(i)]
            zwpm = ws['Y' + str(i)]
            dj = ws['Z' + str(i)]
            dw = ws['AA' + str(i)]
            bz1 = ws['AB' + str(i)]
            xs = ws['AC' + str(i)]
            bz1 = bz1 if bz1 != '' and bz1 != None else 0
            xs = xs if xs != '' and xs != None else 0
            zsl = float(bz1) * float(xs)
            gcmc1 = ws['AE' + str(i)]
            rxfs = ws['AF' + str(i)]
            kh = ws['AG' + str(i)]
            gcdh = ws['AF' + str(i)]
            kkhj1 = ws['AI' + str(i)]
            name = ws['AJ' + str(i)]
            if name == '' or name == None:
                name = user.username
            dlrq = ws['AK' + str(i)]
            dlrq1 = ws['AK' + str(i)]
            if dlrq == '' or dlrq == None:
                dlrq = time.strftime('%Y-%m-%d')
            dj = dj if dj != '' and dj != None else 0
            htzje = zsl * float(dj)
            fpyz = wxfp + '\\' + hthm + '\\' + gcmc1
            yhfy = ws['AL' + str(i)]
            aa = gcmc1
            tqfk = ws['AM' + str(i)]
            if tqfk == '' or tqfk == None:
                tqfk = '否'
            hthm = hthm if hthm != '' and hthm != None else ''
            ywbm = ws['AN' + str(i)]
            sfdl = '否'
            if ywbm == '' or ywbm == None:
                wxfp2 = hthm[:3]
                c = s.query(ywbdzb).all()
                for r in c:
                    if r.dyywzm and r.dyywzm in wxfp2:
                        ywbm = r.ywb
                        break
            if ywbm == '' or ywbm == None:
                ywbm, sfdl = extract_code_optimized(hthm, ywbm)
            sqrq = ws['AO' + str(i)]
            sqje = ws['AP' + str(i)]
            bank1 = ws['AQ' + str(i)]
            yhzh1 = ws['AR' + str(i)]
            # 如果rid为空，结束循环
            if not wxfp:
                break
            xjxh = xjxh + 1
            chrq = format_date_string(chrq)
            dlrq = format_date_string(dlrq)
            sb = str(i-1) + str(name) + str(wxfp) + str(t1)
            org = get_user_path(name)
            path1 = org.get('path', '')
            if path1 != '' and path1 != None:
                path = path1 

            if sh != '' and sh != None:
                c = s.query(newcwcs).filter(newcwcs.company_name == sh).first()
                if c :
                    gcmc =c.shui
                    bank =c.bank if c.bank else ''
                    zh =c.zh
                    province =c.province
                    city =c.city
                    if ('黑名单' in c.hzdj):
                        hmd = hmd + 1
                        tmpstr.append('请注意工厂发票: ' + str(gcfp) + '工厂名称:' + str(gcmc) + str(c.hzdj) + '！！')
                        scqy =c.hzdj
                    if bank != '':
                        c = s.query(yhbm).filter(yhbm.yhdm == bank).first()
                        if c:
                            yhdm1 = c.yhdm
                            yhdz1 = c.yhdz if c.yhdz else ''
                            wyjgh1 = c.wyjgh
                            lhh1 = c.lhh
                            jgh1 = c.jgh
                            sjyh1 = c.sjyh
                            sjlhh1 = c.sjlhh
                            sjhjgh1 = c.sjhjgh

            if (bank1 == '') or (yhzh1 == ''):
                bank = bank1 if bank1 != '' and bank1 != None else ''
                zh = yhzh1 if yhzh1 != '' and yhzh1 != None else ''
                if bank != '':
                    c = s.query(yhbm).filter(yhbm.yhmc == bank).first()
                    if c:
                        yhdm1 = c.yhdm
                        yhdz1 = c.yhdz
                        wyjgh1 = c.wyjgh
                        lhh1 = c.lhh
                        jgh1 = c.jgh
                        sjyh1 = c.sjyh
                        sjlhh1 = c.sjlhh
                        sjhjgh1 = c.sjhjgh
            x = s.query(gchk).filter(gchk.fpyz == fpyz).first()
            if not x or fpyz == '' or fpyz == None:
                a = s.query(gchk.sid).order_by(gchk.sid.desc()).first()
                if a:
                    sb = str(name) + str(t1) + str(a.sid)
                pk = 0
                if hthm != '' and hthm != None:
                    c = s.query(cght).filter(cght.hthm == hthm).first()
                    if c:
                        if jsrm == '' or jsrm == None:
                            if (c.jsfs == '1') and (c.gdry != '' and c.gdry != None):
                                jsrm = c.gdry
                            else:
                                jsrm = c.cgry if c.cgry else ''
                        htje = float(c.htje) if c.htje else 0
                        if gcmc1 == '' or gcmc1 == None:
                            if c.sccj and c.sccj != '':
                                gcmc1 = c.sccj
                            else:
                                gcmc1 = c.sccj1 if c.sccj1 else ''
                        if wxfp == '' or wxfp == None:
                            wxfp = c.wxht if c.wxht else ''
                        if fkdq == '' or fkdq == None:
                            fkdq = c.szdq if c.szdq else ''
                        if province == '' or province == None:
                            province = c.province1 if c.province1 else ''
                        if city == '' or city == None:
                            city = c.city1 if c.city1 else ''
                        if c.wxbm and c.wxbm != '':
                            ywbm = c.wxbm
                        customer = c.khmc if c.khmc else ''
                        ywry = c.ywry if c.ywry else ''

                gcmc2 = gcmc1
                c = s.query(newcwcs).filter(newcwcs.company_name == gcmc2).first()
                if c:
                    if '黑名单' in c.hzdj:
                        hmd = hmd + 1
                        tmpstr.append('请注意工厂发票: ' + str(gcfp) + '工厂名称:' + str(gcmc2) + str(c.hzdj) + '！！')
                        scqy = c.hzdj 
                zsrq = None
                sz = 0
                if chrq != '' and chrq != None:
                    rq_date = datetime.strptime(str(chrq), '%Y-%m-%d')
                    # 计算起始日期
                    current_date = rq_date + timedelta(days=jsts)
                    xjxh = get_date_weekday(current_date)
                    i1 = 0
                    for i2 in range(1, 8):
                        if sz != sz1 :
                            # 计算起始日期
                            current_date = rq_date + timedelta(days=(jsts+i2))
                            xjxh = get_date_weekday(current_date)
                            i1 = i1 + 1
                    zsrq = rq_date + timedelta(days=(jsts+i1))
                    zsrq = zsrq.strftime('%Y-%m-%d')

                cxbg = '不需要'
                org = get_user_path(jsrm)
                uid = org.get('rid', '')
                m = gchk()
                pid = get_uuid()
                m.rid = pid
                m.uid = uid
                m.ctime = time.strftime('%Y-%m-%d %H:%M:%S')
                m.fpbh = gcfp
                m.gcfp = gcfp
                m.yjfk = zsrq
                m.jdstate = '无'
                m.cght = hthm
                m.chrq = chrq
                m.zzsl = int(zzsl) if zzsl else 0
                m.tsl = int(tsl) if tsl else 0
                m.hkrq = hkrq if hkrq else None
                m.gcmc = gcmc if gcmc else ''
                m.fkbz = fkbz if fkbz else ''
                m.hkje = float(hkje) if hkje else 0
                m.htje = float(htje) if htje else 0
                m.bhsj = float(bhsj) if bhsj else 0
                m.fpje = float(fpje) if fpje else 0
                m.se = float(se) if se else 0
                m.tse = float(tse) if tse else 0
                m.tse = round((float(fpje) / (1 + int(zzsl) / 100) * int(tsl) / 100) * 100) / 100
                m.jsrm = jsrm if jsrm else ''
                m.sh = sh if sh else ''
                m.gcmc = gcmc if gcmc else ''
                m.bank = bank if bank else ''
                m.zh = zh if zh else ''
                m.province = province if province else ''
                m.city = city if city else ''
                m.yt = yt if yt else ''
                m.bzsm = bzsm if bzsm else ''
                m.fpyz = fpyz if fpyz else ''
                m.wxfp = wxfp if wxfp else ''
                m.ysfp = wxfp if wxfp else ''
                m.ysfp1 = wxfp if wxfp else ''
                m.yfph = wxfp if wxfp else ''
                m.fpwk = fpwk if fpwk else ''
                m.fkdq = fkdq if fkdq else ''
                m.sb = sb if sb else ''
                m.kh = kh if kh else ''
                m.gcdh = gcdh if gcdh else ''
                m.gcmc1 = gcmc1 if gcmc1 else ''
                m.rxfs = rxfs if rxfs else ''
                m.dlrq = dlrq if dlrq else None
                # m.xjxh = str(i-1) if i else '' # 这里的i是行数减1，可以根据需要调整
                m.sfjq = '否'
                m.kkhj1 = float(kkhj1) if kkhj1 else 0
                m.yhdm = yhdm1 if yhdm1 else ''
                m.yhdz = yhdz1 if yhdz1 else ''
                m.wyjgh = wyjgh1 if wyjgh1 else ''                
                m.lhh = lhh1 if lhh1 else ''
                m.jgh = jgh1 if jgh1 else ''
                m.sjyh = sjyh1 if sjyh1 else ''
                m.sjlhh = sjlhh1 if sjlhh1 else ''                
                m.sjhjgh = sjhjgh1 if sjhjgh1 else ''
                m.fksb = fksb
                m.dkje = 0
                m.sqrq = sqrq if sqrq else None
                m.ywbm = ywbm if ywbm else ''
                m.sqje = float(sqje) if sqje else 0
                m.tqfk = tqfk if tqfk else ''
                # m.zsfkrq = zsfkrq if zsfkrq else None 没有赋值
                m.sfdl = sfdl
                m.fkhb = '否'
                m.pk = float(pk) if pk else 0
                m.customer = customer if customer else ''
                m.ywry = ywry if ywry else ''
                m.scqy = scqy if scqy else ''
                m.fjq = '否'
                m.cxnf = '2017'
                m.cxbg = cxbg if cxbg else ''
                m.wstt = ''

                n = gchksheet()
                n.rid = get_uuid()
                n.pid = pid
                n.yjfk = zsrq if zsrq else None
                n.bz = bz if bz else ''
                n.cphh = cphh if cphh else ''
                n.yfje = float(zsl * dj) if zsl and dj else 0
                n.zwpm = zwpm if zwpm else ''
                n.dj = float(dj) if dj else 0
                n.dw = dw if dw else ''
                n.bz1 = bz1 if bz1 else ''
                n.xs = float(xs) if xs else 0
                n.zsl = float(zsl) if zsl else 0
                n.hkrq = hkrq if hkrq else None
                n.gcmc = gcmc1 if gcmc1 else ''
                n.rxfs = rxfs if rxfs else ''
                m.cght = hthm if hthm else ''
                n.bh = ''
                n.cywyzd = ''
                n.fphm = wxfp if wxfp else ''
                n.fkbz = fkbz if fkbz else ''
                n.yfph = wxfp if wxfp else ''
                n.ysfp = wxfp if wxfp else ''
                n.ysfp1 = wxfp if wxfp else ''
                n.xjsb = '是'
                s.add(n)
                
                while True:
                    bb = ws['AE' + str(i+1)]
                    if bb == '' or bb == None or aa != bb:
                        break
                    i = i + 1
                    hthm = ws['S' + str(i)]
                    xx = ws['W' + str(i)]
                    cphh = ws['X' + str(i)]
                    bz = ws['AH' + str(i)]
                    zwpm = ws['Y' + str(i)]
                    dj = ws['Z' + str(i)]
                    dw = ws['AA' + str(i)]
                    bz1 = ws['AB' + str(i)]
                    xs = ws['AC' + str(i)]
                    zsl = float(bz1) * float(xs)
                    gcmc1 = ws['AE' + str(i)]
                    rxfs = ws['AF' + str(i)]
                    yhfy = ws['AL' + str(i)]

                    yhfy1 = yhfy1 + zsl * float(yhfy)
                    yfhj1 = yfhj1 + zsl * float(dj)
                    zsl1 = zsl1 + zsl
                    xs1 = xs1 + float(xs)

                    n = gchksheet()
                    n.rid = get_uuid()
                    n.pid = pid
                    n.yjfk = zsrq if zsrq else None
                    n.bz = bz if bz else ''
                    n.cphh = cphh if cphh else ''
                    n.yfje = float(zsl * dj) if zsl and dj else 0
                    n.zwpm = zwpm if zwpm else ''
                    n.dj = float(dj) if dj else 0
                    n.dw = dw if dw else ''
                    n.bz1 = bz1 if bz1 else ''
                    n.xs = float(xs) if xs else 0
                    n.zsl = float(zsl) if zsl else 0
                    n.hkrq = hkrq if hkrq else None
                    n.gcmc = gcmc1 if gcmc1 else ''
                    n.rxfs = rxfs if rxfs else ''
                    m.cght = hthm if hthm else ''
                    n.bh = ''
                    n.cywyzd = ''
                    n.fphm = wxfp if wxfp else ''
                    n.fkbz = fkbz if fkbz else ''
                    n.yfph = wxfp if wxfp else ''
                    n.ysfp = wxfp if wxfp else ''
                    n.ysfp1 = wxfp if wxfp else ''
                    n.xjsb = '是'
                    s.add(n)
                
                sjzf1 = float(kkhj1)
                sjzf3 = yfhj1 - sjzf1
                sjzf2 = math.trunc(sjzf3 / 10) * 10
                if htje == '' or htje is None:
                    htje = float(yfhj1)
                if 'YWB' in gcmc2 or '义乌办' in gcmc2 or gcmc2 == '' or 'WM' in gcmc2:
                    cxbg = '不需要'
                else:
                    cxbg = '不需要'
                    if cxje > 0 and (float(fpje) > cxje) or (float(htje) > cxje):
                        if bm1 == 'D':
                            p = s.query(cxgc).filter(cxgc.gcmc == gcmc2 | cxgc.chgc == gcmc2).first()
                            if not p:
                                cxbg = '待提供'
                                z = cxgc()
                                z.rid = get_uuid()
                                z.uid = user.rid
                                z.ctime = time.strftime('%Y-%m-%d %H:%M:%S')
                                z.gcmc = gcmc2
                                z.wxfp = wxfp
                                z.ywry = jsrm
                                z.ywz = ywbm
                                z.yrrq = time.strftime('%Y-%m-%d')
                                z.sfsh = '否'
                                z.yrgs = gsmc
                                z.qrrq = None
                                z.chgc = gcmc2
                                z.fkry = ''
                                z.xgry = ''
                                s.add(z)
                                cxje1 = cxje1 + 1
                            else:
                                qrrq = None
                                if c.qrrq == '' or c.qrrq is None:
                                    qrrq = '1999-01-01'
                                else:
                                    qrrq = c.qrrq

                                if qrrq != '' and qrrq is not None and str(qrrq)[:4] != time.strftime('%Y'):
                                    cxbg = '已提供'
                                else:
                                    cxbg = '待提供'
                                cxje1 = cxje1 + 1
                        else:
                            p = s.query(cxgc).filter(cxgc.gcmc == gcmc2 | cxgc.chgc == gcmc2).first()
                            if c and c.sfsh != '是':
                                cxbg = '待提供'

                        if cxbg == '待提供':
                            tmpstr.append('工厂名称:' + str(gcmc2) + '需提交诚信报告')

                m.yfhj = yfhj1
                m.zsl = zsl1
                m.xshj = xs1
                m.yhfy = yhfy1
                m.sjzf = sjzf2
                m.htje = float(htje) if htje else 0
                m.cxbg = cxbg
                s.add(m)
            else:
                tmpstr.append('发票: ' + fpyz + '已存在!')

            i = i + 1

        out_name = ''
        msg = '导入成功'
        if len(tmpstr) > 0:
            msg = '导入存在异常,正在生成txt文件...'
            out_name = f"导入异常记录_{time.strftime('%Y-%m-%d_%H-%M-%S')}.txt"
            out_path = os.path.join(config.tmp_path, out_name)
            val = '\n'.join(tmpstr)
            write_file(out_path, val, 'w')

        s.commit()
        return json_result(0, msg, out_name)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()



"""
导入发票
对应原Pascal: 导入发票
"""
@any_route('/api/saier/payment/update_invoice', methods=['POST'])
@require_token
async def view_payment_update_invoice(request):
    # 1. 检查用户是否是财务岗位 (对应 tmpcom2 查询)
    s = Session()
    try:
        j = await request.form()
        module = form_value(j, 'module', '采购付款')
        user = request.current_user
        org = get_user_path(user.username)
        path = org.get('path')
        position = org.get('position', '')
        fkdq = '宁波'
        if '义乌' in position:
            fkdq = '义乌'

        t = time.strftime('%Y-%m-%d %H:%M:%S')
        t1 = t[2:4] + t[5:7] + t[8:10] + t[11:13] + t[14:16] + t[17:19]
        jsts = 0
        sz1 = 0
        cxje = 0
        cxje1 = 0
        a1 = 0
        pk = 0
        cjs1 = 0.07
        yhs1 = 0.0003
        cxje = 0
        cxje1 = 0

        gsmc = ''
        c = s.query(zx).filter(zx.ly == '财务付款结算期').first()
        if c and c.mc:
            jzsj = c.cs if c.cs else 0
            sz1 = c.sz1 if c.sz1 else 0
        org = get_user_path('zjnblh')
        position = org.get('position', '')
        bm1 = position[:1] if position else ''
        c = s.query(zx).filter(zx.ly == '诚信金额').first()
        if c and c.mc:
            cxje = c.cs if c.cs else 0
        c = s.query(wfgs).filter(ywrybiao.yhm == 'zjnblh').first()
        if c and c.wfgs:
            gsmc = c.wfgs
        i = 0
        kpjxts = 0
        xjjxts = 0
        kpxjts = 0
        xjxh = 0
        scqy = ''
        if '优胜' in path:
            c = s.query(cwjxrq).filter(cwjxrq.szgs.like('%优胜%')).first()
            if c:
                kpjxts = c.kpjxts if c.kpjxts else 0
                xjjxts = c.xjjxts if c.xjjxts else 0
        # 2. 获取上传的Excel文件
        file = form_value(j,'file',None)
        if is_none(file):
            return json_result(ERR_PARAM_NOT_ENOUGH)
        # 保存临时文件
        temp_dir = config.tmp_path
        temp_file = os.path.join(temp_dir, f"{uuid.uuid4()}_{file.filename}")
        # logger.error(f"Saving uploaded file to: {temp_file}")
        data = await file.read()
        write_file(temp_file, data, 'wb', encoding=None)

        # 3. 读取Excel文件
        wb = load_workbook(temp_file)
        ws = wb.worksheets[0]  # 读取第一个工作表
        # 4. 处理每一行数据
        # duplicates = []  # 存储重复记录信息
        i = 2  # 从第2行开始（第1行是表头）
        xgqd = []
        tmpstr = []
        while True:
            # 读取Excel单元格数据
            gcfp = ws['A' + str(i)]
            wxfp = ws['T' + str(i)]

            # 如果rid为空，结束循环
            if not wxfp:
                break

            hthm = ws['B' + str(i)]
            chrq = ws['C' + str(i)]
            fkbz = ws['D' + str(i)]
            if fkbz == '' or fkbz is None:
                fkbz = 'RMB'
            hkje = ws['E' + str(i)]
            htje = ws['F' + str(i)]
            hkrq = ws['G' + str(i)]
            customer = ''
            ywry = ''
            if hkrq == '' or hkrq is None:
                fksb = '否'
            if hkrq != '' and hkrq is not None:
                fksb = '是'
            fpje = ws['H' + str(i)]
            if htje == '' or htje is None:
                htje = fpje
            bhsj = ws['I' + str(i)]
            zzsl = ws['J' + str(i)]
            se = ws['K' + str(i)]
            tsl = ws['L' + str(i)]
            tse = ws['M' + str(i)]
            jsrm = ws['N' + str(i)]
            sh = ws['O' + str(i)]
            yt = ws['P' + str(i)]
            if yt == '' or yt is None:
                yt = '货款'
            fpbh = ws['Q' + str(i)]
            bzsm = ws['R' + str(i)]
            dprq = ws['S' + str(i)]
            fpyz = fpbh + '-' + gcfp
            fpwk = ws['U' + str(i)]
            if fpwk == '' or fpwk is None:
                fpwk = '否'
            if fpwk == '否':
                kpxjts = kpjxts
            if fpwk == '是':
                kpxjts = xjjxts

            fkdq = ws['V' + str(i)]
            if fkdq == '' or fkdq is None:
                i = i + 1
                continue
                # fkdq = '宁波'

            gcmc1 = ws['W' + str(i)]
            yhfy = ws['X' + str(i)]
            dlrq = ws['Y' + str(i)]
            if dlrq == '' or dlrq is None:
                dlrq = time.strftime('%Y-%m-%d')
            tqfk = ws['Z' + str(i)]
            if tqfk == '' or tqfk is None:
                tqfk = '否'

            ywbm = ws['AA' + str(i)]
            sfdl = '否'
            wxfp2 = hthm[:3] if hthm else ''
            if ywbm == '' or ywbm == None:
                wxfp2 = hthm[:3]
                c = s.query(ywbdzb).all()
                for r in c:
                    if r.dyywzm and r.dyywzm in wxfp2:
                        ywbm = r.ywb
                        break
            if ywbm == '' or ywbm == None:
                ywbm, sfdl = extract_code_optimized(hthm, ywbm)

            sqrq = ws['AB' + str(i)]
            ywbm = ws['AA' + str(i)]
            sqrq = ws['AB' + str(i)]
            sqje = ws['AC' + str(i)]
            sfjq = ws['AD' + str(i)]
            zwpm = ws['AE' + str(i)]
            kpgc = ws['AF' + str(i)]
            wstt = ws['AG' + str(i)]
            dd = ws['AH' + str(i)]
            fpxh = ws['AI' + str(i)]
            sfdh = ws['AJ' + str(i)]
            gzsh = ws['AK' + str(i)]
            htxq1 = ws['AL' + str(i)]
            sfdh = ws['AM' + str(i)]
            gzsh = ws['AN' + str(i)]
            gcmc2 = ws['AO' + str(i)]
            wxh = ws['AP' + str(i)]
            gzrq1 = ws['AQ' + str(i)]
            dhrq1 = ws['AR' + str(i)]
            htsh = ws['AS' + str(i)]
            fkxq = ws['AT' + str(i)]
            szxq = ws['AU' + str(i)]
            zkfy1 = ws['AV' + str(i)]
            ywdq = ws['AW' + str(i)]
            if sfjq == '' or sfjq is None:
                sfjq = '否'
            shui = ''
            gcmc = ''
            bank = ''
            zh = ''
            province = ''
            city = ''
            fhsh = ''
            yhdm1 = ''
            yhdz1 = ''
            wyjgh1 = ''
            lhh1 = ''
            jgh1 = ''
            sjyh1 = ''
            sjlhh1 = ''
            sjhjgh1 = ''
            
            xjxh = xjxh + 1
            chrq = format_date_string(chrq)
            dlrq = format_date_string(dlrq)
            sb = str(i-1) + str(user.username) + str(wxfp) + str(t1)
            org = get_user_path(user.username)
            path1 = org.get('path', '')
            if path1 != '' and path1 != None:
                path = path1 

            if sh != '' and sh != None:
                c = s.query(newcwcs).filter(newcwcs.company_name == sh, newcwcs.shui==shui, newcwcs.company_name== str(sh) + '司').first()
                if c :

                    shui = c.shui
                    gcmc = c.company_name
                    bank = c.bank
                    zh = c.zh
                    province = c.province
                    city = c.city
                    fhsh = c.fhsh
                    yhdm1 = c.yhdm
                    yhdz1 = c.yhdz
                    wyjgh1 = c.wyjgh
                    lhh1 = c.lhh
                    jgh1 = c.jgh
                    sjyh1 = c.sjyh
                    sjlhh1 = c.sjlhh
                    sjhjgh1 = c.sjhjgh

                    if ('黑名单' in c.hzdj):
                        hmd = hmd + 1
                        tmpstr.append('请注意工厂发票: ' + str(gcfp) + '工厂名称:' + str(gcmc) + str(c.hzdj) + '！！')
                        scqy =c.hzdj
            
            if (gcmc == '' or gcmc == None) and (kpgc != '' and kpgc != None):
                gcmc = kpgc
            if fhsh != '' and fhsh != None:
                tmpstr.append('发函税号: ' + str(fhsh) + '工厂名称:' + str(gcmc))
            else:
                fhsh = ''

            if bank != '':
                c = s.query(yhbm).filter(yhbm.yhdm == bank).first()
                if c:
                    yhdm1 = c.yhdm
                    yhdz1 = c.yhdz if c.yhdz else ''
                    wyjgh1 = c.wyjgh
                    lhh1 = c.lhh
                    jgh1 = c.jgh
                    sjyh1 = c.sjyh
                    sjlhh1 = c.sjlhh
                    sjhjgh1 = c.sjhjgh
            wckp = ''
            x = s.query(gchk).filter(gchk.fpyz == fpyz).first()
            if not x or fpyz == '' or fpyz == None:
                a = s.query(gchk.sid).order_by(gchk.sid.desc()).first()
                if a:
                    sb = str(user.username) + str(t1) + str(a.sid)
                pk = 0
                if hkje != '0' and hkje != '' and hkje != None:
                    result = chinese_amount(float(hkje))

                fkn = ''
                fky = ''
                fkr = ''
                year = ''
                month = ''
                day = ''
                if hkrq != '' and hkrq != None:
                    if isinstance(hkrq, datetime):
                        hkrq = hkrq.strftime('%Y-%m-%d')
                    year = hkrq[:4]
                    month = hkrq[5:7]
                    day = hkrq[8:10]
                    fky = ''.join(ArabiaToChinese(ch) for ch in year)
                    fkn = format_month_to_chinese(month)
                    fkr = format_month_to_chinese(day)

                gsfprq = None
                kaipnumber = ''
                ckhwbgdh = ''
                wckp = ''
                ywfp = ''
                htxs = ''
                htxq1 = ''
                sfdh = ''
                gzsh = ''
                gcmc2 = ''
                wxh = ''
                gzrq1 = None
                dhrq1 = None
                htsh = None
                if fpxh != '' and fpxh != None:
                    c = s.query(kaiptzsheet1).filter(kaiptzsheet1.kpxh == fpxh).first()
                    if c:
                        c.sfkp = '是'
                        s.add(c)
                        wckp = '否'
                        ywfp = c.fphm
                        htxs = c.htxs if c.htxs else ''
                        htxq1 = str(htxs)[:241]
                        sfdh = c.sfdh
                        gzsh = c.gzsh
                        gcmc2 = c.sccj2
                        wxh = c.wxh
                        gzrq1 = c.gzrq1
                        dhrq1 = c.dhrq1
                        htsh = c.htsh
                        kaipnumber = c.pid
                        wckp = c.wckp
                        ckhwbgdh = c.ckhwbgdh
            
                    if kaipnumber != '':
                        c = s.query(kaiptz).filter(kaiptz.rid == kaipnumber).first()
                        if c:
                            gsfprq = c.gsfprq

                if gcmc2 == '' or gcmc2 == None:
                    gcmc2 = gcmc1 if gcmc1 else ''

                jsrybm = ''
                c = s.query(ywrybiao).filter(ywrybiao.yhm == jsrm).first()
                if c:
                    jsrybm = c.bm if c.bm else ''

                zsrq = None
                sz = 0
                if chrq != '' and chrq != None:
                    chrq = chrq[0:10]
                    rq_date = datetime.strptime(str(chrq), '%Y-%m-%d')
                    # 计算起始日期
                    current_date = rq_date + timedelta(days=jsts)
                    xjxh = get_date_weekday(current_date)
                    i1 = 0
                    for i2 in range(1, 8):
                        if sz != sz1 :
                            # 计算起始日期
                            current_date = rq_date + timedelta(days=(jsts+i2))
                            xjxh = get_date_weekday(current_date)
                            i1 = i1 + 1
                    zsrq = rq_date + timedelta(days=(jsts+i1))
                    zsrq = zsrq.strftime('%Y-%m-%d')
                    
                cxbg = '不需要'
                org = get_user_path(jsrm)
                uid = org.get('rid', '')
                m = gchk()
                pid = get_uuid()
                m.rid = pid
                m.uid = uid
                m.ctime = time.strftime('%Y-%m-%d %H:%M:%S')
                m.fpbh = gcfp
                m.gcfp = gcfp
                m.yjfk = zsrq
                m.cght = hthm
                m.chrq = chrq
                m.zzsl = int(zzsl) if zzsl else 0
                m.tsl = int(tsl) if tsl else 0
                m.hkrq = hkrq if hkrq else None
                m.gcmc = gcmc if gcmc else ''
                m.fkbz = fkbz if fkbz else ''
                m.hkje = float(hkje) if hkje else 0
                m.htje = float(htje) if htje else 0
                m.bhsj = float(bhsj) if bhsj else 0
                m.fpje = float(fpje) if fpje else 0
                m.se = float(se) if se else 0
                m.jsrybm= jsrybm
                m.gsfprq= gsfprq
                m.ckhwbgdh= ckhwbgdh
                m.wckp= wckp
                m.ysfp= ywfp
                m.jdstate= '无'
                m.scqy= scqy
                if  float(zzsl) == float(tsl):
                    m.tse = float(se) if se else 0
                else:
                    m.tse = round((float(fpje) / (1 + int(zzsl) / 100) * int(tsl) / 100) * 100) / 100
                m.jsrm = jsrm
                m.sh = shui
                m.bank = bank
                m.zh = zh
                m.province = province
                m.city = city
                m.yt = yt
                m.bzsm = bzsm
                m.fpyz = fpyz
                m.wxfp = wxfp
                m.fpwk = fpwk
                m.fkdq = fkdq
                m.sb = sb
                m.gcmc1 = gcmc1
                m.yhfy= float(yhfy) if yhfy != '' and yhfy is not None else 0
                m.dlrq = dlrq
                m.yhdm = yhdm1
                m.yhdz = yhdz1
                m.wyjgh = wyjgh1
                m.lhh = lhh1
                m.jgh = jgh1
                m.sjyh = sjyh1
                m.sjlhh = sjlhh1
                m.sjhjgh = sjhjgh1
                m.fksb = fksb
                m.dkje= 0
                m.fhsh = fhsh
                m.sqrq = sqrq
                m.ywbm = ywbm
                m.sqje= float(sqje) if sqje != '' and sqje is not None else 0
                m.tqfk = tqfk
                # m.zsfkrq = zsfkrq
                m.sfjq = sfjq

                m.sfdl = sfdl
                m.fkdx = result
                m.fkn = fkn
                m.fky = fky
                m.fkr = fkr
                m.fkhj= float(hkje) if hkje != '' and hkje is not None else 0
                m.fkhb = '否'
                m.pk = pk
                m.customer = customer
                m.ywry = ywry
                m.nsz = year
                m.nsy = month
                m.nsr = day
                m.zwpm = zwpm
                m.wstt = wstt
                m.kpxh = fpxh
                m.fjq = '否'
                m.dkje1= float(dd) if dd != '' and dd is not None else 0
                if float(zzsl) == 3:
                    m.cjs= round((float(se) * cjs1 * 100) / 2) / 100
                    m.yhs= round((float(fpje) / 1.03 * yhs1 * 100) / 2) / 100
                    m.sjhj= round(((float(fpje) / 1.03 * yhs1) / 2 + (float(se) * cjs1) / 2 + float(se)) * 100) / 100
                else:
                    m.sjhj= float(se)
                m.dprq = dprq
                m.htxq1 = htxq1
                m.sfdh = sfdh
                m.gzsh = gzsh
                m.cxbg = cxbg
                m.wxh = wxh
                m.gzrq1 = gzrq1
                m.dhrq1 = dhrq1
                m.htsh = htsh
                m.fkxq = fkxq
                m.szxq = szxq
                m.zkfy1= float(zkfy1) if zkfy1 != '' and zkfy1 is not None else 0
                m.ywdq = ywdq

                yfjez = 0
                ysqjez = 0
                htrqz = None
                yfph = ''
                ysfp1 = ''
                if float(hkje) > 0:
                    n = gchksheet1()
                    n.rid = get_uuid()
                    n.uid = uid
                    n.ctime = time.strftime('%Y-%m-%d %H:%M:%S')
                    n.pid = pid
                    n.fkrq = hkrq
                    n.hkje = hkje
                    n.czry = user.username
                    s.add(n)

                if fpxh != '' and fpxh is not None:
                    c = s.query(kaiptzsheet1).filter(kaiptzsheet1.kpxh == fpxh).first()
                    if c:
                        if c.yfph1 != '' and c.yfph1 is not None:
                            yfph = c.yfph1
                            ysfp1 = c.ysfp1
                        fatherkp = c.pid
                        zwpm12 = c.zwpm
                        hyd12 = c.hyd
                        zsl12 = c.zsl
                        SSB = '2'
                        b = s.query(kaiptzsheet).filter(kaiptzsheet.pid == fatherkp, kaiptzsheet.zsl == zsl12, kaiptzsheet.zwpm == zwpm12, kaiptzsheet.hyd == hyd12).first()
                        if b:
                            SSB = '1'
                        if SSB == '1':
                            z = s.query(kaiptzsheet).filter(kaiptzsheet.pid == fatherkp, kaiptzsheet.zsl == zsl12, kaiptzsheet.zwpm == zwpm12, kaiptzsheet.hyd == hyd12).all()
                        else:
                            z = s.query(kaiptzsheet).filter(kaiptzsheet.pid == fatherkp).all()
                        for r in z:
                            izz = izz + 1
                            n = gchksheet()
                            n.rid = get_uuid()
                            n.pid = pid
                            n.uid = uid
                            n.ctime = time.strftime('%Y-%m-%d %H:%M:%S')
                            n.pid = pid
                            n.fphm = r.wxfp
                            n.ysfp = r.yfph
                            n.cphh = r.cpbh
                            n.yjfk = zsrq
                            n.cght = r.cght
                            n.zwpm = r.zwpm
                            n.dj = r.gcjg
                            n.dw = r.jldw
                            n.xs = r.chxs
                            n.bz1 = ''
                            n.zsl = r.chsl
                            n.hkrq = ''
                            n.gcmc = r.sccj
                            n.rxfs = ''
                            n.bz = ''
                            n.ywbm = r.wxbm1
                            n.yfph = r.yfph
                            n.ysfp1 = r.yysfp
                            n.Yfje1 = r.Yfje
                            n.kpds = r.kpds
                            n.ysqje = r.ysqje
                            n.yfkp = r.yfkp
                            n.cywyzd = r.cywyzd
                            n.fktt = r.fktt
                            n.htrq = r.htrq
                            n.sjcy = r.sjcy
                            n.jsfs = r.jsfs
                            n.bh = fpxh
                            n.khhh = r.khhh
                            n.fkbz = fkbz
                            if r.chsl * r.gcjg != 0:
                                n.Yfje = r.chsl * r.gcjg
                            else:
                                n.Yfje = r.gczj
                            n.cgry = r.cgry
                            n.ywry = r.ywry
                            n.gdry = r.gdry
                            n.zzsl = r.zzsl
                            n.tsl = round(r.zsl)
                
                # sjzf1 = float(kkhj1)
                # sjzf3 = yfhj1 - sjzf1
                # sjzf2 = math.trunc(sjzf3 / 10) * 10
                # if htje == '' or htje is None:
                #     htje = float(yfhj1)
                logger.error(f"工厂名称: {gcmc2}, 诚信金额: {cxje}, 发票金额: {fpje}, 合同金额: {htje}")
                if 'YWB' in gcmc2 or '义乌办' in gcmc2 or gcmc2 == '' or 'WM' in gcmc2:
                    cxbg = '不需要'
                else:
                    cxbg = '不需要'
                    if cxje > 0 and (float(fpje) > cxje) or (float(htje) > cxje):
                        if bm1 == 'D':
                            p = s.query(cxgc).filter(cxgc.gcmc == gcmc2 | cxgc.chgc == gcmc2).first()
                            if not p:
                                cxbg = '待提供'
                                z = cxgc()
                                z.rid = get_uuid()
                                z.uid = user.rid
                                z.ctime = time.strftime('%Y-%m-%d %H:%M:%S')
                                z.gcmc = gcmc2
                                z.wxfp = wxfp
                                z.ywry = jsrm
                                z.ywz = ywbm
                                z.yrrq = time.strftime('%Y-%m-%d')
                                z.sfsh = '否'
                                z.yrgs = gsmc
                                z.qrrq = None
                                z.chgc = gcmc2
                                z.fkry = ''
                                z.xgry = ''
                                s.add(z)
                                cxje1 = cxje1 + 1
                            else:
                                qrrq = None
                                if c.qrrq == '' or c.qrrq is None:
                                    qrrq = '1999-01-01'
                                else:
                                    qrrq = c.qrrq
                                if htrqz != '' and htrqz is not None:
                                    if qrrq != '' and qrrq is not None and (htrqz - qrrq).days > 365:
                                        cxbg = '待提供'
                                        p.sfsh = '否'
                                        p.yrrq =  time.strftime('%Y-%m-%d')
                                        p.yrgs =  wfgs
                                        p.qrrq =  None
                                        s.add(p)
                                    else:
                                        cxbg = '已提供'
                                else:
                                    if qrrq != '' and qrrq is not None and str(qrrq)[:4] != time.strftime('%Y'):
                                        cxbg = '已提供'
                                    else:
                                        cxbg = '待提供'
                                        cxje1 = cxje1 + 1
                        else:
                            p = s.query(cxgc).filter(cxgc.gcmc == gcmc2 | cxgc.chgc == gcmc2).first()
                            if c and c.sfsh != '是':
                                cxbg = '待提供'

                        if cxbg == '待提供':
                            tmpstr.append('工厂名称:' + str(gcmc2) + '需提交诚信报告')

                yfjezz = 0
                c = s.query(func.sum(func.ifnull(gchksheet.Yfje, 0)).label('Yfje')).filter(gchksheet.pid == pid).first()
                if c and c.Yfje is not None:
                    yfjezz = c.Yfje

                m.yfhj = yfjezz
                m.yfje1 = yfjez
                m.ysqhj = ysqjez
                m.htrq = htrqz
                m.cxbg = cxbg
                m.yfph = yfph
                m.ysfp1 = ysfp1
                s.add(m)
                izz1 = 0
                fpjezz = float(fpje)
                if yfjezz > 0 :
                    c = s.query(gchksheet).filter(gchksheet.pid == pid).all()
                    for r in c:
                        izz1 = izz1 + 1
                        r.fkdq = fkdq
                        r.fpwk = '否'
                        r.zzsl = int(zzsl) if zzsl else 0
                        r.tsl = int(tsl) if tsl else 0
                        if izz1 == izz:
                            r.fkje = fpjezz
                        else:
                            r.fkje = round((r.Yfje / yfjezz) * float(fpje), 2)
                            fpjezz = fpjezz - round((r.Yfje / yfjezz) * float(fpje), 2)
                        s.add(r)
                s.query(fpgl).filter(fpgl.wxfp == wxfp, fpgl.tssd == '否', fpgl.sfjd == '否').update({'webpdfk': '是'})                
            else:
                tmpstr.append('发票: ' + fpyz + '已存在!')

            i = i + 1

        out_name = ''
        msg = '导入成功'
        if len(tmpstr) > 0:
            msg = '导入发票存在异常,正在生成txt文件...'
            out_name = f"导入发票异常记录_{time.strftime('%Y-%m-%d_%H-%M-%S')}.txt"
            out_path = os.path.join(config.tmp_path, out_name)
            val = '\n'.join(tmpstr)
            write_file(out_path, val, 'w')

        s.commit()
        return json_result(0, msg, out_name)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()


"""
销售帐入帐清单(优胜)_当前选中记录
对应原Pascal: 销售帐入帐清单(优胜)_当前选中记录
"""
@any_route('/api/saier/payment/account_list_rids', methods=['POST'])
@require_token
async def view_payment_account_list_rids(request):
    # 1. 检查用户是否是财务岗位 (对应 tmpcom2 查询)
    s = Session()
    try:
        j = await request.json()
        rids = j.get('rids', [])
        if len(rids) == 0:
            return json_result(-1, '未选中记录')
        fp_json = {}
        for rid in rids:
            c = s.query(gchk.wxfp).filter(gchk.rid == rid).first()
            if not c:
                continue
            if c.wxfp == None or c.wxfp == "":
                continue
            if c.wxfp not in fp_json:
                fp_json[c.wxfp] = [rid]
            else:
                fp_json[c.wxfp].append(rid)

        res = payment_account_export_list(s, fp_json, rids)
        if res.get('code') != 1:
            return json_result(-1, res.get('msg', '导出失败'))
        data = res.get('data')

        return json_result(1, '操作成功', data)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()


"""
销售帐入帐清单(优胜)_excel导入接口
对应原Pascal: 销售帐入帐清单(优胜)_excel导入接口
"""
@any_route('/api/saier/payment/account_list_excel', methods=['POST'])
@require_token
async def view_payment_account_list_excel(request):
    # 1. 检查用户是否是财务岗位 (对应 tmpcom2 查询)
    s = Session()
    try:
        j = await request.form()
        # module = form_value(j, 'module', '采购付款')
        # user = request.current_user
        # org = get_user_path(user.username)
        # path = org.get('path')
        # position = org.get('position', '')
        # fkdq = '宁波'
        # if '义乌' in position:
        #     fkdq = '义乌'
        # 2. 获取上传的Excel文件
        file = form_value(j,'file',None)
        if is_none(file):
            return json_result(ERR_PARAM_NOT_ENOUGH)
        # 保存临时文件
        temp_dir = config.tmp_path
        temp_file = os.path.join(temp_dir, f"{uuid.uuid4()}_{file.filename}")
        # logger.error(f"Saving uploaded file to: {temp_file}")
        data = await file.read()
        write_file(temp_file, data, 'wb', encoding=None)

        # 3. 读取Excel文件
        wb = load_workbook(temp_file)
        ws = wb.worksheets[0]  # 读取第一个工作表
        # 4. 处理每一行数据
        i = 2  # 从第2行开始（第1行是表头）
        rids = []
        tmpstr = []
        fp_json = {}
        while True:
            wxfp = ws['A' + str(i)].value
            logger.error(f"Processing row {i}, wxfp: {wxfp if wxfp else 'None'}")
            if wxfp == '' or wxfp is None:
                break
            else:
                c = s.query(gchk.rid).filter(gchk.wxfp == wxfp).all()
                for r in c:
                    if wxfp not in fp_json:
                        fp_json[wxfp] = [r.rid]
                    else:
                        fp_json[wxfp].append(r.rid)
            i = i + 1

        if len(fp_json) == 0:
            return json_result(-1, '没有找到对应的发票记录')
        
        res = payment_account_export_list(s, fp_json, rids)
        if res.get('code') != 1:
            return json_result(-1, res.get('msg', '导出失败'))
        data = res.get('data')

        return json_result(0, '操作成功', data)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()


def sanitize_filename(filename: str) -> str:
    """
    清理文件名中的非法字符
    返回第一个非法字符之前的部分，如果无非法字符则返回原字符串
    """
    # Windows/Linux 文件名中的非法字符
    illegal_chars = r'<>:"/\|?*'
    
    for i, char in enumerate(filename):
        if char in illegal_chars or ord(char) < 32:  # 控制字符
            return filename[:i]
    return filename

def payment_account_export_list(s, fp_json, rids):
    try:
        wb = Workbook()
        ws = wb.active
        index = 0
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                top=Side(style='thin'), bottom=Side(style='thin'))
        for k,v in fp_json.items():
            cx1z = sanitize_filename(k)
            if index == 0:
                ws.title = cx1z
            else:
                wb.create_sheet(title=cx1z)
            sjchny = ''
            ws = wb.worksheets[index]
            ws.sheet_view.showGridLines = False
            # logger.error(f"正在处理发票: {k} 对应记录数: {ws.title}")
            i = 1
            ws['A' + str(i)] = '销售清单'
            ws.merge_cells('A' + str(i) + ':H' + str(i))
            ws['A' + str(i)].alignment = Alignment(horizontal='center', vertical='center')
            for col in range(1, 10):
                cell = ws.cell(row=i, column=col)
                cell.font = Font(size=12)
            i = i + 1
            ws['A' + str(i)] = '工厂发票'
            ws.column_dimensions['A'].width = 7.2
            ws['B' + str(i)] = '外销发票'
            ws.column_dimensions['B'].width = 15.6
            ws['C' + str(i)] = '生产厂家'
            ws.column_dimensions['C'].width = 30.3
            ws['D' + str(i)] = '发票金额'
            ws.column_dimensions['D'].width = 9.8
            ws['E' + str(i)] = '不含税价'
            ws.column_dimensions['E'].width = 9.8
            ws['F' + str(i)] = '税    额'
            ws.column_dimensions['F'].width = 9.3
            ws['G' + str(i)] = '退税率'
            ws.column_dimensions['G'].width = 8
            ws['H' + str(i)] = '退税额'
            ws.column_dimensions['H'].width = 8.8
            for cell in ws['A2:H2'][0]:
                cell.border = border
                cell.font = Font(size=9)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            fpje = 0
            bhsj = 0
            se = 0
            tse = 0
            bgbz = ''
            a = s.query(bgmxd.hbdm).filter((bgmxd.fphm == k) | (bgmxd.ysfp == k)).first()
            if a:
                bgbz = a.hbdm
            for r in v:
                c = s.query(gchk).filter(gchk.rid == r).first()
                if not c:
                    continue
                i = i + 1
                wxfp = c.wxfp
                bgbz = ''
                if sjchny == '' or sjchny == None:
                    sjchny = c.sjchny
                for cell in ws[f"A{i}:H{i}"][0]:
                    cell.border = border
                ws['A' + str(i)] = c.gcfp
                ws['B' + str(i)] = c.wxfp
                ws['C' + str(i)] = c.gcmc
                ws['D' + str(i)] = c.fpje
                ws['E' + str(i)] = c.bhsj
                ws['F' + str(i)] = c.se
                ws['G' + str(i)] = c.tsl
                ws['H' + str(i)] = c.tse
                fpje = fpje + (c.fpje if c.fpje else 0)
                bhsj = bhsj + (c.bhsj if c.bhsj else 0)
                se = se + (c.se if c.se else 0)
                tse = tse + (c.tse if c.tse else 0)

            i = i + 1
            for cell in ws[f"A{i}:H{i}"][0]:
                cell.border = border
            ws['C' + str(i)] = '合计:'
            ws['D' + str(i)] = fpje
            ws['E' + str(i)] = bhsj
            ws['F' + str(i)] = se
            ws['H' + str(i)] = tse
            i = i + 1
            ws['E' + str(i)] = '成本结转:'
            ws['F' + str(i)] = tse - se
            i = i + 1
            ws['E' + str(i)] = '实际出货年月:'
            ws['F' + str(i)].number_format = '@'
            ws['F' + str(i)] = sjchny
            i = i + 1
            ws['E' + str(i)] = '实际出货汇率:'
            if (bgbz != 'USD$') and (bgbz != 'USD') and (bgbz != 'RMB') :
                a = s.query(cwhshlsheet.hl).filter(cwhshlsheet.hsny == sjchny, cwhshlsheet.hbdm == bgbz).first()
            else:
                a = s.query(cwhshl.hl).filter(cwhshl.hsny == sjchny).first()
            if a:
                ws['F' + str(i)].number_format = '@'
                ws['F' + str(i)] = a.hl
            for row in range(2, i + 2):
                for col in range(1, 10):
                    cell = ws.cell(row=row, column=col)
                    cell.font = Font(size=10)
            index = index + 1

        out_name = f"销售帐入帐清单_{time.strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        out_path = os.path.join(config.tmp_path, out_name)
        wb.save(out_path)

        return {'code': 1, 'msg': '操作成功', 'data': out_name}
    except Exception as e:
        logger.error(trace_error())
        return {'code': -1, 'msg': f'导出失败: {str(e)}'}
    



# ==================== 辅助函数 ====================
def is_leap_year(year: int) -> bool:
    return (year % 4 == 0 and year % 100 != 0) or (year % 400 == 0)


def get_month_last_day(year: int, month: int) -> str:
    if month == 2:
        day = 29 if is_leap_year(year) else 28
    elif month in [4, 6, 9, 11]:
        day = 30
    else:
        day = 31
    return f"{day:02d}"


def get_kpxh_number(kpxh_char: str) -> str:
    mapping = {
        'a': '1', 'A': '1', 'b': '2', 'B': '2',
        'c': '3', 'C': '3', 'd': '4', 'D': '4',
        'e': '5', 'E': '5', 'f': '6', 'F': '6'
    }
    return mapping.get(kpxh_char, '')


def format_index(idx: int) -> str:
    if idx <= 9:
        return '000'
    elif idx <= 99:
        return '00'
    elif idx <= 999:
        return '0'
    return ''


def round_two_decimals(value: float) -> float:
    """保留两位小数"""
    return round(value * 100) / 100


# ==================== 数据库操作函数 ====================
def update_gchk_status(s, da1: str, da2: str, wsttz: str):
    """更新 gchk 表状态"""
    s.query(gchk).filter(
        and_(
            gchk.jdsj >= da1,
            gchk.jdsj <= da2,
            gchk.wstt == wsttz,
            gchk.fpwk == '否',
            gchk.wckp == '是',
            or_(
                gchk.jdstate != '已上传',
                gchk.jdstate == '',
                gchk.jdstate.is_(None),
                gchk.jdstate == '无'
            )
        )
    ).update({gchk.jdstate: '待上传'}, synchronize_session=False)


def get_server_config(s) -> tuple:
    """获取服务器IP和密码"""
    fwq = s.query(fwqgl).filter(fwqgl.fwqm.like('%优景%')).first()
    if fwq:
        return fwq.ipdz, fwq.password
    return '60.12.220.xx,xx', 'xx'


def get_position_info(s, some_id: str) -> str:
    """获取位置信息（原 tmpcom41）"""
    # 根据实际表结构调整
    # result = s.query(SomeTable.position).filter(...).first()
    # return result[0] if result else ''
    return ''  # 占位


def get_exchange_rate(s, n11: str, yl1: str, r11: str, nbyw: str) -> float:
    """获取汇率"""
    result = s.query(xshl.hl).filter(
        and_(
            xshl.hsrq >= f"{n11}-{yl1}-01",
            xshl.hsrq <= f"{n11}-{yl1}-{r11}",
            xshl.ly == f"{nbyw}金蝶"
        )
    ).first()
    hl = result[0] if result else 0
    return hl if hl != 0 else 1


def get_cwhshl_rate(s, sjchny: str) -> float:
    """从 cwhshl 获取汇率"""
    result = s.query(cwhshl.hl).filter(cwhshl.hsny == sjchny).first()
    return result[0] if result else 1


def get_zpbm_by_mc(s, mc: str) -> tuple:
    """根据名称获取账簿编码"""
    z = s.query(zx).filter(
        and_(zx.ly == "金蝶账簿编码", zx.mc == mc)
    ).first()
    if z:
        return z.wb1, z.wb2
    return '', ''


def get_gchk_info_by_wxfp(s, wxfp: str) -> dict:
    """根据发票号获取 gchk 信息"""
    c = s.query(gchk).filter(
        and_(
            gchk.wxfp == wxfp,
            gchk.wstt != '',
            gchk.wstt.isnot(None),
            gchk.kpxh != '',
            gchk.kpxh.isnot(None)
        )
    ).first()
    if c:
        return {
            'wstt': c.wstt,
            'kpxh': c.kpxh,
            'sjchny': c.sjchny
        }
    return {}


def get_invoice_list(s, da1: str, da2: str, wsttz: str) -> list:
    """获取需要处理的发票列表"""
    c = s.query(gchk.wxfp).filter(
        and_(
            gchk.jdsj >= da1,
            gchk.jdsj <= da2,
            gchk.wstt == wsttz,
            gchk.fpwk == '否',
            gchk.wckp == '是',
            or_(
                gchk.jdstate != '已上传',
                gchk.jdstate == '',
                gchk.jdstate.is_(None),
                gchk.jdstate == '无'
            )
        )
    ).distinct().all()
    d = alchemy_object_list_to_dict(c)
    logger.error(f"Found {d} unique invoices")
    return d


def get_summary_by_wxfp(s, wxfp: str) -> list:
    """按发票号分组汇总"""
    results = s.query(
        gchk.gcmc,
        func.sum(gchk.fpje).label('fpje1'),
        func.sum(gchk.se).label('se1'),
        func.sum(gchk.tse).label('tse1'),
        gchk.dlrq
    ).filter(
        and_(gchk.wxfp == wxfp, gchk.fpwk == '否')
    ).group_by(gchk.gcmc, gchk.dlrq).order_by(gchk.dlrq.desc()).all()

    return [{
        'gcmc': r[0],
        'fpje1': float(r[1]) if r[1] else 0,
        'se1': float(r[2]) if r[2] else 0,
        'tse1': float(r[3]) if r[3] else 0
    } for r in results]


def get_company_jdbm(s, company_name: str) -> tuple:
    """获取公司编码"""
    newcw = s.query(newcwcs).filter(
        or_(
            newcwcs.company_name == company_name,
            newcwcs.company_name == f'个体户{company_name}'
        )
    ).first()
    if newcw:
        return newcw.jdbm, newcw.company_name
    return '', ''


def get_bgmxd_info(s, ysfp: str) -> dict:
    """获取 bgmxd 信息"""
    c = s.query(bgmxd).filter(bgmxd.ysfp == ysfp).first()
    if c:
        return {
            'rid': c.rid,
            'bgbgzje': float(c.bgbgzje) if c.bgbgzje else 0,
            'RMBkh': c.RMBkh,
            'CNFyf': float(c.CNFyf) if c.CNFyf else 0,
            'hbdm': c.hbdm
        }
    return {}


def get_cwhshl_rate_by_date(s, hsny: str, hbdm: str) -> float:
    """根据日期和币种获取汇率"""
    result = s.query(cwhshlsheet.hl).filter(
        and_(cwhshlsheet.hsny == hsny, cwhshlsheet.hbdm == hbdm)
    ).first()
    return result[0] if result else 0


# ==================== 业务逻辑函数 ====================
def build_debit_entry(fphm: str, account_code: str, account_name: str, amount: float) -> str:
    """构建借方分录 JSON"""
    return f""",{{ 'FEXPLANATION':'{fphm}货物入库', 'FACCOUNTID':{{'FNumber':'{account_code}','FACCOUNT__NAME':'{account_name}'}},'FCURRENCYID':{{'FNumber':'PRE001','FCURRENCYID__Name': '人民币'}},'FEXCHANGERATETYPE':{{'FNumber':'HLTX01_SYS','FEXCHANGERATETYPE__Name':'固定汇率'}},'FEXCHANGERATE':1,'FAMOUNTFOR':'{amount}','FDEBIT':'{amount}'}}"""


def build_credit_entry(fphm: str, account_code: str, account_name: str, amount: float) -> str:
    """构建贷方分录 JSON"""
    return f""",{{ 'FEXPLANATION':'{fphm}货物入库', 'FACCOUNTID':{{'FNumber':'{account_code}','FACCOUNT__NAME':'{account_name}'}},'FCURRENCYID':{{'FNumber':'PRE001','FCURRENCYID__Name': '人民币'}},'FEXCHANGERATETYPE':{{'FNumber':'HLTX01_SYS','FEXCHANGERATETYPE__Name':'固定汇率'}},'FEXCHANGERATE':1,'FAMOUNTFOR':'{amount}','FCREDIT':'{amount}'}}"""


def build_voucher_data(s, fphm: str, hr: str, nbyw: str) -> dict:
    """构建凭证数据（核心业务逻辑）"""
    # 获取年份月份信息
    n11 = hr[:4]
    y = hr[5:7]
    y1 = str(int(y))

    # 获取当月最后一天
    r11 = get_month_last_day(int(n11), int(y))

    logger.error(f"Processing invoice: {fphm}, Year: {n11}, Month: {y1}, Last Day: {r11}")
    # 获取汇率
    hl = get_exchange_rate(s, n11, y, r11, nbyw)

    # 获取发票对应的 gchk 信息
    gchk_info = get_gchk_info_by_wxfp(s, fphm)
    if not gchk_info:
        return None

    # 如果有 sjchny，更新汇率
    if gchk_info.get('sjchny'):
        hl = get_cwhshl_rate(s, gchk_info['sjchny'])

    # 获取账簿编码
    zpbm, zpbm1 = get_zpbm_by_mc(s, gchk_info['wstt'])

    # 处理 kpxh
    kpxh_raw = gchk_info['kpxh']
    kpxh_first = kpxh_raw[0] if kpxh_raw else ''
    kpxh_num = get_kpxh_number(kpxh_first)
    kpxh_suffix = kpxh_raw[1:11] if len(kpxh_raw) > 1 else ''

    kpxh1 = f"{kpxh_num}{kpxh_suffix}1"
    kpxh2 = f"{kpxh_num}{kpxh_suffix}2"
    wstt = gchk_info['wstt']

    # 获取汇总数据
    summary_list = get_summary_by_wxfp(s, fphm)

    fpje = se = tse = 0
    entries = []
    iz = 1

    # 构建应付账款分录
    for item in summary_list:
        iz += 1
        fpje += round_two_decimals(item['fpje1'])
        se += round_two_decimals(item['se1'])
        tse += round_two_decimals(item['tse1'])

        as1, at1 = get_company_jdbm(s, item['gcmc'])
        if not at1:
            at1 = f"个体户{item['gcmc']}" if len(item['gcmc']) <= 8 else item['gcmc']

        # 应付账款分录
        entry = {
            'type': 'credit',
            'account_code': '2202.01',
            'account_name': '明细应付款',
            'amount': item['fpje1'],
            'detail_code': as1,
            'detail_name': at1
        }
        entries.append(entry)

    # 获取 bgmxd 信息
    bgmxd_info = get_bgmxd_info(s, fphm)

    return {
        'kpxh1': kpxh1,
        'kpxh2': kpxh2,
        'zpbm': zpbm,
        'zpbm1': zpbm1,
        'wstt': wstt,
        'fpje': fpje,
        'se': se,
        'tse': tse,
        'hl': hl,
        'bgmxd': bgmxd_info,
        'entries': entries,
        'hr': hr,
        'n11': n11,
        'y1': y1,
        'fphm': fphm
    }


def send_to_api(voucher_main: dict, voucher_detail: dict, yhm: str, passwd: str):
    """
    发送数据到 API
    
    Args:
        voucher_main: 主凭证数据（kpxh1 对应的凭证）
        voucher_detail: 明细凭证数据（kpxh2 对应的凭证）
        yhm: 用户名
        passwd: 密码
    """
    url = 'http://192.168.1.12:44375/api/appapi/'
    
    # 构建完整的请求数据（参考原代码格式）
    # 原代码: '{''Model'':[' + final_four + ',' + final + ']}'
    request_data = {
        'Model': [voucher_main, voucher_detail]
    }
    
    # 转换为 JSON 字符串（使用双引号，符合 JSON 标准）
    json_str = json.dumps(request_data, ensure_ascii=False)
    
    # 原代码格式: 'a=' + json_str + '?' + yhm + '?' + pass
    # 注意：这种格式可能需要 URL 编码
    payload = f"a={json_str}?{yhm}?{passwd}"
    
    try:
        response = requests.post(
            url, 
            data=payload.encode('utf-8'),
            headers={'Content-Type': 'application/x-www-form-urlencoded; charset=utf-8'},
            timeout=30
        )
        response.raise_for_status()
        return response.text
    except requests.exceptions.RequestException as e:
        print(f"API 请求失败: {e}")
        return None


def build_voucher_json(data: dict, is_main: bool = True) -> dict:
    """
    构建完整的凭证 JSON 对象
    
    Args:
        data: 包含所有凭证数据的字典
        is_main: True 表示构建 kpxh1 的主凭证，False 表示构建 kpxh2 的明细凭证
    """
    kpxh = data['kpxh1'] if is_main else data['kpxh2']
    
    # 基础凭证结构
    voucher = {
        "FBillNo": kpxh,
        "FAccountBookID": {
            "FNumber": data['zpbm'],
            "FAccountBookID__Name": f"{data['wstt']}账簿"
        },
        "FDate": data['hr'],
        "FVOUCHERGROUPID": {
            "FNumber": "PRE001"
        },
        "FATTACHMENTS": 0,
        "FYEAR": data['n11'],
        "FACCBOOKORGID": {
            "FNumber": data['zpbm1'],
            "FACCBOOKORGID__Name": data['wstt']
        },
        "FPERIOD": data['y1'],
        "FEntity": []
    }
    
    if is_main:
        # 构建 kpxh1 的凭证分录（入库分录）
        voucher["FEntity"] = build_main_entries(data)
    else:
        # 构建 kpxh2 的凭证分录（收入、成本等分录）
        voucher["FEntity"] = build_detail_entries(data)
    
    return voucher


def build_main_entries(data: dict) -> list:
    """构建主凭证的分录列表（入库：借库存商品/进项税，贷应付账款）"""
    entries = []
    iz = 1
    
    # 1. 库存商品分录（借方）
    entries.append({
        "FEXPLANATION": f"{data['fphm']}货物入库",
        "FACCOUNTID": {
            "FNumber": "1405.01",
            "FACCOUNT__NAME": "外销"
        },
        "FCURRENCYID": {
            "FNumber": "PRE001",
            "FCURRENCYID__Name": "人民币"
        },
        "FEXCHANGERATETYPE": {
            "FNumber": "HLTX01_SYS",
            "FEXCHANGERATETYPE__Name": "固定汇率"
        },
        "FEXCHANGERATE": 1,
        "FAMOUNTFOR": str(data['fpje'] - data['se']),
        "FDEBIT": str(data['fpje'] - data['se'])
    })
    
    # 2. 进项税额分录（借方）
    entries.append({
        "FEXPLANATION": f"{data['fphm']}货物入库",
        "FACCOUNTID": {
            "FNumber": "2221.01.01.01",
            "FACCOUNT__NAME": "进项税额"
        },
        "FCURRENCYID": {
            "FNumber": "PRE001",
            "FCURRENCYID__Name": "人民币"
        },
        "FEXCHANGERATETYPE": {
            "FNumber": "HLTX01_SYS",
            "FEXCHANGERATETYPE__Name": "固定汇率"
        },
        "FEXCHANGERATE": 1,
        "FAMOUNTFOR": str(data['se']),
        "FDEBIT": str(data['se'])
    })
    
    # 3. 应付账款分录（贷方）- 可能有多个供应商
    for i, entry_data in enumerate(data.get('entries', []), start=3):
        entry = {
            "FEXPLANATION": f"{data['fphm']}货物入库",
            "FACCOUNTID": {
                "FNumber": entry_data['account_code'],
                "FACCOUNT__NAME": entry_data['account_name']
            },
            "FCURRENCYID": {
                "FNumber": "PRE001",
                "FCURRENCYID__Name": "人民币"
            },
            "FEXCHANGERATETYPE": {
                "FNumber": "HLTX01_SYS",
                "FEXCHANGERATETYPE__Name": "固定汇率"
            },
            "FEXCHANGERATE": 1,
            "FAMOUNTFOR": str(entry_data['amount']),
            "FCREDIT": str(entry_data['amount'])
        }
        
        # 如果有辅助核算信息
        if entry_data.get('detail_code'):
            entry["FDetailID"] = {
                "FDetailID__FFlex4": {
                    "Fnumber": entry_data['detail_code'],
                    "FDetailID__FFlex4__Name": entry_data['detail_name']
                }
            }
        
        entries.append(entry)
    
    return entries


def build_detail_entries(data: dict) -> list:
    """构建明细凭证的分录列表（收入、成本、退税等）"""
    entries = []
    
    # 1. 应收账款分录（借方）
    bgmxd = data.get('bgmxd', {})
    hl = data.get('hl', 1)
    
    # 计算外销收入金额
    if bgmxd.get('RMBkh') == '是':
        sales_amount = bgmxd.get('bgbgzje', 0)
    else:
        sales_amount = round_two_decimals(hl * bgmxd.get('bgbgzje', 0))
    
    entries.append({
        "FEXPLANATION": f"{data['fphm']}外销收入",
        "FACCOUNTID": {
            "FNumber": "1122.01",
            "FACCOUNT__NAME": "外销"
        },
        "FCURRENCYID": {
            "FNumber": "PRE001",
            "FCURRENCYID__Name": "人民币"
        },
        "FEXCHANGERATETYPE": {
            "FNumber": "HLTX01_SYS",
            "FEXCHANGERATETYPE__Name": "固定汇率"
        },
        "FEXCHANGERATE": 1,
        "FAMOUNTFOR": str(sales_amount),
        "FDEBIT": str(sales_amount)
    })
    
    # 2. 主营业务收入分录（贷方）
    if data['wstt'] == '宁波景业国际贸易有限公司' and bgmxd.get('CNFyf'):
        # 特殊处理运费
        main_amount = sales_amount - round_two_decimals(hl * bgmxd['CNFyf'])
        entries.append({
            "FEXPLANATION": f"{data['fphm']}外销收入",
            "FACCOUNTID": {
                "FNumber": "6001.01.01",
                "FACCOUNT__NAME": "外销"
            },
            "FCURRENCYID": {
                "FNumber": "PRE001",
                "FCURRENCYID__Name": "人民币"
            },
            "FEXCHANGERATETYPE": {
                "FNumber": "HLTX01_SYS",
                "FEXCHANGERATETYPE__Name": "固定汇率"
            },
            "FEXCHANGERATE": 1,
            "FAMOUNTFOR": str(main_amount),
            "FCREDIT": str(main_amount)
        })
        
        # 运费收入分录
        freight_amount = round_two_decimals(hl * bgmxd['CNFyf'])
        entries.append({
            "FEXPLANATION": f"{data['fphm']}外销收入",
            "FACCOUNTID": {
                "FNumber": "6001.01.03",
                "FACCOUNT__NAME": "外销"
            },
            "FCURRENCYID": {
                "FNumber": "PRE001",
                "FCURRENCYID__Name": "人民币"
            },
            "FEXCHANGERATETYPE": {
                "FNumber": "HLTX01_SYS",
                "FEXCHANGERATETYPE__Name": "固定汇率"
            },
            "FEXCHANGERATE": 1,
            "FAMOUNTFOR": str(freight_amount),
            "FCREDIT": str(freight_amount)
        })
    else:
        entries.append({
            "FEXPLANATION": f"{data['fphm']}外销收入",
            "FACCOUNTID": {
                "FNumber": "6001.01.01",
                "FACCOUNT__NAME": "外销"
            },
            "FCURRENCYID": {
                "FNumber": "PRE001",
                "FCURRENCYID__Name": "人民币"
            },
            "FEXCHANGERATETYPE": {
                "FNumber": "HLTX01_SYS",
                "FEXCHANGERATETYPE__Name": "固定汇率"
            },
            "FEXCHANGERATE": 1,
            "FAMOUNTFOR": str(sales_amount),
            "FCREDIT": str(sales_amount)
        })
    
    # 3. 应收退税款分录（借方）
    if data['tse'] > 0:
        entries.append({
            "FEXPLANATION": f"{data['fphm']}应收退税款",
            "FACCOUNTID": {
                "FNumber": "1812.01",
                "FACCOUNT__NAME": "应收出口退税"
            },
            "FCURRENCYID": {
                "FNumber": "PRE001",
                "FCURRENCYID__Name": "人民币"
            },
            "FEXCHANGERATETYPE": {
                "FNumber": "HLTX01_SYS",
                "FEXCHANGERATETYPE__Name": "固定汇率"
            },
            "FEXCHANGERATE": 1,
            "FAMOUNTFOR": str(data['tse']),
            "FDEBIT": str(data['tse'])
        })
        
        # 出口退税分录（贷方）
        entries.append({
            "FEXPLANATION": f"{data['fphm']}应收退税款",
            "FACCOUNTID": {
                "FNumber": "2221.01.01.03",
                "FACCOUNT__NAME": "出口退税"
            },
            "FCURRENCYID": {
                "FNumber": "PRE001",
                "FCURRENCYID__Name": "人民币"
            },
            "FEXCHANGERATETYPE": {
                "FNumber": "HLTX01_SYS",
                "FEXCHANGERATETYPE__Name": "固定汇率"
            },
            "FEXCHANGERATE": 1,
            "FAMOUNTFOR": str(data['tse']),
            "FCREDIT": str(data['tse'])
        })
    
    # 4. 结转成本分录
    cost_amount = data['fpje'] - data['tse']
    entries.append({
        "FEXPLANATION": f"{data['fphm']}结转成本",
        "FACCOUNTID": {
            "FNumber": "6401.01.01",
            "FACCOUNT__NAME": "外销"
        },
        "FCURRENCYID": {
            "FNumber": "PRE001",
            "FCURRENCYID__Name": "人民币"
        },
        "FEXCHANGERATETYPE": {
            "FNumber": "HLTX01_SYS",
            "FEXCHANGERATETYPE__Name": "固定汇率"
        },
        "FEXCHANGERATE": 1,
        "FAMOUNTFOR": str(cost_amount),
        "FDEBIT": str(cost_amount)
    })
    
    # 库存商品减少（贷方）
    entries.append({
        "FEXPLANATION": f"{data['fphm']}结转成本",
        "FACCOUNTID": {
            "FNumber": "1405.01",
            "FACCOUNT__NAME": "外销"
        },
        "FCURRENCYID": {
            "FNumber": "PRE001",
            "FCURRENCYID__Name": "人民币"
        },
        "FEXCHANGERATETYPE": {
            "FNumber": "HLTX01_SYS",
            "FEXCHANGERATETYPE__Name": "固定汇率"
        },
        "FEXCHANGERATE": 1,
        "FAMOUNTFOR": str(cost_amount),
        "FCREDIT": str(cost_amount)
    })
    
    # 5. 进项税额转出（如果需要）
    if data['se'] != data['tse']:
        transfer_amount = data['se'] - data['tse']
        entries.append({
            "FEXPLANATION": f"{data['fphm']}结转成本",
            "FACCOUNTID": {
                "FNumber": "2221.01.01.02",
                "FACCOUNT__NAME": "进项税额转出"
            },
            "FCURRENCYID": {
                "FNumber": "PRE001",
                "FCURRENCYID__Name": "人民币"
            },
            "FEXCHANGERATETYPE": {
                "FNumber": "HLTX01_SYS",
                "FEXCHANGERATETYPE__Name": "固定汇率"
            },
            "FEXCHANGERATE": 1,
            "FAMOUNTFOR": str(transfer_amount),
            "FCREDIT": str(transfer_amount)
        })
    
    return entries


# 修正后的主处理函数
def process_export_data(s, hr: str, da1: str, da2: str, nbyw: str, yhm: str, passwd: str, wsttz: str):
    """主处理函数"""
    try:
        # 1. 更新状态
        update_gchk_status(s, da1, da2, wsttz)

        # 2. 获取位置信息判断 nbyw
        # position_val = get_position_info(s, 'some_id')
        # nbyw = '义乌' if '义乌' in position_val else '宁波'

        # 3. 获取服务器配置
        ip, ps =  '',''# get_server_config(s)

        # 4. 获取需要处理的发票列表
        invoice_list = get_invoice_list(s, da1, da2, wsttz)
        # 5. 逐个处理发票
        for r in invoice_list:
            # 构建凭证数据
            wxfp = r.get('wxfp')
            voucher_data = build_voucher_data(s, wxfp, hr, nbyw)
            if voucher_data:
                # 构建两个凭证 JSON
                voucher_main = build_voucher_json(voucher_data, is_main=True)
                voucher_detail = build_voucher_json(voucher_data, is_main=False)

                # 发送到 API
                # send_to_api(voucher_main, voucher_detail, yhm, passwd)

        return {'code': 1, 'msg': '操作成功', 'data': invoice_list}
    except Exception as e:
        logger.error(trace_error())
        return {'code': -1, 'msg': f'处理失败: {str(e)}'}


"""
宁波新金蝶入账New
对应原Pascal: 宁波新金蝶入账New
"""
@any_route('/api/saier/payment/voucher_api', methods=['POST'])
@require_token
async def view_payment_voucher_api(request):
    # 1. 检查用户是否是财务岗位 (对应 tmpcom2 查询)
    s = Session()
    try:
        j = await request.json()
        position = j.get('position', '')
        nbyw = '义乌' if '义乌' in position else '宁波'
        tjny = j.get('cxrq')
        ksrq = j.get('ksrq')
        jsrq = j.get('jsrq')
        yhm = j.get('yhm')
        passwd = j.get('passwd')
        wstt = j.get('wstt')
        if tjny == '' or tjny == None:
            tjny = time.strftime('%Y-%m')
        else:
            tjny = tjny[:7]
        if ksrq == '' or ksrq == None:
            ksrq = time.strftime('%Y-%m-01')
        if jsrq == '' or jsrq == None:
            jsrq = time.strftime('%Y-%m-%d')
        if ksrq > jsrq:
            return json_result(-1, '开始日期不能大于结束日期')
        if wstt == '' or wstt == None:
            return json_result(-1, '请选择我司抬头')
        res = process_export_data(
            s,
            hr=tjny,
            da1=ksrq,
            da2=jsrq,
            nbyw=nbyw,
            yhm='username',
            passwd='password',
            wsttz=wstt
        )
        if res.get('code') != 1:
            return json_result(-1, res.get('msg', '导出失败'))
        data = res.get('data')

        # s.commit()  # 如果需要提交事务
        return json_result(1, '操作成功', data)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()


def get_sjchny_by_wxfp(s, wxfp: str, rts: str) -> str:
    """根据发票号获取 sjchny"""
    query = s.query(gchk.sjchny).filter(gchk.wxfp == wxfp)
    if rts == '1':
        query = query.filter(gchk.tsl > 0)
    else:
        query = query.filter(gchk.tsl == 0)
    result = query.first()
    return result[0] if result else ''


def get_summary_by_wxfp_2(s, wxfp: str, rts: str) -> list:
    """按发票号分组汇总"""
    query = s.query(
        gchk.gcmc,
        func.sum(gchk.fpje).label('fpje1'),
        func.sum(gchk.se).label('se1'),
        func.sum(gchk.tse).label('tse1'),
        func.sum(gchk.bhsj).label('bhsj12'),
        gchk.dlrq
    ).filter(
        and_(gchk.wxfp == wxfp, gchk.fpwk == '否')
    )
    
    if rts == '1':
        query = query.filter(gchk.tsl > 0)
    else:
        query = query.filter(gchk.tsl == 0)
    
    results = query.group_by(gchk.gcmc, gchk.dlrq).order_by(gchk.dlrq.desc()).all()
    
    return [{
        'gcmc': r[0],
        'fpje1': float(r[1]) if r[1] else 0,
        'se1': float(r[2]) if r[2] else 0,
        'tse1': float(r[3]) if r[3] else 0,
        'bhsj12': float(r[4]) if r[4] else 0
    } for r in results]


def get_bgmxdhbcp_sum(s, father: str, rts: str) -> dict:
    """获取 bgmxdhbcp 汇总"""
    query = s.query(
        func.sum(bgmxdhbcp.wxzj).label('bgbgzje'),
        func.sum(bgmxdhbcp.CNFyf).label('CNFyf')
    ).filter(bgmxdhbcp.pid == father)
    
    if rts == '1':
        query = query.filter(bgmxdhbcp.tsl > 0)
    else:
        query = query.filter(bgmxdhbcp.tsl == 0)
    
    result = query.first()
    return {
        'bgbgzje': float(result[0]) if result[0] else 0,
        'CNFyf': float(result[1]) if result[1] else 0
    }


def get_gchk_info_2(s, wxfp: str, rts: str) -> dict:
    """获取 gchk 信息（wstt, kpxh, sjchny）"""
    query = s.query(gchk.wstt, gchk.kpxh, gchk.sjchny).filter(
        and_(
            gchk.wxfp == wxfp,
            gchk.wstt != '',
            gchk.wstt.isnot(None),
            gchk.kpxh != '',
            gchk.kpxh.isnot(None)
        )
    )
    
    if rts == '1':
        query = query.filter(gchk.tse > 0)
    else:
        query = query.filter(gchk.tse == 0)
    
    result = query.first()
    if result:
        return {
            'wstt': result[0],
            'kpxh': result[1],
            'sjchny': result[2]
        }
    return {}


# ==================== Excel 操作类 ====================
class ExcelWriter:
    def __init__(self, loadfilename: str, savefilename: str):
        self.loadfilename = loadfilename
        self.savefilename = savefilename
        if os.path.exists(loadfilename):
            self.wb = load_workbook(loadfilename)
        else:
            self.wb = load_workbook(loadfilename)  # 需要创建模板
        
        self.ws = self.wb.active
        self.current_row = 0
    
    def set_cell(self, col: str, row: int, value, number_format=None):
        """设置单元格值"""
        cell = self.ws[f"{col}{row}"]
        cell.value = value
        if number_format:
            cell.number_format = number_format
    
    def get_row(self) -> int:
        """获取当前行号"""
        return self.current_row
    
    def set_row(self, row: int):
        """设置当前行号"""
        self.current_row = row
    
    def save(self):
        """保存文件"""
        self.wb.save(self.savefilename)


# ==================== 主业务逻辑 ====================
def process_export_to_excel(
    s,
    hr: str,           # 会计期间，如 "2026-01-01"
    rts: str,          # '1' 外销，其他内销
    nbyw: str,         # '义乌' 或 '宁波'
    fp_list: list,     # 发票号列表
    sb: str,           # 摘要后缀
    filename: str,     # Excel 文件路径
    yhm: str = '',     # 用户名（如需要）
    passwd: str = ''   # 密码（如需要）
):
    """
    主处理函数
    """
    new_name = ''
    # 保存 Excel
    fp = config.tmp_path
    fn = f"宁波新金蝶入账_{time.strftime('%Y%m%d%H%M%S')}.xlsx"
    savefile = os.path.join(fp, fn)
    # 初始化 Excel
    excel = ExcelWriter(filename, savefile)
    try:
        # 解析日期
        n1 = hr[:4]        # 年
        y = hr[5:7]        # 月
        r = hr[8:10]       # 日
        n = hr[2:4]        # 年后两位
        
        # 获取月份和最后一天
        y1 = str(int(y))
        r1 = get_month_last_day(int(n1), int(y))
        
        # 获取汇率
        hl = get_exchange_rate(s, n1, y, r1, nbyw)
        
        # 初始化行计数器
        iz = 0
        iz4 = 0
        
        # 遍历发票
        for fphm in fp_list:
            iz += 1
            # 获取凭证信息
            gchk_data = get_gchk_info_2(s, fphm, rts)
            if not gchk_data:
                continue
            # fphm = gchk_data['wxfp']  # 发票号码
            # 获取主表信息
            bgmxd_info = get_bgmxd_info(s, fphm)
            father = bgmxd_info['rid']
            RMBkh = bgmxd_info['RMBkh']
            hbdm = bgmxd_info['hbdm']
            
            # 获取 sjchny
            hr1 = get_sjchny_by_wxfp(s, fphm, rts)
            
            # 如果有 sjchny，更新汇率
            if hr1:
                hl = get_cwhshl_rate(s, hr1)
            
            # 解析 hr1
            if hr1:
                nn = hr1[2:4]      # 年后两位
                nn1 = hr1[:4]      # 年
                ny = hr1[5:7]      # 月
                nr = hr1[8:10]     # 日
            
            wstt = gchk_data['wstt']
            kpxh_raw = gchk_data['kpxh']
            sjchny = gchk_data['sjchny']
            
            # 更新汇率
            if sjchny:
                hl = get_cwhshl_rate(s, sjchny)
            
            # 获取账簿编码
            zpbm, zpbm1 = get_zpbm_by_mc(s, wstt)
            
            # 处理 kpxh
            kpxh_first = kpxh_raw[0] if kpxh_raw else ''
            kpxh_num = get_kpxh_number(kpxh_first)
            kpxh_suffix = kpxh_raw[1:11] if len(kpxh_raw) > 1 else ''
            
            kpxh1 = f"{kpxh_num}{kpxh_suffix}1"
            kpxh2 = f"{kpxh_num}{kpxh_suffix}2"
            
            # 获取汇总数据
            summary_list = get_summary_by_wxfp_2(s, fphm, rts)
            
            # ==================== 写入第一行（库存商品借方） ====================
            row_num = iz + 2
            excel.set_cell('A', row_num, kpxh1)
            excel.set_cell('B', row_num, zpbm)
            excel.set_cell('C', row_num, f"{wstt}账簿")
            excel.set_cell('D', row_num, hr)
            excel.set_cell('E', row_num, 'g1')
            excel.set_cell('F', row_num, '总账')
            excel.set_cell('G', row_num, 'PRE001')
            excel.set_cell('H', row_num, '记')
            excel.set_cell('I', row_num, 0)
            excel.set_cell('L', row_num, n1)
            excel.set_cell('M', row_num, zpbm1)
            excel.set_cell('N', row_num, wstt)
            excel.set_cell('O', row_num, y1)
            excel.set_cell('Q', row_num, f"{kpxh1}000{iz}")
            excel.set_cell('R', row_num, f"{fphm}货物入库{sb}")
            excel.set_cell('S', row_num, '1405.01')
            excel.set_cell('T', row_num, '外销')
            excel.set_cell('U', row_num, '外销')
            excel.set_cell('V', row_num, '库存商品_外销')
            excel.set_cell('BE', row_num, 'PRE001')
            excel.set_cell('BF', row_num, '人民币')
            excel.set_cell('BG', row_num, 'HLTX01_SYS')
            excel.set_cell('BH', row_num, '固定汇率')
            excel.set_cell('BI', row_num, 1)
            
            iz4 = row_num
            iz += 1
            
            # ==================== 写入第二行（进项税额借方） ====================
            row_num = iz + 2
            excel.set_cell('Q', row_num, f"{kpxh1}000{iz}")
            excel.set_cell('R', row_num, f"{fphm}货物入库{sb}")
            
            if rts == '1':
                excel.set_cell('S', row_num, '2221.01.01.01')
                excel.set_cell('V', row_num, '应交税费_应交增值税_外销_进项税额')
            else:
                excel.set_cell('S', row_num, '2221.01.02.01')
                excel.set_cell('V', row_num, '应交税费_应交增值税_内销_进项税额')
            
            excel.set_cell('T', row_num, '进项税额')
            excel.set_cell('U', row_num, '进项税额')
            excel.set_cell('BE', row_num, 'PRE001')
            excel.set_cell('BF', row_num, '人民币')
            excel.set_cell('BG', row_num, 'HLTX01_SYS')
            excel.set_cell('BH', row_num, '固定汇率')
            excel.set_cell('BI', row_num, 1)
            
            # 初始化累计变量
            fpje = 0
            se = 0
            tse = 0
            bhsje = 0
            
            # ==================== 写入应付账款分录（贷方） ====================
            for item in summary_list:
                iz += 1
                ws = format_index(iz)
                row_num = iz + 2
                
                excel.set_cell('Q', row_num, f"{kpxh1}{ws}{iz}")
                excel.set_cell('R', row_num, f"{fphm}货物入库{sb}")
                excel.set_cell('S', row_num, '2202.01')
                excel.set_cell('T', row_num, '明细应付款')
                excel.set_cell('U', row_num, '明细应付款')
                excel.set_cell('V', row_num, '应付账款_明细应付款')
                
                # 辅助核算
                if item['gcmc']:
                    jdbm, company_name = get_company_jdbm(s, item['gcmc'])
                    if jdbm:
                        excel.set_cell('AS', row_num, jdbm)
                    
                    if len(item['gcmc'].strip()) > 8:
                        excel.set_cell('AT', row_num, item['gcmc'])
                    else:
                        excel.set_cell('AT', row_num, f"个体户{item['gcmc']}")
                
                excel.set_cell('BE', row_num, 'PRE001')
                excel.set_cell('BF', row_num, '人民币')
                excel.set_cell('BG', row_num, 'HLTX01_SYS')
                excel.set_cell('BH', row_num, '固定汇率')
                excel.set_cell('BI', row_num, 1)
                
                fpje_amount = round_two_decimals(item['fpje1'])
                excel.set_cell('BN', row_num, fpje_amount)
                excel.set_cell('BP', row_num, fpje_amount)
                
                fpje += fpje_amount
                
                if rts == '1':
                    se += round_two_decimals(item['se1'])
                    tse += round_two_decimals(item['tse1'])
                else:
                    se += round_two_decimals(item['se1'])
                    bhsje += round_two_decimals(item['bhsj12'])
            
            # 写入 BN/BO 汇总
            if rts == '1':
                excel.set_cell('BN', 3, fpje - se)
                excel.set_cell('BO', 3, fpje - se)
                excel.set_cell('BN', 4, se)
                excel.set_cell('BO', 4, se)
            else:
                excel.set_cell('BN', iz4, bhsje)
                excel.set_cell('BO', iz4, bhsje)
                excel.set_cell('BN', iz4 + 1, se)
                excel.set_cell('BO', iz4 + 1, se)
            
            iz += 1
            
            # ==================== 写入第二张凭证（收入/成本） ====================
            row_num = iz + 2
            excel.set_cell('A', row_num, kpxh2)
            excel.set_cell('B', row_num, zpbm)
            excel.set_cell('C', row_num, f"{wstt}账簿")
            excel.set_cell('D', row_num, hr)
            excel.set_cell('E', row_num, 'g1')
            excel.set_cell('F', row_num, '总账')
            excel.set_cell('G', row_num, 'PRE001')
            excel.set_cell('H', row_num, '记')
            excel.set_cell('I', row_num, 0)
            excel.set_cell('L', row_num, n1)
            excel.set_cell('M', row_num, zpbm1)
            excel.set_cell('N', row_num, wstt)
            excel.set_cell('O', row_num, y1)
            
            ws = format_index(iz)
            excel.set_cell('Q', row_num, f"{kpxh2}{ws}{iz}")
            excel.set_cell('R', row_num, f"{fphm}外销收入{sb}")
            excel.set_cell('S', row_num, '1122.01')
            excel.set_cell('T', row_num, '外销')
            excel.set_cell('U', row_num, '外销')
            excel.set_cell('V', row_num, '应收账款_外销')
            excel.set_cell('BE', row_num, 'PRE001')
            excel.set_cell('BF', row_num, '人民币')
            excel.set_cell('BG', row_num, 'HLTX01_SYS')
            excel.set_cell('BH', row_num, '固定汇率')
            excel.set_cell('BI', row_num, 1)
            
            # 获取 bgmxdhbcp 汇总
            bgmxdhbcp = get_bgmxdhbcp_sum(s, father, rts)
            bgbgzje = bgmxdhbcp['bgbgzje']
            cnfyf = bgmxdhbcp['CNFyf']
            
            hl1 = hl
            ts1 = 0
            
            # 计算金额
            if RMBkh != '是':
                if hbdm not in ['USD$', 'USD']:
                    hl1 = get_cwhshl_rate_by_date(s, f"{nn1}-{ny}", hbdm)
                    if hl1 == 0:
                        hl1 = hl
                
                if rts == '1':
                    # 应收账款金额
                    excel.set_cell('BN', row_num, round_two_decimals(hl1 * bgbgzje))
                    excel.set_cell('BO', row_num, round_two_decimals(hl1 * bgbgzje))
                    
                    # 收入金额（扣除运费）
                    next_row = row_num + 1
                    excel.set_cell('BN', next_row, 
                                round_two_decimals(hl1 * bgbgzje) - round_two_decimals(hl1 * cnfyf))
                    excel.set_cell('BP', next_row,
                                round_two_decimals(hl1 * bgbgzje) - round_two_decimals(hl1 * cnfyf))
                    
                    # 运费收入
                    if cnfyf > 0 or wstt == '宁波景业国际贸易有限公司':
                        next_row2 = row_num + 2
                        excel.set_cell('BN', next_row2, round_two_decimals(hl1 * cnfyf))
                        excel.set_cell('BP', next_row2, round_two_decimals(hl1 * cnfyf))
                else:
                    ts1 = round_two_decimals(((bgbgzje - cnfyf) * hl1) / 1.13 * 0.13)
                    excel.set_cell('BN', row_num, round_two_decimals(hl1 * bgbgzje))
                    excel.set_cell('BO', row_num, round_two_decimals(hl1 * bgbgzje))
                    
                    next_row = row_num + 1
                    excel.set_cell('BN', next_row, round_two_decimals(hl1 * bgbgzje) - ts1)
                    excel.set_cell('BP', next_row, round_two_decimals(hl1 * bgbgzje) - ts1)
                    
                    if cnfyf > 0 or wstt == '宁波景业国际贸易有限公司':
                        next_row2 = row_num + 2
                        excel.set_cell('BN', next_row2, round_two_decimals(hl1 * cnfyf))
                        excel.set_cell('BP', next_row2, round_two_decimals(hl1 * cnfyf))
            else:
                # RMBkh = '是' 的情况
                if rts == '1':
                    excel.set_cell('BN', row_num, bgbgzje)
                    excel.set_cell('BO', row_num, bgbgzje)
                    
                    next_row = row_num + 1
                    excel.set_cell('BN', next_row, bgbgzje - round_two_decimals(hl * cnfyf))
                    excel.set_cell('BP', next_row, bgbgzje - round_two_decimals(hl * cnfyf))
                    
                    if cnfyf > 0 or wstt == '宁波景业国际贸易有限公司':
                        next_row2 = row_num + 2
                        excel.set_cell('BN', next_row2, round_two_decimals(hl * cnfyf))
                        excel.set_cell('BP', next_row2, round_two_decimals(hl * cnfyf))
                else:
                    ts1 = round_two_decimals((bgbgzje - cnfyf * hl1) / 1.13 * 0.13)
                    excel.set_cell('BN', row_num, round_two_decimals(bgbgzje))
                    excel.set_cell('BO', row_num, round_two_decimals(bgbgzje))
                    
                    next_row = row_num + 1
                    excel.set_cell('BN', next_row, round_two_decimals(bgbgzje) - ts1)
                    excel.set_cell('BP', next_row, round_two_decimals(bgbgzje) - ts1)
                    
                    if cnfyf > 0 or wstt == '宁波景业国际贸易有限公司':
                        next_row2 = row_num + 2
                        excel.set_cell('BN', next_row2, round_two_decimals(hl * cnfyf))
                        excel.set_cell('BP', next_row2, round_two_decimals(hl * cnfyf))
            
            # ==================== 主营业务收入分录 ====================
            iz += 1
            row_num = iz + 2
            ws = format_index(iz)
            excel.set_cell('Q', row_num, f"{kpxh2}{ws}{iz}")
            excel.set_cell('R', row_num, f"{fphm}外销收入{sb}")
            excel.set_cell('S', row_num, '6001.01.01')
            excel.set_cell('T', row_num, '外销')
            excel.set_cell('U', row_num, '外销')
            excel.set_cell('V', row_num, '主营业务收入_商品销售收入_外销')
            excel.set_cell('BE', row_num, 'PRE001')
            excel.set_cell('BF', row_num, '人民币')
            excel.set_cell('BG', row_num, 'HLTX01_SYS')
            excel.set_cell('BH', row_num, '固定汇率')
            excel.set_cell('BI', row_num, 1)
            
            # 运费收入分录
            if cnfyf > 0 or wstt == '宁波景业国际贸易有限公司':
                iz += 1
                row_num = iz + 2
                ws = format_index(iz)
                excel.set_cell('Q', row_num, f"{kpxh2}{ws}{iz}")
                excel.set_cell('R', row_num, f"{fphm}外销收入{sb}")
                excel.set_cell('S', row_num, '6001.01.03')
                excel.set_cell('T', row_num, '外销')
                excel.set_cell('U', row_num, '外销')
                excel.set_cell('V', row_num, '主营业务收入_商品销售收入_CIF运费')
                excel.set_cell('BE', row_num, 'PRE001')
                excel.set_cell('BF', row_num, '人民币')
                excel.set_cell('BG', row_num, 'HLTX01_SYS')
                excel.set_cell('BH', row_num, '固定汇率')
                excel.set_cell('BI', row_num, 1)
            
            # 应收退税款/销项税额分录
            if rts == '1':
                # 应收退税款（借方）
                iz += 1
                row_num = iz + 2
                ws = format_index(iz)
                excel.set_cell('Q', row_num, f"{kpxh2}{ws}{iz}")
                excel.set_cell('R', row_num, f"{fphm}应收退税款{sb}")
                excel.set_cell('S', row_num, '1812.01')
                excel.set_cell('T', row_num, '应收出口退税')
                excel.set_cell('U', row_num, '应收出口退税')
                excel.set_cell('V', row_num, '其他流动资产_应收出口退税')
                excel.set_cell('BE', row_num, 'PRE001')
                excel.set_cell('BF', row_num, '人民币')
                excel.set_cell('BG', row_num, 'HLTX01_SYS')
                excel.set_cell('BH', row_num, '固定汇率')
                excel.set_cell('BI', row_num, 1)
                excel.set_cell('BN', row_num, tse)
                excel.set_cell('BO', row_num, tse)
            
            # 出口退税/销项税额分录（贷方）
            iz += 1
            row_num = iz + 2
            ws = format_index(iz)
            excel.set_cell('Q', row_num, f"{kpxh2}{ws}{iz}")
            excel.set_cell('R', row_num, f"{fphm}应收退税款{sb}")
            
            if rts == '1':
                excel.set_cell('S', row_num, '2221.01.01.03')
                excel.set_cell('T', row_num, '出口退税')
                excel.set_cell('U', row_num, '出口退税')
                excel.set_cell('V', row_num, '应交税费_应交增值税_外销_出口退税')
                excel.set_cell('BN', row_num, tse)
                excel.set_cell('BP', row_num, tse)
            else:
                excel.set_cell('S', row_num, '2221.01.02.03')
                excel.set_cell('T', row_num, '销项税额')
                excel.set_cell('U', row_num, '销项税额')
                excel.set_cell('V', row_num, '应交税费_应交增值税_内销_销项税额')
                excel.set_cell('BN', row_num, ts1)
                excel.set_cell('BP', row_num, ts1)
            
            excel.set_cell('BE', row_num, 'PRE001')
            excel.set_cell('BF', row_num, '人民币')
            excel.set_cell('BG', row_num, 'HLTX01_SYS')
            excel.set_cell('BH', row_num, '固定汇率')
            excel.set_cell('BI', row_num, 1)
            
            # 结转成本分录（借方 - 主营业务成本）
            iz += 1
            row_num = iz + 2
            ws = format_index(iz)
            excel.set_cell('Q', row_num, f"{kpxh2}{ws}{iz}")
            excel.set_cell('R', row_num, f"{fphm}结转成本{sb}")
            excel.set_cell('S', row_num, '6401.01.01')
            excel.set_cell('T', row_num, '外销')
            excel.set_cell('U', row_num, '外销')
            excel.set_cell('V', row_num, '主营业务成本_商品销售成本_外销')
            excel.set_cell('BE', row_num, 'PRE001')
            excel.set_cell('BF', row_num, '人民币')
            excel.set_cell('BG', row_num, 'HLTX01_SYS')
            excel.set_cell('BH', row_num, '固定汇率')
            excel.set_cell('BI', row_num, 1)
            
            if rts == '1':
                excel.set_cell('BN', row_num, fpje - tse)
                excel.set_cell('BO', row_num, fpje - tse)
            else:
                excel.set_cell('BN', row_num, bhsje)
                excel.set_cell('BO', row_num, bhsje)
            
            # 结转成本分录（贷方 - 库存商品减少）
            iz += 1
            row_num = iz + 2
            ws = format_index(iz)
            excel.set_cell('Q', row_num, f"{kpxh2}{ws}{iz}")
            excel.set_cell('R', row_num, f"{fphm}结转成本{sb}")
            excel.set_cell('S', row_num, '1405.01')
            excel.set_cell('T', row_num, '外销')
            excel.set_cell('U', row_num, '外销')
            excel.set_cell('V', row_num, '库存商品_外销')
            excel.set_cell('BE', row_num, 'PRE001')
            excel.set_cell('BF', row_num, '人民币')
            excel.set_cell('BG', row_num, 'HLTX01_SYS')
            excel.set_cell('BH', row_num, '固定汇率')
            excel.set_cell('BI', row_num, 1)
            
            if rts == '1':
                excel.set_cell('BN', row_num, fpje - se)
                excel.set_cell('BP', row_num, fpje - se)
            else:
                excel.set_cell('BN', row_num, bhsje)
                excel.set_cell('BP', row_num, bhsje)
            
            # 进项税额转出分录（如果需要）
            if rts == '1' and tse != se:
                iz += 1
                row_num = iz + 2
                ws = format_index(iz)
                excel.set_cell('Q', row_num, f"{kpxh2}{ws}{iz}")
                excel.set_cell('R', row_num, f"{fphm}结转成本{sb}")
                excel.set_cell('S', row_num, '2221.01.01.02')
                excel.set_cell('T', row_num, '进项税额转出')
                excel.set_cell('U', row_num, '进项税额转出')
                excel.set_cell('V', row_num, '应交税费_应交增值税_外销_进项税额转出')
                excel.set_cell('BE', row_num, 'PRE001')
                excel.set_cell('BF', row_num, '人民币')
                excel.set_cell('BG', row_num, 'HLTX01_SYS')
                excel.set_cell('BH', row_num, '固定汇率')
                excel.set_cell('BI', row_num, 1)
                excel.set_cell('BN', row_num, se - tse)
                excel.set_cell('BP', row_num, se - tse)
        

        excel.save()
        if os.path.exists(os.path.join(fp, fn)):
            new_name = fn

        return {"code": 1, "msg": "处理成功", "data": new_name}
    except Exception as e:
        logger.error(trace_error())
        return {"code": -1, "msg": f"处理失败: {str(e)}"}


"""
宁波新金蝶入账
对应原Pascal: 宁波新金蝶入账
"""
@any_route('/api/saier/payment/voucher_excel', methods=['POST'])
@require_token
async def view_payment_voucher_excel(request):
    # 1. 检查用户是否是财务岗位 (对应 tmpcom2 查询)
    s = Session()
    try:
        j = await request.json()
        position = j.get('position', '')
        # fphm = j.get('fphm', '')
        rids = j.get('rids', [])
        nbyw = '义乌' if '义乌' in position else '宁波'
        tjny = j.get('cxrq')
        kind = j.get('kind')
        temp_dir = config.data_upload_path
        temp_file = os.path.join(temp_dir, f"template/新金蝶导入模板优景.xlsx")
        rts = '1' if kind == '否' else '2'
        sb = '(零退税)' if  rts == '2' else ''
        # ED24UV1246-5714670
        fp_list = []
        for rid in rids:
            c = s.query(gchk).filter(gchk.rid == rid).first()  # 确保 rid 存在
            if c and c.wxfp != None and c.wxfp != '' and c.wxfp not in fp_list:  # 确保发票号码存在
                c.sfrz1 = '是'  # 标记已处理
                s.add(c)
                fp_list.append(c.wxfp)

        logger.error(f"发票列表: {fp_list}")
        logger.error(f"会计期间: {tjny}, 内外销: {rts}, 业务类型: {nbyw}")
        res = process_export_to_excel(
            s,
            hr=tjny,
            rts=rts,                    # '1' 外销，其他内销
            nbyw=nbyw,                 # '义乌' 或 '宁波'
            fp_list=fp_list,
            sb=sb,                      # 摘要后缀
            filename=temp_file,
            yhm='username',
            passwd='password'
        )
        if res.get('code') != 1:
            s.rollback()
            return json_result(-1, res.get('msg', '导出失败'))
        data = res.get('data')

        s.commit()  # 提交事务
        return json_result(0, '操作成功', data)
    except Exception as e:
        logger.error(trace_error())
        s.rollback()
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()


"""
批量更新开票资料已到
对应原Pascal: 批量更新开票资料已到
"""
@any_route('/api/saier/payment/batch/update/invoice_flag', methods=['POST'])
@require_token
async def view_payment_batch_update_invoice_flag(request):
    s = Session()
    try:
        j = await request.json()
        user = request.current_user
        rids = j.get('rids', [])
        if len(rids) == 0:
            return json_result(-1, '未选中记录')
        org = get_user_path(user.username)
        fp_json = {}
        for rid in rids:
            c = s.query(gchk).filter(gchk.rid == rid).first()
            if not c:
                continue
            c.kpzlyd = '是'
            s.add(c)
            # s.query(kaiptz).filter(kaiptz.rid == rid).update({kaiptz.kpzlyd: '是'})
            # s.query(zlbinvoice).filter(zlbinvoice.rid == rid).update({zlbinvoice.kpzlyd: '是'})


        s.commit()
        return json_result(1, '操作成功')
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')
    finally:
        s.close()