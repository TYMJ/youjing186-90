# 批量采购合同（图）
from any import *
import os
import re
import json
import base64
import hashlib
import shutil
import subprocess
from collections import defaultdict
from datetime import datetime, date

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

ILLEGAL_CHARS = set('/\\:*?"<>|\r\n')

from openpyxl import load_workbook as _openpyxl_load_workbook

def load_workbook(*args, **kwargs):
    kwargs.pop('encoding', None)
    return _openpyxl_load_workbook(*args, **kwargs)
def _safe_filename(text):
    s = str(text or '').strip()
    if not s:
        return 'EMPTY'
    s = ''.join('_' if ch in ILLEGAL_CHARS else ch for ch in s)
    return s[:120]


def _to_upload_rel_path(abs_path):
    abs_path = os.path.abspath(abs_path)
    base_upload = getattr(config, 'data_upload_path', '') or ''
    if base_upload:
        base_abs = os.path.abspath(base_upload)
        if abs_path.startswith(base_abs):
            rel = abs_path[len(base_abs):].lstrip('/\\')
            return rel.replace('\\', '/')
    return abs_path.replace('\\', '/')


def _abs_upload(path):
    base = getattr(config, 'data_upload_path', '') or ''
    return os.path.join(base, str(path or '').lstrip('/\\'))


def _parse_json_list(v):
    if isinstance(v, list):
        return v
    if v is None:
        return []
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return []
        try:
            x = json.loads(s)
            if isinstance(x, list):
                return x
        except Exception:
            pass
        return [p.strip() for p in s.split(',') if p.strip()]
    return []


def _to_float(v):
    try:
        return float(v or 0)
    except Exception:
        return 0.0


def _first_row(sql, params):
    rows = run_sql(sql, params)
    return rows[0] if rows else {}


def _convert_excel_to_pdf(excel_path):
    pdf_path = os.path.splitext(excel_path)[0] + '.pdf'

    if not getattr(config, 'java_path', ''):
        return {'success': False, 'error': '未配置 config.java_path'}
    if not getattr(config, 'report_jar', ''):
        return {'success': False, 'error': '未配置 config.report_jar'}
    if not os.path.exists(config.report_jar):
        return {'success': False, 'error': f'JAR文件不存在: {config.report_jar}'}

    cmd = [config.java_path, '-jar', config.report_jar, 'a', 'b', excel_path, pdf_path, '2']
    try:
        rs = subprocess.run(cmd, capture_output=True, text=True, timeout=180)
        if rs.returncode != 0:
            return {'success': False, 'error': rs.stderr or f'returncode={rs.returncode}'}
        if not os.path.exists(pdf_path):
            return {'success': False, 'error': 'PDF未生成'}
        return {'success': True, 'pdf_path': pdf_path}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def _normalize_company(company_raw, gs):
    company = str(company_raw or '').strip()
    if company:
        return company

    gs = str(gs or '').strip()
    if gs in ('', '1'):
        return '宁波优景进出口有限公司'
    if gs == '2':
        return '宁波锐亿进出口有限公司'
    return gs


def _resolve_pdf_sign_policy(pdf_input, has_batch_pdf, has_sign_auth):
    """
    Pascal 逻辑归一化：
    - 有“采购合同图PDF”权限 -> 最终强制批量PDF(3)，且签章开启
    - 无该权限时：用户选2/3 -> 最终批量PDF(3)；若有“采购合同图PDF签名”权限则签章开启
    - 其他 -> Excel(1)
    """
    pdf = str(pdf_input or '1').strip()
    if has_batch_pdf:
        return '3', True
    if pdf in ('2', '3'):
        return '3', bool(has_sign_auth)
    return '1', False


def _resolve_template_path(company, is_tedi):
    def _norm_name(v):
        n = os.path.basename(str(v or '')).strip().lower().replace('\\', '/')
        n = re.sub(r'\s+', '', n)
        return n

    suffix = 'tedi' if is_tedi else ''
    base = f'{company}{suffix}'
    data_upload = getattr(config, 'data_upload_path', '') or ''
    base_dir = os.path.dirname(__file__)

    names = [
        f'{base}_image.xlsx',
        f'{base}image.xlsx',
        f'{base}图.xlsx',
        f'{base}.xlsx',
        '宁波优景进出口有限公司_image.xlsx',
        '宁波锐亿进出口有限公司_image.xlsx',
    ]

    target_full = {_norm_name(x) for x in names}
    target_stem = {os.path.splitext(x)[0] for x in target_full}

    # 1) 先从 sys_attachment 里按 name 做匹配（你截图里那批结果就在这里）
    try:
        rows = run_sql(
            """
            SELECT rid, name, path
            FROM sys_attachment
            WHERE module=:module
              AND IFNULL(path,'')<>''
              AND LOWER(IFNULL(name,'')) LIKE '%.xlsx'
            ORDER BY rid DESC
            """,
            {'module': '采购跟单'}
        )
    except Exception:
        rows = []

    for r in rows:
        raw_name = str(r.get('name') or '').strip()
        raw_path = str(r.get('path') or '').strip()
        if not raw_name or not raw_path:
            continue

        norm_name = _norm_name(raw_name)
        stem = os.path.splitext(norm_name)[0]

        # 避免误用 BOOKING 模板
        if 'booking' in stem:
            continue

        matched = (norm_name in target_full) or any((ts in stem) or (stem in ts) for ts in target_stem)
        if not matched:
            continue

        abs_path = raw_path if os.path.isabs(raw_path) else _abs_upload(raw_path)
        if os.path.exists(abs_path):
            return os.path.abspath(abs_path)

    # 2) 数据库没命中，再走本地目录兜底
    candidates = []
    for fname in names:
        candidates.extend([
            os.path.join(data_upload, 'contract_templates', fname),
            os.path.join(data_upload, 'addonfiles', fname),
            os.path.join(base_dir, '..', 'frontend', 'static', 'contract_templates', fname),
            os.path.join(base_dir, '..', 'frontend', 'static', fname),
            os.path.join(base_dir, fname),
        ])

    for p in candidates:
        ap = os.path.abspath(p)
        if os.path.exists(ap):
            return ap

    return ''

def _get_sign_image_path(cpbh='陈妍科小'):
    rows = run_sql(
        """
        SELECT path
        FROM sys_attachment
        WHERE pid IN (
            SELECT rid FROM tpzx WHERE cpbh=:cpbh
        )
        AND path IS NOT NULL
        LIMIT 1
        """,
        {'cpbh': str(cpbh or '')}
    )
    if not rows:
        return ''
    return _abs_upload(rows[0].get('path'))


def _try_add_image(ws, image_abs_path, anchor_cell, max_width=105, max_height=78):
    try:
        if not image_abs_path or not os.path.exists(image_abs_path):
            return

        img = XLImage(image_abs_path)
        src_w = max(float(img.width or 1), 1.0)
        src_h = max(float(img.height or 1), 1.0)
        scale = min(max_width / src_w, max_height / src_h)
        if scale <= 0:
            scale = 1.0

        img.width = int(src_w * scale)
        img.height = int(src_h * scale)
        img.anchor = anchor_cell

        m = re.search(r'(\d+)$', str(anchor_cell or ''))
        if m:
            row_no = int(m.group(1))
            expected_height = max(65.0, img.height * 0.75 + 6)
            current_height = float(ws.row_dimensions[row_no].height or 0)
            if expected_height > current_height:
                ws.row_dimensions[row_no].height = expected_height

        ws.add_image(img)
    except Exception:
        pass


def _is_probably_base64(s):
    if not s:
        return False
    t = s.strip()
    if len(t) < 100:
        return False
    return re.fullmatch(r'[A-Za-z0-9+/=\s]+', t) is not None


def _save_base64_image(raw_b64, cache_dir, cache_key):
    try:
        os.makedirs(cache_dir, exist_ok=True)
        b = base64.b64decode(raw_b64)
        if not b:
            return ''

        ext = '.jpg'
        if b.startswith(b'\x89PNG'):
            ext = '.png'
        elif b.startswith(b'GIF8'):
            ext = '.gif'
        elif b.startswith(b'RIFF') and b[8:12] == b'WEBP':
            ext = '.webp'

        fp = os.path.join(cache_dir, f'{cache_key}{ext}')
        with open(fp, 'wb') as wf:
            wf.write(b)
        return fp
    except Exception:
        return ''


def _extract_first_image_src(v):
    if v is None:
        return ''

    if isinstance(v, (bytes, bytearray)):
        try:
            v = v.decode('utf-8', errors='ignore')
        except Exception:
            return ''

    if isinstance(v, dict):
        for key in ('src', 'path', 'url', 'name'):
            x = v.get(key)
            if x:
                return str(x)
        return ''

    if isinstance(v, list):
        for item in v:
            got = _extract_first_image_src(item)
            if got:
                return got
        return ''

    s = str(v).strip()
    if not s or s in ('[]', '{}', 'null', 'None'):
        return ''

    if s.startswith('[') or s.startswith('{'):
        for candidate in (s, s.replace("'", '"')):
            try:
                obj = json.loads(candidate)
                got = _extract_first_image_src(obj)
                if got:
                    return got
            except Exception:
                pass

    return s


def _resolve_image_from_raw(raw_value, cache_dir, cache_ns):
    source = _extract_first_image_src(raw_value)
    if not source:
        return ''

    s = str(source).strip()
    if not s:
        return ''

    if s.startswith('data:image'):
        parts = s.split(',', 1)
        if len(parts) == 2:
            key = hashlib.md5((cache_ns + s[:80]).encode('utf-8', errors='ignore')).hexdigest()
            return _save_base64_image(parts[1], cache_dir, key)

    if _is_probably_base64(s):
        key = hashlib.md5((cache_ns + s[:80]).encode('utf-8', errors='ignore')).hexdigest()
        fp = _save_base64_image(s, cache_dir, key)
        if fp:
            return fp

    s = s.replace('\\', '/').strip()
    if os.path.isabs(s) and os.path.exists(s):
        return s

    p = _abs_upload(s)
    if os.path.exists(p):
        return p

    return ''


def _get_line_image_path(line, use_split_source, image_cache_dir, line_image_cache):
    wyzd = str(line.get('wyzd') or '').strip()
    hhbz = str(line.get('hhbz') or '').strip()
    cache_key = ('split' if use_split_source else 'normal') + '|' + wyzd + '|' + hhbz

    if cache_key in line_image_cache:
        return line_image_cache[cache_key]

    raw = ''
    if not use_split_source and hhbz:
        row = _first_row(
            """
            SELECT yytp
            FROM lscp
            WHERE lshh=:lshh AND IFNULL(yytp,'')<>''
            ORDER BY sid DESC
            LIMIT 1
            """,
            {'lshh': hhbz}
        )
        raw = row.get('yytp')

    if (not raw) and use_split_source and wyzd:
        row = _first_row(
            """
            SELECT yytp
            FROM cghtsheet
            WHERE wyzd=:wyzd AND IFNULL(yytp,'')<>''
            LIMIT 1
            """,
            {'wyzd': wyzd}
        )
        raw = row.get('yytp')

    if not raw:
        raw = line.get('yytp')

    img_path = _resolve_image_from_raw(raw, image_cache_dir, cache_key)
    line_image_cache[cache_key] = img_path
    return img_path


def _collect_cggd_rows(rid_list, user_name):
    rows = []
    for rid in rid_list:
        rid_s = str(rid or '').strip()
        if not rid_s:
            continue


        rs = run_sql(
            """
            SELECT rid number, hthm, gdry, sccj, cgry, cs_id, wxht, ywry
            FROM cggd
            WHERE rid=:rid 
            """,
            {'rid': rid_s}
        )

        for r in rs:
            gdry = str(r.get('gdry') or '').strip()
            if user_name and gdry and gdry != user_name:
                continue
            rows.append(r)

    return rows


def _check_customer_flags(wxht_list, user_name):
    is_tedi = False

    for order_id in wxht_list:
        if not order_id:
            continue
        wx = _first_row(
            """
            SELECT customer
            FROM wxht
            WHERE order_id=:order_id
            LIMIT 1
            """,
            {'order_id': str(order_id)}
        )
        customer = str(wx.get('customer') or '').strip()
        if customer == 'TEDi GmbH & Co. KG':
            is_tedi = True

    has_batch_pdf = bool(_first_row(
        """
        SELECT rid
        FROM cyzglsheet
        WHERE xm=:xm AND zm='采购合同图PDF'
        LIMIT 1
        """,
        {'xm': user_name}
    ))

    has_sign_auth = bool(_first_row(
        """
        SELECT rid
        FROM cyzglsheet
        WHERE xm=:xm AND zm='采购合同图PDF签名'
        LIMIT 1
        """,
        {'xm': user_name}
    ))

    return {
        'is_tedi': is_tedi,
        'has_batch_pdf': has_batch_pdf,
        'has_sign_auth': has_sign_auth,
    }


def _get_contract_lines(hthm, factory, buyer):
    run_sql("UPDATE cght SET webpd1='是' WHERE hthm=:hthm", {'hthm': hthm})
    return run_sql(
        """
        SELECT rid, pid, hthm, bjhh, khhh, zwpm, wxrl,
               cgxs, cgsl, cgjg, zje, yjcq, jhrq, sfhs,
               wxht, cpgg, zhwbzh, zwdw, cpsm, hyd, zhwbgpm, zzsl,
               sccj1, cgry, wyzd, hhbz, qtsm1, krddh, yytp, ksdhh, zycpbh,
               zmz, zjz, ztj
        FROM cghtsheet
        WHERE hthm=:hthm AND sccj1=:sccj1 AND cgry=:cgry
        ORDER BY jhrq, rid
        """,
        {'hthm': hthm, 'sccj1': factory, 'cgry': buyer}
    )


def _get_dforder(order_id):
    row = _first_row(
        """
        SELECT dforder_id
        FROM wxht
        WHERE order_id=:order_id
        LIMIT 1
        """,
        {'order_id': str(order_id or '')}
    )
    return str(row.get('dforder_id') or '')


def _calc_totals(lines):
    zxs = sum(_to_float(r.get('cgxs')) for r in lines)
    zsl = sum(_to_float(r.get('cgsl')) for r in lines)
    zjg = sum(_to_float(r.get('zje')) for r in lines)
    zmz = sum(_to_float(r.get('zmz')) for r in lines)
    zjz = sum(_to_float(r.get('zjz')) for r in lines)
    ztj = sum(_to_float(r.get('ztj')) for r in lines)

    sfhs1 = '是'
    for r in lines:
        if str(r.get('sfhs') or '是') != '是':
            sfhs1 = '否'
            break
    return zxs, zsl, zjg, sfhs1, zmz, zjz, ztj


def _build_context(first_line, buyer, company):
    hthm = str(first_line.get('hthm') or '')

    cght_row = _first_row(
        """
        SELECT szdq, gdry, cgry, ywry, sccj1id, lxry, gcdh, sjhm, gccz,
               htrq, jsfs, sffl, bzyq
        FROM cght
        WHERE hthm=:hthm AND cgry=:cgry
        LIMIT 1
        """,
        {'hthm': hthm, 'cgry': buyer}
    )

    szdq = str(cght_row.get('szdq') or '')
    supplier_id = str(cght_row.get('sccj1id') or '')

    htjj = _first_row(
        """
        SELECT dz, lxfs, cgzj, zjl, hcyx, lxdh, gstt
        FROM htjj
        WHERE dq=:dq AND gstt LIKE :gstt
        LIMIT 1
        """,
        {'dq': szdq, 'gstt': f'%{company}%'}
    )

    supplier = _first_row(
        """
        SELECT cslxr, phone, fax, sjhm, address, twhm, fkhm, bank1, zh1
        FROM zycs
        WHERE cs_id=:cs_id
        LIMIT 1
        """,
        {'cs_id': supplier_id}
    )

    def user_profile(yhm):
        if not yhm:
            return {}
        return _first_row(
            """
            SELECT ryxm, lxdh, ydhm, bmjl, jldh
            FROM ywrybiao
            WHERE yhm=:yhm
            LIMIT 1
            """,
            {'yhm': yhm}
        )

    gdry_yhm = str(cght_row.get('gdry') or '')
    cgry_yhm = str(cght_row.get('cgry') or '')
    ywry_yhm = str(cght_row.get('ywry') or '')

    return {
        'htrq': str(cght_row.get('htrq') or ''),
        'szdq': szdq,
        'jsfs': str(cght_row.get('jsfs') or ''),
        'sffl': str(cght_row.get('sffl') or ''),
        'bzyq': str(cght_row.get('bzyq') or ''),
        'supplier_id': supplier_id,
        'gdry_yhm': gdry_yhm,
        'cgry_yhm': cgry_yhm,
        'ywry_yhm': ywry_yhm,
        'htjj': htjj,
        'supplier': supplier,
        'gdry': user_profile(gdry_yhm),
        'cgry': user_profile(cgry_yhm),
        'ywry': user_profile(ywry_yhm),
    }


def _build_line_note(row):
    parts = []
    cpsm = str(row.get('cpsm') or '').strip()
    qtsm1 = str(row.get('qtsm1') or '').strip()
    hyd = str(row.get('hyd') or '').strip()
    krddh = str(row.get('krddh') or '').strip()

    if cpsm:
        parts.append(cpsm)
    if qtsm1:
        parts.append(qtsm1)
    if hyd:
        parts.append(f'请注意开票货源地为：{hyd}')
    if krddh:
        parts.append(f'客人订单号为：{krddh}')

    return '\n'.join(parts)


def _fill_detail_sheet_with_images(
    ws,
    lines,
    company,
    cpbh_show,
    sign_path,
    use_split_source,
    image_cache_dir,
    line_image_cache,
):
    zxs, zsl, zjg, _, _, _, _ = _calc_totals(lines)
    is_jingchi = ('景驰' in (company or ''))

    ws['A1'] = f'合同编号：{cpbh_show}项下附页合同具体内容如下：'

    for idx, row in enumerate(lines, start=1):
        excel_row = 2 + idx
        cphh = str(row.get('khhh') or row.get('ksdhh') or row.get('bjhh') or row.get('zycpbh') or '')
        if not cphh:
            cphh = str(row.get('bjhh') or '')

        ws[f'A{excel_row}'] = idx
        ws[f'B{excel_row}'] = row.get('hthm', '')
        ws[f'C{excel_row}'] = cphh
        ws[f'D{excel_row}'] = row.get('zwpm', '')
        ws[f'E{excel_row}'] = row.get('wxrl', '')
        ws[f'F{excel_row}'] = row.get('cgxs', '')
        ws[f'G{excel_row}'] = row.get('cgsl', '')
        ws[f'H{excel_row}'] = row.get('cgjg', '')
        ws[f'I{excel_row}'] = row.get('zje', '')
        ws[f'J{excel_row}'] = row.get('jhrq', '')
        ws[f'K{excel_row}'] = row.get('yjcq', '')
        ws[f'L{excel_row}'] = _build_line_note(row)

        img_path = _get_line_image_path(row, use_split_source, image_cache_dir, line_image_cache)
        if img_path:
            _try_add_image(ws, img_path, f'M{excel_row}', max_width=105, max_height=78)

        if is_jingchi:
            ws[f'N{excel_row}'] = _get_dforder(row.get('wxht'))

    sum_row = 3 + len(lines)
    ws[f'B{sum_row}'] = '合计'
    ws[f'F{sum_row}'] = zxs
    ws[f'G{sum_row}'] = zsl
    ws[f'I{sum_row}'] = zjg

    ws[f'B{sum_row + 3}'] = '需 方（签字盖章）'
    ws[f'F{sum_row + 3}'] = '供 方（签字盖章）'
    if sign_path:
        _try_add_image(ws, sign_path, f'B{sum_row + 2}', max_width=90, max_height=58)


def _fill_main_sheet(ws, first_line, totals, line_count, ctx, sign_path, anti_corruption_text, contract_notice):
    zxs, _, zjg, sfhs1, _, _, _ = totals
    htjj = ctx.get('htjj', {})
    supplier = ctx.get('supplier', {})

    gd = ctx.get('gdry', {})
    cg = ctx.get('cgry', {})
    yw = ctx.get('ywry', {})

    ws['D10'] = ctx.get('htrq', '')
    ws['I10'] = ctx.get('szdq', '')

    item_code = first_line.get('khhh') or first_line.get('bjhh')
    ws['N10'] = item_code
    ws['U10'] = item_code

    ws['C11'] = first_line.get('sccj1', '')
    ws['L11'] = supplier.get('cslxr', '')
    ws['C12'] = supplier.get('phone', '')
    ws['L12'] = supplier.get('fax', '')

    ws['B15'] = item_code
    ws['D15'] = f"{first_line.get('zwpm', '')}\n{first_line.get('cpgg', '')}"
    ws['J15'] = first_line.get('wxrl', '')
    ws['K15'] = first_line.get('zwdw', '')
    ws['L15'] = zxs
    ws['M15'] = first_line.get('cgsl', '')
    ws['O15'] = first_line.get('cgjg', '')
    ws['P15'] = first_line.get('zje', '')

    ws['F16'] = first_line.get('cpsm', '')
    ws['N16'] = f"请注意开票货源地为：{first_line.get('hyd', '')}\n本合同为附页子合同的总合同，共计{line_count}个子合同"

    bzyq = str(ctx.get('bzyq') or '').strip()
    if bzyq:
        ws['B24'] = '四、外箱包装要求：' + bzyq

    ws['B30'] = '七、结算方式：' + str(ctx.get('jsfs') or '')

    if htjj.get('dz'):
        ws['B7'] = htjj.get('dz')
    if htjj.get('lxfs'):
        ws['B8'] = htjj.get('lxfs')

    if sfhs1 != '是':
        ws['B19'] = f"收款户名:{supplier.get('fkhm', '')}  开户银行:{supplier.get('bank1', '')}  银行账号:{supplier.get('zh1', '')}"

    ws['F17'] = f'金额合计:{zjg:.2f}'
    ws['P17'] = zjg

    if first_line.get('zhwbgpm') and str(first_line.get('zhwbgpm')) != '无':
        ws['O18'] = f"开票品名: {first_line.get('zhwbgpm', '')}{first_line.get('zzsl', '')}%"

    ws['D25'] = f'{zxs}箱'

    ws['B48'] = f"跟单人员: {gd.get('ryxm', '')}({gd.get('lxdh', '')}/{gd.get('ydhm', '')})"
    ws['J48'] = f"采购员: {cg.get('ryxm', '')}({cg.get('lxdh', '')}/{cg.get('ydhm', '')})"
    ws['B49'] = f"外销负责: {yw.get('ryxm', '')}({yw.get('lxdh', '')}/{yw.get('ydhm', '')})"
    ws['J49'] = htjj.get('cgzj', '')
    ws['B50'] = htjj.get('zjl', '')
    ws['J50'] = anti_corruption_text

    if contract_notice:
        ws['A52'] = contract_notice

    if sign_path:
        _try_add_image(ws, sign_path, 'B47', max_width=90, max_height=58)


def _get_credit_threshold():
    row = _first_row(
        "SELECT cs,sz1 FROM zx WHERE mc=:mc LIMIT 1",
        {'mc': '诚信金额'}
    )
    single_limit = _to_float(row.get('cs'))
    yearly_limit = _to_float(row.get('sz1'))
    if yearly_limit <= 0:
        yearly_limit = 1000000.0
    return single_limit, yearly_limit


def _get_year_contract_total(supplier_id):
    if not supplier_id:
        return 0.0
    qsrq = f"{datetime.now().year}-01-01"
    row = _first_row(
        """
        SELECT SUM(htje) AS htje1
        FROM cght
        WHERE htrq>=:htrq AND sccj1id=:sccj1id AND sfhz='通过'
        """,
        {'htrq': qsrq, 'sccj1id': supplier_id}
    )
    return _to_float(row.get('htje1'))


def _check_credit(factory):
    warns = []
    row = _first_row(
        """
        SELECT qrrq
        FROM cxgc
        WHERE (gcmc=:gcmc) OR (chgc=:chgc)
        LIMIT 1
        """,
        {'gcmc': factory, 'chgc': factory}
    )

    if not row:
        warns.append(f'工厂名称:{factory}需提交诚信报告')
        return warns

    qrrq = str(row.get('qrrq') or '').strip()
    if not qrrq:
        warns.append(f'工厂名称:{factory}需提交诚信报告')
        return warns

    try:
        qd = datetime.strptime(qrrq[:10], '%Y-%m-%d').date()
    except Exception:
        warns.append(f'工厂名称:{factory}需提交诚信报告')
        return warns

    days = (date.today() - qd).days
    if days > 365:
        warns.append(f'工厂名称:{factory}需提交诚信报告')
    elif days > 330:
        warns.append(f'工厂名称:{factory}还有一个月需提交诚信报告')

    return warns


def _sync_contract_header(hthm, totals, sfsh):
    zxs, zsl, zje, _, zmz, zjz, ztj = totals
    run_sql(
        """
        UPDATE cght
        SET webpd1='是',
            htje=:htje,
            htzsl=:htzsl,
            htzxs=:htzxs,
            htzmz=:htzmz,
            htzjz=:htzjz,
            htztj=:htztj,
            sfsh=:sfsh
        WHERE hthm=:hthm
        """,
        {
            'hthm': hthm,
            'htje': zje,
            'htzsl': zsl,
            'htzxs': zxs,
            'htzmz': zmz,
            'htzjz': zjz,
            'htztj': ztj,
            'sfsh': sfsh
        }
    )


@any_route('/api/Ravencloud/generate_contract_batch_image', methods=['POST'])
async def api_generate_contract_batch_image(request):
    image_cache_dir = ''
    try:
        form = await request.form()

        rid_list = _parse_json_list(form.get('rid_list') or form.get('rids') or '[]')
        gs = str(form.get('gs') or '').strip()
        company = _normalize_company(form.get('company') or '', gs)
        pdf_input = str(form.get('pdf') or '1').strip()

        if not rid_list:
            return json_result(code=-1, msg='请先选择要操作的记录')

        current_user = ''
        try:
            current_user = str(request.current_user.name or '').strip()
        except Exception:
            pass

        cggd_rows = _collect_cggd_rows(rid_list, current_user)
        if not cggd_rows:
            return json_result(code=-1, msg='未找到可处理数据，或当前用户无权限')

        wxht_list = []
        for r in cggd_rows:
            wxht = str(r.get('wxht') or '').strip()
            if wxht and wxht not in wxht_list:
                wxht_list.append(wxht)

        flags = _check_customer_flags(wxht_list, current_user)
        pdf_mode, sign_enabled = _resolve_pdf_sign_policy(
            pdf_input,
            flags.get('has_batch_pdf', False),
            flags.get('has_sign_auth', False)
        )

        template_path = _resolve_template_path(company, flags.get('is_tedi', False))
        if not template_path:
            return json_result(code=-1, msg=f'未找到图合同模板，请上传: {company}_image.xlsx 或 {company}.xlsx')

        anti_corruption_text = str(
            _first_row("SELECT wb1 FROM zx WHERE ly=:ly LIMIT 1", {'ly': '反腐专线'}).get('wb1') or ''
        )
        contract_notice = str(
            _first_row("SELECT nr FROM zx WHERE ly=:ly LIMIT 1", {'ly': '采购合同签订注意要点'}).get('nr') or ''
        )

        sign_path = _get_sign_image_path('陈妍科小') if sign_enabled else ''

        output_root = os.path.join(config.get_today_upload_path(), '采购合同图')
        os.makedirs(output_root, exist_ok=True)

        image_cache_dir = os.path.join(output_root, '.tmp_contract_images')
        os.makedirs(image_cache_dir, exist_ok=True)

        thresholds = _get_credit_threshold()
        files = []
        warnings = []
        line_image_cache = {}

        targets = []
        seen = set()
        for r in cggd_rows:
            hthm = str(r.get('hthm') or '').strip()
            factory = str(r.get('sccj') or '').strip()
            buyer = str(r.get('cgry') or '').strip()
            if not hthm or not factory or not buyer:
                continue
            key = (hthm, factory, buyer)
            if key not in seen:
                seen.add(key)
                targets.append(key)

        if not targets:
            return json_result(code=-1, msg='无可处理合同，请检查选中记录')

        for hthm, factory, buyer in targets:
            lines = _get_contract_lines(hthm, factory, buyer)
            if not lines:
                continue

            totals_all = _calc_totals(lines)
            ctx_all = _build_context(lines[0], buyer, company)
            supplier_id = str(ctx_all.get('supplier_id') or '')
            year_total = _get_year_contract_total(supplier_id)

            sfsh = '不需要'
            if totals_all[2] > thresholds[0] or year_total > thresholds[1]:
                credit_warns = _check_credit(factory)
                warnings.extend(credit_warns)
                sfsh = '待提供' if any('需提交诚信报告' in w for w in credit_warns) else '已提供'

            _sync_contract_header(hthm, totals_all, sfsh)

            grouped = defaultdict(list)
            for row in lines:
                key_item = str(row.get('bjhh') or row.get('khhh') or row.get('ksdhh') or 'UNKNOWN').strip() or 'UNKNOWN'
                grouped[key_item].append(row)

            factory_dir = os.path.join(output_root, _safe_filename(factory))
            os.makedirs(factory_dir, exist_ok=True)

            for _, item_lines in grouped.items():
                first_line = item_lines[0]
                cpbh_show = str(first_line.get('khhh') or first_line.get('bjhh') or first_line.get('ksdhh') or '')

                wb = load_workbook(template_path)
                ws_main = wb.worksheets[0]
                ws_detail = wb.worksheets[1] if len(wb.worksheets) > 1 else wb.worksheets[0]

                totals = _calc_totals(item_lines)
                ctx = _build_context(first_line, buyer, company)
                use_split_source = str(ctx.get('sffl') or '') == '是'

                _fill_detail_sheet_with_images(
                    ws_detail,
                    item_lines,
                    company,
                    cpbh_show,
                    sign_path,
                    use_split_source,
                    image_cache_dir,
                    line_image_cache,
                )
                _fill_main_sheet(
                    ws_main,
                    first_line,
                    totals,
                    len(item_lines),
                    ctx,
                    sign_path,
                    anti_corruption_text,
                    contract_notice
                )

                date_tag = datetime.now().strftime('%Y%m%d')
                ywry_yhm = str(ctx.get('ywry_yhm') or '')
                base_name = f"{_safe_filename(ywry_yhm)}-{_safe_filename(factory)}-{_safe_filename(cpbh_show[:80])}-{date_tag}"
                xlsx_path = os.path.join(factory_dir, base_name + '.xlsx')

                wb.save(xlsx_path)
                wb.close()

                if pdf_mode == '1':
                    files.append(_to_upload_rel_path(xlsx_path))
                else:
                    pdf_res = _convert_excel_to_pdf(xlsx_path)
                    if pdf_res.get('success'):
                        files.append(_to_upload_rel_path(pdf_res.get('pdf_path')))
                    else:
                        warnings.append(f'{base_name} 转PDF失败: {pdf_res.get("error")}')
                    try:
                        os.remove(xlsx_path)
                    except Exception:
                        pass

        warning_file = ''
        if warnings:
            warning_name = datetime.now().strftime('%Y-%m-%d') + '.txt'
            warning_abs = os.path.join(output_root, warning_name)
            with open(warning_abs, 'w', encoding='utf-8') as fp:
                fp.write('\n'.join(dict.fromkeys(warnings)))
            warning_file = _to_upload_rel_path(warning_abs)

        if not files:
            return json_result(code=-1, msg='未生成任何文件，请检查数据/模板', data={
                'warnings': list(dict.fromkeys(warnings)),
                'warning_file': warning_file
            })

        return json_result(code=1, msg=f'成功生成{len(files)}个文件', data={
            'files': files,
            'count': len(files),
            'warnings': list(dict.fromkeys(warnings)),
            'warning_file': warning_file,
            'pdf_mode': pdf_mode,
            'notice': contract_notice
        })

    except Exception:
        logger.error(trace_error())
        return json_result(code=-1, msg=trace_error())
    finally:
        if image_cache_dir and os.path.isdir(image_cache_dir):
            try:
                shutil.rmtree(image_cache_dir, ignore_errors=True)
            except Exception:
                pass