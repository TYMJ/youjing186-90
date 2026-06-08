from types import ClassMethodDescriptorType
from webbrowser import get

from any import *
from .model import *
from sqlalchemy.sql import func,not_,or_,and_
import time
from datetime import datetime
import json,os,time
from .__default__ import get_user_path
from pypinyin import pinyin, Style
from openpyxl import load_workbook
from openpyxl import Workbook

# 新财务工厂资料审核确认按钮  
@any_route('/api/saier/new_supplier/confirm/btn', methods=['POST', 'GET'])
@require_token
async def view_saier_new_supplier_confirm_btn(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        org = get_user_path(user.username)
        # if not '财务' in org.get('path',''):
        #     return json_result(-1, '不好意思,您没有权限执行此操作',org)
        
        rids = j.get('rids', [])
        cwqr = j.get('cwqr', '')
        for rid in rids:
            d = s.query(newcwcs).filter(newcwcs.rid==rid).first()
            if not d:
                return json_result(-1, '没有找到相关记录')
            if d.cwqr != cwqr:
                d.cwqr = cwqr
                d.qrr1 = user.username
                d.modi_uid = user.rid
                d.mtime = time.strftime('%Y-%m-%d %H:%M:%S')
                s.add(d)

        s.commit()
        return json_result(1, '操作成功')
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


def get_first_letter_pinyin(text):
    """
    获取字符串第一个字符的拼音首字母
    """
    if not text:
        return 'T'  # 如果字符串为空，返回默认值
    
    # 获取第一个字符的拼音首字母
    first_char = text[0]
    result = pinyin(first_char, style=Style.FIRST_LETTER)
    
    if result and result[0]:
        return result[0][0].upper()  # 转为大写
    
    return 'T'  # 如果无法获取拼音首字母，返回默认值

@any_route('/api/saier/new_supplier/save/before/check', methods=['POST', 'GET'])
@require_token
async def view_saier_new_supplier_save_before_check(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        csmc = j.get('csmc', '')
        csbh = j.get('csbh', '')
        xydm = j.get('xydm', '')
        rid = j.get('rid', '')
        yhzh = j.get('yhzh', '')
        yhmc = j.get('yhmc', '')
        csdm = ''
        bhxh = 1
        
        if csbh == '' or csbh == None:
            csdm = get_first_letter_pinyin(csmc)
            d = run_sql(f"select cs_id from newcwcs where (cs_id like '{csdm}%') and (cs_id<>'zk') and (cs_id<>'1zk') order by cs_id desc limit 1")
            if len(d) > 0:
                bhxh = d[0].get('cs_id')[1:] if len(d[0].get('cs_id'))>1 else 0
                bhxh = int(bhxh) + 1
            csdm = csdm + "{:05d}".format(bhxh)
        if xydm == '' or xydm == None:
            d = run_sql(f"select sid from newcwcs order by sid desc limit 1")
            if len(d) > 0:
                xydm = str(user.username) + str(time.strftime('%Y')[2:4]) + str(time.strftime('%m%d')) + str(d[0].get('sid'))

        if csmc!= '' and csmc != None:
            d = run_sql(f"select rid from newcwcs where (company_name='{csmc}' and rid<>'{rid}') limit 1")
            if len(d) > 0:
                if csmc == '1zk':
                    return json_result(-1, '厂商名称重复,请核对,这个工厂在金蝶里手工建过 麻烦 将相关信息 发李辉 包括金蝶编码（要金蝶里查的）')
                else:
                    return json_result(-1, '厂商名称重复,请核对')
        if yhzh!= '' and yhzh != None:
            d = run_sql(f"select rid from newcwcs where (bank='{yhmc}' and zh='{yhzh}' and rid<>'{rid}') limit 1")
            if len(d) > 0:
                return json_result(-1, '银行账户重复,请核对')
        if xydm!= '' and xydm != None:
            d = run_sql(f"select rid from newcwcs where (shui='{xydm}' and rid<>'{rid}') limit 1")
        if len(d) > 0:
              return json_result(-1, '社会统一信用代码重复,请核对')

        return json_result(1, '操作成功', {'csdm': csdm, 'xydm': xydm})
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


@any_route('/api/finance/import_factory_excel', methods=['POST'])
@require_token
async def api_import_factory_excel(request):
    """
    导入财务工厂资料Excel
    对应原Pascal: 财务工厂资料导入功能
    """
    try:
        user = request.current_user
        
        # 1. 检查用户是否是财务岗位 (对应 tmpcom2 查询)
        s = Session()
        try:
            org = get_user_path(user.username)
            position = org.get('position')
            if '财务' not in position:
                return json_result(-1, '只有财务岗位用户才能执行此操作')
            
            # 2. 获取上传的Excel文件
            j = await request.form()
            file = form_value(j,'file',None)
            if is_none(file):
                return json_result(ERR_PARAM_NOT_ENOUGH)
            # 保存临时文件
            temp_dir = config.tmp_path
            temp_file = os.path.join(temp_dir, f"{uuid.uuid4()}_{file.filename}")
            logger.error(f"Saving uploaded file to: {temp_file}")
            data = await file.read()
            write_file(temp_file, data, 'wb', encoding=None)
            
            # 3. 读取Excel文件
            wb = load_workbook(temp_file)
            ws = wb.worksheets[0]  # 读取第一个工作表
            
            # 4. 处理每一行数据
            duplicates = []  # 存储重复记录信息
            row_idx = 2  # 从第2行开始（第1行是表头）
            
            while True:
                # 读取Excel单元格数据
                mc = str(ws.cell(row=row_idx, column=1).value or '').strip()  # A列 - 公司名称
                bank = str(ws.cell(row=row_idx, column=2).value or '')  # B列 - 银行
                zh = str(ws.cell(row=row_idx, column=3).value or '')  # C列 - 账号
                shui = str(ws.cell(row=row_idx, column=4).value or '')  # D列 - 税号
                
                # 如果公司名称为空，结束循环
                if not mc:
                    break
                
                # 5. 生成cs_id编号 (对应 tmpcom 查询 newcwcs)
                # 获取拼音首字母
                s_initial = get_pinyin_initial(mc)
                
                d = s.query(newcwcs.cs_id).filter(newcwcs.cs_id.like(f'{s_initial}%'), newcwcs.cs_id != 'zk', newcwcs.cs_id != '1zk').order_by(newcwcs.cs_id.desc()).first()
                if d:
                    max_id = d.cs_id
                    # 提取数字部分并加1
                    num_part = max_id[len(s_initial):]
                    try:
                        next_num = int(num_part) + 1
                    except:
                        next_num = 1
                else:
                    next_num = 1
                
                # 格式化编号 (5位数字，不足补0)
                cs_id = f"{s_initial}{next_num:05d}"
                
                # 6. 查询yhbm表获取银行信息 (对应 tmpcom 查询 yhbm)
                yhdm = yhdz = wyjgh = lhh = jgh = sjyh = sjlhh = yhmc = sjhjgh = city = province = ''
                
                sql_yhbm = """
                    SELECT yhdm, yhdz, wyjgh, lhh, jgh, sjyh, sjlhh, yhmc, sjhjgh, city, province 
                    FROM yhbm 
                    WHERE yhmc = :yhmc
                """
                result = s.execute(text(sql_yhbm), {'yhmc': bank})
                row_yhbm = result.fetchone()
                
                if row_yhbm:
                    yhdm = row_yhbm.yhdm or ''
                    yhdz = row_yhbm.yhdz or ''
                    wyjgh = row_yhbm.wyjgh or ''
                    lhh = row_yhbm.lhh or ''
                    jgh = row_yhbm.jgh or ''
                    sjyh = row_yhbm.sjyh or ''
                    sjlhh = row_yhbm.sjlhh or ''
                    yhmc = row_yhbm.yhmc or ''
                    sjhjgh = row_yhbm.sjhjgh or ''
                    city = row_yhbm.city or ''
                    province = row_yhbm.province or ''
                
                # 7. 检查newcwcs表是否已存在 (对应 tmpcom1 查询)
                sql_check_dup = """
                    SELECT * FROM newcwcs 
                    WHERE company_name = :company_name OR shui = :shui
                """
                result = s.execute(text(sql_check_dup), {
                    'company_name': mc,
                    'shui': shui
                })
                row_dup = result.fetchone()
                
                now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                owner = user.rid
                rid = get_uuid()
                if row_dup is None:
                    # 8. 不存在则INSERT
                    sql_insert = """
                        INSERT INTO newcwcs (
                            ctime, uid, cs_id, company_name, rid,
                            sjhm, cslxr, phone, fax, bank, zh, shui, 
                            province, city, address, 
                            xgry, yhdm, yhdz, wyjgh, lhh, jgh, sjyh, sjlhh, sjhjgh, webpd
                        ) VALUES (
                            :ctime, :uid, :cs_id, :company_name, :rid,
                            :sjhm, :cslxr, :phone, :fax, :bank, :zh, :shui,
                            :province, :city, :address,
                            :xgry, :yhdm, :yhdz, :wyjgh, :lhh, :jgh, :sjyh, :sjlhh, :sjhjgh, :webpd
                        )
                    """
                    s.execute(text(sql_insert), {
                        'ctime': now_str,
                        'uid': owner,
                        'rid': rid,
                        'cs_id': cs_id,
                        'company_name': mc,
                        'sjhm': '',
                        'cslxr': '',
                        'phone': '',
                        'fax': '',
                        'bank': bank,
                        'zh': zh,
                        'shui': shui,
                        'province': province,
                        'city': city,
                        'address': '',
                        'xgry': 'zjnblh',
                        'yhdm': yhdm,
                        'yhdz': yhdz,
                        'wyjgh': wyjgh,
                        'lhh': lhh,
                        'jgh': jgh,
                        'sjyh': sjyh,
                        'sjlhh': sjlhh,
                        'sjhjgh': sjhjgh,
                        'webpd': '是'
                    })
                else:
                    # 9. 存在则UPDATE (保留原有值如果新值为空)
                    sql_update = """
                        UPDATE newcwcs SET
                            mtime = :mtime,
                            modi_uid = :modi_uid,
                            webpd = :webpd,
                            city = CASE WHEN :city_new != '' THEN :city_new ELSE city END,
                            province = CASE WHEN :province_new != '' THEN :province_new ELSE province END,
                            zh = :zh,
                            shui = :shui,
                            bank = :bank,
                            yhdm = CASE WHEN :yhdm_new != '' THEN :yhdm_new ELSE yhdm END,
                            yhdz = CASE WHEN :yhdz_new != '' THEN :yhdz_new ELSE yhdz END,
                            wyjgh = CASE WHEN :wyjgh_new != '' THEN :wyjgh_new ELSE wyjgh END,
                            lhh = CASE WHEN :lhh_new != '' THEN :lhh_new ELSE lhh END,
                            jgh = CASE WHEN :jgh_new != '' THEN :jgh_new ELSE jgh END,
                            sjyh = CASE WHEN :sjyh_new != '' THEN :sjyh_new ELSE sjyh END,
                            sjlhh = CASE WHEN :sjlhh_new != '' THEN :sjlhh_new ELSE sjlhh END,
                            sjhjgh = CASE WHEN :sjhjgh_new != '' THEN :sjhjgh_new ELSE sjhjgh END
                        WHERE company_name = :company_name
                    """
                    s.execute(text(sql_update), {
                        'mtime': now_str,
                        'modi_uid': owner,
                        'webpd': '是',
                        'city_new': city,
                        'province_new': province,
                        'zh': zh,
                        'shui': shui,
                        'bank': bank,
                        'yhdm_new': yhdm,
                        'yhdz_new': yhdz,
                        'wyjgh_new': wyjgh,
                        'lhh_new': lhh,
                        'jgh_new': jgh,
                        'sjyh_new': sjyh,
                        'sjlhh_new': sjlhh,
                        'sjhjgh_new': sjhjgh,
                        'company_name': mc
                    })
                    
                    # 记录重复信息
                    duplicates.append(f'请注意工厂名: {mc} 或税号: {shui} 重复！')
                
                row_idx += 1
            
            # 提交事务
            s.commit()
            
            # 10. 如果有重复记录，生成文本文件
            result_data = {
                'has_duplicates': len(duplicates) > 0,
            }
            
            if duplicates:
                # 生成重复记录文件
                dup_filename = f"重复财务工厂_{datetime.now().strftime('%Y-%m-%d')}.txt"
                dup_filepath = os.path.join(temp_dir, dup_filename)
                # with open(dup_filepath, 'wb', encoding=None) as f:
                #     f.write('\n'.join(duplicates))
                val ='\n'.join(duplicates)
                logger.error(f"Writing duplicates to: {dup_filepath}")
                write_file(dup_filepath, val, 'w')

                # 这里应该上传文件到可下载的位置，返回URL
                # 简化处理：返回文件内容或保存到特定目录
                result_data['duplicate_file_name'] = dup_filename
            
            # 清理临时文件
            # os.remove(temp_file)
            
            return json_result(0, data=result_data)
            
        except Exception as e:
            s.rollback()
            logger.error(trace_error())
            return json_result(-1, f'导入失败: {str(e)}')
        finally:
            s.close()
            wb.close()
            
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'导入失败: {str(e)}')


def get_pinyin_initial(chinese_str):
    """
    获取中文字符串的拼音首字母
    例如：'阿里巴巴' -> 'A'
    """
    # 这里需要引入pypinyin库或类似工具
    # 简化实现：返回第一个字符的大写
    if not chinese_str:
        return 'X'
    
    first_char = chinese_str[0]
    # 如果是中文，需要转换为拼音首字母
    # 这里使用简单的映射或引入pypinyin
    try:
        from pypinyin import lazy_pinyin
        pinyin_list = lazy_pinyin(first_char)
        if pinyin_list:
            return pinyin_list[0][0].upper()
    except:
        pass
    
    #  fallback：返回X
    return 'X'
