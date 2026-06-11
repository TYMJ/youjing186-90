# 批量唛头生成
from any import *
import os
import re
import json
import subprocess
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage


ILLEGAL_CHARS = set('/\\:*?"<>|\r\n')

TMTP_ANCHOR = {
    "内外箱唛头": "I5",
    "常规唛头": "A4",
    "港加特标唛头": "A6",
    "港加特标加警语唛头": "A6",
    "多货号唛头": "A10",
}
GK_ANCHOR = {
    "内外箱唛头": "D6",
    "常规唛头": "E2",
    "港加特标唛头": "F8",
    "港加特标加警语唛头": "G6",
    "多货号唛头": "G10",
}
# 模板单元格图片区域参考尺寸（像素），对齐 Pascal AddPicture 缩放留白
ANCHOR_IMAGE_BOX = {
    "I5": (110, 85),
    "A4": (95, 75),
    "A6": (95, 75),
    "A10": (95, 75),
    "D6": (52, 48),
    "E2": (48, 42),
    "F8": (52, 48),
    "G6": (52, 48),
    "G10": (52, 48),
    "F3": (42, 38),
    "D2": (42, 38),
    "E29": (48, 42),
    "G45": (48, 42),
    "G43": (48, 42),
    "G47": (48, 42),
    "L24": (48, 42),
}
CHECKMARK_IMAGE_BOX = (20, 20)
PX_TO_EMU = 9525


def _safe_filename(text):
    s = str(text or "").strip()
    if not s:
        return "EMPTY"
    s = "".join("_" if ch in ILLEGAL_CHARS else ch for ch in s)
    return s[:120]


def _to_upload_rel_path(abs_path):
    abs_path = os.path.abspath(abs_path)
    base_upload = getattr(config, "data_upload_path", "") or ""
    if base_upload:
        base_abs = os.path.abspath(base_upload)
        if abs_path.startswith(base_abs):
            rel = abs_path[len(base_abs) :].lstrip("/\\")
            return rel.replace("\\", "/")
    return abs_path.replace("\\", "/")


def _convert_excel_to_pdf(excel_path):
    pdf_path = os.path.splitext(excel_path)[0] + ".pdf"

    if not getattr(config, "java_path", ""):
        return {"success": False, "error": "未配置 config.java_path"}
    if not getattr(config, "report_jar", ""):
        return {"success": False, "error": "未配置 config.report_jar"}
    if not os.path.exists(config.report_jar):
        return {"success": False, "error": f"JAR文件不存在: {config.report_jar}"}

    cmd = [config.java_path, "-jar", config.report_jar, "a", "b", excel_path, pdf_path, "2"]
    try:
        rs = subprocess.run(cmd, capture_output=True, text=True, timeout=180)
        if rs.returncode != 0:
            return {"success": False, "error": rs.stderr or f"returncode={rs.returncode}"}
        if not os.path.exists(pdf_path):
            return {"success": False, "error": "PDF未生成"}
        return {"success": True, "pdf_path": pdf_path}
    except Exception as e:
        return {"success": False, "error": str(e)}


def _split_multi_cpbh(text):
    src = str(text or "").strip()
    if not src:
        return []
    parts = re.split(r"[,\，]+", src)
    return [x.strip() for x in parts if x.strip()]


def _get_template_path(mark_type, env_opt, region_opt):
    # env_opt: 1环保, 2不加, 3PP, 4LDEP-1, 5LDPE-2
    # region_opt: 1BP, 2拉脱维亚
    # mark_type: 内外箱唛头/常规唛头/港加特标唛头/港加特标加警语唛头/多货号唛头
    # 按 Pascal 命名映射成模板文件名（你可继续补齐实际文件）
    # base = getattr(config, 'data_upload_path', '')
    # tpl_dir = os.path.join(base, 'template')
    tpl_dir = os.path.join(config.data_upload_path, "template")
    is_lv = str(region_opt) == "2"
    # prefix = "拉港" if is_lv else ""

    if str(env_opt) == "1":
        suffix = "环保"
    elif str(env_opt) == "2":
        suffix = ""
    elif str(env_opt) in ("3", "4", "5"):
        suffix = "环保图"
    else:
        suffix = "环保"

    # mapping = {
    #     "内外箱唛头": f"{prefix}双联{suffix}.xlsx",
    #     "常规唛头": f"半港{suffix}.xlsx",
    #     "港加特标唛头": f"{prefix}加标{suffix}.xlsx",
    #     "港加特标加警语唛头": f"{prefix}加标加警{suffix}.xlsx",
    #     "多货号唛头": f"{prefix}多货号{suffix}.xlsx",
    # }
    # 按地区拆分，避免 "拉港" + "港加标" = "拉港港加标" 的重复问题
    if is_lv:
        mapping = {
            "内外箱唛头": f"拉港双联{suffix}.xlsx",
            "常规唛头": f"半港{suffix}.xlsx",
            "港加特标唛头": f"拉港加标{suffix}.xlsx",
            "港加特标加警语唛头": f"拉港加标加警{suffix}.xlsx",
            "多货号唛头": f"拉港多货号{suffix}.xlsx",
        }
    else:
        mapping = {
            "内外箱唛头": f"双联{suffix}.xlsx",
            "常规唛头": f"半港{suffix}.xlsx",
            "港加特标唛头": f"港加标{suffix}.xlsx",
            "港加特标加警语唛头": f"港加标加警{suffix}.xlsx",
            "多货号唛头": f"多货号{suffix}.xlsx",
        }

    name = mapping.get(mark_type, "")
    if not name:
        return ""
    return os.path.join(tpl_dir, name)


def _get_selected_order_numbers(number_list):
    # number_list 里是 cggd.rid 列表
    nums = []
    for rid in number_list:
        rows = run_sql(
            """
            SELECT rid number, wxht
            FROM cggd
            WHERE rid=:rid
        """,
            {"rid": str(rid)},
        )
        for r in rows:
            no = str(r.get("number") or "").strip()
            wxht = str(r.get("wxht") or "").strip()
            if no and no not in nums:
                nums.append({"number": no, "wxht": wxht})
    return nums


def _get_order_dforder(order_id):
    rows = run_sql(
        """
        SELECT dforder_id, order_id
        FROM wxht
        WHERE order_id=:order_id
        LIMIT 1
    """,
        {"order_id": str(order_id)},
    )
    if not rows:
        return str(order_id or "")
    r = rows[0]
    return str(r.get("dforder_id") or r.get("order_id") or order_id or "")


def _get_detail_rows(order_number):
    return run_sql(
        """
        SELECT cpbh, zxl, htzxs, nxrl
        FROM cggdsheet
        WHERE pid=:father
    """,
        {"father": str(order_number)},
    )


def _get_product_info(cpbh):
    rows = run_sql(
        """
        SELECT krtm,djpmw,tsbj,khlb,cxmc,jgby,tmtp,cpbh,jldw
        FROM zscp
        WHERE (cpbh=:cpbh) OR ((krhh=:cpbh1) AND (krhh<>'') AND (krhh IS NOT NULL))
        LIMIT 1
    """,
        {"cpbh": str(cpbh), "cpbh1": str(cpbh)},
    )
    return rows[0] if rows else {}


def _get_wypm_name(cpbh):
    rows = run_sql(
        """
        SELECT djpmw
        FROM zscp
        WHERE (cpbh=:cpbh) OR ((krhh=:cpbh1) AND (krhh<>'') AND (krhh IS NOT NULL))
        LIMIT 1
    """,
        {"cpbh": str(cpbh), "cpbh1": str(cpbh)},
    )
    if rows:
        return str(rows[0].get("djpmw") or "").strip()
    return ""


def _get_tpzx_image_path_by_sb(sb):
    rows = run_sql(
        """
        SELECT tpmc
        FROM tpzx
        WHERE sb=:sb AND tpmc IS NOT NULL AND LENGTH(tpmc) > 5
        LIMIT 1
        """,
        {"sb": str(sb or "")},
    )
    if rows:
        return _image_path(rows[0].get("tpmc"))
    return ""


def _get_tpzx_image_path_by_cpbh(cpbh):
    rows = run_sql(
        """
        SELECT tpmc
        FROM tpzx
        WHERE cpbh=:cpbh AND tpmc IS NOT NULL AND LENGTH(tpmc) > 5
        LIMIT 1
        """,
        {"cpbh": str(cpbh or "")},
    )
    if rows:
        return _image_path(rows[0].get("tpmc"))
    return ""


def _abs_upload(path):
    base = getattr(config, "data_upload_path", "") or ""
    return os.path.join(base, str(path or "").lstrip("/\\"))


def _normalize_image_path(path):
    fp = str(path or "").strip()
    if not fp:
        return ""
    if os.path.isfile(fp):
        return fp
    p1 = _abs_upload(fp)
    if os.path.isfile(p1):
        return p1
    rel = fp.lstrip("/\\").replace("\\", "/")
    if rel != fp:
        p2 = _abs_upload(rel)
        if os.path.isfile(p2):
            return p2
    return ""


def _image_path(raw):
    s = str(raw or "").strip()
    if len(s) <= 5:
        return ""
    if s.startswith("[") or s.startswith("{"):
        try:
            obj = json.loads(s.replace("'", '"'))
            src = obj[0].get("src", "") if isinstance(obj, list) and obj else obj.get("src", "") if isinstance(obj, dict) else ""
            return _normalize_image_path(src)
        except Exception:
            return ""
    return _normalize_image_path(s)


def _cell_box_px(ws, anchor_cell, pad=3):
    m = re.match(r"([A-Z]+)(\d+)", str(anchor_cell or "").upper())
    if not m:
        return 80, 60
    col_letter, row_num = m.group(1), int(m.group(2))
    col_dim = ws.column_dimensions.get(col_letter)
    col_w = col_dim.width if col_dim and col_dim.width else 8.43
    row_dim = ws.row_dimensions.get(row_num)
    row_h = row_dim.height if row_dim and row_dim.height else 15.0
    w = int(col_w * 7 + 5) - pad * 2
    h = int(row_h * 96 / 72) - pad * 2
    return max(10, w), max(10, h)


def _try_add_image(ws, image_abs_path, anchor_cell, max_width=None, max_height=None):
    path = _normalize_image_path(image_abs_path)
    if not path:
        return
    try:
        anchor = str(anchor_cell or "").upper()
        box = ANCHOR_IMAGE_BOX.get(anchor)
        if max_width is None and box:
            max_width, max_height = box
        if max_width is None or max_height is None:
            cw, ch = _cell_box_px(ws, anchor_cell)
            max_width = max_width or cw
            max_height = max_height or ch

        img = XLImage(path)
        src_w = max(float(img.width or 1), 1.0)
        src_h = max(float(img.height or 1), 1.0)
        scale = min(max_width / src_w, max_height / src_h)
        if scale > 0:
            img.width = max(1, int(src_w * scale))
            img.height = max(1, int(src_h * scale))

        m = re.match(r"([A-Z]+)(\d+)", anchor)
        if not m:
            img.anchor = anchor_cell
            ws.add_image(img)
            return

        col_str, row_num = m.group(1), int(m.group(2))
        col_idx = 0
        for ch in col_str:
            col_idx = col_idx * 26 + (ord(ch) - 64)
        col_idx -= 1
        row_idx = row_num - 1

        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
        from openpyxl.drawing.xdr import XDRPositiveSize2D

        col_off = int(max(0, (max_width - img.width) / 2) * PX_TO_EMU)
        row_off = int(max(0, (max_height - img.height) / 2) * PX_TO_EMU)
        marker = AnchorMarker(col=col_idx, row=row_idx, colOff=col_off, rowOff=row_off)
        img.anchor = OneCellAnchor(
            _from=marker,
            ext=XDRPositiveSize2D(cx=int(img.width * PX_TO_EMU), cy=int(img.height * PX_TO_EMU)),
        )
        ws.add_image(img)
    except Exception:
        pass


def _fill_by_mark_type(ws, mark_type, data, extra):
    cpbh = str(data.get("cpbh") or "")
    czkrhh = str(extra.get("czkrhh") or "")
    show_code = czkrhh if czkrhh else cpbh
    zxl = str(data.get("zxl") or "")
    htzxs = str(data.get("htzxs") or "")
    nxrl = str(data.get("nxrl") or "")
    krtm = str(extra.get("krtm") or "")
    khlb = str(extra.get("khlb") or "")
    cxmc = str(extra.get("cxmc") or "")
    jgby = str(extra.get("jgby") or "")
    jldw = str(extra.get("jldw") or "")
    ship_date = str(extra.get("ship_date") or "")
    khht = str(extra.get("khht") or "")
    djpmw = str(extra.get("djpmw") or "")

    if jldw.lower().endswith("s"):
        jldw = jldw[:-1]

    if mark_type == "内外箱唛头":
        ws["C1"] = show_code
        ws["R1"] = show_code
        ws["J2"] = show_code
        ws["S2"] = show_code
        ws["C14"] = f"{zxl} {jldw.upper()}S"
        ws["C24"] = f"OF   {htzxs}"
        ws["C26"] = khht
        ws["C21"] = khlb
        ws["R21"] = khlb
        ws["C29"] = cxmc
        ws["R29"] = cxmc
        ws["C31"] = ship_date
        ws["AB31"] = ship_date
        ws["J19"] = khlb
        ws["S19"] = khlb
        ws["J24"] = cxmc
        ws["S24"] = cxmc
        ws["A9"] = djpmw
        ws["P9"] = djpmw
        ws["I11"] = djpmw
        ws["Q11"] = djpmw
        ws["C12"] = krtm[:7]
        ws["K17"] = krtm[:7]
        ws["F8"] = f"{zxl}/{max(float(nxrl or 1), 1)} TPK" if str(nxrl or "").strip() else f"{zxl}/TPK"

    elif mark_type == "常规唛头":
        ws["B1"] = show_code
        ws["I1"] = show_code
        ws["A9"] = djpmw
        ws["J9"] = djpmw
        ws["C12"] = krtm[:7]
        ws["C14"] = f"{zxl} {jldw.upper()}S"
        ws["C21"] = khlb
        ws["K21"] = khlb
        ws["C24"] = f"OF   {htzxs}"
        ws["C26"] = khht
        ws["C29"] = cxmc
        ws["K29"] = cxmc
        ws["C31"] = ship_date

    elif mark_type == "港加特标唛头":
        ws["B2"] = show_code
        ws["K2"] = show_code
        ws["A12"] = djpmw
        ws["L12"] = djpmw
        ws["C28"] = krtm[:7]
        ws["C30"] = f"{zxl} {jldw.upper()}S"
        ws["B37"] = khlb
        ws["L37"] = khlb
        ws["B40"] = f"OF   {htzxs}"
        ws["B42"] = khht
        ws["B45"] = cxmc
        ws["L45"] = cxmc
        ws["B47"] = ship_date

    elif mark_type == "港加特标加警语唛头":
        ws["B2"] = show_code
        ws["K2"] = show_code
        ws["A12"] = djpmw
        ws["L12"] = djpmw
        ws["A17"] = jgby
        ws["L17"] = jgby
        ws["C28"] = krtm[:7]
        ws["C30"] = f"{zxl} {jldw.upper()}S"
        ws["C37"] = khlb
        ws["K37"] = khlb
        ws["C40"] = f"OF   {htzxs}"
        ws["C42"] = khht
        ws["C45"] = cxmc
        ws["K45"] = cxmc
        ws["C47"] = ship_date

    elif mark_type == "多货号唛头":
        codes = _split_multi_cpbh(show_code)
        ck = str(extra.get("checkmark_path") or "")
        row_start = 2 if len(codes) < 15 else 1
        row = row_start
        col_pairs = [("B", "C"), ("F", "G")]
        i = 0
        for code in codes:
            c_mark, c_text = col_pairs[i % 2]
            if ck:
                _try_add_image(
                    ws,
                    ck,
                    f"{c_mark}{row}",
                    max_width=CHECKMARK_IMAGE_BOX[0],
                    max_height=CHECKMARK_IMAGE_BOX[1],
                )
            ws[f"{c_text}{row}"] = code
            if i % 2 == 1:
                row += 1
            i += 1
        ws["A16"] = djpmw
        ws["K16"] = djpmw
        ws["C28"] = krtm[:7]
        ws["C30"] = f"{zxl} {jldw.upper()}S"
        ws["C37"] = khlb
        ws["L37"] = khlb
        ws["C40"] = f"OF   {htzxs}"
        ws["C42"] = khht
        ws["C45"] = cxmc
        ws["L45"] = cxmc
        ws["C47"] = ship_date


@any_route("/api/Ravencloud/batch_mark_generate", methods=["POST"])
async def api_batch_mark_generate(request):
    try:
        form = await request.form()

        mark_type = (form.get("mark_type") or "").strip()
        env_opt = (form.get("env_opt") or "1").strip()
        region_opt = (form.get("region_opt") or "1").strip()
        out_type = (form.get("out_type") or "1").strip()  # 1/pdf, 2/jpg, 3/低版本pdf, 4/批量pdf
        mode = (form.get("mode") or "current").strip()  # current / batch
        ship_date = (form.get("ship_date") or "").strip()
        filename_prefix = _safe_filename(form.get("filename_prefix") or "批量唛头_")

        number_list_raw = form.get("number_list", "[]")
        try:
            number_list = json.loads(number_list_raw) if isinstance(number_list_raw, str) else (number_list_raw or [])
        except Exception:
            number_list = []

        if not mark_type:
            return json_result(code=-1, msg="唛头类型不能为空")
        if mode == "batch" and not number_list:
            return json_result(code=-1, msg="批量模式请先选择订单")

        template_path = _get_template_path(mark_type, env_opt, region_opt)
        if not template_path or not os.path.exists(template_path):
            return json_result(code=-1, msg=f"未找到模板文件: {template_path}")

        output_dir = config.get_today_upload_path()
        os.makedirs(output_dir, exist_ok=True)

        if mode == "current":
            current_no = form.get("current_number")
            current_wxht = form.get("current_wxht", "")
            selected_orders = [{"number": str(current_no or ""), "wxht": str(current_wxht or "")}] if current_no else []
        else:
            selected_orders = _get_selected_order_numbers(number_list)

        if not selected_orders:
            return json_result(code=-1, msg="未找到可处理订单")

        checkmark_path = _get_tpzx_image_path_by_cpbh("打勾") if mark_type == "多货号唛头" else ""

        files = []
        idx = 0

        for order in selected_orders:
            order_no = order["number"]
            wxht = order.get("wxht", "")
            khht = _get_order_dforder(wxht or order_no)

            details = _get_detail_rows(order_no)
            for detail in details:
                idx += 1
                cpbh = str(detail.get("cpbh") or "").strip()
                prod = _get_product_info(cpbh)

                djpmw = _get_wypm_name(cpbh) or str(prod.get("djpmw") or "")
                tsbj = str(prod.get("tsbj") or "").strip()
                khlb = str(prod.get("khlb") or "").strip()
                cxmc = str(prod.get("cxmc") or "").strip()
                jgby = str(prod.get("jgby") or "").strip()
                krtm = str(prod.get("krtm") or "").strip()
                czkrhh = str(prod.get("czkrhh") or "").strip()
                jldw = str(prod.get("jldw") or "").strip()
                tmtp_path = _image_path(prod.get("tmtp"))

                wb = load_workbook(template_path)
                ws = wb.active

                _fill_by_mark_type(
                    ws,
                    mark_type,
                    detail,
                    {
                        "czkrhh": czkrhh,
                        "djpmw": djpmw,
                        "ship_date": ship_date,
                        "khht": khht,
                        "krtm": krtm,
                        "khlb": khlb,
                        "cxmc": cxmc,
                        "jgby": jgby,
                        "jldw": jldw,
                        "checkmark_path": checkmark_path,
                    },
                )

                # 条码图片（zscp.tmtp）
                if tmtp_path:
                    _try_add_image(ws, tmtp_path, TMTP_ANCHOR.get(mark_type, "A4"))

                # 港口图（根据 khht 前2位/前1位在 tpzx.sb 查图）
                gk_path = _get_tpzx_image_path_by_sb(khht[:2]) or _get_tpzx_image_path_by_sb(khht[:1])
                if gk_path:
                    _try_add_image(ws, gk_path, GK_ANCHOR.get(mark_type, "F8"))

                # 特殊标记图 tsbj
                if tsbj and mark_type in ("港加特标唛头", "港加特标加警语唛头", "内外箱唛头"):
                    ts_path = _get_tpzx_image_path_by_cpbh(tsbj)
                    if ts_path:
                        _try_add_image(ws, ts_path, "F3" if mark_type != "内外箱唛头" else "D2")

                # 环保图标（PP/LDEP-1/LDPE-2）
                if env_opt in ("3", "4", "5"):
                    flag = {"3": "PP标志", "4": "LDEP-1标志", "5": "LDPE-2标志"}[env_opt]
                    flag_path = _get_tpzx_image_path_by_cpbh(flag)
                    if flag_path:
                        if mark_type in ("内外箱唛头", "常规唛头"):
                            _try_add_image(ws, flag_path, "E29")
                            if mark_type == "内外箱唛头":
                                _try_add_image(ws, flag_path, "L24")
                        elif mark_type in ("港加特标唛头", "港加特标加警语唛头", "多货号唛头"):
                            flag_anchor = "G45"
                            if str(region_opt) == "2" and mark_type == "多货号唛头":
                                flag_anchor = "G47"
                            elif str(region_opt) == "2" and mark_type in ("港加特标唛头", "港加特标加警语唛头"):
                                flag_anchor = "G43"
                            _try_add_image(ws, flag_path, flag_anchor)

                cpbh_show = _safe_filename(czkrhh or cpbh)
                xls_name = f"{filename_prefix}{idx};{cpbh_show} IN {khht}.xlsx"
                xls_path = os.path.join(output_dir, xls_name)
                wb.save(xls_path)
                wb.close()

                # 输出类型：当前统一返回 PDF/JPG 文件（WhaleCloud 服务端稳态）
                if out_type in ("1", "3", "4"):
                    pdf_res = _convert_excel_to_pdf(xls_path)
                    try:
                        os.remove(xls_path)
                    except Exception:
                        pass
                    if pdf_res.get("success"):
                        files.append(_to_upload_rel_path(pdf_res["pdf_path"]))
                else:
                    # out_type == '2'：先转PDF返回（如必须JPG可后续加专门渲染器）
                    pdf_res = _convert_excel_to_pdf(xls_path)
                    try:
                        os.remove(xls_path)
                    except Exception:
                        pass
                    if pdf_res.get("success"):
                        files.append(_to_upload_rel_path(pdf_res["pdf_path"]))

        return json_result(code=1, msg=f"成功生成{len(files)}个唛头文件", data={"files": files, "count": len(files)})

    except Exception:
        logger.error(trace_error())
        return json_result(code=-1, msg=trace_error())
