"""
优景 I.R 报表导出。

对照原 Delphi「优景I.R」；模版：``template/uvIR.xlsx``（或 ``data/saier20050822/addonfiles/uvIR.xls``）。
"""

from any import *
from .__default__ import get_user_path
from .model import *

import json
import math
import os
from copy import copy

from openpyxl.cell import Cell
from openpyxl.cell.cell import MergedCell
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.drawing.image import Image as Image_Get
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import coordinate_to_tuple
from sqlalchemy import text as sql_text

TEMPLATE_IR = "uvIR.xlsx"
TEMPLATE_IR_LEGACY = "uvIR.xls"

HEADER_UNION = (
    "NINGBO UNION VISION IMP&EXP CO.LTD",
    "BUILDING No. 1, No. 1 (17-4), YINCHEN BUSINESS CENTER, YINZHOU DISTRICT, "
    "NINGBO,ZHEJIANG,  315103, CHINA  TEL: 86-574-27833605",
)
HEADER_SOURCE = (
    "Union Source Co.,Ltd.",
    "20F, NO.1 Building, Ningbo R & D Park, 399 Juxian Road, National Hi-Tech Zone, Ningbo, China 315103 ",
)

CGGD_IMAGE_FIELDS = ("wxmt1", "mzzp1", "kxzp1", "kxzp2", "cpzp1", "txm1", "tmsm1", "wxsm1")
CGGD_IMAGE_COLS = ("B", "D", "F", "H", "I", "K", "M", "N")
ZGTPSHEET_IMAGE_COLS = ("B", "D", "F", "H", "I", "K", "M")

LIST_TEMPLATE_ROW = 12
DESC_TEMPLATE_ROW = 12
BLOCK_TEMPLATE_START = 12
BLOCK_TEMPLATE_ROWS = 8


def _resolve_writable_cell(ws, row, col):
    """插入行后合并区可能变成 MergedCell，需解析为可写 Cell。"""
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            row, col = mr.min_row, mr.min_col
            break

    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return cell

    for mr in list(ws.merged_cells.ranges):
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            anchor = ws.cell(row=mr.min_row, column=mr.min_col)
            if not isinstance(anchor, MergedCell):
                return anchor
            ws.unmerge_cells(str(mr))
            break

    key = (row, col)
    if key in ws._cells and isinstance(ws._cells[key], MergedCell):
        del ws._cells[key]
    cell = Cell(ws, row=row, column=col)
    ws._cells[key] = cell
    return cell


def safe_write(ws, coord, value, num_format=None):
    row, col = coordinate_to_tuple(coord)
    target_cell = _resolve_writable_cell(ws, row, col)
    target_cell.value = value
    if num_format:
        target_cell.number_format = num_format


def offset_img(img, col, row, x_pad=2, y_pad=2):
    size = XDRPositiveSize2D(pixels_to_EMU(img.width), pixels_to_EMU(img.height))
    marker = AnchorMarker(col=col - 1, colOff=pixels_to_EMU(x_pad), row=row - 1, rowOff=pixels_to_EMU(y_pad))
    img.anchor = OneCellAnchor(_from=marker, ext=size)


def get_field_value(row, field_name, default="", as_type=str):
    if row is None:
        return default if as_type is str else (0.0 if as_type is float else 0)
    if hasattr(row, "_mapping"):
        value = row._mapping.get(field_name, default)
    elif hasattr(row, field_name):
        value = getattr(row, field_name, default)
    else:
        value = row.get(field_name, default) if isinstance(row, dict) else default
    if as_type is float:
        try:
            return float(value) if value not in ("", None) else 0.0
        except (TypeError, ValueError):
            return 0.0
    if value is None:
        return ""
    return str(value).strip()


def copy_row_style(ws, source_row, target_row):
    for col in range(1, ws.max_column + 1):
        src = ws.cell(row=source_row, column=col)
        dst = ws.cell(row=target_row, column=col)
        dst.value = src.value
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.alignment = copy(src.alignment)


def _safe_filename(s):
    s = str(s or "").strip() or "export"
    for c in r'\/:*?"<>|':
        s = s.replace(c, "_")
    return s


def _format_zgrq(val):
    if val is None or val == "":
        return ""
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if " " in s:
        s = s.split()[0]
    if "T" in s and len(s) > 10:
        s = s[:10]
    return s


def _round1(v):
    return round(float(v or 0) * 10) / 10


def _round2(v):
    return round(float(v or 0) * 100) / 100


def _find_template_path():
    candidates = [
        os.path.join(config.data_upload_path, "template", TEMPLATE_IR),
        os.path.join(config.data_upload_path, "template", TEMPLATE_IR_LEGACY),
        os.path.join(config.data_upload_path, "data", "saier20050822", "addonfiles", TEMPLATE_IR),
        os.path.join(config.data_upload_path, "data", "saier20050822", "addonfiles", TEMPLATE_IR_LEGACY),
    ]
    for p in candidates:
        if os.path.isfile(p):
            return p
    return None


def _user_can_export(user):
    info = get_user_path(user.username) or {}
    path = info.get("path", "") or ""
    position = info.get("position", "") or ""
    return ("外销" in path) or ("单证" in position) or ("外销" in position)


def _detail_filter_sql():
    return "((chxs > 0) OR (sfpx = '是'))"


def _esc_sql(v):
    return str(v or "").replace("'", "''")


def _cggd_match_params(detail):
    cpbh = get_field_value(detail, "cpbh")
    zycpbh = get_field_value(detail, "zycpbh") or cpbh
    hthm = get_field_value(detail, "cght")
    wxwyzd = get_field_value(detail, "wxwyzd")
    return cpbh, zycpbh, hthm, wxwyzd


def _parse_image_field_first_src(raw):
    """解析附件 JSON，与 customer_claim_report._parse_tpzx_tpmc_first_src 一致。"""
    if raw is None:
        return None
    if isinstance(raw, dict):
        return raw.get("src") or raw.get("path")
    if isinstance(raw, list) and len(raw) > 0:
        return _parse_image_field_first_src(raw[0])
    s = str(raw).strip()
    if s in ("", "[]", "null", "None", "{}"):
        return None
    try:
        data = json.loads(s)
    except (json.JSONDecodeError, TypeError):
        return s if not s.startswith("[") and not s.startswith("{") else None
    if isinstance(data, list) and len(data) > 0:
        first = data[0]
        if isinstance(first, dict):
            return first.get("src") or first.get("path")
    if isinstance(data, dict):
        return data.get("src") or data.get("path")
    return None


def _image_src_to_local_path(file_path):
    """将 JSON src 或绝对/相对路径转为可读本地文件路径。"""
    if not file_path:
        return None
    fp = str(file_path).strip()
    if not fp:
        return None
    if os.path.isabs(fp) and os.path.isfile(fp):
        return fp
    upload_root = str(getattr(config, "data_upload_path", "") or "").rstrip("/\\")
    if upload_root:
        if fp.replace("\\", "/").startswith(upload_root.replace("\\", "/")) and os.path.isfile(fp):
            return fp
        norm = fp.replace("\\", "/")
        for marker in ("/data/upload/", "data/upload/"):
            idx = norm.find(marker)
            if idx >= 0:
                rel = norm[idx + len(marker) :].lstrip("/")
                candidate = os.path.join(upload_root, rel)
                if os.path.isfile(candidate):
                    return candidate
        candidate = os.path.join(upload_root, fp.lstrip("/\\"))
        if os.path.isfile(candidate):
            return candidate
    rel = fp.lstrip("/\\").replace("\\", "/")
    if rel != fp and upload_root:
        fn2 = os.path.join(upload_root, rel)
        if os.path.isfile(fn2):
            return fn2
    return None


def _resolve_image_field(raw_value):
    src = _parse_image_field_first_src(raw_value)
    return _image_src_to_local_path(src)


def _cggd_image_path_from_db(raw_value):
    """cggdsheet 字段可能是 JSON 或直链路径。"""
    if raw_value is None:
        return None
    s = str(raw_value).strip()
    if len(s) <= 5:
        return None
    if s.startswith("[") or s.startswith("{"):
        return _resolve_image_field(raw_value)
    return _image_src_to_local_path(s)


def _add_cell_image(ws, col_letter, row, abs_path, default_h=90):
    """参考 customer_claim_report._add_signature_image。"""
    if not abs_path or not os.path.isfile(abs_path):
        return False
    try:
        img = Image_Get(abs_path)
        col_dim = ws.column_dimensions[col_letter].width or 12
        col_width_px = float(col_dim) * 7
        row_dim = ws.row_dimensions[row].height or default_h
        img.width = int(max(20, col_width_px - 5))
        img.height = int(max(20, float(row_dim) - 5))
        col_idx = ord(col_letter.upper()) - ord("A") + 1
        offset_img(img, col_idx, row, x_pad=2, y_pad=2)
        ws.add_image(img)
        return True
    except Exception:
        logger.warning(f"插入图片失败 {col_letter}{row}: {trace_error()}")
        return False


def apply_subscript(text, djyxb_str):
    """平替 Delphi Characters[].Font.Subscript。"""
    if not text or not djyxb_str:
        return text
    indices = set()
    for part in str(djyxb_str).replace("，", ",").split(","):
        part = part.strip()
        if part.isdigit():
            indices.add(int(part) - 1)
    if not indices:
        return str(text).upper()
    sub_font = InlineFont(vertAlign="subscript")
    elements, curr = [], ""
    for i, char in enumerate(str(text)):
        if i in indices:
            if curr:
                elements.append(curr)
                curr = ""
            elements.append(TextBlock(sub_font, char))
        else:
            curr += char
    if curr:
        elements.append(curr)
    return CellRichText(*elements)


def _product_code(detail):
    cz = get_field_value(detail, "czkrhh")
    if cz:
        return cz
    return get_field_value(detail, "cpbh")


def _query_zscp_barcode(s, cpbh, fallback):
    row = s.execute(sql_text("SELECT IRcptm, cctm FROM zscp WHERE cpbh=:cpbh LIMIT 1"), {"cpbh": cpbh}).fetchone()
    if not row:
        return fallback
    cctm = get_field_value(row, "cctm")
    if cctm:
        return cctm
    irc = get_field_value(row, "IRcptm")
    if irc:
        return irc
    return fallback


def _build_name_line(prefix, djpmy, djyxb=None):
    body = f"{prefix}:{djpmy.upper()}" if prefix else djpmy.upper()
    if djyxb:
        return apply_subscript(body, djyxb)
    return body


def _name_line_from_detail(detail, prefix=None):
    """从 cymxsheet 明细取英文品名；新系统无 tspmb，对应 Delphi 的 else 分支。"""
    if prefix is None:
        prefix = _product_code(detail)
    djpmy = get_field_value(detail, "djpmy")
    if not djpmy:
        return f"{prefix}:" if prefix else ""
    djyxb = get_field_value(detail, "djyxb")
    if djyxb:
        return _build_name_line(prefix, djpmy, djyxb)
    return f"{prefix}: {str(djpmy).upper()}" if prefix else str(djpmy).upper()


def _write_mirror(ws, row, col_a, col_mirror, value):
    safe_write(ws, f"{col_a}{row}", value)
    safe_write(ws, f"{col_mirror}{row}", value)


def _insert_row_from_template(ws, insert_at, template_row):
    ws.insert_rows(insert_at)
    copy_row_style(ws, template_row, insert_at)


def _sum_sheet_weights(s, pid):
    rows = s.execute(
        sql_text(f"SELECT zmz, zjz, ztj FROM cymxsheet WHERE pid=:pid AND {_detail_filter_sql()}"), {"pid": pid}
    ).fetchall()
    mzhj = sum(get_field_value(r, "zmz", 0, float) for r in rows)
    jzhj = sum(get_field_value(r, "zjz", 0, float) for r in rows)
    tjhj = sum(get_field_value(r, "ztj", 0, float) for r in rows)
    return _round1(mzhj), _round1(jzhj), _round2(tjhj)


def _fetch_cggd_image_field(detail, field):
    """对照 Pascal：逐字段查 cggdsheet；用 run_sql 便于排查，条件逐级放宽。"""
    cpbh, zycpbh, hthm, wxwyzd = _cggd_match_params(detail)
    if not cpbh and not zycpbh:
        return None
    cpbh_e, zy_e = _esc_sql(cpbh), _esc_sql(zycpbh)
    hthm_e, wx_e = _esc_sql(hthm), _esc_sql(wxwyzd)
    base = f"((cpbh='{cpbh_e}') OR (cpbh='{zy_e}')) AND (LENGTH(IFNULL({field},''))>5)"
    sql_list = []
    if hthm and wxwyzd:
        sql_list.append(
            f"SELECT {field} FROM cggdsheet WHERE {base} AND (hthm='{hthm_e}') AND (wxwyzd='{wx_e}') LIMIT 1"
        )
    if hthm:
        sql_list.append(f"SELECT {field} FROM cggdsheet WHERE {base} AND (hthm='{hthm_e}') ORDER BY sid DESC LIMIT 1")
    sql_list.append(f"SELECT {field} FROM cggdsheet WHERE {base} ORDER BY sid DESC LIMIT 1")

    for sql in sql_list:
        try:
            rows = run_sql(sql)
        except Exception:
            logger.warning(f"cggdsheet 查询失败 {field}: {trace_error()}")
            continue
        if not rows:
            continue
        raw = rows[0].get(field) if isinstance(rows[0], dict) else None
        if raw is None and hasattr(rows[0], field):
            raw = getattr(rows[0], field, None)
        if raw is not None and len(str(raw).strip()) > 5:
            logger.info(f"cggdsheet {field} hit: sql={sql[:120]}... path={_cggd_image_path_from_db(raw)}")
            return raw
    logger.warning(f"cggdsheet {field} 未命中: cpbh={cpbh}, zycpbh={zycpbh}, hthm={hthm}, wxwyzd={wxwyzd}")
    return None


def _fill_header(ws, s, main):
    khmc = get_field_value(main, "khmc")
    ssgs = ""
    kh_row = s.execute(
        sql_text("SELECT ssgs FROM kh WHERE company_name=:company_name LIMIT 1"), {"company_name": khmc}
    ).fetchone()
    if kh_row:
        ssgs = get_field_value(kh_row, "ssgs")

    kh_id = get_field_value(main, "kh_id")
    tpl_row = s.execute(
        sql_text("SELECT xm FROM cyzglsheet WHERE zm='优景报关模板' AND xm=:xm LIMIT 1"), {"xm": kh_id}
    ).fetchone()
    if tpl_row:
        safe_write(ws, "A1", HEADER_UNION[0])
        safe_write(ws, "A3", HEADER_UNION[1])
    else:
        safe_write(ws, "A1", HEADER_SOURCE[0])
        safe_write(ws, "A3", HEADER_SOURCE[1])

    safe_write(ws, "D9", ssgs)
    safe_write(ws, "K9", _format_zgrq(main.zgrq if hasattr(main, "zgrq") else main.get("zgrq")))
    fphm = get_field_value(main, "fphm")
    safe_write(ws, "D10", f"#{fphm}")
    return ssgs


def _fill_list_section(ws, s, pid):
    details = s.execute(
        sql_text(f"SELECT * FROM cymxsheet WHERE pid=:pid AND {_detail_filter_sql()} ORDER BY seq, sid"), {"pid": pid}
    ).fetchall()
    k = 0
    i4 = 0
    template_row = LIST_TEMPLATE_ROW
    for detail in details:
        k += 1
        row = template_row + k
        _insert_row_from_template(ws, row, template_row)
        ws.row_dimensions[row].height = 24
        code = _product_code(detail)
        _write_mirror(ws, row, "A", "O", code)
        chsl = get_field_value(detail, "chsl", 0, float)
        chxs = int(get_field_value(detail, "chxs", 0, float))
        safe_write(ws, f"C{row}", chsl)
        safe_write(ws, f"F{row}", chxs)
        safe_write(ws, f"I{row}", chsl)
        safe_write(ws, f"L{row}", chxs)
        i4 += chxs
    ws.delete_rows(template_row, 1)
    return details, k, i4


def _fill_sampling_text(ws, k, i4):
    k = k + 14
    safe_write(ws, f"A{11 + k}", f"From the total of {i4}CTNS, we had picked 0CTNS randomly")
    k += 1
    safe_write(
        ws,
        f"A{11 + k}",
        "from the batch of above items and this was subjected to qualitative and quantitative control.",
    )
    k += 1
    return k


def _fill_desc_section(ws, s, details, k):
    template_row = LIST_TEMPLATE_ROW + k
    for detail in details:
        k += 1
        row = LIST_TEMPLATE_ROW + k
        _insert_row_from_template(ws, row, template_row)
        ws.row_dimensions[row].height = 30
        val = _name_line_from_detail(detail)
        _write_mirror(ws, row, "A", "P", val)
        chxs = int(get_field_value(detail, "chxs", 0, float))
        jldw = get_field_value(detail, "jldw").upper()
        wxrl = int(get_field_value(detail, "wxrl", 0, float))
        safe_write(ws, f"F{row}", chxs)
        safe_write(ws, f"G{row}", "CTNS, ")
        safe_write(ws, f"H{row}", wxrl)
        safe_write(ws, f"I{row}", f"{jldw}S/CTN")
        safe_write(ws, f"L{row}", f"{jldw}S")
    if details:
        ws.delete_rows(template_row, 1)
    k += 2
    return k


def _fill_detail_blocks(ws, s, details, k):
    template_start = LIST_TEMPLATE_ROW + k
    item_no = 0
    for detail in details:
        k += BLOCK_TEMPLATE_ROWS
        block_start = LIST_TEMPLATE_ROW + k
        for offset in range(BLOCK_TEMPLATE_ROWS):
            _insert_row_from_template(ws, block_start + offset, template_start + offset)
        item_no += 1

        zycpbh = get_field_value(detail, "zycpbh") or get_field_value(detail, "cpbh")
        ircptm = _query_zscp_barcode(s, zycpbh, get_field_value(detail, "krtm"))
        title = f"{_name_line_from_detail(detail)}       BARCODE NUMBER: {ircptm}"
        row_title = block_start
        safe_write(ws, f"A{row_title}", str(item_no))
        _write_mirror(ws, row_title, "B", "Q", title)
        safe_write(ws, f"B{row_title + 1}", f"MATERIAL: {get_field_value(detail, 'caizi')}")
        safe_write(ws, f"Q{row_title + 1}", f"MATERIAL: {get_field_value(detail, 'caizi')}")
        safe_write(ws, f"B{row_title + 2}", f"PACKING: {get_field_value(detail, 'ywbz')}")
        safe_write(ws, f"Q{row_title + 2}", f"PACKING: {get_field_value(detail, 'ywbz')}")

        col_idx = 0
        img_row_num = row_title + 3
        for field in CGGD_IMAGE_FIELDS:
            raw = _fetch_cggd_image_field(detail, field)
            img_path = _cggd_image_path_from_db(raw)
            if not img_path:
                continue
            if col_idx >= len(CGGD_IMAGE_COLS):
                break
            col_letter = CGGD_IMAGE_COLS[col_idx]
            ws.row_dimensions[img_row_num].height = 90
            _add_cell_image(ws, col_letter, img_row_num, img_path, default_h=90)
            col_idx += 1

    if details:
        ws.delete_rows(template_start, BLOCK_TEMPLATE_ROWS)
    return k


def _fill_zgtp_section(ws, s, main, k):
    fphm = get_field_value(main, "fphm")
    tp_count = s.execute(sql_text("SELECT COUNT(*) AS cnt FROM zgtpsheet WHERE fphp=:fphp"), {"fphp": fphm}).fetchone()
    tpz = int(get_field_value(tp_count, "cnt", 0, float)) if tp_count else 0

    qfh_row = s.execute(
        sql_text("SELECT qfh FROM zgtp WHERE fphp=:fphp AND LENGTH(IFNULL(qfh,''))>5 LIMIT 1"), {"fphp": fphm}
    ).fetchone()
    qfh_path = _cggd_image_path_from_db(get_field_value(qfh_row, "qfh")) if qfh_row else None

    zld_row = s.execute(
        sql_text("SELECT zld FROM zgtp WHERE fphp=:fphp AND LENGTH(IFNULL(zld,''))>5 LIMIT 1"), {"fphp": fphm}
    ).fetchone()
    zld_path = _cggd_image_path_from_db(get_field_value(zld_row, "zld")) if zld_row else None

    tps = 1.0 if tpz == 0 else tpz / 7.0
    tph = int(math.ceil(tps)) if tps > 1 else 1
    photo_row_base = 13 + k
    if tph > 1:
        for _ in range(tph - 1):
            insert_at = photo_row_base + 1
            _insert_row_from_template(ws, insert_at, photo_row_base)
            photo_row_base = insert_at

    tp_rows = s.execute(
        sql_text("SELECT tp, rid, sid FROM zgtpsheet WHERE fphp=:fphp AND LENGTH(IFNULL(tp,''))>5 ORDER BY seq, sid"),
        {"fphp": fphm},
    ).fetchall()
    tp3 = 0
    current_k = k
    for tp_row in tp_rows:
        if tp3 >= 7:
            tp3 = 0
            current_k += 1
        tp3 += 1
        col = ZGTPSHEET_IMAGE_COLS[tp3 - 1]
        row_num = 13 + current_k
        ws.row_dimensions[row_num].height = 90
        img_path = _cggd_image_path_from_db(get_field_value(tp_row, "tp"))
        if img_path:
            _add_cell_image(ws, col, row_num, img_path, default_h=90)

    k = current_k + 3
    safe_write(ws, f"E{14 + k}", get_field_value(main, "hglx"))
    k += 4
    safe_write(ws, f"B{14 + k}", get_field_value(main, "xh"), num_format="@")
    safe_write(ws, f"E{14 + k}", get_field_value(main, "fh"), num_format="@")
    if qfh_path:
        ws.row_dimensions[14 + k].height = 90
        _add_cell_image(ws, "H", 14 + k, qfh_path, default_h=90)
    safe_write(ws, f"K{14 + k}", f"TOTAL {get_field_value(main, 'xshj')} LOADED ")
    k += 6
    if zld_path:
        ws.row_dimensions[14 + k].height = 90
        _add_cell_image(ws, "B", 14 + k, zld_path, default_h=90)
    safe_write(ws, f"F{14 + k}", "KGS")
    safe_write(ws, f"H{14 + k}", "KGS")
    safe_write(ws, f"I{14 + k}", "KGS")
    return k


def _fill_footer_weights(ws, k, mzhj, jzhj):
    safe_write(ws, f"K{14 + k}", f"{mzhj}KGS")
    safe_write(ws, f"M{14 + k}", f"{jzhj}KGS")


def build_ir_workbook(s, main, pid):
    template_path = _find_template_path()
    if not template_path:
        raise FileNotFoundError("模版文件缺失：uvIR.xlsx / uvIR.xls")

    ext = os.path.splitext(template_path)[1].lower()
    if ext == ".xls":
        raise ValueError("请将模版 uvIR.xls 转为 uvIR.xlsx 后放在 template 目录")
    wb = load_workbook(template_path)
    ws = wb.worksheets[0]
    logger.error(f"test: {main}")

    mzhj, jzhj, tjhj = _sum_sheet_weights(s, pid)
    _ = tjhj
    _fill_header(ws, s, main)

    details, k, i4 = _fill_list_section(ws, s, pid)
    k = _fill_sampling_text(ws, k, i4)
    k = _fill_desc_section(ws, s, details, k)
    k = _fill_detail_blocks(ws, s, details, k)
    k = _fill_zgtp_section(ws, s, main, k)
    _fill_footer_weights(ws, k, mzhj, jzhj)
    return wb


def export_ir_for_rid(s, rid, user):
    if not _user_can_export(user):
        return None, "您没有权限导出 I.R 报表（需外销或单证岗位）"

    main = s.execute(sql_text("SELECT * FROM cymx WHERE rid=:rid LIMIT 1"), {"rid": rid}).fetchone()
    if not main:
        return None, "未找到出运明细记录"

    detail_cnt = s.execute(
        sql_text(f"SELECT COUNT(*) AS cnt FROM cymxsheet WHERE pid=:pid AND {_detail_filter_sql()}"), {"pid": rid}
    ).fetchone()
    if not detail_cnt or int(get_field_value(detail_cnt, "cnt", 0, float)) == 0:
        return None, "没有可导出的出运明细行（需出货箱数>0 或 是否拼箱=是）"

    pid = rid
    try:
        wb = build_ir_workbook(s, main, pid)
    except FileNotFoundError as e:
        return None, str(e)
    except Exception:
        logger.error(trace_error())
        return None, trace_error()

    fphm = _safe_filename(get_field_value(main, "fphm"))
    save_dir = config.tmp_path
    os.makedirs(save_dir, exist_ok=True)
    out_name = f"{fphm} IR.xlsx"
    out_path = os.path.join(save_dir, out_name)
    wb.save(out_path)
    wb.close()

    s.execute(
        sql_text("UPDATE cymx SET IRdyrq=:dt WHERE rid=:rid AND (IFNULL(IRdyrq,'')='')"),
        {"dt": time.strftime("%Y-%m-%d"), "rid": rid},
    )
    return out_name, None


@any_route("/api/shipping/export/detailed/IR/list", methods=["POST", "GET"])
@require_token
async def view_saier_youjing_ir_export(request):
    """
    优景 I.R 报表导出。

    请求 JSON：
    - rid: 出运明细 rid（必填）
    """
    s = Session()
    user = request.current_user
    try:
        j = await request.json()
        rid = (j.get("rid") or "").strip()
        if not rid:
            return json_result(-1, "缺少出运明细 rid")

        out_name, err = export_ir_for_rid(s, rid, user)
        if err:
            s.rollback()
            return json_result(-1, err)
        s.commit()
        return json_result(1, "导出成功", out_name)
    except Exception:
        logger.error(trace_error())
        s.rollback()
        return json_result(-1, trace_error())
    finally:
        s.close()
