"""
快件与办公费用报表输出。

对照原 Delphi「快件与办公费用报表输出」；模版：``template/优景快件空白表格.xlsx``、``CSD付款审批单.xlsx``。
"""

from any import *
from .model import *
from .payment import chinese_amount

import json
import os
import subprocess
import time
import zipfile

try:
    import xlwings as xw
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "xlwings"])
    import xlwings as xw


TEMPLATE_EXPRESS_OFFICE = "优景快件空白表格.xlsx"
TEMPLATE_CSD = "CSD付款审批单.xlsx"
EXPRESS_OFFICE_HKLX = ("办公费用", "快件费用")


def safe_write(ws, coord, value, num_format=None):
    """合并单元格安全写入：coord 若在合并区内则写入该区域左上角主单元格。"""
    rng = ws.range(coord)
    area = rng.merge_area
    rows, cols = area.shape
    if rows * cols > 1:
        cell = area(1, 1)
        cell.value = value
        if num_format:
            cell.number_format = num_format
    else:
        if num_format:
            ws[coord].number_format = num_format
        ws[coord].value = value


def _safe_write_center(ws, coord, value, num_format=None):
    """写入并居中（合并单元格取左上角）。"""
    safe_write(ws, coord, value, num_format)
    _set_cell_center(ws, coord)


def copy_row_style(ws, source_row, target_row):
    ws.range(source_row, source_row).copy()
    ws.range(target_row, target_row).paste(paste="formats")


def insert_row_at(ws, row):
    ws.range(row, row).api.EntireRow.Insert()


def _add_picture_fitted_to_cell(ws, coord, path, pad=2):
    """将图片缩放并居中放入单元格，保持纵横比（允许放大）。"""
    area = ws.range(coord).merge_area
    pic = ws.pictures.add(path, left=area.left, top=area.top)
    if pic.width and pic.height:
        scale = min(1, (area.width - pad * 2) / pic.width, (area.height - pad * 2) / pic.height)
        if scale < 1:
            pic.width *= scale
    pic.left = area.left + (area.width - pic.width) / 2
    pic.top = area.top + (area.height - pic.height) / 2
    return pic


def _set_cell_center(ws, coord):
    rng = ws.range(coord)
    area = rng.merge_area
    rows, cols = area.shape
    cell = area(1, 1) if rows * cols > 1 else rng
    api = cell.api
    api.WrapText = True
    api.HorizontalAlignment = -4108
    api.VerticalAlignment = -4108


def _merge_main_amount_cell(ws, row):
    """模版主表金额列为 C:D 合并；勿 UnMerge，否则数字只占半格且无法居中。"""
    rng = ws.range(f"C{row}:D{row}")
    try:
        if not rng.api.MergeCells:
            rng.api.Merge()
    except Exception:
        try:
            rng.merge()
        except Exception:
            pass


def _openpyxl_center_cells(excel_path, coords):
    """保存后写入 xlsx 对齐样式，供 whale_report.jar 转 PDF 时读取。"""
    if not coords:
        return
    try:
        from openpyxl import load_workbook
        from openpyxl.cell.cell import MergedCell
        from openpyxl.styles import Alignment
        from openpyxl.utils import coordinate_to_tuple
    except ImportError:
        return

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wb = load_workbook(excel_path)
    ws = wb.active
    for coord in coords:
        row, col = coordinate_to_tuple(coord)
        target = None
        for merged in ws.merged_cells.ranges:
            if (
                merged.min_row <= row <= merged.max_row
                and merged.min_col <= col <= merged.max_col
            ):
                target = ws.cell(merged.min_row, merged.min_col)
                break
        if target is None:
            cell = ws.cell(row=row, column=col)
            if not isinstance(cell, MergedCell):
                target = cell
        if target is not None:
            target.alignment = center
    wb.save(excel_path)


def _start_excel_app():
    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    app.screen_updating = False
    return app


def get_field_value(row, field_name, default="", as_type=str):
    value = row.get(field_name, default)
    if as_type is float:
        return float(value) if value not in ("", None) else 0.0
    if value is None:
        return ""
    return str(value).strip()


def _safe_filename(s):
    s = str(s or "").strip() or "export"
    for c in r'\/:*?"<>|':
        s = s.replace(c, "_")
    return s


def convert_excel_to_pdf(excel_path, output_dir=None):
    try:
        if output_dir is None:
            output_dir = os.path.dirname(excel_path)
        pdf_path = os.path.splitext(excel_path)[0] + ".pdf"
        if not getattr(config, "java_path", None):
            return {"success": False, "error": "Java路径未配置（config.java_path）"}
        if not getattr(config, "report_jar", None):
            return {"success": False, "error": "whale_report.jar路径未配置（config.report_jar）"}
        if not os.path.exists(config.report_jar):
            return {"success": False, "error": f"whale_report.jar文件不存在: {config.report_jar}"}
        cmd = [config.java_path, "-jar", config.report_jar, "a", "b", excel_path, pdf_path, "2"]
        result = subprocess.run(cmd, capture_output=True, timeout=120, text=True)
        if result.returncode != 0:
            return {
                "success": False,
                "error": f"PDF转换失败（返回码 {result.returncode}）：{result.stderr or result.stdout}",
            }
        if not os.path.exists(pdf_path):
            return {"success": False, "error": f"PDF文件未生成（期望路径：{pdf_path}）"}
        return {"success": True, "pdf_path": pdf_path, "error": None}
    except subprocess.TimeoutExpired:
        return {"success": False, "error": "PDF转换超时（超过2分钟）"}
    except Exception:
        logger.error(trace_error())
        return {"success": False, "error": trace_error()}


def _parse_tpzx_tpmc_first_src(tpmc_raw):
    if tpmc_raw is None:
        return None
    raw = str(tpmc_raw).strip()
    if raw in ("", "[]", "null"):
        return None
    try:
        data = json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        return None
    if not isinstance(data, list) or len(data) == 0:
        return None
    first = data[0]
    if not isinstance(first, dict):
        return None
    return first.get("src")


def _tpzx_src_to_local_path(file_path):
    if not file_path:
        return None
    fp = str(file_path).strip()
    if not fp:
        return None
    fn = os.path.join(config.data_upload_path, fp)
    if os.path.isfile(fn):
        return fn
    rel = fp.lstrip("/\\").replace("\\", "/")
    if rel != fp:
        fn2 = os.path.join(config.data_upload_path, rel)
        if os.path.isfile(fn2):
            return fn2
    if os.path.isabs(fp) and os.path.isfile(fp):
        return fp
    return None


def _format_sqrq1(dt_val):
    if dt_val is None or dt_val == "":
        return ""
    if hasattr(dt_val, "strftime"):
        s = dt_val.strftime("%Y-%m-%d")
    else:
        s = str(dt_val).strip()
        if " " in s:
            s = s.split()[0]
        if "T" in s and len(s) > 10:
            s = s[:10]
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return f"{s[0:4]}年{s[5:7]}月{s[8:10]}日"
    return s


def _gsjc_for_wfgs(wfgs_name):
    d = run_sql(f"select gsjc from wfgs where wfgs='{wfgs_name}' limit 1")
    return get_field_value(d[0], "gsjc") if d else ""


def _sig_path(cpbh):
    if not cpbh:
        return None
    rows = run_sql(f"select tpmc from tpzx where (cpbh='{cpbh}') and (LENGTH(tpmc) > 5) limit 1")
    if not rows:
        return None
    src = _parse_tpzx_tpmc_first_src(rows[0].get("tpmc"))
    return _tpzx_src_to_local_path(src)


def _add_signature_or_text(ws, col_letter, row, img_path, text, watermark_text=None):
    coord = f"{col_letter}{row}"
    if img_path and os.path.isfile(img_path):
        if watermark_text:
            safe_write(ws, coord, watermark_text)
            rng = ws.range(coord)
            area = rng.merge_area
            rows, cols = area.shape
            cell = area(1, 1) if rows * cols > 1 else rng
            cell.api.Font.Color = 0xC0C0C0
            cell.api.Font.Size = 8
        _add_picture_fitted_to_cell(ws, coord, img_path, pad=2)
        return
    safe_write(ws, coord, text)
    _set_cell_center(ws, coord)


def _apply_invoice_row_borders(ws, row):
    """外销发票明细行边框（对照 Delphi Borders[7..10]）；表头行不调用。"""
    rng = ws.range(f"B{row}:E{row}").api
    for border_id in (7, 8, 9, 10):
        bd = rng.Borders(border_id)
        bd.LineStyle = 1
    for col in ("B", "C", "D"):
        ws.range(f"{col}{row}").api.Borders(10).LineStyle = 1


def _header_hsbm(row):
    if get_field_value(row, "sfrbmfy") == "是":
        return get_field_value(row, "hsbm")
    return ""


def _update_print_status(session, row, user_name):
    rid = get_field_value(row, "rid")
    sid = row.get("sid")
    today = time.strftime("%Y-%m-%d")
    dyrq = f"{user_name}:{today}"
    if rid:
        session.query(fysq).filter(fysq.rid == rid).update(
            {fysq.ywdy: "有", fysq.dyrq: dyrq}, synchronize_session=False
        )
    elif sid is not None:
        session.query(fysq).filter(fysq.sid == sid).update(
            {fysq.ywdy: "有", fysq.dyrq: dyrq}, synchronize_session=False
        )


def _save_workbook_as_pdf(wb, excel_path, save_dir, want_pdf, center_coords=None):
    os.makedirs(os.path.dirname(excel_path) or save_dir, exist_ok=True)
    wb.save(excel_path)
    wb.close()
    _openpyxl_center_cells(excel_path, center_coords)
    if not want_pdf:
        return excel_path, None
    pdf_res = convert_excel_to_pdf(excel_path, save_dir)
    if not pdf_res.get("success"):
        raise RuntimeError(pdf_res.get("error") or "PDF转换失败")
    return excel_path, pdf_res["pdf_path"]


def _export_csd(row, ctx):
    if "可思达" not in get_field_value(row, "wfgs"):
        return None
    tpl = os.path.join(ctx["base_path"], TEMPLATE_CSD)
    fkbh = get_field_value(row, "fkbh")
    cwgc = get_field_value(row, "cwgc")
    name = f"可思达{cwgc}付款审批单({fkbh})"
    excel_path = os.path.join(ctx["save_dir"], _safe_filename(name) + ".xlsx")

    app = _start_excel_app()
    wb = None
    pdf_path = None
    try:
        wb = app.books.open(tpl)
        sheet_names = [s.name for s in wb.sheets]
        ws = wb.sheets["审批单"] if "审批单" in sheet_names else wb.sheets[0]
        sprq1 = row.get("sprq1")
        if hasattr(sprq1, "strftime"):
            sprq1 = sprq1.strftime("%Y-%m-%d")
        else:
            sprq1 = get_field_value(row, "sprq1")
        safe_write(ws, "A3", sprq1)
        safe_write(ws, "B4", get_field_value(row, "cwgc"))
        safe_write(ws, "B5", get_field_value(row, "yhzh"))
        safe_write(ws, "B6", get_field_value(row, "khh"))
        seje1 = round(get_field_value(row, "seje", as_type=float), 2)
        safe_write(ws, "B10", seje1)
        _, pdf_path = _save_workbook_as_pdf(wb, excel_path, ctx["save_dir"], ctx["want_pdf"])
        wb = None
    finally:
        if wb is not None:
            wb.close()
        try:
            app.quit()
        except Exception:
            pass

    if pdf_path and os.path.isfile(excel_path):
        try:
            os.remove(excel_path)
        except OSError:
            pass
    return pdf_path or excel_path


def _record_eligible(row):
    """对照 Delphi: zjlhz=通过 且 hklx 为办公费用/快件费用。"""
    if get_field_value(row, "zjlhz") != "通过":
        return False
    return get_field_value(row, "hklx") in EXPRESS_OFFICE_HKLX


def _spbm_for_sqry(sqry):
    if not sqry:
        return ""
    d = run_sql(f"select spbm from ywrybiao where yhm='{sqry}' limit 1")
    return get_field_value(d[0], "spbm") if d else ""


def _load_dept_groups(pid):
    return run_sql(
        f"select sqbm, max(hsbm) as hsbm, sum(seje) as sejez, tjzg, sqry "
        f"from fysqsheet where pid='{pid}' "
        f"group by sqbm, tjzg, sqry order by tjzg"
    )


def _load_detail_lines(pid):
    return run_sql(f"select wxfp, ysfp, seje, sqry, khmc from fysqsheet where pid='{pid}' order by wxfp, tjzg")


def _preload_signatures(pid):
    rows = run_sql(f"select distinct tjzg from fysqsheet where pid='{pid}' and ifnull(tjzg,'')<>''")
    for r in rows:
        _sig_path(get_field_value(r, "tjzg"))


def _export_express_office_main(row, ctx):
    tpl = os.path.join(ctx["base_path"], TEMPLATE_EXPRESS_OFFICE)
    if not os.path.isfile(tpl):
        raise RuntimeError(f"模版文件缺失：{TEMPLATE_EXPRESS_OFFICE}")

    pid = get_field_value(row, "rid")
    _preload_signatures(pid)

    groups = _load_dept_groups(pid)
    i1 = len(groups)
    fkbh = get_field_value(row, "fkbh")
    name = f"{time.strftime('%Y-%m-%d')}优景快件表格({fkbh})"
    excel_path = os.path.join(ctx["save_dir"], _safe_filename(name) + ".xlsx")

    app = _start_excel_app()
    wb = None
    pdf_path = None
    try:
        wb = app.books.open(tpl)
        ws = wb.sheets[0]

        template_row = 4
        if i1 > 2:
            for i3 in range(1, i1 - 1):
                insert_row_at(ws, template_row + i3)
                copy_row_style(ws, template_row, template_row + i3)

        detail_end = 3 + i1 if i1 >= 1 else template_row
        if detail_end >= template_row:
            try:
                ws.range(f"A{template_row}:A{detail_end}").merge()
            except Exception:
                pass
            for r in range(template_row, detail_end + 1):
                _merge_main_amount_cell(ws, r)

        i = 3
        sig_cache = {}
        for grp in groups:
            i += 1
            _merge_main_amount_cell(ws, i)
            sejez = round(get_field_value(grp, "sejez", as_type=float), 2)
            _safe_write_center(ws, f"C{i}", sejez, "#,##0.##")

            tjzg = get_field_value(grp, "tjzg")
            if tjzg not in sig_cache:
                sig_cache[tjzg] = _sig_path(tjzg)
            _add_signature_or_text(ws, "E", i, sig_cache[tjzg], tjzg)
            sqbm = get_field_value(grp, "sqbm")
            safe_write(ws, f"B{i}", sqbm or _spbm_for_sqry(get_field_value(grp, "sqry")))

        if i1 <= 1:
            summary_row = detail_end + 3
        else:
            summary_row = detail_end + 2

        gsjc = ctx.get("gsjc") or _gsjc_for_wfgs(get_field_value(row, "wfgs"))
        wfgs = get_field_value(row, "wfgs")
        spny = _format_sqrq1(row.get("sqrq1"))
        fynr = get_field_value(row, "fynr")
        seje = round(get_field_value(row, "seje", as_type=float), 2)
        jbry = get_field_value(row, "jbry")

        safe_write(ws, "A1", f"{gsjc}快件与办公费用报销单")
        safe_write(ws, "A2", spny)
        _set_cell_center(ws, "A2")
        safe_write(ws, f"A{template_row}", fynr)
        _set_cell_center(ws, f"A{template_row}")

        chinese_str = chinese_amount(seje)
        if not chinese_str.endswith("整"):
            chinese_str += "整"
        safe_write(ws, f"B{summary_row}", chinese_str)
        _set_cell_center(ws, f"B{summary_row}")
        _merge_main_amount_cell(ws, summary_row)
        _safe_write_center(ws, f"C{summary_row}", seje, "#,##0.00")

        sig_row = summary_row + 1
        tjzjl = get_field_value(row, "tjzjl")
        tjcw = get_field_value(row, "tjcw")
        fkbh = get_field_value(row, "fkbh")
        _add_signature_or_text(ws, "B", sig_row, _sig_path(tjzjl), tjzjl, watermark_text=fkbh)
        _add_signature_or_text(ws, "D", sig_row, _sig_path(tjcw), tjcw, watermark_text=fkbh)
        safe_write(ws, f"E{sig_row}", f"报销人:{jbry}")
        _set_cell_center(ws, f"E{sig_row}")

        main_amount_coords = [f"C{r}" for r in range(template_row, detail_end + 1)]
        main_amount_coords.append(f"C{summary_row}")

        # 财务工厂 / 开户银行 / 银行帐号（签名行下方，空一行后）
        cwgc = get_field_value(row, "cwgc")
        khh = get_field_value(row, "khh")
        yhzh = get_field_value(row, "yhzh")

        # 适当加宽 B/D 列，防止长文本自动换行
        ws.range("B:B").api.ColumnWidth = 38
        ws.range("D:D").api.ColumnWidth = 32

        bank_info_row = sig_row + 2
        if cwgc:
            cell = ws.range(f"B{bank_info_row}")
            cell.value = f"财务工厂：{cwgc}"
            cell.api.WrapText = False

        if khh or yhzh:
            bank_info_row2 = bank_info_row + 1
            if khh:
                cell = ws.range(f"B{bank_info_row2}")
                cell.value = f"开户银行：{khh}"
                cell.api.WrapText = False
            if yhzh:
                cell = ws.range(f"D{bank_info_row2}")
                cell.value = f"银行帐号：{yhzh}"
                cell.api.WrapText = False
                cell.api.NumberFormat = "@"
            bank_info_row = bank_info_row2

        wx_header_row = bank_info_row + 2
        safe_write(ws, f"B{wx_header_row}", "外销发票")
        safe_write(ws, f"C{wx_header_row}", "申请金额")
        safe_write(ws, f"D{wx_header_row}", "申请人员")
        safe_write(ws, f"E{wx_header_row}", "客人名称")

        i = wx_header_row
        for det in _load_detail_lines(pid):
            i += 1
            wx = get_field_value(det, "ysfp") or get_field_value(det, "wxfp")
            safe_write(ws, f"B{i}", wx, "@")
            safe_write(ws, f"C{i}", round(get_field_value(det, "seje", as_type=float), 2))
            safe_write(ws, f"D{i}", get_field_value(det, "sqry"))
            safe_write(ws, f"E{i}", get_field_value(det, "khmc"))
            _apply_invoice_row_borders(ws, i)

        # 核算部门在外销发票明细下方，空一行（无标签，B列部门+C列金额）
        i += 2
        hsbm = _header_hsbm(row)
        if hsbm:
            safe_write(ws, f"B{i}", hsbm)
            safe_write(ws, f"C{i}", seje)

        _, pdf_path = _save_workbook_as_pdf(
            wb, excel_path, ctx["save_dir"], ctx["want_pdf"], center_coords=main_amount_coords
        )
        wb = None
    finally:
        if wb is not None:
            wb.close()
        try:
            app.quit()
        except Exception:
            pass

    if ctx["want_pdf"] and pdf_path and os.path.isfile(excel_path):
        try:
            os.remove(excel_path)
        except OSError:
            pass
    return pdf_path or excel_path


def collect_eligible_records(mode, rid, rids):
    keys = []
    seen = set()

    def add_row(row):
        key = get_field_value(row, "rid") or str(row.get("sid", ""))
        if not key or key in seen:
            return
        if not _record_eligible(row):
            return
        seen.add(key)
        keys.append(row)

    if mode == "2":
        id_list = rids if rids else ([rid] if rid else [])
        for one_rid in id_list:
            if not one_rid:
                continue
            d = run_sql(f"select * from fysq where rid='{one_rid}' limit 1")
            if d:
                add_row(d[0])
    else:
        if not rid:
            return []
        d = run_sql(f"select * from fysq where rid='{rid}' limit 1")
        if d:
            add_row(d[0])
    return keys


def _export_one_record(row, ctx, session, user_name):
    ctx = dict(ctx)
    ctx["gsjc"] = _gsjc_for_wfgs(get_field_value(row, "wfgs"))
    output_files = []

    if "可思达" in get_field_value(row, "wfgs") and ctx.get("has_csd"):
        try:
            p = _export_csd(row, ctx)
            if p:
                output_files.append(p)
        except Exception:
            logger.error(trace_error())

    try:
        p = _export_express_office_main(row, ctx)
        if p:
            output_files.append(p)
    except Exception:
        logger.error(trace_error())
        raise

    if output_files:
        _update_print_status(session, row, user_name)
    return output_files


@any_route("/api/saier/express/office/expense/export", methods=["POST", "GET"])
@require_token
async def view_saier_express_office_expense_export(request):
    """
    快件与办公费用报表导出。

    请求 JSON：
    - mode: ``1`` 当前单条（默认），``2`` 批量（传 rids）
    - rid / rids: 记录 rid
    - tp: ``1`` 仅 PDF（默认），``2`` 批量 PDF（打包 zip）
    """
    s = Session()
    user = request.current_user
    j = await request.json()

    try:
        mode = j.get("mode", "1")
        if mode != "2":
            mode = "1"
        tp = j.get("tp", "1")
        if tp != "2":
            tp = "1"
        rid = j.get("rid", "")
        rids = j.get("rids") or []
        if isinstance(rids, str):
            rids = [rids] if rids else []

        records = collect_eligible_records(mode, rid, rids)
        if not records:
            return json_result(-1, "没有符合导出条件的费用申请（需总经理审批通过，且货款类型为办公费用或快件费用）")

        base_path = os.path.join(config.data_upload_path, "template")
        save_dir = config.tmp_path
        os.makedirs(save_dir, exist_ok=True)

        for tpl in (TEMPLATE_EXPRESS_OFFICE, TEMPLATE_CSD):
            if not os.path.isfile(os.path.join(base_path, tpl)):
                return json_result(-1, f"模版文件缺失：{tpl}")

        has_csd = any("可思达" in get_field_value(r, "wfgs") for r in records)
        ctx = {"base_path": base_path, "save_dir": save_dir, "want_pdf": True, "has_csd": has_csd}

        output_files = []
        for row in records:
            paths = _export_one_record(row, ctx, s, user.username)
            for path in paths:
                if path and os.path.isfile(path):
                    output_files.append({"name": os.path.basename(path), "path": path})

        if not output_files:
            return json_result(-1, "未生成任何报表，请检查模版与数据")

        if len(output_files) == 1:
            s.commit()
            return json_result(1, "导出成功", output_files[0]["name"])

        zip_name = time.strftime("%Y%m%d%H%M%S") + "_express_office_expense.zip"
        zip_path = os.path.join(save_dir, zip_name)
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for f in output_files:
                zf.write(f["path"], f["name"])

        s.commit()
        return json_result(1, "导出成功", zip_name)

    except Exception:
        logger.error(trace_error())
        s.rollback()
        return json_result(-1, trace_error())
    finally:
        s.close()
