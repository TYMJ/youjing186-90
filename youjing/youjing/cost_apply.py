
from any import *
from .model import *
from .__default__ import get_user_path, module_share_new, module_xxck_new, user_task_delete, user_task_new
import os,math,re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU, pixels_to_points
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from sqlalchemy import text, func
from decimal import Decimal,ROUND_HALF_UP
from typing import List, Dict, Any
from datetime import datetime, timedelta
import calendar


"""
费用申请.货款类型.change
对应原Pascal: 货款类型.change
"""
@any_route('/api/saier/cost_apply/hklx/change', methods=['POST'])
@require_token
async def view_cost_apply_hklx_change(request):
    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        hklx = j.get('hklx', '')
        code = 1
        msg = '操作成功'
        data = {}
        d = s.query(zx.mc,zx.wb1,zx.wb10).filter(zx.ly=='费用类型', zx.mc==hklx).first()
        if not d:
            code = -1
            msg = '货款类型错误,请重新选择(必需从下拉框中选)'
            return json_result(code, msg)
        data = {
            'mc': d.mc,
            'wb1': d.wb1,
            'wb10': d.wb10
        }
        return json_result(code, msg, data)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'货款类型校验失败: {str(e)}')
    finally:
        s.close()


"""
费用申请.经办人员.change
对应原Pascal: 经办人员.change
"""
@any_route('/api/saier/cost_apply/jbry/change', methods=['POST'])
@require_token
async def view_cost_apply_jbry_change(request):
    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        jbry = j.get('jbry', '')
        code = 1
        msg = '操作成功'
        data = ''
        d = s.query(ywrybiao.bm).filter(ywrybiao.yhm==jbry).first()
        if not d:
            code = -1
            msg = '无此人员,请重新输入'
            return json_result(code, msg)
        
        data = d.bm
        return json_result(code, msg, data)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'经办人员校验失败: {str(e)}')
    finally:
        s.close()


"""
费用申请.外销发票号.change
对应原Pascal: 外销发票号.change
"""
@any_route('/api/saier/cost_apply/fphm/change', methods=['POST'])
@require_token
async def view_cost_apply_fphm_change(request):
    s = Session()
    data = {'htje': 0.0, 'khrmbhj': 0.0, 'khpd': '', 'khmc': '', 'RMBkh': '', 'ysfp': '', 'khmc': '', 'RMBkh': '', 'ysfp': '', 'cymx': 0}
    try:
        user = request.current_user
        j = await request.json()
        fphm = j.get('fphm', '')
        flag = j.get('flag', 0)
        code = 1
        msg = '操作成功'

        if flag == 1:
            d = s.query(cymx.khmc, cymx.RMBkh, cymx.ysfp).filter(((cymx.fphm == fphm) | (cymx.ysfp == fphm))).first()
            if d:
                data.update({
                    'khmc': d.khmc,
                    'RMBkh': d.RMBkh if d.RMBkh is not None else '',
                    'ysfp': d.ysfp,
                    'cymx': 1
                })

        d = s.query(fpgl.rid).filter(((fpgl.wxfp == fphm) | (fpgl.hsfp == fphm)), fpgl.sfjd == "是").first()
        if d:
            code = -1
            msg = '请注意此发票已结算,不能申请,请与财务联系'
            return json_result(code, msg, data)
        d = s.query(wxht.htje,wxht.khrmbhj,wxht.khpd).filter((wxht.order_id == fphm)).first()
        if d:
            data = {
                'htje': float(d.htje) if d.htje is not None else 0.0,
                'khrmbhj': float(d.khrmbhj) if d.khrmbhj is not None else 0.0,
                'khpd': d.khpd
            }

        c = s.query(cymx.khmc, cymx.RMBkh, cymx.ysfp).filter((cymx.fphm == fphm) | (cymx.ysfp == fphm)).first()
        if c:
            data.update({
                'khmc': c.khmc,
                'RMBkh': c.RMBkh if c.RMBkh is not None else '',
                'ysfp': c.ysfp
            })
        
        return json_result(code, msg, data)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'外销发票号校验失败: {str(e)}', data)
    finally:
        s.close()


"""
费用申请.费用收回.change
对应原Pascal: 费用收回.change
"""
@any_route('/api/saier/cost_apply/fysh/change', methods=['POST'])
@require_token
async def view_cost_apply_fysh_change(request):
    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        bfdh = j.get('bfdh', '')
        fysh = j.get('fysh', '')
        code = 1
        msg = '操作成功'
        if bfdh == '' or fysh == '':
            code = 0
            msg = '无需更新'
            return json_result(code, msg)
        
        d = s.query(gckm).filter(gckm.cpbh == bfdh).first()
        if d:
            d.cdfysh = fysh
            d.djcw = user.username
            d.shcw = user.username if fysh == '已收回' else ''
            s.add(d)
            if fysh == '已收回':
                res = user_task_delete('工厂开模', d.rid, s, [])
                if res.get('code') != 1:
                    s.rollback()

            s.commit() 
        return json_result(code, msg, data)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'经办人员校验失败: {str(e)}')
    finally:
        s.close()


"""
费用申请.费用详情.insert
对应原Pascal: 费用详情.insert
"""
@any_route('/api/saier/cost_apply/child/new', methods=['POST'])
@require_token
async def view_cost_apply_child_new(request):
    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        hklx = j.get('hklx', '')
        code = 1
        msg = '操作成功'
        data = ''
        d = s.query(zx.mc,zx.wb1,zx.wb10).filter(zx.ly=='费用类型', zx.mc==hklx).first()
        if d:
            data = d.wb1

        return json_result(code, msg, data)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'经办人员校验失败: {str(e)}')
    finally:
        s.close()

"""
费用申请.费用详情.申请部门.change
对应原Pascal: 费用详情.申请部门.change
"""
@any_route('/api/saier/cost_apply/sqbm/change', methods=['POST'])
@require_token
async def view_cost_apply_sqbm_change(request):
    s = Session()
    try:
        user = request.current_user
        j = await request.json()
        sqbm = j.get('sqbm', '')
        code = 1
        msg = '操作成功'
        data = ''
        d = s.query(ywrybiao.fysp).filter(ywrybiao.fysp!='侯柳红', ywrybiao.spbm==sqbm).first()
        if d:
            data = d.fysp

        return json_result(code, msg, data)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'经办人员校验失败: {str(e)}')
    finally:
        s.close()


"""
费用申请.记录保存前
对应原Pascal: 费用详情审批.主管确认.change
"""
@any_route('/api/saier/cost_apply/save/before', methods=['POST'])
@require_token
async def view_cost_apply_save_before(request):
    s = Session()
    try:
        j = await request.json()
        user = request.current_user
        x = j.get('main', {})
        ysfp = x.get('ysfp', '')
        sccj = x.get('sccj', '')
        rid = x.get('rid', '')
        kfdy = x.get('kfdy', '')
        lines = j.get('lines', []) # 费用详情审批列表
        costs = j.get('costs', []) # 费用详情列表
        jbry = x.get('jbry', '')
        jbpath = ''
        cwsb = ''
        if jbry != '' and jbry is not None:
            jbpath = get_user_path(jbry).get('path', '')
        org = get_user_path(user.username)
        position = org.get('position', '')
        if '财务' in position:
            cwsb = '1'
        elif '综合' in position:
            cwsb = '2'

        code = 1
        msg = '操作成功'
        flag = 0
        spzt = 1 # 当前用户待审批的是否完成
        zdzt = 1 # 整个费用申请是否有待审批记录
        tjzg = ''
        rid_json = {}
        data = {'rid_json': rid_json, 'flag': flag, 'zdzt': zdzt, 'gsmc': '','jbpath': jbpath, 'cwsb': cwsb}

        if (x.get('jlhz')!='待审批' and x.get('tjzg') == user.username):
            c = s.query(fysq.fkbh).filter(fysq.rid == x.get('rid'), func.ifnull(fysq.tjjl, '') == '', fysq.jbry != user.username).first()
            if c:
                return json_result(-1, '业务撤单，不能保存', data)
            
        if x.get('tjzg', '') != '' and x.get('tjzg', '') is not None and (x.get('tjfk', '') == '' or x.get('tjfk', '') == None) and (x.get('tjzjl', '') =='' or x.get('tjzjl', '') == None) and (x.get('tjcw', '') =='' or x.get('tjcw', '') == None):
            d = s.query(zx.wb2, zx.wb3, zx.wb4, zx.wb5, zx.wb6, zx.wb7).filter(zx.ly == '付款审批', zx.wb1 == x.get('tjzg')).first()
            if d:
                data.update({
                    'tjfk': d.wb2,
                    'tjzjl': d.wb3,
                    'tjcw': d.wb6,
                    'tjzg': d.wb5
                })

        for l in lines:
            tjzg = l.get('tjzg', '') if l.get('tjzg', '') else ''
            zgqr = l.get('zgqr', '') if l.get('zgqr', '') else ''
            fywyzd = l.get('fywyzd', '') if l.get('fywyzd', '') else ''
            a = [r for r in costs if r.get('fywyzd') == fywyzd and r.get('tjzg') == tjzg]
            for r in a:
                if zgqr != r.get('zgqr'):
                    r['zgqr'] = zgqr
                    rid_json[r.get('rid')] = zgqr
                    flag = 1
                if zgqr != '通过':
                    spzt = 0

        # 如果是当前用户无待审批记录了,则删除待审批任务
        if spzt == 1 and flag == user.username:
            res = user_task_delete('费用申请', rid, s, [user.username])
            if res.get('code') != 1:
                s.rollback()
                return json_result(res.get('code'), res.get('msg'))
        
        for r in costs:
            if r.get('zgqr') == '待审批' and r.get('tjzg') != None and r.get('tjzg') != '':
                zdzt = 0
                data['zdzt'] = 0
                break

        if zdzt == 1:
            xxnr = str(user.username) + '费用申请编号:' + str(x.get('fkbh', '')) + '全通过,日期:' + time.strftime('%Y-%m-%d')
            res = user_task_new('费用申请', rid, '付款单号', '费用申请审批全通过', xxnr, user, s, [x.get('jbry')])
            if res.get('code') != 1:
                s.rollback()
                return json_result(res.get('code'), res.get('msg'))
            
            row = {
                "xxly": '费用申请',
                "wxht": '',
                "gdht": str(x.get('fkbh', '')),
                "yhdh": '',
                "xxnr": xxnr,
                "jsr": str(x.get('jbry')),
                "sys_path": "",
                "spsq": user.username
            }
            res = module_xxck_new([row],user,s)
            if res.get('code')!=1:
                s.rollback()
                return json_result(res.get('code'), res.get('code'))
        if x.get('sfkp') == '是':
            c = s.query(cyzglsheet.rid).filter(cyzglsheet.xm == x.get('khmc'), func.ifnull(cyzglsheet.xm, '') != '', cyzglsheet.zm=="我方公司不锁定").first()
            if not c:
                a = s.query(kh.Vendorgs).filter(kh.company_name == x.get('khmc'), func.ifnull(kh.Vendorgs, '') != '').first()
                if a:
                    data['gsmc'] = a.Vendorgs

        s.commit()

        return json_result(code, msg, data)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'保存校验失败: {str(e)}')
    finally:
        s.close()


"""
费用申请.记录保存后
对应原Pascal: 查看人员
"""
@any_route('/api/saier/cost_apply/save/after', methods=['POST'])
@require_token
async def view_cost_apply_save_after(request):
    s = Session()
    try:
        j = await request.json()
        user = request.current_user
        rid = j.get('rid')
        x = j.get('main', {})
        d = s.query(fysq).filter(fysq.rid == rid).first()

        if not d:
            return json_result(0, '无需处理')
        fields = [
            '经办人员', '提交主管', '提交总监', '提交总经理', '提交财务', '领用人员', '打印财务', '侯柳红'
        ]
        user_list = []
        m = get_module('费用申请')
        for field in fields:
            f = m.field_by_full_name('费用申请.' + field)
            if not f:
                continue
            name = f.db.name
            if hasattr(d, name) and getattr(d, name) and getattr(d, name) not in user_list:
                user_list.append(getattr(d, name))
        d = s.query(fysqsheet.tjzg).filter(fysqsheet.pid == rid, func.ifnull(fysqsheet.tjzg, '') != '').group_by(fysqsheet.tjzg).all()
        for u in d:
            if u.tjzg and u.tjzg not in user_list:
                user_list.append(u.tjzg)

        res = module_share_new('费用申请',user_list, rid, user, s, ['all'])
        if res.get('code')!=1:
            s.rollback()
            return json_result(res.get('code'),res.get('msg'))

        s.query(gckm).filter(gckm.cpbh == x.get('fkbh')).update({
            gckm.zj: '',
            gckm.zjhz1: '待审批',
            gckm.zjhzrq: ''
        }, synchronize_session=False)
        s.commit()

        return json_result(1, '操作成功')
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'保存校验失败: {str(e)}')
    finally:
        s.close()


"""
费用申请.记录加载后
对应原Pascal: 修改清单查看
"""
@any_route('/api/saier/cost_apply/load/check', methods=['POST'])
@require_token
async def view_cost_apply_load_check(request):
    s = Session()
    try:
        j = await request.json()
        gsmc = j.get('gsmc', '')
        user = request.current_user
        data = {'position': '', 'wfgs': gsmc}
        org = get_user_path(user.username)
        position = org.get('position', '')
        code = 1
        msg = '操作成功'
        if gsmc == '' or gsmc is None:
            c = s.query(ywrybiao.wfgs,ywrybiao.bm).filter(ywrybiao.yhm == user.username).first()
            if c:
                data['wfgs'] = c.wfgs
                data['bm'] = c.bm

        return json_result(code, msg, data)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'保存校验失败: {str(e)}')
    finally:
        s.close()

"""
费用申请.刷新已出按钮
对应原Pascal: 刷新已出
"""
@any_route('/api/saier/cost_apply/update_sfyc', methods=['POST'])
@require_token
async def view_cost_apply_update_sfyc(request):
    s = Session()
    try:
        j = await request.json()
        user = request.current_user
        code = 1
        msg = '操作成功'
        d = s.query(fysqsheet).filter(func.ifnull(fysqsheet.wxfp, '') != '', fysqsheet.sfyc != '是').all()
        for r in d:
            c = s.query(cymx.rid).filter(cymx.fphm == r.wxfp, func.ifnull(cymx.sjcy1, '') != '').first()
            if not c:
                continue
            r.sfyc = '是'
            r.mtime = time.strftime('%Y-%m-%d %H:%M:%S')
            r.modi_uid = user.rid
            s.add(r)

        s.commit()
        return json_result(code, msg)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'保存校验失败: {str(e)}')
    finally:
        s.close()


"""
费用申请.撤销审批按钮
对应原Pascal: 撤销审批
"""
@any_route('/api/saier/cost_apply/update_audit', methods=['POST'])
@require_token
async def view_cost_apply_update_audit(request):
    s = Session()
    try:
        j = await request.json()
        rid = j.get('rid', '')
        user = request.current_user
        code = 1
        msg = '操作成功'
        d = s.query(fysq).filter(func.ifnull(fysq.jlhz, '') == '待审批', fysq.rid == rid).first()
        if not d:
            return json_result(-1, '已审批的记录不能撤销')
        if d.jbry != user.username:
            return json_result(-1, '只有提交人可以撤销审批')
        
        d.tjjl=""
        d.tjfk=""
        d.tjzjl=""
        d.tjcw=""
        d.cwsp="待审批"
        d.zjlhz="待审批"
        d.zjhz="待审批"
        d.mtime = time.strftime('%Y-%m-%d %H:%M:%S')
        d.modi_uid = user.rid
        s.add(d)
        res = user_task_delete('费用申请', rid, s, [])
        if res.get('code') != 1:
            s.rollback()
            return json_result(res.get('code'), res.get('msg'))
        
        s.commit()
        return json_result(code, msg)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'撤销审批失败: {str(e)}')
    finally:
        s.close()



"""
费用申请.生成预付货款按钮
对应原Pascal: 生成预付货款
"""
@any_route('/api/saier/cost_apply/advance_payment_new', methods=['POST'])
@require_token
async def view_cost_apply_advance_payment_new(request):
    s = Session()
    try:
        j = await request.json()
        rid = j.get('rid', '')
        user = request.current_user
        code = 1
        msg = '操作成功'
        org = get_user_path(user.username)
        position = org.get('position', '')
        if '财务' not in position :
            return json_result(-1, '只有财务或综合岗位可以生成预付货款')
        
        d = s.query(fysq).filter(fysq.rid == rid).first()
        if not d:
            return json_result(-1, '请选择需要生成预付款的记录')
        
        if d.fkbh == '' or d.fkbh is None or d.tjcw != user.username or d.cwsp != '通过':
            return json_result(-1, '只有提交财务且财务审批通过的记录可以生成预付款')

        fkbh = 'yf' + str(d.fkbh)
        c = s.query(yfhk.rid).filter(yfhk.fkbh == fkbh).first()
        if c:
            return json_result(-1, '付款编号已有预付款记录,请勿重复生成')
        m = yfhk()
        x = alchemy_object_to_dict(d)
        for k, v in x.items():
            if hasattr(m, k):
                setattr(m, k, v)
        pid = get_uuid()
        m.rid = pid
        m.fkbh = fkbh
        m.uid = user.rid
        m.ctime = time.strftime('%Y-%m-%d %H:%M:%S')
        m.wxht = d.wxfp
        m.wxfp = d.fphm
        m.cght = d.cght
        m.htje = d.fpje
        m.sqrq = d.sqrq1
        m.sqje = d.seje
        m.fkbz = d.hbdm
        m.chrq = d.yjcq
        m.jsrm = d.jbry
        m.sccj = d.cwgc
        m.gcmc1 = d.csmc
        m.yhzh = d.yhzh
        m.khh = d.khh
        m.gcbh = d.gcbh
        m.bzsm = d.bz1
        m.fkdq = None
        m.zwpm = d.zwpm
        u = s.query(ywrybiao.bm).filter(ywrybiao.yhm == d.jbry).first()
        if u:
            m.ywbm = u.bm
        else:
            m.ywbm = None
        m.rxfs = None
        m.hklx = d.hklx
        m.fkxs = d.fkxs
        m.wyzd = m.rid
        m.dlrq = time.strftime('%Y-%m-%d')
        m.sb = m.rid
        s.add(m)
        c = s.query(fysqsheet).filter(fysqsheet.pid == rid).all()
        for r in c:
            y = gcdjsheet()
            y.rid = get_uuid()
            y.pid = pid
            y.fphm = ''
            y.cght = r.cght
            y.sydje = r.seje
            y.cphh = r.cpbh
            y.zwpm = r.zwpm
            y.dj = 0
            y.dw = ''
            y.bz1 = ''
            y.xs = 0
            y.zsl = 0
            y.gcmc = r.csmc
            y.rxfs = ''
            y.rxr = ''
            y.sjhm = ''
            y.bjhh = ''
            y.gchh = ''
            y.fkbh = fkbh
            s.add(y)
        
        s.commit()
        return json_result(code, msg, pid)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'撤销审批失败: {str(e)}')
    finally:
        s.close()


"""
费用申请.详情审批提交按钮
对应原Pascal: 详情审批提交
"""
@any_route('/api/saier/cost_apply/detail_audit_apply', methods=['POST'])
@require_token
async def view_cost_apply_detail_audit_apply(request):
    s = Session()
    try:
        j = await request.json()
        rid = j.get('rid', '')
        kind = j.get('kind', '正常提交')
        user = request.current_user
        fkbh = j.get('fkbh', '')
        jbry = j.get('jbry', '')
        code = 1
        msg = '操作成功'
        filters = []
        if kind != '正常提交':
            filters.append(or_(fysqsheet.fssb == '否', func.ifnull(fysqsheet.fssb, '') == ''))
        d = s.query(fysqsheet).filter(fysqsheet.pid == rid, fysqsheet.tjzg != None, fysqsheet.tjzg != '').filter(*filters).all()
        if not d:
            return json_result(-1, '请选择需要提交审批的记录，未找到符合条件的费用详情记录')
        xx = str(user.username) + '的费用审请:' + str(fkbh) + '需审批,日期:' + str(time.strftime('%Y-%m-%d'))
        i = 0
        user_list = []
        for r in d:
            i = i + 1
            tjzg = r.tjzg if r.tjzg else ''
            if tjzg == '' or tjzg is None:
                continue
            if tjzg in user_list:
                r.fssb = '是'
                s.add(r)
                continue
            user_list.append(tjzg)
            row = {
                "xxly": '费用申请',
                "wxht": '',
                "gdht": str(fkbh),
                "yhdh": '',
                "xxnr": xx,
                "jsr": str(tjzg),
                "sys_path": "",
                "spsq": user.username
            }
            res = module_xxck_new([row], user, s)
            if res.get('code') != 1:
                s.rollback()
                return json_result(res.get('code'), res.get('msg'))
            r.fssb = '是'
            s.add(r)
        
        if len(user_list) > 0:
            res = user_task_new('费用申请', rid, '付款单号', '费用申请审批', xx, user, s, user_list, ['all'])
            if res.get('code') != 1:
                s.rollback()
                return json_result(res.get('code'), res.get('msg'))
            s.query(fysq).filter(fysq.rid == rid).update({
                fysq.xqspzt: '没通过'
            }, synchronize_session=False)
    
        s.commit()
        return json_result(code, msg)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'撤销审批失败: {str(e)}')
    finally:
        s.close()


"""
费用申请.主管审批
对应原Pascal: 主管审批
"""
@any_route('/api/saier/cost_apply/zgsp/change', methods=['POST'])
@require_token
async def view_cost_apply_zgsp_change(request):
    s = Session()
    try:
        j = await request.json()
        rid = j.get('rid', '')
        hklx = j.get('hklx', '')
        tjzg = j.get('tjzg', '')
        user = request.current_user
        code = 1
        msg = '操作成功'
        cxje = 0
        d = s.query(zx.cs).filter(zx.mc.like('%' + hklx + '%'), zx.ly == '费用审请额度', zx.wb1 == tjzg).first()
        if d:
            cxje = float(d.cs) if d.cs is not None else 0.0

        return json_result(code, msg, cxje)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f'撤销审批失败: {str(e)}')
    finally:
        s.close()


"""
费用申请.财务识别
对应原Pascal: 财务识别
"""
@any_route('/api/saier/cost_apply/cwsp/change', methods=['POST'])
@require_token
async def view_cost_apply_cwsp_change(request):
    s = Session()
    try:
        j = await request.json()
        rid = j.get('rid', '')
        hklx = j.get('hklx', '')
        cwsp = j.get('cwsp', '')
        tjcw = j.get('tjcw', '')
        cwyj = j.get('cwyj', '')
        fkbh = j.get('fkbh', '')
        jbry = j.get('jbry', '')
        khmc = j.get('khmc', '')
        gsmc = j.get('gsmc', '')
        wxfp = j.get('wxfp', '')
        lines = j.get('lines', [])
        user = request.current_user
        cwzj = ''
        sb1 = ''
        cw1 = ''
        sqje = j.get('sqje', 0.0)
        code = 1
        msg = '操作成功'
        stnumber = ''
        k = s.query(fysq.rid).filter(fysq.rid == rid).first()
        if k:
            stnumber = k.rid
        if stnumber != '' and stnumber != None:
            res = user_task_delete('费用申请', rid, s, [])
            if res.get('code') != 1:
                s.rollback()
                return json_result(res.get('code'), res.get('msg'))
            
        if khmc != '' and khmc is not None:
            khmc = khmc.upper()
        if cwsp == '不通过':
            xxnr = str(user.username) + '付款审批:' + str(fkbh) + '财务不能通过,没通过原因:' + str(cwyj)

            res = user_task_new('费用申请', rid, '付款单号', '费用申请审批未通过', xxnr, user, s, [jbry])
            if res.get('code') != 1:
                s.rollback()
                return json_result(res.get('code'), res.get('msg'))
            t = {"type":"success","title":"财务不能通过通知",
                "msg": xxnr,
                "module": "费用申请",
                "rid": rid, "rids":[],
                "btns":[{"title":"打开","icon":"any-rmb-full","name":"open","color":"red"}]}
            await messages.message_to_user(jbry,t,MSG_KIND_NOTICE_RECORD,request)

        if cwsp == '通过' or cwsp == '通过返回':
            d = s.query(zx.cs).filter(zx.mc == hklx, zx.ly == '费用类型').first()
            if d and d.cs and float(d.cs) > 0 and sqje > float(d.cs):
                sb1 = '1'
            if 'AMAZON' in khmc or 'AMZ' in khmc or gsmc == '宁波景业国际贸易有限公司':
                c = s.query(zx.cwzj).filter(zx.wb8 == tjcw, zx.ly == '付款审批').first()
                if c and c.cwzj:
                    cwzj = c.cwzj
            elif sb1 == '1':
                c = s.query(zx.wb7).filter(zx.wb6 == tjcw, zx.ly == '付款审批').first()
                if c and c.wb7:
                    cwzj = c.wb7
            # if cwzj == '' or cwzj == None:
            d = s.query(zxsheet.dymc).filter(zxsheet.ly == '付款审批', zxsheet.sj == user.username, zxsheet.lx == '费用审批财务', zxsheet.xtxh > 0).first()
            if d and d.dymc:
                cwzj = d.dymc
 
            if (cwzj == '' or cwzj == None) or (cwzj != user.username and cwzj != '' and cwzj is not None):
                xxnr = '费用申请:' + str(fkbh) + '财务通过请查看,日期:' + str(time.strftime('%Y-%m-%d'))
                row = {
                    "xxly": '费用申请',
                    "wxht": '',
                    "gdht": str(fkbh),
                    "yhdh": '',
                    "xxnr": xxnr,
                    "jsr": str(jbry),
                    "sys_path": "",
                    "spsq": user.username
                }
                res = module_xxck_new([row],user,s)
                if res.get('code')!=1:
                    s.rollback()
                    return json_result(res.get('code'), res.get('code'))
                for l in lines:
                    if l.get('wxfp', '') == None or l.get('wxfp', '') == '':
                        continue
                    f = s.query(fpgl).filter(fpgl.wxfp == l.get('wxfp'), fpgl.sfjd=="否").update({
                        fpgl.webpdfy: '是'
                    }, synchronize_session=False)

                if 'AMAZON' in khmc  or 'AMZ' in wxfp or gsmc == '宁波景业国际贸易有限公司':
                    c = s.query(zx.wb5,zx.wb6,zx.wb7,zx.wb8).filter(zx.wb8 == tjcw, zx.ly == '电商付款审批付款财务').first()
                    if c:
                        if wxfp != '' and wxfp is not None:
                            if c.wb5 == '发票':
                                cw1 = c.wb6
                        else:                            
                            if c.wb7 == '部门':
                                cw1 = c.wb8
                else:
                    c = s.query(zx.wb5,zx.wb6,zx.wb7,zx.wb8).filter(zx.wb8 == tjcw, zx.ly == '付款审批付款财务').first()
                    if c:
                        if wxfp != '' and wxfp is not None:
                            if c.wb5 == '发票':
                                cw1 = c.wb6
                        else:                            
                            if c.wb7 == '部门':
                                cw1 = c.wb8

                if cw1 != '' and cw1 != None and stnumber != '' and stnumber != None:
                    xxnr = str(user.username) + '费用申请:' + str(fkbh) + '通过请安排'
                    row = {
                        "xxly": '费用申请',
                        "wxht": '',
                        "gdht": str(fkbh),
                        "yhdh": '',
                        "xxnr": xxnr,
                        "jsr": str(cw1),
                        "sys_path": "",
                        "spsq": user.username
                    }
                    res = module_xxck_new([row],user,s)
                    if res.get('code')!=1:
                        s.rollback()
                    res = user_task_new('费用申请', rid, '付款单号', '费用申请审批', xxnr, user, s, [cw1])
                    if res.get('code') != 1:
                        s.rollback()
                        return json_result(res.get('code'), res.get('msg'))

                if jbry != '' and jbry is not None and jbry != user.username:
                    xxnr = '费用申请:' + fkbh + '财务通过请查看,日期:' + str(time.strftime('%Y-%m-%d'))
                    t = {"type":"success","title":"财务审批通过通知",
                        "msg": xxnr,
                        "module": "费用申请",
                        "rid": rid, "rids":[],
                        "btns":[{"title":"打开","icon":"any-rmb-full","name":"open","color":"red"}]}
                    await messages.message_to_user(jbry,t,MSG_KIND_NOTICE_RECORD,request)

        s.commit()
        return json_result(code, msg, {'cwzj': cwzj, 'sb1': sb1, 'cw1': cw1})
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'撤销审批失败: {str(e)}')
    finally:
        s.close()


"""
费用申请.提交总监
对应原Pascal: 提交总监
"""
@any_route('/api/saier/cost_apply/tjzj/change', methods=['POST'])
@require_token
async def view_cost_apply_tjzj_change(request):
    s = Session()
    try:
        j = await request.json()
        rid = j.get('rid', '')
        tjzj = j.get('tjzj', '')
        zgsp = j.get('zgsp', '')
        wysp = j.get('wysp', '')
        fkbh = j.get('fkbh', '')
        user = request.current_user
        code = 1
        msg = '操作成功'
        cxje = 0
        yzj = ''
        d = s.query(fysq.tjfk).filter(fysq.rid == rid).first()
        if d and d.tjfk:
            yzj = d.tjfk
        
        if yzj != tjzj and tjzj != '' and tjzj is not None:
            d = s.query(zx.wb2).filter(zx.ly == '付款审批', zx.wb2 == tjzj).first()
            if not d:
                return json_result(-1, '此人没有审批权限,请重新选择')
            if zgsp == '通过':
                    # return json_result(-1, '请先提交主管审批通过再选择总监审批人')
                if tjzj == user.username:
                    res = user_task_delete('费用申请', rid, s, [])
                    if res.get('code') != 1:
                        s.rollback()
                        return json_result(res.get('code'), res.get('msg'))
                else:
                    if wysp != tjzj and tjzj != '' and tjzj is not None:
                        if tjzj != user.username:
                            xxnr = str(user.username) + '的费用申请:' + str(fkbh) + '需审批,日期:' + str(time.strftime('%Y-%m-%d'))
                            row = {
                                "xxly": '费用申请',
                                "wxht": '',
                                "gdht": str(fkbh),
                                "yhdh": '',
                                "xxnr": xxnr,
                                "jsr": str(tjzj),
                                "sys_path": "",
                                "spsq": user.username
                            }
                            res = module_xxck_new([row], user, s)
                            if res.get('code') != 1:
                                s.rollback()
                                return json_result(res.get('code'), res.get('msg'))
                            res = user_task_delete('费用申请', rid, s, [])
                            if res.get('code') != 1:
                                s.rollback()
                                return json_result(res.get('code'), res.get('msg'))
                            res = user_task_new('费用申请', rid, '付款单号', '费用申请审批', xxnr, user, s, [tjzj], ['all'])
                            if res.get('code') != 1:
                                s.rollback()
                                return json_result(res.get('code'), res.get('msg'))
                            
                            t = {"type":"success","title":"财务审批通过通知",
                            "msg": xxnr,
                            "module": "费用申请",
                            "rid": rid, "rids":[],
                            "btns":[{"title":"打开","icon":"any-rmb-full","name":"open","color":"red"}]}
                        await messages.message_to_user(tjzj, t, MSG_KIND_NOTICE_RECORD, request)

            s.commit()
        return json_result(code, msg, cxje)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'撤销审批失败: {str(e)}')
    finally:
        s.close()


"""
费用申请.提交财务
对应原Pascal: 提交财务
"""
@any_route('/api/saier/cost_apply/tjcw/change', methods=['POST'])
@require_token
async def view_cost_apply_tjcw_change(request):
    s = Session()
    try:
        j = await request.json()
        rid = j.get('rid', '')
        tjcw = j.get('tjcw', '')
        zjsp = j.get('zgsp', '')
        fkbh = j.get('fkbh', '')
        user = request.current_user
        code = 1
        msg = '操作成功'
        cxje = 0
        ycw = ''
        d = s.query(fysq.tjcw).filter(fysq.rid == rid).first()
        if d and d.tjcw:
            ycw = d.tjcw
        if tjcw == user.username:
            return json_result(-1, '自已不能审批自已,请重新选择')
        if tjcw == ycw and tjcw != '' and tjcw is not None:
            return json_result(0, '财务和数据库一致，无需修改')
        d = s.query(zx.wb5,zx.wb4).filter(zx.ly == '付款审批', or_(zx.wb6 == tjcw, zx.wb7 == tjcw, zx.wb8 == tjcw, zx.wb9 == tjcw)).first()
        if not d:
            return json_result(-1, '此人没有审批权限,请重新选择')
        
        if zjsp == '通过':
            if tjcw != user.username:
                xxnr = str(user.username) + '的费用申请:' + str(fkbh) + '需审批,日期:' + str(time.strftime('%Y-%m-%d'))
                row = {
                    "xxly": '费用申请',
                    "wxht": '',
                    "gdht": str(fkbh),
                    "yhdh": '',
                    "xxnr": xxnr,
                    "jsr": str(tjcw),
                    "sys_path": "",
                    "spsq": user.username
                }
                res = module_xxck_new([row], user, s)
                if res.get('code') != 1:
                    s.rollback()
                    return json_result(res.get('code'), res.get('msg'))
                res = user_task_delete('费用申请', rid, s, [])
                if res.get('code') != 1:
                    s.rollback()
                    return json_result(res.get('code'), res.get('msg'))
                res = user_task_new('费用申请', rid, '付款单号', '费用申请审批', xxnr, user, s, [tjcw], ['all'])
                if res.get('code') != 1:
                    s.rollback()
                    return json_result(res.get('code'), res.get('msg'))
                
                t = {"type":"success","title":"财务审批通过通知",
                    "msg": xxnr,
                    "module": "费用申请",
                    "rid": rid, "rids":[],
                    "btns":[{"title":"打开","icon":"any-rmb-full","name":"open","color":"red"}]}
                await messages.message_to_user(tjcw, t, MSG_KIND_NOTICE_RECORD, request)

        s.commit()
            
        return json_result(code, msg, cxje)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'撤销审批失败: {str(e)}')
    finally:
        s.close()

"""
费用申请.提交总经理
对应原Pascal: 提交总经理
"""
@any_route('/api/saier/cost_apply/tjzjl/change', methods=['POST'])
@require_token
async def view_cost_apply_tjzjl_change(request):
    s = Session()
    try:
        j = await request.json()
        rid = j.get('rid', '')
        tjzjl = j.get('tjzjl', '')
        zjsp = j.get('zgsp', '')
        fkbh = j.get('fkbh', '')
        wysp = j.get('wysp', '')
        user = request.current_user
        code = 1
        msg = '操作成功'
        yzjl = ''
        d = s.query(fysq.tjzjl).filter(fysq.rid == rid).first()
        if d and d.tjzjl:
            yzjl = d.tjzjl
        if tjzjl == yzjl:
            return json_result(0, '总经理和数据库一致，无需修改')
        
        if tjzjl == user.username:
            return json_result(-1, '自已不能审批自已,请重新选择')
        d = s.query(zx.wb3).filter(zx.ly == '付款审批', zx.wb3 == tjzjl).first()
        if not d:
            return json_result(-1, '此人没有审批权限,请重新选择')
        
        if zjsp == '通过':
                # return json_result(-1, '请先提交总监审批通过再选择总经理审批人')
            if tjzjl == user.username:
                res = user_task_delete('费用申请', rid, s, [])
                if res.get('code') != 1:
                    s.rollback()
                    return json_result(res.get('code'), res.get('msg'))
                return json_result(code, msg)
            
            if tjzjl != user.username:
                if wysp != tjzjl and tjzjl != '' and tjzjl is not None:
                    xxnr = str(user.username) + '的费用申请:' + str(fkbh) + '需审批,日期:' + str(time.strftime('%Y-%m-%d'))
                    row = {
                        "xxly": '费用申请',
                        "wxht": '',
                        "gdht": str(fkbh),
                        "yhdh": '',
                        "xxnr": xxnr,
                        "jsr": str(tjzjl),
                        "sys_path": "",
                        "spsq": user.username
                    }
                    res = module_xxck_new([row], user, s)
                    if res.get('code') != 1:
                        s.rollback()
                        return json_result(res.get('code'), res.get('msg'))
                    res = user_task_delete('费用申请', rid, s, [])
                    if res.get('code') != 1:
                        s.rollback()
                        return json_result(res.get('code'), res.get('msg'))
                    res = user_task_new('费用申请', rid, '付款单号', '费用申请审批', xxnr, user, s, [tjzjl], ['all'])
                    if res.get('code') != 1:
                        s.rollback()
                        return json_result(res.get('code'), res.get('msg'))
                    
                    t = {"type":"success","title":"财务审批通过通知",
                        "msg": xxnr,
                        "module": "费用申请",
                        "rid": rid, "rids":[],
                        "btns":[{"title":"打开","icon":"any-rmb-full","name":"open","color":"red"}]}
                    await messages.message_to_user(tjzjl, t, MSG_KIND_NOTICE_RECORD, request)

        s.commit()
            
        return json_result(code, msg)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'撤销审批失败: {str(e)}')
    finally:
        s.close()



"""
费用申请.提交主管
对应原Pascal: 提交主管
"""
@any_route('/api/saier/cost_apply/tjzg/change', methods=['POST'])
@require_token
async def view_cost_apply_tjzg_change(request):
    s = Session()
    try:
        j = await request.json()
        rid = j.get('rid', '')
        tjjl = j.get('tjzg', '')
        fkbh = j.get('fkbh', '')
        sb12 = j.get('sb12', '')
        wysp = j.get('wysp', '')
        user = request.current_user
        code = 1
        msg = '操作成功'
        yjl = ''
        bm = ''
        data = {'bm': '', 'user':{}, 'user2': {}}
        d = s.query(fysq.tjjl).filter(fysq.rid == rid).first()
        if d and d.tjjl:
            yjl = d.tjjl
        if tjjl == yjl:
            return json_result(0, '主管和数据库一致，无需修改')
        
        if tjjl == user.username:
            return json_result(-1, '自已不能审批自已,请重新选择')
        
        if tjjl != user.username and sb12 != '1':
            d = s.query(zx).filter(zx.ly == '付款审批', zx.wb1 == tjjl).first()
            if not d:
                return json_result(-1, '此人没有审批权限,请重新选择')
            data['user'] = alchemy_object_to_dict(d) 
            c = s.query(ywrybiao.bm).filter(ywrybiao.yhm == tjjl).first()
            if c:
                data['bm'] = c.bm
                bm = c.bm
            if wysp != tjjl : 
                d = s.query(zx.wb2).filter(zx.ly == '付款审批', zx.wb2 == tjjl, zx.mc == bm).first()
                if d:
                    data['user2'] = alchemy_object_to_dict(d)
                xxnr = str(user.username) + '的费用申请:' + str(fkbh) + '需审批,日期:' + str(time.strftime('%Y-%m-%d'))
                if user.username != tjjl and tjjl != '' and tjjl is not None:
                    row = {
                        "xxly": '费用申请',
                        "wxht": '',
                        "gdht": str(fkbh),
                        "yhdh": '',
                        "xxnr": xxnr,
                        "jsr": str(tjjl),
                        "sys_path": "",
                        "spsq": user.username
                    }
                    res = module_xxck_new([row], user, s)
                    if res.get('code') != 1:
                        s.rollback()
                        return json_result(res.get('code'), res.get('msg'))
                    t = {"type":"success","title":"主管审批通过通知",
                        "msg": xxnr,
                        "module": "费用申请",
                        "rid": rid, "rids":[],
                        "btns":[{"title":"打开","icon":"any-rmb-full","name":"open","color":"red"}]}
                    await messages.message_to_user(tjjl, t, MSG_KIND_NOTICE_RECORD, request)

                res = user_task_delete('费用申请', rid, s, [])
                if res.get('code') != 1:
                    s.rollback()
                    return json_result(res.get('code'), res.get('msg'))
                
                res = user_task_new('费用申请', rid, '付款单号', '费用申请审批', xxnr, user, s, [tjjl], ['all'])
                if res.get('code') != 1:
                    s.rollback()
                    return json_result(res.get('code'), res.get('msg'))
                
        else:
            d = s.query(zx).filter(zx.ly == '付款审批', zx.wb1 == tjjl).first()
            if not d:
                return json_result(-1, '此人没有审批权限,请重新选择')
            data['user'] = alchemy_object_to_dict(d)
            res = user_task_delete('费用申请', rid, s, [])
            if res.get('code') != 1:
                s.rollback()
                return json_result(res.get('code'), res.get('msg'))

            s.commit()
        return json_result(code, msg, data)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'提交主管失败: {str(e)}')
    finally:
        s.close()


"""
费用申请.经理识别
对应原Pascal: 经理识别
"""
@any_route('/api/saier/cost_apply/jlsb/change', methods=['POST'])
@require_token
async def view_cost_apply_jlsb_change(request):
    s = Session()
    try:
        j = await request.json()
        rid = j.get('rid', '')
        tjjl = j.get('tjzg', '')
        fkbh = j.get('fkbh', '')
        zgsp = j.get('zgsp', '')
        sb12 = j.get('sb12', '')
        wysp = j.get('wysp', '')
        hklx = j.get('hklx', '')
        sqje = j.get('sqje', 0.0)
        jbry = j.get('jbry', '')
        spyj = j.get('spyj', '')
        tjzj = j.get('tjzj', '')
        user = request.current_user
        code = 1
        msg = '操作成功'
        yjl = ''
        sb = ''
        data = {'sb': ''}

        cxje = 0
        d = s.query(zx.cs).filter(zx.mc.like('%' + hklx + '%'), zx.ly == '费用审请额度', zx.wb1 == user.username).first()
        if d:
            cxje = float(d.cs) if d.cs is not None else 0.0
        if cxje > 0 and sqje > cxje:
            sb = '1'
            data['sb'] = '1'

        d = s.query(fysq).filter(fysq.rid == rid).first()
        if d:
            res = user_task_delete('费用申请', rid, s, [])
            if res.get('code') != 1:
                s.rollback()
                return json_result(res.get('code'), res.get('msg'))
        if zgsp == '不通过':
            if jbry != '' and jbry is not None and jbry != user.username:
                xxnr = str(user.username) + '的费用申请:' + str(fkbh) + '主管不能通过,原因:' + str(spyj) + ',日期:' + str(time.strftime('%Y-%m-%d'))
                res = user_task_new('费用申请', rid, '付款单号', '费用申请审批未通过', xxnr, user, s, [jbry])
                if res.get('code') != 1:
                    s.rollback()
                    return json_result(res.get('code'), res.get('msg'))
                t = {"type":"success","title":"主管审批未通过通知",
                    "msg": xxnr,
                    "module": "费用申请",
                    "rid": rid, "rids":[],
                    "btns":[{"title":"打开","icon":"any-rmb-full","name":"open","color":"red"}]}
                await messages.message_to_user(jbry,t,MSG_KIND_NOTICE_RECORD,request)
        elif zgsp == '通过':
            xxnr = str(user.username) + '的费用申请:' + str(fkbh) + '需审批,日期:' + str(time.strftime('%Y-%m-%d'))
            if sb == '1' and wysp != tjzj and tjzj != '' and tjzj!=user.username and tjzj is not None:
                row = {
                    "xxly": '费用申请',
                    "wxht": '',
                    "gdht": str(fkbh),
                    "yhdh": '',
                    "xxnr": xxnr,
                    "jsr": str(tjzj),
                    "sys_path": "",
                    "spsq": user.username
                }
                res = module_xxck_new([row], user, s)
                if res.get('code') != 1:
                    s.rollback()
                    return json_result(res.get('code'), res.get('msg'))

                res = user_task_delete('费用申请', rid, s, [])
                if res.get('code') != 1:
                    s.rollback()
                    return json_result(res.get('code'), res.get('msg'))
                
                res = user_task_new('费用申请', rid, '付款单号', '费用申请审批', xxnr, user, s, [tjzj], ['all'])
                if res.get('code') != 1:
                    s.rollback()
                    return json_result(res.get('code'), res.get('msg'))

            s.commit()
        return json_result(code, msg, data)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'撤销审批失败: {str(e)}')
    finally:
        s.close()


"""
费用申请.总监识别
对应原Pascal: 总监识别
"""
@any_route('/api/saier/cost_apply/zjsb/change', methods=['POST'])
@require_token
async def view_cost_apply_zjsb_change(request):
    s = Session()
    try:
        j = await request.json()
        rid = j.get('rid', '')
        fkbh = j.get('fkbh', '')
        zjsp = j.get('zjsp', '')
        wysp = j.get('wysp', '')
        hklx = j.get('hklx', '')
        sqje = j.get('sqje', 0.0)
        jbry = j.get('jbry', '')
        spyj = j.get('spyj', '')
        tjzj = j.get('tjzj', '')
        tjzjl = j.get('tjzjl', '')
        user = request.current_user
        code = 1
        msg = '操作成功'
        yjl = ''
        sb = ''
        data = {'sb': ''}

        cxje = 0
        d = s.query(cyzglsheet.bz, cyzglsheet.sz).filter(cyzglsheet.xm == '总监费用审请额度', cyzglsheet.xm == user.username).first()
        if d:
            if hklx in d.bz:
                cxje = float(d.cs) if d.cs is not None else 0.0

        if cxje > 0 and sqje > cxje:
            sb = '1'
            data['sb'] = '1'

        res = user_task_delete('费用申请', rid, s, [])
        if res.get('code') != 1:
            s.rollback()
            return json_result(res.get('code'), res.get('msg'))
        
        if zjsp == '不通过':
            if jbry != '' and jbry is not None and jbry != user.username:
                xxnr = str(user.username) + '的费用申请:' + str(fkbh) + '总监不能通过,原因:' + str(spyj) + ',日期:' + str(time.strftime('%Y-%m-%d'))
                res = user_task_new('费用申请', rid, '付款单号', '费用申请审批未通过', xxnr, user, s, [jbry])
                if res.get('code') != 1:
                    s.rollback()
                    return json_result(res.get('code'), res.get('msg'))
                t = {"type":"success","title":"总监审批未通过通知",
                    "msg": xxnr,
                    "module": "费用申请",
                    "rid": rid, "rids":[],
                    "btns":[{"title":"打开","icon":"any-rmb-full","name":"open","color":"red"}]}
                await messages.message_to_user(jbry,t,MSG_KIND_NOTICE_RECORD,request)
        elif zjsp == '通过':
            xxnr = str(user.username) + '的费用申请:' + str(fkbh) + '需审批,日期:' + str(time.strftime('%Y-%m-%d'))
            if sb == '1' and wysp != tjzjl and tjzjl != '' and tjzjl!=user.username and tjzjl is not None:
                row = {
                    "xxly": '费用申请',
                    "wxht": '',
                    "gdht": str(fkbh),
                    "yhdh": '',
                    "xxnr": xxnr,
                    "jsr": str(tjzjl),
                    "sys_path": "",
                    "spsq": user.username
                }
                res = module_xxck_new([row], user, s)
                if res.get('code') != 1:
                    s.rollback()
                    return json_result(res.get('code'), res.get('msg'))

                res = user_task_delete('费用申请', rid, s, [])
                if res.get('code') != 1:
                    s.rollback()
                    return json_result(res.get('code'), res.get('msg'))
                
                res = user_task_new('费用申请', rid, '付款单号', '费用申请审批', xxnr, user, s, [tjzjl], ['all'])
                if res.get('code') != 1:
                    s.rollback()
                    return json_result(res.get('code'), res.get('msg'))

            s.commit()
        return json_result(code, msg, data)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'撤销审批失败: {str(e)}')
    finally:
        s.close()


"""
费用申请.总经理识别
对应原Pascal: 总经理识别
"""
@any_route('/api/saier/cost_apply/zjlsb/change', methods=['POST'])
@require_token
async def view_cost_apply_zjlsb_change(request):
    s = Session()
    try:
        j = await request.json()
        rid = j.get('rid', '')
        fkbh = j.get('fkbh', '')
        zjlsp = j.get('zjlsp', '')
        jbry = j.get('jbry', '')
        spyj = j.get('spyj', '')
        tjcw = j.get('tjcw', '')
        tjzjl = j.get('tjzjl', '')
        user = request.current_user
        code = 1
        msg = '操作成功'

        res = user_task_delete('费用申请', rid, s, [])
        if res.get('code') != 1:
            s.rollback()
            return json_result(res.get('code'), res.get('msg'))
        zjlmc = '总经理'
        if tjzjl == '谢培雅':
            zjlmc = '谢培雅'
        if zjlsp == '不通过':
            if jbry != '' and jbry is not None and jbry != user.username:
                xxnr = str(user.username) + '的费用申请:' + str(fkbh) + str(zjlmc) + '不能通过,原因:' + str(spyj) + ',日期:' + str(time.strftime('%Y-%m-%d'))
                res = user_task_new('费用申请', rid, '付款单号', '费用申请审批未通过', xxnr, user, s, [jbry])
                if res.get('code') != 1:
                    s.rollback()
                    return json_result(res.get('code'), res.get('msg'))
                data = {"type":"success","title": str(zjlmc) + "审批未通过通知",
                    "msg": xxnr,
                    "module": "费用申请",
                    "rid": rid, "rids":[],
                    "btns":[{"title":"打开","icon":"any-rmb-full","name":"open","color":"red"}]}
                await messages.message_to_user(jbry,data,MSG_KIND_NOTICE_RECORD,request)
        elif zjlsp == '通过':
            xxnr = str(user.username) + '的费用申请:' + str(fkbh) + str(zjlmc) + '通过请查看,日期:' + str(time.strftime('%Y-%m-%d'))
            if tjcw != '' and tjcw!=user.username and tjcw is not None:
                row = {
                    "xxly": '费用申请',
                    "wxht": '',
                    "gdht": str(fkbh),
                    "yhdh": '',
                    "xxnr": xxnr,
                    "jsr": str(tjcw),
                    "sys_path": "",
                    "spsq": user.username
                }
                res = module_xxck_new([row], user, s)
                if res.get('code') != 1:
                    s.rollback()
                    return json_result(res.get('code'), res.get('msg'))

                res = user_task_delete('费用申请', rid, s, [])
                if res.get('code') != 1:
                    s.rollback()
                    return json_result(res.get('code'), res.get('msg'))
                
                res = user_task_new('费用申请', rid, '付款单号', '费用申请审批', xxnr, user, s, [tjcw], ['all'])
                if res.get('code') != 1:
                    s.rollback()
                    return json_result(res.get('code'), res.get('msg'))
                
            t = {"type":"success","title": str(zjlmc) + "审批未通过通知",
                "msg": xxnr,
                "module": "费用申请",
                "rid": rid, "rids":[],
                "btns":[{"title":"打开","icon":"any-rmb-full","name":"open","color":"red"}]}
            if jbry != '' and jbry is not None and jbry != user.username:
                await messages.message_to_user(jbry,t,MSG_KIND_NOTICE_RECORD,request)
            if tjcw != '' and tjcw is not None and tjcw != user.username:
                await messages.message_to_user(tjcw,t,MSG_KIND_NOTICE_RECORD,request)

            s.commit()
        return json_result(code, msg)
    except Exception as e:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, f'撤销审批失败: {str(e)}')
    finally:
        s.close()
