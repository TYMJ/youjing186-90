from hmac import new

from any import *
from .model import *
from sqlalchemy.sql import func,not_,or_,and_
import time
import shutil,json,os
from starlette.background import BackgroundTask
from PIL import Image as PILImage  # 重命名避免冲突
from .__default__ import get_user_path, module_share_new

# jpeg图片出错的解决方法
from PIL import JpegImagePlugin
JpegImagePlugin._getmp = lambda x:None
try:
    # 需要安装 pip install openpyxl
    from openpyxl import Workbook
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Border, Side, Alignment, Font, Alignment
    from openpyxl.drawing.image import Image as Image_Get
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.worksheet.protection import SheetProtection
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", 'openpyxl'])

SYS_FIELDS = ['sid','rid','uid','ctime','mtime','has_att','modi_uid','wf_status','wf_unit','pid','archived']


# 脚本日志新增，如果 s=None,则自动提交事务，否则返回Session对象
async def insert_script_log(module, keyNo, kind, status, user, logText,photos=None, s=None):
    flag = False
    if not s:
        flag = True
        s = Session()
    try:
        m = ScriptLog()
        m.rid = get_uuid()
        m.uid = user.rid
        m.ModuleName = module
        m.KeyNo = keyNo
        m.Kind = kind
        m.Status = status
        m.UserName = user.username
        m.Date = time.strftime('%Y-%m-%d %H:%M:%S')
        m.LogText = logText
        if  photos!='' and photos!='[]' and photos!=None and photos!='None':
            logger.error(photos)
            logger.error('-----aaaaaaaaa-----')
            photos = json.loads(photos)
            new_photos = []
            data_path = config.data_upload_path
            file_path = config.get_today_upload_path()
            if not os.path.exists(file_path):
                make_dirs(file_path)
            for photo in photos:
                # src = photo.get('src','')[:10]
                fp = os.path.join(data_path, photo.get('src'))
                if os.path.exists(fp):
                    fn = photo.get('name')
                    # if os.path.join(data_path,src)==file_path:
                    fn = str(get_uuid()) + '_' + str(fn)
                    shutil.copy2(fp, os.path.join(file_path, fn))
                    sub_path = file_path[-10:]
                    new_photos.append({'name': fn, 'src': sub_path+'/'+fn})
            if len(new_photos) > 0:
                m.cptp = str(new_photos).replace("'",'"')

        s.add(m)
        if flag:
            s.commit()
        return {'code':1,'msg':'日志新增成功'}
    except:
        if flag:
            s.rollback()
        logging.info(trace_error())
        return {'code':-1,'msg':trace_error()}
    finally:
        if flag:
            s.close()
    
#专业产品记录修改权限校验
@any_route('/api/saier/items/user/check', methods=['POST', 'GET'])
@require_token
async def view_saier_items_user_check(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        i = 0
        wfgs_n = j.get('wfgs', '')
        cpbh = j.get('cpbh', '')
        rid = j.get('rid', '')
        res = items_cpbh_check(rid,cpbh,s)
        if res.get('code') != 1:
            return json_result(res.get('code'), res.get('msg'))
    
        d = run_sql(f"select count(rid) as qty from bjsheet where bjhh='{str(cpbh)}' and (bjry<>'{str(user.username)}' or ifnull(bjbjwyzd1,'')<>'') limit 1")
        if len(d)>0:
            i = d[0].get('qty')

        if wfgs_n == '宁波优景进出口有限公司' and i >= 1:
            d = run_sql(f"select rid from sys_user where username='{str(user.username)}' and (position like '%外销%' or memo like '%外销%' or path like '%样品%')")
            if len(d)==0:
                return json_result(-1,'不好意思,您没有权利修改此记录,请和相关业务人员联系!')
        
        return json_result(1, '校验成功')
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 专业产品产品分类检验
@any_route('/api/saier/items/cpfl/check', methods=['POST', 'GET'])
@require_token
async def view_saier_items_cpfl_check(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        msg = '校验成功'
        where_list = []
        data = {'cpfl':[],'yjfl':[],'ejfl':[],'sjfl':[]}
        cpfl = j.get('cpfl', '')
        if cpfl!=None and cpfl!='':
            where_list.append(zycpfenglb.cpfl==str(cpfl))
            msg = '产品大类不存在'
        yjfl = j.get('yjfl', '')
        if yjfl!=None and yjfl!='':
            where_list.append(zycpfenglb.yjfl==str(yjfl))
            msg = '一级分类不存在'
        ejfl = j.get('ejfl', '')
        if ejfl!=None and ejfl!='':
            where_list.append(zycpfenglb.ejfl==str(ejfl))
            msg = '二级分类不存在'
        sjfl = j.get('sjfl', '')
        if sjfl!=None and sjfl!='':
            where_list.append(zycpfenglb.sjfl==str(sjfl))
            msg = '三级分类不存在'
        d = s.query(zycpfenglb).filter(*where_list).all()
        if not d:
            return json_result(-1,msg)
        data['cpfl'] = [r.cpfl for r in d]
        data['yjfl'] = [r.yjfl for r in d]
        data['ejfl'] = [r.ejfl for r in d]
        data['sjfl'] = [r.sjfl for r in d]
        
        return json_result(1, msg, data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 专业产品产品编号检验主函数  
def items_cpbh_check(rid,cpbh,s):
    try:
        d = s.query(cjcp.rid).filter(cjcp.cpbh==str(cpbh),cjcp.rid!=str(rid)).first()
        if d:
            return {'code': -1, 'msg': f"产品编号【{cpbh}】已存在,请重新修改!"}

        d = s.query(zscp.rid).filter(zscp.cpbh==str(cpbh)).first()
        if d:
            return {'code': -1, 'msg': f"产品编号【{cpbh}】已存在,请重新修改!"}
        
        return {'code': 1, 'msg': '校验成功'}
    except:
        logger.error(trace_error())
        return {'code': -1, 'msg': trace_error()}

# 专业专属产品产品编号检验  
@any_route('/api/saier/items/cpbh/check', methods=['POST', 'GET'])
@require_token
async def view_saier_items_cpbh_check(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        rid = j.get('rid', '')
        cpbh = j.get('cpbh', '')
        res = items_cpbh_check(rid,cpbh,s)
        
        return json_result(res.get('code'), res.get('msg'))
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 专业产品条形码检验主函数  
def items_txm_check(rid,txm,s):
    try:
        d = s.query(cjcp.rid).filter(cjcp.rid!=rid,cjcp.lxm==str(txm)).first()
        if d:
            return {'code': -1, 'msg': '条 形 码已存在'}

        d = s.query(zscp.rid).filter(zscp.lxm==str(txm)).first()
        if d:
            return {'code': -1, 'msg': '专属产品中已存在'}
        
        return {'code': 1, 'msg': '校验成功'}
    except:
        logger.error(trace_error())
        return {'code': -1, 'msg': trace_error()}

# 专业产品条形码检验  
@any_route('/api/saier/items/txm/check', methods=['POST', 'GET'])
@require_token
async def view_saier_items_txm_check(request):
    s = Session()
    # user = request.current_user
    j = await request.json()
    try:
        rid = j.get('rid', '')
        txm = j.get('txm', '')
        res = items_txm_check(rid,txm,s)
        
        return json_result(res.get('code'), res.get('msg'))
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 专业产品记录加载
@any_route('/api/saier/items/zy/load/check', methods=['POST', 'GET'])
@require_token
async def view_saier_items_zy_load_check(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        logger.error(user.orgs)
        data = {}
        data['check'] = 0
        wfgs_n = j.get('wfgs')
        module = j.get('module')
        d = run_sql(f"select sys_user.rid from sys_user where sys_user.username='{user.username}' and ((position like '%外销%') or (position like '%总经理%') or (memo like '%外销%') or (memo like '%总经理%')) limit 1")
        if len(d)>0:
            data['check'] = 1
        if wfgs_n==None or wfgs_n=='':
            d = s.query(wfgs).first()
            if d:
                data['wfgs'] = d.wfgs

        return json_result(1, '取数成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 专属产品记录加载
@any_route('/api/saier/items/zs/load/check', methods=['POST', 'GET'])
@require_token
async def view_saier_items_zs_load_check(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        data = {}
        data['ywry'] = 0
        data['gckm'] = 0
        data['wfgs'] = ''
        data['cpbh'] = 0
        data['tsml'] = 0
        data['gxyw'] = 0
        data['price'] = 0
        data['cgsb'] = 0
        data['dygc'] = 1
        data['cgry'] = 0
        data['path'] = ''

        module = j.get('module')
        cpbh = j.get('cpbh')
        ywry = j.get('ywry')
        wfgs_n = j.get('wfgs')
        data['wfgs'] = wfgs_n
        d = s.query(ywrybiao.yhm).filter(or_(ywrybiao.bmjl==user.username,ywrybiao.sybzj==user.username,ywrybiao.sybdzj==user.username)).first()
        if d:
            data['ywry'] = 1
        d = s.query(gckm.rid).filter(gckm.kmbh==cpbh).first()
        if d:
            data['gckm'] = 1
        if wfgs_n==None or wfgs_n=='':
            d = s.query(wfgs).first()
            if d:
                data['wfgs'] = d.wfgs
        d = s.query(func.count(bjsheet.rid).label('qty')).filter(bjsheet.bjhh==cpbh,func.ifnull(bjsheet.bjhh,'')!='',or_(bjsheet.bjry!=user.username,bjsheet.bjbjwyzd1==None,bjsheet.bjbjwyzd1=='')).first()
        if d and d.qty and int(d.qty)>=1:
            data['cpbh'] = 1
        d = s.query(cyzglsheet.rid).filter(cyzglsheet.xm==user.username,cyzglsheet.zm=="特殊毛利率").first()
        if d:
            data['tsml'] = 1

        d = s.query(cyzgl.rid).filter(cyzgl.qxzl=="开启",cyzglsheet.zm=="专属产品共享权限").first()
        if d:
            c = s.query(cyzglsheet.rid).filter(cyzglsheet.xm==user.username,cyzglsheet.zm=="专属产品共享权限").first()
            if c:
                data['gxyw'] = 1
        else:
            c = s.query(ywrybiao.yhm).filter(ywrybiao.yhm==ywry).filter(or_(ywrybiao.yhm==user.username,ywrybiao.bmjl==user.username,
                ywrybiao.sybzj==user.username,ywrybiao.sybdzj==user.username,ywrybiao.zjl==user.username)).first()
            if c:
                data['gxyw'] = 1
            if ywry==user.username or ywry=='' or ywry==None or user.username=='zjnblh':
                data['gxyw'] = 1 
        # d = run_sql(f"select rid from sys_user where username='{user.username}'  and (path like '%爱马士%' limit 1")
        # if len(d)>0:
        #     data['gxyw'] = 1
        org = get_user_path(user.username," and ((position like '%外销%') or (position like '%总经理%') or (memo like '%外销%') or (memo like '%总经理%'))")
        data['path'] = org.get('path')
        # d = run_sql(f"select sys_user.rid from sys_user where sys_user.username='{user.username}' and ((position like '%外销%') or (position like '%总经理%')) limit 1")
        if org.get('rid')!=None and org.get('rid')!='':
            data['price'] = 1
        else:
            c = s.query(cyzglsheet.rid).filter(cyzglsheet.xm==user.username,cyzglsheet.zm=="采购新增对应工厂").first()
            if not c:
                data['dygc'] = 0
            data['cgsb'] = 1

        d = s.query(cyzglsheet.rid).filter(cyzglsheet.xm==user.username,cyzglsheet.zm=="专业专属更改采购人员").first()
        if not d:
            data['cgry'] = 1

        return json_result(1, '取数成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 专业产品获取体积毛重比
@any_route('/api/saier/items/get/size', methods=['POST', 'GET'])
@require_token
async def view_saier_items_get_size(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        v = 0
        f = 1
        xm = j.get('xm', '')
        sz = j.get('sz', 0)
        d = s.query(cyzglsheet.sz).filter(cyzglsheet.xm==xm,cyzglsheet.zm=='体积毛重比',cyzglsheet.sz1<=sz,cyzglsheet.sz2>=sz).first()
        if d:
            v = float(d.sz)
            f = 0
        return json_result(1, '取数成功', {'v': v, 'f': f})
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 模块记录解锁
@any_route('/api/saier/module/unlock', methods=['POST', 'GET'])
@require_token
async def view_saier_module_unlock(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        rid = j.get('rid', '')
        module = j.get('module', '')
        d = s.query(sys_record_lock).filter(sys_record_lock.rid==rid,sys_record_lock.module==module,sys_record_lock.uid!=user.rid).all()
        for r in d:
            s.delete(r) 
        s.commit()
        res = {'code': 1, 'msg': '解锁成功'}
        return json_result(res.get('code'), res.get('msg'))
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 专业产品更新货号主函数
async def items_update_cpbh(module,rid,cpbh_new,user,s):
    try:
        cpbh = ''
        pid = ''
        m = get_module(module)
        o = get_model_by_table_name(m.table_name)

        filters=[o.rid==rid]
        d = s.query(o).filter(*filters).first()
        if d:
            cpbh = d.cpbh
            pid = d.rid
            if str(d.cpbh)==str(cpbh_new):
                return {'code': -1, 'msg': f"{module}中已存在"}
        
        if module == '专属产品':
            c = s.query(cjcp.rid,cjcp.cpbh).filter(cjcp.cpbh==cpbh_new).first()
            if c:
                logger.error(f"{cpbh_new}已存在")
                return {'code': -1, 'msg': '专业产品中已存在'}
            logger.error(f"{cpbh_new}不已存在")
        elif module == '专业产品':
            c = s.query(zscp.rid,zscp.cpbh).filter(zscp.cpbh==cpbh_new).first()
            if c:
                return {'code': -1, 'msg': '专属产品中已存在'}
            c = s.query(cjcp.rid,cjcp.cpbh).filter(cjcp.cpbh==cpbh_new,cjcp.rid!=rid).first()
            if c:
                return {'code': -1, 'msg': '专业产品中已存在'}

        org = get_user_path(user.username)
        path = org.get('path')
        position = org.get('position')
        if '样品间' not in path and '外销' not in position and '总经理' not in position:
            return {'code': -1, 'msg': '您没有权限'}

        if module == '专属产品':
            t = s.query(wxhtsheet.rid).filter(wxhtsheet.bjhh==str(cpbh)).first()
            if t:
                return {'code': -1, 'msg': '外销合同中已存在此货号'}
        # 因为取消了保存后修改，所以注销了
        # d.cpbh = cpbh_new
        # d.lxm = ''
        # s.add(d)
        # res = await insert_script_log(module,cpbh_new,'产品编号修改','成功',user,str({"原产品编号":cpbh,"新产品编号":cpbh_new}),photos=d.yytp,s=s)
        # if res.get('code')!=1:
        #     return res
        
        return {'code': 1, 'msg': '校验成功'}
    except:
        logger.error(trace_error())
        return {'code': -1, 'msg': trace_error()}

# 专业产品更新货号
@any_route('/api/saier/items/update/cpbh', methods=['POST', 'GET'])
@require_token
async def view_saier_items_update_cpbh(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        rid = j.get('rid', '')
        # cpbh = j.get('cpbh', '')
        cpbh_new = j.get('cpbh_new', '')
        module = j.get('module', '专业产品')
        res = await items_update_cpbh(module,rid,cpbh_new,user,s)
        if res.get('code')==1:
            s.commit()
        else:
            s.rollback()
        return json_result(res.get('code'), res.get('msg'))
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 专业产品保存后
@any_route('/api/saier/items/zy/save/after', methods=['POST', 'GET'])
@require_token
async def view_saier_items_zy_save_after(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        cgry = j.get('cgry', '')
        uid = j.get('uid', '')
        cpbh = j.get('cpbh', '')
        path = ''
        sccj = j.get('sccj', '')
        gcID = j.get('gcID', '')
        cs_id = ''
        module = j.get('module', '专业产品')
        o = cjcp
        org = get_user_path(cgry)
        # d = s.query(sys_user.path,sys_user.rid).filter(sys_user.username==cgry).first()
        # if d:
            # if d.path != None and d.path != '':
                # path = str(d.path)[:100]
            # uid = str(d.rid)
        path = org.get('path')
        uid = org.get('rid')
        if uid == None or uid == '':
            uid = user.rid
        d= s.query(cyzglsheet.bz).filter(cyzglsheet.xm==cgry,cyzglsheet.zm=="报价产品查看权限").first()
        if d and d.bz != None and d.bz != '':
            path = str(d.bz)[:100]
        if module == '专属产品':
            o = zscp
            d = s.query(zycs.cs_id).filter(zycs.company_name==sccj).first()
            if d:
                cs_id = str(d.cs_id)

        # if path != '':
        cgry_old = ''
        d = s.query(o).filter(o.rid==str(j.get('rid', ''))).first()
        if d and str(d.uid) != uid:
            d.uid = uid
            u = s.query(sys_user.username).filter(sys_user.rid==d.uid).first()
            if u:
                cgry_old = u.username
            if cs_id!=None and cs_id != '' and cs_id != gcID:
                d.gcID = cs_id
            # if path != '':
            d.path = path
            # d.cgry = cgry
            s.add(d)
            ssdq = ''
            c = s.query(ywrybiao.ssdq).filter(ywrybiao.yhm==cgry).first()
            if c:
                ssdq = str(c.ssdq)
            c = s.query(cgjhsheet).filter(cgjhsheet.bjhh==d.cpbh).all()
            for r in c:
                r.cgry = d.cgry
                r.ssdq = ssdq
                s.add(r)
            c = s.query(zscpsheet5).filter(zscpsheet5.cpbh==d.cpbh,zscpsheet5.sccj==d.sccj).all()
            for r in c:
                r.cgry = d.cgry
                s.add(r)
            await insert_script_log(module,cpbh,"采购人员修改",'成功',user,str({"产品编号":d.cpbh,"新采购人员":cgry,"原采购人员":cgry_old}),s=s)
            s.commit()

        return json_result(1, '更新成功', uid)
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

async def zycssheet2_new(j, pid, cpbh, user, s, check=True, modi=False):
    try:
        d = s.query(zycssheet2).filter(zycssheet2.cpbh==cpbh,zycssheet2.pid==pid).first()
        if not d:
            if check==False:
                return {'code':1,'msg':'操作成功'}
            m = zycssheet2()
            for k,v in j.items():
                if k in SYS_FIELDS:
                    continue
                setattr(m,k,v)
            m.rid = get_uuid()
            m.pid = pid
            m.uid = user.rid
            m.ctime = time.strftime("%Y-%m-%d %H:%M:%S")
            m.lshh = j.get("gchh1", "")
            m.syq = j.get("cgry", '')
            m.cs_id = j.get("gcID", '')
            s.add(m)
        elif modi:
            for k,v in j.items():
                if k in SYS_FIELDS:
                    continue
                if hasattr(d,k):
                    setattr(d,k,v)
            d.mtime = time.strftime("%Y-%m-%d %H:%M:%S")
            d.modi_uid = user.rid
            s.add(d)

        return {'code':1,'msg':'操作成功'}
    except:
        logger.error(trace_error())
        return {'code':-1,'msg':trace_error()}
    
async def zycssheet5_new(j, pid, cpbh, gcID, sccj, user, s, check=True, modi=False, filters=[]):
    try:
        flag = 0
        old_cgry = ''
        new_cgry = ''
        rid = ''
        d= s.query(zscpsheet5).filter(zscpsheet5.cpbh==cpbh,or_(zscpsheet5.gcID==gcID,zscpsheet5.sccj==sccj)).filter(*filters).first()
        if not d:
            if check==False:
                return {'code':1,'msg':'操作成功'}
            m = zscpsheet5()
            for k,v in j.items():
                if k in SYS_FIELDS:
                    continue
                setattr(m,k,v)
            m.rid = get_uuid()
            m.uid = user.rid
            m.pid = pid
            m.ctime = time.strftime("%Y-%m-%d %H:%M:%S")
            m.cpbh = cpbh
            m.sccj = sccj
            m.ggbz = str(user.username) + "_" + time.strftime("%Y-%m-%d %H:%M:%S")
            s.add(m)
            flag = 1
            rid = m.rid
            new_cgry = m.cgry
        elif modi:
            old_cgry = d.cgry
            for k,v in j.items():
                if k in SYS_FIELDS:
                    continue
                if hasattr(d,k):
                    setattr(d,k,v)
            d.mtime = time.strftime("%Y-%m-%d %H:%M:%S")
            d.modi_uid = user.rid
            new_cgry = d.cgry
            rid = d.rid
            s.add(d)
        res = module_share_new('对应工厂', [new_cgry], rid, user, s, [old_cgry])
        if res.get('code')!=1:
            return res
        
        return {'code':1,'msg':'操作成功','data':flag}
    except:
        logger.error(trace_error())
        return {'code':-1,'msg':trace_error()}
    
# 专属产品保存后
@any_route('/api/saier/items/zs/save/after', methods=['POST', 'GET'])
@require_token
async def view_saier_items_zs_save_after(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        user_list = []
        rid = j.get('rid', '')
        d = s.query(zscp.rid,zscp.cgry).filter(zscp.rid==rid).first()
        if d:
            if d.cgry not in user_list:
                user_list.append(d.cgry)
            c = s.query(zscpsheet5.cgry).filter(zscpsheet5.pid==d.rid).all()
            for r in c:
                if r.cgry not in user_list:
                    user_list.append(r.cgry)
        o = get_module('专属产品')
        table = get_model_by_table_name(str('sys_')+o.table_name+str('_share'))
        s.query(table).filter(table.record_id==rid).delete(synchronize_session=False)
        res = module_share_new('专属产品', user_list, rid, user, s)
        if res.get('code')!=1:
            return res
        s.commit()
        # if sccj!=None and sccj != '':
        #     return json_result(-1, f"未找到专业工厂{sccj}记录")
        # pid = str(d.rid)
        # cs_id= str(d.cs_id)
        # res = await zycssheet2_new(j, pid, cpbh, user, s)
        # if res.get('code')!=1:
        #     return json_result(res.get('code'), res.get('msg'))
        
        # res = await zycssheet5_new(j, pid, cpbh, cs_id, sccj, user, s)
        # if res.get('code')!=1:
        #     return json_result(res.get('code'), res.get('msg'))
            
        # await insert_script_log(module,cpbh,"采购人员修改",'成功',user,str({"产品编号":d.cpbh,"新采购人员":cgry,"原采购人员":cgry_old}),s=s)
        # s.commit()
        return json_result(1, '操作成功')
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 专业产品转专属产品处理主表
async def items_trans_main_new(j,module,path,cpbh,uid,c,user,s):
    try:
        new_photos = []
        ywry = j.get('ywry', '')
        retain = j.get('retain',False)
        module = j.get('module','专业产品')
        if module == '专业产品':
            m = get_module('专属产品')
        elif module == '专属产品':
            m = get_module('专业产品')
        else:
            m = get_module(module)
        o = get_model_by_table_name(m.table_name)
        # logger.error(config.__dict__)
        data_path = config.data_upload_path
        
        file_path = config.get_today_upload_path()
        if not os.path.exists(file_path):
            make_dirs(file_path)

        m = o()
        new_rid = get_uuid()
        m.rid = new_rid
        m.uid = uid
        for k,v in c.items():
            if k in SYS_FIELDS:
                continue
            if len(k)>5 and (k.upper())[:5]=='FIELD':
                continue
            if hasattr(m, k):
                setattr(m, k, v)
        m.zrsj = time.strftime('%Y-%m-%d %H:%M:%S')
        m.ywcpcs = '无'
        m.ywpath = path
        m.cpbh = cpbh
        m.fljg = 0
        m.zycpbh = c.get('cpbh')
        m.zycphhp = c.get('cpbh')
        m.Twxdj = c.get('mjfob',0)
        m.pkRMB = c.get('rmbdj',0)
        if c.get('bgpm','') != '':
            m.bgpm = c.get('bgpm','')
        else:
            m.bgpm = '无'
        if c.get('kmxx','') != '':
            m.kmxx = c.get('kmxx','')
        else:
            m.kmxx = '无'
        customer = c.get('customer','')        
        if customer != None and customer != '':
            l = s.query(kh.kh_id).filter(kh.company_name == customer).first()
            if l:
                m.kh_id = l.kh_id

        m.hssc = '否'
        m.bjyw = ywry
        m.wxdj2=round(c.get('wxdj2',0)*100)/100
        m.nhzcm=round(c.get('nhzcm',0)*10)/10
        m.nhkcm=round(c.get('nhkcm',0)*10)/10
        m.nhgcm=round(c.get('nhgcm',0)*10)/10
        m.nhdj2=round(c.get('nhdj2',0)*100)/100
        m.xzzcb=round(c.get('xzzcb',0)*100)/100
        m.yfjg=round(c.get('yfjg',0)*100)/100
        m.jcf=round(c.get('jcf',0)*100)/100
        m.yfcb=round(c.get('yfcb',0)*100)/100
        m.zkzcm=round(c.get('zkzcm',0)*10)/10
        m.zkkcm=round(c.get('zkkcm',0)*10)/10
        m.zkdj2=round(c.get('zkdj2',0)*100)/100
        m.zkdj=round(c.get('zkdj',0)*100)/100
        m.zkpkzcm=round(c.get('zkpkzcm',0)*10)/10
        m.zkpkkcm=round(c.get('zkpkkcm',0)*10)/10
        m.zkphcm=round(c.get('zkphcm',0)*10)/10
        m.zkpkjg2=round(c.get('zkpkjg2',0)*100)/100
        m.pkzcm=round(c.get('pkzcm',0)*10)/10
        m.pkkcm=round(c.get('pkkcm',0)*10)/10
        m.pkgcm=round(c.get('pkgcm',0)*10)/10
        m.pkhdcm=round(c.get('pkhdcm',0)*10)/10
        m.pkdj2=round(c.get('pkdj2',0)*100)/100
        m.dzzcm=round(c.get('dzzcm',0)*10)/10
        m.dzkcm=round(c.get('dzkcm',0)*10)/10
        m.dzhdcm=round(c.get('dzhdcm',0)*1000)/1000
        m.md=round(c.get('md',0)*100)/100
        m.bfs=round(c.get('bfs',0)*100)/100
        m.dzdj=round(c.get('dzdj',0)*100)/100
        m.hzzcm=round(c.get('hzzcm',0)*10)/10
        m.hzkcm=round(c.get('hzkcm',0)*10)/10
        m.hzgcm=round(c.get('hzgcm',0)*10)/10
        m.zkdj3=round(c.get('zkdj3',0)*100)/100
        m.hzdj=round(c.get('hzdj',0)*100)/100
        m.jcfy=round(c.get('jcfy',0)*100)/100
        m.zsbh=str(cpbh)+'_'+str(new_rid)
        if c.get('kmxx','') != '':
            m.kmxx = c.get('kmxx','')
        else:
            m.kmxx = '无'
        photos = str(c.get('yytp',''))
        if retain==True and c.get('yytp') != None and c.get('yytp') != '' and c.get('yytp') != '[]':
            photos = json.loads(c.get('yytp'))
            for photo in photos:
                src = photo.get('src','')[:10]
                fp = os.path.join(data_path, photo.get('src'))
                if os.path.exists(fp):
                    fn = photo.get('name')
                    # if os.path.join(data_path,src)==file_path:
                    #     fn = str(ywry) + '_' + str(fn)
                    # shutil.copy2(fp, os.path.join(file_path, fn))
                    # sub_path = file_path[-10:]
                    # new_photos.append({'name': fn, 'src': sub_path+'/'+fn})
                    ext = get_suffix(fp)
                    filename = get_unique(16) + ext
                    logger.error(f"复制文件{fp}到{os.path.join(file_path, filename)}")
                    shutil.copy2(fp, os.path.join(file_path, filename))
                    sub_path = file_path[-10:]
                    new_photos.append({'name': fn, 'src': sub_path+'/'+filename})

            if len(new_photos) > 0:
                m.yytp = str(new_photos).replace("'",'"')
                photos = str(new_photos).replace("'",'"')
        s.add(m)
        res = module_share_new('专属产品', c.get('cgry'), new_rid, user,s)
        if res.get('code') != 1:
            return res

        return {'code':1,'msg':'主表生成成功','rid':new_rid,"photos":str(photos)}
    except:
        logger.error(trace_error())
        return  {'code':-1,'msg':trace_error()}

# 专业产品转专属产品处理子表
async def items_trans_line(j,o,old_pid,uid,c,new_pid,s):
    try:
        retain = j.get('retain',0)
        d = s.query(o).filter(o.pid==old_pid).all()
        d = alchemy_object_list_to_dict(d)
        for r in d:
            m = get_model_by_table_name(c)()
            for k,v in r.items():
                if k in SYS_FIELDS:
                    continue
                # if len(k)>5 and (k.upper())[:5]=='FIELD':
                #     continue
                if hasattr(m, k):
                    setattr(m, k, v)
            m.rid = get_uuid()
            m.pid = new_pid
            m.uid = uid
            s.add(m)
        if retain==False:
            s.query(o).filter(o.pid==old_pid).delete()

        return {'code':1,'msg':'子表生成成功'}
    except:
        logger.error(trace_error())
        return  {'code':-1,'msg':trace_error()}
    
# 专业产品转专属产品主函数
async def items_trans_zscp(j,user,s):
    try:
        flag = 0
        ywbm = ''
        rybh = ''
        path = ''
        ywry = j.get('ywry', '')
        module = j.get('module', '')
        kind = '专业转专属'
        org = get_user_path(ywry)
        uid = org.get('rid')
        path = org.get('path')
        # u = run_sql(f"select rid,path from sys_user where username='{str(ywry)}' limit 1")
        if uid==None or uid=="":
            await messages.message(user.username,{'msg': res.get('msg'),'kind':'2'},MSG_KIND_NORMAL)
            return {'code': -1, 'msg': '业务人员不存在,请重新输入'}
        org = get_user_path(ywry," and ((position like '%外销总监%') or (position like '%总经理%') or (memo like '%外销总监%') or (memo like '%总经理%'))")
        # u = run_sql(f"select rid from sys_user where (username='{str(ywry)}') and ((position like '%外销总监%') or (position like '%总经理%')) limit 1")
        if org.get('rid')!=None and org.get('rid')!="":
            d = s.query(ywrybiao).filter(ywrybiao.yhm==ywry).first()
            if d:
                rybh = str(d.rybh)
                ywbm = str(d.bm)
                flag = 1
        else:
            logger.error(f"{ywry}不是外销总监或总经理")
            d = s.query(ywrybiao).filter(ywrybiao.yhm==ywry).filter(or_(ywrybiao.bmjl==user.username,ywrybiao.sybzj==user.username,
                ywrybiao.tscpzy.like("%," + str(user.username) +",%"))).first()
            if d:
                rybh = str(d.rybh)
                ywbm = str(d.bm)
                flag = 1

        if flag == 0:
            return {'code': -1, 'msg': '不好意思,此人非本部门员工,请重新选择,谢谢'}
        
        flag = 0
        retain = j.get('retain',False)
        errors = []
        empty_str = {'zwpm':{'title':'中文品名','kind':'str'},'ywpm':{'title':'英文品名','kind':'str'},'jldw':{'title':'计量单位','kind':'str'},'zwdw':{'title':'中文计量单位','kind':'str'},
            'bgpm':{'title':'报关品名','kind':'str'},'sfhs':{'title':'是否含税','kind':'str'},'cpgg':{'title':'产品规格','kind':'str'},'cpggz':{'title':'规格英语','kind':'str'},
            'nhrl':{'title':'内盒容量','kind':'float'},'nhwx':{'title':'内盒/外箱','kind':'float'},'bzkd':{'title':'包装宽度','kind':'float'},'bzcd':{'title':'包装长度','kind':'float'},
            'bzgd':{'title':'包装高度','kind':'float'},'bztj':{'title':'外箱体积','kind':'float'},'mxmz':{'title':'毛重','kind':'float'},'mxjz':{'title':'净重','kind':'float'},
            'zhwbzh':{'title':'中文包装','kind':'str'},'bzhfsh':{'title':'英文包装','kind':'str'},'caizi':{'title':'材质中文','kind':'str'},'caiziz':{'title':'材质英文','kind':'str'},
            'topcz':{'title':'产品来源','kind':'str'},'gcID':{'title':'厂商编号','kind':'str'},'sccj':{'title':'生产厂家','kind':'str'},'cgry':{'title':'采购人员','kind':'str'},
            'flmc':{'title':'分类名称','kind':'str'},'rkdd':{'title':'入库地点','kind':'str'}}

        rids = j.get('rids', [])
        for rid in rids:
            d = s.query(cjcp).filter(or_(cjcp.rid==rid,cjcp.cpbh==rid)).first()
            if not d:
                return {'code': -1, 'msg': '专业产品中不存在'}
            c = alchemy_object_to_dict(d)
            cpbh = c.get('cpbh')
            old_pid = c.get('rid')
            for k,v in c.items():
                if k in empty_str:
                    if empty_str[k]['kind']=='str':
                        if v == '':
                            errors.append(empty_str[k]['title'])
                            flag = 1
                    elif empty_str[k]['kind']=='float':
                        if v == '' or v == 0:
                            errors.append(empty_str[k]['title'])
                            flag = 1
            if c.get('kmxx')=='是' and c.get('mjfcdf')=='':
                errors.append('模具费承担方')
                flag = 1

            if cpbh!='' and flag == 0:
                if int(retain) == True:
                    cpbh = str(cpbh) + '-' + str(rybh)
            f = s.query(zscp).filter(zscp.cpbh==str(cpbh)).first()
            if f:
                return {'code': -1, 'msg': f"不好意思,{cpbh}已有此专属产品,谢谢"}
            
            logger.error(cpbh)
            res = await items_trans_main_new(j,module,path,cpbh,uid,c,user,s)
            if res.get('code')!=1:
                return res
            new_rid = res.get('rid')
            photos = res.get('photos')

            o = zycjcpsheet
            res = await items_trans_line(j,o,old_pid,uid,'zscpsheet',new_rid,s)
            if res.get('code')!=1:
                return res
            o = zscpsheet5()
            o.rid = get_uuid()
            o.pid = new_rid
            o.uid = uid
            for k,v in c.items():
                if k in SYS_FIELDS:
                    continue
                if hasattr(o, k):
                    setattr(o, k, v)
            o.cpbh = cpbh
            s.add(o)
            if d.cgry != None and d.cgry != '':
                res = module_share_new('对应工厂', [d.cgry], o.rid, user,s)
                if res.get('code') != 1:
                    return res
                
            l = s.query(zycs.cs_id,zycs.rid).filter(zycs.company_name==c.get('sccj')).first()
            if l:
                o = zycssheet2()
                o.rid = get_uuid()
                o.pid = l.rid
                for k,v in c.items():
                    if k in SYS_FIELDS:
                        continue
                    if hasattr(o, k):
                        setattr(o, k, v)
                o.kfkp = c.get('ljrk')
                o.kbgpm = c.get('bgpm')
                o.syq = c.get('cgry')
                o.cs_id = l.cs_id
                s.add(o)

            o = zscpsheet6()
            o.rid = get_uuid()
            o.pid = new_rid
            o.cpbh = c.get('cpbh')
            o.ywry = ywry
            o.ywbm = ywbm
            o.syspath = path
            o.gxqx = '完全'
            o.cpyw = ywry
            s.add(o)

            l = s.query(zx).filter(zx.ly=="最低毛利率").all()
            for r in l:
                o = zscpsheet1()
                o.rid = get_uuid()
                o.pid = new_rid
                o.khmc = r.mc
                if r.cs != None:
                    o.mll = float(r.cs)
                else:
                    o.mll = 0
                o.jzrq = '2999-12-31'
                o.cpbh = cpbh
                s.add(o)
            if d.cgry != None and d.cgry != '':
                res = module_share_new('专属产品', [d.cgry], new_rid, user,s)
                if res.get('code') != 1:
                    return res

            if int(retain)==False:
                s.delete(d)

            res = await insert_script_log(module,cpbh,kind,'成功',user,str({"产品编号":cpbh}),photos=photos,s=s)
            if res.get('code')!=1:
                return res

        return {'code': 1, 'msg': '生成成功'}
    except:
        logger.error(trace_error())
        return {'code': -1, 'msg': trace_error()}

# 专业产品转专属产品
@any_route('/api/saier/items/trans/zscp', methods=['POST', 'GET'])
@require_token
async def view_saier_items_trans_zscp(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        res = await items_trans_zscp(j,user,s)
        if res.get('code')==1:
            s.commit()
        else:
            s.rollback()

        return json_result(res.get('code'), res.get('msg'))
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

        
async def items_trans_zycp_check(rid,user,s):
    try: 
        u = s.query(cyzglsheet.rid).filter(cyzglsheet.xm==user.username,cyzglsheet.zm=="专属转专业").first()
        if not u:
            return {'code': -1, 'msg': '权限校验失败'}
        d = s.query(zscp.uid).filter(zscp.rid==rid).first()
        if not d:
            return {'code': -1, 'msg': '专属产品记录不存在'}
        uid = d.uid
        ywry = ''
        org = get_user_path(user.username)
        # uid = org.get('rid')
        path = org.get('path')
        # d = run_sql(f"select ifnull(path,'') path,username from sys_user where rid='{str(uid)}' limit 1")
        if org.get('rid')==None or org.get('rid')=='':
            return {'code': -1, 'msg': '业务人员不存在'}
        # path = d[0].get('path')[:100]
        ywry = user.username
        if path=='':
            return {'code': 0, 'msg': '业务人员不存在'}
        
        return {'code': 1, 'msg': '校验成功','data':ywry}
    except:
        logger.error(trace_error())
        return {'code': -1, 'msg': trace_error()}
    
# 专业产品转专属产品
@any_route('/api/saier/items/trans/zycp/check', methods=['POST', 'GET'])
@require_token
async def view_saier_items_trans_zycp(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        rid = j.get('rid')
        res = await items_trans_zycp_check(rid,user,s)

        return json_result(res.get('code'), res.get('msg'),res.get('data'))
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 专业产品转专属产品主函数
async def items_trans_zycp(j,user,s):
    try:
        flag = 0
        ywbm = ''
        rybh = ''
        path = ''
        ywry = j.get('ywry', '')
        module = j.get('module', '')
        kind = '专属转专业'
                                       
        errors = []
        rid = j.get('rid')
        d = s.query(zscp).filter(zscp.rid==rid).first()
        if not d:
            return {'code': -1, 'msg': '专属产品记录不存在'}
        
        # u = run_sql(f"select rid,path from sys_user where username='{str(ywry)}' limit 1")
        # if len(u)==0:
        #     return {'code': -1, 'msg': '业务人员不存在,请重新输入'}
        # uid = u[0].get('rid')
        # path = u[0].get('path')[:100]
        org = get_user_path(user.username)
        uid = org.get('rid')
        path = org.get('path')
        if uid==None or uid=="":
            return {'code': -1, 'msg': '业务人员不存在,请重新输入'}
        
        c = alchemy_object_to_dict(d)
        zycpbh = c.get('zycpbh')
        old_pid = c.get('rid')

        f = s.query(cjcp.rid).filter(cjcp.cpbh==str(zycpbh)).first()
        if f:
            return {'code': -1, 'msg': f"不好意思,{zycpbh}已有此专业产品,谢谢"}
        
        res = await items_trans_main_new(j,module,path,zycpbh,uid,c,user,s)
        if res.get('code')!=1:
            return res
        new_rid = res.get('rid')
        photos = res.get('photos')

        o = zscpsheet
        res = await items_trans_line(j,o,old_pid,uid,'zycjcpsheet',new_rid,s)
        if res.get('code')!=1:
            return res
        s.query(zscpsheet1).filter(zscpsheet1.pid==old_pid).delete()
        s.query(zscpsheet2).filter(zscpsheet2.pid==old_pid).delete()
        s.query(zscpsheet3).filter(zscpsheet3.pid==old_pid).delete()
        res = await insert_script_log(module,zycpbh,kind,'成功',user,str({"产品编号":zycpbh}),photos=photos,s=s)
        if res.get('code')!=1:
            return res
        s.delete(d)

        return {'code': 1, 'msg': '生成成功'}
    except:
        logger.error(trace_error())
        return {'code': -1, 'msg': trace_error()}

# 专业产品转专属产品
@any_route('/api/saier/items/trans/zycp', methods=['POST', 'GET'])
@require_token
async def view_saier_items_trans_zycp(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        res = await items_trans_zycp(j,user,s)
        if res.get('code')==1:
            s.commit()
        else:
            s.rollback()

        return json_result(res.get('code'), res.get('msg'))
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 专业产品更新货号主函数
async def items_import_cpbh(fn,user,module,ywry,retain):
    j = {}
    j['ywry'] = ywry
    j['retain'] = retain
    if not os.path.exists(fn):
        return {'code':-1,'msg':'无法找到文件'}
    wb = load_workbook(fn)
    path = config.get_today_upload_path()
    if not os.path.exists(path):
        make_dirs(path)
    cpbh_list = []
    for sheet in wb.worksheets:
        i = 2
        value = sheet.cell(i,1).value
        while value !="" and value != None:
            cpbh_list.append(str(value).strip())
            i = i +1
            value = sheet.cell(i,1).value
        break
                
    s = Session()
    j['rids'] = cpbh_list
    j['module'] = module
    try:
        res = await items_trans_zscp(j,user,s)
        if res.get('code')!=1:
            s.rollback()
            return res
        
        s.commit()
        return {'code':1,'msg':'操作完成'}
    except Exception as e:
        logger.error(trace_error())
        s.rollback()
        return {'code':-1,'msg':trace_error()}
    finally:
        s.close()
        wb.close()

# 专业产品更新货号
@any_route('/api/saier/items/import/cpbh', methods=['POST', 'GET'])
@require_token
async def api_saier_items_import_cpbh(request):
    try:
        form = await request.form()
        user = request.current_user
        file  = form.get('file.raw',None)
        file_name = form.get('file.name','')

        if is_none(file):
            return json_result(ERR_PARAM_NOT_ENOUGH)
        fn = get_tmp_abs_file_name(get_suffix(file_name))

        data = await file.read()
        write_file(fn,data,'wb',encoding=None)
        ywry = form.get('ywry')
        module = form.get('module')
        retain = form.get('retain',False)
        await asyncio.sleep(0.2)
        await messages.message(user.username,{'msg':'正在处理数据。。。','total':100,'progress':1,'kind':'0'},MSG_KIND_NORMAL) 

        res = await items_import_cpbh(fn,user,module,ywry,retain)
        if res.get('code')!=1:
            await asyncio.sleep(0.2)
            await messages.message(user.username,{'msg': res.get('msg'),'kind':'2'},MSG_KIND_NORMAL)
        else:
            await asyncio.sleep(0.2)
            await messages.message(user.username,{'msg':'数据处理完成。。。','total':100,'progress':100,'kind':'1'},MSG_KIND_NORMAL) 

        return json_result(res.get('code'),res.get('msg'),res)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1,trace_error())
    pass

# 专业产品更新单据品名英
async def items_update_djpmy_from_excel(fn,module,kind,user):
    if not os.path.exists(fn):
        return {'code':-1,'msg':'无法找到文件'}
    
    wb = load_workbook(fn)
    s = Session()
    try:
        for sheet in wb.worksheets:
            i = 2
            value = sheet.cell(i,1).value
            while value !="" and value != None:
                cpbh = str(value).strip()
                djpmy = sheet.cell(i,2).value
                d = s.query(cjcp).filter(cjcp.cpbh==cpbh,func.ifnull(cjcp.djpmy,'')=='').first()
                if d:
                    d.djpmy=djpmy
                    d.modi_uid = user.rid
                    d.mtime = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime())
                    s.add(d)
                    await insert_script_log(module, cpbh, kind, '成功', user, str({"产品编号:":cpbh,"单据品名英":djpmy}),s=s)
                d = s.query(zscp).filter(zscp.zycpbh==cpbh,func.ifnull(zscp.djpmy,'')=='').first()
                if d:
                    d.djpmy=djpmy
                    d.modi_uid = user.rid
                    d.mtime = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime())
                    s.add(d)
                    await insert_script_log('专属产品', cpbh, kind, '成功', user, str({"产品编号:":cpbh,"单据品名英":djpmy}),s=s)
                i = i +1
                value = sheet.cell(i,1).value

            s.commit()
            break
    
        return {'code':1,'msg':'操作完成'}
    except Exception as e:
        logger.error(trace_error())
        s.rollback()
        return {'code':-1,'msg':trace_error()}
    finally:
        s.close()
        wb.close()

# 专业产品更新原始采购价级退税点
async def items_update_yscgj_from_excel(fn,module,kind,user):
    if not os.path.exists(fn):
        return {'code':-1,'msg':'无法找到文件'}
    wb = load_workbook(fn)
    s = Session()
    try:
        m = get_module(module)
        o = get_model_by_table_name(m.table_name)
        for sheet in wb.worksheets:
            i = 2
            value = sheet.cell(i,1).value
            while value !="" and value != None:
                cpbh = str(value).strip()
                yscgj = sheet.cell(i,2).value
                gctd = sheet.cell(i,3).value
                flag = False
                d = s.query(o).filter(o.cpbh==cpbh,func.ifnull(o.uid,'')==user.rid).first()
                if d:
                    old_yscgj = float(d.yscgj)
                    if yscgj!=None and yscgj!='':
                        d.yscgj = yscgj
                        flag = True
                    old_gctd = float(gctd)
                    if gctd!=None and gctd!='':
                        d.gctd = gctd
                        flag = True
                    d.modi_uid = user.rid
                    d.mtime = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime())
                    if flag:
                        s.add(d)
                        await insert_script_log(module, cpbh, kind, '成功', user, str({"产品编号:":cpbh,"更新前原始采购价":old_yscgj,"更新后原始采购价":yscgj,"更新前退点":old_gctd,"更新后退点":gctd}),s=s)
                i = i +1
                value = sheet.cell(i,1).value

            s.commit()
            break
    
        return {'code':1,'msg':'操作完成'}
    except Exception as e:
        logger.error(trace_error())
        s.rollback()
        return {'code':-1,'msg':trace_error()}
    finally:
        s.close()
        wb.close()

# 专属产品批量低毛利
async def items_update_zdml_from_excel(fn,module,kind,user):
    if not os.path.exists(fn):
        return {'code':-1,'msg':'无法找到文件'}
    wb = load_workbook(fn)
    s = Session()
    try:
        m = get_module(module)
        o = get_model_by_table_name(m.table_name)
        for sheet in wb.worksheets:
            i = 2
            value = sheet.cell(i,1).value
            while value !="" and value != None:
                cpbh = str(value).strip()
                zdml = sheet.cell(i,2).value
                if (float(zdml) > 1) or (float(zdml) < -1):
                    mll = float(zdml)/100
                else:
                    mll = float(zdml)
                flag = False
                d = s.query(zscpsheet1).filter(zscpsheet1.cpbh==cpbh,zscpsheet1.khmc.like("BEST PRICE%%")).all()
                for r in d:
                    r.mll = mll*100
                    r.jzrq = '2999-12-31'
                    r.khmc = 'BEST PRICE LLC'
                    s.add(r)
                    flag = True
                d = s.query(cgjhsheet).filter(cgjhsheet.bjhh==cpbh,cgjhsheet.zdml>mll,cgjhsheet.mll>=mll,cgjhsheet.ggxm!='是').all()
                for r in d:
                    r.zdml = mll*100
                    r.ggxm = '是'
                    s.add(r)
                    flag = True
                if flag:
                    res = await insert_script_log(module, cpbh, kind, '成功', user, str({"产品编号:":cpbh,"毛 利 率":mll}),s=s)
                i = i +1
                value = sheet.cell(i,1).value

            s.commit()
            break
    
        return {'code':1,'msg':'操作完成'}
    except Exception as e:
        logger.error(trace_error())
        s.rollback()
        return {'code':-1,'msg':trace_error()}
    finally:
        s.close()
        wb.close()

# 专属产品更新EAN条码
async def items_update_barcode_from_excel(fn,module,kind,user):
    if not os.path.exists(fn):
        return {'code':-1,'msg':'无法找到文件'}
    wb = load_workbook(fn)
    s = Session()
    try:
        # m = get_module(module)
        # o = get_model_by_table_name(m.table_name)
        for sheet in wb.worksheets:
            i = 2
            value = sheet.cell(i,1).value
            while value !="" and value != None:
                cpbh = str(value).strip()
                krtm = sheet.cell(i,2).value
                lvkrtm = sheet.cell(i,3).value
                RUkrtm = sheet.cell(i,4).value
                flag = False
                filters = []
                data = {'产品编号':cpbh,'lv客人条码':lvkrtm,'客人条码':krtm}
                d = s.query(zscp).filter(zscp.cpbh==cpbh).filter(*filters).all()
                for r in d:
                    r.krtm = krtm
                    r.lvkrtm = lvkrtm
                    if RUkrtm!=None and RUkrtm!="":
                        r.RUkrtm = RUkrtm
                        data['RU客人条码']=RUkrtm
                    s.add(r)
                    flag = True

                if flag:
                    res = await insert_script_log(module, cpbh, kind, '成功', user, str(data),s=s)
                i = i +1
                value = sheet.cell(i,1).value

            s.commit()
            break
    
        return {'code':1,'msg':'操作完成'}
    except Exception as e:
        logger.error(trace_error())
        s.rollback()
        return {'code':-1,'msg':trace_error()}
    finally:
        s.close()
        wb.close()


# 专属产品更新EAN条码
async def chyjh_update_from_excel(fn,module,kind,user):
    if not os.path.exists(fn):
        return {'code':-1,'msg':'无法找到文件'}
    wb = load_workbook(fn)
    s = Session()
    try:
        # m = get_module(module)
        # o = get_model_by_table_name(m.table_name)
        for sheet in wb.worksheets:
            i = 2
            value = sheet.cell(i,1).value
            while value !="" and value != None:
                wxfp = str(value).strip()
                khmc = sheet.cell(i,2).value
                hglx = sheet.cell(i,3).value
                dq1 = sheet.cell(i,4).value
                cdmc = sheet.cell(i,5).value
                etd = sheet.cell(i,6).value
                cyka = sheet.cell(i,7).value
                mdka = sheet.cell(i,8).value
                zgrq = sheet.cell(i,9).value
                dzry = sheet.cell(i,10).value
                bzsm = sheet.cell(i,11).value

                filters = []
                data = {'外销发票':wxfp,'客户名称':khmc,'货柜类型':hglx,'地区1':dq1,'产地名称':cdmc,'ETD':etd,'出运口岸':cyka,'目的口岸':mdka,'装柜日期':zgrq,'单证人员':dzry,'备注说明':bzsm}
                r = s.query(chyjh).filter(chyjh.wxfp==wxfp).filter(*filters).first()
                if r:
                    r.khmc = khmc 
                    r.hglx = hglx 
                    r.zgdd = dq1 
                    r.cdmc = cdmc 
                    r.etd = etd 
                    r.zgrq = zgrq 
                    r.dzry = dzry 
                    r.mtime = time.strftime('%Y-%m-%d %H:%M:%S')
                    r.modi_uid = user.rid
                    r.cyka = cyka 
                    r.mdka = mdka 
                    s.add(r)
                    flag = True

                r = s.query(wxht).filter(wxht.order_id==wxfp).filter(*filters).first()
                if r:
                    r.hdmc = cdmc 
                    r.hglx = hglx 
                    r.yjcq = etd 
                    r.cdmc = cdmc 
                    r.zysx = bzsm 
                    r.zgrq = zgrq 
                    r.dzry = dzry 
                    r.mtime = time.strftime('%Y-%m-%d %H:%M:%S')
                    r.modi_uid = user.rid
                    r.cyka = cyka 
                    r.mdka = mdka
                    s.add(r)
                    flag = True

                if flag:
                    res = await insert_script_log(module, wxfp, kind, '成功', user, str(data),s=s)
                i = i +1
                value = sheet.cell(i,1).value

            s.commit()
            break
    
        return {'code':1,'msg':'操作完成'}
    except Exception as e:
        logger.error(trace_error())
        s.rollback()
        return {'code':-1,'msg':trace_error()}
    finally:
        s.close()
        wb.close()

# 专业产品更新数据
@any_route('/api/saier/module/update/from/excel', methods=['POST', 'GET'])
@require_token
async def api_saier_module_update_from_excel(request):
    try:
        form = await request.form()
        user = request.current_user
        file  = form.get('file.raw',None)
        file_name = form.get('file.name','')
        module = form.get('module')
        kind = form.get('kind')
        if is_none(file):
            return json_result(-1,'文件上传失败')
        fn = get_tmp_abs_file_name(get_suffix(file_name))

        data = await file.read()
        write_file(fn,data,'wb',encoding=None)

        await asyncio.sleep(0.2)
        await messages.message(user.username,{'msg':'正在处理数据。。。','total':100,'progress':1,'kind':'0'},MSG_KIND_NORMAL) 

        if (module == '专业产品' or module=='专属产品') and kind == '更新报关品名英':
            res = await items_update_djpmy_from_excel(fn,module,kind,user)
        elif (module == '专业产品' or module=='专属产品') and kind == '更新原始采购价及点数':
            res = await items_update_yscgj_from_excel(fn,module,kind,user)
        elif module == '专属产品' and kind == '批量低毛利':
            res = await items_update_zdml_from_excel(fn,module,kind,user)
        elif module == '专属产品' and kind == '批量引入EAN条码':
            res = await items_update_barcode_from_excel(fn,module,kind,user)
        elif module == '出运计划' and kind == '订柜信息':
            res = await chyjh_update_from_excel(fn,module,kind,user)
        else:
            res = {'code':-1,'msg':'参数错误'}

        if res.get('code')!=1:
            await asyncio.sleep(0.2)
            await messages.message(user.username,{'msg': res.get('msg'),'kind':'2'},MSG_KIND_NORMAL)
        else:
            await asyncio.sleep(0.2)
            await messages.message(user.username,{'msg':'数据处理完成。。。','total':100,'progress':100,'kind':'1'},MSG_KIND_NORMAL) 

        return json_result(res.get('code'),res.get('msg'),res)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1,trace_error())
    pass

# 更新报关品名英权限校验
@any_route('/api/saier/module/user/check', methods=['POST', 'GET'])
@require_token
async def view_saier_module_user_check(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        position = j.get('position')
        d = run_sql(f"select rid from sys_user where username='{user.username}' and (position like '%{position}%' or memo like '%{position}%')")
        # 测试时暂时注销
        if len(d) == 0:
            return json_result(-1, '用户权限校验失败')

        return json_result(1, '校验成功')
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 采购人员修改时校验数据
@any_route('/api/saier/items/cgry/change', methods=['POST', 'GET'])
@require_token
async def view_saier_items_cgry_change(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        cgry = j.get('cgry', '')
        path = ''
        cpbh = j.get('cpbh', '')
        cgry_old = ''
        flag = False
        rid = j.get('rid', '')
        d = s.query(cjcp.cgry).filter(cjcp.rid==rid).first()
        if d:
            cgry_old = str(d.cgry)
        if cgry == cgry_old:
            return json_result(1,'新采购员和原来的一致',cgry)
        
        # t = run_sql(f"select rid,path from sys_user where username='{cgry}'")
        # if len(t)>0:
        #     path = t[0].get('path')
        org = get_user_path(cgry)
        # uid = org.get('rid')
        path = org.get('path')
        if path == '':
            return json_result(-1,'无此人员,更改无效',cgry_old)

        if cgry!=None and cgry!='':
            c = s.query(cyzglsheet.bz).filter(cyzglsheet.xm==user.username,cyzglsheet.zm=="专业专属更改采购人员").first()
            if not c:
                flag = False
                if cpbh!='' and cpbh!=None:
                    m = s.query(cgjhsheet).filter(cgjhsheet.bjhh==str(cpbh), func.ifnull(cgjhsheet.bjhh,'')!='', cgjhsheet.cgry!="待定").first()
                    if m:
                        flag = True
                    m = s.query(cghtsheet).filter(cghtsheet.bjhh==str(cpbh), func.ifnull(cghtsheet.bjhh,'')!='').first()
                    if m:
                        flag = True
                    if flag:
                        u = s.query(ywrybiao).filter(ywrybiao.yhm==cgry, ywrybiao.rybh.not_like("%LZ%")).first()
                        if u:
                            return json_result(-1,'这个产品已有采购合同或采购计划，您无权更改采购人员请通知相关人员更改',cgry_old)
        
        return json_result(1, '校验成功')
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error(),'')
    finally:
        s.close()

# 专业产品从excel导入数据时权限校验
@any_route('/api/saier/module/user/group/check', methods=['POST', 'GET'])
@require_token
async def view_saier_module_user_group_check(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        group = j.get('group','专业产品引入')
        d = run_sql(f"select rid from cyzglsheet where xm='{user.username}' and zm='{group}'")
        if len(d) == 0:
            return json_result(-1, '用户权限校验失败')

        return json_result(1, '校验成功')
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# excel导入数据时对应模块的表名和条件字段
exel_module = {
    "专业产品":{"key_fields":{'产品编号':{}},"child":[]},
    "专属产品":{"key_fields":{'产品编号':{}},"child":[{"name":"工厂报价"}]}
}

# excel导入数据时获取表头字段
async def get_excel_fields(sheet,t,name='专业产品'):
    try: 
        # return {"code":-1,"msg":f"无法找到{name}对应的表"}
        names = name.split('.')
        if len(names)==1:
            table = t.table_name
        elif(len(names)==2):
            group = t.group_by_name(names[1])
            if not group:
                return {"code":-1,"msg":f"无法找到{name}对应的表"}
            if not group.is_table:
                return {"code":-1,"msg":f"{name}不是子表"}
            table = group.db.table_name
        
        j = 1
        fields = []
        columns= []
        key_fields = exel_module.get(name).get('key_fields')
        child = exel_module.get(name).get('child')
        _name = sheet.cell(1,j).value
        while _name !="" and _name != None:
            if _name in columns:
                return {"code":-1,"msg":f"{name}的{_name}字段重复"}
            columns.append(_name)
            field = t.field_by_full_name(f'{name}.{_name}')
            if field!=None and field!="":
                fields.append({
                    'name':_name,
                    'field':field.db.name,
                    'col':j,
                    'kind':field.db.kind,
                    'child':False,
                    'check':False
                })
                if _name in key_fields:
                    key_fields[_name] = {'field':field.db.name,'col':j}
            else:
                fields.append({
                    'name':_name,
                    'field':None,
                    'col':j,
                    'kind':None,
                    'child':True,
                    'check':False
                })
            j+=1
            _name = sheet.cell(1,j).value

        return {"code":1,"msg":"获取成功","fields":fields,"columns":columns,"key_fields":key_fields,"child":child}
    except Exception as e:
        logger.error(trace_error())
        return {'code':-1,'msg':trace_error()}

# 专业产品从excel导入数据主方法
async def items_insert_from_excel(fn,module,user):
    if not os.path.exists(fn):
        return {'code':-1,'msg':'无法找到文件'}

    wb = load_workbook(fn)
    s = Session()
    try:
        for sheet in wb.worksheets:
            i = 2
            t = get_module(module)
            res = await get_excel_fields(sheet,t,module)
            if res.get('code')==-1:
                return res
            table = t.table_name
            fields = res.get('fields')
            columns = res.get('columns')
            key_fields = res.get('key_fields')
            child = res.get('child')
            while_flag = True
            insert_flag = True 
            error_list = []
            o = get_model_by_table_name(table)
            filters = []
            for k,v in key_fields.items():
                c = v.get('col')
                if sheet.cell(i,c).value==None or sheet.cell(i,c).value=='':
                    while_flag = False
                    break
                else:
                    filters.append(getattr(o,v.get('field'))==sheet.cell(i,c).value)

            while while_flag:
                '''下面的数据检验和另外的字段数据取数可以到时封装一个函数'''
                cgry = ''
                cpfl = ''
                yjfl = ''
                ejfl = ''
                path = ''
                cgry_i = columns.index('采购人员')
                cpfl_i = columns.index('产品大类')
                yjfl_i = columns.index('一级分类')
                ejfl_i = columns.index('二级分类')
                cpbh_i = columns.index('产品编号') 
                if cgry_i>=0:
                    cgry = sheet.cell(i,cgry_i+1).value
                if cpfl_i>=0:
                    cpfl = sheet.cell(i,cpfl_i+1).value
                if yjfl_i>=0:
                    yjfl = sheet.cell(i,yjfl_i+1).value
                if ejfl_i>=0:
                    ejfl = sheet.cell(i,ejfl_i+1).value
                if cpbh_i>=0:
                    cpbh = sheet.cell(i,cpbh_i+1).value

                if (cgry==None or cgry==''):
                    insert_flag = False
                    error_list.append(f"第{i-1}行数据产品编号{cpbh}采购人员为空")
                    await insert_script_log(module, cpbh, 'Excel导入', '采购人员为空', user, str({"产品编号:":cpbh, "采购人员":cgry}))
                else:
                    org = get_user_path(cgry)
                    uid = org.get('rid')
                    path = org.get('path')
                    # u = run_sql(f"select path from sys_user where username='{cgry}'")
                    if uid!=None and uid!="":
                        path = u[0].get('path')
                    else:
                        insert_flag = False
                        error_list.append(f"第{i-1}行数据产品编号{cpbh}采购人员不存在")
                        await insert_script_log(module, cpbh, 'Excel导入', '采购人员不存在', user, str({"产品编号:":cpbh, "采购人员":cgry}))

                if (cpfl==None or cpfl=='') and (yjfl==None or yjfl=='') and (ejfl==None or ejfl==''):
                    insert_flag = False
                    error_list.append(f"第{i-1}行数据产品编号{cpbh}产品分类错误")
                    await insert_script_log(module, cpbh, 'Excel导入', '分类数据为空', user, str({"产品编号:":cpbh,"产品大类": cpfl,"一级分类":yjfl,"二级分类":ejfl}))

                l = s.query(zycpfenglb.rid).filter(or_(and_(zycpfenglb.cpfl==cpfl,func.ifnull(yjfl,"")=="",func.ifnull(ejfl,"")=="",),
                    and_(zycpfenglb.cpfl==cpfl,zycpfenglb.yjfl==yjfl,func.ifnull(ejfl,"")==""),and_(zycpfenglb.cpfl==cpfl,zycpfenglb.yjfl==yjfl,zycpfenglb.ejfl==ejfl))).first()
                if not l:
                    insert_flag = False
                    error_list.append(f"第{i-1}行数据产品编号{cpbh}产品分类错误")
                    await insert_script_log(module, cpbh, 'Excel导入', '分类记录未找到', user, str({"产品编号:":cpbh,"产品大类": cpfl,"一级分类":yjfl,"二级分类":ejfl}))

                d = s.query(o.rid).filter(*filters).first()
                if d:
                    insert_flag = False
                    error_list.append(f"第{i-1}行数据产品编号{cpbh}产品编号重复")
                    await insert_script_log(module, cpbh, 'Excel导入', '记录重复', user, str({"产品编号:":cpbh}))
                '''上面的数据检验和另外的字段数据取数可以到时封装一个函数'''
                # 执行数据库数据插入
                if insert_flag:
                    m = o()
                    rid = get_uuid()
                    m.rid = rid
                    m.uid = user.rid
                    m.path = path
                    for r in fields:
                        if r.get('child')==True:
                            continue
                        if sheet.cell(i,r.get('col')).value!=None and sheet.cell(i,r.get('col')).value!='':
                            setattr(m,r.get('field'),sheet.cell(i,r.get('col')).value)

                        flmc = str(cpfl)
                        if (str(cpfl)  != '') and (str(yjfl) != ''):
                            flmc = str(cpfl) + '\\' +  str(yjfl)
                        if (str(cpfl)  != '') and (str(yjfl) != '') and (str(ejfl) != ''):
                            flmc = str(cpfl) + '\\' +  str(yjfl) + '\\' +  str(ejfl)
                        setattr(m,'flmc',flmc)

                    await insert_script_log(module, cpbh, 'Excel导入', '成功', user, str({"产品编号:":cpbh,"rid": rid}),s=s)
                    s.add(m)

                    for l in child:
                        g = t.group_by_name(l.get('name'))
                        khhh_i = columns.index('客人货号') 
                        cgdj_i = columns.index('采购单价')
                        bztj_i = columns.index('外箱体积') 
                        sccj_i = columns.index('生产厂家')
                        khhh = ''
                        cgdj = 0
                        bztj = 0
                        sccj = ''
                        if khhh_i>=0:
                            khhh = sheet.cell(i,khhh_i+1).value
                        if cgdj_i>=0:
                            cgdj = sheet.cell(i,cgdj_i+1).value
                        if bztj_i>=0:
                            bztj = sheet.cell(i,bztj_i+1).value
                        if sccj_i>=0:
                            sccj = sheet.cell(i,sccj_i+1).value
                        x = get_model_by_table_name(g.db.table_name) 
                        m = x()
                        m.rid = get_uuid()
                        m.uid = user.rid
                        m.pid = rid
                        for r in fields:
                            if r.get('child')==True and r.get('check')==False and r.get('field')==None:
                                field = t.field_by_full_name(f"{module}.工厂报价.{r.get('name')}")
                                if field:
                                    r['field'] = field.db.name
                                    r['kind'] = field.db.kind
                                r['check'] = True
                            if sheet.cell(i,r.get('col')).value!=None and sheet.cell(i,r.get('col')).value!='':
                                setattr(m,r.get('field'),sheet.cell(i,r.get('col')).value)
                        m.xjry = cpbh
                        m.khhh = khhh
                        m.dj = cgdj
                        m.tj = bztj
                        m.sccj1 = sccj
                        m.sccj11 = bztj
                        s.add(m)

                i = i +1
                filters = []
                insert_flag = True
                for k,v in key_fields.items():
                    c = v.get('col')
                    if sheet.cell(i,c).value==None or sheet.cell(i,c).value=='':
                        while_flag = False
                        break
                    else:
                        filters.append(getattr(o,v.get('field'))==sheet.cell(i,c).value)

            s.commit()
        if len(error_list)>0:
            return {'code':-1,'msg':"导入存在错误,详细错误请查看脚本日志," + "\n".join(error_list)}
        return {'code':1,'msg':'操作完成'}
    except Exception as e:
        logger.error(trace_error())
        s.rollback()
        return {'code':-1,'msg':trace_error()}
    finally:
        s.close()
        wb.close()

# 专业产品从excel导入数据
@any_route('/api/saier/module/insert/from/excel', methods=['POST', 'GET'])
@require_token
async def api_saier_module_insert_from_excel(request):
    try:
        form = await request.form()
        user = request.current_user
        file  = form.get('file.raw',None)
        file_name = form.get('file.name','')
        module = form.get('module')
        kind = form.get('kind')
        if is_none(file):
            return json_result(ERR_PARAM_NOT_ENOUGH)
        fn = get_tmp_abs_file_name(get_suffix(file_name))

        data = await file.read()
        write_file(fn,data,'wb',encoding=None)

        await asyncio.sleep(0.2)
        await messages.message(user.username,{'msg':'正在处理数据。。。','total':100,'progress':1,'kind':'0'},MSG_KIND_NORMAL) 

  
        res = await items_insert_from_excel(fn,module,user)

        if res.get('code')!=1:
            await asyncio.sleep(0.2)
            await messages.message(user.username,{'msg': res.get('msg'),'kind':'2'},MSG_KIND_NORMAL)
        else:
            await asyncio.sleep(0.2)
            await messages.message(user.username,{'msg':'数据处理完成。。。','total':100,'progress':100,'kind':'1'},MSG_KIND_NORMAL) 

        return json_result(res.get('code'),res.get('msg'),res)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1,trace_error())
    pass

# 专业产品工厂报价更新出货数
@any_route('/api/saier/items/supplier/update/qty', methods=['POST', 'GET'])
@require_token
async def api_saier_items_supplier_update_qty(request):
    s = Session()
    try:
        j = await request.json()
        cpbh = j.get('cpbh')
        user = request.current_user
        lines = j.get('lines')
        for r in lines:
            cght = r.get('cght')
            d = s.query(func.ifnull(func.sum(func.ifnull(cymxsheet.chsl,0)),0).label('chsl')).filter(cymxsheet.cght==cght,cymxsheet.fpsb1=='是',func.ifnull(cymxsheet.cpbh,'')==cpbh).first()
            if d:
               r['chsl'] = float(d.chsl)
            else:
                r['chsl'] = 0

        return json_result(1,'取数成功',lines)
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1,trace_error())
    finally:
        s.close()
    pass


# 专属产品保存前获取杂项数据
@any_route('/api/saier/items/get/zx/data', methods=['POST', 'GET'])
@require_token
async def view_saier_items_get_zx_data(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        data = []
        rid = j.get('rid', '')
        cpbh = j.get('cpbh', '')
        job = j.get('job', 1)
        gc_data = j.get('xzgc')
        wfgs_n = j.get('wfgs')
        sbbz = j.get('sbbz')
        zsbh = j.get('zsbh')
        cgry = j.get('cgry', '')
        dycp = 0
        d = s.query(zscp.rid).filter(zscp.rid!=rid,zscp.cpbh==cpbh).first()
        if d:
            return json_result(-1, f"产品编号【{cpbh}】已存在,请重新输入")
        
        if job==0:
            d = s.query(zx).filter(zx.ly=='最低毛利率').all()
            for r in d:
                o = {}
                o['rid'] = get_uuid()
                o['pid'] = rid
                o['khmc'] = r.mc
                if r.cs != None:
                    o['mll'] = float(r.cs)
                else:
                    o['mll'] = 0
                o['cpbh'] = cpbh
                data.append(o)

        ywry = j.get('ywry', '')
        bjyw = ''
        cpbh_o = ''
        d = s.query(zscp.bjyw,zscp.cpbh).filter(zscp.rid==rid).first()
        if d:
            bjyw = str(d.bjyw)
            cpbh_n = str(d.cpbh)
            d = run_sql(f"select rid from sys_user where username='{bjyw}'")
            if len(d) == 0:
                bjyw = ''

        if cpbh_o!="" and cpbh!=cpbh_o:
            t = s.query(wxhtsheet.rid).filter(wxhtsheet.bjhh==str(cpbh_o)).first()
            if t:
                return {'code': -1, 'msg': '外销合同中已存在此货号'}
            m = get_module(module)
            o = get_model_by_table_name(m.table_name)
            filters=[o.rid==rid]
            d = s.query(o).filter(*filters).first()
            if d:
                cpbh = d.cpbh
                pid = d.rid
                c = s.query(zscpsheet5).filter(zscpsheet5.cpbh==cpbh_o).all()
                for r in c:
                    r.cpbh = cpbh
                    r.pid = rid
                    s.add(r)


        xzcj = gc_data.get('xzcj')
        xzbh = gc_data.get('cs_id')
        if xzcj != None and xzcj != '' and xzcj!='待定':
            d = s.query(zycs.cs_id,zycs.rid).filter(zycs.company_name==xzcj).first()
            if d:
                pid = d.rid
                xzcg = False
                syq = gc_data.get('syq')
                if syq!=None and syq!="":
                    xzcg = True
                res = await zycssheet2_new(gc_data, pid, cpbh, user, s, xzcg, True)
                if res.get('code')!=1:
                    return json_result(res.get('code'),res.get('msg'))

        if xzcj != None and xzcj != '' and xzbh!=None and xzbh!='':
            filters = []
            syq = gc_data.get('syq')
            d = s.query(cyzglsheet.rid).filter(cyzglsheet.xm==user.username,cyzglsheet.zm=="采购新增对应工厂").first()
            if d:
                filters.append(zscpsheet5.cgry==syq)
            xzcg = False
            if syq!=None and syq!="":
                xzcg = True
            res = await zycssheet5_new(gc_data, rid, cpbh, xzbh, xzcj, user, s, xzcg, True, filters)
            if res.get('code')!=1:
                return json_result(res.get('code'),res.get('msg'))
            else:
                dycp = res.get('data')
                
        if wfgs_n == None or wfgs_n == '':
            d = s.query(wfgs).first()
            if d:
                wfgs_n = d.wfgs
        zs_sid = 1
        if zsbh == None or zsbh == '':
            d = s.query(zscp.sid).order_by(zscp.sid.desc()).first()
            if d:
                zs_sid = int(d.sid) + 1

        test_data={}
        if sbbz == '是':
            d = s.query(zscp.csgj,zscp.cszl,zscp.tgrq,zscp.yxqx).filter(zscp.cpbh==cpbh).first()
            if d:
                test_data = alchemy_object_to_dict(d)

        s.commit()
        return json_result(1, '取数成功', {'tsml':data,'ywry':bjyw,'wfgs':wfgs_n,'zs_sid':zs_sid,'dycp':dycp,'test_data':test_data})
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error(),'')
    finally:
        s.close()


# 专属产品更新采购人员主函数
async def items_update_purchase(j,user,s):
    try:
        cpbh = j.get('cpbh')
        module = j.get('module')
        csbh = j.get('csbh')
        cgry = j.get('cgry')
        path = ''
        uid = ''
        # d = run_sql(f"select path,rid from sys_user where username='{cgry}'")
        # if len(d) == 0:
        #     return {'code': -1, 'msg':'无此人员,请重新输入'}
        org = get_user_path(cgry)
        path = org.get('path')
        logger.error('aa')
        logger.error(path)
        if path=='':
            return {'code': -1, 'msg':'无此人员,请重新输入'}
        uid = org.get('rid')
        if uid == '' or uid == None:
            uid = user.rid
        if cgry == '':
            cgry = user.username
        d = s.query(bjsheet.rid).filter(bjsheet.bjhh==cpbh).first()
        if d:
            return {'code':-1,'msg':"该产品已经有客户报价,采购人员修改失败"}
        d = s.query(wxhtsheet.rid).filter(wxhtsheet.bjhh==cpbh).first()
        if d:
            return {'code':-1,'msg':"该产品已经有外销合同,采购人员修改失败"}
        d = s.query(cghtsheet.rid).filter(cghtsheet.bjhh==cpbh).first()
        if d:
            return {'code':-1,'msg':"该产品已经有采购合同,采购人员修改失败"}
        m = get_module(module)
        o = get_model_by_table_name(m.table_name)
        # filters=[o.rid==rid]
        d = s.query(o).filter(o.cpbh==cpbh).first()
        if not d:
            return {'code': -1, 'msg': f"未找到产品编号{cpbh}厂商编号{csbh}的记录"}
        cgry_old = d.cgry
        c = s.query(zscpsheet5).filter(zscpsheet5.cpbh==cpbh,zscpsheet5.gcID==csbh).all()
        for r in c:
            res = module_share_new('对应工厂', [cgry], r.rid, user, s, [r.cgry])
            if res.get('code')!=1:
                return res
            r.cgry = cgry
            s.add(r)
        s.query(zycssheet2).filter(zycssheet2.cpbh==cpbh,zycssheet2.cs_id==csbh).all()
        for r in c:
            r.syq = cgry
            s.add(r)
        c = s.query(ywrybiao.ssdq).filter(ywrybiao.yhm==cgry).first()
        if c:
            ssdq = c.ssdq
            l = s.query(cgjhsheet).filter(cgjhsheet.bjhh==cpbh,cgjhsheet.sccj1id==csbh).all()
            for r in l:
                r.cgry = cgry
                r.xddd = ssdq
                s.add(r)
        res = module_share_new(module, [cgry], d.rid, user, s, [d.cgry])
        if res.get('code')!=1:
            return res
        if d.gcID == csbh:
            d.cgry = cgry
            d.uid = uid
            s.add(d)

        res = await insert_script_log(module,cpbh,'采购人员修改','成功',user,str({"产品编号":cpbh,"厂商编号":csbh,"原采购人员":cgry_old, "新采购人员":cgry}),None,s)
        if res.get('code')!=1:
            return res
        
        return {'code': 1, 'msg': '更新成功'}
    except:
        logger.error(trace_error())
        return {'code': -1, 'msg': trace_error()}

# 专属产品更新采购人员
@any_route('/api/saier/items/update/purchase', methods=['POST', 'GET'])
@require_token
async def view_saier_items_update_purchase(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        res = await items_update_purchase(j,user,s)
        if res.get('code')==1:
            s.commit()
        else:
            s.rollback()
        return json_result(res.get('code'), res.get('msg'))
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 更新客人货号权限校验
@any_route('/api/saier/items/krhh/check', methods=['POST', 'GET'])
@require_token
async def view_saier_items_krhh_check(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        krhh = j.get('krhh')
        d = run_sql(f"select rid from cjcp where krhh='{krhh}' limit 1")
        if len(d) > 0:
            return json_result(-1, '专业产品中已存在')
        d = run_sql(f"select rid from zscp where krhh='{krhh}' limit 1")
        if len(d) > 0:
            return json_result(-1, '专属产品中已存在')
        return json_result(1, '校验成功')
    except:
        # s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 专属产品开票申请
@any_route('/api/saier/items/gckm/insert', methods=['POST', 'GET'])
@require_token
async def view_saier_items_gckm_insert(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        res = await items_gckm_insert(j,user,s)
        if res.get('code')!=1:
            s.rollback()
        else:
            s.commit()

        return json_result(res.get('code'), res.get('msg'),res.get('data'))
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 专属产品开票申请
async def items_gckm_insert(j,user,s):
    try:
        cpbh = j.get('cpbh')
        d = s.query(gckm.rid).filter(gckm.kmbh==cpbh).first()
        if d:
            return {'code':0, 'msg':'请注意此产品已有开模审请','data':d.rid}
        ywbm = ''
        wfgs_n = ''
        d = s.query(ywrybiao.bm,ywrybiao.ssdq,ywrybiao.wfgs).filter(ywrybiao.yhm==user.username).first()
        if d:
            ywbm = d.bm
            wfgs_n = d.wfgs

        m = gckm()
        for k,v in j.items():
            if k in SYS_FIELDS:
                continue
            setattr(m,k,v)
        rid = get_uuid()
        m.rid = rid
        m.uid = user.rid
        m.ctime = time.strftime("%Y-%m-%d %H:%M:%S")
        m.zshh = cpbh
        m.ywbm = ywbm
        m.wfgs = wfgs_n
        m.kmbh = cpbh
        m.wyzd = str(cpbh) + "_" + str(j.get("sccj", ''))
        s.add(m)
        res = await insert_script_log('专属产品', cpbh, '开模申请', '成功', user, str({"产品编号:":cpbh}),s=s)
        if res.get('code')!=1:
            return res
        return {'code':1,'msg':'操作完成', 'data':rid}
    except Exception as e:
        logger.error(trace_error())
        return {'code':-1,'msg':trace_error()}
    
# 专属产品生成客户报价
async def items_quotation_insert(j,user,s):
    try:
        bj_id = j.get('bj_id')
        rids = j.get('rids')
        ywbm = ''
        wfgs_n = ''
        d = s.query(ywrybiao.bm,ywrybiao.ssdq,ywrybiao.wfgs).filter(ywrybiao.yhm==user.username).first()
        if d:
            ywbm = d.bm
            wfgs_n = d.wfgs
        bzyq_n = ''
        d = s.query(bzyq.bzyq).first()
        if d:
            bzyq_n = d.bzyq
        flag = 0
        m = s.query(bj).filter(bj.bj_id==bj_id).first()
        if m:
            pid = m.rid
            flag = 1
        else:
            cyka_n = ''
            d = s.query(cyka.ywmc).first()
            if d:
                cyka_n = d.ywmc
            ysfs_n = ''
            d = s.query(ysfs.ysfs).first()
            if d:
                ysfs_n = d.ysfs
            hhl = 0
            d = s.query(hbdm.hhl).filter(hbdm.hbdm=="USD$").first()
            if d:
                hhl = float(d.hhl)
            pid = get_uuid()
            m = bj()
            m.bj_id = bj_id
            m.rid = pid
            m.uid = user.rid
            m.dateis = time.strftime('%Y-%m-%d')
            m.ctime = time.strftime('%Y-%m-%d %H:%M:%S')
            m.wyzd = 'bj' + bj_id + user.username + time.strftime('%Y-%m-%d %H:%M:%S')
            m.qsn = time.strftime('%Y')
            m.qsy = time.strftime('%m')
            m.wfgs = wfgs_n
            m.ywz = ywbm
            m.yxqx = time.strftime('%Y-%m-%d')
            m.bjzt = '使用中'
            m.hbdm = 'USD$'
            m.bjjg= '待审批'
            m.bjql= '待审批'
            m.ykbz= '否'
            m.yjbl = 1
            m.bjql1= '待审批'
            m.tgsb= '否'
            m.fxsb= '否'
            m.sfms= '否'
            m.xgck= '否'
            m.cgqx= '有'
            m.cyka=  cyka
            m.ysfs = ysfs
            m.huilv = hhl
            m.bxjc= 110
            m.bxbl = 0.78
            m.spsq = ''
            m.bjsp = ''
            if wfgs_n == '宁波优景进出口有限公司':
                m.bjsh = '123'
            else:
                m.bjsh = '1'
            
        spsq = m.spsq
        bjsp = m.bjsp
        if (spsq!=None and spsq != '') and (bjsp!=None and bjsp != ''):
            return {'code':-1,'msg':'不好意思,此报价已提交,不能追加'}
        
        m.mtime = time.strftime('%Y-%m-%d %H:%M:%S')
        m.modi_uid = user.rid
        s.add(m)
        item_list = []
        for rid in rids:
            c = s.query(zscp).filter(zscp.rid==rid).first()
            if not c:
                continue
            cpbh = str(c.cpbh)
            if flag == 1:
                d = s.query(bjsheet.rid).filter(bjsheet.pid==pid,bjsheet.bjhh==cpbh).first()
                if d:
                    continue
            c = alchemy_object_to_dict(c)
            m = bjsheet()
            for k,v in c.items():
                if k in SYS_FIELDS:
                    continue
                setattr(m,k,v)
            rid = get_uuid()
            m.rid = rid
            m.pid = pid
            m.uid = user.rid
            m.ctime = time.strftime("%Y-%m-%d %H:%M:%S")
            m.dw = c.get('bzdw')
            m.jg = c.get('mjfob',0)
            m.nhrl = c.get('nhrl',0) if c.get('nhrl',0)<1 else 1
            m.wxrl = c.get('bzrl',0)
            m.tj = c.get('bztj',0)
            m.mz = c.get('mxmz',0)
            m.jz = c.get('mxjz',0)
            m.mxmz = c.get('chpkzh',0)
            m.cz = c.get('caiziz')
            m.cply1 = c.get('topcz')
            m.xdsb = '无'
            m.nhwx = 1
            m.hbdm = 'USD$'
            m.jsfs = jsfs
            m.bjrq = time.strftime('%Y-%m-%d')
            m.yxqx = time.strftime('%Y-%m-%d')
            m.bzyq = bzyq_n
            m.hhbz = cpbh
            m.sfdr = '否'
            m.tps = 1
            m.gcpf = 3
            m.pfsb = '无'
            m.sfxz = '否'
            m.xdsb1 = '无'
            m.wxxz = '否'
            m.ywbf = '无'
            m.bjjd = 1
            s.add(m)

            item_list.append(cpbh)

        res = await insert_script_log('专属产品', bj_id, '当前产品生成报价', '成功', user, str({"产品编号:":item_list,"报价单号":bj_id}),s=s)
        if res.get('code')!=1:
            return res
        
        return {'code':1,'msg':'操作完成', 'data':rid}
    except Exception as e:
        logger.error(trace_error())
        return {'code':-1,'msg':trace_error()}
    

# 专属产品生成客户报价
@any_route('/api/saier/items/quotation/insert', methods=['POST', 'GET'])
@require_token
async def view_saier_items_quotation_insert(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        res = await items_quotation_insert(j,user,s)
        if res.get('code')!=1:
            s.rollback()
        else:
            s.commit()

        return json_result(res.get('code'), res.get('msg'),res.get('data'))
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


# 专属产品获取客户报价单号
@any_route('/api/saier/items/quotation/get', methods=['POST', 'GET'])
@require_token
async def view_saier_items_quotation_get(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        data = []
        d = s.query(bj.bj_id).filter(bj.uid==user.rid).order_by(bj.bj_id.desc()).limit(100).all()
        data = [r.bj_id for r in d]
        return json_result(1, '取数成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()


def pixels_to_points(value, dpi=96):
    """96 dpi, 72i"""
    return value * 72 / dpi


def points_to_pixels(value, dpi=96):
    return int(math.ceil(value * dpi / 72))

def offset_img(img, col, row, x_pad=4, y_pad=25):
    """精确设置图片位置，偏移量以万为单位进行微调吧，具体计算公式太麻烦了
    row column 的索引都是从0开始的，我这里要把图片插入到单元格A17
    """
    p2e = pixels_to_EMU
    h, w = img.height, img.width
    # 图像等比例缩放因子
    resize_factor = 0.8
    w_h_ratio = w/h
    resize_H = int(resize_factor * h)
    resize_W = int(resize_factor * resize_H * w_h_ratio)
    #
    # x_pad = 4
    # y_pad = 25
    # 注意这里的行、列索引从0开始, 所以需要减1
    size = XDRPositiveSize2D(p2e(w), p2e(h))
    marker = AnchorMarker(col=col-1, colOff=pixels_to_EMU(x_pad), row=row-1, rowOff=pixels_to_EMU(y_pad))
    img.anchor = OneCellAnchor(_from=marker, ext=size)

    # 图像在Excel里面的大小
    # image_size_excel = XDRPositiveSize2D(pixels_to_EMU(resize_W), pixels_to_EMU(resize_H))
    ##############################
    # 设置单元格大小，单元格默认宽度单位：字符；高度单位：point(磅)
    # cell_height = int(pixels_to_points(resize_H+10, dpi=96)) #高度上增加10个像素放大单元格
    # cell_width = int(resize_W/8) + 2 # 宽度上增加16（2*8）个像素放大单元格

    # marker = AnchorMarker(col=col, colOff=pixels_to_EMU(x_pad), row=row, rowOff=pixels_to_EMU(y_pad))
    # img.anchor = OneCellAnchor(_from=marker, ext=image_size_excel)



# 专属产品生成客户报价
async def items_quotation_export(res,user,s):
    try:
        path = config.get_today_upload_path()
        if not os.path.exists(path):
            make_dirs(path)
        col = {}
        module = res.get('module')
        m = get_module(module)
        table = m.table_name
        columns = res.get('columns')
        # index = 1
        # for c in columns:
        #     if not c.get('params'):
        #         continue
        #     if not c.get('visible'):
        #         continue
        #     field = c.get('params').get('field')
        #     if not field:
        #         continue
        #     if field.table!=table:
        #         continue
        #     if field.get('data_type')==5:
        #         continue
        #     index = index + 1
        #     col[c.get('title')] = {'name':field.get('field'),'index':index}

        i = 1
        j = 1
        wb = Workbook() # 新建Excel文件
        ws = wb.active # wb.create_sheet() # 默认激活第一个sheet
        ws.title = "客户报价"
        ws.cell(i,j).value='产品图片'
        for k,v in columns.items():
            j += 1
            ws.cell(i,j).value=k

        ws.column_dimensions['A'].width = 23
        i += 1
        rids = res.get('rids')

        o = get_model_by_table_name(table)
        for rid in rids:
            c = s.query(o).filter(o.rid==rid).first()
            if not c:
                continue
            ws.row_dimensions[i].height = 76
            c = alchemy_object_to_dict(c)
            if c.get('yytp')!=None and c['yytp']!='' and c.get('yytp')!='[]':
                photo = json.loads(c.get('yytp'))
                if photo != None:
                    file_path = photo[0]['src']
                    fn = os.path.join(config.data_upload_path, str(file_path))
                    if (os.path.exists(fn)):
                        img = Image_Get(fn) #选择图片
                        # img.width = 150  # 设置图像宽度
                        # img.height = 68  # 设置图像高度
                        col_width = ws.column_dimensions['A'].width*7
                        row_height = (ws.row_dimensions[i].height)*1.2
                        img.width=col_width-4 # 转换为像素
                        img.height=row_height-4 # 转换为像素
                        x_offset = 8 # (col_width-img.width)/2
                        y_offset = 6 # (row_height-img.height)/2
                        col = 1
                        offset_img(img, col, i, x_offset, y_offset) #col为列位置，i为行位置,x_offset为左边边距,y_offset为上边距
                        ws.add_image(img)  #添加图片

            for k,v in columns.items():
                ws.cell(i,v.get('index')).value=c.get(v.get('name'))
            i += 1

        report_rid = get_uuid()
        fn = os.path.join(path, str(report_rid)+'.xlsx')
        wb.save(fn)
        sbs_path = path[-10:]

        return {'code':1,'msg':'操作完成', 'data':{'path':sbs_path + '/'+ str(report_rid)+'.xlsx','name': report_rid}}
    except Exception as e:
        logger.error(trace_error())
        return {'code':-1,'msg':trace_error()}
    

# 专属产品生成客户报价
@any_route('/api/saier/items/quotation/export', methods=['POST', 'GET'])
@require_token
async def view_saier_items_quotation_export(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        res = await items_quotation_export(j,user,s)

        return json_result(res.get('code'), res.get('msg'),res.get('data'))
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

# 专属产品业务人员change
@any_route('/api/saier/items/ywry/change', methods=['POST', 'GET'])
@require_token
async def view_saier_items_ywry_change(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        ywry = j.get('ywry', '')
        cpbh = j.get('cpbh', '')
        rid = j.get('rid', '')
        bjyw = ''
        d = s.query(zscp.bjyw).filter(zscp.rid==rid).first()
        if d:
            bjyw = str(d.bjyw)
            d = run_sql(f"select rid from sys_user where username='{bjyw}'")
            if len(d) == 0:
                bjyw = ''
        
        d = run_sql(f"select rid from sys_user where username='{ywry}'")
        if len(d) == 0:
            return json_result(-1, '无此人员,请重新输入', bjyw)
        
        return json_result(1, '校验成功')
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error(),'')
    finally:
        s.close()

# 专属产品特殊毛利率change
@any_route('/api/saier/items/mll/change', methods=['POST', 'GET'])
@require_token
async def view_saier_items_mll_change(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        cpbh = j.get('cpbh', '')
        rid = j.get('rid', '')
        mll = j.get('mll',0)
        d = s.query(cgjhsheet).filter(cgjhsheet.bjhh==cpbh,cgjhsheet.zdml>mll,cgjhsheet.mll>=mll,cgjhsheet.ggxm!='是').all()
        for r in d:
            r.zdml = mll
            r.ggxm = '是'
            s.add(r)
            res = await insert_script_log(module, cpbh, '低毛利变更', '成功', user, str({"产品编号:":cpbh,"毛 利 率":mll}),s=s)
            s.commit()

        return json_result(1, '更新成功')
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error(),'')
    finally:
        s.close()


# 专属产品特殊毛利率change
@any_route('/api/saier/items/supplier/change', methods=['POST', 'GET'])
@require_token
async def view_saier_items_supplier_change(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        cpbh = j.get('cpbh', '')
        rid = j.get('rid', '')
        cgry = j.get('cgry','')
        sccj = j.get('sccj')
        csid = j.get('gcID')
        cgry_old = ''
        path = ''
        flag = False
        d = s.query(zscpsheet5.cgry).filter(zscpsheet5.rid==rid).first()
        if d:
            cgry_old = d.cgry

        org = get_user_path(cgry)
        uid = org.get('rid')
        path = org.get('path')
        if uid==None or uid=='':
            return json_result(-1, '无此人员更改无效', cgry_old)
        # d = run_sql(f"select path,rid from sys_user where username='{cgry}'")
        # if len(d) == 0:
        #     return json_result(-1, '无此人员更改无效', cgry_old)
        # path = d[0]['path'][:100]
        # uid = d[0]['rid']
        d = s.query(cyzglsheet.rid).filter(cyzglsheet.xm==user.username,cyzglsheet.zm=="专业专属更改采购人员").first()
        if not d:
            if cpbh!='' and cpbh!=None:
                c = s.query(cgjhsheet.cgry).filter(cgjhsheet.bjhh==cpbh, func.ifnull(cgjhsheet.bjhh,'')!='', cgjhsheet.sccj1==sccj).first()
                if c:
                    return json_result(-1, '这个产品已有采购计划，您无权更改采购人员,请通知相关人员更改', cgry_old)
                c = s.query(cghtsheet.cgry).filter(cghtsheet.bjhh==cpbh, func.ifnull(cghtsheet.bjhh,'')!='', cghtsheet.sccj1==sccj).first()
                if c:
                    u = s.query(ywrybiao).filter(ywrybiao.yhm==cgry, ywrybiao.rybh.not_like("%LZ%")).first()
                    if u:
                        return json_result(-1, '这个产品已有采购合同，您无权更改采购人员,请通知相关人员更改', cgry_old)
            
        d = s.query(zscpsheet5).filter(zscpsheet5.cpbh==cpbh,zscp.sccj==sccj).all()
        for r in d:
            r.cgry = cgry
            s.add(r)
            flag = True
        d = s.query(zycssheet2).filter(zycssheet2.cpbh==cpbh,zycssheet2.cs_id==csid).first()
        if d:
            d.syq = cgry
            flag = True
            s.add(d)
        d = s.query(zscp).filter(zscp.cpbh==cpbh,zscp.sccj==sccj).all()
        for r in d:
            r.cgry = cgry
            r.uid = uid
            flag = True
            s.add(r)
        if flag:
            res = await insert_script_log('专属产品', cpbh, '对应工厂的采购人员修改', '成功', user, str({"产品编号:":cpbh,"生产厂家":sccj,"新采购人员":cgry,"原采购人员":cgry_old}),s=s)
            s.commit()

        return json_result(1, '更新成功')
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error(),'')
    finally:
        s.close()


# 专属产品特殊毛利率change
@any_route('/api/saier/items/supplier/update/jsfs', methods=['POST', 'GET'])
@require_token
async def view_saier_items_supplier_update_jsfs(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        r = j.get('data',{})
        # rid = j.get('rid')
        # logger.error(user.__dict__)
        # path = organizations.get_org_name_path(rid)
        d = s.query(zscpsheet5).filter(zscpsheet5.rid==r.get('rid')).first()
        if d:
            for k,v in r.items():
                if k in SYS_FIELDS:
                    continue
                if hasattr(d,k):
                    setattr(d,k,v)
            d.modi_uid = user.rid
            d.mtime = time.strftime('%Y-%m-%d %H:%M:%S')
            s.add(d)
            res = await insert_script_log('专属产品', r.get('cpbh'), '对应工厂的字段值修改', '成功', user, str({"产品编号:":r.get('cpbh'),"data":r}),s=s)
            s.commit()

        return json_result(1, '更新成功')
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error(),'')
    finally:
        s.close()