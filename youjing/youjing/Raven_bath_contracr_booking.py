# BOOKING导出（固定列版，业务口径按原 Pascal）
from any import *
import os
import re
import json
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


ILLEGAL_CHARS = set('/\\:*?"<>|\r\n')


def _safe_filename(text):
    s = str(text or '').strip()
    if not s:
        return 'EMPTY'
    s = ''.join('_' if ch in ILLEGAL_CHARS else ch for ch in s)
    return s[:120]


def _upper(v):
    return str(v or '').upper()


def _to_int(v, default_value=-1):
    try:
        return int(float(v))
    except Exception:
        return default_value


def _to_float(v):
    try:
        return float(v or 0)
    except Exception:
        return 0.0


def _parse_json_list(v):
    if isinstance(v, list):
        return v
    if not v:
        return []
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return []
        try:
            x = json.loads(s)
            if isinstance(x, list):
                return x
        except Exception:
            pass
        return [i.strip() for i in s.split(',') if i.strip()]
    return []


def _to_upload_rel_path(abs_path):
    abs_path = os.path.abspath(abs_path)
    base_upload = os.path.abspath(getattr(config, 'data_upload_path', '') or '')
    if base_upload and abs_path.startswith(base_upload):
        rel = abs_path[len(base_upload):].lstrip('/\\')
        return rel.replace('\\', '/')
    return abs_path.replace('\\', '/')


def _build_unique_path(folder, base_name, ext):
    name = f'{base_name}{ext}'
    path = os.path.join(folder, name)
    idx = 1
    while os.path.exists(path):
        name = f'{base_name}_{idx}{ext}'
        path = os.path.join(folder, name)
        idx += 1
    return path


def _first_row(sql, params):
    rows = run_sql(sql, params)
    return rows[0] if rows else {}


def _try_rows(sql, params):
    try:
        return run_sql(sql, params)
    except Exception:
        return []


def _query_chyjh_by_key(key):
    key_s = str(key or '').strip()
    if not key_s:
        return {}

    # 优先按 rid
    row = _first_row(
        "SELECT * FROM chyjh WHERE rid=:v LIMIT 1",
        {'v': key_s}
    )
    if row:
        return row

    # 回退按 number（兼容仍保留 number 的库）
    row = _first_row(
        "SELECT * FROM chyjh WHERE number=:v LIMIT 1",
        {'v': key_s}
    )
    return row or {}


def _collect_target_headers(raw_ids, mode):
    ids = raw_ids if mode == '2' else (raw_ids[:1] if raw_ids else [])
    out = []
    seen = set()

    for x in ids:
        h = _query_chyjh_by_key(x)
        if not h:
            continue

        uniq = str(h.get('rid') or h.get('number') or '').strip()
        if not uniq:
            continue

        if uniq in seen:
            continue
        seen.add(uniq)
        out.append(h)

    return out


def _get_details_for_header(header):
    rid_val = str(header.get('rid') or '').strip()
    number_val = str(header.get('number') or '').strip()

    sql_tpl = """
        SELECT cpbh, kpgc, zhwbgpm, djpmy, htzxs, nxrl, mjzj, wxzj
        FROM chyjhsheet
        WHERE {field}=:v
        ORDER BY rid
    """

    tries = []
    # 按你之前口径，优先 pid/rid
    if rid_val:
        tries.extend([
            ('pid', rid_val),
            ('father', rid_val),   # 兼容旧库
        ])
    if number_val:
        tries.extend([
            ('pid', number_val),
            ('father', number_val),
        ])

    for field, val in tries:
        rows = _try_rows(sql_tpl.format(field=field), {'v': val})
        if rows:
            return rows

    return []


def _collect_booking_texts(details):
    seen = set()
    kpgc_list, zh_list, en_list = [], [], []

    for r in details:
        kpgc = str(r.get('kpgc') or '').strip()
        zh = str(r.get('zhwbgpm') or '').strip()
        en = str(r.get('djpmy') or '').strip()

        if not (kpgc or zh or en):
            continue

        key = (kpgc, zh, en)
        if key in seen:
            continue
        seen.add(key)

        if kpgc:
            kpgc_list.append(kpgc)
        if zh:
            zh_list.append(zh)
        if en:
            en_list.append(en)

    return '\n'.join(kpgc_list), '\n'.join(zh_list), '\n'.join(en_list)


def _top_product_names(header, pms):
    pms = max(1, min(_to_int(pms, 1), 20))
    khpd = str(header.get('khpd') or '')
    sum_col = 'mjzj' if khpd == '是' else 'wxzj'

    rid_val = str(header.get('rid') or '').strip()
    number_val = str(header.get('number') or '').strip()

    sql_tpl = f"""
        SELECT djpmy, SUM({sum_col}) AS amt
        FROM chyjhsheet
        WHERE {{field}}=:v
        GROUP BY djpmy
        ORDER BY amt DESC
        LIMIT {pms}
    """

    tries = []
    if rid_val:
        tries.extend([
            ('pid', rid_val),
            ('father', rid_val),
        ])
    if number_val:
        tries.extend([
            ('pid', number_val),
            ('father', number_val),
        ])

    rows = []
    for field, val in tries:
        rows = _try_rows(sql_tpl.format(field=field), {'v': val})
        if rows:
            break

    names = [str(r.get('djpmy') or '').strip() for r in rows if str(r.get('djpmy') or '').strip()]
    return ';\n'.join(names)


def _checkbox_flags(header, kh_row):
    sxfw = str(header.get('sxfw') or '')
    fkfs = str(header.get('fkfs') or '')
    jgtk = str(header.get('jgtk') or '').upper()
    hxdate = str(header.get('hxdate') or '')
    ysfs = str(header.get('ysfs') or '').upper()
    ins = str(header.get('Insurancesfzbx') or '')
    xybx = str(header.get('xybx') or '')
    xbtt = str((kh_row or {}).get('xbtt') or '').strip()

    yes_259 = bool(xbtt) or (xybx == '是')

    return {
        'FCL': sxfw == 'FCL柜货',
        'LCL': sxfw == 'LCL散货',
        '到付': fkfs == '到付',
        '预付': fkfs == '预付',
        'FOB': jgtk == 'FOB',
        'CNF': jgtk == 'CNF',
        'CIF': jgtk == 'CIF',
        '回箱有': hxdate == '是',
        '回箱无': hxdate in ('否', ''),
        '空运': ysfs == 'BY AIR',
        '海运': ysfs == 'BY SEA',
        '铁路': ysfs == 'BY TRAIN',
        '保险有': ins == '是',
        '保险无': ins in ('否', ''),
        '协议有': yes_259,
        '协议无': not yes_259,
    }


def _tick(v):
    return '√' if v else ''


@any_route('/api/Ravencloud/export_booking_latest', methods=['POST'])
async def api_export_booking_latest(request):
    try:
        form = await request.form()

        rid_list = _parse_json_list(form.get('rid_list') or '[]')
        mode = '2' if str(form.get('mode') or '1').strip() == '2' else '1'
        pms = _to_int(form.get('pms') or '1', 1)
        if pms <= 0:
            pms = 1
        if pms > 20:
            pms = 20

        if not rid_list:
            return json_result(code=-1, msg='请先选择记录')

        headers_data = _collect_target_headers(rid_list, mode)
        if not headers_data:
            return json_result(code=-1, msg='未找到可导出数据')

        wb = Workbook()
        ws = wb.active
        ws.title = 'BOOKING主表'

        main_headers = [
             '客人全称', '客人ID', '提货人', '业务员', '业务电话', '业务传真',
            '收货人', '通知人', '外销发票', '出运口岸', '目的口岸',
            '柜型', '付款方式', '价格条款', '运输方式', '回箱日期', '保险',
            '唛头文本', '总箱数', '总毛重', '总体积', '客判', '美金总价', '外销总额',
            '开票工厂聚合', '中文报关品名聚合', '英文品名聚合', 'TOP英文品名',
            'FCL', 'LCL', '到付', '预付', 'FOB', 'CNF', 'CIF',
            '回箱有', '回箱无', '空运', '海运', '铁路', '保险有', '保险无', '协议有', '协议无'
        ]
        ws.append(main_headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws_detail = wb.create_sheet('BOOKING明细')
        detail_headers = [
            '记录键(rid/number)', '产品编号', '开票工厂', '中文报关品名', '英文品名',
            '装箱率', '合同总箱数', '内箱容量', '明细美金金额', '明细外销金额'
        ]
        ws_detail.append(detail_headers)
        for cell in ws_detail[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        warnings = []

        for header in headers_data:
            key_show = str(header.get('rid') or header.get('number') or '').strip()
            details = _get_details_for_header(header)
            kpgc_txt, zh_txt, en_txt = _collect_booking_texts(details)
            top_en = _top_product_names(header, pms)

            user_row = _first_row(
                "SELECT lxdh, czhm FROM ywrybiao WHERE yhm=:yhm LIMIT 1",
                {'yhm': str(header.get('ywry') or '')}
            )
            kh_row = _first_row(
                "SELECT xbtt, xybx FROM kh WHERE kh_id=:kh_id LIMIT 1",
                {'kh_id': str(header.get('kh_id') or '')}
            )
            flags = _checkbox_flags(header, kh_row)

            ws.append([
                # key_show,
                _upper(header.get('khmc')),
                _upper(header.get('kh_id')),
                _upper(header.get('ttr')),
                str(header.get('ywry') or ''),
                _upper(user_row.get('lxdh') if user_row else ''),
                str(user_row.get('czhm') if user_row else '') or '0574-27903600',
                _upper(header.get('shr')),
                _upper(header.get('tzr')),
                _upper(header.get('wxfp')),
                _upper(header.get('cyka')),
                _upper(header.get('mdka')),
                str(header.get('sxfw') or ''),
                str(header.get('fkfs') or ''),
                str(header.get('jgtk') or ''),
                str(header.get('ysfs') or ''),
                str(header.get('hxdate') or ''),
                str(header.get('Insurancesfzbx') or ''),
                _upper(header.get('wxmt')),
                str(header.get('xshj2') or ''),
                _to_float(header.get('mzhj')),
                _to_float(header.get('tjhj')),
                str(header.get('khpd') or ''),
                _to_float(header.get('mjzj')),
                _to_float(header.get('wxje')),
                kpgc_txt,
                zh_txt,
                en_txt,
                top_en,
                _tick(flags['FCL']),
                _tick(flags['LCL']),
                _tick(flags['到付']),
                _tick(flags['预付']),
                _tick(flags['FOB']),
                _tick(flags['CNF']),
                _tick(flags['CIF']),
                _tick(flags['回箱有']),
                _tick(flags['回箱无']),
                _tick(flags['空运']),
                _tick(flags['海运']),
                _tick(flags['铁路']),
                _tick(flags['保险有']),
                _tick(flags['保险无']),
                _tick(flags['协议有']),
                _tick(flags['协议无']),
            ])

            # 多行文本列自动换行
            row_idx = ws.max_row
            for col in ('Z', 'AA', 'AB', 'AC'):
                ws[f'{col}{row_idx}'].alignment = Alignment(wrapText=True, vertical='top')

            if not details:
                warnings.append(f'{key_show} 无明细数据')
                continue

            for d in details:
                ws_detail.append([
                    key_show,
                    str(d.get('cpbh') or ''),
                    str(d.get('kpgc') or ''),
                    str(d.get('zhwbgpm') or ''),
                    str(d.get('djpmy') or ''),
                    # str(d.get('zxl') or ''),
                    str(d.get('htzxs') or ''),
                    str(d.get('nxrl') or ''),
                    _to_float(d.get('mjzj')),
                    _to_float(d.get('wxzj')),
                ])

        if ws.max_row <= 1:
            return json_result(code=-1, msg='未生成任何有效数据')

        # 主表列宽（固定列）
        width_main = {
            'A': 16, 'B': 24, 'C': 16, 'D': 16, 'E': 14, 'F': 16, 'G': 16, 'H': 16, 'I': 16, 'J': 18,
            'K': 14, 'L': 14, 'M': 12, 'N': 12, 'O': 10, 'P': 12, 'Q': 10, 'R': 10, 'S': 20, 'T': 10,
            'U': 10, 'V': 10, 'W': 10, 'X': 12, 'Y': 12, 'Z': 30, 'AA': 30, 'AB': 30, 'AC': 30,
            'AD': 6, 'AE': 6, 'AF': 6, 'AG': 6, 'AH': 6, 'AI': 6, 'AJ': 6, 'AK': 6, 'AL': 6,
            'AM': 6, 'AN': 6, 'AO': 6, 'AP': 6, 'AQ': 6, 'AR': 6
        }
        for col, w in width_main.items():
            ws.column_dimensions[col].width = w

        # 明细列宽
        width_detail = {'A': 16, 'B': 16, 'C': 24, 'D': 26, 'E': 26, 'F': 12, 'G': 12, 'H': 12, 'I': 14, 'J': 14}
        for col, w in width_detail.items():
            ws_detail.column_dimensions[col].width = w

        output_root = os.path.join(config.get_today_upload_path(), 'BOOKING导出')
        os.makedirs(output_root, exist_ok=True)

        out_base = _safe_filename(f"BOOKING固定列_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
        out_abs = _build_unique_path(output_root, out_base, '.xlsx')
        wb.save(out_abs)
        wb.close()

        data = {
            'files': [_to_upload_rel_path(out_abs)],
            'count': 1
        }
        if warnings:
            data['warnings'] = warnings

        return json_result(code=1, msg='BOOKING导出成功', data=data)

    except Exception:
        logger.error(trace_error())
        return json_result(code=-1, msg=trace_error())