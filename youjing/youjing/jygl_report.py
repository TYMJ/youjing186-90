# import win32com.client as win32
# import os
import subprocess
# import pyodbc
from PIL import Image
from ast import mod
from webbrowser import get

from any import *
from .model import *
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Optional, Tuple
import os
from .__default__ import module_share_new, module_xxck_new, get_user_path, user_task_new
from sqlalchemy.sql import func, not_, or_, and_
from collections import OrderedDict,defaultdict
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from datetime import datetime
import time
# from openpyxl.copy import copy_worksheet
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from PIL import Image as PILImage   
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.drawing.image import Image as Image_Get




def convert_excel_to_pdf(excel_path, output_dir=None):
    """
    将Excel文件转换为PDF（使用 whale_report.jar）
    
    Args:
        excel_path: Excel文件的完整路径
        output_dir: PDF输出目录，默认与Excel同目录
        
    Returns:
        dict: {'success': bool, 'pdf_path': str, 'error': str}
    """
    try:
        # 确定输出目录
        if output_dir is None:
            output_dir = os.path.dirname(excel_path)
        
        # 生成PDF文件路径
        base_name = os.path.splitext(excel_path)[0]
        pdf_path = base_name + '.pdf'
        
        print(f'========== 开始PDF转换 ==========')
        print(f'源文件: {excel_path}')
        print(f'目标文件: {pdf_path}')
        
        # 检查 Java 和 JAR 文件是否存在
        if not hasattr(config, 'java_path') or not config.java_path:
            return {
                'success': False,
                'error': 'Java路径未配置（config.java_path）'
            }
        
        if not hasattr(config, 'report_jar') or not config.report_jar:
            return {
                'success': False,
                'error': 'whale_report.jar路径未配置（config.report_jar）'
            }
        
        if not os.path.exists(config.report_jar):
            return {
                'success': False,
                'error': f'whale_report.jar文件不存在: {config.report_jar}'
            }
        
        print(f'Java路径: {config.java_path}')
        print(f'JAR路径: {config.report_jar}')
        
        # 构建命令
        # console_run(config.java_path, ['-jar', config.report_jar, 'a', 'b', template, output_file, '2'])
        cmd = [
            config.java_path,
            '-jar',
            config.report_jar,
            'a',  # 占位参数
            'b',  # 占位参数
            excel_path,      # 源Excel文件
            pdf_path,        # 目标PDF文件
            '2'              # 转换类型参数
        ]
        
        print(f'执行命令: {" ".join(cmd)}')
        
        # 执行转换
        result = subprocess.run(
            cmd,
            capture_output=True,
            timeout=120,  # 2分钟超时
            text=True
        )
        
        if result.returncode != 0:
            print(f'转换失败，返回码: {result.returncode}')
            print(f'错误输出: {result.stderr}')
            return {
                'success': False,
                'error': f'PDF转换失败（返回码 {result.returncode}）：{result.stderr}'
            }
        
        # 检查PDF文件是否生成
        if not os.path.exists(pdf_path):
            return {
                'success': False,
                'error': f'PDF文件未生成（期望路径：{pdf_path}）'
            }
        
        file_size = os.path.getsize(pdf_path)
        print(f'✓ PDF转换成功')
        print(f'✓ 文件大小: {file_size:,} 字节')
        print(f'========== PDF转换完成 ==========')
        
        return {
            'success': True,
            'pdf_path': pdf_path,
            'error': None
        }
        
    except subprocess.TimeoutExpired:
        return {
            'success': False,
            'error': 'PDF转换超时（超过2分钟）'
        }
    except Exception as e:
        logger.error(trace_error())
        return {
            'success': False,
            'error': f'PDF转换异常：{str(e)}'
        }

def copy_worksheet_full(source_ws, target_ws):
    # 1. 复制所有单元格的值和样式（跳过合并单元格，避免报错）
    for row in source_ws.iter_rows():
        for cell in row:
            # 跳过合并单元格对象，只处理普通单元格
            if cell.__class__.__name__ == "MergedCell":
                continue

            # 写入值
            new_cell = target_ws.cell(
                row=cell.row,
                column=cell.column,  # 用 column 替代 col_idx，兼容所有版本
                value=cell.value
            )

            # 复制所有格式
            if cell.has_style:
                new_cell.font = cell.font.copy()
                new_cell.border = cell.border.copy()
                new_cell.fill = cell.fill.copy()
                new_cell.number_format = cell.number_format
                new_cell.alignment = cell.alignment.copy()
                new_cell.protection = cell.protection.copy()

    # 2. 复制列宽
    for col in source_ws.column_dimensions:
        target_ws.column_dimensions[col].width = source_ws.column_dimensions[col].width

    # 3. 复制行高
    for row in source_ws.row_dimensions:
        target_ws.row_dimensions[row].height = source_ws.row_dimensions[row].height

    # 4. 复制合并单元格
    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    # 5. 复制冻结窗格
    target_ws.freeze_panes = source_ws.freeze_panes

    return target_ws


@any_route("/api/khjy/generate_sample_tag_report", methods=["POST"])
@require_token
async def generate_sample_tag_report(request):
    jyrq = ''
    khmc = ''
    zts = 0
    l = 0
    ln = 0
    lg = 0
    bz = 0.0
    bz1 = 0
    
    j = await request.json()
    rid_list = j.get("rids")
    if not rid_list:
        return json_result(-1, "未选择导出记录", None)
    template_path = os.path.join(config.data_upload_path, "template/寄样吊牌稿.xlsx")
    if not os.path.exists(template_path):
        return json_result(-1, "[寄样吊牌稿]模板文件不存在", None)
    try:
        s = Session()
        # 查询主记录信息
        d = s.query(jygl.jyrq,jygl.khmc).filter(jygl.rid==rid_list).first()
        if not d:
            return json_result(-1, f"请检查数据是否正确!")
        jyrq = d.jyrq
        khmc = d.khmc
        # 查询产品数量
        zts = s.query(func.count(jyglcp.rid)).filter(jyglcp.pid == rid_list).scalar() or 0
        # 计算工作表数量
        ys1 = zts // 4
        iz = zts // 4
        ys2 = zts / 4
        if ys1 == ys2:
            ys3 = ys1
            iz = iz - 1
        else:
            ys3 = ys1 + 1
        wb = load_workbook(template_path)
        ws = wb.worksheets[0]
        ws.title = "US寄样打印"
        # 复制工作表
        a1z = 1
        for az in range(iz):
            source_sheet = wb.worksheets[0]
            new_sheet = wb.create_sheet("US寄样打印")
            copy_worksheet_full(source_sheet, new_sheet)
        rows = s.query(jyglcp).filter(jyglcp.pid == rid_list).all()
        for row in rows:
            l += 1
            ln += 1
            if ln == 1:
                lg += 1
                bz = 0.0
                msexcelworksheet = wb.worksheets[lg-1]             
                msexcelworksheet['B1'] = 'US'
                msexcelworksheet['D1'] = jyrq
                msexcelworksheet['B2'] = row.hwbh
                msexcelworksheet['B3'] = row.ywpm
                msexcelworksheet['B5'] = row.cpgge
                if row.mjdj1 > 0:
                    msexcelworksheet['A6'] = 'price RMB'
                    msexcelworksheet['B6'] = row.mjdj1
                else:
                    msexcelworksheet['A6'] = 'price USD'
                    msexcelworksheet['B6'] = row.jg
                msexcelworksheet['D6'] = row.bzhfsh
                msexcelworksheet['B7'] = row.czyy
                msexcelworksheet['B8'] = row.bz
                hhbz = row.hwbh
                if hhbz:
                    insert_image_to_excel(hhbz, row.rid, wb, msexcelworksheet, 2,4)
            elif ln == 2:
                # 右侧上半部分
                msexcelworksheet['H1'] = 'US'
                msexcelworksheet['J1'] = jyrq
                msexcelworksheet['H2'] = row.hwbh
                msexcelworksheet['H3'] = row.ywpm
                msexcelworksheet['H5'] = row.cpgge
                if row.mjdj1 > 0:
                    msexcelworksheet['G6'] = 'price RMB'
                    msexcelworksheet['H6'] = row.mjdj1
                else:
                    msexcelworksheet['G6'] = 'price USD'
                    msexcelworksheet['H6'] = row.jg
                msexcelworksheet['J6'] = row.bzhfsh
                msexcelworksheet['H7'] = row.czyy
                msexcelworksheet['H8'] = row.bz
                hhbz = row.hwbh
                hhbz = row.hwbh
                if hhbz:
                    insert_image_to_excel(hhbz, row.rid, wb, msexcelworksheet, 8,4)
            
            elif ln == 3:
                msexcelworksheet['B14'] = 'US'
                msexcelworksheet['D14'] = jyrq
                msexcelworksheet['B15'] = row.hwbh
                msexcelworksheet['B16'] = row.ywpm
                msexcelworksheet['B18'] = row.cpgge
                if row.mjdj1 > 0:
                    msexcelworksheet['A19'] = 'price RMB'
                    msexcelworksheet['B19'] = row.mjdj1
                else:
                    msexcelworksheet['A19'] = 'price USD'
                    msexcelworksheet['B19'] = row.jg
                msexcelworksheet['D19'] = row.bzhfsh
                msexcelworksheet['B20'] = row.czyy
                msexcelworksheet['B21'] = row.bz
                hhbz = row.hwbh
                hhbz = row.hwbh
                if hhbz:
                    insert_image_to_excel(hhbz, row.rid, wb, msexcelworksheet, 2,17)
            elif ln == 4:
                # 右侧下半部分
                msexcelworksheet['H14'] = 'US'
                msexcelworksheet['J14'] = jyrq
                msexcelworksheet['H15'] = row.hwbh
                msexcelworksheet['H16'] = row.ywpm
                msexcelworksheet['H18'] = row.cpgge
                if row.mjdj1 > 0:
                    msexcelworksheet['G19'] = 'price RMB'
                    msexcelworksheet['H19'] = row.mjdj1
                else:
                    msexcelworksheet['G19'] = 'price USD'
                    msexcelworksheet['H19'] = row.jg
                msexcelworksheet['J19'] = row.bzhfsh
                msexcelworksheet['H20'] = row.czyy
                msexcelworksheet['H21'] = row.bz
                hhbz = row.hwbh
                hhbz = row.hwbh
                if hhbz:
                    insert_image_to_excel(hhbz, row.rid, wb, msexcelworksheet, 8,17)
                ln = 0  
        path = config.tmp_path
        logger.error(f"临时路径: {path}")
        os.makedirs(path, exist_ok=True)
        report_rid = get_uuid()
        output_filename = f'{report_rid}_寄样吊牌稿.xlsx'
        full_output_path = os.path.join(path, output_filename)
        logger.error(f"导出路径: {full_output_path}")
        wb.save(full_output_path)
        # return {'code': 1, 'msg': '报表导出成功', 'data': output_filename}
        return JSONResponse({
        'code': 1,
        'msg': '报表导出成功',
        'data': {
            'filename': output_filename,
            'full_path': full_output_path,
            'total_products': zts
        }
    })
    except Exception as e:
        print(f"发生错误: {e}")
        try:
            wb.Close()
            wb.Quit()
        except:
            pass

def offset_img(img, col, row, x_pad=4, y_pad=25):
    p2e = pixels_to_EMU
    h, w = img.height, img.width
    # 图像等比例缩放因子
    resize_factor = 0.8
    w_h_ratio = w/h
    resize_H = int(resize_factor * h)
    resize_W = int(resize_factor * resize_H * w_h_ratio)
    #
    # x_pad = 4
    # y_pad = 25
    # 注意这里的行、列索引从0开始, 所以需要减1
    size = XDRPositiveSize2D(p2e(w), p2e(h))
    marker = AnchorMarker(col=col-1, colOff=pixels_to_EMU(x_pad), row=row-1, rowOff=pixels_to_EMU(y_pad))
    img.anchor = OneCellAnchor(_from=marker, ext=size)

def insert_image_to_excel(hhbz, wyzd, wb, msexcelworksheet, c,r):
    """
    插入图片到Excel指定位置
    """
    Swyzd=wyzd
    g = run_sql(f"select yytp  from jyglcp where (rid='{Swyzd}')")
    if len(g)>0 and g[0].get('yytp','') != '' and g[0].get('yytp','') != None and g[0].get('yytp','') != '[]':
        cpbh = g[0].get('yytp','')
        Photo = json.loads(str(cpbh))
        if Photo != None:
            file_path = Photo[0]['src']
            fn = os.path.join(config.data_upload_path, str(file_path))
            if (os.path.exists(fn)):
                img = Image_Get(fn) #选择图片
                # img.width = 150  # 设置图像宽度
                # img.height = 68  # 设置图像高度
                col_width = (msexcelworksheet.column_dimensions['C'].width + msexcelworksheet.column_dimensions['E'].width)*7
                row_height = 75   # (msexcelworksheet.row_dimensions[cell].height)*1.5
                img.width=col_width-4 # 转换为像素
                img.height=row_height-4 # 转换为像素
                x_offset = 15 # (col_width-img.width)/2
                y_offset = 5 # (row_height-img.height)/2
                col = c  
                row = r
                offset_img(img, col, row, x_offset, y_offset) #col为列位置，row为行位置,x_offset为左边边距,y_offset为上边距
                msexcelworksheet.add_image(img)  #添加图片




@any_route("/api/khjy/generate_sample_tag_report_UV", methods=["POST"])
@require_token
async def generate_sample_tag_report_UV(request):
    j = await request.json()
    rid_list = j.get("rids")
    if not rid_list:
        return json_result(-1, "未选择导出记录", None)
    template_path = os.path.join(config.data_upload_path, "template/寄样吊牌稿.xlsx")
    if not os.path.exists(template_path):
        return json_result(-1, "[寄样吊牌稿]模板文件不存在", None)
    
    s = Session()
    # 查询主记录信息
    d = s.query(jygl.jyrq,jygl.khmc).filter(jygl.rid==rid_list).first()
    if not d:
        return json_result(-1, f"请检查数据是否正确!")
    jyrq = d.jyrq
    khmc = d.khmc
    # 读取产品总数
    zts = s.query(func.count(jyglcp.rid)).filter(jyglcp.pid == rid_list).scalar() or 0
    if zts == 0:
        return json_result(-1, f"无产品明细!")
    # 计算页数：每页4个
    ys3 = (zts + 3) // 4
    iz = ys3 - 1
    wb = load_workbook(template_path)
    ws = wb.worksheets[0]
    ws.title = "UV寄样打印"
    
    # 复制工作表
    a1z = 1
    for az in range(iz):
        source_sheet = wb.worksheets[0]
        new_sheet = wb.create_sheet("UV寄样打印")
        copy_worksheet_full(source_sheet, new_sheet)
        # wb.add_sheet(new_sheet)
    rows = s.query(jyglcp).filter(jyglcp.pid == rid_list).all()
    current_sheet_index = 0
    ln = 0  # 1~4 计数
    for prod in rows:
        ln += 1
        if ln == 1:
            current_sheet = wb.worksheets[current_sheet_index]
            current_sheet_index += 1
        hwbh = prod.hwbh
        ywpm = prod.ywpm
        cpgge = prod.cpgge
        mjdj1 = prod.mjdj1
        jg = prod.jg
        bzhfsh = prod.bzhfsh
        czyy = prod.czyy
        bz = prod.bz
        wyzd = prod.wyzd

        ws = current_sheet

        # ========== 填充 4 个位置 ==========
        if ln == 1:
            ws["B1"] = "US"
            ws["D1"] = jyrq
            ws["B2"] = hwbh
            ws["B3"] = ywpm
            ws["B5"] = cpgge
            if mjdj1 > 0:
                ws["A6"] = "price RMB"
                ws["B6"] = mjdj1
            else:
                ws["A6"] = "price USD"
                ws["B6"] = jg
            ws["D6"] = bzhfsh
            ws["B7"] = czyy
            ws["B8"] = bz
            img_cell = "B4"
            if hwbh:
                insert_image_to_excel(hwbh, prod.rid, wb, ws, 2,4)

        elif ln == 2:
            ws["H1"] = "US"
            ws["J1"] = jyrq
            ws["H2"] = hwbh
            ws["H3"] = ywpm
            ws["H5"] = cpgge
            if mjdj1 > 0:
                ws["G6"] = "price RMB"
                ws["H6"] = mjdj1
            else:
                ws["G6"] = "price USD"
                ws["H6"] = jg
            ws["J6"] = bzhfsh
            ws["H7"] = czyy
            ws["H8"] = bz
            img_cell = "H4"
            if hwbh:
                insert_image_to_excel(hwbh, prod.rid, wb, ws, 8,4)

        elif ln == 3:
            ws["B14"] = "US"
            ws["D14"] = jyrq
            ws["B15"] = hwbh
            ws["B16"] = ywpm
            ws["B18"] = cpgge
            if mjdj1 > 0:
                ws["A19"] = "price RMB"
                ws["B19"] = mjdj1
            else:
                ws["A19"] = "price USD"
                ws["B19"] = jg
            ws["D19"] = bzhfsh
            ws["B20"] = czyy
            ws["B21"] = bz      
            img_cell = "B17"
            if hwbh:
                insert_image_to_excel(hwbh, prod.rid, wb, ws, 2,17)

        elif ln == 4:
            ws["H14"] = "US"
            ws["J14"] = jyrq
            ws["H15"] = hwbh    
            ws["H16"] = ywpm
            ws["H18"] = cpgge
            if mjdj1 > 0:
                ws["G19"] = "price RMB"
                ws["H19"] = mjdj1
            else:
                ws["G19"] = "price USD"
                ws["H19"] = jg
            ws["J19"] = bzhfsh
            ws["H20"] = czyy
            ws["H21"] = bz
            img_cell = "H17"
            ln = 0
            if hwbh:
                insert_image_to_excel(hwbh, prod.rid, wb, ws, 8,17)

    path = config.tmp_path
    logger.error(f"临时路径: {path}")
    os.makedirs(path, exist_ok=True)
    report_rid = get_uuid()
    output_filename = f'{report_rid}_寄样吊牌稿.xlsx'
    full_output_path = os.path.join(path, output_filename)
    logger.error(f"导出路径: {full_output_path}")
    wb.save(full_output_path)
    # return {'code': 1, 'msg': '报表导出成功', 'data': output_filename}
    return JSONResponse({
    'code': 1,
    'msg': '报表导出成功',
    'data': {
        'filename': output_filename,
        'full_path': full_output_path,
        'total_products': zts
        }
    })



@any_route("/api/khjy/generate_sample_tag_report_jydy8", methods=["POST"])
@require_token
async def generate_sample_tag_report_jydy8(request):
    j = await request.json()
    rid_list = j.get("rids")
    if not rid_list:
        return json_result(-1, "未选择导出记录", None)
    template_path = os.path.join(config.data_upload_path, "template/寄样吊牌稿.xlsx")
    if not os.path.exists(template_path):
        return json_result(-1, "[寄样吊牌稿]模板文件不存在", None)
    s = Session()
    # 查询主记录信息
    d = s.query(jygl.jyrq,jygl.khmc).filter(jygl.rid==rid_list).first()
    if not d:
        return json_result(-1, f"请检查数据是否正确!")
    jyrq = d.jyrq
    khmc = d.khmc
    # 查询产品数量
    zts = s.query(func.count(jyglcp.rid)).filter(jyglcp.pid == rid_list).scalar() or 0
    ys1 = zts // 4
    iz = ys1
    ys2 = zts / 4
    if ys1 == ys2:
        ys3 = ys1
        iz = iz - 1
    else:
        ys3 = ys1 + 1
    wb = load_workbook(template_path)
    ws = wb.worksheets[0]
    ws.title = "US寄样打印"
    # 复制工作表
    a1z = 1
    for az in range(iz):
        source_sheet = wb.worksheets[0]
        new_sheet = wb.create_sheet("US寄样打印")
        copy_worksheet_full(source_sheet, new_sheet)
    rows = s.query(jyglcp).filter(jyglcp.pid == rid_list).all()
    current_sheet_idx = 0
    ln = 0

    for prod in rows:
        ln += 1

        # 每到第1个，切换新 sheet
        if ln == 1:
            ws = wb.worksheets[current_sheet_idx]
            current_sheet_idx += 1

        hwbh = prod.hwbh
        ywpm = prod.ywpm
        cpgge = prod.cpgge
        mjdj1 = prod.mjdj1
        jg = prod.jg
        bzhfsh = prod.bzhfsh
        czyy = prod.czyy
        bz = prod.bz

        # ==============================================
        # 第 1 个位置：B1~B8
        # ==============================================
        if ln == 1:
            ws["B1"] = "US"
            ws["D1"] = jyrq
            ws["B2"] = hwbh
            ws["B3"] = ywpm
            ws["B5"] = cpgge
            if mjdj1 > 0:
                ws["A6"] = "price RMB"
                ws["B6"] = mjdj1
            else:
                ws["A6"] = "price USD"
                ws["B6"] = jg
            ws["D6"] = bzhfsh
            ws["B7"] = czyy
            ws["B8"] = bz
            if hwbh:
                insert_image_to_excel(hwbh, prod.rid, wb, ws, 2,4)

        # ==============================================
        # 第 2 个位置：H1~H8
        # ==============================================
        elif ln == 2:
            ws["H1"] = "US"
            ws["J1"] = jyrq
            ws["H2"] = hwbh
            ws["H3"] = ywpm
            ws["H5"] = cpgge
            if mjdj1 > 0:
                ws["G6"] = "price RMB"
                ws["H6"] = mjdj1
            else:
                ws["G6"] = "price USD"
                ws["H6"] = jg
            ws["J6"] = bzhfsh
            ws["H7"] = czyy
            ws["H8"] = bz
            if hwbh:
                insert_image_to_excel(hwbh, prod.rid, wb, ws, 8,4)

        # ==============================================
        # 第 3 个位置：B14~B21
        # ==============================================
        elif ln == 3:
            ws["B14"] = "US"
            ws["D14"] = jyrq
            ws["B15"] = hwbh
            ws["B16"] = ywpm
            ws["B18"] = cpgge
            if mjdj1 > 0:
                ws["A19"] = "price RMB"
                ws["B19"] = mjdj1
            else:
                ws["A19"] = "price USD"
                ws["B19"] = jg
            ws["D19"] = bzhfsh
            ws["B20"] = czyy
            ws["B21"] = bz
            if hwbh:
                insert_image_to_excel(hwbh, prod.rid, wb, ws, 2,17)

        # ==============================================
        # 第 4 个位置：H14~H21 
        # ==============================================
        elif ln == 4:
            ws["H14"] = "US"
            ws["J14"] = jyrq
            ws["H15"] = hwbh
            ws["H16"] = ywpm
            ws["H18"] = cpgge
            if mjdj1 > 0:
                ws["G19"] = "price RMB"
                ws["H19"] = mjdj1
            else:
                ws["G19"] = "price USD"
                ws["H19"] = jg
            ws["J19"] = bzhfsh
            ws["H20"] = czyy
            ws["H21"] = bz
            ln = 0
            if hwbh:
                insert_image_to_excel(hwbh, prod.rid, wb, ws, 8,17)
    path = config.tmp_path
    logger.error(f"临时路径: {path}")
    os.makedirs(path, exist_ok=True)
    report_rid = get_uuid()
    output_filename = f'{report_rid}_寄样吊牌稿.xlsx'
    full_output_path = os.path.join(path, output_filename)
    logger.error(f"导出路径: {full_output_path}")
    wb.save(full_output_path)
    # return {'code': 1, 'msg': '报表导出成功', 'data': output_filename}
    return JSONResponse({
        'code': 1,
        'msg': '报表导出成功',
        'data': {
            'filename': output_filename,
            'full_path': full_output_path,
            'total_products': zts
        }
    })



@any_route("/api/khjy/generate_sample_tag_report_jydy1", methods=["POST"])
@require_token
async def generate_sample_tag_report_jydy1(request):
    j = await request.json()
    rid_list = j.get("rids")
    pdf = j.get("pdf")
    da2 = j.get("da2")
    logger.error(f"导出参数: pdf={pdf}, da2={da2}")
    if not rid_list:
        return json_result(-1, "未选择导出记录", None)
    template_path = os.path.join(config.data_upload_path, "template/客户寄样-1.xlsx")
    if not os.path.exists(template_path):
        return json_result(-1, "[客户寄样-1]模板文件不存在", None)
    s = Session()
    d = s.query(jygl.jyrq,jygl.khmc).filter(jygl.rid==rid_list).first()
    if not d:
        return json_result(-1, f"请检查数据是否正确!")
    jyrq = d.jyrq
    khmc = d.khmc
    if pdf != "2":
        pdf = "1"
    if da2 == "2":
        UV = "US"
    else:
        UV = "UV"
    zts = s.query(func.count(jyglcp.rid)).filter(jyglcp.pid == rid_list).scalar() or 0
    ys1 = zts
    iz = zts
    ys2 = zts
    if ys1 == ys2:
        ys3 = ys1
        iz = iz - 1
    else:
        ys3 = ys1 + 1
    wb = load_workbook(template_path)
    ws = wb.worksheets[0]
    ws.title = "客户寄样"
    a1z = 1
    for az in range(iz):
        source_sheet = wb.worksheets[0]
        new_sheet = wb.create_sheet("客户寄样")
        copy_worksheet_full(source_sheet, new_sheet)
    rows = s.query(jyglcp).filter(jyglcp.pid == rid_list).all()
    for idx, prod in enumerate(rows):
        ws = wb.worksheets[idx]
        hwbh = prod.hwbh
        ywpm = prod.ywpm
        cpgge = prod.cpgge
        mjdj1 = prod.mjdj1
        jg = prod.jg
        bzhfsh = prod.bzhfsh
        czyy = prod.czyy
        bz = prod.bz
        ws["B1"] = UV
        ws["D1"] = jyrq
        ws["B2"] = hwbh
        ws["B3"] = ywpm
        ws["B5"] = cpgge
        # ws["G5"] = cpgge

        if mjdj1 > 0:
            ws["A6"] = "price RMB"
            ws["B6"] = mjdj1
        else:
            ws["A6"] = "price USD"
            ws["B6"] = jg

        ws["D6"] = bzhfsh
        ws["B7"] = czyy
        ws["B8"] = bz
        insert_image_to_excel(hwbh, prod.rid, wb, ws, 2,4)
    path = config.tmp_path
    logger.error(f"临时路径: {path}")
    os.makedirs(path, exist_ok=True)
    report_rid = get_uuid()
    output_filename = f'{report_rid}_客户寄样.xlsx'
    full_output_path = os.path.join(path, output_filename)
    logger.error(f"导出路径: {full_output_path}")
    wb.save(full_output_path)
    final_file_name=''
    if pdf == "1":
        print(f'\n========== 开始PDF转换（类型: {pdf}） ==========')
        
        pdf_result = convert_excel_to_pdf(full_output_path, path)
        
        if pdf_result['success']:
            pdf_full_path = pdf_result['pdf_path']
            pdf_file_name = output_filename.replace('.xlsx', '.pdf')
            
            final_file_path = pdf_full_path
            final_file_name = pdf_file_name
            sbs_path = path[-10:]
            file_path = sbs_path + '/' + final_file_name
            logger.error(f"PDF路径: {file_path}")
            print(f'✓ PDF生成成功')
        else:
            print(f'✗ PDF转换失败: {pdf_result["error"]}')
            return json_result(code=-1, msg=pdf_result['error'])
    return JSONResponse({
        'code': 1,
        'msg': '报表导出成功',
        'data': {
            'filename': output_filename,
            'pdf_filename': final_file_name,
            'full_path': full_output_path,
            'total_products': zts
        }
    })
