from any import *
from sqlalchemy.sql import func,not_,or_,and_
import json,os
from .model import *
from .__default__ import get_user_path, user_task_new, module_xxck_new
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image as OpenpyxlImage
import traceback
from sqlalchemy import text
from PIL import Image as PILImage
import tempfile

from copy import copy
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU

# 报价统计表 
@any_route('/api/saier/quotation/stats/export', methods=['POST'])
@require_token
async def api_saier_quotation_stats_export(request):
    j = await request.json()
    try:
        # 获取前端传来的参数
        qsrq = j.get('start_date')
        jsrq = j.get('end_date')
        cgryaa = j.get('purchaser', '')
        
        # 获取当前用户信息 (假设挂载在 request 上)
        user = request.current_user
        user_name = user.username

        
        res = generate_quotation_excel(qsrq, jsrq, cgryaa, user_name)
        return json_result(res.get('code'), res.get('msg'), res.get('data'))
    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, trace_error())


def generate_quotation_excel(qsrq, jsrq, cgryaa, user_name):
    s = Session()
    try:
        # 【权限校验模拟】检查是否为经理/总监/采购
        # 实际开发中建议用权限中间件拦截，此处保留 Delphi 业务逻辑
        user_info = s.execute(
            "SELECT path FROM sys_user WHERE username = :name",
            {'name': user_name}
        ).fetchone()
        if not user_info:
            return {'code': -1, 'msg': '未找到当前用户信息，请重新登录'}
        org=get_user_path(user_name)
        user_path=org.get('path', '')
       
        # is_manager = s.execute(
        #     """SELECT 1 FROM sys_user
        #        WHERE name=:name AND (position LIKE '%经理%' OR position LIKE '%总%' OR position LIKE '%采购%')""",
        #     {'name': user_name}
        # ).fetchone()
        
        # if not is_manager:
        #     return {'code': -1, 'msg': '您没有权限导出该报表'}

        # =========================================================
        # 第一步：优化查询 —— 仅用两次 SQL 获取所有必要数据
        # =========================================================
        
        # 1. 查询符合条件的主表数据 (bj)
        # 注意：此处用原生 SQL 示范，以便无缝对照 Delphi 的 SQL
        bj_sql = """
            SELECT ywry, customer, kh_id, rid 
            FROM bj 
            WHERE dateis >= :qsrq AND dateis <= :jsrq 
              AND bjjg = '通过' 
              AND ywry LIKE :cgryaa
        """
        bj_records = s.execute(bj_sql, {
            'qsrq': qsrq, 'jsrq': jsrq, 
             'cgryaa': f'%{cgryaa}%'
        }).fetchall()

        if not bj_records:
            return {'code': -1, 'msg': '该日期范围内没有符合条件的报价数据'}

        # 提取所有的 bj rid 用于 IN 查询子表
        bj_rid = [r.rid for r in bj_records]
        
        # 2. 批量查询子表数据 (bjsheet)
    
        bjsheet_sql = """
            SELECT pid, bjry, wxxz, bjbjwyzd, cply1, wxbm 
            FROM bjsheet 
            WHERE pid IN :rids
        """
        # 注意：如果 rids 非常大，SQLAlchemy 会自动处理，但极大并发时考虑分片
        bjsheet_records = s.execute(bjsheet_sql, {'rids': tuple(bj_rid)}).fetchall()

        # 将子表数据按 father (主表rid) 进行 Hash 映射，提升匹配速度 O(1)
        detail_map = {}
        for row in bjsheet_records:
            detail_map.setdefault(row.pid, []).append(row)

        # =========================================================
        # 第二步：内存聚合数据 (替代 Delphi 的各种 StringList indexOf)
        # =========================================================
        
        # 数据结构: stats_data[ywry][customer][kh_id] = { 统计指标 }
        stats_data = {}
        
        for bj in bj_records:
            ywry = bj.ywry
            customer = bj.customer
            kh_id = bj.kh_id
            
            if ywry not in stats_data:
                stats_data[ywry] = {}
            if customer not in stats_data[ywry]:
                stats_data[ywry][customer] = {}
            if kh_id not in stats_data[ywry][customer]:
                # 初始化统计结构
                stats_data[ywry][customer][kh_id] = {
                    'cgbj': 0, 'wxbj': 0,
                    'FD': 0, 'FD1': 0, 'XJ': 0, 'XJ1': 0,
                    'CGTJ': 0, 'CGTJ1': 0, 'WXTJ': 0, 'WXTJ1': 0,
                    'KHZX': 0, 'KHZX1': 0, 'YPLL': 0, 'YPLL1': 0,
                    'wxbm_set': set(), 'wxbm_dict': {} # 用于外销部门统计
                }
            
            node = stats_data[ywry][customer][kh_id]
            details = detail_map.get(bj.rid, [])
            
            for d in details:
                # 只统计 bjry 匹配的明细
                if d.bjry != ywry:
                    continue
                    
                node['cgbj'] += 1
                is_selected = (d.wxxz == '是')
                if is_selected:
                    node['wxbj'] += 1
                
                # 外销部门合并统计逻辑 (对应 Delphi 的 wxbmbj 拼接)
                if is_selected and d.bjbjwyzd:
                    if d.bjbjwyzd not in node['wxbm_set']:
                        node['wxbm_set'].add(d.bjbjwyzd) # 记录该字段，避免后续循环重复累加
                        wxbm = str(d.wxbm).strip() if d.wxbm else ''
                        node['wxbm_dict'][wxbm] = node['wxbm_dict'].get(wxbm, 0) + 1
                
                  

                # 来源统计 (通过 Python 字符串 in 模拟 Delphi 的 POS 函数)
                cply = str(d.cply1 or '')
                if '返单' in cply:
                    node['FD'] += 1
                    if is_selected: node['FD1'] += 1
                if '询价' in cply:
                    node['XJ'] += 1
                    if is_selected: node['XJ1'] += 1
                if '采购推荐' in cply:
                    node['CGTJ'] += 1
                    if is_selected: node['CGTJ1'] += 1
                if '外销推荐' in cply:
                    node['WXTJ'] += 1
                    if is_selected: node['WXTJ1'] += 1
                if '客户自选' in cply:
                    node['KHZX'] += 1
                    if is_selected: node['KHZX1'] += 1
                if '样品录入' in cply:
                    node['YPLL'] += 1
                    if is_selected: node['YPLL1'] += 1

        # =========================================================
        # 第三步：生成 Excel (使用 openpyxl)
        # =========================================================
        wb = Workbook()
        ws = wb.active
        
        # 居中样式
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 表头配置 (对应 Delphi 代码 A1 到 P1)
        headers = [
            '姓名', '报价产品数', '外销选中数', '外销报价命中率', '返单', '询价',
            '采购推荐', '外销推荐', '客户自选', '样品录入', '外销合同选中数', 
            '外销合同命中率', '采购合同选中数', '采购合同命中率', '客户名称', '客户详情'
        ]
        ws.append(headers)
        
        # 设置列宽
        for col_idx in range(1, 16):  # A 到 O
            ws.column_dimensions[get_column_letter(col_idx)].width = 14.3
        ws.column_dimensions['P'].width = 120 # P列客户详情较长
        
        row_idx = 2
        
        # 遍历组装好的字典并写入 Excel
        for ywry, customers in stats_data.items():
            for customer, kh_dict in customers.items():
                
                # 汇总该 业务员+客户 级别的数据
                total_cgbj = sum(kh['cgbj'] for kh in kh_dict.values())
                total_wxbj = sum(kh['wxbj'] for kh in kh_dict.values())
                
                # 其他累加项 (利用字典推导式快速累加)
                sums = {k: sum(kh[k] for kh in kh_dict.values()) for k in ['FD', 'FD1', 'XJ', 'XJ1', 'CGTJ', 'CGTJ1', 'WXTJ', 'WXTJ1', 'KHZX', 'KHZX1', 'YPLL', 'YPLL1']}
                
                # 计算命中率
                hit_rate = round((total_wxbj / total_cgbj), 3) if total_cgbj > 0 else 0.000
                
                # 组装外销选中数拼接字符串 (C列)
                wxbmbj_list = []
                # 遍历收集所有子节点的 wxbm_dict
                for kh in kh_dict.values():
                    for k, v in kh['wxbm_dict'].items():
                        wxbmbj_list.append(f"{k} {v}")
                wxbmbj_str = str(total_wxbj) + ("\n" + "\n".join(wxbmbj_list) if wxbmbj_list else "")
                
                # 组装客户详情 (P列)
                khxq_lines = []
                for kh_id, kh in kh_dict.items():
                    mz = round((kh['wxbj'] / kh['cgbj']), 3) if kh['cgbj'] > 0 else 0.000
                    khxq_lines.append(
                        f"客户编号:{kh_id},报价产品数:{kh['cgbj']},外销选中数:{kh['wxbj']},外销报价命中率:{mz:.3f},"
                        f"返单:{kh['FD']}({kh['FD1']}),询价:{kh['XJ']}({kh['XJ1']}),"
                        f"采购推荐:{kh['CGTJ']}({kh['CGTJ1']}),外销推荐:{kh['WXTJ']}({kh['WXTJ1']}),"
                        f"客户自选:{kh['KHZX']}({kh['KHZX1']}),样品录入:{kh['YPLL']}({kh['YPLL1']});"
                    )
                khxq_str = "\n".join(khxq_lines)
                
                # 写入行数据 (外销/采购合同数量遵循原逻辑硬编码0)
                row_data = [
                    ywry, total_cgbj, wxbmbj_str, f"{hit_rate:.3f}",
                    f"{sums['FD']}({sums['FD1']})", f"{sums['XJ']}({sums['XJ1']})",
                    f"{sums['CGTJ']}({sums['CGTJ1']})", f"{sums['WXTJ']}({sums['WXTJ1']})",
                    f"{sums['KHZX']}({sums['KHZX1']})", f"{sums['YPLL']}({sums['YPLL1']})",
                    0, "0.000", 0, "0.000", customer, khxq_str
                ]
                
                ws.append(row_data)
                
                # 设置整行居中对齐
                for cell in ws[row_idx]:
                    cell.alignment = center_align
                    
                row_idx += 1

        # 保存文件
        path = config.tmp_path
        report_rid = get_uuid()
        wb.save(path + '/'+ str(report_rid)+'.xlsx')
        return {'code':1,'msg':'生成报表成功','data': str(report_rid)+'.xlsx'}
    

    except Exception as e:
        logger.error(trace_error())
        return {'code': -1, 'msg': str(e)}
    finally:
        s.close()




# 特殊报表  请求获取zm 列表
@any_route('/api/saier/quotation/special/get_list', methods=['POST'])
@require_token
async def get_report_template_list(request):
    s = Session()
    try:
        # 对应原代码：select zm from cyzgl where qxzl="特殊报表打印"
        sql = "SELECT zm FROM cyzgl WHERE qxzl = :category"
        result = s.execute(text(sql), {"category": "特殊报表打印"}).fetchall()
        
        # 将结果转为简单的字符串列表供前端渲染
        zm_list = [row[0] for row in result]
        return json_result(1, '', zm_list)
    finally:
        s.close()




# ==============  特殊报表打印 ======================

# 常量：EMU 每像素的换算（常用近似值）
EMU_PER_PIXEL = 9525  # 近似：1 px = 9525 EMU

def get_row_value(row, key):
    # Try attribute
    if hasattr(row, key):
        return getattr(row, key)
    # Try mapping (Row)
    try:
        return row[key]
    except Exception:
        try:
            return row._mapping.get(key)
        except Exception:
            return None

def excel_col_width_to_pixels(col_width):
    """
    估算 openpyxl 列宽到像素的换算（近似）。
    Excel 的精确计算比较复杂；此处使用常用近似：
      pixels = int(col_width * 7 + 5)
    这个在多数常见场景下能给出可接受的近似值。
    """
    if col_width is None:
        col_width = 8.43  # Excel 默认宽度
    try:
        return int(col_width * 7 + 5)
    except Exception:
        return int(8.43 * 7 + 5)

def excel_row_height_to_pixels(row_height):
    """
    openpyxl row.height 单位是 points。1 point = 1.333333 px (假定屏幕 96 DPI)。
    """
    if row_height is None:
        row_height = 15  # Excel 默认行高（points）
    try:
        return int(row_height * 96 / 72)  # = row_height * 1.3333
    except Exception:
        return int(row_height * 1.3333)

def scale_and_center_image(img_path, target_w_px, target_h_px, margin_px=5):
    """
    计算缩放后宽高（像素）与偏移（像素），用于在目标像素尺寸区域内居中显示图片。
    margin_px: 对应你 Delphi 中的 -5 留白
    返回 (new_w_px, new_h_px, offset_x_px, offset_y_px)
    """
    with PILImage.open(img_path) as pil:
        img_w, img_h = pil.size

    # 避免负或零尺寸
    tw = max(1, target_w_px - margin_px)
    th = max(1, target_h_px - margin_px)

    ratio = min(tw / img_w, th / img_h)
    ratio = min(ratio, 1.0)  # 通常不强制放大原图（如果需要放大可去掉）

    new_w = int(round(img_w * ratio))
    new_h = int(round(img_h * ratio))

    offset_x = (target_w_px - new_w) // 2
    offset_y = (target_h_px - new_h) // 2

    return new_w, new_h, offset_x, offset_y

def insert_image_into_cell_ws(ws, cell_address, img_path, target_cell_col=None, target_cell_row=None):
    """
    使用 openpyxl 插入图片到指定单元格，并尽量应用像素级偏移以实现居中显示。
    - cell_address: Excel address string like 'B5' or template anchor string supplied in DB.
    - 如果 target_cell_col/row 已知可直接传入以避免解析。
    说明：openpyxl 不像 Excel COM 那样完美支持像素级控制，但我们通过修改 img.anchor._from 的 colOff/rowOff 来接近。
    """
    # 解析目标单元格
    if target_cell_col is None or target_cell_row is None:
        # 解析 cell_address
        from openpyxl.utils import coordinate_from_string
        (col_letter, row_idx) = coordinate_from_string(cell_address)
        col_idx = column_index_from_string(col_letter)
        row_idx = row_idx
    else:
        col_idx = target_cell_col
        row_idx = target_cell_row
        col_letter = get_column_letter(col_idx)

    # 估算单元格像素尺寸（列宽 * 列像素 + 行高像素）
    # 获取列宽（openpyxl 存储在 ws.column_dimensions）
    col_dim = ws.column_dimensions.get(get_column_letter(col_idx))
    col_width = None
    if col_dim is not None and hasattr(col_dim, "width"):
        col_width = col_dim.width
    # fallback: try workbook default or None
    # 估算像素宽度
    cell_w_px = excel_col_width_to_pixels(col_width)

    # 获取行高
    row_dim = ws.row_dimensions.get(row_idx)
    row_height = None
    if row_dim is not None and hasattr(row_dim, "height"):
        row_height = row_dim.height
    cell_h_px = excel_row_height_to_pixels(row_height)

    new_w, new_h, off_x_px, off_y_px = scale_and_center_image(img_path, cell_w_px, cell_h_px)

    img = OpenpyxlImage(img_path)
    img.width = new_w
    img.height = new_h

    # Set anchor to the target cell
    img.anchor = cell_address  # e.g. 'B5'

    # Set internal offset via EMU (may work in many cases)
    try:
        # Ensure anchor._from exists and set offsets
        img.anchor._from.col = col_idx - 1
        img.anchor._from.row = row_idx - 1
        img.anchor._from.colOff = int(off_x_px * EMU_PER_PIXEL)
        img.anchor._from.rowOff = int(off_y_px * EMU_PER_PIXEL)
    except Exception:
        # 如果内部结构不同，则忽略（openpyxl 版本/对象差异）
        pass

    ws.add_image(img)  # 在 openpyxl 中：如果 img.anchor 给定，则不需要指定第二个参数

# 特殊报表打印
@any_route('/api/saier/quotation/special_print', methods=['POST','GET'])
@require_token
async def api_saier_quotation_special_print(request):
    s = Session()
    j = await request.json()
    rid = j.get('rid')
    zm = j.get('zm')
    try:
        # 0. 基本校验
        if not rid:
            return json_result(-1, "缺少 rid 参数")
        if not zm:
            return json_result(-1, "缺少 template_name (zm) 参数")
  
        # 1. 审批/权限检查（保持与 Delphi 逻辑一致：检查 bj 表中的 bjql（是否通过））
        bj_row = s.execute(text("SELECT bjql FROM bj WHERE rid=:rid"), {"rid": rid}).fetchone()
        if not bj_row:
            return json_result(-1, "未找到报价单记录")

        bjql=bj_row.bjql
        if bjql != '通过':
            return json_result(-1, "报价单未通过审批，不能打印特殊报表")

        # 2. 查询模板配置 (cyzgl)
        tpl_row = s.execute(text("SELECT tpl, qsh, sfdhy, ywxh, scmb FROM cyzgl WHERE zm=:zm"), {"zm": zm}).fetchone()
        if not tpl_row:
            return json_result(-1, f"未找到模板配置: {zm}")

        tpl_config = tpl_row  # keep name parity

        # 3. 读取字段映射明细 (cyzglsheet)
        mappings = s.execute(text("SELECT xm, bz, bz3, sz, sz1 FROM cyzglsheet WHERE zm=:zm"), {"zm": zm}).fetchall()

        # 4. 获取报价单数据 (bjsheet)
        items = s.execute(text("SELECT * FROM bjsheet WHERE pid=:rid"), {"rid": rid}).fetchall()
        if not items:
            return json_result(-1, "没有找到报价单明细")

        # 5. 模板来源：可能存 DB(scmb) 或 文件系统
        template_tempfile = None

        rid_row = s.execute(text("SELECT rid FROM cyzgl WHERE qxzl=:qxzl AND zm=:zm"), {"qxzl": "特殊报表打印", "zm": zm}).fetchone()
        rid_val = rid_row.rid
        report_path = ''
        if rid_val:
            att = s.execute(text("SELECT path FROM sys_attachment WHERE pid=:rid AND path IS NOT NULL LIMIT 1"), {"rid": rid_val}).fetchone()
            if att:
                report_path = att.path   
          
        if not report_path:
            return json_result(-1, f"未找到报表模板: {zm}")
        r_path = os.path.join(config.data_upload_path,report_path)

        if not os.path.exists(r_path):
            return {"code":-1,"msg":"未找到报表模板"}


        # 6. 转换 xls 为 xlsx
        # if report_path.endswith('.xls'):
        #     xlsx_path = convert_xls_to_xlsx_via_soffice(r_path)
        #     r_path = xlsx_path
        
        # print(f"转换后的 xlsx 文件路径: {xlsx_path}")
       
        wb = load_workbook(r_path) 
        # logger.error(f"成功加载工作簿aaa: {r_path}")
        # # wb = load_workbook(r_path)
        # logger.error(f"成功加载工作簿aaaa: {r_path}")
        # Helper: safe get maximum column index used by mappings
        max_col_idx = 0
        for m in mappings:
            try:
                idx = column_index_from_string(m.xm)
                max_col_idx = max(max_col_idx, idx)
            except Exception:
                return json_result(-1, f"映射字段 {m.xm} 格式错误")
        
        # --- 模式一：单页模式 (sfdhy == '是') ---
        if get_row_value(tpl_config, "sfdhy") == '是':
            for i, item in enumerate(items):
                if i == 0:
                    ws = wb.active
                else:
                    ws = wb.copy_worksheet(wb.worksheets[0])
                ws.title = f"产品_{i+1}"

                # 填充字段值
                for m in mappings:
                    raw = get_row_value(item, m.bz)
                    if get_row_value(m, "bz3") == 'FLOAT' and (get_row_value(m, "sz") or 0) > 0:
                        divisor = get_row_value(m, "sz") or 1
                        try:
                            val = float(raw or 0) / float(divisor)
                        except Exception:
                            val = raw
                    else:
                        val = raw
                    target_cell = f"{m.xm}{int(m.sz1)}"
                    ws[target_cell] = val

                # 处理图片（tphh 字段）
                img_code = get_row_value(item, "tphh")
                if img_code:
                    # 支持多种扩展名查找
                    for ext in (".jpg", ".jpeg", ".png", ".JPG", ".JPEG", ".PNG"):
                        img_path = f"/data/images/{img_code}{ext}"
                        if os.path.exists(img_path):
                            break
                    else:
                        img_path = None

                    # tpl_config.tpl 储存了目标单元格地址（例如 'B6'），或者你们的实现可能是 'Sheet1!B6'
                    anchor = get_row_value(tpl_config, "tpl")
                    if img_path and anchor:
                        # 如果 tpl 带 sheet 名，移除 sheet 前缀
                        if "!" in anchor:
                            _, anchor = anchor.split("!", 1)
                        try:
                            insert_image_into_cell_ws(ws, anchor, img_path)
                        except Exception as e_img:
                            # 记录错误，但不致命（可改为抛出）
                            print("插入图片失败:", e_img)

        # --- 模式二：列表模式 (sfdhy != '是') ---
        else:
            ws = wb.active
            qsh = int(get_row_value(tpl_config, "qsh") or 1)
            # 复制模板行格式的帮助函数
            def copy_row_format(ws, src_row, dst_row, max_col):
                # 插入后逐单元格复制 style
                for col in range(1, max_col + 1):
                    src = ws.cell(row=src_row, column=col)
                    dst = ws.cell(row=dst_row, column=col)
                    # 复制 style/格式
                    try:
                        dst.font = copy(src.font)
                        dst.border = copy(src.border)
                        dst.fill = copy(src.fill)
                        dst.number_format = copy(src.number_format)
                        dst.protection = copy(src.protection)
                        dst.alignment = copy(src.alignment)
                    except Exception:
                        pass
                # 复制行高
                if src_row in ws.row_dimensions:
                    try:
                        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
                    except Exception:
                        pass
                # 复制合并区域，如果源行中有合并单元格，按 dst_row 创建相同列范围的合并
                merges_to_add = []
                for merged in list(ws.merged_cells.ranges):
                    min_col, min_row, max_col, max_row = merged.bounds
                    # 如果合并区域在 src_row 行范围内（整行）
                    if min_row <= src_row <= max_row:
                        width_rows = max_row - min_row
                        # 构造新的合并，偏移到 dst_row（简单策略：如果合并仅在一行内则映射到 dst_row）
                        if width_rows == 0:
                            a = get_column_letter(min_col) + str(dst_row)
                            b = get_column_letter(max_col) + str(dst_row)
                            merges_to_add.append(f"{a}:{b}")
                # 应在复制后添加合并（避免干扰循环）
                for rng in merges_to_add:
                    try:
                        ws.merge_cells(rng)
                    except Exception:
                        pass

            for i, item in enumerate(items):
                curr_row = qsh + i
                if i > 0:
                    ws.insert_rows(curr_row)
                    # 复制模板 qsh 行格式到新插入行
                    try:
                        copy_row_format(ws, qsh, curr_row, max_col_idx)
                    except Exception:
                        pass

                # 写入数据
                for m in mappings:
                    raw = get_row_value(item, m.bz)
                    if get_row_value(m, "bz3") == 'FLOAT' and (get_row_value(m, "sz") or 0) > 0:
                        divisor = get_row_value(m, "sz") or 1
                        try:
                            val = float(raw or 0) / float(divisor)
                        except Exception:
                            val = raw
                    else:
                        val = raw
                    ws[f"{m.xm}{curr_row}"] = val

                # 序号处理
                if get_row_value(tpl_config, "ywxh") == '有':
                    ws[f"A{curr_row}"] = i + 1

            # 如 Delphi 在结束后删除模板行 qsh（如果需要）
            try:
                ws.delete_rows(qsh)
            except Exception:
                pass

        # 7. 保存结果到临时目录并返回 uuid 文件名
        path = config.tmp_path
        os.makedirs(path, exist_ok=True)
        report_rid = str(uuid.uuid4())
        out_name = f"{report_rid}.xlsx"
        out_path = os.path.join(path, out_name)
        wb.save(out_path)
        
        # 清理临时模板文件（如果有）
        if template_tempfile and os.path.exists(template_tempfile):
            try:
                os.remove(template_tempfile)
            except Exception:
                pass
        print(f"生成报表成功: {out_path}")
        return json_result(1, '生成报表成功', out_name)

    except Exception as e:
        tb = traceback.format_exc()
        # 打印到服务日志，方便运维查看
        print("Exception in special_print:\n", tb)
        return json_result(-1, str(e))
    finally:
        s.close()





# 自定义报价表 导出

# --- 辅助工具函数：搬运合同模块的神级图片精确偏移算法 ---
def offset_img(img, col, row, x_pad=4, y_pad=25):
    """精确设置图片位置，偏移量以像素为单位进行微调"""
    p2e = pixels_to_EMU
    h, w = img.height, img.width
    size = XDRPositiveSize2D(p2e(w), p2e(h))
    # 注意 openpyxl anchor 的行列索引是从 0 开始的
    marker = AnchorMarker(col=col-1, colOff=pixels_to_EMU(x_pad), row=row-1, rowOff=pixels_to_EMU(y_pad))
    img.anchor = OneCellAnchor(_from=marker, ext=size)

@any_route('/api/saier/quotation/custom/export', methods=['POST'])
@require_token
async def api_custom_quotation_export(request):
    j = await request.json()
    record_id = j.get('rid')
    include_image = j.get('include_image', 1)

    if not record_id:
        return json_result(-1, "缺少单据ID")

    # 提取当前登录用户名 (WhaleCloud 标准)
    user = request.current_user
    user_name = user.username 
    s = Session()
    try:
        # ================= [ Part 1: 严格还原 Delphi 权限与人员校验 ] =================
        # 使用 default_function.py 的 get_user_path 方法判断权限
        user_info = get_user_path(user_name)
        user_path = user_info.get('path', '')
        
        if '优景外销' not in user_path and user_name not in ('zjnblh', '侯柳红'):
            return json_result(-1, "您没有使用此导出功能的权限")

        # ================= [ Part 2: 获取用户专属模板 (参照 special_report 风格) ] =================
        rid_row = s.execute(
            sql_text("SELECT rid FROM cyzglsheet WHERE zm='报价自定义模板' AND xm=:xm"), 
            {'xm': user_name}
        ).fetchone()

        if not rid_row or not rid_row.rid:
            return json_result(-1, "未找到您的专属报价模板，请先在系统中配置")

        rid_val = rid_row.rid
        
        # ================= [ Part 3: 获取报表附件路径 ] =================
        att = s.execute(
            sql_text("SELECT path FROM sys_attachment WHERE pid=:rid AND path IS NOT NULL LIMIT 1"), 
            {"rid": rid_val}
        ).fetchone()

        if not att or not att.path:
            return json_result(-1, "未找到报表模板附件")

        tpl_path = os.path.join(config.data_upload_path, att.path)
        if not os.path.exists(tpl_path):
            return json_result(-1, "未找到报表模板物理文件")

        # ================= [ Part 4: 加载模板并侦测动态表头 ] =================
        wb = load_workbook(tpl_path)
        ws = wb.worksheets[0]
        
        headers = []
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col_idx).value
            if not cell_value or str(cell_value).strip() == '':
                break
            headers.append(str(cell_value).strip())

        if not headers:
            return json_result(-1, "模板第一行未读取到有效的表头字段名")

        # ================= [ Part 5: 查询明细数据 ] =================
        # ⚠️ 注意：此处需替换为真实报价单明细表名 (例如 bjmxsheet)
        items = s.execute(
            sql_text("SELECT * FROM bjmxsheet WHERE pid=:rid ORDER BY id ASC"), 
            {'rid': record_id}
        ).fetchall()

        # ================= [ Part 6: 填入数据与图片排版 ] =================
        row_cursor = 2 
        # 原逻辑：d:\zjnblh图片\
        img_base_dir = os.path.join(config.data_upload_path, 'zjnblh图片') 

        for item in items:
            # 严格设定行高，配合图片尺寸
            ws.row_dimensions[row_cursor].height = 123.75 
            item_dict = item._mapping if hasattr(item, '_mapping') else dict(item)

            for col_idx, field_name in enumerate(headers, start=1):
                #  Delphi 逻辑：第一列 (z=1) 为图片列
                if col_idx == 1 and include_image == 1:
                    # 兼容不同字段命名
                    cpbh12 = str(item_dict.get('专业货号') or item_dict.get('cpbh') or '') 
                    if cpbh12:
                        img_path = os.path.join(img_base_dir, f"{cpbh12}.JPG")
                        if os.path.exists(img_path):
                            try:
                                img = XLImage(img_path)
                                # 借助 PIL 获取原尺寸计算等比缩放
                                with Image.open(img_path) as pil_img:
                                    orig_w, orig_h = pil_img.size
                                    ratio = min(150 / orig_w, 150 / orig_h)
                                    img.width = int(orig_w * ratio)
                                    img.height = int(orig_h * ratio)
                                
                                # 💡 核心升级：调用精准偏移函数 (左移6px，下移6px防压线)
                                offset_img(img, col=col_idx, row=row_cursor, x_pad=6, y_pad=6)
                                ws.add_image(img)
                            except Exception as img_err:
                                logger.warning(f"图片插入失败 {img_path}: {trace_error()}")
                else:
                    # 文本写入与格式化
                    val = item_dict.get(field_name, '')
                    ws.cell(row=row_cursor, column=col_idx, value=val)
                    ws.cell(row=row_cursor, column=col_idx).alignment = Alignment(
                        horizontal='center', vertical='center', wrap_text=True
                    )
            
            row_cursor += 1

        # ================= [ Part 7: 生成、保存与清理 ] =================
        s_path = config.tmp_path
        # os.makedirs(s_path, exist_ok=True)
        filename = "报价自定义模板.xlsx"
        final_path = os.path.join(s_path, filename)
        wb.save(final_path)

        return json_result(1, '单据生成成功', filename)

    except Exception as e:
        logger.error(trace_error())
        return json_result(-1, f"生成自定义报价单报错: {str(e)}")
    finally:
        s.close()