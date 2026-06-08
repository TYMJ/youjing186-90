import asyncio
from concurrent.futures import ThreadPoolExecutor
from any import *  
from .model import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors, Alignment
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
import traceback
import io
from openpyxl.utils import coordinate_to_tuple
from email import errors
from math import e
from pdb import run
from webbrowser import get
from any import *
from .model import *
from sqlalchemy.sql import func,not_,or_,and_
import time
import json,os
from .__default__ import get_user_path,module_xxck_new,module_share_new
from openpyxl.drawing.image import Image as Image_Get
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, Alignment
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from PIL import Image as PILImage
from datetime import datetime, timedelta
import re
def clean_filename(filename):
        # 替换 Windows 文件名中的非法字符
        illegal_chars = r'[<>:"/\\|?*]'
        return re.sub(illegal_chars, '_', filename)


def safe_write(ws, coord, value, num_format=None):
    """
    智能穿透写入：如果目标单元格被合并了，自动找到该合并区域的左上角主单元格进行写入。
    """
    row, col = coordinate_to_tuple(coord)
    target_cell = ws[coord]
    
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            target_cell = ws.cell(row=mr.min_row, column=mr.min_col)
            break
            
    target_cell.value = value
    if num_format:
        target_cell.number_format = num_format


def crate_excel_logic(recordset, user):
    """
    生成 BP CO 清单 Excel 报表（按 Delphi 逻辑调整）
    当前时间：2026-05-11 16:03:50 星期一
    """
    # ========== 第一步：初始化变量 ==========
    k = 0          # 行计数器
    khje = 0.0     # 客户总金额（人民币）
    mjje = 0.0     # 美金总金额
    fphm = ''      # 发票号码
    sb = ''        # 临时变量
    lj = ''        # 保存路径
    hl1 = ''       # 默认汇率
    dysb = ''      # 权限标志
    
    cpdj = []      # 存储 "产品编号;单价" 组合，用于去重
    cpbh = []      # 存储不重复的产品编号
    djpmy = []     # 存储不重复的单价
    hl = recordset.get('hl')
    if hl is None or hl == 0:
            hl = 7
    else:
            hl = recordset.get('hl')
    # ========== 第二步：检查模板文件 ==========
    r_path = os.path.join(config.data_upload_path, 'template')
    template_file = 'BP CO清单.xlsx'
    fn = os.path.join(r_path, template_file)
    
    if not os.path.exists(fn):
        return {'code': -1, 'msg': f"未找到报表模板: {template_file}", 'data': ''}
    
    wb = load_workbook(fn)
    ws = wb.active
    
    s = Session()
    try:
        # ========== 第三步：权限验证（对应 Delphi 中的 tmpcomcgs）==========
        # g = run_sql(
        #     "select rid, position from sys_user where (username=:user) and (position like '%单证%')",
        #     {"user": user}
        # )
        
        # if g:
        #     dysb = '1'
        
        # if dysb != '1':
        #     return {'code': -1, 'msg': f"没有权限: {user}", 'data': ''}
        
        # ========== 第四步：读取配置（对应 Delphi 中的 cyzgl 查询）==========
        c = run_sql("select wb, ywxh from cyzgl where (zm='BP CO清单抬头')")
        if c:
            lj = c[0].get('wb', '')
            hl1 = c[0].get('ywxh', '')
            lj=lj.replace('$', 'usd')
            
        # ========== 第五步：获取汇率 ==========
        
        
        # ========== 第六步：查询并去重产品数据（对应 Delphi 中的 while 循环）==========
        a = run_sql(
            "select cpbh, djpmy from cymxsheet where (pid=:rid) and (chxs>0) order by rid",
            {"rid": recordset.get('rids')[0]}
        )
        
        logger.error(f"查询结果 a: {a}")
        
        # Delphi 逻辑：遍历结果集，用 cpdj.indexof 去重
        for row in a:
            combination = f"{row['cpbh']};{row['djpmy']}"
            
            # 模拟 Delphi 中的 indexof 方法
            j1 = -1
            for index, item in enumerate(cpdj):
                if item == combination:
                    j1 = index
                    break
            
            # 如果不存在则添加（对应 Delphi 中的 if j1 < 0）
            if j1 < 0:
                cpdj.append(combination)
                cpbh.append(row['cpbh'])
                djpmy.append(row['djpmy'])
        
       
        
        # ========== 第七步：设置 J2 单元格 ==========
        ws['J2'] = f'客户总美金(人民币除以{hl})'
        
        # ========== 第八步：循环处理每个不重复的产品（对应 Delphi 中的 for l:=0 to cpdj.count-1）==========
        for l in range(len(cpdj)):
            # 查询当前产品的汇总数据（对应 Delphi 中的 tmpcom 查询）
            b = run_sql(
                """
                select sum(chxs) as chxsz1, sum(chsl) as chslz1, sum(mjzj) as mjzjz1,
                       cpbh, djpmy
                from cymxsheet
                where (pid=:rid) and (chxs>0) and (cpbh=:cpbh) and (djpmy=:djpmy)
                group by cpbh, djpmy
                """,
                {
                    "rid": recordset.get('rids')[0],
                    "cpbh": cpbh[l],
                    "djpmy": djpmy[l]
                }
            )
            logger.error(f"查询结果 b  : {b}")
            
            
            if b:  # 对应 Delphi 中的 if tmpcom.isempty=false
                # 查询产品详细信息（对应 Delphi 中的 tmpcom1 查询）
                d = run_sql(
                    """
                    select cpbh, djpmy, zhwbgpm, krcode, jldw
                    from cymxsheet
                    where (pid=:rid) and (cpbh=:cpbh) and (djpmy=:djpmy) and (chxs>0)
                    order by rid limit 1
                    """,
                    {
                        "rid": recordset.get('rids')[0],
                        "cpbh": cpbh[l],
                        "djpmy": djpmy[l]
                    }
                )
                logger.error(f"查询结果 d  : {d}")
                
                
                # 行号计算（对应 Delphi 中的 inttostr(3+k)）
                k += 1
                row_num = 3 + k
                
                # 插入新行并复制格式（对应 Delphi 中的 Insert + Copy + PasteSpecial）
                ws.insert_rows(row_num)
                # 复制第3行的样式
                for col in range(1, ws.max_column + 1):
                    source_cell = ws.cell(row=3, column=col)
                    target_cell = ws.cell(row=row_num, column=col)
                    if source_cell.has_style:
                        target_cell.font = source_cell.font.copy()
                        target_cell.border = source_cell.border.copy()
                        target_cell.fill = source_cell.fill.copy()
                        target_cell.number_format = source_cell.number_format
                        target_cell.protection = source_cell.protection.copy()
                        target_cell.alignment = source_cell.alignment.copy()
                
                # 写入数据到 Excel（对应 Delphi 中的 Range['B'+inttostr(3+k)].Value）
                if d:
                    ws[f'B{row_num}'] = d[0].get('cpbh', '')
                    ws[f'C{row_num}'] = d[0].get('djpmy', 0)
                    ws[f'D{row_num}'] = d[0].get('zhwbgpm', '')
                    ws[f'E{row_num}'] = d[0].get('krcode', '')
                
                ws[f'F{row_num}'] = b[0].get('chxsz1', 0)
                ws[f'G{row_num}'] = b[0].get('chslz1', 0)
                
                if d:
                    ws[f'H{row_num}'] = d[0].get('jldw', '')
                
                ws[f'I{row_num}'] = b[0].get('mjzjz1', 0)
                
                # 累加人民币总金额
                khje += float(b[0].get('mjzjz1', 0))
                
                
                usd_amount = round(float(b[0].get('mjzjz1', 0)) /hl, 2)
                ws[f'J{row_num}'] = usd_amount
                
                # 累加美金总金额
                mjje += usd_amount
        
        # ========== 第九步：写入汇总行（对应 Delphi 中的 Range['I'+inttostr(4+k)]）==========
        total_row = 3 + k + 1
        ws[f'I{total_row}'] = khje
        ws[f'J{total_row}'] = mjje
        
        # 删除模板中的示例行（对应 Delphi 中的 Rows[inttostr(3)].Delete）
        ws.delete_rows(3)
        
        # 调整行计数器（对应 Delphi 中的 k:=k+3）
        k += 3
        
        # ========== 第十步：填充单据抬头信息（对应 Delphi 中的 cymx 查询）==========
        fphm = ''
        
        c = run_sql(
            "select fphm, fprq, wftt, shr, mdka, kagj, ysfs, zdry, htjx from cymx where rid=:rid",
            {"rid": recordset.get('rids')[0]}
        )
        
        if c:
            fphm = c[0].get('htjx', '')
            
            # 写入抬头信息（对应 Delphi 中的 Range['C'+inttostr(3+k)]）
            ws[f'C{3 + k}'] = '宁波优胜国际贸易有限公司'
            ws[f'C{4 + k}'] = c[0].get('htjx', '')
            ws[f'C{5 + k}'] = c[0].get('fprq', '')
            
            # 查询并匹配公司抬头（对应 Delphi 中的 cyzglsheet 查询和 POS 匹配）
            f = run_sql("select xm, bz, bz1 from cyzglsheet where (zm='BP CO清单抬头')")
            
            if f:
                for row in f:
                    xm_value = row.get('xm', '')
                    fphm_prefix = c[0].get('fphm', '')[:8]
                    
                    # 对应 Delphi 中的 POS(xm, copy(fphm,1,8)) > 0
                    if xm_value in fphm_prefix:
                        ws[f'C{3 + k}'] = row.get('bz', '')
                        ws[f'C{6 + k}'] = row.get('bz1', '')
            
            # 写入运输信息
            ws[f'C{8 + k}'] = 'NINGBO，CHINA'
            ws[f'C{9 + k}'] = f"{c[0].get('mdka', '')}, {c[0].get('kagj', '')}"
            ws[f'C{10 + k}'] = c[0].get('ysfs', '')
            
            # 日期加一天（对应 Delphi 中的 strtodatetime + 1）
            fprq_str = c[0].get('fprq', '')
            if fprq_str:
                try:
                    fprq_date = datetime.strptime(fprq_str, '%Y-%m-%d')
                    next_day = fprq_date + timedelta(days=1)
                    ws[f'C{11 + k}'] = next_day.strftime('%Y-%m-%d')
                except:
                    ws[f'C{11 + k}'] = fprq_str
            else:
                ws[f'C{11 + k}'] = '/'
            
            ws[f'C{12 + k}'] = '/'
            ws[f'C{13 + k}'] = 'N/M'
            ws[f'C{14 + k}'] = f"FROM NINGBO，CHINA TO {c[0].get('mdka', '')}  {c[0].get('kagj', '')}  {c[0].get('ysfs', '')}"
            ws[f'C{15 + k}'] = c[0].get('zdry', '')
        if hasattr(user, 'username'):
            username = user.username
        elif hasattr(user, 'name'):
            username = user.name
        elif isinstance(user, str):
            username = user
        else:
            username = str(user) 
        # ========== 第十一步：保存文件 ==========
        path = config.tmp_path
        os.makedirs(path, exist_ok=True) # 确保目录存在
        
        output_filename = f'{lj}{fphm}{username}cymx.xlsx'
        output_filename = clean_filename(output_filename)
        full_output_path = os.path.join(path, output_filename)
        s_path = config.tmp_path
               
        
        wb.save(full_output_path)
        
        return {'code': 1, 'msg': '报表生成成功', 'data': output_filename}
    
    except Exception as e:
        logger.error(f"生成Excel失败: {str(e)}")
        logger.error(traceback.format_exc())
        return {'code': -1, 'msg': f'生成Excel失败: {str(e)}', 'data': ''}
    finally:
        s.close()


@any_route('/api/shipping/export/detailed/BP_OC/list', methods=['POST'])
@require_token
async def BP_OC_list(request):
    """
    导出BP OC清单详细数据
    """
    user = request.current_user
    j = await request.json()
    
    try:
        result = crate_excel_logic(j, user)
        logger.error(result)
        if result['code'] == 1:
            return json_result(1, result['msg'], result['data'])
        else:
            return json_result(result['code'], result['msg'], {})
    
    except Exception as e:
        logger.error(f"API调用失败: {str(e)}")
        logger.error(traceback.format_exc())
        return json_result(-1, f'API调用失败: {str(e)}', {})