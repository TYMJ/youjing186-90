# 批量总合同（rid/pid版）
from any import *
import os
import json
import subprocess
from collections import defaultdict
from datetime import datetime, date
import io

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

from openpyxl import load_workbook as _openpyxl_load_workbook


def load_workbook(*args, **kwargs):
    kwargs.pop("encoding", None)
    return _openpyxl_load_workbook(*args, **kwargs)


ILLEGAL_CHARS = set('/\\:*?"<>|\r\n')


def _safe_filename(text):
    s = str(text or "").strip()
    if not s:
        return "EMPTY"
    s = "".join("_" if ch in ILLEGAL_CHARS else ch for ch in s)
    return s[:120]


def _to_upload_rel_path(abs_path):
    abs_path = os.path.abspath(abs_path)
    base_upload = getattr(config, "data_upload_path", "") or ""
    if base_upload:
        base_abs = os.path.abspath(base_upload)
        if abs_path.startswith(base_abs):
            rel = abs_path[len(base_abs) :].lstrip("/\\")
            return rel.replace("\\", "/")
    return abs_path.replace("\\", "/")


def _abs_upload(path):
    base = getattr(config, "data_upload_path", "") or ""
    return os.path.join(base, str(path or "").lstrip("/\\"))


def _parse_json_list(v):
    if isinstance(v, list):
        return v
    if v is None:
        return []
    try:
        return json.loads(v) if isinstance(v, str) else []
    except Exception:
        return []


def _to_float(v):
    try:
        return float(v or 0)
    except Exception:
        return 0.0


def _convert_excel_to_pdf(excel_path):
    pdf_path = os.path.splitext(excel_path)[0] + ".pdf"

    if not getattr(config, "java_path", ""):
        return {"success": False, "error": "未配置 config.java_path"}
    if not getattr(config, "report_jar", ""):
        return {"success": False, "error": "未配置 config.report_jar"}
    if not os.path.exists(config.report_jar):
        return {"success": False, "error": f"JAR文件不存在: {config.report_jar}"}

    cmd = [config.java_path, "-jar", config.report_jar, "a", "b", excel_path, pdf_path, "2"]
    try:
        rs = subprocess.run(cmd, capture_output=True, text=True, timeout=180)
        if rs.returncode != 0:
            return {"success": False, "error": rs.stderr or f"returncode={rs.returncode}"}
        if not os.path.exists(pdf_path):
            return {"success": False, "error": "PDF未生成"}
        return {"success": True, "pdf_path": pdf_path}
    except Exception as e:
        return {"success": False, "error": str(e)}


def _first_row(sql, params):
    rows = run_sql(sql, params)
    return rows[0] if rows else {}


def _resolve_template_path(company, is_tedi):
    suffix = "tedi" if is_tedi else ""
    fname = f"{company}{suffix}.xlsx"
    data_upload = getattr(config, "data_upload_path", "") or ""
    base_dir = os.path.dirname(__file__)

    candidates = [
        os.path.join(data_upload, "template", fname),
        os.path.join(data_upload, "addonfiles", fname),
        os.path.join(base_dir, "..", "frontend", "static", "contract_templates", fname),
        os.path.join(base_dir, "..", "frontend", "static", fname),
        os.path.join(base_dir, fname),
    ]

    for path in candidates:
        p = os.path.abspath(path)
        if os.path.exists(p):
            return p
    return ""


def _get_sign_image_path(cpbh):
    rows = run_sql(
        """
        SELECT path
        FROM sys_attachment
        WHERE pid IN (
            SELECT rid FROM tpzx WHERE cpbh=:cpbh
        )
        AND path IS NOT NULL
        LIMIT 1
        """,
        {"cpbh": str(cpbh or "")},
    )
    if not rows:
        return ""
    return _abs_upload(rows[0].get("path"))


def _try_add_image(ws, image_abs_path, anchor_cell, max_width=None, max_height=None):
    try:
        if not image_abs_path or not os.path.exists(image_abs_path):
            return
        img = XLImage(image_abs_path)
        if max_width and max_height:
            src_w = max(float(img.width or 1), 1.0)
            src_h = max(float(img.height or 1), 1.0)
            scale = min(max_width / src_w, max_height / src_h)
            if scale > 0:
                img.width = int(src_w * scale)
                img.height = int(src_h * scale)

            # 居中：根据单元格可用空间计算偏移量
            from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
            from openpyxl.drawing.xdr import XDRPositiveSize2D

            m = re.match(r"([A-Z]+)(\d+)", str(anchor_cell))
            if m:
                col_str, row_num = m.group(1), int(m.group(2))
                col_idx = 0
                for ch in col_str:
                    col_idx = col_idx * 26 + (ord(ch) - 64)
                col_idx -= 1
                row_idx = row_num - 1
                PX_TO_EMU = 9525
                col_off = int(max(0, (max_width - img.width) / 2) * PX_TO_EMU)
                row_off = int(max(0, (max_height - img.height) / 2) * PX_TO_EMU)
                marker = AnchorMarker(col=col_idx, row=row_idx, colOff=col_off, rowOff=row_off)
                img.anchor = OneCellAnchor(
                    _from=marker, ext=XDRPositiveSize2D(cx=int(img.width * PX_TO_EMU), cy=int(img.height * PX_TO_EMU))
                )
                ws.add_image(img)
                return

        img.anchor = anchor_cell
        ws.add_image(img)
    except Exception:
        pass


def _resolve_yytp_to_path(yytp_raw):
    if not yytp_raw:
        return ""
    s = str(yytp_raw).strip()
    if s in ("[]", "{}", "null", "None"):
        return ""
    if s.startswith("[") or s.startswith("{"):
        try:
            obj = json.loads(s.replace("'", '"'))
            if isinstance(obj, list) and obj:
                src = obj[0].get("src", "")
                if src:
                    p = _abs_upload(src)
                    return p if os.path.exists(p) else ""
            elif isinstance(obj, dict):
                for k in ("src", "path", "url"):
                    v = obj.get(k)
                    if v:
                        p = _abs_upload(v)
                        return p if os.path.exists(p) else ""
        except Exception:
            pass
    p = _abs_upload(s)
    return p if os.path.exists(p) else ""


def _get_cell_display_pixels(ws, anchor_cell):
    import re
    from openpyxl.utils import get_column_letter

    m = re.match(r"([A-Z]+)(\d+)", str(anchor_cell or ""))
    if not m:
        return 120, 100
    col_letter, row_no = m.group(1), int(m.group(2))

    def _col_px(c):
        cd = ws.column_dimensions.get(get_column_letter(c))
        return int((cd.width if cd and cd.width else 8.43) * 7.6)

    def _row_px(r):
        rd = ws.row_dimensions.get(r)
        return int((rd.height if rd and rd.height else 15) * 1.333)

    for merged in ws.merged_cells.ranges:
        if str(anchor_cell) in merged:
            col_px = sum(_col_px(c) for c in range(merged.min_col, merged.max_col + 1))
            row_px = sum(_row_px(r) for r in range(merged.min_row, merged.max_row + 1))
            return max(20, col_px - 8), max(20, row_px - 8)

    cd = ws.column_dimensions.get(col_letter)
    rd = ws.row_dimensions.get(row_no)
    return (
        max(20, int((cd.width if cd and cd.width else 8.43) * 7.6) - 8),
        max(20, int((rd.height if rd and rd.height else 15) * 1.333) - 8),
    )


def _collect_cggd_rows(rid_list, user_name):
    rows = []
    for rid in rid_list:
        rs = run_sql(
            """
            SELECT rid, hthm, gdry, sccj, cgry, cs_id, wxht, ywry, dforder_id
            FROM cggd
            WHERE rid=:rid
            """,
            {"rid": str(rid)},
        )
        for r in rs:
            gdry = str(r.get("gdry") or "").strip()
            if user_name and gdry and gdry != user_name:
                continue
            rows.append(r)
    return rows


def _check_customer_flags(wxht_list, user_name):
    is_tedi = False
    missing_customer_cfg = False

    for order_id in wxht_list:
        if not order_id:
            continue

        wx = _first_row(
            """
            SELECT customer
            FROM wxht
            WHERE order_id=:order_id
            LIMIT 1
            """,
            {"order_id": str(order_id)},
        )
        customer = str(wx.get("customer") or "").strip()

        if customer == "TEDi GmbH & Co. KG":
            is_tedi = True

        cfg = _first_row(
            """
            SELECT rid
            FROM cyzglsheet
            WHERE xm=:xm AND zm='EXCEL采购客人'
            LIMIT 1
            """,
            {"xm": customer},
        )
        if not cfg:
            missing_customer_cfg = True

    has_batch_pdf = bool(
        _first_row(
            """
        SELECT rid
        FROM cyzglsheet
        WHERE xm=:xm AND zm='采购合同批量总PDF'
        LIMIT 1
        """,
            {"xm": user_name},
        )
    )

    has_sign_auth = bool(
        _first_row(
            """
        SELECT rid
        FROM cyzglsheet
        WHERE xm=:xm AND zm='采购合同图PDF签名'
        LIMIT 1
        """,
            {"xm": user_name},
        )
    )

    return {
        "is_tedi": is_tedi,
        "missing_customer_cfg": missing_customer_cfg,
        "has_batch_pdf": has_batch_pdf,
        "has_sign_auth": has_sign_auth,
    }


def _get_contract_lines(hthm_list, factory, buyer):
    lines = []
    for hthm in hthm_list:
        run_sql("UPDATE cght SET webpd1='是' WHERE hthm=:hthm", {"hthm": hthm})
        rs = run_sql(
            """
            SELECT rid, pid, hthm, bjhh, khhh, zwpm, wxrl,
                   cgxs, cgsl, cgjg, zje, yjcq, jhrq, sfhs,
                   wxht, cpgg, zhwbzh, zwdw, cpsm, hyd, zhwbgpm, zzsl,
                   sccj1, cgry
            FROM cghtsheet
            WHERE hthm=:hthm AND sccj1=:sccj1 AND cgry=:cgry
            ORDER BY jhrq, rid
            """,
            {"hthm": hthm, "sccj1": factory, "cgry": buyer},
        )
        lines.extend(rs)
    return lines


def _get_dforder(order_id):
    row = _first_row(
        """
        SELECT dforder_id
        FROM wxht
        WHERE order_id=:order_id
        LIMIT 1
        """,
        {"order_id": str(order_id or "")},
    )
    return str(row.get("dforder_id") or "")


def _calc_totals(lines):
    zxs = int(sum(_to_float(r.get("cgxs")) for r in lines))  # 箱数取整
    zsl = sum(_to_float(r.get("cgsl")) for r in lines)
    zjg = sum(_to_float(r.get("zje")) for r in lines)
    sfhs1 = "是"
    for r in lines:
        if str(r.get("sfhs") or "是") != "是":
            sfhs1 = "否"
            break
    return zxs, zsl, zjg, sfhs1


def _build_context(first_line, buyer, company):
    hthm = str(first_line.get("hthm") or "")

    cght_row = _first_row(
        """
        SELECT szdq, gdry, cgry, ywry, sccj1id, lxry, gcdh, sjhm, gccz, htrq, jsfs
        FROM cght
        WHERE hthm=:hthm AND cgry=:cgry
        LIMIT 1
        """,
        {"hthm": hthm, "cgry": buyer},
    )

    szdq = str(cght_row.get("szdq") or "")
    cs_id = str(cght_row.get("sccj1id") or "")

    htjj = _first_row(
        """
        SELECT dz, lxfs, cgzj, zjl, hcyx, lxdh, gstt
        FROM htjj
        WHERE dq=:dq AND gstt LIKE :gstt
        LIMIT 1
        """,
        {"dq": szdq, "gstt": f"%{company}%"},
    )

    supplier = _first_row(
        """
        SELECT cslxr, phone, fax, sjhm, address, twhm, fkhm, bank1, zh1
        FROM zycs
        WHERE cs_id=:cs_id
        LIMIT 1
        """,
        {"cs_id": cs_id},
    )

    def user_profile(yhm):
        if not yhm:
            return {}
        return _first_row(
            """
            SELECT ryxm, lxdh, ydhm, bmjl, jldh
            FROM ywrybiao
            WHERE yhm=:yhm
            LIMIT 1
            """,
            {"yhm": yhm},
        )

    gdry_yhm = str(cght_row.get("gdry") or "")
    cgry_yhm = str(cght_row.get("cgry") or "")
    ywry_yhm = str(cght_row.get("ywry") or "")

    return {
        "htrq": str(cght_row.get("htrq") or ""),
        "szdq": szdq,
        "jsfs": str(cght_row.get("jsfs") or ""),
        "gdry_yhm": gdry_yhm,
        "cgry_yhm": cgry_yhm,
        "ywry_yhm": ywry_yhm,
        "htjj": htjj,
        "supplier": supplier,
        "gdry": user_profile(gdry_yhm),
        "cgry": user_profile(cgry_yhm),
        "ywry": user_profile(ywry_yhm),
    }


def _fill_detail_sheet(ws, lines, company, cpbh_show, sign_path, hthm_khht_list):
    zxs, zsl, zjg, _ = _calc_totals(lines)
    is_jingchi = "景驰" in (company or "")

    for idx, row in enumerate(lines, start=1):
        excel_row = 2 + idx
        ws[f"A{excel_row}"] = idx
        ws[f"B{excel_row}"] = row.get("hthm", "")
        ws[f"C{excel_row}"] = hthm_khht_list.get(row.get("hthm", ""), "")
        ws[f"D{excel_row}"] = row.get("khhh") or row.get("bjhh")
        ws[f"E{excel_row}"] = row.get("zwpm", "")
        ws[f"F{excel_row}"] = row.get("wxrl", "")
        ws[f"G{excel_row}"] = row.get("cgxs", "")
        ws[f"H{excel_row}"] = row.get("cgsl", "")
        ws[f"I{excel_row}"] = row.get("cgjg", "")
        ws[f"J{excel_row}"] = row.get("zje", "")
        # ws[f'K{excel_row}'] = str(row.get('jhrq', ''))
        ws[f"K{excel_row}"] = str(row.get("jhrq", "")).split()[0]  # 交货日期只显示年月日
        ws[f"L{excel_row}"] = row.get("yjcq", "")

        if is_jingchi:
            ws[f"M{excel_row}"] = _get_dforder(row.get("wxht"))

    head_cell = "N1" if is_jingchi else "A1"
    ws[head_cell] = f"合同编号：{cpbh_show}项下附页合同具体内容如下："
    ws["A1"] = f"合同编号：{cpbh_show}项下附页合同具体内容如下："

    sum_row = 3 + len(lines)
    ws[f"B{sum_row}"] = "合计"
    ws[f"G{sum_row}"] = zxs
    ws[f"H{sum_row}"] = zsl
    ws[f"J{sum_row}"] = zjg

    ws[f"B{sum_row + 3}"] = "需 方（签字盖章）"
    ws[f"F{sum_row + 3}"] = "供 方（签字盖章）"

    if sign_path:
        _try_add_image(ws, sign_path, f"C{sum_row + 2}")


def _fill_main_sheet(ws, first_line, totals, line_count, ctx, sign_path, anti_corruption_text):
    zxs, zsl, zjg, sfhs1 = totals
    htjj = ctx.get("htjj", {})
    supplier = ctx.get("supplier", {})

    gd = ctx.get("gdry", {})
    cg = ctx.get("cgry", {})
    yw = ctx.get("ywry", {})

    ws["D10"] = ctx.get("htrq", "")
    ws["I10"] = ctx.get("szdq", "")
    ws["N10"] = first_line.get("khhh") or first_line.get("bjhh")
    ws["U10"] = first_line.get("khhh") or first_line.get("bjhh")

    ws["C11"] = first_line.get("sccj1", "")
    ws["L11"] = supplier.get("cslxr", "")
    ws["C12"] = supplier.get("phone", "")
    ws["L12"] = supplier.get("fax", "")

    ws["B15"] = first_line.get("khhh") or first_line.get("bjhh")
    ws["D15"] = f"{first_line.get('zwpm', '')}\n{first_line.get('cpgg', '')}"
    ws["J15"] = first_line.get("wxrl", "")
    ws["K15"] = first_line.get("zwdw", "")
    ws["L15"] = zxs
    ws["M15"] = zsl
    ws["O15"] = first_line.get("cgjg", "")
    ws["P15"] = zjg

    ws["F16"] = first_line.get("cpsm", "")
    ws["N16"] = f"请注意开票货源地为：{first_line.get('hyd', '')}\n本合同为附页子合同的总合同，共计{line_count}个子合同"

    # B16插入产品图片：先查专属产品表，没有再查专业产品表
    item_code = first_line.get("khhh") or first_line.get("bjhh")
    if item_code:
        raw_img = ""
        row_img = _first_row(
            "SELECT yytp FROM zscp WHERE cpbh=:cpbh AND IFNULL(yytp,'')<>'' LIMIT 1", {"cpbh": item_code}
        )
        if row_img:
            raw_img = row_img.get("yytp") or ""
        if not raw_img:
            row_img = _first_row(
                "SELECT yytp FROM cjcp WHERE cpbh=:cpbh AND IFNULL(yytp,'')<>'' LIMIT 1", {"cpbh": item_code}
            )
            if row_img:
                raw_img = row_img.get("yytp") or ""
        img_path = _resolve_yytp_to_path(raw_img)
        if img_path:
            max_w, max_h = _get_cell_display_pixels(ws, "B16")
            _try_add_image(ws, img_path, "B16", max_width=max_w, max_height=max_h)

    ws["B30"] = "七、结算方式：" + str(ctx.get("jsfs") or "")

    if htjj.get("dz"):
        ws["B7"] = htjj.get("dz")
    if htjj.get("lxfs"):
        ws["B8"] = htjj.get("lxfs")

    if sfhs1 != "是":
        ws["B19"] = (
            f"收款户名:{supplier.get('fkhm', '')}  开户银行:{supplier.get('bank1', '')}  银行账号:{supplier.get('zh1', '')}"
        )

    ws["F17"] = f"金额合计:{zjg:.2f}"
    ws["P17"] = zjg

    if first_line.get("zhwbgpm") and str(first_line.get("zhwbgpm")) != "无":
        ws["O18"] = f"开票品名: {first_line.get('zhwbgpm', '')}{first_line.get('zzsl', '')}%"

    ws["D25"] = f"{zxs}箱"

    ws["B48"] = f"跟单人员: {gd.get('ryxm', '')}({gd.get('lxdh', '')}/{gd.get('ydhm', '')})"
    ws["J48"] = f"采购员: {cg.get('ryxm', '')}({cg.get('lxdh', '')}/{cg.get('ydhm', '')})"
    ws["B49"] = f"外销负责: {yw.get('ryxm', '')}({yw.get('lxdh', '')}/{yw.get('ydhm', '')})"
    ws["J49"] = htjj.get("cgzj", "")
    ws["B50"] = htjj.get("zjl", "")
    ws["J50"] = anti_corruption_text

    if sign_path:
        _try_add_image(ws, sign_path, "B47")


def _check_credit(factory):
    warns = []
    row = _first_row(
        """
        SELECT qrrq
        FROM cxgc
        WHERE (gcmc=:gcmc) OR (chgc=:chgc)
        LIMIT 1
        """,
        {"gcmc": factory, "chgc": factory},
    )

    if not row:
        warns.append(f"工厂名称:{factory}需提交诚信报告")
        return warns

    qrrq = str(row.get("qrrq") or "").strip()
    if not qrrq:
        warns.append(f"工厂名称:{factory}需提交诚信报告")
        return warns

    try:
        qd = datetime.strptime(qrrq[:10], "%Y-%m-%d").date()
    except Exception:
        warns.append(f"工厂名称:{factory}需提交诚信报告")
        return warns

    days = (date.today() - qd).days
    if days > 365:
        warns.append(f"工厂名称:{factory}需提交诚信报告")
    elif days > 330:
        warns.append(f"工厂名称:{factory}还有一个月需提交诚信报告")

    return warns


@any_route("/api/Ravencloud/generate_contract_batch", methods=["POST"])
async def api_generate_contract_batch(request):
    try:
        form = await request.form()

        company = (form.get("company") or "").strip() or "宁波优景进出口有限公司"
        pdf_mode = str(form.get("pdf") or "1").strip()
        rid_list = _parse_json_list(form.get("rid_list") or "[]")

        if pdf_mode not in ("1", "2", "3"):
            pdf_mode = "1"

        if not rid_list:
            return json_result(code=-1, msg="请先选择要操作的记录")

        current_user = ""
        try:
            current_user = str(request.current_user.name or "").strip()
        except Exception:
            pass

        cggd_rows = _collect_cggd_rows(rid_list, current_user)
        if not cggd_rows:
            return json_result(code=-1, msg="未找到可处理数据，或当前用户无权限")

        hthm_list = []
        wxht_list = []
        factories = []
        buyers = []
        hthm_khht_list = {}  # 客户合同

        for r in cggd_rows:
            hthm = str(r.get("hthm") or "").strip()
            wxht = str(r.get("wxht") or "").strip()
            factory = str(r.get("sccj") or "").strip()
            buyer = str(r.get("cgry") or "").strip()
            dforder_id = str(r.get("dforder_id") or "").strip()

            if hthm and hthm not in hthm_list:
                hthm_list.append(hthm)
                hthm_khht_list[hthm] = dforder_id
            if wxht and wxht not in wxht_list:
                wxht_list.append(wxht)
            if factory and factory not in factories:
                factories.append(factory)
            if buyer and buyer not in buyers:
                buyers.append(buyer)

        flags = _check_customer_flags(wxht_list, current_user)
        if flags.get("has_batch_pdf") and pdf_mode == "1":
            pdf_mode = "3"

        template_path = _resolve_template_path(company, flags.get("is_tedi", False))
        if not template_path:
            return json_result(code=-1, msg=f"未找到合同模板，请上传: {company}.xlsx（或{company}tedi.xlsx）")

        anti_corruption_text = str(
            _first_row("SELECT wb1 FROM zx WHERE ly=:ly LIMIT 1", {"ly": "反腐专线"}).get("wb1") or ""
        )

        sign_path = ""
        if flags.get("has_sign_auth") or flags.get("has_batch_pdf"):
            sign_path = _get_sign_image_path("陈妍科小")

        output_root = os.path.join(config.get_today_upload_path(), "总合同")
        os.makedirs(output_root, exist_ok=True)

        files = []
        warnings = []
        date_tag = datetime.now().strftime("%Y%m%d")

        for factory in factories:
            factory_dir = os.path.join(output_root, _safe_filename(factory))
            os.makedirs(factory_dir, exist_ok=True)

            warnings.extend(_check_credit(factory))

            for buyer in buyers:
                all_lines = _get_contract_lines(hthm_list, factory, buyer)
                if not all_lines:
                    continue

                grouped = defaultdict(list)
                for row in all_lines:
                    key = str(row.get("bjhh") or "").strip() or "UNKNOWN"
                    grouped[key].append(row)

                for bjhh, lines in grouped.items():
                    first_line = lines[0]
                    cpbh_show = str(first_line.get("khhh") or first_line.get("bjhh") or bjhh)

                    wb = load_workbook(template_path)
                    ws_main = wb.worksheets[0]
                    ws_detail = wb.worksheets[1] if len(wb.worksheets) > 1 else wb.worksheets[0]

                    totals = _calc_totals(lines)
                    _fill_detail_sheet(ws_detail, lines, company, cpbh_show, sign_path, hthm_khht_list)

                    ctx = _build_context(first_line, buyer, company)
                    _fill_main_sheet(ws_main, first_line, totals, len(lines), ctx, sign_path, anti_corruption_text)

                    ywry_yhm = str(ctx.get("ywry_yhm") or "")
                    base_name = f"{_safe_filename(ywry_yhm)}-{_safe_filename(factory)}-{_safe_filename(cpbh_show[:100])}-{date_tag}"
                    xlsx_path = os.path.join(factory_dir, base_name + ".xlsx")

                    wb.save(xlsx_path)
                    wb.close()

                    if pdf_mode == "1":
                        files.append(_to_upload_rel_path(xlsx_path))
                    else:
                        pdf_res = _convert_excel_to_pdf(xlsx_path)
                        if pdf_res.get("success"):
                            files.append(_to_upload_rel_path(pdf_res.get("pdf_path")))
                        try:
                            os.remove(xlsx_path)
                        except Exception:
                            pass

        warning_file = ""
        if warnings:
            warning_name = datetime.now().strftime("%Y-%m-%d") + ".txt"
            warning_abs = os.path.join(output_root, warning_name)
            with io.open(warning_abs, "w", encoding="utf-8") as fp:
                fp.write("\n".join(dict.fromkeys(warnings)))
            warning_file = _to_upload_rel_path(warning_abs)

        return json_result(
            code=1,
            msg=f"成功生成{len(files)}个文件",
            data={
                "files": files,
                "count": len(files),
                "warnings": list(dict.fromkeys(warnings)),
                "warning_file": warning_file,
                "pdf_mode": pdf_mode,
            },
        )
    except Exception:
        logger.error(trace_error())
        return json_result(code=-1, msg=trace_error())
