import asyncio
from concurrent.futures import ThreadPoolExecutor
from any import *  # 假设此模块包含了 Session, config, logger, get_uuid, require_token, any_route, json_result
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import os
import traceback

# 用于异步执行耗时任务的线程池
executor = ThreadPoolExecutor(max_workers=1)

def _generate_detailed_contract_list_on_thread(selected_contract_rids, user_name):
    """
    在后台线程中执行的核心逻辑，接受一个选定的合同号列表。
    """
    s = Session()
    try:
        # --- 1. 权限检查 (修正版，仅检查 cyzglsheet 表) ---
        permission_check_query = """
            SELECT 1 FROM cyzglsheet WHERE xm = :username AND zm = '跟单统计表' LIMIT 1;
        """
        permission_result = s.execute(permission_check_query, {'username': user_name}).fetchone()
        
        if not permission_result:
            return {'code': -1, 'msg': '权限不足，无法导出。', 'data': None}

        # --- 2. 输入验证 ---
        if not selected_contract_rids:
            return {'code': -1, 'msg': '未选择任何合同进行导出', 'data': None}

        # 将字符串列表转换为适合SQL IN子句的格式
        placeholders = ','.join([f':num_{i}' for i in range(len(selected_contract_rids))])
        params = {f'num_{i}': num for i, num in enumerate(selected_contract_rids)}

        # --- 3. 生成 Excel 报告 ---
        wb = Workbook()
        ws = wb.active
        ws.title = "合同明细清单"

        # 设置标题行和列宽 (对应Delphi代码中的Excel列设置)
        headers = [
            'Order NO.', 'Costomer NO.', 'Item NO.', 'item Name', '款式', '颜色', 'Pcs/Ctn', 'UNIT', 
            'Inner CTNS in 1 CTNt', 'кол-во кор (CTNS)', 'TOTAL QTY', 'USD', 'AMOUNT', 'ETD', 'G.W.', 
            'N,W.', 'Volume', '工厂', 'RMB', 'Amount(RMB)', '下单地点', '跟单人员', '工厂回签日期', 
            '工厂交期', '产前确认', 'AW确认', 'AW', 'SM确认', 'SM', 'PR-M SAMPLE', 'INSPECTION PHOTS', 
            'LOADING TIME', 'IN ORDER', 'QTY', '备注', '采购业务员', '业务员', '船期'
        ]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14.3

        current_row = 2
        
        # 查询并处理指定的合同号
        # 注意: 根据您的澄清，cght表的主键是rid，cghtsheet通过pid字段关联到cght.rid
        query_main_contracts = f"""
        SELECT wxht, hthm, szdq, gdry, ywry, rid FROM cght WHERE rid IN ({placeholders})
        """
        main_contracts = s.execute(query_main_contracts, params).fetchall()

        for main_contract in main_contracts:
            # 2. 获取业务员信息 (wxht)
            salesperson = ''
            if main_contract.wxht:
                salesperson_query = "SELECT ywry FROM wxht WHERE order_id = :order_id"
                salesperson_result = s.execute(salesperson_query, {'order_id': main_contract.wxht}).fetchone()
                if salesperson_result:
                    salesperson = salesperson_result.ywry or ''

            # 3. 获取合同明细 (cghtsheet)
            # 关键修正: WHERE pid = :contract_rid, 这里的 :contract_rid 就是 cght.rid
            details_query = """
            SELECT mjdj1, khhh, cpbh, bjhh, ywpm, ks, yse, wxrl, jldw, nhwx, cgxs, cgsl, Twxdj, zje, 
                   yjcq, mz, jz, wxwyzd, bz3, tj, csmc, sccj1, cgjg, jhrq, cqqr, AWqr, SMqr, rid, pid
            FROM cghtsheet WHERE pid = :contract_rid
            """
            details = s.execute(details_query, {'contract_rid': main_contract.rid}).fetchall()

            for detail in details:
                # 4. 获取采购跟单信息 (cggd)
                factory_sign_date = ''
                if main_contract.hthm:
                    factory_sign_query = "SELECT gdrq FROM cggd WHERE hthm = :hthm"
                    factory_sign_result = s.execute(factory_sign_query, {'hthm': main_contract.hthm}).fetchone()
                    if factory_sign_result:
                        factory_sign_date = factory_sign_result.gdrq or ''

                # 5. 获取采购明细信息 (cymxsheet) 用于填充 AH (QTY) 和 Q (Volume)
                chxs_combined = '' # 对应 AH 列
                tj_combined = ''   # 对应 Q 列
                cymx_query = """
                SELECT chxs, wxtj FROM cymxsheet WHERE wxwyzd = :wxwyzd AND chsl1 > 0 AND cght = :cght
                """
                cymx_results = s.execute(cymx_query, {
                    'wxwyzd': detail.wxwyzd or '',
                    'cght': main_contract.hthm or ''
                }).fetchall()
                
                for idx, cymx_item in enumerate(cymx_results):
                    if idx == 0:
                        chxs_combined = str(cymx_item.chxs)
                        if cymx_item.wxtj > 0:
                            tj_combined = str(cymx_item.wxtj)
                    else:
                        chxs_combined += '/' + str(cymx_item.chxs)
                        if cymx_item.wxtj > 0:
                            tj_combined += '/' + str(cymx_item.wxtj)

                # 6. 获取发票信息 (cymx) 用于填充 AF (LOADING TIME), AG (IN ORDER)
                # 关键修正: cymx表通过rid字段关联到cghtsheet.rid
                loading_time = '' # 对应 AF 列
                in_order = ''     # 对应 AG 列
                if detail.rid: # 使用 detail.rid 作为 pid
                    invoice_query = "SELECT zgrq, fphm FROM cymx WHERE rid = :rid"
                    invoice_result = s.execute(invoice_query, {'rid': detail.rid}).fetchone() # 关联字段是 rid (即 pid)
                    if invoice_result:
                        loading_time = invoice_result.zgrq or ''
                        in_order = invoice_result.fphm or ''

                # 7. 填充 Excel 行
                # A: Order NO.
                ws.cell(row=current_row, column=1, value=main_contract.wxht or '').alignment = Alignment(horizontal='center', vertical='center')
                # ... (其余列的填充逻辑保持不变) ...
                # B: Costomer NO.
                ws.cell(row=current_row, column=2, value=detail.khhh or '').alignment = Alignment(horizontal='center', vertical='center')
                # C: Item NO. (优先使用 bjhh)
                item_no = detail.bjhh or detail.cpbh
                ws.cell(row=current_row, column=3, value=item_no or '').alignment = Alignment(horizontal='center', vertical='center')
                # D: item Name
                ws.cell(row=current_row, column=4, value=detail.ywpm or '').alignment = Alignment(horizontal='center', vertical='center')
                # E: 款式
                ws.cell(row=current_row, column=5, value=detail.ks or '').alignment = Alignment(horizontal='center', vertical='center')
                # F: 颜色
                ws.cell(row=current_row, column=6, value=detail.yse or '').alignment = Alignment(horizontal='center', vertical='center')
                # G: Pcs/Ctn
                ws.cell(row=current_row, column=7, value=float(detail.wxrl or 0)).alignment = Alignment(horizontal='center', vertical='center')
                # H: UNIT
                ws.cell(row=current_row, column=8, value=detail.jldw or '').alignment = Alignment(horizontal='center', vertical='center')
                # I: Inner CTNS in 1 CTNt
                ws.cell(row=current_row, column=9, value=int(detail.nhwx or 0)).alignment = Alignment(horizontal='center', vertical='center')
                # J: кол-во кор (CTNS)
                ws.cell(row=current_row, column=10, value=float(detail.cgxs or 0)).alignment = Alignment(horizontal='center', vertical='center')
                # K: TOTAL QTY
                ws.cell(row=current_row, column=11, value=float(detail.cgsl or 0)).alignment = Alignment(horizontal='center', vertical='center')
                # L: USD (价格)
                unit_price = detail.mjdj1 if detail.mjdj1 > 0 else detail.Twxdj
                ws.cell(row=current_row, column=12, value=float(unit_price or 0)).alignment = Alignment(horizontal='center', vertical='center')
                # M: AMOUNT (金额)
                amount = (unit_price or 0) * (detail.cgsl or 0)
                ws.cell(row=current_row, column=13, value=float(amount)).alignment = Alignment(horizontal='center', vertical='center')
                # N: ETD
                ws.cell(row=current_row, column=14, value=detail.yjcq or '').alignment = Alignment(horizontal='center', vertical='center')
                # O: G.W.
                ws.cell(row=current_row, column=15, value=float(detail.mz or 0)).alignment = Alignment(horizontal='center', vertical='center')
                # P: N,W.
                ws.cell(row=current_row, column=16, value=float(detail.jz or 0)).alignment = Alignment(horizontal='center', vertical='center')
                # Q: Volume (来自cymxsheet的组合)
                ws.cell(row=current_row, column=17, value=tj_combined).alignment = Alignment(horizontal='center', vertical='center')
                # R: 工厂 (优先使用 sccj1)
                factory_name = detail.sccj1 or detail.csmc
                ws.cell(row=current_row, column=18, value=factory_name or '').alignment = Alignment(horizontal='center', vertical='center')
                # S: RMB
                ws.cell(row=current_row, column=19, value=float(detail.cgjg or 0)).alignment = Alignment(horizontal='center', vertical='center')
                # T: Amount(RMB)
                ws.cell(row=current_row, column=20, value=float(detail.zje or 0)).alignment = Alignment(horizontal='center', vertical='center')
                # U: 下单地点
                ws.cell(row=current_row, column=21, value=main_contract.szdq or '').alignment = Alignment(horizontal='center', vertical='center')
                # V: 跟单人员
                ws.cell(row=current_row, column=22, value=main_contract.gdry or '').alignment = Alignment(horizontal='center', vertical='center')
                # W: 工厂回签日期
                ws.cell(row=current_row, column=23, value=factory_sign_date).alignment = Alignment(horizontal='center', vertical='center')
                # X: 工厂交期
                ws.cell(row=current_row, column=24, value=detail.jhrq or '').alignment = Alignment(horizontal='center', vertical='center')
                # Y: 产前确认
                ws.cell(row=current_row, column=25, value=detail.cqqr or '').alignment = Alignment(horizontal='center', vertical='center')
                # Z: AW确认
                ws.cell(row=current_row, column=26, value=detail.AWqr or '').alignment = Alignment(horizontal='center', vertical='center')
                # AA: AW (图片占位，这里留空)
                ws.cell(row=current_row, column=27, value='').alignment = Alignment(horizontal='center', vertical='center')
                # AB: SM确认
                ws.cell(row=current_row, column=28, value=detail.SMqr or '').alignment = Alignment(horizontal='center', vertical='center')
                # AC: SM (图片占位，这里留空)
                ws.cell(row=current_row, column=29, value='').alignment = Alignment(horizontal='center', vertical='center')
                # AD: PR-M SAMPLE (Delphi代码中未填充，这里留空)
                ws.cell(row=current_row, column=30, value='').alignment = Alignment(horizontal='center', vertical='center')
                # AE: INSPECTION PHOTS (Delphi代码中未填充，这里留空)
                ws.cell(row=current_row, column=31, value='').alignment = Alignment(horizontal='center', vertical='center')
                # AF: LOADING TIME (来自cymx的zgrq)
                ws.cell(row=current_row, column=32, value=loading_time).alignment = Alignment(horizontal='center', vertical='center')
                # AG: IN ORDER (来自cymx的fphm)
                ws.cell(row=current_row, column=33, value=in_order).alignment = Alignment(horizontal='center', vertical='center')
                # AH: QTY (来自cymxsheet的组合)
                ws.cell(row=current_row, column=34, value=chxs_combined).alignment = Alignment(horizontal='center', vertical='center')
                # AI: 备注
                ws.cell(row=current_row, column=35, value=detail.bz3 or '').alignment = Alignment(horizontal='center', vertical='center')
                # AJ: 采购业务员 (固定为当前登录用户)
                ws.cell(row=current_row, column=36, value=user_name).alignment = Alignment(horizontal='center', vertical='center')
                # AK: 业务员
                ws.cell(row=current_row, column=37, value=salesperson).alignment = Alignment(horizontal='center', vertical='center')
                # AL: 船期
                ws.cell(row=current_row, column=38, value=detail.yjcq or '').alignment = Alignment(horizontal='center', vertical='center')

                current_row += 1


        # --- 8. 保存文件并返回路径 ---
        path = config.tmp_path
        os.makedirs(path, exist_ok=True)
        report_rid = get_uuid()
        output_filename = f'{report_rid}_合同明细清单.xlsx'
        full_output_path = os.path.join(path, output_filename)
        
        wb.save(full_output_path)

        return {'code': 1, 'msg': '报表导出成功', 'data': output_filename}

    except Exception as e:
        logger.error(traceback.format_exc())
        return {'code': -1, 'msg': f'导出报表时发生错误: {str(e)}', 'data': None}
    finally:
        s.close()

# --- FastAPI 路由处理器 ---
@any_route('/api/saier/export/detailed/contract/list', methods=['POST'])
@require_token
async def api_saier_export_detailed_contract_list(request):
    """
    API 接口，接收前端传来的选中合同号列表，并触发后台的 Excel 生成任务。
    """
    try:
        # 从请求体中获取 JSON 数据
        body_data = await request.json()
        selected_rids = body_data.get('rids', []) # 假设前端传递的是一个名为 'rids' 的数组
        user_name = request.state.user.get('name') # 请根据实际情况修改

        # if not user_name:
        #     raise AttributeError("无法从请求上下文获取当前用户名。")

        # 将耗时的数据库和文件操作放到后台线程执行
        loop = asyncio.get_event_loop()
        res = await loop.run_in_executor(
            executor,
            _generate_detailed_contract_list_on_thread,
            selected_rids, # 将选中的合同号列表传递给后台函数
            user_name
        )
        return json_result(res.get('code'), res.get('msg'), res.get('data'))
    
    except AttributeError as ae:
        error_msg = f"获取用户信息失败: {str(ae)}"
        logger.error(error_msg)
        return json_result(-1, error_msg)
    except Exception as e:
        logger.error(traceback.format_exc())
        return json_result(-1, traceback.format_exc())