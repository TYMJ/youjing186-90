# from pdb import run
from any import *
from .model import *
from sqlalchemy.sql import func,not_,or_,and_
import time
from datetime import datetime, timedelta
import time, json, os, math, csv, tempfile, shutil, zipfile, re, subprocess, sys
from collections import defaultdict
from .__default__ import get_user_path,module_xxck_new,module_share_new,user_task_delete,user_task_new,get_user_path
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle, PatternFill, Color
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import load_workbook
    from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle, PatternFill, Color
    

SYS_FIELDS = ['sid','rid','uid','ctime','mtime','has_att','modi_uid','wf_status','wf_unit','pid','archived']

@any_route('/api/saier/declaration/load/check', methods=['POST', 'GET'])
@require_token
async def view_saier_declaration_load_check(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        fphm = j.get('fphm','')
        zdry = j.get('zdry','')

        sb1 = ''
        sb = ''
        ry1 = ''
        bm1 = ''
        dzdata = 0
        fphmdata = 0
        dzzgdata = 0
        dzcydata = 0
        if fphm != '':
            d = run_sql(f"select ry from bgmxd where (fphm='{str(fphm)}') and (zxpd='是')")
            if len(d)>0:
                if d[0].get('ry','') != user.username and d[0].get('ry','') != '':
                    sb1 = '1'
                    ry1 = d[0].get('ry','')
        
        if sb1 == '':
            org = get_user_path('zjnblh')
            bm1 = org.get('position', '')[:1]

            d = run_sql(f"select rid from sys_user where username='{str(user.username)}'")
            if len(d)>0:
                ds = run_sql(f"select role_id from sys_user_role where user_id = '{str(d[0].get('rid',''))}'")
                if len(ds)>0:
                    dt = run_sql(f"select rid from sys_role where rid = '{str(ds[0].get('role_id',''))}' and name like '%单证%'")
                    if len(dt)>0:
                        dzdata = 1
                        if not (zdry != '' and zdry == user.username):
                            if zdry == '':
                                zdry = user.username
                    else:
                        sb = '1'
                        dzdata = 2
                        
            if sb == '1':
                org = get_user_path(user.username)
                postion = org.get('position', '')
                if '财务' in postion:
                    if bm1 != 'C':
                        dzdata = 3
        
        if fphm != '' and user.username != '谢晓霞':
            d = run_sql(f"select fphm from kaiptz where (fphm='{str(fphm)}')")
            if len(d)>0:
                fphmdata = 1
        
        d = run_sql(f"select xm from cyzglsheet where xm='{str(user.username)}' and zm='单证主管复核' and bz='{str(zdry)}'")
        if len(d)>0:
            dzzgdata = 1
        else:
            dzzgdata = 2

        d = run_sql(f"select * from cyzglsheet where xm='{str(user.username)}' and zm='单证查验'")
        if len(d)>0:
            dzcydata = 1

        data = {'dzdata':dzdata,'fphmdata': fphmdata, 'dzzgdata': dzzgdata, 'dzcydata': dzcydata}

        return json_result(1, '查询成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        
def _f(v, d=0.0):
    try:
        if v in (None, ''):
            return d
        return float(v)
    except:
        return d

def _i(v, d=0):
    try:
        if v in (None, ''):
            return d
        return int(float(v))
    except:
        return d

def _r2(v):
    return round(_f(v) * 100) / 100

def _r3(v):
    return round(_f(v) * 1000) / 1000

def _dt(v):
    if not v:
        return None
    s = str(v).strip()
    for f in ('%Y-%m-%d', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S'):
        try:
            return datetime.strptime(s, f)
        except:
            pass
    return None

def _set_if_has(obj, data: dict):
    for k, v in data.items():
        if hasattr(obj, k):
            setattr(obj, k, v)

@any_route('/api/saier/declaration/save/before', methods=['POST', 'GET'])
@require_token
async def view_saier_declaration_save_before(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        # ===== 主表入参 =====
        rid = _s(j.get('rid', ''))                 # self.getnumber => recordset.val('rid')
        fphm = _s(j.get('fphm', ''))
        hsfp = _s(j.get('hsfp', ''))
        cyrq = _s(j.get('cyrq', ''))
        sfxb = _s(j.get('sfxb', ''))
        ssgs = _s(j.get('ssgs', ''))
        bgsb = _s(j.get('bgsb', ''))
        rmbkh = _s(j.get('rmbkh', ''))
        hbdm = _s(j.get('hbdm', ''))
        sz = _f(j.get('sz', 0))
        cnfyf = _f(j.get('cnfyf', 0))
        bgtjhj = _f(j.get('bgtjhj', 0))
        bggs = _s(j.get('bggs', ''))
        sjcy = j.get('sjcy')
        htdate = _s(j.get('htdate', ''))
        zgrq = _s(j.get('zgrq', ''))
        hl = _f(j.get('hl', 0))
        zmjhl = _f(j.get('zmjhl', 1), 1)
        jgtk = _s(j.get('jgtk', ''))
        bgje = _f(j.get('bgje', 0))
        khmc = _s(j.get('khmc', ''))
        zdry = _s(j.get('zdry', ''))
        bgdh = _s(j.get('bgdh', ''))
        tdh = _s(j.get('tdh', ''))
        fttdh = _s(j.get('fttdh', ''))
        xh = _s(j.get('xh', ''))
        fh = _s(j.get('fh', ''))
        sfcy = _s(j.get('sfcy', ''))
        mygb = _s(j.get('mygb', ''))

        # 电商费用相关（AMZ）
        ds_hl = _f(j.get('ds_hl', 1), 1)
        ds_hyf_usd = _f(j.get('ds_hyf_usd', 0))
        ds_hyf_rmb = _f(j.get('ds_hyf_rmb', 0))

        # ===== 子表入参 =====
        hbcplist = j.get('hbcplist', [])      # 合并产品
        cpzllist = j.get('cpzllist', [])      # 产品资料
        bgdlist = j.get('bgdlist', [])        # 报关单
        bbqdlist = j.get('bbqdlist', [])      # 补报清单
        ytxxlist = j.get('ytxxlist', [])      # 预填信息

        patch = {}
        errors = []
        warnings = []

        # ===== 主表回填 1：核算发票 =====
        if fphm != '' and hsfp == '':
            hsfp = fphm
        patch['hsfp'] = hsfp

        # ===== 主表回填 2：合同日期 -> kaiptz.htrq(-3天) =====
        if htdate != '' and fphm != '':
            d = _dt(htdate)
            if d:
                htrq = (d - timedelta(days=3)).strftime('%Y-%m-%d')
                s.query(kaiptz).filter(kaiptz.fphm == fphm).update({'htrq': htrq})

        # ===== 主表回填 3：装柜日期加(+10) + 外表 =====
        zgrqj = ''
        if zgrq != '':
            d = _dt(zgrq)
            if d:
                zgrqj = (d + timedelta(days=10)).strftime('%Y-%m-%d')
                s.query(cymx).filter(cymx.fphm == fphm).update({'zgrq': zgrq, 'zgrqj': zgrqj})
                s.query(cymxsheet).filter(cymxsheet.fphm == fphm).update({'dlrq': zgrq, 'zgrqj': zgrqj})
        patch['zgrqj'] = zgrqj

        # ===== 主表回填 4：申报日期/年月 =====
        patch['sbrq'] = htdate or None
        patch['sbny'] = (htdate or '')[:7]

        # ===== fkhl =====
        fkhl = 0
        if cyrq != '':
            d = run_sql(f"select hl from fkhlb where djny='{cyrq[:7]}'")
            if len(d) > 0:
                fkhl = _f(d[0].get('hl', 0))

        # ===== bgbz =====
        bgbz = 1
        d = run_sql("select cs from zx where mc='信保买单'") if sfxb == '是' else run_sql("select cs from zx where mc='非信保买单'")
        if len(d) > 0:
            bgbz = _f(d[0].get('cs', 1), 1)
        patch['bgbz'] = bgbz

        # ===== 角色判断：宁波优胜 / 财务 =====
        org = get_user_path(user.username)
        path = org.get('path','')
        postion = org.get('position', '')
        is_nbys = '宁波优胜' in path
        is_cw = '财务' in postion
        
        # ===== 合并产品自动报关单价 + CNF分摊 =====
        if not is_nbys and not is_cw:
            for row in hbcplist:
                price = _f(row.get('price', 0))
                lock = _s(row.get('yfsd', ''))
                if price < 0.0000001 and lock != '是':
                    zzsl = _i(row.get('zzsl', 0))
                    tsl = _i(row.get('tsl', 0))

                    if bgsb == '是':
                        base = _f(row.get('mjdj1', 0)) if rmbkh == '是' else _f(row.get('wxjg', 0))
                        row['price'] = _r3(base)
                        row['ybgdj'] = _r3(base)
                        row['wxzj'] = _f(row.get('price',0) * row.get('chsl', 0))
                    else:
                        if sz != 0:
                            base = _f(row.get('mjdj1', 0)) if rmbkh == '是' else _f(row.get('wxjg', 0))
                            row['price'] = _r3(base * sz)
                            row['ybgdj'] = _r3(base * sz)
                            row['wxzj'] = _f(row.get('price',0) * row.get('chsl', 0))
                        else:
                            bz = 0
                            bz1 = 0
                            sb = ''
                            if rmbkh == '是':
                                sssgs = ssgs if 'RMB' in ssgs else (ssgs + 'RMB')
                                d = run_sql(f"select * from bgdjjs where ssgs like '%{sssgs}%' and zzsl={zzsl} and tsl={tsl}")
                                if len(d) == 0:
                                    d = run_sql(f"select * from bgdjjs where ssgs like '%优景RMB%' and zzsl={zzsl} and tsl={tsl}")
                            else:
                                if hbdm in ('USD$', 'USD'):
                                    d = run_sql(f"select * from bgdjjs where ssgs='{ssgs}' and zzsl={zzsl} and tsl={tsl}")
                                else:
                                    d = run_sql(f"select zzslbz zzsbz,tslbz tsbz from bgdjjssheet where ssgs='{ssgs}' and zzsl={zzsl} and tsl={tsl} and hbdm='{hbdm}'")

                            if len(d) > 0:
                                bz = _f(d[0].get('zzsbz', 0))
                                bz1 = _f(d[0].get('tsbz', 0))
                                sb = _s(d[0].get('sb', ''))
                            bzz = bz1 if bz1 > 0 else bz
                            if bzz > 0:
                                cgdj = _f(row.get('cgdj', 0))
                                if bzz == 1 and rmbkh != '是':
                                    bgzj = _f(row.get('wxjg', 0))
                                else:
                                    bgzj = cgdj * bzz if sb == '乘' else (cgdj / bzz if bzz != 0 else 0)
                                row['price'] = _r3(bgzj)
                                row['wxzj'] = _f(row.get('price',0) * row.get('chsl', 0))
                                row['ybgdj'] = _r3(bgzj)

        cnf_used = 0
        if cnfyf > 0 and bgtjhj > 0:
            for row in hbcplist:
                lock = _s(row.get('yfsd', ''))
                row['bggs'] = bggs
                row['sjcy1'] = sjcy
                row['RMBkh'] = rmbkh
                row['CNFyf'] = math.trunc((_f(row.get('ztj', 0)) / bgtjhj) * cnfyf * 100) / 100
                cnf_used += _f(row.get('CNFyf', 0))
                if cnfyf - cnf_used <= 1:
                    row['CNFyf'] = _f(row.get('CNFyf', 0)) + (cnfyf - cnf_used)
                chsl = _f(row.get('chsl', 0))
                row['CNFdj'] = _r3(_f(row.get('CNFyf', 0)) / chsl) if chsl > 0 else 0
                if lock != '是':
                    row['price'] = _f(row.get('CNFdj', 0)) + _f(row.get('ybgdj', 0))
                    row['wxzj'] = _f(row.get('price',0) * row.get('chsl', 0))
                    row['ybgdj'] = row['price']
        
        if rid != '':
            s.query(bgmxdsheet1).filter(bgmxdsheet1.pid == rid).delete()

        # ===== 预填信息 + 产品资料回填 + 补报清单 =====
        ytxxlist = []
        bbqdlist = []
        ytmx_lines = []
        
        fksb_idx = []
        chzje = chzmz = chzjz = chztj = chzxs = 0
        zmz = zjz = ztj = 0
        zsl = zxs = 0
        cywyzd_sum = {}
        for idx, row in enumerate(cpzllist, start=1):
            row['new_flag'] = 0
            if _s(row.get('yfph', '')) != '':
                k = _s(row.get('cywyzd', ''))
                if k != '':
                    cywyzd_sum[k] = cywyzd_sum.get(k, 0) + _f(row.get('gczj', 0))

            if _s(row.get('tsl', '')) == '':
                row['tsl'] = 0

            cght = _s(row.get('cght', ''))
            sccj = _s(row.get('gcmc', ''))
            hyd = _s(row.get('hyd', ''))
            zsl_v = _s(row.get('tsl', '0'))

            # ytfysheet
            d = run_sql(f"select ytbh,fktt,pid from ytfysheet where cght='{cght}' and sccj='{sccj}' and hyd='{hyd}' and zsl='{zsl_v}'")
            for x in d:
                row['fktt'] = _s(x.get('fktt', ''))
                ytxxlist.append({
                    'rid': get_uuid(), 'pid': rid,'ctime': time.strftime("%Y-%m-%d %H:%M:%S"), 'uid': user.rid,
                    'cght': cght, 'ytbh': _s(x.get('ytbh', '')),
                    'ytly': '预填费用', 'ytbm': 'ytfy','new_flag': 1
                })
                ytmx_lines.append(f"采购合同:{cght}预填编号:{_s(x.get('ytbh',''))}")

            # dsytfysheet
            d = run_sql(f"select ytbh,fktt,pid from dsytfysheet where cght='{cght}' and sccj='{sccj}' and hyd='{hyd}' and zsl='{zsl_v}'")
            for x in d:
                row['fktt'] = _s(x.get('fktt', ''))
                ytxxlist.append({
                    'rid': get_uuid(), 'pid': rid,'ctime': time.strftime("%Y-%m-%d %H:%M:%S"), 'uid': user.rid,
                    'cght': cght, 'ytbh': _s(x.get('ytbh', '')),
                    'ytly': '电商预填', 'ytbm': 'dsytfy'
                })
                ytmx_lines.append(f"采购合同:{cght}预填编号:{_s(x.get('ytbh',''))}")

            # 发票/核算发票回填
            row['fphm'] = fphm
            row['ysfp'] = hsfp

            # 原发票号为空 -> 装柜日期继承
            if _s(row.get('yfph', '')) == '':
                row['zgrq'] = zgrq if zgrq else None
                row['zgrqj'] = zgrqj if zgrqj else None

            # 付款抬头校验
            if bggs != '' and _s(row.get('fktt', '')) == '':
                row['fktt'] = bggs
            if bggs != '' and _s(row.get('fktt', '')) != '' and _s(row.get('fktt', '')) != bggs and _i(row.get('zzsl', 0)) > 0:
                fksb_idx.append(str(idx))

            # 出运唯一字段
            if _s(row.get('cywyzd', '')) == '':
                row['cywyzd'] = f"{idx};{fphm}{user.username}{time.strftime('%Y-%m-%d %H:%M:%S')}"

            # 汇总（非补报）
            if _s(row.get('yfph', '')) == '':
                chzje += _f(row.get('wxzj', 0))
                chzmz += _f(row.get('zmz', 0))
                chzjz += _f(row.get('zjz', 0))
                chztj += _f(row.get('ztj', 0))
                chzxs += _f(row.get('chxs', 0))

            if _i(row.get('zzsl', 0)) == 0:
                zmz += _f(row.get('zmz', 0))
                zjz += _f(row.get('zjz', 0))
                ztj += _f(row.get('ztj', 0))
                zsl += _i(row.get('chsl', 0))
                zxs += _i(row.get('chxs', 0))

            # 补报逻辑（原发票号不为空）
            yfph = _s(row.get('yfph', ''))
            if yfph != '':
                cght = _s(row.get('cght', row.get('采购合同', '')))
                cpbh = _s(row.get('cpbh', row.get('产品编号', '')))
                cywyzd = _s(row.get('cywyzd', row.get('出运唯一字段', '')))
                wxwyzd = _s(row.get('wxwyzd', row.get('外销唯一字段', '')))

                # 先删旧
                s.query(dbbqd).filter(dbbqd.cght == cght, dbbqd.cpbh == cpbh, dbbqd.cywyzd == cywyzd, dbbqd.wxwyzd == wxwyzd, dbbqd.fphm == yfph).delete()

                cgdj1 = 0
                d = run_sql(f"select gcjg from cymxsheet where cywyzd='{cywyzd}' and fphm='{yfph}'")
                if len(d) > 0:
                    cgdj1 = _f(d[0].get('gcjg', 0))

                chsl1 = _i(row.get('chsl', 0))
                if _f(row.get('gczj', 0)) > 0 and _f(row.get('gcjg', 0)) > 0:
                    chsl1 = _i(_f(row.get('gczj', 0)) / _f(row.get('gcjg', 0)))
                if cgdj1 > 0:
                    chsl1 = _i(round(_f(row.get('gczj', 0)) / cgdj1))
                
                if cgdj1 > 0 and cywyzd in cywyzd_sum and _f(cywyzd_sum.get(cywyzd, 0)) > 0:
                    chsl1 = _i(round(_f(cywyzd_sum.get(cywyzd, 0)) / cgdj1))

                gczj1 = 0
                d = run_sql(f"select chsl,gczj,gcjg from bgmxdsheet where cght='{cght}' and yfph='{yfph}' and cpbh='{cpbh}' and cywyzd='{cywyzd}' and wxwyzd='{wxwyzd}' and fphm<>'{fphm}' and yfph<>''")
                for x in d:
                    gczj1 += _f(x.get('gczj', 0))
                    if cgdj1 > 0:
                        chsl1 += _i(round(_f(x.get('gczj', 0)) / cgdj1))
                    elif _f(x.get('gcjg', 0)) > 0:
                        chsl1 += _i(_f(x.get('gczj', 0)) / _f(x.get('gcjg', 0)))
                    else:
                        chsl1 += _i(x.get('chsl', 0))

                chslz = 0
                fktt1 = ''
                RMBkh1 = ''
                gczj2 = 0
                d = run_sql(f"select chsl,fktt,RMBkh,gczj from cymxsheet where cywyzd='{cywyzd}' and cght='{cght}' and fphm='{yfph}'")
                if len(d) > 0:
                    chslz = _f(d[0].get('chsl', 0))
                    fktt1 = _s(d[0].get('fktt', ''))
                    RMBkh1 = _s(d[0].get('RMBkh', ''))
                    gczj2 = _f(d[0].get('gczj', 0))
                    
                # 补报清单子表
                bb_row = {
                    'rid': get_uuid(), 
                    'pid': rid,
                    'ctime': time.strftime("%Y-%m-%d %H:%M:%S"), 
                    'uid': user.rid,
                    'fphm': fphm,
                    'yfph': yfph,
                    'zwpm': _s(row.get('zhwbgpm', '')),
                    'zsl': _i(row.get('tsl', 0)),
                    'hgjldw': _s(row.get('hgjldw', '')),
                    'gczj': _f(row.get('gczj', row.get('采购总价', 0))),
                    'sccj': _s(row.get('gcmc', row.get('工厂名称', ''))),
                    'ywrya': _s(row.get('ywry', row.get('业务人员', ''))),
                    'tel': _s(row.get('kpgc', row.get('开票工厂', ''))),
                    'new_flag': 1
                }
                bbqdlist.append(bb_row)

                if _s(row.get('bbwc', row.get('补报完成', ''))) != '是':
                    if not (chsl1 >= chslz or gczj1 == gczj2):
                        if (chslz - chsl1) > 2:
                            # dbbqd落库（按model字段存在性赋值）
                            m = dbbqd()
                            m.rid = get_uuid()
                            m.uid = user.rid          # sys_owner -> uid
                            m.ctime = time.strftime("%Y-%m-%d %H:%M:%S")   # sys_created -> ctime
                            m.mtime = time.strftime("%Y-%m-%d %H:%M:%S")   # sys_lastmodified -> mtime

                            rem = _f(chslz - chsl1)
                            cgr = cgdj1 if cgdj1 > 0 else _f(row.get('gcjg', 0))

                            sfpx = _s(row.get('sfpx', ''))
                            wxrl = _f(row.get('wxrl', 0))
                            if (sfpx in ('否', '')) and wxrl != 0:
                                chxs = rem / wxrl
                                zmz_v = _r2(chxs * _f(row.get('wxmz', 0)))
                                zjz_v = _r2(chxs * _f(row.get('wxjz', 0)))
                                ztj_v = _r3(chxs * _f(row.get('wxtj', 0)))
                            else:
                                chxs = 0
                                zmz_v = 0
                                zjz_v = 0
                                ztj_v = 0

                            _set_if_has(m, {
                                'kh_id': _s(j.get('khbh', '')),
                                'khmc': khmc,
                                'chyrq': _s(row.get('chyrq', '')),
                                'fphm': yfph,
                                'wxht': _s(row.get('wxht', '')),
                                'bzcd': _f(row.get('bzcd', 0)),
                                'bzkd': _f(row.get('bzkd', 0)),
                                'bzgd': _f(row.get('bzgd', 0)),
                                'sjcy1': _s(row.get('sjcy1', '')),
                                'xddd': _s(row.get('xddd', '')),
                                'khht': _s(row.get('khht', '')),
                                'khhh': _s(row.get('khhh', '')),
                                'cpbh': _s(row.get('cpbh', '')),
                                'djpmy': _s(row.get('djpmy', '')),
                                'zwpm': _s(row.get('zwpm', '')),
                                'mjdj1': _f(row.get('mjdj1', 0)),
                                'wxjg': _f(row.get('wxjg', 0)),
                                'krpp': _s(row.get('krpp', '')),
                                'ysfp': _s(row.get('yysfp', '')),
                                
                                'gcjg': cgr,
                                
                                'jldw': _s(row.get('jldw', '')),
                                'wxrl': _f(row.get('wxrl', 0)),
                                'wxmz': _f(row.get('wxmz', 0)),
                                'wxjz': _f(row.get('wxjz', 0)),
                                'wxtj': _f(row.get('wxtj', 0)),
                                'gcmc': _s(row.get('gcmc', '')),
                                'ytsb': _s(row.get('ytsb', '')),
                                'zhwbgpm': _s(row.get('zhwbgpm', '')),
                                'zzsl': _i(row.get('zzsl', 0)),
                                'gcdh': _s(row.get('gcdh', '')),
                                'jhrq': _s(row.get('jhrq', '')),
                                'ywrya': _s(row.get('ywrya', '')),
                                'ywry': _s(row.get('ywry', '')),
                                'gdry': _s(row.get('gdry', '')),
                                'wxbm1': _s(row.get('wxbm1', '')),
                                'cgbm1': _s(row.get('cgbm1', '')),
                                'sccj1id': _s(row.get('sccj1id', '')),
                                'zkfy': _f(row.get('zkfy', 0)),
                                'bz1': _s(row.get('bz1', '')),
                                
                                'sfpx': sfpx,
                                'krcode': _s(row.get('krcode', '')),
                                'djpmw': _s(row.get('djpmw', '')),
                                'sfsj': _s(row.get('sfsj', '')),
                                'djje': _f(row.get('djje', 0)),
                                'tsl': _i(row.get('tsl', 0)),
                                'scrq': _s(row.get('scrq', '')),
                                'kpgc': _s(row.get('kpgc', '')),
                                'zzjgdm': _s(row.get('zzjgdm', '')),
                                'kplxr': _s(row.get('kplxr', '')),
                                'kpdh': _s(row.get('kpdh', '')),
                                'yjcq': _s(row.get('yjcq', '')),
                                'caiziz': _s(row.get('caiziz', '')),
                                'hgbm': _s(row.get('hgbm', '')),
                                'hgjldw': _s(row.get('hgjldw', '')),
                                'dzbz': _s(row.get('dzbz', '')),
                                'cght': cght,
                                'dq1': _s(row.get('dq1', '')),
                                'sbys': _s(row.get('sbys', '')),
                                'cywyzd': cywyzd,
                                'wxwyzd': wxwyzd,
                                
                                'kpdj': _f(row.get('kpdj', 0)),
                                'kpds': _f(row.get('kpds', 0)),
                                'kpsl': _f(row.get('kpsl', 0)),
                                'tse': _f(row.get('tse', 0)),
                                'ewchy': _s(row.get('ewchy', '')),
                                'name': 'name',
                                'hyd': _s(row.get('hyd', '')),
                                'fkxq': _s(row.get('fkxq', '')),
                                'szxq': _s(row.get('szxq', '')),
                                'zkfy1': _f(row.get('zkfy1', 0)),
                                'yfje': _f(row.get('yfje', 0)),
                                'fktt': fktt1,
                                'htrq': _s(row.get('htrq', '')),
                                'ysqje': _f(row.get('ysqje', 0)),
                                'jsfs': _s(row.get('jsfs', '')),
                                'chsl': rem,
                                'ck': '是',
                                'wxfp': '',
                                'chxs': chxs,
                                'zmz': zmz_v,
                                'zjz': zjz_v,
                                'ztj': ztj_v,
                                'mjzj': rem * _f(row.get('mjdj1', 0)),
                                'wxzj': rem * _f(row.get('wxjg', 0)),
                                'gczj': rem * cgr,
                                'zgrq': _s(row.get('zgrq', '')),
                                'zgrqj': _s(row.get('zgrqj', '')),
                            })
                            s.add(m)

        # ===== 报关单子表重建 =====
        bgdlist = []
        for r in hbcplist:
            r['RMBkh'] = rmbkh
            r['bggs'] = bggs
            r['sbrq'] = patch.get('sbrq', '')

            # 行汇率为0时，用主表汇率
            if _f(r.get('hl', 0)) == 0 and hl != 0:
                r['hl'] = hl

            # 风控金额
            bgzj_row = _f(r.get('bgzj', 0))
            if rmbkh == '是':
                if fkhl != 0:
                    r['fkje'] = _r2(bgzj_row / fkhl)
                elif hl != 0:
                    r['fkje'] = _r2(bgzj_row / hl)
                else:
                    r['fkje'] = 0
            else:
                r['fkje'] = _r2(bgzj_row * zmjhl)
            
            one = {
                'rid': get_uuid(),
                'pid': rid,
                'ctime': time.strftime("%Y-%m-%d %H:%M:%S"),
                'uid': user.rid,
                'hgbm': _s(r.get('hgbm', '')),
                'zwpm': _s(r.get('zwpm','')),
                'ywpm': _s(r.get('ywpm', '')),
                'chsl': _f(r.get('chsl', 0)),
                'jldw': _s(r.get('jldw', '')),
                'wxzj1': _r3(_f(r.get('wxzj1', 0)) * bgbz),
                'wxzj': _f(r.get('wxzj', 0)),
                'sbys': _s(r.get('sbys', '')),
                'hyd': _s(r.get('hyd', '')),
                'zmz': round(_f(r.get('zjz', 0)) * (10 if 'BEST PRICE' in khmc else 100)) / (10 if 'BEST PRICE' in khmc else 100),
                'hgjldw': _s(r.get('hgjldw', '')),
                'mjzj': _f(r.get('wxzj', 0)),
                'mjzj1': _r3(_f(r.get('mjzj', 0)) * bgbz),
                'new_flag': 1
            }
            bgdlist.append(one)
        bgdlist.append({'rid': get_uuid(), 'pid': rid,'ctime': time.strftime("%Y-%m-%d %H:%M:%S"), 'uid': user.rid, 'hgbm': '', 'zwpm': '', 'ywpm': '', 'chsl': 0, 'jldw': '', 'wxzj1': 0, 'wxzj': 0, 'zmz': 0, 'mjzj': 0, 'mjzj1': 0, 'hgjldw': ''})

        # ===== 贸易国别中文 =====
        if mygb != '':
            d = run_sql(f"select zwmc from mygb where mygb='{mygb}'")
            if len(d) > 0:
                patch['mygbz'] = _s(d[0].get('zwmc', ''))

        # ===== 中文品名总 + 预填明细 + 汇总回填 =====
        zw_set = []
        tshj13 = tshj9 = tshj3 = tshj1 = tshj0 = 0
        ij = 0

        for r in hbcplist:
            ij += 1
            
            if _s(r.get('zwpm1', '')) == '':
                r['zwpm1'] = _s(r.get('zwpm', ''))

            # 退税率合计
            tsl = _i(r.get('tsl', 0))
            bgzj = _f(r.get('wxzj', 0))
            if tsl == 13: tshj13 += bgzj
            if tsl == 9:  tshj9  += bgzj
            if tsl == 3:  tshj3  += bgzj
            if tsl == 1:  tshj1  += bgzj
            if tsl == 0:  tshj0  += bgzj

            # 出口货物报关单号 = 主表报关单号 + 3位序号
            if bgdh != '' and _s(r.get('ckhwbgdh', '')) == '':
                no = f"{bgdh}{ij:03d}"
                r['ckhwbgdh'] = no

            # 报关量：KGS/千克 -> 总净重，否则 -> 出货数量
            hgjldw = _s(r.get('hgjldw', '')).upper()
            if hgjldw in ('KGS', '千克'):
                r['bgl'] = _f(r.get('zjz', 0))
            else:
                r['bgl'] = _f(r.get('chsl', 0))
            r['new_flag'] = 0

            # 中文品名总（去重）
            pm = _s(r.get('zwpm', ''))
            if pm != '' and pm not in zw_set:
                zw_set.append(pm)

            # 中文品名变化，同步外表 cymxsheet.zhwbgpm（Pascal: father=self.getnumber）
            zwpm = _s(r.get('zwpm', ''))
            zwpm1 = _s(r.get('zwpm1', ''))
            if zwpm != zwpm1 and zwpm != '' and zwpm1 != '' and rid != '':
                s.query(cymxsheet).filter(cymxsheet.pid == rid,cymxsheet.zhwbgpm == zwpm1).update({'zhwbgpm': zwpm})

        patch['zwpmz'] = '\r\n'.join(zw_set)
        patch['ytmx'] = '\r\n'.join(ytmx_lines)
        patch['nbg_zmz'] = _r2(zmz)
        patch['nbg_zjz'] = _r2(zjz)
        patch['nbg_ztj'] = round(ztj * 10000) / 10000
        patch['nbg_zsl'] = zsl
        patch['nbg_zxs'] = zxs
        patch['hj_xs'] = chzxs
        patch['hj_je'] = chzje
        patch['hj_mz'] = chzmz
        patch['hj_jz'] = chzjz
        patch['hj_tj'] = chztj
        patch['tshj13'] = tshj13
        patch['tshj9'] = tshj9
        patch['tshj3'] = tshj3
        patch['tshj1'] = tshj1
        patch['tshj0'] = tshj0
        if ds_hl == 0:
            ds_hl = 1
            patch['ds_hl'] = ds_hl

        # ===== AMZ费用回写 =====
        if 'AMZ' in fphm:
            d = run_sql(f"select fyzj1,mdgfy,gnnlf,hybfM,qggsM,`qgzsM`,`mdgewfyM` from cymx where fphm='{fphm}'")
            if len(d) > 0:
                rr = d[0]
                if 'CNF' in jgtk.upper():
                    chhj = _f(rr.get('fyzj1', 0))/ds_hl + ds_hyf_usd + ds_hyf_rmb/ds_hl + _f(rr.get('mdgfy',0)) + _f(rr.get('gnnlf',0))/ds_hl + _f(rr.get('hybfM',0)) + bgje
                    dszfy = chhj + _f(rr.get('qggsM',0)) + _f(rr.get('qgzsM',0)) + _f(rr.get('mdgewfyM',0))
                else:
                    chhj = _f(rr.get('fyzj1', 0))/ds_hl + _f(rr.get('mdgfy',0)) + _f(rr.get('gnnlf',0))/ds_hl + _f(rr.get('hybfM',0)) + bgje
                    dszfy = chhj + _f(rr.get('qggsM',0)) + _f(rr.get('qgzsM',0)) + _f(rr.get('mdgewfyM',0))
                s.query(cymx).filter(cymx.fphm == fphm).update({'chhj1': chhj, 'dszfy': dszfy, 'bgje': bgje, 'sfbg': '是'})

        # ===== 报关金额风控 =====
        if zmjhl <= 0:
            zmjhl = 1
        if rmbkh == '是':
            patch['hzdr'] = _r2(bgje)
            patch['hzdm'] = 0
            if fkhl != 0:
                fkje = _r2(bgje / fkhl)
            elif hl != 0:
                fkje = _r2(bgje / hl)
            else:
                fkje = 0
        else:
            patch['hzdr'] = 0
            patch['hzdm'] = _r2(bgje * zmjhl)
            fkje = _r2(bgje * zmjhl)
        patch['fkje'] = fkje

        # ===== 外表回写 =====
        s.query(fpgl).filter(or_(fpgl.wxfp == fphm, fpgl.hsfp == fphm), fpgl.sfjd == '否').update({'webpdbg': '是', 'CNFyf': cnfyf})

        if zdry != '':
            d = s.query(sys_user).filter(sys_user.rid == zdry).first()
            if d:
                s.query(bgmxd).filter(bgmxd.fphm == fphm, bgmxd.uid != str(d.rid)).update({'uid': str(d.rid)})

        wxjes = (_f(patch['hzdr']) / hl) if (rmbkh == '是' and hl != 0) else _f(patch['hzdm'])
        s.query(cymx).filter(cymx.fphm == fphm).update({
            'bggs': bggs, 'bgRMBkh': rmbkh, 'bgbgzje': bgje, 'fttdh': fttdh, 'bgdh': bgdh,
            'tdh': tdh, 'xh': xh, 'fh': fh, 'sfcy': sfcy, 'sbdate': patch['sbrq'],
            'bghjDR': _f(patch['hzdr']), 'bghjDM': _f(patch['hzdm']), 'wxjes': wxjes
        })
        if tdh != '':
            s.query(cymxsheet9).filter(cymxsheet9.fphm == fphm).update({'tdh': tdh})
        s.query(cymxsheet).filter(cymxsheet.fphm == fphm).update({'tdh': tdh})

        # ===== 校验 =====
        if len(fksb_idx) > 0:
            errors.append('第' + ';'.join(fksb_idx) + '条记录付款抬头有误!!')
        if jgtk == 'CNF' and cnfyf == 0:
            errors.append('请填写报关合计信息——CNF运费$')
            
        old_xgxq = _s(j.get('xgxq', ''))
        add_xgxq = f"{user.username}{time.strftime('%Y-%m-%d %H:%M:%S')};{_f(bgje)}"
        patch['xgxq'] = (old_xgxq + '\r\n' + add_xgxq) if old_xgxq else add_xgxq

        # 信保差额校验（Pascal: ShowError，不拦截保存）
        if sfxb == '是' and ('BEST PRICE' not in khmc):
            myhj = _f(j.get('myhj', j.get('hj_my', 0)))
            hz_rmb = _f(j.get('hzjhc', j.get('hj_hz_rmb', 0)))
            hz_usd = _f(j.get('wxzes', j.get('hj_hz_usd', 0)))

            if sz == 0:
                base = (hz_rmb - myhj) if ('RMB' in ssgs) else (hz_usd - myhj)
            else:
                base = ((hz_rmb - myhj) * sz) if ('RMB' in ssgs) else ((hz_usd - myhj) * sz)

            diff = _f(base) - _f(bgje)
            if abs(diff) > 0.0001:
                unit = '￥' if ('RMB' in ssgs) else '$'
                warnings.append(f'请注意出运明细货值合计和报关金额差值为{round(diff, 2)}{unit}')

        # ===== sys_alarm/xxck/task（按项目方法）=====
        if fkje > 60000:
            msg = f'发票[{fphm}]报关总和超过6万美金,请检查是否正确'
            warnings.append(msg)
            
        if len(errors) > 0:
            s.rollback()
            return json_result(0, '保存前校验失败', {
                'errors': errors,
                'warnings': warnings,
                'patch': patch,
                'hbcplist': hbcplist,
                'cpzllist': cpzllist,
                'bgdlist': bgdlist,
                'bbqdlist': bbqdlist,
                'ytxxlist': ytxxlist
            })
# hbcplist、cpzllist
# bgdlist、bbqdlist、ytxxlist

        s.commit()
        return json_result(1, '处理成功', {
            'errors': [],
            'warnings': warnings,
            'patch': patch,
            'hbcplist': hbcplist,
            'cpzllist': cpzllist,
            'bgdlist': bgdlist,
            'bbqdlist': bbqdlist,
            'ytxxlist': ytxxlist
        })
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
 
@any_route('/api/saier/declaration/ssgs/change', methods=['POST', 'GET'])
@require_token
async def view_saier_declaration_ssgs_change(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        ssgs = j.get('ssgs','')
        fphm = j.get('fphm','')
        khmc = j.get('khmc','')
        rmbkh = j.get('rmbkh','')
        
        msg = ''
        ssgsdata = 0
        msg1 = ''
        rmbdata = 0
        bggs = 0
        if ssgs != '':
            if 'CSD-' in ssgs:
                if '可思达' not in fphm:
                    ssgsdata = 1
                    msg = '请注意此客人锁定可思达报关'
                    ssgs = ''
            else:
                d = run_sql(f"select bz,bz1 from cyzglsheet where (xm='{str(khmc)}') and (zm='预付款锁定付款抬头')")
                if len(d)>0:
                    if d[0].get('bz1','') not in ssgs:
                        ssgsdata = 1
                        msg = '请注意此客人锁定' + d[0].get('bz1','') + '报关'
                        ssgs = ''
                if ssgs != '':
                    gssb = ''
                    d = run_sql(f"select * from cyzglsheet where (xm='{str(khmc)}') and (zm='我方公司不锁定') and (xm<>'') and (xm is not null)")
                    if len(d)>0:
                        gssb = '1'
                    if gssb != '1':
                        d = run_sql(f"select gsjc from kh where (company_name='{str(khmc)}') and (gsjc<>'') and (gsjc is not null)")
                        if len(d)>0:
                            if d[0].get('gsjc','') not in ssgs:
                                ssgsdata = 1
                                msg = '请注意此客人锁定' + d[0].get('gsjc','') + '报关'
                                ssgs = ''
            if ssgs != '':
                if rmbkh == '是':
                    if 'RMB' not in ssgs:
                        rmbdata = 1
                        msg1 = '请注意此为RMB客人'
                else:
                    if 'RMB' in ssgs:
                        rmbdata = 1
                        msg1 = '请注意此为RMB客人'
                
                d = run_sql(f"select bz from bgdjjs where ssgs='{ssgs}' limit 1")
                if len(d)>0:
                    bggs = d[0].get('bz','')
        
        data = {'ssgsdata':ssgsdata,'msg':msg,'msg1':msg1,'rmbdata':rmbdata,'bggs':bggs}
        return json_result(1, '查询成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        
@any_route('/api/saier/declaration/fphm/change', methods=['POST'])
@require_token
async def view_saier_declaration_fphm_change(request):
    """
    用于“报关明细.业务人员.change/发票号码.change”联动
    入参: fphm, ywry
    返回: 
      fphmdata: 1=无此发票号
      chydh: 出运单号
      bgdxxdata: 1=有报关单信息
      sz: 数值1
      xbtd: 信保特定
      fkbbdata: 1=有风控补报
      ywxz: 业务小组
    """
    user = request.current_user
    j = await request.json()
    fphm = _s(j.get('fphm'))
    ywry = _s(j.get('ywry'))
    try:
        data = {
            "fphmdata": 0,    # 1=无此发票号
            "chydh": "",
            "bgdxxdata": 0,   # 1=有报关单信息
            "sz": 0,
            "xbtd": "",
            "fkbbdata": 0,    # 1=有风控补报
            "ywxz": ""        # 业务小组
        }
        
        d = run_sql("select * from cyzglsheet where (xm=:name) and (zm='报关明细excel引入')", {"name": user.username}) or []
        if len(d) == 0:
            return json_result(1, 'ok', data)
        # 1. 校验发票号是否存在
        cy = run_sql("select chydh from cymx where fphm=:f limit 1", {"f": fphm}) or []
        if not cy:
            data["fphmdata"] = 1
            return json_result(1, 'ok', data)
        chydh = _s(cy[0].get('chydh'))
        data["chydh"] = chydh

        # 2. 若有出运单号，查 chyjh 表
        if chydh:
            chyjh = run_sql("select xbtd,sz from chyjh where wxfp=:wxfp limit 1", {"wxfp": chydh}) or []
            if chyjh:
                data["bgdxxdata"] = 1
                data["sz"] = _f(chyjh[0].get('sz'))
                data["xbtd"] = _s(chyjh[0].get('xbtd'))

            # 3. 风控补报（dbbqd）
            fkbb = run_sql("select rid from dbbqd where bbfp=:bbfp limit 1", {"bbfp": fphm}) or []
            if fkbb:
                data["fkbbdata"] = 1

        # 4. 业务小组（ywrybiao.bm 拆分大写字母后半段）
        if ywry:
            ywbm_row = run_sql("select bm from ywrybiao where yhm=:yhm limit 1", {"yhm": ywry}) or []
            ywbm = _s(ywbm_row[0].get('bm')) if ywbm_row else ''
            ywxz = ''
            if ywbm:
                # 拆分大写字母为分界
                zs3 = 0
                for idx, ch in enumerate(ywbm):
                    if ch.isupper():
                        zs3 = idx
                        break
                if zs3 > 0:
                    ywxz = ywbm[zs3:]
                else:
                    ywxz = ywbm
            data["ywxz"] = ywxz

        return json_result(1, 'ok', data)
    except Exception:
        logger.error(trace_error())
        return json_result(-1, trace_error())
        
@any_route('/api/saier/declaration/bggs/change', methods=['POST', 'GET'])
@require_token
async def view_saier_declaration_bggs_change(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        bggs = j.get('bggs','')
        cpzl = j.get('cpzl',[])

        rid_list = []
        for row in cpzl:
            fktt = ''
            if row.get('cywyzd','') != '' and row.get('fktt','') != '':
                d = run_sql(f"select fktt from cymxsheet where (fpsb1='是') and (cywyzd='{row.get('cywyzd','')}') and (zzsl>0)")
                if len(d)>0:
                    fktt = d[0].get('fktt','').strip()
                if bggs != '' and fktt =='':
                    rid_list.append(row.get('rid',''))
        
        data = {'rid_list':rid_list}
        return json_result(1, '查询成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        
@any_route('/api/saier/declaration/cpzl/yfph/change', methods=['POST', 'GET'])
@require_token
async def view_saier_declaration_cpzl_yfph_change(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        yfph = j.get('yfph','')

        sjcy1 = ''
        sjcydata = 0
        d = run_sql(f"select sjcy1 from cymx where fphm='{yfph}'")
        if len(d) > 0:
            sjcy1 = d[0].get('sjcy1','')
            if sjcy1 != '':
                sjcydata = 1

        data = {'sjcydata':sjcydata,'sjcy1':sjcy1}
        return json_result(1, '查询成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        
@any_route('/api/saier/declaration/hbcp/cwxg/change', methods=['POST', 'GET'])
@require_token
async def view_saier_declaration_hbcp_cwxg_change(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        rid = j.get('rid','')
        cwxgdata = 0
        
        org = get_user_path(user.username)
        path = org.get('path','')
        postion = org.get('position', '')
        if '财务' in postion:
            cwxgdata = 1
            
        data = {'cwxgdata':cwxgdata}
        return json_result(1, '查询成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        
@any_route('/api/saier/declaration/hbcp/hgbm/change', methods=['POST', 'GET'])
@require_token
async def view_saier_declaration_hbcp_hgbm_change(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        rid = j.get('rid','')
        hgbm = j.get('hgbm','')

        sb = ''
        if hgbm != '':
            d = run_sql("select * from cyzglsheet where :hgbm like concat(xm, '%') and zm='禁止出运海关编码'",{"hgbm": str(hgbm)})
            if len(d) > 0:
                sb = '禁止出运'

            d = run_sql(f"select * from cyzglsheet where (xm='{str(hgbm)}') and (zm='禁止报关海关编码')")
            if len(d) > 0:
                sb = sb + '禁止出运'

        data = {'sb':sb}
        return json_result(1, '查询成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

@any_route('/api/saier/declaration/items/delete', methods=['POST', 'GET'])
@require_token
async def view_saier_declaration_items_delete(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        line = j.get('line',{})
        zfphm = j.get('zfphm','')
        khbh = j.get('khbh','')
        khmc = j.get('khmc','')
        
        yfph = line.get('yfph','')
        cywyzd = line.get('cywyzd','')
        wxwyzd = line.get('wxwyzd','')
        cght = line.get('cght','')
        cpbh = line.get('cpbh','')
        fphm = line.get('fphm','')
        
        cgdj = 0
        if yfph != '':
            chsl = 0
            chslz = 0
            gczj = 0
            gczjz = 0
            if cywyzd != '':
                d = run_sql(f"select gcjg from cymxsheet where (cywyzd='{cywyzd}') and (fphm='{yfph}')")
                if len(d) > 0:
                    cgdj = d[0].get('gcjg',0)
                else:
                    d = run_sql(f"select UnitPurchasePrice from rtspurchaseordergoods where (cgwyzd='{cywyzd}') and (PurchaseOrderNo='{cght}') and (PurchaseOrderNo='{yfph}')")
                    if len(d) > 0:
                        cgdj = d[0].get('UnitPurchasePrice',0)
            bbxq = ''
            d = run_sql(f"select chsl,cywyzd,fphm,gczj,gcjg,gcmc,tsl,zhwbgpm from bgmxdsheet \
                where cght='{cght}' and cpbh='{cpbh}' and cywyzd='{cywyzd}' and wxwyzd='{wxwyzd}' and fphm<>'{zfphm}' and yfph='{yfph}' and yfph<>'' and yfph is not null")
            if len(d) > 0:
                for row in d:
                    item_str = (
                        f"{row.get('gcmc','')}退{int(row.get('tsl',0))};"
                        f"{row.get('zhwbgpm','')};"
                        f"补发票:{zfphm};"
                        f"出货数量:{row.get('chsl',0):.2f};"
                        f"出货金额:{row.get('gczj',0):.2f}"
                    )
                    if not bbxq:
                        bbxq = item_str
                    else:
                        bbxq += '/' + item_str

                    gczj += float(row.get('gczj',0))
                    
                    if cgdj > 0:
                        qty_add = round(float(row.get('gczj',0)) / cgdj)
                    elif float(row.get('gcjg',0)) > 0:
                        qty_add = int(float(row.get('gczj',0)) / float(row.get('gcjg',0)))
                    else:
                        qty_add = int(float(row.get('chsl',0)))

                    chsl += qty_add
            fktt1 = 0
            RMBkh = 0
            d = run_sql(f"select chsl,fktt,RMBkh,gczj from cymxsheet where (cywyzd='{cywyzd}') and (cght='{cght}') and (fphm='{yfph}')")
            if len(d) > 0:
                chslz = d[0].get('chsl',0)
                fktt1 = d[0].get('fktt','')
                RMBkh = d[0].get('RMBkh','')
                gczjz = d[0].get('gczj',0)
            else:
                d = run_sql(f"select Quantity,PurchaseAmount from rtspurchaseordergoods where (cgwyzd='{cywyzd}') and (PurchaseOrderNo='{cght}') and (PurchaseOrderNo='{yfph}')")
                if len(d) > 0:
                    chslz = d[0].get('Quantity',0)
                    gczjz = d[0].get('PurchaseAmount',0)
            update_json = {'IsSupplementary':'否','SupplementaryAmount':0}
            s.query(rtspurchaseordergoods).filter(rtspurchaseordergoods.cgwyzd == str(cywyzd),rtspurchaseordergoods.PurchaseOrderNo == str(cght),rtspurchaseordergoods.PurchaseOrderNo == str(yfph)).update(update_json)
            
            update_json1 = {'bbxq':bbxq}
            s.query(cymxsheet).filter(cymxsheet.cywyzd == str(cywyzd),cymxsheet.cght == str(cght),cymxsheet.fphm == str(yfph)).update(update_json1)
            s.query(dbbqd).filter(dbbqd.cght==str(cght),dbbqd.cpbh==str(cpbh),dbbqd.cywyzd==str(cywyzd),dbbqd.wxwyzd==str(wxwyzd),dbbqd.fphm==str(yfph)).delete()

            if not (chsl >= chslz or gczjz == gczj):
                # 原逻辑应为“剩余数量 > 2”
                if chslz - chsl > 2:
                    m1 = dbbqd()
                    rid = get_uuid()
                    m1.rid = rid
                    m1.uid = user.rid
                    m1.ctime = time.strftime("%Y-%m-%d %H:%M:%S")
                    m1.RMBkh = RMBkh
                    
                    for k,v in line.items():
                        if k in SYS_FIELDS:
                            continue
                        if hasattr(m1,k):
                            setattr(m1,k,v)
                    
                    m1.kh_id = khbh
                    m1.khmc = khmc
                    m1.fphm = yfph
                    m1.ysfp = line.get('yysfp','')
                    m1.name = 'name'
                    m1.fktt = fktt1
                    m1.chsl = chslz - chsl
                    m1.ck = '是'
                    m1.wxfp = ''

                    if line.get('sfpx','') == '否' or line.get('sfpx','') == '':
                        if line.get('wxrl',0) != 0:
                            m1.chxs = (chslz - chsl) / line.get('wxrl',0)
                            m1.zmz = (chslz - chsl) / line.get('wxrl',0) * line.get('wxmz',0)
                            m1.zjz = (chslz - chsl) / line.get('wxrl',0) * line.get('wxjz',0)
                            m1.ztj = (chslz - chsl) / line.get('wxrl',0) * line.get('wxtj',0)
                        else:
                            m1.chxs = 0
                            m1.zmz = 0
                            m1.zjz = 0
                            m1.ztj = 0
                    else:
                        m1.chxs = 0
                        m1.zmz = 0
                        m1.zjz = 0
                        m1.ztj = 0

                    m1.mjzj = (chslz - chsl) * line.get('mjdj1',0)
                    m1.wxzj = (chslz - chsl) * line.get('wxjg',0)
                    
                    if cgdj > 0 :
                        m1.gcjg = cgdj
                        m1.gczj = (chslz - chsl) * cgdj
                    else:
                        m1.gczj = (chslz - chsl) * line.get('gczj',0)
                    s.add(m1)

        # m = deletetable()
        # rid = get_uuid()
        # m.rid = rid
        # m.uid = user.rid
        # m.ctime = time.strftime("%Y-%m-%d %H:%M:%S")
        
        # m.ly = '报关明细详情表'
        # m.scsj = time.strftime("%Y-%m-%d %H:%M:%S")
        # m.scr = user.username
        # m.scbz = cpbh
        # m.scnr = '发票号' + fphm + '原发票号' + yfph + '出运维一' + cywyzd
        # s.add(m)
        
        s.commit()
        
        return json_result(1, '处理成功！')
    except:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

@any_route('/api/saier/declaration/button/hbcpyj', methods=['POST', 'GET'])
@require_token
async def view_saier_declaration_button_hbcpyj(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        rid = j.get('rid','')

        hbcpyjdata = 1
        # org = get_user_path(user.username)
        # path = org.get('path','')
        # postion = org.get('position', '')
        # if '财务' in postion:
        #     hbcpyjdata = 1
        data = {'hbcpyjdata':hbcpyjdata}
        
        return json_result(1, '查询成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

@any_route('/api/saier/declaration/button/sxhyd', methods=['POST', 'GET'])
@require_token
async def view_saier_declaration_button_sxhyd(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        rid = j.get('rid','')
        datalist = j.get('datalist',[])

        redatalist = []
        if len(datalist) > 0:
            for r in datalist:
                if r.get('kpgc','') != '' and r.get('hyd','') == '':
                    d = run_sql(f"select hyd from ozycs where (company_name='{r.get('kpgc','')}')")
                    if len(d)>0:
                        data_json = {'rid':r.get('rid',''),'hyd':d[0].get('hyd','')}
                        redatalist.append(data_json)
            
        data = {'redatalist':redatalist}
        
        return json_result(1, '查询成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        
@any_route('/api/saier/declaration/button/plfktt', methods=['POST', 'GET'])
@require_token
async def view_saier_declaration_button_plfktt(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        rid = j.get('rid','')
        datalist = j.get('datalist',[])
        bggs = j.get('bggs','')
        yfktt = j.get('yfktt','')

        redatalist = []
        if len(datalist) > 0:
            for r in datalist:
                fktt = '' 
                if r.get('cywyzd','') != '':
                    d = run_sql(f"select fktt from cymxsheet where  (fpsb1='是') and (cywyzd='{r.get('cywyzd','')}') and (zzsl>0)")
                    if len(d)>0:
                        fktt = d[0].get('fktt','').strip()

                if bggs != '' and r.get('fktt','') == yfktt and fktt == '':
                    data_json = {'rid':r.get('rid','')}
                    redatalist.append(data_json)

        data = {'redatalist':redatalist}
        
        return json_result(1, '查询成功', data)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()

@any_route('/api/saier/declaration/button/cwplfj', methods=['POST', 'GET'])
@require_token
async def view_saier_declaration_button_cwplfj(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        rids = j.get('rids',[])
        path = config.tmp_path
        data_path = config.data_upload_path
        files = []
        for r in rids:
            d = run_sql(f"select rid,fphm from bgmxd where ((tzcw='{user.username}') or (tzcw='') or (tzcw is null)) and (rid = '"+str(r)+"')")
            if len(d) > 0:
                ds = run_sql(f"select * from sys_attachment where pid = '{d[0].get('rid','')}' and module='报关明细'")
                if len(ds)>0:
                    i = 1
                    for row in d:
                        file_path = os.path.join(data_path, row.get('path',''))
                        if os.path.exists(file_path):
                            files.append({'name': str(row.get('name','')), 'path': file_path + '/'+ str(row.get('path',''))})
        
        filename = ''
        if len(files) > 0:
            filename = '报关明细附件' + time.strftime("%Y%m%d%H%M%S") + '.zip'
            sZipPath = os.path.join(path, filename) # 压缩包路径
            zipFile = zipfile.ZipFile(sZipPath, 'w') #生成一个压缩包，第二个参数默认值为'r'，表示读已经存在的zip文件，'w'表示新建一个zip文档或覆盖一个已经存在的zip文档
            for f in files:
                file_path = os.path.join(path,str(f.get("path")))
                if os.path.exists(file_path):
                    zipFile.write(file_path, f.get('name'), zipfile.ZIP_DEFLATED) #将file_path的文件重命名为sfilename

            zipFile.close()
        
        return json_result(1, '下载成功！', filename)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        
@any_route('/api/saier/declaration/button/drxjsc', methods=['POST', 'GET'])
@require_token
async def view_saier_declaration_button_drxjsc(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        r_path = os.path.join(config.data_upload_path,'template')
        fn = os.path.join(r_path, '现金导入.xlsx')
        
        if not os.path.exists(fn):
            return {"code":-1,"msg":"未找到报表模板"}
        wb = load_workbook(fn) 
        ws = wb.get_sheet_by_name('Sheet1')
        
        rids = j.get('rids',[])
        org = get_user_path('zjnblh')
        path = org.get('path','')
        postion = org.get('position', '')
        bm1 = org.get('position', '')[:1]
        
        uv = ''
        if bm1 == 'D':
            uv = '1'
        files = []
        d = run_sql(f"select rid from sys_user where username='{user.username}' ")
        if len(d) == 0:
            return json_result(-1, '未找到用户信息！')
        
        for r in rids:
            d = run_sql(f"select fphm,ysfp,rid,chyrq,khmc from bgmxd where rid = '{r}'")
            if len(d) > 0:
                ds = run_sql(f"select * from cymxsheet where pid = '{r}' and (tsl = 0 or tsl is null) order by gcmc desc")
                if len(ds)>0:
                    i = 1
                    for row in ds:
                        i += 1

                        ws['N' + str(i)] = '是'
                        ws['O' + str(i)] = '义乌'
                        ws['R' + str(i)] = str(d[0].get('fphm','')).strip()
                        ws['S' + str(i)] = str(row.get('cght','')).strip()
                        ws['T' + str(i)] = str(row.get('gczj','')).strip()
                        ws['U' + str(i)] = str(d[0].get('chyrq','')).strip()
                        if uv == '1':
                            ws['V' + str(i)] = str(row.get('gdry','')).strip()
                        else:
                            ws['V' + str(i)] = str(row.get('ywrya','')).strip()
                        ws['X' + str(i)] = str(row.get('cpbh','')).strip()
                        ws['Y' + str(i)] = str(row.get('zwpm','')).strip()
                        ws['Z' + str(i)] = str(row.get('gcjg','')).strip()
                        
                        ws['AA' + str(i)] = str(row.get('jldw','')).strip()
                        ws['AB' + str(i)] = row.get('wxrl',0)
                        ws['AC' + str(i)] = row.get('chxs',0)
                        ws['AD' + str(i)] = row.get('chsl',0)
                        ws['AE' + str(i)] = str(row.get('gcmc','')).strip()
                        ws['AF' + str(i)] = str(row.get('gcdh','')).strip()
                        ws['AG' + str(i)] = str(d[0].get('khmc','')).strip()
                        ws['AH' + str(i)] = str(row.get('bz1','')).strip()
                        ws['AJ' + str(i)] = str(user.username).strip()
                        
                    wb.save(path + '/'+ str(username)+ time.strftime("%Y%m%d%H%M%S") + '现金导入.xlsx')
                    files.append({'name': fn, 'path': path + '/'+ str(username)+ time.strftime("%Y%m%d%H%M%S") + '现金导入.xlsx'})

        filename = ''
        if len(files) > 0:
            filename = '现金导入' + time.strftime("%Y%m%d%H%M%S") + '.zip'
            sZipPath = os.path.join(path, filename) # 压缩包路径
            zipFile = zipfile.ZipFile(sZipPath, 'w') #生成一个压缩包，第二个参数默认值为'r'，表示读已经存在的zip文件，'w'表示新建一个zip文档或覆盖一个已经存在的zip文档
            for f in files:
                file_path = os.path.join(path,str(f.get("path")))
                if os.path.exists(file_path):
                    zipFile.write(file_path, f.get('name'), zipfile.ZIP_DEFLATED) #将file_path的文件重命名为sfilename

            zipFile.close()
        
        return json_result(1, '下载成功！', filename)
    except:
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()
        
@any_route('/api/saier/declaration/button/dddrmb/options', methods=['POST', 'GET'])
@require_token
async def api_saier_declaration_button_dddrmb_options(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        rids = j.get('rids', [])
        auth = s.query(cyzglsheet).filter(cyzglsheet.xm == user.username,cyzglsheet.zm == '订单导入模板').first()
        if not auth:
            return json_result(-1, '无权限：订单导入模板')

        rows = s.query(dxkzsheet.xzwb).filter(dxkzsheet.xznr == '订单导入报关公司').all()
        bggs_list = []
        if rows:
            for r in rows:
                if r.xzwb is not None and str(r.xzwb).strip() != '':
                    bggs_list.append(str(r.xzwb).strip())

        return json_result(1, 'ok', {'bggs_list': bggs_list})
    except Exception as e:
        return json_result(-1, str(e))
    finally:
        s.close()
        
@any_route('/api/saier/declaration/button/dddrmb', methods=['POST'])
@require_token
async def api_saier_declaration_button_dddrmb(request):
    s = Session()
    user = request.current_user
    j = await request.json()
    try:
        rids = j.get('rids', [])
        bggs_keyword = j.get('bggs', '')
        allow_imported = j.get('allow_imported', '2')  # '1'允许已导，'2'不允许
        
        if not isinstance(rids, list) or len(rids) == 0:
            return json_result(-1, '请先选择要操作的记录')
        if bggs_keyword is None or str(bggs_keyword).strip() == '':
            return json_result(-1, '请选择报关公司')

        # 权限校验
        auth = s.query(cyzglsheet).filter(cyzglsheet.xm == user.username,cyzglsheet.zm == '订单导入模板').first()
        if not auth:
            return json_result(-1, '无权限：订单导入模板')

        # 模板
        template_path = os.path.join(config.data_upload_path, 'template', '订单导入模板.xlsx')
        if not os.path.exists(template_path):
            return json_result(-1, '模板不存在: 订单导入模板.xlsx')

        # 先取选中 number（对应 Pascal 的 datacenter.getnumberlist + MultiReport）
        num_rows = s.query(bgmxd.rid).filter(bgmxd.rid.in_(rids)).all()
        rid_list = []
        for n in num_rows:
            if n.rid is not None:
                rid_list.append(str(n.rid))

        if len(rid_list) == 0:
            return json_result(-1, '未找到可处理的报关明细')

        # 根据条件过滤可导出的主单
        q = s.query(bgmxd).filter(bgmxd.rid.in_(rid_list),bgmxd.bggs.like(f'%{bggs_keyword}%'))
        if allow_imported != '1':
            q = q.filter(or_(bgmxd.dddrmbsb == '', bgmxd.dddrmbsb.is_(None)))

        head_rows = q.all()
        if len(head_rows) == 0:
            return json_result(-1, '无符合条件的数据')

        # 去重 number（保持 Pascal 行为）
        rid_seen = set()
        export_heads = []
        for r in head_rows:
            key = '' if r.rid is None else str(r.rid)
            if key not in rid_seen:
                rid_seen.add(key)
                export_heads.append(r)

        wb = load_workbook(template_path)
        ws = wb.worksheets[0]

        row_idx = 1  # 对齐 Pascal: i := 1; 之后先 i+1，从第2行开始写
        update_rids = []

        for head in export_heads:
            # 明细：bgmxdhbcp，子表统一 pid
            items = s.query(bgmxdhbcp).filter(bgmxdhbcp.pid == head.rid).all()

            if len(items) == 0:
                continue

            first_line = True
            for it in items:
                row_idx += 1

                if first_line:
                    # A: ysfp
                    ws[f'A{row_idx}'] = '' if head.ysfp is None else str(head.ysfp)

                    # B: kpnr.hgbh by wfgs=bggs
                    kp = s.query(kpnr).filter(kpnr.wfgs == head.bggs).first()
                    if kp and kp.hgbh is not None:
                        ws[f'B{row_idx}'] = str(kp.hgbh)

                    # C/D: 国别代码
                    gb = s.query(mygb).filter(or_(mygb.mygb == head.mygb, mygb.zwmc == head.mygb)
                    ).first()
                    if gb and gb.gbdm is not None:
                        ws[f'C{row_idx}'] = str(gb.gbdm)
                        ws[f'D{row_idx}'] = str(gb.gbdm)

                    # E: 接单日期(jdrq)；无则出运日期-55并避开周末
                    cy = s.query(cymx).filter(cymx.fphm == head.fphm).first()
                    if cy and cy.jdrq is not None and str(cy.jdrq) != '':
                        ws[f'E{row_idx}'] = str(cy.jdrq)
                    else:
                        if head.chyrq is not None and str(head.chyrq) != '':
                            try:
                                d = datetime.strptime(str(head.chyrq)[:10], '%Y-%m-%d') - timedelta(days=55)
                                if d.weekday() == 5:  # 周六
                                    d = d - timedelta(days=1)
                                elif d.weekday() == 6:  # 周日
                                    d = d + timedelta(days=1)
                                ws[f'E{row_idx}'] = d.strftime('%Y-%m-%d')
                            except:
                                ws[f'E{row_idx}'] = ''

                    # F: 出运日期
                    ws[f'F{row_idx}'] = '' if head.chyrq is None else str(head.chyrq)[:10]

                    # G: 币种
                    if head.rmbkh == '是':
                        ws[f'G{row_idx}'] = 'CNY:人民币'
                    else:
                        hbdm = '' if head.hbdm is None else str(head.hbdm)
                        if hbdm == 'EUR':
                            ws[f'G{row_idx}'] = 'EUR:欧元'
                        elif hbdm == 'JPY':
                            ws[f'G{row_idx}'] = 'JPY:日元'
                        elif hbdm == 'HKD':
                            ws[f'G{row_idx}'] = 'HKD:港币'
                        else:
                            ws[f'G{row_idx}'] = 'USD:美元'

                    first_line = False

                # H/I/J/K 明细
                ws[f'H{row_idx}'] = '' if it.hgbm is None else str(it.hgbm)
                ws[f'I{row_idx}'] = '' if it.zwpm is None else str(it.zwpm)
                ws[f'J{row_idx}'] = '否'
                try:
                    ws[f'K{row_idx}'] = float(it.wxzj) if it.wxzj is not None else 0
                except:
                    ws[f'K{row_idx}'] = 0

            update_rids.append(head.rid)

        if row_idx <= 1:
            return json_result(-1, '无可导出明细')

        # 边框（A2:K{row_idx}）
        thin = Side(style='thin', color='000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for r in range(2, row_idx + 1):
            for c in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                ws[f'{c}{r}'].border = border

        # 保存到临时目录
        out_name = f"{user.username}_{datetime.now().strftime('%Y-%m-%d')}_订单导入模板.xlsx"
        out_path = os.path.join(config.tmp_path, out_name)
        os.makedirs(config.tmp_path, exist_ok=True)
        wb.save(out_path)

        # 回写 dddrmbsb（仅回写这次实际导出的主单）
        today_str = datetime.now().strftime('%Y-%m-%d')
        if len(update_rids) > 0:
            s.query(bgmxd).filter(bgmxd.rid.in_(update_rids)).update({'dddrmbsb': today_str})

        s.commit()
        return json_result(1, '导出成功', out_name)

    except Exception as e:
        s.rollback()
        return json_result(-1, str(e))
    finally:
        s.close()
        
def _s(v):
    return '' if v is None else str(v).strip()

def _f(v, d=0.0):
    try:
        return float(v if v not in (None, '') else d)
    except Exception:
        return d

def _i(v, d=0):
    try:
        return int(float(v if v not in (None, '') else d))
    except Exception:
        return d
    
def _kptz_get_cs(ly: str, default_value: str = "") -> str:
    rs = run_sql("select cs from zx where ly=:ly limit 1", {"ly": ly})
    if rs and rs[0].get("cs") not in (None, ""):
        return str(rs[0].get("cs"))
    return default_value

def _kptz_get_user_rid(username: str) -> str:
    d = run_sql(
        "select rid from sys_user where username=:username limit 1",
        {"username": str(username)}
    )
    return str(d[0].get("rid", "")) if d else ""

def _kptz_has_role(username: str, role_name_like: str, role_path_like: str = "") -> bool:
    """
    强制按三步：sys_user -> sys_user_role -> sys_role（禁止join）
    """
    user_rid = _kptz_get_user_rid(username)
    if not user_rid:
        return False

    ds = run_sql(
        "select role_id from sys_user_role where user_id=:user_id",
        {"user_id": user_rid}
    )
    if not ds:
        return False

    for x in ds:
        role_id = str(x.get("role_id", ""))
        if not role_id:
            continue
        if role_path_like:
            dt = run_sql(
                "select rid from sys_role where rid=:rid and name like :name and path like :path limit 1",
                {"rid": role_id, "name": role_name_like, "path": role_path_like}
            )
        else:
            dt = run_sql(
                "select rid from sys_role where rid=:rid and name like :name limit 1",
                {"rid": role_id, "name": role_name_like}
            )
        if dt:
            return True
    return False

def _kptz_trunc10(v):
    v = _f(v)
    return int(v / 10) * 10 if v >= 0 else -int(abs(v) / 10) * 10

def _kptz_wm_rate_pascal(zsl_val, kp3_val, kpds_val) -> float:
    zsl = _f(zsl_val, 0.0)
    return _f(kp3_val, 0.0) if abs(zsl - 3.0) < 1e-9 else _f(kpds_val, 0.0)

def _kptz_upsert_fpgl_full(fphm: str, ysfp: str, user, now_dt, main_row: dict, db_session):
    """
    Pascal尾段对齐版：
    - 仅 UV/VC 且非 AMZ 执行
    - CSD- 跳过
    - fpgl 已存在则跳过（不覆盖）
    - 补齐 sbz/chydh/khpk/gcpk/fyUSD/fyRMB 口径
    """
    def S(v): return '' if v is None else str(v).strip()
    def F(v, d=0.0):
        try: return float(v if v not in (None, '') else d)
        except: return d
    def trunc10(v):
        v = F(v)
        return (int(v / 10) * 10) if v >= 0 else (-int(abs(v) / 10) * 10)

    if db_session is None:
        return {"rid": "", "hsml": 0, "skipped": True}

    bh = S(fphm)
    hsfp = S(ysfp) or bh
    if not bh:
        return {"rid": "", "hsml": 0, "skipped": True}

    # 1) Pascal触发门槛：UV/VC 且非 AMZ
    up = bh.upper()
    dssb1 = '1' if bh[:3].upper() == 'AMZ' else ''
    if not (('UV' in up or 'VC' in up) and dssb1 != '1'):
        return {"rid": "", "hsml": 0, "skipped": True}

    # 2) CSD- 直接跳过
    if 'CSD-' in up or 'CSD-' in hsfp.upper():
        return {"rid": "", "hsml": 0, "skipped": True}

    # 3) 已存在则不覆盖（Pascal：bh1='1'）
    ex = run_sql("select rid from fpgl where wxfp=:w or wxfp1=:h limit 1", {"w": bh, "h": hsfp})
    if ex:
        return {"rid": S(ex[0].get("rid")), "hsml": 0, "skipped": True}

    # 主单
    bg = run_sql("""
        select rid,fphm,ysfp,khmc,RMBkh,bggs,ywry,ywbm,zdry,dlmc,zwpmz,
               ifnull(bgbgzje,0) as bgje,ifnull(CNFyf,0) as CNFyf,
               ifnull(bgdh,'') as bgdh
          from bgmxd
         where fphm=:f limit 1
    """, {"f": bh})
    if not bg:
        return {"rid": "", "hsml": 0, "skipped": True}
    b = bg[0]

    # 出运主
    cy = run_sql("""
        select rid,ifnull(chydh,'') as chydh,
               ifnull(ysdj,0) as ysdj,ifnull(xybx,'') as xybx,ifnull(xbfl,0) as xbfl,
               ifnull(htje,0) as htje,ifnull(htjer,0) as htjer,ifnull(htjem,0) as htjem,
               ifnull(yjje,0) as yjje,ifnull(myjje,0) as myjje,ifnull(ayjje,0) as ayjje,
               ifnull(zgrqj,'') as zgrqj,ifnull(jgtk,'') as jgtk,ifnull(jhfs,'') as jhfs,
               ifnull(shqx,'') as shqx,ifnull(dfrq,'') as dfrq,ifnull(mdka,'') as mdka,
               ifnull(cdmc,'') as cdmc,ifnull(mygb,'') as mygb,ifnull(jxKHRMB,0) as jxKHRMB,
               ifnull(ygnlf,0) as ygnlf,ifnull(jxUSD,0) as jxUSD,ifnull(jxRMB,0) as jxRMB,
               ifnull(kpfyz,0) as kpfyz,ifnull(qtrmb,0) as qtrmb,ifnull(ewhj,0) as ewhj,
               ifnull(hybfM,0) as hyf,
               ifnull(bglx,'') as bglx,ifnull(fpbz12,'') as fpbz12,
               ifnull(jxje1,0) as jxje1,ifnull(jxje2,0) as jxje2,ifnull(jxje3,0) as jxje3,
               ifnull(jjxje1,0) as jjxje1,ifnull(jjxje2,0) as jjxje2,ifnull(jjxje3,0) as jjxje3
          from cymx where fphm=:f limit 1
    """, {"f": bh})
    c = cy[0] if cy else {}
    cy_rid = S(c.get("rid"))
    chydh = S(c.get("chydh"))
    jjxje = F(c.get("jxje1")) + F(c.get("jxje2")) + F(c.get("jxje3")) - F(c.get("jjxje1")) - F(c.get("jjxje2")) - F(c.get("jjxje3"))

    # sbz 判断（Pascal A~S）
    sbz = ''
    if chydh:
        candidates = [chydh + x for x in "ABCDEFGHIJKLMNOPQRS"]
        ph = ",".join([f":p{i}" for i in range(len(candidates))])
        ps = {f"p{i}": candidates[i] for i in range(len(candidates))}
        top = run_sql(f"select fphm from cymx where fphm in ({ph}) order by ifnull(htje,0) desc limit 1", ps)
        if top:
            sbz = '1' if S(top[0].get('fphm')).upper() == bh.upper() else ''
        else:
            sbz = '1'

    # 汇率
    hl = run_sql("select tjhl from kpnr where tjhl>0 order by rid desc limit 1")
    yhhl = F((hl[0] if hl else {}).get("tjhl"))

    # 到账
    djr = run_sql("select ifnull(sum(sydje),0) v from krshsheet where fphm=:f and hbdm in ('RMB','RMB￥')", {"f": bh})
    dju = run_sql("select ifnull(sum(sydje),0) v from krshsheet where fphm=:f and hbdm not in ('RMB','RMB￥')", {"f": bh})
    djjer = F((djr[0] if djr else {}).get("v"))
    djjem = F((dju[0] if dju else {}).get("v"))
    shhjz = djjer + (djjem * yhhl if yhhl > 0 else 0)
    sydje = 0.0
    shrq1 = ""
    father = ""
    kr_rows = run_sql("select sydje2,pid from krshsheet where fphm=:f", {"f": bh}) or []
    for r in kr_rows:
        sydje += F(r.get("sydje2"))
        if not father:
            father = S(r.get("pid"))
    if father:
        kr = run_sql("select djrq from krsh where number=:number limit 1", {"number": father})
        if kr:
            shrq1 = S(kr[0].get("djrq"))
    shje = sydje
    if not shrq1:
        shrq1 = None

    # 保险
    xbfy1 = 0.0
    xbdrsj = ""
    xb = run_sql("select ifnull(xbfy,0) xbfy, ifnull(yrrq,'') yrrq from xbsf where fphm=:f limit 1", {"f": bh})
    if xb:
        xbfy1 = F(xb[0].get("xbfy"))
        xbdrsj = S(xb[0].get("yrrq"))

    # hdfy
    yzaf = hyf_hdfy = sjf1 = 0.0
    cwqrqk = ""
    hdfy_rid = ""
    hd = run_sql("""
        select rid,ifnull(zjfy,0) zjfy,ifnull(zjhy1,0) zjhy1,ifnull(sbje,0) sbje,ifnull(cwqrqk,'') cwqrqk
          from hdfy where fphm=:f limit 1
    """, {"f": bh})
    if hd:
        hz = hd[0]
        yzaf = F(hz.get("zjfy"))
        hyf_hdfy = F(hz.get("zjhy1"))
        sjf1 = F(hz.get("sbje"))
        cwqrqk = S(hz.get("cwqrqk"))
        hdfy_rid = S(hz.get("rid"))

    zgfy = ckfy = sjf = jgd = 0.0
    if hdfy_rid:
        hs = run_sql("""
            select ifnull(sum(sjf),0) sjf1,ifnull(sum(ckfy),0) ckfy1,
                   ifnull(sum(jcfy),0) jcfy1,ifnull(sum(JGD),0) JGD1
              from hdfysheet where pid=:pid
        """, {"pid": hdfy_rid})
        if hs:
            sjf = F(hs[0].get("sjf1"))
            zgfy = F(hs[0].get("ckfy1"))
            ckfy = F(hs[0].get("jcfy1"))
            jgd = F(hs[0].get("JGD1"))

    # BEST PRICE 修正
    if 'BEST PRICE' in S(b.get("khmc")):
        yzaf = yzaf - jgd

    # khpk/gcpk + 费用（仅 sbz=1 且按 chydh）
    khpk = gcpk = fyUSD = fyRMB = ywfy = 0.0
    if sbz == '1' and chydh:
        d1 = run_sql("select ifnull(sum(pkje),0) v from khspsheet1 where hsfp=:f and shjg='通过'", {"f": chydh})
        d2 = run_sql("select ifnull(sum(gcpk),0) v from gongcspsheet1 where hsfp=:f and shjg='通过'", {"f": chydh})
        khpk = F((d1[0] if d1 else {}).get("v"))
        gcpk = F((d2[0] if d2 else {}).get("v"))

        du = run_sql("""select ifnull(sum(seje),0) v from fysqsheet
                         where wxfp=:f and cwsp='通过' and sfrbmfy='否'
                           and (hbdm like '%USD%' or hbdm like '%usd%' or hbdm like '%$%')""", {"f": chydh})
        dr = run_sql("""select ifnull(sum(seje),0) v from fysqsheet
                         where wxfp=:f and cwsp='通过' and sfrbmfy='否'
                           and (hbdm like '%RMB%' or hbdm like '%rmb%')""", {"f": chydh})
        fyUSD = F((du[0] if du else {}).get("v"))
        fyRMB = F((dr[0] if dr else {}).get("v"))
        ywfy = fyRMB + (fyUSD * yhhl if yhhl > 0 else 0)

    # ywdz/ywxz/zzsw
    tags = _kptz_get_yw_tags(S(b.get("ywry")), S(b.get("ywbm")))
    ywdz = tags.get("ywdz", "")
    ywxz = tags.get("ywxz", "")

    # cyrq：优先 kaiptz(cpmc=正常) chrq，否则 cymx.zgrqj
    cyrq = ""
    kq = run_sql("select chrq from kaiptz where fphm=:f and cpmc='正常' and ifnull(chrq,'')<>'' limit 1", {"f": bh})
    if kq:
        cyrq = S(kq[0].get("chrq"))
    if not cyrq:
        cyrq = S(c.get("zgrqj"))

    # ywry123/ywbm123
    ywry123 = ywbm123 = ""
    if cy_rid:
        yy = run_sql("""
            select ywry,wxbm1 from cymxsheet
             where pid=:pid and ifnull(chxs,0)>1
             order by chxs desc limit 1
        """, {"pid": cy_rid})
        if yy:
            ywry123 = S(yy[0].get("ywry"))
            ywbm123 = S(yy[0].get("wxbm1"))

    # 采购拆分
    tshj0=tshj1=tshj3=tshj6=tshj10=tshj13=tshj16=0.0
    ywhjr=ywhjm=nbhjr=nbhjm=0.0
    if cy_rid:
        cs = run_sql("""
            select ifnull(mlsb,'') mlsb,ifnull(tsl,0) tsl,ifnull(zzsl,0) zzsl,
                   ifnull(cgdq,'') cgdq,ifnull(cghbdm,'') cghbdm,
                   ifnull(sum(gczj),0) gczj1,ifnull(sum(gczjrmb),0) gczjrmb1
              from cymxsheet
             where pid=:pid
             group by mlsb,tsl,zzsl,cgdq,cghbdm
        """, {"pid": cy_rid}) or []
        for r in cs:
            mlsb = S(r.get("mlsb"))
            tsl = F(r.get("tsl"))
            zzsl = _i(r.get("zzsl"))
            cgdq = S(r.get("cgdq"))
            cghbdm = S(r.get("cghbdm")).upper()
            gczj1 = F(r.get("gczj1"))
            gczjrmb1 = F(r.get("gczjrmb1"))
            gczj_used = trunc10(gczj1) if mlsb == '是' else gczj1
            gczjrmb_used = trunc10(gczjrmb1) if mlsb == '是' else gczjrmb1

            if zzsl == 0:
                is_usd = ('USD' in cghbdm) or ('$' in cghbdm)
                if cgdq == '义乌':
                    ywhjm += gczj_used if is_usd else 0
                    ywhjr += gczj_used if not is_usd else 0
                else:
                    nbhjm += gczj_used if is_usd else 0
                    nbhjr += gczj_used if not is_usd else 0

            if tsl == 16: tshj16 += gczjrmb_used
            elif tsl == 13: tshj13 += gczjrmb_used
            elif tsl == 10: tshj10 += gczjrmb_used
            elif tsl == 6: tshj6 += gczjrmb_used
            elif tsl == 3: tshj3 += gczjrmb_used
            elif tsl == 1: tshj1 += gczjrmb_used
            elif tsl == 0 and zzsl > 1: tshj0 += gczjrmb_used

    ywxj = ywhjr + (ywhjm * yhhl if yhhl > 0 else 0)
    nbxj = nbhjr + (nbhjm * yhhl if yhhl > 0 else 0)
    xjhj = ywxj + nbxj
    cghjzje = tshj0 + tshj16 + tshj13 + tshj10 + tshj6 + tshj3 + tshj1 + ywxj + nbxj
    cbzje = cghjzje + yzaf + F(c.get("qtrmb"))

    ts16 = round((tshj16 / 1.16 * 0.16), 2) if tshj16 > 0 else 0
    if ts16 > 0:
        ts13 = round((tshj13 / 1.16 * 0.13), 2) if tshj13 > 0 else 0
        ts10 = round((tshj10 / 1.16 * 0.10), 2) if tshj10 > 0 else 0
        ts6 = round((tshj6 / 1.16 * 0.06), 2) if tshj6 > 0 else 0
    else:
        ts13 = round((tshj13 / 1.13 * 0.13), 2) if tshj13 > 0 else 0
        ts10 = round((tshj10 / 1.13 * 0.10), 2) if tshj10 > 0 else 0
        ts6 = round((tshj6 / 1.13 * 0.06), 2) if tshj6 > 0 else 0
    ts3 = round((tshj3 / 1.03 * 0.03), 2) if tshj3 > 0 else 0
    ts1 = round((tshj1 / 1.01 * 0.01), 2) if tshj1 > 0 else 0
    tszje = ts13 + ts10 + ts6 + ts3 + ts16 + ts1

    bgje = F(b.get("bgje"))
    CNFyf = F(b.get("CNFyf"))
    xbfy = (F(c.get("htje")) - F(c.get("ysdj"))) * F(c.get("xbfl"))
    if sbz != '1':
        xbfy = 0

    ayjje = F(c.get("ayjje"))
    hyf = F(c.get("hyf"))
    kpfyz = F(c.get("kpfyz"))
    jklx = 0
    xjzk = 0

    if S(b.get("RMBkh")) == "是":
        hsml = round(shhjz - hyf * (yhhl if yhhl > 0 else 0) - xbfy1 - cbzje + tszje - ayjje - jklx - kpfyz - ywfy + xjzk, 2)
        yghsml = round(F(c.get("htje")) - hyf * (yhhl if yhhl > 0 else 0) - xbfy - cbzje + tszje - ayjje - jklx - kpfyz - ywfy + xjzk, 2)
        mll = round((hsml / F(c.get("htje")) * 100), 2) if F(c.get("htje")) != 0 else 0
        ygmll = round((yghsml / F(c.get("htje")) * 100), 2) if F(c.get("htje")) != 0 else 0
    else:
        hsml = round(shhjz - (ayjje + hyf) * (yhhl if yhhl > 0 else 0) - cbzje - xbfy1 - ywfy + tszje - jklx - kpfyz + xjzk, 2)
        yghsml = round((F(c.get("htje")) - ayjje - hyf - xbfy) * (yhhl if yhhl > 0 else 0) - cbzje - ywfy + tszje - jklx - kpfyz + xjzk, 2)
        den = F(c.get("htje")) * (yhhl if yhhl > 0 else 0)
        mll = round((hsml / den * 100), 2) if den != 0 else 0
        ygmll = round((yghsml / den * 100), 2) if den != 0 else 0

    shz_dollar = round((djjem + (djjer / yhhl if yhhl > 0 else 0)), 2) if yhhl > 0 else 0
    htjez1 = round(F(c.get("htje")) / yhhl, 2) if (S(b.get("RMBkh")) == '是' and yhhl > 0) else F(c.get("htje"))
    bgje1 = round(bgje / yhhl, 2) if (S(b.get("RMBkh")) == '是' and yhhl > 0) else bgje
    dfrq = S(c.get("dfrq"))
    if not dfrq:
        dfrq = None
    if not xbdrsj:
        xbdrsj = None
    vals = {
        "rid": get_uuid(), "uid": user.rid, "ctime": now_dt, "mtime": now_dt,
        "wxfp": bh, "wxfp1": hsfp, "hsfp": hsfp, "fphm": hsfp,
        "khmc": S(b.get("khmc")), "RMBkh": S(b.get("RMBkh")), "bggs": S(b.get("bggs")),
        "ywry": ywry123 or S(b.get("ywry")), "ywbm": ywbm123 or S(b.get("ywbm")),
        "zdry": S(b.get("zdry")), "dlmc": S(b.get("dlmc")), "zwpmz": S(b.get("zwpmz")),
        "bglx": S(c.get("bglx")), "fpbz": S(c.get("fpbz12")),
        "sfnm": "是" if S(c.get("bglx")) == "内销" else "否",

        "bgje": bgje, "bgje1": bgje1, "CNFyf": CNFyf,
        "yhhl": yhhl, "djjer": djjer, "djjem": djjem, "shhjz": shhjz,
        "fyUSD": fyUSD, "fyRMB": fyRMB, "ywfy": ywfy,
        "fxhj": xjhj, "xjhj": xjhj, "ywhj": ywxj, "nbhj": nbxj, "ywxj": ywxj, "nbxj": nbxj,
        "ywhjr": ywhjr, "ywhjm": ywhjm, "nbhjr": nbhjr, "nbhjm": nbhjm,
        "tshj0": tshj0, "tshj16": tshj16, "tshj13": tshj13, "tshj10": tshj10, "tshj6": tshj6, "tshj3": tshj3, "tshj1": tshj1,
        "tszje": tszje, "cghjzje": cghjzje, "cbzje": cbzje,
        "yzaf": (yzaf - sjf - zgfy - ckfy), "zgfy": zgfy, "ckfy": ckfy, "sjf": sjf, "sjf1": sjf1,
        "ysdj": F(c.get("ysdj")), "xybx": S(c.get("xybx")), "xbfl": F(c.get("xbfl")),
        "htje": F(c.get("htje")), "htjer": F(c.get("htjer")), "htjem": F(c.get("htjem")),
        "ayjje": ayjje, "myjje": F(c.get("myjje")), "yjje": F(c.get("yjje")),
        "hyf": hyf, "ygnlf": F(c.get("ygnlf")), "kpfyz": kpfyz, "qtrmb": F(c.get("qtrmb")),
        "ewhj": F(c.get("ewhj")), "jxUSD": F(c.get("jxUSD")), "jxRMB": F(c.get("jxRMB")), "jxKHRMB": F(c.get("jxKHRMB")),
        "xbfy": xbfy1, "ygxbfy": round(xbfy, 2),
        "hsml": hsml, "mll": mll, "yghsml": yghsml, "ygmll": ygmll,
        "bgdh": S(b.get("bgdh")), "hdfyxq": cwqrqk, "xbdrsj": xbdrsj,
        "khpk": khpk, "gcpk": gcpk,
        "cyrq": cyrq, "jgtk": S(c.get("jgtk")), "jhfs": S(c.get("jhfs")), "shqx": S(c.get("shqx")),
        "dfrq": dfrq, "mdka": S(c.get("mdka")), "cdmc": S(c.get("cdmc")), "mygb": S(c.get("mygb")),
        "ywdz": ywdz, "ywxz": ywxz,
        "qesh": "否", "hkjq": "否", "shsd": "否", "tssd": "否", "fpsq": "否", "cysd": "否", "cygg": "否", "sfjd": "否",
        "shz_dollar": shz_dollar, "htjez1": htjez1,
        "jjxje": jjxje, "shrq1": shrq1, "shje": shje,
        "wxce": round(F(c.get("htje")) - shje, 2), "JGD": jgd
    }

    obj = fpgl()
    for k, v in vals.items():
        if hasattr(obj, k):
            setattr(obj, k, v)
    db_session.add(obj)

    return {"rid": obj.rid, "hsml": hsml, "skipped": False}

def _kptz_append_duplicate_log(wyzd: str, username: str = ""):
    try:
        # 写入 tmp 目录，不再创建 C:\417
        os.makedirs(config.tmp_path, exist_ok=True)
        fn = os.path.join(config.tmp_path, '开票重复记录.txt')
        ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        with open(fn, 'a', encoding='utf-8') as f:
            f.write(f'{ts}\t{username}\t{wyzd}\n')
        return fn
    except Exception:
        return None

def _kptz_sync_mlsb_backwrite(fphm: str, s, now_dt):
    """
    按 Pascal 口径联动回写 cymxsheet.mlsb：
    1) 工厂在 bmlgc => mlsb='否'
    2) 不在 bmlgc => mlsb='是'
    3) 若 gczj <= 抹零金额阈值 => 强制 mlsb='否'
    """
    # 抹零金额阈值
    ml_limit = 0.0
    d_ml = run_sql("select sz from cyzglsheet where xm='公司' and zm='抹零金额' limit 1")
    if d_ml:
        try:
            ml_limit = float(d_ml[0].get('sz') or 0)
        except:
            ml_limit = 0.0

    rows = run_sql("""
        select rid,
               ifnull(kpgc,'') as kpgc,
               ifnull(gcmc,'') as gcmc,
               ifnull(gczj,0) as gczj
          from cymxsheet
         where fphm=:f
    """, {"f": fphm}) or []

    for r in rows:
        rid = str(r.get('rid') or '')
        if not rid:
            continue

        sccj2 = (str(r.get('kpgc') or '').strip() or str(r.get('gcmc') or '').strip())
        gczj = 0.0
        try:
            gczj = float(r.get('gczj') or 0)
        except:
            gczj = 0.0

        ex = run_sql("select rid from bmlgc where gcmc=:g limit 1", {"g": sccj2})
        mlsb = '否' if ex else '是'
        if gczj <= ml_limit:
            mlsb = '否'

        s.query(cymxsheet).filter(cymxsheet.rid == rid).update({
            "mlsb": mlsb,
            "mtime": now_dt
        }, synchronize_session=False)

def _kptz_get_yw_tags(ywry: str, ywbm: str) -> dict:
    """
    统一补齐 ywdz / ywxz / zzsw
    Pascal口径：ywbdzb 按 ywb 查询，字段 ywdz / ywzjc / dyywzm
    """
    ywbm = _s(ywbm)
    out = {"ywdz": "", "ywxz": "", "zzsw": ""}

    d1 = []
    if ywbm:
        d1 = run_sql("""
            select ifnull(ywdz,'') ywdz, ifnull(ywzjc,'') ywxz, ifnull(dyywzm,'') zzsw
              from ywbdzb
             where ywb=:ywb
             limit 1
        """, {"ywb": ywbm}) or []

    if d1:
        out["ywdz"] = _s(d1[0].get("ywdz"))
        out["ywxz"] = _s(d1[0].get("ywxz"))
        out["zzsw"] = _s(d1[0].get("zzsw"))

    return out

def _kptz_norm_zzsw(zzsw: str, zzsl_val) -> str:
    z = _s(zzsw)
    if z:
        return z
    return '免税' if _i(zzsl_val) == 0 else '征税'

def _norm_zzsw(v, zzsl_val):
    t = _s(v)
    if t:
        return t
    return '免税' if _i(zzsl_val) == 0 else '征税'

def _pascal_wm_text(zsl_val, zzsl_val, kp3_val, kpds_val):
    """
    Pascal口径（最终）:
    - 免税: zzsl=0 或税点<=0
    - zsl=3 用 kp3
    - 其它用 kpds
    """
    zz = _i(zzsl_val, 0)
    if zz == 0:
        return 'WM免税'
    zsl = _f(zsl_val, 0.0)
    rate = _f(kp3_val, 0.0) if abs(zsl - 3.0) < 1e-9 else _f(kpds_val, 0.0)
    if rate <= 0:
        return 'WM免税'
    return f"WM{rate:g}税点"

def _strip_wm_text(v):
    s = _s(v)
    s = re.sub(r'WM免税', '', s, flags=re.I)
    s = re.sub(r'WM[\d.]+税点', '', s, flags=re.I)
    return s.strip(' ;；，,')

def _kptz_calc_yqrq_pascal(cpmc: str, main_zgrq: str, has_zgrqj: bool, yq3_val, now_dt: datetime) -> str:
    """
    Pascal口径：
    1) 先以 bgmxd.zgrq 为基准 + yq3
    2) 若为“补报”且存在 zgrqj，则改为 zgrq+10 再 +yq3
    3) 无有效日期则 now + yq3
    """
    yq_days = _i(yq3_val, 30)
    base_dt = None

    try:
        zgrq_s = _s(main_zgrq)
        if zgrq_s:
            base_dt = datetime.strptime(zgrq_s[:10], '%Y-%m-%d')
    except Exception:
        base_dt = None

    if base_dt is None:
        base_dt = now_dt

    if _s(cpmc) == '补报' and has_zgrqj:
        try:
            zgrq_s = _s(main_zgrq)
            if zgrq_s:
                base_dt = datetime.strptime(zgrq_s[:10], '%Y-%m-%d') + timedelta(days=10)
        except Exception:
            pass

    return (base_dt + timedelta(days=yq_days)).strftime('%Y-%m-%d')

def _fpgl_update_safe(db_session, fp_rid, values: dict):
    if not db_session or not fp_rid:
        return
    safe = {}
    for k, v in (values or {}).items():
        if hasattr(fpgl, k):
            safe[k] = v
    if safe:
        db_session.query(fpgl).filter(fpgl.rid == fp_rid).update(safe, synchronize_session=False)

def _kptz_calc_yqrq(cpmc: str, fphm: str, yq3_val, now_dt: datetime) -> str:
    yq_days = _i(yq3_val, 30)
    cy = run_sql("""
        select ifnull(zgrq,'') zgrq, ifnull(zgrqj,'') zgrqj
          from cymx where fphm=:f limit 1
    """, {"f": _s(fphm)}) or []

    zgrq = _s((cy[0] if cy else {}).get('zgrq'))
    zgrqj = _s((cy[0] if cy else {}).get('zgrqj'))
    base_dt = None

    try:
        if _s(cpmc) == '补报':
            # Pascal口径：仅当 zgrqj 非空时，优先使用 zgrq+10
            if zgrqj:
                if zgrq:
                    base_dt = datetime.strptime(zgrq[:10], '%Y-%m-%d') + timedelta(days=10)
                else:
                    base_dt = datetime.strptime(zgrqj[:10], '%Y-%m-%d')
            else:
                # zgrqj 为空时沿用原基准（不+10）
                if zgrq:
                    base_dt = datetime.strptime(zgrq[:10], '%Y-%m-%d')
        else:
            if zgrq:
                base_dt = datetime.strptime(zgrq[:10], '%Y-%m-%d')
    except Exception:
        base_dt = None

    if base_dt is None:
        base_dt = now_dt
    return (base_dt + timedelta(days=yq_days)).strftime('%Y-%m-%d')

def _pascal_wm_sccj(zsl_val, zzsl_val, kp3_val, kpds_val):
    zsl_num = _f(zsl_val, 0.0)
    zz = _i(zzsl_val, 0)
    if zz == 0:
        return "WM免税"
    rate = _f(kp3_val, 0.0) if abs(zsl_num - 3.0) < 1e-9 else _f(kpds_val, 0.0)
    if rate <= 0:
        return "WM免税"
    zsl_txt = str(int(zsl_num)) if abs(zsl_num - int(zsl_num)) < 1e-9 else f"{zsl_num:g}"
    return f"WM{zsl_txt}%*{rate:g}"

@any_route('/api/saier/declaration/button/kptz/check', methods=['POST'])
@require_token
async def view_saier_declaration_button_kptz_check(request):
    j = await request.json()
    user = request.current_user
    s = Session()
    try:
        rid = _s(j.get('rid'))
        bg = s.query(bgmxd).filter(bgmxd.rid == rid).first()
        if not bg:        
            return json_result(-1, '报关明细记录未找到')
        bg = alchemy_object_to_dict(bg)
        sbdate = _s(bg.get('sbdate'))
        ysfp = _s(bg.get('ysfp'))
        ywrya = _s(bg.get('ywrya'))
        
        return json_result(0, '取数成功',{'sbdate': sbdate, 'ysfp': ysfp, 'ywrya': ywrya})
    except Exception as e:
        return json_result(-1, f'接口异常: {str(e)}', {})
    finally:
        s.close()

@any_route('/api/saier/declaration/button/kptz/run_v2', methods=['POST'])
@require_token
async def view_saier_declaration_button_kptz_run_v2(request):
    import os, math
    from collections import defaultdict
    from openpyxl import Workbook
    from openpyxl.styles import Font

    j = await request.json()
    user = request.current_user
    s = Session()
    now = datetime.now()

    def _k5(a, b, c, d, e): return f"{_s(a)}|{_s(b)}|{_s(c)}|{_s(d)}|{_s(e)}"
    def _set_if_has(obj, data):
        for k, v in data.items():
            if hasattr(obj, k):
                setattr(obj, k, v)

    rid = _s(j.get('rid'))
    bg = s.query(bgmxd).filter(bgmxd.rid == rid).first()
    if not bg:        
        return json_result(-1, '报关明细记录未找到')
    bg = alchemy_object_to_dict(bg)
    sbdate = _s(j.get('sbdate'))
    gsfprq = _s(j.get('gsfprq'))
    ysfp_in = _s(bg.get('ysfp'))
    ywrya_in = _s(bg.get('ywrya'))
    if not rid:
        return json_result(1, '缺少rid', {})
    
    if not gsfprq:
        gsfprq = None

    warnings = []
    duplicated = []
    ytxxck = []
    download_filename = ''

    try:
        # ---------- 0. 参数与权限 ----------
        def _zx(ly, dv=''):
            r = run_sql("select cs from zx where ly=:ly limit 1", {"ly": ly})
            return _s(r[0].get('cs')) if r and _s(r[0].get('cs')) else dv

        kp3 = _zx('开票点数3', '3.39')
        htsh = _zx('合同收回金额', '')
        yq3 = _zx('开票逾期', '30')

        run_sql("""
            update kaiptzsheet1
               set fpyq='是'
             where ifnull(yqrq,'')<>'' and yqrq<:d and ifnull(fpyq,'')=''
        """, {"d": now.strftime('%Y-%m-%d')})

        cu = run_sql("select rid from sys_user where username=:u limit 1", {"u": user.username})
        if not cu:
            return json_result(1, '当前用户不存在', {})
        uid = _s(cu[0].get('rid'))
        org = get_user_path(user.username)
        path = org.get('path','')[:100]
        postion = org.get('position', '')

        # 角色：财务 + 优景
        ur = run_sql("select role_id from sys_user_role where user_id=:uid", {"uid": uid}) or []
        ok_role = False
        for x in ur:
            role_id = _s(x.get('role_id'))
            if not role_id:
                continue
            rr = run_sql("""
                select rid from sys_role
                where rid=:rid and name like '%财务%' 
                limit 1
            """, {"rid": role_id})
            # and path like '%优景%'
            if rr:
                ok_role = True
                break
        if not ok_role:
            return json_result(1, '非跨境财务不能生成跨境开票通知', {})

        # ---------- 1. 主单/出运校验 ----------
        main = run_sql("""
            select rid,fphm,ysfp,khmc,ywbm,ywry,RMBkh,bggs,sbdate,htrq,zgrq,
                   ifnull(bgbgzje,0) as bgje
              from bgmxd where rid=:rid limit 1
        """, {"rid": rid})
        if not main:
            return json_result(1, '报关主单不存在', {})
        m = main[0]
        wxfp = _s(m.get('fphm'))
        ysfp = ysfp_in or _s(m.get('ysfp'))
        bgje = _f(m.get('bgje'))

        # 业务标签（优先 ywbdzb，回退 cyzglsheet）
        yw_tags = {"ywdz": "", "ywxz": "", "zzsw": ""}

        fp3 = wxfp[:3].upper()
        if fp3 != 'AMZ':
            t1 = run_sql("""
                select ifnull(ywdz,'') ywdz, ifnull(ywzjc,'') ywxz, ifnull(dyywzm,'') zzsw
                from ywbdzb where ywb=:ywb limit 1
            """, {"ywb": _s(m.get('ywbm'))}) or []
            if t1:
                yw_tags["ywdz"] = _s(t1[0].get("ywdz"))
                yw_tags["ywxz"] = _s(t1[0].get("ywxz"))
                yw_tags["zzsw"] = _s(t1[0].get("zzsw"))
        else:
            t1 = run_sql("""
                select ifnull(qxzl,'') qxzl, ifnull(xm,'') xm
                from cyzglsheet where bz=:bz and zm='跨境电商' limit 1
            """, {"bz": _s(m.get('khmc'))}) or []
            if t1:
                qxzl = _s(t1[0].get("qxzl"))
                xm = _s(t1[0].get("xm"))
                yw_tags["ywdz"] = qxzl + xm
                yw_tags["ywxz"] = xm

        cy = run_sql("select zgrq,bglx,fpbz12,sjcy1 from cymx where fphm=:fphm limit 1", {"fphm": wxfp})
        if not cy or _s(cy[0].get('zgrq')) == '':
            return json_result(1, '开票通知不能生成，请先通知单证填写装柜日期或进仓回执', {})

        # Pascal 弹窗语义：改为要求前端补录再重提
        sbdate_final = sbdate or _s(m.get('sbdate'))
        need = []
        if not sbdate_final:
            need.append("sbdate")
        if ('义乌财务' in postion) and not _s(gsfprq):
            need.append("gsfprq")
        if need:
            return json_result(1001, '需要补充输入', {
                "required": need,
                "defaults": {"sbdate": sbdate_final, "gsfprq": _s(gsfprq)}
            })

        # AMZ/UVVC 口径（可配置）
        amz_cfg = run_sql("select ifnull(bz,'') p from cyzglsheet where zm='KPTZ_AMZ_ALLOW_PATH'") or []
        amz_paths = [_s(x.get("p")) for x in amz_cfg if _s(x.get("p"))]
        uvvc_cfg = run_sql("select ifnull(bz,'') p from cyzglsheet where zm='KPTZ_UVVC_ALLOW_PATH'") or []
        uvvc_paths = [_s(x.get("p")) for x in uvvc_cfg if _s(x.get("p"))]

        fp3 = wxfp[:3].upper()
        yb2 = _s(m.get('ywbm'))[:2].upper()

        if fp3 == 'AMZ':
            if amz_paths:
                if not any(p in path for p in amz_paths):
                    return json_result(1, '非跨境财务不能生成跨境开票通知', {})
            else:
                if '宁波景业国际贸易有限公司' not in path:
                    return json_result(1, '非跨境财务不能生成跨境开票通知', {})

        if yb2 in ('UV', 'VC') and uvvc_paths:
            if not any(p in path for p in uvvc_paths):
                return json_result(1, '当前账号无UV/VC开票权限', {})

        # ---------- 2. 三分支构建 kaiptz/kaiptzsheet ----------
        dims = run_sql("""
            select zwpm,hyd,hgbm,bgbh
              from bgmxdhbcp
             where pid=:pid and ifnull(zwpm,'')<>''
             group by zwpm,hyd,hgbm,bgbh
        """, {"pid": rid}) or []

        all_new_head_rids = []

        def _combo_key(r):
            return f"{_s(r.get('kpgc'))}-{_s(r.get('gdry'))}-{_s(r.get('hyd'))}-{_s(r.get('hgbm'))}-{_s(r.get('bgbh'))}"

        for d in dims:
            for bgtj in (1, 2, 3):
                sql = """
                    select *
                      from bgmxdsheet
                     where pid=:pid
                       and ifnull(sfdb,'')<>'是'
                       and ifnull(zzsl,0)<>0
                       and ifnull(zhwbgpm,'')=:zwpm
                       and ifnull(hyd,'')=:hyd
                       and ifnull(hgbm,'')=:hgbm
                       and ifnull(bgbh,'')=:bgbh
                """
                p = {"pid": rid, "zwpm": _s(d.get('zwpm')), "hyd": _s(d.get('hyd')), "hgbm": _s(d.get('hgbm')), "bgbh": _s(d.get('bgbh'))}
                if bgtj == 1:
                    sql += " and ifnull(yfph,'')<>'' "
                elif bgtj == 2:
                    sql += " and ifnull(yfph,'')='' and ifnull(ewchy,'')<>'是' "
                else:
                    sql += " and ifnull(yfph,'')='' and ifnull(ewchy,'')='是' "

                rs = run_sql(sql, p) or []
                if not rs:
                    continue

                grp = defaultdict(list)
                for r in rs:
                    grp[_combo_key(r)].append(r)

                for ck, arr in grp.items():
                    cpmc = '补报' if bgtj == 1 else ('正常' if bgtj == 2 else '非正')
                    wyzd = f"{wxfp}|{_s(d.get('zwpm'))}|{_s(d.get('hyd'))}|{_s(d.get('hgbm'))}|{_s(d.get('bgbh'))}|{ck}|{cpmc}"

                    mid = f"{_s(d.get('zwpm'))}{_s(d.get('hyd'))}{_s(d.get('hgbm'))}{_s(d.get('bgbh'))}{ck}"
                    if bgtj == 1:  # 补报：Pascal历史串中段重复
                        wyzd_legacy = f"{wxfp}{mid}{mid}{cpmc}"
                    else:
                        wyzd_legacy = f"{wxfp}{mid}{cpmc}"
                        
                    ex = run_sql("""
                        select rid from kaiptz
                        where wyzd=:w1 or wyzd=:w2
                        limit 1
                    """, {"w1": wyzd, "w2": wyzd_legacy})

                    if ex:
                        duplicated.append(wyzd)
                        try:
                            _kptz_append_duplicate_log(wyzd, user.username)
                        except:
                            pass
                        continue
                    
                    logger.error('==================11111111111111======')
                    h = kaiptz()
                    _set_if_has(h, {
                        "rid": get_uuid(), "uid": user.rid, "ctime": now, "mtime": now,
                        "pid": rid, "fphm": wxfp, "ysfp": ysfp, "khmc": _s(m.get('khmc')),
                        "ywbm": _s(m.get('ywbm')), "ywry": _s(m.get('ywry')), "RMBkh": _s(m.get('RMBkh')),
                        "bggs": _s(m.get('bggs')), "path": path, "wyzd": wyzd, "cpmc": cpmc,
                        "kprq": now.strftime('%Y-%m-%d'), "gsfprq": gsfprq,
                        "ywdz": yw_tags.get("ywdz", ""),
                        "ywxz": yw_tags.get("ywxz", ""),
                        "zzsw": yw_tags.get("zzsw", "")
                    })
                    s.add(h)

                    slhj = xshj = cgze = wxhj = yfhj = ysqhj = 0.0
                    ytsb = ''

                    for r in arr:
                        kd = kaiptzsheet()
                        wxje = _f(r.get('mjzj')) if _s(m.get('RMBkh')) == '是' else _f(r.get('wxzj'))
                        row_zzsw = _s(r.get('zzsw')) or yw_tags.get("zzsw", "") or ('免税' if _i(r.get('zzsl')) == 0 else '征税')

                        _set_if_has(kd, {
                            "rid": get_uuid(), "uid": user.rid, "ctime": now, "mtime": now, "pid": h.rid,
                            "fphm": wxfp, "ysfp": _s(r.get('ysfp')), "wxfp": _s(r.get('fphm')),
                            "zwpm": _s(r.get('zhwbgpm')), "zwpm1": _s(r.get('zwpm')),
                            "hyd": _s(r.get('hyd')), "hgbm": _s(r.get('hgbm')), "bgbh": _s(r.get('bgbh')),
                            "chsl": _f(r.get('chsl')), "chxs": _f(r.get('chxs')), "gczj": _f(r.get('gczj')),
                            "wxzj": _f(r.get('wxzj')), "mjzj": _f(r.get('mjzj')), "wxje": wxje,
                            "zjz": _f(r.get('zjz')), "zsl": _s(r.get('tsl')), "zzsl": _i(r.get('zzsl')),
                            "jldw": _s(r.get('jldw')), "hgjldw": _s(r.get('hgjldw')),
                            "hbdm": 'RMB' if _s(m.get('RMBkh')) == '是' else 'USD$',
                            "cghbdm": _s(r.get('cghbdm')), "wxbm1": _s(r.get('wxbm1')),
                            "sccj": _s(r.get('gcmc')), "sccj1": _s(r.get('kpgc')) or _s(r.get('gcmc')),
                            "dq1": _s(r.get('dq1')) or '宁波',
                            "ywrya": _s(r.get('gdry')), "cgry": _s(r.get('ywrya')), "ywry": _s(r.get('ywry')),
                            "gdry": _s(r.get('gdry')), "cght": _s(r.get('cght')),
                            "kpds": _f(r.get('kpds')), "kpsl": _f(r.get('kpsl')),
                            "cpbh": _s(r.get('cpbh')), "wxrl": _f(r.get('wxrl')),
                            "fkxq": _s(r.get('fkxq')), "szxq": _s(r.get('szxq')), "zkfy1": _f(r.get('zkfy1')),
                            "wxwyzd": _s(r.get('wxwyzd')), "yfkp": _s(r.get('yfkp')),
                            "cywyzd": _s(r.get('cywyzd')), "fktt": _s(r.get('fktt')), "htrq": _s(r.get('htrq')),
                            "ywdz": yw_tags.get("ywdz", ""), "ywxz": yw_tags.get("ywxz", ""), "zzsw": row_zzsw
                        })

                        scrq = _s(r.get('scrq'))
                        jsfs = _s(r.get('jsfs'))
                        bz1 = _s(r.get('bz1'))
                        zzsl_i = _i(r.get('zzsl'))
                        wm_tax = f"WM{zzsl_i}税点" if zzsl_i > 0 else "WM税点"
                        
                        if bgtj == 3:
                            # Pascal原逻辑：
                            # scrq有值: "进仓日期{scrq};{bz1}结算方式{jsfs}"
                            # scrq无值: "{bz1}结算方式{jsfs}"
                            bz1_raw = _s(r.get('bz1'))
                            if scrq:
                                bzsm_txt = f"进仓日期{scrq};{bz1_raw}结算方式{jsfs}"
                            else:
                                bzsm_txt = f"{bz1_raw}结算方式{jsfs}"

                            _set_if_has(kd, {
                                "bzsm": bzsm_txt,
                                "bzsm1": "",
                                "bzsm2": bzsm_txt,
                                "sccj1": _s(r.get('kpgc')) or _s(r.get('gcmc'))
                            })
                        else:
                            parts = []
                            if scrq:
                                parts.append(f"进仓日期{scrq}")
                            if jsfs:
                                parts.append(f"结算方式{jsfs}")
                            kd.bzsm = ';'.join(parts)

                        if bgtj == 1:
                            kd.yfph = _s(r.get('yfph'))
                            kd.yysfp = _s(r.get('yysfp'))
                            kd.bzsm1 = f"补报{kd.yfph}中{_f(r.get('gczj'))}元"

                        ytbh = ''
                        cght = _s(r.get('cght'))
                        if cght:
                            y1 = run_sql("""select ytbh from ytfysheet where cght=:c and sccj=:s and hyd=:h and zzsl=:z limit 1""",
                                         {"c": cght, "s": _s(r.get('gcmc')), "h": _s(r.get('hyd')), "z": _i(r.get('zzsl'))})
                            y2 = run_sql("""select ytbh from dsytfysheet where cght=:c and sccj=:s and hyd=:h and zzsl=:z limit 1""",
                                         {"c": cght, "s": _s(r.get('kpgc')) or _s(r.get('gcmc')), "h": _s(r.get('hyd')), "z": _i(r.get('zzsl'))})
                            if y1:
                                ytbh = _s(y1[0].get('ytbh')); ytsb = '1'; ytxxck.append(f"工厂{_s(r.get('gcmc'))}预填提醒")
                            if y2:
                                ytbh = _s(y2[0].get('ytbh')) or ytbh; ytsb = '1'; ytxxck.append(f"工厂{_s(r.get('kpgc')) or _s(r.get('gcmc'))}预填提醒")
                        _set_if_has(kd, {"ytbh": ytbh})
                        s.add(kd)

                        slhj += _f(r.get('chsl')); xshj += _f(r.get('chxs')); cgze += _f(r.get('gczj'))
                        wxhj += wxje; yfhj += _f(r.get('yfje')); ysqhj += _f(r.get('ysqje'))

                    s.query(kaiptz).filter(kaiptz.rid == h.rid).update({
                        "slhj": round(slhj, 4), "xshj": round(xshj, 4), "cgze": round(cgze, 4),
                        "wxhj": round(wxhj, 4), "yfhj": round(yfhj, 4), "ysqhj": round(ysqhj, 4),
                        "ytsb": "有预填" if ytsb == '1' else "", "mtime": now
                    }, synchronize_session=False)

                    all_new_head_rids.append(h.rid)

        # ---------- 3. 生成 kaiptzsheet1 ----------
        org = get_user_path('zjnblh')
        postion = org.get('position', '')
        bm1 = _s(postion)[:1] if _s(postion) else ''
        kpxh_prefix = f"{bm1}{now.strftime('%y%m')}"
        ml_limit = 0
        ml = run_sql("""select sz from cyzglsheet where xm='公司' and zm='抹零金额' limit 1""")
        if ml: ml_limit = _f(ml[0].get('sz'))

        for pid in all_new_head_rids:
            hinfo = run_sql("select ifnull(cpmc,'') cpmc from kaiptz where rid=:rid limit 1", {"rid": pid})
            cpmc_head = _s((hinfo[0] if hinfo else {}).get('cpmc'))

            rows = run_sql("select * from kaiptzsheet where pid=:pid", {"pid": pid}) or []
            grp = defaultdict(list)
            for r in rows:
                grp[_k5(r.get('zwpm'), r.get('zsl'), r.get('hyd'), r.get('hgbm'), r.get('bgbh'))].append(r)

            for _, arr in grp.items():
                a0 = arr[0]
                chsl = sum(_f(x.get('chsl')) for x in arr)
                gczj = sum(_f(x.get('gczj')) for x in arr)
                zjz = sum(_f(x.get('zjz')) for x in arr)
                wxje = sum(_f(x.get('wxje')) for x in arr)

                old = run_sql("select kpxh from kaiptzsheet1 where kpxh like :p order by kpxh desc limit 1", {"p": f"{kpxh_prefix}%"})
                if old and _s(old[0].get('kpxh')):
                    part = _s(old[0].get('kpxh'))[5:10]
                    seq = (int(part) + 1) if part.isdigit() else 1
                else:
                    seq = 1
                sec_ones = now.strftime('%S')[-1:]  # Pascal 第19位：秒个位
                kpxh = f"{kpxh_prefix}{seq:05d}{sec_ones}"

                zsl = _f(a0.get('zsl'))
                zzsl = 1.01 if zsl == 1 else 1.03
                tsl13 = 0.01 if zsl == 1 else 0.03

                sccj2 = _s(a0.get('sccj1')) or _s(a0.get('sccj'))
                gc50 = run_sql("select rid from bmlgc where gcmc=:g limit 1", {"g": sccj2})
                mlsb = '否' if gc50 else '是'
                if gczj <= ml_limit:
                    mlsb = '否'

                # 修复：补报按 yfph 分摊后再抹零汇总
                sjkpze = 0.0
                if cpmc_head == '补报':
                    yf_map = defaultdict(float)
                    for x in arr:
                        yf_key = _s(x.get('yfph')) or '__NO_YFPH__'
                        yf_map[yf_key] += _f(x.get('gczj'))
                    for _, val in yf_map.items():
                        if mlsb == '否':
                            sjkpze += round(val, 2)
                        else:
                            sjkpze += _kptz_trunc10(val)
                else:
                    sjkpze = round(gczj, 2) if mlsb == '否' else float(int(gczj / 10) * 10)

                cgdj1 = (sjkpze / zzsl / chsl) if (chsl and zzsl) else 0
                cgdj2 = (sjkpze / zzsl / zjz) if (zjz and zzsl) else 0
                gcjg = cgdj2 if _s(a0.get('hgjldw')) in ('KGS', '千克', '公斤', '克') else cgdj1
                bhsje = (sjkpze / zzsl) if zzsl else 0
                se = bhsje * tsl13 if (zzsl * tsl13) else 0
                row_zzsw = _norm_zzsw(_s(a0.get('zzsw')) or yw_tags.get("zzsw", ""), a0.get('zzsl'))

                yq_days = _i(yq3, 30)

                has_zgrqj = False
                try:
                    q_zgrqj = run_sql("""
                        select rid
                        from bgmxdsheet
                        where pid=:pid
                        and ifnull(zhwbgpm,'')=:zwpm
                        and ifnull(hyd,'')=:hyd
                        and ifnull(hgbm,'')=:hgbm
                        and ifnull(bgbh,'')=:bgbh
                        and ifnull(zgrqj,'')<>''
                        limit 1
                    """, {
                        "pid": rid,
                        "zwpm": _s(a0.get('zwpm')),
                        "hyd": _s(a0.get('hyd')),
                        "hgbm": _s(a0.get('hgbm')),
                        "bgbh": _s(a0.get('bgbh')),
                    }) or []
                    has_zgrqj = bool(q_zgrqj)
                except Exception:
                    has_zgrqj = False

                yqrq = _kptz_calc_yqrq_pascal(cpmc_head, _s(m.get('zgrq')), has_zgrqj, yq3, now)
                
                hyd_val = _s(a0.get('hyd'))
                yhyd = hyd_val
                hydgh = _s(a0.get('hydgh')) or hyd_val

                # Pascal口径近似：抹零差额
                if cpmc_head == '补报':
                    if mlsb == '否':
                        kplt = 0.0
                    else:
                        # Pascal：按整组 gczj 计算余数，不按 yfph 分组余数相加
                        kplt = round(max(_f(gczj) - _kptz_trunc10(_f(gczj)), 0), 2)
                else:
                    if mlsb == '否':
                        kplt = 0.0
                    else:
                        kplt = round(max(_f(gczj) - _f(sjkpze), 0), 2)

                xgzf = kplt

                # 合同收回标记（阈值判断）
                htsh_limit = _f(htsh, 0.0)
                htsh_flag = '待收回' if (htsh_limit > 0 and _f(sjkpze) >= htsh_limit) else '不需要'
                
                base_bzsm = _s(a0.get('bzsm'))
                k1_bzsm = base_bzsm
                k1_bzsm2 = base_bzsm
                k1_sccj = _s(a0.get('sccj'))

                if cpmc_head == '非正':
                    zsl_num = _f(a0.get('zsl'), 0.0)
                    zsl_txt = str(int(zsl_num)) if abs(zsl_num - int(zsl_num)) < 1e-9 else f"{zsl_num:g}"

                    kp3_txt = _s(kp3) or '0'
                    kpds_txt = _s(a0.get('kpds')) or '0'

                    if abs(zsl_num - 3.0) < 1e-9:
                        # zsl=3: bzsm始终追加 WM*kp3；sccj=WMzsl%*kp3
                        wm_tag = f"WM*{kp3_txt}"
                        k1_bzsm = f"{base_bzsm},{wm_tag}" if base_bzsm else wm_tag
                        k1_bzsm2 = base_bzsm
                        k1_sccj = f"WM{zsl_txt}%*{kp3_txt}"
                    else:
                        # zsl!=3:
                        # kpds='0' -> bzsm不追加
                        # kpds!='0' -> bzsm追加 WM*kpds
                        if kpds_txt != '0':
                            wm_tag = f"WM*{kpds_txt}"
                            k1_bzsm = f"{base_bzsm},{wm_tag}" if base_bzsm else wm_tag
                        else:
                            k1_bzsm = base_bzsm
                        k1_bzsm2 = base_bzsm
                        # Pascal里这里无论kpds是否为0，都写 WMzsl%*kpds
                        k1_sccj = f"WM{zsl_txt}%*{kpds_txt}"

                k1 = kaiptzsheet1()
                _set_if_has(k1, {
                    "rid": get_uuid(), "uid": user.rid, "ctime": now, "mtime": now, "pid": pid,
                    "fphm": wxfp, "ysfp": ysfp, "zwpm": _s(a0.get('zwpm')), "zwpm1": _s(a0.get('zwpm')),
                    "zsl": _s(a0.get('zsl')), "zzsl": _i(a0.get('zzsl')), "hyd": hyd_val,
                    "hgbm": _s(a0.get('hgbm')), "bgbh": _s(a0.get('bgbh')),
                    "hgjldw": ('千克' if _s(a0.get('hgjldw')) == 'KGS' else _s(a0.get('hgjldw'))),
                    "chsl": chsl, "bgsl": chsl, "bgsl1": chsl, "bgsl2": zjz,
                    "jldw": _s(a0.get('jldw')), "gczj": gczj, "jehj": gczj, "zjz": zjz,
                    "wxje": wxje, "hbdm": _s(a0.get('hbdm')), "cghbdm": _s(a0.get('cghbdm')),
                    "wxbm1": _s(a0.get('wxbm1'))[:200], "kpxh": kpxh, "kprq": now.strftime('%Y-%m-%d'),
                    "sjkpze": sjkpze, "cgdj1": cgdj1, "cgdj2": cgdj2, "gcjg": gcjg,
                    "bhsje": bhsje, "se": se, "sfkp": '否', "wckp": '否', "ykfp": '否', "sfdp": '否',
                    "mlsb": mlsb, "fkxq": _s(a0.get('fkxq')), "szxq": _s(a0.get('szxq')), "zkfy1": _f(a0.get('zkfy1')),
                    "yfph": _s(a0.get('yfph')), "yfph1": _s(a0.get('yfph')), "ysfp1": _s(a0.get('yysfp')),
                    "bzsm": k1_bzsm, "bzsm2": k1_bzsm2,
                    "ywrya": _s(a0.get('ywrya')), "cgry": _s(a0.get('cgry')), "lxdh": _s(a0.get('lxdh')),
                    "wstt": _s(m.get('bggs')), "sccj2": sccj2, "sccj": k1_sccj,
                    "ywdz": yw_tags.get("ywdz", ""), "ywxz": yw_tags.get("ywxz", ""), "zzsw": row_zzsw,
                    # 补齐字段（第2点）
                    "yqrq": yqrq, "fpyq": "", "htsh": htsh_flag, "bgje": bgje,
                    "kplt": kplt, "xgzf": xgzf, "spfl": _s(a0.get('spfl')),
                    "hydgh": hydgh, "yhyd": yhyd
                })
                s.add(k1)

        # ---------- 4. 差异检查 + cz反写 ----------
        bg = run_sql("""
            select zwpm,tsl,hyd,hgbm,bgbh,ifnull(cz,'') cz
              from bgmxdhbcp
             where pid=:pid and ifnull(zwpm,'')<>''
        """, {"pid": rid}) or []
        bg_keys, bg_cz = set(), {}
        for r in bg:
            k = _k5(r.get('zwpm'), r.get('tsl'), r.get('hyd'), r.get('hgbm'), r.get('bgbh'))
            bg_keys.add(k)
            if _s(r.get('cz')):
                bg_cz[k] = _s(r.get('cz'))

        ks = run_sql("select rid,zwpm,zsl,hyd,hgbm,bgbh from kaiptzsheet1 where fphm=:f", {"f": wxfp}) or []
        for r in ks:
            k = _k5(r.get('zwpm'), r.get('zsl'), r.get('hyd'), r.get('hgbm'), r.get('bgbh'))
            if k not in bg_keys:
                warnings.append(f"{k}(待补报或出运明细与报关明细品名,货源地,退税率,海关编码不一)")
            if k in bg_cz:
                s.query(kaiptzsheet1).filter(kaiptzsheet1.rid == r.get('rid')).update({"cz": bg_cz[k], "mtime": now}, synchronize_session=False)

        c_only = run_sql("""
            select ifnull(yfph,'') yfph,ifnull(gcmc,'') gcmc,zhwbgpm,tsl,hyd,hgbm,bgbh
              from bgmxdsheet
             where pid=:pid and ifnull(zhwbgpm,'')<>'' and ifnull(sfdb,'')<>'是'
        """, {"pid": rid}) or []
        for r in c_only:
            k = _k5(r.get('zhwbgpm'), r.get('tsl'), r.get('hyd'), r.get('hgbm'), r.get('bgbh'))
            if k not in bg_keys:
                warnings.append(f"{_s(r.get('yfph'))}{_s(r.get('gcmc'))}{k}(报关详情和报关合并品名,货源地,退税率,海关编码不一)")

        # ---------- 5. 海关编码多税率标记 ----------
        bad = run_sql("""
            select hgbm
              from kaiptzsheet1
             where fphm=:f
             group by hgbm
            having count(distinct zsl)>1
        """, {"f": wxfp}) or []
        if bad:
            hgbm_list = [x.get('hgbm') for x in bad if x.get('hgbm')]
            s.query(kaiptzsheet1).filter(kaiptzsheet1.fphm == wxfp, kaiptzsheet1.hgbm.in_(hgbm_list)).update({"bmts": "是", "mtime": now}, synchronize_session=False)
            warnings.append('请注意海关编码退税有异常')

        # ---------- 6. 导出（xlsx红字） ----------
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(['商品名称', '数量', '单位', '总价', '退税率', '货 源 地', '海关编码', '报关编号'])
        red = Font(color="FF0000")

        b = run_sql("""
            select zwpm,zsl,hyd,hgbm,bgbh,ifnull(sum(bgsl),0) bgsl,ifnull(sum(gczj),0) gczj,ifnull(max(hgjldw),'') hgjldw
              from kaiptzsheet1 where fphm=:f
             group by zwpm,zsl,hyd,hgbm,bgbh
        """, {"f": wxfp}) or []
        b_map = {_k5(x.get('zwpm'), x.get('zsl'), x.get('hyd'), x.get('hgbm'), x.get('bgbh')): x for x in b}

        for k in sorted(bg_keys):
            if k in b_map:
                x = b_map[k]
                ws.append([_s(x.get('zwpm')), _f(x.get('bgsl')), _s(x.get('hgjldw')), _f(x.get('gczj')), _s(x.get('zsl')), _s(x.get('hyd')), _s(x.get('hgbm')), _s(x.get('bgbh'))])
            else:
                ws.append([f"{k}(补报或出运明细与报关明细品名不一)", '', '', '', '', '', '', ''])
                ws[f"A{ws.max_row}"].font = red

        for k in sorted(set(b_map.keys()) - bg_keys):
            ws.append([f"{k}(待补报或出运明细与报关明细品名,货源地,退税率，海关编码不一)", '', '', '', '', '', '', ''])
            ws[f"A{ws.max_row}"].font = red

        os.makedirs(config.tmp_path, exist_ok=True)
        download_filename = f"{wxfp}汇总表.xlsx"
        xlsx_path = os.path.join(config.tmp_path, download_filename)
        wb.save(xlsx_path)

        # 将汇总表与重复记录（若存在）打包成 zip，供前端下载
        zip_name = f"{wxfp}汇总表.zip"
        zip_path = os.path.join(config.tmp_path, zip_name)
        try:
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                # 添加汇总表
                if os.path.exists(xlsx_path):
                    zf.write(xlsx_path, arcname=download_filename)
                # 添加重复记录（若有）
                dup_log_path = os.path.join(config.tmp_path, '开票重复记录.txt')
                if os.path.exists(dup_log_path):
                    zf.write(dup_log_path, arcname=os.path.basename(dup_log_path))
            # 返回 zip 文件名给前端
            download_filename = zip_name
        except Exception:
            # 若打包失败，保留 xlsx 可单独下载
            logger.error('打包汇总表失败：' + trace_error())
        
        # ---------- 7. fpgl（按 Pascal 范围：UV/VC 且非 AMZ） ----------
        up_fphm = (wxfp or '').upper()
        fp_prefix3 = up_fphm[:3]
        need_full_fpgl = (('UV' in up_fphm or 'VC' in up_fphm) and fp_prefix3 != 'AMZ')

        if need_full_fpgl:
            _kptz_upsert_fpgl_full(wxfp, ysfp, user, now, m, s)

            zs_sjkpze = zs_bhsje = zs_se = 0.0
            ms_sjkpze = ms_bhsje = ms_se = 0.0
            qt_sjkpze = qt_bhsje = qt_se = 0.0

            fp = run_sql("select rid from fpgl where wxfp=:w or hsfp=:h limit 1", {"w": wxfp, "h": ysfp or wxfp})
            fp_rid = _s((fp[0] if fp else {}).get('rid'))

            if not fp_rid:
                obj = fpgl()
                _set_if_has(obj, {
                    "rid": get_uuid(), "uid": user.rid, "ctime": now, "mtime": now,
                    "wxfp": wxfp, "wxfp1": ysfp or wxfp, "hsfp": ysfp or wxfp, "fphm": ysfp or wxfp
                })
                s.add(obj)
                fp_rid = obj.rid

            # # 采购拆分（补齐 ywhjr/ywhjm/nbhjr/nbhjm、fxhj、ts*）
            # split = run_sql("""
            #     select ifnull(mlsb,'') mlsb, ifnull(cgdq,'') cgdq, ifnull(cghbdm,'') cghbdm,
            #         ifnull(zsl,0) zsl, ifnull(sum(gczj),0) gczj, ifnull(sum(gczjrmb),0) gczjrmb
            #     from cymxsheet
            #     where fphm=:f
            #     group by ifnull(mlsb,''), ifnull(cgdq,''), ifnull(cghbdm,''), ifnull(zsl,0)
            # """, {"f": wxfp}) or []

            ywhjr = ywhjm = nbhjr = nbhjm = 0.0
            tshj0 = tshj1 = tshj3 = tshj6 = tshj10 = tshj13 = tshj16 = 0.0
            # for r in split:
            #     mlsb_v = _s(r.get('mlsb'))
            #     cgdq_v = _s(r.get('cgdq'))
            #     cghbdm_v = _s(r.get('cghbdm')).upper()
            #     zsl_v = _f(r.get('zsl'))
            #     gczj_v = _f(r.get('gczj'))
            #     gczjrmb_v = _f(r.get('gczjrmb'))

            #     gczj_use = _kptz_trunc10(gczj_v) if mlsb_v == '是' else gczj_v
            #     gczjrmb_use = _kptz_trunc10(gczjrmb_v) if mlsb_v == '是' else gczjrmb_v
            #     is_usd = ('USD' in cghbdm_v) or ('$' in cghbdm_v)

            #     if cgdq_v == '义乌':
            #         if is_usd: ywhjm += gczj_use
            #         else: ywhjr += gczj_use
            #     else:
            #         if is_usd: nbhjm += gczj_use
            #         else: nbhjr += gczj_use

            #     if zsl_v == 16: tshj16 += gczjrmb_use
            #     elif zsl_v == 13: tshj13 += gczjrmb_use
            #     elif zsl_v == 10: tshj10 += gczjrmb_use
            #     elif zsl_v == 6: tshj6 += gczjrmb_use
            #     elif zsl_v == 3: tshj3 += gczjrmb_use
            #     elif zsl_v == 1: tshj1 += gczjrmb_use
            #     elif zsl_v == 0: tshj0 += gczjrmb_use

            hlr = run_sql("select tjhl from kpnr where tjhl>0 order by rid desc limit 1")
            hhhl1 = _f((hlr[0] if hlr else {}).get('tjhl'))

            ywxj = ywhjr + (ywhjm * hhhl1 if hhhl1 > 0 else 0)
            nbxj = nbhjr + (nbhjm * hhhl1 if hhhl1 > 0 else 0)
            fxhj = ywxj + nbxj

            _fpgl_update_safe(s, fp_rid, {
                "ywdz": yw_tags.get("ywdz", ""),
                "ywxz": yw_tags.get("ywxz", ""),
                "zzsw": yw_tags.get("zzsw", ""),
                "zs_sjkpze": zs_sjkpze, "zs_bhsje": zs_bhsje, "zs_se": zs_se,
                "ms_sjkpze": ms_sjkpze, "ms_bhsje": ms_bhsje, "ms_se": ms_se,
                "qt_sjkpze": qt_sjkpze, "qt_bhsje": qt_bhsje, "qt_se": qt_se,
                "zzsw_zsje": zs_sjkpze, "zzsw_msje": ms_sjkpze, "zzsw_qtje": qt_sjkpze,
                "ywhjr": ywhjr, "ywhjm": ywhjm, "nbhjr": nbhjr, "nbhjm": nbhjm,
                "tshj0": tshj0, "tshj1": tshj1, "tshj3": tshj3, "tshj6": tshj6, "tshj10": tshj10, "tshj13": tshj13, "tshj16": tshj16,
                "ywxj": ywxj, "nbxj": nbxj, "fxhj": fxhj, "xjhj": fxhj,
                "mtime": now
            })

        # ---------- 8. hhcb ----------
        hlr = run_sql("select tjhl from kpnr where tjhl>0 order by rid desc limit 1")
        hhhl1 = _f((hlr[0] if hlr else {}).get('tjhl'))  # 修复原代码 NameError
        rows = run_sql("select rid,zwpm1,zsl,zzsl,hyd,bgbh,sjkpze from kaiptzsheet1 where fphm=:f", {"f": wxfp}) or []
        hhhl = 1 if _s(m.get('RMBkh')) == '是' else 0

        for r in rows:
            sm = run_sql("""
                select ifnull(sum(sjkpze),0) v from kaiptzsheet1
                 where fphm=:f and zwpm1=:z and zsl=:s and hyd=:h and bgbh=:b
            """, {"f": wxfp, "z": _s(r.get('zwpm1')), "s": _s(r.get('zsl')), "h": _s(r.get('hyd')), "b": _s(r.get('bgbh'))})
            total = _f((sm[0] if sm else {}).get('v'))
            bl = (_f(r.get('sjkpze')) / total) if total > 0 else 0

            bgq = run_sql("""
                select ifnull(sum(wxzj - ifnull(CNFyf,0)),0) v from bgmxdhbcp
                 where pid=:pid and zwpm=:z and tsl=:s and bgbh=:b and hyd like :h
            """, {"pid": rid, "z": _s(r.get('zwpm1')), "s": _s(r.get('zsl')), "b": _s(r.get('bgbh')), "h": f"%{_s(r.get('hyd'))}"})
            hhbgje = _f((bgq[0] if bgq else {}).get('v')) * bl

            sj = math.trunc(_f(r.get('sjkpze')) * 100) / 100
            zsl = _f(r.get('zsl')); zzsl = _f(r.get('zzsl'))
            if zsl == 3:
                hhtse = (sj / 1.03) * 0.03
            elif zsl == 1:
                hhtse = (sj / 1.01) * 0.01
            else:
                hhtse = (sj / (1 + zzsl / 100)) * zsl / 100 if zzsl > 0 else 0

            hhtse = math.trunc(hhtse * 100) / 100
            hhbgje = math.trunc(hhbgje * 100) / 100

            hhcb = 0
            if hhhl != 1 and hhbgje > 0:
                hhcb = round((sj - hhtse) / hhbgje, 2)
            elif hhhl == 1 and hhbgje > 0 and hhhl1 > 0:
                base = math.trunc((hhbgje / hhhl1) * 100) / 100
                hhcb = round((sj - hhtse) / base, 2) if base > 0 else 0

            s.query(kaiptzsheet1).filter(kaiptzsheet1.rid == r.get('rid')).update({"hhcb": hhcb, "mtime": now}, synchronize_session=False)

        # ---------- 9. 归属与提醒 ----------
        gd_path = path
        if ywrya_in:
            u2 = run_sql("select path from sys_user where username=:u limit 1", {"u": ywrya_in})
            if u2 and _s(u2[0].get('path')):
                gd_path = _s(u2[0].get('path'))[:100]

        if all_new_head_rids:
            s.query(kaiptz).filter(kaiptz.rid.in_(all_new_head_rids)).update({
                "uid": user.rid, "gdry": ywrya_in, "mtime": now
            }, synchronize_session=False)

        if _f(bgje) > 60000:
            warnings.append(f'发票[{wxfp}]报关总和超过6万美金,请检查是否正确')

        try:
            _kptz_sync_mlsb_backwrite(wxfp, s, now)
        except:
            pass

        s.commit()

        warnings = list(dict.fromkeys([x for x in warnings if x]))
        ytxxck = list(dict.fromkeys([x for x in ytxxck if x]))
        if ytxxck:
            warnings = ['预填提醒'] + ytxxck + warnings

        return json_result(0, 'ok', {
            "fphm": wxfp,
            "created_count": len(all_new_head_rids),
            "duplicated": duplicated,
            "download_filename": download_filename,
            "warnings": warnings,
            "kp3": kp3,
            "htsh": htsh,
            "yq3": yq3
        })
    except Exception:
        s.rollback()
        logger.error(trace_error())
        return json_result(-1, trace_error())
    finally:
        s.close()